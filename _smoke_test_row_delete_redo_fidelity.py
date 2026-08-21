"""Actual-App 2/3-way row-delete undo/redo/save fidelity gate.

All workbooks are disposable.  The source files are hashed before/after and
the two-way case injects a failed structural save before retrying the same
operation journal.
"""

from __future__ import annotations

import copy
import os
import tempfile

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side

import sow_merge_tool as mod
from _smoke_test_cross_workbook_row_style_replay import (
    _assert_metadata,
    _force_openpyxl_replay,
    _sha256,
    _wait_exact_ready,
)


def _ids(worksheet):
    return [worksheet.cell(row, 1).value for row in range(1, worksheet.max_row + 1)]


def _assert_hashes(expected: dict[str, str]) -> None:
    actual = {path: _sha256(path) for path in expected}
    assert actual == expected, (expected, actual)


def _assert_absent(path: str) -> None:
    assert not os.path.lexists(path), f"row-delete TemporaryDirectory not removed: {path}"


def _write_typed_book(path: str, *, include_new: bool) -> None:
    """Create the local typed fixture; the first two rows are schema only."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "S1"
    schema = [
        ["id@id", "formula@pm", "link@pm"],
        ["string", "formula", "string"],
    ]
    rows = schema + [["A", '=A3&"-formula"', "plain"]]
    if include_new:
        rows.append(["NEW", '=A4&"-formula"', "https://example.invalid/new"])
    rows.append(["C", '=A5&"-formula"' if include_new else '=A4&"-formula"', "tail"])
    for row_idx, values in enumerate(rows, start=1):
        for column_idx, value in enumerate(values, start=1):
            worksheet.cell(row_idx, column_idx).value = value
    if include_new:
        cell = worksheet["A4"]
        cell.font = Font(name="Arial", bold=True, italic=True, color="FF112233")
        cell.fill = PatternFill(fill_type="solid", fgColor="FF44AA55")
        cell.border = Border(left=Side(style="thin", color="FF334455"))
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.number_format = "0000.000"
        cell.protection = Protection(locked=False, hidden=True)
        cell.comment = Comment("row metadata comment", "SOW")
        worksheet["C4"].hyperlink = "https://example.invalid/new"
        worksheet["C4"].style = "Hyperlink"
        dim = worksheet.row_dimensions[4]
        dim.height = 29.5
        dim.hidden = True
        dim.outlineLevel = 2
        dim.collapsed = True
        dim.thickTop = True
        dim.thickBot = True
    workbook.save(path)
    workbook.close()


def _select_missing_pair(view, *, missing_side: str):
    for index, (row_a, row_b) in enumerate(view.row_pairs):
        if missing_side == "A" and row_a is None and row_b == 4:
            return index
        if missing_side == "BASE" and row_a == 4 and view._base_row_for_pair(index, (row_a, row_b)) is None:
            return index
    raise AssertionError(("missing structural pair", missing_side, tuple(view.row_pairs)))


def _assert_deleted_target(worksheet):
    assert _ids(worksheet) == ["id@id", "string", "A", "C"]
    # The adjacent formula must retain the exact result produced by the first
    # accepted delete through redo and save/reopen.
    assert worksheet.cell(4, 2).value == '=A5&"-formula"'
    assert worksheet.cell(4, 1).comment is None
    assert worksheet.cell(4, 3).hyperlink is None


def _run_two_way(root: str) -> None:
    mine = os.path.join(root, "two-mine.xlsx")
    theirs = os.path.join(root, "two-theirs.xlsx")
    _write_typed_book(mine, include_new=False)
    _write_typed_book(theirs, include_new=True)
    source_hashes = {path: _sha256(path) for path in (mine, theirs)}
    prompt = mod.SowMergeApp._schedule_formula_cache_prompt
    excel_builder = mod._build_manual_merge_output_with_excel
    mod.SowMergeApp._schedule_formula_cache_prompt = lambda _self: None
    app = None
    primary_error = None
    cleanup_errors = []

    def _cleanup(label, callback):
        try:
            callback()
        except BaseException as exc:
            cleanup_errors.append((label, exc))

    try:
        app = mod.SowMergeApp(mine, theirs)
        view = _wait_exact_ready(app)
        view.force_align_var.set(1)
        view._toggle_force_align()
        pair = _select_missing_pair(view, missing_side="A")
        topology_before = int(view._row_model_version)
        assert view._copy_selected_row("A2B", override_pair_idx=pair)
        assert app.manual_b_row_ops[-1]["kind"] == "delete_rows"
        journal_after_apply = copy.deepcopy(app.manual_b_row_ops)
        cells_after_apply = copy.deepcopy(app.manual_b_cell_ops)
        topology_after_apply = int(view._row_model_version)
        assert topology_after_apply > topology_before
        _assert_deleted_target(app.ws_b_edit("S1"))

        view._undo_last_action()
        assert _ids(app.ws_b_edit("S1")) == ["id@id", "string", "A", "NEW", "C"]
        _assert_metadata(app.ws_b_edit("S1"), row=4, expected_formula='=A4&"-formula"')
        assert app.manual_b_row_ops == []
        topology_after_undo = int(view._row_model_version)
        assert topology_after_undo > topology_after_apply

        view._redo_last_action()
        _assert_deleted_target(app.ws_b_edit("S1"))
        assert app.manual_b_row_ops == journal_after_apply
        assert app.manual_b_cell_ops == cells_after_apply
        assert int(view._row_model_version) > topology_after_undo

        mod._build_manual_merge_output_with_excel = lambda *_args, **_kwargs: False
        try:
            app.build_manual_b_output_file()
        except RuntimeError:
            pass
        else:
            raise AssertionError("injected row-delete save failure unexpectedly succeeded")
        assert {path: _sha256(path) for path in source_hashes} == source_hashes
        assert app.manual_b_row_ops == journal_after_apply

        _force_openpyxl_replay(excel_builder)
        output = app.build_manual_b_output_file()
        reopened = load_workbook(output, data_only=False)
        try:
            _assert_deleted_target(reopened["S1"])
        finally:
            reopened.close()
        assert {path: _sha256(path) for path in source_hashes} == source_hashes
    except BaseException as exc:
        primary_error = exc
        raise
    finally:
        _cleanup(
            "restore two-way formula prompt scheduler",
            lambda: setattr(mod.SowMergeApp, "_schedule_formula_cache_prompt", prompt),
        )
        _cleanup(
            "restore two-way Excel builder",
            lambda: setattr(mod, "_build_manual_merge_output_with_excel", excel_builder),
        )
        if app is not None:
            _cleanup("shutdown two-way app", app._shutdown_root)
        _cleanup("verify two-way immutable inputs", lambda: _assert_hashes(source_hashes))
        if primary_error is not None:
            for label, cleanup_exc in cleanup_errors:
                primary_error.add_note(
                    f"secondary cleanup failure [{label}]: {type(cleanup_exc).__name__}: {cleanup_exc}"
                )
        elif cleanup_errors:
            label, cleanup_exc = cleanup_errors[0]
            raise AssertionError(f"row-delete two-way cleanup failure [{label}]: {cleanup_exc}") from cleanup_exc


def _run_three_way(root: str) -> None:
    base = os.path.join(root, "three-base.xlsx")
    mine = os.path.join(root, "three-mine.xlsx")
    theirs = os.path.join(root, "three-theirs.xlsx")
    merged = os.path.join(root, "three-merged.xlsx")
    _write_typed_book(base, include_new=False)
    _write_typed_book(mine, include_new=True)
    _write_typed_book(theirs, include_new=False)
    source_hashes = {path: _sha256(path) for path in (base, mine, theirs)}
    prompt = mod.SowMergeApp._schedule_formula_cache_prompt
    excel_builder = mod._build_manual_merge_output_with_excel
    mod.SowMergeApp._schedule_formula_cache_prompt = lambda _self: None
    app = None
    primary_error = None
    cleanup_errors = []

    def _cleanup(label, callback):
        try:
            callback()
        except BaseException as exc:
            cleanup_errors.append((label, exc))

    try:
        app = mod.SowMergeApp(mine, theirs, merge_mode=True, merged_path=merged, base_path=base)
        view = _wait_exact_ready(app)
        view.force_align_var.set(1)
        view._toggle_force_align()
        pair = _select_missing_pair(view, missing_side="BASE")
        assert view._base_row_for_pair(pair, view.row_pairs[pair]) is None
        topology_before = int(view._row_model_version)
        assert view._copy_selected_row("BASE2A", override_pair_idx=pair)
        assert app.manual_a_row_ops[-1]["kind"] == "delete_rows"
        journal_after_apply = copy.deepcopy(app.manual_a_row_ops)
        _assert_deleted_target(app.ws_a_edit("S1"))
        topology_after_apply = int(view._row_model_version)
        assert topology_after_apply > topology_before

        view._undo_last_action()
        assert _ids(app.ws_a_edit("S1")) == ["id@id", "string", "A", "NEW", "C"]
        _assert_metadata(app.ws_a_edit("S1"), row=4, expected_formula='=A4&"-formula"')
        assert app.manual_a_row_ops == []
        topology_after_undo = int(view._row_model_version)
        assert topology_after_undo > topology_after_apply

        view._redo_last_action()
        _assert_deleted_target(app.ws_a_edit("S1"))
        assert app.manual_a_row_ops == journal_after_apply
        assert int(view._row_model_version) > topology_after_undo

        _force_openpyxl_replay(excel_builder)
        output = app.build_manual_merge_output_file()
        reopened = load_workbook(output, data_only=False)
        try:
            _assert_deleted_target(reopened["S1"])
        finally:
            reopened.close()
        assert {path: _sha256(path) for path in source_hashes} == source_hashes
    except BaseException as exc:
        primary_error = exc
        raise
    finally:
        _cleanup(
            "restore three-way formula prompt scheduler",
            lambda: setattr(mod.SowMergeApp, "_schedule_formula_cache_prompt", prompt),
        )
        _cleanup(
            "restore three-way Excel builder",
            lambda: setattr(mod, "_build_manual_merge_output_with_excel", excel_builder),
        )
        if app is not None:
            _cleanup("shutdown three-way app", app._shutdown_root)
        _cleanup("verify three-way immutable inputs", lambda: _assert_hashes(source_hashes))
        if primary_error is not None:
            for label, cleanup_exc in cleanup_errors:
                primary_error.add_note(
                    f"secondary cleanup failure [{label}]: {type(cleanup_exc).__name__}: {cleanup_exc}"
                )
        elif cleanup_errors:
            label, cleanup_exc = cleanup_errors[0]
            raise AssertionError(f"row-delete three-way cleanup failure [{label}]: {cleanup_exc}") from cleanup_exc


def main() -> None:
    primary_error = None
    cleanup_errors = []
    temporary = tempfile.TemporaryDirectory(prefix="sow_row_delete_redo_fidelity_")
    root = temporary.name

    def _cleanup(label, callback):
        try:
            callback()
        except BaseException as exc:
            cleanup_errors.append((label, exc))

    try:
        _run_two_way(root)
        _run_three_way(root)
    except BaseException as exc:
        primary_error = exc
        raise
    finally:
        _cleanup("remove own TemporaryDirectory", temporary.cleanup)
        _cleanup("verify TemporaryDirectory removal", lambda: _assert_absent(root))
        if primary_error is not None:
            for label, cleanup_exc in cleanup_errors:
                primary_error.add_note(
                    f"secondary cleanup failure [{label}]: {type(cleanup_exc).__name__}: {cleanup_exc}"
                )
        elif cleanup_errors:
            label, cleanup_exc = cleanup_errors[0]
            raise AssertionError(f"row-delete main cleanup failure [{label}]: {cleanup_exc}") from cleanup_exc
    print("SMOKE_ROW_DELETE_REDO_FIDELITY_OK")


if __name__ == "__main__":
    main()

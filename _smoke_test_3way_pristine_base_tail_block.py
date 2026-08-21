"""B3 GUI regression: pristine Base tail tooltip remains immutable and bound."""
from __future__ import annotations

import argparse
from contextlib import contextmanager
import hashlib
import json
import os
import sys
import tempfile
import time
from pathlib import Path
from types import SimpleNamespace

from openpyxl import Workbook
from openpyxl.worksheet._read_only import ReadOnlyWorksheet
from openpyxl.worksheet.worksheet import Worksheet

import sow_merge_tool as sm

_CASE = "pristine-base-tail-tooltip"
_SHEET = "S1"
_ROWS = 2200
_SCHEMA_ROWS = 2
_TAIL_ID = _ROWS
_TAIL_ROW = _SCHEMA_ROWS + _TAIL_ID
_CASE_TIMEOUT = 90.0


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _path_snapshot(path: Path) -> tuple[bool, bytes | None]:
    return (True, path.read_bytes()) if path.exists() else (False, None)


def _canon(value):
    if value is None or isinstance(value, (bool, int, float, str, bytes)):
        return value
    if isinstance(value, dict):
        return tuple(sorted(((_canon(key), _canon(item)) for key, item in value.items()), key=repr))
    if isinstance(value, (tuple, list, set, frozenset)):
        items = tuple(_canon(item) for item in value)
        return tuple(sorted(items, key=repr)) if isinstance(value, (set, frozenset)) else items
    return ("opaque", type(value).__name__, id(value))


def _digest(value) -> str:
    return hashlib.sha256(repr(_canon(value)).encode("utf-8")).hexdigest()


def _fact(value) -> dict:
    try:
        length = len(value)
    except Exception:
        length = None
    preview = repr(_canon(value))
    return {"type": type(value).__name__, "len": length, "digest": _digest(value), "preview": preview[:180]}


def _diff(before, after, path: str = "hard") -> list[dict]:
    if before == after:
        return []
    if isinstance(before, dict) and isinstance(after, dict):
        result = []
        for key in sorted(set(before) | set(after), key=repr):
            child = f"{path}[{key!r}]"
            if key not in before or key not in after:
                result.append({"path": child, "before": _fact(before.get(key)), "after": _fact(after.get(key))})
            else:
                result.extend(_diff(before[key], after[key], child))
        return result
    if isinstance(before, (tuple, list)) and isinstance(after, (tuple, list)):
        result = []
        if len(before) != len(after):
            result.append({"path": f"{path}.len", "before": _fact(before), "after": _fact(after)})
        for index, (left, right) in enumerate(zip(before, after)):
            result.extend(_diff(left, right, f"{path}[{index}]"))
        return result
    return [{"path": path, "before": _fact(before), "after": _fact(after)}]


def _assert_same(before, after, action: str) -> None:
    changes = _diff(before, after)
    assert not changes, f"{action}: immutable hard drift {json.dumps(changes, ensure_ascii=False, sort_keys=True)}"


def _model_projection_fact(view) -> dict:
    cache = view.column_comparison_cache
    projection = view.column_projection
    model = cache.model
    assert projection.model is model
    slots = tuple(
        (
            int(slot.logical_idx),
            getattr(slot, "mine_col", None),
            getattr(slot, "base_col", None),
            getattr(slot, "theirs_col", None),
            str(getattr(slot, "state", "")),
            str(getattr(slot, "origin_side", "") or ""),
        )
        for slot in model.slots
    )
    semantics = {
        "slots": slots,
        "structural": tuple(sorted(cache.structural_diff_cols)),
        "unresolved": tuple(sorted(cache.unresolved_cols)),
        "projection_blocks": tuple(projection.block_ordinal_by_slot),
        "maps": tuple(
            (name, tuple(getattr(getattr(model, name), "entries", ())))
            for name in (
                "mine_physical_to_logical", "base_physical_to_logical", "theirs_physical_to_logical",
                "mine_logical_to_physical", "base_logical_to_physical", "theirs_logical_to_physical",
            )
        ),
    }
    return {
        "identity": (id(cache), id(model), id(projection)),
        "bounds": (view.max_row, view.max_col, view.col_max_a, view.col_max_b, view.col_max_base),
        "flags": (view._align_rows_enabled, view._sheet_structural_diff, view._only_diff_source_version),
        "semantics": _fact(semantics),
    }


def _hard_snapshot(app, view, paths: dict[str, Path]) -> dict:
    overlays = app.sheet_operation_overlays or {}
    overlay = overlays.get(view.sheet)
    row_maps = {
        "row_pairs": tuple(tuple(pair) for pair in view.row_pairs),
        "a_to_pair": dict(view.row_a_to_pair_idx),
        "b_to_pair": dict(view.row_b_to_pair_idx),
        "mine_to_base": dict(view.mine_to_base_row),
        "theirs_to_base": dict(view.theirs_to_base_row),
        "base_override": dict(view.pair_base_row_override),
        "missing_base": dict(view._missing_base_row_map),
    }
    return {
        "input_sha": tuple(sorted((name, _sha256(path)) for name, path in paths.items())),
        "manual": _fact({name: getattr(app, name) for name in (
            "manual_a_cell_ops", "manual_b_cell_ops", "manual_a_formula_cache_ops", "manual_b_formula_cache_ops",
            "manual_a_row_ops", "manual_b_row_ops", "manual_a_column_ops", "manual_b_column_ops",
            "manual_sheet_ops", "auto_sheet_ops",
        )}),
        "undo_redo": _fact((app.undo_stack, app.redo_stack)),
        "modified_touched": _fact((
            app.modified_a, app.modified_b, app.modified_sheets_a, app.modified_sheets_b,
            app.user_touched_conflicts, view.touched_rows,
        )),
        "overlay": _fact({
            name: (getattr(item, "topology_generation", None), getattr(item, "mutation_generation", None), getattr(item, "cells", None))
            for name, item in overlays.items()
        }),
        "row_maps": _fact(row_maps),
        "prepared": {
            "raw_a": _fact(dict(view.pair_raw_parts_a)), "raw_b": _fact(dict(view.pair_raw_parts_b)),
            "raw_base": _fact(dict(view.pair_raw_parts_base)), "ab_diffs": _fact(dict(view.pair_diff_cols)),
            "base_diffs": _fact(dict(view.pair_base_diff_cols)),
            "exactness": (view._prepared_complete, view._data_ready, view._row_model_exact, view._pair_diff_full_exact, view._base_diff_full_exact),
        },
        "model_projection": _model_projection_fact(view),
        "generations": (
            app._sheet_compute_generation.get(view.sheet), view._row_model_version, view._column_model_version,
            view._column_projection_generation, view._virtual_column_window_generation, view._data_version,
            view._virtual_publish_generation, getattr(overlay, "topology_generation", None), getattr(overlay, "mutation_generation", None),
        ),
        "edit_handles": tuple(
            (name, id(getattr(app, name, None)), type(getattr(app, name, None)).__name__, getattr(getattr(app, name, None), "read_only", None))
            for name in ("_wb_a_edit", "_wb_b_edit", "_wb_base_edit")
        ),
    }


def _make_book(path: Path, *, include_tail: bool) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _SHEET
    sheet.append(["id@id", "value"])
    sheet.append(["int32", "string"])
    for row_id in range(1, _ROWS + 1):
        if row_id == _TAIL_ID and not include_tail:
            continue
        sheet.append([row_id, "pristine-tail" if row_id == _TAIL_ID else f"value-{row_id}"])
    workbook.save(path)
    workbook.close()


def _pump(app) -> None:
    app.root.update()
    app.root.update_idletasks()


def _wait(app, predicate, deadline: float, stage: str) -> None:
    while time.monotonic() < deadline:
        _pump(app)
        if predicate():
            return
        view = app.sheet_views.get(_SHEET)
        if view is not None and view._derive_lifecycle_state() in {"FAILED", "UNRESOLVED", "CANCELED", "CLOSING"}:
            raise AssertionError(f"{stage}: lifecycle={view._derive_lifecycle_state()} error={getattr(view, '_lifecycle_error', None)!r} entry={app._sheet_exact_entry(_SHEET)!r}")
        time.sleep(0.005)
    view = app.sheet_views.get(_SHEET)
    facts = None if view is None else {
        "prepared": view._prepared_complete, "data_ready": view._data_ready, "row_model_exact": view._row_model_exact,
        "pending": view._pending_exact_render, "lifecycle": view._derive_lifecycle_state(),
    }
    raise AssertionError(f"timeout {stage}: selected={app.selected_sheet!r} entry={app._sheet_exact_entry(_SHEET)!r} view={facts!r}")


def _new_app(mine: Path, theirs: Path, base: Path):
    original_scheduler = sm.SowMergeApp._schedule_formula_cache_prompt
    sm.SowMergeApp._schedule_formula_cache_prompt = lambda _self: None
    try:
        return (sm.SowMergeApp(str(mine), str(theirs), merge_mode=True, base_path=str(base), initial_sheet=_SHEET), original_scheduler)
    except Exception:
        sm.SowMergeApp._schedule_formula_cache_prompt = original_scheduler
        raise


def _cancel_debounce(app) -> None:
    if app is None:
        return
    for view in tuple(getattr(app, "sheet_views", {}).values()):
        if view is None:
            continue
        for attr in ("_settings_save_id", "_hover_debounce_id", "_diff_map_debounce_id"):
            after_id = getattr(view, attr, None)
            if after_id:
                try:
                    view.frame.after_cancel(after_id)
                finally:
                    setattr(view, attr, None)


def _shutdown(app) -> None:
    if app is not None:
        app._shutdown_root()


@contextmanager
def _view_only_traps(app, view):
    hits = []
    originals = []

    def forbid(label):
        def fail(*_args, **_kwargs):
            hits.append(label)
            raise AssertionError(f"view-only tail hover accessed {label}")
        return fail

    targets = (
        (app, "ws_a_val"), (app, "ws_b_val"), (app, "ws_base_val"), (app, "ws_a_edit"), (app, "ws_b_edit"), (app, "ws_base_edit"),
        (app, "_request_edit_preload"), (app, "_ensure_edit_loaded"), (app, "_load_edit_workbooks_owned"), (app, "_start_background_thread"),
        (app, "_atomic_save"), (app, "_atomic_save_with_retry"), (app, "_atomic_replace_file_with_retry"), (app, "_try_alt_save"),
        (app, "build_manual_b_output_file"), (app, "build_manual_merge_output_file"),
        (app, "save_a_inplace"), (app, "save_b_inplace"), (app, "save_merged_and_exit"),
        (view, "refresh"), (view, "_refresh_mode_switch_preserving_selection"), (view, "_start_async_large_only_diff_build"),
        (view, "_run_copy_action_by_mode"), (view, "_apply_global_sheet_overwrite"), (view, "_apply_selected_column_block"),
        (sm, "_align_selected_sheet_snapshots"), (sm, "_compare_selected_sheet_snapshots"), (sm, "_atomic_save_wb"),
    )
    original_cell, original_iter, original_ro_iter = Worksheet.cell, Worksheet.iter_rows, ReadOnlyWorksheet.iter_rows
    try:
        for owner, name in targets:
            if hasattr(owner, name):
                originals.append((owner, name, getattr(owner, name)))
                setattr(owner, name, forbid(f"{type(owner).__name__}.{name}"))
        Worksheet.cell = forbid("Worksheet.cell")
        Worksheet.iter_rows = forbid("Worksheet.iter_rows")
        ReadOnlyWorksheet.iter_rows = forbid("ReadOnlyWorksheet.iter_rows")
        yield hits
    finally:
        Worksheet.cell, Worksheet.iter_rows, ReadOnlyWorksheet.iter_rows = original_cell, original_iter, original_ro_iter
        for owner, name, original in reversed(originals):
            setattr(owner, name, original)


def _event_for_main_cell(app, widget, view, pair_idx: int, logical_col: int, deadline: float) -> SimpleNamespace:
    _wait(app, lambda: pair_idx in view.display_rows and not view._virtual_publishing, deadline, "tail visible")
    line = int(view.row_to_line[pair_idx])
    start, end = view._spans_for_line()[logical_col]
    index = f"{line}.{start + 1 if end - start > 1 else start}"
    widget.see(index)
    _pump(app)
    bbox = widget.bbox(index) or widget.dlineinfo(index)
    assert bbox is not None, (index, widget.index("@0,0"))
    x, y, width, height = bbox
    event = SimpleNamespace(x=int(x + max(1, width // 2)), y=int(y + max(1, height // 2)))
    target_line, target_col = map(int, str(widget.index(f"@{event.x},{event.y}")).split("."))
    assert target_line == line and start <= target_col < max(start + 1, end), (target_line, target_col, line, start, end)
    return event


def _event_for_cursor(app, view, logical_col: int) -> SimpleNamespace:
    widget = view.cursor_cmp
    start, end = view._spans_for_line()[logical_col]
    index = f"1.{start + 1 if end - start > 1 else start}"
    widget.see(index)
    _pump(app)
    bbox = widget.bbox(index) or widget.dlineinfo(index)
    assert bbox is not None, (index, widget.index("@0,0"))
    x, y, width, height = bbox
    event = SimpleNamespace(x=int(x + max(1, width // 2)), y=int(y + max(1, height // 2)))
    _line, target_col = map(int, str(widget.index(f"@{event.x},{event.y}")).split("."))
    assert start <= target_col < max(start + 1, end), (target_col, start, end)
    return event


def _panel_text(view) -> str:
    return str(view.hover_cmp_text.get("1.0", "end-1c"))


def _selection_fact(view):
    return (
        view.selected_pair_idx, view.selected_excel_row, view.selected_excel_row_a, view.selected_excel_row_b,
        view._main_sel_line, view._main_sel_col, view._cursor_cmp_sel_line, view._cursor_cmp_sel_col,
    )


def _tail_pair(view) -> int:
    candidates = [
        int(pair_idx) for pair_idx, pair in enumerate(view.row_pairs)
        if pair[0] is None
        and tuple(view.pair_raw_parts_a.get(pair_idx, ())) == ("【此侧缺行】", "")
        and tuple(view.pair_raw_parts_b.get(pair_idx, ())) == (str(_TAIL_ID), "pristine-tail")
        and tuple(view.pair_raw_parts_base.get(pair_idx, ())) == (str(_TAIL_ID), "pristine-tail")
    ]
    assert len(candidates) == 1, candidates
    pair_idx = candidates[0]
    assert tuple(view.row_pairs[pair_idx]) == (None, _TAIL_ROW)
    assert view.pair_base_row_override.get(pair_idx) == _TAIL_ROW
    assert view._pair_has_visual_diff(pair_idx)
    return pair_idx


def _next_pair(view) -> int:
    expected = str(_TAIL_ID - 1)
    candidates = [
        int(pair_idx) for pair_idx, pair in enumerate(view.row_pairs)
        if pair[0] is not None and pair[1] is not None
        and tuple(view.pair_raw_parts_a.get(pair_idx, ())) == (expected, f"value-{_TAIL_ID - 1}")
        and tuple(view.pair_raw_parts_b.get(pair_idx, ())) == (expected, f"value-{_TAIL_ID - 1}")
        and tuple(view.pair_raw_parts_base.get(pair_idx, ())) == (expected, f"value-{_TAIL_ID - 1}")
    ]
    assert len(candidates) == 1, candidates
    return candidates[0]


def _tail_vminimap_publication_current(view, tail_pair: int, before_publications: int) -> bool:
    active = getattr(view, "_viewport_request_active", None) or {}
    publication = getattr(view, "_last_virtual_publication_telemetry", None) or {}
    window = publication.get("window", {}) if isinstance(publication, dict) else {}
    row_start = int(window.get("row_start", -1))
    row_count = int(window.get("row_count", -1))
    return bool(
        tail_pair in tuple(view.display_rows) and not view._virtual_publishing
        and int(getattr(view, "_virtual_scroll_publications", 0)) > before_publications
        and isinstance(active, dict) and active.get("status") == "complete" and active.get("reason") == "vminimap"
        and active.get("selected_sheet") == _SHEET and int(active.get("actual_row_start", -1)) == int(view._virtual_window_start)
        and isinstance(publication, dict) and publication.get("selected_sheet") == _SHEET
        and int(publication.get("request_id", -1)) == int(active.get("id", -2))
        and row_start == int(view._virtual_window_start) and row_start <= tail_pair < row_start + row_count
    )


def _run_case() -> None:
    # Global natural-case budget includes fixture construction and cleanup-sensitive setup.
    deadline = time.monotonic() + _CASE_TIMEOUT
    original_settings_path = sm._SETTINGS_PATH
    user_settings = Path(original_settings_path)
    user_before = _path_snapshot(user_settings)
    app = None
    root_path = None
    scheduler = None
    input_paths: dict[str, Path] = {}
    input_before = {}
    primary = None
    try:
        with tempfile.TemporaryDirectory(prefix="sow_pristine_tail_b3_") as raw_root:
            root_path = Path(raw_root)
            mine, theirs, base = root_path / "mine.xlsx", root_path / "theirs.xlsx", root_path / "base.xlsx"
            _make_book(mine, include_tail=False)
            _make_book(theirs, include_tail=True)
            _make_book(base, include_tail=True)
            input_paths = {"mine": mine, "theirs": theirs, "base": base}
            input_before = {name: _sha256(path) for name, path in input_paths.items()}
            settings = root_path / "settings.json"
            settings.write_text(json.dumps({"only_diff": 0}) + "\n", encoding="utf-8")
            sm._SETTINGS_PATH = str(settings)
            try:
                print("PRISTINE_TAIL_STAGE open-current-exact", flush=True)
                app, scheduler = _new_app(mine, theirs, base)
                view = app.sheet_views[_SHEET]
                _wait(app, lambda: (
                    app.selected_sheet == _SHEET and app._is_sheet_exact_current(_SHEET)
                    and bool(app._sheet_exact_entry(_SHEET).get("full_detail_terminal"))
                    and view._prepared_complete and view._data_ready and view._row_model_exact and not view._pending_exact_render
                    and view._is_exact_immutable_view_ready() and view._derive_lifecycle_state() == "EDIT_DEFERRED"
                    and not app._edit_workbooks_ready() and not view.only_diff_var.get()
                ), deadline, "selected immutable exact full detail")
                assert len(view.row_pairs) == _ROWS + _SCHEMA_ROWS
                assert view._virtual_mode_active()
                tail_pair, next_pair = _tail_pair(view), _next_pair(view)
                selection_before = _selection_fact(view)
                requests_before = tuple(app._edit_load_requests)
                hard_before = _hard_snapshot(app, view, input_paths)
                payload_calls = 0
                c_renders = 0
                original_payload = view._cmp_tooltip_payload_by_pair_col
                original_cursor = view._update_cursor_lines

                def counted_payload(*args, **kwargs):
                    nonlocal payload_calls
                    payload_calls += 1
                    return original_payload(*args, **kwargs)

                def counted_cursor(*args, **kwargs):
                    nonlocal c_renders
                    c_renders += 1
                    return original_cursor(*args, **kwargs)

                view._cmp_tooltip_payload_by_pair_col = counted_payload
                view._update_cursor_lines = counted_cursor
                try:
                    with _view_only_traps(app, view) as hits:
                        print("PRISTINE_TAIL_STAGE bound-tail-hover", flush=True)
                        vdiff_map = view.vdiff_map
                        assert str(vdiff_map.bind("<Button-1>") or "").strip()
                        publications_before = int(view._virtual_scroll_publications)
                        vdiff_map.event_generate("<Button-1>", x=1, y=max(1, int(vdiff_map.winfo_height())))
                        _wait(app, lambda: _tail_vminimap_publication_current(view, tail_pair, publications_before), deadline, "bound diff-map tail publication")
                        _pump(app)
                        _assert_same(hard_before, _hard_snapshot(app, view, input_paths), "tail diff-map route")
                        assert tuple(app._edit_load_requests) == requests_before and not app._edit_workbooks_ready()
                        # Publication may render C while no hover payload exists; all hover assertions are relative to this real baseline.
                        payload_base, cursor_base = payload_calls, c_renders
                        assert payload_base == 0, payload_base
                        assert cursor_base >= 1, cursor_base
                        assert str(view.right.bind("<Motion>") or "").strip()
                        tail_event = _event_for_main_cell(app, view.right, view, tail_pair, 2, deadline)
                        view.right.event_generate("<Motion>", x=tail_event.x, y=tail_event.y)
                        _wait(app, lambda: payload_calls == payload_base + 1 and c_renders == cursor_base + 1 and view.hover_pair_idx == tail_pair, deadline, "first bound tail hover")
                        _pump(app)
                        tail_panel = _panel_text(view)
                        assert f"base[{_TAIL_ROW}]: pristine-tail" in tail_panel, tail_panel
                        assert "mine[-]: <missing>" in tail_panel, tail_panel
                        assert f"theirs[{_TAIL_ROW}]: pristine-tail" in tail_panel, tail_panel
                        _assert_same(hard_before, _hard_snapshot(app, view, input_paths), "first tail Motion")
                        assert _selection_fact(view) == selection_before
                        assert tuple(app._edit_load_requests) == requests_before and not app._edit_workbooks_ready()

                        view.right.event_generate("<Motion>", x=tail_event.x, y=tail_event.y)
                        _pump(app)
                        _assert_same(hard_before, _hard_snapshot(app, view, input_paths), "repeat tail Motion")
                        assert payload_calls == payload_base + 1 and c_renders == cursor_base + 1
                        assert _panel_text(view) == tail_panel and _selection_fact(view) == selection_before

                        view.right.event_generate("<Leave>")
                        _pump(app)
                        _assert_same(hard_before, _hard_snapshot(app, view, input_paths), "tail Leave")
                        assert _panel_text(view) == tail_panel and payload_calls == payload_base + 1 and c_renders == cursor_base + 1
                        assert tuple(app._edit_load_requests) == requests_before and not app._edit_workbooks_ready()

                        assert str(view.cursor_cmp.bind("<Motion>") or "").strip()
                        cursor_event = _event_for_cursor(app, view, 2)
                        payload_before_cursor, c_before_cursor = payload_calls, c_renders
                        assert (payload_before_cursor, c_before_cursor) == (payload_base + 1, cursor_base + 1)
                        view.cursor_cmp.event_generate("<Motion>", x=cursor_event.x, y=cursor_event.y)
                        _wait(app, lambda: view.hover_pair_idx == tail_pair and view.hover_side == "C", deadline, "bound C hover suppress")
                        _pump(app)
                        _assert_same(hard_before, _hard_snapshot(app, view, input_paths), "C Motion suppress")
                        assert c_renders == cursor_base + 1, (cursor_base, c_renders)
                        assert payload_calls in {payload_base + 1, payload_base + 2}, (payload_base, payload_calls)
                        assert _panel_text(view) == tail_panel and _selection_fact(view) == selection_before

                        payload_before_next, c_before_next = payload_calls, c_renders
                        next_event = _event_for_main_cell(app, view.right, view, next_pair, 2, deadline)
                        view.right.event_generate("<Motion>", x=next_event.x, y=next_event.y)
                        _wait(app, lambda: payload_calls == payload_before_next + 1 and c_renders == c_before_next + 1 and view.hover_pair_idx == next_pair, deadline, "new bound hover")
                        _pump(app)
                        _assert_same(hard_before, _hard_snapshot(app, view, input_paths), "new tail-neighbor Motion")
                        assert _selection_fact(view) == selection_before
                        assert tuple(app._edit_load_requests) == requests_before and not app._edit_workbooks_ready()
                    assert not hits, hits
                finally:
                    view._cmp_tooltip_payload_by_pair_col = original_payload
                    view._update_cursor_lines = original_cursor

                assert {name: _sha256(path) for name, path in input_paths.items()} == input_before
                assert time.monotonic() <= deadline
                print("PRISTINE_TAIL_DIAGNOSTICS " + json.dumps({
                    "tail_pair": tail_pair, "tail_row": _TAIL_ROW, "next_pair": next_pair,
                    "payload_calls": payload_calls, "c_render_calls": c_renders, "deadline_seconds": _CASE_TIMEOUT,
                }, sort_keys=True), flush=True)
            except BaseException as exc:
                primary = exc
                raise
            finally:
                cleanup_errors = []

                def checked(label, callback):
                    try:
                        callback()
                    except BaseException as exc:
                        cleanup_errors.append(f"{label}: {type(exc).__name__}: {exc}")

                checked("input SHA before shutdown", lambda: assert_input_hashes(input_paths, input_before))
                checked("temporary settings path", lambda: _assert_settings_path(sm._SETTINGS_PATH, settings))
                checked("cancel settings/hover debounce", lambda: _cancel_debounce(app))
                checked("shutdown", lambda: _shutdown(app))
                checked("input SHA after shutdown", lambda: assert_input_hashes(input_paths, input_before))
                checked("user settings unchanged", lambda: _assert_user_settings(user_settings, user_before))
                if scheduler is not None:
                    sm.SowMergeApp._schedule_formula_cache_prompt = scheduler
                    scheduler = None
                if cleanup_errors:
                    text = "; ".join(cleanup_errors)
                    if primary is not None:
                        primary.add_note("cleanup secondary failures: " + text)
                    else:
                        raise AssertionError(text)
    finally:
        if scheduler is not None:
            sm.SowMergeApp._schedule_formula_cache_prompt = scheduler
        sm._SETTINGS_PATH = original_settings_path
        active_exception = primary if primary is not None else sys.exc_info()[1]

        def checked_outer(label, callback) -> None:
            try:
                callback()
            except BaseException as exc:
                if active_exception is not None:
                    active_exception.add_note(f"outer cleanup secondary {label}: {type(exc).__name__}: {exc}")
                else:
                    raise

        checked_outer("user settings", lambda: _assert_user_settings(user_settings, user_before))
        if root_path is not None:
            checked_outer("temporary root removed", lambda: _assert_absent(root_path))


def _assert_settings_path(actual, expected) -> None:
    assert os.fspath(actual) == os.fspath(expected), (actual, expected)


def _assert_user_settings(path: Path, before: tuple[bool, bytes | None]) -> None:
    assert _path_snapshot(path) == before


def _assert_absent(path: Path) -> None:
    assert not path.exists(), path


def assert_input_hashes(paths: dict[str, Path], before: dict[str, str]) -> None:
    assert {name: _sha256(path) for name, path in paths.items()} == before


def main(argv=None) -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--list-cases", action="store_true")
    parser.add_argument("--case", choices=(_CASE,))
    args = parser.parse_args(argv)
    if args.list_cases:
        if args.case:
            parser.error("--list-cases cannot be combined with --case")
        print(_CASE, flush=True)
        return
    selected = (args.case,) if args.case else (_CASE,)
    for case in selected:
        print(f"PRISTINE_TAIL_CASE_START {case}", flush=True)
        _run_case()
        print(f"PRISTINE_TAIL_CASE_OK {case}", flush=True)
    print(f"PRISTINE_TAIL_SUITE_OK ({len(selected)} cases)", flush=True)


if __name__ == "__main__":
    main()

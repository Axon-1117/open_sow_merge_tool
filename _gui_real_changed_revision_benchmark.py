"""Stable, disposable 11.7 changed-revision GUI release gate.

The revision exports are treated as immutable input evidence.  Every selected
case copies them to its own temporary root, redirects settings there, drives
only public Tk routes, and leaves its JSON report inside that root.  The former
diagnostic harness is retained beside this file as a hash-pinned compatibility
module for its independent snapshot/Oracle helpers and their focused pure
tests; it is never used to drive a GUI interaction here.
"""
from __future__ import annotations

import argparse
import contextlib
import copy
import hashlib
import importlib.util
import json
import math
import os
import shutil
import stat
import tempfile
import time
import traceback
from pathlib import Path

import psutil
from openpyxl.worksheet._read_only import ReadOnlyWorksheet
from openpyxl.worksheet.worksheet import Worksheet

_ENV_CALLBACK_TIMING_BEFORE = os.environ.get("SOW_CALLBACK_TIMING_DIAGNOSTIC")
os.environ.setdefault("SOW_CALLBACK_TIMING_DIAGNOSTIC", "1")
import sow_merge_tool as sm

_ROOT = Path(__file__).resolve().parent
_LEGACY_PATH = _ROOT / "_gui_real_changed_revision_benchmark.codex_37E_20260820_022900.py"
_LEGACY_SHA256 = "37E82027D76F916412A891214C0575915C18FB771CB8CB1021DDDC710A492E09"
_SOURCE_SHA256 = "6E76313ED17FDA66333A8AD90BE8059ADBDF31BAFD3B61146A255229F2FC2A47"
REVISION_DIR = Path(r"C:\Users\dd\AppData\Local\Temp\sow_revision_export_39264_39265")
OLD_REVISION, NEW_REVISION = 39264, 39265
SHEETS = ("Dungeon@design", "MonsterGroup@design")
GATE_EXACT_MS, GATE_CALLBACK_MS, GATE_E2E_MAX_MS = 15_000.0, 33.0, 66.0
GATE_HEARTBEAT_MS, CASE_DEADLINE_SECONDS = 200.0, 90.0

_CASES = {
    "two-way-dungeon-release": {"mode": "2way", "anchor": SHEETS[0]},
    "two-way-monstergroup-release": {"mode": "2way", "anchor": SHEETS[1]},
    "three-way-dungeon-release": {"mode": "3way", "anchor": SHEETS[0]},
    "three-way-monstergroup-release": {"mode": "3way", "anchor": SHEETS[1]},
}
_CASE_ORDER = tuple(_CASES)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest().upper()


def _file_fact(path: Path) -> dict:
    resolved = path.resolve()
    info = resolved.stat()
    return {"path": str(resolved), "size": int(info.st_size), "sha256": _sha256(resolved)}


def _same_bytes(left: dict, right: dict) -> bool:
    return (int(left["size"]), str(left["sha256"])) == (int(right["size"]), str(right["sha256"]))


def _load_legacy():
    assert _LEGACY_PATH.is_file() and _sha256(_LEGACY_PATH) == _LEGACY_SHA256
    spec = importlib.util.spec_from_file_location("_changed_revision_legacy_37e", _LEGACY_PATH)
    assert spec is not None and spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


_legacy = _load_legacy()


def __getattr__(name):
    """Keep legacy immutable-oracle focused pure imports source-compatible."""
    return getattr(_legacy, name)


def _p95(values) -> float:
    ordered = sorted(float(value) for value in values)
    assert ordered
    return ordered[min(len(ordered) - 1, int((len(ordered) - 1) * .95))]


def _canonical(value):
    if value is None or isinstance(value, (bool, int, float, str)):
        return value
    if isinstance(value, (tuple, list)):
        return tuple(_canonical(item) for item in value)
    if isinstance(value, dict):
        return tuple(sorted((str(key), _canonical(item)) for key, item in value.items()))
    if isinstance(value, (set, frozenset)):
        return tuple(sorted(_canonical(item) for item in value))
    return ("opaque", type(value).__name__, id(value))


def _bounded_lookup_digest(lookup, *, limit: int = 96) -> dict:
    """Bounded, canonical evidence for immutable column lookup directions."""
    entries = tuple(
        (int(key), int(value))
        for key, value in tuple(getattr(lookup, "entries", ()) or ())
    )
    assert entries == tuple(sorted(entries)), entries
    if len(entries) <= int(limit):
        return {"entry_count": len(entries), "entries": entries}
    edge = max(1, int(limit) // 2)
    return {
        "entry_count": len(entries),
        "head": entries[:edge],
        "tail": entries[-edge:],
    }


def _view_hard_snapshot(app, view, inputs: dict) -> dict:
    """Immutable/view-only state; presentation and heartbeat stay outside it."""
    names = (
        "row_pairs", "row_a_to_pair_idx", "row_b_to_pair_idx", "pair_base_row_override",
        "pair_diff_cols", "pair_base_diff_cols", "pair_raw_parts_a", "pair_raw_parts_b",
        "pair_raw_parts_base", "_full_display_rows", "_only_diff_rows_cache",
        "_only_diff_rows_cache_key", "_sheet_structural_diff", "_missing_base_row_map",
        "_align_rows_enabled", "_data_version", "modified", "touched_cells",
        "_row_model_version", "_column_model_version", "_column_projection_generation",
        "_col_widths_version",
    )
    app_names = (
        "manual_a_ops", "manual_b_ops", "manual_a_row_ops", "manual_b_row_ops",
        "manual_a_column_ops", "manual_b_column_ops", "manual_sheet_ops", "auto_sheet_ops",
        "undo_stack", "redo_stack", "_edit_load_requests",
        "_sheet_compute_generation", "_sheet_topology_generation", "_sheet_mutation_generation",
    )
    overlays = {
        str(sheet): {
            "topology_generation": int(getattr(overlay, "topology_generation", 0) or 0),
            "mutation_generation": int(getattr(overlay, "mutation_generation", 0) or 0),
            "cells": _canonical(getattr(overlay, "cells", {})),
        }
        for sheet, overlay in sorted(
            dict(getattr(app, "sheet_operation_overlays", {}) or {}).items()
        )
    }
    cache = getattr(view, "column_comparison_cache", None)
    projection = getattr(view, "column_projection", None)
    model = getattr(cache, "model", None)
    slots = tuple(
        (
            int(getattr(slot, "logical_idx", -1)),
            getattr(slot, "mine_col", None),
            getattr(slot, "base_col", None),
            getattr(slot, "theirs_col", None),
            str(getattr(slot, "state", "")),
            _canonical(getattr(slot, "confidence", None)),
        )
        for slot in tuple(getattr(model, "slots", ()) or ())
    )
    blocks = tuple(
        (
            int(getattr(block, "ordinal", -1)),
            tuple(int(item) for item in tuple(getattr(block, "slot_indices", ()) or ())),
            str(getattr(block, "state", "")),
            _canonical(getattr(block, "confidence", None)),
        )
        for block in tuple(getattr(model, "blocks", ()) or ())
    )
    physical = {
        side: {
            "physical_to_logical": _bounded_lookup_digest(
                getattr(model, f"{side}_physical_to_logical", None)
            ),
            "logical_to_physical": _bounded_lookup_digest(
                getattr(model, f"{side}_logical_to_physical", None)
            ),
        }
        for side in ("mine", "base", "theirs")
    }
    return {
        "input": {name: _file_fact(path) for name, path in sorted(inputs.items())},
        "view": tuple((name, _canonical(getattr(view, name, None))) for name in names),
        "app": tuple((name, _canonical(getattr(app, name, None))) for name in app_names),
        "handles": tuple((name, type(getattr(app, name, None)).__name__, id(getattr(app, name, None)))
                         for name in ("_wb_a_val", "_wb_b_val", "_wb_base_val", "_wb_a_edit", "_wb_b_edit", "_wb_base_edit")),
        "overlays": overlays,
        "cache_projection": {
            "cache_identity": id(cache),
            "projection_identity": id(projection),
            "slots": slots,
            "blocks": blocks,
            "physical_lookup": physical,
            "structural": tuple(sorted(int(col) for col in (
                getattr(cache, "structural_diff_cols", ()) or ()
            ))),
            "unresolved": tuple(sorted(int(col) for col in (
                getattr(cache, "unresolved_cols", ()) or ()
            ))),
        },
    }


@contextlib.contextmanager
def _forbid_view_only(app, view):
    """Fail on any workbook/edit/save/operation fallback during public routes."""
    hits, originals = [], []
    def forbid(label):
        def blocked(*_args, **_kwargs):
            hits.append(label)
            raise AssertionError("view-only forbidden: " + label)
        return blocked
    app_names = (
        "ws_a_val", "ws_b_val", "ws_base_val", "ws_a_edit", "ws_b_edit", "ws_base_edit",
        "_request_edit_preload", "_ensure_edit_loaded", "_load_edit_workbooks_owned",
        "_start_background_thread", "save_a_inplace", "save_b_inplace", "save_merged_and_exit",
        "_atomic_save", "_atomic_save_with_retry", "_atomic_replace_file_with_retry",
        "build_manual_b_output_file", "_try_alt_save",
        "_enqueue_sheet", "_kick_worker", "_queue_ui_task",
    )
    view_names = ("_refresh_mode_switch_preserving_selection", "refresh", "rescan")
    module_names = (
        "_excel_com_cell_op", "_atomic_save_wb", "_align_selected_sheet_snapshots",
        "align_column_signatures_2way", "align_column_signatures_3way",
        "build_logical_column_comparison_cache_2way",
        "build_logical_column_comparison_cache_3way",
        "compare_logical_row_2way", "compare_logical_row_3way",
    )
    for owner, names in ((app, app_names), (view, view_names)):
        for name in names:
            if hasattr(owner, name):
                originals.append((owner, name, getattr(owner, name)))
                setattr(owner, name, forbid(name))
    for name in module_names:
        if hasattr(sm, name):
            originals.append((sm, name, getattr(sm, name)))
            setattr(sm, name, forbid(name))
    old = (Worksheet.cell, Worksheet.iter_rows, ReadOnlyWorksheet.iter_rows)
    Worksheet.cell, Worksheet.iter_rows, ReadOnlyWorksheet.iter_rows = (forbid("Worksheet.cell"), forbid("Worksheet.iter_rows"), forbid("ReadOnlyWorksheet.iter_rows"))
    try:
        yield hits
    finally:
        Worksheet.cell, Worksheet.iter_rows, ReadOnlyWorksheet.iter_rows = old
        for owner, name, original in reversed(originals):
            setattr(owner, name, original)


def _pump(root):
    root.update()
    root.update_idletasks()


def _wait(root, predicate, deadline: float, label: str):
    while time.monotonic() < deadline:
        _pump(root)
        if predicate():
            return
        time.sleep(.005)
    raise AssertionError("deadline: " + label)


def _exact_ready(app, sheet: str) -> bool:
    view = app.sheet_views.get(sheet)
    entry = app._sheet_exact_entry(sheet)
    request_ms = entry.get("request_to_full_detail_ms")
    return bool(view and getattr(view, "_data_ready", False) and app._is_sheet_exact_current(sheet)
                and entry.get("full_detail_terminal") is True
                and isinstance(request_ms, (int, float)) and float(request_ms) <= GATE_EXACT_MS
                and str(app.selected_sheet) == sheet)


def _cell_event(widget, line: int = 1):
    widget.update_idletasks()
    index = f"{max(1, int(line))}.1"
    box = widget.bbox(index)
    assert box is not None, (widget, index)
    x, y, width, height = box
    return {"x": int(x + max(1, width // 2)), "y": int(y + max(1, height // 2)), "index": index}


def _dispatch_text_event(widget, sequence: str, meta: dict, delta: int | None = None) -> dict:
    """Dispatch a real Text event without leaking benchmark metadata into Tcl."""
    event = dict(meta)
    x, y = event.get("x"), event.get("y")
    assert isinstance(x, int) and not isinstance(x, bool), (sequence, event)
    assert isinstance(y, int) and not isinstance(y, bool), (sequence, event)
    if "index" in event:
        assert str(widget.index(f"@{x},{y}")) == str(event["index"]), (sequence, event)
    tk_kwargs = {"x": int(x), "y": int(y)}
    if delta is not None:
        assert isinstance(delta, int) and not isinstance(delta, bool), (sequence, delta)
        tk_kwargs["delta"] = int(delta)
    widget.event_generate(str(sequence), **tk_kwargs)
    return {
        "tk_kwargs": dict(tk_kwargs),
        "meta": {key: value for key, value in event.items() if key not in {"x", "y"}},
    }


def _window_state(view) -> dict:
    return {
        "row_start": int(getattr(view, "_virtual_window_start", 0) or 0),
        "rows": tuple(getattr(view, "display_rows", ()) or ()),
        "column_start": int(getattr(view, "_virtual_column_window_start", 0) or 0),
        "columns": tuple(view._rendered_logical_columns()),
    }


def _presentation_state(view) -> dict:
    return {
        "virtual_publish_generation": int(getattr(view, "_virtual_publish_generation", 0) or 0),
        "virtual_column_window_generation": int(
            getattr(view, "_virtual_column_window_generation", 0) or 0
        ),
        "request_seq": int(getattr(view, "_viewport_request_seq", 0) or 0),
        "terminal_count": len(tuple(getattr(view, "_viewport_request_terminal", ()) or ())),
    }


def _scrollbar_slider_span(widget, *, vertical: bool) -> tuple[tuple[int, int], dict]:
    """Find the actual thumb/slider span before emitting a real drag."""
    widget.update_idletasks()
    width, height = max(3, int(widget.winfo_width())), max(3, int(widget.winfo_height()))
    axis = height if vertical else width
    cross = max(1, (width // 2) if vertical else (height // 2))
    hits = []
    for point in range(1, axis - 1):
        x, y = (cross, point) if vertical else (point, cross)
        hits.append((point, str(widget.identify(x, y) or "")))
    slider_points = [
        point for point, element in hits
        if element.lower() in {"slider", "thumb"}
    ]
    assert slider_points, {"axis": axis, "cross": cross, "identify": hits}
    spans = []
    first = previous = slider_points[0]
    for point in slider_points[1:]:
        if point == previous + 1:
            previous = point
            continue
        spans.append((first, previous))
        first = previous = point
    spans.append((first, previous))
    span = max(spans, key=lambda item: item[1] - item[0])
    return span, {
        "axis": axis,
        "cross": cross,
        "identify": tuple(hits),
        "spans": tuple(spans),
    }


def _drag_scrollbar(widget, fraction: float, *, vertical: bool) -> dict:
    """Use the real Scrollbar press/drag/release binding; command is evidence only."""
    span, diagnostic = _scrollbar_slider_span(widget, vertical=vertical)
    axis, cross = int(diagnostic["axis"]), int(diagnostic["cross"])
    start = (int(span[0]) + int(span[1])) // 2
    target = max(1, min(axis - 2, round(float(fraction) * (axis - 2))))
    point = (cross, start) if vertical else (start, cross)
    moved = (cross, target) if vertical else (target, cross)
    assert str(widget.identify(point[0], point[1]) or "").lower() in {"slider", "thumb"}, (
        point, diagnostic,
    )
    widget.event_generate("<ButtonPress-1>", x=point[0], y=point[1])
    widget.event_generate("<B1-Motion>", x=moved[0], y=moved[1])
    widget.event_generate("<ButtonRelease-1>", x=moved[0], y=moved[1])
    return {
        "fraction": float(fraction),
        "orientation": "vertical" if vertical else "horizontal",
        "command_diagnostic": str(widget.cget("command") or ""),
        "press": point,
        "release": moved,
        "slider_span": span,
        "identify": diagnostic,
    }


def _vertical_thumb_geometry(widget) -> dict:
    """Read the actual ttk vertical thumb/trough geometry without moving it."""
    widget.update_idletasks()
    width, height = max(3, int(widget.winfo_width())), max(3, int(widget.winfo_height()))
    cross = max(1, width // 2)
    hits = tuple((point, str(widget.identify(cross, point) or "")) for point in range(1, height - 1))
    thumb_points = [
        point for point, element in hits if element.lower() in {"slider", "thumb"}
    ]
    trough_points = [
        point for point, element in hits if element.lower().startswith("trough")
    ]
    assert thumb_points and trough_points, {"axis": height, "cross": cross, "identify": hits}
    thumb_start, thumb_end = min(thumb_points), max(thumb_points)
    track_start = min(thumb_points + trough_points)
    track_end = max(thumb_points + trough_points)
    press_center = (thumb_start + thumb_end) // 2
    left_radius, right_radius = press_center - thumb_start, thumb_end - press_center
    travel_start, travel_end = track_start + left_radius, track_end - right_radius
    assert travel_start <= press_center <= travel_end and travel_end > travel_start, {
        "axis": height, "thumb": (thumb_start, thumb_end), "track": (track_start, track_end),
        "travel": (travel_start, travel_end), "identify": hits,
    }
    trough_spans = []
    if trough_points:
        first = previous = trough_points[0]
        for point in trough_points[1:]:
            if point != previous + 1:
                trough_spans.append((first, previous))
                first = point
            previous = point
        trough_spans.append((first, previous))
    return {
        "axis": height,
        "cross": cross,
        "identify": hits,
        "thumb_span": (thumb_start, thumb_end),
        "trough_spans": tuple(trough_spans),
        "track_span": (track_start, track_end),
        "travel_span": (travel_start, travel_end),
        "press_center": press_center,
    }


def _drag_vertical_thumb_to_pair(view, pair_idx: int) -> dict:
    """One real thumb press/drag/release that centers a prepared logical pair."""
    full_rows = tuple(getattr(view, "_full_display_rows", ()) or ())
    pair_idx = int(pair_idx)
    assert pair_idx in full_rows, (pair_idx, full_rows[:4], full_rows[-4:])
    total = len(full_rows)
    cap = min(int(sm._VIRTUAL_VIEWPORT_MAX_ROWS), total)
    position = full_rows.index(pair_idx)
    wanted_start = max(0, min(position - cap // 2, max(0, total - cap)))
    moveto_fraction = float(wanted_start) / float(max(1, total - cap))
    widget = view.vsb_left
    before_get = tuple(float(value) for value in widget.get())
    geometry = _vertical_thumb_geometry(widget)
    travel_start, travel_end = geometry["travel_span"]
    target_center = int(round(
        int(travel_start) + moveto_fraction * (int(travel_end) - int(travel_start))
    ))
    target_center = max(int(travel_start), min(int(travel_end), target_center))
    press = (int(geometry["cross"]), int(geometry["press_center"]))
    release = (int(geometry["cross"]), target_center)
    assert str(widget.identify(*press) or "").lower() in {"slider", "thumb"}, {
        "press": press, "geometry": geometry,
    }
    widget.event_generate("<ButtonPress-1>", x=press[0], y=press[1])
    widget.event_generate("<B1-Motion>", x=release[0], y=release[1])
    widget.event_generate("<ButtonRelease-1>", x=release[0], y=release[1])
    return {
        "pair_idx": pair_idx,
        "pair_position": position,
        "total_rows": total,
        "cap": cap,
        "wanted_start": wanted_start,
        "moveto_fraction": moveto_fraction,
        "before_get": before_get,
        "geometry": geometry,
        "press": press,
        "release": release,
    }


def _span_prepare_state(app, view, inputs: dict) -> dict:
    """Capture semantic state that public Text layout preparation must not change."""
    sheet = str(view.sheet)
    exact_entry = dict(app._sheet_exact_entry(sheet) or {})
    active = getattr(view, "_viewport_request_active", None)
    if isinstance(active, dict):
        active = dict(active)
    selection_names = (
        "selected_pair_idx", "selected_excel_row", "selected_excel_row_a",
        "selected_excel_row_b", "_main_sel_col", "_main_sel_line",
        "_cursor_cmp_sel_col", "_cursor_cmp_sel_line", "_last_cursor_cmp_pair_idx",
        "hover_pair_idx", "hover_col_idx", "hover_side", "selected_column_block_ordinal",
        "selected_column_logical_range", "selected_column_source_side",
    )
    terminal = tuple(
        _canonical(dict(item))
        for item in tuple(getattr(view, "_viewport_request_terminal", ()) or ())
    )
    return {
        "hard": _view_hard_snapshot(app, view, inputs),
        "window": _window_state(view),
        "request": {
            "seq": int(getattr(view, "_viewport_request_seq", 0) or 0),
            "terminal": terminal,
            "publication_seq": int(getattr(view, "_virtual_scroll_publications", 0) or 0),
            "last_publication": _canonical(
                getattr(view, "_last_virtual_publication_telemetry", {}) or {}
            ),
        },
        "exact": {
            "selected_sheet": str(getattr(app, "selected_sheet", "") or ""),
            "view_sheet": sheet,
            "generation": _generation(getattr(app, "_sheet_compute_generation", {}).get(sheet)),
            "current": bool(app._is_sheet_exact_current(sheet)),
            "entry": _canonical(exact_entry),
        },
        "selection": {
            "row_to_line": tuple(sorted(
                (int(pair), int(line))
                for pair, line in dict(getattr(view, "row_to_line", {}) or {}).items()
            )),
            **{name: _canonical(getattr(view, name, None)) for name in selection_names},
        },
        "virtual": {
            "pending_row": getattr(view, "_virtual_pending_start", None),
            "pending_column": getattr(view, "_virtual_pending_column_start", None),
            "publish_after_id": getattr(view, "_virtual_publish_after_id", None),
            "column_publish_after_id": getattr(view, "_virtual_column_publish_after_id", None),
            "publishing": bool(getattr(view, "_virtual_publishing", False)),
            "active": _canonical(active),
        },
    }


def _span_event(app, view, widget, pair_idx: int, logical_col: int, inputs: dict) -> dict:
    """Derive a public Text coordinate without changing semantic viewport state."""
    assert int(pair_idx) in tuple(view.display_rows), (pair_idx, view.display_rows)
    assert int(logical_col) in tuple(view._rendered_logical_columns()), (
        logical_col, view._rendered_logical_columns(),
    )
    spans = dict(view._base_spans())
    assert int(logical_col) in spans, (logical_col, spans)
    line = tuple(view.display_rows).index(int(pair_idx)) + 1
    start, end = spans[int(logical_col)]
    char = max(int(start), (int(start) + int(end)) // 2)
    index = f"{line}.{char}"
    before = _span_prepare_state(app, view, inputs)
    # `see` is public Tk layout preparation only.  The semantic snapshot makes
    # any accidental virtual queue/publication/selection side effect fail closed.
    widget.see(index)
    _pump(app.root)
    after = _span_prepare_state(app, view, inputs)
    assert after == before, ("span prepare changed semantic state", before, after)
    box = widget.bbox(index)
    lineinfo = widget.dlineinfo(index)
    assert box is not None, (line, char, pair_idx, logical_col, lineinfo)
    assert lineinfo is not None, (line, char, pair_idx, logical_col, box)
    x, y, width, height = box
    event_x = int(x + max(1, width // 2))
    event_y = int(y + max(1, height // 2))
    found_line, found_col = map(int, str(widget.index(f"@{event_x},{event_y}")).split("."))
    assert found_line == int(line) and int(start) <= found_col < max(int(start) + 1, int(end)), (
        found_line, found_col, line, start, end, index,
    )
    return {
        "x": event_x,
        "y": event_y,
        "index": index,
        "pair_idx": int(pair_idx),
        "logical_col": int(logical_col),
        "coordinate_prepare": {
            "window": before["window"],
            "request_seq": before["request"]["seq"],
            "terminal_count": len(before["request"]["terminal"]),
            "publication_seq": before["request"]["publication_seq"],
        },
    }


@contextlib.contextmanager
def _delegating_payload_spy(view):
    """Observe the real payload consumer while public Motion drives it."""
    calls = []
    original = view._cmp_tooltip_payload_by_pair_col

    def wrapped(pair_idx, target_col, *args, **kwargs):
        result = original(pair_idx, target_col, *args, **kwargs)
        raw_values = None
        if isinstance(result, tuple) and len(result) >= 2:
            try:
                raw_values = tuple(
                    (getattr(view, "_hover_payload_cache", {}) or {}).get(result[1], {}).get("values") or ()
                )
            except Exception:
                raw_values = None
        calls.append({
            "pair_idx": int(pair_idx),
            "logical_col": int(target_col),
            "result": result,
            "payload_values": raw_values,
        })
        return result

    view._cmp_tooltip_payload_by_pair_col = wrapped
    try:
        yield calls
    finally:
        view._cmp_tooltip_payload_by_pair_col = original


def _tooltip_value(value) -> str:
    """Use the production tooltip's line-break normalization exactly."""
    return str(value).replace("\r\n", "⏎").replace("\r", "⏎").replace("\n", "⏎")


def _generation(value) -> int:
    """Canonical benchmark generation parsing: preserve zero, reject invalid values."""
    if value is None or isinstance(value, bool):
        return -1
    try:
        parsed = int(value)
    except (TypeError, ValueError, OverflowError):
        return -1
    if isinstance(value, float) and (not math.isfinite(value) or float(parsed) != value):
        return -1
    return parsed


def _terminal_for(view, request_id: int):
    return [dict(item) for item in tuple(getattr(view, "_viewport_request_terminal", ()) or ())
            if int(item.get("id") or -1) == int(request_id)]


def _resource_sample(app, stage: str, *, child_owner=None) -> dict:
    """Record validity-only parent/child resource facts; never set a numeric cap."""
    process = psutil.Process()
    cpu = process.cpu_times()
    sample = {
        "stage": str(stage),
        "at": time.monotonic(),
        "parent": {
            "pid": int(process.pid),
            "rss_bytes": int(process.memory_info().rss),
            "cpu_user_s": float(cpu.user),
            "cpu_system_s": float(cpu.system),
        },
        "runtime": tuple(dict(item) for item in tuple(
            getattr(app, "_runtime_resource_samples", ()) or ()
        )) if app is not None else (),
    }
    assert sample["parent"]["rss_bytes"] > 0
    assert all(math.isfinite(float(value)) and float(value) >= 0.0
               for value in sample["parent"].values() if not isinstance(value, int) or value != process.pid)
    if isinstance(child_owner, dict):
        pid = int(child_owner.get("pid") or getattr(child_owner.get("process"), "pid", 0) or 0)
        assert pid > 0 and psutil.pid_exists(pid), child_owner
        child = psutil.Process(pid)
        child_cpu = child.cpu_times()
        sample["child"] = {
            "pid": pid,
            "request_token": str(child_owner.get("request_token") or child_owner.get("token") or ""),
            "sheet": str(child_owner.get("sheet") or ""),
            "generation": _generation(child_owner.get("generation")),
            "rss_bytes": int(child.memory_info().rss),
            "cpu_user_s": float(child_cpu.user),
            "cpu_system_s": float(child_cpu.system),
        }
        assert sample["child"]["rss_bytes"] > 0 and sample["child"]["request_token"]
    return sample


def _resource_summary(samples: list[dict], *, app, preempt: dict, child_metrics: dict) -> dict:
    """Summarize parent and child telemetry by validity/schema, never by a resource cap."""
    parent_samples = [dict(item["parent"]) for item in samples if isinstance(item, dict) and "parent" in item]
    assert parent_samples and int(parent_samples[0]["rss_bytes"]) > 0
    for item in parent_samples:
        assert int(item["rss_bytes"]) > 0
        assert all(math.isfinite(float(item[key])) and float(item[key]) >= 0.0
                   for key in ("cpu_user_s", "cpu_system_s"))
    first, last = parent_samples[0], parent_samples[-1]
    child_samples = [
        dict(item["child"]) for item in samples
        if isinstance(item, dict) and isinstance(item.get("child"), dict)
    ]
    for item in child_samples:
        assert int(item["rss_bytes"]) > 0 and str(item["request_token"])
        assert all(math.isfinite(float(item[key])) and float(item[key]) >= 0.0
                   for key in ("cpu_user_s", "cpu_system_s"))
    old_live = dict(preempt["live_resource"]["child"])
    old_terminal = dict(preempt["terminal"])
    old = (int(preempt["pid"]), str(preempt["token"]), int(preempt["generation"]))
    assert (
        int(old_live["pid"]), str(old_live["request_token"]), int(old_live["generation"])
    ) == old
    assert (
        int(old_terminal.get("pid") or 0), str(old_terminal.get("request_token") or ""),
        _generation(old_terminal.get("generation")),
    ) == old
    children = [{
        "sheet": str(preempt["peer"]), "pid": old[0], "token": old[1],
        "generation": old[2], "live_n": 1,
        "terminal_or_completed": {
            "kind": "terminal", "at": float(old_terminal["at"]),
            "reason": str(old_terminal.get("reason") or ""),
        },
        "cleanup": {
            "result_exists_after_cleanup": bool(old_terminal.get("result_exists_after_cleanup")),
            "partial_exists_after_cleanup": bool(old_terminal.get("partial_exists_after_cleanup")),
        },
        "rss_bytes": int(old_live["rss_bytes"]),
        "cpu_ms": (float(old_live["cpu_user_s"]) + float(old_live["cpu_system_s"])) * 1000.0,
    }]
    for sheet in (str(preempt["anchor"]), str(preempt["peer"])):
        metric = dict(child_metrics[sheet])
        current_generation = _generation(app._sheet_compute_generation[sheet])
        pid = int(metric.get("pid") or 0)
        token = str(metric.get("request_token") or "")
        assert pid > 0 and token and _generation(metric.get("generation")) == current_generation, metric
        assert int(metric.get("peak_rss_bytes") or 0) > 0
        assert math.isfinite(float(metric.get("last_cpu_ms") or 0)) and float(metric["last_cpu_ms"]) >= 0.0
        if sheet == str(preempt["peer"]):
            assert token == str(preempt["resumed_token"]), (metric, preempt)
        children.append({
            "sheet": sheet, "pid": pid, "token": token,
            "generation": current_generation, "live_n": 0,
            "terminal_or_completed": {"kind": "completed", "exact_state": str(
                app._sheet_exact_entry(sheet).get("state") or ""
            )},
            "cleanup": None,
            "rss_bytes": int(metric["peak_rss_bytes"]),
            "cpu_ms": float(metric["last_cpu_ms"]),
        })
    return {
        "parent_before": first,
        "parent_peak_rss_bytes": max(int(item["rss_bytes"]) for item in parent_samples),
        "parent_after": last,
        "parent_rss_delta_bytes": int(last["rss_bytes"]) - int(first["rss_bytes"]),
        "parent_cpu_delta_s": (
            float(last["cpu_user_s"]) + float(last["cpu_system_s"])
            - float(first["cpu_user_s"]) - float(first["cpu_system_s"])
        ),
        "parent_sample_count": len(parent_samples),
        "children": tuple(children),
    }


def _public_route(
    app, view, label: str, callback, deadline: float, hard, inputs: dict, routes: list[dict],
    *, require_request: bool = False, target_window=None, resources: list[dict] | None = None,
    require_column_generation: bool = False,
):
    assert not require_column_generation or require_request
    before_seq = int(getattr(view, "_viewport_request_seq", 0) or 0)
    before = _view_hard_snapshot(app, view, inputs)
    before_window = _window_state(view)
    before_presentation = _presentation_state(view)
    started = time.perf_counter()
    interaction = callback()
    callback_ms = (time.perf_counter() - started) * 1000.0
    assert callback_ms <= GATE_CALLBACK_MS, (label, callback_ms)
    _pump(app.root)
    request_id = int(getattr(view, "_viewport_request_seq", 0) or 0)
    if require_request:
        assert request_id > before_seq, (label, before_seq, request_id, before_window)
    if request_id == before_seq:
        record = {"route": label, "status": "no-request", "callback_ms": callback_ms, "elapsed_ms": callback_ms, "surface_changed": False}
    else:
        _wait(app.root, lambda: bool(_terminal_for(view, request_id)), deadline, f"{label} terminal#{request_id}")
        record = _terminal_for(view, request_id)[-1]
        assert record.get("status") == "complete", record
        assert int(record.get("generation", -1)) == int(app._sheet_compute_generation[view.sheet]), record
        assert str(record.get("selected_sheet") or "") == str(view.sheet), record
        assert record.get("counted") and record.get("surface_changed"), record
        publication = dict(record.get("publication") or {})
        assert int(publication.get("request_id") or -1) == request_id, record
        assert str(publication.get("selected_sheet") or "") == str(view.sheet), record
    _pump(app.root)
    after_window = _window_state(view)
    after_presentation = _presentation_state(view)
    if target_window is not None:
        assert bool(target_window(before_window, after_window)), (
            label, before_window, after_window, record,
        )
    assert _view_hard_snapshot(app, view, inputs) == before, (label, "immutable state changed")
    if request_id != before_seq:
        assert after_presentation["request_seq"] == request_id, (
            label, before_presentation, after_presentation, record,
        )
        assert after_presentation["terminal_count"] >= before_presentation["terminal_count"] + 1, (
            label, before_presentation, after_presentation, record,
        )
    if require_column_generation:
        assert (
            int(after_presentation["virtual_column_window_generation"])
            > int(before_presentation["virtual_column_window_generation"])
        ), (label, before_presentation, after_presentation, record)
    record.update({
        "route": label,
        "callback_ms": callback_ms,
        "request_id": request_id,
        "window_before": before_window,
        "window_after": after_window,
        "interaction": _canonical(interaction),
        "presentation_before": before_presentation,
        "presentation_after": after_presentation,
    })
    if resources is not None:
        resources.append(_resource_sample(app, f"route:{label}"))
    routes.append(record)
    return record


def _bounded_metrics(app, routes, heartbeat_counter_before: int, heartbeat_buffer, heartbeat_maxlen: int):
    changed = [item for item in routes if item.get("surface_changed")]
    assert changed and len(routes) >= 10, routes
    e2e = [float(item["elapsed_ms"]) for item in changed]
    assert _p95(e2e) <= GATE_CALLBACK_MS and max(e2e) <= GATE_E2E_MAX_MS, changed
    current_buffer = getattr(app, "_ui_heartbeat_gaps_ms", None)
    assert current_buffer is heartbeat_buffer, (id(heartbeat_buffer), id(current_buffer))
    assert int(getattr(current_buffer, "maxlen", 0) or 0) == int(heartbeat_maxlen), (
        getattr(current_buffer, "maxlen", None), heartbeat_maxlen,
    )
    counter_after = int(getattr(app, "_ui_heartbeat_samples", 0) or 0)
    delta = counter_after - int(heartbeat_counter_before)
    gaps = tuple(current_buffer)[-delta:] if delta > 0 else ()
    print(
        "REAL_RELEASE_HEARTBEAT_DIAGNOSTICS " + json.dumps(
            {
                "counter_before": int(heartbeat_counter_before),
                "counter_after": counter_after,
                "delta": delta,
                "buffer_id": id(current_buffer),
                "buffer_len": len(current_buffer),
                "buffer_maxlen": int(heartbeat_maxlen),
                "gaps_ms": tuple(float(value) for value in gaps),
            },
            sort_keys=True,
        ),
        flush=True,
    )
    assert delta >= 3 and delta <= len(current_buffer), (
        heartbeat_counter_before, counter_after, delta, len(current_buffer), heartbeat_maxlen,
    )
    assert len(gaps) == delta, (delta, len(gaps), len(current_buffer))
    assert max(float(value) for value in gaps) <= GATE_HEARTBEAT_MS, gaps
    return {"route_n": len(routes), "changed_n": len(changed), "e2e_p95_ms": round(_p95(e2e), 3),
            "e2e_max_ms": round(max(e2e), 3), "heartbeat_n": len(gaps),
            "heartbeat_p95_ms": round(_p95(gaps), 3), "heartbeat_max_ms": round(max(gaps), 3),
            "heartbeat_counter_before": int(heartbeat_counter_before),
            "heartbeat_counter_after": counter_after, "heartbeat_delta": delta,
            "heartbeat_buffer_id": id(heartbeat_buffer), "heartbeat_buffer_len": len(current_buffer),
            "heartbeat_buffer_maxlen": int(heartbeat_maxlen), "heartbeat_samples_ms": tuple(gaps)}


def _live_snapshot_owner(app, sheet: str, generation: int, *,
                         exclude_token: str | None = None, expected=None):
    """Return the current, actually started child owner or ``None``.

    The production runner publishes a ``spawn`` owner before ``Process.start``.
    A benchmark preemption is meaningful only after that same owner has a PID
    and is still alive, so take and re-check the identity under its lock.
    """
    lock = getattr(app, "_snapshot_child_lock", None)
    manager = lock if lock is not None else contextlib.nullcontext()
    with manager:
        owner = getattr(app, "_snapshot_child_owner", None)
        if not isinstance(owner, dict) or (expected is not None and owner is not expected):
            return None
        token = str(owner.get("token") or "")
        if (
            str(owner.get("sheet") or "") != str(sheet)
            or _generation(owner.get("generation")) != _generation(generation)
            or _generation(app._sheet_compute_generation.get(sheet)) != _generation(generation)
            or not token or token == str(exclude_token or "")
            or str(owner.get("phase") or "") != "snapshot-compare-adapter"
        ):
            return None
        process = owner.get("process")
        pid = int(getattr(process, "pid", 0) or 0)
        try:
            alive = bool(process is not None and process.is_alive() and psutil.pid_exists(pid))
        except Exception:
            alive = False
        return owner if pid > 0 and alive else None


def _live_owner_diagnostic(app, sheet: str, generation: int) -> dict:
    """Stable fail-fast evidence for a peer that never reaches a live owner."""
    lock = getattr(app, "_snapshot_child_lock", None)
    manager = lock if lock is not None else contextlib.nullcontext()
    with manager:
        raw_owner = getattr(app, "_snapshot_child_owner", None)
        owner = dict(raw_owner) if isinstance(raw_owner, dict) else {}
        process = owner.get("process")
        pid = int(getattr(process, "pid", 0) or owner.get("pid") or 0)
        try:
            alive = bool(process is not None and process.is_alive() and pid > 0 and psutil.pid_exists(pid))
        except Exception:
            alive = False
        try:
            exact_entry = dict(app._sheet_exact_entry(sheet) or {})
        except Exception as exc:
            exact_entry = {"diagnostic_error": f"{type(exc).__name__}: {exc}"}
        return {
            "selected_sheet": str(getattr(app, "selected_sheet", "") or ""),
            "target_sheet": str(sheet),
            "target_generation": _generation(generation),
            "current_generation": _generation(getattr(app, "_sheet_compute_generation", {}).get(sheet)),
            "owner": {
                "sheet": str(owner.get("sheet") or ""),
                "generation": _generation(owner.get("generation")),
                "token": str(owner.get("token") or ""),
                "phase": str(owner.get("phase") or ""),
                "pid": pid,
                "alive": alive,
            },
            "exact_entry": _canonical(exact_entry),
        }


def _wait_live_owner(app, sheet: str, generation: int, deadline: float, *,
                     exclude_token: str | None = None, stage: str):
    """Wait for one identity-stable, actually started child within 15 seconds."""
    candidates = []

    def owner_live():
        candidate = _live_snapshot_owner(
            app, sheet, generation, exclude_token=exclude_token,
        )
        if candidate is None:
            return False
        candidates[:] = [candidate]
        return True

    live_deadline = min(float(deadline), time.monotonic() + 15.0)
    try:
        _wait(app.root, owner_live, live_deadline, f"live {sheet} {stage} child")
    except AssertionError as exc:
        detail = _live_owner_diagnostic(app, sheet, generation)
        raise AssertionError(
            f"{exc}; live-owner={json.dumps(detail, ensure_ascii=False, sort_keys=True)}"
        ) from None
    owner = _live_snapshot_owner(
        app, sheet, generation, exclude_token=exclude_token, expected=candidates[0],
    )
    assert owner is not None, candidates
    return owner


def _preempt_peer(app, anchor: str, peer: str, deadline: float, resources: list[dict]):
    """Use public tabs to prove cooperative peer yield, child cleanup, and a fresh ticket."""
    assert str(app.selected_sheet) == str(anchor)
    event_start = len(tuple(getattr(app, "_snapshot_child_events", ()) or ()))
    yield_start = len(tuple(getattr(app, "_background_yield_events", ()) or ()))
    foreground_start = len(tuple(getattr(app, "_foreground_resume_events", ()) or ()))

    def tab_confirmed(sheet: str):
        return [
            dict(item) for item in tuple(getattr(app, "_foreground_resume_events", ()) or ())[foreground_start:]
            if str(dict(item).get("event") or "") == "tab-confirmed"
            and str(dict(dict(item).get("visit") or {}).get("sheet") or "") == str(sheet)
        ]

    app.nb.select(app._sheet_containers[peer])
    _wait(app.root, lambda: str(app.selected_sheet) == peer, deadline, f"{peer} tab")
    _wait(app.root, lambda: bool(tab_confirmed(peer)), deadline, f"{peer} tab-confirmed")
    generation = _generation(app._sheet_compute_generation[peer])
    assert generation >= 0, {"peer": peer, "generation": generation}
    owner = _wait_live_owner(app, peer, generation, deadline, stage="initial")
    pid = int(getattr(owner.get("process"), "pid", 0) or 0)
    token = str(owner.get("token") or "")
    assert pid > 0 and token and psutil.pid_exists(pid), owner
    live_resource = _resource_sample(app, f"peer-live:{peer}", child_owner=owner)
    resources.append(live_resource)
    app.nb.select(app._sheet_containers[anchor])
    _wait(app.root, lambda: str(app.selected_sheet) == anchor, deadline, f"{anchor} preempt tab")
    _wait(app.root, lambda: bool(tab_confirmed(anchor)), deadline, f"{anchor} tab-confirmed")

    def yielded():
        return [
            tuple(item)
            for item in tuple(getattr(app, "_background_yield_events", ()) or ())[yield_start:]
            if len(tuple(item)) >= 4
            and str(tuple(item)[1]) in {"selected-preempt", "ui-activity"}
            and str(tuple(item)[2]) == str(peer)
            and str(tuple(item)[3]) == str(anchor)
        ]

    def terminal():
        return [item for item in tuple(getattr(app, "_snapshot_child_events", ()) or ())[event_start:]
                if item.get("event") == "terminated" and item.get("reason") == "cancel-or-preempt"
                and int(item.get("pid") or 0) == pid and str(item.get("request_token") or "") == token
                and str(item.get("sheet") or "") == peer and _generation(item.get("generation")) == generation]
    _wait(app.root, lambda: bool(yielded()) and bool(terminal()), deadline, f"{peer} preempt terminal")
    yield_event = yielded()[-1]
    event = dict(terminal()[-1])
    assert float(yield_event[0]) <= float(event.get("at") or float("inf")), (yield_event, event)
    assert not event.get("result_exists_after_cleanup") and not event.get("partial_exists_after_cleanup"), event
    terminal_binding = {
        "pid": pid,
        "token": token,
        "generation": generation,
        "live_sample": dict(live_resource["child"]),
        "terminal_at": float(event["at"]),
        "result_exists_after_cleanup": bool(event.get("result_exists_after_cleanup")),
        "partial_exists_after_cleanup": bool(event.get("partial_exists_after_cleanup")),
    }
    assert (
        int(terminal_binding["live_sample"]["pid"]),
        str(terminal_binding["live_sample"]["request_token"]),
        int(terminal_binding["live_sample"]["generation"]),
    ) == (pid, token, generation), terminal_binding
    old_apply = [
        item for item in tuple(getattr(app, "_snapshot_child_events", ()) or ())[event_start:]
        if str(item.get("event") or "") in {"finished", "published", "result-applied", "applied"}
        and int(item.get("pid") or 0) == pid and str(item.get("request_token") or "") == token
    ]
    assert not old_apply, old_apply
    app.nb.select(app._sheet_containers[peer])
    _wait(app.root, lambda: str(app.selected_sheet) == peer, deadline, f"{peer} resume tab")
    _wait(app.root, lambda: len(tab_confirmed(peer)) >= 2, deadline, f"{peer} resume tab-confirmed")
    fresh_owner = _wait_live_owner(
        app, peer, generation, deadline, exclude_token=token, stage="fresh",
    )
    fresh_token = str(fresh_owner["token"])
    fresh_started = {
        "at": float(fresh_owner.get("started") or 0.0),
        "pid": int(fresh_owner.get("pid") or getattr(fresh_owner.get("process"), "pid", 0) or 0),
        "token": fresh_token,
        "generation": _generation(fresh_owner.get("generation")),
    }
    assert fresh_started["at"] > 0.0 and fresh_started["generation"] == generation, fresh_owner
    _wait(app.root, lambda: _exact_ready(app, peer), deadline, f"{peer} fresh full detail")
    resumed = [
        item for item in tuple(getattr(app, "_snapshot_child_events", ()) or ())
        if item.get("event") == "finished" and str(item.get("sheet") or "") == peer
        and _generation(item.get("generation")) == generation
        and str(item.get("request_token") or "") == fresh_token
        and float(item.get("at") or 0.0) >= fresh_started["at"]
    ]
    assert resumed, f"{peer} did not receive a fresh post-preempt completion ticket"
    fresh_finished = dict(resumed[-1])
    assert float(event["at"]) < fresh_started["at"] <= float(fresh_finished["at"]), (
        event, fresh_started, fresh_finished,
    )
    assert _generation(app._sheet_exact_entry(peer).get("generation")) == generation
    foreground = []
    for raw in tuple(getattr(app, "_foreground_resume_events", ()) or ())[foreground_start:]:
        item = dict(raw)
        request = dict(item.get("request") or {})
        if str(request.get("sheet") or "") == peer and _generation(
            request.get("generation")
        ) == generation:
            foreground.append(item)
    queued = [item for item in foreground if item.get("event") == "priority-queued"]
    fired = [item for item in foreground if item.get("event") == "priority-fire"]
    if queued or fired:
        assert len(queued) == len(fired) == 1, foreground
        assert (
            float(event["at"]) <= float(queued[0]["at"]) <= float(fired[0]["at"])
            <= float(fresh_started["at"]) <= float(fresh_finished["at"])
        ), (event, queued, fired, fresh_started, fresh_finished)
        disposition = "priority-resume"
    else:
        peer_tabs = tab_confirmed(peer)
        assert not queued and not fired and len(peer_tabs) >= 2, foreground
        assert float(peer_tabs[-1]["at"]) <= float(fresh_started["at"]), (
            peer_tabs, fresh_started,
        )
        disposition = "direct-tab-resume"
    return {
        "anchor": anchor, "peer": peer, "pid": pid, "token": token, "generation": generation,
        "yield": yield_event, "live_resource": live_resource, "terminal": event,
        "terminal_binding": terminal_binding,
        "fresh_started": fresh_started, "fresh_finished": fresh_finished,
        "resumed_token": fresh_token, "foreground_resume": foreground,
        "tab_confirmed": {"anchor": tab_confirmed(anchor), "peer": tab_confirmed(peer)},
        "resume_disposition": disposition,
    }


def _physical_base_target(view, three_way: bool) -> dict:
    if not three_way:
        return {"required": False}
    _projection, slots = _readonly_projection_slots(view)
    overrides = dict(getattr(view, "pair_base_row_override", {}) or {})
    for pair_idx, cols in sorted(dict(view.pair_base_diff_cols).items()):
        if int(pair_idx) not in overrides:
            continue
        for logical_col in sorted(int(item) for item in cols if int(item) >= 0):
            slot = next((item for item in slots
                         if int(getattr(item, "logical_idx", -1)) == logical_col), None)
            if slot is not None and getattr(slot, "base_col", None) is not None:
                return {"required": True, "pair": int(pair_idx), "logical_col": logical_col + 1,
                        "base_physical_col": int(slot.base_col),
                        "base_override_row": int(overrides[int(pair_idx)])}
    raise AssertionError("three-way Base override/physical target missing")


def _exercise_three_way_base_target(
    app, view, deadline: float, hard: dict, inputs: dict, routes: list[dict],
    resources: list[dict],
) -> dict:
    """Bring one complete Base target into the real 2-D surface and consume it by Motion."""
    target = _physical_base_target(view, True)
    full_rows = tuple(getattr(view, "_full_display_rows", ()) or ())
    column_count = max(1, int(getattr(view, "_logical_slot_count")() or 1))
    column_fraction = float(int(target["logical_col"]) - 1) / float(max(1, column_count - 1))
    assert int(target["pair"]) not in tuple(view.display_rows), {
        "target": target, "before": _window_state(view),
    }
    vertical_record = _public_route(
        app, view, "base-target-v-thumb",
        lambda: _drag_vertical_thumb_to_pair(view, int(target["pair"])),
        deadline, hard, inputs, routes, require_request=True,
        target_window=lambda _before, after: int(target["pair"]) in tuple(after["rows"]),
        resources=resources,
    )
    vertical_drag = dict(vertical_record.get("interaction") or {})
    vertical_drag.update({
        "after_get": tuple(float(value) for value in view.vsb_left.get()),
        "after_window": _window_state(view),
        "request_id": int(vertical_record["request_id"]),
        "request_generation": int(vertical_record["generation"]),
        "publication_window": dict((vertical_record.get("publication") or {}).get("window") or {}),
    })
    assert int(target["pair"]) in tuple(vertical_drag["after_window"]["rows"]), vertical_drag
    vertical_record["base_vthumb_drag"] = vertical_drag
    _public_route(
        app, view, "base-target-h-thumb",
        lambda: _drag_scrollbar(view.hsb_left, column_fraction, vertical=False),
        deadline, hard, inputs, routes, require_request=True,
        target_window=lambda _before, after: int(target["logical_col"]) in tuple(after["columns"]),
        resources=resources, require_column_generation=True,
    )
    raw_base = tuple(dict(view.pair_raw_parts_base)[int(target["pair"])])
    raw_value = raw_base[int(target["base_physical_col"]) - 1]
    prepared_value = view._prepared_value_for_logical_cell(
        int(target["pair"]), "BASE", int(target["logical_col"])
    )
    assert int(dict(view.pair_base_row_override)[int(target["pair"])]) == int(
        target["base_override_row"]
    )
    event = _span_event(
        app, view, view.base, int(target["pair"]), int(target["logical_col"]), inputs
    )
    with _delegating_payload_spy(view) as calls:
        _public_route(
            app, view, "base-target-base-motion",
            lambda: _dispatch_text_event(view.base, "<Motion>", event),
            deadline, hard, inputs, routes, resources=resources,
        )
        _wait(
            app.root,
            lambda: any(
                int(item["pair_idx"]) == int(target["pair"])
                and int(item["logical_col"]) == int(target["logical_col"])
                for item in calls
            ),
            deadline,
            "three-way Base payload",
        )
    _projection, slots = _readonly_projection_slots(view)
    slot = next(
        item for item in slots
        if int(getattr(item, "logical_idx", -1)) == int(target["logical_col"]) - 1
    )
    assert int(getattr(slot, "base_col")) == int(target["base_physical_col"]), (target, slot)
    matching = [
        item for item in calls
        if int(item["pair_idx"]) == int(target["pair"])
        and int(item["logical_col"]) == int(target["logical_col"])
    ]
    assert len(matching) == 1, (target, calls)
    payload_values = tuple(matching[0]["payload_values"] or ())
    assert payload_values, (matching, prepared_value, raw_value)
    raw_text = _tooltip_value(raw_value)
    prepared_text = _tooltip_value(prepared_value)
    payload_text = _tooltip_value(payload_values[0])
    assert raw_text == prepared_text == payload_text, (
        matching, raw_text, prepared_text, payload_text,
    )
    target["payload_calls"] = tuple(calls)
    target["event"] = event
    target["base_raw_value"] = raw_value
    target["prepared_base_value"] = prepared_value
    target["base_raw_tooltip_value"] = raw_text
    target["prepared_base_tooltip_value"] = prepared_text
    target["payload_tooltip_value"] = payload_text
    return target


def _readonly_projection_slots(view):
    """Read the already-installed immutable column projection without rebuilding it."""
    cache = getattr(view, "column_comparison_cache", None)
    projection = getattr(view, "column_projection", None)
    assert isinstance(cache, sm.LogicalColumnComparisonCache), type(cache).__name__
    assert isinstance(projection, sm.LogicalColumnProjection), type(projection).__name__
    assert projection.model is cache.model
    slots = tuple(projection.slots)
    assert slots == tuple(cache.model.slots)
    return projection, slots


@contextlib.contextmanager
def _case_files(revision_dir: Path, case: str):
    old = revision_dir / f"Dungeon-r{OLD_REVISION}.xlsx"
    new = revision_dir / f"Dungeon-r{NEW_REVISION}.xlsx"
    assert old.is_file() and new.is_file(), (old, new)
    original = {"mine": _file_fact(new), "theirs": _file_fact(old)}
    old_settings = Path(sm._SETTINGS_PATH)
    settings_before = old_settings.read_bytes() if old_settings.exists() else None
    previous_env = _ENV_CALLBACK_TIMING_BEFORE
    audit = {"settings_before_exists": settings_before is not None, "errors": []}
    with tempfile.TemporaryDirectory(prefix=f"sow_real_release_{case}_") as raw_root:
        root = Path(raw_root)
        inputs = {"mine": root / "Dungeon-r39265.xlsx", "theirs": root / "Dungeon-r39264.xlsx"}
        shutil.copy2(new, inputs["mine"]); shutil.copy2(old, inputs["theirs"])
        for path in inputs.values():
            path.chmod(stat.S_IREAD | stat.S_IRGRP | stat.S_IROTH)
        copied_before = {name: _file_fact(path) for name, path in inputs.items()}
        assert all(_same_bytes(copied_before[name], original[name]) for name in inputs)
        settings = root / "settings.json"
        settings.write_text(json.dumps({"only_diff": 0}), encoding="utf-8")
        sm._SETTINGS_PATH = settings
        try:
            yield root, inputs, original, copied_before, audit
        finally:
            try:
                sm._SETTINGS_PATH = old_settings
                if old_settings.exists():
                    assert old_settings.read_bytes() == settings_before
                else:
                    assert settings_before is None
            except Exception as exc:
                audit["errors"].append(f"settings:{type(exc).__name__}:{exc}")
            try:
                if previous_env is None:
                    os.environ.pop("SOW_CALLBACK_TIMING_DIAGNOSTIC", None)
                else:
                    os.environ["SOW_CALLBACK_TIMING_DIAGNOSTIC"] = previous_env
            except Exception as exc:
                audit["errors"].append(f"environment:{type(exc).__name__}:{exc}")
            try:
                assert _file_fact(new) == original["mine"] and _file_fact(old) == original["theirs"]
            except Exception as exc:
                audit["errors"].append(f"original-input:{type(exc).__name__}:{exc}")


def _cancel_debounces(app) -> list[str]:
    """Cancel every known debounce/after owner before the Tk root is shut down."""
    cancelled = []
    owners = [app] + [view for view in tuple(app.sheet_views.values()) if view is not None]
    explicit = {
        "_settings_save_id", "_hover_debounce_id", "_diff_map_debounce_id",
        "_only_diff_progress_show_after_id", "_only_diff_progress_watchdog_after_id",
        "_only_diff_progress_confirm_after_id", "_viewport_publish_after_id",
        "_viewport_interaction_after_id", "_remaining_sheet_resume_after_id",
        "_foreground_resume_after_id",
    }
    for owner in owners:
        for name in dir(owner):
            if name not in explicit and not (
                name.endswith("_id") and ("debounce" in name or "after" in name)
            ):
                continue
            try:
                after_id = getattr(owner, name)
                if after_id:
                    app.root.after_cancel(after_id)
                    cancelled.append(f"{type(owner).__name__}.{name}")
            except Exception:
                continue
    return cancelled


def _assert_child_ipc_clean(app) -> dict:
    lock = getattr(app, "_snapshot_child_lock", None)
    if lock is not None:
        with lock:
            owner = getattr(app, "_snapshot_child_owner", None)
            paths = tuple(getattr(app, "_snapshot_child_temp_paths", ()) or ())
            events = tuple(getattr(app, "_snapshot_child_events", ()) or ())
    else:
        owner = getattr(app, "_snapshot_child_owner", None)
        paths = tuple(getattr(app, "_snapshot_child_temp_paths", ()) or ())
        events = tuple(getattr(app, "_snapshot_child_events", ()) or ())
    assert owner is None and not paths, {"owner": owner, "ipc": paths}
    return {"terminal_events": tuple(_canonical(item) for item in events)}


def _cleanup_app(app, effective_inputs) -> dict:
    """Cleanup runs after a primary failure too; callers decide whether secondary errors fail it."""
    detail, errors = {"cancelled": []}, []
    try:
        detail["cancelled"] = _cancel_debounces(app)
    except Exception as exc:
        errors.append(f"cancel-debounce:{type(exc).__name__}:{exc}")
    try:
        app._shutdown_root()
    except Exception as exc:
        errors.append(f"shutdown:{type(exc).__name__}:{exc}")
    owned_cleanup = tuple(
        dict(item) for item in tuple(
            getattr(app, "_owned_startup_temp_cleanup_evidence", ()) or ()
        )
        if isinstance(item, dict)
    )
    detail["owned_startup_temp_cleanup"] = owned_cleanup
    if not owned_cleanup:
        errors.append("owned-startup-temp:no-cleanup-evidence")
    for item in owned_cleanup:
        if not bool(item.get("removed")) or bool(item.get("exists_after")) or item.get("error"):
            errors.append(
                "owned-startup-temp:"
                f"path={item.get('path')} removed={item.get('removed')} "
                f"exists_after={item.get('exists_after')} error={item.get('error')}"
            )
    try:
        detail["ipc"] = _assert_child_ipc_clean(app)
    except Exception as exc:
        errors.append(f"ipc:{type(exc).__name__}:{exc}")
    try:
        detail["effective_inputs"] = _legacy._assert_effective_live_inputs_after_shutdown(
            effective_inputs or {"roles": {}}
        )
    except Exception as exc:
        errors.append(f"effective-inputs:{type(exc).__name__}:{exc}")
    detail["errors"] = errors
    return detail


def _run_release_case(case: str, revision_dir: Path) -> dict:
    config = dict(_CASES[case]); three_way = config["mode"] == "3way"
    anchor = str(config["anchor"])
    peer = next(name for name in SHEETS if name != anchor)
    deadline = time.monotonic() + CASE_DEADLINE_SECONDS
    assert _sha256(Path(sm.__file__).resolve()) == _SOURCE_SHA256
    payload = {
        "schema": "real-changed-revision-release-v4",
        "case": case,
        "mode": config["mode"],
        "anchor": anchor,
        "peer": peer,
        "status": "FAILED",
    }
    app = None
    root_after = None
    audit = {}
    try:
        with _case_files(revision_dir, case) as (root, inputs, original, copies, audit):
            root_after = root
            effective, resources, routes = {}, [_resource_sample(None, "parent-before")], []
            primary_error = False
            try:
                app = _legacy._new_benchmark_app(
                    str(inputs["mine"]), str(inputs["theirs"]),
                    merge_mode=three_way,
                    base_path=(str(inputs["theirs"]) if three_way else None),
                    initial_sheet=anchor,
                )
                resources.append(_resource_sample(app, "parent-after-constructor"))
                effective = _legacy._capture_effective_live_inputs(
                    app, inputs["mine"], inputs["theirs"],
                    inputs["theirs"] if three_way else None,
                )
                _wait(app.root, lambda: _exact_ready(app, anchor), deadline, f"{anchor} full exact")
                resources.append(_resource_sample(app, f"anchor-full:{anchor}"))
                assert not app._edit_workbooks_ready() and not app._edit_load_requests

                # The preemption transition is intentionally outside the view-only
                # route sentinels: it is the one stage where a peer compute is legal.
                preempt = _preempt_peer(app, anchor, peer, deadline, resources)
                app.nb.select(app._sheet_containers[SHEETS[0]])
                _wait(app.root, lambda: _exact_ready(app, SHEETS[0]), deadline, "Dungeon restored full")
                dungeon = app.sheet_views[SHEETS[0]]
                assert not app._edit_workbooks_ready() and not app._edit_load_requests
                resources.append(_resource_sample(app, "dungeon-restored"))

                heartbeat0 = int(getattr(app, "_ui_heartbeat_samples", 0) or 0)
                _wait(
                    app.root,
                    lambda: int(getattr(app, "_ui_heartbeat_samples", 0) or 0) >= heartbeat0 + 2,
                    deadline,
                    "heartbeat prime",
                )
                heartbeat_counter = int(getattr(app, "_ui_heartbeat_samples", 0) or 0)
                heartbeat_buffer = getattr(app, "_ui_heartbeat_gaps_ms", None)
                heartbeat_maxlen = int(getattr(heartbeat_buffer, "maxlen", 0) or 0)
                assert heartbeat_buffer is not None and heartbeat_maxlen > 0, heartbeat_buffer
                with _forbid_view_only(app, dungeon) as hits:
                    hard = _view_hard_snapshot(app, dungeon, inputs)
                    main = _cell_event(dungeon.left)
                    cursor = _cell_event(dungeon.cursor_cmp)
                    _public_route(app, dungeon, "main-motion",
                                  lambda: _dispatch_text_event(dungeon.left, "<Motion>", main),
                                  deadline, hard, inputs, routes, resources=resources)
                    _public_route(app, dungeon, "c-motion",
                                  lambda: _dispatch_text_event(dungeon.cursor_cmp, "<Motion>", cursor),
                                  deadline, hard, inputs, routes, resources=resources)
                    _public_route(app, dungeon, "c-click",
                                  lambda: _dispatch_text_event(dungeon.cursor_cmp, "<Button-1>", cursor),
                                  deadline, hard, inputs, routes, resources=resources)
                    _public_route(
                        app, dungeon, "wheel",
                        lambda: _dispatch_text_event(dungeon.left, "<MouseWheel>", main, delta=-120),
                        deadline, hard, inputs, routes, resources=resources,
                    )
                    assert str(dungeon.vsb_left.cget("command") or "") and str(
                        dungeon.hsb_left.cget("command") or ""
                    )
                    _public_route(
                        app, dungeon, "v-thumb-drag",
                        lambda: _drag_scrollbar(dungeon.vsb_left, .91, vertical=True),
                        deadline, hard, inputs, routes, require_request=True,
                        target_window=lambda before, after: after["row_start"] != before["row_start"],
                        resources=resources,
                    )
                    vmap = dungeon.vdiff_map; vmap.update_idletasks()
                    _public_route(
                        app, dungeon, "v-minimap",
                        lambda: vmap.event_generate(
                            "<Button-1>", x=1, y=max(1, vmap.winfo_height() // 2)
                        ),
                        deadline, hard, inputs, routes, require_request=True,
                        target_window=lambda before, after: after["row_start"] != before["row_start"],
                        resources=resources,
                    )
                    _public_route(
                        app, dungeon, "h-thumb-last-drag",
                        lambda: _drag_scrollbar(dungeon.hsb_left, 1.0, vertical=False),
                        deadline, hard, inputs, routes, require_request=True,
                        target_window=lambda before, after: after["column_start"] != before["column_start"],
                        resources=resources, require_column_generation=True,
                    )
                    _public_route(
                        app, dungeon, "h-thumb-first-drag",
                        lambda: _drag_scrollbar(dungeon.hsb_left, 0.0, vertical=False),
                        deadline, hard, inputs, routes, require_request=True,
                        target_window=lambda before, after: after["column_start"] != before["column_start"],
                        resources=resources, require_column_generation=True,
                    )
                    hmap = dungeon.hdiff_right; hmap.update_idletasks()
                    _public_route(
                        app, dungeon, "h-minimap",
                        lambda: hmap.event_generate(
                            "<Button-1>", x=max(1, hmap.winfo_width() // 2), y=1
                        ),
                        deadline, hard, inputs, routes, require_request=True,
                        target_window=lambda before, after: after["column_start"] != before["column_start"],
                        resources=resources, require_column_generation=True,
                    )
                    _public_route(
                        app, dungeon, "combined-2d-drag",
                        lambda: (
                            _drag_scrollbar(dungeon.vsb_left, .07, vertical=True),
                            _drag_scrollbar(dungeon.hsb_left, .5, vertical=False),
                        ),
                        deadline, hard, inputs, routes, require_request=True,
                        target_window=lambda before, after: (
                            after["row_start"] != before["row_start"]
                            and after["column_start"] != before["column_start"]
                        ),
                        resources=resources, require_column_generation=True,
                    )
                    base_target = (
                        _exercise_three_way_base_target(
                            app, dungeon, deadline, hard, inputs, routes, resources
                        )
                        if three_way else _physical_base_target(dungeon, False)
                    )
                    assert not hits, hits
                assert _view_hard_snapshot(app, dungeon, inputs) == hard

                # The release contract always drives MonsterGroup explicitly
                # after the Dungeon wide routes; peer is only preemption proof.
                monster_sheet = SHEETS[1]
                app.nb.select(app._sheet_containers[monster_sheet])
                _wait(
                    app.root, lambda: _exact_ready(app, monster_sheet), deadline,
                    "MonsterGroup final exact",
                )
                monster_view = app.sheet_views[monster_sheet]
                with _forbid_view_only(app, monster_view) as hits:
                    monster_hard = _view_hard_snapshot(app, monster_view, inputs)
                    monster_event = _cell_event(monster_view.left)
                    _public_route(
                        app, monster_view, "monster-main-motion",
                        lambda: _dispatch_text_event(monster_view.left, "<Motion>", monster_event),
                        deadline, monster_hard, inputs, routes, resources=resources,
                    )
                    assert not hits, hits
                    assert _view_hard_snapshot(app, monster_view, inputs) == monster_hard
                _wait(
                    app.root,
                    lambda: int(getattr(app, "_ui_heartbeat_samples", 0) or 0) >= heartbeat_counter + 3,
                    deadline,
                    "heartbeat suffix",
                )
                metrics = _bounded_metrics(
                    app, routes, heartbeat_counter, heartbeat_buffer, heartbeat_maxlen
                )
                resources.append(_resource_sample(app, "parent-end"))
                child = {
                    sheet: dict(getattr(app.sheet_views[sheet], "_snapshot_child_metrics", {}) or {})
                    for sheet in SHEETS if app.sheet_views.get(sheet) is not None
                }
                assert child and all(
                    int(item.get("peak_rss_bytes") or 0) > 0
                    and math.isfinite(float(item.get("last_cpu_ms") or 0))
                    and float(item.get("last_cpu_ms") or 0) >= 0.0
                    for item in child.values()
                ), child
                resource_summary = _resource_summary(
                    resources, app=app, preempt=preempt, child_metrics=child,
                )
                payload.update({
                    "status": "PASS",
                    "original_inputs": original,
                    "copied_inputs_before": copies,
                    "copied_inputs_after": {name: _file_fact(path) for name, path in inputs.items()},
                    "effective_inputs": effective,
                    "routes": routes,
                    "metrics": metrics,
                    "preempt": preempt,
                    "base_target": base_target,
                    "resource_samples": resources,
                    "resource_summary": resource_summary,
                    "child_resources": child,
                    "edit_load_requests": list(app._edit_load_requests),
                    "manifest": {
                        "source_sha256": _SOURCE_SHA256,
                        "legacy_sha256": _LEGACY_SHA256,
                        "support": {str(_LEGACY_PATH.name): _file_fact(_LEGACY_PATH)},
                    },
                })
                assert payload["copied_inputs_after"] == copies and not payload["edit_load_requests"]
            except Exception as exc:
                primary_error = True
                payload.update({
                    "status": "FAILED",
                    "error": f"{type(exc).__name__}: {exc}",
                    "traceback": traceback.format_exc(limit=20),
                })
            finally:
                secondary = []
                if app is not None:
                    cleanup = _cleanup_app(app, effective)
                    payload["cleanup"] = cleanup
                    secondary.extend(cleanup["errors"])
                try:
                    payload["copies_final"] = {name: _file_fact(path) for name, path in inputs.items()}
                    if payload["copies_final"] != copies:
                        secondary.append("copied input changed")
                except Exception as exc:
                    secondary.append(f"copies-final:{type(exc).__name__}:{exc}")
                if secondary:
                    payload.setdefault("secondary", []).extend(secondary)
                    if not primary_error:
                        payload.update({"status": "FAILED", "error": "cleanup: " + "; ".join(secondary)})
                (root / "report.json").write_text(
                    json.dumps(payload, ensure_ascii=False, indent=2, default=str),
                    encoding="utf-8",
                )
    except Exception as exc:
        payload.update({
            "status": "FAILED",
            "error": f"outer:{type(exc).__name__}: {exc}",
            "traceback": traceback.format_exc(limit=20),
        })
    if audit.get("errors"):
        payload.setdefault("secondary", []).extend(audit["errors"])
        if "error" not in payload:
            payload.update({"status": "FAILED", "error": "case-file cleanup: " + "; ".join(audit["errors"])})
    if root_after is not None:
        payload["temp_root_removed"] = not root_after.exists()
        if not payload["temp_root_removed"] and "error" not in payload:
            payload.update({"status": "FAILED", "error": "temporary root remained"})
    return payload


def _changed_revision_case_plan(mode: str | None = None, sheet: str | None = None):
    """Compatibility plan used by the existing pure test; stable CLI uses case IDs."""
    modes = (str(mode),) if mode else ("2way", "3way")
    sheets = (str(sheet),) if sheet else tuple(SHEETS)
    return tuple((item, name) for item in modes for name in sheets)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--case", choices=_CASE_ORDER)
    parser.add_argument("--list-cases", action="store_true")
    parser.add_argument("--revision-dir", default=str(REVISION_DIR))
    args = parser.parse_args()
    if args.list_cases:
        print("\n".join(_CASE_ORDER)); return
    selected = args.case or _CASE_ORDER[0]
    print("START " + selected, flush=True)
    result = _run_release_case(selected, Path(args.revision_dir))
    print(json.dumps(result, ensure_ascii=False, sort_keys=True, default=str), flush=True)
    raise SystemExit(0 if result.get("status") == "PASS" else 1)


if __name__ == "__main__":
    main()

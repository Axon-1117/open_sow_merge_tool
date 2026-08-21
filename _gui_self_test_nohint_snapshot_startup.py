"""Regression: ordinary GUI launches use the selected-Sheet snapshot gate.

No caller supplies ``initial_sheet`` in this test.  The first tab therefore
models the production entry point, while the second tab proves the catalog is
available without scanning sibling worksheet XML.  The companion small-sheet
case verifies the conservative legacy fallback still restores writable value
and edit workbooks before the first operation/save.
"""

from __future__ import annotations

import os
import sys
import tempfile
import time

from openpyxl import Workbook, load_workbook

import sow_merge_tool as sm


def _make_book(path: str, *, large: bool, suffix: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Large" if large else "Small"
    ws.append(["id @id", "value"])
    rows = 420 if large else 4
    for number in range(1, rows + 1):
        ws.append([number, f"{suffix}-{number}"])
    if large:
        unopened = wb.create_sheet("Unopened")
        unopened.append(["id @id", "value"])
        unopened.append([1, "must stay unparsed"])
    wb.save(path)
    wb.close()


def _pump_until(app, predicate, message: str, timeout: float = 25.0) -> None:
    deadline = time.monotonic() + timeout
    while time.monotonic() < deadline:
        app.root.update_idletasks()
        app.root.update()
        if predicate():
            return
        time.sleep(0.01)
    raise AssertionError(message)


def _save_and_assert(app, path: str, expected: str) -> None:
    app._atomic_save(app._wb_a_edit, path)
    reopened = load_workbook(path, data_only=False)
    try:
        assert reopened[app.selected_sheet]["B2"].value == expected
    finally:
        reopened.close()


def _run_large_no_hint(tmp: str) -> None:
    mine = os.path.join(tmp, "large-mine.xlsx")
    theirs = os.path.join(tmp, "large-theirs.xlsx")
    out = os.path.join(tmp, "large-out.xlsx")
    _make_book(mine, large=True, suffix="mine")
    _make_book(theirs, large=True, suffix="theirs")

    app = sm.SowMergeApp(mine, theirs)
    try:
        assert app.selected_sheet == "Large", app.selected_sheet
        _pump_until(
            app,
            lambda: (
                app.sheet_views.get("Large") is not None
                and bool(getattr(app.sheet_views["Large"], "_data_ready", False))
                and getattr(app.sheet_views["Large"], "_cache_source", "") == "snapshot"
            ),
            "no-hint first large tab did not become snapshot-ready",
        )
        view = app.sheet_views["Large"]
        assert app._snapshot_startup_lightweight is True
        assert bool(getattr(app._wb_a_val, "read_only", False))
        assert bool(getattr(app._wb_b_val, "read_only", False))
        assert app._wb_a_edit is None and app._wb_b_edit is None
        assert app.sheet_views.get("Unopened") is None
        assert "Unopened" not in app._sheet_cache_store

        # First real edit is the explicit demand signal: promote both normal
        # value and formula workbooks, then preserve the ordinary save path.
        app.ws_a_edit("Large")
        assert app._edit_workbooks_ready()
        assert not getattr(app._wb_a_val, "read_only", False)
        view._copy_single_cell_by_pair(1, "B2A", 2)
        assert app._wb_a_edit["Large"]["B2"].value == "theirs-1"
        _save_and_assert(app, out, "theirs-1")
    finally:
        app._shutdown_root()


def _run_small_no_hint(tmp: str) -> None:
    mine = os.path.join(tmp, "small-mine.xlsx")
    theirs = os.path.join(tmp, "small-theirs.xlsx")
    out = os.path.join(tmp, "small-out.xlsx")
    _make_book(mine, large=False, suffix="mine")
    _make_book(theirs, large=False, suffix="theirs")

    app = sm.SowMergeApp(mine, theirs)
    try:
        assert app.selected_sheet == "Small", app.selected_sheet
        _pump_until(
            app,
            lambda: (
                app.sheet_views.get("Small") is not None
                and bool(getattr(app.sheet_views["Small"], "_data_ready", False))
                and getattr(app.sheet_views["Small"], "_cache_source", "") == "legacy"
                and app._edit_workbooks_ready()
            ),
            "no-hint small first tab did not take the writable legacy fallback",
        )
        view = app.sheet_views["Small"]
        assert not getattr(app._wb_a_val, "read_only", False)
        view._copy_single_cell_by_pair(1, "B2A", 2)
        assert app._wb_a_edit["Small"]["B2"].value == "theirs-1"
        _save_and_assert(app, out, "theirs-1")
    finally:
        app._shutdown_root()


def main() -> None:
    prior_flag = sm._LARGE_SHEET_SNAPSHOT_ENGINE_ENABLED
    prior_threshold = sm._LARGE_SHEET_ROW_THRESHOLD
    sm._LARGE_SHEET_SNAPSHOT_ENGINE_ENABLED = True
    sm._LARGE_SHEET_ROW_THRESHOLD = 200
    try:
        with tempfile.TemporaryDirectory(prefix="sow_nohint_snapshot_") as tmp:
            _run_large_no_hint(tmp)
            _run_small_no_hint(tmp)
    finally:
        sm._LARGE_SHEET_SNAPSHOT_ENGINE_ENABLED = prior_flag
        sm._LARGE_SHEET_ROW_THRESHOLD = prior_threshold
    print("GUI_SELF_TEST_NOHINT_SNAPSHOT_STARTUP_OK")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"GUI_SELF_TEST_NOHINT_SNAPSHOT_STARTUP_FAIL: {exc}", file=sys.stderr)
        raise

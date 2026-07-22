"""Exercise Excel COM structural replay with an explicit blank cell operation."""

import os

from openpyxl import Workbook, load_workbook

import sow_merge_tool as mod
from _test_temp_utils import make_temp_dir


def _make_book(path: str, inserted_source: bool = False):
    wb = Workbook()
    ws = wb.active
    ws.title = "S1"
    ws.append(["id", "blank", "number", "boolean", "text"])
    if inserted_source:
        ws.append(["inserted", "must-clear", 9, True, "source-text"])
    else:
        ws.append(["original", "keep", 7, True, "original-text"])
    wb.save(path)
    wb.close()


def main():
    root = make_temp_dir(prefix="sow_excel_com_blank_")
    source = os.path.join(root, "source.xlsx")
    theirs = os.path.join(root, "theirs.xlsx")
    output = os.path.join(root, "output.xlsx")
    _make_book(source)
    _make_book(theirs, inserted_source=True)

    ok = mod._build_manual_merge_output_with_excel(
        source,
        output,
        {
            ("S1", 2, 2): None,
        },
        row_ops=[{
            "kind": "insert_rows",
            "sheet": "S1",
            "row": 2,
            "count": 1,
            "source_side": "B",
            "source_rows": [2],
        }],
        source_paths={"B": theirs},
    )
    if not ok:
        print("SMOKE_EXCEL_COM_BLANK_CELL_SKIPPED (Excel COM unavailable)")
        return

    wb = load_workbook(output, data_only=False)
    ws = wb["S1"]
    assert ws["A2"].value == "inserted", ws["A2"].value
    assert ws["B2"].value is None, ws["B2"].value
    assert ws["C2"].value == 9, ws["C2"].value
    assert ws["D2"].value is True, ws["D2"].value
    assert ws["E2"].value == "source-text", ws["E2"].value
    assert ws["A3"].value == "original", ws["A3"].value
    wb.close()
    print("SMOKE_EXCEL_COM_BLANK_CELL_OK")


if __name__ == "__main__":
    main()

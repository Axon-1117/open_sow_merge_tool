"""OpenSpec 4.5: native column replay and fidelity acceptance tests.

The default suite mocks only the Excel process boundary.  It still exercises
the real JSON payload construction and public A/B save orchestration, so a
machine without an interactive Excel COM session can validate 4.1--4.4.

Run the optional real-Excel layer explicitly:
  $env:SOW_RUN_REAL_EXCEL_COLUMN_REPLAY = "1"
  python _smoke_test_column_native_save_replay.py

Until 4.1--4.4 are implemented this file is expected to fail at the missing
``column_ops`` native-builder contract.  Do not weaken that failure into a
skip: it is the implementation gate for 4.5.
"""

from __future__ import annotations

import hashlib
import inspect
import json
import os
import re
import shutil
import subprocess
import tempfile
import zipfile
from dataclasses import dataclass
from types import SimpleNamespace
from xml.etree import ElementTree as ET

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

import sow_merge_tool as smt


_RUN_REAL_EXCEL = os.environ.get("SOW_RUN_REAL_EXCEL_COLUMN_REPLAY", "").strip() == "1"
_CUSTOM_PART = "customXml/item1.xml"
_VBA_PART = "xl/vbaProject.bin"
_SHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


def _sha256(path: str) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


@dataclass
class _OwnedFixture:
    temporary: tempfile.TemporaryDirectory
    root: str
    mine: str
    theirs: str
    expected: str | None
    input_hashes: dict[str, str]


def _finish_cleanup(primary: BaseException | None, errors: list[str]) -> None:
    if not errors:
        return
    detail = "; ".join(errors)
    if primary is not None:
        primary.add_note(f"cleanup failure: {detail}")
        return
    raise AssertionError(f"cleanup failure: {detail}")


def _cleanup_owned_fixture(
    fixture: _OwnedFixture,
    primary: BaseException | None,
    *,
    reader_closers=(),
    restore=None,
) -> None:
    """Keep fixture cleanup deterministic without masking a test failure."""
    errors: list[str] = []
    for label, closer in reader_closers:
        try:
            closer()
        except BaseException as error:
            errors.append(f"{label} close: {error!r}")
    if restore is not None:
        try:
            restore()
        except BaseException as error:
            errors.append(f"restore monkeypatch: {error!r}")
    try:
        actual = {path: _sha256(path) for path in fixture.input_hashes}
        if actual != fixture.input_hashes:
            errors.append(
                f"fixture input SHA changed: expected={fixture.input_hashes!r}, actual={actual!r}"
            )
    except BaseException as error:
        errors.append(f"input SHA check: {error!r}")
    try:
        fixture.temporary.cleanup()
    except BaseException as error:
        errors.append(f"temporary cleanup: {error!r}")
    if os.path.lexists(fixture.root):
        errors.append(f"temporary root remains: {fixture.root!r}")
    _finish_cleanup(primary, errors)


def _close_workbook(workbook):
    """Close both the workbook reader and openpyxl's keep_vba memory ZIP."""
    if workbook is None:
        return
    try:
        workbook.close()
    finally:
        vba_archive = getattr(workbook, "vba_archive", None)
        if vba_archive is not None:
            vba_archive.close()


def _rewrite_zip(path: str, replacements=None, additions=None):
    replacements = dict(replacements or {})
    additions = dict(additions or {})
    temp_path = path + ".rewrite"
    with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(temp_path, "w") as target:
        names = set()
        for info in source.infolist():
            names.add(info.filename)
            payload = replacements.get(info.filename, source.read(info.filename))
            target.writestr(info, payload)
        for name, payload in additions.items():
            if name not in names:
                target.writestr(name, payload, compress_type=zipfile.ZIP_DEFLATED)
    os.replace(temp_path, path)


def _zip_bytes(path: str, name: str) -> bytes:
    with zipfile.ZipFile(path, "r") as archive:
        return archive.read(name)


def _inject_advanced_formula_records(path: str, *, include_vba=False):
    part = "xl/worksheets/sheet2.xml"
    root = ET.fromstring(_zip_bytes(path, part))
    q = lambda name: f"{{{_SHEET_NS}}}{name}"
    cells = {
        node.attrib.get("r"): node
        for node in root.iter(q("c"))
        if node.attrib.get("r")
    }

    def formula(ref: str):
        node = cells[ref]
        result = node.find(q("f"))
        if result is None:
            result = ET.SubElement(node, q("f"))
        return result

    shared_master = formula("A1")
    shared_master.attrib.clear()
    shared_master.attrib.update({"t": "shared", "ref": "A1:A2", "si": "0"})
    shared_master.text = "ROW()"
    shared_follower = formula("A2")
    shared_follower.attrib.clear()
    shared_follower.attrib.update({"t": "shared", "si": "0"})
    shared_follower.text = None

    array_formula = formula("C1")
    array_formula.attrib.clear()
    array_formula.attrib.update({"t": "array", "ref": "C1:C2"})
    array_formula.text = "ROW(C1:C2)"

    data_table = formula("E1")
    data_table.attrib.clear()
    data_table.attrib.update({"t": "dataTable", "ref": "E1:F2", "r1": "A1"})
    data_table.text = None

    additions = {_CUSTOM_PART: b"<codex-sentinel version='4.5'/>"}
    if include_vba:
        additions[_VBA_PART] = b"CODEx-VBA-SENTINEL-DO-NOT-DROP"
    _rewrite_zip(
        path,
        replacements={part: ET.tostring(root, encoding="utf-8", xml_declaration=True)},
        additions=additions,
    )


def _make_fidelity_book(
    path: str,
    headers,
    *,
    decorated=(),
    include_vba=False,
    inject_advanced=True,
):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "S1"
    decorated = set(int(value) for value in decorated)
    for col, header in enumerate(headers, start=1):
        sheet.cell(1, col).value = header
        for row in range(2, 9):
            sheet.cell(row, col).value = (
                f"={chr(64 + max(1, col - 1))}{row}*2"
                if row == 3 else f"{header}-{row}"
            )
        if col in decorated:
            letter = chr(64 + col)
            sheet.column_dimensions[letter].width = 18.0 + col
            sheet.column_dimensions[letter].hidden = (col == min(decorated))
            cell = sheet.cell(2, col)
            cell.fill = PatternFill("solid", fgColor="22AA66")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.comment = Comment(f"comment-{header}", "Codex")
            cell.hyperlink = f"https://example.test/{header}"
    if {2, 3}.issubset(decorated):
        validation = DataValidation(type="list", formula1='"red,green,blue"', allow_blank=True)
        validation.add("B2:C8")
        sheet.add_data_validation(validation)
        sheet.merge_cells("B5:C5")
        sheet.conditional_formatting.add(
            "B2:C8",
            CellIsRule(
                operator="equal",
                formula=["1"],
                fill=PatternFill("solid", fgColor="FFFF00"),
            ),
        )

    untouched = workbook.create_sheet("Untouched")
    untouched["A1"] = "=ROW()"
    untouched["A2"] = "=ROW()"
    untouched["C1"] = "=ROW(C1)"
    untouched["C2"] = 2
    untouched["E1"] = "=A1"
    untouched["F1"] = 1
    untouched["E2"] = 2
    untouched["F2"] = 3
    untouched["H1"] = "untouched-zip-sentinel"
    workbook.save(path)
    _close_workbook(workbook)
    if inject_advanced:
        _inject_advanced_formula_records(path, include_vba=include_vba)


def _transplant_untouched_parts(source: str, target: str, *, include_vba=False):
    names = ["xl/worksheets/sheet2.xml", _CUSTOM_PART]
    if include_vba:
        names.append(_VBA_PART)
    with zipfile.ZipFile(target, "r") as archive:
        target_names = set(archive.namelist())
    payloads = {name: _zip_bytes(source, name) for name in names}
    _rewrite_zip(
        target,
        replacements={name: payload for name, payload in payloads.items() if name in target_names},
        additions={name: payload for name, payload in payloads.items() if name not in target_names},
    )


def _column_operations(target_side="A", *, start_order=1):
    source_side = "B" if target_side == "A" else "A"
    operations = [
        {
            "kind": "insert_cols",
            "sheet": "S1",
            "target_side": target_side,
            "target_logical_slot": 2,
            "target_physical_anchor": 2,
            "count": 2,
            "source_side": source_side,
            "source_physical_cols": [2, 3],
            "metadata_scope": list(smt._COLUMN_ACTION_METADATA_SCOPE),
            "batch_id": "column-action-insert",
            "action_id": "column-action-insert",
            "order": start_order,
        },
        {
            "kind": "copy_cols",
            "sheet": "S1",
            "target_side": target_side,
            "target_logical_slot": 2,
            "target_physical_anchor": 2,
            "count": 2,
            "source_side": source_side,
            "source_physical_cols": [2, 3],
            "metadata_scope": list(smt._COLUMN_ACTION_METADATA_SCOPE),
            "batch_id": "column-action-insert",
            "action_id": "column-action-insert",
            "order": start_order + 1,
        },
        {
            "kind": "delete_cols",
            "sheet": "S1",
            "target_side": target_side,
            "target_logical_slot": 5,
            "target_physical_anchor": 5,
            "count": 1,
            "source_side": source_side,
            "source_physical_cols": [],
            "metadata_scope": list(smt._COLUMN_ACTION_METADATA_SCOPE),
            "batch_id": "column-action-delete",
            "action_id": "column-action-delete",
            "order": start_order + 2,
        },
    ]
    return operations


def _fake_app(file_a: str, file_b: str):
    app = object.__new__(smt.SowMergeApp)
    app.file_a = file_a
    app.file_b = file_b
    app.base_path = None
    app._merge_mine_snapshot = None
    app.sheet_views = {}
    app.manual_a_cell_ops = {}
    app.manual_b_cell_ops = {}
    app.manual_a_formula_cache_ops = {}
    app.manual_b_formula_cache_ops = {}
    app.manual_a_row_ops = []
    app.manual_b_row_ops = []
    app.manual_a_column_ops = []
    app.manual_b_column_ops = []
    app.manual_sheet_ops = []
    app.auto_sheet_ops = []
    app._wb_a_val = None
    app._wb_b_val = None
    app._wb_base_val = None
    app._wb_a_edit = None
    app._wb_b_edit = None
    app._wb_base_edit = None
    return app


def _fixture_set(extension=".xlsx"):
    temporary = tempfile.TemporaryDirectory(prefix="sow_native_column_replay_")
    root = temporary.name
    include_vba = extension.lower() == ".xlsm"
    mine = os.path.join(root, "mine" + extension)
    theirs = os.path.join(root, "theirs" + extension)
    expected = os.path.join(root, "expected" + extension)
    try:
        _make_fidelity_book(mine, ("A", "B", "C", "D"), include_vba=include_vba)
        _make_fidelity_book(
            theirs,
            ("A", "X", "Y", "B", "D"),
            decorated=(2, 3),
            include_vba=include_vba,
        )
        _make_fidelity_book(
            expected,
            ("A", "X", "Y", "B", "D"),
            decorated=(2, 3),
            include_vba=include_vba,
        )
        workbook = load_workbook(expected, data_only=False, keep_vba=include_vba)
        try:
            workbook["S1"]["E5"] = "pre-structure-edit"
            workbook.save(expected)
        finally:
            _close_workbook(workbook)
        _inject_advanced_formula_records(expected, include_vba=include_vba)
        _transplant_untouched_parts(mine, expected, include_vba=include_vba)
        inputs = (mine, theirs, expected)
        return _OwnedFixture(
            temporary=temporary,
            root=root,
            mine=mine,
            theirs=theirs,
            expected=expected,
            input_hashes={path: _sha256(path) for path in inputs},
        )
    except BaseException as error:
        errors: list[str] = []
        try:
            temporary.cleanup()
        except BaseException as cleanup_error:
            errors.append(f"fixture construction cleanup: {cleanup_error!r}")
        if os.path.lexists(root):
            errors.append(f"fixture construction root remains: {root!r}")
        _finish_cleanup(error, errors)
        raise


def _real_excel_fixture_set():
    """Create an Excel-openable fixture without synthetic OOXML records.

    Shared/array/data-table XML and custom ZIP-part preservation are covered by
    the package-level mocked gates.  The real COM gate deliberately starts from
    ordinary openpyxl workbooks so a malformed synthetic source cannot be
    mistaken for a native column-replay failure.
    """
    temporary = tempfile.TemporaryDirectory(prefix="sow_native_column_replay_real_")
    root = temporary.name
    mine = os.path.join(root, "mine.xlsx")
    theirs = os.path.join(root, "theirs.xlsx")
    try:
        _make_fidelity_book(
            mine,
            ("A", "B", "C", "D"),
            inject_advanced=False,
        )
        _make_fidelity_book(
            theirs,
            ("A", "X", "Y", "B", "D"),
            decorated=(2, 3),
            inject_advanced=False,
        )
        inputs = (mine, theirs)
        return _OwnedFixture(
            temporary=temporary,
            root=root,
            mine=mine,
            theirs=theirs,
            expected=None,
            input_hashes={path: _sha256(path) for path in inputs},
        )
    except BaseException as error:
        errors: list[str] = []
        try:
            temporary.cleanup()
        except BaseException as cleanup_error:
            errors.append(f"real fixture construction cleanup: {cleanup_error!r}")
        if os.path.lexists(root):
            errors.append(f"real fixture construction root remains: {root!r}")
        _finish_cleanup(error, errors)
        raise


def _assert_fidelity_output(path: str, source: str, *, include_vba=False):
    workbook = load_workbook(path, data_only=False, keep_vba=include_vba)
    try:
        sheet = workbook["S1"]
        assert [sheet.cell(1, col).value for col in range(1, 6)] == ["A", "X", "Y", "B", "D"]
        assert sheet["E5"].value == "pre-structure-edit"
        assert sheet["B3"].value == "=A3*2"
        assert sheet["C3"].value == "=B3*2"
        assert sheet.column_dimensions["B"].width == 20.0
        assert sheet.column_dimensions["B"].hidden is True
        assert sheet["B2"].fill.fgColor.rgb in ("0022AA66", "FF22AA66")
        assert sheet["B2"].font.bold is True
        assert sheet["B2"].comment.text == "comment-X"
        assert sheet["B2"].hyperlink.target.endswith("/X")
        assert any(str(item.sqref) == "B2:C8" for item in sheet.data_validations.dataValidation)
        assert "B5:C5" in {str(value) for value in sheet.merged_cells.ranges}
        assert [str(key.sqref) for key in sheet.conditional_formatting._cf_rules] == ["B2:C8"]
    finally:
        _close_workbook(workbook)

    assert _zip_bytes(path, "xl/worksheets/sheet2.xml") == _zip_bytes(
        source, "xl/worksheets/sheet2.xml"
    )
    assert _zip_bytes(path, _CUSTOM_PART) == _zip_bytes(source, _CUSTOM_PART)
    if include_vba:
        assert _zip_bytes(path, _VBA_PART) == _zip_bytes(source, _VBA_PART)


def _assert_real_excel_output(path: str, source: str):
    """Assert semantic fidelity after a real Excel SaveCopyAs round trip."""
    workbook = load_workbook(path, data_only=False)
    source_workbook = load_workbook(source, data_only=False)
    try:
        sheet = workbook["S1"]
        assert [sheet.cell(1, col).value for col in range(1, 6)] == ["A", "X", "Y", "B", "D"]
        assert sheet["E5"].value == "pre-structure-edit"
        assert sheet["B3"].value == "=A3*2"
        assert sheet["C3"].value == "=B3*2"
        assert sheet.column_dimensions["B"].width == 20.0
        assert sheet.column_dimensions["B"].hidden is True
        assert sheet["B2"].fill.fgColor.rgb in ("0022AA66", "FF22AA66")
        assert sheet["B2"].font.bold is True
        assert sheet["B2"].comment.text == "comment-X"
        assert sheet["B2"].hyperlink.target.endswith("/X")
        assert any(str(item.sqref) == "B2:C8" for item in sheet.data_validations.dataValidation)
        assert "B5:C5" in {str(value) for value in sheet.merged_cells.ranges}
        assert [str(key.sqref) for key in sheet.conditional_formatting._cf_rules] == ["B2:C8"]
        expected_untouched = {
            cell.coordinate: cell.value
            for row in source_workbook["Untouched"].iter_rows()
            for cell in row
            if cell.value is not None
        }
        actual_untouched = {
            cell.coordinate: cell.value
            for row in workbook["Untouched"].iter_rows()
            for cell in row
            if cell.value is not None
        }
        assert actual_untouched == expected_untouched
    finally:
        _close_workbook(workbook)
        _close_workbook(source_workbook)


def _ops_path_from_powershell(script: str) -> str:
    match = re.search(r"\$opsPath='((?:[^']|'')*)';", script)
    assert match is not None, script[:1000]
    return match.group(1).replace("''", "'")


def _capture_native_payload(*, row_first: bool):
    signature = inspect.signature(smt._build_manual_merge_output_with_excel)
    assert "column_ops" in signature.parameters, signature
    fixture = _fixture_set(".xlsx")
    mine, theirs = fixture.mine, fixture.theirs
    output = os.path.join(
        fixture.root,
        "payload-row-first.xlsx" if row_first else "payload-column-first.xlsx",
    )
    captured = {}
    original_run = smt.subprocess.run

    def fake_run(command, **_kwargs):
        script = command[-1]
        ops_path = _ops_path_from_powershell(script)
        with open(ops_path, "r", encoding="utf-8") as stream:
            captured["payload"] = json.load(stream)
        captured["script"] = script
        # Model only the row/column crossing.  The row is copied from BASE and
        # the columns from B, so the operation replayed last owns B3/C3.
        crossing_winner = "mine"
        for operation in captured["payload"]["structural_ops"]:
            if operation["kind"] == "copy_cols":
                crossing_winner = "column:B"
            elif operation["kind"] == "insert_rows":
                crossing_winner = "row:BASE"
        captured["crossing_winner"] = crossing_winner
        shutil.copy2(mine, output)
        return SimpleNamespace(returncode=0, stdout="", stderr="")

    if row_first:
        row_ops = [{
            "kind": "insert_rows",
            "sheet": "S1",
            "row": 3,
            "count": 1,
            "source_side": "BASE",
            "source_rows": [3],
            "order": 1,
        }]
        column_ops = _column_operations("A", start_order=2)
    else:
        column_ops = _column_operations("A", start_order=1)
        row_ops = [{
            "kind": "insert_rows",
            "sheet": "S1",
            "row": 3,
            "count": 1,
            "source_side": "BASE",
            "source_rows": [3],
            "order": 4,
        }]

    primary = None
    try:
        smt.subprocess.run = fake_run
        ok = smt._build_manual_merge_output_with_excel(
            mine,
            output,
            {("S1", 5, 5): "final-cell"},
            row_ops=row_ops,
            column_ops=column_ops,
            source_paths={"B": theirs, "BASE": mine},
        )
        assert ok is True
        return captured, row_ops, column_ops
    except BaseException as error:
        primary = error
        raise
    finally:
        _cleanup_owned_fixture(
            fixture,
            primary,
            restore=lambda: setattr(smt.subprocess, "run", original_run),
        )


def _assert_native_payload_order(*, row_first: bool):
    captured, row_ops, column_ops = _capture_native_payload(row_first=row_first)
    payload = captured["payload"]
    assert payload["column_ops"] == column_ops
    assert payload["row_ops"] == row_ops
    assert captured["payload"]["cell_ops"][0]["r"] == 5
    assert captured["payload"]["cell_ops"][0]["c"] == 5
    expected_kinds = (
        ["insert_rows", "insert_cols", "copy_cols", "delete_cols"]
        if row_first else
        ["insert_cols", "copy_cols", "delete_cols", "insert_rows"]
    )
    assert [op["kind"] for op in payload["structural_ops"]] == expected_kinds
    assert [op["order"] for op in payload["structural_ops"]] == [1, 2, 3, 4]
    assert captured["crossing_winner"] == (
        "column:B" if row_first else "row:BASE"
    )
    script = captured["script"]
    assert "payload.structural_ops" in script
    assert all(kind in script for kind in ("insert_cols", "delete_cols", "copy_cols"))
    assert script.index("payload.structural_ops") < script.index("payload.cell_ops")
    # The COM script must not regroup rows/columns, which would destroy the
    # explicitly captured winner at their intersection.
    assert "foreach($op in @($payload.column_ops" not in script
    assert "foreach($op in @($payload.row_ops" not in script


def test_native_builder_replays_column_then_row_before_cells():
    _assert_native_payload_order(row_first=False)


def test_native_builder_replays_row_then_column_before_cells():
    _assert_native_payload_order(row_first=True)


def test_action_time_column_remap_produces_final_save_coordinate():
    app = _fake_app("mine.xlsx", "theirs.xlsx")
    app.manual_a_cell_ops = {("S1", 5, 4): "pre-structure-edit"}
    app.manual_a_formula_cache_ops = {("S1", 5, 4): "cached-value"}
    insert_plan = smt.ColumnBlockActionPlan(
        action_id="column-action-insert",
        sheet="S1",
        block_ordinal=1,
        logical_start=2,
        logical_end=3,
        source_side="B",
        target_side="A",
        source_physical_cols=(2, 3),
        target_physical_cols=(),
        target_physical_anchor=2,
        count=2,
        action_kind="insert_copy",
    )
    delete_plan = smt.ColumnBlockActionPlan(
        action_id="column-action-delete",
        sheet="S1",
        block_ordinal=2,
        logical_start=5,
        logical_end=5,
        source_side="B",
        target_side="A",
        source_physical_cols=(),
        target_physical_cols=(5,),
        target_physical_anchor=5,
        count=1,
        action_kind="delete",
    )
    smt.SowMergeApp.remap_manual_cell_operations_for_column_action(app, "A", insert_plan)
    assert app.manual_a_cell_ops == {("S1", 5, 6): "pre-structure-edit"}
    smt.SowMergeApp.remap_manual_cell_operations_for_column_action(app, "A", delete_plan)
    assert app.manual_a_cell_ops == {("S1", 5, 5): "pre-structure-edit"}
    assert app.manual_a_formula_cache_ops == {("S1", 5, 5): "cached-value"}


def _run_public_route(extension: str):
    fixture = _fixture_set(extension)
    mine, theirs, expected = fixture.mine, fixture.theirs, fixture.expected
    assert expected is not None
    app = _fake_app(mine, theirs)
    app.manual_a_column_ops = _column_operations("A")
    # The live row action has already shifted r4 -> r5 and the column action
    # helper has already shifted c4 -> c6 -> c5.  Save consumes the final key;
    # it must not remap it a second time.
    app.manual_a_cell_ops = {("S1", 5, 5): "pre-structure-edit"}
    app.manual_a_row_ops = [
        {"kind": "insert_rows", "sheet": "S1", "row": 3, "count": 1, "order": 4},
    ]
    captured = {}
    original_builder = smt._build_manual_merge_output_with_excel
    original_reopen = smt._excel_reopen_validate

    def fake_builder(
        src,
        out,
        manual_ops,
        row_ops=None,
        column_ops=None,
        sheet_ops=None,
        source_paths=None,
    ):
        captured.update({
            "src": src,
            "manual_ops": dict(manual_ops),
            "row_ops": list(row_ops or []),
            "column_ops": list(column_ops or []),
            "sheet_ops": list(sheet_ops or []),
            "source_paths": dict(source_paths or {}),
        })
        shutil.copy2(expected, out)
        return True

    primary = None
    try:
        smt._build_manual_merge_output_with_excel = fake_builder
        smt._excel_reopen_validate = lambda _path: True
        output = smt.SowMergeApp.build_manual_merge_output_file(app)
        assert captured["src"] == mine
        assert captured["column_ops"] == _column_operations("A")
        assert captured["source_paths"]["B"] == theirs
        assert captured["manual_ops"] == {("S1", 5, 5): "pre-structure-edit"}
        assert captured["row_ops"] == app.manual_a_row_ops
        _assert_fidelity_output(output, mine, include_vba=extension == ".xlsm")
    except BaseException as error:
        primary = error
        raise
    finally:
        def restore():
            smt._build_manual_merge_output_with_excel = original_builder
            smt._excel_reopen_validate = original_reopen
        _cleanup_owned_fixture(fixture, primary, restore=restore)


def test_public_xlsx_route_preserves_fidelity_and_remaps_mixed_coordinates():
    _run_public_route(".xlsx")


def test_public_xlsm_route_preserves_macro_and_untouched_zip_parts():
    _run_public_route(".xlsm")


def test_native_failure_never_replaces_target_and_same_batch_can_retry():
    fixture = _fixture_set(".xlsx")
    root, mine, theirs, expected = fixture.root, fixture.mine, fixture.theirs, fixture.expected
    assert expected is not None
    app = _fake_app(mine, theirs)
    app.manual_a_column_ops = _column_operations("A")
    user_target = os.path.join(root, "user-target.xlsx")
    shutil.copy2(mine, user_target)
    target_hash = _sha256(user_target)
    calls = []
    original_native = smt._build_manual_merge_output_with_excel
    original_fallback = smt._build_manual_merge_output_with_openpyxl
    original_reopen = smt._excel_reopen_validate

    def fake_native(
        _src,
        out,
        _manual_ops,
        row_ops=None,
        column_ops=None,
        sheet_ops=None,
        source_paths=None,
    ):
        calls.append(list(column_ops or []))
        if len(calls) == 1:
            return False
        shutil.copy2(expected, out)
        return True

    def forbidden_fallback(*_args, **_kwargs):
        raise AssertionError("column replay must never use the openpyxl structural fallback")

    primary = None
    try:
        smt._build_manual_merge_output_with_excel = fake_native
        smt._build_manual_merge_output_with_openpyxl = forbidden_fallback
        smt._excel_reopen_validate = lambda _path: True
        try:
            smt.SowMergeApp.build_manual_merge_output_file(app)
        except RuntimeError as exc:
            assert "原生" in str(exc) or "列" in str(exc), str(exc)
        else:
            raise AssertionError("native replay failure unexpectedly returned a save artifact")
        assert _sha256(user_target) == target_hash
        assert app.manual_a_column_ops == _column_operations("A")
        assert calls == [_column_operations("A")]

        retry_output = smt.SowMergeApp.build_manual_merge_output_file(app)
        assert calls == [_column_operations("A"), _column_operations("A")]
        smt.SowMergeApp._atomic_replace_file_with_retry(
            app, retry_output, user_target, retries=1, delay_sec=0
        )
        assert _sha256(user_target) == _sha256(expected)
        _assert_fidelity_output(user_target, mine)
    except BaseException as error:
        primary = error
        raise
    finally:
        def restore():
            smt._build_manual_merge_output_with_excel = original_native
            smt._build_manual_merge_output_with_openpyxl = original_fallback
            smt._excel_reopen_validate = original_reopen
        _cleanup_owned_fixture(fixture, primary, restore=restore)


def test_reopen_validation_failure_removes_artifact_and_fails_closed():
    fixture = _fixture_set(".xlsx")
    root, mine, theirs, expected = fixture.root, fixture.mine, fixture.theirs, fixture.expected
    assert expected is not None
    app = _fake_app(mine, theirs)
    app.manual_a_column_ops = _column_operations("A")
    user_target = os.path.join(root, "validation-target.xlsx")
    shutil.copy2(mine, user_target)
    target_hash = _sha256(user_target)
    captured = {}
    original_native = smt._build_manual_merge_output_with_excel
    original_fallback = smt._build_manual_merge_output_with_openpyxl
    original_reopen = smt._excel_reopen_validate

    def fake_native(
        _src,
        out,
        _manual_ops,
        row_ops=None,
        column_ops=None,
        sheet_ops=None,
        source_paths=None,
    ):
        captured["out"] = out
        captured["column_ops"] = list(column_ops or [])
        shutil.copy2(expected, out)
        # Native replay now performs the Excel reopen inside the builder. A
        # failed reopen is represented by False; the public route must still
        # fail closed and remove the untrusted artifact.
        os.remove(out)
        return False

    def forbidden_fallback(*_args, **_kwargs):
        raise AssertionError("failed native validation must not use openpyxl fallback")

    primary = None
    try:
        smt._build_manual_merge_output_with_excel = fake_native
        smt._build_manual_merge_output_with_openpyxl = forbidden_fallback
        try:
            smt.SowMergeApp.build_manual_merge_output_file(app)
        except RuntimeError as exc:
            assert "原生列结构回放失败" in str(exc) and "未替换" in str(exc), str(exc)
        else:
            raise AssertionError("failed Excel reopen validation unexpectedly returned output")
        assert captured["column_ops"] == _column_operations("A")
        assert not os.path.exists(captured["out"])
        assert _sha256(user_target) == target_hash
        assert app.manual_a_column_ops == _column_operations("A")
    except BaseException as error:
        primary = error
        raise
    finally:
        def restore():
            smt._build_manual_merge_output_with_excel = original_native
            smt._build_manual_merge_output_with_openpyxl = original_fallback
            smt._excel_reopen_validate = original_reopen
        _cleanup_owned_fixture(fixture, primary, restore=restore)


def test_real_excel_native_column_replay_optional():
    if not _RUN_REAL_EXCEL:
        return "skipped"
    fixture = _real_excel_fixture_set()
    mine, theirs = fixture.mine, fixture.theirs
    primary = None
    try:
        assert smt._excel_reopen_validate(mine), "Excel cannot reopen pristine mine fixture"
        assert smt._excel_reopen_validate(theirs), "Excel cannot reopen pristine theirs fixture"
        output = os.path.join(fixture.root, "real-excel-output.xlsx")
        ok = smt._build_manual_merge_output_with_excel(
            mine,
            output,
            {("S1", 5, 5): "pre-structure-edit"},
            row_ops=[],
            column_ops=_column_operations("A"),
            source_paths={"B": theirs},
        )
        assert ok is True, "real Excel native column replay failed"
        assert smt._excel_reopen_validate(output), "Excel cannot reopen native replay output"
        _assert_real_excel_output(output, mine)
    except BaseException as error:
        primary = error
        raise
    finally:
        _cleanup_owned_fixture(fixture, primary)


def main():
    tests = (
        test_native_builder_replays_column_then_row_before_cells,
        test_native_builder_replays_row_then_column_before_cells,
        test_action_time_column_remap_produces_final_save_coordinate,
        test_public_xlsx_route_preserves_fidelity_and_remaps_mixed_coordinates,
        test_public_xlsm_route_preserves_macro_and_untouched_zip_parts,
        test_native_failure_never_replaces_target_and_same_batch_can_retry,
        test_reopen_validation_failure_removes_artifact_and_fails_closed,
        test_real_excel_native_column_replay_optional,
    )
    passed = 0
    skipped = 0
    for test in tests:
        result = test()
        if result == "skipped":
            skipped += 1
            print(
                "SKIP: test_real_excel_native_column_replay_optional "
                "(set SOW_RUN_REAL_EXCEL_COLUMN_REPLAY=1)"
            )
            continue
        passed += 1
        print(f"PASS: {test.__name__}")
    print(
        "PASS: native column save/reopen regression "
        f"({passed} passed, {skipped} skipped)"
    )


if __name__ == "__main__":
    main()

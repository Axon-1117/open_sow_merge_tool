"""Focused GUI regression for loading/read-only mutation gates.

The test intentionally calls command/event handlers directly.  A disabled Tk
button is not enough protection because row headers, the comparison panel,
keyboard commands, and app-level save commands can invoke the same mutations.
"""

from __future__ import annotations

import argparse
import copy
from dataclasses import fields, is_dataclass
import hashlib
import inspect
import json
import os
import tempfile
import time
from contextlib import contextmanager
from pathlib import Path
from types import SimpleNamespace

from openpyxl import Workbook

import sow_merge_tool as smt


_CASES = (
    "non-ready-zero-write",
    "only-diff-pending",
    "edit-ready-no-rescan",
    "stale-generation",
    "hidden-cache",
)
_ACTIVE_CASE_DEADLINE: float | None = None


def _make_book(path: str, rows) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    for row in rows:
        worksheet.append(list(row))
    workbook.save(path)
    workbook.close()


def _pump(root, seconds: float = 0.05) -> None:
    deadline = time.monotonic() + seconds
    while time.monotonic() < deadline:
        root.update_idletasks()
        root.update()
        time.sleep(0.005)


def _pump_one_tk_turn(root) -> None:
    """Run one real Tk turn without extending an action's deadline window."""
    root.update_idletasks()
    root.update()


def _wait_until(root, predicate, message: str, timeout: float = 15.0) -> None:
    deadline = time.monotonic() + timeout
    if _ACTIVE_CASE_DEADLINE is not None:
        deadline = min(deadline, _ACTIVE_CASE_DEADLINE)
    while time.monotonic() < deadline:
        _pump(root, 0.025)
        if predicate():
            return
    raise AssertionError(message)


def _open_ready_view(case):
    mine = str(case.root / "mine.xlsx")
    theirs = str(case.root / "theirs.xlsx")
    _make_book(
        mine,
        (
            ("id@id", "value", "note"),
            ("string", "string", "string"),
            ("row-1", "same-1", "same"),
            ("row-2", "mine-2", "same"),
            ("row-3", "mine-3", "same"),
            ("row-4", "same-4", "same"),
        ),
    )
    _make_book(
        theirs,
        (
            ("id@id", "value", "note"),
            ("string", "string", "string"),
            ("row-1", "same-1", "same"),
            ("row-2", "theirs-2", "same"),
            ("row-3", "theirs-3", "same"),
            ("row-4", "same-4", "same"),
        ),
    )
    case.track_inputs(mine, theirs)
    app = smt.SowMergeApp(mine, theirs)
    case.track_app(app)
    app.root.deiconify()
    app.root.geometry("940x760")
    app.nb.select(app._sheet_containers["Data"])

    _wait_until(
        app.root,
        lambda: (
            app.sheet_views.get("Data") is not None
            and bool(getattr(app.sheet_views["Data"], "_data_ready", False))
            and app._is_sheet_exact_current("Data")
        ),
        "small test workbook did not publish current exact data",
    )
    view = app.sheet_views["Data"]
    assert app._is_sheet_exact_current("Data")
    assert view._data_ready
    # B3 is view-only: exact snapshot readiness must not eagerly materialize
    # editable workbooks or force a foreground rescan.
    assert not app._edit_workbooks_ready(), "test setup eagerly loaded edit workbooks"
    _settle_initial_tab_watchdog(app, view, (mine, theirs))
    return app, view, (mine, theirs)


def _file_digest(path: str) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as handle:
        for chunk in iter(lambda: handle.read(64 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _settle_initial_tab_watchdog(app, view, source_paths) -> None:
    """Let the selected-tab retry fire while its real exact surface is healthy.

    ``_on_tab_changed`` installs a 700 ms selected-tab watchdog while the
    initial exact request is still pending.  The non-ready matrix later forces
    ``_data_ready`` false to exercise a guard; without this fence, that old
    watchdog can be dispatched by the action's one Tk turn and be mistaken for
    a mutation-owned recompute.  The fence runs before any synthetic lifecycle
    state while the production snapshot is still current and full-detail.
    """
    entry_before = app._sheet_exact_entry("Data") or {}
    assert str(getattr(app, "selected_sheet", "") or "") == "Data"
    assert app._is_sheet_exact_current("Data"), entry_before
    assert bool(entry_before.get("full_detail_terminal", False)), entry_before
    assert bool(getattr(view, "_data_ready", False))
    assert bool(getattr(view, "_prepared_complete", False))
    assert bool(getattr(view, "_row_model_exact", False))
    assert bool(getattr(view, "_pair_diff_full_exact", False))
    assert not bool(getattr(view, "_pending_exact_render", False))
    assert not app._edit_workbooks_ready()
    before = _mutation_snapshot(app, view, source_paths)
    assert not before["scheduler_semantics"]["compute_queue"], before
    assert not before["scheduler_semantics"]["compute_inflight"], before
    assert before["scheduler_semantics"]["active_compute_sheet"] is None, before
    assert not before["scheduler_semantics"]["ui_tasks"], before

    fired = {"value": False}

    def _fence_reached():
        fired["value"] = True

    started = time.monotonic()
    deadline = started + 0.9
    if _ACTIVE_CASE_DEADLINE is not None:
        deadline = min(deadline, _ACTIVE_CASE_DEADLINE)
    assert deadline - started > 0.7, "case deadline left no safe watchdog fence"
    after_id = view.root.after(725, _fence_reached)
    try:
        while time.monotonic() < deadline:
            _pump_one_tk_turn(view.root)
            if fired["value"]:
                break
            time.sleep(0.005)
    finally:
        if not fired["value"]:
            try:
                view.root.after_cancel(after_id)
            except Exception:
                pass
    elapsed = time.monotonic() - started
    assert fired["value"], f"initial tab watchdog fence exceeded 0.9s ({elapsed:.3f}s)"
    assert 0.7 < elapsed <= 0.9, elapsed

    entry_after = app._sheet_exact_entry("Data") or {}
    assert str(getattr(app, "selected_sheet", "") or "") == "Data"
    assert app._is_sheet_exact_current("Data"), entry_after
    assert bool(entry_after.get("full_detail_terminal", False)), entry_after
    assert bool(getattr(view, "_data_ready", False))
    assert bool(getattr(view, "_prepared_complete", False))
    assert bool(getattr(view, "_row_model_exact", False))
    assert bool(getattr(view, "_pair_diff_full_exact", False))
    assert not bool(getattr(view, "_pending_exact_render", False))
    assert not app._edit_workbooks_ready()
    after = _mutation_snapshot(app, view, source_paths)
    _assert_snapshot_section_unchanged(
        before["hard_mutation"], after["hard_mutation"],
        "initial-tab-watchdog fence", "hard_mutation",
    )
    _assert_snapshot_section_unchanged(
        before["scheduler_semantics"], after["scheduler_semantics"],
        "initial-tab-watchdog fence", "scheduler_semantics",
    )


def _path_snapshot(path: Path) -> tuple[bool, bytes | None]:
    if path.exists():
        return True, path.read_bytes()
    return False, None


class _CaseContext:
    def __init__(self, root: Path):
        self.root = root
        self._input_before: dict[str, str] = {}
        self._apps: list[object] = []

    def track_inputs(self, *paths: str) -> None:
        for path in paths:
            path_s = str(path)
            self._input_before.setdefault(path_s, _file_digest(path_s))

    def verify_inputs(self) -> None:
        after = {path: _file_digest(path) for path in self._input_before}
        assert after == self._input_before, (
            "synthetic source inputs changed", self._input_before, after,
        )

    def input_provenance(self) -> list[dict[str, str]]:
        after = {path: _file_digest(path) for path in self._input_before}
        return [
            {
                "path": path,
                "before_sha256": digest,
                "after_sha256": after[path],
            }
            for path, digest in sorted(self._input_before.items())
        ]

    def track_app(self, app) -> None:
        self._apps.append(app)

    def close_app(self, app) -> None:
        if app is None:
            return
        try:
            for view in tuple(getattr(app, "sheet_views", {}).values()):
                after_id = getattr(view, "_settings_save_id", None)
                if after_id:
                    try:
                        view.frame.after_cancel(after_id)
                    finally:
                        view._settings_save_id = None
            app._shutdown_root()
        finally:
            if app in self._apps:
                self._apps.remove(app)

    def close_remaining(self) -> None:
        for app in tuple(self._apps):
            self.close_app(app)


def _canonical_primitive(value):
    """Normalize prepared/cache data without descending into Tk or workbook state."""
    if value is None or isinstance(value, (bool, int, float, str, bytes)):
        return value
    if isinstance(value, (tuple, list)):
        return tuple(_canonical_primitive(item) for item in value)
    if isinstance(value, (set, frozenset)):
        return tuple(sorted((_canonical_primitive(item) for item in value), key=repr))
    if isinstance(value, dict):
        items = [
            (_canonical_primitive(key), _canonical_primitive(item))
            for key, item in value.items()
        ]
        return tuple(sorted(items, key=repr))
    if is_dataclass(value):
        return (
            "dataclass",
            f"{type(value).__module__}.{type(value).__qualname__}",
            tuple(
                (field.name, _canonical_primitive(getattr(value, field.name)))
                for field in fields(value)
            ),
        )
    # Unknown objects are identities only. This deliberately covers neither
    # Tk widgets/locks nor OpenPyXL objects, so the view-only assertion can
    # never turn a read-only workbook into a source of reads.
    return ("opaque", f"{type(value).__module__}.{type(value).__qualname__}", id(value))


def _workbook_handle_identity(workbook):
    if workbook is None:
        return None
    return (
        id(workbook),
        f"{type(workbook).__module__}.{type(workbook).__qualname__}",
        bool(getattr(workbook, "read_only", False)),
    )


def _confidence_snapshot(confidence):
    return (
        float(getattr(confidence, "score", 0.0)),
        bool(getattr(confidence, "ambiguous", False)),
        str(getattr(confidence, "reason", "")),
        tuple(str(item) for item in tuple(getattr(confidence, "evidence", ()) or ())),
        tuple(str(item) for item in tuple(getattr(confidence, "cause_codes", ()) or ())),
    )


def _column_model_snapshot(model):
    if model is None:
        return None
    slots = tuple(
        (
            int(getattr(slot, "logical_idx", -1)),
            getattr(slot, "mine_col", None),
            getattr(slot, "base_col", None),
            getattr(slot, "theirs_col", None),
            str(getattr(slot, "state", "")),
            _confidence_snapshot(getattr(slot, "confidence", None)),
            getattr(slot, "base_boundary", None),
            getattr(slot, "origin_side", None),
        )
        for slot in tuple(getattr(model, "slots", ()) or ())
    )
    blocks = tuple(
        (
            int(getattr(block, "ordinal", -1)),
            tuple(int(index) for index in tuple(getattr(block, "slot_indices", ()) or ())),
            str(getattr(block, "state", "")),
            _confidence_snapshot(getattr(block, "confidence", None)),
        )
        for block in tuple(getattr(model, "blocks", ()) or ())
    )
    mappings = tuple(
        (
            name,
            tuple((int(left), int(right)) for left, right in tuple(
                getattr(getattr(model, name, None), "entries", ()) or ()
            )),
        )
        for name in (
            "mine_physical_to_logical", "base_physical_to_logical", "theirs_physical_to_logical",
            "mine_logical_to_physical", "base_logical_to_physical", "theirs_logical_to_physical",
        )
    )
    cache_key = getattr(model, "cache_key", None)
    return (
        _canonical_primitive(cache_key),
        slots,
        blocks,
        mappings,
        _confidence_snapshot(getattr(model, "confidence", None)),
    )


def _immutable_view_cache_snapshot(view):
    """Capture the prepared data/projection facts that an action must not change."""
    cache = getattr(view, "column_comparison_cache", None)
    projection = getattr(view, "column_projection", None)
    return {
        "prepared_raw": (
            _canonical_primitive(getattr(view, "pair_raw_parts_a", {})),
            _canonical_primitive(getattr(view, "pair_raw_parts_base", {})),
            _canonical_primitive(getattr(view, "pair_raw_parts_b", {})),
        ),
        "row_base_maps": (
            tuple(tuple(pair) for pair in tuple(getattr(view, "row_pairs", ()) or ())),
            _canonical_primitive(getattr(view, "row_a_to_pair_idx", {})),
            _canonical_primitive(getattr(view, "row_b_to_pair_idx", {})),
            _canonical_primitive(getattr(view, "mine_to_base_row", {})),
            _canonical_primitive(getattr(view, "theirs_to_base_row", {})),
            _canonical_primitive(getattr(view, "pair_base_row_override", {})),
            _canonical_primitive(getattr(view, "_missing_base_row_map", {})),
        ),
        "diff_maps": (
            _canonical_primitive(getattr(view, "pair_diff_cols", {})),
            _canonical_primitive(getattr(view, "pair_base_diff_cols", {})),
        ),
        "exactness": {
            name: bool(getattr(view, name, False))
            for name in (
                "_data_ready", "_prepared_complete", "_row_model_exact",
                "_pair_diff_full_exact", "_base_diff_full_exact",
                "_cache_formula_aware", "_only_diff_rows_exact",
            )
        },
        "column_model": (
            _column_model_snapshot(getattr(projection, "model", None)),
            _column_model_snapshot(getattr(cache, "model", None)),
            tuple(sorted(int(col) for col in (getattr(cache, "structural_diff_cols", ()) or ()))),
            tuple(sorted(int(col) for col in (getattr(cache, "unresolved_cols", ()) or ()))),
        ),
        "data_projection_versions": {
            name: int(getattr(view, name, 0) or 0)
            for name in (
                "_row_model_version", "_column_model_version", "_data_version",
                "_mine_edit_version", "_base_edit_version", "_theirs_edit_version",
                "_column_projection_generation", "_accepted_common_insertion_mutation_generation",
                "_virtual_publish_generation", "_virtual_column_window_generation",
            )
        },
    }


def _task_descriptor(task):
    return (
        getattr(task, "__qualname__", type(task).__name__),
        getattr(task, "_sow_cache_sheet", None),
    )


def _scheduler_semantics_snapshot(app):
    """Capture work semantics, intentionally excluding opaque Tk after tokens."""
    try:
        with app._compute_lock:
            compute_queue = _canonical_primitive(
                tuple(getattr(app, "_compute_queue", ()) or ())
            )
            compute_inflight = tuple(sorted(
                str(sheet) for sheet in (getattr(app, "_compute_inflight", set()) or set())
            ))
            active_compute_sheet = _canonical_primitive(
                getattr(app, "_active_compute_sheet", None)
            )
    except Exception as exc:
        raise AssertionError("unable to capture compute scheduler semantics") from exc
    try:
        with app._ui_task_lock:
            queued_tasks = tuple(
                _task_descriptor(task)
                for task in tuple(getattr(app, "_ui_tasks", ()) or ())
            )
    except Exception as exc:
        raise AssertionError("unable to capture UI task scheduler semantics") from exc
    return {
        "compute_queue": compute_queue,
        "compute_inflight": compute_inflight,
        "active_compute_sheet": active_compute_sheet,
        "ui_tasks": queued_tasks,
    }


def _edit_owner_snapshot(app, preload_thread):
    try:
        preload_thread_alive = bool(preload_thread is not None and preload_thread.is_alive())
    except Exception:
        preload_thread_alive = False
    request_audit = []
    for entry in tuple(getattr(app, "_edit_load_requests", ()) or ()):
        if isinstance(entry, dict):
            request_audit.append((
                str(entry.get("reason") or ""),
                str(entry.get("caller") or ""),
                bool(entry.get("ready", False)),
            ))
        else:
            request_audit.append(("invalid-audit-entry", _canonical_primitive(entry), None))
    return {
        "loading_started": bool(getattr(app, "_edit_loading_started", False)),
        "thread_identity": id(preload_thread) if preload_thread is not None else None,
        "thread_alive": preload_thread_alive,
        "preload_active": bool(getattr(app, "_edit_preload_active_event", None)
                               and app._edit_preload_active_event.is_set()),
        "loaded": bool(getattr(app, "_edit_loaded_event", None)
                       and app._edit_loaded_event.is_set()),
        "requests": tuple(request_audit),
        "interactive_owner": _canonical_primitive(
            tuple(getattr(app, "_interactive_owner_stack", ()) or ())
        ),
        "interactive_active": bool(getattr(app, "_interactive_action_event", None)
                                   and app._interactive_action_event.is_set()),
    }


def _guard_ui_snapshot(app, view, preload_thread):
    try:
        with app._ui_activity_lock:
            activity = (
                int(getattr(app, "_ui_activity_seq", 0) or 0),
                str(getattr(app, "_ui_activity_reason", "") or ""),
            )
    except Exception as exc:
        raise AssertionError("unable to capture UI activity telemetry") from exc
    return {
        "lifecycle": (
            str(getattr(view, "_lifecycle_state", "") or ""),
            str(getattr(view, "_lifecycle_error", "") or ""),
            bool(getattr(view, "_lifecycle_canceled", False)),
            int(getattr(view, "_lifecycle_generation", 0) or 0),
        ),
        "render": {
            "pending_exact_render": bool(getattr(view, "_pending_exact_render", False)),
            "virtual_window_start": getattr(view, "_virtual_window_start", None),
            "virtual_pending_start": getattr(view, "_virtual_pending_start", None),
            "virtual_column_window_start": getattr(view, "_virtual_column_window_start", None),
            "virtual_pending_column_start": getattr(view, "_virtual_pending_column_start", None),
            "mode_switch_seq": int(getattr(view, "_mode_switch_seq", 0) or 0),
            "only_diff_async_build_seq": int(getattr(view, "_only_diff_async_build_seq", 0) or 0),
        },
        "selection": (
            getattr(view, "selected_pair_idx", None),
            getattr(view, "selected_excel_row", None),
            getattr(view, "selected_excel_row_a", None),
            getattr(view, "selected_excel_row_b", None),
            getattr(view, "_cursor_cmp_sel_line", None),
            getattr(view, "_cursor_cmp_sel_col", None),
            getattr(view, "_last_cursor_cmp_pair_idx", None),
            getattr(view, "hover_pair_idx", None),
            getattr(view, "hover_col_idx", None),
            getattr(view, "hover_side", None),
        ),
        "activity": activity,
        "mutation_preload_marked": bool(
            getattr(view, "_last_mutation_started_edit_preload", False)
        ),
        "edit_owner": _edit_owner_snapshot(app, preload_thread),
    }


def _mutation_snapshot(app, view, source_paths):
    """Split hard write invariants from scheduler and benign guard/UI telemetry."""
    operation_attrs = (
        "manual_a_cell_ops",
        "manual_b_cell_ops",
        "manual_a_formula_cache_ops",
        "manual_b_formula_cache_ops",
        "manual_a_row_ops",
        "manual_b_row_ops",
        "manual_a_column_ops",
        "manual_b_column_ops",
        "manual_sheet_ops",
        "auto_sheet_ops",
        "undo_stack",
        "redo_stack",
    )
    overlays = []
    for sheet, overlay in sorted((getattr(app, "sheet_operation_overlays", {}) or {}).items()):
        overlays.append(
            (
                str(sheet),
                int(getattr(overlay, "topology_generation", 0)),
                int(getattr(overlay, "mutation_generation", 0)),
                _canonical_primitive(getattr(overlay, "cells", {})),
            )
        )
    preload_thread = getattr(app, "_edit_preload_thread", None)
    return {
        "hard_mutation": {
            "source_hashes": tuple(_file_digest(path) for path in source_paths),
            "workbook_handles": (
                _workbook_handle_identity(getattr(app, "_wb_a_val", None)),
                _workbook_handle_identity(getattr(app, "_wb_base_val", None)),
                _workbook_handle_identity(getattr(app, "_wb_b_val", None)),
                _workbook_handle_identity(getattr(app, "_wb_a_edit", None)),
                _workbook_handle_identity(getattr(app, "_wb_base_edit", None)),
                _workbook_handle_identity(getattr(app, "_wb_b_edit", None)),
            ),
            "edit_backend": (
                bool(app._edit_workbooks_ready()),
                app._wb_a_edit is None,
                app._wb_b_edit is None,
            ),
            "operations": {
                name: _canonical_primitive(getattr(app, name, ()))
                for name in operation_attrs
            },
            "overlays": tuple(overlays),
            "modified": (
                bool(app.modified_a),
                bool(app.modified_b),
                frozenset(app.modified_sheets_a),
                frozenset(app.modified_sheets_b),
            ),
            "touched_rows": tuple(sorted(
                int(row) for row in (getattr(view, "touched_rows", set()) or set())
            )),
            "immutable_view_cache": _immutable_view_cache_snapshot(view),
            "compute_generation": tuple(sorted(
                (str(sheet), int(generation))
                for sheet, generation in (getattr(app, "_sheet_compute_generation", {}) or {}).items()
            )),
        },
        "scheduler_semantics": _scheduler_semantics_snapshot(app),
        "guard_ui": _guard_ui_snapshot(app, view, preload_thread),
    }


def _diff_value_summary(value):
    try:
        rendered = repr(value)
    except Exception as exc:
        rendered = f"<unrepresentable {type(exc).__name__}>"
    try:
        size = len(value)
    except Exception:
        size = None
    return {
        "type": f"{type(value).__module__}.{type(value).__qualname__}",
        "length": size,
        "sha256": hashlib.sha256(rendered.encode("utf-8", "backslashreplace")).hexdigest(),
        "preview": rendered[:240] + ("..." if len(rendered) > 240 else ""),
    }


def _canonical_field_diff(before, after, *, limit: int = 48):
    """Return bounded path-level diagnostics without dumping full prepared rows."""
    diffs = []
    missing = object()

    def _walk(left, right, path):
        if len(diffs) >= limit or left == right:
            return
        if isinstance(left, dict) and isinstance(right, dict):
            for key in sorted(set(left) | set(right), key=repr):
                _walk(left.get(key, missing), right.get(key, missing), path + (str(key),))
            return
        if isinstance(left, (tuple, list)) and isinstance(right, (tuple, list)):
            if len(left) != len(right):
                diffs.append({
                    "path": ".".join(path),
                    "before": _diff_value_summary(left),
                    "after": _diff_value_summary(right),
                })
                return
            if len(left) > 64:
                diffs.append({
                    "path": ".".join(path),
                    "before": _diff_value_summary(left),
                    "after": _diff_value_summary(right),
                })
                return
            for index, (left_item, right_item) in enumerate(zip(left, right)):
                _walk(left_item, right_item, path + (str(index),))
            return
        diffs.append({
            "path": ".".join(path),
            "before": _diff_value_summary(left),
            "after": _diff_value_summary(right),
        })

    _walk(before, after, ())
    return tuple(diffs)


def _assert_snapshot_section_unchanged(before, after, action: str, section: str) -> None:
    differences = _canonical_field_diff(before, after)
    assert not differences, (
        f"{action} changed {section}: "
        + json.dumps(differences, ensure_ascii=False, sort_keys=True)
    )


def _assert_only_diff_cache_publish_hard_invariants(before, after, *, action: str) -> None:
    """Permit only the bounded render-cache facts an exact only-diff publish owns."""
    before_stable = copy.deepcopy(before)
    after_stable = copy.deepcopy(after)
    before_cache = before_stable["immutable_view_cache"]
    after_cache = after_stable["immutable_view_cache"]
    before_exactness = before_cache["exactness"]
    after_exactness = after_cache["exactness"]
    before_diff_maps = before_cache["diff_maps"]
    after_diff_maps = after_cache["diff_maps"]
    assert not bool(before_exactness["_only_diff_rows_exact"]), before
    assert bool(after_exactness["_only_diff_rows_exact"]), after
    before_versions = before_cache["data_projection_versions"]
    after_versions = after_cache["data_projection_versions"]
    assert int(after_versions["_data_version"]) == int(before_versions["_data_version"]) + 3, (
        before_versions,
        after_versions,
    )
    assert int(after_versions["_virtual_publish_generation"]) == int(
        before_versions["_virtual_publish_generation"]
    ), (before_versions, after_versions)
    before_cache["exactness"] = {
        name: value
        for name, value in before_exactness.items()
        if name not in {"_only_diff_rows_exact", "_base_diff_full_exact"}
    }
    after_cache["exactness"] = {
        name: value
        for name, value in after_exactness.items()
        if name not in {"_only_diff_rows_exact", "_base_diff_full_exact"}
    }
    # The caller has already proved the deterministic two-way completion's
    # sparse A/B map. Preserve the Base map path below; only this exact map
    # representation is intentionally replaced by the worker payload.
    assert len(before_diff_maps) == len(after_diff_maps) == 2
    before_cache["diff_maps"] = ("only-diff-pair-map-asserted", before_diff_maps[1])
    after_cache["diff_maps"] = ("only-diff-pair-map-asserted", after_diff_maps[1])
    before_cache["data_projection_versions"] = {
        name: value
        for name, value in before_versions.items()
        if name not in {"_data_version", "_virtual_publish_generation"}
    }
    after_cache["data_projection_versions"] = {
        name: value
        for name, value in after_versions.items()
        if name not in {"_data_version", "_virtual_publish_generation"}
    }
    _assert_snapshot_section_unchanged(
        before_stable, after_stable, action, "hard_mutation except only-diff cache publication"
    )

def _assert_request_audit_append(before, after, action: str, expected_request: dict) -> None:
    before_requests = tuple(before["requests"])
    after_requests = tuple(after["requests"])
    assert after_requests[:len(before_requests)] == before_requests, (
        f"{action} rewrote edit request audit", before_requests, after_requests,
    )
    assert len(after_requests) == len(before_requests) + 1, (
        f"{action} did not append exactly one edit request", before_requests, after_requests,
    )
    assert after_requests[-1] == (
        expected_request["request_reason"],
        expected_request["request_caller"],
        False,
    ), (action, expected_request, after_requests[-1])


def _assert_edit_owner_policy(before, after, action: str, policy: str, expected_request: dict | None) -> None:
    if policy == "unchanged":
        _assert_snapshot_section_unchanged(before, after, action, "guard_ui.edit_owner")
        return

    stable_before = dict(before)
    stable_after = dict(after)
    before_requests = stable_before.pop("requests")
    after_requests = stable_after.pop("requests")
    if policy == "existing-owner-audit":
        _assert_snapshot_section_unchanged(
            stable_before, stable_after, action, "existing edit owner"
        )
        assert bool(before["loading_started"]), (action, before)
        assert expected_request is not None
        _assert_request_audit_append(
            {"requests": before_requests}, {"requests": after_requests}, action, expected_request
        )
        return

    if policy == "first-demand":
        assert before["loading_started"] is False and before["thread_identity"] is None, (
            action, before,
        )
        assert after["loading_started"] is True and after["thread_identity"] is not None, (
            action, after,
        )
        assert after["thread_alive"] is True, (action, after)
        for name in ("preload_active", "loaded", "interactive_owner", "interactive_active"):
            assert after[name] == before[name], (action, name, before[name], after[name])
        assert expected_request is not None
        _assert_request_audit_append(
            {"requests": before_requests}, {"requests": after_requests}, action, expected_request
        )
        return
    raise AssertionError(f"unknown edit owner policy: {policy}")


def _assert_guard_ui_policy(
    before,
    after,
    action: str,
    *,
    edit_policy: str,
    expected_request: dict | None,
    ui_activity_reason: str | None = None,
    expected_pair_idx: int | None = None,
) -> None:
    if edit_policy == "first-demand":
        before_lifecycle = tuple(before["lifecycle"])
        after_lifecycle = tuple(after["lifecycle"])
        assert len(before_lifecycle) == len(after_lifecycle) == 4, (
            action, before_lifecycle, after_lifecycle,
        )
        assert before_lifecycle[0] == "EDIT_DEFERRED", (action, before_lifecycle)
        # ``_request_edit_preload`` records the demand and establishes its
        # one owner, but deliberately does not project that CAS to widgets.
        # The stored lifecycle must therefore remain the exact deferred
        # snapshot until the test explicitly drives the real gate below.
        assert after_lifecycle == before_lifecycle, (
            action, before_lifecycle, after_lifecycle,
        )
    elif edit_policy == "first-demand-projection":
        before_lifecycle = tuple(before["lifecycle"])
        after_lifecycle = tuple(after["lifecycle"])
        assert len(before_lifecycle) == len(after_lifecycle) == 4, (
            action, before_lifecycle, after_lifecycle,
        )
        assert before_lifecycle[0] == "EDIT_DEFERRED", (action, before_lifecycle)
        assert after_lifecycle[0] == "EDIT_LOADING", (action, after_lifecycle)
        assert after_lifecycle[1:3] == before_lifecycle[1:3], (
            action, before_lifecycle, after_lifecycle,
        )
        assert after_lifecycle[3] == before_lifecycle[3] + 1, (
            action, before_lifecycle, after_lifecycle,
        )
    else:
        _assert_snapshot_section_unchanged(
            before["lifecycle"], after["lifecycle"], action, "guard_ui.lifecycle"
        )
    _assert_snapshot_section_unchanged(before["render"], after["render"], action, "guard_ui.render")
    _assert_edit_owner_policy(
        before["edit_owner"],
        after["edit_owner"],
        action,
        "unchanged" if edit_policy == "first-demand-projection" else edit_policy,
        expected_request,
    )
    if edit_policy == "first-demand":
        assert before["mutation_preload_marked"] is False
        assert after["mutation_preload_marked"] is True
    else:
        assert after["mutation_preload_marked"] == before["mutation_preload_marked"], (
            action, before["mutation_preload_marked"], after["mutation_preload_marked"],
        )

    if ui_activity_reason is None:
        _assert_snapshot_section_unchanged(
            before["selection"], after["selection"], action, "guard_ui.selection"
        )
        _assert_snapshot_section_unchanged(
            before["activity"], after["activity"], action, "guard_ui.activity"
        )
        return

    if expected_pair_idx is None:
        _assert_snapshot_section_unchanged(
            before["selection"], after["selection"], action, "guard_ui.selection"
        )
    else:
        assert after["selection"][0] == int(expected_pair_idx), (
            action, before["selection"], after["selection"], expected_pair_idx,
        )
    assert after["activity"][0] == before["activity"][0] + 1, (
        action, before["activity"], after["activity"],
    )
    assert after["activity"][1] == ui_activity_reason, (
        action, before["activity"], after["activity"],
    )


def _drain_current_sheet_scheduler(app, view):
    """Drain the one-Sheet app to semantic quiescence before each action."""
    deadline = time.monotonic() + 1.0
    last = None
    while time.monotonic() < deadline:
        view.root.update_idletasks()
        view.root.update()
        last = _scheduler_semantics_snapshot(app)
        if (
            not last["compute_queue"]
            and not last["compute_inflight"]
            and last["active_compute_sheet"] is None
            and not last["ui_tasks"]
        ):
            return last
        time.sleep(0.002)
    raise AssertionError(f"one-Sheet scheduler did not quiesce before guarded action: {last}")


@contextmanager
def _observe_action_scheduler(app, view):
    """Split synchronous action work from its following real Tk turn.

    The app's recurring heartbeat rearms an ``after`` callback independently of
    a button click.  It is meaningful to reject ``after`` work created during
    the synchronous action, but attributing a global heartbeat rearm from the
    following Tk turn to that action would be false evidence.  Compute/UI work
    remains fail-closed in both phases.
    """
    events = {
        "action": {
            "ui_tasks": [],
            "after": [],
            "compute_enqueues": [],
            "worker_kicks": [],
        },
        "turn": {
            "ui_tasks": [],
            "compute_enqueues": [],
            "worker_kicks": [],
            "after_observation": "disabled-global-rearm",
        },
    }
    phase = {"value": "action"}
    after_restorers = []
    compute_restorers = []

    def _callback_name(callback):
        return getattr(callback, "__qualname__", type(callback).__name__)

    def _install_after(owner, label):
        original_after = getattr(owner, "after")

        def _observed_after(delay_ms, callback=None, *args):
            events["action"]["after"].append(
                (label, int(delay_ms), _callback_name(callback))
            )
            return original_after(delay_ms, callback, *args)

        setattr(owner, "after", _observed_after)
        after_restorers.append(lambda: setattr(owner, "after", original_after))

    original_queue = getattr(app, "_queue_ui_task")
    original_safe_after = getattr(app, "_safe_root_after")
    original_enqueue = getattr(app, "_enqueue_sheet")
    original_kick = getattr(app, "_kick_worker")

    def _phase_events():
        return events[phase["value"]]

    def _observed_queue(task, *args, **kwargs):
        _phase_events()["ui_tasks"].append(_task_descriptor(task))
        return original_queue(task, *args, **kwargs)

    def _observed_safe_after(delay_ms, callback, *args, **kwargs):
        events["action"]["after"].append(
            ("app._safe_root_after", int(delay_ms), _callback_name(callback))
        )
        return original_safe_after(delay_ms, callback, *args, **kwargs)

    def _compute_context(sheet, *, front, exact_only_diff, force_recompute):
        sheet = str(sheet or "")
        queued_view = getattr(app, "sheet_views", {}).get(sheet)
        try:
            generation = int((getattr(app, "_sheet_compute_generation", {}) or {}).get(sheet, -1))
        except Exception:
            generation = -1
        try:
            exact_entry = dict((getattr(app, "_sheet_exact_entry")(sheet) or {}))
        except Exception as exc:
            exact_entry = {"capture_error": f"{type(exc).__name__}: {exc}"}
        return {
            "sheet": sheet,
            "front": bool(front),
            "exact_only_diff": None if exact_only_diff is None else bool(exact_only_diff),
            "force_recompute": bool(force_recompute),
            "selected_sheet": str(getattr(app, "selected_sheet", "") or ""),
            "generation": generation,
            "exact_entry": {
                key: exact_entry.get(key)
                for key in (
                    "state", "generation", "stage", "request_started_at",
                    "full_detail_terminal", "full_detail_terminal_at",
                )
            },
            "data_ready": None if queued_view is None else bool(getattr(queued_view, "_data_ready", False)),
        }

    def _observed_enqueue(sheet, *args, **kwargs):
        front = kwargs.get("front", args[0] if len(args) >= 1 else False)
        exact_only_diff = kwargs.get("exact_only_diff", args[1] if len(args) >= 2 else None)
        force_recompute = kwargs.get("force_recompute", args[2] if len(args) >= 3 else False)
        record = _compute_context(
            sheet,
            front=front,
            exact_only_diff=exact_only_diff,
            force_recompute=force_recompute,
        )
        _phase_events()["compute_enqueues"].append(record)
        raise AssertionError(f"guarded action attempted exact recompute enqueue: {record}")

    def _observed_kick(*args, **kwargs):
        try:
            with app._compute_lock:
                queue = tuple(str(sheet) for sheet in (getattr(app, "_compute_queue", ()) or ()))
                inflight = tuple(sorted(str(sheet) for sheet in (getattr(app, "_compute_inflight", set()) or set())))
        except Exception as exc:
            queue = (f"capture_error:{type(exc).__name__}:{exc}",)
            inflight = ()
        record = {
            "selected_sheet": str(getattr(app, "selected_sheet", "") or ""),
            "queue": queue,
            "inflight": inflight,
        }
        _phase_events()["worker_kicks"].append(record)
        raise AssertionError(f"guarded action attempted compute worker kick: {record}")

    setattr(app, "_queue_ui_task", _observed_queue)
    compute_restorers.append(lambda: setattr(app, "_queue_ui_task", original_queue))
    setattr(app, "_safe_root_after", _observed_safe_after)
    after_restorers.append(lambda: setattr(app, "_safe_root_after", original_safe_after))
    setattr(app, "_enqueue_sheet", _observed_enqueue)
    compute_restorers.append(lambda: setattr(app, "_enqueue_sheet", original_enqueue))
    setattr(app, "_kick_worker", _observed_kick)
    compute_restorers.append(lambda: setattr(app, "_kick_worker", original_kick))

    def _restore(restorers):
        while restorers:
            restorers.pop()()

    try:
        _install_after(view.root, "root.after")
        frame = getattr(view, "frame", None)
        if frame is not None and frame is not view.root:
            _install_after(frame, "view.frame.after")
        yield events
    except BaseException:
        try:
            _restore(after_restorers)
        finally:
            _restore(compute_restorers)
        raise
    else:
        try:
            # Keep compute/UI traps installed for exactly one turn, but restore
            # global after wrappers first so heartbeat rearm is not misattributed.
            _restore(after_restorers)
            phase["value"] = "turn"
            _pump_one_tk_turn(view.root)
        finally:
            _restore(compute_restorers)


def _assert_mutation_unchanged(
    before,
    app,
    view,
    source_paths,
    action: str,
    *,
    scheduler_before,
    scheduler_events,
    edit_policy: str = "unchanged",
    expected_request: dict | None = None,
    ui_activity_reason: str | None = None,
    expected_pair_idx: int | None = None,
):
    after = _mutation_snapshot(app, view, source_paths)
    _assert_snapshot_section_unchanged(
        before["hard_mutation"], after["hard_mutation"], action, "hard_mutation"
    )
    _assert_snapshot_section_unchanged(
        scheduler_before, after["scheduler_semantics"], action, "scheduler_semantics"
    )
    action_events = scheduler_events["action"]
    turn_events = scheduler_events["turn"]
    assert not action_events["ui_tasks"], (
        f"{action} synchronously queued UI work", action_events["ui_tasks"],
    )
    assert not action_events["after"], (
        f"{action} synchronously queued an after callback", action_events["after"],
    )
    assert not action_events["compute_enqueues"], (
        f"{action} synchronously queued an exact recompute", action_events["compute_enqueues"],
    )
    assert not action_events["worker_kicks"], (
        f"{action} synchronously kicked the compute worker", action_events["worker_kicks"],
    )
    assert not turn_events["ui_tasks"], (
        f"{action} queued UI work in its immediate Tk turn", turn_events["ui_tasks"],
    )
    assert not turn_events["compute_enqueues"], (
        f"{action} queued an exact recompute in its immediate Tk turn", turn_events["compute_enqueues"],
    )
    assert not turn_events["worker_kicks"], (
        f"{action} kicked the compute worker in its immediate Tk turn", turn_events["worker_kicks"],
    )
    _assert_guard_ui_policy(
        before["guard_ui"], after["guard_ui"], action,
        edit_policy=edit_policy,
        expected_request=expected_request,
        ui_activity_reason=ui_activity_reason,
        expected_pair_idx=expected_pair_idx,
    )
    return after


def _event_for_text_index(widget, index: str):
    widget.see(index)
    widget.update_idletasks()
    box = widget.bbox(index)
    assert box is not None, f"no visible bbox for {widget!r} at {index}"
    x, y, width, height = box
    return SimpleNamespace(
        x=int(x + max(1, width // 2)),
        y=int(y + max(1, height // 2)),
    )


def _prepare_direct_handler_events(view):
    pair_idx, diff_cols = next(
        (idx, set(cols))
        for idx, cols in sorted(view.pair_diff_cols.items())
        if cols
        and idx < len(view.row_pairs)
        and view.row_pairs[idx][0] is not None
        and view.row_pairs[idx][1] is not None
        and min(view.row_pairs[idx]) >= 3
    )
    logical_col = min(diff_cols)
    main_line = int(view.row_to_line[pair_idx])
    view._select_line(main_line)
    view._set_main_selected_cell(main_line, logical_col)
    view._cursor_cmp_sel_col = logical_col
    view._cursor_cmp_sel_line = 2
    view.hover_pair_idx = pair_idx
    view.hover_col_idx = logical_col
    view.hover_side = "B"
    view._last_cursor_cmp_pair_idx = pair_idx
    view._update_cursor_lines()
    _pump(view.root)

    row_header_event = _event_for_text_index(view.right_ln, f"{main_line}.0")
    line_text = view.cursor_cmp.get("2.0", "2.end")
    start, end = view._spans_for_line(line_text)[logical_col]
    char_pos = start + 1 if end - start > 1 else start
    comparison_event = _event_for_text_index(view.cursor_cmp, f"2.{char_pos}")
    return pair_idx, logical_col, main_line, row_header_event, comparison_event


@contextmanager
def _forced_lifecycle(view, state: str):
    app = view.app
    original = {
        "_data_ready": view._data_ready,
        "_row_model_exact": view._row_model_exact,
        "_cache_formula_aware": view._cache_formula_aware,
        "_lifecycle_error": view._lifecycle_error,
        "_lifecycle_canceled": view._lifecycle_canceled,
        "_suppress_bg_apply": view._suppress_bg_apply,
        "_edit_workbooks_ready": app._edit_workbooks_ready,
        "_is_sheet_exact_current": app._is_sheet_exact_current,
        "_edit_loading_started": app._edit_loading_started,
        "_edit_preload_thread": getattr(app, "_edit_preload_thread", None),
        "_edit_preload_active": app._edit_preload_active_event.is_set(),
        "_interactive_event_set": app._interactive_action_event.is_set(),
    }
    view.only_diff_var.set(0)
    view._last_only_diff_value = 0
    view._data_ready = True
    view._row_model_exact = True
    view._cache_formula_aware = True
    view._lifecycle_error = None
    view._lifecycle_canceled = False
    view._suppress_bg_apply = True
    # Exercise every non-ready state without loading a real editable workbook.
    app._edit_workbooks_ready = lambda: True

    if state == "LOADING":
        view._data_ready = False
    elif state == "DIFFING":
        view._row_model_exact = False
    elif state == "EDIT_LOADING":
        app._edit_workbooks_ready = lambda: False
        app._edit_loading_started = True
        app._edit_preload_thread = SimpleNamespace(is_alive=lambda: True)
        app._edit_preload_active_event.set()
    elif state == "EDIT_DEFERRED":
        app._edit_workbooks_ready = lambda: False
        app._edit_loading_started = False
        app._edit_preload_thread = None
        app._edit_preload_active_event.clear()
    elif state == "FAILED":
        view._lifecycle_error = "forced failure"
    elif state == "CANCELED":
        view._lifecycle_canceled = True
    elif state == "BUSY":
        app._interactive_action_event.set()
    else:
        raise AssertionError(f"unsupported forced lifecycle: {state}")

    if state in ("LOADING", "DIFFING", "FAILED", "CANCELED", "BUSY"):
        app._is_sheet_exact_current = lambda _sheet: False

    view._refresh_interaction_gate()
    assert view._lifecycle_state == state, (
        state,
        view._lifecycle_state,
        view._mutation_block_message(),
    )
    try:
        yield
    finally:
        view._data_ready = original["_data_ready"]
        view._row_model_exact = original["_row_model_exact"]
        view._cache_formula_aware = original["_cache_formula_aware"]
        view._lifecycle_error = original["_lifecycle_error"]
        view._lifecycle_canceled = original["_lifecycle_canceled"]
        view._suppress_bg_apply = original["_suppress_bg_apply"]
        app._edit_workbooks_ready = original["_edit_workbooks_ready"]
        app._is_sheet_exact_current = original["_is_sheet_exact_current"]
        app._edit_loading_started = original["_edit_loading_started"]
        app._edit_preload_thread = original["_edit_preload_thread"]
        if original["_edit_preload_active"]:
            app._edit_preload_active_event.set()
        else:
            app._edit_preload_active_event.clear()
        if original["_interactive_event_set"]:
            app._interactive_action_event.set()
        else:
            app._interactive_action_event.clear()
        view._refresh_interaction_gate()


def _widget_state(widget):
    return None if widget is None else str(widget.cget("state"))


def _assert_non_ready_control_policy(view, state: str) -> None:
    """Check the deliberate UI policy without mistaking it for the mutation gate."""
    expected_only_diff_state = (
        "normal" if view._is_exact_immutable_view_ready() else "disabled"
    )
    expected_by_forced_lifecycle = {
        "EDIT_DEFERRED": "normal",
        "EDIT_LOADING": "normal",
        "LOADING": "disabled",
        "DIFFING": "disabled",
        "FAILED": "disabled",
        "CANCELED": "disabled",
        "BUSY": "disabled",
    }
    if state in expected_by_forced_lifecycle:
        assert expected_only_diff_state == expected_by_forced_lifecycle[state], state
    assert _widget_state(getattr(view, "only_diff_cb", None)) == expected_only_diff_state, state
    assert _widget_state(getattr(view, "force_align_cb", None)) == "disabled", state

    assert getattr(view.app, "selected_sheet", None) == view.sheet
    assert _widget_state(getattr(view.app, "recalc_btn", None)) == "disabled", state

    # Non-closing operations remain clickable so their real public commands can
    # reach _guard_mutation_ready and present the shared readiness explanation.
    for name in (
        "use_left_btn", "use_right_btn", "use_left_menu_btn", "use_right_menu_btn",
        "use_base_btn", "undo_btn", "redo_btn", "save_a_btn", "save_b_btn", "three_way_cb",
    ):
        widget = getattr(view, name, None)
        if widget is not None:
            assert _widget_state(widget) == "normal", (name, state, _widget_state(widget))

    expected_rescan = "normal" if state in ("FAILED", "CANCELED") else "disabled"
    assert _widget_state(getattr(view, "manual_rescan_btn", None)) == expected_rescan, state

    selected = view._selected_column_block() is not None
    expected_structural = "normal" if selected else "disabled"
    for name in ("use_mine_col_btn", "use_theirs_col_btn"):
        widget = getattr(view, name, None)
        if widget is not None:
            assert _widget_state(widget) == expected_structural, (name, state, selected)
    base_button = getattr(view, "use_base_col_btn", None)
    if base_button is not None:
        expected_base = expected_structural if view._is_three_way_enabled() else "disabled"
        assert _widget_state(base_button) == expected_base, (state, selected)


@contextmanager
def _readiness_modal_and_edit_traps(app):
    """Keep real public commands observable while making accidental edit I/O fatal."""
    original = {
        name: getattr(app, name)
        for name in (
            "_show_exact_readiness_modal", "_request_edit_preload",
            "_start_background_thread",
            "_ensure_edit_loaded", "_load_edit_workbooks_owned",
            "ws_a_val", "ws_b_val", "ws_base_val",
            "ws_a_edit", "ws_b_edit", "ws_base_edit",
            "_atomic_save", "_atomic_save_with_retry", "_atomic_replace_file_with_retry",
            "build_manual_merge_output_file", "build_manual_b_output_file", "_try_alt_save",
        )
        if hasattr(app, name)
    }
    original_showerror = smt.messagebox.showerror
    evidence = {
        "modals": [],
        "preload_calls": [],
        "owner_starts": [],
        "loader_calls": [],
        "native_save_calls": [],
        "error_dialogs": [],
    }

    def _modal(action, sheets):
        normalized = sheets if isinstance(sheets, (tuple, list, set)) else (sheets,)
        sheet_names = tuple(str(item) for item in normalized)
        details = []
        for sheet in sheet_names:
            view = getattr(app, "sheet_views", {}).get(sheet)
            entry = getattr(app, "_sheet_exact_entry", lambda _sheet: {})(sheet) or {}
            details.append(
                {
                    "sheet": sheet,
                    "state": None if view is None else str(view._derive_lifecycle_state()),
                    "stage": str(entry.get("stage") or ""),
                    "retry": bool(getattr(view, "_last_mutation_started_edit_preload", False)),
                }
            )
        evidence["modals"].append({"action": str(action), "sheets": sheet_names, "details": tuple(details)})

    def _preload(*args, **kwargs):
        evidence["preload_calls"].append((tuple(args), dict(kwargs)))
        # Keep production request accounting and its `loading_started` early
        # return. Only the actual worker creation is replaced below.
        return original["_request_edit_preload"](*args, **kwargs)

    class _FakePreloadOwner:
        def __init__(self, name):
            self.name = str(name)

        def is_alive(self):
            return True

    def _start_background_thread(target, *, name: str):
        owner = _FakePreloadOwner(name)
        evidence["owner_starts"].append(
            {"name": str(name), "target": getattr(target, "__qualname__", type(target).__name__), "owner": owner}
        )
        # Do not run the target: this is the case-scope boundary that prevents
        # workbook parsing while still exercising the production owner CAS.
        return owner

    def _forbid_loader(label):
        def _blocked(*args, **kwargs):
            evidence["loader_calls"].append((label, tuple(args), dict(kwargs)))
            raise AssertionError(f"blocked action reached {label}")
        return _blocked

    def _forbid_save(label):
        def _blocked(*args, **kwargs):
            evidence["native_save_calls"].append((label, tuple(args), dict(kwargs)))
            raise AssertionError(f"blocked action reached native save {label}")
        return _blocked

    try:
        app._show_exact_readiness_modal = _modal
        app._request_edit_preload = _preload
        app._start_background_thread = _start_background_thread
        for name in (
            "_ensure_edit_loaded", "_load_edit_workbooks_owned",
            "ws_a_val", "ws_b_val", "ws_base_val",
            "ws_a_edit", "ws_b_edit", "ws_base_edit",
        ):
            if name in original:
                setattr(app, name, _forbid_loader(name))
        for name in (
            "_atomic_save", "_atomic_save_with_retry", "_atomic_replace_file_with_retry",
            "build_manual_merge_output_file", "build_manual_b_output_file", "_try_alt_save",
        ):
            if name in original:
                setattr(app, name, _forbid_save(name))
        smt.messagebox.showerror = lambda *args, **kwargs: evidence["error_dialogs"].append((args, kwargs))
        yield evidence
    finally:
        for name, value in original.items():
            setattr(app, name, value)
        smt.messagebox.showerror = original_showerror


def _assert_modal_context(
    evidence,
    previous_count: int,
    *,
    action: str,
    state: str,
    stage: str,
    retry: bool,
) -> None:
    assert len(evidence["modals"]) == previous_count + 1, evidence["modals"][previous_count:]
    observed = evidence["modals"][-1]
    assert observed["action"] == action, observed
    assert observed["sheets"] == ("Data",), observed
    assert observed["details"] == (
        {"sheet": "Data", "state": state, "stage": stage, "retry": retry},
    ), observed


def _assert_preload_call_count(evidence, previous_count: int, *, expected_request: dict | None) -> None:
    expected_count = 0 if expected_request is None else 1
    observed = evidence["preload_calls"][previous_count:]
    assert len(observed) == expected_count, observed
    if expected_request is None:
        return
    args, kwargs = observed[0]
    assert args == (), observed[0]
    assert kwargs == {
        "reason": expected_request["request_reason"],
        "caller": expected_request["request_caller"],
    }, observed[0]


def test_non_ready_direct_handlers_have_no_write_side_effects(case):
    app, view, source_paths = _open_ready_view(case)
    try:
        pair_idx, logical_col, main_line, row_header_event, comparison_event = _prepare_direct_handler_events(view)
        assert set(view.pair_diff_cols.get(pair_idx, set())) and logical_col in view.pair_diff_cols[pair_idx]
        # The event intentionally has no hover-arrow precondition. It is a real
        # selection-only path, not false evidence that a row mutation was gated.
        assert getattr(view, "_hover_ln_line_right", None) != main_line

        # Sentinels prove the public Undo/Redo invokes reach their common guard
        # rather than no-op because both stacks happen to be empty.
        app.undo_stack.append({"kind": "gate-test-sentinel", "sheet": "Data"})
        app.redo_stack.append({"kind": "gate-test-redo-sentinel", "sheet": "Data", "target": "B2A"})
        actions = (
            ("toolbar row action", lambda: view._run_copy_action_by_mode("B2A")),
            (
                "direct row primitive",
                lambda: view._copy_selected_row(
                    "B2A",
                    override_pair_idx=pair_idx,
                    override_cols={logical_col},
                ),
            ),
            ("direct region handler", lambda: view._copy_selected_region("B2A")),
            (
                "direct cell handler",
                lambda: view._copy_single_cell_by_pair(pair_idx, "B2A", logical_col),
            ),
            (
                "toolbar use-left invoke",
                view.use_left_btn.invoke,
            ),
            (
                "toolbar use-right invoke",
                view.use_right_btn.invoke,
            ),
            (
                "toolbar undo invoke",
                view.undo_btn.invoke,
            ),
            (
                "toolbar redo invoke",
                view.redo_btn.invoke,
            ),
            (
                "forged guarded cell primitive",
                lambda: view._copy_single_cell_by_pair(
                    pair_idx,
                    "B2A",
                    logical_col,
                    _guarded=True,
                ),
            ),
            (
                "forged batch row primitive",
                lambda: view._copy_selected_row(
                    "B2A",
                    override_pair_idx=pair_idx,
                    override_cols={logical_col},
                    suppress_refresh=True,
                ),
            ),
            (
                "row-header selector without arrow",
                lambda: view._on_row_header_click(
                    view.right_ln,
                    row_header_event,
                    "B2A",
                ),
            ),
            (
                "C-area double-click handler",
                lambda: view._on_cursor_cmp_double_click(comparison_event),
            ),
            ("column action handler", lambda: view._on_column_action_button("B")),
        )
        public_action_expectations = {
            "toolbar row action": {
                "modal_action": "采用所选内容",
                "request_reason": "mutation:采用所选内容",
                "request_caller": "SheetView._guard_mutation_ready",
            },
            "direct row primitive": {
                "modal_action": "行覆盖",
                "request_reason": "mutation:行覆盖",
                "request_caller": "SheetView._guard_mutation_ready",
            },
            "direct region handler": {
                "modal_action": "区域覆盖",
                "request_reason": "mutation:区域覆盖",
                "request_caller": "SheetView._guard_mutation_ready",
            },
            "direct cell handler": {
                "modal_action": "单元格覆盖",
                "request_reason": "mutation:单元格覆盖",
                "request_caller": "SheetView._guard_mutation_ready",
            },
            "toolbar use-left invoke": {
                "modal_action": "采用所选内容",
                "request_reason": "mutation:采用所选内容",
                "request_caller": "SheetView._guard_mutation_ready",
            },
            "toolbar use-right invoke": {
                "modal_action": "采用所选内容",
                "request_reason": "mutation:采用所选内容",
                "request_caller": "SheetView._guard_mutation_ready",
            },
            "toolbar undo invoke": {
                "modal_action": "撤销",
                "request_reason": "mutation:撤销",
                "request_caller": "SheetView._guard_mutation_ready",
            },
            "toolbar redo invoke": {
                "modal_action": "重做",
                "request_reason": "mutation:重做",
                "request_caller": "SheetView._guard_mutation_ready",
            },
            "C-area double-click handler": {
                "modal_action": "单元格覆盖",
                "request_reason": "mutation:单元格覆盖",
                "request_caller": "SheetView._guard_mutation_ready",
            },
            "column action handler": {
                "modal_action": "列结构操作",
                "request_reason": "mutation:列结构操作",
                "request_caller": "SheetView._guard_mutation_ready",
            },
        }
        for name in ("save_a_btn", "save_b_btn"):
            button = getattr(view, name, None)
            if button is not None:
                action_name = f"{name} invoke"
                actions += ((action_name, button.invoke),)
                modal_action = "保存 A" if name == "save_a_btn" else "保存 B"
                public_action_expectations[action_name] = {
                    "modal_action": modal_action,
                    "request_reason": f"save:{modal_action}",
                    "request_caller": "_guard_save_readiness",
                }
        selector_activity = {
            "row-header selector without arrow": ("row-header-click", pair_idx),
            "C-area double-click handler": ("c-area-double-click", None),
        }

        with _readiness_modal_and_edit_traps(app) as evidence:
            for state in (
                "LOADING",
                "DIFFING",
                "EDIT_LOADING",
                "FAILED",
                "CANCELED",
                "BUSY",
            ):
                with _forced_lifecycle(view, state):
                    _assert_non_ready_control_policy(view, state)
                    owner_starts_before = len(evidence["owner_starts"])
                    for action_name, action in actions:
                        assert view._derive_lifecycle_state() == state, (
                            state,
                            action_name,
                            view._derive_lifecycle_state(),
                        )
                        scheduler_before = _drain_current_sheet_scheduler(app, view)
                        before = _mutation_snapshot(app, view, source_paths)
                        _assert_snapshot_section_unchanged(
                            scheduler_before,
                            before["scheduler_semantics"],
                            f"{state}: {action_name}",
                            "pre-action scheduler_semantics",
                        )
                        modals_before = len(evidence["modals"])
                        preload_before = len(evidence["preload_calls"])
                        modal_stage = str((app._sheet_exact_entry("Data") or {}).get("stage") or "")
                        with _observe_action_scheduler(app, view) as scheduler_events:
                            action()
                        expected_request = public_action_expectations.get(action_name)
                        # Verify the public guard evidence before checking data
                        # invariants, so a future failure cannot hide a wrong
                        # action/sheet/state/stage/retry modal.
                        if expected_request is None:
                            assert len(evidence["modals"]) == modals_before, (
                                state, action_name, evidence["modals"][modals_before:],
                            )
                        else:
                            _assert_modal_context(
                                evidence,
                                modals_before,
                                action=expected_request["modal_action"],
                                state=state,
                                stage=modal_stage,
                                retry=False,
                            )
                        expected_preload_request = (
                            expected_request if state == "EDIT_LOADING" else None
                        )
                        _assert_preload_call_count(
                            evidence, preload_before, expected_request=expected_preload_request
                        )
                        _assert_mutation_unchanged(
                            before,
                            app,
                            view,
                            source_paths,
                            f"{state}: {action_name}",
                            scheduler_before=scheduler_before,
                            scheduler_events=scheduler_events,
                            edit_policy=(
                                "existing-owner-audit"
                                if state == "EDIT_LOADING" and expected_request is not None
                                else "unchanged"
                            ),
                            expected_request=expected_request,
                            ui_activity_reason=selector_activity.get(action_name, (None, None))[0],
                            expected_pair_idx=selector_activity.get(action_name, (None, None))[1],
                        )
                    assert len(evidence["owner_starts"]) == owner_starts_before, (
                        state, evidence["owner_starts"][owner_starts_before:],
                    )

            # The first demand path is distinct from the already-loading
            # retry path above. It must execute production request accounting
            # and the single-owner CAS exactly once without parsing a workbook.
            with _forced_lifecycle(view, "EDIT_DEFERRED"):
                _assert_non_ready_control_policy(view, "EDIT_DEFERRED")
                scheduler_before = _drain_current_sheet_scheduler(app, view)
                before = _mutation_snapshot(app, view, source_paths)
                _assert_snapshot_section_unchanged(
                    scheduler_before,
                    before["scheduler_semantics"],
                    "deferred first demand",
                    "pre-action scheduler_semantics",
                )
                modal_stage = str((app._sheet_exact_entry("Data") or {}).get("stage") or "")
                modals_before = len(evidence["modals"])
                preloads_before = len(evidence["preload_calls"])
                starts_before = len(evidence["owner_starts"])
                deferred_first_request = public_action_expectations["toolbar use-right invoke"]
                with _observe_action_scheduler(app, view) as scheduler_events:
                    view.use_right_btn.invoke()
                _assert_modal_context(
                    evidence,
                    modals_before,
                    action=deferred_first_request["modal_action"],
                    state="EDIT_LOADING",
                    stage=modal_stage,
                    retry=True,
                )
                _assert_preload_call_count(
                    evidence, preloads_before, expected_request=deferred_first_request
                )
                first_after = _assert_mutation_unchanged(
                    before,
                    app,
                    view,
                    source_paths,
                    "deferred first demand",
                    scheduler_before=scheduler_before,
                    scheduler_events=scheduler_events,
                    edit_policy="first-demand",
                    expected_request=deferred_first_request,
                )
                assert len(evidence["owner_starts"]) == starts_before + 1, evidence["owner_starts"]
                first_owner = app._edit_preload_thread
                assert first_owner is evidence["owner_starts"][-1]["owner"]
                assert first_after["guard_ui"]["lifecycle"] == before["guard_ui"]["lifecycle"], (
                    first_after["guard_ui"]["lifecycle"], before["guard_ui"]["lifecycle"],
                )
                assert first_after["guard_ui"]["lifecycle"][0] == "EDIT_DEFERRED", first_after
                assert view._derive_lifecycle_state() == "EDIT_LOADING"
                first_entry = app._sheet_exact_entry("Data") or {}
                assert app._is_sheet_exact_current("Data"), first_entry
                assert str(first_entry.get("stage") or "") == modal_stage, (
                    modal_stage, first_entry,
                )
                # The real modal clears its one-shot first-demand annotation;
                # this record-only fake intentionally leaves UI state alone.
                view._last_mutation_started_edit_preload = False
                projection_scheduler_before = _drain_current_sheet_scheduler(app, view)
                projection_before = _mutation_snapshot(app, view, source_paths)
                _assert_snapshot_section_unchanged(
                    projection_scheduler_before,
                    projection_before["scheduler_semantics"],
                    "deferred first-demand projection",
                    "pre-action scheduler_semantics",
                )
                with _observe_action_scheduler(app, view) as projection_scheduler_events:
                    # Exercise the actual production projection; do not rely
                    # on a fake-thread poll or an arbitrary wait to change the
                    # stored lifecycle after the one-owner preload CAS.
                    view._refresh_interaction_gate()
                post_first = _assert_mutation_unchanged(
                    projection_before,
                    app,
                    view,
                    source_paths,
                    "deferred first-demand projection",
                    scheduler_before=projection_scheduler_before,
                    scheduler_events=projection_scheduler_events,
                    edit_policy="first-demand-projection",
                    expected_request=None,
                )
                projection_entry = app._sheet_exact_entry("Data") or {}
                assert app._is_sheet_exact_current("Data"), projection_entry
                assert str(projection_entry.get("stage") or "") == modal_stage, (
                    modal_stage, projection_entry,
                )
                assert post_first["guard_ui"]["edit_owner"]["thread_identity"] == id(first_owner), post_first
                assert len(post_first["guard_ui"]["edit_owner"]["requests"]) == (
                    len(before["guard_ui"]["edit_owner"]["requests"]) + 1
                ), post_first
                repeat_baseline = post_first

                repeat_actions = [
                    (
                        "deferred repeat use-left",
                        view.use_left_btn.invoke,
                        public_action_expectations["toolbar use-left invoke"],
                    ),
                ]
                save_button = getattr(view, "save_a_btn", None)
                if save_button is not None:
                    repeat_actions.append((
                        "deferred repeat save-a",
                        save_button.invoke,
                        public_action_expectations["save_a_btn invoke"],
                    ))
                for action_name, action, expected_request in repeat_actions:
                    scheduler_before = _drain_current_sheet_scheduler(app, view)
                    before_repeat = _mutation_snapshot(app, view, source_paths)
                    _assert_snapshot_section_unchanged(
                        repeat_baseline["hard_mutation"], before_repeat["hard_mutation"],
                        action_name, "post-first hard_mutation baseline",
                    )
                    assert before_repeat["guard_ui"]["lifecycle"] == (
                        repeat_baseline["guard_ui"]["lifecycle"]
                    ), (action_name, repeat_baseline["guard_ui"], before_repeat["guard_ui"])
                    assert before_repeat["guard_ui"]["edit_owner"] == (
                        repeat_baseline["guard_ui"]["edit_owner"]
                    ), (action_name, repeat_baseline["guard_ui"], before_repeat["guard_ui"])
                    _assert_snapshot_section_unchanged(
                        scheduler_before,
                        before_repeat["scheduler_semantics"],
                        action_name,
                        "pre-action scheduler_semantics",
                    )
                    modals_before = len(evidence["modals"])
                    preloads_before = len(evidence["preload_calls"])
                    starts_before = len(evidence["owner_starts"])
                    modal_stage = str((app._sheet_exact_entry("Data") or {}).get("stage") or "")
                    with _observe_action_scheduler(app, view) as scheduler_events:
                        action()
                    _assert_modal_context(
                        evidence,
                        modals_before,
                        action=expected_request["modal_action"],
                        state="EDIT_LOADING",
                        stage=modal_stage,
                        retry=False,
                    )
                    _assert_preload_call_count(
                        evidence, preloads_before, expected_request=expected_request
                    )
                    after_repeat = _assert_mutation_unchanged(
                        before_repeat,
                        app,
                        view,
                        source_paths,
                        action_name,
                        scheduler_before=scheduler_before,
                        scheduler_events=scheduler_events,
                        edit_policy="existing-owner-audit",
                        expected_request=expected_request,
                    )
                    repeat_entry = app._sheet_exact_entry("Data") or {}
                    assert app._is_sheet_exact_current("Data"), repeat_entry
                    assert str(repeat_entry.get("stage") or "") == modal_stage, (
                        action_name, modal_stage, repeat_entry,
                    )
                    assert after_repeat["guard_ui"]["lifecycle"] == (
                        repeat_baseline["guard_ui"]["lifecycle"]
                    ), (action_name, repeat_baseline["guard_ui"], after_repeat["guard_ui"])
                    assert after_repeat["guard_ui"]["edit_owner"]["thread_identity"] == id(first_owner), (
                        action_name, after_repeat["guard_ui"]["edit_owner"], id(first_owner),
                    )
                    assert app._edit_preload_thread is first_owner
                    assert len(evidence["owner_starts"]) == starts_before, evidence["owner_starts"]
                    assert view._derive_lifecycle_state() == "EDIT_LOADING"
                    repeat_baseline = after_repeat

            assert not evidence["loader_calls"], evidence["loader_calls"]
            assert not evidence["native_save_calls"], evidence["native_save_calls"]
            assert not evidence["error_dialogs"], evidence["error_dialogs"]

        # The B3 case stops at the non-ready zero-write boundary; accepted
        # actions and editable-workbook materialization belong to operation
        # coverage, not this view-only gate.
        app.undo_stack[:] = [
            action for action in app.undo_stack
            if action.get("kind") != "gate-test-sentinel"
        ]
        app.redo_stack[:] = [
            action for action in app.redo_stack
            if action.get("kind") != "gate-test-redo-sentinel"
        ]
        assert not app.manual_a_cell_ops
        assert not app.manual_b_cell_ops
        assert not app.modified_a
        assert not app.modified_b
    finally:
        case.close_app(app)


def test_only_diff_pending_stays_checked_locked_and_keeps_stable_view(case):
    app, view, source_paths = _open_ready_view(case)
    original_start = view._start_async_large_only_diff_build
    original_cache_from_maps = view._cache_only_diff_rows_from_exact_pair_maps
    try:
        with _readiness_modal_and_edit_traps(app) as evidence:
            view._invalidate_only_diff_snapshot_cache()
            # The exact full view is intentionally EDIT_DEFERRED: only-diff
            # must still be a public immutable-view action, never an edit-load
            # demand.  Only its dedicated exact row snapshot is absent here.
            view._pair_diff_full_exact = True
            view._only_diff_rows_exact = False
            view._refresh_interaction_gate()
            assert app._is_sheet_exact_current("Data")
            assert view._lifecycle_state == "EDIT_DEFERRED", view._lifecycle_state
            assert str(view.only_diff_cb.cget("state")) == "normal"

            stable = (
                tuple(view.display_rows),
                view.left.get("1.0", "end-1c"),
                view.right.get("1.0", "end-1c"),
            )
            starts = []

            def _progress_visible(build_seq: int) -> bool:
                token = (view, int(build_seq))
                win = getattr(app, "_only_diff_progress_win", None)
                cancel = getattr(app, "_only_diff_progress_cancel_btn", None)
                try:
                    return bool(
                        app._only_diff_progress_owner == token
                        and app._only_diff_progress_visible_token == token
                        and win is not None
                        and str(win.state()) == "normal"
                        and bool(win.winfo_ismapped())
                        and bool(win.winfo_viewable())
                        and win.grab_current() == win
                        and cancel is not None
                        and str(cancel.cget("state")) == "normal"
                    )
                except Exception:
                    return False

            def _start_pending_once(*, user_initiated=False):
                assert user_initiated is True
                view._only_diff_async_build_seq += 1
                build_seq = int(view._only_diff_async_build_seq)
                build_key = view._current_only_diff_cache_key()
                current_generation = int(app._sheet_compute_generation[view.sheet])
                prior_terminal = copy.deepcopy(dict(app._sheet_exact_entry(view.sheet) or {}))
                assert app._is_sheet_exact_current(view.sheet), prior_terminal
                assert int(prior_terminal.get("generation", -1)) == current_generation
                assert bool(view._begin_only_diff_exact_transition(build_seq, build_key))
                current = dict(app._sheet_exact_entry(view.sheet) or {})
                prior_record = getattr(view, "_only_diff_async_prior_exact", None)
                assert int(current.get("generation", -1)) == current_generation, current
                assert str(current.get("state") or "") == "CALCULATING", current
                assert isinstance(prior_record, dict), prior_record
                assert int(prior_record.get("build_seq", -1)) == build_seq, prior_record
                assert int(prior_record.get("generation", -1)) == current_generation, prior_record
                assert dict(prior_record.get("entry") or {}) == prior_terminal, prior_record
                assert prior_record.get("entry") is not prior_terminal, prior_record
                assert bool(view._only_diff_async_building)
                assert view._only_diff_async_build_key == build_key
                assert bool(view._only_diff_preview_full)
                assert not bool(view._pending_exact_render)
                assert int(view._only_diff_async_requested_value) == 1
                starts.append(
                    {
                        "build_seq": build_seq,
                        "build_key": build_key,
                        "generation": current_generation,
                        "prior_terminal": prior_terminal,
                    }
                )
                app._begin_only_diff_progress(view, build_seq)
                return True

            view._start_async_large_only_diff_build = _start_pending_once
            view._cache_only_diff_rows_from_exact_pair_maps = lambda: False
            # Exercise the real Checkbutton command; never pre-set its IntVar
            # or call a private production entry to manufacture a request.
            hard_before_pending = _mutation_snapshot(app, view, source_paths)
            view.only_diff_cb.invoke()
            _wait_until(
                app.root,
                lambda: bool(starts) and _progress_visible(starts[0]["build_seq"]),
                "only-diff progress was not visible within 500ms",
                timeout=0.5,
            )
            pending_snapshot = _mutation_snapshot(app, view, source_paths)
            _assert_snapshot_section_unchanged(
                hard_before_pending["hard_mutation"],
                pending_snapshot["hard_mutation"],
                "only-diff public pending transition",
                "hard_mutation",
            )
            build_seq = int(starts[0]["build_seq"])
            pending_entry = dict(app._sheet_exact_entry(view.sheet) or {})
            assert int(pending_entry.get("generation", -1)) == int(starts[0]["generation"])
            assert str(pending_entry.get("state") or "") == "CALCULATING", pending_entry
            assert app._only_diff_progress_owner == (view, build_seq)
            assert app._only_diff_progress_visible_token == (view, build_seq)

            assert int(view.only_diff_var.get()) == 1
            # The stable preference is committed only after exact publication;
            # cancellation must still know it should return to full view.
            assert view._last_only_diff_value == 0
            assert view._lifecycle_state == "DIFFING", view._lifecycle_state
            assert str(view.only_diff_cb.cget("state")) == "disabled"
            assert str(view.only_diff_cb.cget("text")) == "只看差异内容"
            assert "正在后台生成精确差异行" in str(view.info.cget("text"))
            assert starts and len(starts) == 1
            assert (
                tuple(view.display_rows),
                view.left.get("1.0", "end-1c"),
                view.right.get("1.0", "end-1c"),
            ) == stable
            assert not evidence["preload_calls"], evidence["preload_calls"]
            assert not evidence["owner_starts"], evidence["owner_starts"]
            assert not evidence["loader_calls"], evidence["loader_calls"]
            assert not evidence["native_save_calls"], evidence["native_save_calls"]
            assert not evidence["modals"], evidence["modals"]
            assert not app.manual_a_cell_ops and not app.manual_b_cell_ops
            assert not app.modified_a and not app.modified_b

            # A disabled public Checkbutton cannot start another request. Also
            # exercise the handler's queued-event branch, which must preserve
            # the current request rather than cancel or duplicate the builder.
            view.only_diff_cb.invoke()
            view._toggle_only_diff()
            _pump(app.root)
            assert int(view.only_diff_var.get()) == 1
            assert view._lifecycle_state == "DIFFING"
            assert len(starts) == 1
            assert view._only_diff_async_building is True
            assert _progress_visible(build_seq)
            repeat_pending_snapshot = _mutation_snapshot(app, view, source_paths)
            _assert_snapshot_section_unchanged(
                pending_snapshot["hard_mutation"],
                repeat_pending_snapshot["hard_mutation"],
                "only-diff repeated pending request",
                "hard_mutation",
            )
            assert not evidence["preload_calls"], evidence["preload_calls"]
            assert not evidence["loader_calls"], evidence["loader_calls"]
            assert not evidence["native_save_calls"], evidence["native_save_calls"]
            assert not app.manual_a_cell_ops and not app.manual_b_cell_ops
            assert not app.modified_a and not app.modified_b

            exact_rows = sorted(
                pair_idx
                for pair_idx in range(len(view.row_pairs))
                if view._pair_has_visual_diff(pair_idx)
            )
            assert not view._is_three_way_enabled()
            assert not bool(getattr(app, "has_base", False))
            pair_domain = tuple(range(len(view.row_pairs)))
            fixture_pair_diff = {3: {2}, 4: {2}}
            expected_prior_pair_diff = {
                pair_idx: set(fixture_pair_diff.get(pair_idx, set()))
                for pair_idx in pair_domain
            }
            assert exact_rows == sorted(fixture_pair_diff), exact_rows
            assert {
                int(pair_idx): set(cols)
                for pair_idx, cols in view.pair_diff_cols.items()
            } == expected_prior_pair_diff
            base_diff_before = {
                int(pair_idx): set(cols)
                for pair_idx, cols in view.pair_base_diff_cols.items()
            }
            assert set(base_diff_before) == set(pair_domain), base_diff_before
            assert all(not cols for cols in base_diff_before.values()), base_diff_before
            base_raw_before = copy.deepcopy(view.pair_raw_parts_base)
            assert base_raw_before == {}, base_raw_before
            assert view._base_diff_full_exact is True
            # Complete through the extracted production worker-result entry.
            # Do not manufacture a visible result by calling the legacy
            # worksheet-aware refresh path: this B3 case keeps every value,
            # edit, loader, save, alignment, and comparison sentinel armed.
            build_seq = int(starts[0]["build_seq"])
            build_key = starts[0]["build_key"]
            hard_before_completion = _mutation_snapshot(app, view, source_paths)
            assert not view._virtual_mode_active()
            assert app._only_diff_progress_owner == (view, build_seq)
            assert int(view._only_diff_async_build_seq) == build_seq
            assert view._only_diff_async_build_key == build_key
            payload = {
                "build_seq": build_seq,
                "build_key": build_key,
                "sheet": view.sheet,
                "diff_pair_indices": list(exact_rows),
                "pair_diff_cols": {
                    pair_idx: set(cols)
                    for pair_idx, cols in fixture_pair_diff.items()
                },
                "pair_base_diff_cols": {},
                "pair_parts_a": {
                    pair_idx: view.pair_raw_parts_a[pair_idx]
                    for pair_idx in exact_rows
                    if pair_idx in view.pair_raw_parts_a
                },
                "pair_parts_b": {
                    pair_idx: view.pair_raw_parts_b[pair_idx]
                    for pair_idx in exact_rows
                    if pair_idx in view.pair_raw_parts_b
                },
                "pair_parts_base": {},
            }
            expected_pair_diff = {
                int(pair_idx): set(cols)
                for pair_idx, cols in payload["pair_diff_cols"].items()
                if cols
            }
            assert expected_pair_diff == fixture_pair_diff
            original_refresh = view.refresh
            original_mode_refresh = view._refresh_mode_switch_preserving_selection
            original_align = smt._align_selected_sheet_snapshots
            original_compare = smt._compare_selected_sheet_snapshots
            original_publish = view._publish_prepared_cache_surface
            original_append_rows = view._append_rows
            publisher_calls = []
            append_calls = []

            def _forbidden_view_only(*_args, **_kwargs):
                raise AssertionError("only-diff completion entered legacy worksheet refresh")

            def _forbidden_snapshot_work(*_args, **_kwargs):
                raise AssertionError("only-diff completion recomputed immutable comparison")

            def _record_cache_publish(*args, **kwargs):
                record = {
                    "prepared_rows": tuple(
                        int(pair_idx)
                        for pair_idx in (kwargs.get("prepared_rows") or ())
                    ),
                }
                publisher_calls.append(record)
                result = original_publish(*args, **kwargs)
                record["result"] = result
                return result

            def _record_append_rows(*args, **kwargs):
                bound = inspect.signature(original_append_rows).bind(*args, **kwargs)
                bound.apply_defaults()
                record = {
                    "rows": tuple(
                        int(pair_idx)
                        for pair_idx in (bound.arguments.get("new_rows") or ())
                    ),
                    "refresh_block_ui": bool(bound.arguments["refresh_block_ui"]),
                    "returned": False,
                }
                append_calls.append(record)
                result = original_append_rows(*args, **kwargs)
                record["returned"] = True
                record["result"] = result
                return result

            try:
                view.refresh = _forbidden_view_only
                view._refresh_mode_switch_preserving_selection = _forbidden_view_only
                smt._align_selected_sheet_snapshots = _forbidden_snapshot_work
                smt._compare_selected_sheet_snapshots = _forbidden_snapshot_work
                view._publish_prepared_cache_surface = _record_cache_publish
                view._append_rows = _record_append_rows
                result = view._apply_async_only_diff_result(
                    payload,
                    build_seq=build_seq,
                    has_base=False,
                )
                assert result == "publish-scheduled", result
                _wait_until(
                    app.root,
                    lambda: (
                        not bool(getattr(view, "_mode_switch_pending", False))
                        and getattr(app, "_only_diff_progress_owner", None) is None
                    ),
                    "only-diff immutable cache publisher did not reach one terminal publication",
                    timeout=3.0,
                )
            finally:
                view.refresh = original_refresh
                view._refresh_mode_switch_preserving_selection = original_mode_refresh
                smt._align_selected_sheet_snapshots = original_align
                smt._compare_selected_sheet_snapshots = original_compare
                view._publish_prepared_cache_surface = original_publish
                view._append_rows = original_append_rows

            assert len(publisher_calls) == 1, publisher_calls
            assert publisher_calls[0]["result"] is True, publisher_calls
            assert publisher_calls[0]["prepared_rows"] == tuple(exact_rows), publisher_calls
            assert len(append_calls) == 1, append_calls
            assert append_calls[0]["rows"] == tuple(exact_rows), append_calls
            assert append_calls[0]["refresh_block_ui"] is False, append_calls
            assert append_calls[0]["returned"] is True, append_calls
            assert not view._virtual_mode_active()
            assert {
                int(pair_idx): set(cols)
                for pair_idx, cols in view.pair_diff_cols.items()
            } == expected_pair_diff
            assert set(view.pair_diff_cols) == set(exact_rows)
            assert view._pair_diff_full_exact is True
            assert {
                int(pair_idx): set(cols)
                for pair_idx, cols in view.pair_base_diff_cols.items()
            } == base_diff_before
            assert all(not cols for cols in view.pair_base_diff_cols.values())
            assert view.pair_raw_parts_base == base_raw_before
            assert view._base_diff_full_exact is False
            hard_after_completion = _mutation_snapshot(app, view, source_paths)
            _assert_only_diff_cache_publish_hard_invariants(
                hard_before_completion["hard_mutation"],
                hard_after_completion["hard_mutation"],
                action="only-diff immutable cache completion",
            )
            terminal_entry = dict(app._sheet_exact_entry(view.sheet) or {})
            assert app._is_sheet_exact_current("Data"), terminal_entry
            assert int(terminal_entry.get("generation", -1)) == int(starts[0]["generation"])
            assert bool(terminal_entry.get("full_detail_terminal", False)), terminal_entry
            assert getattr(view, "_only_diff_async_prior_exact", None) is None
            assert app._only_diff_progress_owner is None
            assert app._only_diff_progress_visible_token is None
            assert app._only_diff_progress_show_after_id is None
            assert app._only_diff_progress_watchdog_after_id is None
            assert app._only_diff_progress_confirm_after_id is None
            assert not bool(view._only_diff_async_building)
            assert view._only_diff_async_build_key is None
            assert not bool(view._only_diff_preview_full)
            assert not bool(view._pending_exact_render)
            assert tuple(view._only_diff_rows_cache or ()) == tuple(exact_rows)
            assert view._only_diff_rows_cache_key == view._current_only_diff_cache_key()
            assert view._has_valid_only_diff_snapshot_cache()

            assert app._is_sheet_exact_current("Data")
            assert int(view.only_diff_var.get()) == 1
            assert str(view.only_diff_cb.cget("state")) == "normal"
            assert "计算中" not in str(view.only_diff_cb.cget("text"))
            assert tuple(view.display_rows) == tuple(exact_rows)
            blocks = tuple(view._ensure_full_diff_blocks())
            assert blocks and all(
                int(pair_idx) in exact_rows
                for block in blocks
                for pair_idx in range(block.start_pair_idx, block.end_pair_idx + 1)
            ), blocks
            assert str(view.use_right_btn.cget("state")) == "normal"
            assert not evidence["preload_calls"], evidence["preload_calls"]
            assert not evidence["loader_calls"], evidence["loader_calls"]
            assert not evidence["native_save_calls"], evidence["native_save_calls"]
            assert not app.manual_a_cell_ops and not app.manual_b_cell_ops
            assert not app.modified_a and not app.modified_b
    finally:
        view._start_async_large_only_diff_build = original_start
        view._cache_only_diff_rows_from_exact_pair_maps = original_cache_from_maps
        case.close_app(app)


def test_edit_ready_callback_never_calls_refresh_rescan_true(case):
    app, view, _source_paths = _open_ready_view(case)
    original_refresh = view.refresh
    original_enqueue = app._enqueue_sheet
    original_kick = app._kick_worker
    original_missing = view._is_missing_sheet_view
    original_edit_ready = app._edit_workbooks_ready
    try:
        refresh_calls = []
        enqueue_calls = []

        def _record_refresh(*args, **kwargs):
            refresh_calls.append((args, dict(kwargs)))

        view.refresh = _record_refresh
        app._enqueue_sheet = lambda *args, **kwargs: enqueue_calls.append(
            (args, dict(kwargs))
        )
        app._kick_worker = lambda: None
        app._edit_workbooks_ready = lambda: True

        # Ordinary materialized views must be requeued for prepared background
        # data rather than synchronously rescanned by the edit-ready callback.
        view._data_ready = True
        view._cache_formula_aware = False
        app._refresh_loaded_views_after_edit_ready()
        assert refresh_calls == [], refresh_calls
        assert enqueue_calls, "non-formula-aware view was not requeued"

        # The no-rescan contract is universal, including missing-Sheet views.
        # The legacy foreground refresh branch must stay bypassed here.
        refresh_calls.clear()
        enqueue_calls.clear()
        view._data_ready = True
        view._cache_formula_aware = False
        view._is_missing_sheet_view = lambda: True
        app._refresh_loaded_views_after_edit_ready()
        assert not any(
            bool(kwargs.get("rescan"))
            for _args, kwargs in refresh_calls
        ), refresh_calls
    finally:
        view.refresh = original_refresh
        app._enqueue_sheet = original_enqueue
        app._kick_worker = original_kick
        view._is_missing_sheet_view = original_missing
        app._edit_workbooks_ready = original_edit_ready
        case.close_app(app)


def test_stale_exact_generation_cannot_publish(case):
    app, view, _source_paths = _open_ready_view(case)
    original_queue_ui = app._queue_ui_task
    try:
        queued_ui = []
        before_diff = copy.deepcopy(view.pair_diff_cols)
        before_base_diff = copy.deepcopy(view.pair_base_diff_cols)
        view._invalidate_only_diff_snapshot_cache()
        view.only_diff_var.set(1)
        app._queue_ui_task = lambda fn: queued_ui.append(fn) or True
        assert view._start_async_large_only_diff_build() is True
        deadline = time.monotonic() + 10.0
        if _ACTIVE_CASE_DEADLINE is not None:
            deadline = min(deadline, _ACTIVE_CASE_DEADLINE)
        while time.monotonic() < deadline:
            with app._exact_broker_lock:
                running = bool(app._exact_broker_running)
            if not running and queued_ui:
                break
            time.sleep(0.01)
        assert queued_ui, "exact worker did not queue a result"

        # Supersede the worker after it computed but before Tk publication.
        view._only_diff_async_build_seq += 1
        for callback in queued_ui:
            callback()
        assert view.pair_diff_cols == before_diff
        assert view.pair_base_diff_cols == before_base_diff
        assert not view._has_valid_only_diff_snapshot_cache()
        assert view._derive_lifecycle_state() != "READY"
    finally:
        app._queue_ui_task = original_queue_ui
        case.close_app(app)


def test_hidden_cache_completeness_is_retained_without_view_render(case):
    mine = str(case.root / "hidden-mine.xlsx")
    theirs = str(case.root / "hidden-theirs.xlsx")
    for path, suffix in ((mine, "mine"), (theirs, "theirs")):
        workbook = Workbook()
        data = workbook.active
        data.title = "Data"
        data.append(("id@id", "value"))
        data.append(("string", "string"))
        data.append(("row-1", suffix))
        hidden = workbook.create_sheet("Hidden")
        hidden.append(("id@id", "value"))
        hidden.append(("string", "string"))
        hidden.append(("hidden-1", suffix))
        workbook.save(path)
        workbook.close()
    case.track_inputs(mine, theirs)

    app = smt.SowMergeApp(mine, theirs)
    case.track_app(app)
    try:
        app.nb.select(app._sheet_containers["Data"])
        _wait_until(
            app.root,
            lambda: "Hidden" in app._sheet_cache_store,
            "hidden Sheet cache was not retained",
        )
        assert app.sheet_views["Hidden"] is None
        cache = app._sheet_cache_store["Hidden"]
        completeness = cache.get("completeness") or {}
        assert set(completeness) >= {
            "formula_aware",
            "row_model_exact",
            "column_projection_exact",
            "ab_diff_exact",
            "base_diff_exact",
            "only_diff_rows_exact",
            "mode",
        }, completeness
        assert completeness["formula_aware"] is True
        assert completeness["row_model_exact"] is True
        assert completeness["column_projection_exact"] is True
        with app._compute_lock:
            app._compute_queue[:] = ["Hidden", "Data"]
        app._enqueue_sheet("Data", front=True)
        with app._compute_lock:
            assert app._compute_queue[0] == "Data", app._compute_queue
    finally:
        case.close_app(app)


def _run_case(case_name: str, callback) -> None:
    global _ACTIVE_CASE_DEADLINE
    original_settings_path = smt._SETTINGS_PATH
    user_settings_path = Path(original_settings_path)
    user_settings_before = _path_snapshot(user_settings_path)
    environment_before = {
        key: value for key, value in os.environ.items() if key.startswith("SOW_")
    }
    root_path = None
    previous_deadline = _ACTIVE_CASE_DEADLINE
    primary_error: BaseException | None = None

    def _record_secondary(
        failures: list[tuple[str, BaseException]],
        label: str,
        check,
    ) -> None:
        try:
            check()
        except BaseException as exc:
            failures.append((label, exc))

    def _assert_cleanup(condition: bool, detail) -> None:
        if not condition:
            raise AssertionError(detail)

    def _preserve_primary_or_raise_cleanup(
        failures: list[tuple[str, BaseException]],
    ) -> None:
        if not failures:
            return
        details = "; ".join(
            f"{label}={type(exc).__name__}: {exc}"
            for label, exc in failures
        )
        if primary_error is not None:
            try:
                primary_error.add_note(f"secondary cleanup verification: {details}")
            except Exception:
                pass
            return
        raise AssertionError(f"{case_name} cleanup verification failed: {details}") from failures[0][1]

    try:
        with tempfile.TemporaryDirectory(prefix=f"sow_loading_readonly_{case_name}_") as raw_root:
            root_path = Path(raw_root)
            temp_settings_path = root_path / "settings.json"
            temp_settings_path.write_text(json.dumps({"only_diff": 0}), encoding="utf-8")
            smt._SETTINGS_PATH = str(temp_settings_path)
            context = _CaseContext(root_path)
            _ACTIVE_CASE_DEADLINE = time.monotonic() + 90.0
            print(f"LOADING_READONLY_CASE_START {case_name}", flush=True)
            try:
                callback(context)
                context.verify_inputs()
                if time.monotonic() > _ACTIVE_CASE_DEADLINE:
                    raise AssertionError(f"{case_name} exceeded the 90s case deadline")
                print(
                    "LOADING_READONLY_CASE_OK "
                    + json.dumps(
                        {
                            "case": case_name,
                            "deadline_seconds": 90,
                            "inputs": context.input_provenance(),
                            "settings_path_is_temp": str(smt._SETTINGS_PATH)
                            == str(temp_settings_path),
                            "user_settings_unchanged": _path_snapshot(user_settings_path)
                            == user_settings_before,
                            "sow_environment_unchanged": {
                                key: value
                                for key, value in os.environ.items()
                                if key.startswith("SOW_")
                            }
                            == environment_before,
                            "monkeypatches_restored": True,
                        },
                        ensure_ascii=False,
                        sort_keys=True,
                    ),
                    flush=True,
                )
            except BaseException as exc:
                primary_error = exc
                raise
            finally:
                cleanup_failures: list[tuple[str, BaseException]] = []

                def _assert_temp_settings_path() -> None:
                    _assert_cleanup(
                        str(smt._SETTINGS_PATH) == str(temp_settings_path),
                        smt._SETTINGS_PATH,
                    )

                def _assert_user_settings_unchanged() -> None:
                    _assert_cleanup(
                        _path_snapshot(user_settings_path) == user_settings_before,
                        user_settings_path,
                    )

                def _restore_settings_path() -> None:
                    smt._SETTINGS_PATH = original_settings_path

                def _assert_sow_environment_unchanged() -> None:
                    _assert_cleanup(
                        {
                            key: value
                            for key, value in os.environ.items()
                            if key.startswith("SOW_")
                        }
                        == environment_before,
                        "test changed SOW environment",
                    )

                _record_secondary(
                    cleanup_failures, "app shutdown", context.close_remaining,
                )
                _record_secondary(
                    cleanup_failures, "synthetic input SHA", context.verify_inputs,
                )
                _record_secondary(
                    cleanup_failures,
                    "temporary settings path",
                    _assert_temp_settings_path,
                )
                _record_secondary(
                    cleanup_failures,
                    "user settings before restore",
                    _assert_user_settings_unchanged,
                )
                _record_secondary(
                    cleanup_failures,
                    "settings restore",
                    _restore_settings_path,
                )
                _record_secondary(
                    cleanup_failures,
                    "user settings after restore",
                    _assert_user_settings_unchanged,
                )
                _record_secondary(
                    cleanup_failures,
                    "SOW environment",
                    _assert_sow_environment_unchanged,
                )
                _preserve_primary_or_raise_cleanup(cleanup_failures)
    finally:
        _ACTIVE_CASE_DEADLINE = previous_deadline
        smt._SETTINGS_PATH = original_settings_path
        if root_path is not None:
            root_cleanup_error = None
            try:
                assert not root_path.exists(), root_path
            except BaseException as exc:
                root_cleanup_error = exc
            if root_cleanup_error is not None:
                _preserve_primary_or_raise_cleanup(
                    [("TemporaryDirectory removal", root_cleanup_error)]
                )


def main(argv=None):
    parser = argparse.ArgumentParser()
    parser.add_argument("--list-cases", action="store_true")
    parser.add_argument("--case", choices=_CASES)
    args = parser.parse_args(argv)
    if args.list_cases:
        if args.case:
            parser.error("--list-cases cannot be combined with --case")
        for case_name in _CASES:
            print(case_name, flush=True)
        return

    callbacks = {
        "non-ready-zero-write": test_non_ready_direct_handlers_have_no_write_side_effects,
        "only-diff-pending": test_only_diff_pending_stays_checked_locked_and_keeps_stable_view,
        "edit-ready-no-rescan": test_edit_ready_callback_never_calls_refresh_rescan_true,
        "stale-generation": test_stale_exact_generation_cannot_publish,
        "hidden-cache": test_hidden_cache_completeness_is_retained_without_view_render,
    }
    selected = (args.case,) if args.case else _CASES
    for case_name in selected:
        _run_case(case_name, callbacks[case_name])
    print(f"PASS: loading/read-only gate regression ({len(selected)} cases)", flush=True)


if __name__ == "__main__":
    main()

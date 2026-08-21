"""Fresh-process, read-only real-workbook corpus benchmark (OpenSpec 9.1-9.5).

Only this program's ``TemporaryDirectory`` roots contain Mine/Base/Theirs.
The source corpus is recursively enumerated as input-only.  A file/mode worker
copies a workbook once, then gives every Sheet worker those same disposable
copies.  JSON is atomically checkpointed after every file/mode and JSONL is
flushed after every Sheet so a crash or a slow outlier cannot hide evidence.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import os
import shutil
import subprocess
import sys
import tempfile
import time
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path

import psutil
from openpyxl import load_workbook

ROOT = Path(r"C:\GM15\design\sheets\develop")
EXTS = {".xlsx", ".xlsm"}
SCHEMA = "large-sheet-corpus-v2"


def write_json(path, value):
    path = Path(path); path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    os.replace(temporary, path)


def append_jsonl(path, value):
    with Path(path).open("a", encoding="utf-8", newline="\n") as stream:
        stream.write(json.dumps(value, ensure_ascii=False, sort_keys=True) + "\n")
        stream.flush(); os.fsync(stream.fileno())


def signature(path):
    stat = path.stat(); digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return {"size_bytes": stat.st_size, "mtime_ns": stat.st_mtime_ns, "sha256": digest.hexdigest()}


def inventory(root):
    items = []
    for path in sorted(Path(root).rglob("*"), key=lambda entry: str(entry).casefold()):
        if not path.is_file():
            continue
        item = {"relative_path": str(path.relative_to(root)), "source_path": str(path), "extension": path.suffix.casefold()}
        item["classification"] = "temporary" if path.name.startswith("~$") else "supported" if item["extension"] in EXTS else "unsupported"
        try:
            stat = path.stat(); item.update(size_bytes=stat.st_size, mtime_ns=stat.st_mtime_ns)
        except OSError as exc:
            item.update(classification="inventory_error", error=f"{type(exc).__name__}: {exc}")
        items.append(item)
    return {"source_root": str(root), "read_only_contract": "source files are never test sides or mutation targets", "files": items,
            "counts": {kind: sum(item["classification"] == kind for item in items) for kind in ("supported", "temporary", "unsupported", "inventory_error")}}


def copy_sides(source, root, mode):
    source, root = Path(source), Path(root)
    if root.resolve().is_relative_to(ROOT.resolve()):
        raise AssertionError("disposable benchmark root cannot be under corpus source")
    paths = [root / "Mine" / source.name, root / "Theirs" / source.name]
    if mode == "3way": paths.append(root / "Base" / source.name)
    for path in paths:
        path.parent.mkdir(parents=True, exist_ok=True); shutil.copy2(source, path)
    return paths[0], paths[1], paths[2] if len(paths) == 3 else None


def formula_count(snapshot):
    return sum(cell.formula_kind != "literal" for row in snapshot.rows for cell in row.cells)


def sheet_worker(args):
    """No source input is opened: mine/theirs/base have already been copied once."""
    import sow_merge_tool as sm
    process = psutil.Process(); rss_before = process.memory_info().rss; started = time.perf_counter()
    result = {"schema": SCHEMA, "mode": args.mode, "source_path": args.source, "sheet": args.sheet,
              "engine": "snapshot-paired-read-only", "fallback": None, "status": "FAILED", "final_state": "FAILED",
              "oracle_zero_difference": False, "oracle_zero_conflict": False, "request_to_final_exact_ms": None,
              "revisit_ms": None, "error": None, "gate_seconds": args.gate}
    try:
        paths = [Path(args.mine), Path(args.theirs)] + ([Path(args.base)] if args.base else [])
        if any(path.resolve().is_relative_to(ROOT.resolve()) for path in paths):
            raise AssertionError("Sheet worker was passed a source file rather than a disposable copy")
        request = time.perf_counter()
        with ThreadPoolExecutor(max_workers=len(paths)) as pool:
            futures = [pool.submit(sm._stream_selected_sheet_snapshot, str(path), str(path), args.sheet, side)
                       for path, side in zip(paths, ("mine", "theirs", "base"))]
            mine, theirs = futures[0].result(), futures[1].result()
            base = futures[2].result() if len(futures) == 3 else None
        comparison = sm._compare_selected_sheet_snapshots(mine, theirs, base)
        elapsed = (time.perf_counter() - request) * 1000
        diffs = sum(bool(value) for value in comparison.pair_diff_cols) + sum(bool(value) for value in comparison.pair_base_diff_cols)
        conflicts = sum(bool(value) for value in comparison.conflict_cols)
        exact_same = diffs == conflicts == 0 and not comparison.unresolved
        result.update({"dimensions": {"rows": mine.max_row, "columns": mine.max_col}, "formula_cells": formula_count(mine),
                       "request_to_final_exact_ms": round(elapsed, 3), "diff_pair_count": diffs, "conflict_pair_count": conflicts,
                       "unresolved": bool(comparison.unresolved), "oracle_zero_difference": diffs == 0,
                       "oracle_zero_conflict": conflicts == 0,
                       "final_state": "EXACT_SAME" if exact_same else "UNRESOLVED" if comparison.unresolved else "EXACT_CHANGED",
                       "status": "PASS" if exact_same and elapsed <= args.gate * 1000 else "TIMEOUT" if elapsed > args.gate * 1000 else "ORACLE_FAILURE"})
        revisit = time.perf_counter(); _ = comparison.row_pairs, mine.rows, theirs.rows
        result["revisit_ms"] = round((time.perf_counter() - revisit) * 1000, 3)
    except Exception as exc:
        result.update(error=f"{type(exc).__name__}: {exc}", fallback="none: snapshot result unavailable")
    finally:
        result.update(rss_before_bytes=rss_before, rss_after_bytes=process.memory_info().rss,
                      rss_delta_bytes=process.memory_info().rss-rss_before, rss_peak_bytes=process.memory_info().rss,
                      startup_and_request_ms=round((time.perf_counter()-started)*1000, 3))
    return result


def run_sheet_child(args, mine, theirs, base, sheet):
    command = [sys.executable, str(Path(__file__).resolve()), "--sheet-worker", "--source", args.source, "--mode", args.mode,
               "--sheet", sheet, "--mine", str(mine), "--theirs", str(theirs), "--gate", str(args.gate)]
    if base: command += ["--base", str(base)]
    child = subprocess.Popen(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
    try:
        stdout, stderr = child.communicate(timeout=args.gate + args.process_grace)
    except subprocess.TimeoutExpired as exc:
        child.kill(); stdout, stderr = child.communicate()
        return {"schema": SCHEMA, "mode": args.mode, "source_path": args.source, "sheet": sheet, "engine": "snapshot-paired-read-only",
                "fallback": "process-watchdog", "status": "PROCESS_TIMEOUT", "final_state": "FAILED", "request_to_final_exact_ms": None,
                "revisit_ms": None, "oracle_zero_difference": False, "oracle_zero_conflict": False, "child_pid": child.pid,
                "child_cleanup": "killed-and-waited", "error": f"process exceeded {args.gate + args.process_grace:.1f}s: {exc}"}
    lines = [line for line in stdout.splitlines() if line.strip()]
    if child.returncode or not lines:
        return {"schema": SCHEMA, "mode": args.mode, "source_path": args.source, "sheet": sheet, "engine": "snapshot-paired-read-only",
                "fallback": "worker-crash", "status": "WORKER_FAILURE", "final_state": "FAILED", "request_to_final_exact_ms": None,
                "revisit_ms": None, "oracle_zero_difference": False, "oracle_zero_conflict": False, "child_pid": child.pid,
                "child_cleanup": "waited", "error": (stderr or stdout or f"exit {child.returncode}")[-4000:]}
    try:
        parsed = json.loads(lines[-1]); parsed.update(child_pid=child.pid, child_cleanup="waited"); return parsed
    except json.JSONDecodeError as exc:
        return {"schema": SCHEMA, "mode": args.mode, "source_path": args.source, "sheet": sheet, "engine": "snapshot-paired-read-only",
                "fallback": "worker-protocol", "status": "WORKER_PROTOCOL_FAILURE", "final_state": "FAILED", "request_to_final_exact_ms": None,
                "revisit_ms": None, "oracle_zero_difference": False, "oracle_zero_conflict": False, "child_pid": child.pid,
                "child_cleanup": "waited", "error": f"{exc}: {lines[-1][-500:]}"}


def file_worker(args):
    import_started = time.perf_counter(); process = psutil.Process(); before = process.memory_info().rss
    result = {"schema": SCHEMA, "source_path": args.source, "relative_path": args.relative_path, "mode": args.mode, "status": "FAILED",
              "startup_ms": None, "catalog_ms": None, "whole_summary_ms": None, "sheets": [], "error": None, "rss_before_bytes": before}
    try:
        import sow_merge_tool  # startup evidence includes production import
        result["startup_ms"] = round((time.perf_counter()-import_started)*1000, 3)
        catalog_started = time.perf_counter()
        workbook = load_workbook(args.source, read_only=True, data_only=False, keep_vba=Path(args.source).suffix.casefold() == ".xlsm")
        try: sheets = list(workbook.sheetnames)
        finally: workbook.close()
        result["catalog_ms"] = round((time.perf_counter()-catalog_started)*1000, 3)
        with tempfile.TemporaryDirectory(prefix="sow_corpus_file_") as temporary:
            mine, theirs, base = copy_sides(args.source, temporary, args.mode)
            result["disposable_sides"] = True
            for sheet in sheets:
                sheet_result = run_sheet_child(args, mine, theirs, base, sheet)
                result["sheets"].append(sheet_result)
                append_jsonl(args.jsonl, {"event": "sheet-complete", "relative_path": args.relative_path, "mode": args.mode, "sheet_result": sheet_result})
        result["status"] = "PASS" if all(sheet["status"] == "PASS" for sheet in result["sheets"]) else "FAIL"
    except Exception as exc:
        result["error"] = f"{type(exc).__name__}: {exc}"
    finally:
        result.update(whole_summary_ms=round((time.perf_counter()-import_started)*1000, 3), rss_after_bytes=process.memory_info().rss,
                      rss_delta_bytes=process.memory_info().rss-before, rss_peak_bytes=process.memory_info().rss)
    return result


def file_child(source, relative_path, mode, args):
    command = [sys.executable, str(Path(__file__).resolve()), "--file-worker", "--source", str(source), "--relative-path", relative_path,
               "--mode", mode, "--gate", str(args.gate), "--process-grace", str(args.process_grace), "--jsonl", str(args.jsonl)]
    complete = subprocess.run(command, capture_output=True, text=True)
    lines = [line for line in complete.stdout.splitlines() if line.strip()]
    if complete.returncode or not lines:
        return {"source_path": str(source), "relative_path": relative_path, "mode": mode, "status": "FILE_WORKER_FAILURE", "sheets": [],
                "error": (complete.stderr or complete.stdout or f"exit {complete.returncode}")[-4000:]}
    return json.loads(lines[-1])


def rankings(runs, mode):
    rows = [{"relative_path": run.get("relative_path"), "sheet": sheet.get("sheet"), "status": sheet.get("status"),
             "final_state": sheet.get("final_state"), "request_to_final_exact_ms": sheet.get("request_to_final_exact_ms"), "error": sheet.get("error")}
            for run in runs if run.get("mode") == mode for sheet in run.get("sheets", [])]
    return sorted(rows, key=lambda row: (row["request_to_final_exact_ms"] is None, -(row["request_to_final_exact_ms"] or -1)))


def write_report(path, evidence):
    lines = ["# Large-sheet corpus slowest Sheet rankings", ""]
    for mode in ("2way", "3way"):
        lines += [f"## {mode}", "", "| Rank | Workbook | Sheet | Exact-ready ms | State | Status |", "| ---: | --- | --- | ---: | --- | --- |"]
        for number, row in enumerate(evidence["slowest_sheets"][mode], 1):
            elapsed = "-" if row["request_to_final_exact_ms"] is None else f"{row['request_to_final_exact_ms']:.3f}"
            lines.append(f"| {number} | {row['relative_path']} | {row['sheet']} | {elapsed} | {row['final_state']} | {row['status']} |")
        lines.append("")
    Path(path).write_text("\n".join(lines), encoding="utf-8")


def controller(args):
    root, out = Path(args.source_root), Path(args.out)
    if not root.is_dir(): raise FileNotFoundError(f"source corpus not found: {root}")
    if args.resume and out.exists():
        evidence = json.loads(out.read_text(encoding="utf-8")); evidence.setdefault("runs", [])
    else:
        before = inventory(root)
        evidence = {"schema": SCHEMA, "created_at_epoch": time.time(), "gate_seconds": args.gate, "fresh_process": True,
                    "inventory_before": before, "runs": [], "source_signature_before": {}, "source_signature_after": {},
                    "slowest_sheets": {"2way": [], "3way": []}}
    completed = {(run.get("relative_path"), run.get("mode")) for run in evidence["runs"]} if args.resume else set()
    supported = [item for item in evidence["inventory_before"]["files"] if item["classification"] == "supported"]
    if args.max_files is not None: supported = supported[:args.max_files]
    for item in supported:
        source, relative = Path(item["source_path"]), item["relative_path"]
        if relative not in evidence["source_signature_before"]:
            try: evidence["source_signature_before"][relative] = signature(source)
            except Exception as exc: evidence["runs"].append({"relative_path": relative, "source_path": str(source), "status": "SOURCE_SIGNATURE_FAILURE", "error": f"{type(exc).__name__}: {exc}", "sheets": []}); continue
        for mode in ("2way", "3way"):
            if (relative, mode) in completed: continue
            run = file_child(source, relative, mode, args); evidence["runs"].append(run)
            evidence["slowest_sheets"] = {item_mode: rankings(evidence["runs"], item_mode) for item_mode in ("2way", "3way")}
            append_jsonl(args.jsonl, {"event": "file-mode-checkpoint", "relative_path": relative, "mode": mode, "status": run.get("status")})
            write_json(out, evidence)  # atomic durable file/mode checkpoint
    for item in supported:
        try: evidence["source_signature_after"][item["relative_path"]] = signature(Path(item["source_path"]))
        except Exception as exc: evidence.setdefault("source_signature_errors", {})[item["relative_path"]] = f"{type(exc).__name__}: {exc}"
    evidence["inventory_after"] = inventory(root)
    evidence["source_unchanged"] = evidence["source_signature_before"] == evidence["source_signature_after"] and not evidence.get("source_signature_errors")
    evidence["slowest_sheets"] = {mode: rankings(evidence["runs"], mode) for mode in ("2way", "3way")}
    write_json(out, evidence); write_report(out.with_name(out.stem + "_slowest.md"), evidence)
    failed = [sheet for run in evidence["runs"] for sheet in run.get("sheets", []) if sheet.get("status") != "PASS"]
    print(json.dumps({"out": str(out), "jsonl": str(args.jsonl), "runs": len(evidence["runs"]), "sheet_failures": len(failed), "source_unchanged": evidence["source_unchanged"]}, ensure_ascii=False))
    return 0 if not failed and evidence["source_unchanged"] else 1


def main():
    parser = argparse.ArgumentParser(); parser.add_argument("--source-root", default=str(ROOT)); parser.add_argument("--out", default="benchmark_results/large_sheet_corpus.json")
    parser.add_argument("--jsonl", default="benchmark_results/large_sheet_corpus.jsonl"); parser.add_argument("--gate", type=float, default=15.0); parser.add_argument("--process-grace", type=float, default=30.0)
    parser.add_argument("--resume", action="store_true"); parser.add_argument("--max-files", type=int); parser.add_argument("--file-worker", action="store_true"); parser.add_argument("--sheet-worker", action="store_true")
    parser.add_argument("--source"); parser.add_argument("--relative-path", default=""); parser.add_argument("--mode", choices=("2way", "3way")); parser.add_argument("--sheet"); parser.add_argument("--mine"); parser.add_argument("--theirs"); parser.add_argument("--base")
    args = parser.parse_args()
    if args.sheet_worker: print(json.dumps(sheet_worker(args), ensure_ascii=False, sort_keys=True)); return
    if args.file_worker: print(json.dumps(file_worker(args), ensure_ascii=False, sort_keys=True)); return
    raise SystemExit(controller(args))


if __name__ == "__main__": main()

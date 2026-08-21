"""Disposable fixtures for the large-sheet exact-result Oracle.

The real Excel files listed here are *never* modified.  Synthetic variants
are intentionally written below ``_test_temp_utils.make_temp_dir`` so a
failed or interrupted test cannot leave any artifact in the design directory.
"""

from __future__ import annotations

import argparse
import os
import shutil
import zipfile
from dataclasses import dataclass
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import Workbook, load_workbook

from _test_temp_utils import make_temp_dir


REAL_SOURCE_ROOT = Path(r"C:\GM15\design\sheets\develop")


@dataclass(frozen=True)
class RealFixture:
    name: str
    filename: str
    sheet: str
    min_rows: int = 3000

    @property
    def path(self) -> Path:
        return REAL_SOURCE_ROOT / self.filename


REAL_FIXTURES = (
    RealFixture("Skill", "Skill.xlsx", "SkillLogicBuff@design"),
    RealFixture("WorldMonster", "WorldMonster.xlsx", "WorldMonsterSurvivor@design", 18000),
    RealFixture("Dungeon", "Dungeon.xlsx", "MonsterGroup@design"),
    RealFixture("Language", "Language.xlsx", "default@design@na_TLanguageCn", 20000),
    # The Sheet name includes localized text and can be locale-display sensitive;
    # callers use ``resolve_real_fixture_sheet`` if the literal differs.
    RealFixture("IdleBuildingComposite", "IdleBuilding.xlsx", "IdleBuildingEquip备份@pm", 6000),
)


def assert_real_sources_read_only() -> None:
    missing = [str(item.path) for item in REAL_FIXTURES if not item.path.is_file()]
    if missing:
        raise FileNotFoundError("Missing read-only real fixture(s): " + ", ".join(missing))


def resolve_real_fixture_sheet(item: RealFixture) -> str:
    """Return the declared Sheet, or the largest qualifying Sheet as fallback."""
    assert_real_sources_read_only()
    wb = load_workbook(item.path, read_only=True, data_only=False)
    try:
        if item.sheet in wb.sheetnames:
            return item.sheet
        candidates = [ws for ws in wb.worksheets if int(ws.max_row or 0) >= item.min_rows]
        if not candidates:
            raise AssertionError(f"{item.path} has no Sheet with >= {item.min_rows} rows")
        return max(candidates, key=lambda ws: int(ws.max_row or 0)).title
    finally:
        wb.close()


def copy_real_fixture(item: RealFixture, target_root: str | os.PathLike[str] | None = None) -> tuple[Path, str]:
    """Copy one source to a disposable root and return (copy, resolved_sheet)."""
    assert_real_sources_read_only()
    root = Path(target_root or make_temp_dir("sow_large_sheet_real_"))
    root.mkdir(parents=True, exist_ok=True)
    target = root / item.filename
    if target.resolve().is_relative_to(REAL_SOURCE_ROOT.resolve()):
        raise AssertionError("real fixture mutation target must be disposable")
    shutil.copy2(item.path, target)
    return target, resolve_real_fixture_sheet(item)


def _write_book(path: Path, rows: list[list[object]], *, sheet: str = "Data") -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    for row in rows:
        ws.append(row)
    wb.save(path)
    wb.close()


def _set_formula_cache(path: Path, cell_ref: str, value: object) -> None:
    """Set a formula's cached ``<v>`` in a disposable xlsx package only."""
    ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    with zipfile.ZipFile(path, "r") as src:
        members = {name: src.read(name) for name in src.namelist()}
    root = ET.fromstring(members["xl/worksheets/sheet1.xml"])
    cell = next((node for node in root.iter(ns + "c") if node.attrib.get("r") == cell_ref), None)
    if cell is None or cell.find(ns + "f") is None:
        raise AssertionError(f"formula cell missing: {cell_ref}")
    cached = cell.find(ns + "v")
    if cached is None:
        cached = ET.SubElement(cell, ns + "v")
    cached.text = str(value)
    members["xl/worksheets/sheet1.xml"] = ET.tostring(root, encoding="utf-8", xml_declaration=True)
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as dst:
        for name, payload in members.items():
            dst.writestr(name, payload)


def build_adversarial_fixture_set(target_root: str | os.PathLike[str] | None = None) -> dict[str, dict[str, object]]:
    """Create every 1.4 case below a disposable directory and return metadata."""
    root = Path(target_root or make_temp_dir("sow_large_sheet_oracle_"))
    root.mkdir(parents=True, exist_ok=True)
    if root.resolve().is_relative_to(REAL_SOURCE_ROOT.resolve()):
        raise AssertionError("adversarial fixtures must not be written into design sources")
    cases: dict[str, dict[str, object]] = {}

    def case(name: str, mine: list[list[object]], theirs: list[list[object]], base=None, **meta):
        folder = root / name
        folder.mkdir(exist_ok=True)
        mine_path, theirs_path = folder / "mine.xlsx", folder / "theirs.xlsx"
        _write_book(mine_path, mine)
        _write_book(theirs_path, theirs)
        item = {"root": str(folder), "mine": str(mine_path), "theirs": str(theirs_path), "sheet": "Data", **meta}
        if base is not None:
            base_path = folder / "base.xlsx"
            _write_book(base_path, base)
            item["base"] = str(base_path)
        cases[name] = item
        return item

    header = [["id@id", "level@id", "value"], ["int", "int", "string"]]
    case("duplicate_missing_keys", header + [[1, 1, "a"], [1, 1, "duplicate"], [None, 2, "missing"]], header + [[1, 1, "b"], [1, 1, "duplicate"], [None, 2, "missing"]], ambiguity=True)
    case("composite_key", header + [[1, 1, "a"], [1, 2, "b"]], header + [[1, 1, "a"], [1, 2, "changed"]], composite_key=("id", "level"))
    case("blank_continuation", [["id@id", "text"], ["int", "string"], [1, "head"], [None, "continuation-a"], [None, "continuation-b"], [2, "next"]], [["id@id", "text"], ["int", "string"], [1, "head"], [None, "continuation-a changed"], [None, "continuation-b"], [2, "next"]], continuation=True)
    case("equal_count_insert_delete", [["id@id", "value"], ["int", "string"], [1, "one"], [3, "three"], [4, "four"]], [["id@id", "value"], ["int", "string"], [1, "one"], [2, "two"], [4, "four"]], equal_count=True)
    case("reorder", [["id@id", "value"], ["int", "string"], [1, "one"], [2, "two"], [3, "three"]], [["id@id", "value"], ["int", "string"], [3, "three"], [1, "one"], [2, "two"]], reorder=True)
    formula = case("formula_cache", [["id@id", "result"], ["int", "int"], [1, "=A3+1"]], [["id@id", "result"], ["int", "int"], [1, "=A3+1"]], [["id@id", "result"], ["int", "int"], [1, "=A3+1"]], formula_cache=True)
    _set_formula_cache(Path(formula["mine"]), "B3", 2)
    _set_formula_cache(Path(formula["theirs"]), "B3", 3)
    _set_formula_cache(Path(formula["base"]), "B3", 2)
    case("three_way_conflict", [["id@id", "value"], ["int", "string"], [1, "mine"]], [["id@id", "value"], ["int", "string"], [1, "theirs"]], [["id@id", "value"], ["int", "string"], [1, "base"]], direct_conflict=True)
    case("column_structure", [["id@id", "left", "right"], ["int", "string", "string"], [1, "a", "b"]], [["id@id", "added", "left", "right"], ["int", "string", "string", "string"], [1, "new", "a", "b"]], column_structure=True)
    # Generation/cancellation are deterministic publication-gate adversaries;
    # they intentionally do not need an Excel package to prove stale rejection.
    cases["stale_generation"] = {"root": str(root), "generation_before": 7, "generation_after": 8, "expected_publish": False}
    cases["cancellation"] = {"root": str(root), "generation_before": 9, "cancelled": True, "expected_publish": False}
    return cases


def verify_adversarial_fixture_set(cases: dict[str, dict[str, object]]) -> None:
    """Validate key semantics, including a distinct formula cache, before use."""
    for name, item in cases.items():
        if name in {"stale_generation", "cancellation"}:
            assert item["expected_publish"] is False
            continue
        for side in ("mine", "theirs"):
            assert Path(str(item[side])).is_file(), (name, side)
        if name == "formula_cache":
            values = []
            for side in ("mine", "theirs"):
                wb = load_workbook(item[side], data_only=True, read_only=True)
                values.append(wb["Data"]["B3"].value)
                wb.close()
            assert values == [2, 3], values
        if name == "composite_key":
            wb = load_workbook(item["mine"], read_only=True, data_only=False)
            assert tuple(wb["Data"].iter_rows(min_row=1, max_row=1, values_only=True).__next__()[:2]) == ("id@id", "level@id")
            wb.close()
        if name == "three_way_conflict":
            assert Path(str(item["base"])).is_file()


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--out")
    args = parser.parse_args()
    cases = build_adversarial_fixture_set(args.out)
    verify_adversarial_fixture_set(cases)
    print("LARGE_SHEET_ORACLE_FIXTURES_OK", cases["duplicate_missing_keys"]["root"])


if __name__ == "__main__":
    main()

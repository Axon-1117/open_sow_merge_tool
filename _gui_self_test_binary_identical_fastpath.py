"""Fresh real-GUI regressions for binary-identical snapshot comparison.

Each selector owns its files, temporary settings redirect, startup-copy ledger,
and cleanup.  Positive selectors are view-only: normal exact compute is allowed,
but an isolated snapshot child, editable load, operation, undo, or save is not.
"""

from __future__ import annotations

import argparse
import contextlib
import hashlib
import json
import math
import os
import shutil
import sys
import tempfile
import time
import zipfile

from openpyxl import Workbook, load_workbook

import sow_merge_tool as sm


_ROWS = 2200
_TOTAL_ROWS = _ROWS + 2  # declaration/type rows belong to the immutable row model.
_CASE_2WAY = "binary-identical-2way"
_CASE_3WAY = "binary-identical-3way"
_CASE_NEGATIVE = "binary-identical-negative"
_CASE_STALE = "binary-identical-stale"
_CASE_HIDDEN_2WAY = "binary-identical-hidden-2way"
_CASE_HIDDEN_3WAY = "binary-identical-hidden-3way"
_CASES = (_CASE_2WAY, _CASE_3WAY, _CASE_NEGATIVE, _CASE_STALE, _CASE_HIDDEN_2WAY, _CASE_HIDDEN_3WAY)


def _sha(path: str) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as source:
        for block in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _abs(path: str) -> str:
    return os.path.normcase(os.path.abspath(os.fspath(path)))


def _note_or_raise(primary: BaseException | None, errors: list[str]) -> None:
    if primary is not None:
        for error in errors:
            try:
                primary.add_note(error)
            except Exception:
                pass
    elif errors:
        raise AssertionError("cleanup failures: " + " | ".join(errors))


def _pump(app, deadline: float, seconds: float = 0.025) -> None:
    until = min(deadline, time.monotonic() + seconds)
    while time.monotonic() < until:
        app.root.update_idletasks()
        app.root.update()
        time.sleep(0.003)


def _wait(app, predicate, message: str, deadline: float) -> None:
    while time.monotonic() < deadline:
        _pump(app, deadline)
        if predicate():
            return
    selected = str(getattr(app, "selected_sheet", "") or "")
    view = app.sheet_views.get(selected)
    diagnostic = {
        "selected": selected,
        "entry": app._sheet_exact_entry(selected) if selected else None,
        "queue": tuple(getattr(app, "_compute_queue", ()) or ()),
        "inflight": tuple(sorted(getattr(app, "_compute_inflight", set()) or set())),
        "child": repr(getattr(app, "_snapshot_child_owner", None)),
        "child_events": tuple(getattr(app, "_snapshot_child_events", ()) or ()),
        "prepared": getattr(view, "_prepared_complete", None),
        "data_ready": getattr(view, "_data_ready", None),
        "pending": getattr(view, "_pending_exact_render", None),
        "cache": getattr(view, "_cache_source", None),
    }
    raise AssertionError(f"{message}: {json.dumps(diagnostic, ensure_ascii=False, default=str)}")


def _write_book(path: str, sheets: dict[str, int], *, duplicate: bool) -> None:
    book = Workbook()
    initial = book.active
    initial.title = next(iter(sheets))
    for index, (name, count) in enumerate(sheets.items()):
        ws = initial if index == 0 else book.create_sheet(name)
        ws.append(["id@id", "value", "note"])
        ws.append(["int32", "string", "string"])
        for row in range(1, int(count) + 1):
            ws.append([1 if duplicate else row, f"value-{row}", "same"])
    book.save(path)
    book.close()


class _Fixture:
    """Explicit ownership without scanning or deleting arbitrary TEMP paths."""

    def __init__(self, case: str):
        self.temporary = tempfile.TemporaryDirectory(prefix=f"sow_binary_identity_{case}_")
        self.root = self.temporary.name
        self.previous_settings_path = sm._SETTINGS_PATH
        self.user_settings_exists = os.path.exists(self.previous_settings_path)
        self.user_settings_bytes = None
        try:
            if self.user_settings_exists:
                with open(self.previous_settings_path, "rb") as source:
                    self.user_settings_bytes = source.read()
            self.settings_path = os.path.join(self.root, "settings.json")
            with open(self.settings_path, "w", encoding="utf-8") as target:
                json.dump({"only_diff": 0}, target)
            sm._SETTINGS_PATH = self.settings_path
            self.inputs: dict[str, str] = {}
            self.outputs: set[str] = set()
        except BaseException:
            sm._SETTINGS_PATH = self.previous_settings_path
            self.temporary.cleanup()
            raise

    def track(self, path: str) -> str:
        path = _abs(path)
        assert os.path.isfile(path), path
        self.inputs[path] = _sha(path)
        return path

    def sidecar(self, workbook: str, label: str, revision: int) -> str:
        raw = os.path.join(self.root, f"{label}.xlsx.r{revision}")
        shutil.copy2(workbook, raw)
        return self.track(raw)

    def identical(self, name: str, *, three_way: bool, sheets: dict[str, int] | None = None, duplicate: bool = True) -> dict[str, str]:
        template = os.path.join(self.root, f"{name}.xlsx")
        _write_book(template, sheets or {"Data": _ROWS}, duplicate=duplicate)
        result = {"A": self.sidecar(template, f"{name}-mine", 39265), "B": self.sidecar(template, f"{name}-theirs", 39264)}
        if three_way:
            result["BASE"] = self.sidecar(template, f"{name}-base", 39263)
        assert len({_sha(path) for path in result.values()}) == 1, result
        return result

    def output(self, name: str) -> str:
        path = _abs(os.path.join(self.root, name))
        assert _abs(os.path.dirname(path)) == _abs(self.root) and not os.path.lexists(path)
        self.outputs.add(path)
        return path

    def assert_inputs(self) -> None:
        for path, expected in self.inputs.items():
            assert os.path.isfile(path), f"input disappeared: {path}"
            assert _sha(path) == expected, f"input changed: {path}"

    def cleanup(self) -> list[str]:
        errors = []
        try:
            self.assert_inputs()
        except Exception as exc:
            errors.append(f"input SHA: {type(exc).__name__}: {exc}")
        try:
            assert all(not os.path.lexists(path) for path in self.outputs), self.outputs
            with open(self.settings_path, encoding="utf-8") as source:
                assert json.load(source) == {"only_diff": 0}
        except Exception as exc:
            errors.append(f"temporary fixture: {type(exc).__name__}: {exc}")
        try:
            sm._SETTINGS_PATH = self.previous_settings_path
            assert os.path.exists(self.previous_settings_path) == self.user_settings_exists
            if self.user_settings_exists:
                with open(self.previous_settings_path, "rb") as source:
                    assert source.read() == self.user_settings_bytes
        except Exception as exc:
            errors.append(f"user settings: {type(exc).__name__}: {exc}")
        try:
            self.temporary.cleanup()
            assert not os.path.lexists(self.root), self.root
        except Exception as exc:
            errors.append(f"temporary root: {type(exc).__name__}: {exc}")
        return errors


class _App:
    """A per-app shared ledger, including constructor-failure recovery."""

    def __init__(self, fixture: _Fixture, inputs: dict[str, str], sheet: str):
        self.fixture, self.inputs, self.sheet = fixture, dict(inputs), str(sheet)
        self.ledger: set[str] = set()
        self.app = None
        self.effective: tuple[str, ...] = ()

    def __enter__(self):
        kwargs = {"initial_sheet": self.sheet, "startup_owned_paths": self.ledger}
        if "BASE" in self.inputs:
            kwargs.update(merge_mode=True, base_path=self.inputs["BASE"], merged_path=self.fixture.output(f"{self.sheet}-merged.xlsx"))
        try:
            self.app = sm.SowMergeApp(self.inputs["A"], self.inputs["B"], **kwargs)
        except BaseException:
            expected = {_abs(path) for path in self.ledger}
            evidence: list[dict] = []
            if expected:
                sm._consume_owned_startup_temp_paths(self.ledger, evidence)
                observed = {_abs(item.get("path", "")): item for item in evidence}
                assert set(observed) == expected, (observed, expected)
                assert all(item.get("removed") and not item.get("exists_after") and not item.get("error") for item in observed.values()), observed
            assert not self.ledger
            raise
        assert self.app._owned_startup_temp_paths is self.ledger
        paths = [self.app.file_a, self.app.file_b]
        if "BASE" in self.inputs:
            paths.append(self.app.base_path)
        self.effective = tuple(_abs(path) for path in paths)
        assert len(set(self.effective)) == len(self.effective)
        assert set(self.effective) == {_abs(path) for path in self.ledger}, (self.effective, self.ledger)
        assert all(path.endswith(".xlsx") and os.path.isfile(path) for path in self.effective)
        return self.app

    def __exit__(self, exc_type, exc, _tb):
        errors = []
        try:
            _cancel_view_debounces(self.app)
            self.app._shutdown_root()
        except Exception as shutdown_exc:
            errors.append(f"shutdown: {type(shutdown_exc).__name__}: {shutdown_exc}")
        try:
            evidence = tuple(self.app._owned_startup_temp_cleanup_evidence)
            assert not self.ledger
            assert len(evidence) == len(self.effective), evidence
            observed = {_abs(item.get("path", "")): item for item in evidence}
            assert set(observed) == set(self.effective), (observed, self.effective)
            for path in self.effective:
                item = observed[path]
                assert item.get("removed") and item.get("exists_after") is False and not item.get("error"), item
                assert not os.path.lexists(path), path
        except Exception as evidence_exc:
            errors.append(f"startup ledger: {type(evidence_exc).__name__}: {evidence_exc}")
        try:
            self.fixture.assert_inputs()
        except Exception as sha_exc:
            errors.append(f"input SHA after app: {type(sha_exc).__name__}: {sha_exc}")
        _note_or_raise(exc, errors)
        return False


@contextlib.contextmanager
def _threshold(value: int):
    original = sm._LARGE_SHEET_ROW_THRESHOLD
    sm._LARGE_SHEET_ROW_THRESHOLD = int(value)
    try:
        yield
    finally:
        sm._LARGE_SHEET_ROW_THRESHOLD = original


@contextlib.contextmanager
def _view_only_traps(app):
    originals, hits = {}, []
    original_popen = sm.subprocess.Popen

    def blocked(*_args, **_kwargs):
        hits.append("edit-or-save")
        raise AssertionError("binary identity view-only path reached edit/save backend")

    for name in ("_request_edit_preload", "_load_edit_workbooks_owned", "save_a_inplace", "save_b_inplace", "save_merged_and_exit", "_atomic_save_wb"):
        if hasattr(app, name):
            originals[name] = getattr(app, name)
            setattr(app, name, blocked)
    sm.subprocess.Popen = blocked
    primary = None
    try:
        yield hits
    except BaseException as exc:
        primary = exc
        raise
    finally:
        for name, original in originals.items():
            setattr(app, name, original)
        sm.subprocess.Popen = original_popen
        _note_or_raise(primary, [f"forbidden view-only hits: {hits}"] if hits else [])


def _cancel_view_debounces(app) -> None:
    """Cancel test-owned scheduled settings/UI callbacks before app shutdown."""
    for view in tuple(getattr(app, "sheet_views", {}).values()):
        if view is None:
            continue
        for attr in ("_settings_save_id", "_hover_debounce_id", "_diff_map_debounce_id"):
            after_id = getattr(view, attr, None)
            if not after_id:
                continue
            try:
                view.frame.after_cancel(after_id)
            except Exception:
                pass
            finally:
                setattr(view, attr, None)


def _assert_no_edit_backend(app, view) -> None:
    assert tuple(app._edit_load_requests) == ()
    assert not app._edit_loading_started and getattr(app, "_edit_preload_thread", None) is None
    assert not app._edit_preload_active_event.is_set() and not app._edit_workbooks_ready()
    assert all(getattr(app, field, None) is None for field in ("_wb_a_edit", "_wb_b_edit", "_wb_base_edit"))
    assert view._derive_lifecycle_state() == "EDIT_DEFERRED"
    assert getattr(app, "_snapshot_child_owner", None) is None
    assert not set(getattr(app, "_snapshot_child_temp_paths", set()) or set())
    assert not view._pending_exact_render
    assert not app.manual_a_cell_ops and not app.manual_b_cell_ops
    assert not app.manual_a_formula_cache_ops and not app.manual_b_formula_cache_ops
    assert not app.manual_a_row_ops and not app.manual_b_row_ops
    assert not app.manual_a_column_ops and not app.manual_b_column_ops
    assert not app.undo_stack and not app.redo_stack


def _assert_zero_backend(app, view) -> None:
    _assert_no_edit_backend(app, view)
    # The physical-identity path must not spawn the isolated ordinary-snapshot
    # child; non-identical inputs use that child under the default production
    # policy and are checked separately below.
    assert tuple(getattr(app, "_snapshot_child_events", ()) or ()) == ()


def _assert_ordinary_snapshot_child(app) -> None:
    events = tuple(getattr(app, "_snapshot_child_events", ()) or ())
    if not bool(getattr(sm, "_ISOLATED_SNAPSHOT_PROCESS_ENABLED", False)):
        assert events == (), events
        return
    assert len(events) == 1, events
    event = events[0]
    assert event["event"] == event["original_event"] == "finished", event
    assert event["reason"] == "normal-result-verified" and not event["exception"], event
    assert event["terminal_record_error"] in (None, "") and event["result_decoded"] is True, event
    assert event["sheet"] == "Data" and int(event["generation"]) == 0, event
    assert int(event["pid"]) > 0 and int(event["exitcode"]) == 0, event
    assert str(event["token"] or event["request_token"]), event
    assert not event["result_exists_after_cleanup"] and not event["partial_exists_after_cleanup"], event


def _assert_identity(app, view, *, three_way: bool, sheet: str = "Data", rows: int = _TOTAL_ROWS) -> None:
    assert app._is_sheet_exact_current(sheet)
    assert app._sheet_exact_entry(sheet)["state"] == sm._SHEET_EXACT_SAME
    assert view._prepared_complete and view._data_ready and not view._pending_exact_render
    assert view._row_model_exact and view._pair_diff_full_exact and not view.only_diff_var.get()
    assert tuple(view.row_pairs) == tuple((row, row) for row in range(1, rows + 1))
    assert tuple(view._full_display_rows) == tuple(range(rows))
    assert set(view.pair_diff_cols) == set(range(rows)) and all(not item for item in view.pair_diff_cols.values())
    assert set(view.pair_base_diff_cols) == set(range(rows)) and all(not item for item in view.pair_base_diff_cols.values())
    cache = view.column_comparison_cache
    assert cache is not None and not cache.unresolved_cols
    slots = [(slot.logical_idx, slot.mine_col, slot.base_col, slot.theirs_col, slot.state) for slot in cache.model.slots]
    assert slots == [(0, 1, 1 if three_way else None, 1, "retained"), (1, 2, 2 if three_way else None, 2, "retained"), (2, 3, 3 if three_way else None, 3, "retained")]
    if three_way:
        expected = {row: row for row in range(1, rows + 1)}
        assert view.mine_to_base_row == expected and view.theirs_to_base_row == expected
        assert view.pair_base_row_override == {index: index + 1 for index in range(rows)}
    else:
        assert not view.mine_to_base_row and not view.theirs_to_base_row
    assert view._snapshot_metrics_ms.get("binary_identity_fast_path") is True and view._cache_source == "snapshot"
    _assert_zero_backend(app, view)


@contextlib.contextmanager
def _terminal_guard(
    sheet: str,
    *,
    three_way: bool,
    rows: int,
    full_terminal: list[str],
    accepted_records: list[dict],
):
    original = sm.SowMergeApp._set_sheet_exact_state

    def guarded(self, changed_sheet, state, *args, **kwargs):
        selected_terminal = (
            str(changed_sheet) == sheet
            and str(state) in sm._SHEET_EXACT_TERMINAL
            and self.selected_sheet == sheet
        )
        pre = dict(self._sheet_exact_entry(sheet)) if selected_terminal else None
        current_generation = int(self._sheet_compute_generation.get(sheet, 0))
        requested_generation = int(kwargs.get("generation", current_generation))
        accepted = original(self, changed_sheet, state, *args, **kwargs)
        if not selected_terminal:
            return accepted
        post = dict(self._sheet_exact_entry(sheet))
        if accepted:
            assert post["state"] == str(state)
            assert int(post["generation"]) == requested_generation == current_generation
            record = {
                "state": str(state),
                "generation": requested_generation,
                "stage": str(post.get("stage") or ""),
                "pre_state": str(pre.get("state") or ""),
                "pre_full_detail": bool(pre.get("full_detail_terminal")),
                "post_full_detail": bool(post.get("full_detail_terminal")),
            }
            accepted_records.append(record)
            if bool(post.get("full_detail_terminal")):
                view = self.sheet_views[sheet]
                assert view._prepared_complete and view._data_ready and not view._pending_exact_render
                assert len(view.row_pairs) == rows and view.column_comparison_cache is not None
                if three_way:
                    assert view.mine_to_base_row.get(2) == 2 and view.theirs_to_base_row.get(2) == 2
                full_terminal.append(str(state))
            return accepted
        # A duplicate selected terminal is legal only when it is the same
        # current full-detail fact; source drops it without changing registry
        # timestamps, stage, or readiness. It is diagnostic, never a pass item.
        assert requested_generation == current_generation
        assert str(state) == sm._SHEET_EXACT_SAME
        assert pre["state"] == sm._SHEET_EXACT_SAME and bool(pre.get("full_detail_terminal"))
        assert post == pre and bool(post.get("full_detail_terminal"))
        return accepted

    sm.SowMergeApp._set_sheet_exact_state = guarded
    try:
        yield
    finally:
        sm.SowMergeApp._set_sheet_exact_state = original


@contextlib.contextmanager
def _identity_compare_traps():
    """A proven package identity must never re-enter ordinary alignment."""
    original_align = sm._align_selected_sheet_snapshots
    original_compare = sm._compare_selected_sheet_snapshots
    original_atomic_save = sm._atomic_save_wb
    hits = []

    def forbidden(name):
        def invoke(*_args, **_kwargs):
            hits.append(name)
            raise AssertionError(f"binary identity unexpectedly called {name}")
        return invoke

    sm._align_selected_sheet_snapshots = forbidden("_align_selected_sheet_snapshots")
    sm._compare_selected_sheet_snapshots = forbidden("_compare_selected_sheet_snapshots")
    sm._atomic_save_wb = forbidden("_atomic_save_wb")
    primary = None
    try:
        yield hits
    except BaseException as exc:
        primary = exc
        raise
    finally:
        sm._align_selected_sheet_snapshots = original_align
        sm._compare_selected_sheet_snapshots = original_compare
        sm._atomic_save_wb = original_atomic_save
        _note_or_raise(primary, [f"ordinary comparison hits: {hits}"] if hits else [])


def _positive(fixture: _Fixture, deadline: float, *, three_way: bool) -> dict:
    inputs, full_terminal, accepted_records = fixture.identical("positive", three_way=three_way), [], []
    with _terminal_guard(
        "Data",
        three_way=three_way,
        rows=_TOTAL_ROWS,
        full_terminal=full_terminal,
        accepted_records=accepted_records,
    ):
        with _identity_compare_traps():
            with _App(fixture, inputs, "Data") as app:
                view = app.sheet_views["Data"]
                with _view_only_traps(app):
                    _wait(app, lambda: app._is_sheet_exact_current("Data") and view._prepared_complete and view._data_ready and not view._pending_exact_render, "identity full terminal timeout", deadline)
                    _assert_identity(app, view, three_way=three_way)
                assert full_terminal == [sm._SHEET_EXACT_SAME], full_terminal
                assert accepted_records, accepted_records
                assert all(
                    record["state"] == sm._SHEET_EXACT_SAME
                    and record["generation"] == 0
                    for record in accepted_records
                ), accepted_records
                full_indices = [
                    index
                    for index, record in enumerate(accepted_records)
                    if record["post_full_detail"]
                ]
                assert full_indices == [len(accepted_records) - 1], accepted_records
                assert all(
                    not record["post_full_detail"]
                    for record in accepted_records[:full_indices[0]]
                ), accepted_records
                assert accepted_records[full_indices[0]] == {
                    "state": sm._SHEET_EXACT_SAME,
                    "generation": 0,
                    "stage": "精确比较完成",
                    "pre_state": sm._SHEET_EXACT_CALCULATING,
                    "pre_full_detail": False,
                    "post_full_detail": True,
                }, accepted_records
    return {"mode": "3way" if three_way else "2way", "rows": _TOTAL_ROWS, "terminal": full_terminal, "accepted_records": accepted_records, "input_hashes": fixture.inputs}


def _rewrite(path: str, member: str, transform) -> None:
    replacement = path + ".rewrite.xlsx"
    with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(replacement, "w", compression=zipfile.ZIP_DEFLATED) as target:
        seen = False
        for info in source.infolist():
            payload = source.read(info.filename)
            if info.filename == member:
                payload, seen = transform(payload), True
            target.writestr(info, payload)
        if not seen:
            assert member == "xl/sharedStrings.xml"
            target.writestr(member, b'<?xml version="1.0" encoding="UTF-8"?><sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="0" uniqueCount="0"/>')
    os.replace(replacement, path)


def _negative_one(fixture: _Fixture, left: str, right: str, expected: str, deadline: float) -> None:
    with _App(fixture, {"A": left, "B": right}, "Data") as app:
        view = app.sheet_views["Data"]
        with _view_only_traps(app):
            _wait(app, lambda: app._is_sheet_exact_current("Data") and view._prepared_complete and view._data_ready and not view._pending_exact_render, "negative terminal timeout", deadline)
            metrics = dict(view._snapshot_metrics_ms)
            if bool(getattr(sm, "_ISOLATED_SNAPSHOT_PROCESS_ENABLED", False)):
                assert metrics["isolated_process"] is True
                assert "binary_identity_fast_path" not in metrics
                parent_total = float(metrics["parent_total"])
                assert math.isfinite(parent_total) and parent_total >= 0.0
            else:
                assert metrics.get("binary_identity_fast_path") is False
                assert "isolated_process" not in metrics
            assert view._cache_source == "snapshot"
            assert app._sheet_exact_entry("Data")["state"] == expected
            _assert_no_edit_backend(app, view)
            _assert_ordinary_snapshot_child(app)


def _negative(fixture: _Fixture, deadline: float) -> dict:
    completed = []
    with _threshold(10_000):  # ordinary snapshot comparator, never binary identity.
        source_book = os.path.join(fixture.root, "negative-source.xlsx")
        _write_book(source_book, {"Data": _ROWS}, duplicate=False)
        left = fixture.sidecar(source_book, "negative-left", 39265)
        variants = (("byte", "docProps/core.xml", sm._SHEET_EXACT_SAME), ("global", "xl/workbook.xml", sm._SHEET_EXACT_SAME), ("shared", "xl/sharedStrings.xml", sm._SHEET_EXACT_SAME), ("sheet", None, sm._SHEET_EXACT_CHANGED))
        for name, member, expected in variants:
            workbook = os.path.join(fixture.root, f"negative-{name}.xlsx")
            shutil.copy2(source_book, workbook)
            if member:
                _rewrite(workbook, member, lambda payload: payload + b"\n")
            else:
                book = load_workbook(workbook)
                book["Data"]["B8"] = "changed"
                book.save(workbook)
                book.close()
            _negative_one(fixture, left, fixture.sidecar(workbook, f"negative-{name}", 39264), expected, deadline)
            completed.append(name)
    return {"negative_cases": completed, "input_hashes": fixture.inputs}


def _stale(fixture: _Fixture, deadline: float) -> dict:
    inputs = fixture.identical("stale", three_way=False)
    original_snapshot, original_state = sm.SowMergeApp.selected_sheet_snapshot, sm.SowMergeApp._set_sheet_exact_state
    touched, invalidated, accepted = set(), [], []

    def stale_after_a(self, side, sheet, **kwargs):
        snapshot = original_snapshot(self, side, sheet, **kwargs)
        key = (id(self), str(side).upper(), str(sheet))
        if str(side).upper() == "A" and key not in touched:
            touched.add(key)
            with self._compute_lock:
                generation = int(self._sheet_compute_generation.get(sheet, 0))
                self._sheet_compute_generation[sheet] = generation + 1
            invalidated.append(generation)
            self._enqueue_sheet(sheet, front=True, force_recompute=True)
        return snapshot

    def capture(self, sheet, state, *args, **kwargs):
        result = original_state(self, sheet, state, *args, **kwargs)
        if result and str(sheet) == "Data" and str(state) in sm._SHEET_EXACT_TERMINAL:
            accepted.append(int(kwargs.get("generation")))
        return result

    sm.SowMergeApp.selected_sheet_snapshot, sm.SowMergeApp._set_sheet_exact_state = stale_after_a, capture
    try:
        with _identity_compare_traps():
            with _App(fixture, inputs, "Data") as app:
                view = app.sheet_views["Data"]
                with _view_only_traps(app):
                    _wait(app, lambda: app._is_sheet_exact_current("Data") and view._prepared_complete and view._data_ready, "stale retry timeout", deadline)
                    assert invalidated == [0] and accepted == [1], (invalidated, accepted)
                    assert app._sheet_exact_entry("Data")["generation"] == 1
                    _assert_identity(app, view, three_way=False)
    finally:
        sm.SowMergeApp.selected_sheet_snapshot, sm.SowMergeApp._set_sheet_exact_state = original_snapshot, original_state
    return {"invalidated": invalidated, "accepted": accepted, "input_hashes": fixture.inputs}


def _hidden(fixture: _Fixture, deadline: float, *, three_way: bool) -> dict:
    inputs = fixture.identical("hidden", three_way=three_way, sheets={"S1": 160, "S2": _ROWS})
    original_identity, original_state, original_loading = sm._physical_identity_snapshot_comparison, sm.SowMergeApp._set_sheet_exact_state, sm.SheetView._show_loading
    terminal, states, transitions, errors, queued = [], [], [], [], []

    def delayed(snapshot, **kwargs):
        if snapshot.sheet == "S2":
            time.sleep(0.18)
        return original_identity(snapshot, **kwargs)

    def state_hook(self, sheet, state, *args, **kwargs):
        selected_terminal = str(sheet) == "S2" and str(state) in sm._SHEET_EXACT_TERMINAL and self.selected_sheet == "S2"
        if selected_terminal:
            view = self.sheet_views["S2"]
            assert view._prepared_complete and view._data_ready and not view._pending_exact_render and len(view.row_pairs) == _TOTAL_ROWS
        result = original_state(self, sheet, state, *args, **kwargs)
        if result and str(sheet) == "S2" and self.selected_sheet == "S2":
            states.append((str(state), int(kwargs.get("generation", self._sheet_exact_entry("S2").get("generation", -1)))) )
        if selected_terminal and result:
            terminal.append(str(state))
        return result

    def loading_hook(self, *args, **kwargs):
        result = original_loading(self, *args, **kwargs)
        if self.sheet == "S2" and self.app.selected_sheet == "S2" and self._pending_exact_render and not queued:
            queued.append(True)
            def observe():
                try:
                    state = str(self.app._sheet_exact_entry("S2").get("state"))
                    if state == sm._SHEET_EXACT_CALCULATING:
                        assert self._pending_exact_render and not self._prepared_complete and not self._data_ready
                        transitions.append("CALCULATING_MASKED")
                    elif state not in sm._SHEET_EXACT_TERMINAL:
                        raise AssertionError(f"unexpected promotion state: {state}")
                except Exception as exc:
                    errors.append(f"{type(exc).__name__}: {exc}")
            self.app._queue_ui_task(observe)
        return result

    sm._physical_identity_snapshot_comparison, sm.SowMergeApp._set_sheet_exact_state, sm.SheetView._show_loading = delayed, state_hook, loading_hook
    try:
        with _identity_compare_traps():
            with _App(fixture, inputs, "S1") as app:
                with _view_only_traps(app):
                    _wait(app, lambda: app._is_sheet_exact_current("S2") and "S2" in app._sheet_cache_store, "hidden summary timeout", deadline)
                    summary_cache = dict(app._sheet_cache_store["S2"])
                    summary_entry = dict(app._sheet_exact_entry("S2"))
                    summary_generation = int(app._sheet_compute_generation["S2"])
                    summary_completeness = dict(summary_cache.get("completeness") or {})
                    assert app.nb.tab(app.nb.select(), "text") == "S1"
                    assert app.selected_sheet != "S2"
                    assert app.sheet_views.get("S2") is None and app._sheet_loaded.get("S2") is False
                    assert int(summary_cache.get("generation", -1)) == summary_generation
                    assert summary_cache.get("snapshot_engine") is True
                    assert summary_cache.get("prepared_complete") is True
                    assert summary_completeness == {
                        "formula_aware": True,
                        "row_model_exact": True,
                        "column_projection_exact": True,
                        "sheet_summary_exact": True,
                        "ab_diff_exact": True,
                        "base_diff_exact": True,
                        "only_diff_rows_exact": True,
                        "mode": "snapshot-full",
                    }
                    assert dict(summary_cache.get("snapshot_metrics_ms") or {}).get("binary_identity_fast_path") is True
                    assert tuple(summary_cache.get("row_pairs") or ()) == tuple(
                        (row, row) for row in range(1, _TOTAL_ROWS + 1)
                    )
                    assert set(summary_cache.get("pair_diff_cols") or {}) == set(range(_TOTAL_ROWS))
                    assert all(not value for value in (summary_cache.get("pair_diff_cols") or {}).values())
                    assert set(summary_cache.get("pair_base_diff_cols") or {}) == set(range(_TOTAL_ROWS))
                    assert all(not value for value in (summary_cache.get("pair_base_diff_cols") or {}).values())
                    if three_way:
                        expected_base_rows = {row: row for row in range(1, _TOTAL_ROWS + 1)}
                        assert summary_cache.get("mine_to_base_row") == expected_base_rows
                        assert summary_cache.get("theirs_to_base_row") == expected_base_rows
                        assert summary_cache.get("pair_base_row_override") == {
                            index: index + 1 for index in range(_TOTAL_ROWS)
                        }
                    else:
                        assert not summary_cache.get("mine_to_base_row") and not summary_cache.get("theirs_to_base_row")
                    assert summary_entry.get("generation") == summary_generation
                    assert summary_entry.get("state") == sm._SHEET_EXACT_SAME
                    assert summary_entry.get("full_detail_terminal") is False
                    app.nb.select(app._sheet_containers["S2"])
                    _wait(
                        app,
                        lambda: app.nb.tab(app.nb.select(), "text") == "S2"
                        and app.selected_sheet == "S2",
                        "hidden real Notebook selection confirmation",
                        deadline,
                    )
                    view = app.sheet_views.get("S2")
                    assert view is not None and app._sheet_loaded.get("S2") is True
                    _wait(app, lambda: errors or transitions or terminal, "hidden promotion did not expose transition", deadline)
                    assert not errors, errors
                    _wait(
                        app,
                        lambda: app._is_sheet_exact_current("S2")
                        and int(app._sheet_exact_entry("S2").get("generation", -1)) == summary_generation
                        and app._sheet_exact_entry("S2").get("state") == sm._SHEET_EXACT_SAME
                        and app._sheet_exact_entry("S2").get("full_detail_terminal") is True
                        and view._prepared_complete
                        and view._data_ready
                        and not view._pending_exact_render,
                        "hidden full detail timeout",
                        deadline,
                    )
                    _wait(app, lambda: "S2" not in tuple(app._compute_queue) and "S2" not in set(app._compute_inflight), "hidden queue ownership retained", deadline)
                    assert transitions in ([], ["CALCULATING_MASKED"]) and terminal == [sm._SHEET_EXACT_SAME], (transitions, terminal, states)
                    _assert_identity(app, view, three_way=three_way, sheet="S2", rows=_TOTAL_ROWS)
    finally:
        sm._physical_identity_snapshot_comparison, sm.SowMergeApp._set_sheet_exact_state, sm.SheetView._show_loading = original_identity, original_state, original_loading
    return {"mode": "3way" if three_way else "2way", "terminal": terminal, "transitions": transitions, "states": states, "input_hashes": fixture.inputs}


def _dispatch(case: str, fixture: _Fixture, deadline: float) -> dict:
    if case == _CASE_2WAY:
        return _positive(fixture, deadline, three_way=False)
    if case == _CASE_3WAY:
        return _positive(fixture, deadline, three_way=True)
    if case == _CASE_NEGATIVE:
        return _negative(fixture, deadline)
    if case == _CASE_STALE:
        return _stale(fixture, deadline)
    if case == _CASE_HIDDEN_2WAY:
        return _hidden(fixture, deadline, three_way=False)
    if case == _CASE_HIDDEN_3WAY:
        return _hidden(fixture, deadline, three_way=True)
    raise AssertionError(f"unknown case: {case}")


def _run(case: str) -> dict:
    deadline, fixture, primary = time.monotonic() + 90.0, None, None
    try:
        fixture = _Fixture(case)
        detail = _dispatch(case, fixture, deadline)
        assert time.monotonic() <= deadline, f"90-second deadline exceeded: {case}"
        return detail
    except BaseException as exc:
        primary = exc
        raise
    finally:
        if fixture is not None:
            _note_or_raise(primary, fixture.cleanup())


def main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--list-cases", action="store_true")
    parser.add_argument("--case", choices=_CASES)
    args = parser.parse_args(argv)
    if args.list_cases:
        print("\n".join(_CASES))
        return
    selected = (args.case,) if args.case else (_CASE_2WAY,)
    for case in selected:
        print(f"START {case}", flush=True)
        detail = _run(case)
        print(f"BINARY_IDENTICAL_FASTPATH_OK {json.dumps({'case': case, **detail}, ensure_ascii=False, sort_keys=True)}", flush=True)
    print(f"BINARY_IDENTICAL_FASTPATH_SUITE_OK ({len(selected)} cases)", flush=True)


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"GUI_SELF_TEST_BINARY_IDENTICAL_FASTPATH_FAIL: {exc}", file=sys.stderr)
        raise

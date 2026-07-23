"""Exact full-scan oracle for the conservative 2-way column diff seed."""

from __future__ import annotations

from types import SimpleNamespace

from openpyxl import Workbook

import sow_merge_tool as mod


class _Var:
    def get(self):
        return 0


def _cache(name, a_rows, b_rows):
    width_a = max(len(row) for row in a_rows)
    width_b = max(len(row) for row in b_rows)
    return mod.build_logical_column_comparison_cache_2way(
        mod.ColumnModelCacheKey(name, 1, 1),
        a_rows,
        b_rows,
        a_rows,
        b_rows,
        mine_max_col=width_a,
        theirs_max_col=width_b,
    )


def _worksheets(rows, edit_rows=None):
    value_book = Workbook()
    edit_book = Workbook()
    value = value_book.active
    edit = edit_book.active
    for row_idx, row in enumerate(rows, start=1):
        for col_idx, cell_value in enumerate(row, start=1):
            if cell_value is not None:
                value.cell(row_idx, col_idx).value = cell_value
    for row_idx, row in enumerate(edit_rows or rows, start=1):
        for col_idx, cell_value in enumerate(row, start=1):
            if cell_value is not None:
                edit.cell(row_idx, col_idx).value = cell_value
    return value_book, edit_book, value, edit


def _row(ws, row_idx, width):
    if row_idx is None:
        return (None,) * width
    cells = getattr(ws, "_cells", {}) or {}
    return tuple(
        getattr(cells.get((int(row_idx), col)), "value", None)
        for col in range(1, width + 1)
    )


def _full_map(cache, row_pairs, a_value, b_value, a_edit, b_edit):
    width = max(
        [1]
        + [slot.mine_col or 0 for slot in cache.model.slots]
        + [slot.theirs_col or 0 for slot in cache.model.slots]
    )
    result = {}
    for pair_idx, (ra, rb) in enumerate(row_pairs):
        comparison = mod.compare_logical_row_2way(
            cache,
            _row(a_value, ra, width),
            _row(b_value, rb, width),
            _row(a_edit, ra, width),
            _row(b_edit, rb, width),
            mine_row=ra,
            theirs_row=rb,
            mine_present=ra is not None,
            theirs_present=rb is not None,
        )
        if comparison.diff_cols:
            result[pair_idx] = set(comparison.diff_cols)
    return result


def _view(cache, row_pairs, old_map):
    view = object.__new__(mod.SheetView)
    view.sheet = cache.model.cache_key.sheet_name
    view.app = SimpleNamespace(has_base=False, merge_conflict_mode=False)
    view.three_way_var = _Var()
    view._is_large_sheet = True
    view._data_ready = True
    view._pair_diff_full_exact = True
    view._column_mapping_stale_reason = ""
    view._row_model_version = 1
    view._column_model_version = 1
    view._mine_edit_version = 0
    view._base_edit_version = 0
    view._theirs_edit_version = 0
    view._column_projection_generation = 1
    view.row_pairs = list(row_pairs)
    view.pair_diff_cols = {idx: set(cols) for idx, cols in old_map.items()}
    view.column_comparison_cache = cache
    view.column_projection = mod.LogicalColumnProjection.from_model(cache.model)
    view.max_col = max(
        [1]
        + [slot.mine_col or 0 for slot in cache.model.slots]
        + [slot.theirs_col or 0 for slot in cache.model.slots]
    )
    view._column_diff_seed_last = {}
    return view


def _plan(kind, source, target, source_cols, target_cols, anchor):
    return mod.ColumnBlockActionPlan(
        action_id=f"oracle-{kind}",
        sheet="SeedOracle",
        block_ordinal=1,
        logical_start=2,
        logical_end=2,
        source_side=source,
        target_side=target,
        source_physical_cols=tuple(source_cols),
        target_physical_cols=tuple(target_cols),
        target_physical_anchor=anchor,
        count=1,
        action_kind=kind,
    )


def _exercise_case(name, old_rows_a, old_edit_a, old_rows_b, old_edit_b,
                   new_rows_a, new_edit_a, new_rows_b, new_edit_b, plan):
    # Pair 7 is deliberately one-sided: sentinel {-1} must survive both the
    # apply remap and exact restore. Rows also include formulas and blank cells.
    row_pairs = [(row, row) for row in range(1, 7)] + [(7, None)]
    old_cache = _cache(name + "-old", old_rows_a, old_rows_b)
    new_cache = _cache(name + "-new", new_rows_a, new_rows_b)
    old_books = _worksheets(old_rows_a, old_edit_a) + _worksheets(old_rows_b, old_edit_b)
    new_books = _worksheets(new_rows_a, new_edit_a) + _worksheets(new_rows_b, new_edit_b)
    old_av, old_ae = old_books[2], old_books[3]
    old_bv, old_be = old_books[6], old_books[7]
    new_av, new_ae = new_books[2], new_books[3]
    new_bv, new_be = new_books[6], new_books[7]
    old_exact = _full_map(old_cache, row_pairs, old_av, old_bv, old_ae, old_be)
    new_exact = _full_map(new_cache, row_pairs, new_av, new_bv, new_ae, new_be)
    view = _view(old_cache, row_pairs, old_exact)
    state = view._capture_column_action_diff_seed(plan)
    assert state is not None, view._column_diff_seed_last
    detached_hash = hash(state["pair_diff_bitmap"]["bits"])
    view.col_max_a = len(new_rows_a[0])
    view.col_max_b = len(new_rows_b[0])
    assert view._try_install_column_action_projection_seed(
        {"state": state, "plan": plan, "mode": "apply"}
    ), view._column_projection_seed_last
    seeded_cache = view.column_comparison_cache
    adopted_confidence = mod.ColumnMappingConfidence(
        1.0, False, "exact-anchor", ("intrinsic",)
    )
    normalized_full_slots = []
    for slot in new_cache.model.slots:
        logical_col = slot.logical_idx + 1
        normalized_full_slots.append(mod.ColumnSlot(
            logical_idx=slot.logical_idx,
            mine_col=slot.mine_col,
            base_col=slot.base_col,
            theirs_col=slot.theirs_col,
            state="retained" if plan.logical_start <= logical_col <= plan.logical_end else slot.state,
            confidence=(
                adopted_confidence
                if plan.logical_start <= logical_col <= plan.logical_end
                else slot.confidence
            ),
            base_boundary=slot.base_boundary,
            origin_side=slot.origin_side,
        ))
    normalized_full_slots = tuple(normalized_full_slots)
    assert tuple(seeded_cache.model.slots) == normalized_full_slots, (
        name, seeded_cache.model.slots, normalized_full_slots
    )
    assert tuple(seeded_cache.model.blocks) == mod._build_column_blocks(normalized_full_slots)
    expected_structural = frozenset(
        slot.logical_idx + 1 for slot in normalized_full_slots
        if slot.mine_col is None or slot.theirs_col is None
    )
    expected_unresolved = frozenset(
        slot.logical_idx + 1 for slot in normalized_full_slots
        if slot.state == "unresolved" or slot.confidence.ambiguous
    )
    assert seeded_cache.structural_diff_cols == expected_structural
    assert seeded_cache.unresolved_cols == expected_unresolved
    for side in ("A", "B"):
        for physical_col in range(1, max(view.col_max_a, view.col_max_b) + 1):
            assert view.column_projection.logical_col(side, physical_col) == mod.LogicalColumnProjection.from_model(
                new_cache.model
            ).logical_col(side, physical_col)
    view.column_comparison_cache = new_cache
    view.column_projection = mod.LogicalColumnProjection.from_model(new_cache.model)
    view.max_col = max(len(new_rows_a[0]), len(new_rows_b[0]))
    assert view._try_apply_column_action_diff_seed(
        {"state": state, "plan": plan, "mode": "apply"},
        ws_a_val=new_av,
        ws_b_val=new_bv,
        ws_a_edit=new_ae,
        ws_b_edit=new_be,
    ), view._column_diff_seed_last
    assert view.pair_diff_cols == new_exact, (name, view.pair_diff_cols, new_exact)

    # Simulate only-diff/background replacement; the detached undo bitmap is
    # immutable and must still restore the original full map exactly.
    view.pair_diff_cols.clear()
    view.pair_diff_cols[0] = {1}
    assert hash(state["pair_diff_bitmap"]["bits"]) == detached_hash
    view.column_comparison_cache = old_cache
    view.column_projection = mod.LogicalColumnProjection.from_model(old_cache.model)
    assert view._try_apply_column_action_diff_seed(
        {"state": state, "plan": plan, "mode": "restore"},
        ws_a_val=old_av,
        ws_b_val=old_bv,
        ws_a_edit=old_ae,
        ws_b_edit=old_be,
    )
    assert view.pair_diff_cols == old_exact
    view.col_max_a = len(old_rows_a[0])
    view.col_max_b = len(old_rows_b[0])
    assert view._try_install_column_action_projection_seed(
        {"state": state, "plan": plan, "mode": "restore"}
    )
    assert tuple(view.column_projection.model.slots) == tuple(old_cache.model.slots)


def _fallback_guards():
    rows_a = [("A", "C"), (1, 3)]
    rows_b = [("A", "B", "C"), (1, 2, 3)]
    cache = _cache("fallback", rows_a, rows_b)
    row_pairs = [(1, 1), (2, 2)]
    books_a = _worksheets(rows_a)
    books_b = _worksheets(rows_b)
    exact = _full_map(cache, row_pairs, books_a[2], books_b[2], books_a[3], books_b[3])
    plan = _plan("insert_copy", "B", "A", (2,), (), 2)
    view = _view(cache, row_pairs, exact)
    state = view._capture_column_action_diff_seed(plan)
    assert state is not None
    view.col_max_a = 3
    view.col_max_b = 3
    request = {"state": state, "plan": plan, "mode": "apply"}

    view.app.has_base = True
    assert not view._try_install_column_action_projection_seed(request)
    view.app.has_base = False
    view.app.merge_conflict_mode = True
    assert not view._try_install_column_action_projection_seed(request)
    view.app.merge_conflict_mode = False

    state_not_exact = dict(state, pair_diff_full_exact=False)
    assert not view._try_install_column_action_projection_seed(
        {"state": state_not_exact, "plan": plan, "mode": "apply"}
    )
    view._row_model_version += 1
    assert not view._try_install_column_action_projection_seed(request)
    view._row_model_version -= 1
    unresolved_plan = mod.ColumnBlockActionPlan(
        **{**plan.__dict__, "unresolved": True}
    ) if hasattr(plan, "__dict__") else mod.ColumnBlockActionPlan(
        action_id=plan.action_id,
        sheet=plan.sheet,
        block_ordinal=plan.block_ordinal,
        logical_start=plan.logical_start,
        logical_end=plan.logical_end,
        source_side=plan.source_side,
        target_side=plan.target_side,
        source_physical_cols=plan.source_physical_cols,
        target_physical_cols=plan.target_physical_cols,
        target_physical_anchor=plan.target_physical_anchor,
        count=plan.count,
        action_kind=plan.action_kind,
        unresolved=True,
    )
    assert not view._try_install_column_action_projection_seed(
        {"state": state, "plan": unresolved_plan, "mode": "apply"}
    )


def main():
    _fallback_guards()
    old_a = [("A", "C"), (1, 3), (2, 4), (None, None), (5, 8), (6, 9), (7, 10)]
    old_a_edit = [("A", "C"), (1, 3), (2, 4), (None, None), (5, 8), (6, 9), (7, 10)]
    old_b = [("A", "B", "C"), (1, 2, 3), (2, 3, 4), (None, None, None), (5, 7, 8), (6, 8, 9)]
    old_b_edit = [("A", "B", "C"), (1, 2, 3), (2, "=A3+1", 4), (None, None, None), (5, 7, 8), (6, 8, 9)]
    new_a = old_b + [(7, 9, 10)]
    new_a_edit = old_b_edit + [(7, "=A7+2", 10)]
    _exercise_case(
        "insert", old_a, old_a_edit, old_b, old_b_edit,
        new_a, new_a_edit, old_b, old_b_edit,
        _plan("insert_copy", "B", "A", (2,), (), 2),
    )

    _exercise_case(
        "delete", new_a, new_a_edit, old_a, old_a_edit,
        old_a, old_a_edit, old_a, old_a_edit,
        _plan("delete", "B", "A", (), (2,), 2),
    )

    copy_a = [("A", "B", "C"), (1, 99, 3), (2, 99, 4), (None, None, None), (5, 99, 8), (6, 99, 9), (7, 99, 10)]
    copy_a_edit = [("A", "B", "C"), (1, 99, 3), (2, "=A3+97", 4), (None, None, None), (5, 99, 8), (6, 99, 9), (7, 99, 10)]
    copy_b = [("A", "B", "C"), (1, 2, 3), (2, 3, 4), (None, None, None), (5, 7, 8), (6, 8, 9)]
    copy_b_edit = [("A", "B", "C"), (1, 2, 3), (2, "=A3+1", 4), (None, None, None), (5, 7, 8), (6, 8, 9)]
    copy_new_a = copy_b + [(7, None, 10)]
    copy_new_a_edit = copy_b_edit + [(7, None, 10)]
    _exercise_case(
        "copy", copy_a, copy_a_edit, copy_b, copy_b_edit,
        copy_new_a, copy_new_a_edit, copy_b, copy_b_edit,
        _plan("copy", "B", "A", (2,), (2,), 2),
    )
    print("PERFORMANCE_TEST_COLUMN_DIFF_SEED_ORACLE_OK")


if __name__ == "__main__":
    main()

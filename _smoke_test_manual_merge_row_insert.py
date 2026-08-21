import hashlib
import os
import tempfile
import time
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

import sow_merge_tool as mod


def _make_book(path: str, rows: list[list[object]]):
    wb = Workbook()
    ws = wb.active
    ws.title = "S1"
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx).value = value
    wb.save(path)
    wb.close()


def _sha256(path: str) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _assert_input_sha(paths: dict[str, str], expected: dict[str, str]) -> None:
    actual = {name: _sha256(path) for name, path in paths.items()}
    assert actual == expected, (expected, actual)


def _assert_absent(path: str) -> None:
    assert not os.path.lexists(path), f"manual row insert TemporaryDirectory not removed: {path}"


def _run(root: str, state: dict[str, object]) -> None:
    base = os.path.join(root, "base.xlsx")
    mine = os.path.join(root, "mine.xlsx")
    theirs = os.path.join(root, "theirs.xlsx")
    merged = os.path.join(root, "merged.xlsx")

    # Snapshot comparison treats the first two physical rows as the immutable
    # declaration/type schema.  Keep every side's four-column schema identical
    # and let the unique @id declaration anchor the theirs-only B data row.
    schema = [
        ["id@id", "formula@pm", "literal@pm", "stamp@pm"],
        ["string", "formula", "string", "string"],
    ]
    _make_book(base, schema + [["A", None, None, None], ["C", None, None, None]])
    _make_book(mine, schema + [["A", None, None, None], ["C", None, None, None]])
    _make_book(theirs, schema + [["A", None, None, None], ["B", None, None, None], ["C", None, None, None]])

    for path in (base, mine, theirs):
        wb = load_workbook(path, data_only=False)
        ws = wb["S1"]
        for row in range(3, ws.max_row + 1):
            ws.cell(row=row, column=2).value = "=1"
        if path == theirs:
            ws.row_dimensions[4].height = 27
            ws.cell(row=4, column=1).number_format = "0000"
            ws.cell(row=4, column=1).fill = PatternFill(fill_type="solid", fgColor="00FF00")
            ws.cell(row=4, column=3).value = "=literal-text"
            ws.cell(row=4, column=3).data_type = "s"
            ws.cell(row=4, column=4).value = datetime(2026, 7, 21, 8, 30)
            ws.cell(row=4, column=4).number_format = "yyyy-mm-dd hh:mm"
        wb.save(path)
        wb.close()
    input_paths = {"base": base, "mine": mine, "theirs": theirs}
    state["input_paths"] = input_paths
    state["input_before"] = {name: _sha256(path) for name, path in input_paths.items()}

    # This is a non-interactive structural replay regression.  Preserve the
    # production confirmation UI, but do not let its asynchronous formula
    # cache modal block the test's Tk pump before the action is exercised.
    original_prompt_scheduler = mod.SowMergeApp._schedule_formula_cache_prompt
    mod.SowMergeApp._schedule_formula_cache_prompt = lambda _self: None
    app = None
    primary_error = None
    cleanup_errors = []

    def _cleanup(label, callback):
        try:
            callback()
        except BaseException as exc:
            cleanup_errors.append((label, exc))

    try:
        app = mod.SowMergeApp(
            mine, theirs, merge_mode=True, merged_path=merged, base_path=base,
        )
        # View-only startup intentionally defers editable workbooks.  This
        # regression exercises a structural mutation, so explicitly request
        # the single-owner backend and wait for both its completion and the
        # selected Sheet's exact comparison before touching an operation.
        app._request_edit_preload()
        deadline = time.monotonic() + 30.0
        while time.monotonic() < deadline:
            app.root.update_idletasks()
            app.root.update()
            if app._edit_loaded_event.is_set() and app._edit_workbooks_ready() and app._is_sheet_exact_current("S1"):
                break
            time.sleep(0.01)
        assert app._edit_loaded_event.is_set() and app._edit_workbooks_ready(), "editable backend did not become ready"
        assert app._is_sheet_exact_current("S1"), "selected Sheet did not reach exact-ready"

        view = app.sheet_views.get("S1")
        if view is None:
            app.nb.select(app._sheet_containers["S1"])
            for _ in range(50):
                app.root.update_idletasks()
                app.root.update()
                time.sleep(0.02)
            view = app.sheet_views["S1"]

        view.force_align_var.set(1)
        view._toggle_force_align()
        view.refresh(row_only=None, rescan=True)

        insert_pair = None
        for pair_idx, (ra, rb) in enumerate(view.row_pairs):
            if ra is None and rb == 4 and app.ws_b_val("S1").cell(rb, 1).value == "B":
                insert_pair = pair_idx
                break

        assert insert_pair is not None, f"did not find inserted row: {view.row_pairs}"

        # A Base-side failure must roll back both in-memory insertion and save ops.
        base_edit_ws = app.ws_base_edit("S1")
        original_base_insert = base_edit_ws.insert_rows
        original_showerror = mod.messagebox.showerror

        def _fail_base_insert(*_args, **_kwargs):
            raise RuntimeError("injected base insert failure")

        base_edit_ws.insert_rows = _fail_base_insert
        mod.messagebox.showerror = lambda *_args, **_kwargs: None
        try:
            assert not view._copy_selected_row("B2A", override_pair_idx=insert_pair)
        finally:
            base_edit_ws.insert_rows = original_base_insert
            mod.messagebox.showerror = original_showerror
        assert [app.ws_a_val("S1").cell(row=r, column=1).value for r in range(1, 5)] == ["id@id", "string", "A", "C"]
        assert not app.manual_a_row_ops
        assert not app.manual_a_cell_ops
        assert not app.undo_stack

        assert view._copy_selected_row("B2A", override_pair_idx=insert_pair), "row insert copy failed"

        out = app.build_manual_merge_output_file()
        wb = load_workbook(out, data_only=False)
        try:
            ws = wb["S1"]
            values = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
            assert values == ["id@id", "string", "A", "B", "C"], f"unexpected merged rows: {values}"
            assert ws.cell(row=4, column=2).value == "=1", "inserted identical formula was dropped"
            assert ws.row_dimensions[4].height == 27
            assert ws.cell(row=4, column=1).number_format == "0000"
            assert ws.cell(row=4, column=1).fill.fgColor.rgb in ("FF00FF00", "0000FF00", "00FF00")
            assert ws.cell(row=4, column=3).value == "=literal-text"
            assert ws.cell(row=4, column=3).data_type == "s"
            assert ws.cell(row=4, column=4).value == datetime(2026, 7, 21, 8, 30)
        finally:
            wb.close()
    except BaseException as exc:
        primary_error = exc
        raise
    finally:
        _cleanup(
            "restore formula prompt scheduler",
            lambda: setattr(mod.SowMergeApp, "_schedule_formula_cache_prompt", original_prompt_scheduler),
        )
        if app is not None:
            _cleanup("shutdown app", app._shutdown_root)
        if primary_error is not None:
            for label, exc in cleanup_errors:
                primary_error.add_note(f"secondary cleanup failure [{label}]: {type(exc).__name__}: {exc}")
        elif cleanup_errors:
            label, exc = cleanup_errors[0]
            raise AssertionError(f"manual row insert cleanup failure [{label}]: {exc}") from exc


def main():
    primary_error = None
    state: dict[str, object] = {}
    cleanup_errors = []
    temporary = tempfile.TemporaryDirectory(prefix="sow_manual_row_insert_")
    root_path = temporary.name

    def _cleanup(label, callback):
        try:
            callback()
        except BaseException as exc:
            cleanup_errors.append((label, exc))

    try:
        _run(root_path, state)
    except BaseException as exc:
        primary_error = exc
        raise
    finally:
        input_paths = state.get("input_paths")
        input_before = state.get("input_before")
        if isinstance(input_paths, dict) and isinstance(input_before, dict):
            _cleanup(
                "verify immutable inputs",
                lambda: _assert_input_sha(input_paths, input_before),
            )
        _cleanup("remove own TemporaryDirectory", temporary.cleanup)
        _cleanup("verify TemporaryDirectory removal", lambda: _assert_absent(root_path))
        if primary_error is not None:
            for label, exc in cleanup_errors:
                primary_error.add_note(f"secondary cleanup failure [{label}]: {type(exc).__name__}: {exc}")
        elif cleanup_errors:
            label, exc = cleanup_errors[0]
            raise AssertionError(f"manual row insert cleanup failure [{label}]: {exc}") from exc

    print("SMOKE_MANUAL_ROW_INSERT_OK")


if __name__ == "__main__":
    main()

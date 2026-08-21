"""Regression: structural rows must never carry foreign openpyxl style IDs.

The test uses only disposable workbooks.  It exercises the user-facing 2-way
and 3-way row-copy action, undo/redo, the openpyxl structural replay path, a
failed save attempt followed by retry, and normal-mode reopen of the output.
"""

from __future__ import annotations

import copy
import hashlib
import os
import tempfile
import time

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side

import sow_merge_tool as mod


def _sha256(path: str) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _optional_excel_reopen(path: str) -> None:
    """Use the production hidden COM open/close gate only when requested."""
    if os.environ.get("SOW_RUN_REAL_EXCEL_ROW_REPLAY", "").strip() != "1":
        return
    assert mod._excel_reopen_validate(path), (
        "Excel COM could not open the saved disposable row-replay workbook "
        "without UI; treat this as a package/repair failure"
    )


def _write_book(path: str, *, include_new: bool) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "S1"
    rows = [
        ["id@id", "formula@pm", "link@pm"],
        ["string", "formula", "string"],
        ["A", "=A3&\"-formula\"", "plain"],
    ]
    if include_new:
        rows.append(["NEW", "=A4&\"-formula\"", "https://example.invalid/new"])
    rows.append(["C", "=A5&\"-formula\"" if include_new else "=A4&\"-formula\"", "tail"])
    for row_idx, values in enumerate(rows, start=1):
        for column_idx, value in enumerate(values, start=1):
            worksheet.cell(row_idx, column_idx).value = value
    if include_new:
        cell = worksheet["A4"]
        cell.font = Font(name="Arial", bold=True, italic=True, color="FF112233")
        cell.fill = PatternFill(fill_type="solid", fgColor="FF44AA55")
        cell.border = Border(left=Side(style="thin", color="FF334455"))
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.number_format = "0000.000"
        cell.protection = Protection(locked=False, hidden=True)
        cell.comment = Comment("row metadata comment", "SOW")
        worksheet["C4"].hyperlink = "https://example.invalid/new"
        worksheet["C4"].style = "Hyperlink"
        dim = worksheet.row_dimensions[4]
        dim.height = 29.5
        dim.hidden = True
        dim.outlineLevel = 2
        dim.collapsed = True
        dim.thickTop = True
        dim.thickBot = True
    workbook.save(path)
    workbook.close()


def _assert_metadata(
    worksheet,
    row: int = 4,
    *,
    assert_formula: bool = True,
    expected_link: str = "https://example.invalid/new",
    expected_formula: str = "=A4&\"-formula\"",
) -> None:
    cell = worksheet.cell(row, 1)
    assert cell.font.bold and cell.font.italic and cell.font.color.rgb == "FF112233"
    assert cell.fill.fill_type == "solid" and cell.fill.fgColor.rgb == "FF44AA55"
    assert cell.border.left.style == "thin" and cell.border.left.color.rgb == "FF334455"
    assert cell.alignment.horizontal == "center" and cell.alignment.wrap_text is True
    assert cell.number_format == "0000.000"
    assert cell.protection.locked is False and cell.protection.hidden is True
    assert cell.comment is not None and cell.comment.text == "row metadata comment"
    assert worksheet.cell(row, 3).hyperlink is not None
    assert worksheet.cell(row, 3).hyperlink.target == expected_link
    dimension = worksheet.row_dimensions[row]
    assert dimension.height == 29.5 and dimension.hidden is True
    assert dimension.outlineLevel == 2 and dimension.collapsed is True
    assert dimension.thickTop is True and dimension.thickBot is True
    if assert_formula:
        assert worksheet.cell(row, 2).value == expected_formula


def _wait_exact_ready(app: mod.SowMergeApp):
    app._request_edit_preload()
    view = None
    for _ in range(400):
        app.root.update_idletasks()
        app.root.update()
        view = app.sheet_views.get("S1")
        if (
            view is not None
            and view._data_ready
            and app._is_sheet_exact_current("S1")
            and app._edit_workbooks_ready()
        ):
            return view
        time.sleep(0.02)
    raise AssertionError(f"sheet never reached exact/edit ready: {app._sheet_exact_entry('S1')!r}")


def _finish_cleanup(primary: BaseException | None, errors: list[str]) -> None:
    if not errors:
        return
    detail = "; ".join(errors)
    if primary is not None:
        primary.add_note(f"cleanup failure: {detail}")
        return
    raise AssertionError(f"cleanup failure: {detail}")


def _assert_input_hashes(expected: dict[str, str], errors: list[str]) -> None:
    actual = {path: _sha256(path) for path in expected}
    if actual != expected:
        errors.append(f"input SHA changed: expected={expected!r}, actual={actual!r}")


def _restore_run(
    primary: BaseException | None,
    *,
    app: mod.SowMergeApp | None,
    prompt,
    excel_builder,
    source_hashes: dict[str, str],
) -> None:
    errors: list[str] = []
    try:
        mod.SowMergeApp._schedule_formula_cache_prompt = prompt
        mod._build_manual_merge_output_with_excel = excel_builder
    except BaseException as error:
        errors.append(f"restore patches: {error!r}")
    if app is not None:
        try:
            app._shutdown_root()
        except BaseException as error:
            errors.append(f"shutdown: {error!r}")
    try:
        _assert_input_hashes(source_hashes, errors)
    except BaseException as error:
        errors.append(f"input SHA check: {error!r}")
    _finish_cleanup(primary, errors)


def _copy_inserted_row(view, app: mod.SowMergeApp, direction: str) -> None:
    view.force_align_var.set(1)
    view._toggle_force_align()
    source_is_a = direction == "A2B"
    pair = next(
        index
        for index, (row_a, row_b) in enumerate(view.row_pairs)
        if (
            (row_a == 4 and row_b is None)
            if source_is_a else (row_a is None and row_b == 4)
        )
    )
    assert view._copy_selected_row(direction, override_pair_idx=pair)
    if source_is_a:
        assert app.manual_b_row_ops and app.manual_b_row_ops[-1]["source_side"] == "A"
    else:
        assert app.manual_a_row_ops and app.manual_a_row_ops[-1]["source_side"] == "B"
    view._undo_last_action()
    view._redo_last_action()


def _force_openpyxl_replay(original_excel):
    def builder(src, out, manual_ops, row_ops, sheet_ops=None, source_paths=None, column_ops=None):
        assert not column_ops
        return mod._build_manual_merge_output_with_openpyxl(
            src, out, manual_ops, row_ops, sheet_ops=sheet_ops, source_paths=source_paths,
        )
    mod._build_manual_merge_output_with_excel = builder
    return original_excel


def _direct_cross_workbook_copy_probe(root: str) -> None:
    source = os.path.join(root, "direct-source.xlsx")
    target = os.path.join(root, "direct-target.xlsx")
    _write_book(source, include_new=True)
    _write_book(target, include_new=False)
    source_hash = _sha256(source)
    src_wb = dst_wb = reopened = None
    primary = None
    try:
        src_wb = load_workbook(source, data_only=False)
        dst_wb = load_workbook(target, data_only=False)
        dst_wb["S1"].insert_rows(4)
        mod._copy_row_metadata(src_wb["S1"], dst_wb["S1"], 4, 4, 3)
        # Accessing this used to fail immediately with an unregistered style ID.
        _assert_metadata(dst_wb["S1"], 4, assert_formula=False)
        dst_wb.save(target)
        src_wb.close()
        src_wb = None
        dst_wb.close()
        dst_wb = None
        reopened = load_workbook(target, data_only=False)
        _assert_metadata(reopened["S1"], 4, assert_formula=False)
    except BaseException as error:
        primary = error
        raise
    finally:
        errors: list[str] = []
        for label, workbook in (("source", src_wb), ("target", dst_wb), ("reopened target", reopened)):
            if workbook is None:
                continue
            try:
                workbook.close()
            except BaseException as error:
                errors.append(f"{label} close: {error!r}")
        try:
            _assert_input_hashes({source: source_hash}, errors)
        except BaseException as error:
            errors.append(f"source SHA check: {error!r}")
        _finish_cleanup(primary, errors)


def _write_batch_source(path: str) -> None:
    """Make two consecutive inserted rows, both with rich row metadata."""
    _write_book(path, include_new=True)
    workbook = load_workbook(path, data_only=False)
    try:
        worksheet = workbook["S1"]
        worksheet.insert_rows(5)
        worksheet.cell(5, 1).value = "NEW2"
        worksheet.cell(5, 2).value = "=A5&\"-formula\""
        worksheet.cell(5, 3).value = "https://example.invalid/new2"
        mod._copy_row_metadata(worksheet, worksheet, 4, 5, 3)
        worksheet.cell(5, 3).hyperlink = "https://example.invalid/new2"
        workbook.save(path)
    finally:
        workbook.close()


def _run_two_way_batch(root: str) -> None:
    mine = os.path.join(root, "batch-mine.xlsx")
    theirs = os.path.join(root, "batch-theirs.xlsx")
    _write_batch_source(mine)
    _write_book(theirs, include_new=False)
    mine_hash, theirs_hash = _sha256(mine), _sha256(theirs)
    prompt = mod.SowMergeApp._schedule_formula_cache_prompt
    excel_builder = mod._build_manual_merge_output_with_excel
    mod.SowMergeApp._schedule_formula_cache_prompt = lambda _self: None
    app = None
    primary = None
    try:
        app = mod.SowMergeApp(mine, theirs)
        view = _wait_exact_ready(app)
        view.force_align_var.set(1)
        view._toggle_force_align()
        run = [
            (pair_idx, row_a)
            for pair_idx, (row_a, row_b) in enumerate(view.row_pairs)
            if row_a in (4, 5) and row_b is None
        ]
        assert [row for _pair, row in run] == [4, 5], view.row_pairs
        # Batch primitives are intentionally callable only beneath the same
        # interactive mutation lease that the public region action acquires.
        app._begin_interactive_action(view)
        try:
            assert view._batch_insert_row_copy(run, "A2B", suppress_refresh=False, anchor=None)
        finally:
            app._end_interactive_action()
        assert app.manual_b_row_ops[-1]["source_rows"] == [4, 5]
        view._undo_last_action()
        view._redo_last_action()
        _force_openpyxl_replay(excel_builder)
        output = app.build_manual_b_output_file()
        _optional_excel_reopen(output)
        reopened = load_workbook(output, data_only=False)
        try:
            worksheet = reopened["S1"]
            assert [worksheet.cell(r, 1).value for r in range(1, 7)] == ["id@id", "string", "A", "NEW", "NEW2", "C"]
            _assert_metadata(worksheet, 4)
            _assert_metadata(
                worksheet,
                5,
                expected_link="https://example.invalid/new2",
                expected_formula="=A5&\"-formula\"",
            )
            assert worksheet.cell(5, 2).value == "=A5&\"-formula\""
            assert worksheet.cell(5, 3).hyperlink.target == "https://example.invalid/new2"
        finally:
            reopened.close()
        assert _sha256(mine) == mine_hash and _sha256(theirs) == theirs_hash
    except BaseException as error:
        primary = error
        raise
    finally:
        _restore_run(
            primary, app=app, prompt=prompt, excel_builder=excel_builder,
            source_hashes={mine: mine_hash, theirs: theirs_hash},
        )


def _run_two_way(root: str) -> None:
    mine = os.path.join(root, "two-mine.xlsx")
    theirs = os.path.join(root, "two-theirs.xlsx")
    _write_book(mine, include_new=True)
    _write_book(theirs, include_new=False)
    mine_hash, theirs_hash = _sha256(mine), _sha256(theirs)
    prompt = mod.SowMergeApp._schedule_formula_cache_prompt
    excel_builder = mod._build_manual_merge_output_with_excel
    mod.SowMergeApp._schedule_formula_cache_prompt = lambda _self: None
    app = None
    primary = None
    try:
        app = mod.SowMergeApp(mine, theirs)
        view = _wait_exact_ready(app)
        _copy_inserted_row(view, app, "A2B")
        operations_before_failure = copy.deepcopy(app.manual_b_row_ops)
        cells_before_failure = copy.deepcopy(app.manual_b_cell_ops)
        mod._build_manual_merge_output_with_excel = lambda *_args, **_kwargs: False
        try:
            app.build_manual_b_output_file()
        except RuntimeError:
            pass
        else:
            raise AssertionError("injected structural save failure unexpectedly succeeded")
        assert _sha256(mine) == mine_hash and _sha256(theirs) == theirs_hash
        assert app.manual_b_row_ops == operations_before_failure
        assert app.manual_b_cell_ops == cells_before_failure
        _force_openpyxl_replay(excel_builder)
        output = app.build_manual_b_output_file()
        _optional_excel_reopen(output)
        reopened = load_workbook(output, data_only=False)
        try:
            worksheet = reopened["S1"]
            assert [worksheet.cell(r, 1).value for r in range(1, 6)] == ["id@id", "string", "A", "NEW", "C"]
            _assert_metadata(worksheet)
        finally:
            reopened.close()
        assert _sha256(mine) == mine_hash and _sha256(theirs) == theirs_hash
    except BaseException as error:
        primary = error
        raise
    finally:
        _restore_run(
            primary, app=app, prompt=prompt, excel_builder=excel_builder,
            source_hashes={mine: mine_hash, theirs: theirs_hash},
        )


def _run_three_way(root: str) -> None:
    base = os.path.join(root, "three-base.xlsx")
    mine = os.path.join(root, "three-mine.xlsx")
    theirs = os.path.join(root, "three-theirs.xlsx")
    merged = os.path.join(root, "three-merged.xlsx")
    _write_book(base, include_new=False)
    _write_book(mine, include_new=False)
    _write_book(theirs, include_new=True)
    source_hashes = {path: _sha256(path) for path in (base, mine, theirs)}
    prompt = mod.SowMergeApp._schedule_formula_cache_prompt
    excel_builder = mod._build_manual_merge_output_with_excel
    mod.SowMergeApp._schedule_formula_cache_prompt = lambda _self: None
    app = None
    primary = None
    try:
        app = mod.SowMergeApp(mine, theirs, merge_mode=True, merged_path=merged, base_path=base)
        view = _wait_exact_ready(app)
        _copy_inserted_row(view, app, "B2A")
        _force_openpyxl_replay(excel_builder)
        output = app.build_manual_merge_output_file()
        _optional_excel_reopen(output)
        reopened = load_workbook(output, data_only=False)
        try:
            worksheet = reopened["S1"]
            assert [worksheet.cell(r, 1).value for r in range(1, 6)] == ["id@id", "string", "A", "NEW", "C"]
            _assert_metadata(worksheet)
        finally:
            reopened.close()
        assert {path: _sha256(path) for path in source_hashes} == source_hashes
    except BaseException as error:
        primary = error
        raise
    finally:
        _restore_run(
            primary, app=app, prompt=prompt, excel_builder=excel_builder,
            source_hashes=source_hashes,
        )


def main() -> None:
    temporary = tempfile.TemporaryDirectory(prefix="sow_cross_workbook_row_style_")
    root = temporary.name
    primary = None
    try:
        _direct_cross_workbook_copy_probe(root)
        _run_two_way_batch(root)
        _run_two_way(root)
        _run_three_way(root)
    except BaseException as error:
        primary = error
        raise
    finally:
        errors: list[str] = []
        try:
            temporary.cleanup()
        except BaseException as error:
            errors.append(f"temporary cleanup: {error!r}")
        if os.path.lexists(root):
            errors.append(f"temporary root remains: {root!r}")
        _finish_cleanup(primary, errors)
    print("SMOKE_CROSS_WORKBOOK_ROW_STYLE_REPLAY_OK")


if __name__ == "__main__":
    main()

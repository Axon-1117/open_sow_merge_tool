"""Recoverable multi-branch SVN submission workbench for Excel configs.

The workflow deliberately keeps every repository-changing action behind a
visible TortoiseSVN dialog.  This module owns discovery, semantic projection,
write-ahead recovery, and post-dialog reconciliation; it never treats a
TortoiseProc exit code as proof that a commit happened.
"""

from __future__ import annotations

import copy
import hashlib
import json
import os
import queue
import re
import shutil
import sqlite3
import subprocess
import sys
import tempfile
import threading
import time
import uuid
from dataclasses import asdict, dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Callable, Iterable

from openpyxl import load_workbook

from svn_status_provider import (
    SvnStatusError,
    SvnStatusRecord,
    record_for_path,
    records_by_path,
    scan_status,
)


SUPPORTED_EXTENSIONS = (".xlsx",)
DEFAULT_BRANCHES = ("develop", "release", "sandbox")
TERMINAL_ACTION_STATES = {"committed", "already_applied", "restored"}
BLOCKING_NODE_STATES = {"conflicted", "obstructed", "replaced", "incomplete", "status-callback-failed"}
SOURCE_CHANGE_STATES = {"modified", "added", "deleted", "missing", "unversioned"}
STATE_VERSION = 2


def _sha256(path: str) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _safe_copy(src: str, dst: str) -> None:
    os.makedirs(os.path.dirname(os.path.abspath(dst)), exist_ok=True)
    tmp = dst + f".tmp-{os.getpid()}-{uuid.uuid4().hex}"
    shutil.copy2(src, tmp)
    os.replace(tmp, dst)


def _safe_json_write(path: str, payload: dict) -> None:
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    tmp = path + f".tmp-{os.getpid()}-{uuid.uuid4().hex}"
    with open(tmp, "w", encoding="utf-8") as stream:
        json.dump(payload, stream, ensure_ascii=False, indent=2)
    os.replace(tmp, path)


def _artifact_path(root: str, category: str, relative_path: str) -> str:
    return os.path.join(root, category, *relative_path.replace("\\", "/").split("/"))


def _is_within(path: str, parent: str) -> bool:
    try:
        return os.path.normcase(os.path.commonpath((os.path.abspath(path), os.path.abspath(parent)))) == os.path.normcase(os.path.abspath(parent))
    except ValueError:
        return False


def settings_dir() -> str:
    root = os.environ.get("LOCALAPPDATA") or tempfile.gettempdir()
    return os.path.join(root, "SowMergeTool", "branch_submit")


def load_settings() -> dict:
    try:
        with open(os.path.join(settings_dir(), "settings.json"), "r", encoding="utf-8") as stream:
            value = json.load(stream)
        return value if isinstance(value, dict) else {}
    except (OSError, ValueError):
        return {}


def save_settings(data: dict) -> None:
    _safe_json_write(os.path.join(settings_dir(), "settings.json"), data)


@dataclass
class _WcNode:
    local_relpath: str
    repos_id: int | None
    repos_path: str
    revision: int | None
    presence: str
    kind: str
    moved_here: bool = False
    moved_to: str = ""
    changed_revision: int | None = None
    repo_root: str = ""
    repo_uuid: str = ""


def _wc_db_path(wc_root: str) -> str:
    return os.path.join(os.path.abspath(wc_root), ".svn", "wc.db")


def _node_for_path(wc_root: str, path: str) -> _WcNode | None:
    rel = os.path.relpath(os.path.abspath(path), os.path.abspath(wc_root)).replace("\\", "/")
    if rel == ".":
        rel = ""
    try:
        with sqlite3.connect(f"file:{_wc_db_path(wc_root)}?mode=ro", uri=True) as conn:
            row = conn.execute(
                """
                select n.local_relpath, n.repos_id, coalesce(n.repos_path,''), n.revision,
                       n.presence, n.kind, coalesce(n.moved_here,0), coalesce(n.moved_to,''),
                       n.changed_revision, coalesce(r.root,''), coalesce(r.uuid,'')
                from NODES n left join REPOSITORY r on r.id=n.repos_id
                where n.local_relpath=?
                order by n.op_depth desc limit 1
                """,
                (rel,),
            ).fetchone()
    except sqlite3.Error as exc:
        raise RuntimeError(f"无法读取 SVN 工作副本数据库：{exc}") from exc
    if not row:
        return None
    return _WcNode(
        local_relpath=str(row[0] or ""), repos_id=row[1], repos_path=str(row[2] or ""),
        revision=int(row[3]) if row[3] is not None else None,
        presence=str(row[4] or ""), kind=str(row[5] or ""), moved_here=bool(row[6]),
        moved_to=str(row[7] or ""), changed_revision=int(row[8]) if row[8] is not None else None,
        repo_root=str(row[9] or "").rstrip("/"), repo_uuid=str(row[10] or ""),
    )


def repository_metadata(wc_root: str) -> tuple[str, str]:
    try:
        with sqlite3.connect(f"file:{_wc_db_path(wc_root)}?mode=ro", uri=True) as conn:
            row = conn.execute("select root, uuid from REPOSITORY order by id limit 1").fetchone()
    except sqlite3.Error as exc:
        raise RuntimeError(f"无法读取 SVN 仓库信息：{exc}") from exc
    if not row or not row[0] or not row[1]:
        raise RuntimeError("SVN 工作副本缺少仓库 URL 或 UUID")
    return str(row[0]).rstrip("/"), str(row[1])


@dataclass
class BranchCandidate:
    name: str
    path: str
    url: str
    repo_root: str
    repo_uuid: str
    enabled: bool = True
    reason: str = ""
    favorite: bool = False
    last_changed_at: float = 0.0


def _contains_excel(path: str) -> bool:
    for root, dirs, files in os.walk(path):
        dirs[:] = [
            name for name in dirs
            if name != ".svn" and not os.path.islink(os.path.join(root, name))
        ]
        if any(name.lower().endswith(SUPPORTED_EXTENSIONS) and not name.startswith("~$") for name in files):
            return True
    return False


def _branch_last_changed_times(wc_root: str, repo_root: str, repo_uuid: str) -> dict[str, float]:
    """Return each branch's newest SVN changed_date as a Unix timestamp."""
    try:
        with sqlite3.connect(f"file:{_wc_db_path(wc_root)}?mode=ro", uri=True) as conn:
            rows = conn.execute(
                """
                select case
                         when instr(n.local_relpath, '/')=0 then n.local_relpath
                         else substr(n.local_relpath, 1, instr(n.local_relpath, '/')-1)
                       end as branch_name,
                       max(coalesce(n.changed_date, 0))
                from NODES n join REPOSITORY r on r.id=n.repos_id
                where n.local_relpath<>'' and n.op_depth=0 and n.presence='normal'
                  and coalesce(n.file_external,0)=0 and r.root=? and r.uuid=?
                group by branch_name
                """,
                (repo_root, repo_uuid),
            ).fetchall()
    except sqlite3.Error:
        return {}
    return {
        str(name): (float(changed_date) / 1_000_000.0)
        for name, changed_date in rows if name and changed_date
    }


def discover_branch_candidates(wc_root: str, *, favorites: Iterable[str] = ()) -> list[BranchCandidate]:
    root = os.path.abspath(wc_root)
    if not os.path.isfile(_wc_db_path(root)):
        raise ValueError(f"不是 SVN 工作副本根目录：{root}")
    favorite_set = {str(item) for item in favorites}
    expected_root, expected_uuid = repository_metadata(root)
    changed_times = _branch_last_changed_times(root, expected_root, expected_uuid)
    candidates: list[BranchCandidate] = []
    for entry in os.scandir(root):
        if not entry.is_dir(follow_symlinks=False) or entry.name.startswith(".") or entry.is_symlink():
            continue
        node = _node_for_path(root, entry.path)
        if (
            not node or node.kind != "dir" or node.presence != "normal"
            or node.repo_root != expected_root or node.repo_uuid != expected_uuid
        ):
            continue
        if not _contains_excel(entry.path):
            continue
        url = node.repo_root + ("/" + node.repos_path.strip("/") if node.repos_path else "")
        candidates.append(BranchCandidate(
            name=entry.name, path=os.path.abspath(entry.path), url=url,
            repo_root=node.repo_root, repo_uuid=node.repo_uuid,
            favorite=entry.name in favorite_set,
            last_changed_at=changed_times.get(entry.name, entry.stat(follow_symlinks=False).st_mtime),
        ))
    candidates.sort(key=lambda item: (-item.last_changed_at, item.name.lower()))
    return candidates


def discover_branches(wc_root: str, allowed: Iterable[str] | None = None) -> list[str]:
    """Compatibility API returning dynamic, versioned branch directory names."""
    try:
        names = [item.name for item in discover_branch_candidates(wc_root)]
    except (RuntimeError, ValueError, sqlite3.Error):
        # Tiny legacy unit fixtures contain only ``.svn`` without a wc.db.
        names = sorted(
            entry.name for entry in os.scandir(wc_root)
            if entry.is_dir() and entry.name in DEFAULT_BRANCHES
        )
    if allowed is not None:
        allowed_set = {str(item).strip() for item in allowed if str(item).strip()}
        names = [name for name in names if name in allowed_set]
    return sorted(dict.fromkeys(names), key=lambda item: (item != "develop", item.lower()))


def _validate_branch_name(branch: str, branches: Iterable[str]) -> str:
    value = str(branch or "").strip()
    if value not in set(branches):
        raise ValueError(f"分支不在已验证候选中：{value or '<empty>'}")
    return value


def _validate_relative_file(relative_path: str) -> str:
    value = str(relative_path or "").replace("\\", "/").strip("/")
    if not value or value.startswith("../") or "/../" in f"/{value}/":
        raise ValueError(f"非法配置相对路径：{relative_path!r}")
    if Path(value).suffix.lower() not in SUPPORTED_EXTENSIONS:
        raise ValueError(f"只支持 .xlsx：{relative_path}")
    return value


@dataclass
class BranchContext:
    wc_root: str
    source_branch: str
    scope_path: str
    initial_paths: list[str] = field(default_factory=list)


def _find_wc_root(path: str) -> str:
    probe = os.path.abspath(path if os.path.isdir(path) else os.path.dirname(path))
    while probe:
        if os.path.isfile(os.path.join(probe, ".svn", "wc.db")):
            return probe
        parent = os.path.dirname(probe)
        if parent == probe:
            break
        probe = parent
    raise ValueError(f"路径不在 SVN 工作副本中：{path}")


def infer_context(initial_paths: Iterable[str]) -> BranchContext:
    paths = [os.path.abspath(str(path)) for path in initial_paths if str(path).strip()]
    if not paths:
        raise ValueError("没有可用的右键路径")
    first = paths[0]
    if not os.path.exists(first):
        raise ValueError(f"右键路径不存在：{first}")
    wc_root = _find_wc_root(first)
    source = ""
    for path in paths:
        if not os.path.exists(path):
            raise ValueError(f"右键路径不存在：{path}")
        current_root = _find_wc_root(path)
        rel = os.path.relpath(path, current_root).replace("\\", "/")
        parts = rel.split("/")
        if not parts or parts[0] in ("", "."):
            raise ValueError("请在具体分支目录内启动多分支提交")
        current_source = parts[0]
        if not source:
            source = current_source
        if os.path.normcase(current_root) != os.path.normcase(wc_root) or current_source != source:
            raise ValueError("右键路径必须位于同一 SVN 工作副本和同一源分支")
    candidates = discover_branch_candidates(wc_root)
    _validate_branch_name(source, [item.name for item in candidates])
    scope = first if os.path.isdir(first) else os.path.dirname(first)
    if any(not _is_within(path, scope) for path in paths):
        scope = os.path.commonpath(paths)
        if os.path.isfile(scope):
            scope = os.path.dirname(scope)
    return BranchContext(wc_root=wc_root, source_branch=source, scope_path=os.path.abspath(scope), initial_paths=paths)


def infer_context_from_files(initial_paths: Iterable[str]) -> tuple[str, str, list[str]]:
    context = infer_context(initial_paths)
    files = [path for path in context.initial_paths if os.path.isfile(path)]
    return context.wc_root, context.source_branch, files


@dataclass
class SvnChangeItem:
    path: str
    relative_path: str
    extension: str
    node_kind: str
    node_status: str
    text_status: str
    prop_status: str
    versioned: bool
    conflicted: bool = False
    switched: bool = False
    file_external: bool = False
    wc_locked: bool = False
    lock_owner: str = ""
    changelist: str = ""
    moved_from: str = ""
    moved_to: str = ""
    revision: int | None = None
    checked: bool = False
    selectable: bool = False
    reason: str = ""


def _record_reason(record: SvnStatusRecord, extension: str) -> str:
    if record.conflicted or record.node_status == "conflicted" or record.prop_status == "conflicted":
        return "存在 SVN 冲突"
    if record.node_status in BLOCKING_NODE_STATES:
        return f"不支持的 SVN 状态：{record.node_status}"
    if record.switched:
        return "路径已 switched，禁止自动跨分支提交"
    if record.file_external or record.node_status == "external":
        return "svn:externals 不进入批次"
    if record.prop_status not in {"none", "normal"}:
        return f"存在属性修改：{record.prop_status}"
    if extension not in SUPPORTED_EXTENSIONS:
        return "仅显示，不支持提交此文件类型"
    if os.path.basename(record.path).startswith("~$"):
        return "Excel 临时锁文件不可提交"
    if record.node_kind == "dir":
        return "目录仅显示，不作为 Excel 提交项"
    if record.node_status not in SOURCE_CHANGE_STATES:
        return f"不是可提交的 Excel 变更：{record.node_status}"
    return ""


def scan_changes(wc_root: str, source_branch: str, scope_path: str, *, cancel_event: threading.Event | None = None) -> list[SvnChangeItem]:
    branch_root = os.path.abspath(os.path.join(wc_root, source_branch))
    scope = os.path.abspath(scope_path)
    if not _is_within(scope, branch_root):
        raise ValueError("扫描范围必须位于源分支内")
    records = scan_status(scope, cancel_event=cancel_event)
    items: list[SvnChangeItem] = []
    for record in records:
        if not _is_within(record.path, scope) or not _is_within(record.path, branch_root):
            continue
        if record.node_status in {"ignored", "external"} or record.file_external:
            continue
        relative = os.path.relpath(record.path, branch_root).replace("\\", "/")
        extension = Path(record.path).suffix.lower()
        reason = _record_reason(record, extension)
        selectable = not reason and extension in SUPPORTED_EXTENSIONS and record.node_kind != "dir"
        checked = selectable and record.versioned and record.node_status in {"modified", "added", "deleted", "missing"}
        if record.changelist == "ignore-on-commit":
            checked = False
        items.append(SvnChangeItem(
            path=os.path.abspath(record.path), relative_path=relative, extension=extension,
            node_kind=record.node_kind, node_status=record.node_status,
            text_status=record.text_status, prop_status=record.prop_status,
            versioned=record.versioned, conflicted=record.conflicted,
            switched=record.switched, file_external=record.file_external,
            wc_locked=record.wc_locked, lock_owner=record.lock_owner,
            changelist=record.changelist, moved_from=record.moved_from,
            moved_to=record.moved_to, revision=record.revision,
            checked=checked, selectable=selectable, reason=reason,
        ))
    items.sort(key=lambda item: (item.node_kind == "dir", item.relative_path.lower()))
    return items


def read_recent_messages(repo_uuid: str, limit: int = 25) -> list[str]:
    if os.name != "nt" or not repo_uuid:
        return []
    import winreg
    keys = [
        rf"Software\TortoiseSVN\History\commit{repo_uuid}",
        r"Software\TortoiseSVN\History\commit",
    ]
    messages: list[tuple[int, str]] = []
    for key_name in keys:
        try:
            with winreg.OpenKey(winreg.HKEY_CURRENT_USER, key_name) as key:
                index = 0
                while True:
                    try:
                        name, value, _kind = winreg.EnumValue(key, index)
                    except OSError:
                        break
                    index += 1
                    match = re.fullmatch(r"logmsgs(\d+)", str(name), re.IGNORECASE)
                    text = str(value or "").strip()
                    if match and text:
                        messages.append((int(match.group(1)), text))
        except OSError:
            continue
        if messages:
            break
    seen: set[str] = set()
    result: list[str] = []
    for _index, message in sorted(messages, key=lambda pair: pair[0]):
        if message not in seen:
            seen.add(message)
            result.append(message)
        if len(result) >= max(1, int(limit)):
            break
    return result


def _json_atom(value):
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return repr(value)


def _workbook_semantic_digest(path: str) -> str:
    """Hash configuration semantics while ignoring OOXML timestamps/caches."""
    wb = load_workbook(path, read_only=False, data_only=False, keep_links=False)
    try:
        digest = hashlib.sha256()
        for ws in wb.worksheets:
            digest.update(json.dumps(["sheet", ws.title, ws.sheet_state], ensure_ascii=False).encode("utf-8"))
            digest.update(json.dumps(sorted(str(item) for item in ws.merged_cells.ranges), ensure_ascii=False).encode("utf-8"))
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    payload = [cell.coordinate, cell.data_type, _json_atom(cell.value), cell.number_format]
                    digest.update(json.dumps(payload, ensure_ascii=False, sort_keys=True).encode("utf-8"))
        return digest.hexdigest()
    finally:
        wb.close()


def _semantic_equal(left: str, right: str) -> bool:
    try:
        return _workbook_semantic_digest(left) == _workbook_semantic_digest(right)
    except Exception as exc:
        raise RuntimeError(f"Excel 语义比较失败：{exc}") from exc


def _wc_revision(core, path: str) -> int | None:
    root = core._find_svn_wc_root_for_path(path)
    if not root:
        return None
    relative = os.path.relpath(os.path.abspath(path), root).replace("\\", "/")
    revision, _author, _reason = core._wc_node_metadata(root, relative)
    return revision


def _has_conflict(core, path: str) -> bool:
    try:
        return bool(core._detect_svn_conflict_files(path)) or bool(core._has_svn_conflict_artifacts(path))
    except Exception as exc:
        raise RuntimeError(f"SVN 冲突探测失败：{path}：{exc}") from exc


@dataclass
class BatchFileAction:
    branch: str
    relative_path: str
    operation: str
    state: str = "planned"
    reason: str = ""
    target_before_hash: str = ""
    candidate_hash: str = ""
    backup_path: str = ""
    candidate_path: str = ""
    revision_before: int | None = None
    revision_after: int | None = None
    prepared_at: str = ""


@dataclass
class FilePlan:
    relative_path: str
    operation: str = "modify"
    source_before: str = ""
    source_after: str = ""
    source_before_hash: str = ""
    source_after_hash: str = ""
    source_revision: int | None = None
    source_state: str = "planned"
    source_committed_revision: int | None = None
    actions: dict[str, BatchFileAction] = field(default_factory=dict)
    target_summaries: dict[str, dict] = field(default_factory=dict)


@dataclass
class BranchSubmitBatch:
    batch_id: str
    wc_root: str
    source_branch: str
    target_branches: list[str]
    files: list[FilePlan]
    message: str
    scope_path: str = ""
    source_status: str = "pending"
    target_status: dict[str, str] = field(default_factory=dict)
    created_at: str = field(default_factory=lambda: datetime.now().isoformat(timespec="seconds"))
    updated_at: str = ""
    source_revision_after: int | None = None
    error: str = ""
    abandoned: bool = False
    superseded_by: str = ""
    journal: list[dict] = field(default_factory=list)

    @property
    def folder(self) -> str:
        return os.path.join(settings_dir(), "batches", self.batch_id)

    @property
    def state_path(self) -> str:
        return os.path.join(self.folder, "batch.json")

    def event(self, kind: str, **details) -> None:
        self.journal.append({"time": datetime.now().isoformat(timespec="seconds"), "kind": kind, **details})
        self.save()

    def save(self) -> None:
        self.updated_at = datetime.now().isoformat(timespec="seconds")
        payload = asdict(self)
        payload["state_version"] = STATE_VERSION
        _safe_json_write(self.state_path, payload)

    @classmethod
    def load(cls, path: str) -> "BranchSubmitBatch":
        with open(path, "r", encoding="utf-8") as stream:
            payload = json.load(stream)
        version = int(payload.pop("state_version", 1))
        if version != STATE_VERSION:
            raise ValueError(f"批次状态版本不兼容：{version}")
        plans: list[FilePlan] = []
        for raw in payload.pop("files", []):
            actions = {name: BatchFileAction(**action) for name, action in raw.pop("actions", {}).items()}
            plans.append(FilePlan(actions=actions, **raw))
        return cls(files=plans, **payload)


def list_unfinished_batches() -> list[BranchSubmitBatch]:
    root = os.path.join(settings_dir(), "batches")
    if not os.path.isdir(root):
        return []
    result: list[BranchSubmitBatch] = []
    for entry in os.scandir(root):
        state_path = os.path.join(entry.path, "batch.json")
        if not entry.is_dir() or not os.path.isfile(state_path):
            continue
        try:
            batch = BranchSubmitBatch.load(state_path)
        except Exception:
            continue
        if batch.abandoned or batch.superseded_by:
            continue
        complete = batch.source_status == "committed" and all(
            state in {"committed", "already_present"} for state in batch.target_status.values()
        )
        if not complete:
            result.append(batch)
    return sorted(result, key=lambda item: item.updated_at, reverse=True)


def list_corrupt_batch_files() -> list[str]:
    root = os.path.join(settings_dir(), "batches")
    if not os.path.isdir(root):
        return []
    corrupt: list[str] = []
    for entry in os.scandir(root):
        state_path = os.path.join(entry.path, "batch.json")
        if not entry.is_dir() or not os.path.isfile(state_path):
            continue
        try:
            BranchSubmitBatch.load(state_path)
        except Exception:
            corrupt.append(state_path)
    return sorted(corrupt)


def _status_for_exact_path(wc_root: str, path: str, status_map: dict[str, SvnStatusRecord]) -> SvnStatusRecord:
    record = status_map.get(os.path.normcase(os.path.abspath(path)))
    if record:
        return record
    node = _node_for_path(wc_root, path)
    exists = os.path.exists(path)
    if node and node.presence == "normal":
        return SvnStatusRecord(
            path=os.path.abspath(path), node_kind=node.kind, node_status="normal" if exists else "missing",
            text_status="normal" if exists else "none", prop_status="normal",
            versioned=True, revision=node.revision, repos_root_url=node.repo_root,
            repos_uuid=node.repo_uuid, repos_relpath=node.repos_path,
            moved_to=node.moved_to,
        )
    return SvnStatusRecord(
        path=os.path.abspath(path), node_kind="file" if exists else "none",
        node_status="unversioned" if exists else "none", versioned=False,
    )


def _status_block_reason(record: SvnStatusRecord, *, require_clean: bool) -> str:
    if record.conflicted or record.node_status == "conflicted" or record.prop_status == "conflicted":
        return "存在 SVN 冲突"
    if record.node_status in BLOCKING_NODE_STATES:
        return f"不支持的 SVN 状态：{record.node_status}"
    if record.switched:
        return "路径已 switched"
    if record.file_external or record.node_status == "external":
        return "路径属于 svn:externals"
    if record.prop_status not in {"none", "normal"}:
        return f"存在属性修改：{record.prop_status}"
    if require_clean and record.node_status not in {"normal", "none"}:
        return f"目标工作副本不干净：{record.node_status}"
    return ""


class BranchSubmitEngine:
    def __init__(
        self,
        wc_root: str,
        *,
        allowed_branches: Iterable[str] | None = None,
        runner: Callable | None = None,
        status_scanner: Callable[[str], list[SvnStatusRecord]] | None = None,
    ):
        self.wc_root = os.path.abspath(wc_root)
        self.allowed_branches = tuple(allowed_branches) if allowed_branches is not None else None
        self.core = None
        self.runner = runner or self._default_runner
        self.status_scanner = status_scanner or scan_status

    def _load_core(self):
        if self.core is None:
            import sow_merge_tool as core
            self.core = core
        return self.core

    @staticmethod
    def _default_runner(args, *, timeout=300):
        return subprocess.run(args, capture_output=True, text=True, errors="replace", timeout=timeout)

    def _tortoise(self, command: str, paths: list[str], *, message: str | None = None) -> int:
        core = self._load_core()
        exe = core._find_tortoise_proc_exe()
        if not exe or (os.path.dirname(exe) and not os.path.isfile(exe)):
            raise RuntimeError("未找到 TortoiseProc.exe")
        temp_paths: list[str] = []
        args = [exe, f"/command:{command}"]
        if len(paths) == 1:
            args.append(f"/path:{paths[0]}")
        else:
            pathfile = os.path.join(tempfile.gettempdir(), f"sow_pathfile_{uuid.uuid4().hex}.txt")
            with open(pathfile, "w", encoding="utf-16-le", newline="") as stream:
                stream.write("\n".join(paths))
            temp_paths.append(pathfile)
            args.extend((f"/pathfile:{pathfile}", "/deletepathfile"))
        if message is not None:
            logmsg = os.path.join(tempfile.gettempdir(), f"sow_logmsg_{uuid.uuid4().hex}.txt")
            with open(logmsg, "w", encoding="utf-8-sig", newline="") as stream:
                stream.write(message)
            temp_paths.append(logmsg)
            args.append(f"/logmsgfile:{logmsg}")
        args.append("/closeonend:1")
        try:
            result = self.runner(args, timeout=3600 if command == "commit" else 900)
            return int(getattr(result, "returncode", 1))
        finally:
            for path in temp_paths:
                try:
                    os.remove(path)
                except OSError:
                    pass

    def show_log(self, path: str) -> int:
        return self._tortoise("log", [path])

    def show_diff(self, path: str) -> int:
        return self._tortoise("diff", [path])

    def _update(self, paths: list[str]) -> None:
        if paths and self._tortoise("update", sorted(dict.fromkeys(paths))) != 0:
            raise RuntimeError("SVN update 未成功或被取消")

    def _source_snapshot(self, batch: BranchSubmitBatch, item: SvnChangeItem) -> FilePlan:
        core = self._load_core()
        relative = _validate_relative_file(item.relative_path)
        source_path = os.path.join(batch.wc_root, batch.source_branch, *relative.split("/"))
        if item.reason or not item.selectable:
            raise RuntimeError(f"{relative}：{item.reason or '不可提交'}")
        if item.moved_from or item.moved_to:
            raise RuntimeError(f"{relative}：检测到移动/重命名，请先在 TortoiseSVN 中 Repair Move")
        if item.node_status == "modified":
            operation = "modify"
        elif item.node_status in {"added", "unversioned"}:
            operation = "add"
        elif item.node_status in {"deleted", "missing"}:
            operation = "delete"
        else:
            raise RuntimeError(f"{relative}：不支持的源状态 {item.node_status}")
        plan = FilePlan(relative_path=relative, operation=operation, source_revision=item.revision)
        if operation in {"modify", "delete"}:
            before = core._try_export_svn_base_from_working_copy(source_path)
            if not before:
                raise RuntimeError(f"{relative}：无法读取源文件 SVN pristine")
            plan.source_before = _artifact_path(batch.folder, "source-before", relative)
            _safe_copy(before, plan.source_before)
            plan.source_before_hash = _sha256(plan.source_before)
        if operation in {"modify", "add"}:
            if not os.path.isfile(source_path):
                raise RuntimeError(f"{relative}：源文件不存在")
            plan.source_after = _artifact_path(batch.folder, "source-after", relative)
            _safe_copy(source_path, plan.source_after)
            plan.source_after_hash = _sha256(plan.source_after)
        return plan

    @staticmethod
    def _detect_rename_pairs(plans: list[FilePlan]) -> None:
        added = [plan for plan in plans if plan.operation == "add"]
        deleted = [plan for plan in plans if plan.operation == "delete"]
        for old in deleted:
            for new in added:
                if _semantic_equal(old.source_before, new.source_after):
                    raise RuntimeError(
                        f"检测到高置信度重命名：{old.relative_path} → {new.relative_path}；"
                        "请使用 TortoiseSVN Repair Move 保留历史"
                    )

    def _preflight_target_action(
        self,
        batch: BranchSubmitBatch,
        plan: FilePlan,
        target: str,
        status_map: dict[str, SvnStatusRecord],
    ) -> BatchFileAction:
        core = self._load_core()
        target_path = os.path.join(batch.wc_root, target, *plan.relative_path.split("/"))
        record = _status_for_exact_path(batch.wc_root, target_path, status_map)
        action = BatchFileAction(branch=target, relative_path=plan.relative_path, operation=plan.operation)
        reason = _status_block_reason(record, require_clean=plan.operation != "add" or record.versioned)
        if reason:
            action.state, action.reason = "blocked", reason
            return action
        if _has_conflict(core, target_path):
            action.state, action.reason = "blocked", "目标文件存在 SVN 冲突或冲突残留"
            return action
        action.revision_before = record.revision
        if plan.operation == "add":
            if os.path.isfile(target_path):
                if record.versioned and record.node_status == "normal" and _semantic_equal(plan.source_after, target_path):
                    action.state = "already_applied"
                    return action
                action.state, action.reason = "blocked", "目标路径已存在且内容不同"
                return action
            parent = os.path.dirname(target_path)
            parent_node = _node_for_path(batch.wc_root, parent)
            parent_record = _status_for_exact_path(batch.wc_root, parent, status_map)
            parent_reason = _status_block_reason(parent_record, require_clean=True)
            if not parent_node or parent_node.kind != "dir" or parent_node.presence != "normal" or parent_reason:
                action.state, action.reason = "blocked", parent_reason or "目标父目录未版本化"
                return action
            candidate = _artifact_path(batch.folder, os.path.join("candidates", target), plan.relative_path)
            _safe_copy(plan.source_after, candidate)
            action.candidate_path, action.candidate_hash = candidate, _sha256(candidate)
            return action
        if not os.path.isfile(target_path):
            if plan.operation == "delete" and not record.versioned:
                action.state = "already_applied"
            else:
                action.state, action.reason = "blocked", "目标文件不存在"
            return action
        if not record.versioned or record.node_status != "normal":
            action.state, action.reason = "blocked", f"目标文件不干净：{record.node_status}"
            return action
        action.target_before_hash = _sha256(target_path)
        if plan.operation == "delete":
            if not _semantic_equal(plan.source_before, target_path):
                action.state, action.reason = "blocked", "目标文件已有独立内容变化，不能安全删除"
            return action
        conflicts, candidate, _mapping, summary, merge_reason = core._cross_branch_source_delta_premerge(
            plan.source_before, target_path, plan.source_after
        )
        plan.target_summaries[target] = dict(summary)
        if merge_reason or conflicts:
            action.state, action.reason = "blocked", merge_reason or f"存在 {len(conflicts)} 个待解决冲突"
            return action
        if not summary.get("applied_count", 0):
            action.state = "already_applied"
            return action
        candidate_copy = _artifact_path(batch.folder, os.path.join("candidates", target), plan.relative_path)
        _safe_copy(candidate, candidate_copy)
        action.candidate_path, action.candidate_hash = candidate_copy, _sha256(candidate_copy)
        return action

    def preflight(
        self,
        source_branch: str,
        target_branches: Iterable[str],
        selected_files: Iterable[SvnChangeItem | str],
        message: str,
        *,
        scope_path: str | None = None,
    ) -> BranchSubmitBatch:
        candidates = discover_branch_candidates(self.wc_root)
        branch_names = [item.name for item in candidates]
        if self.allowed_branches is not None:
            branch_names = [name for name in branch_names if name in self.allowed_branches]
        source = _validate_branch_name(source_branch, branch_names)
        targets: list[str] = []
        for raw in target_branches:
            target = _validate_branch_name(raw, branch_names)
            if target == source:
                raise ValueError("目标分支不能与源分支相同")
            if target not in targets:
                targets.append(target)
        if not targets:
            raise ValueError("至少选择一个目标分支")
        frozen_message = str(message or "").strip()
        if not frozen_message:
            raise ValueError("SVN 提交说明不能为空")
        raw_selected = list(selected_files)
        if not raw_selected:
            raise ValueError("至少选择一个 Excel 变更")
        if all(isinstance(item, SvnChangeItem) for item in raw_selected):
            items = [item for item in raw_selected if isinstance(item, SvnChangeItem)]
        else:
            paths = [os.path.abspath(str(item)) for item in raw_selected]
            scan_scope = scope_path or os.path.commonpath([path if os.path.isdir(path) else os.path.dirname(path) for path in paths])
            all_items = scan_changes(self.wc_root, source, scan_scope)
            selected_norm = {os.path.normcase(path) for path in paths}
            items = [item for item in all_items if os.path.normcase(item.path) in selected_norm]
        batch = BranchSubmitBatch(
            batch_id=datetime.now().strftime("%Y%m%d-%H%M%S") + "-" + uuid.uuid4().hex[:8],
            wc_root=self.wc_root, source_branch=source, target_branches=targets,
            files=[], message=frozen_message,
            scope_path=os.path.abspath(scope_path or os.path.join(self.wc_root, source)),
            target_status={target: "pending" for target in targets},
        )
        try:
            for item in items:
                batch.files.append(self._source_snapshot(batch, item))
            self._detect_rename_pairs(batch.files)
            target_maps: dict[str, dict[str, SvnStatusRecord]] = {}
            for target in targets:
                target_scope = os.path.join(self.wc_root, target)
                target_maps[target] = records_by_path(self.status_scanner(target_scope))
            blocked = False
            for plan in batch.files:
                for target in targets:
                    action = self._preflight_target_action(batch, plan, target, target_maps[target])
                    plan.actions[target] = action
                    if action.state == "blocked":
                        blocked = True
                        batch.target_status[target] = "blocked"
            for target in targets:
                if batch.target_status[target] != "blocked":
                    states = {plan.actions[target].state for plan in batch.files}
                    batch.target_status[target] = "already_present" if states == {"already_applied"} else "ready"
            if blocked:
                batch.error = "预检查未通过：任一不兼容项都会阻止整批提交"
            else:
                batch.source_status = "ready"
            batch.event("preflight", source_status=batch.source_status, target_status=dict(batch.target_status))
            return batch
        except Exception as exc:
            batch.error = str(exc)
            batch.event("preflight-failed", error=batch.error)
            raise

    def _source_status_map(self, batch: BranchSubmitBatch) -> dict[str, SvnStatusRecord]:
        scope = batch.scope_path if _is_within(batch.scope_path, os.path.join(batch.wc_root, batch.source_branch)) else os.path.join(batch.wc_root, batch.source_branch)
        return records_by_path(self.status_scanner(scope))

    def _verify_source_before_commit(self, batch: BranchSubmitBatch) -> None:
        status_map = self._source_status_map(batch)
        for plan in batch.files:
            path = os.path.join(batch.wc_root, batch.source_branch, *plan.relative_path.split("/"))
            record = _status_for_exact_path(batch.wc_root, path, status_map)
            reason = _status_block_reason(record, require_clean=False)
            if reason:
                raise RuntimeError(f"{plan.relative_path}：{reason}")
            if plan.operation == "modify":
                valid = os.path.isfile(path) and _sha256(path) == plan.source_after_hash and record.node_status == "modified"
            elif plan.operation == "add":
                valid = os.path.isfile(path) and _sha256(path) == plan.source_after_hash and record.node_status in {"added", "unversioned"}
            else:
                valid = not os.path.exists(path) and record.node_status in {"deleted", "missing"}
            if not valid:
                raise RuntimeError(f"源文件已偏离预检查结果：{plan.relative_path}")

    def _reconcile_source(self, batch: BranchSubmitBatch) -> tuple[int, int, int]:
        status_map = self._source_status_map(batch)
        committed = pending = unknown = 0
        revisions: list[int] = []
        core = self._load_core()
        for plan in batch.files:
            path = os.path.join(batch.wc_root, batch.source_branch, *plan.relative_path.split("/"))
            record = _status_for_exact_path(batch.wc_root, path, status_map)
            node = _node_for_path(batch.wc_root, path)
            if plan.operation in {"modify", "add"}:
                pristine = core._try_export_svn_base_from_working_copy(path) if os.path.isfile(path) else None
                if (
                    os.path.isfile(path) and record.node_status == "normal" and record.versioned
                    and pristine and _sha256(path) == plan.source_after_hash and _sha256(pristine) == plan.source_after_hash
                ):
                    plan.source_state = "committed"
                elif os.path.isfile(path) and _sha256(path) == plan.source_after_hash and record.node_status in {"modified", "added", "unversioned"}:
                    plan.source_state = "prepared"
                else:
                    plan.source_state = "unknown"
            else:
                if not os.path.exists(path) and (not node or node.presence != "normal") and record.node_status not in {"deleted", "missing"}:
                    plan.source_state = "committed"
                elif not os.path.exists(path) and record.node_status in {"deleted", "missing"}:
                    plan.source_state = "prepared"
                else:
                    plan.source_state = "unknown"
            if plan.source_state == "committed":
                committed += 1
                revision = node.changed_revision if node else None
                plan.source_committed_revision = revision
                if revision:
                    revisions.append(revision)
            elif plan.source_state == "prepared":
                pending += 1
            else:
                unknown += 1
        batch.source_revision_after = max(revisions, default=None)
        return committed, pending, unknown

    def _create_committed_sub_batch(self, batch: BranchSubmitBatch) -> BranchSubmitBatch:
        """Split a source partial commit into one explicitly resumable child."""
        if batch.superseded_by:
            existing = os.path.join(settings_dir(), "batches", batch.superseded_by, "batch.json")
            return BranchSubmitBatch.load(existing)
        committed_plans = [copy.deepcopy(plan) for plan in batch.files if plan.source_state == "committed"]
        if not committed_plans:
            raise RuntimeError("源分支没有可拆分的已提交文件")
        child_id = batch.batch_id + "-committed"
        child_status = {}
        for target in batch.target_branches:
            states = {plan.actions[target].state for plan in committed_plans}
            child_status[target] = "already_present" if states == {"already_applied"} else "ready"
        child = BranchSubmitBatch(
            batch_id=child_id, wc_root=batch.wc_root, source_branch=batch.source_branch,
            target_branches=list(batch.target_branches), files=committed_plans,
            message=batch.message, scope_path=batch.scope_path, source_status="committed",
            target_status=child_status, source_revision_after=batch.source_revision_after,
        )
        child.event("created-from-partial-source", parent_batch=batch.batch_id)
        batch.superseded_by = child.batch_id
        batch.event("partial-source-split", child_batch=child.batch_id, committed_files=len(committed_plans))
        return child

    @staticmethod
    def _has_prepare_intent(batch: BranchSubmitBatch, target: str, relative_path: str) -> bool:
        return any(
            event.get("kind") == "prepare-intent"
            and event.get("target") == target
            and event.get("path") == relative_path
            for event in batch.journal
        )

    def _audit_incomplete_intents(self, batch: BranchSubmitBatch) -> None:
        """Recover the crash window between a filesystem write and state save."""
        changed = False
        for plan in batch.files:
            for target, action in plan.actions.items():
                if action.state != "planned" or not self._has_prepare_intent(batch, target, plan.relative_path):
                    continue
                path = os.path.join(batch.wc_root, target, *plan.relative_path.split("/"))
                if plan.operation in {"modify", "add"} and os.path.isfile(path) and action.candidate_hash and _sha256(path) == action.candidate_hash:
                    action.state = "prepared"; changed = True
                elif plan.operation == "delete" and not os.path.exists(path) and action.backup_path:
                    action.state = "prepared"; changed = True
                elif action.backup_path and os.path.isfile(path) and _sha256(path) == action.target_before_hash:
                    # The intent was durable but the working-copy write never happened.
                    continue
                elif plan.operation == "add" and not os.path.exists(path):
                    continue
                else:
                    action.state, action.reason = "unknown", "写前日志存在，但工作副本无法匹配候选或备份"
                    changed = True
        if changed:
            batch.event("prepare-intents-audited")

    def _fresh_target_action(self, batch: BranchSubmitBatch, plan: FilePlan, target: str, status_map: dict[str, SvnStatusRecord]) -> BatchFileAction:
        action = plan.actions[target]
        target_path = os.path.join(batch.wc_root, target, *plan.relative_path.split("/"))
        record = _status_for_exact_path(batch.wc_root, target_path, status_map)
        # A candidate left by this same batch is the sole non-clean state that
        # can be resumed without an update/reprojection.
        if action.state == "prepared":
            if plan.operation == "delete" and not os.path.exists(target_path):
                return action
            if os.path.isfile(target_path) and action.candidate_hash and _sha256(target_path) == action.candidate_hash:
                return action
            action.state, action.reason = "unknown", "工作副本不再等于本批次候选"
            return action
        reason = _status_block_reason(record, require_clean=plan.operation != "add" or record.versioned)
        if reason:
            action.state, action.reason = "blocked", reason
            return action
        core = self._load_core()
        if plan.operation == "add":
            if os.path.isfile(target_path):
                if record.versioned and record.node_status == "normal" and _semantic_equal(plan.source_after, target_path):
                    action.state = "already_applied"
                else:
                    action.state, action.reason = "blocked", "更新后目标路径已存在且内容不同"
                return action
            candidate = _artifact_path(batch.folder, os.path.join("candidates", target), plan.relative_path)
            _safe_copy(plan.source_after, candidate)
            action.candidate_path, action.candidate_hash = candidate, _sha256(candidate)
            return action
        if plan.operation == "delete":
            if not os.path.exists(target_path) and not record.versioned:
                action.state = "already_applied"
            elif not os.path.isfile(target_path) or record.node_status != "normal" or not _semantic_equal(plan.source_before, target_path):
                action.state, action.reason = "blocked", "更新后目标文件不满足安全删除条件"
            return action
        if not os.path.isfile(target_path) or record.node_status != "normal":
            action.state, action.reason = "blocked", f"更新后目标状态不是 normal：{record.node_status}"
            return action
        conflicts, candidate, _mapping, summary, merge_reason = core._cross_branch_source_delta_premerge(
            plan.source_before, target_path, plan.source_after
        )
        plan.target_summaries[target] = dict(summary)
        if merge_reason or conflicts:
            action.state, action.reason = "blocked", merge_reason or f"更新后出现 {len(conflicts)} 个冲突"
        elif not summary.get("applied_count", 0):
            action.state = "already_applied"
        else:
            candidate_copy = _artifact_path(batch.folder, os.path.join("candidates", target), plan.relative_path)
            _safe_copy(candidate, candidate_copy)
            action.candidate_path, action.candidate_hash = candidate_copy, _sha256(candidate_copy)
        return action

    def _prepare_target_action(self, batch: BranchSubmitBatch, plan: FilePlan, target: str) -> None:
        action = plan.actions[target]
        if action.state in TERMINAL_ACTION_STATES or action.state == "prepared":
            return
        path = os.path.join(batch.wc_root, target, *plan.relative_path.split("/"))
        backup = _artifact_path(batch.folder, os.path.join("backups", target), plan.relative_path)
        if os.path.isfile(path):
            _safe_copy(path, backup)
            action.backup_path = backup
            action.target_before_hash = _sha256(backup)
        action.state = "planned"
        batch.event("prepare-intent", target=target, path=plan.relative_path, operation=plan.operation)
        if plan.operation == "delete":
            os.remove(path)
        else:
            source = action.candidate_path or plan.source_after
            _safe_copy(source, path)
            action.candidate_hash = _sha256(path)
        action.prepared_at = datetime.now().isoformat(timespec="seconds")
        action.state = "prepared"
        batch.event("prepared", target=target, path=plan.relative_path, operation=plan.operation)

    def _reconcile_target(self, batch: BranchSubmitBatch, target: str) -> tuple[int, int, int]:
        status_map = records_by_path(self.status_scanner(os.path.join(batch.wc_root, target)))
        committed = pending = unknown = 0
        core = self._load_core()
        for plan in batch.files:
            action = plan.actions[target]
            if action.state in {"already_applied", "restored"}:
                continue
            path = os.path.join(batch.wc_root, target, *plan.relative_path.split("/"))
            record = _status_for_exact_path(batch.wc_root, path, status_map)
            node = _node_for_path(batch.wc_root, path)
            if plan.operation in {"modify", "add"}:
                pristine = core._try_export_svn_base_from_working_copy(path) if os.path.isfile(path) and record.versioned else None
                if (
                    os.path.isfile(path) and record.node_status == "normal" and record.versioned
                    and action.candidate_hash and _sha256(path) == action.candidate_hash
                    and pristine and _sha256(pristine) == action.candidate_hash
                ):
                    action.state = "committed"
                elif os.path.isfile(path) and action.candidate_hash and _sha256(path) == action.candidate_hash and record.node_status in {"modified", "added", "unversioned"}:
                    action.state = "prepared"
                else:
                    action.state, action.reason = "unknown", "提交后内容或 SVN 状态无法与候选对账"
            else:
                if not os.path.exists(path) and (not node or node.presence != "normal") and record.node_status not in {"deleted", "missing"}:
                    action.state = "committed"
                elif not os.path.exists(path) and record.node_status in {"deleted", "missing"}:
                    action.state = "prepared"
                else:
                    action.state, action.reason = "unknown", "删除提交后状态无法确认"
            if action.state == "committed":
                committed += 1
                action.revision_after = node.changed_revision if node else batch.source_revision_after
            elif action.state == "prepared":
                pending += 1
            else:
                unknown += 1
        batch.event("target-reconciled", target=target, committed=committed, pending=pending, unknown=unknown)
        return committed, pending, unknown

    def commit(self, batch: BranchSubmitBatch, *, stop_on_failure: bool = True) -> BranchSubmitBatch:
        if batch.source_status not in {"ready", "committed"}:
            raise RuntimeError(f"批次不可提交：source_status={batch.source_status}")
        source_paths = [os.path.join(batch.wc_root, batch.source_branch, *plan.relative_path.split("/")) for plan in batch.files]
        if batch.source_status == "ready":
            try:
                exit_code = None
                # Reconcile first: the previous process may have died after
                # the server committed but before batch.json was updated.
                committed, pending, unknown = self._reconcile_source(batch)
                if committed == len(batch.files):
                    batch.source_status = "committed"
                    batch.event("source-recovered-after-crash", committed=committed)
                elif committed:
                    batch.source_status = "partial"
                    child = self._create_committed_sub_batch(batch)
                    batch.error = f"源分支只提交了部分文件；已生成可继续的子批次 {child.batch_id}"
                    batch.save()
                    return batch
                elif unknown:
                    batch.source_status = "unknown"
                    batch.error = "源分支提交结果无法确认；已停止传播"
                    batch.save()
                    return batch
                else:
                    self._verify_source_before_commit(batch)
                    batch.event("source-commit-open", paths=len(source_paths))
                    exit_code = self._tortoise("commit", source_paths, message=batch.message)
                    committed, pending, unknown = self._reconcile_source(batch)
                if committed == len(batch.files):
                    batch.source_status = "committed"
                elif committed:
                    batch.source_status = "partial"
                    child = self._create_committed_sub_batch(batch)
                    batch.error = f"源分支只提交了部分文件；已生成可继续的子批次 {child.batch_id}"
                elif unknown:
                    batch.source_status = "unknown"
                    batch.error = "源分支提交结果无法确认；已停止传播"
                else:
                    batch.source_status = "cancelled" if exit_code == 1 else "failed"
                    batch.error = f"源分支没有文件完成提交（TortoiseSVN 退出码 {exit_code}）"
                batch.event("source-reconciled", committed=committed, pending=pending, unknown=unknown, exit_code=exit_code)
                if batch.source_status != "committed":
                    return batch
            except Exception as exc:
                batch.source_status, batch.error = "unknown", str(exc)
                batch.event("source-error", error=batch.error)
                return batch
        footer = f"[MultiBranchSync] batch={batch.batch_id} source={batch.source_branch}@r{batch.source_revision_after or 'unknown'}"
        target_message = batch.message.rstrip() + "\n\n" + footer
        self._audit_incomplete_intents(batch)
        for target in batch.target_branches:
            if batch.target_status.get(target) in {"committed", "already_present"}:
                continue
            try:
                actions = [plan.actions[target] for plan in batch.files]
                nonterminal = [action for action in actions if action.state not in TERMINAL_ACTION_STATES]
                if nonterminal and all(action.state == "prepared" for action in nonterminal):
                    committed, pending, unknown = self._reconcile_target(batch, target)
                    if unknown:
                        batch.target_status[target] = "unknown"
                        batch.error = f"目标分支 {target} 的中断现场无法确认"
                        break
                    if pending == 0:
                        batch.target_status[target] = "committed"
                        batch.event("target-recovered-after-crash", target=target, committed=committed)
                        continue
                resume_ready = any(action.state == "prepared" for action in actions)
                if not resume_ready:
                    update_paths = []
                    for plan in batch.files:
                        target_path = os.path.join(batch.wc_root, target, *plan.relative_path.split("/"))
                        update_paths.append(os.path.dirname(target_path) if plan.operation == "add" else target_path)
                    self._update(update_paths)
                status_map = records_by_path(self.status_scanner(os.path.join(batch.wc_root, target)))
                for plan in batch.files:
                    action = self._fresh_target_action(batch, plan, target, status_map)
                    if action.state in {"blocked", "unknown"}:
                        raise RuntimeError(f"{target}/{plan.relative_path}：{action.reason}")
                for plan in batch.files:
                    self._prepare_target_action(batch, plan, target)
                commit_paths = [
                    os.path.join(batch.wc_root, target, *plan.relative_path.split("/"))
                    for plan in batch.files if plan.actions[target].state == "prepared"
                ]
                if not commit_paths:
                    batch.target_status[target] = "already_present"
                    batch.event("target-already-present", target=target)
                    continue
                batch.event("target-commit-open", target=target, paths=len(commit_paths))
                exit_code = self._tortoise("commit", commit_paths, message=target_message)
                committed, pending, unknown = self._reconcile_target(batch, target)
                active = len([action for action in actions if action.state != "already_applied"])
                if committed == active and unknown == 0:
                    batch.target_status[target] = "committed"
                elif committed:
                    batch.target_status[target] = "partial"
                    batch.error = f"目标分支 {target} 只提交了部分文件，已停止后续分支"
                elif unknown:
                    batch.target_status[target] = "unknown"
                    batch.error = f"目标分支 {target} 提交结果无法确认"
                else:
                    batch.target_status[target] = "cancelled" if exit_code == 1 else "failed"
                    batch.error = f"目标分支 {target} 没有文件完成提交（退出码 {exit_code}）"
                batch.save()
                if batch.target_status[target] != "committed" and stop_on_failure:
                    break
            except Exception as exc:
                batch.target_status[target] = "unknown"
                batch.error = str(exc)
                batch.event("target-error", target=target, error=batch.error)
                if stop_on_failure:
                    break
        return batch

    def restore_uncommitted(self, batch: BranchSubmitBatch) -> BranchSubmitBatch:
        """Restore only target files still proven to be this batch's candidate."""
        self._audit_incomplete_intents(batch)
        for plan in batch.files:
            for target, action in plan.actions.items():
                if action.state != "prepared":
                    continue
                path = os.path.join(batch.wc_root, target, *plan.relative_path.split("/"))
                try:
                    if plan.operation == "modify":
                        if not os.path.isfile(path) or _sha256(path) != action.candidate_hash or not os.path.isfile(action.backup_path):
                            raise RuntimeError("当前文件不再等于候选或备份缺失")
                        _safe_copy(action.backup_path, path)
                    elif plan.operation == "delete":
                        if os.path.exists(path) or not os.path.isfile(action.backup_path):
                            raise RuntimeError("删除现场已变化或备份缺失")
                        _safe_copy(action.backup_path, path)
                    else:
                        if not os.path.isfile(path) or _sha256(path) != action.candidate_hash:
                            raise RuntimeError("新增文件不再等于候选")
                        status = _status_for_exact_path(batch.wc_root, path, records_by_path(self.status_scanner(os.path.dirname(path))))
                        if status.node_status == "added":
                            raise RuntimeError("文件已被 SVN add；请在 TortoiseSVN 中 Undo Add")
                        if status.node_status != "unversioned":
                            raise RuntimeError(f"新增文件状态为 {status.node_status}，不能自动移除")
                        os.remove(path)
                    action.state, action.reason = "restored", ""
                    batch.event("restored", target=target, path=plan.relative_path)
                except Exception as exc:
                    action.state, action.reason = "unknown", str(exc)
                    batch.event("restore-failed", target=target, path=plan.relative_path, error=str(exc))
        if all(action.state in TERMINAL_ACTION_STATES for plan in batch.files for action in plan.actions.values()):
            batch.abandoned = True
            batch.event("restore-complete")
        return batch

    @staticmethod
    def abandon(batch: BranchSubmitBatch) -> None:
        batch.abandoned = True
        batch.event("abandoned")


def _format_batch_result(batch: BranchSubmitBatch) -> str:
    targets = ", ".join(f"{name}={state}" for name, state in batch.target_status.items())
    return f"批次 {batch.batch_id}\n源分支：{batch.source_status}\n目标分支：{targets}\n{batch.error}".strip()


def _choose_recovery_action(root, batches: list[BranchSubmitBatch]):
    import tkinter as tk
    from tkinter import ttk
    result = {"value": None}
    win = tk.Toplevel(root)
    win.title("检测到未完成的多分支提交")
    win.geometry("850x330")
    win.transient(root)
    frame = ttk.Frame(win, padding=12)
    frame.pack(fill="both", expand=True)
    ttk.Label(frame, text="以下批次没有完整结束。继续或恢复前会再次核对 SVN 状态。", font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(0, 8))
    tree = ttk.Treeview(frame, columns=("time", "source", "targets", "state"), show="headings", height=8)
    for key, title, width in (("time", "更新时间", 150), ("source", "源分支", 100), ("targets", "目标分支", 280), ("state", "状态", 220)):
        tree.heading(key, text=title); tree.column(key, width=width, anchor="w")
    for batch in batches:
        tree.insert("", "end", iid=batch.batch_id, values=(batch.updated_at, batch.source_branch, ", ".join(batch.target_branches), batch.source_status))
    tree.pack(fill="both", expand=True)
    if batches:
        tree.selection_set(batches[0].batch_id)
    buttons = ttk.Frame(frame); buttons.pack(fill="x", pady=(10, 0))
    def selected():
        values = tree.selection()
        return next((batch for batch in batches if values and batch.batch_id == values[0]), None)
    def finish(action):
        batch = selected()
        if batch:
            result["value"] = (action, batch)
            win.destroy()
    ttk.Button(buttons, text="继续批次", command=lambda: finish("continue")).pack(side="left")
    ttk.Button(buttons, text="恢复未提交工作副本", command=lambda: finish("restore")).pack(side="left", padx=6)
    ttk.Button(buttons, text="放弃批次", command=lambda: finish("abandon")).pack(side="left")
    ttk.Button(buttons, text="查看批次目录", command=lambda: os.startfile(selected().folder) if selected() else None).pack(side="left", padx=6)
    ttk.Button(buttons, text="稍后处理", command=win.destroy).pack(side="right")
    win.protocol("WM_DELETE_WINDOW", win.destroy)
    win.grab_set(); root.wait_window(win)
    return result["value"]


class BranchSubmitWorkbench:
    """TortoiseSVN-inspired Tk workbench; no repository writes happen here."""

    STATUS_TEXT = {
        "modified": "Modified", "added": "Added", "deleted": "Deleted", "missing": "Missing",
        "unversioned": "Non-versioned", "conflicted": "Conflicted", "normal": "Normal",
    }

    def __init__(self, root, context: BranchContext, *, resume_batch: BranchSubmitBatch | None = None):
        import tkinter as tk
        from tkinter import ttk
        self.tk, self.ttk, self.root = tk, ttk, root
        self.context = context
        self.settings = load_settings()
        favorites = self.settings.get("favorite_branches", list(DEFAULT_BRANCHES))
        self.candidates = discover_branch_candidates(context.wc_root, favorites=favorites)
        self.items: list[SvnChangeItem] = []
        self.scan_generation = 0
        self.scan_cancel = threading.Event()
        self.scan_results: queue.Queue = queue.Queue()
        self.scan_polling = False
        self.closing = False
        self.current_batch = resume_batch
        self.engine = BranchSubmitEngine(context.wc_root)
        self.target_vars: dict[str, object] = {}
        self.source_var = tk.StringVar(value=context.source_branch)
        self.scope_var = tk.StringVar(value=context.scope_path)
        self.repo_url_var = tk.StringVar()
        self.target_search_var = tk.StringVar()
        self.show_unversioned_var = tk.BooleanVar(value=True)
        self.status_var = tk.StringVar(value="正在读取 SVN 状态…")
        self.count_var = tk.StringVar(value="0 个文件")
        self._build_style()
        self._build_ui()
        self._refresh_source_metadata()
        self._rebuild_targets()
        if resume_batch:
            self.message.delete("1.0", tk.END); self.message.insert("1.0", resume_batch.message)
            self._apply_resume_targets(resume_batch)
            self.submit_button.state(["!disabled"])
            self.status_var.set(f"已载入未完成批次 {resume_batch.batch_id}，继续前会重新核对状态")
        self._start_scan()

    def _build_style(self):
        style = self.ttk.Style(self.root)
        try:
            style.theme_use("vista")
        except self.tk.TclError:
            pass
        style.configure("Treeview", rowheight=24, font=("Segoe UI", 9))
        style.configure("Treeview.Heading", font=("Segoe UI", 9, "bold"))
        style.configure("Title.TLabel", font=("Segoe UI", 10, "bold"))

    def _build_ui(self):
        tk, ttk = self.tk, self.ttk
        self.root.title("Excel 合并器 · 多分支 SVN 提交")
        self.root.geometry(self.settings.get("window_geometry", "1120x760"))
        self.root.minsize(900, 620)
        outer = ttk.Frame(self.root, padding=12); outer.pack(fill="both", expand=True)
        info = ttk.LabelFrame(outer, text="提交范围", padding=10); info.pack(fill="x")
        ttk.Label(info, text="源分支").grid(row=0, column=0, sticky="w")
        sources = [item.name for item in self.candidates if item.enabled]
        self.source_box = ttk.Combobox(info, textvariable=self.source_var, values=sources, state="readonly", width=22)
        self.source_box.grid(row=0, column=1, sticky="w", padx=(8, 24)); self.source_box.bind("<<ComboboxSelected>>", self._source_changed)
        ttk.Label(info, text="Commit to").grid(row=0, column=2, sticky="w")
        ttk.Label(info, textvariable=self.repo_url_var, foreground="#124a8a").grid(row=0, column=3, sticky="w", padx=(8, 0))
        ttk.Label(info, text="扫描范围").grid(row=1, column=0, sticky="w", pady=(8, 0))
        ttk.Entry(info, textvariable=self.scope_var, state="readonly").grid(row=1, column=1, columnspan=3, sticky="ew", padx=(8, 0), pady=(8, 0))
        info.columnconfigure(3, weight=1)

        paned = ttk.Panedwindow(outer, orient="horizontal"); paned.pack(fill="both", expand=True, pady=(10, 0))
        target_box = ttk.LabelFrame(paned, text="目标分支", padding=8); paned.add(target_box, weight=1)
        ttk.Entry(target_box, textvariable=self.target_search_var).pack(fill="x")
        self.target_search_var.trace_add("write", lambda *_: self._rebuild_targets())
        quick = ttk.Frame(target_box); quick.pack(fill="x", pady=6)
        ttk.Button(quick, text="全选", command=lambda: self._set_targets("all")).pack(side="left")
        ttk.Button(quick, text="全不选", command=lambda: self._set_targets("none")).pack(side="left", padx=4)
        ttk.Button(quick, text="常用", command=lambda: self._set_targets("favorite")).pack(side="left")
        canvas_holder = ttk.Frame(target_box); canvas_holder.pack(fill="both", expand=True)
        self.target_canvas = tk.Canvas(canvas_holder, highlightthickness=0, width=210)
        target_scroll = ttk.Scrollbar(canvas_holder, orient="vertical", command=self.target_canvas.yview)
        self.target_frame = ttk.Frame(self.target_canvas)
        self.target_window = self.target_canvas.create_window((0, 0), window=self.target_frame, anchor="nw")
        self.target_canvas.configure(yscrollcommand=target_scroll.set)
        self.target_canvas.pack(side="left", fill="both", expand=True); target_scroll.pack(side="right", fill="y")
        self.target_frame.bind("<Configure>", lambda _e: self.target_canvas.configure(scrollregion=self.target_canvas.bbox("all")))
        self.target_canvas.bind("<Configure>", lambda e: self.target_canvas.itemconfigure(self.target_window, width=e.width))

        main = ttk.Frame(paned); paned.add(main, weight=5)
        message_box = ttk.LabelFrame(main, text="提交说明", padding=8); message_box.pack(fill="x")
        message_tools = ttk.Frame(message_box); message_tools.pack(fill="x", pady=(0, 6))
        ttk.Button(message_tools, text="Recent messages", command=self._show_recent_messages).pack(side="left")
        ttk.Button(message_tools, text="粘贴文件名", command=self._paste_filenames).pack(side="left", padx=5)
        ttk.Button(message_tools, text="Show log", command=self._show_log).pack(side="left")
        self.message = tk.Text(message_box, height=5, wrap="word", font=("Segoe UI", 9), undo=True)
        self.message.pack(fill="x")

        changes = ttk.LabelFrame(main, text="Changes made（双击文件查看差异）", padding=8); changes.pack(fill="both", expand=True, pady=(10, 0))
        filters = ttk.Frame(changes); filters.pack(fill="x", pady=(0, 6))
        ttk.Label(filters, text="Check:", style="Title.TLabel").pack(side="left")
        for text, mode in (("All", "all"), ("None", "none"), ("Versioned", "versioned"), ("Added", "added"), ("Deleted", "deleted"), ("Modified", "modified"), ("Files", "files")):
            label = tk.Label(filters, text=text, fg="#0645ad", cursor="hand2", font=("Segoe UI", 9, "underline"))
            label.pack(side="left", padx=(8, 0)); label.bind("<Button-1>", lambda _e, value=mode: self._quick_check(value))
        ttk.Checkbutton(filters, text="Show unversioned files", variable=self.show_unversioned_var, command=self._render_items).pack(side="right")
        self.scan_stop_button = ttk.Button(filters, text="停止扫描", command=self._stop_scan)
        self.scan_stop_button.pack(side="right", padx=6)
        ttk.Button(filters, text="刷新", command=self._start_scan).pack(side="right")
        columns = ("check", "path", "extension", "status", "property", "lock", "switched", "changelist")
        self.tree = ttk.Treeview(changes, columns=columns, show="headings", selectmode="extended")
        headings = {"check":"✓", "path":"Path", "extension":"Extension", "status":"Status", "property":"Property status", "lock":"Lock", "switched":"Switched", "changelist":"Changelist"}
        widths = {"check":38, "path":390, "extension":72, "status":105, "property":105, "lock":80, "switched":70, "changelist":120}
        widths.update({key: int(value) for key, value in self.settings.get("column_widths", {}).items() if key in widths})
        for key in columns:
            self.tree.heading(key, text=headings[key]); self.tree.column(key, width=widths[key], stretch=key == "path", anchor="w" if key != "check" else "center")
        yscroll = ttk.Scrollbar(changes, orient="vertical", command=self.tree.yview)
        xscroll = ttk.Scrollbar(changes, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        self.tree.pack(side="left", fill="both", expand=True); yscroll.pack(side="right", fill="y"); xscroll.pack(side="bottom", fill="x")
        self.tree.bind("<Button-1>", self._tree_click); self.tree.bind("<Double-1>", self._tree_double_click); self.tree.bind("<Button-3>", self._tree_menu)

        bottom = ttk.Frame(outer); bottom.pack(fill="x", pady=(10, 0))
        self.scan_progress = ttk.Progressbar(bottom, mode="indeterminate", length=110)
        self.scan_progress.pack(side="left", padx=(0, 8))
        ttk.Label(bottom, textvariable=self.count_var).pack(side="left")
        ttk.Label(bottom, textvariable=self.status_var, foreground="#555555").pack(side="left", padx=18)
        ttk.Button(bottom, text="取消", command=self.root.destroy).pack(side="right")
        self.submit_button = ttk.Button(bottom, text="开始提交", command=self._submit); self.submit_button.pack(side="right", padx=6); self.submit_button.state(["disabled"])
        self.preflight_button = ttk.Button(bottom, text="预检查", command=self._preflight); self.preflight_button.pack(side="right")
        self.root.protocol("WM_DELETE_WINDOW", self._close)

    def _candidate_for_source(self):
        return next((item for item in self.candidates if item.name == self.source_var.get()), None)

    def _refresh_source_metadata(self):
        candidate = self._candidate_for_source()
        self.repo_url_var.set(candidate.url if candidate else "")

    def _source_changed(self, _event=None):
        source = self.source_var.get()
        self.context.source_branch = source
        self.context.scope_path = os.path.join(self.context.wc_root, source)
        self.scope_var.set(self.context.scope_path)
        self.current_batch = None; self.submit_button.state(["disabled"])
        self._refresh_source_metadata(); self._rebuild_targets(); self._start_scan()

    def _rebuild_targets(self):
        tk, ttk = self.tk, self.ttk
        existing = {name: bool(var.get()) for name, var in self.target_vars.items()}
        remembered = set(self.settings.get("last_targets", {}).get(self.source_var.get(), []))
        for widget in self.target_frame.winfo_children(): widget.destroy()
        self.target_vars = {}
        query = self.target_search_var.get().strip().lower()
        for candidate in self.candidates:
            if candidate.name == self.source_var.get() or (query and query not in candidate.name.lower()):
                continue
            selected = existing.get(candidate.name, candidate.name in remembered)
            var = tk.BooleanVar(value=selected and candidate.enabled)
            self.target_vars[candidate.name] = var
            text = candidate.name + ("  ★" if candidate.favorite else "") + ("  （禁用）" if not candidate.enabled else "")
            row = ttk.Frame(self.target_frame); row.pack(fill="x", pady=1)
            ttk.Checkbutton(row, text=text, variable=var, state="normal" if candidate.enabled else "disabled", command=self._invalidate_batch).pack(side="left", fill="x", expand=True, anchor="w")
            ttk.Button(row, text="★" if candidate.favorite else "☆", width=2, command=lambda name=candidate.name: self._toggle_favorite(name)).pack(side="right")

    def _toggle_favorite(self, name: str):
        favorites = set(self.settings.get("favorite_branches", list(DEFAULT_BRANCHES)))
        if name in favorites: favorites.remove(name)
        else: favorites.add(name)
        self.settings["favorite_branches"] = sorted(favorites)
        for candidate in self.candidates:
            candidate.favorite = candidate.name in favorites
        self._rebuild_targets()

    def _set_targets(self, mode: str):
        candidates = {item.name: item for item in self.candidates}
        for name, var in self.target_vars.items():
            candidate = candidates[name]
            var.set(bool(candidate.enabled and (mode == "all" or (mode == "favorite" and candidate.favorite))))
        self._invalidate_batch()

    def _apply_resume_targets(self, batch: BranchSubmitBatch):
        for name, var in self.target_vars.items(): var.set(name in batch.target_branches)

    def _start_scan(self):
        self.scan_generation += 1; generation = self.scan_generation
        self.scan_cancel.set(); self.scan_cancel = threading.Event()
        source_branch = self.source_var.get()
        scope_path = self.scope_var.get()
        cancel_event = self.scan_cancel
        self.status_var.set("正在递归读取 SVN 状态…"); self.preflight_button.state(["disabled"])
        self.scan_stop_button.state(["!disabled"]); self.scan_progress.start(12)
        def worker():
            try:
                result = scan_changes(self.context.wc_root, source_branch, scope_path, cancel_event=cancel_event)
                error = None
            except Exception as exc:
                result, error = [], str(exc)
            self.scan_results.put((generation, result, error))
        threading.Thread(target=worker, daemon=True, name="branch-submit-status").start()
        if not self.scan_polling:
            self.scan_polling = True
            self.root.after(30, self._poll_scan_results)

    def _poll_scan_results(self):
        if self.closing:
            self.scan_polling = False
            return
        try:
            while True:
                generation, items, error = self.scan_results.get_nowait()
                self._finish_scan(generation, items, error)
        except queue.Empty:
            pass
        self.root.after(30, self._poll_scan_results)

    def _stop_scan(self):
        self.scan_generation += 1
        self.scan_cancel.set()
        self.scan_progress.stop(); self.scan_progress.configure(value=0); self.scan_stop_button.state(["disabled"])
        self.preflight_button.state(["!disabled"])
        self.status_var.set("已停止等待本次扫描结果；可点击刷新重新扫描")

    def _finish_scan(self, generation: int, items: list[SvnChangeItem], error: str | None):
        if generation != self.scan_generation:
            return
        self.scan_progress.stop(); self.scan_progress.configure(value=0); self.scan_stop_button.state(["disabled"])
        self.preflight_button.state(["!disabled"])
        if error:
            self.items = []; self._render_items(); self.status_var.set(error); return
        self.items = items; self._render_items()
        self.status_var.set("状态扫描完成；灰色或带原因的项目不会进入批次")

    def _render_items(self):
        for iid in self.tree.get_children(): self.tree.delete(iid)
        visible = [item for item in self.items if self.show_unversioned_var.get() or item.node_status != "unversioned"]
        for index, item in enumerate(visible):
            mark = "☑" if item.checked else "☐" if item.selectable else "—"
            status = self.STATUS_TEXT.get(item.node_status, item.node_status)
            if item.reason: status += f" · {item.reason}"
            self.tree.insert("", "end", iid=str(self.items.index(item)), values=(mark, item.relative_path, item.extension, status, item.prop_status, item.lock_owner or ("locked" if item.wc_locked else ""), "Yes" if item.switched else "", item.changelist))
        selected = sum(item.checked for item in self.items)
        self.count_var.set(f"{selected} files selected, {len(visible)} files shown, {len(self.items)} total")

    def _tree_click(self, event):
        if self.tree.identify_region(event.x, event.y) != "cell" or self.tree.identify_column(event.x) != "#1": return
        iid = self.tree.identify_row(event.y)
        if not iid: return
        item = self.items[int(iid)]
        if item.selectable:
            item.checked = not item.checked; self._invalidate_batch(); self._render_items()
        return "break"

    def _tree_double_click(self, event):
        iid = self.tree.identify_row(event.y)
        if iid:
            item = self.items[int(iid)]
            if item.node_status == "modified" and os.path.isfile(item.path):
                threading.Thread(target=lambda: self.engine.show_diff(item.path), daemon=True).start()

    def _tree_menu(self, event):
        tk = self.tk; iid = self.tree.identify_row(event.y)
        if not iid: return
        self.tree.selection_set(iid); item = self.items[int(iid)]
        menu = tk.Menu(self.root, tearoff=False)
        menu.add_command(label="查看差异", command=lambda: threading.Thread(target=lambda: self.engine.show_diff(item.path), daemon=True).start(), state="normal" if item.node_status == "modified" else "disabled")
        menu.add_command(label="打开文件", command=lambda: os.startfile(item.path), state="normal" if os.path.isfile(item.path) else "disabled")
        menu.add_command(label="打开所在目录", command=lambda: subprocess.Popen(["explorer.exe", "/select,", item.path] if os.path.exists(item.path) else ["explorer.exe", os.path.dirname(item.path)]))
        menu.add_command(label="复制路径", command=lambda: (self.root.clipboard_clear(), self.root.clipboard_append(item.path)))
        menu.tk_popup(event.x_root, event.y_root)

    def _quick_check(self, mode: str):
        for item in self.items:
            if not item.selectable: continue
            item.checked = mode == "all" or (mode == "versioned" and item.versioned) or (mode == "files" and item.node_kind != "dir") or item.node_status == mode
            if mode == "none": item.checked = False
        self._invalidate_batch(); self._render_items()

    def _invalidate_batch(self):
        self.current_batch = None; self.submit_button.state(["disabled"])

    def _selected_targets(self):
        return [name for name, var in self.target_vars.items() if var.get()]

    def _selected_items(self):
        return [item for item in self.items if item.checked and item.selectable]

    def _show_recent_messages(self):
        tk, ttk = self.tk, self.ttk
        try: _root_url, repo_uuid = repository_metadata(self.context.wc_root); messages = read_recent_messages(repo_uuid)
        except Exception as exc: messages = []; self.status_var.set(str(exc))
        win = tk.Toplevel(self.root); win.title("Recent messages"); win.geometry("720x420"); win.transient(self.root)
        frame = ttk.Frame(win, padding=10); frame.pack(fill="both", expand=True)
        query = tk.StringVar(); ttk.Entry(frame, textvariable=query).pack(fill="x", pady=(0, 6))
        listing = tk.Listbox(frame, exportselection=False); listing.pack(fill="both", expand=True)
        def render(*_):
            listing.delete(0, tk.END); needle=query.get().lower()
            for value in messages:
                if not needle or needle in value.lower(): listing.insert(tk.END, value)
        def use():
            if listing.curselection():
                value=listing.get(listing.curselection()[0]); self.message.delete("1.0", tk.END); self.message.insert("1.0", value); win.destroy()
        query.trace_add("write", render); render(); listing.bind("<Double-1>", lambda _e: use())
        ttk.Button(frame, text="使用所选消息", command=use).pack(anchor="e", pady=(6,0))

    def _paste_filenames(self):
        names = "\n".join(item.relative_path for item in self._selected_items())
        if names: self.message.insert(self.tk.END, ("\n" if self.message.get("1.0", self.tk.END).strip() else "") + names)

    def _show_log(self):
        threading.Thread(target=lambda: self.engine.show_log(self.scope_var.get()), daemon=True).start()

    def _matrix_dialog(self, batch: BranchSubmitBatch) -> bool:
        tk, ttk = self.tk, self.ttk
        result={"ok":False}; win=tk.Toplevel(self.root); win.title("多分支提交预检查"); win.geometry("920x520"); win.transient(self.root)
        frame=ttk.Frame(win,padding=10); frame.pack(fill="both",expand=True)
        ttk.Label(frame,text=f"批次 {batch.batch_id} · 这是一组可恢复的分步提交，不是跨分支原子事务。",style="Title.TLabel").pack(anchor="w",pady=(0,8))
        tree=ttk.Treeview(frame,columns=("branch","file","operation","state","reason"),show="headings")
        for key,title,width in (("branch","目标分支",100),("file","文件",310),("operation","动作",80),("state","状态",110),("reason","说明",260)):
            tree.heading(key,text=title);tree.column(key,width=width,anchor="w")
        for plan in batch.files:
            for target in batch.target_branches:
                action=plan.actions[target]; summary=plan.target_summaries.get(target,{})
                reason=action.reason or (f"应用 {summary.get('applied_count',0)} 项" if plan.operation=="modify" else "")
                tree.insert("","end",values=(target,plan.relative_path,plan.operation,action.state,reason))
        tree.pack(fill="both",expand=True)
        buttons=ttk.Frame(frame);buttons.pack(fill="x",pady=(8,0))
        def accept():result["ok"]=True;win.destroy()
        ttk.Button(buttons,text="关闭",command=win.destroy).pack(side="right")
        if batch.source_status=="ready":ttk.Button(buttons,text="确认预检查结果",command=accept).pack(side="right",padx=6)
        win.grab_set();self.root.wait_window(win);return result["ok"]

    def _preflight(self):
        from tkinter import messagebox
        try:
            selected=self._selected_items(); targets=self._selected_targets(); message=self.message.get("1.0",self.tk.END)
            self.status_var.set("正在生成分支 × 文件预检查矩阵…"); self.root.update_idletasks()
            batch=self.engine.preflight(self.source_var.get(),targets,selected,message,scope_path=self.scope_var.get())
            self.current_batch=batch
            if self._matrix_dialog(batch) and batch.source_status=="ready":
                self.submit_button.state(["!disabled"]);self.status_var.set(f"批次 {batch.batch_id} 已冻结，点击开始提交")
            else:
                self.submit_button.state(["disabled"]);self.status_var.set(batch.error or "预检查结果未确认")
        except Exception as exc:
            self.status_var.set(str(exc));messagebox.showerror("预检查失败",str(exc),parent=self.root)

    def _submit(self):
        from tkinter import messagebox
        batch=self.current_batch
        if not batch:return
        if not messagebox.askyesno("开始分步提交","将依次打开源分支和目标分支的 TortoiseSVN 提交窗口。\n任何取消、部分勾选或未知结果都会停止后续分支。\n\n继续吗？",parent=self.root):return
        self.submit_button.state(["disabled"]);self.status_var.set("等待 TortoiseSVN 提交与逐文件对账…");self.root.update_idletasks()
        result=self.engine.commit(batch);self.current_batch=result
        self.status_var.set(_format_batch_result(result).replace("\n"," · "))
        messagebox.showinfo("批次结果",_format_batch_result(result),parent=self.root)
        if result.superseded_by:
            child_path=os.path.join(settings_dir(),"batches",result.superseded_by,"batch.json")
            if os.path.isfile(child_path) and messagebox.askyesno("载入已提交子批次",f"源分支部分文件已成功提交。\n是否载入子批次 {result.superseded_by}，稍后再次点击“开始提交”传播这些已提交文件？",parent=self.root):
                self.current_batch=BranchSubmitBatch.load(child_path)
                self.submit_button.state(["!disabled"])
                self.status_var.set(f"已载入子批次 {result.superseded_by}；开始提交前仍会逐文件对账")

    def _close(self):
        self.closing = True
        self.scan_cancel.set()
        data=dict(self.settings);data["wc_root"]=self.context.wc_root;data["window_geometry"]=self.root.geometry()
        data["column_widths"]={key:self.tree.column(key,"width") for key in self.tree["columns"]}
        data.setdefault("last_targets",{})[self.source_var.get()]=self._selected_targets()
        data["favorite_branches"]=list(dict.fromkeys(data.get("favorite_branches",[])))
        save_settings(data);self.root.destroy()


def launch_ui(initial_paths: Iterable[str] | None = None) -> None:
    if os.name == "nt":
        try:
            import ctypes
            ctypes.windll.shcore.SetProcessDpiAwareness(1)
        except Exception:
            try:
                ctypes.windll.user32.SetProcessDPIAware()
            except Exception:
                pass
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk
    root = tk.Tk(); root.withdraw()
    resume_batch = None
    corrupt = list_corrupt_batch_files()
    if corrupt:
        messagebox.showwarning(
            "发现损坏的批次状态",
            "以下批次状态无法解析，工具不会猜测或自动处理：\n\n" + "\n".join(corrupt[:8]),
            parent=root,
        )
    unfinished = list_unfinished_batches()
    if unfinished:
        choice = _choose_recovery_action(root, unfinished)
        if choice:
            action, batch = choice
            engine = BranchSubmitEngine(batch.wc_root)
            if action == "restore":
                engine.restore_uncommitted(batch); messagebox.showinfo("恢复结果", _format_batch_result(batch), parent=root)
            elif action == "abandon":
                engine.abandon(batch)
            elif action == "continue":
                resume_batch = batch
    paths = [os.path.abspath(str(path)) for path in (initial_paths or []) if str(path).strip()]
    try:
        if resume_batch:
            context = BranchContext(resume_batch.wc_root, resume_batch.source_branch, resume_batch.scope_path or os.path.join(resume_batch.wc_root, resume_batch.source_branch))
        elif paths:
            context = infer_context(paths)
        else:
            settings = load_settings(); initial = settings.get("wc_root") or os.getcwd()
            folder = filedialog.askdirectory(parent=root, title="选择某个 SVN 分支内的文件夹", initialdir=initial)
            if not folder:
                root.destroy(); return
            context = infer_context([folder])
    except Exception as exc:
        messagebox.showerror("无法打开多分支提交", str(exc), parent=root); root.destroy(); return
    root.deiconify()
    try:
        BranchSubmitWorkbench(root, context, resume_batch=resume_batch)
        root.mainloop()
    except Exception as exc:
        messagebox.showerror("多分支提交启动失败", str(exc), parent=root)
        root.destroy()


def prompt_mode() -> str:
    import tkinter as tk
    from tkinter import ttk
    root=tk.Tk();root.title("Excel 合并器");result={"mode":"legacy"}
    frame=ttk.Frame(root,padding=20);frame.pack(fill="both",expand=True)
    ttk.Label(frame,text="请选择工作模式",font=("Microsoft YaHei",12,"bold")).pack(pady=(0,16))
    def choose(mode):result["mode"]=mode;root.destroy()
    ttk.Button(frame,text="Excel 差异 / 冲突合并",command=lambda:choose("legacy"),width=28).pack(pady=5)
    ttk.Button(frame,text="多分支 SVN 提交",command=lambda:choose("branch"),width=28).pack(pady=5)
    root.protocol("WM_DELETE_WINDOW",root.destroy);root.mainloop();return result["mode"]

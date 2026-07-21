import os
import time

from openpyxl import Workbook, load_workbook

import sow_merge_tool as mod
from _test_temp_utils import make_temp_dir


def _make_book(path: str, sheets: list[tuple[str, list[object]]]):
    wb = Workbook()
    ws = wb.active
    ws.title = sheets[0][0]
    for idx, value in enumerate(sheets[0][1], start=1):
        ws.cell(row=idx, column=1).value = value
    for title, values in sheets[1:]:
        ws2 = wb.create_sheet(title)
        for idx, value in enumerate(values, start=1):
            ws2.cell(row=idx, column=1).value = value
    wb.save(path)
    wb.close()


def _open_sheet(app, sheet: str):
    app.nb.select(app._sheet_containers[sheet])
    for _ in range(60):
        app.root.update_idletasks()
        app.root.update()
        time.sleep(0.02)
    view = app.sheet_views[sheet]
    if view is None:
        raise RuntimeError(f"view missing for {sheet}")
    view.refresh(row_only=None, rescan=True)
    return view


def main():
    root = make_temp_dir("sow_sheet_ops_")

    a = os.path.join(root, "a.xlsx")
    b = os.path.join(root, "b.xlsx")
    out_b = os.path.join(root, "out_b.xlsx")
    _make_book(a, [("Common", ["base"]), ("OnlyA", ["A-only"])])
    _make_book(b, [("Common", ["mine"])])
    app = mod.SowMergeApp(a, b)
    try:
        view = _open_sheet(app, "OnlyA")
        assert view._is_missing_sheet_view()
        assert app.get_sheet_meta("OnlyA").get("view_mode") == "missing_sheet"
        view._copy_missing_sheet("A2B")
        app._ensure_edit_loaded()
        app._atomic_save(app._wb_b_edit, out_b)
        wb = load_workbook(out_b, data_only=False)
        try:
            assert "OnlyA" in wb.sheetnames, wb.sheetnames
            assert wb["OnlyA"]["A1"].value == "A-only"
        finally:
            wb.close()
    finally:
        try:
            app._shutdown_root()
        except Exception:
            pass

    base = os.path.join(root, "base.xlsx")
    mine = os.path.join(root, "mine.xlsx")
    theirs = os.path.join(root, "theirs.xlsx")
    merged = os.path.join(root, "merged.xlsx")
    _make_book(base, [("Common", ["base"]), ("DeleteMe", ["same-as-base"]), ("BaseOnly", ["from-base"])])
    _make_book(mine, [("Common", ["mine"]), ("DeleteMe", ["same-as-base"])])
    _make_book(theirs, [("Common", ["theirs"]), ("Added", ["from-theirs"])])

    app = mod.SowMergeApp(mine, theirs, merge_mode=True, merged_path=merged, base_path=base)
    try:
        out = app.build_manual_merge_output_file()
        wb = load_workbook(out, data_only=False)
        try:
            assert "Added" in wb.sheetnames, wb.sheetnames
            assert "DeleteMe" not in wb.sheetnames, wb.sheetnames
            assert wb["Added"]["A1"].value == "from-theirs"
        finally:
            wb.close()

        base_only_view = _open_sheet(app, "BaseOnly")
        assert base_only_view._is_missing_sheet_view()
        base_only_view._copy_missing_sheet("BASE2A")
        out = app.build_manual_merge_output_file()
        wb = load_workbook(out, data_only=False)
        try:
            assert "BaseOnly" in wb.sheetnames, wb.sheetnames
            assert wb["BaseOnly"]["A1"].value == "from-base"
        finally:
            wb.close()
    finally:
        try:
            app._shutdown_root()
        except Exception:
            pass

    # Mine deleted a base sheet while theirs kept it unchanged: preserve mine's
    # deletion. If theirs modified that sheet, surface a sheet-level conflict.
    base2 = os.path.join(root, "base_mine_delete.xlsx")
    mine2 = os.path.join(root, "mine_delete.xlsx")
    theirs2 = os.path.join(root, "theirs_unchanged.xlsx")
    merged2 = os.path.join(root, "merged_mine_delete.xlsx")
    _make_book(base2, [("Common", ["base"]), ("MineDeleted", ["same"])])
    _make_book(mine2, [("Common", ["mine"])])
    _make_book(theirs2, [("Common", ["theirs"]), ("MineDeleted", ["same"])])
    app = mod.SowMergeApp(mine2, theirs2, merge_mode=True, merged_path=merged2, base_path=base2)
    try:
        assert not any(op.get("sheet") == "MineDeleted" for op in app.auto_sheet_ops), app.auto_sheet_ops
        out = app.build_manual_merge_output_file()
        wb = load_workbook(out, data_only=False)
        try:
            assert "MineDeleted" not in wb.sheetnames, wb.sheetnames
        finally:
            wb.close()
    finally:
        try:
            app._shutdown_root()
        except Exception:
            pass

    theirs3 = os.path.join(root, "theirs_modified.xlsx")
    _make_book(theirs3, [("Common", ["theirs"]), ("MineDeleted", ["changed"])])
    app = mod.SowMergeApp(mine2, theirs3, merge_mode=True, merged_path=merged2, base_path=base2)
    try:
        assert any(
            item.get("kind") == "sheet_deleted_in_mine_modified_in_theirs"
            and item.get("sheet") == "MineDeleted"
            for item in app.sheet_level_conflicts
        ), app.sheet_level_conflicts
        assert not any(op.get("sheet") == "MineDeleted" for op in app.auto_sheet_ops), app.auto_sheet_ops
    finally:
        try:
            app._shutdown_root()
        except Exception:
            pass

    print("SMOKE_SHEET_LEVEL_OPS_OK")


if __name__ == "__main__":
    main()

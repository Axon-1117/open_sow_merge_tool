"""Focused coverage for operations-fixture attribution before ZIP output build."""

from __future__ import annotations

import argparse
import shutil
import tempfile
import time
from pathlib import Path
from types import SimpleNamespace

from openpyxl import Workbook

import _large_sheet_excel_fidelity_gate as gate


_CASE = "fidelity-operations-attribution"
_SHEET = "Data"
_MARKER = "__SOW_FIDELITY_OPERATION_TARGET__"


def _write_book(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _SHEET
    sheet["A1"] = "id@id"
    sheet["B1"] = "value@pm"
    sheet["A2"] = "string"
    sheet["B2"] = "string"
    for row in range(3, 8):
        sheet.cell(row, 1).value = f"row-{row}"
        sheet.cell(row, 2).value = f"value-{row}"
    workbook.save(path)
    workbook.close()


def _test_selector_is_zero_write(root: Path) -> None:
    source = root / "selector.xlsx"
    _write_book(source)
    before = gate._sha256(source)
    target = gate._select_overlay_target(source, _SHEET, _MARKER)
    assert target == gate.MutationTarget("value", 5, 2, _MARKER)
    assert gate._sha256(source) == before


def _test_operations_never_call_mutator(root: Path) -> None:
    original = root / "original.xlsx"
    _write_book(original)
    fixture = SimpleNamespace(name="Synthetic", path=original)
    copy_calls = []
    exercise_calls = []
    original_copy = gate.copy_real_fixture
    original_mutate = gate._mutate_value
    original_value_mutator = gate._MUTATORS_BY_LABEL["value"]
    original_exercise = gate._exercise_overlay_save

    def copy_fixture(item, target_root):
        copy_calls.append(item)
        target = Path(target_root) / "source.xlsx"
        target.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(item.path, target)
        return target, _SHEET

    def forbidden_mutator(*_args, **_kwargs):
        raise AssertionError("operations called _mutate_value")

    def capture_exercise(_root, source, sheet, target, **kwargs):
        exercise_calls.append((Path(source), sheet, target, kwargs))
        return {"captured": True}

    gate.copy_real_fixture = copy_fixture
    gate._mutate_value = forbidden_mutator
    gate._MUTATORS_BY_LABEL["value"] = forbidden_mutator
    gate._exercise_overlay_save = capture_exercise
    try:
        result = gate._run_fixture(
            root / "run", fixture, timeout=90.0, real_excel=False,
            variant_labels=(), include_operations=True,
        )
    finally:
        gate.copy_real_fixture = original_copy
        gate._mutate_value = original_mutate
        gate._MUTATORS_BY_LABEL["value"] = original_value_mutator
        gate._exercise_overlay_save = original_exercise

    assert copy_calls == [fixture]
    assert result["operations"] == {"captured": True}
    assert len(exercise_calls) == 1
    source, sheet, target, kwargs = exercise_calls[0]
    assert sheet == _SHEET and target == gate.MutationTarget("value", 5, 2, _MARKER)
    assert kwargs["absolute_deadline"] is None
    assert gate._sha256(original) == gate._sha256(source)


def _test_prebuild_and_output_reopen_share_deadline(root: Path) -> None:
    source = root / "operations-source.xlsx"
    _write_book(source)
    baseline = gate._sha256(source)
    calls = []
    original_reopen = gate.sm._excel_reopen_validate

    def reopen(path, *, absolute_deadline=None):
        calls.append((Path(path), absolute_deadline))
        return True

    deadline = time.monotonic() + 30.0
    gate.sm._excel_reopen_validate = reopen
    try:
        result = gate._exercise_overlay_save(
            root / "operations", source, _SHEET,
            gate.MutationTarget("value", 5, 2, _MARKER),
            real_excel=True, absolute_deadline=deadline,
        )
    finally:
        gate.sm._excel_reopen_validate = original_reopen

    working = root / "operations" / "operations-source.xlsx"
    output = root / "operations" / "operations-output.xlsx"
    assert result["prebuild_excel_reopen"] == "ok"
    assert gate._sha256(source) == baseline == gate._sha256(working)
    assert calls == [(working, deadline), (output, deadline)]


def _test_prebuild_reopen_failure_stops_builder(root: Path) -> None:
    source = root / "prebuild-failure-source.xlsx"
    _write_book(source)
    calls = []
    original_reopen = gate.sm._excel_reopen_validate
    original_builder = gate.sm._build_manual_merge_xlsx_via_zip

    def reopen(path, *, absolute_deadline=None):
        calls.append((Path(path), absolute_deadline))
        return False

    def forbidden_builder(*_args, **_kwargs):
        raise AssertionError("builder ran after pre-build Excel reopen failure")

    deadline = time.monotonic() + 30.0
    gate.sm._excel_reopen_validate = reopen
    gate.sm._build_manual_merge_xlsx_via_zip = forbidden_builder
    try:
        try:
            gate._exercise_overlay_save(
                root / "prebuild-failure", source, _SHEET,
                gate.MutationTarget("value", 5, 2, _MARKER),
                real_excel=True, absolute_deadline=deadline,
            )
        except AssertionError as exc:
            assert str(exc) == (
                f"Excel cannot reopen pre-build fixture: "
                f"{root / 'prebuild-failure' / 'operations-source.xlsx'}"
            )
        else:
            raise AssertionError("pre-build Excel reopen failure was accepted")
    finally:
        gate.sm._excel_reopen_validate = original_reopen
        gate.sm._build_manual_merge_xlsx_via_zip = original_builder

    working = root / "prebuild-failure" / "operations-source.xlsx"
    assert calls == [(working, deadline)]


def _test_output_reopen_failure_follows_one_builder(root: Path) -> None:
    source = root / "output-failure-source.xlsx"
    _write_book(source)
    calls = []
    builders = []
    original_reopen = gate.sm._excel_reopen_validate
    original_builder = gate.sm._build_manual_merge_xlsx_via_zip

    def reopen(path, *, absolute_deadline=None):
        calls.append((Path(path), absolute_deadline))
        return len(calls) == 1

    def builder(working, output, _manual):
        builders.append((Path(working), Path(output)))
        shutil.copy2(working, output)

    deadline = time.monotonic() + 30.0
    gate.sm._excel_reopen_validate = reopen
    gate.sm._build_manual_merge_xlsx_via_zip = builder
    try:
        try:
            gate._exercise_overlay_save(
                root / "output-failure", source, _SHEET,
                gate.MutationTarget("value", 5, 2, _MARKER),
                real_excel=True, absolute_deadline=deadline,
            )
        except AssertionError as exc:
            output = root / "output-failure" / "operations-output.xlsx"
            assert str(exc) == f"Excel cannot reopen disposable package: {output}"
        else:
            raise AssertionError("output Excel reopen failure was accepted")
    finally:
        gate.sm._excel_reopen_validate = original_reopen
        gate.sm._build_manual_merge_xlsx_via_zip = original_builder

    working = root / "output-failure" / "operations-source.xlsx"
    output = root / "output-failure" / "operations-output.xlsx"
    assert builders == [(working, output)]
    assert calls == [(working, deadline), (output, deadline)]


def run_case() -> None:
    with tempfile.TemporaryDirectory(prefix="sow_fidelity_operations_attribution_") as raw_root:
        root = Path(raw_root)
        _test_selector_is_zero_write(root)
        _test_operations_never_call_mutator(root)
        _test_prebuild_and_output_reopen_share_deadline(root)
        _test_prebuild_reopen_failure_stops_builder(root)
        _test_output_reopen_failure_follows_one_builder(root)


def main(argv=None) -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--case", choices=(_CASE,))
    parser.add_argument("--list-cases", action="store_true")
    args = parser.parse_args(argv)
    if args.list_cases:
        print(_CASE)
        return
    run_case()
    print(f"PASS {_CASE}")


if __name__ == "__main__":
    main()

"""Actual-App ZIP save failure/retry gate on disposable workbooks only."""

from __future__ import annotations

import hashlib
import json
import os
import shutil
import tempfile
import time
from types import SimpleNamespace

from openpyxl import Workbook, load_workbook

import sow_merge_tool as sm


def _sha256(path: str) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _write(path: str, value: str) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "S1"
    worksheet.append(["id@id", "value"])
    worksheet.append(["int", "string"])
    worksheet.append([1, value])
    workbook.save(path)
    workbook.close()


def _pump_operation_ready(app, deadline: float):
    view = None
    while time.monotonic() < deadline:
        app._request_edit_preload()
        app.root.update_idletasks()
        app.root.update()
        view = app.sheet_views.get("S1")
        if (
            view is not None
            and view._data_ready
            and app._is_sheet_exact_current("S1")
            and app._edit_workbooks_ready()
            and view._derive_lifecycle_state() == "READY"
        ):
            return view
        time.sleep(0.01)
    raise AssertionError((
        "S1 did not reach operation readiness",
        app._sheet_exact_entry("S1"),
        app._edit_workbooks_ready(),
    ))


def _is_test_owned_os_temp_output(path: str) -> bool:
    path = os.path.abspath(path)
    temp_dir = os.path.abspath(tempfile.gettempdir())
    return bool(
        os.path.dirname(path) == temp_dir
        and os.path.isfile(path)
        and not os.path.islink(path)
        and os.path.basename(path).startswith(
            f"{sm.APP_NAME}_merged_output_{os.getpid()}_"
        )
    )


def main() -> None:
    temporary = tempfile.TemporaryDirectory(prefix="sow_zip_save_failure_retry_")
    root = temporary.name
    case_deadline = time.monotonic() + 90.0
    original_prompt = sm.SowMergeApp._schedule_formula_cache_prompt
    original_zip = sm._build_manual_merge_xlsx_via_zip
    original_native = sm._EXCEL_NATIVE_SAVE_ON_MERGE
    original_settings_path = sm._SETTINGS_PATH
    original_settings_exists = os.path.lexists(original_settings_path)
    if original_settings_exists:
        with open(original_settings_path, "rb") as stream:
            original_settings_bytes = stream.read()
    else:
        original_settings_bytes = None
    state = SimpleNamespace(app=None, immutable_hashes={}, retry_output=None)
    primary = None
    try:
        settings_path = os.path.join(root, "settings.json")
        with open(settings_path, "w", encoding="utf-8") as stream:
            json.dump({"only_diff": 0}, stream)
        sm._SETTINGS_PATH = settings_path
        sm.SowMergeApp._schedule_formula_cache_prompt = lambda _self: None

        mine = os.path.join(root, "mine.xlsx")
        theirs = os.path.join(root, "theirs.xlsx")
        target = os.path.join(root, "user-target.xlsx")
        _write(mine, "mine")
        _write(theirs, "theirs")
        shutil.copy2(mine, target)
        state.immutable_hashes = {mine: _sha256(mine), theirs: _sha256(theirs)}
        target_hash = _sha256(target)

        state.app = sm.SowMergeApp(mine, theirs)
        app = state.app
        view = _pump_operation_ready(app, case_deadline)
        pair = next(index for index, rows in enumerate(view.row_pairs) if rows == (3, 3))
        assert view._copy_selected_row("B2A", override_pair_idx=pair, override_cols={2})
        recorded = dict(app.manual_a_cell_ops)
        assert recorded == {("S1", 3, 2): "theirs"}

        def fail_zip(*_args, **_kwargs):
            raise RuntimeError("injected ZIP patch failure")

        sm._build_manual_merge_xlsx_via_zip = fail_zip
        sm._EXCEL_NATIVE_SAVE_ON_MERGE = False
        try:
            app.build_manual_merge_output_file()
        except RuntimeError as exc:
            assert "ZIP" in str(exc) or "保存" in str(exc), str(exc)
        else:
            raise AssertionError("injected ZIP failure unexpectedly produced output")
        assert _sha256(mine) == state.immutable_hashes[mine]
        assert _sha256(theirs) == state.immutable_hashes[theirs]
        assert _sha256(target) == target_hash
        assert app.manual_a_cell_ops == recorded

        sm._build_manual_merge_xlsx_via_zip = original_zip
        state.retry_output = app.build_manual_merge_output_file()
        assert _is_test_owned_os_temp_output(state.retry_output)
        app._atomic_replace_file_with_retry(state.retry_output, target, retries=1, delay_sec=0)
        assert _sha256(mine) == state.immutable_hashes[mine]
        assert _sha256(theirs) == state.immutable_hashes[theirs]
        assert _sha256(target) != target_hash
        assert app.manual_a_cell_ops == recorded
        workbook = load_workbook(target, data_only=False, read_only=True)
        try:
            assert workbook["S1"].cell(3, 2).value == "theirs"
        finally:
            workbook.close()
    except BaseException as exc:
        primary = exc
        raise
    finally:
        cleanup_errors = []
        if state.app is not None:
            try:
                for candidate in (state.app, *getattr(state.app, "sheet_views", {}).values()):
                    if candidate is None:
                        continue
                    after_id = getattr(candidate, "_settings_save_id", None)
                    if after_id is not None:
                        state.app.root.after_cancel(after_id)
                        candidate._settings_save_id = None
                state.app._shutdown_root()
            except BaseException as exc:
                cleanup_errors.append(f"shutdown: {exc!r}")
        try:
            sm.SowMergeApp._schedule_formula_cache_prompt = original_prompt
            sm._build_manual_merge_xlsx_via_zip = original_zip
            sm._EXCEL_NATIVE_SAVE_ON_MERGE = original_native
            sm._SETTINGS_PATH = original_settings_path
            if original_settings_exists:
                with open(original_settings_path, "rb") as stream:
                    assert stream.read() == original_settings_bytes
            else:
                assert not os.path.lexists(original_settings_path)
        except BaseException as exc:
            cleanup_errors.append(f"patch/settings restore: {exc!r}")
        for path, before_hash in state.immutable_hashes.items():
            try:
                assert _sha256(path) == before_hash, path
            except BaseException as exc:
                cleanup_errors.append(f"immutable input SHA {path!r}: {exc!r}")
        if state.retry_output is not None:
            try:
                assert _is_test_owned_os_temp_output(state.retry_output), state.retry_output
                os.remove(state.retry_output)
                assert not os.path.lexists(state.retry_output), state.retry_output
            except BaseException as exc:
                cleanup_errors.append(f"retry output cleanup: {exc!r}")
        try:
            temporary.cleanup()
            assert not os.path.lexists(root), root
        except BaseException as exc:
            cleanup_errors.append(f"owned temporary root: {exc!r}")
        if cleanup_errors:
            message = "ZIP save failure/retry cleanup failed: " + "; ".join(cleanup_errors)
            if primary is not None:
                primary.add_note(message)
            else:
                raise AssertionError(message)
    print("SMOKE_ZIP_SAVE_FAILURE_RETRY_OK")


if __name__ == "__main__":
    main()

import json
import os
import threading
import time
import xml.etree.ElementTree as ET
from datetime import datetime
from types import SimpleNamespace

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula
from openpyxl.worksheet.datavalidation import DataValidation

import sow_merge_tool as mod
from _test_temp_utils import make_temp_dir


def _test_only_diff_region_boundaries():
    class DummyView:
        display_rows = [1, 4, 5, 9]
        _diff_blocks_cache = None

        @staticmethod
        def _pair_has_visual_diff(_pair_idx):
            return True

    view = DummyView()
    blocks = mod.SheetView._compute_diff_blocks(view)
    assert blocks == [(1, 1), (2, 3), (4, 4)], blocks


def _test_logical_region_extends_beyond_render_limit():
    class DummyView:
        # The UI rendered only a small window inside a much larger logical block.
        display_rows = [3, 4, 5]
        row_pairs = [(idx + 1, idx + 1) for idx in range(12)]

        @staticmethod
        def _normalize_pair_idx(pair_idx):
            return int(pair_idx) if pair_idx is not None and 0 <= int(pair_idx) < 12 else None

        @staticmethod
        def _pair_has_visual_diff(pair_idx):
            return 2 <= pair_idx <= 9

        _logical_diff_pair_block_for_pair = mod.SheetView._logical_diff_pair_block_for_pair

    view = DummyView()
    block = mod.SheetView._logical_diff_pair_block_for_line(view, 2)
    assert block == list(range(2, 10)), block


def _test_region_anchor_uses_selected_pair_before_trailing_insert_line():
    class DummyText:
        @staticmethod
        def index(_mark):
            # Tk Text uses the line after the rendered content as its end mark.
            return "801.0"

    class DummyView:
        display_rows = list(range(800))
        row_pairs = [(idx + 1, idx + 1) for idx in range(1714)]
        selected_pair_idx = 152
        hover_pair_idx = 152
        _last_cursor_cmp_pair_idx = 152
        _main_sel_line = 153
        _main_sel_col = 1
        _cursor_cmp_sel_line = None
        _cursor_cmp_sel_col = None
        left = DummyText()

        @staticmethod
        def _normalize_pair_idx(pair_idx):
            return int(pair_idx) if pair_idx is not None and 0 <= int(pair_idx) < 1714 else None

        @staticmethod
        def _pair_idx_for_line(line):
            return int(line) - 1 if 1 <= int(line) <= 800 else None

        @staticmethod
        def _pair_has_visual_diff(pair_idx):
            return 151 <= int(pair_idx) <= 1351

        has_explicit_cell_selection = mod.SheetView.has_explicit_cell_selection
        resolved_pair_idx_for_c_area = mod.SheetView.resolved_pair_idx_for_c_area
        _logical_diff_pair_block_for_pair = mod.SheetView._logical_diff_pair_block_for_pair

    view = DummyView()
    anchor = view.resolved_pair_idx_for_c_area()
    assert anchor == 152, anchor
    block = view._logical_diff_pair_block_for_pair(anchor)
    assert block == list(range(151, 1352)), (block[:2], block[-2:], len(block))


def _test_full_only_diff_block_model_is_stable_and_uncapped():
    rows = [10, 11, 50, 51, 800, 801, 1200]
    pending_pairs = {50, 51, 1200}
    blocks = mod.SheetView._group_diff_pair_rows(
        rows,
        pending_predicate=lambda pair_idx: pair_idx in pending_pairs,
    )
    assert [(b.start_pair_idx, b.end_pair_idx) for b in blocks] == [
        (10, 11),
        (50, 51),
        (800, 801),
        (1200, 1200),
    ]
    assert [b.ordinal for b in blocks] == [1, 2, 3, 4]
    assert [b.pending for b in blocks] == [False, True, False, True]
    assert blocks[0].pair_indices == (10, 11)
    assert blocks[-1].start_pair_idx > 800

    # Duplicate/invalid entries must not create phantom blocks; a retained
    # resolved row keeps the original consecutive snapshot block intact.
    normalized = mod.SheetView._group_diff_pair_rows(
        [5, 6, 6, -1, 7],
        pending_predicate=lambda pair_idx: pair_idx == 7,
    )
    assert len(normalized) == 1
    assert normalized[0].pair_indices == (5, 6, 7)
    assert normalized[0].pending is True


def _test_diff_block_updates_never_read_workbook_state():
    class BombApp:
        def __getattribute__(self, name):
            raise AssertionError(f"Unexpected app/workbook access: {name}")

    class DummyVar:
        @staticmethod
        def get():
            return 1

    class DummyText:
        @staticmethod
        def index(_position):
            return "1.0"

    class DummyView:
        app = BombApp()
        only_diff_var = DummyVar()
        left = DummyText()
        row_pairs = [(idx + 1, idx + 1) for idx in range(1300)]
        display_rows = [10, 11, 1200]
        _full_display_rows = [10, 11, 1200]
        _only_diff_source_version = 3
        _data_version = 7
        _data_ready = True
        _only_diff_async_building = False
        _full_diff_blocks_cache_key = None
        _full_diff_blocks = []
        _pair_to_full_diff_block = {}
        selected_pair_idx = 1200
        _last_selected_line = 3
        _main_sel_line = None
        _main_sel_col = None
        _cursor_cmp_sel_line = None
        _cursor_cmp_sel_col = None

        @staticmethod
        def _pair_has_visual_diff(pair_idx):
            return int(pair_idx) in {10, 1200}

        _normalize_pair_idx = mod.SheetView._normalize_pair_idx
        has_explicit_cell_selection = mod.SheetView.has_explicit_cell_selection
        _invalidate_diff_block_model = mod.SheetView._invalidate_diff_block_model
        _diff_block_model_ready = mod.SheetView._diff_block_model_ready
        _group_diff_pair_rows = staticmethod(mod.SheetView._group_diff_pair_rows)
        _ensure_full_diff_blocks = mod.SheetView._ensure_full_diff_blocks
        _full_diff_block_index_for_pair = mod.SheetView._full_diff_block_index_for_pair
        _full_diff_block_for_pair = mod.SheetView._full_diff_block_for_pair
        _active_full_diff_block_index = mod.SheetView._active_full_diff_block_index
        _logical_diff_pair_block_for_pair = mod.SheetView._logical_diff_pair_block_for_pair
        _pair_idx_for_line = mod.SheetView._pair_idx_for_line

    view = DummyView()
    blocks = view._ensure_full_diff_blocks()
    assert [(block.start_pair_idx, block.end_pair_idx) for block in blocks] == [
        (10, 11),
        (1200, 1200),
    ]
    assert view._active_full_diff_block_index() == 1
    assert view._logical_diff_pair_block_for_pair(10) == [10, 11]


def _test_tail_identical_append_stays_paired():
    wb_mine = Workbook()
    wb_theirs = Workbook()
    try:
        ws_mine = wb_mine.active
        ws_theirs = wb_theirs.active
        ws_mine.append(["base"])
        ws_mine.append(["same append"])
        ws_theirs.append(["base"])
        ws_theirs.append(["same append"])
        pairs = mod._split_tail_independent_append_pairs(
            [(1, 1), (2, 2)],
            {1: 1},
            {1: 1},
            ws_mine,
            ws_theirs,
            1,
        )
        assert pairs == [(1, 1), (2, 2)], pairs

        ws_theirs["A2"] = "different append"
        ws_mine["A3"] = "mine append 2"
        ws_theirs["A3"] = "theirs append 2"
        pairs = mod._split_tail_independent_append_pairs(
            [(1, 1), (2, 2), (3, 3)],
            {1: 1},
            {1: 1},
            ws_mine,
            ws_theirs,
            1,
        )
        assert pairs == [(1, 1), (None, 2), (None, 3), (2, None), (3, None)], pairs
    finally:
        wb_mine.close()
        wb_theirs.close()


def _test_shared_formula_is_not_destroyed():
    ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    xml = f"""
    <worksheet xmlns="{ns}">
      <sheetData>
        <row r="1"><c r="A1"><f t="shared" ref="A1:A2" si="7">SUM(B1:C1)</f><v>3</v></c></row>
        <row r="2"><c r="A2"><f t="shared" si="7"/><v>7</v></c></row>
      </sheetData>
    </worksheet>
    """
    root = ET.fromstring(xml)
    mod._sheet_xml_set_cell(root, 1, 1, "=SUM(B1:C1)", 9)
    f = root.find(f".//{{{ns}}}c[@r='A1']/{{{ns}}}f")
    v = root.find(f".//{{{ns}}}c[@r='A1']/{{{ns}}}v")
    assert f is not None and f.attrib == {"t": "shared", "ref": "A1:A2", "si": "7"}, f.attrib
    assert v is not None and v.text == "9"

    # A cache-only adoption must preserve even a shared member formula whose
    # formula text is stored only on the group master.
    mod._sheet_xml_set_cell(
        root,
        2,
        1,
        "=IGNORED_FOR_CACHE_ONLY()",
        11,
        preserve_existing_formula=True,
    )
    member_f = root.find(f".//{{{ns}}}c[@r='A2']/{{{ns}}}f")
    member_v = root.find(f".//{{{ns}}}c[@r='A2']/{{{ns}}}v")
    assert member_f is not None and member_f.attrib == {"t": "shared", "si": "7"}
    assert member_f.text in (None, "")
    assert member_v is not None and member_v.text == "11"

    # A value/formula replacement may safely detach one member by expanding
    # the shared group into ordinary formulas first.
    mod._sheet_xml_set_cell(root, 1, 1, "=SUM(B1:D1)", 4)
    master_f = root.find(f".//{{{ns}}}c[@r='A1']/{{{ns}}}f")
    member_f = root.find(f".//{{{ns}}}c[@r='A2']/{{{ns}}}f")
    master_v = root.find(f".//{{{ns}}}c[@r='A1']/{{{ns}}}v")
    member_v = root.find(f".//{{{ns}}}c[@r='A2']/{{{ns}}}v")
    assert master_f is not None and master_f.attrib == {}
    assert master_f.text == "SUM(B1:D1)"
    assert member_f is not None and member_f.attrib == {}
    assert member_f.text == "SUM(B2:C2)"
    assert master_v is not None and master_v.text == "4"
    assert member_v is not None and member_v.text == "11"

    array_root = ET.fromstring(
        f'<worksheet xmlns="{ns}"><sheetData><row r="1">'
        '<c r="A1"><f t="array" ref="A1:A2">ROW(A1:A2)</f><v>1</v></c>'
        '</row></sheetData></worksheet>'
    )
    try:
        mod._sheet_xml_set_cell(array_root, 1, 1, "=SUM(B1:C1)", 4)
    except RuntimeError as exc:
        assert "special formula" in str(exc)
    else:
        raise AssertionError("unsafe array-formula replacement was not rejected")


def _test_formula_and_value_comparison_is_conservative():
    assert mod._same_formula("=sum(a1)", "=SUM(A1)")
    assert not mod._same_formula('="a b"', '="A B"')
    assert not mod._same_formula('="a b"', '="ab"')
    assert not mod._same_formula("=A:A B:B", "=A:AB:B")
    assert mod._merge_cmp_value("001") != mod._merge_cmp_value(1)
    assert mod._merge_cmp_value("value ") != mod._merge_cmp_value("value")
    assert mod._merge_cmp_value(1) == mod._merge_cmp_value(1.0)
    literal = mod._choose_edit_value("=not-a-formula", "=not-a-formula")
    assert isinstance(literal, mod._LiteralText)
    assert mod._formula_text(literal) is None

    ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    root = ET.fromstring(f'<worksheet xmlns="{ns}"><sheetData/></worksheet>')
    mod._sheet_xml_set_cell(root, 1, 1, literal)
    cell = root.find(f".//{{{ns}}}c[@r='A1']")
    assert cell.attrib.get("t") == "inlineStr"
    assert cell.find(f"{{{ns}}}f") is None
    assert cell.find(f"{{{ns}}}is/{{{ns}}}t").text == "=not-a-formula"

    array_a = ArrayFormula("A1:A2", "=ROW(A1:A2)")
    array_same = ArrayFormula("A1:A2", "=ROW(A1:A2)")
    array_other = ArrayFormula("A1:A3", "=ROW(A1:A3)")
    assert mod._same_formula(array_a, array_same)
    assert not mod._same_formula(array_a, array_other)
    data_a = DataTableFormula("B1:C3", r1="A1")
    data_same = DataTableFormula("B1:C3", r1="A1")
    assert mod._same_formula(data_a, data_same)
    try:
        mod._choose_edit_value(1, data_a)
    except RuntimeError as exc:
        assert "数据表公式" in str(exc)
    else:
        raise AssertionError("data-table formula was silently converted to a scalar")

    shown_a, shown_b, equal = mod._cell_display_and_equal_from_values(
        None, None, "=A1+1", "=A1+2"
    )
    assert not equal
    assert shown_a == "=A1+1" and shown_b == "=A1+2"
    shown_a, shown_b, equal = mod._cell_display_and_equal_from_values(
        None, None, "=A1+1", "=a1+1"
    )
    assert equal


def _test_ooxml_datetime_is_numeric():
    ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    root = ET.fromstring(f'<worksheet xmlns="{ns}"><sheetData/></worksheet>')
    value = datetime(2026, 7, 21, 12, 30, 0)
    mod._sheet_xml_set_cell(root, 1, 1, value)
    cell = root.find(f".//{{{ns}}}c[@r='A1']")
    cached = cell.find(f"{{{ns}}}v")
    assert cell.attrib.get("t") is None
    assert float(cached.text) > 40000
    assert mod._excel_cached_value_payload("#comment") == ("str", "#comment")
    assert mod._excel_cached_value_payload("#N/A") == ("e", "#N/A")
    try:
        mod._excel_cached_value_payload(float("nan"))
    except RuntimeError as exc:
        assert "non-finite" in str(exc)
    else:
        raise AssertionError("NaN was written as an invalid OOXML numeric value")

    root_dir = make_temp_dir("sow_date1904_")
    source = os.path.join(root_dir, "source.xlsx")
    output = os.path.join(root_dir, "output.xlsx")
    wb = Workbook()
    wb.epoch = mod.CALENDAR_MAC_1904
    ws = wb.active
    ws.title = "S1"
    ws["A1"] = datetime(2020, 1, 1)
    ws["A1"].number_format = "yyyy-mm-dd"
    wb.save(source)
    wb.close()
    expected = datetime(2026, 7, 21)
    mod._build_manual_merge_xlsx_via_zip(source, output, {("S1", 1, 1): expected})
    wb = load_workbook(output, data_only=False)
    try:
        assert wb["S1"]["A1"].value == expected
    finally:
        wb.close()


def _test_formula_noop_filter():
    root = make_temp_dir("sow_formula_noop_")
    path = os.path.join(root, "source.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "S1"
    ws["A1"] = "=SUM(B1:C1)"
    ws["B1"] = 1
    ws["C1"] = 2
    wb.save(path)
    wb.close()

    ops = {
        ("S1", 1, 1): "=SUM(B1:C1)",
        ("S1", 1, 2): 10,
    }
    filtered = mod._filter_noop_manual_ops(path, ops)
    assert ("S1", 1, 1) not in filtered, filtered
    assert filtered[("S1", 1, 2)] == 10
    row_aware = mod._prepare_manual_ops_for_save(
        path,
        {("S1", 1, 1): "=SUM(B1:C1)"},
        row_ops=[{"sheet": "S1", "kind": "insert_rows", "row": 1, "count": 1}],
    )
    assert row_aware == {("S1", 1, 1): "=SUM(B1:C1)"}

    replay_sources = mod._replay_formula_source_paths(
        "mine.xlsx",
        row_ops=[{"kind": "insert_rows", "source_side": "B"}],
        source_paths={"B": "theirs.xlsx", "BASE": "base.xlsx"},
    )
    assert replay_sources == ["mine.xlsx", "theirs.xlsx"]

    text_out = os.path.join(root, "text-output.xlsx")
    mod._build_manual_merge_xlsx_via_zip(
        path,
        text_out,
        {("S1", 2, 1): "1.0"},
    )
    wb = load_workbook(text_out, data_only=False)
    try:
        assert wb["S1"]["A2"].value == "1.0"
        assert isinstance(wb["S1"]["A2"].value, str)
    finally:
        wb.close()


def _test_formula_only_tail_is_not_trimmed():
    root = make_temp_dir("sow_formula_tail_bounds_")
    path = os.path.join(root, "formula-tail.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "S1"
    ws["A1"] = "header"
    ws["C10"] = "=IF(1=1,\"\",\"x\")"
    wb.save(path)
    wb.close()

    wb_val = load_workbook(path, data_only=True)
    wb_edit = load_workbook(path, data_only=False)
    try:
        assert mod._effective_bounds(wb_val["S1"])[0] == 1
        assert mod._effective_bounds_with_edit(wb_val["S1"], wb_edit["S1"]) == (10, 3)
    finally:
        wb_val.close()
        wb_edit.close()


def _test_large_alignment_fast_and_exact():
    sig_base = [f"id-{idx}" for idx in range(20000)]
    sig_theirs = list(sig_base)
    sig_theirs[16153:16158] = [f"changed-{idx}" for idx in range(5)]
    started = time.perf_counter()
    pairs = mod._compute_row_pairs_from_signatures(sig_base, sig_theirs)
    elapsed = time.perf_counter() - started
    assert len(pairs) == 20000, len(pairs)
    assert pairs[0] == (1, 1) and pairs[-1] == (20000, 20000)
    assert elapsed < 2.0, elapsed

    # Equal total row counts can still contain one insertion and one deletion.
    # Large-sheet auto alignment must not degrade this into thousands of value diffs.
    structural_a = [f"row-{idx}" for idx in range(20000)]
    structural_b = structural_a[:500] + ["inserted"] + structural_a[500:15000] + structural_a[15001:]
    assert len(structural_a) == len(structural_b)
    assert mod._should_auto_row_align(len(structural_a), len(structural_b))
    started = time.perf_counter()
    structural_pairs = mod._compute_row_pairs_from_signatures(structural_a, structural_b)
    structural_elapsed = time.perf_counter() - started
    assert sum(1 for left, right in structural_pairs if left is None and right is not None) == 1
    assert sum(1 for left, right in structural_pairs if left is not None and right is None) == 1
    assert structural_elapsed < 2.0, structural_elapsed

    # A formula/result-heavy sheet can produce a huge equal-length replace
    # block. Pairing is unambiguously 1:1 and must not run per-row fuzzy scans.
    changed_a = [f"left-{idx}-" + ("x" * 120) for idx in range(20000)]
    changed_b = [f"right-{idx}-" + ("y" * 120) for idx in range(20000)]
    started = time.perf_counter()
    changed_pairs = mod._compute_row_pairs_from_signatures(changed_a, changed_b)
    changed_elapsed = time.perf_counter() - started
    assert changed_pairs[0] == (1, 1) and changed_pairs[-1] == (20000, 20000)
    assert len(changed_pairs) == 20000
    assert changed_elapsed < 2.0, changed_elapsed


def _test_stable_copy_waits_for_complete_zip():
    root = make_temp_dir("sow_partial_svn_export_")
    complete = os.path.join(root, "complete.xlsx")
    partial = os.path.join(root, "WorldMonster.xlsx-revBASE.tmp.xlsx")
    wb = Workbook()
    wb.active["A1"] = "complete"
    wb.save(complete)
    wb.close()

    with open(complete, "rb") as src:
        payload = src.read()
    split_at = len(payload) // 2
    with open(partial, "wb") as dst:
        dst.write(payload[:split_at])
        dst.flush()

    def _finish_export():
        time.sleep(0.35)
        with open(partial, "ab") as dst:
            dst.write(payload[split_at:])
            dst.flush()

    writer = threading.Thread(target=_finish_export, daemon=True)
    writer.start()
    stable = mod._ensure_stable_copy(partial)
    writer.join(timeout=2.0)
    assert stable != partial
    assert os.path.getsize(stable) == len(payload)
    assert mod._workbook_package_ready(stable)


def _test_background_recalc_policy_is_explicit():
    old_always = mod._AUTO_RECALC_FORMULAS_ALWAYS
    old_missing = mod._AUTO_RECALC_MISSING_CACHE
    old_recalc = mod._recalc_and_prepare_val_path
    calls = []
    try:
        mod._AUTO_RECALC_FORMULAS_ALWAYS = False
        mod._AUTO_RECALC_MISSING_CACHE = False
        mod._recalc_and_prepare_val_path = lambda path: calls.append(path) or "recalc.xlsx"
        assert mod._maybe_recalc_and_prepare_val_path("formula.xlsx", force=False) is None
        assert not calls
        assert mod._maybe_recalc_and_prepare_val_path("formula.xlsx", force=True) == "recalc.xlsx"
        assert calls == ["formula.xlsx"]
    finally:
        mod._AUTO_RECALC_FORMULAS_ALWAYS = old_always
        mod._AUTO_RECALC_MISSING_CACHE = old_missing
        mod._recalc_and_prepare_val_path = old_recalc


def _test_large_only_diff_disk_worker_is_blocked_after_user_edits():
    view = SimpleNamespace(
        _is_large_sheet=True,
        touched_rows={1401},
        sheet="S1",
        app=SimpleNamespace(modified_sheets_a={"S1"}, modified_sheets_b=set()),
    )
    view._has_user_edits_for_current_sheet = lambda: mod.SheetView._has_user_edits_for_current_sheet(view)
    assert mod.SheetView._start_async_large_only_diff_build(view) is False

    view.touched_rows.clear()
    view.app.modified_sheets_a.clear()
    assert not mod.SheetView._has_user_edits_for_current_sheet(view)


def _test_three_way_formula_structure_is_compared_without_cache():
    root = make_temp_dir("sow_formula_structure_3way_")

    def _book(name, formula):
        path = os.path.join(root, name)
        wb = Workbook()
        ws = wb.active
        ws.title = "S1"
        ws["A1"] = formula
        wb.save(path)
        wb.close()
        return path

    base = _book("base.xlsx", "=SUM(1,1)")
    mine = _book("mine.xlsx", "=SUM(1,2)")
    theirs = _book("theirs.xlsx", "=SUM(1,3)")
    conflicts, conflict_map = mod._scan_three_way_conflicts(base, mine, theirs)
    assert any(sheet == "S1" and row == 1 and col == 1 for sheet, row, col, _vm, _vt in conflicts), conflicts
    assert conflict_map["S1"][1] == {1}

    mine_unchanged = _book("mine-unchanged.xlsx", "=SUM(1,1)")
    theirs_changed = _book("theirs-changed.xlsx", "=SUM(1,4)")
    out = os.path.join(root, "merged.xlsx")
    conflicts, _preview, conflict_map = mod._merge_three_way(
        base,
        mine_unchanged,
        theirs_changed,
        out,
        save_merged=True,
    )
    assert conflicts == [] and conflict_map == {}
    wb = load_workbook(out, data_only=False)
    try:
        assert wb["S1"]["A1"].value == "=SUM(1,4)"
    finally:
        wb.close()

    assert mod._merge_cell_compare_key("=literal", "=literal")[0] == "VALUE"
    assert mod._merge_cell_compare_key(None, "=literal")[0] == "FORMULA"


def _test_formula_copy_translates_relative_references():
    translated = mod._copy_edit_value_for_destination(
        None,
        "=A11+$C$1+D$2+$E11",
        None,
        src_row=11,
        src_col=2,
        dst_row=10,
        dst_col=2,
    )
    assert translated == "=A10+$C$1+D$2+$E10", translated

    literal = mod._copy_edit_value_for_destination(
        "=A11",
        "=A11",
        None,
        src_row=11,
        src_col=2,
        dst_row=10,
        dst_col=2,
    )
    assert isinstance(literal, mod._LiteralText)
    assert str(literal) == "=A11"

    # Formula structure remains significant even when current cached results match.
    _da, _db, equal = mod._cell_display_and_equal_from_values(
        2,
        2,
        "=1+1",
        "=SUM(1,1)",
    )
    assert not equal

    shifted = mod._translate_normal_formula_for_compare(
        None,
        "=A11",
        11,
        2,
        10,
        2,
    )
    _da, _db, equal = mod._cell_display_and_equal_from_values(
        None,
        None,
        "=A10",
        shifted,
    )
    assert equal


def _test_same_formula_copy_only_records_changed_cache():
    view = SimpleNamespace(_formula_copy_skips_pending=0)
    mode = mod.SheetView._same_formula_copy_mode(view, "=A1", "=A1", 10, 10)
    assert mode == "noop"
    assert view._formula_copy_skips_pending == 0
    mode = mod.SheetView._same_formula_copy_mode(view, "=A1", "=A1", 11, 10)
    assert mode == "cache"
    assert view._formula_copy_skips_pending == 1
    mode = mod.SheetView._same_formula_copy_mode(view, "=A2", "=A1", 11, 10)
    assert mode is None


def _test_excel_row_replay_uses_full_paste():
    captured = {}
    original_run = mod.subprocess.run
    original_validate = mod._validate_xlsx_package
    fixture_root = make_temp_dir("sow_review_native_row_replay_")
    source_path = os.path.join(fixture_root, "source.xlsx")
    theirs_path = os.path.join(fixture_root, "theirs.xlsx")
    output_path = os.path.join(fixture_root, "output.xlsx")
    for path in (source_path, theirs_path):
        workbook = Workbook()
        workbook.active.title = "S1"
        workbook.active["A1"] = "fixture"
        workbook.save(path)
        workbook.close()
    try:
        def _fake_run(command, **_kwargs):
            captured["command"] = command
            script = command[-1]
            marker = "$opsPath='"
            start = script.index(marker) + len(marker)
            end = script.index("';", start)
            ops_path = script[start:end].replace("''", "'")
            with open(ops_path, "r", encoding="utf-8") as f:
                captured["payload"] = json.load(f)
            return SimpleNamespace(returncode=0, stdout="", stderr="")

        mod.subprocess.run = _fake_run
        mod._validate_xlsx_package = lambda _path: (True, "")
        ok = mod._build_manual_merge_output_with_excel(
            source_path,
            output_path,
            {
                ("S1", 2, 2): None,
                ("S1", 2, 3): "",
                ("S1", 2, 4): 0,
                ("S1", 2, 5): False,
            },
            row_ops=[{
                "kind": "insert_rows",
                "sheet": "S1",
                "row": 2,
                "count": 1,
                "source_side": "B",
                "source_rows": [2],
            }],
            source_paths={"B": theirs_path},
        )
        assert ok
        script = captured["command"][-1]
        assert "PasteSpecial(-4104)" in script
        assert "PasteSpecial(-4122)" not in script
        assert "value_kind -eq 'blank'" in script
        assert "$cell.ClearContents()" in script
        payload_ops = {
            (op["r"], op["c"]): op
            for op in captured["payload"]["cell_ops"]
        }
        assert payload_ops[(2, 2)] == {
            "sheet": "S1", "r": 2, "c": 2, "value_kind": "blank"
        }, payload_ops[(2, 2)]
        assert payload_ops[(2, 3)]["value_kind"] == "text"
        assert payload_ops[(2, 3)]["value"] == ""
        assert payload_ops[(2, 4)]["value_kind"] == "typed"
        assert payload_ops[(2, 4)]["value"] == 0
        assert payload_ops[(2, 5)]["value_kind"] == "typed"
        assert payload_ops[(2, 5)]["value"] is False
    finally:
        mod.subprocess.run = original_run
        mod._validate_xlsx_package = original_validate


def _test_structural_replay_risk_detection():
    root = make_temp_dir("sow_structural_risk_")
    plain = os.path.join(root, "plain.xlsx")
    risky = os.path.join(root, "risky.xlsx")

    wb = Workbook()
    wb.active["A1"] = "plain"
    wb.save(plain)
    wb.close()
    assert not mod._xlsx_requires_native_structural_replay(plain)

    wb = Workbook()
    ws = wb.active
    validation = DataValidation(type="list", formula1='"A,B"')
    ws.add_data_validation(validation)
    validation.add("A1:A3")
    wb.save(risky)
    wb.close()
    assert mod._xlsx_requires_native_structural_replay(risky)


def main():
    _test_only_diff_region_boundaries()
    _test_logical_region_extends_beyond_render_limit()
    _test_region_anchor_uses_selected_pair_before_trailing_insert_line()
    _test_full_only_diff_block_model_is_stable_and_uncapped()
    _test_diff_block_updates_never_read_workbook_state()
    _test_tail_identical_append_stays_paired()
    _test_shared_formula_is_not_destroyed()
    _test_formula_and_value_comparison_is_conservative()
    _test_ooxml_datetime_is_numeric()
    _test_formula_noop_filter()
    _test_formula_only_tail_is_not_trimmed()
    _test_large_alignment_fast_and_exact()
    _test_stable_copy_waits_for_complete_zip()
    _test_background_recalc_policy_is_explicit()
    _test_large_only_diff_disk_worker_is_blocked_after_user_edits()
    _test_three_way_formula_structure_is_compared_without_cache()
    _test_formula_copy_translates_relative_references()
    _test_same_formula_copy_only_records_changed_cache()
    _test_excel_row_replay_uses_full_paste()
    _test_structural_replay_risk_detection()
    print("SMOKE_REVIEW_REGRESSIONS_OK")


if __name__ == "__main__":
    main()

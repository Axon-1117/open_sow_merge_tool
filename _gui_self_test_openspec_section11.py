"""OpenSpec section 11 acceptance: CJK pixels, compact actions, Excel labels.

The real Gunships workbooks are immutable inputs.  Synthetic workbooks are
created only in temporary directories.

Run:
  python _gui_self_test_openspec_section11.py
"""

from __future__ import annotations

import argparse
import re
import time

import tkinter as tk
import tkinter.font as tkfont
from openpyxl.utils import get_column_letter

import sow_merge_tool as smt
import _gui_self_test_openspec_section10 as section10
from _gui_diagnose_gunships_pixel_alignment import (
    _document_xpixels,
    _pixel_boundary_records,
    _rect,
    _root_top_button_inventory,
)
from _gui_self_test_latest_gunships_feedback import (
    _pump,
    _real_gunships_app,
    _wait_until,
)
from _gui_self_test_openspec_section10 import _synthetic_view


_VISIBLE_INTERNAL_COLUMN_RE = re.compile(r"\bL\d+(?::L\d+)?\b")


def _widget_rect(widget) -> tuple[int, int, int, int]:
    return _rect(widget)


def _contains(outer, inner, *, tolerance: int = 0) -> bool:
    return (
        outer[0] - tolerance <= inner[0]
        and outer[1] - tolerance <= inner[1]
        and inner[2] <= outer[2] + tolerance
        and inner[3] <= outer[3] + tolerance
    )


def _horizontal_overlap(left, right) -> int:
    return max(0, min(left[2], right[2]) - max(left[0], right[0]))


def _vertical_overlap(left, right) -> int:
    return max(0, min(left[3], right[3]) - max(left[1], right[1]))


def _nearest_common_ancestor(left, right):
    left_ancestors = []
    current = left
    while current is not None:
        left_ancestors.append(current)
        current = getattr(current, "master", None)
    right_ancestors = set()
    current = right
    while current is not None:
        right_ancestors.add(current)
        current = getattr(current, "master", None)
    return next(
        (ancestor for ancestor in left_ancestors if ancestor in right_ancestors),
        None,
    )


def _header_boundaries(view):
    result = {}
    widgets = (
        ("A", view.left_colhdr),
        ("BASE", view.base_colhdr),
        ("B", view.right_colhdr),
    )
    for side, widget in widgets:
        line = widget.get("1.0", "1.end")
        spans = view._spans_for_line(line)
        result[side] = {
            get_column_letter(logical_col): _document_xpixels(
                widget,
                1,
                spans[logical_col][1],
            )
            for logical_col in range(1, 6)
        }
    return result


def _record_boundary_map(records):
    return {
        (
            str(record["side"]),
            int(record["mine_row"]),
            str(column),
        ): int(pixels)
        for record in records
        for column, pixels in record["boundaries_px"].items()
    }


def _assert_boundary_tolerance(records, headers, tolerance: int = 1) -> None:
    by_side_column = {}
    by_row_column = {}
    for record in records:
        side = str(record["side"])
        row = int(record["mine_row"])
        for column, raw_pixels in record["boundaries_px"].items():
            pixels = int(raw_pixels)
            by_side_column.setdefault((side, column), []).append(pixels)
            by_row_column.setdefault((row, column), []).append(pixels)
            assert abs(pixels - int(headers[side][column])) <= tolerance, (
                "data/header pixel drift",
                side,
                row,
                column,
                pixels,
                headers[side][column],
            )
    for key, pixels in by_side_column.items():
        assert max(pixels) - min(pixels) <= tolerance, (
            "cross-row pixel drift",
            key,
            pixels,
        )
    for key, pixels in by_row_column.items():
        assert max(pixels) - min(pixels) <= tolerance, (
            "cross-pane pixel drift",
            key,
            pixels,
        )


def _font_metrics(widget):
    font = tkfont.Font(root=widget, font=widget.cget("font"))
    return {
        "family": str(font.actual("family")),
        "size": int(font.actual("size")),
        "linespace": int(font.metrics("linespace")),
        "space": int(font.measure(" ")),
        "latin": int(font.measure("A")),
        "cjk": int(font.measure("中")),
        "sample": int(font.measure("A中 0")),
    }


def _assert_tag_font_metrics_unchanged(widget, tag_name: str) -> None:
    tagged_font = str(widget.tag_cget(tag_name, "font") or "").strip()
    if not tagged_font:
        return
    base = tkfont.Font(root=widget, font=widget.cget("font"))
    tagged = tkfont.Font(root=widget, font=tagged_font)
    for sample in (" ", "A", "中", "A中 0"):
        assert abs(int(base.measure(sample)) - int(tagged.measure(sample))) <= 1, (
            widget,
            tag_name,
            tagged_font,
            sample,
            base.measure(sample),
            tagged.measure(sample),
        )


def _assert_accented_latin_and_combining_pixel_boundaries(root, source_widget) -> None:
    probe = tk.Text(
        root,
        width=80,
        height=5,
        wrap="none",
        borderwidth=0,
        highlightthickness=0,
        padx=0,
        pady=0,
        font=source_widget.cget("font"),
    )
    probe.place(x=-10000, y=-10000, width=1600, height=120)
    samples = (
        ("ascii", "ae"),
        ("precomposed", "áé"),
        ("decomposed", "a\u0301e\u0301"),
        ("uncomposed_mark", "a\u0338e"),
    )
    try:
        rendered = {}
        for line_no, (name, value) in enumerate(samples, start=1):
            first = smt._format_cell(value, 18)
            second = smt._format_cell("tail", 18)
            assert len(first) == 18, (name, value, repr(first), len(first))
            probe.insert(f"{line_no}.0", first + second + "\n")
            rendered[name] = first
        root.update_idletasks()
        boundaries = {
            name: (
                _document_xpixels(probe, line_no, 18),
                _document_xpixels(probe, line_no, 36),
            )
            for line_no, (name, _value) in enumerate(samples, start=1)
        }
        assert rendered["precomposed"] == rendered["decomposed"], rendered
        assert "\\u0338" in rendered["uncomposed_mark"], rendered
        baseline = boundaries["ascii"]
        for name, pixels in boundaries.items():
            assert max(
                abs(int(pixels[index]) - int(baseline[index]))
                for index in (0, 1)
            ) <= 1, (
                "accented Latin/combining mark changed a Tk cell boundary",
                name,
                rendered[name],
                pixels,
                baseline,
            )
    finally:
        probe.destroy()


def test_real_gunships_a1_a5_pixels_match_headers_panes_and_tags():
    with _real_gunships_app() as (app, view, _analysis):
        app.root.geometry("1450x860")
        _pump(app.root, 0.2)
        main_and_headers = (
            view.left,
            view.base,
            view.right,
            view.cursor_cmp,
            view.left_colhdr,
            view.base_colhdr,
            view.right_colhdr,
            view.cursor_cmp_colhdr,
        )
        metrics = [_font_metrics(widget) for widget in main_and_headers]
        assert len(
            {
                (
                    item["family"],
                    item["size"],
                    item["linespace"],
                    item["space"],
                    item["latin"],
                    item["cjk"],
                )
                for item in metrics
            }
        ) == 1, metrics
        common = metrics[0]
        assert common["space"] == common["latin"], common
        assert abs(common["cjk"] - common["space"] * 2) <= 1, common
        _assert_accented_latin_and_combining_pixel_boundaries(
            app.root,
            view.left,
        )

        headers_before = _header_boundaries(view)
        records_before = _pixel_boundary_records(app, view)
        _assert_boundary_tolerance(records_before, headers_before)
        boundary_before = _record_boundary_map(records_before)

        selected_pair = view.row_a_to_pair_idx[3]
        view._select_line(view.row_to_line[selected_pair])
        view._update_cursor_lines()
        for widget in (view.left, view.base, view.right):
            for mine_row in range(1, 6):
                pair_idx = view.row_a_to_pair_idx[mine_row]
                line_no = view.row_to_line[pair_idx]
                line_text = widget.get(f"{line_no}.0", f"{line_no}.end")
                spans = view._spans_for_line(line_text)
                for logical_col in range(1, 6):
                    start, end = spans[logical_col]
                    widget.tag_add(
                        "diffcell",
                        f"{line_no}.{start}",
                        f"{line_no}.{end}",
                    )
            for tag_name in ("selrow", "diffrow", "diffcell", "selcell"):
                _assert_tag_font_metrics_unchanged(widget, tag_name)
        _pump(app.root, 0.08)

        headers_after = _header_boundaries(view)
        records_after = _pixel_boundary_records(app, view)
        _assert_boundary_tolerance(records_after, headers_after)
        boundary_after = _record_boundary_map(records_after)
        assert headers_after == headers_before
        for key, pixels in boundary_before.items():
            assert abs(int(boundary_after[key]) - int(pixels)) <= 1, (
                "selection/difference tag changed pixel geometry",
                key,
                pixels,
                boundary_after[key],
            )


def _structural_cluster_rect(view):
    status = _widget_rect(view.column_action_status_group)
    buttons = _widget_rect(view.column_action_button_group)
    return (
        min(status[0], buttons[0]),
        min(status[1], buttons[1]),
        max(status[2], buttons[2]),
        max(status[3], buttons[3]),
    )


def _assert_column_buttons_visible(view, owner_rect) -> None:
    for button in (
        view.use_mine_col_btn,
        view.use_base_col_btn,
        view.use_theirs_col_btn,
    ):
        assert bool(button.winfo_ismapped()), button
        rect = _widget_rect(button)
        assert _contains(owner_rect, rect, tolerance=1), (
            owner_rect,
            button.cget("text"),
            rect,
        )
        assert rect[2] - rect[0] >= int(button.winfo_reqwidth()), (
            button.cget("text"),
            rect,
            button.winfo_reqwidth(),
        )


def test_real_gunships_combined_action_row_is_centered_or_collision_safe():
    with _real_gunships_app() as (app, view, _analysis):
        for geometry in ("1450x860", "1024x760"):
            app.root.geometry(geometry)
            _pump(app.root, 0.2)
            nav = _widget_rect(view.diff_nav_group)
            structural = _structural_cluster_rect(view)
            shared = _nearest_common_ancestor(
                view.diff_nav_group,
                view.column_action_button_group,
            )
            assert shared is not None
            shared_rect = _widget_rect(shared)
            assert shared not in (view.frame, app.root), (
                "navigation and structural actions still occupy separate rows",
                geometry,
                shared,
                nav,
                structural,
            )
            assert shared_rect[3] - shared_rect[1] <= 44, (
                geometry,
                shared_rect,
            )
            assert _contains(shared_rect, nav, tolerance=1)
            assert _contains(shared_rect, structural, tolerance=1)
            assert _vertical_overlap(nav, structural) > 0, (
                geometry,
                nav,
                structural,
            )
            assert _horizontal_overlap(nav, structural) == 0, (
                geometry,
                nav,
                structural,
            )
            assert structural[0] - nav[2] >= 6, (
                geometry,
                nav,
                structural,
            )
            _assert_column_buttons_visible(view, shared_rect)
            for child in (
                view.prev_diff_btn,
                view.diff_block_status,
                view.next_diff_btn,
            ):
                assert bool(child.winfo_ismapped())
                assert _contains(nav, _widget_rect(child), tolerance=1)
            centered_left = (
                shared_rect[0]
                + (
                    (shared_rect[2] - shared_rect[0])
                    - (nav[2] - nav[0])
                )
                // 2
            )
            centered_right = centered_left + (nav[2] - nav[0])
            if centered_right + 10 <= structural[0]:
                assert abs(
                    (nav[0] + nav[2]) - (shared_rect[0] + shared_rect[2])
                ) <= 4, (geometry, nav, shared_rect)
            else:
                assert nav[0] >= shared_rect[0]
                assert 8 <= structural[0] - nav[2] <= 12, (
                    geometry,
                    nav,
                    structural,
                )


def test_real_gunships_root_utilities_are_left_aligned_in_stable_order():
    deadline = time.monotonic() + 90.0
    with _real_gunships_app(absolute_deadline=deadline) as (app, view, _analysis):
        snapshots = {}
        for phase in ("before", "after"):
            snapshots[phase] = {}
            for geometry in ("1450x860", "1024x760"):
                app.root.geometry(geometry)
                _pump(app.root, 0.15)
                inventory = _root_top_button_inventory(app)
                assert inventory["order_matches"], inventory
                assert inventory["first_left_gap_px"] <= 10, inventory
                records = inventory["buttons"]
                for previous, current in zip(records, records[1:]):
                    gap = int(current["rect"][0]) - int(previous["rect"][2])
                    assert 7 <= gap <= 9, (previous, current, gap)
                snapshots[phase][geometry] = inventory
            if phase == "before":
                assert view._derive_lifecycle_state() == "EDIT_DEFERRED"
                assert not app._edit_workbooks_ready()
                assert view.selected_column_logical_range == (14, 14)
                undo_before = len(app.undo_stack)
                preload_calls = []
                original_preload = app._request_edit_preload

                def _tracked_preload(*args, **kwargs):
                    preload_calls.append((args, dict(kwargs)))
                    return original_preload(*args, **kwargs)

                app._request_edit_preload = _tracked_preload
                try:
                    view.use_base_col_btn.invoke()
                    assert preload_calls == [
                        (
                            (),
                            {
                                "reason": "mutation:列结构操作",
                                "caller": "SheetView._guard_mutation_ready",
                            },
                        )
                    ], preload_calls
                    assert view.selected_column_logical_range == (14, 14)
                    assert len(app.undo_stack) == undo_before
                    _wait_until(
                        app.root,
                        lambda: (
                            app._edit_workbooks_ready()
                            and view._derive_lifecycle_state() == "READY"
                            and view.selected_column_logical_range == (14, 14)
                        ),
                        "Gunships public column action did not reach edit READY",
                        timeout=max(0.001, deadline - time.monotonic()),
                    )
                    view.use_base_col_btn.invoke()
                finally:
                    app._request_edit_preload = original_preload
                assert len(preload_calls) == 1, preload_calls
                _wait_until(
                    app.root,
                    lambda: (
                        view._derive_lifecycle_state() == "READY"
                        and view.selected_column_logical_range == (20, 20)
                        and len(app.undo_stack) == undo_before + 1
                    ),
                    "Gunships did not reach its second structural block",
                    timeout=max(0.001, deadline - time.monotonic()),
                )
        for geometry in snapshots["before"]:
            before = snapshots["before"][geometry]
            after = snapshots["after"][geometry]
            assert before["buttons"] == after["buttons"], (
                geometry,
                before,
                after,
            )
        assert time.monotonic() <= deadline


def _mapped_guidance_texts(view):
    result = []
    widgets = (
        ("status_prefix", view.column_action_status_prefix_label),
        ("status_token", view.column_action_selection_label),
        ("status_suffix", view.column_action_status_suffix_label),
        ("sheet_info", view.info),
    )
    for name, widget in widgets:
        if not bool(widget.winfo_ismapped()):
            continue
        text = str(widget.cget("text") or "")
        if not text:
            try:
                variable = str(widget.cget("textvariable") or "")
                text = str(widget.getvar(variable)) if variable else ""
            except Exception:
                text = ""
        result.append((name, text))
        detail = str(getattr(widget, "_identity_detail_text", "") or "")
        if detail:
            result.append((f"{name}:tooltip", detail))
    return result


def _assert_no_visible_l_notation(texts) -> None:
    offenders = [
        (source, text, _VISIBLE_INTERNAL_COLUMN_RE.findall(str(text)))
        for source, text in texts
        if _VISIBLE_INTERNAL_COLUMN_RE.search(str(text))
    ]
    assert not offenders, offenders


def _excel_label_rows():
    headers = tuple(
        (
            f"field_{get_column_letter(index)}"
            f"{'@id' if index == 2 else '@pm'}"
        )
        for index in range(1, 28)
    )
    base = [
        headers,
        tuple(f"base-{index}" for index in range(1, 28)),
        tuple(f"same-{index}" for index in range(1, 28)),
    ]
    mine_keep = tuple(
        value
        for index, value in enumerate(headers, start=1)
        if index not in (1, 26, 27)
    )
    mine = [
        mine_keep,
        tuple(
            f"base-{index}"
            for index in range(1, 28)
            if index not in (1, 26, 27)
        ),
        tuple(
            f"same-{index}"
            for index in range(1, 28)
            if index not in (1, 26, 27)
        ),
    ]
    assert base[0][0].endswith("@pm") and base[0][1].endswith("@id")
    assert mine[0][0] == "field_B@id"
    assert sum(field.endswith("@id") for field in mine[0]) == 1
    return mine, base, [list(row) for row in base]


def test_visible_column_guidance_uses_excel_a_z_aa_without_l_notation():
    mine, base, theirs = _excel_label_rows()
    visible_evidence = []
    with _synthetic_view(
        mine,
        theirs,
        base_rows=base,
        geometry="1450x860",
    ) as (app, view, dialogs):
        expected_steps = (
            (1, 1, "A"),
            (26, 27, "Z:AA"),
        )

        view._set_copy_scope_mode("global")
        view._run_copy_action_by_mode("B2A")
        _pump(app.root, 0.08)
        blocker = str(view.info.cget("text") or "")
        visible_evidence.append(("global_blocker", blocker))
        for label in ("A", "Z", "AA"):
            assert label in blocker, blocker

        for expected_start, expected_end, expected_label in expected_steps:
            assert view.selected_column_logical_range == (
                expected_start,
                expected_end,
            ), (
                (expected_start, expected_end),
                view.selected_column_logical_range,
                view._column_structure_summary(),
            )
            pending = _mapped_guidance_texts(view)
            visible_evidence.extend(
                (f"{expected_label}:{source}", text)
                for source, text in pending
            )
            assert view.column_action_selection_var.get() == expected_label, (
                expected_label,
                view.column_action_selection_var.get(),
            )
            assert any(expected_label in text for _source, text in pending), pending
            _assert_no_visible_l_notation(pending)
            plan = view._apply_selected_column_block("BASE", "A")
            assert (plan.logical_start, plan.logical_end) == (
                expected_start,
                expected_end,
            )
            _wait_until(
                app.root,
                lambda expected=(expected_start, expected_end): (
                    view._derive_lifecycle_state() == "READY"
                    and (
                        view.selected_column_logical_range is None
                        or view.selected_column_logical_range != expected
                    )
                ),
                f"column action did not advance after {expected_label}",
                timeout=25.0,
            )

        completion = _mapped_guidance_texts(view)
        visible_evidence.extend(
            (f"completion:{source}", text)
            for source, text in completion
        )
        assert any("AA" in text for _source, text in completion), completion

        for logical_col, label in ((1, "A"), (26, "Z"), (27, "AA")):
            conflict_text = app._conflict_cell_location_text(
                ("Data", 2, logical_col)
            )
            visible_evidence.append((f"conflict_location:{label}", conflict_text))
            assert label in conflict_text, conflict_text

    ordinary_right = [list(row) for row in base]
    assert str(base[0][2]).endswith("@pm")
    ordinary_right[1][2] = "theirs-C"
    with _synthetic_view(
        base,
        ordinary_right,
        base_rows=base,
        geometry="1450x860",
    ) as (app, view, dialogs):
        confirmation = []
        original_askyesno = smt.messagebox.askyesno

        def _cancel_confirmation(*args, **kwargs):
            confirmation.append(
                " ".join(str(value) for value in args)
                + " "
                + " ".join(str(value) for value in kwargs.values())
            )
            return False

        smt.messagebox.askyesno = _cancel_confirmation
        try:
            view._set_copy_scope_mode("global")
            view._run_copy_action_by_mode("B2A")
            _pump(app.root, 0.08)
        finally:
            smt.messagebox.askyesno = original_askyesno
        assert confirmation, dialogs
        visible_evidence.append(("global_confirmation", confirmation[-1]))

    _assert_no_visible_l_notation(visible_evidence)
    combined = "\n".join(text for _source, text in visible_evidence)
    for label in ("A", "Z", "AA"):
        assert label in combined


_CASES = (
    ("gunships-header-pane-pixels", test_real_gunships_a1_a5_pixels_match_headers_panes_and_tags),
    ("gunships-combined-action-row", test_real_gunships_combined_action_row_is_centered_or_collision_safe),
    ("gunships-root-utility-layout", test_real_gunships_root_utilities_are_left_aligned_in_stable_order),
    ("excel-column-guidance", test_visible_column_guidance_uses_excel_a_z_aa_without_l_notation),
)


def _list_cases() -> None:
    for case_id, _test in _CASES:
        print(case_id)


def main(argv=None) -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--list-cases", action="store_true")
    parser.add_argument("--case", choices=tuple(case_id for case_id, _ in _CASES))
    args = parser.parse_args(argv)
    if args.list_cases:
        _list_cases()
        return
    selected = tuple(
        (case_id, test)
        for case_id, test in _CASES
        if args.case is None or case_id == args.case
    )
    # Default-suite execution remains serial and fail-fast.  A test exception
    # is deliberately re-raised instead of being accumulated and masked.
    for case_id, test in selected:
        started = time.perf_counter()
        case_deadline = time.monotonic() + 90.0
        previous_deadline = section10._ACTIVE_CASE_DEADLINE
        section10._ACTIVE_CASE_DEADLINE = case_deadline
        try:
            test()
            assert time.monotonic() <= case_deadline, (
                case_id,
                case_deadline,
                time.monotonic(),
            )
        except Exception:
            print(f"FAIL {case_id}", flush=True)
            raise
        finally:
            section10._ACTIVE_CASE_DEADLINE = previous_deadline
        print(f"PASS {case_id} elapsed_sec={time.perf_counter() - started:.3f}")
    print(f"PASS: OpenSpec section 11 acceptance ({len(selected)} case(s))")


if __name__ == "__main__":
    main()

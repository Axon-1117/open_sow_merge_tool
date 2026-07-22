import os
import sys
import argparse
import re
import bisect
import difflib
import tempfile
import subprocess
import traceback
import atexit
import copy
import gc
import math
from datetime import date, datetime, time as datetime_time, timedelta
import time
import stat
import shutil
import zipfile
import posixpath
import platform
import xml.etree.ElementTree as ET

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import json
import threading

from openpyxl import load_workbook as _openpyxl_load_workbook, Workbook
from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula
from openpyxl.formula.translate import Translator
# Note: formulas will be treated as cached values only (data_only), with fallback when cache is missing.
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import CALENDAR_MAC_1904, CALENDAR_WINDOWS_1900, to_excel


APP_NAME = "sow_merge_tool"
APP_VERSION = "2026-07-22.update52"
APP_BUILD_TAG = "new129-fast-ui-and-sheet-premark-fix"
_SUPPORTED_WORKBOOK_EXTS = (".xlsx", ".xlsm")

# Debug logging (writes to %TEMP%\sow_merge_tool_debug.log)
_DEBUG_LOG_PATH = os.path.join(tempfile.gettempdir(), f"{APP_NAME}_debug.log")
_DEBUG_ENABLED = True
_LAUNCH_TRACE_PATH = os.path.join(tempfile.gettempdir(), f"{APP_NAME}_launch_trace.log")

# Save safety: default to atomic "write tmp + os.replace" so an interrupted save
# can never leave a half-written / corrupted target. The fast direct-write path
# (faster, but not crash-safe) remains available behind this opt-in flag.
_FAST_SAVE_ENABLED = False
# Save correctness: keep workbook fidelity (styles/formulas/metadata).
# values-only fast save can make unrelated sheets look modified in SVN diff.
_FAST_SAVE_VALUES_ONLY = False
# Open performance: skip background preloads and global scans (loads on demand)
_FAST_OPEN_ENABLED = True
# Global mode: compare and save cached values only (ignore formulas)
_USE_CACHED_VALUES_ONLY = True
# When cached values are missing for formulas, try to recalc via Excel (if available)
_AUTO_RECALC_MISSING_CACHE = False
_AUTO_RECALC_FORMULAS_ALWAYS = False
# Recalculation is explicit. Automatic Excel recalculation is both expensive on
# large workbooks and unsafe during merge because a late result can overwrite
# cache values the user has just adopted from theirs.
_AUTO_RECALC_ON_OPEN = False
_EXCEL_NATIVE_SAVE_ON_MERGE = True
_CACHE_CHECK_MAX_CELLS = 3000
# Render performance: limit initial rows rendered (user can load full)
_FAST_RENDER_ROW_LIMIT = 800
_FAST_RENDER_BATCH = 500
_LARGE_SHEET_ROW_THRESHOLD = 2000
_LARGE_SHEET_INITIAL_ROWS = 200
_LARGE_SHEET_BLOCK_ROWS = 1000
_ROW_ALIGN_MAX_ROWS = 2000
_ROW_ALIGN_SOFT_MAX_ROWS = 50000
_TABMARK_QUICK_TAIL_ROWS = 2000
# Retained for the disabled legacy row-tail scanner below.
_FAST_TABMARK_SCAN_SKIP_MB = 25
# Stage-1 tab coloring reads only XLSX ZIP metadata and never parses worksheet
# XML. Exact background compute later confirms bright-yellow state or clears it.
_FAST_TABMARK_ENABLED = True
_FAST_TABMARK_PHASE2_ENABLED = False
_SVN_EXPORT_TIMEOUT_SECS = 15
# Grid display: max chars shown per cell before truncation, and column separator
_COL_MAX_DISPLAY_WIDTH = 30
_COL_SEP = " \u2502 "    # 3-char separator between columns (U+2502 BOX DRAWINGS LIGHT VERTICAL)
_COL_SEP_LEN = 3

# Unified pane colors (main 3-way panes and C-area rows)
_MINE_BG = "#F6C16B"
_BASE_BG = "#E3E3FF"
_THEIRS_BG = "#FFF176"
_DIFF_CELL_BG = "#FF2D2D"

# Unified row-header hover/action arrows (keep one visual family).
_ROW_ARROW_RIGHT = "➡"
_ROW_ARROW_LEFT = "⬅"

# Settings (persist UI prefs)
_SETTINGS_PATH = os.path.join(os.environ.get("LOCALAPPDATA", tempfile.gettempdir()), APP_NAME, "settings.json")
_STARTUP_PROGRESS_ROOT = None


def _workbook_ext(path: str | None, default: str = ".xlsx") -> str:
    """Infer workbook extension even for SVN sidecar/temp filenames."""
    base = os.path.basename(str(path or ""))
    ext = os.path.splitext(base)[1].lower()
    if ext in _SUPPORTED_WORKBOOK_EXTS:
        return ext
    m = re.search(r"\.(xlsx|xlsm)", base, flags=re.IGNORECASE)
    if m:
        return "." + m.group(1).lower()
    return default


def _is_macro_enabled_workbook(path: str | None) -> bool:
    return _workbook_ext(path) == ".xlsm"


def _dlog(msg: str):
    if not _DEBUG_ENABLED:
        return
    try:
        ts = datetime.now().strftime("%H:%M:%S.%f")[:-3]
        with open(_DEBUG_LOG_PATH, "a", encoding="utf-8") as f:
            f.write(f"[{ts}] {msg}\n")
    except Exception:
        pass


_WORKBOOK_REPAIR_CACHE: dict[tuple[str, int, int], str] = {}


def _cleanup_repair_cache():
    """Remove all temporary repaired workbook copies on process exit."""
    for path in list(_WORKBOOK_REPAIR_CACHE.values()):
        try:
            if os.path.exists(path):
                os.remove(path)
        except Exception:
            pass
    _WORKBOOK_REPAIR_CACHE.clear()


atexit.register(_cleanup_repair_cache)


def _workbook_sig(path: str) -> tuple[str, int, int]:
    p = os.path.abspath(path)
    try:
        st = os.stat(p)
        return p, int(getattr(st, "st_mtime_ns", 0)), int(st.st_size)
    except Exception:
        return p, 0, 0


def _repair_missing_shared_strings_part(xlsx_path: str) -> str | None:
    """Create a temporary repaired copy when sharedStrings.xml is referenced but missing.

    Some third-party exports leave the sharedStrings part declared in [Content_Types].xml
    but omit xl/sharedStrings.xml entirely. openpyxl raises KeyError on open in that case.
    We inject an empty sharedStrings.xml into a temp copy and retry the load.
    """
    if not xlsx_path or (not os.path.isfile(xlsx_path)):
        return None
    key = _workbook_sig(xlsx_path)
    cached = _WORKBOOK_REPAIR_CACHE.get(key)
    if cached and os.path.isfile(cached):
        return cached
    try:
        with zipfile.ZipFile(xlsx_path, "r") as zf:
            names = zf.namelist()
            if "xl/sharedStrings.xml" in names:
                return xlsx_path
            original_infos = {n: zf.getinfo(n) for n in names}
            all_bytes = {n: zf.read(n) for n in names}
    except Exception as e:
        _dlog(f"repair sharedStrings open failed: path={xlsx_path} err={e}")
        return None

    all_bytes["xl/sharedStrings.xml"] = (
        b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        b'<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="0" uniqueCount="0"/>'
    )

    ct_name = "[Content_Types].xml"
    ct_bytes = all_bytes.get(ct_name)
    if ct_bytes is not None:
        try:
            ct_root = ET.fromstring(ct_bytes)
            if ct_root.tag.startswith("{"):
                ns_uri = ct_root.tag[1:].split("}", 1)[0]
                q = lambda tag: f"{{{ns_uri}}}{tag}"
            else:
                q = lambda tag: tag
            has_override = False
            for node in ct_root.findall(q("Override")):
                if node.get("PartName") == "/xl/sharedStrings.xml":
                    has_override = True
                    break
            if not has_override:
                ET.SubElement(
                    ct_root,
                    q("Override"),
                    {
                        "PartName": "/xl/sharedStrings.xml",
                        "ContentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml",
                    },
                )
                all_bytes[ct_name] = ET.tostring(ct_root, encoding="utf-8", xml_declaration=True)
        except Exception as e:
            _dlog(f"repair sharedStrings content-types failed: path={xlsx_path} err={e}")

    base = os.path.basename(xlsx_path)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    repaired = os.path.join(tempfile.gettempdir(), f"{APP_NAME}_repair_sst_{os.getpid()}_{ts}_{base}")
    ext = _workbook_ext(xlsx_path)
    if not repaired.lower().endswith(ext):
        repaired += ext
    try:
        with zipfile.ZipFile(repaired, "w") as zf:
            for name, payload in all_bytes.items():
                info = original_infos.get(name)
                comp = info.compress_type if info else zipfile.ZIP_DEFLATED
                zf.writestr(name, payload, compress_type=comp)
        _WORKBOOK_REPAIR_CACHE[key] = repaired
        _dlog(f"repair sharedStrings created: src={xlsx_path} repaired={repaired}")
        return repaired
    except Exception as e:
        _dlog(f"repair sharedStrings write failed: path={xlsx_path} err={e}")
        try:
            if os.path.exists(repaired):
                os.remove(repaired)
        except Exception:
            pass
        return None


def load_workbook(filename, *args, **kwargs):
    if "keep_vba" not in kwargs and _is_macro_enabled_workbook(str(filename)):
        kwargs["keep_vba"] = True
    try:
        return _openpyxl_load_workbook(filename, *args, **kwargs)
    except KeyError as e:
        if "sharedStrings.xml" not in str(e):
            raise
        path = str(filename)
        repaired = _repair_missing_shared_strings_part(path)
        if not repaired:
            raise
        _dlog(f"load_workbook retry with repaired sharedStrings: src={path} repaired={repaired}")
        return _openpyxl_load_workbook(repaired, *args, **kwargs)


def _blank_worksheet(title: str = "Sheet"):
    wb = Workbook()
    ws = wb.active
    ws.title = title
    return wb, ws


def _row_sig_list_for_ws(ws, max_row_local: int, max_col: int):
    try:
        all_rows = ws.iter_rows(
            min_row=1,
            max_row=max_row_local,
            min_col=1,
            max_col=max_col,
            values_only=True,
        )
    except Exception:
        return []
    sigs = []
    for row in all_rows:
        sigs.append(_row_signature(row or ()))
    return sigs


def _row_signature(row_values) -> str:
    return "\x1f".join(
        _merge_cmp_value(value).replace("\x1f", "\x1e")
        for value in (row_values or ())
    )


def _compute_row_pairs_from_signatures(sig_a: list[str], sig_b: list[str]):
    """Align precomputed row signatures without rescanning worksheets.

    Trimming the exact common prefix/suffix first avoids SequenceMatcher's
    quadratic behavior on large sheets containing many repeated rows while
    preserving exact alignment inside the changed middle section.
    """
    len_a = len(sig_a)
    len_b = len(sig_b)
    prefix = 0
    common_len = min(len_a, len_b)
    while prefix < common_len and sig_a[prefix] == sig_b[prefix]:
        prefix += 1

    suffix = 0
    while (
        suffix < (common_len - prefix)
        and sig_a[len_a - suffix - 1] == sig_b[len_b - suffix - 1]
    ):
        suffix += 1

    pairs: list[tuple[int | None, int | None]] = [
        (idx + 1, idx + 1) for idx in range(prefix)
    ]

    end_a = len_a - suffix
    end_b = len_b - suffix
    mid_a = sig_a[prefix:end_a]
    mid_b = sig_b[prefix:end_b]

    def _sim_score(sa: str, sb: str) -> float:
        if sa == sb:
            return 2.0
        if (not sa) or (not sb):
            return 0.0
        try:
            return difflib.SequenceMatcher(a=sa, b=sb, autojunk=False).ratio()
        except Exception:
            return 0.0

    sm = difflib.SequenceMatcher(a=mid_a, b=mid_b, autojunk=False)
    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        ai1 = prefix + i1
        ai2 = prefix + i2
        bj1 = prefix + j1
        bj2 = prefix + j2
        if tag == "equal":
            for i, j in zip(range(ai1, ai2), range(bj1, bj2)):
                pairs.append((i + 1, j + 1))
        elif tag == "replace":
            block_len_a = ai2 - ai1
            block_len_b = bj2 - bj1
            common = min(block_len_a, block_len_b)
            # Equal-length replacement blocks have exactly the same row pairing
            # whether aligned from the head or tail. Avoid expensive per-row
            # character SequenceMatcher calls (a major cost on formula-heavy
            # workbooks such as WorldMonster.xlsx).
            if block_len_a == block_len_b:
                for k in range(common):
                    pairs.append((ai1 + k + 1, bj1 + k + 1))
                continue
            head_score = 0.0
            tail_score = 0.0
            # For very large unequal replace blocks the choice is still only
            # head-vs-tail. A bounded, evenly distributed sample preserves that
            # decision without quadratic character matching across every row.
            if common <= 128:
                sample_offsets = range(common)
            else:
                sample_count = 64
                sample_offsets = sorted({
                    int(round(idx * (common - 1) / (sample_count - 1)))
                    for idx in range(sample_count)
                })
            for k in sample_offsets:
                head_score += _sim_score(sig_a[ai1 + k], sig_b[bj1 + k])
                tail_score += _sim_score(sig_a[ai2 - common + k], sig_b[bj2 - common + k])
            use_tail = tail_score >= head_score
            if use_tail:
                extra_a = block_len_a - common
                extra_b = block_len_b - common
                for k in range(extra_a):
                    pairs.append((ai1 + k + 1, None))
                for k in range(extra_b):
                    pairs.append((None, bj1 + k + 1))
                a_start = ai2 - common
                b_start = bj2 - common
                for k in range(common):
                    pairs.append((a_start + k + 1, b_start + k + 1))
            else:
                for k in range(common):
                    pairs.append((ai1 + k + 1, bj1 + k + 1))
                for k in range(common, block_len_a):
                    pairs.append((ai1 + k + 1, None))
                for k in range(common, block_len_b):
                    pairs.append((None, bj1 + k + 1))
        elif tag == "delete":
            for i in range(ai1, ai2):
                pairs.append((i + 1, None))
        elif tag == "insert":
            for j in range(bj1, bj2):
                pairs.append((None, j + 1))

    if suffix:
        start_a = len_a - suffix
        start_b = len_b - suffix
        for k in range(suffix):
            pairs.append((start_a + k + 1, start_b + k + 1))
    return pairs


def _compute_row_pairs_generic(ws_a, ws_b, max_col: int, force: bool = False):
    """Compute row alignment pairs between two worksheets."""
    max_row_a = ws_a.max_row or 1
    max_row_b = ws_b.max_row or 1
    max_row = max(max_row_a, max_row_b)
    if max_row <= 0:
        return []
    if not _should_auto_row_align(max_row_a, max_row_b, force=force):
        return [(r if r <= max_row_a else None, r if r <= max_row_b else None) for r in range(1, max_row + 1)]

    sig_a = _row_sig_list_for_ws(ws_a, max_row_a, max_col)
    sig_b = _row_sig_list_for_ws(ws_b, max_row_b, max_col)

    return _compute_row_pairs_from_signatures(sig_a, sig_b)


def _row_map_from_pairs(pairs: list[tuple[int | None, int | None]]) -> dict[int, int]:
    out = {}
    for left, right in pairs:
        if left is not None and right is not None:
            out[left] = right
    return out


def _split_tail_independent_append_pairs(
    display_pairs: list[tuple[int | None, int | None]],
    mine_to_base: dict[int, int] | None,
    theirs_to_base: dict[int, int] | None,
    ws_mine=None,
    ws_theirs=None,
    max_col: int | None = None,
) -> list[tuple[int | None, int | None]]:
    """Split safe 3-way tail append pairs into independent theirs/mine blocks.

    Conservative rule:
    - only consider pairs where both mine/theirs rows exist
    - neither row maps to base
    - both rows are strictly after their respective last mapped base row
    - require each side to have at least one mapped base row first

    Output order is fixed as: theirs block first, mine block second.
    """
    pairs = list(display_pairs or [])
    if not pairs:
        return pairs
    mine_map = mine_to_base or {}
    theirs_map = theirs_to_base or {}
    if not mine_map or not theirs_map:
        return pairs
    try:
        last_mine_mapped = max(int(r) for r in mine_map.keys() if r is not None)
        last_theirs_mapped = max(int(r) for r in theirs_map.keys() if r is not None)
    except Exception:
        return pairs

    candidate_rows: list[tuple[int, int, int]] = []
    for idx, (ra, rb) in enumerate(pairs):
        if ra is None or rb is None:
            continue
        if ra in mine_map or rb in theirs_map:
            continue
        if int(ra) <= last_mine_mapped or int(rb) <= last_theirs_mapped:
            continue
        candidate_rows.append((idx, int(ra), int(rb)))

    split_indices: set[int] = set()
    if candidate_rows:
        mine_rows = {}
        theirs_rows = {}
        if ws_mine is not None and ws_theirs is not None and max_col:
            # A read_only worksheet may parse from the start of its XML stream for
            # every high-row iter_rows() call. Read all candidate rows once per
            # side so multiple tail appends stay linear rather than O(rows * XML).
            mine_rows = _read_rows_into_cache(
                ws_mine,
                [ra for _idx, ra, _rb in candidate_rows],
                int(max_col),
                require_complete=True,
            )
            theirs_rows = _read_rows_into_cache(
                ws_theirs,
                [rb for _idx, _ra, rb in candidate_rows],
                int(max_col),
                require_complete=True,
            )
        for idx, ra, rb in candidate_rows:
            # Identical independent appends are one logical row, not two
            # competing blocks. If equality cannot be proven, retain the
            # conservative split.
            if mine_rows and theirs_rows:
                mine_sig = _row_signature(_row_from_cache(mine_rows, ra, int(max_col)))
                theirs_sig = _row_signature(_row_from_cache(theirs_rows, rb, int(max_col)))
                if mine_sig == theirs_sig:
                    continue
            split_indices.add(idx)

    if not split_indices:
        return pairs

    out: list[tuple[int | None, int | None]] = []
    idx = 0
    while idx < len(pairs):
        if idx not in split_indices:
            out.append(pairs[idx])
            idx += 1
            continue
        end = idx
        while end + 1 < len(pairs) and (end + 1) in split_indices:
            end += 1
        # Keep the agreed deterministic order for an independent tail block:
        # all theirs rows first, followed by all mine rows.
        out.extend((None, pairs[pos][1]) for pos in range(idx, end + 1))
        out.extend((pairs[pos][0], None) for pos in range(idx, end + 1))
        idx = end + 1
    return out


def _split_low_similarity_tail_pairs(
    display_pairs: list[tuple[int | None, int | None]],
    mine_to_base: dict[int, int] | None,
    theirs_to_base: dict[int, int] | None,
    ws_mine,
    ws_theirs,
    max_col: int,
    *,
    sim_threshold: float = 0.75,
) -> list[tuple[int | None, int | None]]:
    """Split misaligned tail pairs before a mapped one-sided tail block.

    Conservative rule:
    - only consider the very end of the sheet
    - require a trailing mapped one-sided block on one side
    - only split immediately preceding paired rows that map to the matching
      base rows but have low row similarity

    This preserves exact behavior for the body of the sheet while allowing
    "local tail append vs remote tail append" cases to stay visible as
    independent rows in 3-way mode.
    """
    pairs = list(display_pairs or [])
    mine_map = mine_to_base or {}
    theirs_map = theirs_to_base or {}
    if not pairs or not mine_map or not theirs_map:
        return pairs

    def _row_sim_cache(ws_a, rows_a, ws_b, rows_b):
        rows_a_cache = _read_rows_into_cache(ws_a, rows_a, max_col)
        rows_b_cache = _read_rows_into_cache(ws_b, rows_b, max_col)
        sim_cache: dict[tuple[int, int], float] = {}
        for ra, rb in zip(rows_a, rows_b):
            row_a = _row_from_cache(rows_a_cache, ra, max_col)
            row_b = _row_from_cache(rows_b_cache, rb, max_col)
            sig_a = _row_signature(row_a)
            sig_b = _row_signature(row_b)
            if sig_a == sig_b:
                sim_cache[(ra, rb)] = 2.0
                continue
            if (not sig_a) or (not sig_b):
                sim_cache[(ra, rb)] = 0.0
                continue
            try:
                sim_cache[(ra, rb)] = difflib.SequenceMatcher(a=sig_a, b=sig_b, autojunk=False).ratio()
            except Exception:
                sim_cache[(ra, rb)] = 0.0
        return sim_cache

    def _split_one_side(
        cur_pairs: list[tuple[int | None, int | None]],
        *,
        trailing_side: str,
    ) -> list[tuple[int | None, int | None]]:
        if trailing_side == "B":
            trailing_idx = 1
            lead_idx = 0
            trailing_map = theirs_map
            lead_map = mine_map
        else:
            trailing_idx = 0
            lead_idx = 1
            trailing_map = mine_map
            lead_map = theirs_map

        cursor = len(cur_pairs) - 1
        suffix_start = len(cur_pairs)
        found_one_sided = False
        candidate_rows_a: list[int] = []
        candidate_rows_b: list[int] = []
        candidate_idx_pairs: list[tuple[int, int, int]] = []
        while cursor >= 0:
            pair = cur_pairs[cursor]
            lead_row = pair[lead_idx]
            trail_row = pair[trailing_idx]
            if lead_row is None and trail_row is not None:
                base_row = trailing_map.get(trail_row)
                if base_row is None:
                    break
                found_one_sided = True
                suffix_start = cursor
                cursor -= 1
                continue
            ra, rb = pair
            if ra is None or rb is None:
                break
            base_a = mine_map.get(ra)
            base_b = theirs_map.get(rb)
            if base_a is None or base_b is None or base_a != base_b:
                break
            candidate_rows_a.append(ra)
            candidate_rows_b.append(rb)
            candidate_idx_pairs.append((cursor, ra, rb))
            suffix_start = cursor
            cursor -= 1
        if (not found_one_sided) or (not candidate_idx_pairs):
            return cur_pairs

        sim_cache = _row_sim_cache(ws_mine, candidate_rows_a, ws_theirs, candidate_rows_b)

        split_indices: set[int] = set()
        for global_idx, ra, rb in sorted(candidate_idx_pairs, key=lambda x: x[0], reverse=True):
            if sim_cache.get((ra, rb), 0.0) >= sim_threshold:
                break
            split_indices.add(global_idx)
        if not split_indices:
            return cur_pairs

        out: list[tuple[int | None, int | None]] = list(cur_pairs[:suffix_start])
        seen_one_sided = False
        for global_idx in range(suffix_start, len(cur_pairs)):
            ra, rb = cur_pairs[global_idx]
            lead_row = cur_pairs[global_idx][lead_idx]
            trail_row = cur_pairs[global_idx][trailing_idx]
            if lead_row is None and trail_row is not None:
                out.append(cur_pairs[global_idx])
                seen_one_sided = True
                continue
            if global_idx not in split_indices:
                out.append(cur_pairs[global_idx])
                continue
            if trailing_side == "B":
                if seen_one_sided:
                    out.append((ra, None))
                    out.append((None, rb))
                else:
                    out.append((None, rb))
                    out.append((ra, None))
            else:
                if seen_one_sided:
                    out.append((None, rb))
                    out.append((ra, None))
                else:
                    out.append((ra, None))
                    out.append((None, rb))
        return out

    pairs = _split_one_side(pairs, trailing_side="B")
    pairs = _split_one_side(pairs, trailing_side="A")
    return pairs


def _build_pair_base_row_overrides(
    row_pairs: list[tuple[int | None, int | None]],
    mine_to_base: dict[int, int] | None,
    theirs_to_base: dict[int, int] | None,
    ws_base,
    ws_mine,
    ws_theirs,
    max_col: int,
    *,
    sim_delta: float = 0.05,
) -> dict[int, int | None]:
    """Choose which side should own the base row for adjacent split pairs."""
    pairs = list(row_pairs or [])
    mine_map = mine_to_base or {}
    theirs_map = theirs_to_base or {}
    if (not pairs) or ws_base is None or (not mine_map and not theirs_map):
        return {}

    twin_specs: list[tuple[int, int, int, int, int]] = []
    base_rows_needed: set[int] = set()
    mine_rows_needed: set[int] = set()
    theirs_rows_needed: set[int] = set()

    for idx in range(len(pairs) - 1):
        p1 = pairs[idx]
        p2 = pairs[idx + 1]
        base_1 = mine_map.get(p1[0]) if p1[0] is not None else theirs_map.get(p1[1]) if p1[1] is not None else None
        base_2 = mine_map.get(p2[0]) if p2[0] is not None else theirs_map.get(p2[1]) if p2[1] is not None else None
        if base_1 is None or base_2 is None or base_1 != base_2:
            continue
        if p1[0] is None and p1[1] is not None and p2[0] is not None and p2[1] is None:
            theirs_pair_idx, mine_pair_idx = idx, idx + 1
            theirs_row, mine_row = p1[1], p2[0]
        elif p1[0] is not None and p1[1] is None and p2[0] is None and p2[1] is not None:
            mine_pair_idx, theirs_pair_idx = idx, idx + 1
            mine_row, theirs_row = p1[0], p2[1]
        else:
            continue
        twin_specs.append((mine_pair_idx, theirs_pair_idx, mine_row, theirs_row, base_1))
        base_rows_needed.add(base_1)
        mine_rows_needed.add(mine_row)
        theirs_rows_needed.add(theirs_row)

    if not twin_specs:
        return {}

    rows_base = _read_rows_into_cache(ws_base, sorted(base_rows_needed), max_col)
    rows_mine = _read_rows_into_cache(ws_mine, sorted(mine_rows_needed), max_col)
    rows_theirs = _read_rows_into_cache(ws_theirs, sorted(theirs_rows_needed), max_col)

    def _sim_against_base(side_cache, side_row: int, base_row: int) -> float:
        row_side = _row_from_cache(side_cache, side_row, max_col)
        row_base = _row_from_cache(rows_base, base_row, max_col)
        sig_side = "\x1f".join(_merge_cmp_value(v) for v in row_side)
        sig_base = "\x1f".join(_merge_cmp_value(v) for v in row_base)
        if sig_side == sig_base:
            return 2.0
        if (not sig_side) or (not sig_base):
            return 0.0
        try:
            return difflib.SequenceMatcher(a=sig_side, b=sig_base, autojunk=False).ratio()
        except Exception:
            return 0.0

    overrides: dict[int, int | None] = {}
    for mine_pair_idx, theirs_pair_idx, mine_row, theirs_row, base_row in twin_specs:
        sim_mine = _sim_against_base(rows_mine, mine_row, base_row)
        sim_theirs = _sim_against_base(rows_theirs, theirs_row, base_row)
        if abs(sim_mine - sim_theirs) < sim_delta:
            continue
        if sim_mine > sim_theirs:
            overrides[mine_pair_idx] = base_row
            overrides[theirs_pair_idx] = None
        else:
            overrides[mine_pair_idx] = None
            overrides[theirs_pair_idx] = base_row
    return overrides


def _copy_sheet_basic(src_ws, dst_ws):
    """Copy enough worksheet structure for merge/save fallback paths."""
    for row in src_ws.iter_rows():
        for src_cell in row:
            try:
                dst_cell = dst_ws.cell(row=src_cell.row, column=src_cell.column)
                src_value = src_cell.value
                if src_cell.data_type == "s" and isinstance(src_value, str) and src_value.startswith("="):
                    src_value = _LiteralText(src_value)
                _assign_edit_cell_value(dst_cell, src_value)
                if src_cell.has_style:
                    dst_cell._style = copy.copy(src_cell._style)
                if src_cell.number_format:
                    dst_cell.number_format = src_cell.number_format
                if src_cell.font:
                    dst_cell.font = copy.copy(src_cell.font)
                if src_cell.fill:
                    dst_cell.fill = copy.copy(src_cell.fill)
                if src_cell.border:
                    dst_cell.border = copy.copy(src_cell.border)
                if src_cell.alignment:
                    dst_cell.alignment = copy.copy(src_cell.alignment)
                if src_cell.protection:
                    dst_cell.protection = copy.copy(src_cell.protection)
                if src_cell.comment is not None:
                    dst_cell.comment = copy.copy(src_cell.comment)
                if src_cell.hyperlink is not None:
                    dst_cell._hyperlink = copy.copy(src_cell.hyperlink)
            except Exception as exc:
                raise RuntimeError(
                    f"复制 Sheet 单元格失败：{src_ws.title}!{src_cell.coordinate}: {exc}"
                ) from exc
    try:
        for key, dim in src_ws.row_dimensions.items():
            dst_ws.row_dimensions[key] = copy.copy(dim)
    except Exception:
        pass
    try:
        for key, dim in src_ws.column_dimensions.items():
            dst_ws.column_dimensions[key] = copy.copy(dim)
    except Exception:
        pass
    try:
        for merged in list(src_ws.merged_cells.ranges):
            dst_ws.merge_cells(str(merged))
    except Exception:
        pass
    for attr in (
        "sheet_format",
        "sheet_properties",
        "sheet_view",
        "views",
        "page_margins",
        "page_setup",
        "print_options",
        "sheet_state",
    ):
        try:
            setattr(dst_ws, attr, copy.copy(getattr(src_ws, attr)))
        except Exception:
            pass
    try:
        dst_ws.freeze_panes = src_ws.freeze_panes
    except Exception:
        pass
    try:
        dst_ws.auto_filter = copy.copy(src_ws.auto_filter)
    except Exception:
        pass


def _remove_sheet_if_exists(wb, sheet_name: str):
    if wb is None or (not sheet_name):
        return False
    if sheet_name not in wb.sheetnames:
        return False
    wb.remove(wb[sheet_name])
    return True


def _create_sheet_from_source(dst_wb, src_ws, title: str, index: int | None = None):
    if dst_wb is None or src_ws is None or not title:
        return None
    if title in dst_wb.sheetnames:
        _remove_sheet_if_exists(dst_wb, title)
    if index is None:
        dst_ws = dst_wb.create_sheet(title=title)
    else:
        dst_ws = dst_wb.create_sheet(title=title, index=max(0, int(index)))
    _copy_sheet_basic(src_ws, dst_ws)
    return dst_ws


def _copy_row_metadata(src_ws, dst_ws, src_row: int, dst_row: int, max_col: int):
    """Copy row/cell presentation metadata without changing cell values."""
    if src_ws is None or dst_ws is None:
        return
    src_row = int(src_row)
    dst_row = int(dst_row)
    for col in range(1, max(1, int(max_col)) + 1):
        src_cell = src_ws.cell(row=src_row, column=col)
        dst_cell = dst_ws.cell(row=dst_row, column=col)
        if src_cell.has_style:
            dst_cell._style = copy.copy(src_cell._style)
        if src_cell.comment is not None:
            dst_cell.comment = copy.copy(src_cell.comment)
        if src_cell.hyperlink is not None:
            dst_cell._hyperlink = copy.copy(src_cell.hyperlink)
    src_dim = src_ws.row_dimensions.get(src_row)
    if src_dim is not None:
        dst_dim = copy.copy(src_dim)
        try:
            dst_dim.index = dst_row
        except Exception:
            pass
        dst_ws.row_dimensions[dst_row] = dst_dim


def _val_to_str(v):
    """Render a cell value as single-line text for the Text widget.

    IMPORTANT: We must keep each Excel row rendered as exactly ONE line in tk.Text.
    So we sanitize embedded newlines/tabs that would otherwise break line alignment
    and cause diff highlights to drift.
    """
    if v is None:
        return ""
    s = str(v)
    # Normalize line breaks and tabs to keep one-row-per-line invariant
    s = s.replace("\r\n", " ").replace("\r", " ").replace("\n", " ")
    s = s.replace("\t", "    ")
    return s


def _format_cell(val_str: str, width: int) -> str:
    """Left-justify val_str in exactly `width` chars; truncate with ellipsis if too long."""
    if len(val_str) > width:
        return val_str[:width - 1] + "\u2026"  # … (single char)
    return val_str.ljust(width)


def _should_auto_row_align(max_row_a: int, max_row_b: int, force: bool = False) -> bool:
    """Auto-enable row alignment when it prevents cascading false diffs.

    Row counts can remain equal when one side inserts a row and deletes another, so
    count equality is not evidence that raw row-number comparison is safe. Align all
    sheets under the soft limit and leave larger sheets to their dedicated block/tail
    paths.
    """
    max_row = max(max_row_a, max_row_b)
    if max_row <= 0:
        return False
    if force:
        return True
    return max_row <= _ROW_ALIGN_SOFT_MAX_ROWS


def _effective_bounds(ws):
    """Return (max_row, max_col) using the true last non-empty row.

    Empty strings are treated as empty, so formulas returning "" do not keep a sheet
    artificially large. On read-only worksheets we avoid shrinking max_col from a
    single tail row because that can truncate real columns seen earlier in the sheet.
    """
    max_r = ws.max_row or 1
    max_c = ws.max_column or 1
    last_r = 1
    last_c = 1
    found = False
    found_via_cells = False
    try:
        cells = getattr(ws, "_cells", None)
        if cells:
            for cell in cells.values():
                v = cell.value
                if v not in (None, ""):
                    found = True
                    found_via_cells = True
                    if cell.row > last_r:
                        last_r = cell.row
                    if cell.column > last_c:
                        last_c = cell.column
    except Exception:
        pass
    if not found:
        try:
            for r in range(max_r, max(0, max_r - 5000), -1):
                row = next(ws.iter_rows(min_row=r, max_row=r, min_col=1, max_col=max_c, values_only=True), ())
                if any(v not in (None, "") for v in row):
                    found = True
                    last_r = r
                    break
        except Exception:
            pass
    if not found:
        return 1, max(1, max_c)
    if found_via_cells:
        # data_only mode can read uncached formula cells as None, which would
        # under-estimate columns when relying only on non-empty _cells.
        if _USE_CACHED_VALUES_ONLY and max_c > last_c:
            return max(1, last_r), max(1, max_c)
        if last_c <= 1 and max_c > last_c:
            return max(1, last_r), max(1, max_c)
    return max(1, last_r), max(1, last_c)


def _effective_bounds_with_edit(ws_val, ws_edit=None):
    """Include formula/literal cells whose data-only cache is empty."""
    val_row, val_col = _effective_bounds(ws_val)
    if ws_edit is None or ws_edit is ws_val:
        return val_row, val_col
    try:
        edit_row, edit_col = _effective_bounds(ws_edit)
        return max(val_row, edit_row), max(val_col, edit_col)
    except Exception:
        return val_row, val_col
    return max(1, last_r), max(1, max_c)


def _save_values_only_from_wb(src_wb, target_path: str):
    """Fast save: values only, no styles. Drops formatting."""
    def _trim_bounds_ws(ws):
        # Find the true last non-empty row. Empty strings count as empty so formulas
        # returning "" do not keep trailing blank regions alive. Keep max_c for
        # read-only fallback paths so earlier populated columns are not truncated.
        max_r = ws.max_row or 1
        max_c = ws.max_column or 1
        last_r = 1
        last_c = 1
        found = False
        found_via_cells = False
        try:
            cells = getattr(ws, "_cells", None)
            if cells:
                for cell in cells.values():
                    v = cell.value
                    if v not in (None, ""):
                        found = True
                        found_via_cells = True
                        if cell.row > last_r:
                            last_r = cell.row
                        if cell.column > last_c:
                            last_c = cell.column
        except Exception:
            pass
        if not found:
            for r in range(max_r, max(0, max_r - 5000), -1):
                row = next(ws.iter_rows(min_row=r, max_row=r, min_col=1, max_col=max_c, values_only=True), ())
                if any(v not in (None, "") for v in row):
                    found = True
                    last_r = r
                    break
        if not found:
            return 1, max(1, max_c)
        if found_via_cells:
            # Keep columns conservative in cached-values mode to avoid losing
            # uncached formula columns when _cells appears sparse.
            if _USE_CACHED_VALUES_ONLY and max_c > last_c:
                return max(1, last_r), max(1, max_c)
            if last_c <= 1 and max_c > last_c:
                return max(1, last_r), max(1, max_c)
            return max(1, last_r), max(1, last_c)
        return max(1, last_r), max(1, max_c)

    dst = Workbook(write_only=True)
    # Remove default sheet
    try:
        if dst.sheetnames:
            dst.remove(dst[dst.sheetnames[0]])
    except Exception:
        pass
    for name in src_wb.sheetnames:
        ws_src = src_wb[name]
        ws_dst = dst.create_sheet(title=name)
        max_row, max_col = _trim_bounds_ws(ws_src)
        if max_row <= 0 or max_col <= 0:
            continue
        for row in ws_src.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True):
            ws_dst.append(list(row))
    try:
        dst.save(target_path)
    finally:
        try:
            dst.close()
        except Exception:
            pass


def _capture_external_link_parts(xlsx_path: str):
    """Capture fragile package parts to preserve them across saves.

    Only activate this path when the workbook actually contains externalLinks parts.
    Capturing workbook/content-types metadata for every normal workbook can overwrite
    openpyxl's newly written package metadata and produce invalid archives.
    """
    if not xlsx_path or (not os.path.isfile(xlsx_path)):
        return None
    try:
        with zipfile.ZipFile(xlsx_path, "r") as zf:
            names = zf.namelist()
            target_names = set(
                n for n in names
                if n.startswith("xl/externalLinks/")
            )
            if not target_names:
                return None
            # workbook rels may carry externalLink relationship targets
            if "xl/_rels/workbook.xml.rels" in names:
                target_names.add("xl/_rels/workbook.xml.rels")
            # Preserve workbook/content-types metadata only for real external-link workbooks.
            if "xl/workbook.xml" in names:
                target_names.add("xl/workbook.xml")
            if "[Content_Types].xml" in names:
                target_names.add("[Content_Types].xml")
            parts = {n: zf.read(n) for n in target_names}
            _dlog(f"capture fragile parts: path={xlsx_path} count={len(parts)}")
            return parts
    except Exception:
        return None


def _restore_external_link_parts(xlsx_path: str, parts) -> int:
    """Restore captured fragile package parts into saved workbook."""
    if not parts:
        return 0
    if not xlsx_path or (not os.path.isfile(xlsx_path)):
        return 0
    try:
        with zipfile.ZipFile(xlsx_path, "r") as zf:
            names = zf.namelist()
            original_infos = {n: zf.getinfo(n) for n in names}
            all_bytes = {n: zf.read(n) for n in names}
    except Exception:
        return 0

    # Merge: keep all saved parts, overwrite/add captured external-link parts.
    for n, b in parts.items():
        all_bytes[n] = b

    # Keep package metadata internally consistent. Some originals carry a sharedStrings
    # override while openpyxl rewrites strings inline and omits xl/sharedStrings.xml.
    # If we restore the old [Content_Types].xml, add an empty sharedStrings part back.
    try:
        ct_name = "[Content_Types].xml"
        ct_bytes = all_bytes.get(ct_name)
        if ct_bytes is not None and "xl/sharedStrings.xml" not in all_bytes:
            ct_root = ET.fromstring(ct_bytes)
            for node in ct_root.iter():
                if node.tag.endswith('Override') and node.get('PartName') == '/xl/sharedStrings.xml':
                    all_bytes['xl/sharedStrings.xml'] = (
                        b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                        b'<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="0" uniqueCount="0"/>'
                    )
                    break
    except Exception as e:
        _dlog(f"fragile parts sharedStrings reconcile failed: {xlsx_path} err={e}")

    changed = len(parts)

    tmp = f"{xlsx_path}.extlinkrestore.{os.getpid()}.tmp"
    try:
        with zipfile.ZipFile(tmp, "w") as zf:
            for name in all_bytes.keys():
                info = original_infos.get(name)
                comp = info.compress_type if info else zipfile.ZIP_DEFLATED
                zf.writestr(name, all_bytes[name], compress_type=comp)
        os.replace(tmp, xlsx_path)
        _dlog(f"fragile parts restored: {xlsx_path} changed={changed}")
        return changed
    except Exception as e:
        _dlog(f"fragile parts restore failed: {xlsx_path} err={e}")
        try:
            if os.path.exists(tmp):
                os.remove(tmp)
        except Exception:
            pass
        return 0


def _apply_external_link_parts_on_file(path: str, external_link_parts=None):
    try:
        if path:
            _restore_external_link_parts(path, external_link_parts)
    except Exception as e:
        _dlog(f"apply fragile parts failed: path={path} err={e}")


def _cell_display_and_equal(ws_a_val, ws_b_val, ws_a_edit, ws_b_edit, r: int, c: int):
    va_val = ws_a_val.cell(row=r, column=c).value
    vb_val = ws_b_val.cell(row=r, column=c).value

    va_edit = ws_a_edit.cell(row=r, column=c).value if ws_a_edit is not None else None
    vb_edit = ws_b_edit.cell(row=r, column=c).value if ws_b_edit is not None else None
    return _cell_display_and_equal_from_values(va_val, vb_val, va_edit, vb_edit)


def _cell_display_and_equal_from_values(va_val, vb_val, va_edit=None, vb_edit=None):
    """Compare already-fetched cell values while preserving the existing fallback rules."""
    if _USE_CACHED_VALUES_ONLY:
        def _formula_identity(v_val, v_edit):
            if isinstance(v_edit, str) and v_edit.startswith("=") and v_val == v_edit:
                return None
            special = _special_formula_signature(v_edit)
            if special is not None:
                return ("SPECIAL", special)
            formula = _norm_formula_text(v_edit)
            return ("TEXT", formula) if formula else None

        # If cache missing but edit has a literal value, use it for display/compare.
        identity_a = _formula_identity(va_val, va_edit)
        identity_b = _formula_identity(vb_val, vb_edit)
        if va_val is None and va_edit is not None and identity_a is None:
            va_val = va_edit
        if vb_val is None and vb_edit is not None and identity_b is None:
            vb_val = vb_edit

        formula_a = _formula_text(va_edit) if identity_a is not None else None
        formula_b = _formula_text(vb_edit) if identity_b is not None else None

        # Formula structure is part of the cell value. Equal cached results do
        # not make different formulas (or formula-vs-literal) equivalent.
        if identity_a != identity_b and (identity_a is not None or identity_b is not None):
            display_a = va_val if va_val is not None else formula_a
            display_b = vb_val if vb_val is not None else formula_b
            return display_a, display_b, False

        # If both caches are missing, formula structure is the only available
        # evidence. Never collapse different formulas (or formula-vs-blank) into
        # an apparently equal pair of None values.
        if va_val is None and vb_val is None and (formula_a or formula_b):
            if identity_a is not None and identity_a == identity_b:
                return formula_a, formula_b, True
            return formula_a, formula_b, False

        # If cache missing on one side but both formulas are the same, treat as equal
        # and display the available value.
        if (va_val is None) != (vb_val is None) and identity_a is not None and identity_a == identity_b:
            v = va_val if va_val is not None else vb_val
            return v, v, True

        if va_val is None and formula_a:
            va_val = formula_a
        if vb_val is None and formula_b:
            vb_val = formula_b

        eq = (_merge_cmp_value(va_val) == _merge_cmp_value(vb_val))
        return va_val, vb_val, eq

    eq = (_merge_cmp_value(va_val) == _merge_cmp_value(vb_val))
    return va_val, vb_val, eq


def _pad_row_values(row_vals, max_col: int):
    row = tuple(row_vals or ())
    if len(row) >= max_col:
        return row[:max_col]
    return row + (None,) * (max_col - len(row))


def _read_rows_into_cache(
    ws,
    row_indices,
    max_col: int,
    *,
    require_complete: bool = False,
    cancel_check=None,
):
    """Read a set of rows via one contiguous iter_rows pass."""
    rows = {}
    try:
        needed = sorted({int(r) for r in row_indices if r is not None and int(r) > 0})
    except Exception:
        needed = []
    if not needed:
        return rows

    min_r = needed[0]
    max_r = needed[-1]
    needed_set = set(needed)
    try:
        for idx, row in enumerate(
            ws.iter_rows(
                min_row=min_r,
                max_row=max_r,
                min_col=1,
                max_col=max_col,
                values_only=True,
            ),
            start=min_r,
        ):
            if cancel_check is not None and (idx & 127) == 0:
                cancel_check()
            if idx in needed_set:
                rows[idx] = _pad_row_values(row, max_col)
    except InterruptedError:
        raise
    except Exception:
        if require_complete:
            return {}

    if require_complete and any(r not in rows for r in needed):
        return {}
    for r in needed:
        rows.setdefault(r, (None,) * max_col)
    return rows


def _row_from_cache(rows_cache: dict[int, tuple], row_idx: int | None, max_col: int):
    if row_idx is None:
        return (None,) * max_col
    return rows_cache.get(int(row_idx), (None,) * max_col)


def _cell_display_and_equal_by_row(ws_a_val, ws_b_val, ws_a_edit, ws_b_edit, ra: int | None, rb: int | None, c: int):
    va_val = ws_a_val.cell(row=ra, column=c).value if ra is not None else None
    vb_val = ws_b_val.cell(row=rb, column=c).value if rb is not None else None
    va_edit = ws_a_edit.cell(row=ra, column=c).value if (ra is not None and ws_a_edit is not None) else None
    vb_edit = ws_b_edit.cell(row=rb, column=c).value if (rb is not None and ws_b_edit is not None) else None
    if ra is not None and rb is not None and ra != rb:
        vb_edit = _translate_normal_formula_for_compare(vb_val, vb_edit, rb, c, ra, c)
    return _cell_display_and_equal_from_values(va_val, vb_val, va_edit, vb_edit)


class _LiteralText(str):
    """String beginning with '=' that must remain text, not become a formula."""


def _formula_text(v):
    if isinstance(v, _LiteralText):
        return None
    if isinstance(v, str) and v.startswith("="):
        return v
    if isinstance(v, ArrayFormula):
        return getattr(v, "text", None)
    t = getattr(v, "text", None)
    if isinstance(t, str) and t.startswith("="):
        return t
    return None


def _special_formula_signature(value):
    """Return structural identity for formulas that cannot be copied as a cell string."""
    if isinstance(value, ArrayFormula):
        return (
            "array",
            str(getattr(value, "ref", "") or ""),
            _norm_formula_text(getattr(value, "text", None)),
        )
    if isinstance(value, DataTableFormula):
        try:
            attrs = tuple(sorted((str(key), str(val)) for key, val in value))
        except Exception:
            attrs = tuple(sorted((str(key), repr(val)) for key, val in vars(value).items()))
        return ("dataTable", attrs)
    return None


def _ensure_formula_copy_supported(src_edit, dst_edit=None):
    """Reject copying a special multi-cell formula unless the target already owns it."""
    src_sig = _special_formula_signature(src_edit)
    if src_sig is None:
        return
    if src_sig == _special_formula_signature(dst_edit):
        return
    raise RuntimeError(
        "检测到数组公式或数据表公式。该公式跨越多个单元格，不能按单元格/行直接覆盖；"
        "请改为合并其依赖数据，或在 Excel 中整体处理该公式区域。"
    )


def _norm_formula_text(v):
    f = _formula_text(v)
    if not f:
        return None
    s = str(f).strip()
    if s.startswith("="):
        s = s[1:]
    # Excel identifiers/functions are case-insensitive, but string literals and
    # spaces are semantically significant (a space can be the intersection
    # operator). Normalize case only outside double-quoted string literals.
    out = []
    in_string = False
    idx = 0
    while idx < len(s):
        ch = s[idx]
        if ch == '"':
            out.append(ch)
            if in_string and idx + 1 < len(s) and s[idx + 1] == '"':
                out.append('"')
                idx += 2
                continue
            in_string = not in_string
        else:
            out.append(ch if in_string else ch.upper())
        idx += 1
    return "".join(out)


def _same_formula(a, b):
    special_a = _special_formula_signature(a)
    special_b = _special_formula_signature(b)
    if special_a is not None or special_b is not None:
        return special_a is not None and special_a == special_b
    na = _norm_formula_text(a)
    nb = _norm_formula_text(b)
    return bool(na and nb and na == nb)


def _choose_edit_value(v_val, v_edit):
    """Preserve formula semantics on overwrite, even in cached-value mode."""
    _ensure_formula_copy_supported(v_edit, None)
    if (
        isinstance(v_edit, str)
        and v_edit.startswith("=")
        and v_val == v_edit
    ):
        # A true formula's data-only value is its cached result (or None), never
        # the formula text itself. Matching edit/data-only values therefore mean
        # the source cell is literal text whose first character happens to be '='.
        return _LiteralText(v_edit)
    f = _formula_text(v_edit)
    if f:
        return f
    return v_val if _USE_CACHED_VALUES_ONLY else v_edit


def _assign_edit_cell_value(cell, value):
    """Assign an edit-workbook value without turning literal '=...' text into a formula."""
    if isinstance(value, _LiteralText):
        cell.value = str(value)
        cell.data_type = "s"
        return
    cell.value = value


def _translate_normal_formula_for_compare(
    v_val,
    v_edit,
    src_row: int,
    src_col: int,
    dst_row: int,
    dst_col: int,
):
    """Translate a normal formula to a common coordinate for aligned comparison."""
    if (
        isinstance(v_edit, str)
        and v_edit.startswith("=")
        and v_val == v_edit
    ):
        return v_edit
    if _special_formula_signature(v_edit) is not None:
        return v_edit
    formula = _formula_text(v_edit)
    if not formula or (int(src_row), int(src_col)) == (int(dst_row), int(dst_col)):
        return v_edit
    try:
        origin = f"{get_column_letter(int(src_col))}{int(src_row)}"
        target = f"{get_column_letter(int(dst_col))}{int(dst_row)}"
        return Translator(formula, origin=origin).translate_formula(target)
    except Exception:
        # If translation cannot be proven, retain the original formula so the
        # comparator reports a conservative difference instead of hiding one.
        return v_edit


def _formula_edit_value_map(ws_edit) -> dict[tuple[int, int], object]:
    """Index only formula cells from a normal worksheet for hot 3-way scans."""
    result = {}
    if ws_edit is None:
        return result
    try:
        cells = getattr(ws_edit, "_cells", None)
        if cells is not None:
            for (row_idx, col_idx), cell in cells.items():
                value = cell.value
                if getattr(cell, "data_type", None) == "f" or _special_formula_signature(value) is not None:
                    result[(int(row_idx), int(col_idx))] = value
            return result
    except Exception:
        result = {}
    try:
        for row in ws_edit.iter_rows(values_only=False):
            for cell in row:
                value = cell.value
                if getattr(cell, "data_type", None) == "f" or _special_formula_signature(value) is not None:
                    result[(int(cell.row), int(cell.column))] = value
    except Exception:
        pass
    return result


class _FormulaIdentityKey:
    __slots__ = ("kind", "value")

    def __init__(self, kind: str, value):
        self.kind = kind
        self.value = value

    def __eq__(self, other):
        if not isinstance(other, _FormulaIdentityKey) or self.kind != other.kind:
            return False
        if self.value == other.value:
            return True
        if self.kind == "TEXT":
            return _same_formula(self.value, other.value)
        return False


def _merge_cell_compare_key(v_val, v_edit):
    """Type- and formula-aware key for 3-way change/conflict classification."""
    literal_equals = (
        isinstance(v_edit, str)
        and v_edit.startswith("=")
        and v_val == v_edit
    )
    special = _special_formula_signature(v_edit)
    formula = None if literal_equals else _formula_text(v_edit)
    if special is not None:
        return ("FORMULA", _FormulaIdentityKey("SPECIAL", special), _merge_cmp_value(v_val))
    if formula:
        return ("FORMULA", _FormulaIdentityKey("TEXT", formula), _merge_cmp_value(v_val))
    effective = v_val
    if effective is None and v_edit is not None:
        effective = v_edit
    return ("VALUE", _merge_cmp_value(effective))


def _copy_edit_value_for_destination(
    v_val,
    v_edit,
    dst_edit,
    *,
    src_row: int,
    src_col: int,
    dst_row: int,
    dst_col: int,
):
    """Choose an edit value and translate normal formulas like Excel copy/paste."""
    src_row = int(src_row)
    src_col = int(src_col)
    dst_row = int(dst_row)
    dst_col = int(dst_col)
    special = _special_formula_signature(v_edit)
    if special is not None:
        if (src_row, src_col) != (dst_row, dst_col):
            raise RuntimeError(
                "数组公式或数据表公式不能安全移动到不同坐标；"
                "请在 Excel 中处理该公式区域。"
            )
        _ensure_formula_copy_supported(v_edit, dst_edit)
        return v_edit

    chosen = _choose_edit_value(v_val, v_edit)
    formula = _formula_text(chosen)
    if not formula or (src_row, src_col) == (dst_row, dst_col):
        return chosen
    try:
        origin = f"{get_column_letter(src_col)}{src_row}"
        target = f"{get_column_letter(dst_col)}{dst_row}"
        return Translator(formula, origin=origin).translate_formula(target)
    except Exception as exc:
        raise RuntimeError(
            f"无法安全平移公式 {formula!r}："
            f"{get_column_letter(src_col)}{src_row} -> "
            f"{get_column_letter(dst_col)}{dst_row}"
        ) from exc


def _is_same_formula_copy_noop(src_edit, dst_edit) -> bool:
    """Return True when copying would only replace a formula with itself.

    Cached results can differ because precedent cells differ. Copying the same
    formula cannot adopt that result; Excel will recalculate it from the target
    workbook and make the apparent overwrite revert.
    """
    return _same_formula(src_edit, dst_edit)


def _col_index_from_ref(ref: str) -> int:
    if not ref:
        return 0
    m = re.match(r"([A-Za-z]+)", str(ref))
    if not m:
        return 0
    letters = m.group(1).upper()
    n = 0
    for ch in letters:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def _extract_ns_map(xml_bytes: bytes) -> dict:
    """Extract XML namespace prefix→URI map from the opening bytes of an XML document.
    Used to register namespaces with ET before re-serialization to preserve original prefixes."""
    text = xml_bytes[:4096].decode("utf-8", errors="ignore")
    ns_map = {}
    for m in re.finditer(r'xmlns(?::(\w+))?="([^"]+)"', text):
        prefix = m.group(1) if m.group(1) is not None else ""
        uri = m.group(2)
        ns_map[prefix] = uri
    return ns_map


_MISSING_VALUE = object()
_EXCEL_ERROR_VALUES = {
    "#NULL!", "#DIV/0!", "#VALUE!", "#REF!", "#NAME?", "#NUM!", "#N/A",
    "#GETTING_DATA", "#SPILL!", "#CALC!", "#CONNECT!", "#BLOCKED!",
    "#UNKNOWN!", "#FIELD!",
}


def _excel_cached_value_payload(value, excel_epoch=CALENDAR_WINDOWS_1900):
    """Return (cell_type, text) for an OOXML cached/scalar value."""
    if value is None:
        return None, None
    if isinstance(value, bool):
        return "b", "1" if value else "0"
    if isinstance(value, (datetime, date, datetime_time, timedelta)):
        return None, str(to_excel(value, epoch=excel_epoch))
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if isinstance(value, float) and not math.isfinite(value):
            raise RuntimeError(f"Excel cell cannot store non-finite numeric value: {value!r}")
        return None, str(value)
    text = str(value)
    return ("e" if text.upper() in _EXCEL_ERROR_VALUES else "str"), text


def _set_formula_cached_value(target_cell, q, cached_value, excel_epoch):
    for child in list(target_cell):
        if child.tag == q("v"):
            target_cell.remove(child)
    target_cell.attrib.pop("t", None)
    cell_type, cache_text = _excel_cached_value_payload(cached_value, excel_epoch)
    if cache_text is None:
        return
    if cell_type:
        target_cell.attrib["t"] = cell_type
    v = ET.SubElement(target_cell, q("v"))
    v.text = cache_text


def _sheet_xml_set_cell(
    ws_root,
    row_idx: int,
    col_idx: int,
    value,
    cached_value=_MISSING_VALUE,
    preserve_existing_formula: bool = False,
    excel_epoch=CALENDAR_WINDOWS_1900,
):
    ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    q = lambda t: f"{{{ns}}}{t}"
    xml_space = "{http://www.w3.org/XML/1998/namespace}space"

    sheet_data = ws_root.find(q("sheetData"))
    if sheet_data is None:
        sheet_data = ET.SubElement(ws_root, q("sheetData"))

    target_row = None
    insert_row_at = None
    rows = list(sheet_data.findall(q("row")))
    for i, r in enumerate(rows):
        try:
            rr = int(r.attrib.get("r", "0") or 0)
        except (ValueError, TypeError):
            rr = 0
        if rr == row_idx:
            target_row = r
            break
        if rr > row_idx and insert_row_at is None:
            insert_row_at = i
    if target_row is None:
        target_row = ET.Element(q("row"), {"r": str(row_idx)})
        if insert_row_at is None:
            sheet_data.append(target_row)
        else:
            sheet_data.insert(insert_row_at, target_row)

    cell_ref = f"{get_column_letter(col_idx)}{row_idx}"
    target_cell = None
    insert_cell_at = None
    cells = list(target_row.findall(q("c")))
    for i, c in enumerate(cells):
        ref = c.attrib.get("r", "")
        cc = _col_index_from_ref(ref)
        if ref == cell_ref:
            target_cell = c
            break
        if cc > col_idx and insert_cell_at is None:
            insert_cell_at = i
    if target_cell is None:
        target_cell = ET.Element(q("c"), {"r": cell_ref})
        if insert_cell_at is None:
            target_row.append(target_cell)
        else:
            target_row.insert(insert_cell_at, target_cell)

    existing_formula = target_cell.find(q("f"))
    if preserve_existing_formula:
        if existing_formula is None:
            raise RuntimeError(f"Cache-only update target is not a formula cell: {cell_ref}")
        if cached_value is _MISSING_VALUE:
            raise RuntimeError(f"Cache-only update has no cached value: {cell_ref}")
        _set_formula_cached_value(target_cell, q, cached_value, excel_epoch)
        return

    formula = _formula_text(value)
    if existing_formula is not None:
        existing_text = existing_formula.text or ""
        existing_full = ("=" + existing_text) if existing_text else None
        if formula and existing_full and _norm_formula_text(existing_full) == _norm_formula_text(formula):
            # Preserve shared/array formula metadata while allowing an explicit
            # cached-result adoption for workbooks that use manual calculation.
            if cached_value is not _MISSING_VALUE:
                _set_formula_cached_value(target_cell, q, cached_value, excel_epoch)
            return
        if existing_formula.attrib.get("t") in ("shared", "array", "dataTable"):
            raise RuntimeError(
                f"ZIP fallback cannot safely replace special formula cell {cell_ref}; "
                "Excel native save is required"
            )

    # Keep row/cell attrs (like style index), replace only value payload.
    for child in list(target_cell):
        target_cell.remove(child)

    if formula:
        target_cell.attrib.pop("t", None)
        f = ET.SubElement(target_cell, q("f"))
        f.text = formula[1:] if str(formula).startswith("=") else str(formula)
        if cached_value is not _MISSING_VALUE:
            _set_formula_cached_value(target_cell, q, cached_value, excel_epoch)
        return

    if value is None:
        target_cell.attrib.pop("t", None)
        return

    if isinstance(value, bool):
        target_cell.attrib["t"] = "b"
        v = ET.SubElement(target_cell, q("v"))
        v.text = "1" if value else "0"
        return

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if isinstance(value, float) and not math.isfinite(value):
            raise RuntimeError(f"Excel cell cannot store non-finite numeric value: {value!r}")
        target_cell.attrib.pop("t", None)
        v = ET.SubElement(target_cell, q("v"))
        v.text = str(value)
        return

    if isinstance(value, (datetime, date, datetime_time, timedelta)):
        target_cell.attrib.pop("t", None)
        v = ET.SubElement(target_cell, q("v"))
        v.text = str(to_excel(value, epoch=excel_epoch))
        return

    # Keep text independent from sharedStrings to avoid global reindex churn.
    s = str(value)
    target_cell.attrib["t"] = "inlineStr"
    is_node = ET.SubElement(target_cell, q("is"))
    t_node = ET.SubElement(is_node, q("t"))
    if s and (s[0].isspace() or s[-1].isspace()):
        t_node.attrib[xml_space] = "preserve"
    t_node.text = s


def _filter_noop_manual_ops(src_xlsx: str, manual_ops: dict) -> dict:
    """Drop formula-to-identical-formula writes before any save backend runs."""
    if not manual_ops:
        return {}
    wb = None
    try:
        wb = load_workbook(src_xlsx, data_only=False, read_only=False)
        filtered = {}
        for key, value in manual_ops.items():
            sheet, row_idx, col_idx = key
            if sheet not in wb.sheetnames:
                filtered[key] = value
                continue
            original = wb[sheet].cell(row=int(row_idx), column=int(col_idx)).value
            if _same_formula(original, value):
                continue
            filtered[key] = value
        return filtered
    except Exception as e:
        _dlog(f"manual op no-op filter failed; keeping original ops: {e}")
        return dict(manual_ops)
    finally:
        _wbs_close(wb)


def _prepare_manual_ops_for_save(src_xlsx: str, manual_ops: dict, row_ops=None, sheet_ops=None) -> dict:
    """Filter cell no-ops only when coordinates still match the source snapshot."""
    if row_ops or sheet_ops:
        return dict(manual_ops or {})
    return _filter_noop_manual_ops(src_xlsx, manual_ops or {})


def _xlsx_contains_formulas(path: str) -> bool:
    try:
        with zipfile.ZipFile(path, "r") as zf:
            for name in zf.namelist():
                if name.startswith("xl/worksheets/") and name.endswith(".xml"):
                    if re.search(rb"<(?:[A-Za-z_][\w.-]*:)?f(?:\s|>)", zf.read(name)):
                        return True
    except Exception:
        return True
    return False


def _xlsx_sheet_part_fingerprints(path: str) -> dict[str, tuple[int, int]]:
    """Return cheap per-Sheet XML fingerprints from the ZIP central directory.

    A mismatch is only a provisional difference signal: worksheet XML also
    contains non-value metadata. Exact cell/formula comparison remains the
    authority and will confirm or clear the tab state later.
    """
    fingerprints: dict[str, tuple[int, int]] = {}
    try:
        with zipfile.ZipFile(path, "r") as zf:
            workbook_name = "xl/workbook.xml"
            rels_name = "xl/_rels/workbook.xml.rels"
            if workbook_name not in zf.namelist() or rels_name not in zf.namelist():
                return fingerprints
            workbook_root = ET.fromstring(zf.read(workbook_name))
            rels_root = ET.fromstring(zf.read(rels_name))
            rel_targets = {
                str(rel.attrib.get("Id") or ""): str(rel.attrib.get("Target") or "")
                for rel in rels_root.iter()
                if rel.tag.rsplit("}", 1)[-1] == "Relationship"
            }
            rel_id_attr = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
            for sheet_node in workbook_root.iter():
                if sheet_node.tag.rsplit("}", 1)[-1] != "sheet":
                    continue
                sheet_name = str(sheet_node.attrib.get("name") or "")
                rel_id = str(sheet_node.attrib.get(rel_id_attr) or "")
                if not rel_id:
                    rel_id = str(next(
                        (
                            value for key, value in sheet_node.attrib.items()
                            if key == "id" or key.rsplit("}", 1)[-1] == "id"
                        ),
                        "",
                    ))
                target = rel_targets.get(rel_id, "").replace("\\", "/")
                if not sheet_name or not target:
                    continue
                if target.startswith("/"):
                    part_name = target.lstrip("/")
                else:
                    part_name = posixpath.normpath(posixpath.join("xl", target))
                try:
                    info = zf.getinfo(part_name)
                except KeyError:
                    continue
                fingerprints[sheet_name] = (int(info.CRC), int(info.file_size))
    except Exception as e:
        _dlog(f"sheet fingerprint premark failed: path={path} err={e}")
    return fingerprints


def _xlsx_requires_native_structural_replay(path: str) -> bool:
    """Return True when openpyxl row/sheet replay cannot prove package fidelity."""
    if _xlsx_contains_formulas(path):
        return True
    risky_prefixes = (
        "xl/drawings/", "xl/charts/", "xl/tables/", "xl/pivot",
        "xl/comments", "xl/threadedComments/", "xl/externalLinks/",
        "xl/ctrlProps/", "xl/embeddings/", "xl/activeX/", "xl/slicers/",
        "xl/timelines/", "xl/vbaProject.bin", "xl/macrosheets/",
        "xl/dialogsheets/",
    )
    risky_sheet_tags = re.compile(
        rb"<(?:[A-Za-z_][\w.-]*:)?(?:dataValidations?|conditionalFormatting|mergeCells|hyperlinks)(?:\s|>)"
    )
    try:
        with zipfile.ZipFile(path, "r") as zf:
            names = zf.namelist()
            if any(name.startswith(risky_prefixes) for name in names):
                return True
            for name in names:
                if name.startswith("xl/worksheets/") and name.endswith(".xml"):
                    if risky_sheet_tags.search(zf.read(name)):
                        return True
        return False
    except Exception:
        # Structural replay is destructive; unreadable package metadata must fail closed.
        return True


def _replay_formula_source_paths(
    src_xlsx: str,
    row_ops=None,
    sheet_ops=None,
    source_paths: dict[str, str] | None = None,
) -> list[str]:
    """Return workbook paths whose formulas can enter a row/sheet replay."""
    source_paths = source_paths or {}
    sides = {"A"}
    for op in list(row_ops or []) + list(sheet_ops or []):
        side = str(op.get("source_side") or "").upper()
        if side:
            sides.add(side)
    paths = [src_xlsx]
    if "A" in sides and source_paths.get("A"):
        paths.append(source_paths["A"])
    if sides & {"B", "THEIRS"}:
        path_b = source_paths.get("B") or source_paths.get("THEIRS")
        if path_b:
            paths.append(path_b)
    if "BASE" in sides and source_paths.get("BASE"):
        paths.append(source_paths["BASE"])
    unique = []
    seen = set()
    for path in paths:
        try:
            key = os.path.normcase(os.path.abspath(path))
        except Exception:
            key = str(path)
        if key in seen:
            continue
        seen.add(key)
        unique.append(path)
    return unique


def _validate_xlsx_package(path: str) -> tuple[bool, str]:
    """Validate ZIP/XML integrity and shared-formula master references."""
    try:
        with zipfile.ZipFile(path, "r") as zf:
            bad_part = zf.testzip()
            if bad_part:
                return False, f"ZIP CRC failed: {bad_part}"
            for name in zf.namelist():
                if not name.endswith((".xml", ".rels")):
                    continue
                root = ET.fromstring(zf.read(name))
                if not name.startswith("xl/worksheets/") or not name.endswith(".xml"):
                    continue
                shared: dict[str, dict[str, int]] = {}
                for node in root.iter():
                    if not str(node.tag).endswith("}f") and node.tag != "f":
                        continue
                    if node.attrib.get("t") != "shared":
                        continue
                    si = str(node.attrib.get("si") or "")
                    state = shared.setdefault(si, {"masters": 0, "members": 0})
                    state["members"] += 1
                    if node.attrib.get("ref") and (node.text or "").strip():
                        state["masters"] += 1
                for si, state in shared.items():
                    if state["members"] and state["masters"] != 1:
                        return False, (
                            f"invalid shared formula group in {name}: "
                            f"si={si!r} masters={state['masters']} members={state['members']}"
                        )
        return True, ""
    except Exception as e:
        return False, str(e)


def _build_manual_merge_xlsx_via_zip(
    src_xlsx: str,
    out_xlsx: str,
    manual_ops: dict,
    cached_values: dict | None = None,
    cache_only_keys: set[tuple[str, int, int]] | None = None,
):
    """Patch only selected cells at XML level; keep untouched parts byte-identical."""
    with zipfile.ZipFile(src_xlsx, "r") as zf:
        names = zf.namelist()
        infos = {n: zf.getinfo(n) for n in names}
        files = {n: zf.read(n) for n in names}

    wb_name = "xl/workbook.xml"
    rels_name = "xl/_rels/workbook.xml.rels"
    if wb_name not in files or rels_name not in files:
        shutil.copy2(src_xlsx, out_xlsx)
        return

    wb_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    rel_ns = "http://schemas.openxmlformats.org/package/2006/relationships"
    doc_rel_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

    wb_root = ET.fromstring(files[wb_name])
    wb_pr = wb_root.find(f"{{{wb_ns}}}workbookPr")
    date1904 = False
    if wb_pr is not None:
        date1904 = str(wb_pr.attrib.get("date1904", "0")).strip().lower() in ("1", "true")
    excel_epoch = CALENDAR_MAC_1904 if date1904 else CALENDAR_WINDOWS_1900
    rel_root = ET.fromstring(files[rels_name])
    rid_to_target = {}
    for rel in rel_root.findall(f"{{{rel_ns}}}Relationship"):
        rid = rel.attrib.get("Id")
        tgt = rel.attrib.get("Target")
        if not rid or not tgt:
            continue
        if tgt.startswith("/"):
            norm = tgt.lstrip("/")
        elif tgt.startswith("xl/"):
            norm = tgt
        else:
            norm = f"xl/{tgt}"
        rid_to_target[rid] = norm.replace("\\", "/")

    sheet_to_part = {}
    sheets = wb_root.find(f"{{{wb_ns}}}sheets")
    if sheets is not None:
        for sh in sheets.findall(f"{{{wb_ns}}}sheet"):
            name = sh.attrib.get("name")
            rid = sh.attrib.get(f"{{{doc_rel_ns}}}id")
            if not name or not rid:
                continue
            part = rid_to_target.get(rid)
            if part and part in files:
                sheet_to_part[name] = part

    ops_by_sheet = {}
    cached_values = cached_values or {}
    cache_only_keys = {
        (str(sheet), int(r), int(c))
        for sheet, r, c in (cache_only_keys or set())
    }
    for (sheet, r, c), v in manual_ops.items():
        key = (str(sheet), int(r), int(c))
        cached = cached_values.get((sheet, int(r), int(c)), _MISSING_VALUE)
        ops_by_sheet.setdefault(sheet, []).append((int(r), int(c), v, cached, key in cache_only_keys))

    for sheet, ops in ops_by_sheet.items():
        part = sheet_to_part.get(sheet)
        if not part:
            continue
        # Register original namespace prefixes before parsing to preserve them in output
        ns_map = _extract_ns_map(files[part])
        for pfx, uri in ns_map.items():
            ET.register_namespace(pfx, uri)
        ws_root = ET.fromstring(files[part])
        for r, c, v, cached, cache_only in sorted(ops, key=lambda x: (x[0], x[1])):
            _sheet_xml_set_cell(
                ws_root,
                r,
                c,
                v,
                cached,
                preserve_existing_formula=cache_only,
                excel_epoch=excel_epoch,
            )
        xml_bytes = ET.tostring(ws_root, encoding="utf-8", xml_declaration=True)
        xml_bytes = re.sub(
            rb'<\?xml[^?]*\?>',
            b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
            xml_bytes, count=1
        )
        # ET silently drops namespace declarations for prefixes not used in any element
        # tag (e.g. x14ac/xr/xr2 referenced only as string tokens in mc:Ignorable).
        # Re-inject any such declarations so mc:Ignorable remains valid.
        missing_decls = [
            f'xmlns:{pfx}="{uri}"'
            for pfx, uri in ns_map.items()
            if pfx and pfx != "xml" and f'xmlns:{pfx}="'.encode() not in xml_bytes
        ]
        if missing_decls:
            inject = (" " + " ".join(missing_decls)).encode()
            xml_bytes = re.sub(
                rb'(<\?xml[^?]*\?>\s*)(<[A-Za-z][\w:.-]*)',
                lambda m: m.group(1) + m.group(2) + inject,
                xml_bytes, count=1
            )
        files[part] = xml_bytes

    # Remove calcChain.xml to avoid Excel repair caused by stale calculation chain
    calc_chain = "xl/calcChain.xml"
    if calc_chain in names:
        names = [n for n in names if n != calc_chain]
        files.pop(calc_chain, None)

        # Remove calcChain Override from [Content_Types].xml
        ct_name = "[Content_Types].xml"
        if ct_name in files:
            ct_ns_map = _extract_ns_map(files[ct_name])
            for pfx, uri in ct_ns_map.items():
                ET.register_namespace(pfx, uri)
            ct_root = ET.fromstring(files[ct_name])
            ct_ns = "http://schemas.openxmlformats.org/package/2006/content-types"
            for override in ct_root.findall(f"{{{ct_ns}}}Override"):
                if "/calcChain" in override.attrib.get("PartName", ""):
                    ct_root.remove(override)
            ct_bytes = ET.tostring(ct_root, encoding="utf-8", xml_declaration=True)
            ct_bytes = re.sub(
                rb'<\?xml[^?]*\?>',
                b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
                ct_bytes, count=1
            )
            files[ct_name] = ct_bytes

        # Remove calcChain Relationship from xl/_rels/workbook.xml.rels
        if rels_name in files:
            wbr_ns_map = _extract_ns_map(files[rels_name])
            for pfx, uri in wbr_ns_map.items():
                ET.register_namespace(pfx, uri)
            wbr_root = ET.fromstring(files[rels_name])
            for rel in wbr_root.findall(f"{{{rel_ns}}}Relationship"):
                if "calcChain" in rel.attrib.get("Target", ""):
                    wbr_root.remove(rel)
            wbr_bytes = ET.tostring(wbr_root, encoding="utf-8", xml_declaration=True)
            wbr_bytes = re.sub(
                rb'<\?xml[^?]*\?>',
                b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
                wbr_bytes, count=1
            )
            files[rels_name] = wbr_bytes

    with zipfile.ZipFile(out_xlsx, "w") as zf:
        for n in names:
            info = infos.get(n)
            comp = info.compress_type if info else zipfile.ZIP_DEFLATED
            zf.writestr(n, files[n], compress_type=comp)
    valid, reason = _validate_xlsx_package(out_xlsx)
    if not valid:
        raise RuntimeError(f"unsafe XLSX output rejected: {reason}")


def _excel_com_cell_op(sheet: str, row: int, col: int, value) -> dict:
    """Return a JSON-safe, type-preserving Excel COM cell operation."""
    formula = _formula_text(value)
    if formula:
        return {"sheet": sheet, "r": int(row), "c": int(col), "formula": formula}
    if isinstance(value, datetime):
        return {
            "sheet": sheet,
            "r": int(row),
            "c": int(col),
            "value": float(to_excel(value, epoch=CALENDAR_WINDOWS_1900)),
            "value_kind": "datetime_serial",
        }
    if isinstance(value, date):
        return {
            "sheet": sheet,
            "r": int(row),
            "c": int(col),
            "value": float(to_excel(value, epoch=CALENDAR_WINDOWS_1900)),
            "value_kind": "datetime_serial",
        }
    if isinstance(value, datetime_time):
        seconds = (
            value.hour * 3600
            + value.minute * 60
            + value.second
            + value.microsecond / 1_000_000.0
        )
        value = seconds / 86400.0
    elif isinstance(value, timedelta):
        value = value.total_seconds() / 86400.0
    return {
        "sheet": sheet,
        "r": int(row),
        "c": int(col),
        "value": str(value) if isinstance(value, _LiteralText) else value,
        "value_kind": "text" if isinstance(value, str) else "typed",
    }


def _build_manual_merge_output_with_excel(
    src_xlsx: str,
    out_xlsx: str,
    manual_ops: dict,
    row_ops: list[dict] | None = None,
    sheet_ops: list[dict] | None = None,
    source_paths: dict[str, str] | None = None,
) -> bool:
    """Apply manual ops through Excel COM to preserve workbook fidelity."""
    if not manual_ops and not row_ops and not sheet_ops:
        try:
            shutil.copy2(src_xlsx, out_xlsx)
            return True
        except Exception:
            return False
    row_ops = list(row_ops or [])
    sheet_ops = list(sheet_ops or [])
    ops = []
    for (sheet, r, c), v in manual_ops.items():
        ops.append(_excel_com_cell_op(sheet, r, c, v))
    try:
        ops_json = os.path.join(tempfile.gettempdir(), f"{APP_NAME}_merge_ops_{os.getpid()}_{int(time.time())}.json")
        with open(ops_json, "w", encoding="utf-8") as f:
            json.dump(
                {"row_ops": row_ops, "cell_ops": ops, "sheet_ops": sheet_ops},
                f,
                ensure_ascii=False,
                allow_nan=False,
            )
    except Exception as e:
        _dlog(f"excel native save: write ops json failed: {e}")
        return False
    try:
        source_paths = source_paths or {}
        mine_src = str(source_paths.get("A") or "").replace("'", "''")
        base_src = str(source_paths.get("BASE") or "").replace("'", "''")
        theirs_src = str(source_paths.get("B") or source_paths.get("THEIRS") or "").replace("'", "''")
        ps = (
            "$ErrorActionPreference='Stop';"
            "[Console]::OutputEncoding=[System.Text.Encoding]::UTF8;"
            "$OutputEncoding=[System.Text.Encoding]::UTF8;"
            "[Console]::OutputEncoding=[System.Text.Encoding]::UTF8;"
            "$OutputEncoding=[System.Text.Encoding]::UTF8;"
            "$src='" + src_xlsx.replace("'", "''") + "';"
            "$out='" + out_xlsx.replace("'", "''") + "';"
            "$opsPath='" + ops_json.replace("'", "''") + "';"
            "$mineSrc='" + mine_src + "';"
            "$baseSrc='" + base_src + "';"
            "$theirsSrc='" + theirs_src + "';"
            "$payload=Get-Content -Raw -LiteralPath $opsPath | ConvertFrom-Json;"
            "$xl=$null;$wb=$null;$wbMine=$null;$wbBase=$null;$wbTheirs=$null;"
            "try{"
            "$xl=New-Object -ComObject Excel.Application;"
            "$xl.Visible=$false;"
            "$xl.DisplayAlerts=$false;"
            "$xl.AskToUpdateLinks=$false;"
            "$xl.EnableEvents=$false;"
            "try{$xl.Calculation=-4135}catch{};"
            "try{$xl.CalculateBeforeSave=$false}catch{};"
            "$wb=$xl.Workbooks.Open($src,0,$false);"
            "try{$xl.Calculation=-4135}catch{};"
            "if($mineSrc){$wbMine=$xl.Workbooks.Open($mineSrc,0,$true)};"
            "if($baseSrc){$wbBase=$xl.Workbooks.Open($baseSrc,0,$true)};"
            "if($theirsSrc){$wbTheirs=$xl.Workbooks.Open($theirsSrc,0,$true)};"
            "foreach($op in @($payload.sheet_ops)){"
            "  if($null -eq $op){continue};"
            "  if($op.kind -eq 'delete_sheet'){"
            "    try{"
            "      if($wb.Worksheets.Count -gt 1){$wb.Worksheets.Item($op.sheet).Delete()}"
            "    }catch{};"
            "    continue"
            "  };"
            "  if($op.kind -eq 'copy_sheet'){"
            "    $srcWs=$null;"
            "    if($op.source_side -eq 'BASE' -and $wbBase -ne $null){$srcWs=$wbBase.Worksheets.Item($op.sheet)}"
            "    elseif(($op.source_side -eq 'B' -or $op.source_side -eq 'THEIRS') -and $wbTheirs -ne $null){$srcWs=$wbTheirs.Worksheets.Item($op.sheet)}"
            "    elseif($op.source_side -eq 'A' -and $wbMine -ne $null){$srcWs=$wbMine.Worksheets.Item($op.sheet)}"
            "    elseif($op.source_side -eq 'A'){$srcWs=$wb.Worksheets.Item($op.sheet)};"
            "    if($srcWs -eq $null){continue};"
            "    try{$wb.Worksheets.Item($op.sheet).Delete()}catch{};"
            "    $after=$wb.Worksheets.Item($wb.Worksheets.Count);"
            "    $srcWs.Copy($null, $after);"
            "    $newWs=$wb.Worksheets.Item($wb.Worksheets.Count);"
            "    try{$newWs.Name=$op.sheet}catch{};"
            "  }"
            "};"
            "foreach($op in @($payload.row_ops)){"
            "  if($null -eq $op){continue};"
            "  $ws=$wb.Worksheets.Item($op.sheet);"
            "  if($op.kind -eq 'insert_rows'){"
            "    for($i=0;$i -lt [int]$op.count;$i++){[void]$ws.Rows.Item([int]$op.row).Insert()}"
            "    $srcWs=$null;"
            "    if(($op.source_side -eq 'B' -or $op.source_side -eq 'THEIRS') -and $wbTheirs -ne $null){$srcWs=$wbTheirs.Worksheets.Item($op.sheet)}"
            "    elseif($op.source_side -eq 'BASE' -and $wbBase -ne $null){$srcWs=$wbBase.Worksheets.Item($op.sheet)};"
            "    if($op.source_side -eq 'A' -and $wbMine -ne $null){$srcWs=$wbMine.Worksheets.Item($op.sheet)};"
            "    $srcRows=@($op.source_rows);"
            "    if($srcWs -ne $null -and $srcRows.Count -gt 0){"
            "      $lastCol=[Math]::Max(1,[int]($srcWs.UsedRange.Column+$srcWs.UsedRange.Columns.Count-1));"
            "      for($i=0;$i -lt $srcRows.Count;$i++){"
            "        $srcRow=[int]$srcRows[$i];$dstRow=[int]$op.row+$i;"
            "        $srcRange=$srcWs.Range($srcWs.Cells.Item($srcRow,1),$srcWs.Cells.Item($srcRow,$lastCol));"
            "        $dstRange=$ws.Range($ws.Cells.Item($dstRow,1),$ws.Cells.Item($dstRow,$lastCol));"
            # xlPasteAll preserves comments, hyperlinks, validation and cell
            # metadata. Explicit cell_ops run afterwards and remain the source
            # of truth for translated formulas and cached-value decisions.
            "        [void]$srcRange.Copy();[void]$dstRange.PasteSpecial(-4104);"
            "        try{$ws.Rows.Item($dstRow).RowHeight=$srcWs.Rows.Item($srcRow).RowHeight}catch{}"
            "      }"
            "    }"
            "  }"
            "};"
            "foreach($op in @($payload.cell_ops)){"
            "  if($null -eq $op){continue};"
            "  $ws=$wb.Worksheets.Item($op.sheet);"
            "  $cell=$ws.Cells.Item([int]$op.r,[int]$op.c);"
            "  if($null -ne $op.formula){$cell.Formula=$op.formula}"
            "  elseif($op.value_kind -eq 'datetime_serial'){"
            "    $serial=[double]$op.value;if($wb.Date1904){$serial-=1462};$cell.Value2=$serial"
            "  }"
            "  elseif($op.value_kind -eq 'text'){"
            "    if($cell.NumberFormat -eq 'General'){$cell.NumberFormat='@'};"
            "    $textValue=[string]$op.value;"
            "    $cell.Value=$textValue"
            "  }"
            "  else{$cell.Value2=$op.value}"
            "};"
            "$wb.SaveCopyAs($out);"
            "}finally{"
            "  $cell=$null;$ws=$null;$srcRange=$null;$dstRange=$null;$srcWs=$null;$newWs=$null;$after=$null;"
            "  if($wbTheirs -ne $null){try{$wbTheirs.Close($false)}catch{};try{[void][Runtime.InteropServices.Marshal]::FinalReleaseComObject($wbTheirs)}catch{}};"
            "  if($wbBase -ne $null){try{$wbBase.Close($false)}catch{};try{[void][Runtime.InteropServices.Marshal]::FinalReleaseComObject($wbBase)}catch{}};"
            "  if($wbMine -ne $null){try{$wbMine.Close($false)}catch{};try{[void][Runtime.InteropServices.Marshal]::FinalReleaseComObject($wbMine)}catch{}};"
            "  if($wb -ne $null){try{$wb.Close($false)}catch{};try{[void][Runtime.InteropServices.Marshal]::FinalReleaseComObject($wb)}catch{}};"
            "  if($xl -ne $null){try{$xl.Quit()}catch{};try{[void][Runtime.InteropServices.Marshal]::FinalReleaseComObject($xl)}catch{}};"
            "  [GC]::Collect();[GC]::WaitForPendingFinalizers();[GC]::Collect();[GC]::WaitForPendingFinalizers();"
            "};"
        )
        r = subprocess.run(
            ["powershell", "-NoProfile", "-STA", "-ExecutionPolicy", "Bypass", "-Command", ps],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=180,
        )
        if r.returncode != 0:
            _dlog(f"excel native save failed rc={r.returncode} err={r.stderr.strip()}")
            return False
        valid, reason = _validate_xlsx_package(out_xlsx)
        if not valid:
            _dlog(f"excel native save produced invalid package: {reason}")
            return False
        _dlog(f"excel native save ok: {out_xlsx} sheet_ops={len(sheet_ops)} row_ops={len(row_ops)} cell_ops={len(ops)}")
        return True
    except Exception as e:
        _dlog(f"excel native save exception: {e}")
        return False
    finally:
        try:
            if os.path.exists(ops_json):
                os.remove(ops_json)
        except Exception:
            pass


def _build_manual_merge_output_with_openpyxl(
    src_xlsx: str,
    out_xlsx: str,
    manual_ops: dict,
    row_ops: list[dict] | None = None,
    sheet_ops: list[dict] | None = None,
    source_paths: dict[str, str] | None = None,
) -> bool:
    """Replay manual merge operations with openpyxl when Excel COM is unavailable."""
    wb = None
    wb_a = None
    wb_base = None
    wb_b = None
    ext_parts = _capture_external_link_parts(src_xlsx)
    try:
        wb = load_workbook(src_xlsx, data_only=False)
        source_paths = source_paths or {}
        if source_paths.get("A"):
            wb_a = load_workbook(source_paths["A"], data_only=False)
        if source_paths.get("BASE"):
            wb_base = load_workbook(source_paths["BASE"], data_only=False)
        if source_paths.get("B"):
            wb_b = load_workbook(source_paths["B"], data_only=False)
        for op in list(sheet_ops or []):
            kind = str(op.get("kind") or "")
            sheet = str(op.get("sheet") or "")
            if kind == "delete_sheet":
                _remove_sheet_if_exists(wb, sheet)
                continue
            if kind != "copy_sheet" or not sheet:
                continue
            side = str(op.get("source_side") or "").upper()
            src_ws = None
            if side == "A":
                source_wb = wb_a if wb_a is not None else wb
                src_ws = source_wb[sheet] if sheet in source_wb.sheetnames else None
            elif side == "BASE":
                src_ws = wb_base[sheet] if wb_base is not None and sheet in wb_base.sheetnames else None
            elif side in ("B", "THEIRS"):
                src_ws = wb_b[sheet] if wb_b is not None and sheet in wb_b.sheetnames else None
            if src_ws is None:
                continue
            _create_sheet_from_source(wb, src_ws, sheet)
        for op in list(row_ops or []):
            if op.get("kind") != "insert_rows":
                continue
            sheet = op.get("sheet")
            if not sheet or sheet not in wb.sheetnames:
                continue
            row = max(1, int(op.get("row", 1)))
            count = max(1, int(op.get("count", 1)))
            dst_ws = wb[sheet]
            dst_ws.insert_rows(idx=row, amount=count)
            source_side = str(op.get("source_side") or "").upper()
            source_rows = [int(value) for value in (op.get("source_rows") or [])]
            src_ws = None
            if source_side in ("B", "THEIRS") and wb_b is not None and sheet in wb_b.sheetnames:
                src_ws = wb_b[sheet]
            elif source_side == "A" and wb_a is not None and sheet in wb_a.sheetnames:
                src_ws = wb_a[sheet]
            elif source_side == "BASE" and wb_base is not None and sheet in wb_base.sheetnames:
                src_ws = wb_base[sheet]
            if src_ws is not None:
                max_col = max(1, int(src_ws.max_column or 1))
                for offset, source_row in enumerate(source_rows[:count]):
                    _copy_row_metadata(src_ws, dst_ws, source_row, row + offset, max_col)
        for (sheet, r, c), v in manual_ops.items():
            if sheet not in wb.sheetnames:
                continue
            _assign_edit_cell_value(wb[sheet].cell(row=int(r), column=int(c)), v)
        wb.save(out_xlsx)
        if ext_parts:
            _apply_external_link_parts_on_file(out_xlsx, ext_parts)
        _dlog(f"openpyxl merge save ok: {out_xlsx} sheet_ops={len(list(sheet_ops or []))} row_ops={len(list(row_ops or []))} cell_ops={len(manual_ops)}")
        return True
    except Exception as e:
        _dlog(f"openpyxl merge save failed: {e}")
        return False
    finally:
        _wbs_close(wb, wb_a, wb_base, wb_b)


def _merge_cmp_value(v):
    """Build a type-aware comparison key without collapsing meaningful data."""
    try:
        if v is None:
            return "BLANK:"
        if isinstance(v, bool):
            return "BOOL:1" if v else "BOOL:0"
        if isinstance(v, int) and not isinstance(v, bool):
            return f"NUM:{v}"
        if isinstance(v, float):
            if v == 0:
                v = 0.0
            if v.is_integer():
                return f"NUM:{int(v)}"
            return f"NUM:{repr(v)}"
        if isinstance(v, (datetime, date, datetime_time, timedelta)):
            return f"DATE:{repr(v)}"
        if isinstance(v, str):
            return "TEXT:" + v.replace("\r\n", "\n").replace("\r", "\n")
        return f"{type(v).__name__}:{_val_to_str(v)}"
    except Exception:
        return f"FALLBACK:{repr(v)}"


def _scan_formula_cache(path: str):
    """Return (has_formula, missing_cache) based on a sample scan."""
    wb_val = None
    wb_edit = None
    try:
        wb_val = load_workbook(path, data_only=True, read_only=True)
        wb_edit = load_workbook(path, data_only=False, read_only=True)
    except Exception as e:
        _dlog(f"cache check open failed: {e}")
        _wbs_close(wb_val, wb_edit)
        return False, True

    checked = 0
    has_formula = False
    missing_cache = False
    try:
        for sheet in wb_edit.sheetnames:
            ws_e = wb_edit[sheet]
            if sheet not in wb_val.sheetnames:
                continue
            ws_v = wb_val[sheet]
            max_row = max(int(ws_e.max_row or 1), int(ws_v.max_row or 1))
            max_col = max(int(ws_e.max_column or 1), int(ws_v.max_column or 1))
            edit_rows = ws_e.iter_rows(
                min_row=1, max_row=max_row, min_col=1, max_col=max_col,
                values_only=False,
            )
            value_rows = ws_v.iter_rows(
                min_row=1, max_row=max_row, min_col=1, max_col=max_col,
                values_only=False,
            )
            for edit_row, value_row in zip(edit_rows, value_rows):
                for cell, value_cell in zip(edit_row, value_row):
                    if checked >= _CACHE_CHECK_MAX_CELLS:
                        return has_formula, missing_cache
                    f = _formula_text(cell.value)
                    if not f:
                        continue
                    has_formula = True
                    checked += 1
                    v = value_cell.value
                    if v is None:
                        missing_cache = True
                        return has_formula, missing_cache
    finally:
        _wbs_close(wb_val, wb_edit)
    return has_formula, missing_cache


def _recalc_with_excel(path: str) -> str | None:
    """Use Excel COM to recalc formulas and update cached values in a temp copy."""
    try:
        base = os.path.basename(path)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        tmp = os.path.join(tempfile.gettempdir(), f"{APP_NAME}_recalc_{os.getpid()}_{ts}_{base}")
        ext = _workbook_ext(path)
        if not tmp.lower().endswith(ext):
            tmp += ext
        shutil.copy2(path, tmp)
    except Exception as e:
        _dlog(f"recalc copy failed: {e}")
        return None

    try:
        ps = (
            "$ErrorActionPreference='Stop';"
            "$p='" + tmp.replace("'", "''") + "';"
            "$xl=$null;$wb=$null;"
            "try{"
            "$xl=New-Object -ComObject Excel.Application;"
            "$xl.Visible=$false;"
            "$xl.DisplayAlerts=$false;"
            "$xl.AskToUpdateLinks=$false;"
            "$xl.EnableEvents=$false;"
            "$wb=$xl.Workbooks.Open($p,$false,$false);"
            "try{$xl.Calculation=-4105}catch{};"
            "try{$xl.CalculateFullRebuild()}catch{};"
            "try{$wb.RefreshAll();$xl.CalculateFullRebuild()}catch{};"
            "$wb.Save();"
            "}finally{"
            "  if($wb -ne $null){try{$wb.Close($false)}catch{};try{[void][Runtime.InteropServices.Marshal]::FinalReleaseComObject($wb)}catch{}};"
            "  if($xl -ne $null){try{$xl.Quit()}catch{};try{[void][Runtime.InteropServices.Marshal]::FinalReleaseComObject($xl)}catch{}};"
            "  [GC]::Collect();[GC]::WaitForPendingFinalizers();[GC]::Collect();[GC]::WaitForPendingFinalizers();"
            "};"
        )
        no_window = getattr(subprocess, "CREATE_NO_WINDOW", 0)
        r = subprocess.run(
            ["powershell", "-NoProfile", "-STA", "-Command", ps],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=120,
            creationflags=no_window,
        )
        if r.returncode != 0:
            _dlog(f"excel recalc ps failed: {r.stderr.strip()}")
            try:
                if os.path.exists(tmp):
                    os.remove(tmp)
            except Exception:
                pass
            return None
        return tmp
    except Exception as e:
        _dlog(f"excel recalc failed: {e}")
        try:
            if os.path.exists(tmp):
                os.remove(tmp)
        except Exception:
            pass
        return None


def _prepare_val_path(path: str) -> str:
    if not _USE_CACHED_VALUES_ONLY or not _AUTO_RECALC_MISSING_CACHE:
        return path
    try:
        has_formula, missing_cache = _scan_formula_cache(path)
        if has_formula and (_AUTO_RECALC_FORMULAS_ALWAYS or missing_cache):
            _dlog(f"formula cache recalc: has_formula={has_formula} missing_cache={missing_cache} path={path}")
            tmp = _recalc_with_excel(path)
            if tmp:
                _dlog(f"recalc cache OK: {tmp}")
                return tmp
    except Exception as e:
        _dlog(f"prepare val path failed: {e}")
    return path


def _recalc_and_prepare_val_path(path: str) -> str | None:
    """Force Excel recalc to refresh cached values and return temp path."""
    try:
        tmp = _recalc_with_excel(path)
        return tmp
    except Exception:
        return None


def _maybe_recalc_and_prepare_val_path(path: str, force: bool = False) -> str | None:
    """Recalc only when needed, unless forced by the caller."""
    if not path:
        return None
    try:
        if not force:
            if not (_AUTO_RECALC_FORMULAS_ALWAYS or _AUTO_RECALC_MISSING_CACHE):
                return None
            has_formula, missing_cache = _scan_formula_cache(path)
            if not has_formula:
                return None
            if not (_AUTO_RECALC_FORMULAS_ALWAYS or (_AUTO_RECALC_MISSING_CACHE and missing_cache)):
                return None
        return _recalc_and_prepare_val_path(path)
    except Exception as e:
        _dlog(f"maybe recalc path failed: path={path} err={e}")
        return None


def _launch_deferred_copy(src: str, dst: str, retries: int = 60, delay_ms: int = 500) -> bool:
    """Launch a background copy that retries for a while (to avoid lock issues).

    Returns True if the background process was successfully launched, False
    otherwise. A False result means the deferred copy did NOT start, so callers
    must NOT report success to the user.
    """
    try:
        ps = (
            "$src='" + src.replace("'", "''") + "';"
            "$dst='" + dst.replace("'", "''") + "';"
            f"for($i=0;$i -lt {retries};$i++){{"
            "try{Copy-Item -LiteralPath $src -Destination $dst -Force;"
            "Remove-Item -LiteralPath $src -Force;exit 0}catch{Start-Sleep -Milliseconds "
            f"{delay_ms}}};exit 1"
        )
        creationflags = 0
        try:
            creationflags = subprocess.CREATE_NO_WINDOW
        except Exception:
            creationflags = 0
        subprocess.Popen(["powershell", "-NoProfile", "-Command", ps], creationflags=creationflags)
        return True
    except Exception as e:
        _dlog(f"deferred copy launch failed: {e}")
        return False


def _find_tortoise_merge_exe():
    candidates = [
        os.path.join(os.environ.get("ProgramFiles", r"C:\Program Files"), "TortoiseSVN", "bin", "TortoiseMerge.exe"),
        os.path.join(os.environ.get("ProgramFiles(x86)", r"C:\Program Files (x86)"), "TortoiseSVN", "bin", "TortoiseMerge.exe"),
    ]
    for p in candidates:
        if os.path.exists(p):
            return p
    return "TortoiseMerge.exe"


def _find_tortoise_proc_exe():
    candidates = [
        os.path.join(os.environ.get("ProgramFiles", r"C:\Program Files"), "TortoiseSVN", "bin", "TortoiseProc.exe"),
        os.path.join(os.environ.get("ProgramFiles(x86)", r"C:\Program Files (x86)"), "TortoiseSVN", "bin", "TortoiseProc.exe"),
    ]
    for p in candidates:
        if os.path.exists(p):
            return p
    return "TortoiseProc.exe"


def _find_svn_cli_exe() -> str | None:
    """Resolve svn.exe from PATH first, then common TortoiseSVN install paths."""
    p = shutil.which("svn")
    if p:
        return p
    candidates = [
        os.path.join(os.environ.get("ProgramFiles", r"C:\Program Files"), "TortoiseSVN", "bin", "svn.exe"),
        os.path.join(os.environ.get("ProgramFiles(x86)", r"C:\Program Files (x86)"), "TortoiseSVN", "bin", "svn.exe"),
    ]
    for c in candidates:
        try:
            if os.path.exists(c):
                return c
        except Exception:
            pass
    return None


def _query_svn_version(svn_exe: str) -> str:
    try:
        no_window = getattr(subprocess, "CREATE_NO_WINDOW", 0)
        out = subprocess.run(
            [svn_exe, "--version", "--quiet"],
            capture_output=True,
            text=True,
            errors="replace",
            timeout=8,
            creationflags=no_window,
        )
        if out.returncode == 0:
            return (out.stdout or "").strip()
    except Exception:
        pass
    return ""


def _workbook_package_ready(path: str) -> bool:
    """Return True only for a complete, readable OOXML workbook package."""
    try:
        if not path or not os.path.isfile(path) or os.path.getsize(path) <= 0:
            return False
        with zipfile.ZipFile(path, "r") as zf:
            names = set(zf.namelist())
            if "[Content_Types].xml" not in names or "xl/workbook.xml" not in names:
                return False
            return zf.testzip() is None
    except (OSError, zipfile.BadZipFile, EOFError):
        return False


def _wait_for_complete_workbook(path: str, timeout_seconds: float | None = None) -> bool:
    """Wait for asynchronous SVN exports to finish writing a valid OOXML ZIP."""
    timeout = float(_SVN_EXPORT_TIMEOUT_SECS if timeout_seconds is None else timeout_seconds)
    deadline = time.monotonic() + max(0.1, timeout)
    previous_size = -1
    stable_polls = 0
    while time.monotonic() < deadline:
        try:
            size = os.path.getsize(path) if os.path.isfile(path) else -1
        except OSError:
            size = -1
        if size > 0 and size == previous_size:
            stable_polls += 1
        else:
            previous_size = size
            stable_polls = 0
        # A ZIP central directory is written at the end, so a successful full
        # package test is the definitive signal that TortoiseProc has finished.
        if stable_polls >= 2 and _workbook_package_ready(path):
            return True
        time.sleep(0.1)
    ready = _workbook_package_ready(path)
    if not ready:
        _dlog(f"workbook export incomplete: path={path} size={previous_size}")
    return ready


def _try_export_svn_revision_from_merge_temp(path: str) -> str:
    """If path looks like *.merge-left.r#### or *.merge-right.r####, export that revision from WC.

    Returns replacement path if export succeeded; otherwise returns original path.
    """
    try:
        if not path:
            return path
        p = os.path.abspath(path)
        m = re.match(r"^(?P<base>.+)\.merge-(left|right)\.r(?P<rev>\d+)$", p, flags=re.IGNORECASE)
        if not m:
            return path
        base_path = m.group("base")
        rev = m.group("rev")
        if not os.path.exists(base_path):
            # Try same dir + original base filename
            base_path = os.path.join(os.path.dirname(p), os.path.basename(m.group("base")))
        if not os.path.exists(base_path):
            _dlog(f"svn export skip: base not found for {path}")
            return path

        proc_exe = _find_tortoise_proc_exe()
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_name = os.path.basename(base_path)
        save_path = os.path.join(tempfile.gettempdir(), f"{APP_NAME}_svncat_r{rev}_{ts}_{base_name}")
        ext = _workbook_ext(base_name)
        if not save_path.lower().endswith(ext):
            save_path += ext

        try:
            _dlog(f"svn export: {base_path} r{rev} -> {save_path}")
        except Exception:
            pass

        # TortoiseProc may show UI; run and wait briefly for file to appear.
        try:
            subprocess.Popen([
                proc_exe,
                "/command:cat",
                f"/path:{base_path}",
                f"/revision:{rev}",
                f"/savepath:{save_path}",
                "/closeonend:1",
            ])
        except Exception as e:
            _dlog(f"svn export failed launch: {e}")
            return path

        if _wait_for_complete_workbook(save_path):
            return save_path

        _dlog(f"svn export timeout or invalid workbook: {save_path}")
        return path
    except Exception as e:
        _dlog(f"svn export error: {e}")
        return path



def _try_export_svn_base_from_working_copy(path: str) -> str | None:
    """Export BASE revision for a working-copy file path.

    Returns exported temp .xlsx path when successful, otherwise None.
    """
    try:
        if not path:
            return None
        p = os.path.abspath(path)
        if not os.path.exists(p):
            return None

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_name = os.path.basename(p)
        save_path = os.path.join(tempfile.gettempdir(), f"{APP_NAME}_svncat_BASE_{ts}_{base_name}")
        ext = _workbook_ext(base_name)
        if not save_path.lower().endswith(ext):
            save_path += ext

        # Prefer svn CLI (usually uses WC metadata for BASE).
        svn_exe = shutil.which("svn")
        if svn_exe:
            try:
                no_window = getattr(subprocess, "CREATE_NO_WINDOW", 0)
                with open(save_path, "wb") as f:
                    r = subprocess.run(
                        [svn_exe, "cat", "-r", "BASE", p],
                        stdout=f,
                        stderr=subprocess.PIPE,
                        timeout=30,
                        creationflags=no_window,
                    )
                if r.returncode == 0 and _workbook_package_ready(save_path):
                    _dlog(f"svn base export(cli): {p} -> {save_path}")
                    return save_path
                try:
                    if os.path.exists(save_path):
                        os.remove(save_path)
                except Exception:
                    pass
                _dlog(
                    f"svn base export(cli) failed: rc={r.returncode} "
                    f"err={(r.stderr or b'').decode('utf-8', errors='ignore')}"
                )
            except Exception as e:
                _dlog(f"svn base export(cli) exception: {e}")

        # Prefer the working-copy pristine blob over asynchronous TortoiseProc
        # export. It is the exact WC BASE and copying it is synchronous.
        try:
            import sqlite3

            p_dir = os.path.dirname(p)
            wc_root = None
            probe = p_dir
            while probe:
                wc_db = os.path.join(probe, ".svn", "wc.db")
                if os.path.isfile(wc_db):
                    wc_root = probe
                    break
                parent = os.path.dirname(probe)
                if not parent or parent == probe:
                    break
                probe = parent
            if wc_root is None:
                raise FileNotFoundError(f"working-copy metadata not found for {p}")

            rel_path = os.path.relpath(p, wc_root).replace("\\", "/")
            wc_db = os.path.join(wc_root, ".svn", "wc.db")
            conn = sqlite3.connect(wc_db)
            try:
                cur = conn.cursor()
                cur.execute(
                    """
                    select checksum
                    from NODES
                    where local_relpath = ?
                      and op_depth = 0
                      and kind = 'file'
                      and presence = 'normal'
                    limit 1
                    """,
                    (rel_path,),
                )
                row = cur.fetchone()
            finally:
                try:
                    conn.close()
                except Exception:
                    pass
            checksum = row[0] if row else None
            m = re.match(r"^\$sha1\$([0-9a-fA-F]{40})$", str(checksum or ""))
            if not m:
                raise RuntimeError(f"working-copy pristine checksum not found for {p}")
            sha1 = m.group(1).lower()
            pristine_path = os.path.join(wc_root, ".svn", "pristine", sha1[:2], sha1 + ".svn-base")
            if not os.path.exists(pristine_path):
                raise FileNotFoundError(f"working-copy pristine file not found: {pristine_path}")
            part_path = save_path + ".part"
            shutil.copyfile(pristine_path, part_path)
            os.replace(part_path, save_path)
            if _workbook_package_ready(save_path):
                _dlog(f"svn base export(pristine): {p} -> {save_path}")
                return save_path
            _dlog(f"svn base export(pristine) invalid workbook: {pristine_path}")
        except Exception as e:
            _dlog(f"svn base export(pristine) exception: {e}")

        # Last fallback: TortoiseProc cat BASE. TortoiseProc writes the output
        # asynchronously, so file existence alone is not sufficient here.
        try:
            proc_exe = _find_tortoise_proc_exe()
            subprocess.Popen([
                proc_exe,
                "/command:cat",
                f"/path:{p}",
                "/revision:BASE",
                f"/savepath:{save_path}",
                "/closeonend:1",
            ])
            if _wait_for_complete_workbook(save_path):
                _dlog(f"svn base export(tortoise): {p} -> {save_path}")
                return save_path
            _dlog(f"svn base export(tortoise) incomplete: {save_path}")
        except Exception as e:
            _dlog(f"svn base export(tortoise) exception: {e}")
    except Exception as e:
        _dlog(f"svn base export error: {e}")
        return None
    return None


def _find_handle_exe():
    candidates = [
        os.path.join(os.environ.get("SystemRoot", r"C:\Windows"), "System32", "handle.exe"),
        r"C:\Sysinternals\handle.exe",
        r"C:\Tools\Sysinternals\handle.exe",
        r"D:\Tools\Sysinternals\handle.exe",
    ]
    for p in candidates:
        if os.path.exists(p):
            return p
    return shutil.which("handle.exe")


def _log_lock_holders(path: str) -> bool:
    """Return True if Excel is detected holding the file."""
    excel_found = False
    try:
        handle_exe = _find_handle_exe()
        if not handle_exe:
            _dlog(f"lock holders: handle.exe not found for {path}")
            return False
        no_window = getattr(subprocess, "CREATE_NO_WINDOW", 0)
        r = subprocess.run(
            [handle_exe, "-accepteula", path],
            capture_output=True,
            text=True,
            errors="replace",
            timeout=10,
            creationflags=no_window,
        )
        out = (r.stdout or "") + (r.stderr or "")
        if out.strip():
            for line in out.splitlines():
                if path.lower() in line.lower():
                    _dlog(f"lock holders: {line.strip()}")
                    if "excel.exe" in line.lower():
                        excel_found = True
            return excel_found
        _dlog(f"lock holders: no output for {path}")
    except Exception as e:
        _dlog(f"lock holders: failed {e}")
    return excel_found


def _try_svn_resolve(path: str) -> bool:
    """Attempt to mark conflict as resolved in SVN."""
    try:
        svn_exe = shutil.which("svn")
        if svn_exe:
            subprocess.run(
                [svn_exe, "resolve", "--accept", "working", path],
                check=False,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
            )
            return True
    except Exception:
        pass
    # Fallback to TortoiseProc (may show UI)
    try:
        proc_exe = _find_tortoise_proc_exe()
        subprocess.Popen([proc_exe, "/command:resolve", f"/path:{path}", "/closeonend:1"])
        return True
    except Exception:
        return False


def open_tortoise_merge(left_txt: str, right_txt: str, title: str):
    exe = _find_tortoise_merge_exe()
    args = [exe, "/base", left_txt, "/mine", right_txt, "/title", title]
    subprocess.Popen(args)


def _show_conflict_popup(conflicts):
    try:
        root = tk.Tk()
        root.withdraw()
        win = tk.Toplevel(root)
        win.title("发现冲突")
        win.resizable(False, False)
        win.geometry("+{}+{}".format(root.winfo_screenwidth() // 2 - 220, root.winfo_screenheight() // 2 - 180))
        msg = "与其他同学冲突，请联系确认后再修改保存！！！"
        lbl = tk.Label(win, text=msg, fg="red", font=("Microsoft YaHei", 12, "bold"), padx=16, pady=10)
        lbl.pack()

        detail_lines = []
        for sheet, r, c, _vm, _vt in conflicts[:3]:
            col = get_column_letter(c)
            detail_lines.append(f"{sheet}!{col}{r}")
        if len(conflicts) > 3:
            detail_lines.append("...")
        detail_text = "\n".join(detail_lines) if detail_lines else "（无）"
        txt = tk.Text(win, height=12, width=60)
        txt.insert("1.0", detail_text)
        txt.configure(state="disabled")
        txt.pack(padx=12, pady=(0, 10))

        tk.Button(win, text="确定", command=win.destroy).pack(pady=(0, 10))
        win.grab_set()
        win.wait_window()
        root.destroy()
    except Exception:
        pass


def _run_startup_progress_task(title: str, message: str, fn):
    """Run startup work off the Tk thread while keeping a progress window alive.

    ``fn`` receives ``report(message, detail=None, percent=None)``.  The helper is
    intentionally independent from ``SowMergeApp`` because merge conflict scanning
    happens before the main application object exists.
    """
    global _STARTUP_PROGRESS_ROOT
    root = _STARTUP_PROGRESS_ROOT
    try:
        root_exists = root is not None and bool(root.winfo_exists())
    except Exception:
        root_exists = False
    if not root_exists:
        root = tk.Tk()
        _STARTUP_PROGRESS_ROOT = root
    else:
        for child in list(root.winfo_children()):
            try:
                child.destroy()
            except Exception:
                pass
    root.title(title)
    root.resizable(False, False)
    root.protocol("WM_DELETE_WINDOW", lambda: None)
    root.deiconify()

    frame = ttk.Frame(root, padding=18)
    frame.pack(fill="both", expand=True)
    title_label = ttk.Label(frame, text=message, font=("Microsoft YaHei", 11, "bold"))
    title_label.pack(anchor="w")
    detail_label = ttk.Label(frame, text="正在准备...", foreground="#555", wraplength=500)
    detail_label.pack(
        anchor="w", fill="x", pady=(8, 10)
    )
    progress = ttk.Progressbar(frame, mode="indeterminate", length=500)
    progress.pack(fill="x")
    elapsed_label = ttk.Label(frame, text="已用时 0.0 秒", foreground="#777")
    elapsed_label.pack(anchor="e", pady=(7, 0))
    progress.start(12)

    root.update_idletasks()
    width = 550
    height = 155
    x = max(0, (root.winfo_screenwidth() - width) // 2)
    y = max(0, (root.winfo_screenheight() - height) // 2)
    root.geometry(f"{width}x{height}+{x}+{y}")

    updates_lock = threading.Lock()
    updates: list[tuple[str | None, str | None, float | None]] = []
    state: dict[str, object] = {}
    done = threading.Event()
    started = time.monotonic()

    def report(text=None, detail=None, percent=None):
        pct = None
        if percent is not None:
            try:
                pct = max(0.0, min(100.0, float(percent)))
            except Exception:
                pct = None
        with updates_lock:
            updates.append((text, detail, pct))

    def worker():
        try:
            state["result"] = fn(report)
        except BaseException as exc:
            state["error"] = exc
            state["traceback"] = traceback.format_exc()
        finally:
            done.set()

    # Tk objects must be finalized on the UI thread. openpyxl's XML parsing can
    # trigger cyclic GC in the worker, so defer cyclic collection until the
    # temporary Tk interpreter has been destroyed on this thread.
    gc_was_enabled = gc.isenabled()
    if gc_was_enabled:
        gc.disable()
    thread = threading.Thread(target=worker, daemon=True, name="sow-startup-progress")
    thread.start()

    def poll():
        latest = None
        with updates_lock:
            if updates:
                latest = updates[-1]
                updates.clear()
        if latest is not None:
            text_value, detail_value, percent_value = latest
            if text_value:
                title_label.configure(text=str(text_value))
            if detail_value is not None:
                detail_label.configure(text=str(detail_value))
            if percent_value is not None:
                progress.stop()
                progress.configure(mode="determinate", maximum=100, value=percent_value)
        elapsed_label.configure(text=f"已用时 {time.monotonic() - started:.1f} 秒")
        if done.is_set():
            try:
                progress.stop()
                root.quit()
            except Exception:
                pass
            return
        root.after(80, poll)

    root.after(50, poll)
    try:
        root.mainloop()
    finally:
        try:
            root.withdraw()
        except Exception:
            pass
        if gc_was_enabled:
            gc.enable()
            gc.collect()
    if "error" in state:
        _dlog(f"startup task failed: {state.get('traceback', state['error'])}")
        raise state["error"]
    return state.get("result")


def _destroy_startup_progress_root():
    global _STARTUP_PROGRESS_ROOT
    root = _STARTUP_PROGRESS_ROOT
    _STARTUP_PROGRESS_ROOT = None
    if root is None:
        return
    try:
        for aid in list(root.tk.call("after", "info") or ()):
            try:
                root.after_cancel(aid)
            except Exception:
                pass
    except Exception:
        pass
    try:
        if root.winfo_exists():
            root.destroy()
    except Exception:
        pass


def _take_startup_progress_root():
    """Promote the startup Tk interpreter to the application's main window."""
    global _STARTUP_PROGRESS_ROOT
    root = _STARTUP_PROGRESS_ROOT
    _STARTUP_PROGRESS_ROOT = None
    try:
        root_exists = root is not None and bool(root.winfo_exists())
    except Exception:
        root_exists = False
    if not root_exists:
        return tk.Tk()
    try:
        for aid in list(root.tk.call("after", "info") or ()):
            try:
                root.after_cancel(aid)
            except Exception:
                pass
    except Exception:
        pass
    for child in list(root.winfo_children()):
        try:
            child.destroy()
        except Exception:
            pass
    try:
        root.deiconify()
    except Exception:
        pass
    return root


def excel_to_text(path: str, out_path: str, thick_sep_char: str = "="):
    val_path = _prepare_val_path(path)
    wb = load_workbook(val_path, data_only=True)
    try:
        with open(out_path, "w", encoding="utf-8") as f:
            f.write(f"{APP_NAME} text export\n")
            f.write(f"Source: {path}\n")
            f.write(f"Time: {datetime.now().isoformat(sep=' ', timespec='seconds')}\n\n")

            for idx, name in enumerate(wb.sheetnames):
                ws = wb[name]
                max_row = ws.max_row or 1
                max_col = ws.max_column or 1

                if idx != 0:
                    f.write("\n" + (thick_sep_char * 120) + "\n")

                title = f"SHEET: {name}"
                pad = max(0, 120 - len(title) - 2)
                left = thick_sep_char * (pad // 2)
                right = thick_sep_char * (pad - (pad // 2))
                f.write(f"{left} {title} {right}\n")

                cols = [ws.cell(row=1, column=c).coordinate[:-1] for c in range(1, max_col + 1)]
                f.write("ROW\t" + "\t".join(cols) + "\n")

                for r in range(1, max_row + 1):
                    vals = []
                    for c in range(1, max_col + 1):
                        vals.append(_val_to_str(ws.cell(row=r, column=c).value))
                    f.write(str(r) + "\t" + "\t".join(vals) + "\n")
    finally:
        _wbs_close(wb)
        if val_path != path:
            try:
                os.remove(val_path)
            except Exception:
                pass


def pick_two_files_same_name():
    root = tk.Tk()
    root.withdraw()

    a = filedialog.askopenfilename(title="Select first Excel workbook", filetypes=[("Excel Workbook", "*.xlsx *.xlsm")])
    if not a:
        return None, None
    b = filedialog.askopenfilename(title="Select second Excel workbook (same filename)", filetypes=[("Excel Workbook", "*.xlsx *.xlsm")])
    if not b:
        return None, None

    if os.path.basename(a).lower() != os.path.basename(b).lower():
        messagebox.showerror(
            "Filename mismatch",
            f"The two files must have the same filename.\n\nA: {os.path.basename(a)}\nB: {os.path.basename(b)}",
        )
        return None, None

    return a, b


def _detect_svn_conflict_files(target_path: str):
    # If user selected a conflict artifact directly, map back to merged target first.
    try:
        p = os.path.abspath(target_path)
        m = re.match(r"^(?P<base>.+)\.merge-(left|right)\.r\d+$", p, flags=re.IGNORECASE)
        if m:
            target_path = m.group("base")
    except Exception:
        pass
    folder = os.path.dirname(target_path)
    base_name = os.path.basename(target_path)
    # SVN conflict artifacts:
    # - file.merge-left.r#### / file.merge-right.r#### (newer SVN)
    # - file.r<rev> (older SVN), possibly .mine
    merge_left = []
    merge_right = []
    for name in os.listdir(folder):
        if name.startswith(base_name + ".merge-left.r"):
            suffix = name[len(base_name) + len(".merge-left.r"):]
            if suffix.isdigit():
                merge_left.append((int(suffix), os.path.join(folder, name)))
        elif name.startswith(base_name + ".merge-right.r"):
            suffix = name[len(base_name) + len(".merge-right.r"):]
            if suffix.isdigit():
                merge_right.append((int(suffix), os.path.join(folder, name)))
        elif name == base_name + ".merge-left":
            merge_left.append((0, os.path.join(folder, name)))
        elif name == base_name + ".merge-right":
            merge_right.append((0, os.path.join(folder, name)))
    if merge_left and merge_right:
        merge_left.sort(key=lambda x: x[0])
        merge_right.sort(key=lambda x: x[0])
        base_path = merge_left[-1][1]
        theirs_path = merge_right[-1][1]
        mine_path = target_path
        merged_path = target_path
        return base_path, mine_path, theirs_path, merged_path

    # Older SVN conflict artifacts: file.r<rev> (numeric), possibly .mine
    r_files = []
    for name in os.listdir(folder):
        if not name.startswith(base_name + ".r"):
            continue
        suffix = name[len(base_name) + 2:]
        if suffix.isdigit():
            r_files.append((int(suffix), os.path.join(folder, name)))
    if len(r_files) >= 2:
        r_files.sort(key=lambda x: x[0])
        base_path = r_files[0][1]
        theirs_path = r_files[-1][1]
        mine_path = target_path
        merged_path = target_path
        return base_path, mine_path, theirs_path, merged_path
    # Fallback for rOLD/rNEW naming
    r_old = os.path.join(folder, base_name + ".rOLD")
    r_new = os.path.join(folder, base_name + ".rNEW")
    if os.path.exists(r_old) and os.path.exists(r_new):
        return r_old, target_path, r_new, target_path
    # Fallback: fuzzy match for temp-stable names that still contain "<base>.merge-left/right.r####"
    # e.g. sow_merge_tool_stable_..._<base>.merge-right.r27548_...
    try:
        merge_left_fuzzy = []
        merge_right_fuzzy = []
        key = (base_name + ".merge-").lower()
        for name in os.listdir(folder):
            low = name.lower()
            if key not in low:
                continue
            i_left = low.find((base_name + ".merge-left.r").lower())
            if i_left >= 0:
                j = i_left + len((base_name + ".merge-left.r").lower())
                rev = []
                while j < len(low) and low[j].isdigit():
                    rev.append(low[j])
                    j += 1
                if rev:
                    merge_left_fuzzy.append((int("".join(rev)), os.path.join(folder, name)))
            i_right = low.find((base_name + ".merge-right.r").lower())
            if i_right >= 0:
                j = i_right + len((base_name + ".merge-right.r").lower())
                rev = []
                while j < len(low) and low[j].isdigit():
                    rev.append(low[j])
                    j += 1
                if rev:
                    merge_right_fuzzy.append((int("".join(rev)), os.path.join(folder, name)))
        if merge_left_fuzzy and merge_right_fuzzy:
            merge_left_fuzzy.sort(key=lambda x: x[0])
            merge_right_fuzzy.sort(key=lambda x: x[0])
            return merge_left_fuzzy[-1][1], target_path, merge_right_fuzzy[-1][1], target_path
    except Exception:
        pass
    return None


def _trace_launch(msg: str):
    try:
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
        with open(_LAUNCH_TRACE_PATH, "a", encoding="utf-8") as f:
            f.write(f"[{ts}] {msg}\n")
    except Exception:
        pass


def _has_svn_conflict_artifacts(target_path: str) -> bool:
    try:
        folder = os.path.dirname(target_path)
        base_name = os.path.basename(target_path)
        mine = os.path.join(folder, base_name + ".mine")
        if os.path.exists(mine):
            return True
        for name in os.listdir(folder):
            if name.startswith(base_name + ".merge-left") or name.startswith(base_name + ".merge-right"):
                return True
        r_old = os.path.join(folder, base_name + ".rOLD")
        r_new = os.path.join(folder, base_name + ".rNEW")
        if os.path.exists(r_old) or os.path.exists(r_new):
            return True
        for name in os.listdir(folder):
            if name.startswith(base_name + ".r"):
                suffix = name[len(base_name) + 2:]
                if suffix.isdigit():
                    return True
    except Exception:
        pass
    return False


def _find_conflict_in_dir(folder: str):
    try:
        # If there is exactly one conflicted file in folder, return it.
        base_names = set()
        for name in os.listdir(folder):
            if ".merge-left" in name:
                base = name.split(".merge-left")[0]
                base_names.add(base)
                continue
            if ".merge-right" in name:
                base = name.split(".merge-right")[0]
                base_names.add(base)
                continue
            if ".r" in name:
                base = name.split(".r")[0]
                base_names.add(base)
        candidates = []
        for base in base_names:
            target = os.path.join(folder, base)
            if os.path.exists(target) and _has_svn_conflict_artifacts(target):
                candidates.append(target)
        if len(candidates) == 1:
            return candidates[0]
    except Exception:
        pass
    return None


def _auto_pick_conflict_file():
    # Best-effort: try current working directory
    try:
        cwd = os.getcwd()
        p = _find_conflict_in_dir(cwd)
        if p:
            return p
        try:
            svn_exe = shutil.which("svn")
            if svn_exe:
                no_window = getattr(subprocess, "CREATE_NO_WINDOW", 0)
                r = subprocess.run(
                    [svn_exe, "status", cwd],
                    capture_output=True,
                    text=True,
                    errors="replace",
                    timeout=8,
                    creationflags=no_window,
                )
                if r.returncode == 0:
                    conflicted = []
                    for line in (r.stdout or "").splitlines():
                        if not line:
                            continue
                        if line[0] != "C":
                            continue
                        rel = line[8:].strip() if len(line) > 8 else ""
                        if not rel:
                            continue
                        cand = os.path.abspath(os.path.join(cwd, rel))
                        if os.path.exists(cand):
                            conflicted.append(cand)
                    if len(conflicted) == 1:
                        return conflicted[0]
        except Exception:
            pass
        # Walk up to find SVN working copy root (.svn)
        cur = cwd
        wc_root = None
        while True:
            if os.path.isdir(os.path.join(cur, ".svn")):
                wc_root = cur
                break
            parent = os.path.dirname(cur)
            if parent == cur:
                break
            cur = parent
        if wc_root:
            # If exactly one conflicted file exists in the working copy, auto-pick it
            candidates = []
            for root, _dirs, files in os.walk(wc_root):
                base_names = set()
                for name in files:
                    if ".r" in name:
                        base = name.split(".r")[0]
                        base_names.add(base)
                for base in base_names:
                    target = os.path.join(root, base)
                    if os.path.exists(target) and _has_svn_conflict_artifacts(target):
                        candidates.append(target)
                        if len(candidates) > 1:
                            return None
            if len(candidates) == 1:
                return candidates[0]
    except Exception:
        pass
    return None


def pick_files_or_conflict():
    root = tk.Tk()
    root.withdraw()

    auto = _auto_pick_conflict_file()
    if auto:
        conflict = _detect_svn_conflict_files(auto)
        if conflict:
            return ("merge",) + conflict + (True,)

    a = filedialog.askopenfilename(title="Select Excel workbook", filetypes=[("Excel Workbook", "*.xlsx *.xlsm")])
    if not a:
        return None

    conflict = _detect_svn_conflict_files(a)
    if conflict:
        return ("merge",) + conflict + (True,)

    b = filedialog.askopenfilename(title="Select second Excel workbook (same filename)", filetypes=[("Excel Workbook", "*.xlsx *.xlsm")])
    if not b:
        return None

    if os.path.basename(a).lower() != os.path.basename(b).lower():
        messagebox.showerror(
            "Filename mismatch",
            f"The two files must have the same filename.\n\nA: {os.path.basename(a)}\nB: {os.path.basename(b)}",
        )
        return None

    return ("diff", a, b)


def _atomic_save_wb(wb, target_path: str):
    """Safely overwrite a workbook."""
    folder = os.path.dirname(target_path)
    if folder:
        os.makedirs(folder, exist_ok=True)
    base = os.path.basename(target_path)
    tmp_path = os.path.join(folder, f"~{base}.{os.getpid()}.tmp")
    try:
        if _FAST_SAVE_VALUES_ONLY and _USE_CACHED_VALUES_ONLY:
            _save_values_only_from_wb(wb, tmp_path)
        else:
            wb.save(tmp_path)
        os.replace(tmp_path, target_path)
    except Exception:
        try:
            os.remove(tmp_path)
        except Exception:
            pass
        raise


def _ensure_xlsx_copy(path: str) -> str:
    """If path is not a directly openable workbook path, copy it to a temp workbook path."""
    if not path:
        return path
    if os.path.splitext(path)[1].lower() in _SUPPORTED_WORKBOOK_EXTS:
        return _ensure_stable_copy(path)
    try:
        if not _wait_for_complete_workbook(path):
            raise RuntimeError(f"SVN 临时 Excel 文件尚未完整生成或已损坏：{path}")
        base = os.path.basename(path)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        ext = _workbook_ext(path)
        tmp = os.path.join(tempfile.gettempdir(), f"{APP_NAME}_svn_{base}_{ts}")
        if not tmp.lower().endswith(ext):
            tmp += ext
        shutil.copy2(path, tmp)
        if not _workbook_package_ready(tmp):
            raise RuntimeError(f"SVN 临时 Excel 文件复制后校验失败：{path}")
        return tmp
    except RuntimeError:
        raise
    except Exception as e:
        raise RuntimeError(f"无法准备 SVN 临时 Excel 文件：{path}\n{e}") from e


def _ensure_stable_copy(path: str) -> str:
    """If path looks like a temp/svn artifact, copy to a stable temp file."""
    if not path:
        return path
    try:
        temp_root = os.path.abspath(tempfile.gettempdir()).lower()
        p_abs = os.path.abspath(path)
        p_low = p_abs.lower()
        base = os.path.basename(path)
        looks_temp = p_low.startswith(temp_root)
        looks_svn = ".svn" in p_low or ".r" in base or "revbase" in base.lower() or "rev" in base.lower()
        if looks_temp or looks_svn:
            if not _wait_for_complete_workbook(p_abs):
                raise RuntimeError(f"SVN 临时 Excel 文件尚未完整生成或已损坏：{path}")
            if base.lower().startswith(f"{APP_NAME}_stable_"):
                return p_abs
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            tmp = os.path.join(tempfile.gettempdir(), f"{APP_NAME}_stable_{ts}_{base}")
            ext = _workbook_ext(path)
            if not tmp.lower().endswith(ext):
                tmp += ext
            shutil.copy2(path, tmp)
            if not _workbook_package_ready(tmp):
                raise RuntimeError(f"SVN 临时 Excel 文件稳定副本校验失败：{path}")
            return tmp
    except RuntimeError:
        raise
    except Exception:
        pass
    return path


def _is_temp_base_path(path: str) -> bool:
    if not path:
        return False
    p = os.path.abspath(path).lower()
    base = os.path.basename(path).lower()
    if p.startswith(os.path.abspath(tempfile.gettempdir()).lower()):
        return True
    if "revbase" in base or ".svn" in p or base.endswith(".tmp.xlsx") or base.endswith(".tmp.xlsm") or ".r" in base:
        return True
    return False


def _wbs_close(*wbs):
    """Safely close one or more openpyxl workbooks, ignoring errors."""
    for wb in wbs:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass
            try:
                vba_archive = getattr(wb, "vba_archive", None)
                if vba_archive is not None:
                    vba_archive.close()
                    try:
                        vba_archive.fp = None
                    except Exception:
                        pass
                    try:
                        wb.vba_archive = None
                    except Exception:
                        pass
            except Exception:
                pass


def _merge_three_way(base_path: str, mine_path: str, theirs_path: str, merged_path: str, save_merged: bool = True):
    """3-way merge using row-aligned diff (consistent with _scan_three_way_conflicts).

    Conflict: mine and theirs both changed an aligned cell differently vs base.
    Non-conflicting theirs changes are applied onto mine for 1:1 aligned rows.

    Row alignment (via mine_to_base / theirs_to_base / display_pairs) is used instead
    of comparing by physical row number, so inserted/deleted rows no longer cause
    cross-row mis-comparison or theirs being written to the wrong row. Structural rows
    that exist on only one side (inserts/deletes) are left for the manual 3-way UI
    rather than blindly written by physical position.

    Returns (conflicts, merged_preview_path, conflict_cells_by_sheet).
    """
    base_path = _ensure_xlsx_copy(base_path)
    mine_path = _ensure_xlsx_copy(mine_path)
    theirs_path = _ensure_xlsx_copy(theirs_path)

    base_val_path = _prepare_val_path(base_path)
    mine_val_path = _prepare_val_path(mine_path)
    theirs_val_path = _prepare_val_path(theirs_path)

    wb_base_val = load_workbook(base_val_path, data_only=True)
    wb_mine_val = load_workbook(mine_val_path, data_only=True)
    wb_theirs_val = load_workbook(theirs_val_path, data_only=True)
    wb_mine = load_workbook(mine_path, data_only=False)
    wb_base_edit = load_workbook(base_path, data_only=False)
    wb_theirs_edit = load_workbook(theirs_path, data_only=False)

    # Start merged as mine (copy in-memory by reusing workbook; then save to merged_path)
    wb_merged = wb_mine

    set_base = set(wb_base_val.sheetnames)
    set_mine = set(wb_mine_val.sheetnames)
    set_theirs = set(wb_theirs_val.sheetnames)

    # If a sheet exists only in theirs, copy into merged preserving formulas.
    only_theirs = sorted(set_theirs - set_mine)
    for name in only_theirs:
        ws_t = wb_theirs_edit[name]  # data_only=False preserves formulas
        ws_m = wb_merged.create_sheet(title=name)
        _copy_sheet_basic(ws_t, ws_m)

    conflicts = []
    conflict_cells_by_sheet = {}

    def _cmp_cell(
        ws_val,
        ws_edit,
        formula_map,
        row_idx: int | None,
        col_idx: int,
        compare_row: int | None = None,
    ):
        if ws_val is None or row_idx is None or row_idx <= 0:
            return None, _merge_cell_compare_key(None, None)
        try:
            vv = ws_val.cell(row=row_idx, column=col_idx).value
        except Exception:
            vv = None
        missing_formula = object()
        ve = formula_map.get((int(row_idx), int(col_idx)), missing_formula)
        if ve is missing_formula:
            ve = None
            if vv is None:
                try:
                    if ws_edit is not None:
                        ve = ws_edit.cell(row=row_idx, column=col_idx).value
                except Exception:
                    ve = None
        if compare_row is not None and row_idx != compare_row:
            ve = _translate_normal_formula_for_compare(
                vv, ve, row_idx, col_idx, compare_row, col_idx
            )
        cmp_v = vv
        try:
            if cmp_v is None:
                if ve is not None and not _formula_text(ve):
                    cmp_v = ve
        except Exception:
            pass
        return cmp_v, _merge_cell_compare_key(vv, ve)

    common = sorted(set_mine & set_theirs)
    for name in common:
        ws_b = wb_base_val[name] if name in set_base else None
        ws_m_val = wb_mine_val[name]
        ws_t = wb_theirs_val[name]
        ws_m_edit = wb_mine[name]
        ws_b_edit = wb_base_edit[name] if name in wb_base_edit.sheetnames else None
        ws_t_edit = wb_theirs_edit[name] if name in wb_theirs_edit.sheetnames else None
        formula_map_m = _formula_edit_value_map(ws_m_edit)
        formula_map_b = _formula_edit_value_map(ws_b_edit)
        formula_map_t = _formula_edit_value_map(ws_t_edit)

        max_col = max(ws_m_val.max_column or 1, ws_t.max_column or 1, (ws_b.max_column or 1) if ws_b else 1)

        # Same row-alignment pipeline as _scan_three_way_conflicts so auto-merge and
        # the UI agree on which cells line up across inserted/deleted rows.
        mine_sigs = _row_sig_list_for_ws(ws_m_val, ws_m_val.max_row or 1, max_col)
        theirs_sigs = _row_sig_list_for_ws(ws_t, ws_t.max_row or 1, max_col)
        base_sigs = _row_sig_list_for_ws(ws_b, ws_b.max_row or 1, max_col) if ws_b is not None else []
        mine_to_base = _row_map_from_pairs(
            _compute_row_pairs_from_signatures(mine_sigs, base_sigs)
        ) if ws_b is not None else {}
        theirs_to_base = _row_map_from_pairs(
            _compute_row_pairs_from_signatures(theirs_sigs, base_sigs)
        ) if ws_b is not None else {}
        display_pairs = _compute_row_pairs_from_signatures(mine_sigs, theirs_sigs)
        if ws_b is not None:
            display_pairs = _split_tail_independent_append_pairs(
                display_pairs, mine_to_base, theirs_to_base,
                ws_m_val, ws_t, max_col,
            )
            display_pairs = _split_low_similarity_tail_pairs(
                display_pairs,
                mine_to_base,
                theirs_to_base,
                ws_m_val,
                ws_t,
                max_col,
            )

        for ra, rb in display_pairs:
            base_row_m = mine_to_base.get(ra) if ra is not None else None
            base_row_t = theirs_to_base.get(rb) if rb is not None else None
            for c in range(1, max_col + 1):
                compare_row = ra if ra is not None else rb
                vm_cmp, vm_key = _cmp_cell(ws_m_val, ws_m_edit, formula_map_m, ra, c, compare_row)
                vt_cmp, vt_key = _cmp_cell(ws_t, ws_t_edit, formula_map_t, rb, c, compare_row)
                conflict_row = ra if ra is not None else (rb if rb is not None else 0)

                # Aligned 1:1 row sharing a base row -> full 3-way cell logic.
                if ra is not None and rb is not None and base_row_m is not None and base_row_m == base_row_t:
                    vb_cmp, vb_key = _cmp_cell(ws_b, ws_b_edit, formula_map_b, base_row_m, c, compare_row)
                    if vb_cmp is None and vt_cmp is not None and ws_b_edit is not None and ws_t_edit is not None:
                        try:
                            if _same_formula(ws_b_edit.cell(row=base_row_m, column=c).value, ws_t_edit.cell(row=rb, column=c).value):
                                vb_cmp = vt_cmp
                                vb_key = vt_key
                        except Exception:
                            pass
                    mine_changed = (vm_key != vb_key)
                    theirs_changed = (vt_key != vb_key)
                    if mine_changed and theirs_changed:
                        if vm_key != vt_key:
                            conflicts.append((name, conflict_row, c, vm_cmp, vt_cmp))
                            conflict_cells_by_sheet.setdefault(name, {}).setdefault(conflict_row, set()).add(c)
                        # else: identical change on both sides; keep mine as-is.
                    elif (not mine_changed) and theirs_changed:
                        # Safe to apply theirs onto mine's aligned row (preserve formulas).
                        _t_edit_v = ws_t_edit.cell(row=rb, column=c).value if ws_t_edit is not None else None
                        _m_edit_v = ws_m_edit.cell(row=ra, column=c).value
                        try:
                            _new_edit_v = _copy_edit_value_for_destination(
                                vt_cmp,
                                _t_edit_v,
                                _m_edit_v,
                                src_row=rb,
                                src_col=c,
                                dst_row=ra,
                                dst_col=c,
                            )
                        except RuntimeError:
                            conflicts.append((name, conflict_row, c, vm_cmp, vt_cmp))
                            conflict_cells_by_sheet.setdefault(name, {}).setdefault(conflict_row, set()).add(c)
                            continue
                        if not _is_same_formula_copy_noop(_new_edit_v, _m_edit_v):
                            _assign_edit_cell_value(
                                ws_m_edit.cell(row=ra, column=c),
                                _new_edit_v,
                            )
                    continue

                # Both sides added the same logical row independently (no shared base):
                # differing values are a conflict for manual resolution.
                if ra is not None and rb is not None and base_row_m is None and base_row_t is None:
                    if vm_key != vt_key:
                        conflicts.append((name, conflict_row, c, vm_cmp, vt_cmp))
                        conflict_cells_by_sheet.setdefault(name, {}).setdefault(conflict_row, set()).add(c)
                # ra is None (theirs-only row) or rb is None (mine-only row): structural
                # insert/delete -> left for the manual 3-way UI, never written by physical row.

    _merge_result = None
    try:
        # Always save a preview for UI if needed
        if conflicts or (not save_merged):
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            preview = os.path.join(
                tempfile.gettempdir(),
                f"{APP_NAME}_merged_preview_{os.getpid()}_{ts}{_workbook_ext(merged_path or mine_path)}",
            )
            _atomic_save_wb(wb_merged, preview)
            _merge_result = (conflicts, preview, conflict_cells_by_sheet)
        else:
            # No conflicts: save directly to merged path
            _atomic_save_wb(wb_merged, merged_path)
            _merge_result = ([], None, {})
    finally:
        _wbs_close(wb_base_val, wb_mine_val, wb_theirs_val, wb_mine, wb_base_edit, wb_theirs_edit)
        for _orig, _vp in ((base_path, base_val_path), (mine_path, mine_val_path), (theirs_path, theirs_val_path)):
            if _vp != _orig:
                try:
                    os.remove(_vp)
                except Exception:
                    pass
    return _merge_result


def _scan_three_way_conflicts(base_path: str, mine_path: str, theirs_path: str):
    """Detect 3-way conflicts only; do NOT auto-apply theirs before UI."""
    base_path = _ensure_xlsx_copy(base_path)
    mine_path = _ensure_xlsx_copy(mine_path)
    theirs_path = _ensure_xlsx_copy(theirs_path)

    base_val_path = _prepare_val_path(base_path)
    mine_val_path = _prepare_val_path(mine_path)
    theirs_val_path = _prepare_val_path(theirs_path)

    wb_base_val = None
    wb_mine_val = None
    wb_theirs_val = None
    wb_mine_edit = None
    wb_base_edit = None
    wb_theirs_edit = None
    try:
        wb_base_val = load_workbook(base_val_path, data_only=True)
        wb_mine_val = load_workbook(mine_val_path, data_only=True)
        wb_theirs_val = load_workbook(theirs_val_path, data_only=True)
        wb_mine_edit = load_workbook(mine_path, data_only=False)
        wb_base_edit = load_workbook(base_path, data_only=False)
        wb_theirs_edit = load_workbook(theirs_path, data_only=False)

        conflicts = []
        conflict_cells_by_sheet = {}

        set_base = set(wb_base_val.sheetnames)
        set_mine = set(wb_mine_val.sheetnames)
        set_theirs = set(wb_theirs_val.sheetnames)
        common = sorted(set_mine & set_theirs)

        def _cmp_cell(
            ws_val,
            ws_edit,
            formula_map,
            row_idx: int | None,
            col_idx: int,
            compare_row: int | None = None,
        ):
            if ws_val is None or row_idx is None or row_idx <= 0:
                return None, _merge_cell_compare_key(None, None)
            try:
                vv = ws_val.cell(row=row_idx, column=col_idx).value
            except Exception:
                vv = None
            missing_formula = object()
            ve = formula_map.get((int(row_idx), int(col_idx)), missing_formula)
            if ve is missing_formula:
                ve = None
                if vv is None:
                    try:
                        if ws_edit is not None:
                            ve = ws_edit.cell(row=row_idx, column=col_idx).value
                    except Exception:
                        ve = None
            if compare_row is not None and row_idx != compare_row:
                ve = _translate_normal_formula_for_compare(
                    vv, ve, row_idx, col_idx, compare_row, col_idx
                )
            cmp_v = vv
            try:
                if cmp_v is None:
                    if ve is not None and not _formula_text(ve):
                        cmp_v = ve
            except Exception:
                pass
            return cmp_v, _merge_cell_compare_key(vv, ve)

        for name in common:
            ws_b = wb_base_val[name] if name in set_base else None
            ws_m = wb_mine_val[name]
            ws_t = wb_theirs_val[name]
            ws_m_e = wb_mine_edit[name]
            ws_b_e = wb_base_edit[name] if name in wb_base_edit.sheetnames else None
            ws_t_e = wb_theirs_edit[name] if name in wb_theirs_edit.sheetnames else None
            formula_map_m = _formula_edit_value_map(ws_m_e)
            formula_map_b = _formula_edit_value_map(ws_b_e)
            formula_map_t = _formula_edit_value_map(ws_t_e)

            max_col = max(ws_m.max_column or 1, ws_t.max_column or 1, (ws_b.max_column or 1) if ws_b else 1)
            mine_sigs = _row_sig_list_for_ws(ws_m, ws_m.max_row or 1, max_col)
            theirs_sigs = _row_sig_list_for_ws(ws_t, ws_t.max_row or 1, max_col)
            base_sigs = _row_sig_list_for_ws(ws_b, ws_b.max_row or 1, max_col) if ws_b is not None else []
            mine_to_base = _row_map_from_pairs(
                _compute_row_pairs_from_signatures(mine_sigs, base_sigs)
            ) if ws_b is not None else {}
            theirs_to_base = _row_map_from_pairs(
                _compute_row_pairs_from_signatures(theirs_sigs, base_sigs)
            ) if ws_b is not None else {}
            display_pairs = _compute_row_pairs_from_signatures(mine_sigs, theirs_sigs)
            if ws_b is not None:
                display_pairs = _split_tail_independent_append_pairs(
                    display_pairs, mine_to_base, theirs_to_base,
                    ws_m, ws_t, max_col,
                )
                display_pairs = _split_low_similarity_tail_pairs(
                    display_pairs,
                    mine_to_base,
                    theirs_to_base,
                    ws_m,
                    ws_t,
                    max_col,
                )

            for ra, rb in display_pairs:
                base_row_m = mine_to_base.get(ra) if ra is not None else None
                base_row_t = theirs_to_base.get(rb) if rb is not None else None
                for c in range(1, max_col + 1):
                    compare_row = ra if ra is not None else rb
                    vm_cmp, vm_key = _cmp_cell(ws_m, ws_m_e, formula_map_m, ra, c, compare_row)
                    vt_cmp, vt_key = _cmp_cell(ws_t, ws_t_e, formula_map_t, rb, c, compare_row)
                    conflict_row = ra if ra is not None else (rb if rb is not None else 0)

                    if ra is not None and rb is not None and base_row_m is not None and base_row_m == base_row_t:
                        vb_cmp, vb_key = _cmp_cell(ws_b, ws_b_e, formula_map_b, base_row_m, c, compare_row)
                        if vb_cmp is None and vt_cmp is not None and ws_b_e is not None and ws_t_e is not None:
                            try:
                                if _same_formula(ws_b_e.cell(row=base_row_m, column=c).value, ws_t_e.cell(row=rb, column=c).value):
                                    vb_cmp = vt_cmp
                                    vb_key = vt_key
                            except Exception:
                                pass
                        mine_changed = (vm_key != vb_key)
                        theirs_changed = (vt_key != vb_key)
                        if mine_changed and theirs_changed and vm_key != vt_key:
                            conflicts.append((name, conflict_row, c, vm_cmp, vt_cmp))
                            conflict_cells_by_sheet.setdefault(name, {}).setdefault(conflict_row, set()).add(c)
                        continue

                    if ra is not None and rb is not None and base_row_m is None and base_row_t is None:
                        if vm_key != vt_key:
                            conflicts.append((name, conflict_row, c, vm_cmp, vt_cmp))
                            conflict_cells_by_sheet.setdefault(name, {}).setdefault(conflict_row, set()).add(c)

        return conflicts, conflict_cells_by_sheet
    finally:
        _wbs_close(wb_base_val, wb_mine_val, wb_theirs_val, wb_mine_edit, wb_base_edit, wb_theirs_edit)
        for _orig, _vp in ((base_path, base_val_path), (mine_path, mine_val_path), (theirs_path, theirs_val_path)):
            if _vp != _orig:
                try:
                    os.remove(_vp)
                except Exception:
                    pass


class SheetView:
    """TortoiseMerge-like side-by-side full-sheet viewer.

    Performance notes (optimized for responsiveness):
    - Avoids O(N) tag_remove across the whole document on every click.
    - Avoids per-cell ws.cell access loops during normal interactions.
    - Keeps per-row cached text and per-row diff columns; row merge refreshes only that row.
    """

    def __init__(self, parent, app, sheet_name: str):
        self.parent = parent
        self.app = app
        self.root = getattr(app, "root", None)
        self.sheet = sheet_name
        # Support lazy tab containers: if parent is already a tab frame, reuse it.
        if isinstance(parent, ttk.Frame) and not parent.winfo_children():
            self.frame = parent
        else:
            self.frame = ttk.Frame(parent)

        self.max_row = 1
        self.max_col = 1
        self.col_max_a = 1  # column count of side A (may differ from B)
        self.col_max_b = 1  # column count of side B
        self._bounds_checked = False
        # Per-column max display width (chars), computed during diff scan
        self.col_char_widths: dict[int, int] = {}
        self._rownum_display_width: int = 3  # right-justified row number gutter width
        self._row_header_width: int = 4

        # Cached row text and diff cols
        self.row_text_a: dict[int, str] = {}
        self.row_text_b: dict[int, str] = {}
        self.diff_cols_by_row: dict[int, set[int]] = {}
        self._display_diff_row_count: int = 0
        self._sample_scan_started = False
        # Row alignment (pair-wise) caches
        self.row_pairs: list[tuple[int | None, int | None]] = []
        self.pair_text_a: dict[int, str] = {}
        self.pair_text_b: dict[int, str] = {}
        self.pair_diff_cols: dict[int, set[int]] = {}
        self.pair_base_diff_cols: dict[int, set[int]] = {}
        self.row_a_to_pair_idx: dict[int, int] = {}
        self.row_b_to_pair_idx: dict[int, int] = {}
        self.mine_to_base_row: dict[int, int] = {}
        self.theirs_to_base_row: dict[int, int] = {}
        self.pair_base_row_override: dict[int, int | None] = {}
        self._missing_base_row_map: dict[int, int] = {}

        # Render state
        # display_rows stores pair indices (into row_pairs)
        self.display_rows: list[int] = []
        self._full_display_rows: list[int] = []
        self._render_limit: int = _FAST_RENDER_ROW_LIMIT
        self.row_to_line: dict[int, int] = {}
        self._pending_yview: float | None = None
        self._render_cache = {}
        self._data_version = 0
        # Minimap (diff map) throttle: cache diff rows/cols keyed by data version,
        # debounce redraws triggered by scrolling.
        self._diff_map_cache_version = None
        self._diff_map_cache = None
        self._diff_map_debounce_id = None
        self._diff_map_debounce_ms = 40
        self.selected_excel_row: int | None = None
        self.selected_excel_row_a: int | None = None
        self.selected_excel_row_b: int | None = None
        self.selected_pair_idx: int | None = None
        self._last_selected_line: int | None = None
        self._main_sel_col: int | None = None
        self._main_sel_line: int | None = None
        self._applied_main_sel_col: int | None = None
        self._applied_main_sel_line: int | None = None
        self._cursor_cmp_sel_col: int | None = None
        self._cursor_cmp_sel_line: int | None = None
        self.hover_pair_idx: int | None = None
        self.hover_col_idx: int | None = None
        self.hover_side: str | None = None
        self._last_cursor_cmp_pair_idx: int | None = None
        self._trace_click_until: float = 0.0
        self._click_trace_seq: int = 0
        self._is_large_sheet = False
        self._prefer_only_diff_when_ready = False
        self._pending_only_diff_value: int | None = None
        self._diff_partial = False
        self._align_rows_enabled = True
        self._force_sequence_align = False
        # After user-triggered rescan/toggle, ignore late background cache apply for this sheet
        # to avoid delayed stale overwrite (rows unexpectedly disappear a few seconds later).
        self._suppress_bg_apply = False
        self._formula_copy_skips_pending = 0
        # Set to True once initial diff data has been computed (background or manual).
        # Prevents refresh(rescan=False) from triggering a full rescan on empty initial state.
        self._data_ready = False

        # Rows that were modified via overwrite in this session.
        # In "只看差异" mode, we keep these rows visible even if diffs are resolved.
        self.touched_rows: set[int] = set()

        # Snapshot mode: build the diff row list once, then keep the row list stable.
        # Overwrites only update per-row highlight (to show "已处理") and keep the row visible.
        self.snapshot_only_diff = True
        self._only_diff_source_version = 0
        self._only_diff_rows_cache: list[int] | None = None
        self._only_diff_rows_cache_key: tuple | None = None
        self._only_diff_async_build_key: tuple | None = None
        self._only_diff_async_build_seq = 0
        self._only_diff_async_building = False
        self._only_diff_async_thread: threading.Thread | None = None

        # Toolbar
        bar = ttk.Frame(self.frame)
        bar.pack(fill="x", padx=8, pady=(8, 6))

        ttk.Label(bar, text=f"Sheet: {sheet_name}", font=("Segoe UI", 11, "bold")).pack(side="left")
        self.info = ttk.Label(bar, text="", foreground="#444")
        self.info.pack(side="left", padx=(10, 0))
        self.loading_progress = ttk.Progressbar(bar, mode="indeterminate", length=120)

        # Diff block navigation (fixed position on the right; does not shift with label lengths)
        self.next_diff_btn = tk.Button(bar, text="下一处差异", padx=10, pady=2, command=self._goto_next_diff_block)
        self.prev_diff_btn = tk.Button(bar, text="上一处差异", padx=10, pady=2, command=self._goto_prev_diff_block)
        # Pack on right so it stays at a stable location across sheets
        self.next_diff_btn.pack(side="right", padx=(6, 0))
        self.prev_diff_btn.pack(side="right", padx=(6, 0))
        self._diff_blocks_cache = None  # None = not yet computed; [] = computed, no diffs

        # Some environments fail to toggle BooleanVar reliably; use IntVar with explicit on/off values.
        self.only_diff_var = tk.IntVar(value=int(getattr(self.app, "only_diff_default", 0)))

        self.only_diff_cb = tk.Checkbutton(
            bar,
            text="只看差异内容",
            variable=self.only_diff_var,
            onvalue=1,
            offvalue=0,
            command=self._toggle_only_diff,
            padx=6,
        )
        # Put on the right for a stable position
        self.only_diff_cb.pack(side="right", padx=(6, 0))
        self.force_align_var = tk.IntVar(value=0)
        self.force_align_cb = tk.Checkbutton(
            bar,
            text="强制行对齐(SM)",
            variable=self.force_align_var,
            onvalue=1,
            offvalue=0,
            command=self._toggle_force_align,
            padx=6,
        )
        self.force_align_cb.pack(side="right", padx=(6, 0))
        self.grid_overlay_var = tk.IntVar(value=1)
        self.grid_overlay_cb = tk.Checkbutton(
            bar,
            text="网格显示",
            variable=self.grid_overlay_var,
            onvalue=1,
            offvalue=0,
            command=self._toggle_grid_overlay,
            padx=6,
        )
        self.grid_overlay_cb.pack(side="right", padx=(6, 0))
        if getattr(self.app, "merge_conflict_mode", False):
            try:
                self.only_diff_var.set(1)
                self.only_diff_cb.select()
                self.only_diff_cb.configure(state="disabled")
                self.force_align_var.set(0)
                self.force_align_cb.configure(state="disabled")
            except Exception:
                pass
        self.three_way_var = tk.IntVar(value=1 if getattr(self.app, "merge_mode", False) and getattr(self.app, "has_base", False) else 0)
        if getattr(self.app, "merge_mode", False) and getattr(self.app, "has_base", False):
            tk.Checkbutton(
                bar,
                text="3视图",
                variable=self.three_way_var,
                onvalue=1,
                offvalue=0,
                command=self._toggle_three_way_view,
                padx=6,
            ).pack(side="right", padx=(6, 0))

        # Apply initial visual state from persisted setting
        try:
            if self.only_diff_var.get():
                self.only_diff_cb.select()
            else:
                self.only_diff_cb.deselect()
        except Exception:
            pass
        self._last_only_diff_value = int(self.only_diff_var.get())

        # Debug: provide a force-toggle button to prove the filtering path works even if UI toggling fails.
        if _DEBUG_ENABLED:
            tk.Button(
                bar,
                text="强制切换",
                command=lambda: (self.only_diff_var.set(0 if self.only_diff_var.get() else 1), self._toggle_only_diff()),
                padx=6,
                pady=1,
            ).pack(side="right", padx=(6, 0))

        # Debug: log click + resulting value
        def _log_cb_click(_evt=None):
            _dlog(f"CHECKBOX_CLICK sheet={self.sheet} var={self.only_diff_var.get()}")
            try:
                self.frame.after_idle(lambda: _dlog(f"CHECKBOX_AFTER_IDLE sheet={self.sheet} var={self.only_diff_var.get()}"))
            except Exception:
                pass

        try:
            self.only_diff_cb.bind("<ButtonRelease-1>", _log_cb_click)
        except Exception:
            pass
        # Context merge controls (row-level + region-level)
        # 区域 = 连续的差异行块；以当前鼠标所在行为锚点定位该块。
        # Office-like split button:
        # - Main button executes current mode action.
        # - Arrow menu switches mode only (no overwrite on switch).
        left_row_dir = "BASE2A" if (getattr(self.app, "merge_mode", False) and getattr(self.app, "has_base", False)) else "A2B"
        right_row_dir = "B2A"
        self._left_copy_direction = left_row_dir
        self._right_copy_direction = right_row_dir
        self._copy_scope_mode = "row"
        self._copy_scope_var = tk.StringVar(value="row")

        def _build_split_group(parent, main_text: str, bg: str, command):
            group = tk.Frame(parent, bg=bg, bd=1, relief="solid")
            main_btn = tk.Button(
                group,
                text=main_text,
                bg=bg,
                activebackground=bg,
                padx=10,
                pady=2,
                relief="flat",
                bd=0,
                command=command,
            )
            sep = tk.Frame(group, bg="#9aa7b0", width=1)
            menu_btn = tk.Menubutton(
                group,
                text="▾",
                bg=bg,
                activebackground=bg,
                padx=6,
                pady=2,
                relief="flat",
                bd=0,
                indicatoron=False,
                direction="below",
            )
            main_btn.pack(side="left")
            sep.pack(side="left", fill="y", pady=2)
            menu_btn.pack(side="left")
            return group, main_btn, menu_btn

        self.use_left_group, self.use_left_btn, self.use_left_menu_btn = _build_split_group(
            bar,
            "使用左侧行",
            "#eaf2ff",
            command=lambda: self._run_copy_action_by_mode(self._left_copy_direction),
        )
        self._use_left_menu = tk.Menu(self.use_left_menu_btn, tearoff=0)
        self._use_left_menu.add_radiobutton(
            label="行模式",
            variable=self._copy_scope_var,
            value="row",
            command=lambda: self._set_copy_scope_mode("row"),
        )
        self._use_left_menu.add_radiobutton(
            label="区域模式",
            variable=self._copy_scope_var,
            value="region",
            command=lambda: self._set_copy_scope_mode("region"),
        )
        self.use_left_menu_btn.configure(menu=self._use_left_menu)

        self.use_right_group, self.use_right_btn, self.use_right_menu_btn = _build_split_group(
            bar,
            "使用右侧行",
            "#ffecec",
            command=lambda: self._run_copy_action_by_mode(self._right_copy_direction),
        )
        self._use_right_menu = tk.Menu(self.use_right_menu_btn, tearoff=0)
        self._use_right_menu.add_radiobutton(
            label="行模式",
            variable=self._copy_scope_var,
            value="row",
            command=lambda: self._set_copy_scope_mode("row"),
        )
        self._use_right_menu.add_radiobutton(
            label="区域模式",
            variable=self._copy_scope_var,
            value="region",
            command=lambda: self._set_copy_scope_mode("region"),
        )
        self.use_right_menu_btn.configure(menu=self._use_right_menu)
        self._set_copy_scope_mode("row")

        self.use_base_btn = None
        if getattr(self.app, "merge_mode", False) and getattr(self.app, "has_base", False):
            self.use_base_btn = tk.Button(
                bar,
                text="保留Mine",
                bg="#f3f3ff",
                padx=10,
                pady=2,
                command=lambda: self._copy_selected_row("MINE2A"),
            )
        self.undo_btn = tk.Button(
            bar,
            text="回退",
            bg="#f2f2f2",
            padx=8,
            pady=2,
            command=self._undo_last_action,
        )
        # Keep at top-right (avoid misclick)
        self.use_right_group.pack(side="right", padx=(6, 0))
        if self.use_base_btn is not None:
            self.use_base_btn.pack(side="right", padx=(6, 0))
        self.use_left_group.pack(side="right")
        self.undo_btn.pack(side="right", padx=(6, 0))

        ttk.Button(bar, text="刷新本Sheet", command=self._manual_rescan).pack(side="right", padx=(6, 0))
        self._full_render = False
        self._load_all_btn = ttk.Button(bar, text="加载全部", command=self._load_all_rows)
        if _FAST_OPEN_ENABLED:
            self._load_all_btn.pack(side="right", padx=(6, 0))

        # Path bar (requested red-box area): show full paths above the diff panes
        path_bar = ttk.Frame(self.frame)
        path_bar.pack(fill="x", padx=8, pady=(0, 4))

        self._path_font = ("Segoe UI", 9)
        path_bar.grid_columnconfigure(0, weight=1)
        path_bar.grid_columnconfigure(1, weight=1)
        path_bar.grid_columnconfigure(2, weight=1)

        self._path_font = ("Segoe UI", 9, "bold")

        def _one_line_text(s: str, max_len: int = 120) -> str:
            s = (s or "").replace("\r", " ").replace("\n", " ")
            if len(s) <= max_len:
                return s
            # keep file tail visible when path is long
            return "..." + s[-(max_len - 3):]

        if getattr(self.app, "merge_mode", False) and getattr(self.app, "has_base", False):
            mine_src = getattr(self.app, "raw_mine", None) or self.app.file_a
            base_src = getattr(self.app, "raw_base", None) or getattr(self.app, "base_path", "")
            theirs_src = getattr(self.app, "raw_theirs", None) or self.app.file_b
            label_a = f"mine={self._source_display_name(mine_src)}"
            label_base = f"base={self._source_display_name(base_src)}" if base_src else "base=-"
            label_b = f"theirs={self._source_display_name(theirs_src)}"
        else:
            # Diff mode: keep wording consistent with SVN semantics (left=base, right=mine).
            base_src = getattr(self.app, "raw_base", None) or self.app.file_a
            mine_src = getattr(self.app, "raw_mine", None) or self.app.file_b
            base_disp = _one_line_text(str(base_src or "")) or "-"
            mine_disp = _one_line_text(str(mine_src or "")) or "-"
            label_a = f"base={base_disp}"
            label_b = f"mine={mine_disp}"
            label_base = _one_line_text(getattr(self.app, "base_path", "") or "")
        is_merge_labels = bool(getattr(self.app, "merge_mode", False) and getattr(self.app, "has_base", False))
        path_bg_a = _MINE_BG if is_merge_labels else _BASE_BG
        path_bg_b = _THEIRS_BG if is_merge_labels else _MINE_BG

        self.path_label_a = tk.Label(
            path_bar,
            text=label_a,
            font=self._path_font,
            bg=path_bg_a,
            anchor="w",
            padx=6,
            pady=2,
        )
        self.path_label_a.grid(row=0, column=0, sticky="ew")
        self.path_label_base = tk.Label(
            path_bar,
            text=label_base if label_base else "基础(base): -",
            font=self._path_font,
            bg=_BASE_BG,
            anchor="w",
            padx=6,
            pady=2,
        )
        self.path_label_base.grid(row=0, column=1, sticky="ew")
        self.path_label_b = tk.Label(
            path_bar,
            text=label_b,
            font=self._path_font,
            bg=path_bg_b,
            anchor="w",
            padx=6,
            pady=2,
        )
        self.path_label_b.grid(row=0, column=2, sticky="ew")

        # Extra vertical scrollbar (left side) for convenience; controls both panes.
        # NOTE: must be packed BEFORE the paned window so it remains visible.
        self.vsb_left = ttk.Scrollbar(self.frame, orient="vertical", command=self._yview_both)
        self.vsb_left.pack(side="left", fill="y")
        # Diff minimap next to the left vertical scrollbar (more discoverable).
        self.vdiff_map = tk.Canvas(self.frame, width=10, highlightthickness=0, bg="#ebebeb")
        self.vdiff_map.pack(side="left", fill="y", padx=(0, 2))
        self.vdiff_map.bind("<Button-1>", self._on_vdiff_map_click)

        # Reserve the lower compare area before the main panes so C区 and hover
        # panel do not get squeezed to 1px on first layout.
        self.lower_area = ttk.Frame(self.frame)
        self.lower_area.pack(side="bottom", fill="x")

        # Panes
        paned = ttk.PanedWindow(self.frame, orient="horizontal")
        paned.pack(fill="both", expand=True, padx=8, pady=(0, 8))
        self._main_paned = paned

        left_wrap = ttk.Frame(paned)
        mid_wrap = ttk.Frame(paned)
        right_wrap = ttk.Frame(paned)
        self._left_wrap = left_wrap
        self._mid_wrap = mid_wrap
        self._right_wrap = right_wrap
        paned.add(mid_wrap, weight=1)
        paned.add(left_wrap, weight=1)
        paned.add(right_wrap, weight=1)

        def _keep_panes_equal(_evt=None):
            # Keep A/B content panes at 50:50 to avoid visual width mismatch.
            try:
                total = self._main_paned.winfo_width()
                if total and total > 2:
                    if self._is_three_way_enabled():
                        self._main_paned.sashpos(0, total // 3)
                        self._main_paned.sashpos(1, (total * 2) // 3)
                    else:
                        self._main_paned.sashpos(0, total // 2)
            except Exception:
                pass

        self._keep_panes_equal = _keep_panes_equal
        self._main_paned.bind("<Configure>", self._keep_panes_equal)
        self._main_paned.bind("<ButtonRelease-1>", self._keep_panes_equal)
        self.frame.after(0, self._keep_panes_equal)

        self.left_title = ttk.Label(left_wrap, text="Mine", background=_MINE_BG)
        self.left_title.pack(fill="x")
        self.mid_title = ttk.Label(mid_wrap, text="Base", background=_BASE_BG)
        self.mid_title.pack(fill="x")
        self.right_title = ttk.Label(right_wrap, text="Base", background=_THEIRS_BG)
        self.right_title.pack(fill="x")

        # Font size tuned closer to TortoiseMerge (+~20%)
        self.editor_font = ("Consolas", 11)
        left_body = ttk.Frame(left_wrap)
        left_body.pack(fill="both", expand=True)
        mid_body = ttk.Frame(mid_wrap)
        mid_body.pack(fill="both", expand=True)
        right_body = ttk.Frame(right_wrap)
        right_body.pack(fill="both", expand=True)
        left_footer = ttk.Frame(left_wrap)
        left_footer.pack(side="bottom", fill="x", before=left_body)
        mid_footer = ttk.Frame(mid_wrap)
        mid_footer.pack(side="bottom", fill="x", before=mid_body)
        right_footer = ttk.Frame(right_wrap)
        right_footer.pack(side="bottom", fill="x", before=right_body)

        self.left_ln = tk.Text(left_body, width=self._row_header_width, wrap="none", undo=False, font=self.editor_font, bg="#efefef", fg="#666666", relief="flat", takefocus=0, cursor="arrow")
        self.base_ln = tk.Text(mid_body, width=self._row_header_width, wrap="none", undo=False, font=self.editor_font, bg="#efefef", fg="#666666", relief="flat", takefocus=0, cursor="arrow")
        self.right_ln = tk.Text(right_body, width=self._row_header_width, wrap="none", undo=False, font=self.editor_font, bg="#efefef", fg="#666666", relief="flat", takefocus=0, cursor="arrow")
        self.left = tk.Text(left_body, wrap="none", undo=False, font=self.editor_font, bg="white")
        self.base = tk.Text(mid_body, wrap="none", undo=False, font=self.editor_font, bg="white")
        self.right = tk.Text(right_body, wrap="none", undo=False, font=self.editor_font, bg="white")

        # Scrollbars
        # Per-pane vertical scrollbars (user requested visible scrollbars on both A and B)
        self.vsb_a = ttk.Scrollbar(left_body, orient="vertical", command=self._yview_both)
        self.vsb_m = ttk.Scrollbar(mid_body, orient="vertical", command=self._yview_both)
        self.vsb_b = ttk.Scrollbar(right_body, orient="vertical", command=self._yview_both)
        self.vsb_a.pack(side="right", fill="y")
        self.vsb_m.pack(side="right", fill="y")
        self.vsb_b.pack(side="right", fill="y")

        # Shared vertical scrollbar on far right removed (redundant / low usability).
        self.vsb = None

        # Horizontal scroll sync: keep A/B panes aligned when scrolling horizontally.
        self._xsyncing = False
        # Suppress C-pane -> main-pane x-sync during programmatic C text refresh.
        self._suppress_c_xsync = False

        def _sync_main_headers(first):
            try:
                frac = float(first)
            except Exception:
                return
            try:
                if getattr(self, "left_colhdr", None) is not None:
                    self.left_colhdr.xview_moveto(frac)
                if getattr(self, "right_colhdr", None) is not None:
                    self.right_colhdr.xview_moveto(frac)
                if self._is_three_way_enabled() and getattr(self, "base_colhdr", None) is not None:
                    self.base_colhdr.xview_moveto(frac)
            except Exception:
                pass

        def _xscroll_left(first, last):
            # called when left xview changes
            if self._is_click_trace_active():
                _dlog(f"xscroll_left first={first} last={last} xsyncing={self._xsyncing}")
            if self._xsyncing:
                self.hsb_left.set(first, last)
                return
            self._xsyncing = True
            try:
                self.hsb_left.set(first, last)
                self._sync_main_x_from_widget(self.left, first)
                self._sync_c_x_from_widget(self.left, first)
            finally:
                self._xsyncing = False
            try:
                self._schedule_diff_maps()
            except Exception:
                pass

        def _xscroll_right(first, last):
            if self._is_click_trace_active():
                _dlog(f"xscroll_right first={first} last={last} xsyncing={self._xsyncing}")
            if self._xsyncing:
                self.hsb_right.set(first, last)
                return
            self._xsyncing = True
            try:
                self.hsb_right.set(first, last)
                self._sync_main_x_from_widget(self.right, first)
                self._sync_c_x_from_widget(self.right, first)
            finally:
                self._xsyncing = False
            try:
                self._schedule_diff_maps()
            except Exception:
                pass

        def _xscroll_mid(first, last):
            if self._is_click_trace_active():
                _dlog(f"xscroll_mid first={first} last={last} xsyncing={self._xsyncing}")
            if self._xsyncing:
                self.hsb_mid.set(first, last)
                return
            self._xsyncing = True
            try:
                self.hsb_mid.set(first, last)
                self._sync_main_x_from_widget(self.base, first)
                self._sync_c_x_from_widget(self.base, first)
            finally:
                self._xsyncing = False
            try:
                self._schedule_diff_maps()
            except Exception:
                pass

        def _xview_left(*args):
            # scrollbar drag/click on left
            self._xsyncing = True
            try:
                self.left.xview(*args)
                first, last = self.left.xview()
                self.hsb_left.set(first, last)
                self._sync_main_x_from_widget(self.left, first)
                self._sync_c_x_from_widget(self.left, first)
            finally:
                self._xsyncing = False
            try:
                self._schedule_diff_maps()
            except Exception:
                pass

        def _xview_right(*args):
            self._xsyncing = True
            try:
                self.right.xview(*args)
                first, last = self.right.xview()
                self.hsb_right.set(first, last)
                self._sync_main_x_from_widget(self.right, first)
                self._sync_c_x_from_widget(self.right, first)
            finally:
                self._xsyncing = False
            try:
                self._schedule_diff_maps()
            except Exception:
                pass

        def _xview_mid(*args):
            self._xsyncing = True
            try:
                self.base.xview(*args)
                first, last = self.base.xview()
                self.hsb_mid.set(first, last)
                self._sync_main_x_from_widget(self.base, first)
                self._sync_c_x_from_widget(self.base, first)
            finally:
                self._xsyncing = False
            try:
                self._schedule_diff_maps()
            except Exception:
                pass

        self.hsb_left = ttk.Scrollbar(left_footer, orient="horizontal", command=_xview_left)
        self.hsb_mid = ttk.Scrollbar(mid_footer, orient="horizontal", command=_xview_mid)
        self.hsb_right = ttk.Scrollbar(right_footer, orient="horizontal", command=_xview_right)
        self.hdiff_left = tk.Canvas(left_footer, height=10, highlightthickness=0, bg="#ebebeb")
        self.hdiff_mid = tk.Canvas(mid_footer, height=10, highlightthickness=0, bg="#ebebeb")
        self.hdiff_right = tk.Canvas(right_footer, height=10, highlightthickness=0, bg="#ebebeb")
        self.hdiff_left.bind("<Button-1>", lambda e: self._on_hdiff_map_click(e, "left"))
        self.hdiff_mid.bind("<Button-1>", lambda e: self._on_hdiff_map_click(e, "mid"))
        self.hdiff_right.bind("<Button-1>", lambda e: self._on_hdiff_map_click(e, "right"))
        self.left.configure(xscrollcommand=_xscroll_left)
        self.base.configure(xscrollcommand=_xscroll_mid)
        self.right.configure(xscrollcommand=_xscroll_right)

        self.left.configure(yscrollcommand=self._yscroll_left)
        self.base.configure(yscrollcommand=self._yscroll_mid)
        self.right.configure(yscrollcommand=self._yscroll_right)
        self.vsb_left.configure(command=self._yview_both)
        self.vsb_a.configure(command=self._yview_both)
        self.vsb_m.configure(command=self._yview_both)
        self.vsb_b.configure(command=self._yview_both)

        # Excel-like column header row for A/Base/B panes.
        left_head = ttk.Frame(left_wrap)
        left_head.pack(fill="x", before=left_body)
        mid_head = ttk.Frame(mid_wrap)
        mid_head.pack(fill="x", before=mid_body)
        right_head = ttk.Frame(right_wrap)
        right_head.pack(fill="x", before=right_body)

        self.left_corner_hdr = tk.Text(left_head, width=self._row_header_width, height=1, wrap="none", undo=False,
                                       font=self.editor_font, bg="#efefef", fg="#666666", relief="flat", takefocus=0, cursor="arrow")
        self.base_corner_hdr = tk.Text(mid_head, width=self._row_header_width, height=1, wrap="none", undo=False,
                                       font=self.editor_font, bg="#efefef", fg="#666666", relief="flat", takefocus=0, cursor="arrow")
        self.right_corner_hdr = tk.Text(right_head, width=self._row_header_width, height=1, wrap="none", undo=False,
                                        font=self.editor_font, bg="#efefef", fg="#666666", relief="flat", takefocus=0, cursor="arrow")
        self.left_colhdr = tk.Text(left_head, height=1, wrap="none", undo=False,
                                   font=self.editor_font, bg="#efefef", fg="#555555", relief="flat", takefocus=0, cursor="arrow")
        self.base_colhdr = tk.Text(mid_head, height=1, wrap="none", undo=False,
                                   font=self.editor_font, bg="#efefef", fg="#555555", relief="flat", takefocus=0, cursor="arrow")
        self.right_colhdr = tk.Text(right_head, height=1, wrap="none", undo=False,
                                    font=self.editor_font, bg="#efefef", fg="#555555", relief="flat", takefocus=0, cursor="arrow")

        self.left_corner_hdr.pack(side="left", fill="y")
        ttk.Separator(left_head, orient="vertical").pack(side="left", fill="y")
        self._vsb_hdr_spc_a = tk.Frame(left_head, width=1, bg="#efefef")
        self._vsb_hdr_spc_a.pack(side="right", fill="y")
        self.left_colhdr.pack(side="left", fill="x", expand=True)
        self.base_corner_hdr.pack(side="left", fill="y")
        ttk.Separator(mid_head, orient="vertical").pack(side="left", fill="y")
        self._vsb_hdr_spc_m = tk.Frame(mid_head, width=1, bg="#efefef")
        self._vsb_hdr_spc_m.pack(side="right", fill="y")
        self.base_colhdr.pack(side="left", fill="x", expand=True)
        self.right_corner_hdr.pack(side="left", fill="y")
        ttk.Separator(right_head, orient="vertical").pack(side="left", fill="y")
        self._vsb_hdr_spc_b = tk.Frame(right_head, width=1, bg="#efefef")
        self._vsb_hdr_spc_b.pack(side="right", fill="y")
        self.right_colhdr.pack(side="left", fill="x", expand=True)
        for w in (self.left_corner_hdr, self.base_corner_hdr, self.right_corner_hdr,
                  self.left_colhdr, self.base_colhdr, self.right_colhdr):
            try:
                w.configure(state="normal")
                w.delete("1.0", "end")
                w.insert("1.0", "")
                w.configure(state="disabled")
            except Exception:
                pass
        # Keep colhdr spacers in sync with the actual scrollbar width so that column
        # headers align precisely with the data widgets at all times.
        def _sync_vsb_hdr_spacers(event=None):
            try:
                w = self.vsb_a.winfo_width()
                if w > 1:
                    self._vsb_hdr_spc_a.configure(width=w)
                    self._vsb_hdr_spc_b.configure(width=w)
                    self._vsb_hdr_spc_m.configure(width=w)
            except Exception:
                pass
        self.vsb_a.bind("<Configure>", _sync_vsb_hdr_spacers, "+")

        self.left_ln.pack(side="left", fill="y")
        ttk.Separator(left_body, orient="vertical").pack(side="left", fill="y")
        self.left.pack(fill="both", expand=True)
        self.hsb_left.pack(fill="x")
        self.hdiff_left.pack(fill="x")
        self.base_ln.pack(side="left", fill="y")
        ttk.Separator(mid_body, orient="vertical").pack(side="left", fill="y")
        self.base.pack(fill="both", expand=True)
        self.hsb_mid.pack(fill="x")
        self.hdiff_mid.pack(fill="x")

        # Save action row: keep a fixed height on both sides so horizontal
        # scrollbars stay aligned even when only one side has a button.
        # Also keep middle pane height identical to left/right to avoid row misalignment.
        save_row_height = 34

        # Save A button (bottom-right of A pane)
        save_a_row = ttk.Frame(left_footer, height=save_row_height)
        save_a_row.pack(fill="x", pady=(2, 0))
        save_a_row.pack_propagate(False)
        if getattr(self.app, "merge_mode", False):
            tk.Button(save_a_row, text="保存Merged并退出", bg="#eaf2ff", padx=14, pady=4,
                      command=self.app.save_merged_and_exit).pack(side="right")
        else:
            if not getattr(self.app, "diff_base_mine_mode", False):
                if not _is_temp_base_path(getattr(self.app, "file_a", "")):
                    tk.Button(save_a_row, text="保存A", bg="#eaf2ff", padx=14, pady=4,
                              command=self.app.save_a_inplace).pack(side="right")

        # Base pane spacer: maintain same bottom reserved height as A/B panes.
        save_mid_row = ttk.Frame(mid_footer, height=save_row_height)
        save_mid_row.pack(fill="x", pady=(2, 0))
        save_mid_row.pack_propagate(False)

        self.right_ln.pack(side="left", fill="y")
        ttk.Separator(right_body, orient="vertical").pack(side="left", fill="y")
        self.right.pack(fill="both", expand=True)
        self.hsb_right.pack(fill="x")
        self.hdiff_right.pack(fill="x")

        # Save B button (bottom-right of B pane)
        save_b_row = ttk.Frame(right_footer, height=save_row_height)
        save_b_row.pack(fill="x", pady=(2, 0))
        save_b_row.pack_propagate(False)
        if not getattr(self.app, "merge_mode", False):
            save_b_text = "Save Mine" if getattr(self.app, "diff_base_mine_mode", False) else "Save B"
            tk.Button(save_b_row, text=save_b_text, bg="#ffecec", padx=14, pady=4,
                      command=self.app.save_b_inplace).pack(side="right")

        # Tags (order matters: diffcell should be applied after diffrow)
        # Closer to TortoiseMerge vibe: left diff block = orange, right diff block = yellow
        self.left.tag_configure("diffrow", background=_MINE_BG)
        self.base.tag_configure("diffrow", background=_BASE_BG)
        self.right.tag_configure("diffrow", background=_THEIRS_BG)

        # Cell-level highlight (red) for exact diffs
        self.left.tag_configure("diffcell", background=_DIFF_CELL_BG)
        self.base.tag_configure("diffcell", background=_DIFF_CELL_BG)
        self.right.tag_configure("diffcell", background=_DIFF_CELL_BG)
        self.left.tag_raise("diffcell")
        self.base.tag_raise("diffcell")
        self.right.tag_raise("diffcell")
        # Alignment padding: grey slot for rows that exist only on the other side.
        # tag_raise ensures paddingrow background overrides diffrow on the empty slot.
        self.left.tag_configure("paddingrow", background="#A0A0A0")
        self.base.tag_configure("paddingrow", background="#A0A0A0")
        self.right.tag_configure("paddingrow", background="#A0A0A0")
        self.left.tag_raise("paddingrow")
        self.base.tag_raise("paddingrow")
        self.right.tag_raise("paddingrow")
        # paddingcol: grey span for columns that exist only on the other side (新增列).
        self.left.tag_configure("paddingcol", background="#A0A0A0")
        self.right.tag_configure("paddingcol", background="#A0A0A0")
        self.left.tag_raise("paddingcol")
        self.right.tag_raise("paddingcol")

        # selection should not overwrite diff colors
        self.left.tag_configure("selrow", underline=1, font=("Consolas", 11, "bold"))
        self.base.tag_configure("selrow", underline=1, font=("Consolas", 11, "bold"))
        self.right.tag_configure("selrow", underline=1, font=("Consolas", 11, "bold"))
        # Selected-cell highlight (same blue as C区 cselcell), applied on A/B/(Base).
        self.left.tag_configure("selcell", background="#8EB9FF")
        self.base.tag_configure("selcell", background="#8EB9FF")
        self.right.tag_configure("selcell", background="#8EB9FF")
        self.left.tag_raise("selcell")
        self.base.tag_raise("selcell")
        self.right.tag_raise("selcell")
        self.left_ln.tag_configure("diffrow", background="#ffd9d9")
        self.base_ln.tag_configure("diffrow", background="#ffd9d9")
        self.right_ln.tag_configure("diffrow", background="#ffd9d9")

        # Bindings
        self._syncing = False
        for w in (self.left, self.base, self.right):
            w.bind("<MouseWheel>", self._on_mousewheel)
            w.bind("<Button-4>", self._on_mousewheel)
            w.bind("<Button-5>", self._on_mousewheel)
            w.bind("<KeyRelease>", lambda e: self._update_cursor_lines())
            w.bind("<ButtonRelease-1>", lambda e: self._update_cursor_lines())
            if getattr(self.app, "merge_conflict_mode", False):
                #快捷键：下一处/上一处冲突
                # F4 is reserved for hover panel pin/unpin toggle.
                w.bind("<F4>", self._on_hover_compare_f4_toggle)
                w.bind("<Shift-F4>", lambda e: (self._goto_prev_diff_block(), "break"))
                w.bind("<Control-n>", lambda e: (self._goto_next_diff_block(), "break"))
                w.bind("<Control-p>", lambda e: (self._goto_prev_diff_block(), "break"))

        # Click handling (selection + arrow action)
        left_click_dir = "MINE2A" if (getattr(self.app, "merge_mode", False) and getattr(self.app, "has_base", False)) else "A2B"
        self.left.bind("<Button-1>", lambda e, d=left_click_dir: self._on_click_with_arrow(self.left, e, d))
        self.base.bind("<Button-1>", lambda e: self._on_click_with_arrow(self.base, e, "BASE2A"))
        self.right.bind("<Button-1>", lambda e: self._on_click_with_arrow(self.right, e, "B2A"))
        self.left.bind("<Button-3>", lambda e: self._on_main_pane_right_click(self.left, e, "A"))
        self.base.bind("<Button-3>", lambda e: self._on_main_pane_right_click(self.base, e, "BASE"))
        self.right.bind("<Button-3>", lambda e: self._on_main_pane_right_click(self.right, e, "B"))
        self.left.bind("<Motion>", lambda e: self._on_cell_hover_tooltip(self.left, e, "A"))
        self.base.bind("<Motion>", lambda e: self._on_cell_hover_tooltip(self.base, e, "BASE"))
        self.right.bind("<Motion>", lambda e: self._on_cell_hover_tooltip(self.right, e, "B"))
        self.left.bind("<Leave>", lambda e: self._on_hover_compare_leave())
        self.base.bind("<Leave>", lambda e: self._on_hover_compare_leave())
        self.right.bind("<Leave>", lambda e: self._on_hover_compare_leave())
        # Double-click merge (single cell)
        self.left.bind("<Double-Button-1>", lambda e, d=left_click_dir: self._copy_cell(d, e))
        self.base.bind("<Double-Button-1>", lambda e: self._copy_cell("BASE2A", e))
        self.right.bind("<Double-Button-1>", lambda e: self._copy_cell("B2A", e))
        self.left_ln.bind("<Button-1>", lambda e, d=left_click_dir: self._on_row_header_click(self.left_ln, e, d))
        self.base_ln.bind("<Button-1>", lambda e: self._on_row_header_click(self.base_ln, e, "BASE2A"))
        self.right_ln.bind("<Button-1>", lambda e: self._on_row_header_click(self.right_ln, e, "B2A"))
        # Keep old habit compatibility: right-click on row header also triggers row apply.
        self.left_ln.bind("<Button-3>", lambda e, d=left_click_dir: self._on_row_header_click(self.left_ln, e, d))
        self.base_ln.bind("<Button-3>", lambda e: self._on_row_header_click(self.base_ln, e, "BASE2A"))
        self.right_ln.bind("<Button-3>", lambda e: self._on_row_header_click(self.right_ln, e, "B2A"))
        self.left_ln.bind("<Motion>", lambda e, d=left_click_dir: self._on_row_header_hover(self.left_ln, e, d))
        self.base_ln.bind("<Motion>", lambda e: self._on_row_header_hover(self.base_ln, e, "BASE2A"))
        self.right_ln.bind("<Motion>", lambda e: self._on_row_header_hover(self.right_ln, e, "B2A"))
        self.left_ln.bind("<Leave>", lambda e: self._clear_row_header_hover(self.left_ln))
        self.base_ln.bind("<Leave>", lambda e: self._clear_row_header_hover(self.base_ln))
        self.right_ln.bind("<Leave>", lambda e: self._clear_row_header_hover(self.right_ln))

        # C区: compact cursor compare block + cell-aligned view
        self.c_area = ttk.Notebook(self.lower_area)
        self.c_area.pack(fill="x", padx=8, pady=(0, 4))

        # ---- C1: compact row compare (2 lines in 2-way, 3 lines in 3-way) ----
        c_text_frame = ttk.Frame(self.c_area)
        self.c_area.add(c_text_frame, text="C区-行对比")

        # C区 column header + row-header layout (Excel-like)
        c_head = ttk.Frame(c_text_frame)
        c_head.pack(side="top", fill="x")
        self.cursor_cmp_corner = tk.Text(c_head, width=self._row_header_width, height=1, wrap="none", undo=False,
                                         font=self.editor_font, bg="#efefef", fg="#666666", relief="flat", takefocus=0, cursor="arrow")
        self.cursor_cmp_colhdr = tk.Text(c_head, height=1, wrap="none", undo=False,
                                         font=self.editor_font, bg="#efefef", fg="#555555", relief="flat", takefocus=0, cursor="arrow")
        self.cursor_cmp_corner.pack(side="left", fill="y")
        ttk.Separator(c_head, orient="vertical").pack(side="left", fill="y")
        self.cursor_cmp_colhdr.pack(side="left", fill="x", expand=True)

        c_body = ttk.Frame(c_text_frame)
        c_body.pack(side="top", fill="x")
        self.cursor_cmp_ln = tk.Text(c_body, width=self._row_header_width, height=3 if self._is_three_way_enabled() else 2, wrap="none", undo=False,
                                     font=self.editor_font, bg="#efefef", fg="#666666", relief="flat", takefocus=0, cursor="arrow")
        self.cursor_cmp_ln.pack(side="left", fill="y")
        ttk.Separator(c_body, orient="vertical").pack(side="left", fill="y")

        self.cursor_cmp = tk.Text(
            c_body,
            height=3 if self._is_three_way_enabled() else 2,
            wrap="none",
            font=self.editor_font,
            bd=1,
            relief="solid",
        )
        # Make base colors stronger (user feedback: previous too light)
        self.cursor_cmp.tag_configure("a", background=_MINE_BG)
        self.cursor_cmp.tag_configure("base", background=_BASE_BG)
        self.cursor_cmp.tag_configure("b", background=_THEIRS_BG)
        self.cursor_cmp.tag_configure("missing", background="#a6a6a6")
        # Diff cell highlight (match main panes)
        self.cursor_cmp.tag_configure("diffcell", background=_DIFF_CELL_BG)
        # Explicit C-area click target highlight
        self.cursor_cmp.tag_configure("cselcell", background="#8EB9FF")
        self.cursor_cmp.pack(side="left", fill="x", expand=True)

        for w in (self.cursor_cmp_corner, self.cursor_cmp_colhdr, self.cursor_cmp_ln):
            try:
                w.configure(state="normal")
                w.delete("1.0", "end")
                w.insert("1.0", "")
                w.configure(state="disabled")
            except Exception:
                pass

        # Horizontal scrollbar for C区行对比（sync with main panes）
        self.cursor_hsb = ttk.Scrollbar(c_text_frame, orient="horizontal", command=self._xview_cursor_cmp)
        self.cursor_cmp.configure(xscrollcommand=self._xscroll_cursor_cmp)
        self.cursor_hsb.pack(side="top", fill="x")
        self.cursor_cmp.bind("<Button-1>", self._on_cursor_cmp_click)
        self.cursor_cmp.bind("<Double-Button-1>", self._on_cursor_cmp_double_click)
        self.cursor_cmp.bind("<Button-3>", self._on_cursor_cmp_right_click)
        self.cursor_cmp.bind("<Motion>", self._on_cursor_cmp_hover_tooltip)
        self.cursor_cmp.bind("<Leave>", lambda e: self._on_hover_compare_leave())

        # ---- C2: cell-aligned view (optional; can be hidden if not useful/performance) ----
        self._enable_c_cell = False  # user feedback: not useful; keep hidden by default
        c_cell_frame = ttk.Frame(self.c_area)
        self.c_area.add(c_cell_frame, text="C区-单元格对齐")
        if not self._enable_c_cell:
            try:
                self.c_area.tab(c_cell_frame, state="hidden")
            except Exception:
                pass

        top_row = ttk.Frame(c_cell_frame)
        top_row.pack(fill="x", pady=(2, 2))
        self.c_only_diff_cells = tk.IntVar(value=1)
        tk.Checkbutton(
            top_row,
            text="只显示差异单元格",
            variable=self.c_only_diff_cells,
            onvalue=1,
            offvalue=0,
            command=lambda: self._update_cursor_lines(),
        ).pack(side="left")

        self.cell_cmp_text = tk.Text(c_cell_frame, height=6, wrap="none", font=self.editor_font, bd=1, relief="solid")
        self.cell_cmp_text.tag_configure("a", background=_MINE_BG)
        self.cell_cmp_text.tag_configure("b", background=_THEIRS_BG)
        self.cell_cmp_text.tag_configure("diffcell", background=_DIFF_CELL_BG)

        self.cell_cmp_hsb = ttk.Scrollbar(c_cell_frame, orient="horizontal", command=self._xview_cell_cmp)
        self.cell_cmp_text.configure(xscrollcommand=self._xscroll_cell_cmp)

        self.cell_cmp_text.pack(side="top", fill="x", expand=True)
        self.cell_cmp_hsb.pack(side="top", fill="x")

        # Stable hover compare panel (primary path; independent from popup windows).
        self._enable_hover_popup = False
        self._hover_clear_after_id = None
        self._last_hover_compare_key = None
        self._hover_payload_cache = {}
        # Hover throttle: dedup identical targets and debounce heavy panel refresh.
        self._last_hover_target_key = None
        self._hover_debounce_id = None
        self._pending_hover_args = None
        self._hover_debounce_ms = 30
        self.hover_cmp_host = ttk.Frame(self.lower_area, height=self._hover_compare_reserved_height())
        self.hover_cmp_host.pack(fill="x", padx=8, pady=(0, 4))
        try:
            self.hover_cmp_host.pack_propagate(False)
        except Exception:
            pass
        hover_cmp_frame = ttk.LabelFrame(self.hover_cmp_host, text="悬停完整对比")
        hover_cmp_frame.pack(fill="both", expand=True)
        self.hover_cmp_title_var = tk.StringVar(value="悬停完整对比：-")
        self.hover_cmp_pin_var = tk.IntVar(value=0)
        hover_hdr = ttk.Frame(hover_cmp_frame)
        hover_hdr.pack(fill="x", padx=4, pady=(2, 2))
        ttk.Label(hover_hdr, textvariable=self.hover_cmp_title_var).pack(side="left", anchor="w")
        ttk.Button(hover_hdr, text="清空", command=self._on_hover_compare_clear_click).pack(side="right", padx=(6, 0))
        ttk.Checkbutton(
            hover_hdr,
            text="固定(F4)",
            variable=self.hover_cmp_pin_var,
            onvalue=1,
            offvalue=0,
            command=self._on_hover_compare_pin_toggle,
        ).pack(side="right")
        self.hover_cmp_text = tk.Text(
            hover_cmp_frame,
            height=4 if self._is_three_way_enabled() else 2,
            wrap="none",
            font=self.editor_font,
            bd=1,
            relief="solid",
        )
        # Source row backgrounds align with main panes.
        self.hover_cmp_text.tag_configure("hover_side_base", background=_BASE_BG)
        self.hover_cmp_text.tag_configure("hover_side_mine", background=_MINE_BG)
        self.hover_cmp_text.tag_configure("hover_side_theirs", background=_THEIRS_BG)
        # Missing row (新增行对侧无数据): muted gray, no red highlights.
        self.hover_cmp_text.tag_configure("hover_side_missing", background="#C8C8C8", foreground="#666666")
        # Char-level diff highlight inside each source line.
        self.hover_cmp_text.tag_configure("hover_diffchar", background=_DIFF_CELL_BG, foreground="#ffffff")
        self.hover_cmp_hsb = ttk.Scrollbar(hover_cmp_frame, orient="horizontal", command=self.hover_cmp_text.xview)
        self.hover_cmp_text.configure(xscrollcommand=self.hover_cmp_hsb.set)
        self.hover_cmp_hsb.pack(side="bottom", fill="x")
        self.hover_cmp_text.pack(side="top", fill="both", expand=True)
        self.hover_cmp_text.bind("<Shift-MouseWheel>", self._on_hover_cmp_shift_wheel)
        self.hover_cmp_text.bind("<MouseWheel>", self._on_hover_cmp_mousewheel)
        self.hover_cmp_text.bind("<Shift-Button-4>", self._on_hover_cmp_shift_wheel)
        self.hover_cmp_text.bind("<Shift-Button-5>", self._on_hover_cmp_shift_wheel)
        try:
            self.hover_cmp_text.configure(state="disabled")
        except Exception:
            pass

        # initial render should respect the persisted only-diff setting
        # Defer heavy initial refresh; SowMergeApp will lazy-load the active sheet.
        # (Still create the UI widgets now.)
        # self.refresh(row_only=None, rescan=True)
        # self._update_cursor_lines()
        # Initial panel state (must run after C区 widgets are created)
        self._toggle_three_way_view(init_only=True)

    # ---------- Scrolling sync ----------
    def _is_three_way_enabled(self) -> bool:
        try:
            return bool(getattr(self, "three_way_var", None) and self.three_way_var.get() and getattr(self.app, "merge_mode", False) and getattr(self.app, "has_base", False))
        except Exception:
            return False

    def _hover_compare_reserved_height(self, enabled: bool | None = None) -> int:
        enabled = self._is_three_way_enabled() if enabled is None else bool(enabled)
        line_count = 4 if enabled else 2
        line_px = 18
        try:
            from tkinter import font as tkfont
            line_px = max(16, int(tkfont.Font(font=self.editor_font).metrics("linespace")))
        except Exception:
            pass
        # Reserve enough vertical room for the header row, text chrome and x-scrollbar.
        baseline = 160 if enabled else 118
        return max(baseline, int(line_px * line_count + 88))

    def _sync_hover_compare_reserved_height(self, enabled: bool | None = None):
        enabled = self._is_three_way_enabled() if enabled is None else bool(enabled)
        host = getattr(self, "hover_cmp_host", None)
        if host is None:
            return
        try:
            host.configure(height=self._hover_compare_reserved_height(enabled))
        except Exception:
            pass

    @staticmethod
    def _source_display_name(path_like: str) -> str:
        """Prefer stable workbook names and keep SVN revision hints when available."""
        s = (path_like or "").strip().strip('"')
        if not s:
            return "-"

        # Synthetic marker produced by parser when BASE is exported from .svn.
        m = re.match(r"(.+?)@BASE\(\.svn\)$", s, re.IGNORECASE)
        if m:
            return f"{os.path.basename(m.group(1))}@BASE"

        bn = os.path.basename(s)
        ext_pat = r"(?:xlsx|xlsm)"

        # Exported revision snapshots created by this tool.
        m = re.match(rf"{re.escape(APP_NAME)}_svncat_r(\d+)_\d{{8}}_\d{{6}}_(.+?\.{ext_pat})$", bn, re.IGNORECASE)
        if m:
            return f"{m.group(2)}@r{m.group(1)}"
        m = re.match(rf"{re.escape(APP_NAME)}_svncat_BASE_\d{{8}}_\d{{6}}_(.+?\.{ext_pat})$", bn, re.IGNORECASE)
        if m:
            return f"{m.group(1)}@BASE"

        # Temp wrapper pattern: xxx.xlsx-rev123.svn456.tmp.xlsx -> xxx.xlsx@r123
        m = re.match(rf"(.+?\.{ext_pat})-rev(\d+)\.svn\d+\.tmp\.{ext_pat}$", bn, re.IGNORECASE)
        if m:
            return f"{m.group(1)}@r{m.group(2)}"

        # SVN conflict side files: file.merge-left.r123 -> file@r123
        m = re.match(rf"(.+?)\.merge-(?:left|right)\.r(\d+)$", bn, re.IGNORECASE)
        if m:
            return f"{m.group(1)}@r{m.group(2)}"

        # Generic prefixed temp: keep original workbook tail if present.
        m = re.match(rf".*?_(.+?\.{ext_pat})$", bn, re.IGNORECASE)
        if m and ".merge-" not in m.group(1).lower():
            return m.group(1)
        return bn

    def _sheet_meta(self) -> dict[str, object]:
        return self.app.get_sheet_meta(self.sheet)

    def _is_missing_sheet_view(self) -> bool:
        return str(self._sheet_meta().get("view_mode") or "") == "missing_sheet"

    def _side_has_sheet(self, side: str) -> bool:
        meta = self._sheet_meta()
        side = str(side or "").upper()
        if side == "A":
            return bool(meta.get("has_a"))
        if side in ("B", "THEIRS"):
            return bool(meta.get("has_b"))
        if side == "BASE":
            return bool(meta.get("has_base"))
        return False

    def _display_ws(self, side: str, edit: bool = False):
        side = str(side or "").upper()
        ws = self.app.ws_for_side(side, self.sheet, edit=edit, allow_missing=True)
        if ws is not None:
            return ws
        key = (side, bool(edit))
        cache = getattr(self, "_blank_ws_cache", None)
        if cache is None:
            self._blank_ws_cache = {}
            cache = self._blank_ws_cache
        if key not in cache:
            cache[key] = _blank_worksheet(self.sheet)
        return cache[key][1]

    def _base_row_for_pair(self, pair_idx: int, pair: tuple[int | None, int | None] | None = None) -> int | None:
        if not self._is_three_way_enabled():
            return None
        if not getattr(self.app, "has_base", False):
            return None
        overrides = getattr(self, "pair_base_row_override", {}) or {}
        if pair_idx in overrides:
            return overrides.get(pair_idx)
        if pair is None:
            if not (0 <= pair_idx < len(self.row_pairs)):
                return None
            pair = self.row_pairs[pair_idx]
        ra, rb = pair
        if self._is_missing_sheet_view():
            rows = getattr(self, "_missing_base_row_map", {}) or {}
            return rows.get(pair_idx)
        mine_map = getattr(self, "mine_to_base_row", {}) or {}
        theirs_map = getattr(self, "theirs_to_base_row", {}) or {}
        if ra is not None and ra in mine_map:
            return mine_map.get(ra)
        if rb is not None and rb in theirs_map:
            return theirs_map.get(rb)
        return None

    def _update_sheet_role_labels(self):
        enabled = self._is_three_way_enabled()
        meta = self._sheet_meta()
        try:
            if enabled:
                mine_src = getattr(self.app, "raw_mine", None) or self.app.file_a
                base_src = getattr(self.app, "raw_base", None) or getattr(self.app, "base_path", "")
                theirs_src = getattr(self.app, "raw_theirs", None) or self.app.file_b
                mine_text = f"mine={self._source_display_name(mine_src)}" if meta.get("has_a") else "mine=空白(该侧无此Sheet)"
                base_text = f"base={self._source_display_name(base_src)}" if (base_src and meta.get("has_base")) else "base=空白(该侧无此Sheet)"
                theirs_text = f"theirs={self._source_display_name(theirs_src)}" if meta.get("has_b") else "theirs=空白(该侧无此Sheet)"
                self.path_label_a.configure(text=mine_text, bg=_MINE_BG)
                self.path_label_base.configure(text=base_text, bg=_BASE_BG)
                self.path_label_b.configure(text=theirs_text, bg=_THEIRS_BG)
            else:
                base_src = getattr(self.app, "raw_base", None) or self.app.file_a
                mine_src = getattr(self.app, "raw_mine", None) or self.app.file_b
                left_text = f"base={self._source_display_name(base_src)}" if meta.get("has_a") else "base=空白(该侧无此Sheet)"
                right_text = f"mine={self._source_display_name(mine_src)}" if meta.get("has_b") else "mine=空白(该侧无此Sheet)"
                self.path_label_a.configure(text=left_text, bg=_BASE_BG)
                self.path_label_b.configure(text=right_text, bg=_MINE_BG)
        except Exception:
            pass
        try:
            if self._is_missing_sheet_view():
                if enabled:
                    left_state = "normal" if (meta.get("has_base") or meta.get("has_a")) else "disabled"
                else:
                    left_state = "normal" if meta.get("has_a") else "disabled"
                right_state = "normal" if meta.get("has_b") else "disabled"
                mine_state = "normal" if meta.get("has_a") else "disabled"
            else:
                left_state = "normal"
                right_state = "normal"
                mine_state = "normal"
            self.use_left_btn.configure(state=left_state)
            self.use_right_btn.configure(state=right_state)
            if self.use_base_btn is not None:
                self.use_base_btn.configure(state=mine_state)
        except Exception:
            pass
        try:
            self._refresh_copy_scope_buttons()
        except Exception:
            pass

    def _toggle_three_way_view(self, init_only: bool = False):
        enabled = self._is_three_way_enabled()
        def _one_line_text(s: str, max_len: int = 120) -> str:
            s = (s or "").replace("\r", " ").replace("\n", " ")
            if len(s) <= max_len:
                return s
            return "..." + s[-(max_len - 3):]
        try:
            panes = list(self._main_paned.panes())
            mid_id = str(self._mid_wrap)
            has_mid = mid_id in panes
            if enabled and (not has_mid):
                self._main_paned.insert(0, self._mid_wrap, weight=1)
            elif (not enabled) and has_mid:
                self._main_paned.forget(self._mid_wrap)
        except Exception:
            pass
        try:
            if enabled:
                # Layout: left=Base, center=Mine, right=Theirs — reorder header labels to match
                self.path_label_base.grid(row=0, column=0, sticky="ew")
                self.path_label_a.grid(row=0, column=1, sticky="ew")
                self.path_label_b.grid(row=0, column=2, sticky="ew")
                self.left_title.configure(text="Mine", background=_MINE_BG)
                self.mid_title.configure(text="Base", background=_BASE_BG)
                self.right_title.configure(text="Theirs", background=_THEIRS_BG)
            else:
                self.path_label_base.grid_remove()
                # Restore 2-way labels to left/right columns
                self.path_label_a.grid(row=0, column=0, sticky="ew")
                self.path_label_b.grid(row=0, column=2, sticky="ew")
                self.left_title.configure(text="Base", background=_BASE_BG)
                self.right_title.configure(text="Mine", background=_MINE_BG)
                self.mid_title.configure(text="Base", background=_BASE_BG)
            self._update_sheet_role_labels()
        except Exception:
            pass
        try:
            self.cursor_cmp.configure(height=3 if enabled else 2)
        except Exception:
            pass
        try:
            self.cursor_cmp_ln.configure(height=3 if enabled else 2)
        except Exception:
            pass
        try:
            if hasattr(self, "hover_cmp_text"):
                # 3-way hover panel needs one extra visible line to fully show
                # BASE / mine / theirs together after accounting for widget chrome.
                self.hover_cmp_text.configure(height=4 if enabled else 2)
        except Exception:
            pass
        try:
            self._sync_hover_compare_reserved_height(enabled)
        except Exception:
            pass
        if not init_only:
            try:
                self._invalidate_only_diff_snapshot_cache()
                self.refresh(row_only=None, rescan=False)
                self._update_cursor_lines()
            except Exception:
                pass
        try:
            self.frame.after(0, self._keep_panes_equal)
        except Exception:
            pass

    def _sync_main_x_to_frac(self, first):
        try:
            frac = float(first)
        except Exception:
            return

        prev_xsync = bool(getattr(self, "_xsyncing", False))
        self._xsyncing = True
        try:
            # Use left pane as anchor, then map by pixel position to peers.
            # This avoids drift when panes render different glyph widths.
            self.left.xview_moveto(frac)
            self._sync_main_x_from_widget(self.left, frac)
        except Exception:
            pass
        finally:
            self._xsyncing = prev_xsync

    def _sync_c_x_to_frac(self, first):
        try:
            frac = float(first)
        except Exception:
            return

        # Programmatic C-pane sync must not feed back into main panes.
        prev_suppress = bool(getattr(self, "_suppress_c_xsync", False))
        self._suppress_c_xsync = True
        try:
            try:
                c_first = self._map_xfirst_between_widgets(self.left, self.cursor_cmp, frac)
                self.cursor_cmp.xview_moveto(c_first)
                if getattr(self, "cursor_cmp_colhdr", None) is not None:
                    self.cursor_cmp_colhdr.xview_moveto(c_first)
            except Exception:
                pass
            try:
                cell_first = self._map_xfirst_between_widgets(self.left, self.cell_cmp_text, frac)
                self.cell_cmp_text.xview_moveto(cell_first)
            except Exception:
                pass
            self._set_c_hscrollbars_from_main()
        finally:
            self._suppress_c_xsync = prev_suppress

    def _set_c_hscrollbars_from_main(self):
        try:
            first, last = self.left.xview()
        except Exception:
            try:
                first, last = self.right.xview()
            except Exception:
                first, last = (0.0, 1.0)
        try:
            if getattr(self, "cursor_hsb", None) is not None:
                self.cursor_hsb.set(first, last)
        except Exception:
            pass
        try:
            if getattr(self, "cell_cmp_hsb", None) is not None:
                self.cell_cmp_hsb.set(first, last)
        except Exception:
            pass

    @staticmethod
    def _clamp(value: float, lo: float, hi: float) -> float:
        try:
            v = float(value)
        except Exception:
            return lo
        if v < lo:
            return lo
        if v > hi:
            return hi
        return v

    def _map_xfirst_between_widgets(self, src_widget: tk.Text, dst_widget: tk.Text, src_first) -> float:
        """Map horizontal scroll position by pixel offset, with end-point locking."""
        try:
            sf, sl = src_widget.xview()
            sf = float(sf)
            sl = float(sl)
            df, dl = dst_widget.xview()
            df = float(df)
            dl = float(dl)
            src_first_f = float(src_first)
            src_vis = max(1e-6, sl - sf)
            dst_vis = max(1e-6, dl - df)
            src_max = max(0.0, 1.0 - src_vis)
            dst_max = max(0.0, 1.0 - dst_vis)
            src_first_f = self._clamp(src_first_f, 0.0, src_max)

            # Lock endpoints: when source is at extreme left/right, target should also
            # reach its own extreme to avoid "not fully right" under different viewports.
            eps = 1e-4
            if src_first_f <= eps:
                return 0.0
            if src_max <= eps:
                return 0.0
            if src_first_f >= src_max - eps:
                return dst_max

            src_vp = max(1.0, float(src_widget.winfo_width()))
            dst_vp = max(1.0, float(dst_widget.winfo_width()))
            src_total = src_vp / src_vis
            dst_total = dst_vp / dst_vis
            if src_total <= 1e-6 or dst_total <= 1e-6:
                return self._clamp(src_first_f, 0.0, dst_max)

            # Convert to absolute left-pixel position in source, then map to destination ratio.
            left_px = src_first_f * src_total
            dst_first = left_px / dst_total
            return self._clamp(dst_first, 0.0, dst_max)
        except Exception:
            try:
                return float(src_first)
            except Exception:
                return 0.0

    def _sync_main_x_from_widget(self, src_widget: tk.Text, src_first):
        """Sync A/B(/BASE) using source-widget absolute x position."""
        try:
            left_first = self._map_xfirst_between_widgets(src_widget, self.left, src_first)
            right_first = self._map_xfirst_between_widgets(src_widget, self.right, src_first)
            self.left.xview_moveto(left_first)
            self.right.xview_moveto(right_first)
            if self._is_three_way_enabled():
                base_first = self._map_xfirst_between_widgets(src_widget, self.base, src_first)
                self.base.xview_moveto(base_first)
            if getattr(self, "left_colhdr", None) is not None:
                self.left_colhdr.xview_moveto(left_first)
            if getattr(self, "right_colhdr", None) is not None:
                self.right_colhdr.xview_moveto(right_first)
            if self._is_three_way_enabled() and getattr(self, "base_colhdr", None) is not None:
                base_first = float((self.base.xview() or (0.0, 1.0))[0])
                self.base_colhdr.xview_moveto(base_first)
            lf, ll = self.left.xview()
            rf, rl = self.right.xview()
            self.hsb_left.set(lf, ll)
            self.hsb_right.set(rf, rl)
            if self._is_three_way_enabled():
                mf, ml = self.base.xview()
                self.hsb_mid.set(mf, ml)
        except Exception:
            pass

    def _sync_c_x_from_widget(self, src_widget: tk.Text, src_first):
        """Sync C panes using source-widget absolute x position."""
        prev_suppress = bool(getattr(self, "_suppress_c_xsync", False))
        self._suppress_c_xsync = True
        try:
            try:
                c_first = self._map_xfirst_between_widgets(src_widget, self.cursor_cmp, src_first)
                self.cursor_cmp.xview_moveto(c_first)
                if getattr(self, "cursor_cmp_colhdr", None) is not None:
                    self.cursor_cmp_colhdr.xview_moveto(c_first)
            except Exception:
                pass
            try:
                if hasattr(self, "cell_cmp_text"):
                    cell_first = self._map_xfirst_between_widgets(src_widget, self.cell_cmp_text, src_first)
                    self.cell_cmp_text.xview_moveto(cell_first)
            except Exception:
                pass
            self._set_c_hscrollbars_from_main()
        finally:
            self._suppress_c_xsync = prev_suppress

    def _is_click_trace_active(self) -> bool:
        try:
            return float(getattr(self, "_trace_click_until", 0.0) or 0.0) > time.time()
        except Exception:
            return False

    def _log_click_trace_state(self, stage: str):
        if not self._is_click_trace_active():
            return
        try:
            lx = float((self.left.xview() or (0.0, 1.0))[0])
        except Exception:
            lx = -1.0
        try:
            rx = float((self.right.xview() or (0.0, 1.0))[0])
        except Exception:
            rx = -1.0
        try:
            cx = float((self.cursor_cmp.xview() or (0.0, 1.0))[0])
        except Exception:
            cx = -1.0
        try:
            li = str(self.left.index("insert"))
        except Exception:
            li = "?"
        try:
            ri = str(self.right.index("insert"))
        except Exception:
            ri = "?"
        _dlog(f"click_trace {stage} x(left={lx:.6f},right={rx:.6f},c={cx:.6f}) insert(left={li},right={ri})")

    def _post_click_x_guard(self, saved_x: float, stage: str):
        try:
            now_x = float((self.left.xview() or (0.0, 1.0))[0])
        except Exception:
            now_x = saved_x
        drift = abs(now_x - saved_x)
        if drift > 1e-4:
            try:
                self._sync_main_x_to_frac(saved_x)
                self._sync_c_x_to_frac(saved_x)
            except Exception:
                pass
            _dlog(f"click_guard_restore stage={stage} saved={saved_x:.6f} now={now_x:.6f} drift={drift:.6f}")
        else:
            _dlog(f"click_guard_ok stage={stage} saved={saved_x:.6f} now={now_x:.6f}")
        self._log_click_trace_state(f"post_guard:{stage}")

    def _xview_cursor_cmp(self, *args):
        if self._xsyncing:
            return
        self._xsyncing = True
        try:
            # Drive the main pane directly so the full A/B scroll range is accessible.
            # C区's viewport is wider, so using cursor_cmp as driver would clamp
            # the scroll position and prevent reaching the rightmost content in A/B.
            self.left.xview(*args)
            first, last = self.left.xview()
            self._sync_main_x_to_frac(first)
            self._sync_c_x_from_widget(self.left, first)
            self._set_c_hscrollbars_from_main()
        finally:
            self._xsyncing = False
        try:
            self._schedule_diff_maps()
        except Exception:
            pass

    def _xscroll_cursor_cmp(self, first, last):
        if self._is_click_trace_active():
            _dlog(f"xscroll_cursor_cmp first={first} last={last} xsyncing={self._xsyncing} suppress={getattr(self, '_suppress_c_xsync', False)}")
        # Passive C-pane xscroll callback should never drive main panes.
        # Only explicit C scrollbar command handlers (_xview_cursor_cmp/_xview_cell_cmp)
        # are allowed to sync main xview.
        self._set_c_hscrollbars_from_main()
        return

    def _xview_cell_cmp(self, *args):
        if self._xsyncing:
            return
        self._xsyncing = True
        try:
            # Keep the lower C-area scrollbar as a main-pane controller too.
            # The cell text itself is often narrower than the main sheet viewport,
            # so driving from its own xview would collapse the thumb to 0..1.
            self.left.xview(*args)
            first, last = self.left.xview()
            self._sync_main_x_to_frac(first)
            self._sync_c_x_from_widget(self.left, first)
            self._set_c_hscrollbars_from_main()
        finally:
            self._xsyncing = False
        try:
            self._schedule_diff_maps()
        except Exception:
            pass

    def _xscroll_cell_cmp(self, first, last):
        if self._is_click_trace_active():
            _dlog(f"xscroll_cell_cmp first={first} last={last} xsyncing={self._xsyncing} suppress={getattr(self, '_suppress_c_xsync', False)}")
        # Same rule as cursor_cmp: passive callback updates only its own scrollbar.
        self._set_c_hscrollbars_from_main()
        return

    def _is_grid_overlay_enabled(self) -> bool:
        try:
            return bool(getattr(self, "grid_overlay_var", None) and self.grid_overlay_var.get())
        except Exception:
            return False

    def _gridify_parts(self, parts: list[str]) -> list[str]:
        if not self._is_grid_overlay_enabled():
            return parts
        # Keep tab layout unchanged; prepend a visual splitter in each cell.
        return [f"|{p}" for p in parts]

    def _toggle_grid_overlay(self):
        try:
            self._invalidate_render_cache()
            # pair_text_a/b cache formatted line strings; must be cleared so
            # _build_row_and_diff_pair re-runs with the new grid-on/off separator.
            self.pair_text_a = {}
            self.pair_text_b = {}
            # Rescan widths to avoid stale narrow columns when toggling grid mode.
            self.refresh(row_only=None, rescan=True)
            self._update_cursor_lines()
        except Exception:
            pass

    def _yscroll_all(self, first, last):
        for sb in (self.vsb_left, self.vsb_a, self.vsb_m, self.vsb_b):
            try:
                sb.set(first, last)
            except Exception:
                pass
        try:
            self._schedule_diff_maps()
        except Exception:
            pass

    def _yscroll_left(self, first, last):
        if self._syncing:
            return
        self._syncing = True
        try:
            self.left_ln.yview_moveto(first)
            if self._is_three_way_enabled():
                self.base.yview_moveto(first)
                self.base_ln.yview_moveto(first)
            self.right.yview_moveto(first)
            self.right_ln.yview_moveto(first)
            self._yscroll_all(first, last)
        finally:
            self._syncing = False
        self._maybe_load_more_rows(last)

    def _yscroll_mid(self, first, last):
        if self._syncing:
            return
        self._syncing = True
        try:
            self.left.yview_moveto(first)
            self.left_ln.yview_moveto(first)
            self.right.yview_moveto(first)
            self.right_ln.yview_moveto(first)
            self.base_ln.yview_moveto(first)
            self._yscroll_all(first, last)
        finally:
            self._syncing = False
        self._maybe_load_more_rows(last)

    def _yscroll_right(self, first, last):
        if self._syncing:
            return
        self._syncing = True
        try:
            if self._is_three_way_enabled():
                self.base.yview_moveto(first)
                self.base_ln.yview_moveto(first)
            self.left.yview_moveto(first)
            self.left_ln.yview_moveto(first)
            self.right_ln.yview_moveto(first)
            self._yscroll_all(first, last)
        finally:
            self._syncing = False
        self._maybe_load_more_rows(last)

    def _yview_both(self, *args):
        self._syncing = True
        try:
            self.left.yview(*args)
            self.left_ln.yview(*args)
            if self._is_three_way_enabled():
                self.base.yview(*args)
                self.base_ln.yview(*args)
            self.right.yview(*args)
            self.right_ln.yview(*args)
            try:
                first, last = self.left.yview()
                self._yscroll_all(first, last)
            except Exception:
                pass
        finally:
            self._syncing = False
        try:
            _first, last = self.left.yview()
            self._maybe_load_more_rows(last)
        except Exception:
            pass

    def _on_mousewheel(self, event):
        if getattr(event, "num", None) == 4:
            delta = 120
        elif getattr(event, "num", None) == 5:
            delta = -120
        else:
            delta = event.delta
        steps = int(-1 * (delta / 120))
        self._yview_both("scroll", steps, "units")
        return "break"

    def _on_vdiff_map_click(self, event):
        try:
            h = max(1, self.vdiff_map.winfo_height())
            frac = min(1.0, max(0.0, float(event.y) / float(h)))
            total_pairs = len(self.row_pairs) if getattr(self, "row_pairs", None) else 0
            if total_pairs > 0 and self.display_rows:
                target_pair = min(total_pairs - 1, max(0, int(frac * total_pairs)))
                pos = bisect.bisect_left(self.display_rows, target_pair)
                if pos >= len(self.display_rows):
                    pos = len(self.display_rows) - 1
                elif pos > 0:
                    prev_pair = self.display_rows[pos - 1]
                    next_pair = self.display_rows[pos]
                    if abs(prev_pair - target_pair) <= abs(next_pair - target_pair):
                        pos -= 1
                if pos >= 0:
                    target_frac = float(pos) / float(max(1, len(self.display_rows)))
                    self._yview_both("moveto", target_frac)
                    return
            self._yview_both("moveto", frac)
        except Exception:
            pass

    def _visible_pair_span(self):
        if not self.display_rows:
            return None
        top_line = 1
        bottom_line = len(self.display_rows)
        try:
            top_line = int(str(self.left.index("@0,0")).split(".")[0])
        except Exception:
            top_line = 1
        try:
            bottom_y = max(0, int(self.left.winfo_height()) - 1)
            bottom_line = int(str(self.left.index(f"@0,{bottom_y}")).split(".")[0])
        except Exception:
            bottom_line = len(self.display_rows)
        top_line = max(1, min(top_line, len(self.display_rows)))
        bottom_line = max(top_line, min(bottom_line, len(self.display_rows)))
        start_pair = self.display_rows[top_line - 1]
        end_pair = self.display_rows[bottom_line - 1]
        if end_pair < start_pair:
            start_pair, end_pair = end_pair, start_pair
        return start_pair, min(len(self.row_pairs), end_pair + 1)

    def _on_hdiff_map_click(self, event, pane: str):
        try:
            canvas = self.hdiff_left if pane == "left" else (self.hdiff_mid if pane == "mid" else self.hdiff_right)
            w = max(1, canvas.winfo_width())
            frac = min(1.0, max(0.0, float(event.x) / float(w)))
            self._sync_main_x_to_frac(frac)
            self._sync_c_x_to_frac(frac)
        except Exception:
            pass

    def _compute_diff_map_data(self):
        """Return (total_pairs, diff_rows, diff_cols) cached by data version.

        The expensive O(N) scan over all pairs (``_pair_has_visual_diff``) only
        runs when the underlying data changed (``_data_version`` bump), so plain
        scrolling no longer pays for it on every frame.
        """
        ver = self._data_version
        cached = getattr(self, "_diff_map_cache", None)
        if cached is not None and getattr(self, "_diff_map_cache_version", None) == ver:
            return cached
        total_pairs = len(self.row_pairs) if getattr(self, "row_pairs", None) else 0
        diff_rows = sorted(
            int(pidx)
            for pidx in range(total_pairs)
            if self._pair_has_visual_diff(pidx)
        )
        diff_cols = set()
        for cols in (self.pair_diff_cols or {}).values():
            if cols:
                diff_cols.update(c for c in cols if c > 0)
        data = (total_pairs, diff_rows, sorted(diff_cols))
        self._diff_map_cache = data
        self._diff_map_cache_version = ver
        return data

    def _redraw_diff_markers(self):
        """Redraw the data-dependent diff markers (tagged ``dmark``)."""
        try:
            total_pairs, diff_rows, diff_cols = self._compute_diff_map_data()
        except Exception:
            return
        # Vertical diff map (by logical pair index across the full sheet)
        try:
            self.vdiff_map.delete("dmark")
            h = max(1, self.vdiff_map.winfo_height())
            w = max(1, self.vdiff_map.winfo_width())
            n = max(1, total_pairs)
            diff_count = len(diff_rows)
            # Dynamic marker height:
            # - fewer diffs -> thicker marker for visibility
            # - many diffs -> thinner marker to reduce overlap
            marker_min_h = max(2, min(12, int(h / max(1, diff_count)))) if diff_count > 0 else 2

            # Draw contiguous diff blocks as filled segments for better discoverability.
            i = 0
            while i < diff_count:
                start_pair = diff_rows[i]
                end_pair = start_pair
                while i + 1 < diff_count and diff_rows[i + 1] == end_pair + 1:
                    i += 1
                    end_pair = diff_rows[i]
                y1 = int((start_pair / n) * h)
                y2 = int(((end_pair + 1) / n) * h)
                if y2 - y1 < marker_min_h:
                    y2 = min(h, y1 + marker_min_h)
                self.vdiff_map.create_rectangle(0, y1, w, y2, outline="", fill="#ff2d2d", tags="dmark")
                i += 1
            self.vdiff_map.tag_raise("vpbox")
        except Exception:
            pass

        # Horizontal diff maps (under each pane scrollbar; by diff columns)
        try:
            max_col = max(1, int(self.max_col or 1))
            canvases = [self.hdiff_left, self.hdiff_right]
            if self._is_three_way_enabled():
                canvases.insert(1, self.hdiff_mid)
            for canvas in canvases:
                canvas.delete("dmark")
                cw = max(1, canvas.winfo_width())
                ch = max(1, canvas.winfo_height())
                if diff_cols:
                    marker_min_w = max(2, min(14, int(cw / max(1, len(diff_cols)))))
                    sorted_cols = sorted(diff_cols)
                    seg_start = sorted_cols[0]
                    seg_end = sorted_cols[0]
                    for c in sorted_cols[1:]:
                        if c == seg_end + 1:
                            seg_end = c
                        else:
                            x1 = int(((seg_start - 1) / max_col) * cw)
                            x2 = int((seg_end / max_col) * cw)
                            if x2 - x1 < marker_min_w:
                                x2 = min(cw, x1 + marker_min_w)
                            canvas.create_rectangle(x1, 0, x2, ch, outline="", fill="#c46a00", tags="dmark")
                            seg_start = c
                            seg_end = c
                    x1 = int(((seg_start - 1) / max_col) * cw)
                    x2 = int((seg_end / max_col) * cw)
                    if x2 - x1 < marker_min_w:
                        x2 = min(cw, x1 + marker_min_w)
                    canvas.create_rectangle(x1, 0, x2, ch, outline="", fill="#c46a00", tags="dmark")
                canvas.tag_raise("vpbox")
        except Exception:
            pass

    def _redraw_diff_viewport(self):
        """Redraw only the viewport indicator boxes (tagged ``vpbox``, O(1))."""
        # Vertical viewport box
        try:
            total_pairs = len(self.row_pairs) if getattr(self, "row_pairs", None) else 0
            n = max(1, total_pairs)
            self.vdiff_map.delete("vpbox")
            h = max(1, self.vdiff_map.winfo_height())
            w = max(1, self.vdiff_map.winfo_width())
            span = self._visible_pair_span()
            if span is not None:
                start_pair, end_pair = span
                y1 = int((start_pair / n) * h)
                y2 = max(y1 + 2, int((end_pair / n) * h))
                self.vdiff_map.create_rectangle(0, y1, w, y2, outline="#1e78ff", tags="vpbox")
        except Exception:
            pass

        # Horizontal viewport boxes
        try:
            canvases = [self.hdiff_left, self.hdiff_right]
            if self._is_three_way_enabled():
                canvases.insert(1, self.hdiff_mid)
            lf, ll = self.left.xview()
            for canvas in canvases:
                canvas.delete("vpbox")
                cw = max(1, canvas.winfo_width())
                ch = max(1, canvas.winfo_height())
                canvas.create_rectangle(int(lf * cw), 0, max(int(ll * cw), int(lf * cw) + 2), ch, outline="#1e78ff", tags="vpbox")
        except Exception:
            pass

    def _update_diff_maps(self):
        """Full diff-map redraw (markers + viewport). For data-change callers."""
        self._redraw_diff_markers()
        self._redraw_diff_viewport()

    def _run_diff_markers_debounced(self):
        self._diff_map_debounce_id = None
        try:
            if not self.frame.winfo_exists():
                return
        except Exception:
            return
        self._redraw_diff_markers()

    def _schedule_diff_maps(self):
        """Scroll-driven diff-map update.

        Redraws the cheap viewport box immediately for responsive feedback and
        debounces the (cached) marker redraw to coalesce rapid scroll frames.
        """
        self._redraw_diff_viewport()
        aid = getattr(self, "_diff_map_debounce_id", None)
        if aid is not None:
            try:
                self.frame.after_cancel(aid)
            except Exception:
                pass
        delay = int(getattr(self, "_diff_map_debounce_ms", 40) or 0)
        if delay <= 0:
            self._redraw_diff_markers()
            return
        try:
            self._diff_map_debounce_id = self.frame.after(delay, self._run_diff_markers_debounced)
        except Exception:
            self._diff_map_debounce_id = None
            self._redraw_diff_markers()

    # ---------- Selection + toolbar buttons ----------
    def _widget_line(self, w: tk.Text):
        try:
            idx = w.index("@%d,%d" % (w.winfo_pointerx() - w.winfo_rootx(), w.winfo_pointery() - w.winfo_rooty()))
        except Exception:
            idx = w.index("insert")
        return int(str(idx).split(".")[0])

    def _pair_idx_for_line(self, line: int) -> int | None:
        if not (1 <= line <= len(self.display_rows)):
            return None
        return self.display_rows[line - 1]

    def _pair_for_line(self, line: int):
        idx = self._pair_idx_for_line(line)
        if idx is None or idx >= len(self.row_pairs):
            return None
        return self.row_pairs[idx]

    def _side_for_widget(self, w: tk.Text) -> str:
        if w is self.left:
            return "A"
        if w is self.base:
            return "BASE"
        return "B"

    def _col_from_char(self, char_pos: int) -> int | None:
        """Resolve display column index by character offset in a rendered row line."""
        try:
            ch = int(char_pos)
        except Exception:
            return None
        spans = self._spans_for_line()
        for c, (s, e) in spans.items():
            if s <= ch < e:
                return c
        return None

    def _set_main_selected_cell(self, line: int | None, col: int | None):
        try:
            ln = int(line) if line is not None else None
            cc = int(col) if col is not None else None
        except Exception:
            ln = None
            cc = None
        if ln is None or cc is None or ln < 1 or cc < 1:
            self._main_sel_line = None
            self._main_sel_col = None
            return
        self._main_sel_line = ln
        self._main_sel_col = cc

    def has_explicit_cell_selection(self) -> bool:
        try:
            if self._main_sel_line is not None and self._main_sel_col is not None:
                return True
            if self._cursor_cmp_sel_line is not None and self._cursor_cmp_sel_col is not None:
                return True
        except Exception:
            return False
        return False

    def _clear_selected_line_highlight(self):
        prev = self._last_selected_line
        if prev is None:
            return
        try:
            for t in (self.left, self.base, self.right):
                t.tag_remove("selrow", f"{prev}.0", f"{prev}.end")
        except Exception:
            pass
        self._last_selected_line = None

    def clear_explicit_cell_selection(self):
        self._set_main_selected_cell(None, None)
        self._cursor_cmp_sel_col = None
        self._cursor_cmp_sel_line = None
        self.selected_pair_idx = None
        self.selected_excel_row = None
        self.selected_excel_row_a = None
        self.selected_excel_row_b = None
        self._clear_selected_line_highlight()

    def _clear_selection_visuals(self):
        try:
            for t in (self.left, self.base, self.right):
                t.tag_remove("selrow", "1.0", "end")
                t.tag_remove("selcell", "1.0", "end")
        except Exception:
            pass
        try:
            self.cursor_cmp.tag_remove("cselcell", "1.0", "end")
        except Exception:
            pass
        self._last_selected_line = None
        self._applied_main_sel_line = None
        self._applied_main_sel_col = None

    def _snapshot_explicit_selection_state(self):
        if not self.has_explicit_cell_selection():
            return None
        return {
            "pair_idx": self.selected_pair_idx,
            "row_a": self.selected_excel_row_a,
            "row_b": self.selected_excel_row_b,
            "main_col": self._main_sel_col,
            "cursor_cmp_sel_line": self._cursor_cmp_sel_line,
            "cursor_cmp_sel_col": self._cursor_cmp_sel_col,
        }

    def _pair_idx_from_selection_snapshot(self, snapshot) -> int | None:
        if not snapshot:
            return None
        try:
            row_a = snapshot.get("row_a")
            if row_a is not None:
                pair_idx = self.row_a_to_pair_idx.get(int(row_a))
                pair_idx = self._normalize_pair_idx(pair_idx)
                if pair_idx is not None:
                    return pair_idx
        except Exception:
            pass
        try:
            row_b = snapshot.get("row_b")
            if row_b is not None:
                pair_idx = self.row_b_to_pair_idx.get(int(row_b))
                pair_idx = self._normalize_pair_idx(pair_idx)
                if pair_idx is not None:
                    return pair_idx
        except Exception:
            pass
        return self._normalize_pair_idx(snapshot.get("pair_idx"))

    def _restore_explicit_selection_state(self, snapshot) -> bool:
        pair_idx = self._pair_idx_from_selection_snapshot(snapshot)
        if pair_idx is None or pair_idx not in self.row_to_line:
            return False
        try:
            line = int(self.row_to_line[pair_idx])
        except Exception:
            return False
        try:
            self._highlight_selected_line(line)
            self.selected_pair_idx = int(pair_idx)
            pair = self._pair_for_line(line)
            self.selected_excel_row_a = self._row_for_side(pair, "A")
            self.selected_excel_row_b = self._row_for_side(pair, "B")
            self.selected_excel_row = self.selected_excel_row_a or self.selected_excel_row_b
            main_col = snapshot.get("main_col")
            self._set_main_selected_cell(line, main_col)
            cursor_col = snapshot.get("cursor_cmp_sel_col")
            cursor_line = snapshot.get("cursor_cmp_sel_line")
            self._cursor_cmp_sel_col = int(cursor_col) if cursor_col is not None else None
            self._cursor_cmp_sel_line = int(cursor_line) if cursor_line is not None else None
            self._last_cursor_cmp_pair_idx = int(pair_idx)
            return True
        except Exception:
            return False

    def _clear_hover_state(self, *, clear_panel: bool = False):
        self.hover_pair_idx = None
        self.hover_col_idx = None
        self.hover_side = None
        self._last_cursor_cmp_pair_idx = None
        self._hide_hover_popup()
        if clear_panel and (not self._hover_compare_is_pinned()):
            self._clear_hover_compare_panel()
        self._cell_tip_key = None

    def _normalize_pair_idx(self, pair_idx) -> int | None:
        try:
            if pair_idx is None:
                return None
            pair_idx = int(pair_idx)
            if pair_idx < 0 or pair_idx >= len(self.row_pairs):
                return None
            return pair_idx
        except Exception:
            return None

    def resolved_pair_idx_for_c_area(self) -> int | None:
        candidates = []
        if self.has_explicit_cell_selection():
            candidates.append(self.selected_pair_idx)
        candidates.append(self.hover_pair_idx)
        candidates.append(self.selected_pair_idx)
        candidates.append(getattr(self, "_last_cursor_cmp_pair_idx", None))
        try:
            line_guess = int(str(self.left.index("insert")).split(".")[0])
        except Exception:
            line_guess = None
        if line_guess is not None:
            candidates.append(self._pair_idx_for_line(line_guess))
        for pair_idx in candidates:
            norm = self._normalize_pair_idx(pair_idx)
            if norm is not None:
                return norm
        return None

    def update_hover_driven_panels(
        self,
        pair_idx: int | None,
        col_idx: int | None,
        side: str,
        *,
        force_panel: bool = True,
        popup_force_show: bool = False,
        x_root: int | None = None,
        y_root: int | None = None,
        refresh_c_area: bool = True,
    ):
        pair_idx = self._normalize_pair_idx(pair_idx)
        try:
            col_idx = int(col_idx) if col_idx is not None else None
        except Exception:
            col_idx = None
        self.hover_pair_idx = pair_idx
        self.hover_col_idx = col_idx
        self.hover_side = str(side or "").upper() if side is not None else None

        if pair_idx is None or col_idx is None or col_idx <= 0:
            self._hide_cell_tooltip(clear_panel=False)
            if refresh_c_area and not self.has_explicit_cell_selection():
                self._update_cursor_lines()
            return

        if refresh_c_area and not self.has_explicit_cell_selection():
            self._update_cursor_lines()

        panel_payload = self._cmp_tooltip_payload_by_pair_col(
            pair_idx,
            col_idx,
            force_show=bool(popup_force_show),
            force_panel=bool(force_panel),
        )
        if panel_payload:
            panel_text, panel_key = panel_payload
            self._set_hover_compare_panel(panel_text, panel_key)
        popup_payload = self._cmp_tooltip_payload_by_pair_col(pair_idx, col_idx, force_show=bool(popup_force_show))
        if popup_payload and x_root is not None and y_root is not None:
            tip_text, key = popup_payload
            self._show_cell_tooltip(tip_text, int(x_root), int(y_root), key)
        else:
            self._hide_cell_tooltip(clear_panel=False)

    def _apply_main_selected_cell_highlight(self):
        # Remove previously applied selcell highlight in O(1).
        prev_line = self._applied_main_sel_line
        prev_col = self._applied_main_sel_col
        if prev_line and prev_col:
            try:
                spans_prev = self._spans_for_line()
                if prev_col in spans_prev:
                    s0, e0 = spans_prev[prev_col]
                    for t in (self.left, self.base, self.right):
                        t.tag_remove("selcell", f"{prev_line}.{s0}", f"{prev_line}.{e0}")
            except Exception:
                pass
        self._applied_main_sel_line = None
        self._applied_main_sel_col = None

        line = self._main_sel_line
        col = self._main_sel_col
        if not line or not col:
            return
        if not (1 <= int(line) <= max(1, len(self.display_rows))):
            return

        try:
            spans = self._spans_for_line()
            if col not in spans:
                return
            s, e = spans[col]
            pair_idx = self.display_rows[int(line) - 1]
            pair = self._pair_for_line(int(line))
            ra = self._row_for_side(pair, "A")
            rb = self._row_for_side(pair, "B")
            base_r = self._base_row_for_pair(pair_idx, pair) if self._is_three_way_enabled() else None
            if ra is not None:
                self.left.tag_add("selcell", f"{line}.{s}", f"{line}.{e}")
            if self._is_three_way_enabled() and base_r is not None:
                self.base.tag_add("selcell", f"{line}.{s}", f"{line}.{e}")
            if rb is not None:
                self.right.tag_add("selcell", f"{line}.{s}", f"{line}.{e}")
            self._applied_main_sel_line = int(line)
            self._applied_main_sel_col = int(col)
        except Exception:
            pass

    @staticmethod
    def _row_for_side(pair, side: str) -> int | None:
        if not pair:
            return None
        if side == "A":
            return pair[0]
        if side == "BASE":
            # Base must map to its own aligned row identity (A-side index only).
            # Falling back to B causes duplicated base row numbers in gap regions.
            return pair[0]
        return pair[1]

    def _select_from_widget(self, w: tk.Text, event):
        # Set insert mark to the clicked position so follow-up actions can use it.
        x_before = 0.0
        try:
            x_before = float((self.left.xview() or (0.0, 1.0))[0])
        except Exception:
            x_before = 0.0
        try:
            idx = w.index(f"@{event.x},{event.y}")
            w.mark_set("insert", idx)
        except Exception:
            idx = None

        line = self._widget_line(w)
        col = 0
        try:
            if idx is not None:
                col = int(str(idx).split(".")[1])
        except Exception:
            col = 0

        # Resolve clicked character position to display column index.
        hit_col = self._col_from_char(col)
        self._set_main_selected_cell(line, hit_col)

        # Mirror main-pane cell selection into C区 selected-cell hint.
        # 2-way: A->line1, B->line2
        # 3-way: BASE->line1, A->line2, B->line3
        c_line = None
        if hit_col is not None and int(hit_col) > 0:
            side = self._side_for_widget(w)
            if self._is_three_way_enabled():
                if side == "BASE":
                    c_line = 1
                elif side == "A":
                    c_line = 2
                else:
                    c_line = 3
            else:
                c_line = 1 if side in ("A", "BASE") else 2
        self._cursor_cmp_sel_col = int(hit_col) if c_line is not None else None
        self._cursor_cmp_sel_line = int(c_line) if c_line is not None else None

        # Keep both panes aligned for the cursor compare block:
        # when user clicks either side, keep row+column insert marks consistent.
        for other in (self.left, self.base, self.right):
            if other is w:
                continue
            try:
                other.mark_set("insert", f"{line}.{max(0, col)}")
            except Exception:
                pass

        self._highlight_selected_line(line)
        pair = self._pair_for_line(line)
        self.selected_pair_idx = self._pair_idx_for_line(line)
        self.hover_pair_idx = self._normalize_pair_idx(self.selected_pair_idx)
        self.hover_col_idx = int(hit_col) if hit_col is not None else None
        self.hover_side = self._side_for_widget(w)
        self.selected_excel_row_a = self._row_for_side(pair, "A")
        self.selected_excel_row_b = self._row_for_side(pair, "B")
        self.selected_excel_row = self.selected_excel_row_a or self.selected_excel_row_b
        # No button state updates (performance): buttons are always visible and logic no-ops when no diff.

        self._update_cursor_lines()
        self._update_diff_nav_state()

        try:
            x_after = float((self.left.xview() or (0.0, 1.0))[0])
            _dlog(f"select_from_widget xview before={x_before:.6f} after={x_after:.6f} line={line} col={col}")
        except Exception:
            pass

    def _select_line(self, line: int):
        if line < 1:
            return
        for w in (self.left, self.base, self.right):
            try:
                w.mark_set("insert", f"{line}.0")
            except Exception:
                pass
        self._highlight_selected_line(line)
        pair = self._pair_for_line(line)
        self.selected_pair_idx = self._pair_idx_for_line(line)
        self.hover_pair_idx = self._normalize_pair_idx(self.selected_pair_idx)
        self.hover_col_idx = None
        self.hover_side = None
        self.selected_excel_row_a = self._row_for_side(pair, "A")
        self.selected_excel_row_b = self._row_for_side(pair, "B")
        self.selected_excel_row = self.selected_excel_row_a or self.selected_excel_row_b
        self._set_main_selected_cell(line, None)
        self._update_cursor_lines()
        self._update_diff_nav_state()

    def _on_row_header_click(self, w: tk.Text, event, direction: str):
        try:
            idx = w.index(f"@{event.x},{event.y}")
            line = int(str(idx).split(".")[0])
        except Exception:
            line = 1
        self._clear_row_header_hover(w)
        self._select_line(line)

        # 3-way UX: mine row-header is selection-only (no overwrite action).
        if self._is_three_way_enabled() and w is self.left_ln:
            return "break"

        self._copy_selected_row(direction, row_header=True)
        return "break"

    def _row_header_side(self, w: tk.Text) -> str:
        if w is self.left_ln:
            return "A"
        if w is self.base_ln:
            return "BASE"
        return "B"

    def _set_row_header_text(self, w: tk.Text, line: int, txt: str):
        try:
            # Preserve existing row-header visual tags (e.g. diffrow background)
            # when replacing text for hover arrow rendering.
            keep_tags = []
            try:
                keep_tags = list(w.tag_names(f"{line}.0"))
            except Exception:
                keep_tags = []

            w.configure(state="normal")
            w.delete(f"{line}.0", f"{line}.end")
            w.insert(f"{line}.0", txt)
            for tag in keep_tags:
                try:
                    w.tag_add(tag, f"{line}.0", f"{line}.end")
                except Exception:
                    pass
            w.configure(state="disabled")
        except Exception:
            pass

    def _clear_row_header_hover(self, w: tk.Text):
        line = getattr(self, "_hover_ln_line_left", None) if w is self.left_ln else (
            getattr(self, "_hover_ln_line_mid", None) if w is self.base_ln else getattr(self, "_hover_ln_line_right", None)
        )
        if line is None:
            return
        pair_idx = self._pair_idx_for_line(line)
        if pair_idx is None:
            return
        rn_w = self._sync_row_header_width_widgets()
        side = self._row_header_side(w)
        txt = self._row_label_for_pair_idx(pair_idx, side).rjust(rn_w)
        self._set_row_header_text(w, line, txt)
        try:
            w.configure(cursor="arrow")
        except Exception:
            pass
        if w is self.left_ln:
            self._hover_ln_line_left = None
        elif w is self.base_ln:
            self._hover_ln_line_mid = None
        else:
            self._hover_ln_line_right = None

    def _on_row_header_hover(self, w: tk.Text, event, direction: str):
        try:
            idx = w.index(f"@{event.x},{event.y}")
            line = int(str(idx).split(".")[0])
        except Exception:
            self._clear_row_header_hover(w)
            return
        if not (1 <= line <= len(self.display_rows)):
            self._clear_row_header_hover(w)
            return

        # 3-way UX: mine row-header should not display action arrow.
        if self._is_three_way_enabled() and w is self.left_ln:
            self._clear_row_header_hover(w)
            try:
                w.configure(cursor="arrow")
            except Exception:
                pass
            return

        pair = self._pair_for_line(line)
        pair_idx = self._pair_idx_for_line(line)
        side = self._row_header_side(w)
        r = self._row_for_side(pair, side)
        cols = self.pair_diff_cols.get(pair_idx, set()) if pair_idx is not None else set()
        allow_base_row_action = (self._is_three_way_enabled() and direction == "BASE2A" and w is self.base_ln)
        if r is None or ((not cols) and (not allow_base_row_action)):
            self._clear_row_header_hover(w)
            return
        hover_line = getattr(self, "_hover_ln_line_left", None) if w is self.left_ln else (
            getattr(self, "_hover_ln_line_mid", None) if w is self.base_ln else getattr(self, "_hover_ln_line_right", None)
        )
        if hover_line == line:
            return
        self._clear_row_header_hover(w)
        rn_w = self._sync_row_header_width_widgets()
        arrow = _ROW_ARROW_RIGHT if direction in ("A2B", "MINE2A", "BASE2A") else _ROW_ARROW_LEFT
        self._set_row_header_text(w, line, arrow.rjust(rn_w))
        try:
            w.configure(cursor="hand2")
        except Exception:
            pass
        if w is self.left_ln:
            self._hover_ln_line_left = line
        elif w is self.base_ln:
            self._hover_ln_line_mid = line
        else:
            self._hover_ln_line_right = line

    def _hide_hover_popup(self):
        tip_lbl = getattr(self, "_cell_tip_label", None)
        if tip_lbl is not None:
            try:
                tip_lbl.place_forget()
            except Exception:
                pass
        tip = getattr(self, "_cell_tip_win", None)
        if tip is not None:
            try:
                tip.destroy()
            except Exception:
                pass
        self._cell_tip_win = None

    def _clear_hover_compare_panel(self):
        try:
            if hasattr(self, "hover_cmp_title_var"):
                self.hover_cmp_title_var.set("悬停完整对比：-")
            if hasattr(self, "hover_cmp_text"):
                self.hover_cmp_text.configure(state="normal")
                self.hover_cmp_text.delete("1.0", "end")
                self.hover_cmp_text.configure(state="disabled")
        except Exception:
            pass
        self._last_hover_compare_key = None

    def _hover_compare_is_pinned(self) -> bool:
        try:
            return bool(getattr(self, "hover_cmp_pin_var", None) and self.hover_cmp_pin_var.get())
        except Exception:
            return False

    def _on_hover_compare_pin_toggle(self):
        try:
            if self._hover_compare_is_pinned():
                if getattr(self, "_last_hover_compare_key", None) is None:
                    self.hover_cmp_title_var.set("悬停完整对比 | 已固定（等待下一次悬停内容）")
                elif hasattr(self, "hover_cmp_title_var"):
                    t = str(self.hover_cmp_title_var.get() or "")
                    if "已固定" not in t:
                        self.hover_cmp_title_var.set(f"{t} | 已固定")
            else:
                if hasattr(self, "hover_cmp_title_var"):
                    t = str(self.hover_cmp_title_var.get() or "")
                    self.hover_cmp_title_var.set(t.replace(" | 已固定", ""))
        except Exception:
            pass

    def _on_hover_compare_f4_toggle(self, event=None):
        # Quick keyboard toggle for pin/unpin while reviewing long content.
        # Multiple SheetView instances may register F4 on root; only the active
        # sheet view should consume the key event.
        try:
            nb = getattr(self.app, "nb", None)
            if nb is not None:
                tab_id = nb.select()
                if tab_id:
                    tab_text = str(nb.tab(tab_id, "text") or "")
                    if tab_text != str(self.sheet):
                        return None
        except Exception:
            pass
        try:
            cur = 1 if self._hover_compare_is_pinned() else 0
            self.hover_cmp_pin_var.set(0 if cur else 1)
            self._on_hover_compare_pin_toggle()
        except Exception:
            pass
        return "break"

    def _on_hover_compare_clear_click(self):
        self._hide_hover_popup()
        self._clear_hover_compare_panel()

    def _on_hover_cmp_mousewheel(self, event):
        # Keep default behavior unless Shift is held (then treat as horizontal scroll).
        try:
            if int(getattr(event, "state", 0)) & 0x1:
                return self._on_hover_cmp_shift_wheel(event)
        except Exception:
            pass
        return None

    def _on_hover_cmp_shift_wheel(self, event):
        try:
            delta = int(getattr(event, "delta", 0))
        except Exception:
            delta = 0
        step = 0
        if delta != 0:
            step = -1 if delta > 0 else 1
        else:
            num = getattr(event, "num", None)
            if num == 4:
                step = -1
            elif num == 5:
                step = 1
        if step != 0:
            try:
                self.hover_cmp_text.xview_scroll(step * 3, "units")
            except Exception:
                pass
        return "break"

    def _cancel_hover_compare_clear(self):
        aid = getattr(self, "_hover_clear_after_id", None)
        if aid is not None:
            try:
                self.root.after_cancel(aid)
            except Exception:
                pass
        self._hover_clear_after_id = None

    def _on_hover_compare_leave(self):
        self._hide_hover_popup()
        # Keep panel content visible for manual horizontal drag/inspection.
        self._cancel_hover_compare_clear()

    def _set_hover_compare_panel(self, text: str, key):
        if not text:
            return
        self._cancel_hover_compare_clear()
        if self._hover_compare_is_pinned() and getattr(self, "_last_hover_compare_key", None) is not None:
            return
        if getattr(self, "_last_hover_compare_key", None) == key:
            return
        col_text = "-"
        try:
            if isinstance(key, tuple) and len(key) >= 4 and int(key[3]) > 0:
                ci = int(key[3])
                col_text = f"{get_column_letter(ci)}({ci})"
        except Exception:
            col_text = "-"
        try:
            if hasattr(self, "hover_cmp_title_var"):
                suffix = " | 已固定" if self._hover_compare_is_pinned() else ""
                self.hover_cmp_title_var.set(f"悬停完整对比 | Sheet: {self.sheet} | Col: {col_text}{suffix}")
            self._render_hover_compare_panel(text, key)
        except Exception:
            pass
        self._last_hover_compare_key = key

    @staticmethod
    def _hover_diff_masks(values):
        """Mark non-equal chars for each source value via pairwise sequence compare."""
        vals = ["" if v is None else str(v) for v in (values or ())]
        masks = [set() for _ in vals]
        if len(vals) <= 1:
            return masks
        for i in range(len(vals)):
            for j in range(i + 1, len(vals)):
                a = vals[i]
                b = vals[j]
                sm = difflib.SequenceMatcher(None, a, b)
                for tag, a1, a2, b1, b2 in sm.get_opcodes():
                    if tag == "equal":
                        continue
                    masks[i].update(range(a1, a2))
                    masks[j].update(range(b1, b2))
        return masks

    @staticmethod
    def _side_tag_for_hover_line(side: str, has_base: bool) -> str:
        s = (side or "").upper()
        if s == "BASE":
            return "hover_side_base"
        if s == "A":
            return "hover_side_mine" if has_base else "hover_side_base"
        if s == "B":
            return "hover_side_theirs" if has_base else "hover_side_mine"
        return "hover_side_base"

    @staticmethod
    def _side_label_for_hover_line(side: str, has_base: bool) -> str:
        s = (side or "").upper()
        if s == "BASE":
            return "base"
        if s == "A":
            return "mine" if has_base else "base"
        if s == "B":
            return "theirs" if has_base else "mine"
        return "base"

    def _render_hover_compare_panel(self, text: str, key):
        if not hasattr(self, "hover_cmp_text"):
            return
        w = self.hover_cmp_text
        try:
            w.configure(state="normal")
            w.delete("1.0", "end")
            payload = (getattr(self, "_hover_payload_cache", {}) or {}).get(key)
            if not payload:
                w.insert("1.0", text)
                w.configure(state="disabled")
                return
            sides = list(payload.get("sides") or [])
            rows = list(payload.get("rows") or [])
            values = ["" if v is None else str(v) for v in (payload.get("values") or ())]
            has_base = any((str(s).upper() == "BASE") for s in sides)
            # 新增行对侧（row_no 为 None）用空串参与 diff 计算，
            # 避免 "<missing>" 字面量被当作内容误比较。
            diff_inputs = ["" if row_no is None else v for row_no, v in zip(rows, values)]
            masks = self._hover_diff_masks(diff_inputs)

            for i, (side, row_no, val) in enumerate(zip(sides, rows, values), start=1):
                row_label = "-" if row_no is None else str(row_no)
                side_label = self._side_label_for_hover_line(side, has_base)
                prefix = f"{side_label}[{row_label}]: "
                line_txt = f"{prefix}{val}"
                if i > 1:
                    w.insert("end", "\n")
                line_start = f"{i}.0"
                w.insert("end", line_txt)
                # 无数据侧（新增行对侧）：整行置灰，不显示字符级红色高亮。
                if row_no is None:
                    w.tag_add("hover_side_missing", line_start, f"{i}.end")
                    continue
                # Color the whole source line by provenance.
                side_tag = self._side_tag_for_hover_line(side, has_base)
                w.tag_add(side_tag, line_start, f"{i}.end")

                # Color changed chars in value body.
                mask = sorted(masks[i - 1]) if (i - 1) < len(masks) else []
                if mask:
                    seg_s = None
                    seg_e = None
                    for p in mask:
                        if p < 0 or p >= len(val):
                            continue
                        if seg_s is None:
                            seg_s = p
                            seg_e = p + 1
                            continue
                        if p == seg_e:
                            seg_e = p + 1
                            continue
                        b0 = len(prefix) + seg_s
                        b1 = len(prefix) + seg_e
                        w.tag_add("hover_diffchar", f"{i}.{b0}", f"{i}.{b1}")
                        seg_s = p
                        seg_e = p + 1
                    if seg_s is not None and seg_e is not None:
                        b0 = len(prefix) + seg_s
                        b1 = len(prefix) + seg_e
                        w.tag_add("hover_diffchar", f"{i}.{b0}", f"{i}.{b1}")
            w.configure(state="disabled")
        except Exception:
            try:
                w.configure(state="normal")
                w.delete("1.0", "end")
                w.insert("1.0", text)
                w.configure(state="disabled")
            except Exception:
                pass

    def _hide_cell_tooltip(self, clear_panel: bool = True):
        # Reset hover dedup so re-entering the same cell re-triggers a refresh.
        self._last_hover_target_key = None
        self._hide_hover_popup()
        if clear_panel:
            self._clear_hover_compare_panel()
        self._cell_tip_key = None

    def _schedule_hover_panels(self, pair_idx, target_col, side, *,
                               popup_force_show=False, x_root=None, y_root=None,
                               refresh_c_area=True):
        """Throttle hover-driven panel refreshes.

        - Dedup: skip work entirely when the hovered (pair, col, side) is unchanged.
        - Debounce: coalesce rapid cross-cell motion into a single refresh.
        """
        key = (pair_idx, target_col, str(side), bool(refresh_c_area))
        # Always remember the freshest request so the debounced call uses latest position.
        self._pending_hover_args = (
            pair_idx, target_col, side, bool(popup_force_show),
            x_root, y_root, bool(refresh_c_area),
        )
        if key == getattr(self, "_last_hover_target_key", None):
            return
        self._last_hover_target_key = key
        # Coalesce bursts of Motion events: run the heavy panel refresh once when
        # the event loop next goes idle, instead of on every pixel of movement.
        aid = getattr(self, "_hover_debounce_id", None)
        if aid is not None:
            try:
                self.frame.after_cancel(aid)
            except Exception:
                pass
        try:
            self._hover_debounce_id = self.frame.after_idle(self._run_pending_hover_panels)
        except Exception:
            self._hover_debounce_id = None
            self._run_pending_hover_panels()

    def _run_pending_hover_panels(self):
        self._hover_debounce_id = None
        args = getattr(self, "_pending_hover_args", None)
        if not args:
            return
        try:
            if not self.frame.winfo_exists():
                return
        except Exception:
            return
        pair_idx, target_col, side, popup_force_show, x_root, y_root, refresh_c_area = args
        self.update_hover_driven_panels(
            pair_idx,
            target_col,
            side,
            force_panel=True,
            popup_force_show=popup_force_show,
            x_root=x_root,
            y_root=y_root,
            refresh_c_area=refresh_c_area,
        )

    def _show_cell_tooltip(self, text: str, x_root: int, y_root: int, key):
        if not text:
            self._hide_cell_tooltip(clear_panel=False)
            return
        self._set_hover_compare_panel(text, key)
        self._cell_tip_key = key
        if not bool(getattr(self, "_enable_hover_popup", False)):
            return
        try:
            lbl = getattr(self, "_cell_tip_label", None)
            if lbl is None:
                lbl = tk.Label(
                    self.root,
                    justify="left",
                    relief="solid",
                    borderwidth=1,
                    bg="#fffbe6",
                    fg="#222",
                    font=("Consolas", 10),
                    anchor="w",
                )
                self._cell_tip_label = lbl
            lbl.configure(text=text)
            self._cell_tip_win = None

            rx = int(self.root.winfo_rootx())
            ry = int(self.root.winfo_rooty())
            rw = max(1, int(self.root.winfo_width()))
            rh = max(1, int(self.root.winfo_height()))
            x = int(x_root - rx + 16)
            y = int(y_root - ry + 20)
            self.root.update_idletasks()
            tw = max(1, int(lbl.winfo_reqwidth()))
            th = max(1, int(lbl.winfo_reqheight()))
            x = max(0, min(x, max(0, rw - tw - 4)))
            y = max(0, min(y, max(0, rh - th - 4)))
            lbl.place(x=x, y=y)
            lbl.lift()
        except Exception:
            try:
                tip = tk.Toplevel(self.root)
                tip.wm_overrideredirect(True)
                try:
                    tip.wm_attributes("-topmost", True)
                except Exception:
                    pass
                tip.wm_geometry(f"+{x_root + 14}+{y_root + 18}")
                lbl = tk.Label(tip, text=text, justify="left", relief="solid", borderwidth=1, bg="#fffbe6", fg="#222", font=("Consolas", 10))
                self._cell_tip_win = tip
                self._cell_tip_key = key
                lbl.pack(ipadx=4, ipady=2)
            except Exception:
                self._hide_cell_tooltip(clear_panel=False)

    def _cmp_tooltip_payload_by_pair_col(
        self,
        pair_idx: int,
        target_col: int,
        force_show: bool = False,
        force_panel: bool = False,
    ):
        try:
            if pair_idx is None or int(pair_idx) < 0 or int(pair_idx) >= len(self.row_pairs):
                return None
            target_col = int(target_col)
            if target_col <= 0:
                return None
        except Exception:
            return None

        pair = self.row_pairs[int(pair_idx)]
        is_three = self._is_three_way_enabled()
        sides = ["BASE", "A", "B"] if is_three else ["A", "B"]
        rows = []
        for side in sides:
            if side == "BASE" and is_three:
                rows.append(self._base_row_for_pair(int(pair_idx), pair))
            else:
                rows.append(self._row_for_side(pair, side))
        ws_edit_cache = {}

        values = []
        for side, row_no in zip(sides, rows):
            if row_no is None:
                values.append("<missing>")
                continue
            try:
                if side == "A":
                    ws_val = self.app.ws_a_val(self.sheet)
                elif side == "BASE":
                    ws_val = self.app.ws_base_val(self.sheet)
                else:
                    ws_val = self.app.ws_b_val(self.sheet)
                v_disp = ws_val.cell(row=row_no, column=target_col).value
                # Keep tooltip value source aligned with row rendering:
                # when cached-values mode misses literals, rendering falls back to edit WB.
                if _USE_CACHED_VALUES_ONLY and v_disp is None:
                    ws_edit = ws_edit_cache.get(side)
                    if ws_edit is None:
                        try:
                            if side == "A":
                                ws_edit = self.app.ws_a_edit(self.sheet)
                            elif side == "BASE":
                                ws_edit = self.app.ws_base_edit(self.sheet)
                            else:
                                ws_edit = self.app.ws_b_edit(self.sheet)
                            ws_edit_cache[side] = ws_edit
                        except Exception:
                            ws_edit = None
                    if ws_edit is not None:
                        try:
                            v_edit = ws_edit.cell(row=row_no, column=target_col).value
                            if v_edit is not None and not _formula_text(v_edit):
                                v_disp = v_edit
                        except Exception:
                            pass
                if v_disp is None:
                    v_str = ""
                else:
                    v_str = str(v_disp)
                    v_str = v_str.replace("\r\n", "⏎").replace("\r", "⏎").replace("\n", "⏎")
                values.append(v_str)
            except Exception:
                return None

        width = max(1, int(self.col_char_widths.get(target_col, 1)))
        # 新增行：只要有一侧 row_no 为 None（该侧无数据），面板必须显示
        need_tip = bool(force_show) or any(row_no is None for row_no in rows) or any((row_no is not None and len(v) > width) for row_no, v in zip(rows, values))
        if not need_tip and not force_panel:
            return None

        lines = []
        for side, row_no, v in zip(sides, rows, values):
            row_label = "-" if row_no is None else str(row_no)
            side_label = self._side_label_for_hover_line(side, is_three)
            lines.append(f"{side_label}[{row_label}]: {v}")
        tip_text = "\n".join(lines)
        key = (self.sheet, "CMP", int(pair_idx), target_col, tuple(values))
        try:
            if not isinstance(getattr(self, "_hover_payload_cache", None), dict):
                self._hover_payload_cache = {}
            self._hover_payload_cache[key] = {
                "sides": tuple(sides),
                "rows": tuple(rows),
                "values": tuple(values),
            }
            # Keep cache bounded; only recent hover payloads are needed.
            if len(self._hover_payload_cache) > 128:
                for old_k in list(self._hover_payload_cache.keys())[:-64]:
                    self._hover_payload_cache.pop(old_k, None)
        except Exception:
            pass
        return tip_text, key

    def _should_force_hover_tip(self, target_col: int, rendered_fragment: str = "") -> bool:
        """Heuristics for truncated/likely-truncated cell hover in dense grid rendering."""
        try:
            frag = (rendered_fragment or "").rstrip()
            if frag.endswith("\u2026") or frag.endswith("..."):
                return True
        except Exception:
            pass
        try:
            # When a column has already hit the global display cap, rows near the
            # cap can still feel clipped in practice; force compare tooltip there.
            if int(self.col_char_widths.get(int(target_col), 0)) >= int(_COL_MAX_DISPLAY_WIDTH):
                return True
        except Exception:
            pass
        return False

    def _hit_col_from_char(self, char_no: int):
        """Map a text char position to column span; separator chars belong to the left column."""
        spans = self._spans_for_line()
        last_col = None
        last_span = None
        for c, (s, e) in spans.items():
            last_col = c
            last_span = (s, e)
            if s <= char_no < e:
                return c, s, e
            sep_end = e + _COL_SEP_LEN
            if e <= char_no < sep_end:
                return c, s, e
        if last_col is not None and last_span is not None and char_no >= last_span[1]:
            return last_col, last_span[0], last_span[1]
        return None, 0, 0

    def _on_cell_hover_tooltip(self, w: tk.Text, event, side: str):
        try:
            idx = w.index(f"@{event.x},{event.y}")
            line = int(idx.split(".")[0])
            col_char = int(idx.split(".")[1])
        except Exception:
            self._hide_cell_tooltip(clear_panel=False)
            return
        if not (1 <= line <= len(self.display_rows)):
            self._hide_cell_tooltip(clear_panel=False)
            return
        pair_idx = self._pair_idx_for_line(line)
        if pair_idx is None or pair_idx >= len(self.row_pairs):
            self._hide_cell_tooltip(clear_panel=False)
            return
        target_col, span_s, span_e = self._hit_col_from_char(col_char)
        if target_col is None:
            self._hide_cell_tooltip(clear_panel=False)
            return
        force_show = False
        try:
            line_text = w.get(f"{line}.0", f"{line}.end")
            frag = line_text[span_s:span_e]
            force_show = self._should_force_hover_tip(target_col, frag)
            try:
                tags_here = set(w.tag_names(f"{line}.{max(0, int(col_char))}"))
                if "diffcell" in tags_here:
                    force_show = True
            except Exception:
                pass
        except Exception:
            force_show = False
        self._schedule_hover_panels(
            pair_idx,
            target_col,
            side,
            popup_force_show=force_show,
            x_root=getattr(event, "x_root", None),
            y_root=getattr(event, "y_root", None),
            refresh_c_area=True,
        )

    def _active_pair_idx_for_c_area(self) -> int | None:
        return self.resolved_pair_idx_for_c_area()

    def _cursor_cmp_tooltip_payload(self, char_no: int, force_panel: bool = False):
        target_col, _s, _e = self._hit_col_from_char(char_no)
        if target_col is None:
            return None

        pair_idx = self._active_pair_idx_for_c_area()
        return self._cmp_tooltip_payload_by_pair_col(pair_idx, target_col, force_panel=force_panel)

    def _on_cursor_cmp_hover_tooltip(self, event):
        try:
            idx = self.cursor_cmp.index(f"@{event.x},{event.y}")
            idx_s = str(idx)
            line_no = int(idx_s.split(".")[0])
            char_no = int(idx_s.split(".")[1])
        except Exception:
            self._hide_cell_tooltip(clear_panel=False)
            return
        target_col, span_s, span_e = self._hit_col_from_char(char_no)
        if target_col is None:
            self._hide_cell_tooltip(clear_panel=False)
            return
        force_show = False
        try:
            line_text = self.cursor_cmp.get(f"{line_no}.0", f"{line_no}.end")
            frag = line_text[span_s:span_e]
            force_show = self._should_force_hover_tip(target_col, frag)
            try:
                tags_here = set(self.cursor_cmp.tag_names(f"{line_no}.{max(0, int(char_no))}"))
                if "diffcell" in tags_here:
                    force_show = True
            except Exception:
                pass
        except Exception:
            force_show = False
        pair_idx = self._active_pair_idx_for_c_area()
        self._schedule_hover_panels(
            pair_idx,
            target_col,
            "C",
            popup_force_show=force_show,
            x_root=getattr(event, "x_root", None),
            y_root=getattr(event, "y_root", None),
            refresh_c_area=False,
        )

    def _on_main_pane_right_click(self, w: tk.Text, event, side: str):
        self.clear_explicit_cell_selection()
        self._update_cursor_lines()
        self._update_diff_nav_state()
        self._on_cell_hover_tooltip(w, event, side)
        return "break"

    def _on_cursor_cmp_right_click(self, event):
        self.clear_explicit_cell_selection()
        self._update_cursor_lines()
        self._update_diff_nav_state()
        return "break"

    def _on_click_with_arrow(self, w: tk.Text, event, direction: str):
        # Keep horizontal position stable on click; Tk default Text binding may call see(insert).
        saved_x = 0.0
        try:
            saved_x = float((self.left.xview() or (0.0, 1.0))[0])
        except Exception:
            saved_x = 0.0

        try:
            self._click_trace_seq = int(getattr(self, "_click_trace_seq", 0)) + 1
        except Exception:
            self._click_trace_seq = 1
        seq = self._click_trace_seq
        self._trace_click_until = time.time() + 1.2
        self._log_click_trace_state(f"click_start#{seq}")

        # Select row/column first.
        self._select_from_widget(w, event)
        self._log_click_trace_state(f"after_select#{seq}")

        # Multi-stage guard: catch delayed xview writes from Tk/idle callbacks.
        try:
            self.left.after_idle(lambda sx=saved_x, st=f"idle#{seq}": self._post_click_x_guard(sx, st))
            self.left.after(30, lambda sx=saved_x, st=f"t30#{seq}": self._post_click_x_guard(sx, st))
            self.left.after(120, lambda sx=saved_x, st=f"t120#{seq}": self._post_click_x_guard(sx, st))
        except Exception:
            self._post_click_x_guard(saved_x, f"fallback#{seq}")

        try:
            idx = w.index(f"@{event.x},{event.y}")
            line = int(idx.split(".")[0])
        except Exception:
            return "break"

        if not (1 <= line <= len(self.display_rows)):
            return "break"

        pair = self._pair_for_line(line)
        r = self._row_for_side(pair, self._side_for_widget(w))
        pair_idx = self._pair_idx_for_line(line)
        cols = self.pair_diff_cols.get(pair_idx, set()) if pair_idx is not None else set()
        allow_base_row_action = (
            self._is_three_way_enabled()
            and direction == "BASE2A"
            and w is self.base
        )

        # Row overwrite entry is row-header/button driven; main data-area click should only select.
        # Return break to suppress Tk Text default click handler that can reset xview.
        if (not cols) and (not allow_base_row_action):
            return "break"
        if r is None:
            return "break"
        return "break"

    def _on_hover(self, w: tk.Text, event, direction: str):
        try:
            idx = w.index(f"@{event.x},{event.y}")
            line = int(idx.split(".")[0])
            col = int(idx.split(".")[1])
        except Exception:
            self._clear_hover(w)
            return
        if not (1 <= line <= len(self.display_rows)):
            self._clear_hover(w)
            return
        pair = self._pair_for_line(line)
        r = self._row_for_side(pair, self._side_for_widget(w))
        pair_idx = self._pair_idx_for_line(line)
        cols = self.pair_diff_cols.get(pair_idx, set()) if pair_idx is not None else set()
        allow_base_row_action = (
            self._is_three_way_enabled()
            and direction == "BASE2A"
            and w is self.base
        )
        if (not cols) and (not allow_base_row_action):
            self._clear_hover(w)
            return
        if r is None:
            self._clear_hover(w)
            return
        self._show_hover_arrow(w, line, r, direction)

    def _clear_hover(self, w: tk.Text):
        if w is self.left:
            line = self._hover_line_left
        elif w is self.base:
            line = self._hover_line_mid
        else:
            line = self._hover_line_right
        if line is None:
            return
        self._restore_rownum(w, line)
        try:
            if w is self.left:
                w.configure(cursor=self._left_cursor_default)
            elif w is self.base:
                w.configure(cursor=self._mid_cursor_default)
            else:
                w.configure(cursor=self._right_cursor_default)
        except Exception:
            pass
        if w is self.left:
            self._hover_line_left = None
        elif w is self.base:
            self._hover_line_mid = None
        else:
            self._hover_line_right = None

    def _show_hover_arrow(self, w: tk.Text, line: int, r: int, direction: str):
        if w is self.left:
            if self._hover_line_left == line:
                return
        elif w is self.base:
            if self._hover_line_mid == line:
                return
        else:
            if self._hover_line_right == line:
                return
        # restore previous
        self._clear_hover(w)
        self._replace_rownum_with_arrow(w, line, r, direction)
        try:
            w.configure(cursor="hand2")
        except Exception:
            pass
        if w is self.left:
            self._hover_line_left = line
        elif w is self.base:
            self._hover_line_mid = line
        else:
            self._hover_line_right = line

    def _replace_rownum_with_arrow(self, w: tk.Text, line: int, r: int, direction: str):
        if direction == "B2A":
            arrow = _ROW_ARROW_LEFT
        else:
            arrow = _ROW_ARROW_RIGHT
        rownum = str(r)
        new_label = arrow + (" " * max(0, len(rownum) - 1))
        start = f"{line}.0"
        end = f"{line}.{len(rownum)}"
        try:
            w.delete(start, end)
            w.insert(start, new_label)
        except Exception:
            pass

    def _restore_rownum(self, w: tk.Text, line: int):
        if not (1 <= line <= len(self.display_rows)):
            return
        pair = self._pair_for_line(line)
        r = self._row_for_side(pair, self._side_for_widget(w))
        if r is None:
            return
        rownum = str(r)
        start = f"{line}.0"
        end = f"{line}.{len(rownum)}"
        try:
            w.delete(start, end)
            w.insert(start, rownum)
        except Exception:
            pass

    def _highlight_selected_line(self, line: int):
        # Remove highlight only from the previously selected line (O(1))
        self._clear_selected_line_highlight()
        for t in (self.left, self.base, self.right):
            t.tag_add("selrow", f"{line}.0", f"{line}.end")
        self._last_selected_line = line

    def _capture_view_anchor(self):
        """Capture viewport and selection to restore after heavy refresh."""
        first = 0.0
        x_main = 0.0
        x_c = 0.0
        line = 1
        col = 0
        pair_idx = self.selected_pair_idx
        row_a = self.selected_excel_row_a
        row_b = self.selected_excel_row_b
        try:
            first = float((self.left.yview() or (0.0, 1.0))[0])
        except Exception:
            first = 0.0
        try:
            x_main = float((self.left.xview() or (0.0, 1.0))[0])
        except Exception:
            x_main = 0.0
        try:
            x_c = float((self.cursor_cmp.xview() or (0.0, 1.0))[0])
        except Exception:
            x_c = x_main
        try:
            parts = str(self.left.index("insert")).split(".")
            line = int(parts[0])
            col = int(parts[1])
        except Exception:
            line = 1
            col = 0
        return (first, x_main, x_c, line, col, pair_idx, row_a, row_b)

    def _restore_view_anchor(self, anchor):
        if not anchor:
            return
        # backward compatibility with older anchor tuple shape
        if len(anchor) >= 8:
            first, x_main, x_c, line, col, pair_idx, row_a, row_b = anchor
        else:
            first, line, pair_idx, row_a, row_b = anchor
            x_main = 0.0
            x_c = 0.0
            col = 0
        try:
            self.left.yview_moveto(first)
            if self._is_three_way_enabled():
                self.base.yview_moveto(first)
            self.right.yview_moveto(first)
        except Exception:
            pass
        try:
            self._sync_main_x_to_frac(x_main)
            self._sync_c_x_to_frac(x_main if x_c is None else x_c)
        except Exception:
            pass

        target_line = None
        # Prefer relocating by real excel row id; pair indices may shift after rescan.
        try:
            p = None
            if row_a is not None:
                p = self.row_a_to_pair_idx.get(row_a)
            if p is None and row_b is not None:
                p = self.row_b_to_pair_idx.get(row_b)
            if p is not None and p in self.row_to_line:
                target_line = self.row_to_line.get(p)
        except Exception:
            target_line = None
        try:
            if target_line is None and pair_idx is not None and pair_idx in self.row_to_line:
                target_line = self.row_to_line.get(pair_idx)
        except Exception:
            target_line = None

        if target_line is None:
            try:
                max_line = max(1, len(self.display_rows))
            except Exception:
                max_line = 1
            target_line = max(1, min(int(line or 1), max_line))

        idx = f"{target_line}.{max(0, int(col or 0))}"
        for w in (self.left, self.base, self.right):
            try:
                w.mark_set("insert", idx)
            except Exception:
                pass
        try:
            self._highlight_selected_line(target_line)
            self.selected_pair_idx = self._pair_idx_for_line(target_line)
            pair = self._pair_for_line(target_line)
            self.selected_excel_row_a = self._row_for_side(pair, "A")
            self.selected_excel_row_b = self._row_for_side(pair, "B")
            self.selected_excel_row = self.selected_excel_row_a or self.selected_excel_row_b
        except Exception:
            pass

    def _base_to_mine_diff_cols(self, row_a: int | None, row_b: int | None, max_col: int) -> set[int]:
        """Columns that differ between base and mine for the target row in 3-way mode."""
        cols: set[int] = set()
        if not self._is_three_way_enabled():
            return cols
        if not getattr(self.app, "has_base", False):
            return cols
        if self._is_missing_sheet_view():
            r = row_a if row_a is not None else row_b
        else:
            r = None
            mine_map = getattr(self, "mine_to_base_row", {}) or {}
            if row_a is not None:
                r = mine_map.get(row_a)
        if r is None:
            return cols
        try:
            ws_a = self.app.ws_a_val(self.sheet)
            ws_base = self.app.ws_base_val(self.sheet)
        except Exception:
            return cols
        for c in range(1, max_col + 1):
            try:
                va = ws_a.cell(row=r, column=c).value
                vb = ws_base.cell(row=r, column=c).value
            except Exception:
                va = None
                vb = None
            if _val_to_str(va) != _val_to_str(vb):
                cols.add(c)
        return cols

    def _update_cursor_lines(self):
        """Update compact row compare block.

        2-way: line1=mine(A), line2=theirs(B)
        3-way: line1=base, line2=mine, line3=theirs
        """
        prev_suppress_c_xsync = bool(getattr(self, "_suppress_c_xsync", False))
        self._suppress_c_xsync = True
        cursor_first = 0.0
        cell_first = 0.0
        try:
            try:
                # Keep C area aligned with main panes by default.
                cursor_first = float((self.left.xview() or (0.0, 1.0))[0])
            except Exception:
                cursor_first = 0.0
            if hasattr(self, "cell_cmp_text"):
                try:
                    cell_first = cursor_first
                except Exception:
                    cell_first = 0.0

            pair_idx = self.resolved_pair_idx_for_c_area()
            if pair_idx is not None:
                self._last_cursor_cmp_pair_idx = pair_idx
            else:
                pair_idx = self._normalize_pair_idx(getattr(self, "_last_cursor_cmp_pair_idx", None))
            pair = self.row_pairs[pair_idx] if pair_idx is not None and pair_idx < len(self.row_pairs) else None
            ra = self._row_for_side(pair, "A")
            rb = self._row_for_side(pair, "B")
            diff_cols = self._visual_diff_cols_for_pair(pair_idx) if pair_idx is not None else set()
            a_text = self.pair_text_a.get(pair_idx, "") if pair_idx is not None else ""
            b_text = self.pair_text_b.get(pair_idx, "") if pair_idx is not None else ""
            base_text = ""
            if self._is_three_way_enabled() and pair_idx is not None:
                base_text = self._build_base_line(pair_idx)

            is_three = self._is_three_way_enabled()
            self._render_cursor_row_headers(pair, is_three)
            # Force strict rendering order:
            # 2-way: mine/theirs
            # 3-way: base/mine/theirs
            self.cursor_cmp.configure(state="normal")
            self.cursor_cmp.delete("1.0", "end")
            if is_three:
                self.cursor_cmp.insert("1.0", f"{base_text}\n{a_text}\n{b_text}")
            else:
                self.cursor_cmp.insert("1.0", f"{a_text}\n{b_text}")

            # Clear & apply base tags
            self.cursor_cmp.tag_remove("a", "1.0", "end")
            self.cursor_cmp.tag_remove("base", "1.0", "end")
            self.cursor_cmp.tag_remove("b", "1.0", "end")
            self.cursor_cmp.tag_remove("missing", "1.0", "end")
            self.cursor_cmp.tag_remove("diffcell", "1.0", "end")
            self.cursor_cmp.tag_remove("cselcell", "1.0", "end")
            if is_three:
                base_r = self._base_row_for_pair(pair_idx, pair)
                if base_r is None:
                    self.cursor_cmp.tag_add("missing", "1.0", "1.end")
                else:
                    self.cursor_cmp.tag_add("base", "1.0", "1.end")
                if ra is None:
                    self.cursor_cmp.tag_add("missing", "2.0", "2.end")
                else:
                    self.cursor_cmp.tag_add("a", "2.0", "2.end")
                if rb is None:
                    self.cursor_cmp.tag_add("missing", "3.0", "3.end")
                else:
                    self.cursor_cmp.tag_add("b", "3.0", "3.end")
            else:
                if ra is None:
                    self.cursor_cmp.tag_add("missing", "1.0", "1.end")
                else:
                    self.cursor_cmp.tag_add("a", "1.0", "1.end")
                if rb is None:
                    self.cursor_cmp.tag_add("missing", "2.0", "2.end")
                else:
                    self.cursor_cmp.tag_add("b", "2.0", "2.end")

            spans_a = self._spans_for_line(a_text)
            spans_b = self._spans_for_line(b_text)
            spans_base = self._spans_for_line(base_text) if is_three else {}

            # Cell-level diff highlight
            if diff_cols:
                for c in diff_cols:
                    if is_three and c in spans_base:
                        s, e = spans_base[c]
                        self.cursor_cmp.tag_add("diffcell", f"1.{s}", f"1.{e}")
                    if c in spans_a:
                        s, e = spans_a[c]
                        self.cursor_cmp.tag_add("diffcell", f"{2 if is_three else 1}.{s}", f"{2 if is_three else 1}.{e}")
                    if c in spans_b:
                        s, e = spans_b[c]
                        self.cursor_cmp.tag_add("diffcell", f"{3 if is_three else 2}.{s}", f"{3 if is_three else 2}.{e}")

            # Keep a visible clicked-cell hint in C区 after re-render.
            try:
                sel_col = int(getattr(self, "_cursor_cmp_sel_col", 0) or 0)
                sel_line = int(getattr(self, "_cursor_cmp_sel_line", 0) or 0)
            except Exception:
                sel_col = 0
                sel_line = 0
            if sel_col > 0:
                if is_three:
                    if sel_line == 1 and sel_col in spans_base:
                        s0, e0 = spans_base[sel_col]
                        self.cursor_cmp.tag_add("cselcell", f"1.{s0}", f"1.{e0}")
                    elif sel_line == 2 and sel_col in spans_a:
                        s0, e0 = spans_a[sel_col]
                        self.cursor_cmp.tag_add("cselcell", f"2.{s0}", f"2.{e0}")
                    elif sel_line == 3 and sel_col in spans_b:
                        s0, e0 = spans_b[sel_col]
                        self.cursor_cmp.tag_add("cselcell", f"3.{s0}", f"3.{e0}")
                else:
                    if sel_line == 1 and sel_col in spans_a:
                        s0, e0 = spans_a[sel_col]
                        self.cursor_cmp.tag_add("cselcell", f"1.{s0}", f"1.{e0}")
                    elif sel_line == 2 and sel_col in spans_b:
                        s0, e0 = spans_b[sel_col]
                        self.cursor_cmp.tag_add("cselcell", f"2.{s0}", f"2.{e0}")

            # Apply selected-cell highlight on main panes (A/B and Base in 3-way).
            self._apply_main_selected_cell_highlight()

            self.cursor_cmp.configure(state="disabled")

            # ---- Update C区单元格对齐（可选） ----
            if getattr(self, "_enable_c_cell", False) and hasattr(self, "cell_cmp_text"):
                try:
                    self.cell_cmp_text.configure(state="normal")
                    self.cell_cmp_text.delete("1.0", "end")
                    self.cell_cmp_text.tag_remove("a", "1.0", "end")
                    self.cell_cmp_text.tag_remove("b", "1.0", "end")
                    self.cell_cmp_text.tag_remove("diffcell", "1.0", "end")

                    if pair is not None:
                        ws_a_val = self.app.ws_a_val(self.sheet)
                        ws_b_val = self.app.ws_b_val(self.sheet)
                        show_only_diff = bool(self.c_only_diff_cells.get())
                        cols_to_show = sorted(diff_cols) if show_only_diff else list(range(1, self.max_col + 1))

                        if show_only_diff:
                            parts_a = []
                            parts_b = []
                            widths = []
                            for c in cols_to_show:
                                va = ws_a_val.cell(row=ra, column=c).value if ra is not None else None
                                vb = ws_b_val.cell(row=rb, column=c).value if rb is not None else None
                                a_s = _val_to_str(va)
                                b_s = _val_to_str(vb)
                                parts_a.append(a_s)
                                parts_b.append(b_s)
                                widths.append(max(4, min(max(len(a_s), len(b_s)), _COL_MAX_DISPLAY_WIDTH)))

                            sep = _COL_SEP
                            trail = " \u2502"
                            a_line = sep.join(_format_cell(parts_a[i], widths[i]) for i in range(len(parts_a))) + (trail if parts_a else "")
                            b_line = sep.join(_format_cell(parts_b[i], widths[i]) for i in range(len(parts_b))) + (trail if parts_b else "")
                            self.cell_cmp_text.insert("end", a_line + "\n" + b_line + "\n")

                            self.cell_cmp_text.tag_add("a", "1.0", "1.end")
                            self.cell_cmp_text.tag_add("b", "2.0", "2.end")

                            # All shown columns are diffs (show_only_diff path) — highlight whole lines.
                            if a_line:
                                self.cell_cmp_text.tag_add("diffcell", "1.0", "1.end")
                            if b_line:
                                self.cell_cmp_text.tag_add("diffcell", "2.0", "2.end")
                        else:
                            line_no = 1
                            for c in cols_to_show:
                                va = ws_a_val.cell(row=ra, column=c).value if ra is not None else None
                                vb = ws_b_val.cell(row=rb, column=c).value if rb is not None else None
                                a_s = _val_to_str(va)
                                b_s = _val_to_str(vb)

                                self.cell_cmp_text.insert("end", a_s + "\n")
                                self.cell_cmp_text.insert("end", b_s + "\n")

                                self.cell_cmp_text.tag_add("a", f"{line_no}.0", f"{line_no}.end")
                                self.cell_cmp_text.tag_add("b", f"{line_no+1}.0", f"{line_no+1}.end")

                                if va != vb:
                                    self.cell_cmp_text.tag_add("diffcell", f"{line_no}.0", f"{line_no}.end")
                                    self.cell_cmp_text.tag_add("diffcell", f"{line_no+1}.0", f"{line_no+1}.end")

                                line_no += 2

                    self.cell_cmp_text.configure(state="disabled")
                except Exception:
                    try:
                        self.cell_cmp_text.configure(state="disabled")
                    except Exception:
                        pass

            # Restore C pane horizontal viewport without driving main panes.
            try:
                self.cursor_cmp.xview_moveto(cursor_first)
                cf, cl = self.cursor_cmp.xview()
                self.cursor_hsb.set(cf, cl)
                # Also restore column header: delete+insert resets its xview to 0.
                if getattr(self, "cursor_cmp_colhdr", None) is not None:
                    self.cursor_cmp_colhdr.xview_moveto(cursor_first)
            except Exception:
                pass
            if hasattr(self, "cell_cmp_text"):
                try:
                    self.cell_cmp_text.xview_moveto(cell_first)
                    sf, sl = self.cell_cmp_text.xview()
                    self.cell_cmp_hsb.set(sf, sl)
                except Exception:
                    pass
            self._suppress_c_xsync = prev_suppress_c_xsync
        except Exception:
            try:
                self._suppress_c_xsync = prev_suppress_c_xsync
            except Exception:
                pass
            pass

    def _copy_single_cell_by_pair(self, pair_idx: int, direction: str, c: int):
        try:
            if pair_idx is None or pair_idx >= len(self.row_pairs):
                return
            pair = self.row_pairs[pair_idx]
            ra = self._row_for_side(pair, "A")
            rb = self._row_for_side(pair, "B")
            if direction == "A2B":
                if ra is None or rb is None:
                    return
                src_r, dst_r = ra, rb
            elif direction == "BASE2A":
                if ra is None:
                    return
                src_r = self._base_row_for_pair(pair_idx, pair)
                dst_r = ra
            else:
                if rb is None:
                    return
                src_r = rb
                dst_r = ra if ra is not None else rb

            anchor = self._capture_view_anchor()
            adopted_formula_cache = False
            if direction == "A2B":
                old_edit = self.app.ws_b_edit(self.sheet).cell(row=dst_r, column=c).value
                old_val = self.app.ws_b_val(self.sheet).cell(row=dst_r, column=c).value
                v_edit = self.app.ws_a_edit(self.sheet).cell(row=src_r, column=c).value
                v_val = self.app.ws_a_val(self.sheet).cell(row=src_r, column=c).value
                new_edit = _copy_edit_value_for_destination(
                    v_val, v_edit, old_edit,
                    src_row=src_r, src_col=c, dst_row=dst_r, dst_col=c,
                )
                formula_mode = self._same_formula_copy_mode(new_edit, old_edit, v_val, old_val)
                if formula_mode == "noop":
                    return
                if formula_mode == "cache":
                    self.app.record_manual_b_formula_cache(self.sheet, dst_r, c, v_val)
                    adopted_formula_cache = True
                else:
                    self.app.clear_manual_b_formula_cache(self.sheet, dst_r, c)
                    _assign_edit_cell_value(
                        self.app.ws_b_edit(self.sheet).cell(row=dst_r, column=c),
                        new_edit,
                    )
                    self.app.record_manual_b_cell(self.sheet, dst_r, c, new_edit)
                self.app.ws_b_val(self.sheet).cell(row=dst_r, column=c).value = v_val
                self.app.modified_b = True
                self.app.modified_sheets_b.add(self.sheet)
                self.app.push_undo({"sheet": self.sheet, "target": "B", "cells": [(dst_r, c, old_edit, old_val)]})
            elif direction == "BASE2A":
                old_edit = self.app.ws_a_edit(self.sheet).cell(row=dst_r, column=c).value
                old_val = self.app.ws_a_val(self.sheet).cell(row=dst_r, column=c).value
                if src_r is None:
                    v_edit = None
                    v_val = None
                else:
                    v_edit = self.app.ws_base_edit(self.sheet).cell(row=src_r, column=c).value
                    v_val = self.app.ws_base_val(self.sheet).cell(row=src_r, column=c).value
                new_edit = (
                    _copy_edit_value_for_destination(
                        v_val, v_edit, old_edit,
                        src_row=src_r, src_col=c, dst_row=dst_r, dst_col=c,
                    )
                    if src_r is not None else None
                )
                formula_mode = self._same_formula_copy_mode(new_edit, old_edit, v_val, old_val)
                if formula_mode == "noop":
                    return
                if formula_mode == "cache":
                    new_edit = old_edit
                    self.app.record_manual_a_formula_cache(self.sheet, dst_r, c, v_val)
                    adopted_formula_cache = True
                else:
                    _assign_edit_cell_value(self.app.ws_a_edit(self.sheet).cell(row=dst_r, column=c), new_edit)
                    self.app.record_manual_a_cell(self.sheet, dst_r, c, new_edit)
                self.app.ws_a_val(self.sheet).cell(row=dst_r, column=c).value = v_val
                self.app.modified_a = True
                self.app.modified_sheets_a.add(self.sheet)
                self.app.push_undo({"sheet": self.sheet, "target": "A", "cells": [(dst_r, c, old_edit, old_val)]})
            else:
                old_edit = self.app.ws_a_edit(self.sheet).cell(row=dst_r, column=c).value
                old_val = self.app.ws_a_val(self.sheet).cell(row=dst_r, column=c).value
                v_edit = self.app.ws_b_edit(self.sheet).cell(row=src_r, column=c).value
                v_val = self.app.ws_b_val(self.sheet).cell(row=src_r, column=c).value
                new_edit = _copy_edit_value_for_destination(
                    v_val, v_edit, old_edit,
                    src_row=src_r, src_col=c, dst_row=dst_r, dst_col=c,
                )
                formula_mode = self._same_formula_copy_mode(new_edit, old_edit, v_val, old_val)
                if formula_mode == "noop":
                    return
                if formula_mode == "cache":
                    new_edit = old_edit
                    self.app.record_manual_a_formula_cache(self.sheet, dst_r, c, v_val)
                    adopted_formula_cache = True
                else:
                    _assign_edit_cell_value(self.app.ws_a_edit(self.sheet).cell(row=dst_r, column=c), new_edit)
                    self.app.record_manual_a_cell(self.sheet, dst_r, c, new_edit)
                self.app.ws_a_val(self.sheet).cell(row=dst_r, column=c).value = v_val
                self.app.modified_a = True
                self.app.modified_sheets_a.add(self.sheet)
                self.app.push_undo({"sheet": self.sheet, "target": "A", "cells": [(dst_r, c, old_edit, old_val)]})

            touched_r = ra or rb
            if touched_r is not None:
                self.touched_rows.add(touched_r)
            self._invalidate_only_diff_snapshot_cache()
            self._invalidate_render_cache()
            if bool(self.only_diff_var.get()) and self.snapshot_only_diff:
                self._recalc_row_diff_and_update(dst_r)
            self.refresh(row_only=dst_r, rescan=False)
            self._restore_view_anchor(anchor)
            self._update_cursor_lines()
            if adopted_formula_cache:
                self._show_formula_copy_skip_notice(1)
        except Exception as e:
            messagebox.showerror("Error", f"C区覆盖失败：\n{e}")

    def _on_cursor_cmp_click(self, event):
        """Single-click in C区: map clicked cell back to current pair/column selection."""
        try:
            idx = self.cursor_cmp.index(f"@{event.x},{event.y}")
            line_no = int(str(idx).split(".")[0])
            char_no = int(str(idx).split(".")[1])
        except Exception:
            return "break"

        try:
            line_text = self.cursor_cmp.get(f"{line_no}.0", f"{line_no}.end")
        except Exception:
            line_text = ""
        spans = self._spans_for_line(line_text)
        hit_col = None
        hit_char = 0
        for c, (s0, e0) in spans.items():
            if s0 <= char_no < e0:
                hit_col = c
                hit_char = s0
                break
        if hit_col is None:
            return "break"

        # Resolve current pair/line in the main panes.
        pair_idx = self.resolved_pair_idx_for_c_area()

        target_line = None
        try:
            if pair_idx is not None and pair_idx in self.row_to_line:
                target_line = int(self.row_to_line.get(pair_idx))
        except Exception:
            target_line = None
        if target_line is None:
            target_line = 1
        try:
            target_line = max(1, min(int(target_line), max(1, len(self.display_rows))))
        except Exception:
            target_line = 1

        saved_x = 0.0
        try:
            saved_x = float((self.left.xview() or (0.0, 1.0))[0])
        except Exception:
            saved_x = 0.0

        target_idx = f"{target_line}.{max(0, int(hit_char))}"
        for w in (self.left, self.base, self.right):
            try:
                w.mark_set("insert", target_idx)
            except Exception:
                pass

        try:
            self._highlight_selected_line(target_line)
            self.selected_pair_idx = self._pair_idx_for_line(target_line)
            pair = self._pair_for_line(target_line)
            self.hover_pair_idx = self._normalize_pair_idx(self.selected_pair_idx)
            self.hover_col_idx = int(hit_col)
            self.hover_side = "C"
            self.selected_excel_row_a = self._row_for_side(pair, "A")
            self.selected_excel_row_b = self._row_for_side(pair, "B")
            self.selected_excel_row = self.selected_excel_row_a or self.selected_excel_row_b
        except Exception:
            pass

        # Store C selection so _update_cursor_lines can render visible clicked-cell highlight.
        self._cursor_cmp_sel_col = int(hit_col)
        self._cursor_cmp_sel_line = int(line_no)
        self._set_main_selected_cell(target_line, hit_col)

        # Re-render C block and keep x aligned.
        self._update_cursor_lines()
        try:
            self._sync_main_x_to_frac(saved_x)
            self._sync_c_x_to_frac(saved_x)
        except Exception:
            pass
        self._update_diff_nav_state()
        return "break"

    def _on_cursor_cmp_double_click(self, event):
        try:
            idx = self.cursor_cmp.index(f"@{event.x},{event.y}")
            line_no = int(str(idx).split(".")[0])
            char_no = int(str(idx).split(".")[1])
        except Exception:
            return

        pair_idx = self.resolved_pair_idx_for_c_area()
        if pair_idx is None or pair_idx >= len(self.row_pairs):
            return

        is_three = self._is_three_way_enabled()
        if is_three:
            if line_no == 1:
                direction = "BASE2A"
                diff_cols = self._visual_diff_cols_for_pair(pair_idx)
            elif line_no == 3:
                direction = "B2A"
                diff_cols = set(self.pair_diff_cols.get(pair_idx, set()))
            else:
                return
        else:
            if line_no == 2:
                direction = "B2A"
                diff_cols = set(self.pair_diff_cols.get(pair_idx, set()))
            else:
                return
        if not diff_cols:
            return

        line_text = self.cursor_cmp.get(f"{line_no}.0", f"{line_no}.end")
        spans = self._spans_for_line(line_text)
        hit_col = None
        for c, (s, e) in spans.items():
            if s <= char_no < e:
                hit_col = c
                break
        if hit_col is None or hit_col not in diff_cols:
            return
        self._copy_single_cell_by_pair(pair_idx, direction, hit_col)

    def _set_copy_scope_mode(self, mode: str):
        mode_norm = str(mode or "").strip().lower()
        mode_norm = "region" if mode_norm == "region" else "row"
        self._copy_scope_mode = mode_norm
        try:
            self._copy_scope_var.set(mode_norm)
        except Exception:
            pass
        self._refresh_copy_scope_buttons()

    def _refresh_copy_scope_buttons(self):
        if self._is_missing_sheet_view():
            left_text = "使用左侧Sheet" if self._is_three_way_enabled() else "采用左侧Sheet"
            right_text = "使用右侧Sheet" if self._is_three_way_enabled() else "采用右侧Sheet"
            try:
                self.use_left_btn.configure(text=left_text)
            except Exception:
                pass
            try:
                self.use_right_btn.configure(text=right_text)
            except Exception:
                pass
            return
        mode = getattr(self, "_copy_scope_mode", "row")
        if mode == "region":
            left_text = "使用左侧区域"
            right_text = "使用右侧区域"
        else:
            left_text = "使用左侧行"
            right_text = "使用右侧行"
        try:
            self.use_left_btn.configure(text=left_text)
        except Exception:
            pass
        try:
            self.use_right_btn.configure(text=right_text)
        except Exception:
            pass

    def _run_copy_action_by_mode(self, direction: str):
        if self._is_missing_sheet_view():
            self._copy_missing_sheet(direction)
            return
        mode = getattr(self, "_copy_scope_mode", "row")
        if mode == "region":
            self._copy_selected_region(direction)
        else:
            self._copy_selected_row(direction)

    def _copy_missing_sheet(self, direction: str):
        meta = self._sheet_meta()
        action_text = {
            "A2B": "正在复制左侧整张 Sheet...",
            "B2A": "正在复制 theirs 整张 Sheet 到 mine...",
            "BASE2A": "正在按 Base 恢复或删除整张 Sheet...",
        }.get(direction, "正在处理整张 Sheet...")
        try:
            self.info.configure(text=action_text)
            self.root.configure(cursor="watch")
            self.root.update_idletasks()
        except Exception:
            pass
        try:
            if direction == "A2B":
                if meta.get("has_a") and (not meta.get("has_b")):
                    self.app._copy_sheet_between_sides(self.sheet, "A", "B")
            elif direction == "B2A":
                if meta.get("has_b") and (not meta.get("has_a")):
                    self.app._copy_sheet_between_sides(self.sheet, "B", "A")
            elif direction == "BASE2A":
                if meta.get("has_base"):
                    self.app._copy_sheet_between_sides(self.sheet, "BASE", "A")
                elif meta.get("has_a"):
                    self.app._delete_sheet_on_side(self.sheet, "A")
            elif direction == "MINE2A":
                return
            else:
                return
            self._data_ready = False
            self._bounds_checked = False
            self._invalidate_only_diff_snapshot_cache()
            self._invalidate_render_cache()
            self._update_sheet_role_labels()
            self.app.refresh_sheet_nav()
            self.refresh(row_only=None, rescan=True)
            self._update_cursor_lines()
        except Exception as e:
            messagebox.showerror("Error", f"整Sheet操作失败：\n{e}")
        finally:
            try:
                self.root.configure(cursor="")
            except Exception:
                pass

    def _update_merge_buttons_for_row(self, excel_row: int):
        # Buttons are always visible; no UI updates needed.
        return

    # ---------- Diff block navigation ----------
    def _compute_diff_blocks(self):
        """Return list of (start_line, end_line) diff blocks in current view."""
        blocks = []
        start = None
        previous_diff_pair_idx = None
        for line_idx, pair_idx in enumerate(self.display_rows, start=1):
            has = self._pair_has_visual_diff(pair_idx)
            if has:
                # In only-diff mode, adjacent display lines can be thousands of
                # worksheet rows apart. They must remain separate regions.
                if start is None:
                    start = line_idx
                elif previous_diff_pair_idx is None or int(pair_idx) != int(previous_diff_pair_idx) + 1:
                    blocks.append((start, line_idx - 1))
                    start = line_idx
                previous_diff_pair_idx = pair_idx
            elif (not has) and start is not None:
                blocks.append((start, line_idx - 1))
                start = None
                previous_diff_pair_idx = None
        if start is not None:
            blocks.append((start, len(self.display_rows)))
        self._diff_blocks_cache = blocks
        return blocks

    def _current_line(self) -> int:
        try:
            return int(self.left.index("insert").split(".")[0])
        except Exception:
            return 1

    def _current_diff_block_for_line(self, line: int):
        blocks = self._diff_blocks_cache if self._diff_blocks_cache is not None else self._compute_diff_blocks()
        for start, end in blocks:
            if start <= line <= end:
                return (start, end)
        return None

    def _logical_diff_pair_block_for_line(self, line: int) -> list[int]:
        """Expand the selected diff to its full pair block, beyond render limits."""
        if not (1 <= int(line) <= len(self.display_rows)):
            return []
        anchor_pair_idx = int(self.display_rows[int(line) - 1])
        if not self._pair_has_visual_diff(anchor_pair_idx):
            return []
        start_pair_idx = anchor_pair_idx
        end_pair_idx = anchor_pair_idx
        while start_pair_idx > 0 and self._pair_has_visual_diff(start_pair_idx - 1):
            start_pair_idx -= 1
        last_pair_idx = len(self.row_pairs) - 1
        while end_pair_idx < last_pair_idx and self._pair_has_visual_diff(end_pair_idx + 1):
            end_pair_idx += 1
        return list(range(start_pair_idx, end_pair_idx + 1))

    def _copy_selected_region(self, direction: str):
        """Copy contiguous diff block around current line using diff-cell columns only."""
        started = time.perf_counter()
        processed_rows = 0
        changed_any = False
        previous_bg_suppression = bool(getattr(self, "_suppress_bg_apply", False))
        begin_interactive = getattr(self.app, "_begin_interactive_action", None)
        end_interactive = getattr(self.app, "_end_interactive_action", None)
        if callable(begin_interactive):
            begin_interactive()
        self._suppress_bg_apply = True
        direction_text = {
            "B2A": "右侧区域到 mine",
            "A2B": "左侧区域到 theirs",
            "BASE2A": "Base 区域到 mine",
        }.get(direction, direction)
        try:
            formula_skip_before = int(getattr(self, "_formula_copy_skips_pending", 0))
            line = self._current_line()
            region_pair_indices = self._logical_diff_pair_block_for_line(line)
            if not region_pair_indices:
                _dlog(f"OVERWRITE_REGION_NO_BLOCK sheet={self.sheet} dir={direction} line={line}")
                return
            total_region_rows = len(region_pair_indices)
            _dlog(
                f"OVERWRITE_REGION_START sheet={self.sheet} dir={direction} "
                f"pairs={region_pair_indices[0]}-{region_pair_indices[-1]} rows={total_region_rows}"
            )
            try:
                self.info.configure(text=f"正在采用{direction_text}：0/{total_region_rows} 行...")
                self.root.configure(cursor="watch")
                self.root.update_idletasks()
            except Exception:
                pass
            # Validate the whole region before writing the first cell so a later
            # unsupported multi-cell formula cannot leave a half-applied region.
            self._preflight_region_formula_copy(direction, region_pair_indices)
            anchor = self._capture_view_anchor()
            # Collect all undo cells into one list so the entire region is a
            # single undo entry regardless of how many rows it spans.
            undo_cells_region: list = []
            undo_target = "A" if direction in ("B2A", "BASE2A") else "B"
            # The common large-formula case is thousands of already aligned
            # rows copied from theirs to mine. Resolve worksheet objects once
            # and bypass the full single-row command setup for every row.
            fast_b2a = direction == "B2A"
            ws_a_val_fast = None
            ws_b_val_fast = None
            ws_a_edit_fast = None
            ws_b_edit_fast = None
            if fast_b2a:
                ws_a_val_fast = self.app.ws_a_val(self.sheet)
                ws_b_val_fast = self.app.ws_b_val(self.sheet)
                ws_a_edit_fast = self.app.ws_a_edit(self.sheet)
                ws_b_edit_fast = self.app.ws_b_edit(self.sheet)
            region_pos = 0
            while region_pos < total_region_rows:
                pair_idx = region_pair_indices[region_pos]
                cols = set(self.pair_diff_cols.get(pair_idx, set()))
                if not cols:
                    region_pos += 1
                    continue
                pair = self.row_pairs[pair_idx] if pair_idx < len(self.row_pairs) else None
                ra = self._row_for_side(pair, "A")
                rb = self._row_for_side(pair, "B")

                if direction == "B2A" and ra is None and rb is not None and cols == {-1}:
                    run: list[tuple[int, int]] = [(pair_idx, rb)]
                    probe = region_pos + 1
                    prev_rb = rb
                    while probe < total_region_rows:
                        next_pair_idx = region_pair_indices[probe]
                        next_cols = set(self.pair_diff_cols.get(next_pair_idx, set()))
                        next_pair = self.row_pairs[next_pair_idx] if next_pair_idx < len(self.row_pairs) else None
                        next_ra = self._row_for_side(next_pair, "A")
                        next_rb = self._row_for_side(next_pair, "B")
                        if next_cols != {-1} or next_ra is not None or next_rb is None:
                            break
                        if int(next_rb) != int(prev_rb) + 1:
                            break
                        run.append((next_pair_idx, next_rb))
                        prev_rb = next_rb
                        probe += 1
                    row_changed = self._batch_insert_row_copy(
                        run,
                        direction="B2A",
                        suppress_refresh=True,
                        anchor=anchor,
                    )
                    if row_changed:
                        changed_any = True
                    processed_rows += len(run)
                    region_pos = probe
                    continue

                if direction == "A2B" and rb is None and ra is not None and cols == {-1}:
                    run = [(pair_idx, ra)]
                    probe = region_pos + 1
                    prev_ra = ra
                    while probe < total_region_rows:
                        next_pair_idx = region_pair_indices[probe]
                        next_cols = set(self.pair_diff_cols.get(next_pair_idx, set()))
                        next_pair = self.row_pairs[next_pair_idx] if next_pair_idx < len(self.row_pairs) else None
                        next_ra = self._row_for_side(next_pair, "A")
                        next_rb = self._row_for_side(next_pair, "B")
                        if next_cols != {-1} or next_rb is not None or next_ra is None:
                            break
                        if int(next_ra) != int(prev_ra) + 1:
                            break
                        run.append((next_pair_idx, next_ra))
                        prev_ra = next_ra
                        probe += 1
                    row_changed = self._batch_insert_row_copy(
                        run,
                        direction="A2B",
                        suppress_refresh=True,
                        anchor=anchor,
                    )
                    if row_changed:
                        changed_any = True
                    processed_rows += len(run)
                    region_pos = probe
                    continue

                if (
                    fast_b2a
                    and ra is not None
                    and rb is not None
                    and cols
                    and all(int(c) > 0 for c in cols)
                ):
                    applied_cols = set()
                    row_undo = []
                    for c in sorted(int(c) for c in cols):
                        old_edit = ws_a_edit_fast.cell(row=ra, column=c).value
                        old_val = ws_a_val_fast.cell(row=ra, column=c).value
                        source_edit = ws_b_edit_fast.cell(row=rb, column=c).value
                        source_val = ws_b_val_fast.cell(row=rb, column=c).value
                        new_edit = _copy_edit_value_for_destination(
                            source_val, source_edit, old_edit,
                            src_row=rb, src_col=c, dst_row=ra, dst_col=c,
                        )
                        formula_mode = self._same_formula_copy_mode(
                            new_edit, old_edit, source_val, old_val
                        )
                        if formula_mode == "noop":
                            continue
                        if formula_mode == "cache":
                            ws_a_val_fast.cell(row=ra, column=c).value = source_val
                            self.app.record_manual_a_formula_cache(self.sheet, ra, c, source_val)
                        else:
                            _assign_edit_cell_value(ws_a_edit_fast.cell(row=ra, column=c), new_edit)
                            ws_a_val_fast.cell(row=ra, column=c).value = source_val
                            self.app.record_manual_a_cell(self.sheet, ra, c, new_edit)
                        row_undo.append((ra, c, old_edit, old_val))
                        applied_cols.add(c)
                    if row_undo:
                        undo_cells_region.extend(row_undo)
                        self.app.modified_a = True
                        self.app.modified_sheets_a.add(self.sheet)
                        self.touched_rows.add(int(ra))
                        changed_any = True
                    if getattr(self.app, "merge_conflict_mode", False) and applied_cols:
                        self.app.user_touched_conflicts = True
                        self._resolve_conflict_row(ra, applied_cols)
                    processed_rows += 1
                    if processed_rows % 200 == 0:
                        try:
                            self.info.configure(
                                text=f"正在采用{direction_text}：{processed_rows}/{total_region_rows} 行..."
                            )
                            self.root.update_idletasks()
                        except Exception:
                            pass
                    region_pos += 1
                    continue
                # In 3-way mode, "采用Base" is kept as its own button.
                # Left/Right region actions follow left/right semantics only.
                row_changed = self._copy_selected_row(
                    direction,
                    row_header=False,
                    override_pair_idx=pair_idx,
                    override_cols=cols,
                    suppress_refresh=True,
                    _undo_out=undo_cells_region,
                )
                if row_changed:
                    changed_any = True
                processed_rows += 1
                if processed_rows % 200 == 0:
                    try:
                        self.info.configure(
                            text=f"正在采用{direction_text}：{processed_rows}/{total_region_rows} 行..."
                        )
                        self.root.update_idletasks()
                    except Exception:
                        pass
                region_pos += 1
            if changed_any:
                if undo_cells_region:
                    self.app.push_undo({"sheet": self.sheet, "target": undo_target, "cells": undo_cells_region})
                self._invalidate_only_diff_snapshot_cache()
                self._invalidate_render_cache()
                # pair_text_a/b were not updated during suppress_refresh loop;
                # clear them so refresh rebuilds each row from the new cell values.
                self.pair_text_a = {}
                self.pair_text_b = {}
                self.refresh(row_only=None, rescan=False)
                self._restore_view_anchor(anchor)
                self._update_cursor_lines()
                try:
                    elapsed = time.perf_counter() - started
                    self.info.configure(
                        text=f"已采用{direction_text}：{processed_rows} 行，耗时 {elapsed:.1f} 秒"
                    )
                except Exception:
                    pass
            formula_skipped = int(getattr(self, "_formula_copy_skips_pending", 0)) - formula_skip_before
            if formula_skipped > 0:
                self._show_formula_copy_skip_notice(formula_skipped)
        except Exception as e:
            messagebox.showerror("Error", f"覆盖区域失败：\n{e}")
        finally:
            self._suppress_bg_apply = previous_bg_suppression
            if callable(end_interactive):
                end_interactive()
            try:
                self.root.configure(cursor="")
            except Exception:
                pass
            _dlog(
                f"OVERWRITE_REGION_END sheet={self.sheet} dir={direction} "
                f"processed={processed_rows} changed={int(bool(changed_any))} "
                f"ms={(time.perf_counter() - started) * 1000.0:.1f}"
            )

    def _update_diff_nav_state(self):
        blocks = self._compute_diff_blocks()
        if not blocks:
            self.prev_diff_btn.configure(state="disabled")
            self.next_diff_btn.configure(state="disabled")
            return

        cur = self._current_line()
        has_prev = any(b[1] < cur for b in blocks)  # only enable when a block ends before cursor
        has_next = any(b[0] > cur for b in blocks)
        self.prev_diff_btn.configure(state=("normal" if has_prev else "disabled"))
        self.next_diff_btn.configure(state=("normal" if has_next else "disabled"))

    def _goto_block_start(self, start_line: int):
        # Scroll so the line is visible
        try:
            # Preserve horizontal position: see() would reset xview to column 0.
            saved_x = 0.0
            try:
                saved_x = float((self.left.xview() or (0.0, 1.0))[0])
            except Exception:
                pass
            for w in (self.left, self.right):
                w.mark_set("insert", f"{start_line}.0")
                w.see(f"{start_line}.0")
            # Restore horizontal position after see() reset it.
            try:
                self._sync_main_x_to_frac(saved_x)
                self._sync_c_x_to_frac(saved_x)
            except Exception:
                pass
            self._highlight_selected_line(start_line)
            pair = self._pair_for_line(start_line)
            self.selected_pair_idx = self._pair_idx_for_line(start_line)
            self.selected_excel_row_a = self._row_for_side(pair, "A")
            self.selected_excel_row_b = self._row_for_side(pair, "B")
            self.selected_excel_row = self.selected_excel_row_a or self.selected_excel_row_b
            self._set_main_selected_cell(start_line, None)
            self._update_cursor_lines()
        except Exception:
            pass
        self._update_diff_nav_state()

    def _goto_next_diff_block(self):
        blocks = self._compute_diff_blocks()
        cur = self._current_line()
        for start, _end in blocks:
            if start > cur:
                self._goto_block_start(start)
                return
        self._update_diff_nav_state()

    def _goto_prev_diff_block(self):
        blocks = self._compute_diff_blocks()
        cur = self._current_line()
        prev = None
        for start, end in blocks:
            if end < cur:
                prev = start
            elif start >= cur:
                break
        if prev is not None:
            self._goto_block_start(prev)
        self._update_diff_nav_state()

    # ---------- Diff calculation helpers ----------
    def _get_row_values(self, ws, r: int):
        # Fast row read using iter_rows(values_only=True)
        try:
            row = next(ws.iter_rows(min_row=r, max_row=r, min_col=1, max_col=self.max_col, values_only=True))
        except StopIteration:
            row = ()
        if row is None:
            row = ()
        # Ensure length == max_col
        if len(row) < self.max_col:
            row = tuple(row) + (None,) * (self.max_col - len(row))
        return row

    def _show_loading(self, message: str = "正在后台计算差异，请稍候..."):
        """Show a loading placeholder while background diff computation is in progress."""
        try:
            if not self.loading_progress.winfo_manager():
                self.loading_progress.pack(side="left", padx=(10, 0))
            self.loading_progress.start(12)
            for w in (self.left, self.right):
                w.configure(state="normal")
                w.delete("1.0", "end")
                w.insert("1.0", "正在加载并计算差异...\n")
            for w in (self.left_ln, self.base_ln, self.right_ln):
                w.configure(state="normal")
                w.delete("1.0", "end")
                w.insert("1.0", "\n")
                w.configure(state="disabled")
            self.info.configure(text=message)
        except Exception:
            pass

    def _hide_loading(self):
        try:
            self.loading_progress.stop()
            self.loading_progress.pack_forget()
        except Exception:
            pass

    @staticmethod
    def _row_label(r: int | None) -> str:
        return str(r) if r is not None else ""

    def _build_line_from_row_label(self, label: str, row_vals) -> str:
        parts = [_val_to_str(v) for v in row_vals]
        parts = self._gridify_parts(parts)
        return "\t".join(parts)

    def _build_row_and_diff_pair(self, ws_a_val, ws_b_val, ws_a_edit, ws_b_edit, ra: int | None, rb: int | None):
        """Build display lines and diff columns for a row pair.
        col_char_widths must be pre-populated before calling (see _prescan_col_widths)."""
        raw_a = []
        raw_b = []
        cols = set()
        for c in range(1, self.max_col + 1):
            da, db, eq = _cell_display_and_equal_by_row(ws_a_val, ws_b_val, ws_a_edit, ws_b_edit, ra, rb, c)
            raw_a.append(_val_to_str(da))
            raw_b.append(_val_to_str(db))
            if not eq:
                cols.add(c)
        # 新增行/删除行：整行只存在于一侧，视为行级差异。
        # 用哨兵值 -1 确保空行也出现在"只看差异"中，同时渲染层不对任一侧施加单元格红色高亮。
        if (ra is None) != (rb is None):
            cols = {-1}
        grid_on = self._is_grid_overlay_enabled()
        sep = _COL_SEP if grid_on else "   "
        trail = " \u2502" if grid_on else ""
        cells_a = sep.join(_format_cell(raw_a[i], self.col_char_widths.get(i + 1, 1)) for i in range(len(raw_a))) + trail
        cells_b = sep.join(_format_cell(raw_b[i], self.col_char_widths.get(i + 1, 1)) for i in range(len(raw_b))) + trail
        line_a = cells_a
        line_b = cells_b
        return line_a, line_b, cols

    def _render_line_from_raw_parts(self, raw_parts: list[str]) -> str:
        grid_on = self._is_grid_overlay_enabled()
        sep = _COL_SEP if grid_on else "   "
        trail = " \u2502" if grid_on else ""
        return sep.join(
            _format_cell(raw_parts[i], self.col_char_widths.get(i + 1, 1))
            for i in range(len(raw_parts))
        ) + trail

    def _raw_parts_from_row_values(self, row_vals) -> list[str]:
        row_vals = _pad_row_values(row_vals, self.max_col)
        return [_val_to_str(v) for v in row_vals]

    def _quick_diff_cols_from_value_rows(self, row_left_vals, row_right_vals) -> tuple[set[int], bool]:
        row_left_vals = _pad_row_values(row_left_vals, self.max_col)
        row_right_vals = _pad_row_values(row_right_vals, self.max_col)
        if row_left_vals == row_right_vals:
            return set(), False
        cols: set[int] = set()
        need_exact = False
        for offset in range(self.max_col):
            vl = row_left_vals[offset]
            vr = row_right_vals[offset]
            if _merge_cmp_value(vl) != _merge_cmp_value(vr):
                cols.add(offset + 1)
                if (vl is None) != (vr is None):
                    need_exact = True
        return cols, need_exact

    def _edit_row_values_cached(self, ws_edit, row_idx: int | None, cache: dict[int, tuple]) -> tuple:
        if row_idx is None or ws_edit is None:
            return (None,) * self.max_col
        row_idx = int(row_idx)
        cached = cache.get(row_idx)
        if cached is not None:
            return cached
        try:
            row = next(
                ws_edit.iter_rows(
                    min_row=row_idx,
                    max_row=row_idx,
                    min_col=1,
                    max_col=self.max_col,
                    values_only=True,
                ),
                (),
            )
        except Exception:
            row = ()
        cached = _pad_row_values(row, self.max_col)
        cache[row_idx] = cached
        return cached

    def _build_row_and_diff_pair_from_values(
        self,
        row_a_vals,
        row_b_vals,
        *,
        ra: int | None,
        rb: int | None,
        ws_a_edit=None,
        ws_b_edit=None,
        edit_cache_a=None,
        edit_cache_b=None,
        row_a_edit_vals=None,
        row_b_edit_vals=None,
    ):
        raw_a, raw_b, cols = self._build_row_parts_and_diff_pair_from_values(
            row_a_vals,
            row_b_vals,
            ra=ra,
            rb=rb,
            ws_a_edit=ws_a_edit,
            ws_b_edit=ws_b_edit,
            edit_cache_a=edit_cache_a,
            edit_cache_b=edit_cache_b,
            row_a_edit_vals=row_a_edit_vals,
            row_b_edit_vals=row_b_edit_vals,
        )
        return self._render_line_from_raw_parts(raw_a), self._render_line_from_raw_parts(raw_b), cols

    def _build_row_parts_and_diff_pair_from_values(
        self,
        row_a_vals,
        row_b_vals,
        *,
        ra: int | None,
        rb: int | None,
        ws_a_edit=None,
        ws_b_edit=None,
        edit_cache_a=None,
        edit_cache_b=None,
        row_a_edit_vals=None,
        row_b_edit_vals=None,
    ):
        row_a_vals = _pad_row_values(row_a_vals, self.max_col)
        row_b_vals = _pad_row_values(row_b_vals, self.max_col)
        if row_a_edit_vals is None:
            row_a_edit_vals = self._edit_row_values_cached(ws_a_edit, ra, edit_cache_a if edit_cache_a is not None else {}) if ws_a_edit is not None else (None,) * self.max_col
        else:
            row_a_edit_vals = _pad_row_values(row_a_edit_vals, self.max_col)
        if row_b_edit_vals is None:
            row_b_edit_vals = self._edit_row_values_cached(ws_b_edit, rb, edit_cache_b if edit_cache_b is not None else {}) if ws_b_edit is not None else (None,) * self.max_col
        else:
            row_b_edit_vals = _pad_row_values(row_b_edit_vals, self.max_col)

        raw_a: list[str] = []
        raw_b: list[str] = []
        cols: set[int] = set()
        for offset in range(self.max_col):
            col_idx = offset + 1
            edit_b = row_b_edit_vals[offset]
            if ra is not None and rb is not None and ra != rb:
                edit_b = _translate_normal_formula_for_compare(
                    row_b_vals[offset], edit_b, rb, col_idx, ra, col_idx
                )
            da, db, eq = _cell_display_and_equal_from_values(
                row_a_vals[offset],
                row_b_vals[offset],
                row_a_edit_vals[offset],
                edit_b,
            )
            raw_a.append(_val_to_str(da))
            raw_b.append(_val_to_str(db))
            if not eq:
                cols.add(offset + 1)
        if (ra is None) != (rb is None):
            cols = {-1}
        return raw_a, raw_b, cols

    def _compute_base_diff_cols_from_values(
        self,
        row_a_vals,
        row_base_vals,
        *,
        ra: int | None,
        base_row: int | None,
        ws_a_edit=None,
        ws_base_edit=None,
        edit_cache_a=None,
        edit_cache_base=None,
        row_a_edit_vals=None,
        row_base_edit_vals=None,
    ) -> set[int]:
        if ra is None:
            return set()
        if base_row is None:
            return {-1}
        row_a_vals = _pad_row_values(row_a_vals, self.max_col)
        row_base_vals = _pad_row_values(row_base_vals, self.max_col)
        if row_a_edit_vals is None:
            row_a_edit_vals = self._edit_row_values_cached(ws_a_edit, ra, edit_cache_a if edit_cache_a is not None else {}) if ws_a_edit is not None else (None,) * self.max_col
        else:
            row_a_edit_vals = _pad_row_values(row_a_edit_vals, self.max_col)
        if row_base_edit_vals is None:
            row_base_edit_vals = self._edit_row_values_cached(ws_base_edit, base_row, edit_cache_base if edit_cache_base is not None else {}) if ws_base_edit is not None else (None,) * self.max_col
        else:
            row_base_edit_vals = _pad_row_values(row_base_edit_vals, self.max_col)
        cols: set[int] = set()
        for offset in range(self.max_col):
            col_idx = offset + 1
            edit_base = row_base_edit_vals[offset]
            if ra != base_row:
                edit_base = _translate_normal_formula_for_compare(
                    row_base_vals[offset], edit_base, base_row, col_idx, ra, col_idx
                )
            _va, _vb, eq = _cell_display_and_equal_from_values(
                row_a_vals[offset],
                row_base_vals[offset],
                row_a_edit_vals[offset],
                edit_base,
            )
            if not eq:
                cols.add(offset + 1)
        return cols

    def _compute_base_diff_cols_for_pair(
        self,
        pair_idx: int,
        pair: tuple[int | None, int | None] | None = None,
        *,
        max_col: int | None = None,
        ws_a_val=None,
        ws_a_edit=None,
        ws_base_val=None,
        ws_base_edit=None,
    ) -> set[int]:
        if not self._is_three_way_enabled():
            return set()
        if not getattr(self.app, "has_base", False):
            return set()
        if self._is_missing_sheet_view():
            return set()
        if pair is None:
            if not (0 <= pair_idx < len(self.row_pairs)):
                return set()
            pair = self.row_pairs[pair_idx]
        ra, _rb = pair
        if ra is None:
            return set()
        base_row = self._base_row_for_pair(pair_idx, pair)
        if base_row is None:
            return {-1}
        if ws_a_val is None:
            ws_a_val = self.app.ws_a_val(self.sheet)
        if ws_a_edit is None:
            try:
                ws_a_edit = self.app.ws_a_edit(self.sheet)
            except Exception:
                ws_a_edit = ws_a_val
        if ws_base_val is None:
            ws_base_val = self.app.ws_base_val(self.sheet)
        if ws_base_edit is None:
            try:
                ws_base_edit = self.app.ws_base_edit(self.sheet)
            except Exception:
                ws_base_edit = ws_base_val
        full_max_col = max(
            int(max_col or 1),
            ws_a_val.max_column or 1,
            ws_base_val.max_column or 1,
            ws_a_edit.max_column or 1,
            ws_base_edit.max_column or 1,
        )
        cols: set[int] = set()
        for c in range(1, full_max_col + 1):
            _va, _vb, eq = _cell_display_and_equal_by_row(
                ws_a_val,
                ws_base_val,
                ws_a_edit,
                ws_base_edit,
                ra,
                base_row,
                c,
            )
            if not eq:
                cols.add(c)
        return cols

    def _ensure_base_diff_cache(
        self,
        *,
        pair_indices=None,
        max_col: int | None = None,
        ws_a_val=None,
        ws_a_edit=None,
        ws_base_val=None,
        ws_base_edit=None,
    ):
        if not self._is_three_way_enabled():
            self.pair_base_diff_cols = {}
            return False
        if not getattr(self.app, "has_base", False):
            self.pair_base_diff_cols = {}
            return False
        if self._is_missing_sheet_view():
            self.pair_base_diff_cols = {}
            return False
        if ws_a_val is None:
            ws_a_val = self.app.ws_a_val(self.sheet)
        if ws_a_edit is None:
            try:
                ws_a_edit = self.app.ws_a_edit(self.sheet)
            except Exception:
                ws_a_edit = ws_a_val
        if ws_base_val is None:
            ws_base_val = self.app.ws_base_val(self.sheet)
        if ws_base_edit is None:
            try:
                ws_base_edit = self.app.ws_base_edit(self.sheet)
            except Exception:
                ws_base_edit = ws_base_val
        if pair_indices is None:
            targets = range(len(self.row_pairs))
        else:
            targets = [int(idx) for idx in pair_indices if 0 <= int(idx) < len(self.row_pairs)]
        added_any = False
        for idx in targets:
            if idx in self.pair_base_diff_cols:
                continue
            self.pair_base_diff_cols[idx] = self._compute_base_diff_cols_for_pair(
                idx,
                self.row_pairs[idx],
                max_col=max_col,
                ws_a_val=ws_a_val,
                ws_a_edit=ws_a_edit,
                ws_base_val=ws_base_val,
                ws_base_edit=ws_base_edit,
            )
            added_any = True
        return added_any

    def _visual_diff_cols_for_pair(self, pair_idx: int) -> set[int]:
        cols = set(self.pair_diff_cols.get(pair_idx, set()))
        base_cols = set(self.pair_base_diff_cols.get(pair_idx, set()))
        if base_cols:
            if -1 in base_cols:
                cols.add(-1)
            cols.update(c for c in base_cols if c > 0)
        return cols

    def _pair_has_visual_diff(self, pair_idx: int) -> bool:
        return bool(self._visual_diff_cols_for_pair(pair_idx))

    def _prescan_col_widths(self, ws_a_val, ws_b_val, ws_base_val=None, max_pairs: int = 0):
        """Quick first-pass scan to populate col_char_widths before building formatted lines.
        max_pairs>0 limits scanning to first N pairs (for large sheets)."""
        self.col_char_widths = {}
        self._rownum_display_width = 0
        pairs = self.row_pairs[:max_pairs] if max_pairs > 0 else self.row_pairs
        rows_a_cache = _read_rows_into_cache(ws_a_val, [ra for ra, _rb in pairs if ra is not None], self.max_col)
        rows_b_cache = _read_rows_into_cache(ws_b_val, [rb for _ra, rb in pairs if rb is not None], self.max_col)
        base_rows_needed = []
        if ws_base_val is not None:
            for idx, (ra, rb) in enumerate(pairs):
                r_base = self._base_row_for_pair(idx, (ra, rb))
                if r_base is not None:
                    base_rows_needed.append(r_base)
        rows_base_cache = _read_rows_into_cache(ws_base_val, base_rows_needed, self.max_col) if ws_base_val is not None else {}

        for idx, (ra, rb) in enumerate(pairs):
            row_a = _row_from_cache(rows_a_cache, ra, self.max_col)
            row_b = _row_from_cache(rows_b_cache, rb, self.max_col)
            r_base = self._base_row_for_pair(idx, (ra, rb))
            row_base = _row_from_cache(rows_base_cache, r_base, self.max_col) if ws_base_val is not None else None
            for c in range(1, self.max_col + 1):
                sa = _val_to_str(row_a[c - 1])
                sb = _val_to_str(row_b[c - 1])
                w = min(max(len(sa), len(sb)), _COL_MAX_DISPLAY_WIDTH)
                if row_base is not None and r_base is not None:
                    sv = _val_to_str(row_base[c - 1])
                    w = min(max(w, len(sv)), _COL_MAX_DISPLAY_WIDTH)
                if w > self.col_char_widths.get(c, 0):
                    self.col_char_widths[c] = w
        # Avoid ultra-narrow columns (which render as repeated "...").
        # Keep a readable lower bound after grid on/off toggles.
        for c in range(1, self.max_col + 1):
            self.col_char_widths[c] = max(4, int(self.col_char_widths.get(c, 1)))
        # Invalidate cached base spans: column widths just changed.
        self._col_widths_version = int(getattr(self, "_col_widths_version", 0)) + 1

    def _build_row_pairs(self, ws_a_val, ws_b_val, force: bool = False):
        return _compute_row_pairs_generic(ws_a_val, ws_b_val, self.max_col, force=force)

    @staticmethod
    def _build_row_pairs_direct(max_row_a: int, max_row_b: int):
        """Direct row pairing (1:1 by row number), used for very large sheets."""
        max_row = max(max_row_a, max_row_b)
        pairs: list[tuple[int | None, int | None]] = []
        for r in range(1, max_row + 1):
            ra = r if r <= max_row_a else None
            rb = r if r <= max_row_b else None
            pairs.append((ra, rb))
        return pairs

    def _precompute_large_diff_by_blocks(
        self,
        ws_a_val,
        ws_b_val,
        ws_a_edit,
        ws_b_edit,
        max_row_a: int,
        max_row_b: int,
        ws_base_val=None,
        ws_base_edit=None,
    ):
        """Large-sheet only-diff precompute using tail-first block scan."""
        edit_cache_a: dict[int, tuple] = {}
        edit_cache_b: dict[int, tuple] = {}
        edit_cache_base: dict[int, tuple] = {}
        has_base = bool(self._is_three_way_enabled() and getattr(self.app, "has_base", False) and ws_base_val is not None)

        if self._align_rows_enabled and self.row_pairs:
            pair_count = len(self.row_pairs)
            block = _LARGE_SHEET_BLOCK_ROWS
            for block_end in range(pair_count, 0, -block):
                block_start = max(0, block_end - block)
                block_pairs = self.row_pairs[block_start:block_end]

                rows_a = _read_rows_into_cache(ws_a_val, [ra for ra, _rb in block_pairs if ra is not None], self.max_col)
                rows_b = _read_rows_into_cache(ws_b_val, [rb for _ra, rb in block_pairs if rb is not None], self.max_col)
                base_rows_needed = []
                if has_base:
                    for off, pair in enumerate(block_pairs):
                        pair_idx = block_start + off
                        base_row = self._base_row_for_pair(pair_idx, pair)
                        if base_row is not None:
                            base_rows_needed.append(base_row)
                rows_base = _read_rows_into_cache(ws_base_val, base_rows_needed, self.max_col) if has_base else {}
                exact_rows: list[tuple[int, int | None, int | None, int | None, set[int], bool, set[int], bool]] = []

                for off in range(len(block_pairs) - 1, -1, -1):
                    pair_idx = block_start + off
                    if not (0 <= pair_idx < len(self.row_pairs)):
                        continue
                    ra, rb = self.row_pairs[pair_idx]
                    row_a = _row_from_cache(rows_a, ra, self.max_col)
                    row_b = _row_from_cache(rows_b, rb, self.max_col)
                    if (ra is None) != (rb is None):
                        cols = {-1}
                        need_exact_ab = False
                    else:
                        cols, need_exact_ab = self._quick_diff_cols_from_value_rows(row_a, row_b)
                    base_cols = set()
                    need_exact_base = False
                    base_row = None
                    if has_base and ra is not None:
                        base_row = self._base_row_for_pair(pair_idx, (ra, rb))
                        if base_row is None:
                            base_cols = {-1}
                        else:
                            row_base = _row_from_cache(rows_base, base_row, self.max_col)
                            base_cols, need_exact_base = self._quick_diff_cols_from_value_rows(row_a, row_base)
                    if (not cols) and (not base_cols):
                        continue
                    if need_exact_ab or need_exact_base:
                        exact_rows.append((pair_idx, ra, rb, base_row, cols, need_exact_ab, base_cols, need_exact_base))
                        continue
                    self.pair_diff_cols[pair_idx] = cols
                    if has_base:
                        self.pair_base_diff_cols[pair_idx] = base_cols
                    self.pair_text_a[pair_idx] = self._render_line_from_raw_parts(self._raw_parts_from_row_values(row_a))
                    self.pair_text_b[pair_idx] = self._render_line_from_raw_parts(self._raw_parts_from_row_values(row_b))

                if exact_rows:
                    exact_rows_a = [ra for _pair_idx, ra, _rb, _base_row, _cols, need_ab, _bcols, need_base in exact_rows if ra is not None and (need_ab or need_base)]
                    exact_rows_b = [rb for _pair_idx, _ra, rb, _base_row, _cols, need_ab, _bcols, _need_base in exact_rows if rb is not None and need_ab]
                    exact_rows_base = [base_row for _pair_idx, _ra, _rb, base_row, _cols, _need_ab, _bcols, need_base in exact_rows if base_row is not None and need_base]
                    rows_a_edit = _read_rows_into_cache(ws_a_edit, exact_rows_a, self.max_col) if ws_a_edit is not None else {}
                    rows_b_edit = _read_rows_into_cache(ws_b_edit, exact_rows_b, self.max_col) if ws_b_edit is not None else {}
                    rows_base_edit = _read_rows_into_cache(ws_base_edit, exact_rows_base, self.max_col) if (ws_base_edit is not None and has_base) else {}
                    for pair_idx, ra, rb, base_row, cols, need_exact_ab, base_cols, need_exact_base in exact_rows:
                        row_a = _row_from_cache(rows_a, ra, self.max_col)
                        row_b = _row_from_cache(rows_b, rb, self.max_col)
                        row_a_edit = _row_from_cache(rows_a_edit, ra, self.max_col)
                        row_b_edit = _row_from_cache(rows_b_edit, rb, self.max_col)
                        if need_exact_ab:
                            line_a, line_b, cols = self._build_row_and_diff_pair_from_values(
                                row_a,
                                row_b,
                                ra=ra,
                                rb=rb,
                                ws_a_edit=ws_a_edit,
                                ws_b_edit=ws_b_edit,
                                edit_cache_a=edit_cache_a,
                                edit_cache_b=edit_cache_b,
                                row_a_edit_vals=row_a_edit,
                                row_b_edit_vals=row_b_edit,
                            )
                        else:
                            line_a = self._render_line_from_raw_parts(self._raw_parts_from_row_values(row_a))
                            line_b = self._render_line_from_raw_parts(self._raw_parts_from_row_values(row_b))
                        if need_exact_base and has_base and ra is not None:
                            row_base = _row_from_cache(rows_base, base_row, self.max_col)
                            row_base_edit = _row_from_cache(rows_base_edit, base_row, self.max_col)
                            base_cols = self._compute_base_diff_cols_from_values(
                                row_a,
                                row_base,
                                ra=ra,
                                base_row=base_row,
                                ws_a_edit=ws_a_edit,
                                ws_base_edit=ws_base_edit,
                                edit_cache_a=edit_cache_a,
                                edit_cache_base=edit_cache_base,
                                row_a_edit_vals=row_a_edit,
                                row_base_edit_vals=row_base_edit,
                            )
                        if (not cols) and (not base_cols):
                            continue
                        self.pair_diff_cols[pair_idx] = cols
                        if has_base:
                            self.pair_base_diff_cols[pair_idx] = base_cols
                        self.pair_text_a[pair_idx] = line_a
                        self.pair_text_b[pair_idx] = line_b
            return

        max_row = max(max_row_a, max_row_b)
        block = _LARGE_SHEET_BLOCK_ROWS
        for block_end in range(max_row, 0, -block):
            block_start = max(1, block_end - block + 1)
            block_len = block_end - block_start + 1

            rows_a = _read_rows_into_cache(
                ws_a_val,
                range(block_start, min(block_end, max_row_a) + 1),
                self.max_col,
            )
            rows_b = _read_rows_into_cache(
                ws_b_val,
                range(block_start, min(block_end, max_row_b) + 1),
                self.max_col,
            )
            base_rows_needed = []
            if has_base:
                for r in range(block_start, block_end + 1):
                    pair_idx = self.row_a_to_pair_idx.get(r)
                    if pair_idx is None:
                        pair_idx = self.row_b_to_pair_idx.get(r)
                    if pair_idx is None or pair_idx >= len(self.row_pairs):
                        continue
                    base_row = self._base_row_for_pair(pair_idx, self.row_pairs[pair_idx])
                    if base_row is not None:
                        base_rows_needed.append(base_row)
            rows_base = _read_rows_into_cache(ws_base_val, base_rows_needed, self.max_col) if has_base else {}
            exact_rows: list[tuple[int, int | None, int | None, int | None, set[int], bool, set[int], bool]] = []

            # Tail-first within changed block (newer rows first).
            for off in range(block_len - 1, -1, -1):
                r = block_start + off
                pair_idx = self.row_a_to_pair_idx.get(r)
                if pair_idx is None:
                    pair_idx = self.row_b_to_pair_idx.get(r)
                if pair_idx is None:
                    continue
                if pair_idx >= len(self.row_pairs):
                    continue
                ra, rb = self.row_pairs[pair_idx]
                row_a = _row_from_cache(rows_a, ra, self.max_col)
                row_b = _row_from_cache(rows_b, rb, self.max_col)
                if (ra is None) != (rb is None):
                    cols = {-1}
                    need_exact_ab = False
                else:
                    cols, need_exact_ab = self._quick_diff_cols_from_value_rows(row_a, row_b)
                base_cols = set()
                need_exact_base = False
                base_row = None
                if has_base and ra is not None:
                    base_row = self._base_row_for_pair(pair_idx, (ra, rb))
                    if base_row is None:
                        base_cols = {-1}
                    else:
                        row_base = _row_from_cache(rows_base, base_row, self.max_col)
                        base_cols, need_exact_base = self._quick_diff_cols_from_value_rows(row_a, row_base)
                if (not cols) and (not base_cols):
                    continue
                if need_exact_ab or need_exact_base:
                    exact_rows.append((pair_idx, ra, rb, base_row, cols, need_exact_ab, base_cols, need_exact_base))
                    continue
                self.pair_diff_cols[pair_idx] = cols
                if has_base:
                    self.pair_base_diff_cols[pair_idx] = base_cols
                self.pair_text_a[pair_idx] = self._render_line_from_raw_parts(self._raw_parts_from_row_values(row_a))
                self.pair_text_b[pair_idx] = self._render_line_from_raw_parts(self._raw_parts_from_row_values(row_b))

            if exact_rows:
                exact_rows_a = [ra for _pair_idx, ra, _rb, _base_row, _cols, need_ab, _bcols, need_base in exact_rows if ra is not None and (need_ab or need_base)]
                exact_rows_b = [rb for _pair_idx, _ra, rb, _base_row, _cols, need_ab, _bcols, _need_base in exact_rows if rb is not None and need_ab]
                exact_rows_base = [base_row for _pair_idx, _ra, _rb, base_row, _cols, _need_ab, _bcols, need_base in exact_rows if base_row is not None and need_base]
                rows_a_edit = _read_rows_into_cache(ws_a_edit, exact_rows_a, self.max_col) if ws_a_edit is not None else {}
                rows_b_edit = _read_rows_into_cache(ws_b_edit, exact_rows_b, self.max_col) if ws_b_edit is not None else {}
                rows_base_edit = _read_rows_into_cache(ws_base_edit, exact_rows_base, self.max_col) if (has_base and ws_base_edit is not None) else {}
                for pair_idx, ra, rb, base_row, cols, need_exact_ab, base_cols, need_exact_base in exact_rows:
                    row_a = _row_from_cache(rows_a, ra, self.max_col)
                    row_b = _row_from_cache(rows_b, rb, self.max_col)
                    row_a_edit = _row_from_cache(rows_a_edit, ra, self.max_col)
                    row_b_edit = _row_from_cache(rows_b_edit, rb, self.max_col)
                    if need_exact_ab:
                        line_a, line_b, cols = self._build_row_and_diff_pair_from_values(
                            row_a,
                            row_b,
                            ra=ra,
                            rb=rb,
                            ws_a_edit=ws_a_edit,
                            ws_b_edit=ws_b_edit,
                            edit_cache_a=edit_cache_a,
                            edit_cache_b=edit_cache_b,
                            row_a_edit_vals=row_a_edit,
                            row_b_edit_vals=row_b_edit,
                        )
                    else:
                        line_a = self._render_line_from_raw_parts(self._raw_parts_from_row_values(row_a))
                        line_b = self._render_line_from_raw_parts(self._raw_parts_from_row_values(row_b))
                    if need_exact_base and has_base and ra is not None:
                        row_base = _row_from_cache(rows_base, base_row, self.max_col)
                        row_base_edit = _row_from_cache(rows_base_edit, base_row, self.max_col)
                        base_cols = self._compute_base_diff_cols_from_values(
                            row_a,
                            row_base,
                            ra=ra,
                            base_row=base_row,
                            ws_a_edit=ws_a_edit,
                            ws_base_edit=ws_base_edit,
                            edit_cache_a=edit_cache_a,
                            edit_cache_base=edit_cache_base,
                            row_a_edit_vals=row_a_edit,
                            row_base_edit_vals=row_base_edit,
                        )
                    if (not cols) and (not base_cols):
                        continue
                    self.pair_diff_cols[pair_idx] = cols
                    if has_base:
                        self.pair_base_diff_cols[pair_idx] = base_cols
                    self.pair_text_a[pair_idx] = line_a
                    self.pair_text_b[pair_idx] = line_b

    def _build_row_and_diff(self, ws_a_val, ws_b_val, ws_a_edit, ws_b_edit, r: int):
        parts_a = []
        parts_b = []
        cols = set()
        for c in range(1, self.max_col + 1):
            da, db, eq = _cell_display_and_equal(ws_a_val, ws_b_val, ws_a_edit, ws_b_edit, r, c)
            parts_a.append(_val_to_str(da))
            parts_b.append(_val_to_str(db))
            if not eq:
                cols.add(c)
        line_a = str(r) + "\t" + "\t".join(parts_a)
        line_b = str(r) + "\t" + "\t".join(parts_b)
        return line_a, line_b, cols

    def _compute_diff_cols_from_rows(self, row_a, row_b):
        cols = set()
        # row tuples are 0-indexed; cols are 1-indexed
        for i, (va, vb) in enumerate(zip(row_a, row_b), start=1):
            if va != vb:
                cols.add(i)
        return cols

    def _build_line_from_row(self, r: int, row_vals) -> str:
        return str(r) + "\t" + "\t".join(_val_to_str(v) for v in row_vals)

    def _base_spans(self) -> dict:
        """Unclamped {colIndex: (start, end)} char positions, cached per refresh.

        Depends only on max_col, grid-overlay state and the current column widths,
        so it is recomputed only when those change (``_col_widths_version`` bump),
        not once per rendered line in the tag phase."""
        grid = self._is_grid_overlay_enabled()
        key = (int(self.max_col or 0), bool(grid), int(getattr(self, "_col_widths_version", 0)))
        if getattr(self, "_base_spans_cache", None) is not None \
                and getattr(self, "_base_spans_cache_key", None) == key:
            return self._base_spans_cache
        sep_len = len(_COL_SEP) if grid else 3
        pos = 0
        spans = {}
        for c in range(1, self.max_col + 1):
            w = max(1, self.col_char_widths.get(c, 1))
            spans[c] = (pos, pos + w)
            pos += w
            if c < self.max_col:
                pos += sep_len
        self._base_spans_cache = spans
        self._base_spans_cache_key = key
        return spans

    def _spans_for_line(self, line: str = "") -> dict:
        """Return {colIndex: (start, end)} character positions in the rendered line.

        When a concrete line string is available, clamp spans to the actual text length so
        downstream tag ranges stay aligned with what Tk is really showing."""
        base = self._base_spans()
        text = str(line or "")
        if not text:
            # Copy so callers can't mutate the cached base spans.
            return dict(base)
        text_len = len(text)
        return {c: (min(s, text_len), min(e, text_len)) for c, (s, e) in base.items()}

    def _diffcell_tag_args_for_line(
        self,
        line_no: int,
        pair_idx: int,
        line_a: str = "",
        line_base: str = "",
        line_b: str = "",
    ) -> tuple[list[str], list[str], list[str], list[tuple[int, int]], list[tuple[int, int]], list[tuple[int, int]]]:
        left_args: list[str] = []
        base_args: list[str] = []
        right_args: list[str] = []
        left_ranges: list[tuple[int, int]] = []
        base_ranges: list[tuple[int, int]] = []
        right_ranges: list[tuple[int, int]] = []
        if not (0 <= int(pair_idx) < len(self.row_pairs)):
            return left_args, base_args, right_args, left_ranges, base_ranges, right_ranges
        pair = self.row_pairs[pair_idx]
        ra, rb = pair
        base_r = self._base_row_for_pair(pair_idx, pair) if self._is_three_way_enabled() else None
        cols = self._visual_diff_cols_for_pair(pair_idx)
        if not cols:
            return left_args, base_args, right_args, left_ranges, base_ranges, right_ranges
        spans_a = self._spans_for_line(line_a)
        spans_base = self._spans_for_line(line_base) if self._is_three_way_enabled() else {}
        spans_b = self._spans_for_line(line_b)
        for c in cols:
            if c <= 0:
                continue
            if ra is not None and c in spans_a:
                s, e = spans_a[c]
                if s < e:
                    left_args.extend([f"{line_no}.{s}", f"{line_no}.{e}"])
                    left_ranges.append((s, e))
            if base_r is not None and c in spans_base:
                s, e = spans_base[c]
                if s < e:
                    base_args.extend([f"{line_no}.{s}", f"{line_no}.{e}"])
                    base_ranges.append((s, e))
            if rb is not None and c in spans_b:
                s, e = spans_b[c]
                if s < e:
                    right_args.extend([f"{line_no}.{s}", f"{line_no}.{e}"])
                    right_ranges.append((s, e))
        return left_args, base_args, right_args, left_ranges, base_ranges, right_ranges

    def _clear_diffrow_under_diffcells(self, left_args=None, base_args=None, right_args=None):
        try:
            for widget, args in ((self.left, left_args), (self.base, base_args), (self.right, right_args)):
                if not args:
                    continue
                pairs = list(args)
                for i in range(0, len(pairs), 2):
                    if i + 1 >= len(pairs):
                        break
                    widget.tag_remove("diffrow", pairs[i], pairs[i + 1])
        except Exception:
            pass

    def _apply_rownum_diff_tag_line(self, line_idx: int, pair_idx: int):
        pass  # Row headers are rendered in dedicated widgets (left_ln/base_ln/right_ln).

    # ---------- Only-diff toggle ----------
    def _invalidate_only_diff_snapshot_cache(self):
        self._only_diff_source_version += 1
        self._only_diff_rows_cache = None
        self._only_diff_rows_cache_key = None
        self._only_diff_async_build_key = None
        self._only_diff_async_building = False
        self._only_diff_async_build_seq += 1

    def _has_user_edits_for_current_sheet(self) -> bool:
        if bool(self.touched_rows):
            return True
        modified_a = getattr(self.app, "modified_sheets_a", set())
        modified_b = getattr(self.app, "modified_sheets_b", set())
        return self.sheet in modified_a or self.sheet in modified_b

    def _current_only_diff_cache_key(self) -> tuple:
        return (
            self.sheet,
            int(self._only_diff_source_version),
            int(len(self.row_pairs)),
            int(self.max_col or 0),
            int(bool(self._align_rows_enabled)),
            int(bool(self._force_sequence_align)),
            int(bool(self._is_three_way_enabled() and getattr(self.app, "has_base", False))),
        )

    def _has_valid_only_diff_snapshot_cache(self) -> bool:
        rows = getattr(self, "_only_diff_rows_cache", None)
        if rows is None:
            return False
        return self._only_diff_rows_cache_key == self._current_only_diff_cache_key()

    def _cache_only_diff_rows_snapshot(self, rows):
        normalized = sorted({int(idx) for idx in (rows or []) if 0 <= int(idx) < len(self.row_pairs)})
        self._only_diff_rows_cache = normalized
        self._only_diff_rows_cache_key = self._current_only_diff_cache_key()
        return list(normalized)

    def _only_diff_rows_with_touched(self, rows):
        rows_set = set(int(idx) for idx in (rows or []) if 0 <= int(idx) < len(self.row_pairs))
        for r in self.touched_rows:
            idx = self.row_a_to_pair_idx.get(r)
            if idx is None:
                idx = self.row_b_to_pair_idx.get(r)
            if idx is not None:
                rows_set.add(int(idx))
        return sorted(rows_set)

    def _set_only_diff_pending_info(self):
        try:
            total_rows = len(self.row_pairs) if self.row_pairs else self.max_row
            self.info.configure(
                text=f"只看差异 | 正在后台生成精确差异行...   Rows: {total_rows}   Cols: {self.max_col}"
            )
        except Exception:
            pass

    def _refresh_mode_switch_preserving_selection(self, *, rescan: bool):
        saved_selection = self._snapshot_explicit_selection_state()
        self._clear_selection_visuals()
        self._set_main_selected_cell(None, None)
        self._cursor_cmp_sel_col = None
        self._cursor_cmp_sel_line = None
        self.selected_pair_idx = None
        self.selected_excel_row = None
        self.selected_excel_row_a = None
        self.selected_excel_row_b = None
        self._clear_hover_state(clear_panel=True)
        self._suppress_bg_apply = True
        try:
            self.refresh(row_only=None, rescan=rescan)
        finally:
            self._suppress_bg_apply = False  # Fresh data rendered; allow bg applies again.
        if not self._restore_explicit_selection_state(saved_selection):
            self.clear_explicit_cell_selection()
        self._update_cursor_lines()
        self._update_diff_nav_state()

    def _start_async_large_only_diff_build(self) -> bool:
        # This worker opens disk snapshots. Once the user has adopted any data,
        # those files are stale and must never be allowed to rebuild UI text.
        if self._has_user_edits_for_current_sheet():
            _dlog(f"only-diff async skipped after user edits: sheet={self.sheet}")
            return False
        if not getattr(self, "_data_ready", False):
            self._prefer_only_diff_when_ready = True
            return False
        if not self.row_pairs:
            return False
        cache_key = self._current_only_diff_cache_key()
        if self._has_valid_only_diff_snapshot_cache():
            return False
        if self._only_diff_async_building and self._only_diff_async_build_key == cache_key:
            self._set_only_diff_pending_info()
            return True

        self._only_diff_async_build_seq += 1
        build_seq = int(self._only_diff_async_build_seq)
        self._only_diff_async_build_key = cache_key
        self._only_diff_async_building = True
        self._set_only_diff_pending_info()

        sheet_name = self.sheet
        max_col = int(self.max_col or 1)
        row_pairs = list(self.row_pairs)
        row_a_to_pair_idx = dict(self.row_a_to_pair_idx)
        row_b_to_pair_idx = dict(self.row_b_to_pair_idx)
        mine_to_base_row = dict(self.mine_to_base_row)
        theirs_to_base_row = dict(self.theirs_to_base_row)
        pair_base_row_override = dict(getattr(self, "pair_base_row_override", {}) or {})
        missing_base_row_map = dict(getattr(self, "_missing_base_row_map", {}) or {})
        has_base = bool(self._is_three_way_enabled() and getattr(self.app, "has_base", False))
        align_enabled = bool(self._align_rows_enabled)

        def _base_row_for_snapshot(pair_idx: int, pair: tuple[int | None, int | None]) -> int | None:
            if not has_base:
                return None
            if pair_idx in pair_base_row_override:
                return pair_base_row_override.get(pair_idx)
            ra, rb = pair
            mapped = missing_base_row_map.get(pair_idx)
            if mapped is not None:
                return mapped
            if ra is not None and ra in mine_to_base_row:
                return mine_to_base_row.get(ra)
            if rb is not None and rb in theirs_to_base_row:
                return theirs_to_base_row.get(rb)
            return None

        def _worker_impl():
            wb_a_val = None
            wb_b_val = None
            wb_base_val = None
            wb_a_edit = None
            wb_b_edit = None
            wb_base_edit = None

            def _cancelled():
                return bool(
                    getattr(self.app, "_is_closing", False)
                    or build_seq != self._only_diff_async_build_seq
                )

            if _cancelled():
                return
            try:
                wb_a_val = load_workbook(self.app._file_a_val_path, data_only=True, read_only=True)
                wb_b_val = load_workbook(self.app._file_b_val_path, data_only=True, read_only=True)
                if has_base and getattr(self.app, "_file_base_val_path", None):
                    wb_base_val = load_workbook(self.app._file_base_val_path, data_only=True, read_only=True)
                wb_a_edit = load_workbook(self.app.file_a, data_only=False, read_only=True)
                wb_b_edit = load_workbook(self.app.file_b, data_only=False, read_only=True)
                if has_base and getattr(self.app, "base_path", None):
                    wb_base_edit = load_workbook(self.app.base_path, data_only=False, read_only=True)
                if _cancelled():
                    _wbs_close(wb_a_val, wb_b_val, wb_base_val, wb_a_edit, wb_b_edit, wb_base_edit)
                    return
            except Exception as e:
                _dlog(f"only-diff async open failed sheet={sheet_name}: {e}")
                _wbs_close(wb_a_val, wb_b_val, wb_base_val, wb_a_edit, wb_b_edit, wb_base_edit)

                def _apply_open_fail():
                    if build_seq != self._only_diff_async_build_seq:
                        return
                    self._only_diff_async_building = False
                    self._only_diff_async_build_key = None

                try:
                    self.app._queue_ui_task(_apply_open_fail)
                except Exception:
                    pass
                return

            result = {
                "build_key": cache_key,
                "build_seq": build_seq,
                "sheet": sheet_name,
                "diff_pair_indices": [],
                "pair_diff_cols": {},
                "pair_base_diff_cols": {},
                "pair_parts_a": {},
                "pair_parts_b": {},
            }
            try:
                ws_a_val = wb_a_val[sheet_name]
                ws_b_val = wb_b_val[sheet_name]
                ws_a_edit = wb_a_edit[sheet_name]
                ws_b_edit = wb_b_edit[sheet_name]
                ws_base_val = wb_base_val[sheet_name] if wb_base_val is not None and sheet_name in wb_base_val.sheetnames else None
                ws_base_edit = wb_base_edit[sheet_name] if wb_base_edit is not None and sheet_name in wb_base_edit.sheetnames else ws_base_val
                edit_cache_a: dict[int, tuple] = {}
                edit_cache_b: dict[int, tuple] = {}
                edit_cache_base: dict[int, tuple] = {}

                def _formula_rows(ws_edit):
                    rows = {}
                    if ws_edit is None:
                        return rows
                    for row_idx, cells in enumerate(
                        ws_edit.iter_rows(min_col=1, max_col=max_col, values_only=False),
                        start=1,
                    ):
                        if (row_idx & 255) == 0 and _cancelled():
                            break
                        values = tuple(cell.value for cell in cells)
                        if any(
                            getattr(cell, "data_type", None) == "f"
                            or _special_formula_signature(cell.value) is not None
                            for cell in cells
                        ):
                            rows[row_idx] = _pad_row_values(values, max_col)
                    return rows

                formula_rows_a = _formula_rows(ws_a_edit)
                formula_rows_b = _formula_rows(ws_b_edit)
                formula_rows_base = _formula_rows(ws_base_edit) if has_base else {}

                def _needs_formula_exact(row_left, row_right, edit_left, edit_right):
                    for offset in range(max_col):
                        if row_left[offset] is not None or row_right[offset] is not None:
                            continue
                        left_formula = _formula_text(edit_left[offset]) or _special_formula_signature(edit_left[offset])
                        right_formula = _formula_text(edit_right[offset]) or _special_formula_signature(edit_right[offset])
                        if left_formula or right_formula:
                            return True
                    return False

                def _edit_rows_for(needed_rows, preloaded, ws_edit):
                    needed = sorted({int(row) for row in needed_rows if row is not None})
                    result_rows = {row: preloaded[row] for row in needed if row in preloaded}
                    missing = [row for row in needed if row not in result_rows]
                    if missing and ws_edit is not None:
                        result_rows.update(_read_rows_into_cache(ws_edit, missing, max_col))
                    return result_rows

                if align_enabled and row_pairs:
                    pair_count = len(row_pairs)
                    block = _LARGE_SHEET_BLOCK_ROWS
                    for block_end in range(pair_count, 0, -block):
                        if _cancelled():
                            return
                        block_start = max(0, block_end - block)
                        block_pairs = row_pairs[block_start:block_end]
                        rows_a = _read_rows_into_cache(ws_a_val, [ra for ra, _rb in block_pairs if ra is not None], max_col)
                        rows_b = _read_rows_into_cache(ws_b_val, [rb for _ra, rb in block_pairs if rb is not None], max_col)
                        base_rows_needed = []
                        if has_base and ws_base_val is not None:
                            for off, pair in enumerate(block_pairs):
                                base_row = _base_row_for_snapshot(block_start + off, pair)
                                if base_row is not None:
                                    base_rows_needed.append(base_row)
                        rows_base = _read_rows_into_cache(ws_base_val, base_rows_needed, max_col) if (has_base and ws_base_val is not None) else {}
                        exact_rows: list[tuple[int, int | None, int | None, int | None, set[int], bool, set[int], bool]] = []

                        for off in range(len(block_pairs) - 1, -1, -1):
                            pair_idx = block_start + off
                            ra, rb = block_pairs[off]
                            row_a = _row_from_cache(rows_a, ra, max_col)
                            row_b = _row_from_cache(rows_b, rb, max_col)
                            if (ra is None) != (rb is None):
                                cols = {-1}
                                need_exact_ab = False
                            else:
                                cols, need_exact_ab = self._quick_diff_cols_from_value_rows(row_a, row_b)
                                if not need_exact_ab and (ra in formula_rows_a or rb in formula_rows_b):
                                    need_exact_ab = _needs_formula_exact(
                                        row_a,
                                        row_b,
                                        _row_from_cache(formula_rows_a, ra, max_col),
                                        _row_from_cache(formula_rows_b, rb, max_col),
                                    )
                            base_cols = set()
                            need_exact_base = False
                            base_row = None
                            if has_base and ws_base_val is not None and ra is not None:
                                base_row = _base_row_for_snapshot(pair_idx, (ra, rb))
                                if base_row is None:
                                    base_cols = {-1}
                                else:
                                    row_base = _row_from_cache(rows_base, base_row, max_col)
                                    base_cols, need_exact_base = self._quick_diff_cols_from_value_rows(row_a, row_base)
                                    if not need_exact_base and (ra in formula_rows_a or base_row in formula_rows_base):
                                        need_exact_base = _needs_formula_exact(
                                            row_a,
                                            row_base,
                                            _row_from_cache(formula_rows_a, ra, max_col),
                                            _row_from_cache(formula_rows_base, base_row, max_col),
                                        )
                            if (not cols) and (not base_cols) and (not need_exact_ab) and (not need_exact_base):
                                continue
                            if need_exact_ab or need_exact_base:
                                exact_rows.append((pair_idx, ra, rb, base_row, cols, need_exact_ab, base_cols, need_exact_base))
                                continue
                            result["diff_pair_indices"].append(pair_idx)
                            result["pair_diff_cols"][pair_idx] = cols
                            if has_base and base_cols:
                                result["pair_base_diff_cols"][pair_idx] = base_cols
                            result["pair_parts_a"][pair_idx] = self._raw_parts_from_row_values(row_a)
                            result["pair_parts_b"][pair_idx] = self._raw_parts_from_row_values(row_b)

                        if exact_rows:
                            exact_rows_a = [ra for _pair_idx, ra, _rb, _base_row, _cols, need_ab, _bcols, need_base in exact_rows if ra is not None and (need_ab or need_base)]
                            exact_rows_b = [rb for _pair_idx, _ra, rb, _base_row, _cols, need_ab, _bcols, _need_base in exact_rows if rb is not None and need_ab]
                            exact_rows_base = [base_row for _pair_idx, _ra, _rb, base_row, _cols, _need_ab, _bcols, need_base in exact_rows if base_row is not None and need_base]
                            rows_a_edit = _edit_rows_for(exact_rows_a, formula_rows_a, ws_a_edit)
                            rows_b_edit = _edit_rows_for(exact_rows_b, formula_rows_b, ws_b_edit)
                            rows_base_edit = _edit_rows_for(exact_rows_base, formula_rows_base, ws_base_edit) if has_base else {}
                            for pair_idx, ra, rb, base_row, cols, need_exact_ab, base_cols, need_exact_base in exact_rows:
                                row_a = _row_from_cache(rows_a, ra, max_col)
                                row_b = _row_from_cache(rows_b, rb, max_col)
                                row_a_edit = _row_from_cache(rows_a_edit, ra, max_col)
                                row_b_edit = _row_from_cache(rows_b_edit, rb, max_col)
                                if need_exact_ab:
                                    raw_a, raw_b, cols = self._build_row_parts_and_diff_pair_from_values(
                                        row_a,
                                        row_b,
                                        ra=ra,
                                        rb=rb,
                                        ws_a_edit=ws_a_edit,
                                        ws_b_edit=ws_b_edit,
                                        edit_cache_a=edit_cache_a,
                                        edit_cache_b=edit_cache_b,
                                        row_a_edit_vals=row_a_edit,
                                        row_b_edit_vals=row_b_edit,
                                    )
                                else:
                                    raw_a = self._raw_parts_from_row_values(row_a)
                                    raw_b = self._raw_parts_from_row_values(row_b)
                                if need_exact_base and has_base and ws_base_val is not None and ra is not None:
                                    row_base = _row_from_cache(rows_base, base_row, max_col)
                                    row_base_edit = _row_from_cache(rows_base_edit, base_row, max_col)
                                    base_cols = self._compute_base_diff_cols_from_values(
                                        row_a,
                                        row_base,
                                        ra=ra,
                                        base_row=base_row,
                                        ws_a_edit=ws_a_edit,
                                        ws_base_edit=ws_base_edit,
                                        edit_cache_a=edit_cache_a,
                                        edit_cache_base=edit_cache_base,
                                        row_a_edit_vals=row_a_edit,
                                        row_base_edit_vals=row_base_edit,
                                    )
                                if (not cols) and (not base_cols):
                                    continue
                                result["diff_pair_indices"].append(pair_idx)
                                result["pair_diff_cols"][pair_idx] = cols
                                if has_base and base_cols:
                                    result["pair_base_diff_cols"][pair_idx] = base_cols
                                result["pair_parts_a"][pair_idx] = raw_a
                                result["pair_parts_b"][pair_idx] = raw_b
                else:
                    max_row_a = ws_a_val.max_row or 1
                    max_row_b = ws_b_val.max_row or 1
                    max_row = max(max_row_a, max_row_b)
                    block = _LARGE_SHEET_BLOCK_ROWS
                    for block_end in range(max_row, 0, -block):
                        if _cancelled():
                            return
                        block_start = max(1, block_end - block + 1)
                        rows_a = _read_rows_into_cache(
                            ws_a_val,
                            range(block_start, min(block_end, max_row_a) + 1),
                            max_col,
                        )
                        rows_b = _read_rows_into_cache(
                            ws_b_val,
                            range(block_start, min(block_end, max_row_b) + 1),
                            max_col,
                        )
                        base_rows_needed = []
                        if has_base and ws_base_val is not None:
                            for r in range(block_start, block_end + 1):
                                pair_idx = row_a_to_pair_idx.get(r)
                                if pair_idx is None:
                                    pair_idx = row_b_to_pair_idx.get(r)
                                if pair_idx is None or pair_idx >= len(row_pairs):
                                    continue
                                base_row = _base_row_for_snapshot(pair_idx, row_pairs[pair_idx])
                                if base_row is not None:
                                    base_rows_needed.append(base_row)
                        rows_base = _read_rows_into_cache(ws_base_val, base_rows_needed, max_col) if (has_base and ws_base_val is not None) else {}
                        exact_rows: list[tuple[int, int | None, int | None, int | None, set[int], bool, set[int], bool]] = []

                        for r in range(block_end, block_start - 1, -1):
                            pair_idx = row_a_to_pair_idx.get(r)
                            if pair_idx is None:
                                pair_idx = row_b_to_pair_idx.get(r)
                            if pair_idx is None or pair_idx >= len(row_pairs):
                                continue
                            ra, rb = row_pairs[pair_idx]
                            row_a = _row_from_cache(rows_a, ra, max_col)
                            row_b = _row_from_cache(rows_b, rb, max_col)
                            if (ra is None) != (rb is None):
                                cols = {-1}
                                need_exact_ab = False
                            else:
                                cols, need_exact_ab = self._quick_diff_cols_from_value_rows(row_a, row_b)
                                if not need_exact_ab and (ra in formula_rows_a or rb in formula_rows_b):
                                    need_exact_ab = _needs_formula_exact(
                                        row_a,
                                        row_b,
                                        _row_from_cache(formula_rows_a, ra, max_col),
                                        _row_from_cache(formula_rows_b, rb, max_col),
                                    )
                            base_cols = set()
                            need_exact_base = False
                            base_row = None
                            if has_base and ws_base_val is not None and ra is not None:
                                base_row = _base_row_for_snapshot(pair_idx, (ra, rb))
                                if base_row is None:
                                    base_cols = {-1}
                                else:
                                    row_base = _row_from_cache(rows_base, base_row, max_col)
                                    base_cols, need_exact_base = self._quick_diff_cols_from_value_rows(row_a, row_base)
                                    if not need_exact_base and (ra in formula_rows_a or base_row in formula_rows_base):
                                        need_exact_base = _needs_formula_exact(
                                            row_a,
                                            row_base,
                                            _row_from_cache(formula_rows_a, ra, max_col),
                                            _row_from_cache(formula_rows_base, base_row, max_col),
                                        )
                            if (not cols) and (not base_cols) and (not need_exact_ab) and (not need_exact_base):
                                continue
                            if need_exact_ab or need_exact_base:
                                exact_rows.append((pair_idx, ra, rb, base_row, cols, need_exact_ab, base_cols, need_exact_base))
                                continue
                            result["diff_pair_indices"].append(pair_idx)
                            result["pair_diff_cols"][pair_idx] = cols
                            if has_base and base_cols:
                                result["pair_base_diff_cols"][pair_idx] = base_cols
                            result["pair_parts_a"][pair_idx] = self._raw_parts_from_row_values(row_a)
                            result["pair_parts_b"][pair_idx] = self._raw_parts_from_row_values(row_b)

                        if exact_rows:
                            exact_rows_a = [ra for _pair_idx, ra, _rb, _base_row, _cols, need_ab, _bcols, need_base in exact_rows if ra is not None and (need_ab or need_base)]
                            exact_rows_b = [rb for _pair_idx, _ra, rb, _base_row, _cols, need_ab, _bcols, _need_base in exact_rows if rb is not None and need_ab]
                            exact_rows_base = [base_row for _pair_idx, _ra, _rb, base_row, _cols, _need_ab, _bcols, need_base in exact_rows if base_row is not None and need_base]
                            rows_a_edit = _edit_rows_for(exact_rows_a, formula_rows_a, ws_a_edit)
                            rows_b_edit = _edit_rows_for(exact_rows_b, formula_rows_b, ws_b_edit)
                            rows_base_edit = _edit_rows_for(exact_rows_base, formula_rows_base, ws_base_edit) if has_base else {}
                            for pair_idx, ra, rb, base_row, cols, need_exact_ab, base_cols, need_exact_base in exact_rows:
                                row_a = _row_from_cache(rows_a, ra, max_col)
                                row_b = _row_from_cache(rows_b, rb, max_col)
                                row_a_edit = _row_from_cache(rows_a_edit, ra, max_col)
                                row_b_edit = _row_from_cache(rows_b_edit, rb, max_col)
                                if need_exact_ab:
                                    raw_a, raw_b, cols = self._build_row_parts_and_diff_pair_from_values(
                                        row_a,
                                        row_b,
                                        ra=ra,
                                        rb=rb,
                                        ws_a_edit=ws_a_edit,
                                        ws_b_edit=ws_b_edit,
                                        edit_cache_a=edit_cache_a,
                                        edit_cache_b=edit_cache_b,
                                        row_a_edit_vals=row_a_edit,
                                        row_b_edit_vals=row_b_edit,
                                    )
                                else:
                                    raw_a = self._raw_parts_from_row_values(row_a)
                                    raw_b = self._raw_parts_from_row_values(row_b)
                                if need_exact_base and has_base and ws_base_val is not None and ra is not None:
                                    row_base = _row_from_cache(rows_base, base_row, max_col)
                                    row_base_edit = _row_from_cache(rows_base_edit, base_row, max_col)
                                    base_cols = self._compute_base_diff_cols_from_values(
                                        row_a,
                                        row_base,
                                        ra=ra,
                                        base_row=base_row,
                                        ws_a_edit=ws_a_edit,
                                        ws_base_edit=ws_base_edit,
                                        edit_cache_a=edit_cache_a,
                                        edit_cache_base=edit_cache_base,
                                        row_a_edit_vals=row_a_edit,
                                        row_base_edit_vals=row_base_edit,
                                    )
                                if (not cols) and (not base_cols):
                                    continue
                                result["diff_pair_indices"].append(pair_idx)
                                result["pair_diff_cols"][pair_idx] = cols
                                if has_base and base_cols:
                                    result["pair_base_diff_cols"][pair_idx] = base_cols
                                result["pair_parts_a"][pair_idx] = raw_a
                                result["pair_parts_b"][pair_idx] = raw_b

                result["diff_pair_indices"] = sorted(set(result["diff_pair_indices"]))
            except Exception as e:
                _dlog(f"only-diff async build failed sheet={sheet_name}: {e}")
                result["error"] = str(e)
            finally:
                _wbs_close(wb_a_val, wb_b_val, wb_base_val, wb_a_edit, wb_b_edit, wb_base_edit)

            def _apply_result(res=result):
                if res.get("build_seq") != self._only_diff_async_build_seq:
                    return
                if res.get("build_key") != self._current_only_diff_cache_key():
                    self._only_diff_async_building = False
                    self._only_diff_async_build_key = None
                    return
                if self._has_user_edits_for_current_sheet():
                    self._only_diff_async_building = False
                    self._only_diff_async_build_key = None
                    _dlog(f"only-diff async result dropped after user edits: sheet={self.sheet}")
                    return
                self._only_diff_async_building = False
                self._only_diff_async_build_key = None
                if res.get("error"):
                    return
                for pair_idx, cols in (res.get("pair_diff_cols") or {}).items():
                    self.pair_diff_cols[int(pair_idx)] = set(cols)
                for pair_idx, cols in (res.get("pair_base_diff_cols") or {}).items():
                    self.pair_base_diff_cols[int(pair_idx)] = set(cols)
                for pair_idx, parts in (res.get("pair_parts_a") or {}).items():
                    self.pair_text_a[int(pair_idx)] = self._render_line_from_raw_parts(list(parts))
                for pair_idx, parts in (res.get("pair_parts_b") or {}).items():
                    self.pair_text_b[int(pair_idx)] = self._render_line_from_raw_parts(list(parts))
                self._cache_only_diff_rows_snapshot(res.get("diff_pair_indices") or [])
                self._invalidate_render_cache()
                if bool(self.only_diff_var.get()):
                    self._refresh_mode_switch_preserving_selection(rescan=False)
                else:
                    self._update_diff_nav_state()

            try:
                if not _cancelled():
                    self.app._queue_ui_task(_apply_result)
            except Exception as e:
                _dlog(f"only-diff async queue failed sheet={sheet_name}: {e}")

        def _priority_worker():
            try:
                _worker_impl()
            finally:
                self.app._end_priority_diff_compute()

        self.app._begin_priority_diff_compute()
        self._only_diff_async_thread = self.app._start_background_thread(
            _priority_worker,
            name=f"sow-only-diff-{sheet_name[:30]}",
        )
        if self._only_diff_async_thread is None:
            self.app._end_priority_diff_compute()
            self._only_diff_async_building = False
            self._only_diff_async_build_key = None
            return False
        return True

    def _persist_only_diff_setting_debounced(self):
        try:
            self.app.only_diff_default = int(self.only_diff_var.get())
            if getattr(self, "_settings_save_id", None):
                try:
                    self.frame.after_cancel(self._settings_save_id)
                except Exception:
                    pass
            self._settings_save_id = self.frame.after(1000, self._flush_settings)
        except Exception as e:
            _dlog(f"settings debounce failed: {e}")

    def _toggle_only_diff(self):
        # Snapshot mode confirmed by user: diff rows list is generated once when opening (or manual refresh).
        # Toggling "只看差异" only switches display, without recomputing the diff map.
        try:
            _dlog(f"TOGGLE only_diff={bool(self.only_diff_var.get())} raw={self.only_diff_var.get()} sheet={self.sheet}")
        except Exception:
            pass

        # Always trust current UI value; do not auto-flip state.
        # Auto-flip can invert explicit programmatic toggles and break only-diff filtering.
        cur = int(self.only_diff_var.get())
        self._last_only_diff_value = cur

        # Performance optimization:
        # - For normal sheets with ready diff data, toggling only-diff only changes
        #   display mode and does not need a full rescan.
        # - For large sheets when enabling only-diff, build the exact diff snapshot
        #   in background and keep the current full view responsive until ready.
        if not getattr(self, "_data_ready", False):
            # Never perform a full rescan on the Tk callback. Preserve the
            # checkbox state, prioritize this Sheet's worker, and apply it when
            # the exact cache arrives.
            self._pending_only_diff_value = cur
            self._prefer_only_diff_when_ready = bool(cur)
            if cur:
                self._set_only_diff_pending_info()
            else:
                try:
                    self.info.configure(text="正在后台加载当前 Sheet；完成后显示全量数据...")
                except Exception:
                    pass
            try:
                self.app._enqueue_sheet(self.sheet, front=True)
                self.app._kick_worker()
            except Exception as e:
                _dlog(f"only-diff prioritize failed sheet={self.sheet}: {e}")
            self._update_diff_nav_state()
            self._persist_only_diff_setting_debounced()
            return
        if bool(cur) and not self._has_valid_only_diff_snapshot_cache():
            if self._start_async_large_only_diff_build():
                self._update_diff_nav_state()
                self._persist_only_diff_setting_debounced()
                return
        try:
            _dlog(
                "TOGGLE only_diff refresh: rescan=False "
                f"data_ready={getattr(self, '_data_ready', False)} "
                f"large={getattr(self, '_is_large_sheet', False)} "
                f"sheet={self.sheet}"
            )
        except Exception:
            pass

        if bool(getattr(self, "_is_large_sheet", False)) and not self._full_render:
            # Both directions can expose thousands of rows. Rendering 800 tagged
            # Text lines makes the checkbox callback feel stuck, so switch modes
            # with the same 200-row first window used by large-sheet initial load.
            self._render_limit = min(
                max(1, int(self._render_limit or _LARGE_SHEET_INITIAL_ROWS)),
                _LARGE_SHEET_INITIAL_ROWS,
            )

        self._refresh_mode_switch_preserving_selection(rescan=False)

        # Persist setting (debounced: write 1 s after last toggle to avoid per-keypress I/O)
        self._persist_only_diff_setting_debounced()

    def _toggle_force_align(self):
        """Manual override for large-sheet row pairing accuracy."""
        try:
            self._force_sequence_align = bool(self.force_align_var.get())
            _dlog(f"TOGGLE force_align={self._force_sequence_align} sheet={self.sheet}")
        except Exception:
            self._force_sequence_align = bool(self.force_align_var.get())
        self._invalidate_only_diff_snapshot_cache()
        self._suppress_bg_apply = True
        try:
            self.refresh(row_only=None, rescan=True)
        finally:
            self._suppress_bg_apply = False  # Fresh data rendered; allow bg applies again.
        self._update_cursor_lines()
        self._update_diff_nav_state()

    def _flush_settings(self):
        """Debounced settings write: called 1 s after the last only-diff toggle."""
        try:
            os.makedirs(os.path.dirname(_SETTINGS_PATH), exist_ok=True)
            with open(_SETTINGS_PATH, "w", encoding="utf-8") as f:
                json.dump({"only_diff": int(self.only_diff_var.get())}, f, ensure_ascii=False)
        except Exception as e:
            _dlog(f"settings save failed: {e}")

    def _manual_rescan(self):
        self._invalidate_only_diff_snapshot_cache()
        self._suppress_bg_apply = True
        try:
            self.refresh(row_only=None, rescan=True)
        finally:
            self._suppress_bg_apply = False  # Fresh data rendered; allow bg applies again.
        self._update_cursor_lines()
        self._update_diff_nav_state()

    # ---------- Merge operations ----------
    def _same_formula_copy_mode(self, src_edit, dst_edit, src_val, dst_val) -> str | None:
        """Return ``noop`` or ``cache`` for same-formula copies."""
        if not _is_same_formula_copy_noop(src_edit, dst_edit):
            return None
        if _merge_cmp_value(src_val) == _merge_cmp_value(dst_val):
            return "noop"
        self._formula_copy_skips_pending += 1
        return "cache"

    def _preflight_region_formula_copy(self, direction: str, pair_indices: list[int]):
        if direction not in ("A2B", "B2A", "BASE2A"):
            return
        if direction == "A2B" and getattr(self.app, "merge_conflict_mode", False):
            return
        ws_a_edit = self.app.ws_a_edit(self.sheet)
        ws_b_edit = self.app.ws_b_edit(self.sheet)
        ws_base_edit = self.app.ws_base_edit(self.sheet) if getattr(self.app, "has_base", False) else None
        ws_a_val = self.app.ws_a_val(self.sheet)
        ws_b_val = self.app.ws_b_val(self.sheet)
        ws_base_val = self.app.ws_base_val(self.sheet) if getattr(self.app, "has_base", False) else None
        for pair_idx in pair_indices:
            if not (0 <= int(pair_idx) < len(self.row_pairs)):
                continue
            pair = self.row_pairs[int(pair_idx)]
            ra, rb = pair
            if direction == "A2B":
                src_ws, src_val_ws, dst_ws, src_row, dst_row = ws_a_edit, ws_a_val, ws_b_edit, ra, rb
            elif direction == "B2A":
                src_ws, src_val_ws, dst_ws, src_row, dst_row = ws_b_edit, ws_b_val, ws_a_edit, rb, ra
            else:
                src_ws, src_val_ws, dst_ws = ws_base_edit, ws_base_val, ws_a_edit
                src_row = self._base_row_for_pair(int(pair_idx), pair)
                dst_row = ra
            if src_ws is None or src_row is None:
                continue
            cols = set(self.pair_diff_cols.get(int(pair_idx), set()))
            if dst_row is None or cols == {-1}:
                cols = set(range(1, max(1, int(src_ws.max_column or 1)) + 1))
            for col_idx in cols:
                if int(col_idx) <= 0:
                    continue
                src_edit = src_ws.cell(row=int(src_row), column=int(col_idx)).value
                dst_edit = (
                    dst_ws.cell(row=int(dst_row), column=int(col_idx)).value
                    if dst_ws is not None and dst_row is not None else None
                )
                src_val = (
                    src_val_ws.cell(row=int(src_row), column=int(col_idx)).value
                    if src_val_ws is not None else None
                )
                _copy_edit_value_for_destination(
                    src_val,
                    src_edit,
                    dst_edit,
                    src_row=int(src_row),
                    src_col=int(col_idx),
                    dst_row=int(dst_row) if dst_row is not None else int(src_row),
                    dst_col=int(col_idx),
                )

    def _show_formula_copy_skip_notice(self, count: int):
        if count <= 0:
            return
        msg = (
            f"已保留 {count} 个单元格的原公式，并采用 theirs 的当前计算结果。\n\n"
            "这些结果来自其他输入单元格。为保证以后重新计算仍一致，"
            "请同时在产生差异的依赖 Sheet/输入列中采用对应数据。"
        )
        _dlog(f"FORMULA_RESULT_CACHE_ADOPTED sheet={self.sheet} count={count}")
        try:
            self.info.configure(text=f"已采用 {count} 个公式缓存结果并保留公式；请同步合并依赖数据")
        except Exception:
            pass
        try:
            messagebox.showinfo("已保留公式并采用计算结果", msg)
        except Exception:
            pass

    def _copy_cell(self, direction: str, event):
        previous_bg_suppression = bool(getattr(self, "_suppress_bg_apply", False))
        begin_interactive = getattr(self.app, "_begin_interactive_action", None)
        end_interactive = getattr(self.app, "_end_interactive_action", None)
        if callable(begin_interactive):
            begin_interactive()
        self._suppress_bg_apply = True
        try:
            if self._is_missing_sheet_view():
                self._copy_missing_sheet(direction)
                return
            anchor = self._capture_view_anchor()
            if direction == "A2B":
                src = self.left
            elif direction == "MINE2A":
                src = self.left
            elif direction == "BASE2A":
                src = self.base
            else:
                src = self.right
            idx = src.index(f"@{event.x},{event.y}")
            src.mark_set("insert", idx)
            line = int(idx.split(".")[0])
            col_char = int(idx.split(".")[1])

            if not (1 <= line <= len(self.display_rows)):
                return
            pair = self._pair_for_line(line)
            ra = self._row_for_side(pair, "A")
            rb = self._row_for_side(pair, "B")
            if direction == "A2B":
                if ra is None:
                    return
                if rb is None:
                    # A-only row: insert new row in B then return (full-row insert).
                    pair_idx = self.display_rows[line - 1]
                    self._insert_row_copy(pair_idx, "A2B", ra, False, None, anchor)
                    return
                src_r = ra
                dst_r = rb
            elif direction == "MINE2A":
                if ra is None and rb is None:
                    return
                src_r = ra if ra is not None else rb
                dst_r = ra if ra is not None else rb
            elif direction == "BASE2A":
                if ra is None:
                    return
                pair_idx = self.display_rows[line - 1]
                src_r = self._base_row_for_pair(pair_idx, pair)
                dst_r = ra
            else:
                if rb is None:
                    return
                if ra is None:
                    # B-only row: insert new row in A then return (full-row insert).
                    pair_idx = self.display_rows[line - 1]
                    self._insert_row_copy(pair_idx, "B2A", rb, False, None, anchor)
                    return
                src_r = rb
                dst_r = ra

            # Use strict character spans to resolve clicked column.
            # Separator clicks must not map to adjacent cells.
            spans = self._spans_for_line()
            c = None
            for col_num, (s0, e0) in spans.items():
                if s0 <= col_char < e0:
                    c = col_num
                    break
            if c is None:
                return
            if c > self.max_col:
                c = self.max_col

            # Merge conflict mode:
            # - "A2B" means keep mine, just mark resolved.
            # - "B2A" means apply theirs to mine, then mark resolved.
            if getattr(self.app, "merge_conflict_mode", False):
                if direction == "A2B":
                    self.app.user_touched_conflicts = True
                    self._resolve_conflict_cell(dst_r, c)
                    return

            if direction == "A2B":
                old_edit = self.app.ws_b_edit(self.sheet).cell(row=dst_r, column=c).value
                old_val = self.app.ws_b_val(self.sheet).cell(row=dst_r, column=c).value
                v_edit = self.app.ws_a_edit(self.sheet).cell(row=src_r, column=c).value
                v_val = self.app.ws_a_val(self.sheet).cell(row=src_r, column=c).value
                new_edit = _copy_edit_value_for_destination(
                    v_val, v_edit, old_edit,
                    src_row=src_r, src_col=c, dst_row=dst_r, dst_col=c,
                )
                formula_mode = self._same_formula_copy_mode(new_edit, old_edit, v_val, old_val)
                if formula_mode == "noop":
                    return
                if formula_mode == "cache":
                    self.app.ws_b_val(self.sheet).cell(row=dst_r, column=c).value = v_val
                    self.app.record_manual_b_formula_cache(self.sheet, dst_r, c, v_val)
                    self.app.modified_b = True
                    self.app.modified_sheets_b.add(self.sheet)
                    self.app.push_undo({"sheet": self.sheet, "target": "B", "cells": [(dst_r, c, old_edit, old_val)]})
                    self._show_formula_copy_skip_notice(1)
                else:
                    # Cached-value mode: always write the cached value
                    self.app.clear_manual_b_formula_cache(self.sheet, dst_r, c)
                    _assign_edit_cell_value(
                        self.app.ws_b_edit(self.sheet).cell(row=dst_r, column=c),
                        new_edit,
                    )
                    self.app.record_manual_b_cell(self.sheet, dst_r, c, new_edit)
                    self.app.ws_b_val(self.sheet).cell(row=dst_r, column=c).value = v_val
                    self.app.modified_b = True
                    self.app.modified_sheets_b.add(self.sheet)
                    self.app.push_undo({"sheet": self.sheet, "target": "B", "cells": [(dst_r, c, old_edit, old_val)]})
            elif direction == "MINE2A":
                # Keep mine value; in conflict mode this means "accept mine".
                if getattr(self.app, "merge_conflict_mode", False):
                    self.app.user_touched_conflicts = True
                    self._resolve_conflict_cell(dst_r, c)
                return
            elif direction == "B2A":
                old_edit = self.app.ws_a_edit(self.sheet).cell(row=dst_r, column=c).value
                old_val = self.app.ws_a_val(self.sheet).cell(row=dst_r, column=c).value
                v_edit = self.app.ws_b_edit(self.sheet).cell(row=src_r, column=c).value
                v_val = self.app.ws_b_val(self.sheet).cell(row=src_r, column=c).value
                new_edit = _copy_edit_value_for_destination(
                    v_val, v_edit, old_edit,
                    src_row=src_r, src_col=c, dst_row=dst_r, dst_col=c,
                )
                formula_mode = self._same_formula_copy_mode(new_edit, old_edit, v_val, old_val)
                if formula_mode == "noop":
                    return
                if formula_mode == "cache":
                    self.app.ws_a_val(self.sheet).cell(row=dst_r, column=c).value = v_val
                    self.app.record_manual_a_formula_cache(self.sheet, dst_r, c, v_val)
                    self.app.modified_a = True
                    self.app.modified_sheets_a.add(self.sheet)
                    self.app.push_undo({"sheet": self.sheet, "target": "A", "cells": [(dst_r, c, old_edit, old_val)]})
                    self._show_formula_copy_skip_notice(1)
                else:
                    _assign_edit_cell_value(self.app.ws_a_edit(self.sheet).cell(row=dst_r, column=c), new_edit)
                    self.app.ws_a_val(self.sheet).cell(row=dst_r, column=c).value = v_val
                    self.app.record_manual_a_cell(self.sheet, dst_r, c, new_edit)
                    self.app.modified_a = True
                    self.app.modified_sheets_a.add(self.sheet)
                    self.app.push_undo({"sheet": self.sheet, "target": "A", "cells": [(dst_r, c, old_edit, old_val)]})
                # In conflict mode, B2A applies theirs; mark conflict resolved.
                if getattr(self.app, "merge_conflict_mode", False):
                    self.app.user_touched_conflicts = True
                    self._resolve_conflict_cell(dst_r, c)
                    return
            else:
                old_edit = self.app.ws_a_edit(self.sheet).cell(row=dst_r, column=c).value
                old_val = self.app.ws_a_val(self.sheet).cell(row=dst_r, column=c).value
                if src_r is None:
                    v_edit = None
                    v_val = None
                else:
                    v_edit = self.app.ws_base_edit(self.sheet).cell(row=src_r, column=c).value
                    v_val = self.app.ws_base_val(self.sheet).cell(row=src_r, column=c).value
                new_edit = (
                    _copy_edit_value_for_destination(
                        v_val, v_edit, old_edit,
                        src_row=src_r, src_col=c, dst_row=dst_r, dst_col=c,
                    )
                    if src_r is not None else None
                )
                formula_mode = self._same_formula_copy_mode(new_edit, old_edit, v_val, old_val)
                if formula_mode == "noop":
                    return
                if formula_mode == "cache":
                    self.app.ws_a_val(self.sheet).cell(row=dst_r, column=c).value = v_val
                    self.app.record_manual_a_formula_cache(self.sheet, dst_r, c, v_val)
                    self.app.modified_a = True
                    self.app.modified_sheets_a.add(self.sheet)
                    self.app.push_undo({"sheet": self.sheet, "target": "A", "cells": [(dst_r, c, old_edit, old_val)]})
                    self._show_formula_copy_skip_notice(1)
                else:
                    _assign_edit_cell_value(self.app.ws_a_edit(self.sheet).cell(row=dst_r, column=c), new_edit)
                    self.app.ws_a_val(self.sheet).cell(row=dst_r, column=c).value = v_val
                    self.app.record_manual_a_cell(self.sheet, dst_r, c, new_edit)
                    self.app.modified_a = True
                    self.app.modified_sheets_a.add(self.sheet)
                    self.app.push_undo({"sheet": self.sheet, "target": "A", "cells": [(dst_r, c, old_edit, old_val)]})

            # Mark as touched: keep row visible in "只看差异" even if diffs are resolved.
            pair = self._pair_for_line(line)
            touched_r = self._row_for_side(pair, "A") or self._row_for_side(pair, "B")
            if touched_r is not None:
                self.touched_rows.add(touched_r)
            self._invalidate_only_diff_snapshot_cache()
            self._invalidate_render_cache()

            # Minimize flicker: use row-only incremental refresh after overwrite.
            # Full-sheet rescan can be done manually by user when needed.
            if bool(self.only_diff_var.get()) and self.snapshot_only_diff:
                self._recalc_row_diff_and_update(dst_r)
            self.refresh(row_only=dst_r, rescan=False)
            self._restore_view_anchor(anchor)
            self._update_cursor_lines()
        except Exception as e:
            messagebox.showerror("Error", f"覆盖单元格失败：\n{e}")
        finally:
            self._suppress_bg_apply = previous_bg_suppression
            if callable(end_interactive):
                end_interactive()

    # ---------- Row-insert helpers ----------

    def _find_row_insert_position(self, pair_idx: int, side: str) -> int:
        """Return the worksheet row index at which to insert a new row.

        Scans row_pairs[0..pair_idx-1] to find the last non-None row on the
        given side ('A' or 'B').  Returns last_row + 1, or 1 if no prior rows
        exist on that side.
        """
        last_row = 0
        side_idx = 0 if side == "A" else 1
        for i in range(pair_idx):
            r = self.row_pairs[i][side_idx]
            if r is not None:
                last_row = r
        return last_row + 1

    def _find_base_row_insert_position(self, pair_idx: int) -> int:
        """Return the BASE worksheet row index at which to insert a new row.

        Uses the BASE row mapping (mine_to_base_row / theirs_to_base_row via
        _base_row_for_pair) instead of mine's row number, so that when mine and
        base are not identity-aligned (mine has prior inserts/deletes vs base)
        the blank base row lands at the correct BASE position. Returns
        last_base_row + 1, or 1 if no prior pair maps to a base row.
        """
        last_base = 0
        for i in range(pair_idx):
            try:
                br = self._base_row_for_pair(i, self.row_pairs[i])
            except Exception:
                br = None
            if br is not None:
                last_base = int(br)
        return last_base + 1

    @staticmethod
    def _shift_row_number_for_insert(row_num: int | None, insert_pos: int, amount: int = 1) -> int | None:
        if row_num is None:
            return None
        try:
            row_num = int(row_num)
        except Exception:
            return row_num
        amount = max(0, int(amount))
        return row_num + amount if row_num >= int(insert_pos) else row_num

    def _rebuild_row_pair_lookup_maps(self):
        self.row_a_to_pair_idx = {}
        self.row_b_to_pair_idx = {}
        for idx, (ra, rb) in enumerate(self.row_pairs):
            if ra is not None:
                self.row_a_to_pair_idx[int(ra)] = idx
            if rb is not None:
                self.row_b_to_pair_idx[int(rb)] = idx

    def _update_row_model_after_insert(
        self,
        *,
        pair_idx: int,
        direction: str,
        insert_pos: int,
        base_insert_pos: int | None = None,
        old_pair: tuple[int | None, int | None] | None,
        base_inserted: bool,
    ):
        if not (0 <= int(pair_idx) < len(self.row_pairs)):
            return

        pair_idx = int(pair_idx)
        insert_pos = int(insert_pos)
        # Base rows shift at the base-coordinate insert position (defaults to
        # insert_pos for backward compatibility / identity mine-base alignment).
        base_insert_pos = int(base_insert_pos) if base_insert_pos is not None else insert_pos
        old_pair = old_pair if old_pair is not None else self.row_pairs[pair_idx]
        old_base_row = None
        if self._is_three_way_enabled() and getattr(self.app, "has_base", False):
            try:
                old_base_row = self._base_row_for_pair(pair_idx, old_pair)
            except Exception:
                old_base_row = None

        updated_pairs: list[tuple[int | None, int | None]] = []
        for idx, (ra, rb) in enumerate(self.row_pairs):
            if direction == "B2A":
                ra = self._shift_row_number_for_insert(ra, insert_pos)
                if idx == pair_idx:
                    ra = insert_pos
            else:
                rb = self._shift_row_number_for_insert(rb, insert_pos)
                if idx == pair_idx:
                    rb = insert_pos
            updated_pairs.append((ra, rb))
        self.row_pairs = updated_pairs
        self._rebuild_row_pair_lookup_maps()

        if self._is_three_way_enabled() and getattr(self.app, "has_base", False):
            def _shift_base_row(base_row: int | None) -> int | None:
                if not base_inserted:
                    return base_row
                return self._shift_row_number_for_insert(base_row, base_insert_pos)

            shifted_base_row = _shift_base_row(old_base_row)
            mine_map_old = dict(getattr(self, "mine_to_base_row", {}) or {})
            theirs_map_old = dict(getattr(self, "theirs_to_base_row", {}) or {})
            overrides_old = dict(getattr(self, "pair_base_row_override", {}) or {})

            if direction == "B2A":
                mine_map_new: dict[int, int] = {}
                for ra, base_row in mine_map_old.items():
                    new_ra = self._shift_row_number_for_insert(ra, insert_pos)
                    if new_ra is None:
                        continue
                    mine_map_new[int(new_ra)] = _shift_base_row(base_row)
                if shifted_base_row is not None:
                    mine_map_new[insert_pos] = int(shifted_base_row)
                self.mine_to_base_row = mine_map_new

                theirs_map_new: dict[int, int] = {}
                for rb, base_row in theirs_map_old.items():
                    if rb is None:
                        continue
                    theirs_map_new[int(rb)] = _shift_base_row(base_row)
                self.theirs_to_base_row = theirs_map_new
            else:
                mine_map_new = {}
                for ra, base_row in mine_map_old.items():
                    if ra is None:
                        continue
                    mine_map_new[int(ra)] = _shift_base_row(base_row)
                self.mine_to_base_row = mine_map_new

                theirs_map_new: dict[int, int] = {}
                for rb, base_row in theirs_map_old.items():
                    new_rb = self._shift_row_number_for_insert(rb, insert_pos)
                    if new_rb is None:
                        continue
                    theirs_map_new[int(new_rb)] = _shift_base_row(base_row)
                if shifted_base_row is not None:
                    theirs_map_new[insert_pos] = int(shifted_base_row)
                self.theirs_to_base_row = theirs_map_new

            overrides_new: dict[int, int | None] = {}
            for idx, base_row in overrides_old.items():
                overrides_new[int(idx)] = _shift_base_row(base_row)
            if shifted_base_row is not None:
                overrides_new[pair_idx] = int(shifted_base_row)
            else:
                overrides_new.pop(pair_idx, None)
            self.pair_base_row_override = overrides_new

        try:
            self.max_row = max(
                int(self.max_row or 1),
                int(getattr(self.app.ws_a_val(self.sheet), "max_row", 1) or 1),
                int(getattr(self.app.ws_b_val(self.sheet), "max_row", 1) or 1),
            )
        except Exception:
            pass

    def _prime_pair_cache_after_insert(
        self,
        *,
        pair_idx: int,
        ws_a_val,
        ws_b_val,
        ws_a_edit,
        ws_b_edit,
    ):
        if not (0 <= int(pair_idx) < len(self.row_pairs)):
            return
        pair_idx = int(pair_idx)
        ra, rb = self.row_pairs[pair_idx]
        line_a, line_b, cols = self._build_row_and_diff_pair(ws_a_val, ws_b_val, ws_a_edit, ws_b_edit, ra, rb)
        self.pair_text_a[pair_idx] = line_a
        self.pair_text_b[pair_idx] = line_b
        self.pair_diff_cols[pair_idx] = cols
        if self._is_three_way_enabled() and getattr(self.app, "has_base", False):
            try:
                self.pair_base_diff_cols[pair_idx] = self._compute_base_diff_cols_for_pair(
                    pair_idx,
                    (ra, rb),
                    max_col=self.max_col,
                    ws_a_val=ws_a_val,
                    ws_a_edit=ws_a_edit,
                    ws_base_val=self.app.ws_base_val(self.sheet),
                    ws_base_edit=self.app.ws_base_edit(self.sheet),
                )
            except Exception:
                self.pair_base_diff_cols.pop(pair_idx, None)
        else:
            self.pair_base_diff_cols.pop(pair_idx, None)

    def _try_fast_refresh_after_row_insert(
        self,
        *,
        pair_idx: int,
        direction: str,
        insert_pos: int,
        base_insert_pos: int | None = None,
        old_pair: tuple[int | None, int | None] | None,
        base_inserted: bool,
        suppress_refresh: bool,
        anchor,
        ws_a_val,
        ws_b_val,
        ws_a_edit,
        ws_b_edit,
    ) -> bool:
        if suppress_refresh:
            return False
        if not getattr(self, "_is_large_sheet", False):
            return False
        if not getattr(self, "_data_ready", False):
            return False
        try:
            self._update_row_model_after_insert(
                pair_idx=pair_idx,
                direction=direction,
                insert_pos=insert_pos,
                base_insert_pos=base_insert_pos,
                old_pair=old_pair,
                base_inserted=base_inserted,
            )
            self._prime_pair_cache_after_insert(
                pair_idx=pair_idx,
                ws_a_val=ws_a_val,
                ws_b_val=ws_b_val,
                ws_a_edit=ws_a_edit,
                ws_b_edit=ws_b_edit,
            )
            if direction == "B2A":
                touched_r = insert_pos
            else:
                touched_r = self.row_pairs[pair_idx][0] or self.row_pairs[pair_idx][1]
            if touched_r is not None:
                self.touched_rows.add(int(touched_r))
            self._invalidate_only_diff_snapshot_cache()
            self._invalidate_render_cache()
            self.refresh(row_only=None, rescan=False)
            if anchor is not None:
                self._restore_view_anchor(anchor)
            self._update_cursor_lines()
            return True
        except Exception as e:
            _dlog(f"FAST_INSERT_REFRESH_FAILED sheet={self.sheet} pair_idx={pair_idx} err={e}")
            return False

    def _batch_insert_row_copy(
        self,
        run: list[tuple[int, int]],
        direction: str,
        suppress_refresh: bool,
        anchor,
    ) -> bool:
        if not run:
            return False
        committed = False
        dst_val_inserted = False
        dst_edit_inserted = False
        base_val_inserted = False
        base_edit_inserted = False
        ws_dst_val = None
        ws_dst_edit = None
        ws_bv = None
        ws_be = None
        insert_pos = None
        base_insert_pos = None
        count = len(run)
        manual_a_cells_before = dict(getattr(self.app, "manual_a_cell_ops", {}) or {})
        manual_a_cache_before = dict(getattr(self.app, "manual_a_formula_cache_ops", {}) or {})
        manual_b_cells_before = dict(getattr(self.app, "manual_b_cell_ops", {}) or {})
        manual_b_cache_before = dict(getattr(self.app, "manual_b_formula_cache_ops", {}) or {})
        manual_row_ops_len = len(getattr(self.app, "manual_a_row_ops", []) or [])
        manual_b_row_ops_len = len(getattr(self.app, "manual_b_row_ops", []) or [])
        undo_len = len(getattr(self.app, "undo_stack", []) or [])
        modified_a_before = bool(getattr(self.app, "modified_a", False))
        modified_b_before = bool(getattr(self.app, "modified_b", False))
        modified_sheets_a_before = set(getattr(self.app, "modified_sheets_a", set()) or set())
        modified_sheets_b_before = set(getattr(self.app, "modified_sheets_b", set()) or set())
        row_pairs_before = list(getattr(self, "row_pairs", []) or [])
        mine_to_base_before = dict(getattr(self, "mine_to_base_row", {}) or {})
        theirs_to_base_before = dict(getattr(self, "theirs_to_base_row", {}) or {})
        overrides_before = dict(getattr(self, "pair_base_row_override", {}) or {})
        touched_before = set(getattr(self, "touched_rows", set()) or set())
        max_row_before = getattr(self, "max_row", 1)
        cache_maps = (
            self.pair_text_a,
            self.pair_text_b,
            self.pair_diff_cols,
            self.pair_base_diff_cols,
        )
        affected_pair_indices = [int(pair_idx) for pair_idx, _src_row in run]
        pair_cache_before = [
            {
                idx: (idx in cache_map, cache_map.get(idx))
                for idx in affected_pair_indices
            }
            for cache_map in cache_maps
        ]
        try:
            run = [(int(pair_idx), int(src_row)) for pair_idx, src_row in run]
            ws_a_val = self.app.ws_a_val(self.sheet)
            ws_b_val = self.app.ws_b_val(self.sheet)
            ws_a_edit = self.app.ws_a_edit(self.sheet)
            ws_b_edit = self.app.ws_b_edit(self.sheet)

            first_pair_idx = run[0][0]
            old_pairs = {
                pair_idx: self.row_pairs[pair_idx]
                for pair_idx, _src_row in run
                if 0 <= pair_idx < len(self.row_pairs)
            }
            old_base_rows = {
                pair_idx: self._base_row_for_pair(pair_idx, old_pairs.get(pair_idx))
                for pair_idx, _src_row in run
                if pair_idx in old_pairs
            }

            if direction == "B2A":
                ws_dst_val = ws_a_val
                ws_dst_edit = ws_a_edit
                ws_src_val = ws_b_val
                ws_src_edit = ws_b_edit
                insert_pos = self._find_row_insert_position(first_pair_idx, "A")
            else:
                ws_dst_val = ws_b_val
                ws_dst_edit = ws_b_edit
                ws_src_val = ws_a_val
                ws_src_edit = ws_a_edit
                insert_pos = self._find_row_insert_position(first_pair_idx, "B")

            ws_dst_val.insert_rows(idx=insert_pos, amount=count)
            dst_val_inserted = True
            ws_dst_edit.insert_rows(idx=insert_pos, amount=count)
            dst_edit_inserted = True

            if direction == "B2A":
                for op_map in (
                    self.app.manual_a_cell_ops,
                    self.app.manual_a_formula_cache_ops,
                ):
                    to_shift = {
                        k: v for k, v in op_map.items()
                        if k[0] == self.sheet and k[1] >= insert_pos
                    }
                    for k in to_shift:
                        del op_map[k]
                    for (s, r, c), v in to_shift.items():
                        op_map[(s, r + count, c)] = v
                self.app.record_manual_a_row_insert(
                    self.sheet,
                    insert_pos,
                    count,
                    source_side="B",
                    source_rows=[src_row for _pair_idx, src_row in run],
                )
            else:
                for op_map in (
                    self.app.manual_b_cell_ops,
                    self.app.manual_b_formula_cache_ops,
                ):
                    to_shift = {
                        key: value for key, value in op_map.items()
                        if key[0] == self.sheet and key[1] >= insert_pos
                    }
                    for key in to_shift:
                        del op_map[key]
                    for (sheet_name, row_idx, col_idx), value in to_shift.items():
                        op_map[(sheet_name, row_idx + count, col_idx)] = value
                self.app.record_manual_b_row_insert(
                    self.sheet,
                    insert_pos,
                    count,
                    source_side="A",
                    source_rows=[src_row for _pair_idx, src_row in run],
                )

            base_inserted = False
            # BASE insert position is computed in BASE coordinates (not mine's
            # insert_pos) so that non-identity mine/base alignment inserts the
            # blank base row at the right place; the row-model base shifts below
            # use the same base_insert_pos to stay consistent.
            base_insert_pos = insert_pos
            if direction == "B2A" and self._is_three_way_enabled():
                try:
                    base_insert_pos = self._find_base_row_insert_position(first_pair_idx)
                    ws_bv = self.app.ws_base_val(self.sheet)
                    ws_be = self.app.ws_base_edit(self.sheet)
                    if ws_bv is None or ws_be is None:
                        raise RuntimeError("BASE worksheet is unavailable")
                    ws_bv.insert_rows(idx=base_insert_pos, amount=count)
                    base_val_inserted = True
                    ws_be.insert_rows(idx=base_insert_pos, amount=count)
                    base_edit_inserted = True
                    base_inserted = True
                except Exception as _e_base_ins:
                    _dlog(f"batch base insert failed: sheet={self.sheet} pos={base_insert_pos} err={_e_base_ins}")
                    raise RuntimeError(f"向 BASE 批量插入对齐行失败：{_e_base_ins}") from _e_base_ins

            src_rows = [src_row for _pair_idx, src_row in run]
            max_col = max(1, ws_src_val.max_column or 1, ws_src_edit.max_column or 1)
            src_val_cache = _read_rows_into_cache(ws_src_val, src_rows, max_col)
            src_edit_cache = _read_rows_into_cache(ws_src_edit, src_rows, max_col)
            for offset, (_pair_idx, src_row) in enumerate(run):
                dst_row = insert_pos + offset
                row_val = _row_from_cache(src_val_cache, src_row, max_col)
                row_edit = _row_from_cache(src_edit_cache, src_row, max_col)
                for c in range(1, max_col + 1):
                    v_val = row_val[c - 1]
                    v_edit = row_edit[c - 1]
                    dst_cell_edit = ws_dst_edit.cell(row=dst_row, column=c)
                    new_edit = _copy_edit_value_for_destination(
                        v_val,
                        v_edit,
                        dst_cell_edit.value,
                        src_row=src_row,
                        src_col=c,
                        dst_row=dst_row,
                        dst_col=c,
                    )
                    ws_dst_val.cell(row=dst_row, column=c).value = v_val
                    _assign_edit_cell_value(dst_cell_edit, new_edit)
                    if direction == "B2A":
                        self.app.record_manual_a_cell(self.sheet, dst_row, c, new_edit)
                        if _formula_text(new_edit):
                            self.app.record_manual_a_formula_cache(self.sheet, dst_row, c, v_val)
                    else:
                        self.app.record_manual_b_cell(self.sheet, dst_row, c, new_edit)
                        if _formula_text(new_edit):
                            self.app.record_manual_b_formula_cache(self.sheet, dst_row, c, v_val)
                _copy_row_metadata(ws_src_edit, ws_dst_edit, src_row, dst_row, max_col)
                _copy_row_metadata(ws_src_val, ws_dst_val, src_row, dst_row, max_col)

            if direction == "B2A":
                self.app.modified_a = True
                self.app.modified_sheets_a.add(self.sheet)
                self.app.push_undo({
                    "sheet": self.sheet,
                    "target": "A_INSERT_ROW",
                    "row": insert_pos,
                    "base_row": base_insert_pos,
                    "count": count,
                    "base_inserted": base_inserted,
                })
            else:
                self.app.modified_b = True
                self.app.modified_sheets_b.add(self.sheet)
                self.app.push_undo({
                    "sheet": self.sheet,
                    "target": "B_INSERT_ROW",
                    "row": insert_pos,
                    "count": count,
                    "base_inserted": False,
                })

            amount = count
            run_pair_to_dst = {pair_idx: insert_pos + offset for offset, (pair_idx, _src_row) in enumerate(run)}
            updated_pairs: list[tuple[int | None, int | None]] = []
            for idx, (ra, rb) in enumerate(self.row_pairs):
                if direction == "B2A":
                    ra = self._shift_row_number_for_insert(ra, insert_pos, amount)
                    if idx in run_pair_to_dst:
                        ra = run_pair_to_dst[idx]
                else:
                    rb = self._shift_row_number_for_insert(rb, insert_pos, amount)
                    if idx in run_pair_to_dst:
                        rb = run_pair_to_dst[idx]
                updated_pairs.append((ra, rb))
            self.row_pairs = updated_pairs
            self._rebuild_row_pair_lookup_maps()

            if self._is_three_way_enabled() and getattr(self.app, "has_base", False):
                def _shift_base_row(base_row: int | None) -> int | None:
                    if not base_inserted:
                        return base_row
                    return self._shift_row_number_for_insert(base_row, base_insert_pos, amount)

                mine_map_old = dict(getattr(self, "mine_to_base_row", {}) or {})
                theirs_map_old = dict(getattr(self, "theirs_to_base_row", {}) or {})
                overrides_old = dict(getattr(self, "pair_base_row_override", {}) or {})

                if direction == "B2A":
                    mine_map_new: dict[int, int] = {}
                    for ra, base_row in mine_map_old.items():
                        new_ra = self._shift_row_number_for_insert(ra, insert_pos, amount)
                        if new_ra is None:
                            continue
                        mine_map_new[int(new_ra)] = _shift_base_row(base_row)
                    for pair_idx, dst_row in run_pair_to_dst.items():
                        base_row = _shift_base_row(old_base_rows.get(pair_idx))
                        if base_row is not None:
                            mine_map_new[int(dst_row)] = int(base_row)
                    self.mine_to_base_row = mine_map_new

                    theirs_map_new: dict[int, int] = {}
                    for rb, base_row in theirs_map_old.items():
                        if rb is None:
                            continue
                        theirs_map_new[int(rb)] = _shift_base_row(base_row)
                    self.theirs_to_base_row = theirs_map_new
                else:
                    mine_map_new: dict[int, int] = {}
                    for ra, base_row in mine_map_old.items():
                        if ra is None:
                            continue
                        mine_map_new[int(ra)] = _shift_base_row(base_row)
                    self.mine_to_base_row = mine_map_new

                    theirs_map_new: dict[int, int] = {}
                    for rb, base_row in theirs_map_old.items():
                        new_rb = self._shift_row_number_for_insert(rb, insert_pos, amount)
                        if new_rb is None:
                            continue
                        theirs_map_new[int(new_rb)] = _shift_base_row(base_row)
                    for pair_idx, dst_row in run_pair_to_dst.items():
                        base_row = _shift_base_row(old_base_rows.get(pair_idx))
                        if base_row is not None:
                            theirs_map_new[int(dst_row)] = int(base_row)
                    self.theirs_to_base_row = theirs_map_new

                overrides_new: dict[int, int | None] = {}
                for idx, base_row in overrides_old.items():
                    overrides_new[int(idx)] = _shift_base_row(base_row)
                for pair_idx in run_pair_to_dst:
                    base_row = _shift_base_row(old_base_rows.get(pair_idx))
                    if base_row is not None:
                        overrides_new[int(pair_idx)] = int(base_row)
                    else:
                        overrides_new.pop(int(pair_idx), None)
                self.pair_base_row_override = overrides_new

            try:
                self.max_row = max(
                    int(self.max_row or 1),
                    int(getattr(self.app.ws_a_val(self.sheet), "max_row", 1) or 1),
                    int(getattr(self.app.ws_b_val(self.sheet), "max_row", 1) or 1),
                )
            except Exception:
                pass

            for offset, (pair_idx, src_row) in enumerate(run):
                touched_row = insert_pos + offset if direction == "B2A" else src_row
                self.touched_rows.add(int(touched_row))
                self._prime_pair_cache_after_insert(
                    pair_idx=pair_idx,
                    ws_a_val=ws_a_val,
                    ws_b_val=ws_b_val,
                    ws_a_edit=ws_a_edit,
                    ws_b_edit=ws_b_edit,
                )

            committed = True

            if not suppress_refresh:
                try:
                    self._invalidate_only_diff_snapshot_cache()
                    self._invalidate_render_cache()
                    self.refresh(row_only=None, rescan=False)
                    if anchor is not None:
                        self._restore_view_anchor(anchor)
                    self._update_cursor_lines()
                except Exception as refresh_error:
                    _dlog(
                        f"batch row insert committed but refresh failed: sheet={self.sheet} "
                        f"row={insert_pos} count={count} err={refresh_error}"
                    )
                    try:
                        messagebox.showwarning(
                            "行已插入",
                            "数据已成功批量插入，但界面刷新失败。请点击“刷新本Sheet”查看最新结果。\n"
                            f"详情：{refresh_error}",
                        )
                    except Exception:
                        pass
            return True
        except Exception as e:
            if not committed:
                try:
                    if base_edit_inserted and ws_be is not None and base_insert_pos is not None:
                        ws_be.delete_rows(int(base_insert_pos), int(count))
                    if base_val_inserted and ws_bv is not None and base_insert_pos is not None:
                        ws_bv.delete_rows(int(base_insert_pos), int(count))
                    if dst_edit_inserted and ws_dst_edit is not None and insert_pos is not None:
                        ws_dst_edit.delete_rows(int(insert_pos), int(count))
                    if dst_val_inserted and ws_dst_val is not None and insert_pos is not None:
                        ws_dst_val.delete_rows(int(insert_pos), int(count))
                    self.app.manual_a_cell_ops.clear()
                    self.app.manual_a_cell_ops.update(manual_a_cells_before)
                    self.app.manual_a_formula_cache_ops.clear()
                    self.app.manual_a_formula_cache_ops.update(manual_a_cache_before)
                    self.app.manual_b_cell_ops.clear()
                    self.app.manual_b_cell_ops.update(manual_b_cells_before)
                    self.app.manual_b_formula_cache_ops.clear()
                    self.app.manual_b_formula_cache_ops.update(manual_b_cache_before)
                    del self.app.manual_a_row_ops[manual_row_ops_len:]
                    del self.app.manual_b_row_ops[manual_b_row_ops_len:]
                    del self.app.undo_stack[undo_len:]
                    self.app.modified_a = modified_a_before
                    self.app.modified_b = modified_b_before
                    self.app.modified_sheets_a.clear()
                    self.app.modified_sheets_a.update(modified_sheets_a_before)
                    self.app.modified_sheets_b.clear()
                    self.app.modified_sheets_b.update(modified_sheets_b_before)
                    self.row_pairs = row_pairs_before
                    self.mine_to_base_row = mine_to_base_before
                    self.theirs_to_base_row = theirs_to_base_before
                    self.pair_base_row_override = overrides_before
                    self.touched_rows = touched_before
                    self.max_row = max_row_before
                    self._rebuild_row_pair_lookup_maps()
                    for cache_map, saved_entries in zip(cache_maps, pair_cache_before):
                        for idx, (existed, value) in saved_entries.items():
                            if existed:
                                cache_map[idx] = value
                            else:
                                cache_map.pop(idx, None)
                    _dlog(
                        f"batch row insert rolled back: sheet={self.sheet} "
                        f"row={insert_pos} count={count}"
                    )
                except Exception as rollback_error:
                    _dlog(
                        f"CRITICAL batch row insert rollback failed: sheet={self.sheet} "
                        f"row={insert_pos} count={count} err={rollback_error}"
                    )
            messagebox.showerror("Error", f"批量插入行失败：\n{e}")
            return False

    def _insert_row_copy(
        self,
        pair_idx: int,
        direction: str,
        src_row: int,
        suppress_refresh: bool,
        _undo_out,
        anchor,
    ) -> bool:
        """Insert a new row in the destination worksheet and copy src_row data into it.

        Called by _copy_selected_row when the destination side has no paired row:
        - direction="B2A": rb exists, ra is None → insert new row in A.
        - direction="A2B": ra exists, rb is None → insert new row in B.

        In 3-way mode, B2A also inserts an empty row in Base at the same position
        to keep A-Base row-number alignment intact.
        """
        committed = False
        dst_val_inserted = False
        dst_edit_inserted = False
        base_val_inserted = False
        base_edit_inserted = False
        ws_dst_val = None
        ws_dst_edit = None
        ws_bv = None
        ws_be = None
        insert_pos = None
        base_insert_pos = None
        manual_a_cells_before = dict(getattr(self.app, "manual_a_cell_ops", {}) or {})
        manual_a_cache_before = dict(getattr(self.app, "manual_a_formula_cache_ops", {}) or {})
        manual_b_cells_before = dict(getattr(self.app, "manual_b_cell_ops", {}) or {})
        manual_b_cache_before = dict(getattr(self.app, "manual_b_formula_cache_ops", {}) or {})
        manual_row_ops_len = len(getattr(self.app, "manual_a_row_ops", []) or [])
        manual_b_row_ops_len = len(getattr(self.app, "manual_b_row_ops", []) or [])
        undo_len = len(getattr(self.app, "undo_stack", []) or [])
        modified_a_before = bool(getattr(self.app, "modified_a", False))
        modified_b_before = bool(getattr(self.app, "modified_b", False))
        modified_sheets_a_before = set(getattr(self.app, "modified_sheets_a", set()) or set())
        modified_sheets_b_before = set(getattr(self.app, "modified_sheets_b", set()) or set())
        try:
            ws_a_val = self.app.ws_a_val(self.sheet)
            ws_b_val = self.app.ws_b_val(self.sheet)
            ws_a_edit = self.app.ws_a_edit(self.sheet)
            ws_b_edit = self.app.ws_b_edit(self.sheet)
            old_pair = self.row_pairs[pair_idx] if 0 <= int(pair_idx) < len(self.row_pairs) else None

            if direction == "B2A":
                ws_dst_val = ws_a_val
                ws_dst_edit = ws_a_edit
                ws_src_val = ws_b_val
                ws_src_edit = ws_b_edit
                insert_pos = self._find_row_insert_position(pair_idx, "A")
            else:  # A2B
                ws_dst_val = ws_b_val
                ws_dst_edit = ws_b_edit
                ws_src_val = ws_a_val
                ws_src_edit = ws_a_edit
                insert_pos = self._find_row_insert_position(pair_idx, "B")

            # Insert blank row in destination worksheet.
            ws_dst_val.insert_rows(idx=insert_pos)
            dst_val_inserted = True
            ws_dst_edit.insert_rows(idx=insert_pos)
            dst_edit_inserted = True

            # When inserting into A, shift any existing manual-edit records whose
            # row numbers are >= insert_pos (they moved up by one in the worksheet).
            if direction == "B2A":
                for op_map in (
                    self.app.manual_a_cell_ops,
                    self.app.manual_a_formula_cache_ops,
                ):
                    to_shift = {k: v for k, v in op_map.items()
                                if k[0] == self.sheet and k[1] >= insert_pos}
                    for k in to_shift:
                        del op_map[k]
                    for (s, r, c), v in to_shift.items():
                        op_map[(s, r + 1, c)] = v
                self.app.record_manual_a_row_insert(
                    self.sheet,
                    insert_pos,
                    1,
                    source_side="B",
                    source_rows=[src_row],
                )
            else:
                for op_map in (
                    self.app.manual_b_cell_ops,
                    self.app.manual_b_formula_cache_ops,
                ):
                    to_shift = {
                        key: value for key, value in op_map.items()
                        if key[0] == self.sheet and key[1] >= insert_pos
                    }
                    for key in to_shift:
                        del op_map[key]
                    for (sheet_name, row_idx, col_idx), value in to_shift.items():
                        op_map[(sheet_name, row_idx + 1, col_idx)] = value
                self.app.record_manual_b_row_insert(
                    self.sheet,
                    insert_pos,
                    1,
                    source_side="A",
                    source_rows=[src_row],
                )

            # In 3-way mode, inserting into A also inserts an empty row in Base.
            # The base position is mapped via base coordinates (not mine's
            # insert_pos) so non-identity mine/base alignment stays correct; the
            # row-model update shifts base rows using the same base_insert_pos.
            base_inserted = False
            base_insert_pos = insert_pos
            if direction == "B2A" and self._is_three_way_enabled():
                try:
                    base_insert_pos = self._find_base_row_insert_position(pair_idx)
                    ws_bv = self.app.ws_base_val(self.sheet)
                    ws_be = self.app.ws_base_edit(self.sheet)
                    if ws_bv is None or ws_be is None:
                        raise RuntimeError("BASE worksheet is unavailable")
                    ws_bv.insert_rows(idx=base_insert_pos)
                    base_val_inserted = True
                    ws_be.insert_rows(idx=base_insert_pos)
                    base_edit_inserted = True
                    base_inserted = True
                except Exception as _e_base_ins:
                    _dlog(f"base insert failed: sheet={self.sheet} pos={base_insert_pos} err={_e_base_ins}")
                    raise RuntimeError(f"向 BASE 插入对齐行失败：{_e_base_ins}") from _e_base_ins

            # Copy cell values from src_row into the newly inserted row.
            max_col = max(1, ws_src_val.max_column or 1, ws_src_edit.max_column or 1)
            for c in range(1, max_col + 1):
                v_val = ws_src_val.cell(row=src_row, column=c).value
                v_edit = ws_src_edit.cell(row=src_row, column=c).value
                dst_cell_edit = ws_dst_edit.cell(row=insert_pos, column=c)
                new_edit = _copy_edit_value_for_destination(
                    v_val,
                    v_edit,
                    dst_cell_edit.value,
                    src_row=src_row,
                    src_col=c,
                    dst_row=insert_pos,
                    dst_col=c,
                )
                ws_dst_val.cell(row=insert_pos, column=c).value = v_val
                _assign_edit_cell_value(dst_cell_edit, new_edit)
                if direction == "B2A":
                    self.app.record_manual_a_cell(self.sheet, insert_pos, c, new_edit)
                    if _formula_text(new_edit):
                        self.app.record_manual_a_formula_cache(self.sheet, insert_pos, c, v_val)
                else:
                    self.app.record_manual_b_cell(self.sheet, insert_pos, c, new_edit)
                    if _formula_text(new_edit):
                        self.app.record_manual_b_formula_cache(self.sheet, insert_pos, c, v_val)

            _copy_row_metadata(ws_src_edit, ws_dst_edit, src_row, insert_pos, max_col)
            _copy_row_metadata(ws_src_val, ws_dst_val, src_row, insert_pos, max_col)

            if direction == "B2A":
                self.app.modified_a = True
                self.app.modified_sheets_a.add(self.sheet)
            else:
                self.app.modified_b = True
                self.app.modified_sheets_b.add(self.sheet)

            target_tag = "A_INSERT_ROW" if direction == "B2A" else "B_INSERT_ROW"
            self.app.push_undo({
                "sheet": self.sheet,
                "target": target_tag,
                "row": insert_pos,
                "base_row": base_insert_pos,
                "count": 1,
                "base_inserted": base_inserted,
            })
            committed = True

            if not suppress_refresh:
                try:
                    fast_done = self._try_fast_refresh_after_row_insert(
                        pair_idx=pair_idx,
                        direction=direction,
                        insert_pos=insert_pos,
                        base_insert_pos=base_insert_pos,
                        old_pair=old_pair,
                        base_inserted=base_inserted,
                        suppress_refresh=suppress_refresh,
                        anchor=anchor,
                        ws_a_val=ws_a_val,
                        ws_b_val=ws_b_val,
                        ws_a_edit=ws_a_edit,
                        ws_b_edit=ws_b_edit,
                    )
                    if not fast_done:
                        self._invalidate_only_diff_snapshot_cache()
                        self._invalidate_render_cache()
                        self.refresh(row_only=None, rescan=True)
                        if anchor is not None:
                            self._restore_view_anchor(anchor)
                        self._update_cursor_lines()
                except Exception as refresh_error:
                    _dlog(
                        f"row insert committed but refresh failed: sheet={self.sheet} "
                        f"row={insert_pos} err={refresh_error}"
                    )
                    try:
                        messagebox.showwarning(
                            "行已插入",
                            "数据已成功插入，但界面刷新失败。请点击“刷新本Sheet”查看最新结果。\n"
                            f"详情：{refresh_error}",
                        )
                    except Exception:
                        pass
            return True
        except Exception as e:
            if not committed:
                try:
                    if base_edit_inserted and ws_be is not None and base_insert_pos is not None:
                        ws_be.delete_rows(int(base_insert_pos), 1)
                    if base_val_inserted and ws_bv is not None and base_insert_pos is not None:
                        ws_bv.delete_rows(int(base_insert_pos), 1)
                    if dst_edit_inserted and ws_dst_edit is not None and insert_pos is not None:
                        ws_dst_edit.delete_rows(int(insert_pos), 1)
                    if dst_val_inserted and ws_dst_val is not None and insert_pos is not None:
                        ws_dst_val.delete_rows(int(insert_pos), 1)
                    self.app.manual_a_cell_ops.clear()
                    self.app.manual_a_cell_ops.update(manual_a_cells_before)
                    self.app.manual_a_formula_cache_ops.clear()
                    self.app.manual_a_formula_cache_ops.update(manual_a_cache_before)
                    self.app.manual_b_cell_ops.clear()
                    self.app.manual_b_cell_ops.update(manual_b_cells_before)
                    self.app.manual_b_formula_cache_ops.clear()
                    self.app.manual_b_formula_cache_ops.update(manual_b_cache_before)
                    del self.app.manual_a_row_ops[manual_row_ops_len:]
                    del self.app.manual_b_row_ops[manual_b_row_ops_len:]
                    del self.app.undo_stack[undo_len:]
                    self.app.modified_a = modified_a_before
                    self.app.modified_b = modified_b_before
                    self.app.modified_sheets_a.clear()
                    self.app.modified_sheets_a.update(modified_sheets_a_before)
                    self.app.modified_sheets_b.clear()
                    self.app.modified_sheets_b.update(modified_sheets_b_before)
                    _dlog(f"row insert rolled back: sheet={self.sheet} row={insert_pos}")
                except Exception as rollback_error:
                    _dlog(
                        f"CRITICAL row insert rollback failed: sheet={self.sheet} "
                        f"row={insert_pos} err={rollback_error}"
                    )
            messagebox.showerror("Error", f"插入行失败：\n{e}")
            return False

    def _copy_selected_row(
        self,
        direction: str,
        row_header: bool = False,
        override_pair_idx: int | None = None,
        override_cols: set[int] | None = None,
        suppress_refresh: bool = False,
        _undo_out: list | None = None,
    ) -> bool:
        t0 = datetime.now()
        formula_skip_before = int(getattr(self, "_formula_copy_skips_pending", 0))
        interactive_busy = not suppress_refresh
        previous_bg_suppression = bool(getattr(self, "_suppress_bg_apply", False))
        begin_interactive = getattr(self.app, "_begin_interactive_action", None)
        end_interactive = getattr(self.app, "_end_interactive_action", None)
        if callable(begin_interactive):
            begin_interactive()
        self._suppress_bg_apply = True
        if interactive_busy:
            action_text = {
                "B2A": "正在采用 theirs 行到 mine...",
                "A2B": "正在采用 mine 行到 theirs...",
                "BASE2A": "正在采用 Base 行到 mine...",
            }.get(direction, "正在处理行操作...")
            try:
                self.info.configure(text=action_text)
                self.root.configure(cursor="watch")
                self.root.update_idletasks()
            except Exception:
                pass
        try:
            if self._is_missing_sheet_view():
                self._copy_missing_sheet(direction)
                return True
            anchor = None if suppress_refresh else self._capture_view_anchor()
            resolved_only = False
            changed = False
            # use last selected excel row (set on click); fallback to cursor line
            pair_idx = override_pair_idx if override_pair_idx is not None else self.selected_pair_idx
            if pair_idx is None:
                widget = self.left
                try:
                    focus = self.root.focus_get()
                    if focus == self.right:
                        widget = self.right
                except Exception:
                    pass
                try:
                    line = int((widget.index("insert").split(".")[0]))
                except Exception:
                    line = 1
                if not (1 <= line <= len(self.display_rows)):
                    return False
                pair_idx = self.display_rows[line - 1]
            pair = self.row_pairs[pair_idx] if pair_idx is not None and pair_idx < len(self.row_pairs) else None
            ra = self._row_for_side(pair, "A")
            rb = self._row_for_side(pair, "B")
            if direction == "A2B":
                if ra is None:
                    return False
                src_r = ra
                dst_r = rb  # may be None if A-only row; insert path handled in second block
            elif direction == "MINE2A":
                if ra is None and rb is None:
                    return False
                src_r = ra if ra is not None else rb
                dst_r = ra if ra is not None else rb
            elif direction == "BASE2A":
                if ra is None:
                    return False
                src_r = self._base_row_for_pair(pair_idx, pair)
                dst_r = ra
            else:
                if rb is None:
                    return False
                src_r = rb
                dst_r = ra  # may be None if B-only row; insert path handled in second block
            ws_a_val = self.app.ws_a_val(self.sheet)
            ws_b_val = self.app.ws_b_val(self.sheet)
            ws_base_val = self.app.ws_base_val(self.sheet) if getattr(self.app, "has_base", False) else None
            ws_a_edit = self.app.ws_a_edit(self.sheet)
            ws_b_edit = self.app.ws_b_edit(self.sheet)
            ws_base_edit = self.app.ws_base_edit(self.sheet) if getattr(self.app, "has_base", False) else None

            # Default row action overwrites full row range.
            full_max_col = max(
                self.max_col,
                ws_a_val.max_column or 1,
                ws_b_val.max_column or 1,
                (ws_base_val.max_column or 1) if ws_base_val is not None else 1,
                ws_a_edit.max_column or 1,
                ws_b_edit.max_column or 1,
                (ws_base_edit.max_column or 1) if ws_base_edit is not None else 1,
            )
            action_direction = direction
            cols = set(range(1, full_max_col + 1)) if override_cols is None else set(override_cols)

            # 3-way row-header behavior:
            # - Base row number: apply only diff cells to mine
            # - Theirs row number: apply full row to mine
            if row_header and self._is_three_way_enabled() and direction == "BASE2A":
                # Use base-vs-mine diffs for base-row action (not mine-vs-theirs).
                cols = self._base_to_mine_diff_cols(ra, rb, full_max_col)

            # Recompute src/dst based on final action direction.
            if action_direction == "A2B":
                if ra is None:
                    return False
                if rb is None:
                    # A-only row: insert a new row in B at the corresponding position.
                    return self._insert_row_copy(pair_idx, "A2B", ra, suppress_refresh, _undo_out, anchor)
                src_r = ra
                dst_r = rb
            elif action_direction == "MINE2A":
                if ra is None and rb is None:
                    return False
                src_r = ra if ra is not None else rb
                dst_r = ra if ra is not None else rb
            elif action_direction == "BASE2A":
                if ra is None:
                    return False
                src_r = self._base_row_for_pair(pair_idx, pair)
                dst_r = ra
            else:
                if rb is None:
                    return False
                if ra is None:
                    # B-only row: insert a new row in A at the corresponding position.
                    return self._insert_row_copy(pair_idx, "B2A", rb, suppress_refresh, _undo_out, anchor)
                src_r = rb
                dst_r = ra

            # Merge conflict mode:
            # - "A2B" means keep mine, just mark resolved.
            # - "B2A" means apply theirs to mine, then mark resolved.
            if getattr(self.app, "merge_conflict_mode", False):
                rows = self.app.merge_conflict_cells_by_sheet.get(self.sheet) if getattr(self.app, "merge_conflict_cells_by_sheet", None) else None
                conflict_row = ra or rb
                if rows and conflict_row in rows:
                    cols = set(rows.get(conflict_row, set())) if action_direction == "A2B" else cols
                if action_direction == "A2B":
                    self.app.user_touched_conflicts = True
                    self._resolve_conflict_row(conflict_row, cols)
                    resolved_only = True
                    changed = True
                elif action_direction == "MINE2A":
                    self.app.user_touched_conflicts = True
                    self._resolve_conflict_row(conflict_row, cols)
                    resolved_only = True
                    changed = True

            if not cols:
                return False

            if not resolved_only and action_direction in ("A2B", "B2A", "BASE2A"):
                if action_direction == "A2B":
                    src_edit_ws, src_val_ws, dst_edit_ws = ws_a_edit, ws_a_val, ws_b_edit
                elif action_direction == "B2A":
                    src_edit_ws, src_val_ws, dst_edit_ws = ws_b_edit, ws_b_val, ws_a_edit
                else:
                    src_edit_ws, src_val_ws, dst_edit_ws = ws_base_edit, ws_base_val, ws_a_edit
                if src_r is not None and src_edit_ws is not None:
                    for c in cols:
                        if int(c) <= 0:
                            continue
                        _copy_edit_value_for_destination(
                            src_val_ws.cell(row=src_r, column=c).value if src_val_ws is not None else None,
                            src_edit_ws.cell(row=src_r, column=c).value,
                            dst_edit_ws.cell(row=dst_r, column=c).value,
                            src_row=src_r, src_col=c, dst_row=dst_r, dst_col=c,
                        )

            if action_direction == "A2B":
                if not resolved_only:
                    undo_cells = []
                    for c in cols:
                        old_edit = ws_b_edit.cell(row=dst_r, column=c).value
                        old_val = ws_b_val.cell(row=dst_r, column=c).value
                        v_edit = ws_a_edit.cell(row=src_r, column=c).value
                        v_val = ws_a_val.cell(row=src_r, column=c).value
                        new_edit = _copy_edit_value_for_destination(
                            v_val, v_edit, old_edit,
                            src_row=src_r, src_col=c, dst_row=dst_r, dst_col=c,
                        )
                        formula_mode = self._same_formula_copy_mode(
                            new_edit, old_edit, v_val, old_val
                        )
                        if formula_mode == "noop":
                            continue
                        if formula_mode == "cache":
                            ws_b_val.cell(row=dst_r, column=c).value = v_val
                            self.app.record_manual_b_formula_cache(self.sheet, dst_r, c, v_val)
                            undo_cells.append((dst_r, c, old_edit, old_val))
                            continue
                        self.app.clear_manual_b_formula_cache(self.sheet, dst_r, c)
                        _assign_edit_cell_value(
                            ws_b_edit.cell(row=dst_r, column=c),
                            new_edit,
                        )
                        self.app.record_manual_b_cell(self.sheet, dst_r, c, new_edit)
                        ws_b_val.cell(row=dst_r, column=c).value = v_val
                        undo_cells.append((dst_r, c, old_edit, old_val))
                    self.app.modified_b = True
                    self.app.modified_sheets_b.add(self.sheet)
                    if undo_cells:
                        changed = True
                        if _undo_out is not None:
                            _undo_out.extend(undo_cells)
                        else:
                            self.app.push_undo({"sheet": self.sheet, "target": "B", "cells": undo_cells})
            elif action_direction == "MINE2A":
                # Keep mine row as-is. In conflict mode, row can still be "changed"
                # by conflict resolution metadata updates above.
                return bool(changed)
            elif action_direction == "B2A":
                undo_cells = []
                applied_cols = set()
                for c in cols:
                    old_edit = ws_a_edit.cell(row=dst_r, column=c).value
                    old_val = ws_a_val.cell(row=dst_r, column=c).value
                    v_edit = ws_b_edit.cell(row=src_r, column=c).value
                    v_val = ws_b_val.cell(row=src_r, column=c).value
                    new_edit = _copy_edit_value_for_destination(
                        v_val, v_edit, old_edit,
                        src_row=src_r, src_col=c, dst_row=dst_r, dst_col=c,
                    )
                    formula_mode = self._same_formula_copy_mode(
                        new_edit, old_edit, v_val, old_val
                    )
                    if formula_mode == "noop":
                        continue
                    if formula_mode == "cache":
                        ws_a_val.cell(row=dst_r, column=c).value = v_val
                        self.app.record_manual_a_formula_cache(self.sheet, dst_r, c, v_val)
                        undo_cells.append((dst_r, c, old_edit, old_val))
                        applied_cols.add(c)
                        continue
                    _assign_edit_cell_value(ws_a_edit.cell(row=dst_r, column=c), new_edit)
                    ws_a_val.cell(row=dst_r, column=c).value = v_val
                    self.app.record_manual_a_cell(self.sheet, dst_r, c, new_edit)
                    undo_cells.append((dst_r, c, old_edit, old_val))
                    applied_cols.add(c)
                if undo_cells:
                    self.app.modified_a = True
                    self.app.modified_sheets_a.add(self.sheet)
                    changed = True
                    if _undo_out is not None:
                        _undo_out.extend(undo_cells)
                    else:
                        self.app.push_undo({"sheet": self.sheet, "target": "A", "cells": undo_cells})
                # In conflict mode, B2A applies theirs; mark conflict resolved.
                if getattr(self.app, "merge_conflict_mode", False) and applied_cols:
                    self.app.user_touched_conflicts = True
                    self._resolve_conflict_row(conflict_row, applied_cols)
                    resolved_only = True
                    changed = True
            else:
                undo_cells = []
                applied_cols = set()
                if ws_base_edit is None or ws_base_val is None:
                    return False
                for c in cols:
                    old_edit = ws_a_edit.cell(row=dst_r, column=c).value
                    old_val = ws_a_val.cell(row=dst_r, column=c).value
                    if src_r is None:
                        v_edit = None
                        v_val = None
                    else:
                        v_edit = ws_base_edit.cell(row=src_r, column=c).value
                        v_val = ws_base_val.cell(row=src_r, column=c).value
                    new_edit = (
                        _copy_edit_value_for_destination(
                            v_val, v_edit, old_edit,
                            src_row=src_r, src_col=c, dst_row=dst_r, dst_col=c,
                        )
                        if src_r is not None else None
                    )
                    formula_mode = self._same_formula_copy_mode(
                        new_edit, old_edit, v_val, old_val
                    )
                    if formula_mode == "noop":
                        continue
                    if formula_mode == "cache":
                        ws_a_val.cell(row=dst_r, column=c).value = v_val
                        self.app.record_manual_a_formula_cache(self.sheet, dst_r, c, v_val)
                        undo_cells.append((dst_r, c, old_edit, old_val))
                        applied_cols.add(c)
                        continue
                    _assign_edit_cell_value(ws_a_edit.cell(row=dst_r, column=c), new_edit)
                    ws_a_val.cell(row=dst_r, column=c).value = v_val
                    self.app.record_manual_a_cell(self.sheet, dst_r, c, new_edit)
                    undo_cells.append((dst_r, c, old_edit, old_val))
                    applied_cols.add(c)
                if undo_cells:
                    self.app.modified_a = True
                    self.app.modified_sheets_a.add(self.sheet)
                    changed = True
                    if _undo_out is not None:
                        _undo_out.extend(undo_cells)
                    else:
                        self.app.push_undo({"sheet": self.sheet, "target": "A", "cells": undo_cells})
                if getattr(self.app, "merge_conflict_mode", False) and applied_cols:
                    self.app.user_touched_conflicts = True
                    self._resolve_conflict_row(dst_r, applied_cols)
                    resolved_only = True
                    changed = True

            if not changed:
                formula_skipped = int(getattr(self, "_formula_copy_skips_pending", 0)) - formula_skip_before
                if formula_skipped > 0 and not suppress_refresh:
                    self._show_formula_copy_skip_notice(formula_skipped)
                return False

            # Mark as touched: keep row visible in "只看差异" even if diffs are resolved.
            touched_r = ra or rb
            if touched_r is not None:
                self.touched_rows.add(touched_r)
            if not suppress_refresh:
                self._invalidate_only_diff_snapshot_cache()
                self._invalidate_render_cache()
                # Minimize flicker: use row-only incremental refresh after overwrite.
                if bool(self.only_diff_var.get()) and self.snapshot_only_diff:
                    self._recalc_row_diff_and_update(dst_r)
                self.refresh(row_only=dst_r, rescan=False)
                self._restore_view_anchor(anchor)
                self._update_cursor_lines()
                formula_skipped = int(getattr(self, "_formula_copy_skips_pending", 0)) - formula_skip_before
                if formula_skipped > 0:
                    self._show_formula_copy_skip_notice(formula_skipped)
            return True
        except Exception as e:
            messagebox.showerror("Error", f"覆盖整行失败：\n{e}")
            return False
        finally:
            self._suppress_bg_apply = previous_bg_suppression
            if callable(end_interactive):
                end_interactive()
            try:
                dt = (datetime.now() - t0).total_seconds() * 1000.0
                _dlog(f"OVERWRITE_ROW {self.sheet} dir={direction} ms={dt:.1f}")
            except Exception:
                pass
            if interactive_busy:
                try:
                    self.root.configure(cursor="")
                except Exception:
                    pass

    def _undo_last_action(self):
        try:
            action = self.app.pop_undo()
            if not action:
                return
            sheet = action.get("sheet")
            target = action.get("target")
            if target == "A_APPEND":
                start_row = action.get("start_row")
                count = action.get("count")
                if not start_row or not count:
                    return
                ws_edit = self.app.ws_a_edit(sheet)
                ws_val = self.app.ws_a_val(sheet)
                ws_edit.delete_rows(start_row, count)
                ws_val.delete_rows(start_row, count)
                try:
                    for op_map in (
                        self.app.manual_a_cell_ops,
                        self.app.manual_a_formula_cache_ops,
                    ):
                        keys_to_drop = [k for k in op_map.keys() if k[0] == sheet and k[1] >= int(start_row)]
                        for k in keys_to_drop:
                            op_map.pop(k, None)
                except Exception:
                    pass
                self.app.modified_a = True
                self.app.modified_sheets_a.add(sheet)
                if sheet == self.sheet:
                    self._invalidate_only_diff_snapshot_cache()
                    self._invalidate_render_cache()
                    self.refresh(row_only=None, rescan=True)
                    self._update_cursor_lines()
                return
            if target in ("A_INSERT_ROW", "B_INSERT_ROW"):
                row = action.get("row", 1)
                count = action.get("count", 1)
                base_inserted = action.get("base_inserted", False)
                base_row_del = action.get("base_row", row)
                if target == "A_INSERT_ROW":
                    ws_edit = self.app.ws_a_edit(sheet)
                    ws_val = self.app.ws_a_val(sheet)
                    ws_edit.delete_rows(row, count)
                    ws_val.delete_rows(row, count)
                    try:
                        row_int = int(row)
                        # Drop manual-edit records for the deleted row(s).
                        for op_map in (
                            self.app.manual_a_cell_ops,
                            self.app.manual_a_formula_cache_ops,
                        ):
                            to_drop = [k for k in op_map
                                       if k[0] == sheet and row_int <= k[1] < row_int + count]
                            for k in to_drop:
                                del op_map[k]
                            # Shift records for rows that moved up after the delete.
                            to_shift = {k: v for k, v in op_map.items()
                                        if k[0] == sheet and k[1] >= row_int + count}
                            for k in to_shift:
                                del op_map[k]
                            for (s, r, c), v in to_shift.items():
                                op_map[(s, r - count, c)] = v
                    except Exception:
                        pass
                    self.app.remove_last_manual_a_row_insert(sheet, row, count)
                    self.app.modified_a = True
                    self.app.modified_sheets_a.add(sheet)
                else:
                    ws_edit = self.app.ws_b_edit(sheet)
                    ws_val = self.app.ws_b_val(sheet)
                    ws_edit.delete_rows(row, count)
                    ws_val.delete_rows(row, count)
                    try:
                        row_int = int(row)
                        for op_map in (
                            self.app.manual_b_cell_ops,
                            self.app.manual_b_formula_cache_ops,
                        ):
                            to_drop = [
                                key for key in op_map
                                if key[0] == sheet and row_int <= key[1] < row_int + count
                            ]
                            for key in to_drop:
                                del op_map[key]
                            to_shift = {
                                key: value for key, value in op_map.items()
                                if key[0] == sheet and key[1] >= row_int + count
                            }
                            for key in to_shift:
                                del op_map[key]
                            for (sheet_name, row_idx, col_idx), value in to_shift.items():
                                op_map[(sheet_name, row_idx - count, col_idx)] = value
                    except Exception as cache_shift_error:
                        _dlog(
                            f"manual_b_formula_cache_ops undo shift failed: "
                            f"sheet={sheet} row={row} err={cache_shift_error}"
                        )
                    self.app.remove_last_manual_b_row_insert(sheet, row, count)
                    self.app.modified_b = True
                    self.app.modified_sheets_b.add(sheet)
                if base_inserted:
                    try:
                        ws_bv = self.app.ws_base_val(sheet)
                        ws_be = self.app.ws_base_edit(sheet)
                        if ws_bv is not None and ws_be is not None:
                            ws_bv.delete_rows(base_row_del, count)
                            ws_be.delete_rows(base_row_del, count)
                    except Exception:
                        pass
                if sheet == self.sheet:
                    self._invalidate_only_diff_snapshot_cache()
                    self._invalidate_render_cache()
                    self.refresh(row_only=None, rescan=True)
                    self._update_cursor_lines()
                return
            cells = action.get("cells", [])
            if not cells:
                return
            if target == "A":
                ws_edit = self.app.ws_a_edit(sheet)
                ws_val = self.app.ws_a_val(sheet)
                self.app.modified_a = True
                self.app.modified_sheets_a.add(sheet)
            else:
                ws_edit = self.app.ws_b_edit(sheet)
                ws_val = self.app.ws_b_val(sheet)
                self.app.modified_b = True
                self.app.modified_sheets_b.add(sheet)
            rows = set()
            for r, c, old_edit, old_val in cells:
                restored_edit = _choose_edit_value(old_val, old_edit)
                _assign_edit_cell_value(ws_edit.cell(row=r, column=c), restored_edit)
                ws_val.cell(row=r, column=c).value = old_val
                if target == "A":
                    self.app.record_manual_a_cell(sheet, r, c, restored_edit)
                    if _formula_text(restored_edit):
                        self.app.record_manual_a_formula_cache(sheet, r, c, old_val)
                else:
                    self.app.record_manual_b_cell(sheet, r, c, restored_edit)
                    self.app.clear_manual_b_formula_cache(sheet, r, c)
                    if _formula_text(restored_edit):
                        self.app.record_manual_b_formula_cache(sheet, r, c, old_val)
                rows.add(r)
            # refresh current sheet if applicable
            if sheet == self.sheet:
                for r in rows:
                    self.touched_rows.add(r)
                self._invalidate_only_diff_snapshot_cache()
                if self._align_rows_enabled:
                    # Full refresh supersedes per-row work; avoid N redundant partial renders.
                    self.refresh(row_only=None, rescan=True)
                else:
                    for r in rows:
                        if bool(self.only_diff_var.get()) and self.snapshot_only_diff:
                            self._recalc_row_diff_and_update(r)
                        self.refresh(row_only=r, rescan=False)
                self._update_cursor_lines()
        except Exception as e:
            messagebox.showerror("撤销失败", f"撤销操作失败，数据可能未恢复：\n{e}")

    def _resolve_conflict_cell(self, r: int, c: int):
        try:
            if self.app.resolve_conflict_cell(self.sheet, r, c):
                # update view based on updated conflict map
                self.refresh(row_only=None, rescan=False)
                self._update_cursor_lines()
        except Exception:
            pass

    def _resolve_conflict_row(self, r: int, cols):
        try:
            if self.app.resolve_conflict_row(self.sheet, r, cols):
                self.refresh(row_only=None, rescan=False)
                self._update_cursor_lines()
        except Exception:
            pass

    def _refresh_row_text_only(self, r: int):
        """Update the rendered row text for row r without recomputing diff_cols_by_row."""
        try:
            ws_a = self.app.ws_a_val(self.sheet)
            ws_b = self.app.ws_b_val(self.sheet)
            ws_a_edit = self.app.ws_a_edit(self.sheet)
            ws_b_edit = self.app.ws_b_edit(self.sheet)

            self.max_col = max(ws_a.max_column or 1, ws_b.max_column or 1)
            self.col_max_a = ws_a.max_column or 1
            self.col_max_b = ws_b.max_column or 1

            pair_idx = self.row_a_to_pair_idx.get(r)
            if pair_idx is None:
                pair_idx = self.row_b_to_pair_idx.get(r)
            if pair_idx is None:
                return
            ra, rb = self.row_pairs[pair_idx]
            line_a, line_b, _cols = self._build_row_and_diff_pair(ws_a, ws_b, ws_a_edit, ws_b_edit, ra, rb)
            self.pair_text_a[pair_idx] = line_a
            self.pair_text_b[pair_idx] = line_b

            line = self.row_to_line.get(pair_idx)
            if line is None:
                return

            self.left.delete(f"{line}.0", f"{line}.end")
            self._render_base_line(line, pair_idx)
            self.right.delete(f"{line}.0", f"{line}.end")
            self.left.insert(f"{line}.0", self.pair_text_a[pair_idx])
            self.right.insert(f"{line}.0", self.pair_text_b[pair_idx])
            self._render_row_header_line(line, pair_idx)
        except Exception:
            pass

    def _recalc_row_diff_and_update(self, r: int):
        """Recompute diff for row r and update its highlight, without changing the row list (snapshot mode)."""
        try:
            ws_a = self.app.ws_a_val(self.sheet)
            ws_b = self.app.ws_b_val(self.sheet)
            ws_a_edit = self.app.ws_a_edit(self.sheet)
            ws_b_edit = self.app.ws_b_edit(self.sheet)

            self.max_col = max(ws_a.max_column or 1, ws_b.max_column or 1)
            self.col_max_a = ws_a.max_column or 1
            self.col_max_b = ws_b.max_column or 1
            pair_idx = self.row_a_to_pair_idx.get(r)
            if pair_idx is None:
                pair_idx = self.row_b_to_pair_idx.get(r)
            if pair_idx is None:
                return
            if pair_idx >= len(self.row_pairs):
                return
            ra, rb = self.row_pairs[pair_idx]
            line_a, line_b, cols = self._build_row_and_diff_pair(ws_a, ws_b, ws_a_edit, ws_b_edit, ra, rb)
            self.pair_diff_cols[pair_idx] = cols
            if self._is_three_way_enabled() and getattr(self.app, "has_base", False):
                self.pair_base_diff_cols[pair_idx] = self._compute_base_diff_cols_for_pair(
                    pair_idx,
                    (ra, rb),
                    max_col=self.max_col,
                    ws_a_val=ws_a,
                    ws_a_edit=ws_a_edit,
                    ws_base_val=self.app.ws_base_val(self.sheet),
                    ws_base_edit=self.app.ws_base_edit(self.sheet),
                )
            else:
                self.pair_base_diff_cols.pop(pair_idx, None)
            self.pair_text_a[pair_idx] = line_a
            self.pair_text_b[pair_idx] = line_b

            line = self.row_to_line.get(pair_idx)
            if line is None:
                # if not visible and touched, rebuild snapshot to include it
                if bool(self.only_diff_var.get()) and (r in self.touched_rows):
                    self.refresh(row_only=None, rescan=False)
                return

            # update text
            self.left.delete(f"{line}.0", f"{line}.end")
            self._render_base_line(line, pair_idx)
            self.right.delete(f"{line}.0", f"{line}.end")
            self.left.insert(f"{line}.0", self.pair_text_a[pair_idx])
            self.right.insert(f"{line}.0", self.pair_text_b[pair_idx])
            self._render_row_header_line(line, pair_idx)

            # update tags for this line
            for w in (self.left, self.base, self.right):
                w.tag_remove("diffrow", f"{line}.0", f"{line}.end")
                w.tag_remove("diffcell", f"{line}.0", f"{line}.end")

            visual_cols = self._visual_diff_cols_for_pair(pair_idx)
            if visual_cols:
                self.left.tag_add("diffrow", f"{line}.0", f"{line}.end")
                self.right.tag_add("diffrow", f"{line}.0", f"{line}.end")
                self.base.tag_add("diffrow", f"{line}.0", f"{line}.end")
                base_line = self._build_base_line(pair_idx) if self._is_three_way_enabled() else ""
                left_args, base_args, right_args, _lr, _br, _rr = self._diffcell_tag_args_for_line(
                    line,
                    pair_idx,
                    self.pair_text_a.get(pair_idx, ""),
                    base_line,
                    self.pair_text_b.get(pair_idx, ""),
                )
                if left_args:
                    self.left.tag_add("diffcell", *left_args)
                if base_args:
                    self.base.tag_add("diffcell", *base_args)
                if right_args:
                    self.right.tag_add("diffcell", *right_args)
                self._clear_diffrow_under_diffcells(left_args, base_args, right_args)
        except Exception:
            pass

    def _invalidate_render_cache(self):
        self._data_version += 1
        self._render_cache.clear()
        self._diff_blocks_cache = None

    def _build_base_line(self, pair_idx: int) -> str:
        if not self._is_three_way_enabled():
            return ""
        if not getattr(self.app, "has_base", False):
            return ""
        if pair_idx >= len(self.row_pairs):
            return ""
        pair = self.row_pairs[pair_idx]
        if not pair:
            return ""
        r = self._base_row_for_pair(pair_idx, pair)
        if r is None:
            return ""
        try:
            ws_base = self._display_ws("BASE", edit=False)
        except Exception:
            return ""
        raw = []
        for c in range(1, self.max_col + 1):
            try:
                v = ws_base.cell(row=r, column=c).value
            except Exception:
                v = None
            raw.append(_val_to_str(v))
        grid_on = self._is_grid_overlay_enabled()
        sep = _COL_SEP if grid_on else "   "
        trail = " \u2502" if grid_on else ""
        cells = sep.join(_format_cell(raw[i], self.col_char_widths.get(i + 1, 1)) for i in range(len(raw))) + trail
        return cells

    def _render_base_full(self):
        if not self._is_three_way_enabled():
            try:
                self.base.delete("1.0", "end")
                self.base.tag_remove("selrow", "1.0", "end")
            except Exception:
                pass
            return []
        lines = []
        if self.display_rows and getattr(self, "_is_large_sheet", False):
            try:
                ws_base = self._display_ws("BASE", edit=False)
                base_rows_needed = []
                for pair_idx in self.display_rows:
                    base_row = self._base_row_for_pair(pair_idx)
                    if base_row is not None:
                        base_rows_needed.append(base_row)
                rows_base = _read_rows_into_cache(ws_base, base_rows_needed, self.max_col)
                for pair_idx in self.display_rows:
                    base_row = self._base_row_for_pair(pair_idx)
                    if base_row is None:
                        lines.append("")
                        continue
                    row_base = _row_from_cache(rows_base, base_row, self.max_col)
                    raw = [_val_to_str(v) for v in row_base]
                    lines.append(self._render_line_from_raw_parts(raw))
            except Exception:
                lines = [self._build_base_line(pair_idx) for pair_idx in self.display_rows]
        else:
            lines = [self._build_base_line(pair_idx) for pair_idx in self.display_rows]
        try:
            self.base.delete("1.0", "end")
            self.base.insert("1.0", "\n".join(lines) + ("\n" if lines else ""))
        except Exception:
            pass
        return lines

    def _render_base_line(self, line: int, pair_idx: int):
        if not self._is_three_way_enabled():
            return
        txt = self._build_base_line(pair_idx)
        try:
            self.base.delete(f"{line}.0", f"{line}.end")
            self.base.insert(f"{line}.0", txt)
        except Exception:
            pass

    def _row_label_for_pair_idx(self, pair_idx: int, side: str) -> str:
        try:
            if side == "BASE":
                r = self._base_row_for_pair(pair_idx)
            else:
                pair = self.row_pairs[pair_idx]
                r = self._row_for_side(pair, side)
            return str(r) if r is not None else ""
        except Exception:
            return ""

    def _rownum_render_width(self) -> int:
        max_label = int(self.max_row or 0)
        try:
            mine_base = getattr(self, "mine_to_base_row", {}) or {}
            theirs_base = getattr(self, "theirs_to_base_row", {}) or {}
            missing_base = getattr(self, "_missing_base_row_map", {}) or {}
            if mine_base:
                max_label = max(max_label, max(int(v) for v in mine_base.values() if v is not None))
            if theirs_base:
                max_label = max(max_label, max(int(v) for v in theirs_base.values() if v is not None))
            if missing_base:
                max_label = max(max_label, max(int(v) for v in missing_base.values() if v is not None))
        except Exception:
            pass
        return max(3, len(str(max(0, max_label))))

    def _sync_row_header_width_widgets(self):
        rn_w = self._rownum_render_width()
        header_w = max(4, rn_w + 1)
        self._rownum_display_width = rn_w
        if getattr(self, "_row_header_width", None) == header_w:
            return rn_w
        self._row_header_width = header_w
        widgets = [
            getattr(self, "left_ln", None),
            getattr(self, "base_ln", None),
            getattr(self, "right_ln", None),
            getattr(self, "left_corner_hdr", None),
            getattr(self, "base_corner_hdr", None),
            getattr(self, "right_corner_hdr", None),
            getattr(self, "cursor_cmp_corner", None),
            getattr(self, "cursor_cmp_ln", None),
        ]
        for w in widgets:
            if w is None:
                continue
            try:
                w.configure(width=header_w)
            except Exception:
                pass
        return rn_w

    def _render_row_headers_full(self):
        rn_w = self._sync_row_header_width_widgets()
        left_lines = []
        base_lines = []
        right_lines = []
        for pidx in self.display_rows:
            left_lines.append(self._row_label_for_pair_idx(pidx, "A").rjust(rn_w))
            base_lines.append(self._row_label_for_pair_idx(pidx, "BASE").rjust(rn_w))
            right_lines.append(self._row_label_for_pair_idx(pidx, "B").rjust(rn_w))
        for w, lines in ((self.left_ln, left_lines), (self.base_ln, base_lines), (self.right_ln, right_lines)):
            try:
                w.configure(state="normal")
                w.delete("1.0", "end")
                w.insert("1.0", "\n".join(lines) + ("\n" if lines else ""))
                w.tag_remove("diffrow", "1.0", "end")
                w.configure(state="disabled")
            except Exception:
                pass

    def _render_row_header_line(self, line: int, pair_idx: int):
        rn_w = self._sync_row_header_width_widgets()
        vals = (
            self._row_label_for_pair_idx(pair_idx, "A").rjust(rn_w),
            self._row_label_for_pair_idx(pair_idx, "BASE").rjust(rn_w),
            self._row_label_for_pair_idx(pair_idx, "B").rjust(rn_w),
        )
        for w, txt in ((self.left_ln, vals[0]), (self.base_ln, vals[1]), (self.right_ln, vals[2])):
            try:
                w.configure(state="normal")
                w.delete(f"{line}.0", f"{line}.end")
                w.insert(f"{line}.0", txt)
                w.tag_remove("diffrow", f"{line}.0", f"{line}.end")
                w.configure(state="disabled")
            except Exception:
                pass

    def _build_col_header_line(self) -> str:
        if self.max_col <= 0:
            return ""
        sep = _COL_SEP if self._is_grid_overlay_enabled() else "   "
        trail = " │" if self._is_grid_overlay_enabled() else ""
        parts = []
        for c in range(1, self.max_col + 1):
            label = get_column_letter(c)
            parts.append(_format_cell(label, self.col_char_widths.get(c, 1)))
        return sep.join(parts) + (trail if parts else "")

    def _render_col_headers(self):
        hdr = self._build_col_header_line()
        rn_w = self._sync_row_header_width_widgets()
        corner = "".rjust(rn_w)
        for w in (self.left_corner_hdr, self.base_corner_hdr, self.right_corner_hdr, self.cursor_cmp_corner):
            try:
                w.configure(state="normal")
                w.delete("1.0", "end")
                w.insert("1.0", corner)
                w.configure(state="disabled")
            except Exception:
                pass
        for w in (self.left_colhdr, self.base_colhdr, self.right_colhdr, self.cursor_cmp_colhdr):
            try:
                w.configure(state="normal")
                w.delete("1.0", "end")
                w.insert("1.0", hdr)
                w.configure(state="disabled")
            except Exception:
                pass

    def _render_cursor_row_headers(self, pair, is_three: bool):
        if not hasattr(self, "cursor_cmp_ln"):
            return
        rn_w = self._sync_row_header_width_widgets()
        ra = self._row_for_side(pair, "A") if pair else None
        rb = self._row_for_side(pair, "B") if pair else None
        rows = []
        if is_three:
            pair_idx = self._normalize_pair_idx(getattr(self, "_last_cursor_cmp_pair_idx", None))
            base_r = self._base_row_for_pair(pair_idx, pair) if pair and pair_idx is not None else None
            rows.append(str(base_r) if base_r is not None else "")
            rows.append(str(ra) if ra is not None else "")
        else:
            rows.append(str(ra) if ra is not None else "")
        rows.append(str(rb) if rb is not None else "")
        rows_txt = [r.rjust(rn_w) for r in rows]
        try:
            self.cursor_cmp_ln.configure(state="normal")
            self.cursor_cmp_ln.delete("1.0", "end")
            self.cursor_cmp_ln.insert("1.0", "\n".join(rows_txt) + ("\n" if rows_txt else ""))
            self.cursor_cmp_ln.configure(state="disabled")
        except Exception:
            pass

    # ---------- Rendering ----------
    def _load_all_rows(self):
        self._full_render = True
        self.refresh(row_only=None, rescan=False)

    def _append_rows(self, new_rows: list[int]):
        if not new_rows:
            return
        ws_a = self.app.ws_a_val(self.sheet)
        ws_b = self.app.ws_b_val(self.sheet)
        try:
            wb_a_edit = getattr(self.app, "_wb_a_edit", None)
            ws_a_edit = wb_a_edit[self.sheet] if wb_a_edit is not None else None
        except Exception:
            ws_a_edit = None
        try:
            wb_b_edit = getattr(self.app, "_wb_b_edit", None)
            ws_b_edit = wb_b_edit[self.sheet] if wb_b_edit is not None else None
        except Exception:
            ws_b_edit = None
        try:
            wb_base_edit = getattr(self.app, "_wb_base_edit", None)
            ws_base_edit_ready = wb_base_edit[self.sheet] if wb_base_edit is not None and getattr(self.app, "has_base", False) else None
        except Exception:
            ws_base_edit_ready = None

        missing_pair_text = any(
            pair_idx not in self.pair_text_a or pair_idx not in self.pair_text_b
            for pair_idx in new_rows
        )
        # Only rows not covered by the exact background cache need formula-text
        # fallback. Cached rows must never block scrolling/rendering on workbook I/O.
        if missing_pair_text and _USE_CACHED_VALUES_ONLY and (ws_a_edit is None or ws_b_edit is None):
            try:
                self.app._ensure_edit_loaded()
                ws_a_edit = self.app.ws_a_edit(self.sheet)
                ws_b_edit = self.app.ws_b_edit(self.sheet)
                _dlog(f"formula fallback enabled: loaded edit wbs for sheet={self.sheet}")
            except Exception as e:
                _dlog(f"formula fallback load failed: sheet={self.sheet} err={e}")

        start_line = len(self.display_rows) + 1
        # Preserve current scroll position to avoid jumps
        try:
            first, _last = self.left.yview()
        except Exception:
            first = None

        if self._is_three_way_enabled() and getattr(self.app, "has_base", False):
            self._ensure_base_diff_cache(
                pair_indices=new_rows,
                max_col=self.max_col,
                ws_a_val=ws_a,
                ws_a_edit=ws_a_edit if ws_a_edit is not None else ws_a,
                ws_base_val=self.app.ws_base_val(self.sheet),
                ws_base_edit=ws_base_edit_ready if ws_base_edit_ready is not None else self.app.ws_base_val(self.sheet),
            )

        for idx, pair_idx in enumerate(new_rows, start=0):
            if pair_idx not in self.pair_text_a or pair_idx not in self.pair_text_b:
                ra, rb = self.row_pairs[pair_idx]
                line_a, line_b, cols = self._build_row_and_diff_pair(ws_a, ws_b, ws_a_edit, ws_b_edit, ra, rb)
                self.pair_diff_cols[pair_idx] = cols
                self.pair_text_a[pair_idx] = line_a
                self.pair_text_b[pair_idx] = line_b
            else:
                line_a = self.pair_text_a.get(pair_idx, "")
                line_b = self.pair_text_b.get(pair_idx, "")
            visual_cols = self._visual_diff_cols_for_pair(pair_idx)

            line_no = start_line + idx
            self.left.insert("end", line_a + "\n")
            if self._is_three_way_enabled():
                self.base.insert("end", self._build_base_line(pair_idx) + "\n")
            self.right.insert("end", line_b + "\n")
            self._render_row_header_line(line_no, pair_idx)

            if visual_cols:
                self._display_diff_row_count += 1
                self.left.tag_add("diffrow", f"{line_no}.0", f"{line_no}.end")
                self.base.tag_add("diffrow", f"{line_no}.0", f"{line_no}.end")
                self.right.tag_add("diffrow", f"{line_no}.0", f"{line_no}.end")
                base_line = self._build_base_line(pair_idx) if self._is_three_way_enabled() else ""
                left_args, base_args, right_args, _lr, _br, _rr = self._diffcell_tag_args_for_line(
                    line_no,
                    pair_idx,
                    line_a,
                    base_line,
                    line_b,
                )
                if left_args:
                    self.left.tag_add("diffcell", *left_args)
                if base_args:
                    self.base.tag_add("diffcell", *base_args)
                if right_args:
                    self.right.tag_add("diffcell", *right_args)

        self.display_rows.extend(new_rows)
        for i, pair_idx in enumerate(new_rows, start=start_line):
            self.row_to_line[pair_idx] = i

        # row numbers are rendered in dedicated row-header widgets

        mode = "只看差异" if self.only_diff_var.get() else "全量"
        total_rows = len(self.row_pairs) if self.row_pairs else self.max_row
        self.info.configure(text=f"{mode} | RowsShown: {len(self.display_rows)} / {total_rows}   Cols: {self.max_col}   DiffRows: {self._display_diff_row_count}")

        if first is not None:
            try:
                self.left.yview_moveto(first)
                if self._is_three_way_enabled():
                    self.base.yview_moveto(first)
                self.right.yview_moveto(first)
            except Exception:
                pass

        self._invalidate_render_cache()

    def _maybe_load_more_rows(self, last_fraction: float):
        if not _FAST_OPEN_ENABLED:
            return
        try:
            last_fraction = float(last_fraction)
        except Exception:
            return
        if self._full_render:
            return
        if bool(self.only_diff_var.get()):
            return
        if getattr(self.app, "merge_conflict_mode", False):
            return
        # Only for full-list mode (not only-diff or conflict-only)
        if not self._full_display_rows:
            return
        if last_fraction < 0.98:
            return
        if len(self.display_rows) >= len(self._full_display_rows):
            return
        old_limit = len(self.display_rows)
        new_limit = min(len(self._full_display_rows), self._render_limit + _FAST_RENDER_BATCH)
        self._render_limit = new_limit
        new_rows = self._full_display_rows[old_limit:new_limit]
        self._append_rows(new_rows)

    def _refresh_missing_sheet_view(self, row_only: int | None, rescan: bool):
        meta = self._sheet_meta()
        ws_a = self._display_ws("A", edit=False)
        ws_b = self._display_ws("B", edit=False)
        ws_a_edit = self._display_ws("A", edit=True)
        ws_b_edit = self._display_ws("B", edit=True)
        ws_base = self._display_ws("BASE", edit=False) if getattr(self.app, "has_base", False) else None

        a_r, a_c = _effective_bounds_with_edit(ws_a, ws_a_edit) if meta.get("has_a") else (0, 0)
        b_r, b_c = _effective_bounds_with_edit(ws_b, ws_b_edit) if meta.get("has_b") else (0, 0)
        base_r, base_c = (
            _effective_bounds_with_edit(ws_base, self._display_ws("BASE", edit=True))
            if (ws_base is not None and meta.get("has_base")) else (0, 0)
        )
        self.max_row = max(1, a_r, b_r, base_r)
        self.max_col = max(1, a_c, b_c, base_c)
        self.col_max_a = max(1, a_c)
        self.col_max_b = max(1, b_c)
        self._bounds_checked = True
        self._is_large_sheet = False
        self._align_rows_enabled = False
        self._diff_partial = False
        self._full_render = True
        self._missing_base_row_map = {}
        self.mine_to_base_row = {r: r for r in range(1, min(a_r, base_r) + 1)} if meta.get("has_a") and meta.get("has_base") else {}
        self.theirs_to_base_row = {r: r for r in range(1, min(b_r, base_r) + 1)} if meta.get("has_b") and meta.get("has_base") else {}
        self.pair_base_row_override = {}

        self.row_pairs = []
        self.row_a_to_pair_idx = {}
        self.row_b_to_pair_idx = {}
        self.pair_text_a = {}
        self.pair_text_b = {}
        self.pair_diff_cols = {}
        self.pair_base_diff_cols = {}

        for r in range(1, self.max_row + 1):
            ra = r if meta.get("has_a") and r <= a_r else None
            rb = r if meta.get("has_b") and r <= b_r else None
            idx = len(self.row_pairs)
            self.row_pairs.append((ra, rb))
            if ra is not None:
                self.row_a_to_pair_idx[ra] = idx
            if rb is not None:
                self.row_b_to_pair_idx[rb] = idx
            if meta.get("has_base") and r <= base_r:
                self._missing_base_row_map[idx] = r

        self._prescan_col_widths(ws_a, ws_b, ws_base)
        for idx, (ra, rb) in enumerate(self.row_pairs):
            line_a, line_b, cols = self._build_row_and_diff_pair(ws_a, ws_b, ws_a_edit, ws_b_edit, ra, rb)
            if (ra is None and rb is None and idx in self._missing_base_row_map) or ((ra is None) != (rb is None)):
                cols = {-1}
            self.pair_text_a[idx] = line_a
            self.pair_text_b[idx] = line_b
            self.pair_diff_cols[idx] = cols

        self._data_ready = True
        self._full_display_rows = list(range(0, len(self.row_pairs)))
        self.display_rows = list(self._full_display_rows)
        self.row_to_line = {r: i + 1 for i, r in enumerate(self.display_rows)}

        try:
            self._render_col_headers()
        except Exception:
            pass
        self.left.delete("1.0", "end")
        self.base.delete("1.0", "end")
        self.right.delete("1.0", "end")
        for w in (self.left, self.base, self.right, self.left_ln, self.base_ln, self.right_ln):
            try:
                w.tag_remove("diffrow", "1.0", "end")
                w.tag_remove("diffcell", "1.0", "end")
                w.tag_remove("paddingrow", "1.0", "end")
            except Exception:
                pass
        lines_a = [self.pair_text_a.get(pair_idx, "") for pair_idx in self.display_rows]
        lines_b = [self.pair_text_b.get(pair_idx, "") for pair_idx in self.display_rows]
        self.left.insert("1.0", "\n".join(lines_a) + ("\n" if lines_a else ""))
        lines_base = self._render_base_full()
        self.right.insert("1.0", "\n".join(lines_b) + ("\n" if lines_b else ""))
        self._render_row_headers_full()

        diffrow_args = []
        pad_left = []
        pad_right = []
        for line_idx, pair_idx in enumerate(self.display_rows, start=1):
            if self._pair_has_visual_diff(pair_idx):
                diffrow_args.extend([f"{line_idx}.0", f"{line_idx}.end"])
            ra, rb = self.row_pairs[pair_idx]
            if ra is None:
                pad_left.extend([f"{line_idx}.0", f"{line_idx}.end"])
            if rb is None:
                pad_right.extend([f"{line_idx}.0", f"{line_idx}.end"])
        if diffrow_args:
            for w in (self.left, self.base, self.right, self.left_ln, self.base_ln, self.right_ln):
                w.tag_add("diffrow", *diffrow_args)
        if pad_left:
            self.left.tag_add("paddingrow", *pad_left)
        if pad_right:
            self.right.tag_add("paddingrow", *pad_right)

        diff_count = len(self.display_rows)
        self.info.configure(text=f"缺失Sheet对照 | RowsShown: {len(self.display_rows)} / {len(self.row_pairs)}   Cols: {self.max_col}   DiffRows: {diff_count}")
        self._display_diff_row_count = diff_count
        self.app.set_sheet_has_diff(self.sheet, diff_count > 0, confirmed=True)
        self.app.refresh_sheet_nav()
        self._update_diff_nav_state()
        try:
            self._update_diff_maps()
        except Exception:
            pass
        return

    def refresh(self, row_only: int | None, rescan: bool):
        _dlog(f"REFRESH sheet={self.sheet} row_only={row_only} rescan={rescan} only_diff={bool(self.only_diff_var.get())} raw={self.only_diff_var.get()}")
        if rescan and (not self._full_render):
            self._render_limit = _FAST_RENDER_ROW_LIMIT
        if self._is_missing_sheet_view():
            return self._refresh_missing_sheet_view(row_only, rescan)
        conflict_cells_by_row = None
        if getattr(self.app, "merge_conflict_mode", False):
            rows_map = getattr(self.app, "merge_conflict_cells_by_sheet", None)
            conflict_cells_by_row = rows_map.get(self.sheet) if rows_map else None
        ws_a = self.app.ws_a_val(self.sheet)
        ws_b = self.app.ws_b_val(self.sheet)
        # Non-blocking edit sheets: use loaded edit workbook if already available.
        # Do not trigger expensive load_workbook() during pure view refresh/toggle.
        try:
            wb_a_edit = getattr(self.app, "_wb_a_edit", None)
            ws_a_edit = wb_a_edit[self.sheet] if wb_a_edit is not None else None
        except Exception:
            ws_a_edit = None
        try:
            wb_b_edit = getattr(self.app, "_wb_b_edit", None)
            ws_b_edit = wb_b_edit[self.sheet] if wb_b_edit is not None else None
        except Exception:
            ws_b_edit = None
        try:
            wb_base_edit = getattr(self.app, "_wb_base_edit", None)
            ws_base_edit_ready = wb_base_edit[self.sheet] if wb_base_edit is not None and getattr(self.app, "has_base", False) else None
        except Exception:
            ws_base_edit_ready = None

        if rescan or (not self._bounds_checked):
            a_r, a_c = _effective_bounds_with_edit(ws_a, ws_a_edit)
            b_r, b_c = _effective_bounds_with_edit(ws_b, ws_b_edit)
            self.max_row = max(a_r, b_r)
            self.max_col = max(a_c, b_c)
            self.col_max_a = a_c
            self.col_max_b = b_c
            self._bounds_checked = True
            self._is_large_sheet = self.max_row >= _LARGE_SHEET_ROW_THRESHOLD
            if self._is_large_sheet:
                self._prefer_only_diff_when_ready = True

        # Full rescan diff map + cache row text if requested
        # Use _data_ready flag instead of checking pair_diff_cols emptiness:
        # pair_diff_cols can legitimately be empty (no diffs found) while still being valid data.
        if rescan or not self._data_ready:
            if not rescan:
                # Data not yet ready (background computation still running).
                # Skip this call; _apply_sheet_cache will call refresh() when done.
                return
            self.pair_diff_cols = {}
            self.pair_base_diff_cols = {}
            self.pair_text_a = {}
            self.pair_text_b = {}
            self.row_a_to_pair_idx = {}
            self.row_b_to_pair_idx = {}
            self.pair_base_row_override = {}
            self._diff_partial = False

            if conflict_cells_by_row is not None:
                # Conflict-only fast path: avoid full-sheet diff scan.
                self._align_rows_enabled = False
                conflict_rows = sorted(conflict_cells_by_row.keys())
                self.row_pairs = [(r, r) for r in conflict_rows]
                for idx, (ra, rb) in enumerate(self.row_pairs):
                    if ra is not None:
                        self.row_a_to_pair_idx[ra] = idx
                    if rb is not None:
                        self.row_b_to_pair_idx[rb] = idx
                # Pre-scan column widths before formatting
                ws_base_val_opt = None
                if getattr(self.app, "has_base", False):
                    try:
                        ws_base_val_opt = self.app.ws_base_val(self.sheet)
                    except Exception:
                        pass
                self._prescan_col_widths(ws_a, ws_b, ws_base_val_opt)
                for idx, (ra, rb) in enumerate(self.row_pairs):
                    line_a, line_b, _cols = self._build_row_and_diff_pair(ws_a, ws_b, ws_a_edit, ws_b_edit, ra, rb)
                    cols = set(conflict_cells_by_row.get(ra, set())) if ra is not None else set()
                    if not cols and _cols:  # preserve sentinel {-1} for one-sided rows
                        cols = _cols
                    self.pair_diff_cols[idx] = cols
                    self.pair_text_a[idx] = line_a
                    self.pair_text_b[idx] = line_b
            else:
                max_row_a = ws_a.max_row or 1
                max_row_b = ws_b.max_row or 1

                force_align = bool(getattr(self, "_force_sequence_align", False))
                should_align = (
                    (not getattr(self.app, "merge_conflict_mode", False))
                    and _should_auto_row_align(max_row_a, max_row_b, force=force_align)
                )

                self._align_rows_enabled = should_align
                if self._align_rows_enabled:
                    self.row_pairs = self._build_row_pairs(ws_a, ws_b, force=force_align)
                else:
                    self.row_pairs = self._build_row_pairs_direct(max_row_a, max_row_b)
                self.mine_to_base_row = {}
                self.theirs_to_base_row = {}
                if getattr(self.app, "has_base", False):
                    try:
                        ws_base_map = self.app.ws_base_val(self.sheet)
                    except Exception:
                        ws_base_map = None
                    if ws_base_map is not None:
                        try:
                            self.mine_to_base_row = _row_map_from_pairs(_compute_row_pairs_generic(ws_a, ws_base_map, self.max_col, force=force_align))
                        except Exception:
                            self.mine_to_base_row = {}
                        try:
                            self.theirs_to_base_row = _row_map_from_pairs(_compute_row_pairs_generic(ws_b, ws_base_map, self.max_col, force=force_align))
                        except Exception:
                            self.theirs_to_base_row = {}
                        self.row_pairs = _split_tail_independent_append_pairs(
                            self.row_pairs,
                            self.mine_to_base_row,
                            self.theirs_to_base_row,
                            ws_a,
                            ws_b,
                            self.max_col,
                        )
                        self.row_pairs = _split_low_similarity_tail_pairs(
                            self.row_pairs,
                            self.mine_to_base_row,
                            self.theirs_to_base_row,
                            ws_a,
                            ws_b,
                            self.max_col,
                        )
                        self.pair_base_row_override = _build_pair_base_row_overrides(
                            self.row_pairs,
                            self.mine_to_base_row,
                            self.theirs_to_base_row,
                            ws_base_map,
                            ws_a,
                            ws_b,
                            self.max_col,
                        )
                self.row_a_to_pair_idx = {}
                self.row_b_to_pair_idx = {}
                for idx, (ra, rb) in enumerate(self.row_pairs):
                    if ra is not None:
                        self.row_a_to_pair_idx[ra] = idx
                    if rb is not None:
                        self.row_b_to_pair_idx[rb] = idx

                # Pre-scan column widths for aligned display before building any formatted lines.
                ws_base_val_opt = None
                ws_base_edit_opt = None
                if getattr(self.app, "has_base", False):
                    try:
                        ws_base_val_opt = self.app.ws_base_val(self.sheet)
                    except Exception:
                        pass
                    ws_base_edit_opt = ws_base_edit_ready if ws_base_edit_ready is not None else ws_base_val_opt
                _prescan_limit = _FAST_RENDER_ROW_LIMIT if self._is_large_sheet else 0
                self._prescan_col_widths(ws_a, ws_b, ws_base_val_opt, max_pairs=_prescan_limit)

                # Large-sheet strategy:
                # - full mode: lazy row compute (first 200 visible rows only)
                # - only-diff mode: block scan from tail to head (1000 rows/block)
                if self._is_large_sheet and bool(self.only_diff_var.get()):
                    self._precompute_large_diff_by_blocks(
                        ws_a,
                        ws_b,
                        ws_a_edit,
                        ws_b_edit,
                        max_row_a,
                        max_row_b,
                        ws_base_val=ws_base_val_opt,
                        ws_base_edit=ws_base_edit_opt,
                    )
                elif not self._is_large_sheet:
                    for idx, (ra, rb) in enumerate(self.row_pairs):
                        line_a, line_b, cols = self._build_row_and_diff_pair(ws_a, ws_b, ws_a_edit, ws_b_edit, ra, rb)
                        self.pair_diff_cols[idx] = cols
                        self.pair_text_a[idx] = line_a
                        self.pair_text_b[idx] = line_b

            self._data_ready = True

        if self._is_three_way_enabled() and getattr(self.app, "has_base", False):
            need_full_base_diff = (not self._is_large_sheet) and (bool(self.only_diff_var.get()) or (not self._is_large_sheet))
            if need_full_base_diff:
                added_base_diff = self._ensure_base_diff_cache(
                    max_col=self.max_col,
                    ws_a_val=ws_a,
                    ws_a_edit=ws_a_edit if ws_a_edit is not None else ws_a,
                    ws_base_val=self.app.ws_base_val(self.sheet),
                    ws_base_edit=ws_base_edit_ready if ws_base_edit_ready is not None else self.app.ws_base_val(self.sheet),
                )
                if added_base_diff:
                    self._invalidate_render_cache()

        # Build display rows list (pair indices)
        if conflict_cells_by_row is not None:
            # Always show conflict rows only
            rows = []
            for r in sorted(conflict_cells_by_row.keys()):
                idx = self.row_a_to_pair_idx.get(r)
                if idx is not None:
                    rows.append(idx)
            self._full_display_rows = rows
        elif bool(self.only_diff_var.get()):
            cached_only_diff_rows = self._only_diff_rows_cache if self._has_valid_only_diff_snapshot_cache() else None
            if self.snapshot_only_diff and cached_only_diff_rows is not None and (not rescan):
                self._full_display_rows = self._only_diff_rows_with_touched(cached_only_diff_rows)
            elif self.snapshot_only_diff and self._is_large_sheet and (not rescan):
                started_async = self._start_async_large_only_diff_build()
                if started_async:
                    keep_rows = list(self.display_rows) if self.display_rows else list(self._full_display_rows)
                    self._full_display_rows = keep_rows
                else:
                    # A disk-backed async rebuild is intentionally unavailable
                    # after user edits. Reuse the current in-memory diff maps so
                    # resolved/touched rows remain stable and text is rebuilt
                    # from the edited in-memory workbooks below.
                    fallback_rows = cached_only_diff_rows
                    if fallback_rows is None:
                        fallback_rows = [
                            idx for idx in range(len(self.row_pairs))
                            if self._pair_has_visual_diff(idx)
                        ]
                        self._cache_only_diff_rows_snapshot(fallback_rows)
                    self._full_display_rows = self._only_diff_rows_with_touched(fallback_rows)
            elif (not self.snapshot_only_diff) or rescan or (cached_only_diff_rows is None):
                # Build snapshot: diff rows + touched rows.
                rows = [idx for idx in range(len(self.row_pairs)) if self._pair_has_visual_diff(idx)]

                # Recovery path:
                # In fast-open/background scenarios (especially large sheets), pair_diff_cols can be
                # incomplete when users switch to only-diff. If tab state already says this sheet has
                # diffs but rows is empty, do an on-demand fill so only-diff never becomes an empty page.
                if (not rows) and row_only is None:
                    try:
                        state = int(getattr(self.app, "sheet_diff_state", {}).get(self.sheet, 0))
                    except Exception:
                        state = 0
                    if state > 0:
                        try:
                            _dlog(
                                f"ONLY_DIFF_RECOVERY sheet={self.sheet} "
                                f"large={self._is_large_sheet} pairs={len(self.row_pairs)}"
                            )
                        except Exception:
                            pass
                        if self._is_large_sheet:
                            max_row_a = ws_a.max_row or 1
                            max_row_b = ws_b.max_row or 1
                            self._precompute_large_diff_by_blocks(
                                ws_a, ws_b, ws_a_edit, ws_b_edit, max_row_a, max_row_b
                            )
                        else:
                            for idx, (ra, rb) in enumerate(self.row_pairs):
                                if idx in self.pair_diff_cols:
                                    continue
                                line_a, line_b, cols = self._build_row_and_diff_pair(
                                    ws_a, ws_b, ws_a_edit, ws_b_edit, ra, rb
                                )
                                self.pair_diff_cols[idx] = cols
                                self.pair_text_a[idx] = line_a
                                self.pair_text_b[idx] = line_b
                        rows = [idx for idx in range(len(self.row_pairs)) if self._pair_has_visual_diff(idx)]
                        # If recovery found no diffs, the sheet_diff_state was stale.
                        # Clear it so subsequent refreshes do not re-trigger recovery.
                        if not rows:
                            try:
                                if hasattr(self.app, "sheet_diff_state"):
                                    self.app.sheet_diff_state[self.sheet] = 0
                            except Exception:
                                pass

                self._cache_only_diff_rows_snapshot(rows)
                self._full_display_rows = self._only_diff_rows_with_touched(rows)
            else:
                # snapshot mode: keep existing row list stable
                pass
        else:
            self._full_display_rows = list(range(0, len(self.row_pairs)))

        # Fast render: limit initial rows unless user opted to load all
        if self._full_render or (not _FAST_OPEN_ENABLED):
            self.display_rows = list(self._full_display_rows)
        else:
            # reset render limit if full list shrank
            self._render_limit = min(self._render_limit, len(self._full_display_rows)) if self._full_display_rows else _FAST_RENDER_ROW_LIMIT
            if self._full_display_rows:
                target_limit = min(
                    _LARGE_SHEET_INITIAL_ROWS if self._is_large_sheet else _FAST_RENDER_ROW_LIMIT,
                    len(self._full_display_rows),
                )
                if self._render_limit < target_limit:
                    self._render_limit = target_limit
            if (
                (not self._is_large_sheet)
                and len(self._full_display_rows) > _FAST_RENDER_ROW_LIMIT
                and self._render_limit < _FAST_RENDER_ROW_LIMIT
            ):
                self._render_limit = _FAST_RENDER_ROW_LIMIT
            if self._is_large_sheet and rescan:
                self._render_limit = min(_LARGE_SHEET_INITIAL_ROWS, len(self._full_display_rows)) if self._full_display_rows else _LARGE_SHEET_INITIAL_ROWS
            self.display_rows = self._full_display_rows[:self._render_limit]
        _dlog(f"  build display_rows: {len(self.display_rows)} / {self.max_row} (only_diff={bool(self.only_diff_var.get())} raw={self.only_diff_var.get()})")

        if self._is_three_way_enabled() and getattr(self.app, "has_base", False) and self.display_rows:
            added_base_diff = self._ensure_base_diff_cache(
                pair_indices=self.display_rows,
                max_col=self.max_col,
                ws_a_val=ws_a,
                ws_a_edit=ws_a_edit if ws_a_edit is not None else ws_a,
                ws_base_val=self.app.ws_base_val(self.sheet),
                ws_base_edit=ws_base_edit_ready if ws_base_edit_ready is not None else self.app.ws_base_val(self.sheet),
            )
            if added_base_diff:
                self._invalidate_render_cache()

        # Ensure pair text/diff exists for currently displayed rows (lazy fill)
        if self.display_rows:
            missing = [idx for idx in self.display_rows if idx not in self.pair_text_a or idx not in self.pair_text_b]
            if missing:
                for idx in missing:
                    if idx >= len(self.row_pairs):
                        continue
                    ra, rb = self.row_pairs[idx]
                    line_a, line_b, cols = self._build_row_and_diff_pair(ws_a, ws_b, ws_a_edit, ws_b_edit, ra, rb)
                    self.pair_diff_cols[idx] = cols
                    self.pair_text_a[idx] = line_a
                    self.pair_text_b[idx] = line_b

        self.row_to_line = {r: i + 1 for i, r in enumerate(self.display_rows)}

        if row_only is None:
            try:
                frac = float((self.left.xview() or (0.0, 1.0))[0])
            except Exception:
                frac = 0.0
            try:
                self._render_col_headers()
                # _render_col_headers() resets header xview to 0; restore all panes uniformly.
                self._sync_main_x_to_frac(frac)
                self._sync_c_x_to_frac(frac)
            except Exception:
                pass

        # Partial refresh: update a single excel row if it is visible
        if row_only is not None:
            r = row_only
            pair_idx = self.row_a_to_pair_idx.get(r)
            if pair_idx is None:
                pair_idx = self.row_b_to_pair_idx.get(r)
            if pair_idx is None:
                return
            ra, rb = self.row_pairs[pair_idx]
            # recompute diff cols + cache text for that pair only
            line_a, line_b, cols = self._build_row_and_diff_pair(ws_a, ws_b, ws_a_edit, ws_b_edit, ra, rb)
            self.pair_text_a[pair_idx] = line_a
            self.pair_text_b[pair_idx] = line_b
            if conflict_cells_by_row is not None and ra is not None:
                self.pair_diff_cols[pair_idx] = set(conflict_cells_by_row.get(ra, set()))
            else:
                self.pair_diff_cols[pair_idx] = cols
            if self._is_three_way_enabled() and getattr(self.app, "has_base", False):
                self.pair_base_diff_cols[pair_idx] = self._compute_base_diff_cols_for_pair(
                    pair_idx,
                    (ra, rb),
                    max_col=self.max_col,
                    ws_a_val=ws_a,
                    ws_a_edit=ws_a_edit,
                    ws_base_val=self.app.ws_base_val(self.sheet),
                    ws_base_edit=self.app.ws_base_edit(self.sheet),
                )
            else:
                self.pair_base_diff_cols.pop(pair_idx, None)

            # If only-diff enabled, row might need to be added/removed
            if bool(self.only_diff_var.get()):
                visible = pair_idx in self.row_to_line
                has = self._pair_has_visual_diff(pair_idx)

                # If diffs are resolved but this row was touched, keep it visible as a record.
                keep = (r in self.touched_rows)

                if self.snapshot_only_diff:
                    # Snapshot mode: never auto-remove rows from the list.
                    # If a touched row is not visible (was not in initial snapshot), allow adding it.
                    if (not visible) and keep:
                        self.refresh(row_only=None, rescan=False)
                        return
                else:
                    if visible and (not has) and (not keep):
                        # remove the line
                        line = self.row_to_line[pair_idx]
                        self.left.delete(f"{line}.0", f"{line + 1}.0")
                        if self._is_three_way_enabled():
                            self.base.delete(f"{line}.0", f"{line + 1}.0")
                        self.right.delete(f"{line}.0", f"{line + 1}.0")
                        # rebuild
                        self.refresh(row_only=None, rescan=False)
                        return

                    if (not visible) and (has or keep):
                        # add row: simplest is full rebuild (diff list is small)
                        self.refresh(row_only=None, rescan=False)
                        return

            line = self.row_to_line.get(pair_idx)
            if line is None:
                # not visible
                return

            line_a = self.pair_text_a.get(pair_idx, "")
            line_b = self.pair_text_b.get(pair_idx, "")

            # update text
            self.left.delete(f"{line}.0", f"{line}.end")
            self._render_base_line(line, pair_idx)
            self.right.delete(f"{line}.0", f"{line}.end")
            self.left.insert(f"{line}.0", line_a)
            self.right.insert(f"{line}.0", line_b)

            # clear tags on this line then apply diff highlight (unless touched row resolved)
            for w in (self.left, self.base, self.right):
                w.tag_remove("diffrow", f"{line}.0", f"{line}.end")
                w.tag_remove("diffcell", f"{line}.0", f"{line}.end")

            cols = self._visual_diff_cols_for_pair(pair_idx)
            # If this row was touched and has no diffs anymore, keep it visible but don't show diff highlight.
            show_diff = bool(cols)
            if show_diff:
                self.left.tag_add("diffrow", f"{line}.0", f"{line}.end")
                self.base.tag_add("diffrow", f"{line}.0", f"{line}.end")
                self.right.tag_add("diffrow", f"{line}.0", f"{line}.end")
                base_line = self._build_base_line(pair_idx) if self._is_three_way_enabled() else ""
                left_args, base_args, right_args, _lr, _br, _rr = self._diffcell_tag_args_for_line(
                    line,
                    pair_idx,
                    line_a,
                    base_line,
                    line_b,
                )
                if left_args:
                    self.left.tag_add("diffcell", *left_args)
                if base_args:
                    self.base.tag_add("diffcell", *base_args)
                if right_args:
                    self.right.tag_add("diffcell", *right_args)
                self._clear_diffrow_under_diffcells(left_args, base_args, right_args)
            # keep fast; do not rebuild sheet nav here
            try:
                self._display_diff_row_count = sum(1 for idx in self.display_rows if self._pair_has_visual_diff(idx))
                mode = "只看差异" if self.only_diff_var.get() else "全量"
                total_rows = len(self.row_pairs) if self.row_pairs else self.max_row
                self.info.configure(text=f"{mode} | RowsShown: {len(self.display_rows)} / {total_rows}   Cols: {self.max_col}   DiffRows: {self._display_diff_row_count}")
            except Exception:
                pass
            try:
                self._update_diff_maps()
            except Exception:
                pass
            return

        # Full render (use cache when possible)
        mode_key = "diff" if (conflict_cells_by_row is not None or bool(self.only_diff_var.get())) else "full"
        head = tuple(self.display_rows[:5])
        tail = tuple(self.display_rows[-5:]) if len(self.display_rows) > 5 else tuple(self.display_rows)
        cache_key = (mode_key, self._render_limit, len(self.display_rows), head, tail, self._data_version)
        if row_only is None and (not rescan):
            cached = self._render_cache.get(cache_key)
            if cached is not None:
                text_a, text_b, tag_rows, tag_cells, diff_row_count = cached
                self.left.delete("1.0", "end")
                self.base.delete("1.0", "end")
                self.right.delete("1.0", "end")
                self.left.insert("1.0", text_a)
                self._render_base_full()
                self.right.insert("1.0", text_b)
                self._render_row_headers_full()
                # clear tags
                self.left.tag_remove("diffrow", "1.0", "end")
                self.base.tag_remove("diffrow", "1.0", "end")
                self.right.tag_remove("diffrow", "1.0", "end")
                self.left.tag_remove("diffcell", "1.0", "end")
                self.base.tag_remove("diffcell", "1.0", "end")
                self.right.tag_remove("diffcell", "1.0", "end")
                self.left.tag_remove("paddingrow", "1.0", "end")
                self.base.tag_remove("paddingrow", "1.0", "end")
                self.right.tag_remove("paddingrow", "1.0", "end")
                # apply cached tags in bulk (one Tcl call per tag per widget)
                if tag_rows:
                    cached_diffrow_args = []
                    for line_idx in tag_rows:
                        cached_diffrow_args.extend([f"{line_idx}.0", f"{line_idx}.end"])
                    self.left.tag_add("diffrow", *cached_diffrow_args)
                    self.base.tag_add("diffrow", *cached_diffrow_args)
                    self.right.tag_add("diffrow", *cached_diffrow_args)
                    self.left_ln.tag_add("diffrow", *cached_diffrow_args)
                    self.base_ln.tag_add("diffrow", *cached_diffrow_args)
                    self.right_ln.tag_add("diffrow", *cached_diffrow_args)
                if tag_cells:
                    cached_cell_left = []
                    cached_cell_base = []
                    cached_cell_right = []
                    for line_idx, spans_a, spans_base, spans_b in tag_cells:
                        for s, e in spans_a:
                            cached_cell_left.extend([f"{line_idx}.{s}", f"{line_idx}.{e}"])
                        for s, e in spans_base:
                            cached_cell_base.extend([f"{line_idx}.{s}", f"{line_idx}.{e}"])
                        for s, e in spans_b:
                            cached_cell_right.extend([f"{line_idx}.{s}", f"{line_idx}.{e}"])
                    if cached_cell_left:
                        self.left.tag_add("diffcell", *cached_cell_left)
                    if cached_cell_base:
                        self.base.tag_add("diffcell", *cached_cell_base)
                    if cached_cell_right:
                        self.right.tag_add("diffcell", *cached_cell_right)
                    self._clear_diffrow_under_diffcells(cached_cell_left, cached_cell_base, cached_cell_right)

                # paddingrow: grey slot for one-sided pairs (computed from row_pairs, not cached)
                _padding_left = []
                _padding_right = []
                for _i, _pidx in enumerate(self.display_rows):
                    if _pidx < len(self.row_pairs):
                        _ra, _rb = self.row_pairs[_pidx]
                        _ln = _i + 1
                        if _ra is None:
                            _padding_left.extend([f"{_ln}.0", f"{_ln}.end"])
                        elif _rb is None:
                            _padding_right.extend([f"{_ln}.0", f"{_ln}.end"])
                if _padding_left:
                    self.left.tag_add("paddingrow", *_padding_left)
                if _padding_right:
                    self.right.tag_add("paddingrow", *_padding_right)
                # paddingcol: grey out column positions that don't exist on one side (新增列).
                _col_max_a = getattr(self, "col_max_a", self.max_col)
                _col_max_b = getattr(self, "col_max_b", self.max_col)
                if _col_max_a < self.max_col or _col_max_b < self.max_col:
                    _cspans = self._spans_for_line()
                    _n_lines = len(self.display_rows)
                    _pcol_left = []
                    _pcol_right = []
                    if _col_max_a < self.max_col and _col_max_a in _cspans:
                        _gs = _cspans[_col_max_a][1]
                        for _ln in range(1, _n_lines + 1):
                            _pcol_left.extend([f"{_ln}.{_gs}", f"{_ln}.end"])
                    if _col_max_b < self.max_col and _col_max_b in _cspans:
                        _gs = _cspans[_col_max_b][1]
                        for _ln in range(1, _n_lines + 1):
                            _pcol_right.extend([f"{_ln}.{_gs}", f"{_ln}.end"])
                    if _pcol_left:
                        self.left.tag_add("paddingcol", *_pcol_left)
                    if _pcol_right:
                        self.right.tag_add("paddingcol", *_pcol_right)
                # rownum gutter tags are unused when row headers are separate

                mode = "只看差异" if self.only_diff_var.get() else "全量"
                total_rows = len(self.row_pairs) if self.row_pairs else self.max_row
                self.info.configure(text=f"{mode} | RowsShown: {len(self.display_rows)} / {total_rows}   Cols: {self.max_col}   DiffRows: {diff_row_count}")
                self._display_diff_row_count = diff_row_count
                self.app.set_sheet_has_diff(self.sheet, diff_row_count > 0, confirmed=True)
                self.app.refresh_sheet_nav()
                self._update_diff_nav_state()
                try:
                    self._update_diff_maps()
                except Exception:
                    pass
                return

        # Full render
        self.left.delete("1.0", "end")
        self.base.delete("1.0", "end")
        self.right.delete("1.0", "end")
        self.left.tag_remove("diffrow", "1.0", "end")
        self.base.tag_remove("diffrow", "1.0", "end")
        self.right.tag_remove("diffrow", "1.0", "end")
        self.left.tag_remove("diffcell", "1.0", "end")
        self.base.tag_remove("diffcell", "1.0", "end")
        self.right.tag_remove("diffcell", "1.0", "end")
        self.left.tag_remove("paddingrow", "1.0", "end")
        self.base.tag_remove("paddingrow", "1.0", "end")
        self.right.tag_remove("paddingrow", "1.0", "end")

        # Build full text in memory and insert once (faster)
        lines_a = []
        lines_b = []
        for pair_idx in self.display_rows:
            lines_a.append(self.pair_text_a.get(pair_idx, ""))
            lines_b.append(self.pair_text_b.get(pair_idx, ""))
        self.left.insert("1.0", "\n".join(lines_a) + ("\n" if lines_a else ""))
        lines_base = self._render_base_full()
        self.right.insert("1.0", "\n".join(lines_b) + ("\n" if lines_b else ""))
        self._render_row_headers_full()

        # On some environments/large documents, forcing an idle layout pass improves tag correctness.
        try:
            self.left.update_idletasks()
            self.base.update_idletasks()
            self.right.update_idletasks()
        except Exception:
            pass

        # Restore scroll position if we just appended more rows
        if self._pending_yview is not None:
            try:
                self.left.yview_moveto(self._pending_yview)
                if self._is_three_way_enabled():
                    self.base.yview_moveto(self._pending_yview)
                self.right.yview_moveto(self._pending_yview)
            except Exception:
                pass
            self._pending_yview = None

        if self._is_three_way_enabled() and getattr(self.app, "has_base", False):
            added_base_diff = self._ensure_base_diff_cache(
                pair_indices=self.display_rows,
                max_col=self.max_col,
                ws_a_val=ws_a,
                ws_a_edit=ws_a_edit if ws_a_edit is not None else ws_a,
                ws_base_val=self.app.ws_base_val(self.sheet),
                ws_base_edit=ws_base_edit_ready if ws_base_edit_ready is not None else self.app.ws_base_val(self.sheet),
            )
            if added_base_diff:
                self._invalidate_render_cache()

        diff_row_count = 0
        tag_rows = []
        tag_cells = []
        # Collect all tag ranges first; apply in bulk (one Tcl call per tag instead of N).
        # tag_add(tagName, index1, *args) accepts multiple index pairs in a single call.
        diffrow_args = []
        diffcell_args_left = []
        diffcell_args_base = []
        diffcell_args_right = []
        for line_idx, pair_idx in enumerate(self.display_rows, start=1):
            ra, rb = self.row_pairs[pair_idx] if pair_idx < len(self.row_pairs) else (None, None)
            cols = self._visual_diff_cols_for_pair(pair_idx)
            if cols:
                diff_row_count += 1
                diffrow_args.extend([f"{line_idx}.0", f"{line_idx}.end"])
                tag_rows.append(line_idx)

                line_a = lines_a[line_idx - 1] if (line_idx - 1) < len(lines_a) else ""
                line_b = lines_b[line_idx - 1] if (line_idx - 1) < len(lines_b) else ""
                base_line = (
                    lines_base[line_idx - 1]
                    if self._is_three_way_enabled() and (line_idx - 1) < len(lines_base)
                    else ""
                )
                left_args, base_args, right_args, left_ranges, base_ranges, right_ranges = self._diffcell_tag_args_for_line(
                    line_idx,
                    pair_idx,
                    line_a,
                    base_line,
                    line_b,
                )
                diffcell_args_left.extend(left_args)
                diffcell_args_base.extend(base_args)
                diffcell_args_right.extend(right_args)
                if left_ranges or base_ranges or right_ranges:
                    tag_cells.append((line_idx, left_ranges, base_ranges, right_ranges))

        # Apply all diffrow tags in one call per widget
        if diffrow_args:
            self.left.tag_add("diffrow", *diffrow_args)
            self.base.tag_add("diffrow", *diffrow_args)
            self.right.tag_add("diffrow", *diffrow_args)
            self.left_ln.tag_add("diffrow", *diffrow_args)
            self.base_ln.tag_add("diffrow", *diffrow_args)
            self.right_ln.tag_add("diffrow", *diffrow_args)
            for line_idx in tag_rows:
                try:
                    pidx = self.display_rows[line_idx - 1]
                    self._apply_rownum_diff_tag_line(line_idx, pidx)
                except Exception:
                    pass
        # Apply all diffcell tags in one call per widget
        if diffcell_args_left:
            self.left.tag_add("diffcell", *diffcell_args_left)
        if diffcell_args_base:
            self.base.tag_add("diffcell", *diffcell_args_base)
        if diffcell_args_right:
            self.right.tag_add("diffcell", *diffcell_args_right)
        self._clear_diffrow_under_diffcells(diffcell_args_left, diffcell_args_base, diffcell_args_right)
        # Apply paddingrow (grey) to empty slots of one-sided pairs
        _padding_left = []
        _padding_right = []
        for _i, _pidx in enumerate(self.display_rows):
            if _pidx < len(self.row_pairs):
                _ra, _rb = self.row_pairs[_pidx]
                _ln = _i + 1
                if _ra is None:
                    _padding_left.extend([f"{_ln}.0", f"{_ln}.end"])
                elif _rb is None:
                    _padding_right.extend([f"{_ln}.0", f"{_ln}.end"])
        if _padding_left:
            self.left.tag_add("paddingrow", *_padding_left)
        if _padding_right:
            self.right.tag_add("paddingrow", *_padding_right)
        # paddingcol: grey out column positions that don't exist on one side (新增列).
        _col_max_a = getattr(self, "col_max_a", self.max_col)
        _col_max_b = getattr(self, "col_max_b", self.max_col)
        if _col_max_a < self.max_col or _col_max_b < self.max_col:
            _cspans = self._spans_for_line()
            _n_lines = len(self.display_rows)
            _pcol_left = []
            _pcol_right = []
            if _col_max_a < self.max_col and _col_max_a in _cspans:
                _gs = _cspans[_col_max_a][1]
                for _ln in range(1, _n_lines + 1):
                    _pcol_left.extend([f"{_ln}.{_gs}", f"{_ln}.end"])
            if _col_max_b < self.max_col and _col_max_b in _cspans:
                _gs = _cspans[_col_max_b][1]
                for _ln in range(1, _n_lines + 1):
                    _pcol_right.extend([f"{_ln}.{_gs}", f"{_ln}.end"])
            if _pcol_left:
                self.left.tag_add("paddingcol", *_pcol_left)
            if _pcol_right:
                self.right.tag_add("paddingcol", *_pcol_right)
        # row-number styling handled by dedicated row-header widgets

        mode = "只看差异" if self.only_diff_var.get() else "全量"
        total_rows = len(self.row_pairs) if self.row_pairs else self.max_row
        self.info.configure(text=f"{mode} | RowsShown: {len(self.display_rows)} / {total_rows}   Cols: {self.max_col}   DiffRows: {diff_row_count}")
        self._display_diff_row_count = diff_row_count

        self.app.set_sheet_has_diff(self.sheet, diff_row_count > 0, confirmed=True)
        self.app.refresh_sheet_nav()
        self._update_diff_nav_state()
        try:
            self._update_diff_maps()
        except Exception:
            pass

        # Cache rendered result for fast toggle
        if row_only is None:
            text_a = "\n".join(lines_a) + ("\n" if lines_a else "")
            text_b = "\n".join(lines_b) + ("\n" if lines_b else "")
            self._render_cache[cache_key] = (text_a, text_b, tag_rows, tag_cells, diff_row_count)


class SowMergeApp:
    def __init__(self, file_a: str, file_b: str, merge_mode: bool = False, merged_path: str | None = None,
                 base_path: str | None = None,
                 merge_conflict_cells_by_sheet: dict | None = None, merge_conflict_mode: bool = False,
                 raw_base: str | None = None, raw_mine: str | None = None, raw_theirs: str | None = None):
        # Sequential GUI instances are common in smoke tests and can also occur
        # in host integrations. Collect destroyed Tk object cycles on the UI
        # thread before any new background XML parser can trigger that GC.
        gc.collect()
        self._is_closing = False
        self._background_threads_lock = threading.Lock()
        self._background_threads: set[threading.Thread] = set()
        self._interactive_action_lock = threading.Lock()
        self._interactive_action_depth = 0
        self._interactive_action_event = threading.Event()
        self._priority_diff_lock = threading.Lock()
        self._priority_diff_depth = 0
        self._priority_diff_event = threading.Event()
        self.file_a = file_a
        self.file_b = file_b
        self.base_path = base_path
        self.has_base = bool(base_path and os.path.exists(base_path))
        self.raw_base = raw_base
        self.raw_mine = raw_mine
        self.raw_theirs = raw_theirs
        self.merge_mode = merge_mode
        self.diff_base_mine_mode = bool((not merge_mode) and raw_base and raw_mine)
        self.merged_path = merged_path
        self.merge_conflict_cells_by_sheet = merge_conflict_cells_by_sheet or {}
        self.merge_conflict_mode = merge_conflict_mode
        self.initial_conflict_cell_count = sum(
            len(cols)
            for rows in self.merge_conflict_cells_by_sheet.values()
            for cols in rows.values()
        )
        self.user_touched_conflicts = False
        # In 3-way manual merge mode, persist only explicitly operated A-side cells on save.
        # key: (sheet, row, col) -> edit value to write
        self.manual_a_cell_ops: dict[tuple[str, int, int], object] = {}
        self.manual_b_cell_ops: dict[tuple[str, int, int], object] = {}
        # Same-formula cells can have intentionally adopted cached results in
        # manual-calculation workbooks. Keep them separate from formula edits.
        self.manual_a_formula_cache_ops: dict[tuple[str, int, int], object] = {}
        self.manual_b_formula_cache_ops: dict[tuple[str, int, int], object] = {}
        self.manual_a_row_ops: list[dict[str, object]] = []
        self.manual_b_row_ops: list[dict[str, object]] = []
        self.manual_sheet_ops: list[dict[str, object]] = []
        self.auto_sheet_ops: list[dict[str, object]] = []
        self.sheet_level_conflicts: list[dict[str, object]] = []
        self.sheet_level_summary: list[str] = []
        self._merge_mine_snapshot = None
        self.undo_stack = []
        self._auto_recalc_started = False
        # append debug session marker each run (do not truncate old evidence)
        try:
            with open(_DEBUG_LOG_PATH, "a", encoding="utf-8") as f:
                f.write("\n" + "=" * 72 + "\n")
                f.write(f"SESSION {datetime.now().isoformat(sep=' ', timespec='seconds')}\n")
                f.write(f"{APP_NAME} {APP_VERSION} [{APP_BUILD_TAG}]\n")
                f.write(f"A={self.file_a}\nB={self.file_b}\n")
        except Exception:
            pass
        # load settings
        self.settings = {}
        self.only_diff_default = 0
        try:
            os.makedirs(os.path.dirname(_SETTINGS_PATH), exist_ok=True)
            if os.path.exists(_SETTINGS_PATH):
                with open(_SETTINGS_PATH, "r", encoding="utf-8") as f:
                    self.settings = json.load(f) or {}
            self.only_diff_default = int(self.settings.get("only_diff", 0))
        except Exception as e:
            _dlog(f"settings load failed: {e}")

        _dlog(f"SowMergeApp init only_diff_default={self.only_diff_default}")

        # Fast open: load value workbooks first; defer editable workbooks until first modification/save.
        self._wb_a_edit = None
        self._wb_b_edit = None
        self._wb_base_edit = None
        self._edit_loaded_event = threading.Event()
        self._initial_sheet_ready_event = threading.Event()
        self._edit_loading_started = False
        self._edit_fallback_lock = threading.Lock()
        self._wb_a_val = None
        self._wb_b_val = None
        self._wb_base_val = None

        # Preload editable workbooks in background to make the first overwrite fast.
        # Always run regardless of _FAST_OPEN_ENABLED: fast-open defers value loading
        # but edit workbooks must still be ready before the user's first row override.
        def _preload_edit():
            a_edit = None
            b_edit = None
            base_edit = None
            try:
                _dlog("preload edit workbooks waiting for initial Sheet")
                while not self._initial_sheet_ready_event.wait(timeout=0.05):
                    if self._is_closing:
                        return
                _dlog("preload edit workbooks (background) start")
                t1 = datetime.now()
                a_edit = load_workbook(self.file_a, data_only=False)
                _dlog(f"preload wb_a_edit: {(datetime.now()-t1).total_seconds():.3f}s")
                t2 = datetime.now()
                b_edit = load_workbook(self.file_b, data_only=False)
                _dlog(f"preload wb_b_edit: {(datetime.now()-t2).total_seconds():.3f}s")
                base_edit = None
                if self.has_base:
                    t3 = datetime.now()
                    base_edit = load_workbook(self.base_path, data_only=False)
                    _dlog(f"preload wb_base_edit: {(datetime.now()-t3).total_seconds():.3f}s")
                with self._edit_fallback_lock:
                    if self._is_closing:
                        _wbs_close(a_edit, b_edit, base_edit)
                    else:
                        self._wb_a_edit = a_edit
                        self._wb_b_edit = b_edit
                        self._wb_base_edit = base_edit
            except Exception as e:
                _wbs_close(a_edit, b_edit, base_edit)
                _dlog(f"preload edit failed: {e}")
            finally:
                self._edit_loaded_event.set()
                _dlog("preload edit workbooks (background) done")

        def _load_initial_state(report):
            try:
                report("正在打开 Excel 合并工具", f"加载 mine：{os.path.basename(file_a)}", 8)
                t0 = datetime.now()
                self._file_a_val_path = _prepare_val_path(file_a)
                self._wb_a_val = load_workbook(self._file_a_val_path, data_only=True)
                _dlog(f"load wb_a_val: {(datetime.now()-t0).total_seconds():.3f}s")

                report("正在打开 Excel 合并工具", f"加载 theirs：{os.path.basename(file_b)}", 30)
                t0 = datetime.now()
                self._file_b_val_path = _prepare_val_path(file_b)
                self._wb_b_val = load_workbook(self._file_b_val_path, data_only=True)
                _dlog(f"load wb_b_val: {(datetime.now()-t0).total_seconds():.3f}s")

                if self.has_base:
                    report("正在打开 Excel 合并工具", f"加载 base：{os.path.basename(self.base_path)}", 52)
                    t0 = datetime.now()
                    self._file_base_val_path = _prepare_val_path(self.base_path)
                    self._wb_base_val = load_workbook(self._file_base_val_path, data_only=True)
                    _dlog(f"load wb_base_val: {(datetime.now()-t0).total_seconds():.3f}s")

                    report("正在准备三方合并", "创建 mine 安全快照...", 68)
                    try:
                        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                        snap = os.path.join(
                            tempfile.gettempdir(),
                            f"{APP_NAME}_mine_snapshot_{os.getpid()}_{ts}{_workbook_ext(self.file_a)}",
                        )
                        shutil.copy2(self.file_a, snap)
                        self._merge_mine_snapshot = snap
                        _dlog(f"mine snapshot created: {snap}")
                    except Exception as e:
                        _dlog(f"mine snapshot failed: {e}")

                report("正在分析工作簿结构", "读取 Sheet 列表并判断整表新增、删除...", 76)
                self._refresh_sheet_catalog()
                if self._wb_a_edit is None or self._wb_b_edit is None or (self.has_base and self._wb_base_edit is None):
                    self._edit_loading_started = True
                    self._edit_preload_thread = self._start_background_thread(
                        _preload_edit,
                        name="sow-edit-preload",
                    )
                else:
                    self._edit_loaded_event.set()
                report(
                    "工作簿值数据加载完成",
                    f"共 {len(self.display_sheets)} 个 Sheet；公式、样式和批注将在后台继续预载...",
                    100,
                )
            except Exception:
                _wbs_close(self._wb_a_val, self._wb_b_val, self._wb_base_val)
                raise

        _run_startup_progress_task(
            "Excel 合并工具 - 正在打开",
            "正在读取工作簿，请稍候...",
            _load_initial_state,
        )

        self.modified_a = False
        self.modified_b = False
        self.modified_sheets_a = set()
        self.modified_sheets_b = set()

        self.root = _take_startup_progress_root()
        self._window_title_suffix = f"{APP_NAME} {APP_VERSION} [{APP_BUILD_TAG}]"
        self.root.title(self._window_title_suffix)
        self.root.resizable(True, True)
        ttk.Style().theme_use("clam")
        if self.merge_mode:
            self.root.title(f"{self._window_title_suffix} (SVN Merge)")
        else:
            self.root.title(f"{self._window_title_suffix} (TortoiseMerge-like)")
        self.root.geometry("1450x860")
        try:
            self.root.state("zoomed")
        except Exception:
            pass

        self._root_after_ids: set[str] = set()
        try:
            self.root.protocol("WM_DELETE_WINDOW", self._shutdown_root)
        except Exception:
            pass

        self._build_ui()
        self._schedule_auto_recalc()

    def _sheet_names_for_side(self, side: str) -> list[str]:
        side = str(side or "").upper()
        if side == "A":
            wb = getattr(self, "_wb_a_val", None)
        elif side in ("B", "THEIRS"):
            wb = getattr(self, "_wb_b_val", None)
        elif side == "BASE":
            wb = getattr(self, "_wb_base_val", None)
        else:
            wb = None
        try:
            return list(getattr(wb, "sheetnames", []) or [])
        except Exception:
            return []

    def _refresh_sheet_catalog(self):
        names_a = self._sheet_names_for_side("A")
        names_b = self._sheet_names_for_side("B")
        names_base = self._sheet_names_for_side("BASE")
        set_a = set(names_a)
        set_b = set(names_b)
        set_base = set(names_base)

        ordered = []
        for seq in (names_a, names_b, names_base):
            for name in seq:
                if name not in ordered:
                    ordered.append(name)

        if self.merge_conflict_mode and self.merge_conflict_cells_by_sheet:
            conflict_sheets = set(self.merge_conflict_cells_by_sheet.keys())
            ordered = [s for s in ordered if s in conflict_sheets]

        self.common_sheets = sorted(set_a & set_b)
        if self.merge_conflict_mode and self.merge_conflict_cells_by_sheet:
            self.common_sheets = [s for s in self.common_sheets if s in set(self.merge_conflict_cells_by_sheet.keys())]
        self.only_a = sorted(set_a - set_b)
        self.only_b = sorted(set_b - set_a)
        self.display_sheets = ordered
        self.compare_sheets = [s for s in ordered if (s in set_a and s in set_b)]
        self.sheet_meta: dict[str, dict[str, object]] = {}
        for s in ordered:
            has_a = s in set_a
            has_b = s in set_b
            has_base = s in set_base
            self.sheet_meta[s] = {
                "has_a": has_a,
                "has_b": has_b,
                "has_base": has_base,
                "view_mode": "normal" if (has_a and has_b) else "missing_sheet",
            }
        self.sheet_diff_state = {s: int(getattr(self, "sheet_diff_state", {}).get(s, 0)) for s in ordered}
        previous_confirmed = set(getattr(self, "_sheet_diff_confirmed", set()) or set())
        self._sheet_diff_confirmed = {s for s in previous_confirmed if s in self.sheet_diff_state}
        self._recompute_auto_sheet_ops()

    def _recompute_auto_sheet_ops(self):
        self.auto_sheet_ops = []
        self.sheet_level_conflicts = []
        self.sheet_level_summary = []
        if not (self.merge_mode and self.has_base):
            return
        set_a = set(self._sheet_names_for_side("A"))
        set_b = set(self._sheet_names_for_side("B"))
        set_base = set(self._sheet_names_for_side("BASE"))
        ordered = list(getattr(self, "display_sheets", []) or [])
        for sheet in ordered:
            in_a = sheet in set_a
            in_b = sheet in set_b
            in_base = sheet in set_base
            if in_b and (not in_a) and (not in_base):
                self.auto_sheet_ops.append({"kind": "copy_sheet", "sheet": sheet, "source_side": "B", "target_side": "A"})
                self.sheet_level_summary.append(f"自动并入 theirs 新增 Sheet: {sheet}")
            elif in_base and (not in_a) and (not in_b):
                self.sheet_level_summary.append(f"保留双方均已删除的 Sheet: {sheet}")
            elif in_base and in_a and (not in_b):
                if self._sheet_values_equal("A", "BASE", sheet):
                    self.auto_sheet_ops.append({"kind": "delete_sheet", "sheet": sheet, "target_side": "A"})
                    self.sheet_level_summary.append(f"自动删除 theirs 已删除且 mine 未修改的 Sheet: {sheet}")
                else:
                    self.sheet_level_conflicts.append({"kind": "sheet_deleted_or_missing_in_theirs", "sheet": sheet})
            elif in_base and in_b and (not in_a):
                if self._sheet_values_equal("B", "BASE", sheet):
                    # Mine deleted the sheet while theirs kept base unchanged:
                    # preserve the local deletion instead of silently restoring it.
                    self.sheet_level_summary.append(f"保留 mine 已删除、theirs 未修改的 Sheet: {sheet}")
                else:
                    # Mine deleted while theirs modified: neither side can be
                    # chosen safely. Keep mine by default and surface a conflict.
                    self.sheet_level_conflicts.append({
                        "kind": "sheet_deleted_in_mine_modified_in_theirs",
                        "sheet": sheet,
                    })
            elif in_a and (not in_base) and (not in_b):
                self.sheet_level_summary.append(f"保留 mine 独有 Sheet: {sheet}")

    def get_sheet_meta(self, sheet: str) -> dict[str, object]:
        return dict(getattr(self, "sheet_meta", {}).get(sheet, {}))

    def _sheet_values_equal(self, side_a: str, side_b: str, sheet: str) -> bool:
        ws_a = self.ws_for_side(side_a, sheet, edit=False, allow_missing=True)
        ws_b = self.ws_for_side(side_b, sheet, edit=False, allow_missing=True)
        if ws_a is None or ws_b is None:
            return False
        ws_a_edit = self.ws_for_side(side_a, sheet, edit=True, allow_missing=True)
        ws_b_edit = self.ws_for_side(side_b, sheet, edit=True, allow_missing=True)
        if ws_a_edit is None or ws_b_edit is None:
            return False
        max_row = max(ws_a.max_row or 1, ws_b.max_row or 1)
        max_col = max(
            ws_a.max_column or 1,
            ws_b.max_column or 1,
            ws_a_edit.max_column or 1,
            ws_b_edit.max_column or 1,
        )
        try:
            if {str(rng) for rng in ws_a_edit.merged_cells.ranges} != {
                str(rng) for rng in ws_b_edit.merged_cells.ranges
            }:
                return False
            if str(ws_a_edit.freeze_panes or "") != str(ws_b_edit.freeze_panes or ""):
                return False
            if str(getattr(ws_a_edit.auto_filter, "ref", "") or "") != str(
                getattr(ws_b_edit.auto_filter, "ref", "") or ""
            ):
                return False

            def _row_dim_key(dim):
                return (
                    getattr(dim, "height", None), getattr(dim, "hidden", None),
                    getattr(dim, "outlineLevel", None), getattr(dim, "collapsed", None),
                    getattr(dim, "style_id", None), getattr(dim, "thickTop", None),
                    getattr(dim, "thickBot", None),
                )

            row_keys = set(ws_a_edit.row_dimensions) | set(ws_b_edit.row_dimensions)
            for key in row_keys:
                if _row_dim_key(ws_a_edit.row_dimensions[key]) != _row_dim_key(ws_b_edit.row_dimensions[key]):
                    return False

            def _col_dim_key(dim):
                return (
                    getattr(dim, "width", None), getattr(dim, "hidden", None),
                    getattr(dim, "bestFit", None), getattr(dim, "outlineLevel", None),
                    getattr(dim, "collapsed", None), getattr(dim, "style_id", None),
                    getattr(dim, "min", None), getattr(dim, "max", None),
                )

            col_keys = set(ws_a_edit.column_dimensions) | set(ws_b_edit.column_dimensions)
            for key in col_keys:
                if _col_dim_key(ws_a_edit.column_dimensions[key]) != _col_dim_key(ws_b_edit.column_dimensions[key]):
                    return False
        except Exception:
            # This predicate controls destructive auto-delete. If structure
            # cannot be compared, fail closed and surface a sheet conflict.
            return False

        rows_a = ws_a.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True)
        rows_b = ws_b.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True)
        rows_a_edit = ws_a_edit.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=False)
        rows_b_edit = ws_b_edit.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=False)
        for row_a, row_b, row_a_edit, row_b_edit in zip(rows_a, rows_b, rows_a_edit, rows_b_edit):
            for va, vb, cell_a_edit, cell_b_edit in zip(row_a, row_b, row_a_edit, row_b_edit):
                va_edit = cell_a_edit.value
                vb_edit = cell_b_edit.value
                try:
                    if cell_a_edit.data_type != cell_b_edit.data_type:
                        return False
                    if cell_a_edit._style != cell_b_edit._style:
                        return False
                    comment_a = cell_a_edit.comment
                    comment_b = cell_b_edit.comment
                    if (
                        getattr(comment_a, "text", None), getattr(comment_a, "author", None)
                    ) != (
                        getattr(comment_b, "text", None), getattr(comment_b, "author", None)
                    ):
                        return False
                    link_a = cell_a_edit.hyperlink
                    link_b = cell_b_edit.hyperlink
                    if (
                        getattr(link_a, "target", None), getattr(link_a, "location", None),
                        getattr(link_a, "tooltip", None), getattr(link_a, "display", None),
                    ) != (
                        getattr(link_b, "target", None), getattr(link_b, "location", None),
                        getattr(link_b, "tooltip", None), getattr(link_b, "display", None),
                    ):
                        return False
                except Exception:
                    return False
                formula_a = _formula_text(va_edit)
                formula_b = _formula_text(vb_edit)
                if formula_a or formula_b:
                    if not (formula_a and formula_b and _same_formula(va_edit, vb_edit)):
                        return False
                _display_a, _display_b, equal = _cell_display_and_equal_from_values(
                    va, vb, va_edit, vb_edit
                )
                if not equal:
                    return False
        return True

    def ws_for_side(self, side: str, sheet: str, edit: bool = False, allow_missing: bool = False):
        side = str(side or "").upper()
        if side == "A":
            getter = self.ws_a_edit if edit else self.ws_a_val
        elif side in ("B", "THEIRS"):
            getter = self.ws_b_edit if edit else self.ws_b_val
        elif side == "BASE":
            getter = self.ws_base_edit if edit else self.ws_base_val
        else:
            getter = None
        if getter is None:
            return None
        try:
            return getter(sheet)
        except Exception:
            if allow_missing:
                return None
            raise

    def _start_background_thread(self, target, *, name: str):
        """Start and track a worker so shutdown cannot orphan workbook/Tk state."""
        if getattr(self, "_is_closing", False):
            return None

        def _run_tracked():
            try:
                target()
            finally:
                try:
                    with self._background_threads_lock:
                        self._background_threads.discard(threading.current_thread())
                except Exception:
                    pass

        thread = threading.Thread(target=_run_tracked, daemon=True, name=name)
        with self._background_threads_lock:
            if self._is_closing:
                return None
            self._background_threads.add(thread)
        try:
            thread.start()
        except Exception:
            with self._background_threads_lock:
                self._background_threads.discard(thread)
            raise
        return thread

    def _begin_interactive_action(self):
        """Pause optional background scans while the user is changing data."""
        with self._interactive_action_lock:
            self._interactive_action_depth += 1
            self._interactive_action_event.set()

    def _end_interactive_action(self):
        with self._interactive_action_lock:
            self._interactive_action_depth = max(0, self._interactive_action_depth - 1)
            if self._interactive_action_depth == 0:
                self._interactive_action_event.clear()

    def _begin_priority_diff_compute(self):
        """Pause lower-priority tab scans while the active Sheet builds exact diffs."""
        with self._priority_diff_lock:
            self._priority_diff_depth += 1
            self._priority_diff_event.set()

    def _end_priority_diff_compute(self):
        with self._priority_diff_lock:
            self._priority_diff_depth = max(0, self._priority_diff_depth - 1)
            if self._priority_diff_depth == 0:
                self._priority_diff_event.clear()

    def _join_background_threads(self, timeout: float = 5.0):
        deadline = time.monotonic() + max(0.0, float(timeout))
        with self._background_threads_lock:
            threads = list(self._background_threads)
        current = threading.current_thread()
        for thread in threads:
            if thread is current or not thread.is_alive():
                continue
            remaining = deadline - time.monotonic()
            if remaining <= 0:
                break
            thread.join(timeout=remaining)

    def _safe_root_after(self, delay_ms: int, callback):
        if getattr(self, "_is_closing", False):
            return None
        holder = {"id": None}

        def _wrapped():
            aid = holder.get("id")
            if aid is not None:
                try:
                    self._root_after_ids.discard(aid)
                except Exception:
                    pass
            if getattr(self, "_is_closing", False):
                return
            callback()

        try:
            aid = self.root.after(int(delay_ms), _wrapped)
            holder["id"] = aid
            self._root_after_ids.add(aid)
            return aid
        except Exception:
            return None

    def _cancel_root_afters(self):
        try:
            pending = list(getattr(self, "_root_after_ids", set()) or ())
        except Exception:
            pending = []
        try:
            self._root_after_ids.clear()
        except Exception:
            pass
        for aid in pending:
            try:
                self.root.after_cancel(aid)
            except Exception:
                pass

    def _shutdown_root(self):
        if getattr(self, "_is_closing", False):
            return
        self._is_closing = True
        try:
            self._initial_sheet_ready_event.set()
        except Exception:
            pass
        self._cancel_root_afters()
        try:
            progress = getattr(self, "task_progress", None)
            if progress is not None:
                progress.stop()
        except Exception:
            pass
        try:
            with self._compute_lock:
                self._compute_queue.clear()
                self._sheet_cache_store.clear()
        except Exception:
            pass
        try:
            for view in list(getattr(self, "sheet_views", {}).values()):
                if view is None:
                    continue
                view._only_diff_async_build_seq += 1
                view._only_diff_async_building = False
                view._only_diff_async_build_key = None
                try:
                    view._hide_loading()
                except Exception:
                    pass
                for attr in ("_settings_save_id", "_hover_debounce_id", "_diff_map_debounce_id"):
                    aid = getattr(view, attr, None)
                    if not aid:
                        continue
                    try:
                        view.frame.after_cancel(aid)
                    except Exception:
                        pass
                    try:
                        setattr(view, attr, None)
                    except Exception:
                        pass
        except Exception:
            pass
        # Widgets and ttk progressbars can schedule Tcl-level callbacks that are
        # not registered through _safe_root_after. Cancel every remaining callback
        # before destroying the interpreter to avoid stale-command warnings.
        try:
            for aid in list(self.root.tk.call("after", "info") or ()):
                try:
                    self.root.after_cancel(aid)
                except Exception:
                    pass
        except Exception:
            pass
        try:
            with self._ui_task_lock:
                self._ui_tasks.clear()
        except Exception:
            pass
        self._join_background_threads(timeout=5.0)
        try:
            _destroy_startup_progress_root()
        except Exception:
            pass
        try:
            with self._edit_fallback_lock:
                _wbs_close(
                    getattr(self, "_wb_a_val", None),
                    getattr(self, "_wb_b_val", None),
                    getattr(self, "_wb_base_val", None),
                    getattr(self, "_wb_a_edit", None),
                    getattr(self, "_wb_b_edit", None),
                    getattr(self, "_wb_base_edit", None),
                )
                self._wb_a_val = None
                self._wb_b_val = None
                self._wb_base_val = None
                self._wb_a_edit = None
                self._wb_b_edit = None
                self._wb_base_edit = None
        except Exception:
            pass
        try:
            temp_specs = [
                (getattr(self, "_file_a_val_path", None), getattr(self, "file_a", None)),
                (getattr(self, "_file_b_val_path", None), getattr(self, "file_b", None)),
                (getattr(self, "_file_base_val_path", None), getattr(self, "base_path", None)),
            ]
            for temp_path, original_path in temp_specs:
                if not temp_path:
                    continue
                try:
                    if original_path and os.path.abspath(temp_path) == os.path.abspath(original_path):
                        continue
                except Exception:
                    pass
                try:
                    if os.path.exists(temp_path):
                        os.remove(temp_path)
                except Exception:
                    pass
        except Exception:
            pass
        try:
            snap = getattr(self, "_merge_mine_snapshot", None)
            if snap and os.path.exists(snap):
                os.remove(snap)
        except Exception:
            pass
        try:
            if self.root.winfo_exists():
                # Ensure Tk mainloop exits and the window is actually closed.
                try:
                    self.root.quit()
                except Exception:
                    pass
                self.root.destroy()
        except Exception:
            pass

    def _ensure_edit_loaded(self):
        if self._wb_a_edit is not None and self._wb_b_edit is not None and (not self.has_base or self._wb_base_edit is not None):
            return

        # If background preload is running, wait briefly.
        if getattr(self, "_edit_loading_started", False):
            try:
                self._initial_sheet_ready_event.set()
            except Exception:
                pass
            _dlog("waiting for background edit preload")
            self._edit_loaded_event.wait(timeout=10)
            if self._wb_a_edit is not None and self._wb_b_edit is not None and (not self.has_base or self._wb_base_edit is not None):
                return

        _dlog("loading edit workbooks (fallback)")
        with self._edit_fallback_lock:
            # Re-check under lock: background thread may have just finished.
            if self._wb_a_edit is not None and self._wb_b_edit is not None and (not self.has_base or self._wb_base_edit is not None):
                return
            t0 = datetime.now()
            self._wb_a_edit = load_workbook(self.file_a, data_only=False)
            _dlog(f"load wb_a_edit: {(datetime.now()-t0).total_seconds():.3f}s")
            t0 = datetime.now()
            self._wb_b_edit = load_workbook(self.file_b, data_only=False)
            _dlog(f"load wb_b_edit: {(datetime.now()-t0).total_seconds():.3f}s")
            if self.has_base:
                t0 = datetime.now()
                self._wb_base_edit = load_workbook(self.base_path, data_only=False)
                _dlog(f"load wb_base_edit: {(datetime.now()-t0).total_seconds():.3f}s")

    def ws_a_edit(self, sheet: str):
        self._ensure_edit_loaded()
        if self._wb_a_edit is None:
            raise KeyError("file_a edit workbook not available")
        return self._wb_a_edit[sheet]

    def ws_b_edit(self, sheet: str):
        self._ensure_edit_loaded()
        if self._wb_b_edit is None:
            raise KeyError("file_b edit workbook not available")
        return self._wb_b_edit[sheet]

    def ws_base_edit(self, sheet: str):
        self._ensure_edit_loaded()
        if self._wb_base_edit is None:
            raise KeyError("base workbook not available")
        return self._wb_base_edit[sheet]

    def ws_a_val(self, sheet: str):
        if self._wb_a_val is None:
            raise KeyError("file_a val workbook not available")
        return self._wb_a_val[sheet]

    def ws_b_val(self, sheet: str):
        if self._wb_b_val is None:
            raise KeyError("file_b val workbook not available")
        return self._wb_b_val[sheet]

    def ws_base_val(self, sheet: str):
        if self._wb_base_val is None:
            raise KeyError("base workbook not available")
        return self._wb_base_val[sheet]

    def record_manual_a_cell(self, sheet: str, r: int, c: int, edit_value):
        try:
            key = (sheet, int(r), int(c))
            self.manual_a_cell_ops[key] = edit_value
            self.manual_a_formula_cache_ops.pop(key, None)
        except Exception:
            pass

    def record_manual_a_formula_cache(self, sheet: str, r: int, c: int, cached_value):
        try:
            key = (sheet, int(r), int(c))
            self.manual_a_formula_cache_ops[key] = cached_value
        except Exception:
            pass

    def record_manual_b_cell(self, sheet: str, r: int, c: int, edit_value):
        try:
            key = (sheet, int(r), int(c))
            self.manual_b_cell_ops[key] = edit_value
            self.manual_b_formula_cache_ops.pop(key, None)
        except Exception:
            pass

    def record_manual_b_formula_cache(self, sheet: str, r: int, c: int, cached_value):
        try:
            self.manual_b_formula_cache_ops[(sheet, int(r), int(c))] = cached_value
        except Exception:
            pass

    def clear_manual_b_formula_cache(self, sheet: str, r: int, c: int):
        try:
            self.manual_b_formula_cache_ops.pop((sheet, int(r), int(c)), None)
        except Exception:
            pass

    def record_manual_a_row_insert(
        self,
        sheet: str,
        row: int,
        count: int = 1,
        source_side: str | None = None,
        source_rows: list[int] | None = None,
    ):
        try:
            op = {
                "sheet": sheet,
                "kind": "insert_rows",
                "row": int(row),
                "count": max(1, int(count)),
            }
            if source_side:
                op["source_side"] = str(source_side).upper()
            if source_rows:
                op["source_rows"] = [int(value) for value in source_rows]
            self.manual_a_row_ops.append(op)
        except Exception:
            pass

    def record_manual_b_row_insert(
        self,
        sheet: str,
        row: int,
        count: int = 1,
        source_side: str | None = None,
        source_rows: list[int] | None = None,
    ):
        op = {
            "sheet": sheet,
            "kind": "insert_rows",
            "row": int(row),
            "count": max(1, int(count)),
        }
        if source_side:
            op["source_side"] = str(source_side).upper()
        if source_rows:
            op["source_rows"] = [int(value) for value in source_rows]
        self.manual_b_row_ops.append(op)

    def remove_last_manual_a_row_insert(self, sheet: str, row: int, count: int = 1) -> bool:
        row = int(row)
        count = max(1, int(count))
        for i in range(len(self.manual_a_row_ops) - 1, -1, -1):
            op = self.manual_a_row_ops[i]
            try:
                if (
                    op.get("sheet") == sheet and
                    op.get("kind") == "insert_rows" and
                    int(op.get("row", 0)) == row and
                    int(op.get("count", 1)) == count
                ):
                    del self.manual_a_row_ops[i]
                    return True
            except Exception:
                continue
        return False

    def remove_last_manual_b_row_insert(self, sheet: str, row: int, count: int = 1) -> bool:
        row = int(row)
        count = max(1, int(count))
        for idx in range(len(self.manual_b_row_ops) - 1, -1, -1):
            op = self.manual_b_row_ops[idx]
            if (
                op.get("sheet") == sheet
                and op.get("kind") == "insert_rows"
                and int(op.get("row", 0)) == row
                and int(op.get("count", 1)) == count
            ):
                del self.manual_b_row_ops[idx]
                return True
        return False

    def record_manual_sheet_copy(self, sheet: str, source_side: str, target_side: str):
        try:
            self.manual_sheet_ops.append({
                "kind": "copy_sheet",
                "sheet": sheet,
                "source_side": str(source_side or "").upper(),
                "target_side": str(target_side or "").upper(),
            })
        except Exception:
            pass

    def record_manual_sheet_delete(self, sheet: str, target_side: str):
        try:
            self.manual_sheet_ops.append({
                "kind": "delete_sheet",
                "sheet": sheet,
                "target_side": str(target_side or "").upper(),
            })
        except Exception:
            pass

    def _workbooks_for_side(self, side: str):
        side = str(side or "").upper()
        if side == "A":
            return getattr(self, "_wb_a_val", None), getattr(self, "_wb_a_edit", None)
        if side in ("B", "THEIRS"):
            return getattr(self, "_wb_b_val", None), getattr(self, "_wb_b_edit", None)
        if side == "BASE":
            return getattr(self, "_wb_base_val", None), getattr(self, "_wb_base_edit", None)
        return None, None

    def _copy_sheet_between_sides(self, sheet: str, source_side: str, target_side: str):
        self._ensure_edit_loaded()
        source_side = str(source_side or "").upper()
        target_side = str(target_side or "").upper()
        src_val_wb, src_edit_wb = self._workbooks_for_side(source_side)
        dst_val_wb, dst_edit_wb = self._workbooks_for_side(target_side)
        if src_edit_wb is None or dst_edit_wb is None or sheet not in src_edit_wb.sheetnames:
            raise KeyError(f"sheet copy unavailable: {sheet} {source_side}->{target_side}")
        src_val_ws = src_val_wb[sheet] if src_val_wb is not None and sheet in src_val_wb.sheetnames else src_edit_wb[sheet]
        src_edit_ws = src_edit_wb[sheet]
        _create_sheet_from_source(dst_edit_wb, src_edit_ws, sheet)
        _create_sheet_from_source(dst_val_wb, src_val_ws, sheet)
        if target_side == "A":
            self.modified_a = True
            self.modified_sheets_a.add(sheet)
            self.record_manual_sheet_copy(sheet, source_side, target_side)
        elif target_side == "B":
            self.modified_b = True
            self.modified_sheets_b.add(sheet)
            self.record_manual_sheet_copy(sheet, source_side, target_side)
        self._refresh_sheet_catalog()

    def _delete_sheet_on_side(self, sheet: str, target_side: str):
        self._ensure_edit_loaded()
        target_side = str(target_side or "").upper()
        dst_val_wb, dst_edit_wb = self._workbooks_for_side(target_side)
        changed = _remove_sheet_if_exists(dst_edit_wb, sheet)
        _remove_sheet_if_exists(dst_val_wb, sheet)
        if not changed:
            return False
        if target_side == "A":
            self.modified_a = True
            self.modified_sheets_a.add(sheet)
            self.record_manual_sheet_delete(sheet, target_side)
        elif target_side == "B":
            self.modified_b = True
            self.modified_sheets_b.add(sheet)
            self.record_manual_sheet_delete(sheet, target_side)
        self._refresh_sheet_catalog()
        if sheet not in self.sheet_meta:
            self.display_sheets.append(sheet)
            self.sheet_meta[sheet] = {"has_a": False, "has_b": False, "has_base": False, "view_mode": "missing_sheet"}
            self.sheet_diff_state[sheet] = 0
        return True

    def _sheet_ops_for_target(self, target_side: str, *, include_auto: bool = False):
        target_side = str(target_side or "").upper()
        ops = []
        if include_auto:
            ops.extend(getattr(self, "auto_sheet_ops", []) or [])
        ops.extend(getattr(self, "manual_sheet_ops", []) or [])
        return [
            dict(op)
            for op in ops
            if str(op.get("target_side") or "A").upper() == target_side
        ]

    def _clear_sheet_ops_for_target(self, target_side: str):
        target_side = str(target_side or "").upper()
        self.manual_sheet_ops[:] = [
            op for op in self.manual_sheet_ops
            if str(op.get("target_side") or "A").upper() != target_side
        ]

    def build_manual_merge_output_file(self):
        """Build merge output by XML-level patching from pristine mine snapshot."""
        src = self._merge_mine_snapshot if (self._merge_mine_snapshot and os.path.exists(self._merge_mine_snapshot)) else self.file_a
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out = os.path.join(tempfile.gettempdir(), f"{APP_NAME}_merged_output_{os.getpid()}_{ts}{_workbook_ext(src)}")
        sheet_ops = self._sheet_ops_for_target("A", include_auto=True)
        manual_ops = _prepare_manual_ops_for_save(
            src,
            self.manual_a_cell_ops,
            row_ops=self.manual_a_row_ops,
            sheet_ops=sheet_ops,
        )
        formula_cache_values = dict(getattr(self, "manual_a_formula_cache_ops", {}) or {})
        literal_text_values = {
            key: value for key, value in manual_ops.items()
            if isinstance(value, _LiteralText)
        }
        cached_values = dict(formula_cache_values)
        for sheet, row_idx, col_idx in manual_ops.keys():
            try:
                cached_values[(sheet, row_idx, col_idx)] = self.ws_a_val(sheet).cell(
                    row=int(row_idx), column=int(col_idx)
                ).value
            except Exception:
                pass
        zip_ops = dict(manual_ops)
        for key in formula_cache_values.keys():
            sheet, row_idx, col_idx = key
            try:
                formula = _formula_text(
                    self.ws_a_edit(sheet).cell(row=int(row_idx), column=int(col_idx)).value
                )
                if formula:
                    zip_ops[key] = formula
            except Exception:
                pass
        source_paths = {"B": self.file_b}
        if getattr(self, "base_path", None):
            source_paths["BASE"] = self.base_path
        if not zip_ops and not self.manual_a_row_ops and not sheet_ops:
            shutil.copy2(src, out)
            return out

        # Cell-only replay is safer at OOXML level: untouched formula caches and
        # shared-formula metadata stay intact, and text values keep their type.
        if not self.manual_a_row_ops and not sheet_ops:
            try:
                _build_manual_merge_xlsx_via_zip(
                    src,
                    out,
                    zip_ops,
                    cached_values=cached_values,
                    cache_only_keys=set(formula_cache_values),
                )
                return out
            except Exception as e:
                _dlog(f"validated cell-only ZIP patch failed; trying Excel native save: {e}")

        if _EXCEL_NATIVE_SAVE_ON_MERGE:
            ok = _build_manual_merge_output_with_excel(
                src,
                out,
                manual_ops,
                self.manual_a_row_ops,
                sheet_ops=sheet_ops,
                source_paths=source_paths,
            )
            if ok:
                if formula_cache_values or literal_text_values:
                    cache_out = out + ".formula-cache.xlsx"
                    post_save_ops = {
                        key: zip_ops[key] for key in formula_cache_values
                        if key in zip_ops
                    }
                    post_save_ops.update(literal_text_values)
                    _build_manual_merge_xlsx_via_zip(
                        out,
                        cache_out,
                        post_save_ops,
                        cached_values=formula_cache_values,
                        cache_only_keys=set(formula_cache_values),
                    )
                    os.replace(cache_out, out)
                return out
            _dlog("WARNING: excel native save failed, trying openpyxl replay fallback")
        if self.manual_a_row_ops or sheet_ops:
            unsafe_sources = [
                path for path in _replay_formula_source_paths(
                    src,
                    row_ops=self.manual_a_row_ops,
                    sheet_ops=sheet_ops,
                    source_paths=source_paths,
                )
                if _xlsx_requires_native_structural_replay(path)
            ]
            if unsafe_sources:
                raise RuntimeError(
                    "Excel 原生保存失败。为避免 openpyxl 回退破坏公式缓存或高级 Sheet 对象，"
                    "已停止保存且未替换目标文件；请关闭占用文件的 Excel 后重试。"
                    f"\n需要原生回放的来源文件：{', '.join(os.path.basename(path) for path in unsafe_sources)}"
                )
            ok = _build_manual_merge_output_with_openpyxl(
                src,
                out,
                manual_ops,
                self.manual_a_row_ops,
                sheet_ops=sheet_ops,
                source_paths=source_paths,
            )
            if ok:
                return out
            raise RuntimeError("manual merge output fallback failed for sheet/row operations")
        raise RuntimeError(
            "无法安全保存所选公式单元格：ZIP 校验未通过且 Excel 原生保存不可用。"
            "目标文件未被替换。"
        )

    def build_manual_b_output_file(self):
        """Build a 2-way B-side result by replaying structural operations safely."""
        src = self.file_b
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out = os.path.join(
            tempfile.gettempdir(),
            f"{APP_NAME}_b_output_{os.getpid()}_{ts}{_workbook_ext(src)}",
        )
        row_ops = list(getattr(self, "manual_b_row_ops", []) or [])
        sheet_ops = self._sheet_ops_for_target("B")
        manual_ops = _prepare_manual_ops_for_save(
            src,
            self.manual_b_cell_ops,
            row_ops=row_ops,
            sheet_ops=sheet_ops,
        )
        formula_cache_values = dict(getattr(self, "manual_b_formula_cache_ops", {}) or {})
        literal_text_values = {
            key: value for key, value in manual_ops.items()
            if isinstance(value, _LiteralText)
        }
        cached_values = dict(formula_cache_values)
        for sheet, row_idx, col_idx in manual_ops.keys():
            try:
                cached_values[(sheet, row_idx, col_idx)] = self.ws_b_val(sheet).cell(
                    row=int(row_idx), column=int(col_idx)
                ).value
            except Exception:
                pass
        zip_ops = dict(manual_ops)
        for key in formula_cache_values.keys():
            sheet, row_idx, col_idx = key
            try:
                formula = _formula_text(
                    self.ws_b_edit(sheet).cell(row=int(row_idx), column=int(col_idx)).value
                )
                if formula:
                    zip_ops[key] = formula
            except Exception:
                pass
        if not zip_ops and not row_ops and not sheet_ops:
            shutil.copy2(src, out)
            return out
        if not row_ops and not sheet_ops:
            _build_manual_merge_xlsx_via_zip(
                src,
                out,
                zip_ops,
                cached_values=cached_values,
                cache_only_keys=set(formula_cache_values),
            )
            return out

        source_paths = {"A": self.file_a}
        if _EXCEL_NATIVE_SAVE_ON_MERGE and _build_manual_merge_output_with_excel(
            src,
            out,
            manual_ops,
            row_ops,
            sheet_ops=sheet_ops,
            source_paths=source_paths,
        ):
            if formula_cache_values or literal_text_values:
                cache_out = out + ".formula-cache.xlsx"
                post_save_ops = {
                    key: zip_ops[key] for key in formula_cache_values
                    if key in zip_ops
                }
                post_save_ops.update(literal_text_values)
                _build_manual_merge_xlsx_via_zip(
                    out,
                    cache_out,
                    post_save_ops,
                    cached_values=formula_cache_values,
                    cache_only_keys=set(formula_cache_values),
                )
                os.replace(cache_out, out)
            return out

        unsafe_sources = [
            path for path in _replay_formula_source_paths(
                src,
                row_ops=row_ops,
                sheet_ops=sheet_ops,
                source_paths=source_paths,
            )
            if _xlsx_requires_native_structural_replay(path)
        ]
        if unsafe_sources:
            raise RuntimeError(
                "Excel 原生保存失败。为避免 openpyxl 插行破坏公式引用或高级 Sheet 对象，"
                "已停止保存且未替换目标文件；请关闭占用文件的 Excel 后重试。"
                f"\n需要原生回放的来源文件：{', '.join(os.path.basename(path) for path in unsafe_sources)}"
            )
        if _build_manual_merge_output_with_openpyxl(
            src,
            out,
            manual_ops,
            row_ops,
            sheet_ops=sheet_ops,
            source_paths=source_paths,
        ):
            return out
        raise RuntimeError("B-side structural replay failed")

    def set_sheet_has_diff(self, sheet: str, has: bool, confirmed: bool = True):
        # Keep API: mark sheet diff state
        if sheet not in self.sheet_diff_state:
            return
        confirmed_sheets = getattr(self, "_sheet_diff_confirmed", None)
        if confirmed_sheets is None:
            self._sheet_diff_confirmed = set()
            confirmed_sheets = self._sheet_diff_confirmed
        if not confirmed and sheet in confirmed_sheets:
            # A late Stage-1 callback must never overwrite an exact result.
            return
        if confirmed:
            confirmed_sheets.add(sheet)
        if has:
            self.sheet_diff_state[sheet] = 2 if confirmed else max(self.sheet_diff_state[sheet], 1)
        else:
            # only downgrade when confirmed
            if confirmed:
                self.sheet_diff_state[sheet] = 0

    def _active_sheet_name(self) -> str | None:
        try:
            tab_id = self.nb.select()
            if not tab_id:
                return None
            tab_text = str(self.nb.tab(tab_id, "text") or "").strip()
            return tab_text or None
        except Exception:
            return None

    def _on_global_f4(self, event=None):
        """Route F4 to the currently active sheet view only."""
        try:
            sheet = self._active_sheet_name()
            if not sheet:
                return None
            view = self.sheet_views.get(sheet)
            if view is None and sheet in self._sheet_containers:
                try:
                    self.nb.select(self._sheet_containers[sheet])
                    self.root.update_idletasks()
                    self.root.update()
                    view = self.sheet_views.get(sheet)
                except Exception:
                    view = None
            if view is None:
                return None
            return view._on_hover_compare_f4_toggle(event)
        except Exception:
            return None

    def _set_task_status(
        self,
        message: str,
        *,
        active: bool = False,
        current: int | None = None,
        total: int | None = None,
    ):
        """Update the always-visible application task status strip."""
        label = getattr(self, "task_status_var", None)
        progress = getattr(self, "task_progress", None)
        if label is None or progress is None:
            return
        try:
            label.set(message)
            if active:
                if total and total > 0 and current is not None:
                    progress.stop()
                    progress.configure(mode="determinate", maximum=max(1, total), value=max(0, min(current, total)))
                else:
                    progress.configure(mode="indeterminate")
                    progress.start(12)
            else:
                progress.stop()
                if total and total > 0:
                    progress.configure(mode="determinate", maximum=total, value=total)
                else:
                    progress.configure(mode="determinate", maximum=100, value=0)
        except Exception:
            pass

    def _build_ui(self):
        top = ttk.Frame(self.root)
        top.pack(fill="x", padx=10, pady=8)

        # Keep top area minimal (summary + actions). File-source labels are shown inside each Sheet.
        summary = f"同名Sheet: {len(self.common_sheets)}   仅A: {len(self.only_a)}   仅B: {len(self.only_b)}"
        ttk.Label(top, text=summary).grid(row=0, column=0, columnspan=2, sticky="w", pady=(2, 0))
        detail_row = 1
        if self.merge_mode and (self.raw_mine or self.raw_base or self.raw_theirs):
            raw_line = (
                f"SVN原始传参: mine={os.path.basename(self.raw_mine or '-')}"
                f" | base={os.path.basename(self.raw_base or '-')}"
                f" | theirs={os.path.basename(self.raw_theirs or '-')}"
            )
            ttk.Label(top, text=raw_line, foreground="#555").grid(row=detail_row, column=0, columnspan=3, sticky="w", pady=(4, 0))
            detail_row += 1
            read_line = (
                f"当前实际读取: left(A)={os.path.basename(self.file_a or '-')}"
                f" | base={os.path.basename(self.base_path or '-')}"
                f" | right(B)={os.path.basename(self.file_b or '-')}"
            )
            ttk.Label(top, text=read_line, foreground="#555").grid(row=detail_row, column=0, columnspan=3, sticky="w", pady=(2, 0))
        ttk.Label(top, text=f"Build: {APP_BUILD_TAG}", foreground="#666").grid(row=0, column=3, sticky="ne", padx=(16, 0))

        ttk.Button(top, text="重算并刷新", command=self.recalc_and_refresh).grid(row=0, column=2, sticky="ne", padx=(10, 0))
        ttk.Button(top, text="导出诊断包", command=self.export_diagnostic_bundle).grid(row=0, column=4, sticky="ne", padx=(10, 0))
        ttk.Button(top, text="复制反馈信息", command=self.copy_feedback_info).grid(row=0, column=5, sticky="ne", padx=(10, 0))
        self.update_btn = ttk.Button(top, text="检查更新", command=self._do_svn_update)
        self.update_btn.grid(row=0, column=6, sticky="ne", padx=(10, 0))

        ttk.Separator(self.root, orient="horizontal").pack(fill="x", padx=10, pady=(0, 4))

        self.task_status_frame = ttk.Frame(self.root)
        self.task_status_frame.pack(fill="x", padx=10, pady=(0, 4))
        self.task_status_var = tk.StringVar(value="正在准备 Sheet 数据...")
        ttk.Label(self.task_status_frame, textvariable=self.task_status_var, foreground="#555").pack(
            side="left", fill="x", expand=True
        )
        self.task_progress = ttk.Progressbar(self.task_status_frame, mode="indeterminate", length=240)
        self.task_progress.pack(side="right", padx=(12, 0))
        self.task_progress.start(12)

        self.nb = ttk.Notebook(self.root)
        try:
            self.root.bind("<F4>", self._on_global_f4)
        except Exception:
            pass

        # Bottom bar: sheet nav (only)
        self.bottom = ttk.Frame(self.root)
        self.bottom.pack(side="bottom", fill="x", padx=10, pady=(0, 10))

        self.nav = ttk.Frame(self.bottom)
        self.nav.pack(side="left", fill="x", expand=True)
        ttk.Label(self.nav, text="Sheets（浅黄=预检，亮黄=确认）:").pack(side="left")
        self.nav_canvas = tk.Canvas(self.nav, height=28, highlightthickness=0)
        self.nav_canvas.pack(side="left", fill="x", expand=True, padx=(8, 0))
        self.nav_scroll = ttk.Scrollbar(self.nav, orient="horizontal", command=self.nav_canvas.xview)
        self.nav_scroll.pack(side="bottom", fill="x")
        self.nav_canvas.configure(xscrollcommand=self.nav_scroll.set)
        self.nav_inner = ttk.Frame(self.nav_canvas)
        self.nav_canvas.create_window((0, 0), window=self.nav_inner, anchor="nw")
        self.nav_inner.bind("<Configure>", lambda e: self.nav_canvas.configure(scrollregion=self.nav_canvas.bbox("all")))
        self.nb.pack(side="top", fill="both", expand=True, padx=10, pady=(8, 6))

        # Tabs are created up-front, but heavy SheetView is created lazily on first activation.
        self.sheet_views = {}
        self._sheet_loaded = {}
        self._sheet_containers = {}
        for s in self.display_sheets:
            container = ttk.Frame(self.nb)
            self._sheet_containers[s] = container
            self.nb.add(container, text=s)
            self.sheet_views[s] = None
            self._sheet_loaded[s] = False

        # Background compute queue for sheet diffs
        self._compute_lock = threading.Lock()
        self._compute_queue = []  # list of sheet names
        self._compute_inflight = set()
        self._compute_thread = None
        self._compute_total = max(1, len(self.compare_sheets))
        self._compute_done = 0
        # Keep completed background results for tabs that have not been opened yet.
        # Discarding them forced a second full scan on first tab activation.
        self._sheet_cache_store: dict[str, dict] = {}
        self._ui_task_lock = threading.Lock()
        self._ui_tasks = []

        def _enqueue_sheet(sheet: str, front: bool = False):
            if self._is_closing:
                return
            with self._compute_lock:
                if sheet in self._compute_inflight:
                    return
                if sheet in self._compute_queue:
                    # move to front if requested
                    if front:
                        self._compute_queue.remove(sheet)
                        self._compute_queue.insert(0, sheet)
                    return
                if front:
                    self._compute_queue.insert(0, sheet)
                else:
                    self._compute_queue.append(sheet)

        def _queue_ui_task(fn):
            if self._is_closing:
                return False
            with self._ui_task_lock:
                if self._is_closing:
                    return False
                self._ui_tasks.append(fn)
            return True

        def _drain_ui_tasks():
            if self._is_closing:
                return
            tasks = []
            try:
                with self._ui_task_lock:
                    if self._ui_tasks:
                        tasks = self._ui_tasks
                        self._ui_tasks = []
            except Exception:
                tasks = []
            ran = bool(tasks)
            for fn in tasks:
                try:
                    fn()
                except Exception as e:
                    _dlog(f"ui task failed: {e}")
            try:
                # Adaptive delay: poll frequently while work is flowing, back off when idle.
                delay = 50 if ran else 150
                self._safe_root_after(delay, _drain_ui_tasks)
            except Exception:
                pass

        def _check_bg_cancel():
            if self._is_closing:
                raise InterruptedError("background compute cancelled during shutdown")
            # Openpyxl XML parsing is CPU/GIL heavy. Yield while a user-triggered
            # overwrite/insert or active-Sheet exact diff is running so button
            # feedback and the selected Sheet always win over unopened tabs.
            while self._interactive_action_event.is_set() or self._priority_diff_event.is_set():
                if self._is_closing:
                    raise InterruptedError("background compute cancelled during shutdown")
                time.sleep(0.01)

        def _compute_trim_bounds(ws, row_cache=None):
            # Find the true last non-empty row for this sheet. Empty strings are
            # treated as empty so formulas returning "" do not expand the bounds.
            # Keep max_c on read-only fallback paths so earlier wide rows are preserved.
            max_r = ws.max_row or 1
            max_c = ws.max_column or 1
            last_r = 1
            last_c = 1
            found = False
            found_via_cells = False

            # ReadOnlyWorksheet.cell()/high-min_row iteration reparses the XML
            # stream from the beginning. A single forward pass is both exact and
            # avoids the old worst case of doing that up to 5000 times.
            if ws.__class__.__name__ == "ReadOnlyWorksheet":
                captured_rows = [] if row_cache is not None else None
                for row_idx, row in enumerate(
                    ws.iter_rows(min_row=1, max_row=max_r, min_col=1, max_col=max_c, values_only=True),
                    start=1,
                ):
                    if (row_idx & 127) == 0:
                        _check_bg_cancel()
                    if captured_rows is not None:
                        captured_rows.append(tuple(row or ()))
                    row_last_col = 0
                    for col_idx, value in enumerate(row, start=1):
                        if value not in (None, ""):
                            row_last_col = col_idx
                    if row_last_col:
                        found = True
                        last_r = row_idx
                        last_c = max(last_c, row_last_col)
                if row_cache is not None:
                    row_cache[id(ws)] = captured_rows
                if not found:
                    return 1, max(1, max_c)
                return max(1, last_r), max(1, last_c)

            try:
                cells = getattr(ws, "_cells", None)
                if cells:
                    for cell_idx, cell in enumerate(cells.values(), start=1):
                        if (cell_idx & 1023) == 0:
                            _check_bg_cancel()
                        v = cell.value
                        if v not in (None, ""):
                            found = True
                            found_via_cells = True
                            if cell.row > last_r:
                                last_r = cell.row
                            if cell.column > last_c:
                                last_c = cell.column
            except Exception:
                pass

            if not found:
                for r in range(max_r, max(0, max_r - 5000), -1):
                    if (r & 127) == 0:
                        _check_bg_cancel()
                    row = next(ws.iter_rows(min_row=r, max_row=r, min_col=1, max_col=max_c, values_only=True), ())
                    if any(v not in (None, "") for v in row):
                        found = True
                        last_r = r
                        break
            if not found:
                return 1, max(1, max_c)

            if found_via_cells:
                # Cached-values-only reads can miss formula columns with no
                # cache; prefer worksheet max_col in that situation.
                if _USE_CACHED_VALUES_ONLY and max_c > last_c:
                    return max(1, last_r), max(1, max_c)
                if last_c <= 1 and max_c > last_c:
                    return max(1, last_r), max(1, max_c)
                return max(1, last_r), max(1, last_c)
            return max(1, last_r), max(1, max_c)

        def _compute_row_pairs_bg(
            ws_a,
            ws_b,
            max_row_a: int,
            max_row_b: int,
            max_col: int,
            signature_cache=None,
            rows_cache=None,
        ):
            """Compute row alignment pairs using difflib.SequenceMatcher (background-safe)."""
            if not _should_auto_row_align(max_row_a, max_row_b, force=False):
                max_row = max(max_row_a, max_row_b)
                pairs = []
                for r in range(1, max_row + 1):
                    ra = r if r <= max_row_a else None
                    rb = r if r <= max_row_b else None
                    pairs.append((ra, rb))
                return pairs

            def _bulk_sig_list(ws, max_row_local: int):
                cache_key = (id(ws), int(max_row_local), int(max_col))
                if signature_cache is not None and cache_key in signature_cache:
                    return signature_cache[cache_key]
                cached_rows = rows_cache.get(id(ws)) if rows_cache is not None else None
                if cached_rows is not None and len(cached_rows) >= max_row_local:
                    sigs = []
                    for row_idx, row in enumerate(cached_rows[:max_row_local], start=1):
                        if (row_idx & 127) == 0:
                            _check_bg_cancel()
                        sigs.append(_row_signature(_pad_row_values(row, max_col)))
                else:
                    sigs = []
                    try:
                        for row_idx, row in enumerate(ws.iter_rows(
                            min_row=1,
                            max_row=max_row_local,
                            min_col=1,
                            max_col=max_col,
                            values_only=True,
                        ), start=1):
                            if (row_idx & 127) == 0:
                                _check_bg_cancel()
                            sigs.append(_row_signature(row or ()))
                    except InterruptedError:
                        raise
                    except Exception:
                        return []
                if signature_cache is not None:
                    signature_cache[cache_key] = sigs
                return sigs
            sig_a = _bulk_sig_list(ws_a, max_row_a)
            sig_b = _bulk_sig_list(ws_b, max_row_b)
            return _compute_row_pairs_from_signatures(sig_a, sig_b)

        def _has_diff_by_blocks_bg(ws_a, ws_b, max_row_a: int, max_row_b: int, max_col: int):
            max_row = max(max_row_a, max_row_b)
            block = _LARGE_SHEET_BLOCK_ROWS
            for block_end in range(max_row, 0, -block):
                _check_bg_cancel()
                block_start = max(1, block_end - block + 1)
                rows_a = {}
                rows_b = {}
                if block_start <= max_row_a:
                    for idx, row in enumerate(
                        ws_a.iter_rows(
                            min_row=block_start,
                            max_row=min(block_end, max_row_a),
                            min_col=1,
                            max_col=max_col,
                            values_only=True,
                        ),
                        start=block_start,
                    ):
                        if (idx & 127) == 0:
                            _check_bg_cancel()
                        rows_a[idx] = row or ()
                if block_start <= max_row_b:
                    for idx, row in enumerate(
                        ws_b.iter_rows(
                            min_row=block_start,
                            max_row=min(block_end, max_row_b),
                            min_col=1,
                            max_col=max_col,
                            values_only=True,
                        ),
                        start=block_start,
                    ):
                        if (idx & 127) == 0:
                            _check_bg_cancel()
                        rows_b[idx] = row or ()
                for r in range(block_end, block_start - 1, -1):
                    row_a = rows_a.get(r, ())
                    row_b = rows_b.get(r, ())
                    for ci in range(max_col):
                        va = row_a[ci] if ci < len(row_a) else None
                        vb = row_b[ci] if ci < len(row_b) else None
                        if _merge_cmp_value(va) != _merge_cmp_value(vb):
                            return True
            return False

        def _pairs_have_diff_bg(ws_left, ws_right, row_pairs, max_col: int):
            if not row_pairs:
                return False
            block = _LARGE_SHEET_BLOCK_ROWS
            pair_count = len(row_pairs)
            for block_end in range(pair_count, 0, -block):
                _check_bg_cancel()
                block_start = max(0, block_end - block)
                block_pairs = row_pairs[block_start:block_end]
                left_rows_needed = sorted({left for left, _right in block_pairs if left is not None})
                right_rows_needed = sorted({right for _left, right in block_pairs if right is not None})
                rows_left = {}
                rows_right = {}

                if left_rows_needed:
                    min_left = left_rows_needed[0]
                    max_left = left_rows_needed[-1]
                    for idx, row in enumerate(
                        ws_left.iter_rows(
                            min_row=min_left,
                            max_row=max_left,
                            min_col=1,
                            max_col=max_col,
                            values_only=True,
                        ),
                        start=min_left,
                    ):
                        if (idx & 127) == 0:
                            _check_bg_cancel()
                        rows_left[idx] = row or ()
                if right_rows_needed:
                    min_right = right_rows_needed[0]
                    max_right = right_rows_needed[-1]
                    for idx, row in enumerate(
                        ws_right.iter_rows(
                            min_row=min_right,
                            max_row=max_right,
                            min_col=1,
                            max_col=max_col,
                            values_only=True,
                        ),
                        start=min_right,
                    ):
                        if (idx & 127) == 0:
                            _check_bg_cancel()
                        rows_right[idx] = row or ()

                for left_row, right_row in reversed(block_pairs):
                    row_left = rows_left.get(left_row, ()) if left_row is not None else ()
                    row_right = rows_right.get(right_row, ()) if right_row is not None else ()
                    for ci in range(max_col):
                        vl = row_left[ci] if ci < len(row_left) else None
                        vr = row_right[ci] if ci < len(row_right) else None
                        if _merge_cmp_value(vl) != _merge_cmp_value(vr):
                            return True
            return False

        def _sheet_has_base_diff_bg(
            ws_a,
            ws_b,
            ws_base,
            max_r_a: int,
            max_r_b: int,
            max_r_base: int,
            max_col: int,
            signature_cache=None,
            rows_cache=None,
        ):
            if ws_base is None:
                return True
            row_pairs_a_base = _compute_row_pairs_bg(
                ws_a, ws_base, max_r_a, max_r_base, max_col, signature_cache, rows_cache
            )
            if _pairs_have_diff_bg(ws_a, ws_base, row_pairs_a_base, max_col):
                return True
            row_pairs_b_base = _compute_row_pairs_bg(
                ws_b, ws_base, max_r_b, max_r_base, max_col, signature_cache, rows_cache
            )
            if _pairs_have_diff_bg(ws_b, ws_base, row_pairs_b_base, max_col):
                return True
            return False

        def _compute_sheet_cache(
            wb_a_val,
            wb_b_val,
            wb_a_edit,
            wb_b_edit,
            sheet: str,
            wb_base_val=None,
            wb_base_edit=None,
        ):
            _check_bg_cancel()
            trimmed_rows_cache = {}
            ws_a = wb_a_val[sheet]
            ws_b = wb_b_val[sheet]
            max_r_a, max_c_a = _compute_trim_bounds(ws_a, trimmed_rows_cache)
            max_r_b, max_c_b = _compute_trim_bounds(ws_b, trimmed_rows_cache)
            _check_bg_cancel()
            ws_a_edit = wb_a_edit[sheet]
            ws_b_edit = wb_b_edit[sheet]
            edit_r_a, edit_c_a = _compute_trim_bounds(ws_a_edit)
            edit_r_b, edit_c_b = _compute_trim_bounds(ws_b_edit)
            max_r_a, max_c_a = max(max_r_a, edit_r_a), max(max_c_a, edit_c_a)
            max_r_b, max_c_b = max(max_r_b, edit_r_b), max(max_c_b, edit_c_b)
            ws_base = None
            ws_base_edit = None
            max_r_base = 0
            max_c_base = 0
            if wb_base_val is not None and sheet in wb_base_val.sheetnames:
                ws_base = wb_base_val[sheet]
                max_r_base, max_c_base = _compute_trim_bounds(ws_base, trimmed_rows_cache)
                if wb_base_edit is not None and sheet in wb_base_edit.sheetnames:
                    ws_base_edit = wb_base_edit[sheet]
                    edit_r_base, edit_c_base = _compute_trim_bounds(ws_base_edit)
                    max_r_base = max(max_r_base, edit_r_base)
                    max_c_base = max(max_c_base, edit_c_base)
            _check_bg_cancel()
            max_row = max(max_r_a, max_r_b, max_r_base)
            max_col = max(max_c_a, max_c_b, max_c_base)
            signature_cache = {}

            # Compute row-aligned pairs (same algorithm as SheetView._build_row_pairs)
            row_pairs = _compute_row_pairs_bg(
                ws_a, ws_b, max_r_a, max_r_b, max_col, signature_cache, trimmed_rows_cache
            )
            _check_bg_cancel()

            pair_diff_cols: dict[int, set] = {}
            pair_base_diff_cols: dict[int, set] = {}
            # Keep raw per-cell display parts in cache; render with current grid mode on UI thread.
            pair_parts_a: dict[int, list[str]] = {}
            pair_parts_b: dict[int, list[str]] = {}
            col_char_widths: dict[int, int] = {}
            row_a_to_pair_idx: dict[int, int] = {}
            row_b_to_pair_idx: dict[int, int] = {}
            mine_to_base_row: dict[int, int] = {}
            theirs_to_base_row: dict[int, int] = {}
            pair_base_row_override: dict[int, int | None] = {}

            if ws_base is not None:
                try:
                    mine_to_base_row = _row_map_from_pairs(_compute_row_pairs_bg(
                        ws_a, ws_base, max_r_a, max_r_base, max_col, signature_cache, trimmed_rows_cache
                    ))
                except Exception:
                    mine_to_base_row = {}
                try:
                    theirs_to_base_row = _row_map_from_pairs(_compute_row_pairs_bg(
                        ws_b, ws_base, max_r_b, max_r_base, max_col, signature_cache, trimmed_rows_cache
                    ))
                except Exception:
                    theirs_to_base_row = {}
                row_pairs = _split_tail_independent_append_pairs(
                    row_pairs, mine_to_base_row, theirs_to_base_row,
                    ws_a, ws_b, max_col,
                )
                row_pairs = _split_low_similarity_tail_pairs(
                    row_pairs,
                    mine_to_base_row,
                    theirs_to_base_row,
                    ws_a,
                    ws_b,
                    max_col,
                )
                pair_base_row_override = _build_pair_base_row_overrides(
                    row_pairs,
                    mine_to_base_row,
                    theirs_to_base_row,
                    ws_base,
                    ws_a,
                    ws_b,
                    max_col,
                )
                _check_bg_cancel()

            for idx, (ra, rb) in enumerate(row_pairs):
                if ra is not None:
                    row_a_to_pair_idx[ra] = idx
                if rb is not None:
                    row_b_to_pair_idx[rb] = idx

            # Large-sheet fast open: avoid full cell-by-cell precompute.
            # Still estimate display widths from head + tail samples to prevent 4-char collapse.
            if max_row >= _LARGE_SHEET_ROW_THRESHOLD:
                has_diff = _has_diff_by_blocks_bg(ws_a, ws_b, max_r_a, max_r_b, max_col)
                if (not has_diff) and getattr(self, "has_base", False):
                    has_diff = _sheet_has_base_diff_bg(
                        ws_a,
                        ws_b,
                        ws_base,
                        max_r_a,
                        max_r_b,
                        max_r_base,
                        max_col,
                        signature_cache,
                        trimmed_rows_cache,
                    )

                sample_head = min(_LARGE_SHEET_INITIAL_ROWS, len(row_pairs))
                sample_indices = list(range(sample_head))
                tail_n = 50
                tail_start = max(sample_head, len(row_pairs) - tail_n)
                if tail_start < len(row_pairs):
                    sample_indices.extend(range(tail_start, len(row_pairs)))

                sample_rows_a = _read_rows_into_cache(
                    ws_a,
                    [row_pairs[idx][0] for idx in sample_indices],
                    max_col,
                    cancel_check=_check_bg_cancel,
                )
                sample_rows_b = _read_rows_into_cache(
                    ws_b,
                    [row_pairs[idx][1] for idx in sample_indices],
                    max_col,
                    cancel_check=_check_bg_cancel,
                )

                for idx in sample_indices:
                    if (idx & 31) == 0:
                        _check_bg_cancel()
                    ra, rb = row_pairs[idx]
                    row_a_vals = _row_from_cache(sample_rows_a, ra, max_col)
                    row_b_vals = _row_from_cache(sample_rows_b, rb, max_col)
                    for ci in range(max_col):
                        va = row_a_vals[ci] if ci < len(row_a_vals) else None
                        vb = row_b_vals[ci] if ci < len(row_b_vals) else None
                        sa = _val_to_str(va)
                        sb = _val_to_str(vb)
                        w = min(max(len(sa), len(sb)), _COL_MAX_DISPLAY_WIDTH)
                        col_idx = ci + 1
                        if w > col_char_widths.get(col_idx, 0):
                            col_char_widths[col_idx] = w
            else:
                rows_needed_a = [ra for ra, _rb in row_pairs if ra is not None]
                rows_needed_b = [rb for _ra, rb in row_pairs if rb is not None]
                rows_a_val = _read_rows_into_cache(
                    ws_a,
                    rows_needed_a,
                    max_col,
                    cancel_check=_check_bg_cancel,
                )
                rows_b_val = _read_rows_into_cache(
                    ws_b,
                    rows_needed_b,
                    max_col,
                    cancel_check=_check_bg_cancel,
                )
                rows_a_edit = _read_rows_into_cache(
                    ws_a_edit,
                    rows_needed_a,
                    max_col,
                    cancel_check=_check_bg_cancel,
                )
                rows_b_edit = _read_rows_into_cache(
                    ws_b_edit,
                    rows_needed_b,
                    max_col,
                    cancel_check=_check_bg_cancel,
                )
                base_row_by_pair: dict[int, int | None] = {}
                if ws_base is not None:
                    for idx, (ra, rb) in enumerate(row_pairs):
                        if idx in pair_base_row_override:
                            base_row = pair_base_row_override.get(idx)
                        elif ra is not None and ra in mine_to_base_row:
                            base_row = mine_to_base_row.get(ra)
                        elif rb is not None and rb in theirs_to_base_row:
                            base_row = theirs_to_base_row.get(rb)
                        else:
                            base_row = None
                        base_row_by_pair[idx] = base_row
                    base_rows_needed = [r for r in base_row_by_pair.values() if r is not None]
                    rows_base_val = _read_rows_into_cache(
                        ws_base,
                        base_rows_needed,
                        max_col,
                        cancel_check=_check_bg_cancel,
                    )
                    rows_base_edit = _read_rows_into_cache(
                        ws_base_edit,
                        base_rows_needed,
                        max_col,
                        cancel_check=_check_bg_cancel,
                    ) if ws_base_edit is not None else {}
                else:
                    rows_base_val = {}
                    rows_base_edit = {}
                for idx, (ra, rb) in enumerate(row_pairs):
                    if (idx & 127) == 0:
                        _check_bg_cancel()
                    cols = set()
                    parts_a = []
                    parts_b = []
                    row_a_val = _row_from_cache(rows_a_val, ra, max_col)
                    row_b_val = _row_from_cache(rows_b_val, rb, max_col)
                    row_a_edit = _row_from_cache(rows_a_edit, ra, max_col)
                    row_b_edit = _row_from_cache(rows_b_edit, rb, max_col)
                    for c in range(1, max_col + 1):
                        offset = c - 1
                        va_val = row_a_val[offset]
                        vb_val = row_b_val[offset]
                        va_edit = row_a_edit[offset]
                        vb_edit = row_b_edit[offset]
                        if ra is not None and rb is not None and ra != rb:
                            vb_edit = _translate_normal_formula_for_compare(
                                vb_val,
                                vb_edit,
                                rb,
                                c,
                                ra,
                                c,
                            )
                        da, db, eq = _cell_display_and_equal_from_values(
                            va_val,
                            vb_val,
                            va_edit,
                            vb_edit,
                        )
                        sa = _val_to_str(da)
                        sb = _val_to_str(db)
                        parts_a.append(sa)
                        parts_b.append(sb)
                        w = min(max(len(sa), len(sb)), _COL_MAX_DISPLAY_WIDTH)
                        if w > col_char_widths.get(c, 0):
                            col_char_widths[c] = w
                        if not eq:
                            cols.add(c)
                    if (ra is None) != (rb is None):
                        cols = {-1}
                    pair_parts_a[idx] = parts_a
                    pair_parts_b[idx] = parts_b
                    pair_diff_cols[idx] = cols
                    if ws_base is not None and ra is not None:
                        base_row = base_row_by_pair.get(idx)
                        if base_row is None:
                            pair_base_diff_cols[idx] = {-1}
                        else:
                            row_base_val = _row_from_cache(rows_base_val, base_row, max_col)
                            row_base_edit = _row_from_cache(rows_base_edit, base_row, max_col)
                            base_cols = set()
                            for c in range(1, max_col + 1):
                                offset = c - 1
                                base_edit = row_base_edit[offset]
                                if ra != base_row:
                                    base_edit = _translate_normal_formula_for_compare(
                                        row_base_val[offset],
                                        base_edit,
                                        base_row,
                                        c,
                                        ra,
                                        c,
                                    )
                                _mine_display, _base_display, equal = _cell_display_and_equal_from_values(
                                    row_a_val[offset],
                                    row_base_val[offset],
                                    row_a_edit[offset],
                                    base_edit,
                                )
                                if not equal:
                                    base_cols.add(c)
                            pair_base_diff_cols[idx] = base_cols
                has_diff = any(bool(v) for v in pair_diff_cols.values())
                if (not has_diff) and getattr(self, "has_base", False):
                    has_diff = _sheet_has_base_diff_bg(
                        ws_a,
                        ws_b,
                        ws_base,
                        max_r_a,
                        max_r_b,
                        max_r_base,
                        max_col,
                        signature_cache,
                        trimmed_rows_cache,
                    )

            # Keep a readable lower bound and normalize missing columns.
            for c in range(1, max_col + 1):
                col_char_widths[c] = max(4, int(col_char_widths.get(c, 1)))

            return {
                "sheet": sheet,
                "max_row": max_row,
                "max_col": max_col,
                "col_max_a": max_c_a,
                "col_max_b": max_c_b,
                "row_pairs": row_pairs,
                "pair_diff_cols": pair_diff_cols,
                "pair_base_diff_cols": pair_base_diff_cols,
                "pair_parts_a": pair_parts_a,
                "pair_parts_b": pair_parts_b,
                "col_char_widths": col_char_widths,
                "row_a_to_pair_idx": row_a_to_pair_idx,
                "row_b_to_pair_idx": row_b_to_pair_idx,
                "mine_to_base_row": mine_to_base_row,
                "theirs_to_base_row": theirs_to_base_row,
                "pair_base_row_override": pair_base_row_override,
                "has_diff": has_diff,
            }

        def _apply_sheet_cache(cache: dict):
            sheet = cache["sheet"]
            view = self.sheet_views.get(sheet)
            if view is None:
                # Preserve the exact result for lazy tab creation. The previous
                # implementation discarded it and recomputed the whole Sheet.
                self._sheet_cache_store[sheet] = cache
                self.set_sheet_has_diff(sheet, cache.get("has_diff", False), confirmed=True)
                self.refresh_sheet_nav()
                return
            if getattr(view, "_suppress_bg_apply", False):
                _dlog(f"skip bg cache apply by user action: sheet={sheet}")
                view._hide_loading()
                self.refresh_sheet_nav()
                return
            # Skip if the user has made edits in this view; background data (from read-only copies)
            # would be stale relative to the user's in-memory changes.
            has_user_edits = bool(view.touched_rows)
            has_user_edits = has_user_edits or sheet in getattr(self, "modified_sheets_a", set())
            has_user_edits = has_user_edits or sheet in getattr(self, "modified_sheets_b", set())
            if has_user_edits:
                _dlog(f"skip stale bg cache after user action: sheet={sheet}")
                view._hide_loading()
                self.refresh_sheet_nav()
                return
            # Guard against late background cache downgrading an already rendered sheet to no-diff.
            # This has been observed as a delayed "DiffRows -> 0 / rows disappear" regression.
            try:
                old_diff_count = sum(1 for _k, _v in (view.pair_diff_cols or {}).items() if _v)
            except Exception:
                old_diff_count = 0
            try:
                new_diff_count = sum(1 for _k, _v in (cache.get("pair_diff_cols", {}) or {}).items() if _v)
            except Exception:
                new_diff_count = 0
            if getattr(view, "_data_ready", False) and old_diff_count > 0 and new_diff_count == 0:
                _dlog(f"skip stale cache downgrade: sheet={sheet} old_diff={old_diff_count} new_diff={new_diff_count}")
                view._hide_loading()
                self.refresh_sheet_nav()
                return
            # From this point we will apply this cache to the visible view.
            self.set_sheet_has_diff(sheet, cache.get("has_diff", False), confirmed=True)
            view._invalidate_only_diff_snapshot_cache()
            view.max_row = cache["max_row"]
            view.max_col = cache["max_col"]
            view.col_max_a = max(1, int(cache.get("col_max_a", view.max_col)))
            view.col_max_b = max(1, int(cache.get("col_max_b", view.max_col)))
            view._is_large_sheet = view.max_row >= _LARGE_SHEET_ROW_THRESHOLD
            view._bounds_checked = True

            # Apply row-aligned pair data (computed in background with row alignment)
            view.row_pairs = cache["row_pairs"]
            view.pair_diff_cols = cache["pair_diff_cols"]
            view.pair_base_diff_cols = {
                int(idx): set(cols)
                for idx, cols in (cache.get("pair_base_diff_cols", {}) or {}).items()
            }
            view.col_char_widths = cache.get("col_char_widths", {}) or {c: 4 for c in range(1, view.max_col + 1)}
            # Invalidate cached base spans: column widths replaced by background result.
            view._col_widths_version = int(getattr(view, "_col_widths_version", 0)) + 1
            view.mine_to_base_row = cache.get("mine_to_base_row", {}) or {}
            view.theirs_to_base_row = cache.get("theirs_to_base_row", {}) or {}
            view.pair_base_row_override = cache.get("pair_base_row_override", {}) or {}

            pair_parts_a = cache.get("pair_parts_a", {}) or {}
            pair_parts_b = cache.get("pair_parts_b", {}) or {}
            grid_on = view._is_grid_overlay_enabled()
            sep = _COL_SEP if grid_on else "   "
            trail = " │" if grid_on else ""
            if pair_parts_a or pair_parts_b:
                view.pair_text_a = {}
                view.pair_text_b = {}
                for idx, parts in pair_parts_a.items():
                    view.pair_text_a[idx] = sep.join(
                        _format_cell(parts[i], view.col_char_widths.get(i + 1, 1))
                        for i in range(len(parts))
                    ) + (trail if parts else "")
                for idx, parts in pair_parts_b.items():
                    view.pair_text_b[idx] = sep.join(
                        _format_cell(parts[i], view.col_char_widths.get(i + 1, 1))
                        for i in range(len(parts))
                    ) + (trail if parts else "")
            else:
                # Backward-compatible fallback for older cache shape.
                view.pair_text_a = cache.get("pair_text_a", {})
                view.pair_text_b = cache.get("pair_text_b", {})

            view.row_a_to_pair_idx = cache["row_a_to_pair_idx"]
            view.row_b_to_pair_idx = cache["row_b_to_pair_idx"]
            view._align_rows_enabled = True
            view._diff_partial = False
            # Mark data as ready so refresh(rescan=False) uses it without rescanning
            view._data_ready = True
            view._invalidate_render_cache()
            if not view._is_three_way_enabled() and not view._is_large_sheet:
                view._cache_only_diff_rows_snapshot(
                    idx for idx, cols in view.pair_diff_cols.items() if cols
                )

            requested_only_diff = (
                int(view._pending_only_diff_value)
                if view._pending_only_diff_value is not None
                else int(view.only_diff_var.get())
            )
            needs_exact_only_diff = bool(
                requested_only_diff and not view._has_valid_only_diff_snapshot_cache()
            )
            view._pending_only_diff_value = None
            view._prefer_only_diff_when_ready = False
            # Render a useful full first page while an exact 3-way snapshot is
            # built in the background. Temporarily changing the variable avoids
            # synchronous base-diff expansion inside refresh().
            render_only_diff_value = requested_only_diff
            if needs_exact_only_diff:
                render_only_diff_value = 0
                view._full_render = False
                view._render_limit = min(
                    _LARGE_SHEET_INITIAL_ROWS if view._is_large_sheet else _FAST_RENDER_ROW_LIMIT,
                    view.max_row,
                )
            view.only_diff_var.set(render_only_diff_value)
            # Preserve viewport/cursor when background cache is applied; otherwise
            # user operations (overwrite/resolve) appear to "jump to first row/first column".
            prev_first = 0.0
            prev_x = 0.0
            prev_insert = "1.0"
            try:
                prev_first = float((view.left.yview() or (0.0, 1.0))[0])
                prev_x = float((view.left.xview() or (0.0, 1.0))[0])
                prev_insert = view.left.index("insert")
            except Exception:
                pass
            view.refresh(row_only=None, rescan=False)
            try:
                view.left.yview_moveto(prev_first)
                if view._is_three_way_enabled():
                    view.base.yview_moveto(prev_first)
                view.right.yview_moveto(prev_first)
            except Exception:
                pass
            try:
                view._sync_main_x_to_frac(prev_x)
                view._sync_c_x_to_frac(prev_x)
            except Exception:
                pass
            try:
                line = int(str(prev_insert).split(".")[0])
                col = int(str(prev_insert).split(".")[1])
            except Exception:
                line = 1
                col = 0
            try:
                max_line = max(1, len(view.display_rows))
            except Exception:
                max_line = 1
            if line < 1:
                line = 1
            if line > max_line:
                line = max_line
            try:
                idx = f"{line}.{max(0, col)}"
                view.left.mark_set("insert", idx)
                if view._is_three_way_enabled():
                    view.base.mark_set("insert", idx)
                view.right.mark_set("insert", idx)
            except Exception:
                pass
            view._update_cursor_lines()
            view._hide_loading()
            if needs_exact_only_diff:
                view.only_diff_var.set(requested_only_diff)
                view._start_async_large_only_diff_build()
            self.refresh_sheet_nav()

        def _compute_worker():
            wb_a_ro = None
            wb_b_ro = None
            wb_base_ro = None
            wb_a_e = None
            wb_b_e = None
            wb_base_e = None
            try:
                try:
                    # Use separate read-only workbooks to avoid threading issues
                    wb_a_ro = load_workbook(self._file_a_val_path, data_only=True, read_only=True)
                    wb_b_ro = load_workbook(self._file_b_val_path, data_only=True, read_only=True)
                    if getattr(self, "has_base", False) and getattr(self, "_file_base_val_path", None):
                        wb_base_ro = load_workbook(self._file_base_val_path, data_only=True, read_only=True)
                    wb_a_e = load_workbook(self.file_a, data_only=False, read_only=True)
                    wb_b_e = load_workbook(self.file_b, data_only=False, read_only=True)
                    if getattr(self, "has_base", False) and getattr(self, "base_path", None):
                        wb_base_e = load_workbook(self.base_path, data_only=False, read_only=True)
                except Exception as e:
                    _dlog(f"bg compute open read-only failed: {e}")
                    return
                if wb_a_ro is None or wb_b_ro is None or wb_a_e is None or wb_b_e is None:
                    _dlog("bg compute read-only workbooks not available; skip background compute")
                    return

                while True:
                    if self._is_closing:
                        break
                    with self._compute_lock:
                        if not self._compute_queue:
                            break
                        sheet = self._compute_queue.pop(0)
                        self._compute_inflight.add(sheet)
                        progress_current = min(self._compute_done + 1, self._compute_total)
                        progress_total = self._compute_total
                    _queue_ui_task(
                        lambda s=sheet, cur=progress_current, total=progress_total: self._set_task_status(
                            f"正在加载 Sheet：{s}（{cur}/{total}）",
                            active=True,
                            current=max(0, cur - 1),
                            total=total,
                        )
                    )
                    try:
                        _dlog(f"bg compute sheet: {sheet}")
                        cache = _compute_sheet_cache(
                            wb_a_ro,
                            wb_b_ro,
                            wb_a_e,
                            wb_b_e,
                            sheet,
                            wb_base_ro,
                            wb_base_e,
                        )
                        if self._is_closing:
                            break
                        # Never call tkinter APIs from background threads.
                        _queue_ui_task(lambda c=cache: _apply_sheet_cache(c))
                    except InterruptedError:
                        break
                    except Exception as e:
                        _dlog(f"bg compute failed {sheet}: {e}")
                    finally:
                        with self._compute_lock:
                            self._compute_inflight.discard(sheet)
                            self._compute_done = min(self._compute_total, self._compute_done + 1)
                            progress_done = self._compute_done
                            progress_total = self._compute_total
                            all_done = not self._compute_queue and not self._compute_inflight
                        if all_done:
                            _queue_ui_task(
                                lambda done=progress_done, total=progress_total: self._set_task_status(
                                    f"数据加载完成：已计算 {done}/{total} 个 Sheet",
                                    active=False,
                                    current=done,
                                    total=total,
                                )
                            )
                            # Queue behind the final cache-apply callback. This
                            # keeps editable-workbook parsing from contending
                            # with the first usable render and checkbox actions.
                            _queue_ui_task(self._initial_sheet_ready_event.set)
            finally:
                _wbs_close(wb_a_ro, wb_b_ro, wb_base_ro, wb_a_e, wb_b_e, wb_base_e)
                with self._compute_lock:
                    if self._compute_thread is threading.current_thread():
                        self._compute_thread = None

        def _kick_worker():
            # start a worker if not running
            with self._compute_lock:
                th = self._compute_thread
                if th is not None and th.is_alive():
                    return
                if not self._compute_queue:
                    return
                th = self._start_background_thread(
                    _compute_worker,
                    name="sow-sheet-diff",
                )
                if th is None:
                    return
                self._compute_thread = th
        self._queue_ui_task = _queue_ui_task
        self._enqueue_sheet = _enqueue_sheet
        self._kick_worker = _kick_worker

        # Lazy-create SheetView UI immediately; compute diff in background.
        def _on_tab_changed(_evt=None):
            try:
                tab_id = self.nb.select()
                tab_text = self.nb.tab(tab_id, "text")
                self.selected_sheet = tab_text
                self.refresh_sheet_nav()
                if tab_text in self._sheet_containers and not self._sheet_loaded.get(tab_text, False):
                    _dlog(f"lazy create SheetView (ui only): {tab_text}")
                    view = SheetView(self._sheet_containers[tab_text], self, tab_text)
                    self.sheet_views[tab_text] = view
                    self._sheet_loaded[tab_text] = True
                    cached = self._sheet_cache_store.pop(tab_text, None)
                    if cached is not None:
                        _dlog(f"apply retained sheet cache: {tab_text}")
                        _apply_sheet_cache(cached)
                    else:
                        # The background worker will compute diffs and call
                        # _apply_sheet_cache without blocking the UI thread.
                        view._show_loading()
                if tab_text in self._sheet_containers:
                    # Skip background recompute if data is already ready (no edits pending).
                    # Reopening workbooks on every tab switch is the main perf regression.
                    _view = self.sheet_views.get(tab_text)
                    meta = self.get_sheet_meta(tab_text)
                    if meta.get("view_mode") == "missing_sheet":
                        if _view and (not getattr(_view, "_data_ready", False)):
                            _view.refresh(row_only=None, rescan=True)
                            _view._update_cursor_lines()
                            _view._hide_loading()
                    elif not (_view and getattr(_view, "_data_ready", False)):
                        _enqueue_sheet(tab_text, front=True)
                        _kick_worker()
                        # Watchdog: restart stalled background work, but never fall
                        # back to a synchronous full rescan on the Tk thread.
                        def _force_refresh_if_still_loading(sheet_name=tab_text):
                            try:
                                cur_id = self.nb.select()
                                cur_sheet = self.nb.tab(cur_id, "text")
                                if cur_sheet != sheet_name:
                                    return
                                v = self.sheet_views.get(sheet_name)
                                if not v:
                                    return
                                if getattr(v, "_data_ready", False):
                                    return
                                cached_result = self._sheet_cache_store.pop(sheet_name, None)
                                if cached_result is not None:
                                    _apply_sheet_cache(cached_result)
                                    return
                                with self._compute_lock:
                                    still_background = sheet_name in self._compute_inflight or sheet_name in self._compute_queue
                                with self._ui_task_lock:
                                    pending_ui_result = bool(self._ui_tasks)
                                if not still_background and not pending_ui_result:
                                    _enqueue_sheet(sheet_name, front=True)
                                    _kick_worker()
                                v._show_loading(f"正在后台精确计算 {sheet_name}，界面仍可操作...")
                                self._safe_root_after(1500, _force_refresh_if_still_loading)
                            except Exception as e:
                                _dlog(f"background refresh watchdog failed {sheet_name}: {e}")
                        try:
                            self._safe_root_after(700, _force_refresh_if_still_loading)
                        except Exception:
                            pass
            except Exception as e:
                _dlog(f"tab changed handler failed: {e}")

        try:
            self.nb.bind("<<NotebookTabChanged>>", _on_tab_changed)
        except Exception:
            pass

        # Main-thread UI task pump (for background compute/sample updates).
        try:
            self._safe_root_after(50, _drain_ui_tasks)
        except Exception:
            pass

        # Load the initially selected tab immediately so first-open state is ready.
        _on_tab_changed()

        self.refresh_sheet_nav()

        # Stage-1 tab coloring is provisional. Stage-2 exact background compute
        # changes state 1 (pale) to state 2 (bright) or clears it to state 0.
        def _apply_fast_mark_result(sheet: str, has: bool):
            if has:
                self.set_sheet_has_diff(sheet, True, confirmed=False)
                self.refresh_sheet_nav()

        def _scan_sheet_fingerprint_marks():
            started = time.monotonic()
            try:
                fp_a = _xlsx_sheet_part_fingerprints(self._file_a_val_path)
                fp_b = _xlsx_sheet_part_fingerprints(self._file_b_val_path)
                fp_base = (
                    _xlsx_sheet_part_fingerprints(self._file_base_val_path)
                    if getattr(self, "has_base", False) and getattr(self, "_file_base_val_path", None)
                    else {}
                )
                if not fp_a or not fp_b:
                    _dlog(
                        f"sheet fingerprint premark skipped: "
                        f"a_parts={len(fp_a)} b_parts={len(fp_b)}"
                    )
                    return
                marked = 0
                for sheet in self.compare_sheets:
                    if self._is_closing:
                        return
                    a_sig = fp_a.get(sheet)
                    b_sig = fp_b.get(sheet)
                    if getattr(self, "has_base", False):
                        base_sig = fp_base.get(sheet)
                        if base_sig is not None:
                            has_probable_diff = a_sig != base_sig or b_sig != base_sig
                        else:
                            has_probable_diff = a_sig != b_sig
                    else:
                        has_probable_diff = a_sig != b_sig
                    if has_probable_diff:
                        marked += 1
                        _queue_ui_task(lambda s=sheet: _apply_fast_mark_result(s, True))
                _dlog(
                    f"sheet fingerprint premark done: marked={marked}/{len(self.compare_sheets)} "
                    f"elapsed={time.monotonic() - started:.3f}s"
                )
            except Exception as e:
                _dlog(f"sheet fingerprint premark worker failed: {e}")

        def _sheet_has_diff_fast_tail(ws_a, ws_b, max_row: int, max_col: int, min_row: int = 1):
            none_sig = tuple("" for _ in range(max_col))
            block = _LARGE_SHEET_BLOCK_ROWS
            max_row_a = ws_a.max_row or 1
            max_row_b = ws_b.max_row or 1

            for block_end in range(max_row, 0, -block):
                block_start = max(1, block_end - block + 1)
                if block_end < min_row:
                    break
                if block_start < min_row:
                    block_start = min_row
                end_a = min(block_end, max_row_a)
                end_b = min(block_end, max_row_b)

                rows_a = []
                rows_b = []
                if block_start <= end_a:
                    rows_a = list(ws_a.iter_rows(
                        min_row=block_start,
                        max_row=end_a,
                        min_col=1,
                        max_col=max_col,
                        values_only=True,
                    ))
                if block_start <= end_b:
                    rows_b = list(ws_b.iter_rows(
                        min_row=block_start,
                        max_row=end_b,
                        min_col=1,
                        max_col=max_col,
                        values_only=True,
                    ))

                sig_a = [tuple(_merge_cmp_value(v) for v in (row or ())) for row in rows_a]
                sig_b = [tuple(_merge_cmp_value(v) for v in (row or ())) for row in rows_b]

                for r in range(block_end, block_start - 1, -1):
                    if r <= max_row_a:
                        ia = r - block_start
                        sa = sig_a[ia] if 0 <= ia < len(sig_a) else none_sig
                    else:
                        sa = none_sig
                    if r <= max_row_b:
                        ib = r - block_start
                        sb = sig_b[ib] if 0 <= ib < len(sig_b) else none_sig
                    else:
                        sb = none_sig
                    if sa != sb:
                        return True
            return False

        def _sheet_has_diff_quick_tail(ws_a, ws_b, max_row: int, max_col: int):
            # Phase-1 quick check: scan only the tail window.
            # True means "confirmed diff"; False means "unknown yet".
            quick_rows = min(max_row, _TABMARK_QUICK_TAIL_ROWS)
            if quick_rows <= 0:
                return False
            start = max(1, max_row - quick_rows + 1)
            return _sheet_has_diff_fast_tail(ws_a, ws_b, max_row, max_col, min_row=start)

        def _scan_sheet_has_diff_fast():
            wb_a_ro = None
            wb_b_ro = None
            wb_base_ro = None
            try:
                try:
                    a_sz = os.path.getsize(self._file_a_val_path)
                    b_sz = os.path.getsize(self._file_b_val_path)
                    if max(a_sz, b_sz) >= (_FAST_TABMARK_SCAN_SKIP_MB * 1024 * 1024):
                        _dlog(
                            f"skip fast diff mark scan on large files: "
                            f"a={a_sz} b={b_sz} threshold_mb={_FAST_TABMARK_SCAN_SKIP_MB}"
                        )
                        return
                except Exception:
                    pass

                wb_a_ro = load_workbook(self._file_a_val_path, data_only=True, read_only=True)
                wb_b_ro = load_workbook(self._file_b_val_path, data_only=True, read_only=True)
                if getattr(self, "has_base", False) and getattr(self, "_file_base_val_path", None):
                    try:
                        wb_base_ro = load_workbook(self._file_base_val_path, data_only=True, read_only=True)
                    except Exception:
                        wb_base_ro = None
                ordered = list(self.compare_sheets)
                if ordered:
                    # Prefer currently selected sheet first, then newer tabs first.
                    cur = getattr(self, "selected_sheet", None)
                    if cur in ordered:
                        ordered.remove(cur)
                        ordered = [cur] + list(reversed(ordered))
                    else:
                        ordered = list(reversed(ordered))

                unknown_sheets = []

                # Phase-1: quick tail scan to surface diff tabs early.
                for s in ordered:
                    if self._is_closing:
                        break
                    ws_a = wb_a_ro[s]
                    ws_b = wb_b_ro[s]
                    max_row = max(ws_a.max_row or 1, ws_b.max_row or 1)
                    max_col = max(ws_a.max_column or 1, ws_b.max_column or 1)
                    # Read-only tail iteration still reparses XML from row 1. On a
                    # large sheet this optional pre-mark duplicates the exact
                    # background compute and can outlive the window during close.
                    if max_row >= _LARGE_SHEET_ROW_THRESHOLD:
                        _dlog(f"skip fast tab mark for large sheet: {s} rows={max_row}")
                        continue
                    has_quick = _sheet_has_diff_quick_tail(ws_a, ws_b, max_row, max_col)
                    if (not has_quick) and getattr(self, "has_base", False):
                        if wb_base_ro is not None and s not in wb_base_ro.sheetnames:
                            has_quick = True
                    if has_quick:
                        _queue_ui_task(lambda s=s: _apply_fast_mark_result(s, True))
                    else:
                        unknown_sheets.append((s, ws_a, ws_b, max_row, max_col))

                # Phase-2 exact scan is optional: background compute will always confirm.
                if _FAST_TABMARK_PHASE2_ENABLED:
                    # Re-fetch worksheet objects: read_only iterators from Phase-1 are consumed.
                    for s, _stale_a, _stale_b, max_row, max_col in unknown_sheets:
                        if self._is_closing:
                            break
                        ws_a = wb_a_ro[s]
                        ws_b = wb_b_ro[s]
                        has = _sheet_has_diff_fast_tail(ws_a, ws_b, max_row, max_col)
                        _queue_ui_task(lambda s=s, has=has: _apply_fast_mark_result(s, has))
            except Exception as e:
                _dlog(f"fast diff mark scan failed: {e}")
            finally:
                _wbs_close(wb_a_ro, wb_b_ro, wb_base_ro)

        if _FAST_TABMARK_ENABLED:
            try:
                self._fast_tabmark_thread = self._start_background_thread(
                    _scan_sheet_fingerprint_marks,
                    name="sow-sheet-fingerprint",
                )
            except Exception:
                pass
        else:
            self._fast_tabmark_thread = None

        # Enqueue all sheets for background confirmation (slow compute)
        try:
            for s in self.compare_sheets:
                _enqueue_sheet(s, front=False)
            if self.compare_sheets:
                _kick_worker()
            else:
                self._set_task_status("数据加载完成：没有需要逐行计算的同名 Sheet", active=False)
        except Exception as e:
            _dlog(f"enqueue all sheets failed: {e}")

    def push_undo(self, action: dict):
        try:
            self.undo_stack.append(action)
            if len(self.undo_stack) > 20:
                self.undo_stack.pop(0)
        except Exception:
            pass

    def pop_undo(self) -> dict | None:
        try:
            if not self.undo_stack:
                return None
            return self.undo_stack.pop()
        except Exception:
            return None

    def _add_missing_tab(self, title: str, items):
        frame = ttk.Frame(self.nb)
        self.nb.add(frame, text=title)
        ttk.Label(frame, text=title, font=("Segoe UI", 11, "bold")).pack(anchor="w", padx=8, pady=(8, 4))
        txt = tk.Text(frame, wrap="none", height=10)
        txt.pack(fill="both", expand=True, padx=8, pady=8)
        txt.insert("1.0", "\n".join(items))
        txt.configure(state="disabled")

    def _select_tab(self, tab_text: str):
        for tab_id in self.nb.tabs():
            if self.nb.tab(tab_id, "text") == tab_text:
                self.nb.select(tab_id)
                return

    def refresh_sheet_nav(self):
        # Skip the (expensive) destroy/rebuild of all sheet buttons when nothing
        # that affects them has changed. refresh() calls this on every full
        # render, so for large sheets this avoids rebuilding the whole nav bar
        # on edits/scroll-driven refreshes where the sheet set is unchanged.
        try:
            nav_sig = (
                tuple(self.display_sheets),
                tuple((s, self.get_sheet_meta(s).get("view_mode")) for s in self.display_sheets),
                tuple((s, int(self.sheet_diff_state.get(s, 0))) for s in self.display_sheets),
                getattr(self, "selected_sheet", None),
            )
        except Exception:
            nav_sig = None
        if nav_sig is not None and nav_sig == getattr(self, "_nav_sig", None) \
                and self.nav_inner.winfo_children():
            return
        self._nav_sig = nav_sig

        for child in list(self.nav_inner.winfo_children()):
            child.destroy()

        try:
            from tkinter import font as tkfont
            if not hasattr(self, "_nav_font"):
                self._nav_font = tkfont.nametofont("TkDefaultFont")
                self._nav_font_bold = self._nav_font.copy()
                self._nav_font_bold.configure(weight="bold")
        except Exception:
            self._nav_font = None
            self._nav_font_bold = None

        def add_btn(label: str, tab_text: str, kind: str, state: int = 0):
            if kind == "missing":
                bg = "#FFE5E5"
            else:
                # 0=none, 1=maybe (pale), 2=confirmed (bright)
                if state >= 2:
                    bg = "#FFD400"  # bright yellow
                elif state == 1:
                    bg = "#FFF3B0"  # pale yellow
                else:
                    bg = "#F2F2F2"
            is_selected = (tab_text == getattr(self, "selected_sheet", None))
            if is_selected:
                bg = "#D9D9D9"
            b = tk.Button(self.nav_inner, text=label,
                          relief="sunken" if is_selected else "groove",
                          bd=2 if is_selected else 1,
                          padx=8, pady=2, bg=bg,
                          command=lambda: self._select_tab(tab_text))
            try:
                if is_selected and self._nav_font_bold:
                    b.configure(font=self._nav_font_bold)
                elif self._nav_font:
                    b.configure(font=self._nav_font)
            except Exception:
                pass
            b.pack(side="left", padx=4)

        for s in self.display_sheets:
            meta = self.get_sheet_meta(s)
            kind = "missing" if meta.get("view_mode") == "missing_sheet" else "common"
            add_btn(s, s, kind, state=int(self.sheet_diff_state.get(s, 0)))

        self.nav_canvas.update_idletasks()
        self.nav_canvas.configure(scrollregion=self.nav_canvas.bbox("all"))

    def open_textdiff(self):
        try:
            temp_root = os.path.join(os.environ.get("LOCALAPPDATA", tempfile.gettempdir()), "Temp", "TortoiseXlsTemp")
            os.makedirs(temp_root, exist_ok=True)
        except Exception:
            temp_root = tempfile.gettempdir()

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        left_txt = os.path.join(temp_root, f"{APP_NAME}_left_{ts}.txt")
        right_txt = os.path.join(temp_root, f"{APP_NAME}_right_{ts}.txt")
        excel_to_text(self.file_a, left_txt, thick_sep_char="=")
        excel_to_text(self.file_b, right_txt, thick_sep_char="=")
        open_tortoise_merge(left_txt, right_txt, title=f"{APP_NAME}: {os.path.basename(self.file_a)}")

    def export_diagnostic_bundle(self):
        try:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_name = f"{APP_NAME}_diag_{APP_BUILD_TAG}_{ts}.zip"
            desktop = os.path.join(os.path.expanduser("~"), "Desktop")
            initial_dir = desktop if os.path.isdir(desktop) else tempfile.gettempdir()
            save_path = filedialog.asksaveasfilename(
                title="导出诊断包",
                defaultextension=".zip",
                initialdir=initial_dir,
                initialfile=default_name,
                filetypes=[("Zip Archive", "*.zip")],
            )
            if not save_path:
                return

            notes = []
            notes.append(f"app={APP_NAME}")
            notes.append(f"version={APP_VERSION}")
            notes.append(f"build={APP_BUILD_TAG}")
            notes.append(f"time={datetime.now().isoformat(timespec='seconds')}")
            notes.append(f"python={sys.version.splitlines()[0]}")
            notes.append(f"platform={platform.platform()}")
            notes.append(f"merge_mode={self.merge_mode}")
            notes.append(f"merge_conflict_mode={self.merge_conflict_mode}")
            notes.append(f"file_a={self.file_a}")
            notes.append(f"file_b={self.file_b}")
            notes.append(f"base_path={self.base_path}")
            notes.append(f"raw_mine={self.raw_mine}")
            notes.append(f"raw_base={self.raw_base}")
            notes.append(f"raw_theirs={self.raw_theirs}")

            with zipfile.ZipFile(save_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
                zf.writestr("diagnostic_summary.txt", "\n".join(notes) + "\n")
                for p in (_DEBUG_LOG_PATH, _LAUNCH_TRACE_PATH, _SETTINGS_PATH):
                    try:
                        if p and os.path.exists(p):
                            zf.write(p, arcname=os.path.basename(p))
                    except Exception:
                        pass

            messagebox.showinfo("导出完成", f"诊断包已导出：\n{save_path}")
        except Exception as e:
            messagebox.showerror("导出失败", f"导出诊断包失败：\n{e}")

    def copy_feedback_info(self):
        try:
            selected_sheet = "-"
            try:
                tab_id = self.nb.select()
                if tab_id:
                    selected_sheet = self.nb.tab(tab_id, "text")
            except Exception:
                selected_sheet = "-"

            lines = [
                f"app={APP_NAME}",
                f"version={APP_VERSION}",
                f"build={APP_BUILD_TAG}",
                f"time={datetime.now().isoformat(timespec='seconds')}",
                f"merge_mode={self.merge_mode}",
                f"merge_conflict_mode={self.merge_conflict_mode}",
                f"selected_sheet={selected_sheet}",
                f"file_a={self.file_a}",
                f"file_b={self.file_b}",
                f"base_path={self.base_path}",
                f"raw_mine={self.raw_mine}",
                f"raw_base={self.raw_base}",
                f"raw_theirs={self.raw_theirs}",
                f"debug_log={_DEBUG_LOG_PATH}",
                f"launch_trace={_LAUNCH_TRACE_PATH}",
            ]
            text = "\n".join(lines)
            self.root.clipboard_clear()
            self.root.clipboard_append(text)
            self.root.update()
            messagebox.showinfo("已复制", "反馈信息已复制到剪贴板。")
        except Exception as e:
            messagebox.showerror("复制失败", f"复制反馈信息失败：\n{e}")

    def _find_svn_wc_root(self, start_dir: str) -> str | None:
        try:
            cur = os.path.abspath(start_dir)
            while True:
                if os.path.isdir(os.path.join(cur, ".svn")):
                    return cur
                parent = os.path.dirname(cur)
                if parent == cur:
                    return None
                cur = parent
        except Exception:
            return None

    def _ask_update_confirm_dialog(self, wc_root: str) -> bool:
        dlg = tk.Toplevel(self.root)
        dlg.title("检查更新")
        dlg.transient(self.root)
        dlg.resizable(False, False)

        result = {"ok": False}

        frm = ttk.Frame(dlg, padding=12)
        frm.pack(fill="both", expand=True)

        ttk.Label(frm, text=f"当前版本：{APP_VERSION} [{APP_BUILD_TAG}]").pack(anchor="w")
        ttk.Label(
            frm,
            text=(
                "将从 SVN 获取最新版本。\n"
                "确认后工具将自动关闭，更新完成后请重新启动。"
            ),
            justify="left",
            wraplength=620,
        ).pack(anchor="w", pady=(6, 0))

        tk.Label(
            frm,
            text="注意：更新流程不会保存当前工具中的未保存修改。",
            fg="#C62828",
            justify="left",
            anchor="w",
        ).pack(anchor="w", pady=(10, 0))
        tk.Label(
            frm,
            text="请先保存再更新，避免数据丢失。",
            fg="#C62828",
            justify="left",
            anchor="w",
        ).pack(anchor="w")

        ttk.Label(
            frm,
            text=f"工作副本目录：\n{wc_root}",
            justify="left",
            wraplength=620,
        ).pack(anchor="w", pady=(10, 0))

        btns = ttk.Frame(frm)
        btns.pack(fill="x", pady=(12, 0))

        def _cancel():
            result["ok"] = False
            try:
                dlg.destroy()
            except Exception:
                pass

        def _confirm():
            result["ok"] = True
            try:
                dlg.destroy()
            except Exception:
                pass

        ttk.Button(btns, text="取消", command=_cancel).pack(side="right")
        ttk.Button(btns, text="确认更新", command=_confirm).pack(side="right", padx=(0, 8))

        dlg.protocol("WM_DELETE_WINDOW", _cancel)
        try:
            dlg.update_idletasks()
            rx = self.root.winfo_rootx()
            ry = self.root.winfo_rooty()
            rw = self.root.winfo_width()
            rh = self.root.winfo_height()
            dw = dlg.winfo_width()
            dh = dlg.winfo_height()
            x = max(0, rx + (rw - dw) // 2)
            y = max(0, ry + (rh - dh) // 2)
            dlg.geometry(f"+{x}+{y}")
        except Exception:
            pass

        dlg.grab_set()
        try:
            dlg.focus_force()
        except Exception:
            pass
        dlg.wait_window()
        return bool(result.get("ok", False))

    def _do_svn_update(self):
        if getattr(self, "_update_launching", False):
            return

        def _reenable_update_btn():
            try:
                if getattr(self, "update_btn", None) is not None:
                    self.update_btn.configure(state="normal")
            except Exception:
                pass
            self._update_launching = False

        # Protect users from silently losing in-memory changes.
        if bool(getattr(self, "modified_a", False) or getattr(self, "modified_b", False)):
            if not messagebox.askyesno(
                "未保存改动",
                "检测到当前有未保存改动。\n"
                "如果继续更新并关闭工具，这些改动将丢失。\n\n"
                "仍然继续更新吗？",
            ):
                return

        # Locate tool directory: PyInstaller bundle uses sys.executable, plain script uses __file__.
        if getattr(sys, "frozen", False):
            tool_dir = os.path.dirname(sys.executable)
        else:
            tool_dir = os.path.dirname(os.path.abspath(__file__))
        tool_dir = os.path.abspath(tool_dir)

        wc_root = self._find_svn_wc_root(tool_dir)
        if not wc_root:
            messagebox.showerror(
                "更新失败",
                "当前工具目录不在 SVN 工作副本中。\n"
                "请确认工具是从 SVN checkout 目录启动。",
            )
            return
        # Keep update scope limited to the tool directory.
        update_target = tool_dir
        proc_exe = _find_tortoise_proc_exe()
        proc_exists = bool(proc_exe and os.path.exists(proc_exe))
        svn_exe = _find_svn_cli_exe()
        if not svn_exe and proc_exists:
            if not self._ask_update_confirm_dialog(update_target):
                return
            if not messagebox.askyesno(
                "更新方式",
                "未找到 svn 命令行，将改用 TortoiseSVN 图形更新。\n"
                "确认后工具将关闭，并弹出 TortoiseSVN Update 窗口。\n\n"
                "是否继续？",
            ):
                return
            self._update_launching = True
            try:
                if getattr(self, "update_btn", None) is not None:
                    self.update_btn.configure(state="disabled")
            except Exception:
                pass
            try:
                subprocess.Popen([
                    proc_exe,
                    "/command:update",
                    f"/path:{update_target}",
                    "/closeonend:0",
                ], close_fds=True)
            except Exception as e:
                _reenable_update_btn()
                messagebox.showerror("更新失败", f"无法启动 TortoiseSVN 更新：\n{e}")
                return
            try:
                self._shutdown_root()
            except Exception:
                pass
            sys.exit(0)

        if not svn_exe:
            if not proc_exists:
                messagebox.showerror(
                    "更新失败",
                    "未找到 svn 命令，且未检测到 TortoiseSVN 安装目录。\n"
                    "请安装 TortoiseSVN（含命令行工具）或把 svn.exe 加入 PATH。",
                )
            else:
                messagebox.showerror(
                    "更新失败",
                    "未找到 svn 命令。\n"
                    f"已检测到 TortoiseSVN：{proc_exe}\n"
                    "请确认存在 svn.exe（通常在 TortoiseSVN\\bin）或将其加入 PATH。",
                )
            return

        svn_ver = _query_svn_version(svn_exe)
        if not svn_ver:
            messagebox.showerror(
                "更新失败",
                "检测到 svn 可执行文件，但无法获取版本信息。\n"
                f"路径：{svn_exe}\n"
                "请检查 SVN 安装是否完整。",
            )
            return
        _dlog(f"svn detected: exe={svn_exe} version={svn_ver}")

        ps_exe = shutil.which("powershell") or shutil.which("pwsh")
        if not ps_exe:
            messagebox.showerror("更新失败", "未找到 PowerShell，无法启动更新脚本。")
            return

        no_window = getattr(subprocess, "CREATE_NO_WINDOW", 0)
        try:
            info = subprocess.run(
                [svn_exe, "info", update_target],
                capture_output=True,
                text=True,
                errors="replace",
                timeout=12,
                creationflags=no_window,
            )
        except Exception as e:
            messagebox.showerror("更新失败", f"执行 svn info 失败：\n{e}")
            return
        if info.returncode != 0:
            err = (info.stderr or info.stdout or "").strip()
            messagebox.showerror("更新失败", f"当前目录不是有效 SVN 工作副本或无权限：\n{err}")
            return

        # Ask SVN for canonical wc-root (more robust than .svn upward walk).
        try:
            wc_info = subprocess.run(
                [svn_exe, "info", "--show-item", "wc-root", wc_root],
                capture_output=True,
                text=True,
                errors="replace",
                timeout=10,
                creationflags=no_window,
            )
            if wc_info.returncode == 0:
                wc_root_svn = (wc_info.stdout or "").strip()
                if wc_root_svn:
                    wc_root = os.path.abspath(wc_root_svn)
        except Exception as e:
            _dlog(f"resolve wc-root from svn info failed: {e}")

        # Safety check: tool path must still reside under detected working copy.
        try:
            if os.path.normcase(os.path.commonpath([tool_dir, wc_root])) != os.path.normcase(wc_root):
                messagebox.showerror(
                    "更新失败",
                    "检测到工作副本路径异常。\n"
                    "为避免误更新，已中止本次操作。\n\n"
                    f"工具目录：{tool_dir}\n"
                    f"工作副本：{wc_root}",
                )
                return
        except Exception as e:
            messagebox.showerror("更新失败", f"无法校验工作副本路径关系：\n{e}")
            return

        # Preflight status: block known-bad states and warn risky states.
        # Performance optimization:
        # - First check the tool directory only (most relevant for self-update).
        # - Fallback to wc_root only if the narrow-scope status call fails.
        local_changes = []
        conflict_changes = []
        status_error = None
        status_scope_label = "工具目录"
        status_target = update_target
        try:
            status = subprocess.run(
                [svn_exe, "status", "-q", status_target],
                capture_output=True,
                text=True,
                errors="replace",
                timeout=20,
                creationflags=no_window,
            )
            if status.returncode != 0:
                # Fallback: if status on tool dir failed unexpectedly, retry on wc-root.
                status_scope_label = "工作副本"
                status_target = wc_root
                status = subprocess.run(
                    [svn_exe, "status", "-q", status_target],
                    capture_output=True,
                    text=True,
                    errors="replace",
                    timeout=20,
                    creationflags=no_window,
                )
            if status.returncode == 0:
                for ln in (status.stdout or "").splitlines():
                    if not ln.strip():
                        continue
                    flags = (ln[:7] + "       ")[:7]
                    path_part = ln[8:].strip() if len(ln) > 8 else ln.strip()
                    if "C" in flags:
                        conflict_changes.append(path_part)
                        continue
                    # Ignore unversioned-only lines; these do not block update.
                    if flags[0] == "?":
                        continue
                    # Versioned local mods/add/delete/replace/tree-conflict-ish states.
                    if any(ch in "MADR~!" for ch in flags[:2]):
                        local_changes.append(path_part)
            else:
                status_error = (status.stderr or status.stdout or "").strip() or f"returncode={status.returncode}"
        except Exception as e:
            status_error = str(e)

        if status_error:
            messagebox.showerror(
                "更新前检查失败",
                f"无法完成{status_scope_label}状态检查（svn status）。\n"
                "请先在工作副本中手工执行 svn status 并处理异常后重试。\n\n"
                f"{status_error}",
            )
            return

        if conflict_changes:
            preview = "\n".join(conflict_changes[:6])
            if len(conflict_changes) > 6:
                preview += f"\n... 另有 {len(conflict_changes) - 6} 项"
            messagebox.showerror(
                "更新前检查失败",
                f"检测到{status_scope_label}存在冲突项，请先解决冲突后再更新。\n\n" + preview,
            )
            return

        if local_changes:
            preview = "\n".join(local_changes[:6])
            if len(local_changes) > 6:
                preview += f"\n... 另有 {len(local_changes) - 6} 项"
            if not messagebox.askyesno(
                "检测到本地改动",
                f"{status_scope_label}包含本地版本化改动，继续 update 可能产生冲突。\n\n"
                f"{preview}\n\n仍然继续更新吗？",
            ):
                return

        if not self._ask_update_confirm_dialog(update_target):
            return

        self._update_launching = True
        try:
            if getattr(self, "update_btn", None) is not None:
                self.update_btn.configure(state="disabled")
        except Exception:
            pass

        log_path = os.path.join(update_target, "sow_update.log")
        ps1_path = os.path.join(tempfile.gettempdir(), f"sow_update_{os.getpid()}.ps1")

        update_target_q = update_target.replace("'", "''")
        svn_exe_q = svn_exe.replace("'", "''")
        log_path_q = log_path.replace("'", "''")

        ps_script = (
            "$ErrorActionPreference = 'Continue'\n"
            f"$updateTarget = '{update_target_q}'\n"
            f"$svnExe = '{svn_exe_q}'\n"
            f"$logPath = '{log_path_q}'\n"
            "$maxTry = 3\n"
            "Start-Sleep -Seconds 2\n"
            "if (-not (Test-Path -LiteralPath $updateTarget)) { Write-Host '更新目录不存在，更新终止。'; exit 2 }\n"
            "Set-Location -LiteralPath $updateTarget\n"
            "Add-Content -LiteralPath $logPath -Value (\"===== {0} update start =====\" -f (Get-Date -Format 'yyyy-MM-dd HH:mm:ss'))\n"
            "$ok = $false\n"
            "$lastCode = 0\n"
            "for ($i = 1; $i -le $maxTry; $i++) {\n"
            "  Add-Content -LiteralPath $logPath -Value (\"[try {0}] svn cleanup\" -f $i)\n"
            "  & $svnExe cleanup --non-interactive \"$updateTarget\" >> $logPath 2>&1\n"
            "  Add-Content -LiteralPath $logPath -Value (\"[try {0}] svn update\" -f $i)\n"
            "  & $svnExe update --non-interactive \"$updateTarget\" >> $logPath 2>&1\n"
            "  $lastCode = $LASTEXITCODE\n"
            "  if ($lastCode -eq 0) { $ok = $true; break }\n"
            "  Start-Sleep -Seconds 2\n"
            "}\n"
            "if ($ok) {\n"
            "  Add-Content -LiteralPath $logPath -Value 'update succeeded'\n"
            "  Write-Host ''\n"
            "  Write-Host '更新成功，请重新启动工具。'\n"
            "} else {\n"
            "  Add-Content -LiteralPath $logPath -Value (\"update failed rc={0}\" -f $lastCode)\n"
            "  Write-Host ''\n"
            "  Write-Host '更新失败，请查看日志：'\n"
            "  Write-Host $logPath\n"
            "}\n"
            "Write-Host ''\n"
            "Write-Host '按回车关闭窗口...'\n"
            "[void][System.Console]::ReadLine()\n"
            "Remove-Item -LiteralPath $PSCommandPath -Force -ErrorAction SilentlyContinue\n"
        )

        try:
            with open(ps1_path, "w", encoding="utf-8-sig") as f:
                f.write(ps_script)
            new_console = getattr(subprocess, "CREATE_NEW_CONSOLE", 0)
            subprocess.Popen(
                [ps_exe, "-NoProfile", "-ExecutionPolicy", "Bypass", "-File", ps1_path],
                creationflags=new_console,
                close_fds=True,
            )
        except Exception as e:
            try:
                os.remove(ps1_path)
            except Exception:
                pass
            _reenable_update_btn()
            messagebox.showerror("更新失败", f"无法启动更新脚本：\n{e}")
            return

        try:
            self._shutdown_root()
        except Exception:
            pass
        sys.exit(0)

    def recalc_and_refresh(self):
        # Manual: force Excel recalc to refresh cached values, then reload view.
        if self.modified_a or self.modified_b:
            messagebox.showwarning(
                "存在未保存操作",
                "当前存在尚未保存的覆盖操作。此时重算会基于磁盘旧文件，可能覆盖刚采用的结果。\n\n"
                "为保护合并结果，本次重算已取消。请先完成并保存当前合并。",
            )
            return

        def _do_recalc():
            new_a = _maybe_recalc_and_prepare_val_path(self.file_a, force=True)
            new_b = _maybe_recalc_and_prepare_val_path(self.file_b, force=True)
            new_base = _maybe_recalc_and_prepare_val_path(self.base_path, force=True) if getattr(self, "has_base", False) else None
            self._apply_recalc_results(new_a=new_a, new_b=new_b, new_base=new_base)

        try:
            self._with_progress("重算中", "正在重算并刷新，请稍候...", _do_recalc)
        except Exception as e:
            messagebox.showerror("重算失败", f"重算失败：\n{e}")

    def _schedule_auto_recalc(self):
        if not (_AUTO_RECALC_ON_OPEN and _USE_CACHED_VALUES_ONLY):
            return
        if self.merge_mode:
            _dlog("auto recalc skipped in merge mode")
            return
        if self._auto_recalc_started:
            return
        self._auto_recalc_started = True

        def _worker():
            try:
                new_a = _maybe_recalc_and_prepare_val_path(self.file_a)
                new_b = _maybe_recalc_and_prepare_val_path(self.file_b)
                new_base = _maybe_recalc_and_prepare_val_path(self.base_path) if getattr(self, "has_base", False) else None
            except Exception:
                new_a = None
                new_b = None
                new_base = None

            if not (new_a or new_b or new_base):
                return

            def _apply():
                try:
                    self._apply_recalc_results(
                        new_a=new_a,
                        new_b=new_b,
                        new_base=new_base,
                        respect_user_edits=True,
                    )
                except Exception as e:
                    _dlog(f"auto recalc apply failed: {e}")

            try:
                self._safe_root_after(0, _apply)
            except Exception:
                pass

        self._auto_recalc_thread = self._start_background_thread(
            _worker,
            name="sow-auto-recalc",
        )


    def _with_progress(
        self,
        title: str,
        message: str,
        fn,
        *,
        run_in_background: bool = False,
        pass_reporter: bool = False,
    ):
        dlg = tk.Toplevel(self.root)
        dlg.title(title)
        dlg.transient(self.root)
        dlg.grab_set()
        dlg.resizable(False, False)
        dlg.protocol("WM_DELETE_WINDOW", lambda: None)
        dlg.geometry("+{}+{}".format(self.root.winfo_rootx() + 200, self.root.winfo_rooty() + 150))
        message_label = ttk.Label(dlg, text=message, padding=(12, 12, 12, 4), wraplength=520)
        message_label.pack(fill="x")
        detail_label = ttk.Label(
            dlg,
            text="正在准备...",
            foreground="#555",
            padding=(12, 0, 12, 8),
            wraplength=520,
        )
        detail_label.pack(fill="x")
        pb = ttk.Progressbar(dlg, mode="indeterminate")
        pb.pack(fill="x", padx=12, pady=(0, 6))
        elapsed_label = ttk.Label(dlg, text="已用时 0.0 秒", foreground="#777", padding=(12, 0, 12, 10))
        elapsed_label.pack(anchor="e")
        pb.start(12)
        self.root.update_idletasks()

        if not run_in_background:
            try:
                return fn()
            finally:
                try:
                    pb.stop()
                    dlg.destroy()
                except Exception:
                    pass

        updates_lock = threading.Lock()
        updates: list[tuple[str | None, str | None, float | None]] = []
        state: dict[str, object] = {}
        done = threading.Event()
        started = time.monotonic()

        def report(text=None, detail=None, percent=None):
            pct = None
            if percent is not None:
                try:
                    pct = max(0.0, min(100.0, float(percent)))
                except Exception:
                    pct = None
            with updates_lock:
                updates.append((text, detail, pct))

        def _worker():
            try:
                if pass_reporter:
                    state["result"] = fn(report)
                else:
                    state["result"] = fn()
            except BaseException as exc:
                state["error"] = exc
                state["traceback"] = traceback.format_exc()
            finally:
                done.set()

        gc_was_enabled = gc.isenabled()
        if gc_was_enabled:
            gc.disable()
        thread = self._start_background_thread(_worker, name="sow-progress-task")
        if thread is None:
            if gc_was_enabled:
                gc.enable()
                gc.collect()
            try:
                dlg.destroy()
            except Exception:
                pass
            raise RuntimeError("任务未能启动：应用正在关闭")

        def _poll():
            latest = None
            with updates_lock:
                if updates:
                    latest = updates[-1]
                    updates.clear()
            if latest is not None:
                text_value, detail_value, percent_value = latest
                if text_value:
                    message_label.configure(text=str(text_value))
                if detail_value is not None:
                    detail_label.configure(text=str(detail_value))
                if percent_value is not None:
                    pb.stop()
                    pb.configure(mode="determinate", maximum=100, value=percent_value)
            elapsed_label.configure(text=f"已用时 {time.monotonic() - started:.1f} 秒")
            if done.is_set():
                try:
                    pb.stop()
                    dlg.destroy()
                except Exception:
                    pass
                return
            try:
                dlg.after(80, _poll)
            except Exception:
                pass

        dlg.after(50, _poll)
        try:
            dlg.wait_window()
        finally:
            try:
                pb.stop()
            except Exception:
                pass
            if gc_was_enabled:
                gc.enable()
                gc.collect()
        if "error" in state:
            _dlog(f"progress task failed: {state.get('traceback', state['error'])}")
            raise state["error"]
        return state.get("result")

    def _atomic_save(self, wb, target_path: str):
        """Safely overwrite a workbook.

        Writes to a temp file in the same directory, then os.replace.
        This avoids corrupting the target if the process is interrupted.
        """
        ext_parts = _capture_external_link_parts(target_path)
        # If we need to preserve external-link parts, force tmp-path save to avoid
        # a second replace on the final target (often locked by SVN shell after save).
        use_fast_direct = _FAST_SAVE_ENABLED and not ext_parts
        if use_fast_direct:
            # Fast path: write directly (faster, but not atomic)
            try:
                if _FAST_SAVE_VALUES_ONLY and _USE_CACHED_VALUES_ONLY:
                    _save_values_only_from_wb(wb, target_path)
                else:
                    wb.save(target_path)
                return
            except PermissionError:
                # fallback to safe path below
                pass
        folder = os.path.dirname(target_path)
        base = os.path.basename(target_path)
        tmp_path = os.path.join(folder, f"~{base}.{os.getpid()}.tmp")
        if _FAST_SAVE_VALUES_ONLY and _USE_CACHED_VALUES_ONLY:
            _save_values_only_from_wb(wb, tmp_path)
        else:
            wb.save(tmp_path)
        if ext_parts:
            _apply_external_link_parts_on_file(tmp_path, ext_parts)
        try:
            os.replace(tmp_path, target_path)
            return
        except PermissionError:
            # Try clearing readonly flag then retry a few times (file may be locked briefly)
            try:
                if os.path.exists(target_path):
                    os.chmod(target_path, stat.S_IWRITE | stat.S_IREAD)
            except Exception:
                pass
            # If we can delete the target, try that once (replace may fail on readonly)
            try:
                if os.path.exists(target_path):
                    os.remove(target_path)
            except Exception:
                pass
            for _ in range(8):
                try:
                    os.replace(tmp_path, target_path)
                    return
                except PermissionError:
                    time.sleep(0.5)
            # If replace keeps failing, try overwrite-in-place (requires write but not delete)
            try:
                with open(tmp_path, "rb") as src, open(target_path, "wb") as dst:
                    shutil.copyfileobj(src, dst, length=1024 * 1024)
                return
            except Exception:
                raise
        except Exception:
            # Last-resort fallback to shutil.move
            try:
                shutil.move(tmp_path, target_path)
                return
            except Exception:
                raise
        finally:
            try:
                if os.path.exists(tmp_path):
                    os.remove(tmp_path)
            except Exception:
                pass

    def _atomic_save_with_retry(self, wb, target_path: str, retries: int = 6, delay_sec: float = 0.5):
        """Retry save when target is temporarily locked."""
        last_err = None
        for _ in range(max(1, retries)):
            try:
                self._atomic_save(wb, target_path)
                return
            except Exception as e:
                if getattr(e, "winerror", None) in (5, 32, 33) or isinstance(e, PermissionError):
                    last_err = e
                    time.sleep(delay_sec)
                    continue
                raise
        if last_err:
            raise last_err

    def _atomic_replace_file(self, src_path: str, target_path: str):
        """Safely replace target with an existing source file (no openpyxl roundtrip)."""
        folder = os.path.dirname(target_path)
        base = os.path.basename(target_path)
        tmp_path = os.path.join(folder, f"~{base}.{os.getpid()}.tmp")
        shutil.copy2(src_path, tmp_path)
        try:
            os.replace(tmp_path, target_path)
            return
        except PermissionError:
            try:
                if os.path.exists(target_path):
                    os.chmod(target_path, stat.S_IWRITE | stat.S_IREAD)
            except Exception:
                pass
            try:
                if os.path.exists(target_path):
                    os.remove(target_path)
            except Exception:
                pass
            for _ in range(8):
                try:
                    os.replace(tmp_path, target_path)
                    return
                except PermissionError:
                    time.sleep(0.5)
            with open(tmp_path, "rb") as src, open(target_path, "wb") as dst:
                shutil.copyfileobj(src, dst, length=1024 * 1024)
        finally:
            try:
                if os.path.exists(tmp_path):
                    os.remove(tmp_path)
            except Exception:
                pass

    def _atomic_replace_file_with_retry(self, src_path: str, target_path: str, retries: int = 6, delay_sec: float = 0.5):
        last_err = None
        for attempt in range(max(1, retries)):
            try:
                self._atomic_replace_file(src_path, target_path)
                return
            except Exception as e:
                if getattr(e, "winerror", None) in (5, 32, 33) or isinstance(e, PermissionError):
                    last_err = e
                    time.sleep(delay_sec * (2 ** min(attempt, 4)))
                    continue
                raise
        if last_err:
            raise last_err

    def _alt_save_path(self, path: str, which: str):
        folder = os.path.dirname(path)
        base, ext = os.path.splitext(os.path.basename(path))
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        return os.path.join(folder, f"{base}_{which}_saved_{ts}{ext or '.xlsx'}")

    def _try_alt_save(self, wb, path: str, which: str) -> bool:
        alt = self._alt_save_path(path, which)
        try:
            self._atomic_save(wb, alt)
            messagebox.showinfo("另存为成功", f"无法覆盖原文件，已另存为：\n{alt}")
            return True
        except Exception as e:
            messagebox.showerror("另存为失败", f"另存为失败：\n{e}")
            return False

    def _path_diagnostics(self, path: str) -> str:
        try:
            folder = os.path.dirname(path)
            exists = os.path.exists(path)
            readonly = False
            if exists:
                try:
                    readonly = not os.access(path, os.W_OK)
                except Exception:
                    readonly = False
            dir_writable = False
            test_file = None
            try:
                test_file = os.path.join(folder, f"~perm_test_{os.getpid()}.tmp")
                with open(test_file, "w", encoding="utf-8") as f:
                    f.write("x")
                dir_writable = True
            except Exception:
                dir_writable = False
            finally:
                if test_file is not None:
                    try:
                        if os.path.exists(test_file):
                            os.remove(test_file)
                    except Exception:
                        pass
            return f"exists={exists}, readonly={readonly}, dir_writable={dir_writable}"
        except Exception:
            return "diagnostics_failed"

    def _confirm_overwrite(self, which: str, path: str) -> bool:
        if which == "A":
            modified = self.modified_a
        else:
            modified = self.modified_b

        if not modified:
            return messagebox.askyesno("提示", f"{which} 没有检测到改动。仍然要覆盖保存原文件吗？\n\n{path}")

        return messagebox.askyesno(
            "确认保存",
            f"将直接覆盖保存 {which} 文件（原路径、原文件名）：\n\n{path}\n\n建议确保该 Excel 未在 WPS/Excel 中打开。继续吗？",
        )

    def _refresh_current_view_after_val_reload(self):
        try:
            tab_id = self.nb.select()
            tab_text = self.nb.tab(tab_id, "text")
            view = self.sheet_views.get(tab_text)
            if view:
                view.refresh(row_only=None, rescan=True)
                view._update_cursor_lines()
        except Exception:
            pass
        try:
            self._refresh_sheet_catalog()
            self._sheet_diff_confirmed.clear()
            for s in self.compare_sheets:
                self.sheet_diff_state[s] = 0
            with self._compute_lock:
                self._compute_queue = [s for s in self.compare_sheets if s not in self._compute_inflight]
                self._compute_total = max(1, len(self.compare_sheets))
                self._compute_done = 0
                self._sheet_cache_store.clear()
            self._kick_worker()
        except Exception:
            pass

    def _apply_recalc_results(
        self,
        new_a=None,
        new_b=None,
        new_base=None,
        respect_user_edits: bool = False,
    ):
        if respect_user_edits:
            if self.modified_a and new_a:
                _dlog("skip stale auto-recalc result for mine: user edits already exist")
                new_a = None
            if self.modified_b and new_b:
                _dlog("skip stale auto-recalc result for theirs: user edits already exist")
                new_b = None
        specs = []
        if new_a:
            specs.append(("_wb_a_val", "_file_a_val_path", new_a))
        if new_b:
            specs.append(("_wb_b_val", "_file_b_val_path", new_b))
        if new_base and getattr(self, "has_base", False):
            specs.append(("_wb_base_val", "_file_base_val_path", new_base))
        if not specs:
            return

        loaded_items = []
        try:
            for wb_attr, path_attr, wb_path in specs:
                loaded_items.append((wb_attr, path_attr, wb_path, load_workbook(wb_path, data_only=True)))
        except Exception:
            _wbs_close(*(wb for *_rest, wb in loaded_items))
            raise

        for wb_attr, path_attr, wb_path, loaded_wb in loaded_items:
            old_wb = getattr(self, wb_attr, None)
            setattr(self, path_attr, wb_path)
            setattr(self, wb_attr, loaded_wb)
            _wbs_close(old_wb)

        self._refresh_current_view_after_val_reload()

    def _recalc_saved_path_inplace(self, path: str, which: str) -> bool:
        if not (_USE_CACHED_VALUES_ONLY and path):
            return False
        has_formula, _missing_cache = _scan_formula_cache(path)
        if not has_formula:
            return False
        tmp = _recalc_and_prepare_val_path(path)
        if not tmp:
            raise RuntimeError(f"{which} 文件包含公式，但自动重算未能生成最新缓存值。")
        try:
            self._atomic_replace_file_with_retry(tmp, path)
        finally:
            try:
                if os.path.exists(tmp):
                    os.remove(tmp)
            except Exception:
                pass
        return True

    def _post_save_refresh(self, which: str, path: str) -> str | None:
        warning = None
        try:
            self._recalc_saved_path_inplace(path, which)
        except Exception as e:
            warning = f"{which} 文件已保存，但公式缓存刷新失败：\n{e}"
            _dlog(f"post save recalc failed: which={which} path={path} err={e}")

        # A deliberate same-formula adoption is a cached-result decision. Excel
        # recalculation above is still needed for every other formula, but it can
        # overwrite that decision from local precedents. Reapply only the explicit
        # adopted cache cells after recalculation, then reload the value workbook.
        try:
            self._reapply_formula_cache_overrides(path, which)
        except Exception as e:
            cache_warning = f"{which} 文件已保存，但采用的公式计算结果未能写回：\n{e}"
            warning = f"{warning}\n\n{cache_warning}" if warning else cache_warning
            _dlog(f"post save formula cache patch failed: which={which} path={path} err={e}")

        try:
            if which == "A":
                self._apply_recalc_results(new_a=path)
            elif which == "B":
                self._apply_recalc_results(new_b=path)
            elif which == "BASE" and getattr(self, "has_base", False):
                self._apply_recalc_results(new_base=path)
        except Exception as e:
            _dlog(f"post save reload failed: which={which} path={path} err={e}")
            reload_warning = f"{which} 文件已保存，但界面刷新失败：\n{e}"
            warning = f"{warning}\n\n{reload_warning}" if warning else reload_warning
        return warning

    def _reapply_formula_cache_overrides(self, path: str, which: str) -> bool:
        which = str(which or "").upper()
        if which == "A":
            cache_ops = self.manual_a_formula_cache_ops
            edit_wb = self._wb_a_edit
        elif which == "B":
            cache_ops = self.manual_b_formula_cache_ops
            edit_wb = self._wb_b_edit
        else:
            return False
        if not cache_ops:
            return False

        formula_ops = {}
        cached_values = {}
        applied_keys = []
        for key, cached_value in list(cache_ops.items()):
            sheet, row_idx, col_idx = key
            try:
                if edit_wb is None or sheet not in edit_wb.sheetnames:
                    continue
                formula = _formula_text(
                    edit_wb[sheet].cell(row=int(row_idx), column=int(col_idx)).value
                )
                if not formula:
                    continue
                formula_ops[key] = formula
                cached_values[key] = cached_value
                applied_keys.append(key)
            except Exception:
                continue
        if not formula_ops:
            return False

        suffix = _workbook_ext(path)
        patched = os.path.join(
            tempfile.gettempdir(),
            f"{APP_NAME}_{which.lower()}_formula_cache_{os.getpid()}_{time.time_ns()}{suffix}",
        )
        try:
            _build_manual_merge_xlsx_via_zip(
                path,
                patched,
                formula_ops,
                cached_values=cached_values,
                cache_only_keys=set(formula_ops),
            )
            self._atomic_replace_file_with_retry(patched, path)
            for key in applied_keys:
                cache_ops.pop(key, None)
            _dlog(f"post save formula cache patch ok: which={which} cells={len(applied_keys)}")
            return True
        finally:
            try:
                if os.path.exists(patched):
                    os.remove(patched)
            except Exception:
                pass

    def save_b_inplace(self):
        self._ensure_edit_loaded()
        path = self.file_b
        if not self._confirm_overwrite("B", path):
            return
        try:
            warning = None

            def _do_save():
                nonlocal warning
                replay_out = None
                try:
                    if self.manual_b_row_ops or self._sheet_ops_for_target("B"):
                        replay_out = self.build_manual_b_output_file()
                        self._atomic_replace_file_with_retry(replay_out, path)
                    else:
                        self._atomic_save(self._wb_b_edit, path)
                finally:
                    if replay_out and os.path.exists(replay_out):
                        try:
                            os.remove(replay_out)
                        except Exception:
                            pass
                warning = self._post_save_refresh("B", path)

            self._with_progress("保存中", f"正在保存：\n{path}", _do_save)
            self.modified_b = False
            self.manual_b_cell_ops.clear()
            self.manual_b_row_ops.clear()
            self.manual_b_formula_cache_ops.clear()
            self._clear_sheet_ops_for_target("B")
            if warning:
                messagebox.showwarning("Saved", f"已保存并覆盖：\n{path}\n\n{warning}")
            else:
                messagebox.showinfo("Saved", f"已保存并覆盖：\n{path}")
        except Exception as e:
            # If the file is locked or denied, offer save-as fallback
            if getattr(e, "winerror", None) in (5, 32, 33) or isinstance(e, PermissionError):
                diag = self._path_diagnostics(path)
                if messagebox.askyesno("保存失败", f"保存 B 失败（可能文件被占用/无权限）：\n{e}\n\n诊断：{diag}\n\n是否另存为？"):
                    if self._try_alt_save(self._wb_b_edit, path, "B"):
                        self.modified_b = False
                        return
            messagebox.showerror("保存失败", f"保存 B 失败：\n{e}")

    def save_a_inplace(self):
        self._ensure_edit_loaded()
        path = self.file_a
        if not self._confirm_overwrite("A", path):
            return
        try:
            warning = None

            def _do_save():
                nonlocal warning
                replay_out = None
                try:
                    if self.manual_a_row_ops or self._sheet_ops_for_target("A"):
                        replay_out = self.build_manual_merge_output_file()
                        self._atomic_replace_file_with_retry(replay_out, path)
                    else:
                        self._atomic_save(self._wb_a_edit, path)
                finally:
                    if replay_out and os.path.exists(replay_out):
                        try:
                            os.remove(replay_out)
                        except Exception:
                            pass
                warning = self._post_save_refresh("A", path)

            self._with_progress("保存中", f"正在保存：\n{path}", _do_save)
            self.modified_a = False
            self.manual_a_cell_ops.clear()
            self.manual_a_row_ops.clear()
            self.manual_a_formula_cache_ops.clear()
            self._clear_sheet_ops_for_target("A")
            if warning:
                messagebox.showwarning("Saved", f"已保存并覆盖：\n{path}\n\n{warning}")
            else:
                messagebox.showinfo("Saved", f"已保存并覆盖：\n{path}")
        except Exception as e:
            if getattr(e, "winerror", None) in (5, 32, 33) or isinstance(e, PermissionError):
                diag = self._path_diagnostics(path)
                if messagebox.askyesno("保存失败", f"保存 A 失败（可能文件被占用/无权限）：\n{e}\n\n诊断：{diag}\n\n是否另存为？"):
                    if self._try_alt_save(self._wb_a_edit, path, "A"):
                        self.modified_a = False
                        return
            messagebox.showerror("保存失败", f"保存 A 失败：\n{e}")


    def save_merged_and_exit(self, auto: bool = False):
        if not self.merged_path:
            return
        if not auto:
            if self.merge_mode and self.initial_conflict_cell_count > 0:
                unresolved = sum(
                    len(cols)
                    for rows in self.merge_conflict_cells_by_sheet.values()
                    for cols in rows.values()
                )
                if not messagebox.askyesno(
                    "确认冲突处理",
                    f"三方扫描检测到 {self.initial_conflict_cell_count} 个冲突单元格。"
                    f"\n当前仍标记 {unresolved} 个（手动模式下不会自动清零）。"
                    "\n\n请确认你已完成需要处理的冲突数据。是否继续保存？",
                ):
                    return
            if self.merge_mode and getattr(self, "sheet_level_conflicts", None):
                names = ", ".join(str(item.get("sheet")) for item in self.sheet_level_conflicts[:6] if item.get("sheet"))
                if len(self.sheet_level_conflicts) > 6:
                    names += " ..."
                if not messagebox.askyesno(
                    "确认整Sheet变更",
                    f"检测到 {len(self.sheet_level_conflicts)} 个无法自动归类的整Sheet变更。"
                    f"\n默认将保留 mine 版本继续保存。"
                    f"\n涉及：{names or '-'}"
                    "\n\n是否继续保存？",
                ):
                    return
            if not messagebox.askyesno("确认保存", f"将保存合并结果到：\n\n{self.merged_path}\n\n继续吗？"):
                return
        wb_to_save = None
        merged_source_path = None
        try:
            def _save_merged_task(report):
                nonlocal wb_to_save, merged_source_path
                report("正在保存 merged 文件", "等待可编辑工作簿准备完成...", 5)
                self._ensure_edit_loaded()
                wb_to_save = self._wb_a_edit

                # Give SVN/Tortoise a short chance to release transient locks,
                # but keep the UI event loop alive while waiting.
                report("正在保存 merged 文件", "等待 SVN/Tortoise 释放临时文件锁...", 12)
                time.sleep(1.2)

                if self.merge_mode and self.has_base:
                    # Manual 3-way output is rebuilt from pristine mine plus the
                    # recorded operations. This is the expensive part of save and
                    # must be covered by the visible progress window.
                    report("正在构建合并结果", "重放整 Sheet、插入行、单元格和公式缓存操作...", 25)
                    merged_source_path = self.build_manual_merge_output_file()

                try:
                    by_sheet = {}
                    for (s, _r, _c) in getattr(self, "manual_a_cell_ops", {}).keys():
                        by_sheet[s] = by_sheet.get(s, 0) + 1
                    _dlog(
                        f"SAVE_MERGED path={self.merged_path} "
                        f"merge_mode={self.merge_mode} has_base={self.has_base} "
                        f"manual_a_ops={len(getattr(self, 'manual_a_cell_ops', {}))} "
                        f"manual_sheet_ops={len(getattr(self, 'manual_sheet_ops', []))} "
                        f"auto_sheet_ops={len(getattr(self, 'auto_sheet_ops', []))} "
                        f"manual_ops_by_sheet={by_sheet} "
                        f"snapshot={getattr(self, '_merge_mine_snapshot', None)}"
                    )
                except Exception:
                    pass

                report("正在写入 merged 文件", os.path.basename(self.merged_path), 82)
                if merged_source_path:
                    self._atomic_replace_file_with_retry(merged_source_path, self.merged_path)
                else:
                    self._atomic_save_with_retry(wb_to_save, self.merged_path)

                report("正在校验保存结果", "检查 OOXML/ZIP 结构和文件完整性...", 95)
                if not _workbook_package_ready(self.merged_path):
                    raise RuntimeError(f"保存后的工作簿完整性校验失败：{self.merged_path}")
                report("保存完成", self.merged_path, 100)

            self._begin_interactive_action()
            try:
                self._with_progress(
                    "保存 merged",
                    f"正在保存合并结果：\n{self.merged_path}",
                    _save_merged_task,
                    run_in_background=True,
                    pass_reporter=True,
                )
            finally:
                self._end_interactive_action()
            self.modified_a = False
            try:
                messagebox.showinfo("Saved", f"已保存合并结果：\n{self.merged_path}")
            except Exception:
                pass
        except Exception as e:
            if getattr(e, "winerror", None) in (5, 32, 33) or isinstance(e, PermissionError):
                excel_locked = _log_lock_holders(self.merged_path)
                if excel_locked:
                    try:
                        messagebox.showwarning("文件被占用", "检测到 Excel 正在占用目标文件。\n请关闭 Excel 后再保存。")
                    except Exception:
                        pass
                # In conflict UI, target file might still be locked by SVN/Tortoise.
                # Save to a temp file and schedule a deferred replace.
                try:
                    folder = os.path.dirname(self.merged_path)
                    base = os.path.basename(self.merged_path)
                    tmp_path = os.path.join(folder, f"~{base}.deferred.{os.getpid()}.tmp")
                    if merged_source_path:
                        shutil.copy2(merged_source_path, tmp_path)
                    elif _FAST_SAVE_VALUES_ONLY and _USE_CACHED_VALUES_ONLY:
                        _save_values_only_from_wb(wb_to_save, tmp_path)
                    else:
                        wb_to_save.save(tmp_path)
                    if not os.path.exists(tmp_path) or os.path.getsize(tmp_path) == 0:
                        raise RuntimeError(f"临时文件写入失败或为空：{tmp_path}")
                    if not _launch_deferred_copy(tmp_path, self.merged_path):
                        # Deferred copy could not even be launched: do NOT claim
                        # success or silently exit. Surface the temp path so the
                        # user can recover the merge result manually, then fall
                        # through to the save-as recovery offer below.
                        try:
                            messagebox.showerror(
                                "保存失败",
                                "无法启动后台延迟复制进程，合并结果尚未写入目标文件。\n"
                                f"已保留临时文件：\n{tmp_path}\n\n"
                                "请关闭占用目标文件的程序后，手动将该临时文件复制到：\n"
                                f"{self.merged_path}",
                            )
                        except Exception:
                            pass
                        raise RuntimeError(f"deferred copy launch failed; temp preserved at {tmp_path}")
                    messagebox.showinfo("保存中", f"目标文件被占用，已写入临时文件并将在关闭后自动覆盖：\n{self.merged_path}")
                    try:
                        self._shutdown_root()
                    except Exception:
                        pass
                    sys.exit(0)
                except Exception:
                    diag = self._path_diagnostics(self.merged_path)
                    if messagebox.askyesno("保存失败", f"保存合并结果失败（可能文件被占用/无权限）：\n{e}\n\n诊断：{diag}\n\n是否另存为？"):
                        alt_ok = False
                        if merged_source_path and os.path.exists(merged_source_path):
                            alt = self._alt_save_path(self.merged_path, "MERGED")
                            try:
                                shutil.copy2(merged_source_path, alt)
                                messagebox.showinfo("另存为成功", f"无法覆盖原文件，已另存为：\n{alt}")
                                alt_ok = True
                            except Exception:
                                pass
                        if not alt_ok:
                            alt_ok = self._try_alt_save(wb_to_save, self.merged_path, "MERGED")
                        if alt_ok:
                            try:
                                self._shutdown_root()
                            except Exception:
                                pass
                            sys.exit(0)
                        return
            messagebox.showerror("保存失败", f"保存合并结果失败：\n{e}")
            return
        finally:
            if merged_source_path:
                try:
                    os.remove(merged_source_path)
                except Exception:
                    pass
        # Try auto-resolve in SVN if conflict artifacts exist
        try:
            if _has_svn_conflict_artifacts(self.merged_path):
                _try_svn_resolve(self.merged_path)
        except Exception:
            pass
        try:
            self._shutdown_root()
        except Exception:
            pass
        sys.exit(0)

    def resolve_conflict_cell(self, sheet: str, r: int, c: int) -> bool:
        rows = self.merge_conflict_cells_by_sheet.get(sheet)
        if not rows:
            return False
        cols = rows.get(r)
        if not cols or c not in cols:
            return False
        cols.discard(c)
        if not cols:
            rows.pop(r, None)
        if not rows:
            self.merge_conflict_cells_by_sheet.pop(sheet, None)
        self._auto_save_if_no_conflicts()
        return True

    def resolve_conflict_row(self, sheet: str, r: int, cols) -> bool:
        rows = self.merge_conflict_cells_by_sheet.get(sheet)
        if not rows or r not in rows:
            return False
        for c in list(cols):
            rows[r].discard(c)
        if not rows[r]:
            rows.pop(r, None)
        if not rows:
            self.merge_conflict_cells_by_sheet.pop(sheet, None)
        self._auto_save_if_no_conflicts()
        return True

    def _auto_save_if_no_conflicts(self):
        if not self.merge_conflict_cells_by_sheet:
            # If user has manually touched conflicts, require explicit save.
            if getattr(self, "user_touched_conflicts", False):
                return
            # all conflicts resolved
            self.save_merged_and_exit(auto=True)

    def run(self):
        self.root.mainloop()


def main():
    try:
        try:
            _trace_launch("=" * 72)
            _trace_launch(f"cwd={os.getcwd()}")
            _trace_launch(f"argv={repr(sys.argv)}")
        except Exception:
            pass
        # Log raw args early for troubleshooting (even if argparse fails)
        try:
            _dlog(f"argv: {' '.join(sys.argv[1:])}")
        except Exception:
            pass

        def _parse_slash_args(argv):
            out = {}
            keys = ("path", "path2", "base", "mine", "theirs", "merged")
            i = 0
            n = len(argv)
            while i < n:
                a = argv[i]
                la = a.lower()
                matched = False
                for k in keys:
                    p1 = f"/{k}:"
                    p2 = f"/{k}="
                    p3 = f"-{k}:"
                    p4 = f"-{k}="
                    p5 = f"/{k}"
                    p6 = f"-{k}"
                    if la.startswith(p1) or la.startswith(p3):
                        # Colon-delimited: /key:value or -key:value
                        out[k] = a.split(":", 1)[1]
                        matched = True
                        break
                    elif la.startswith(p2) or la.startswith(p4):
                        # Equals-delimited: /key=value or -key=value
                        # Must split on "=" only; ":" in the value is a drive-letter separator on Windows.
                        out[k] = a.split("=", 1)[1]
                        matched = True
                        break
                    if la == p5 or la == p6:
                        if i + 1 < n:
                            out[k] = argv[i + 1]
                            i += 1
                        matched = True
                        break
                i += 1
            return out

        slash_args = _parse_slash_args(sys.argv[1:])
        try:
            _trace_launch(f"slash_args={repr(slash_args)}")
        except Exception:
            pass

        parser = argparse.ArgumentParser(add_help=True)
        parser.add_argument("file_a", nargs="?")
        parser.add_argument("file_b", nargs="?")
        # SVN/TortoiseSVN style args
        parser.add_argument("--base")
        parser.add_argument("--mine")
        parser.add_argument("--theirs")
        parser.add_argument("--merged")
        parser.add_argument("--title")
        parser.add_argument("--textdiff", action="store_true", help="Only generate text files and open TortoiseMerge")
        args, unknown = parser.parse_known_args()
        try:
            _trace_launch(f"argparse={repr(vars(args))} unknown={repr(unknown)}")
        except Exception:
            pass
        if unknown:
            try:
                _dlog(f"unknown args: {unknown}")
            except Exception:
                pass

        # Map /path:/path2:/base: style args (TortoiseProc)
        if not args.base and "base" in slash_args:
            args.base = slash_args.get("base")
        if not args.mine and "mine" in slash_args:
            args.mine = slash_args.get("mine")
        if not args.theirs and "theirs" in slash_args:
            args.theirs = slash_args.get("theirs")
        if not args.merged and "merged" in slash_args:
            args.merged = slash_args.get("merged")
        if not args.file_a and "path" in slash_args:
            args.file_a = slash_args.get("path")
        if not args.file_b and "path2" in slash_args:
            args.file_b = slash_args.get("path2")
        # Fallback: some launchers pass paths as plain unknown tokens.
        # Try extracting existing filesystem paths from unknown args.
        if (not args.file_a) and unknown:
            path_tokens = []
            for u in unknown:
                if not u:
                    continue
                su = str(u).strip().strip('"')
                if not su or su.startswith("-") or su.startswith("/"):
                    continue
                try:
                    if os.path.exists(su):
                        path_tokens.append(su)
                except Exception:
                    pass
            if path_tokens:
                args.file_a = path_tokens[0]
                if len(path_tokens) >= 2:
                    args.file_b = path_tokens[1]
        try:
            _trace_launch(
                "resolved args: "
                + f"file_a={repr(args.file_a)} file_b={repr(args.file_b)} "
                + f"base={repr(args.base)} mine={repr(args.mine)} "
                + f"theirs={repr(args.theirs)} merged={repr(args.merged)}"
            )
        except Exception:
            pass

        # Map SVN-style args to our 2-pane viewer (diff mode) / merge mode.
        if args.base and args.mine and args.theirs and args.merged:
            # Full 3-way merge args are already provided; do not fall back to file picker.
            a, b = None, None
        elif args.base and args.mine and not args.theirs:
            # Diff mode: keep base(revision/older) on left, mine(working-copy) on right.
            a, b = args.base, args.mine
        elif args.file_a and args.file_b:
            a, b = args.file_a, args.file_b
        elif args.file_a and (not args.file_b) and (not args.base) and (not args.mine) and (not args.theirs):
            # Single file provided (e.g., from Explorer/TortoiseSVN). If it's a conflicted file, auto-merge it.
            conflict = _detect_svn_conflict_files(args.file_a)
            if (not conflict) and args.file_a:
                try:
                    auto_target = _find_conflict_in_dir(os.path.dirname(os.path.abspath(args.file_a)))
                    if auto_target:
                        conflict = _detect_svn_conflict_files(auto_target)
                except Exception:
                    conflict = conflict
            if conflict:
                args.base, args.mine, args.theirs, args.merged = conflict
                args.force_ui = True
            else:
                a, b = args.file_a, None
        else:
            sel = pick_files_or_conflict()
            if not sel:
                return
            if sel[0] == "merge":
                _mode, base_p, mine_p, theirs_p, merged_p, force_ui = sel
                args.base = _ensure_xlsx_copy(base_p)
                args.mine = _ensure_xlsx_copy(mine_p)
                args.theirs = _ensure_xlsx_copy(theirs_p)
                args.merged = merged_p
                args.force_ui = bool(force_ui)
            else:
                _mode, a, b = sel

        if args.file_a and (args.file_b is None) and (not args.base) and b is None:
            # Need second file for diff mode
            root = tk.Tk()
            root.withdraw()
            b = filedialog.askopenfilename(title="Select second Excel workbook (same filename)", filetypes=[("Excel Workbook", "*.xlsx *.xlsm")])
            if not b:
                root.destroy()
                return
            if os.path.basename(args.file_a).lower() != os.path.basename(b).lower():
                messagebox.showerror(
                    "Filename mismatch",
                    f"The two files must have the same filename.\n\nA: {os.path.basename(args.file_a)}\nB: {os.path.basename(b)}",
                )
                root.destroy()
                return
            root.destroy()
            a = args.file_a

        raw_base_arg = args.base
        raw_mine_arg = args.mine
        raw_theirs_arg = args.theirs

        # Normalize SVN merge temp files (merge-left/right.r####) by exporting true revision.
        # IMPORTANT:
        # - base/theirs may legitimately be revision snapshots.
        # - mine must stay as the working-copy side; do NOT rewrite mine to a revision export,
        #   otherwise local edits can be replaced by an old revision file.
        full_merge_args = bool(args.base and args.mine and args.theirs and args.merged)
        if (not full_merge_args) and args.base:
            args.base = _try_export_svn_revision_from_merge_temp(args.base)
        if args.theirs:
            # In merge mode, keep "theirs" exactly as passed by SVN/Tortoise wrapper.
            # This avoids accidental re-export to another revision snapshot and ensures
            # content matches the user-visible *.merge-right.r#### sidecar file.
            if not full_merge_args:
                try:
                    args.theirs = _try_export_svn_revision_from_merge_temp(args.theirs)
                except Exception:
                    args.theirs = args.theirs
        if args.file_a:
            args.file_a = _try_export_svn_revision_from_merge_temp(args.file_a)
        if args.file_b:
            args.file_b = _try_export_svn_revision_from_merge_temp(args.file_b)
        try:
            _trace_launch(
                "normalized args: "
                + f"base={repr(args.base)} mine={repr(args.mine)} "
                + f"theirs={repr(args.theirs)} merged={repr(args.merged)} "
                + f"raw_base={repr(raw_base_arg)} raw_theirs={repr(raw_theirs_arg)}"
            )
        except Exception:
            pass

        # Merge mode (manual 3-way): detect conflicts only; do NOT pre-merge before UI.
        if args.base and args.mine and args.theirs and args.merged:
            conflicts = []
            conflict_map = {}
            try:
                _dlog(f"merge args: base={args.base} mine={args.mine} theirs={args.theirs} merged={args.merged}")
                _dlog(f"merge manual mode unknown={unknown}")
            except Exception:
                pass
            try:
                def _prepare_merge_inputs(report):
                    nonlocal raw_base_arg
                    report("正在读取 SVN 合并来源", "从工作副本的 .svn 数据读取 BASE...", 5)
                    base_source = None
                    try:
                        base_source = _try_export_svn_base_from_working_copy(args.merged)
                    except Exception:
                        base_source = None
                    if base_source:
                        _dlog(f"merge base selected from WC BASE: {base_source}")
                        mine_for_note = raw_mine_arg or args.mine or args.merged or "-"
                        raw_base_arg = f"{mine_for_note}@BASE(.svn)"
                    else:
                        report("正在读取 SVN 合并来源", "准备 TortoiseSVN 提供的 base 版本...", 15)
                        base_source = _try_export_svn_revision_from_merge_temp(args.base)

                    report("正在准备三方合并", "规范化 base 版本文件...", 22)
                    base_path = _ensure_xlsx_copy(base_source)
                    report("正在准备三方合并", "规范化 mine 工作副本...", 30)
                    mine_path = _ensure_xlsx_copy(args.mine)
                    report("正在准备三方合并", "规范化 theirs 版本文件...", 38)
                    theirs_path = _ensure_xlsx_copy(args.theirs)
                    _dlog("merge start: calling _scan_three_way_conflicts (no pre-merge)")
                    report("正在扫描三方冲突", "精确比较 base、mine 和 theirs，请稍候...", 45)
                    scan_conflicts, scan_map = _scan_three_way_conflicts(base_path, mine_path, theirs_path)
                    report("三方扫描完成", f"检测到 {len(scan_conflicts)} 个冲突单元格", 100)
                    return base_path, mine_path, theirs_path, scan_conflicts, scan_map

                (
                    args.base,
                    args.mine,
                    args.theirs,
                    conflicts,
                    conflict_map,
                ) = _run_startup_progress_task(
                    "Excel 合并工具 - 三方扫描",
                    "正在准备 SVN 合并数据...",
                    _prepare_merge_inputs,
                )
                try:
                    _dlog(f"merge scan result: conflicts={len(conflicts)} conflict_sheets={len(conflict_map)}")
                except Exception:
                    pass
            except Exception as e:
                try:
                    _dlog(f"merge exception: {e}")
                except Exception:
                    pass
                try:
                    messagebox.showerror("Merge failed", f"合并失败：\n{e}")
                except Exception:
                    print(f"Merge failed: {e}", file=sys.stderr)
                sys.exit(1)

            if conflicts:
                _show_conflict_popup(conflicts)

                try:
                    messagebox.showinfo(
                        "进入手动处理",
                        f"检测到 {len(conflicts)} 个冲突单元格。\n将进入手动 3 视图处理界面。",
                    )
                except Exception:
                    pass
            else:
                try:
                    messagebox.showinfo(
                        "进入手动处理",
                        "未检测到直接冲突。\n仍将进入手动 3 视图，所有差异由你确认后保存。"
                    )
                except Exception:
                    pass

            app = SowMergeApp(
                args.mine,
                args.theirs,
                merge_mode=True,
                merged_path=args.merged,
                base_path=args.base,
                merge_conflict_cells_by_sheet=conflict_map,
                merge_conflict_mode=False,
                raw_base=raw_base_arg,
                raw_mine=raw_mine_arg,
                raw_theirs=raw_theirs_arg,
            )
            try:
                _dlog("open UI: manual 3-way mode")
            except Exception:
                pass
            app.run()
            sys.exit(0)

        if args.textdiff:
            try:
                temp_root = os.path.join(os.environ.get("LOCALAPPDATA", tempfile.gettempdir()), "Temp", "TortoiseXlsTemp")
                os.makedirs(temp_root, exist_ok=True)
            except Exception:
                temp_root = tempfile.gettempdir()

            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            left_txt = os.path.join(temp_root, f"{APP_NAME}_left_{ts}.txt")
            right_txt = os.path.join(temp_root, f"{APP_NAME}_right_{ts}.txt")
            excel_to_text(a, left_txt, thick_sep_char="=")
            excel_to_text(b, right_txt, thick_sep_char="=")
            open_tortoise_merge(left_txt, right_txt, title=f"{APP_NAME}: {os.path.basename(a)}")
            return

        app = SowMergeApp(
            a,
            b,
            raw_base=raw_base_arg,
            raw_mine=raw_mine_arg,
            raw_theirs=raw_theirs_arg,
        )
        app.run()

    except Exception:
        err = traceback.format_exc()
        try:
            _destroy_startup_progress_root()
        except Exception:
            pass
        try:
            messagebox.showerror("Error", err)
        except Exception:
            print(err, file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()

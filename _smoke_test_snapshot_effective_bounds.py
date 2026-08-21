"""Focused pure coverage for paired snapshot effective-boundary parity."""

from __future__ import annotations

import argparse
import shutil
import tempfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import PatternFill

import sow_merge_tool as sm
from _large_sheet_direct_oracle import capture as capture_direct


_CASE = "snapshot-effective-bounds"
_SHEET = "Data"


def _write_fixture(path: Path, kind: str) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _SHEET
    if kind != "all-empty":
        sheet["A1"] = "id@id"
        sheet["B1"] = "value@pm"
        sheet["A2"] = "string"
        sheet["B2"] = "string"
        sheet["A3"] = "record-1"
        sheet["B3"] = "payload"
    if kind == "formula-tail":
        sheet["F3"] = "=40+2"
    else:
        # A styled-but-empty cell forces a declared physical tail without
        # creating value/formula content.
        sheet["F1"].fill = PatternFill(fill_type="solid", fgColor="00FF00")
    workbook.save(path)
    workbook.close()


def _candidate_columns(path: Path):
    mine = sm._stream_selected_sheet_snapshot(str(path), str(path), _SHEET, "A")
    theirs = sm._stream_selected_sheet_snapshot(str(path), str(path), _SHEET, "B")
    result = sm._compare_selected_sheet_snapshots(mine, theirs)
    return mine, result, sm.snapshot_comparison_oracle_manifest(mine, theirs, result)


def _unresolved_columns(manifest) -> frozenset[int]:
    return frozenset(
        int(slot["logical"])
        for slot in manifest["columns"]
        if slot["state"] == "unresolved" or bool(slot["ambiguous"])
    )


def _assert_direct_column_parity(path: Path, expected_unresolved: frozenset[int]):
    right = path.with_name("right.xlsx")
    shutil.copy2(path, right)
    direct = capture_direct(str(path), str(right), _SHEET)
    if _unresolved_columns(direct) != expected_unresolved:
        raise AssertionError(("direct", _unresolved_columns(direct), expected_unresolved))
    snapshot, result, candidate = _candidate_columns(path)
    if direct["columns"] != candidate["columns"]:
        raise AssertionError((direct["columns"], candidate["columns"]))
    if _unresolved_columns(candidate) != expected_unresolved:
        raise AssertionError(("candidate", _unresolved_columns(candidate), expected_unresolved))
    return snapshot, result


def run_case() -> None:
    with tempfile.TemporaryDirectory(prefix="sow_snapshot_effective_bounds_") as raw_root:
        root = Path(raw_root)

        style_tail = root / "style-tail.xlsx"
        _write_fixture(style_tail, "style-tail")
        snapshot, result = _assert_direct_column_parity(style_tail, frozenset())
        assert (snapshot.max_row, snapshot.max_col) == (3, 2)
        assert not result.unresolved
        assert len(snapshot.rows) == 3
        assert all(len(row.cells) == 2 for row in snapshot.rows)
        assert all(row.row_hash == sm._snapshot_row_hash(row.cells) for row in snapshot.rows)

        formula_tail = root / "formula-tail.xlsx"
        _write_fixture(formula_tail, "formula-tail")
        snapshot, result = _assert_direct_column_parity(
            formula_tail, frozenset((3, 4, 5))
        )
        assert (snapshot.max_row, snapshot.max_col) == (3, 6)
        assert snapshot.rows[2].cells[5].formula_value == "=40+2"
        assert result.unresolved

        all_empty = root / "all-empty.xlsx"
        _write_fixture(all_empty, "all-empty")
        snapshot, result = _assert_direct_column_parity(
            all_empty, frozenset(range(1, 7))
        )
        assert (snapshot.max_row, snapshot.max_col) == (1, 6)
        assert len(snapshot.rows) == 1 and len(snapshot.rows[0].cells) == 6
        assert snapshot.rows[0].row_hash == sm._snapshot_row_hash(snapshot.rows[0].cells)
        assert result.unresolved


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--case", choices=(_CASE,))
    parser.add_argument("--list-cases", action="store_true")
    args = parser.parse_args()
    if args.list_cases:
        print(_CASE)
        return
    run_case()
    print(f"PASS {_CASE}")


if __name__ == "__main__":
    main()

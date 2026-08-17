"""Fast branch-analysis and unique-record merge smoke tests."""

from __future__ import annotations

import os
import shutil
import tempfile
import time

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from sow_merge_tool.fast_branch_merge import (
    analyze_source,
    analyze_target,
    apply_source_change_plan,
)


def _book(path: str, rows: list[list[object]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    workbook.close()


def _large_book(path: str, *, changed: bool = False, target_only: bool = False) -> None:
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("Language")
    sheet.append(["ID", "文本"] + (["目标分支字段"] if target_only else []))
    for index in range(1, 5001):
        value = f"新文本{index}" if changed and index <= 1000 else f"旧文本{index}"
        sheet.append([f"key-{index}", value] + (["保留"] if target_only else []))
    workbook.save(path)


def main() -> None:
    root = tempfile.mkdtemp(prefix="sow-fast-branch-")
    try:
        before = os.path.join(root, "before.xlsx")
        source = os.path.join(root, "source.xlsx")
        target = os.path.join(root, "target.xlsx")
        output = os.path.join(root, "candidate.xlsx")
        _book(before, [["ID", "文本"], ["a", "旧"]])
        _book(source, [["ID", "文本"], ["a", "旧"], ["new", "新增"]])
        shutil.copy2(before, target)
        delta = analyze_source(before, source)
        decision = analyze_target(delta, target)
        assert decision.disposition == "direct", decision
        apply_source_change_plan(before, source, target, output, decision)
        workbook = load_workbook(output, read_only=True, data_only=False)
        rows = list(workbook["Data"].iter_rows(values_only=True))
        workbook.close()
        assert rows[-1] == ("new", "新增"), rows

        two_added = os.path.join(root, "two-added.xlsx")
        _book(two_added, [["ID", "文本"], ["a", "旧"], ["new-1", "一"], ["new-2", "二"]])
        two_delta = analyze_source(before, two_added)
        two_decision = analyze_target(two_delta, target)
        assert two_decision.disposition == "direct", two_decision
        apply_source_change_plan(before, two_added, target, output, two_decision)
        two_book = load_workbook(output, read_only=True, data_only=False)
        assert [row[0] for row in two_book["Data"].iter_rows(min_row=2, values_only=True)] == [
            "a", "new-1", "new-2"
        ]
        two_book.close()

        conflict_target = os.path.join(root, "conflict.xlsx")
        _book(conflict_target, [["ID", "文本"], ["a", "目标独立修改"]])
        changed = os.path.join(root, "changed.xlsx")
        _book(changed, [["ID", "文本"], ["a", "源修改"]])
        changed_delta = analyze_source(before, changed)
        conflict_decision = analyze_target(changed_delta, conflict_target)
        assert conflict_decision.disposition == "confirmation_required"
        apply_source_change_plan(
            before, changed, conflict_target, output, conflict_decision, confirmed=True
        )
        confirmed_book = load_workbook(output, read_only=True, data_only=False)
        assert confirmed_book["Data"]["B2"].value == "源修改"
        confirmed_book.close()

        styled_source = os.path.join(root, "styled-source.xlsx")
        styled_target = os.path.join(root, "styled-target.xlsx")
        styled_book = Workbook()
        styled_sheet = styled_book.active
        styled_sheet.title = "Data"
        styled_sheet.append(["ID", "文本"])
        styled_sheet.append(["a", "旧"])
        styled_sheet.append(["styled", "需要人工"])
        styled_sheet["B3"].font = Font(bold=True)
        styled_book.save(styled_source)
        styled_book.close()
        shutil.copy2(before, styled_target)
        styled_target_book = load_workbook(styled_target)
        styled_target_book["Data"]["B2"].fill = PatternFill("solid", fgColor="00FF00")
        styled_target_book["Data"]["B2"].font = Font(italic=True)
        styled_target_book.save(styled_target)
        styled_target_book.close()
        styled_delta = analyze_source(before, styled_source)
        styled_decision = analyze_target(styled_delta, styled_target)
        assert styled_decision.disposition == "direct", styled_decision
        apply_source_change_plan(before, styled_source, styled_target, output, styled_decision)
        styled_result = load_workbook(output)
        assert styled_result["Data"]["B3"].fill.fgColor.rgb == "0000FF00"
        assert styled_result["Data"]["B3"].font.italic
        assert not styled_result["Data"]["B3"].font.bold
        styled_result.close()

        structural_source = os.path.join(root, "structural-source.xlsx")
        structural_book = load_workbook(before)
        structural_sheet = structural_book["Data"]
        structural_sheet["B2"] = "源修改"
        structural_sheet.merge_cells("B2:C2")
        structural_book.save(structural_source)
        structural_book.close()
        structural_delta = analyze_source(before, structural_source)
        assert structural_delta.unsupported_reason, structural_delta
        assert "合并单元格" in structural_delta.unsupported_reason, structural_delta.unsupported_reason

        validation_source = os.path.join(root, "validation-source.xlsx")
        validation_book = load_workbook(before)
        validation_book["Data"].add_data_validation(DataValidation(type="whole", operator="greaterThan", formula1="0"))
        validation_book["Data"].data_validations.dataValidation[0].add("B2")
        validation_book.save(validation_source)
        validation_book.close()
        assert "数据校验" in analyze_source(before, validation_source).unsupported_reason

        style_only_source = os.path.join(root, "style-only-source.xlsx")
        style_only_book = load_workbook(before)
        style_only_book["Data"]["B2"].fill = PatternFill("solid", fgColor="00FF00")
        style_only_book.save(style_only_source)
        style_only_book.close()
        style_only_delta = analyze_source(before, style_only_source)
        assert not style_only_delta.unsupported_reason and style_only_delta.incoming_count == 0
        assert analyze_target(style_only_delta, target).disposition == "already_applied"

        selection_only_source = os.path.join(root, "selection-only-source.xlsx")
        selection_book = load_workbook(before)
        selection_book["Data"].sheet_view.selection[0].activeCell = "B2"
        selection_book["Data"].sheet_view.selection[0].sqref = "B2"
        selection_book.save(selection_only_source)
        selection_book.close()
        selection_delta = analyze_source(before, selection_only_source)
        assert not selection_delta.unsupported_reason and selection_delta.incoming_count == 0

        crlf_before = os.path.join(root, "crlf-before.xlsx")
        crlf_after = os.path.join(root, "crlf-after.xlsx")
        _book(crlf_before, [["ID", "文本"], ["a", "第一行\r\n第二行"]])
        _book(crlf_after, [["ID", "文本"], ["a", "第一行_x000D_\n第二行"]])
        crlf_delta = analyze_source(crlf_before, crlf_after)
        assert not crlf_delta.unsupported_reason and crlf_delta.incoming_count == 0, crlf_delta

        large_before = os.path.join(root, "large-before.xlsx")
        large_source = os.path.join(root, "large-source.xlsx")
        large_target = os.path.join(root, "large-target.xlsx")
        large_output = os.path.join(root, "large-output.xlsx")
        _large_book(large_before)
        _large_book(large_source, changed=True)
        _large_book(large_target, target_only=True)
        started = time.perf_counter()
        large_delta = analyze_source(large_before, large_source)
        large_decision = analyze_target(large_delta, large_target)
        apply_source_change_plan(
            large_before, large_source, large_target, large_output, large_decision
        )
        elapsed = time.perf_counter() - started
        assert large_delta.incoming_count == 1000
        assert large_decision.disposition == "direct"
        assert elapsed < 8.0, f"large source-change projection took {elapsed:.2f}s"
        large_book = load_workbook(large_output, read_only=True, data_only=False)
        assert large_book["Language"]["B2"].value == "新文本1"
        assert large_book["Language"]["C2"].value == "保留"
        large_book.close()
        print("PASS: fast source-change projection and confirmation")
    finally:
        shutil.rmtree(root, ignore_errors=True)


if __name__ == "__main__":
    main()

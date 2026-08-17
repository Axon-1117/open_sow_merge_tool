"""Regression gate for fast sparse column mutation parity and native replay.

The in-memory fast path is compared with openpyxl's insert/delete cell mover
while both sides use the product's supported metadata-range policy.  A second
layer replays the same structural log through Excel COM and compares the saved
workbook with the fast in-memory result.  COM unavailability is reported as an
environment skip, not disguised as a passing native replay.
"""

from __future__ import annotations

import copy
import os

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

import sow_merge_tool as smt
from _test_temp_utils import make_temp_dir


_SHEET = "Sparse"
_INSERT_ANCHOR = 3
_INSERT_COUNT = 2
_DELETE_ANCHOR = 8
_DELETE_COUNT = 1


def _close_workbook(workbook):
    if workbook is not None:
        workbook.close()


def _make_fixture(path: str):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _SHEET
    for col, header in enumerate(("A", "B", "C", "D", "E", "F"), start=1):
        sheet.cell(1, col).value = header
    sheet["A2"] = "left"
    sheet["B3"] = 7
    sheet["D2"] = "decorated"
    sheet["D2"].font = Font(bold=True, color="FFFFFF")
    sheet["D2"].fill = PatternFill("solid", fgColor="2277AA")
    sheet["D2"].comment = Comment("fast-column-comment", "Codex")
    sheet["D2"].hyperlink = "https://example.test/fast-column"
    sheet["E3"] = "=A3*2"
    sheet["F4"] = "tail-to-delete"
    sheet.column_dimensions["D"].width = 23.0
    sheet.column_dimensions["D"].hidden = True
    sheet.column_dimensions["E"].width = 31.0
    validation = DataValidation(type="list", formula1='"red,green"', allow_blank=True)
    validation.add("D2:E4")
    sheet.add_data_validation(validation)
    sheet.conditional_formatting.add(
        "D2:E4",
        CellIsRule(
            operator="equal",
            formula=["1"],
            fill=PatternFill("solid", fgColor="FFFF00"),
        ),
    )
    sheet.merge_cells("D5:E5")
    sheet["D5"] = "merged"
    workbook.save(path)
    workbook.close()


def _apply_metadata_aware_shift(sheet, *, fast: bool, anchor: int, count: int, insert: bool):
    view = smt.SheetView
    dimensions = view._column_action_dimension_snapshot(sheet)
    validations = view._column_action_validation_ranges(sheet)
    conditionals = view._column_action_conditional_ranges(sheet)
    merges = view._column_action_merged_ranges(sheet)
    view._replace_merged_ranges(sheet, ())
    if fast:
        view._shift_worksheet_columns_fast(sheet, anchor, count, insert=insert)
    elif insert:
        sheet.insert_cols(anchor, count)
    else:
        sheet.delete_cols(anchor, count)
    view._restore_shifted_column_dimensions(
        sheet, dimensions, anchor, count, insert=insert
    )
    view._replace_validation_ranges(
        sheet,
        view._shift_column_validation_ranges(
            validations, anchor, count, insert=insert
        ),
    )
    view._replace_conditional_ranges(
        sheet,
        view._shift_column_validation_ranges(
            conditionals, anchor, count, insert=insert
        ),
    )
    shifted_merges = view._shift_column_validation_ranges(
        [(None, cell_range) for cell_range in merges],
        anchor,
        count,
        insert=insert,
    )
    view._replace_merged_ranges(
        sheet, [cell_range for _none, cell_range in shifted_merges]
    )


def _meaningful_cell_snapshot(sheet):
    result = {}
    for (row, col), cell in sorted(sheet._cells.items()):
        if not (
            cell.value is not None
            or cell.has_style
            or cell.comment is not None
            or cell.hyperlink is not None
        ):
            continue
        result[(int(row), int(col))] = {
            "value": cell.value,
            "data_type": cell.data_type,
            "style": copy.copy(cell._style),
            "comment": (
                (cell.comment.text, cell.comment.author)
                if cell.comment is not None else None
            ),
            "hyperlink": cell.hyperlink.target if cell.hyperlink is not None else None,
            "number_format": cell.number_format,
        }
    return result


def _metadata_snapshot(sheet):
    dimensions = {
        key: (
            dimension.width,
            bool(dimension.hidden),
            int(dimension.style_id),
            int(dimension.min or 0),
            int(dimension.max or 0),
        )
        for key, dimension in sorted(sheet.column_dimensions.items())
    }
    validations = sorted(
        (
            validation.type,
            validation.formula1,
            str(validation.sqref),
        )
        for validation in sheet.data_validations.dataValidation
    )
    conditionals = sorted(
        (
            str(key.sqref),
            tuple(rule.type for rule in rules),
        )
        for key, rules in sheet.conditional_formatting._cf_rules.items()
    )
    merges = sorted(str(cell_range) for cell_range in sheet.merged_cells.ranges)
    return dimensions, validations, conditionals, merges


def _semantic_snapshot(sheet):
    return _meaningful_cell_snapshot(sheet), _metadata_snapshot(sheet)


def _assert_native_semantics_match(output_sheet, fast_sheet):
    # Excel may normalize style-table ids, default dimensions, and other OOXML
    # encodings on SaveCopyAs.  Compare the user-visible/supported semantics,
    # not private serialization identities.
    fast_values = {
        (row, col): cell.value
        for (row, col), cell in fast_sheet._cells.items()
        if cell.value is not None
    }
    output_values = {
        (row, col): cell.value
        for (row, col), cell in output_sheet._cells.items()
        if cell.value is not None
    }
    assert output_values == fast_values
    fast_cell = fast_sheet["F2"]
    output_cell = output_sheet["F2"]
    assert output_cell.font.bold == fast_cell.font.bold is True
    assert output_cell.fill.fill_type == fast_cell.fill.fill_type == "solid"
    assert str(output_cell.fill.fgColor.rgb)[-6:] == str(fast_cell.fill.fgColor.rgb)[-6:]
    assert output_cell.comment.text == fast_cell.comment.text
    assert output_cell.comment.author == fast_cell.comment.author
    assert output_cell.hyperlink.target == fast_cell.hyperlink.target
    for letter in ("F", "G"):
        output_dimension = output_sheet.column_dimensions[letter]
        fast_dimension = fast_sheet.column_dimensions[letter]
        assert abs(float(output_dimension.width) - float(fast_dimension.width)) < 0.01
        assert bool(output_dimension.hidden) == bool(fast_dimension.hidden)
    assert [
        (item.type, item.formula1, str(item.sqref))
        for item in output_sheet.data_validations.dataValidation
    ] == [
        (item.type, item.formula1, str(item.sqref))
        for item in fast_sheet.data_validations.dataValidation
    ]
    assert [
        (str(key.sqref), tuple(rule.type for rule in rules))
        for key, rules in output_sheet.conditional_formatting._cf_rules.items()
    ] == [
        (str(key.sqref), tuple(rule.type for rule in rules))
        for key, rules in fast_sheet.conditional_formatting._cf_rules.items()
    ]
    assert {str(value) for value in output_sheet.merged_cells.ranges} == {
        str(value) for value in fast_sheet.merged_cells.ranges
    }


def _apply_test_sequence(sheet, *, fast: bool):
    _apply_metadata_aware_shift(
        sheet,
        fast=fast,
        anchor=_INSERT_ANCHOR,
        count=_INSERT_COUNT,
        insert=True,
    )
    _apply_metadata_aware_shift(
        sheet,
        fast=fast,
        anchor=_DELETE_ANCHOR,
        count=_DELETE_COUNT,
        insert=False,
    )


def _column_ops():
    common = {
        "sheet": _SHEET,
        "target_side": "A",
        "source_side": "A",
        "source_physical_cols": [],
        "metadata_scope": list(smt._COLUMN_ACTION_METADATA_SCOPE),
    }
    return [
        {
            **common,
            "kind": "insert_cols",
            "target_logical_slot": _INSERT_ANCHOR,
            "target_physical_anchor": _INSERT_ANCHOR,
            "count": _INSERT_COUNT,
            "batch_id": "fast-parity-insert",
            "action_id": "fast-parity-insert",
            "order": 1,
        },
        {
            **common,
            "kind": "delete_cols",
            "target_logical_slot": _DELETE_ANCHOR,
            "target_physical_anchor": _DELETE_ANCHOR,
            "count": _DELETE_COUNT,
            "batch_id": "fast-parity-delete",
            "action_id": "fast-parity-delete",
            "order": 2,
        },
    ]


def _prepare_fast_and_reference():
    root = make_temp_dir("sow_fast_column_parity_")
    source = os.path.join(root, "source.xlsx")
    _make_fixture(source)
    fast_workbook = load_workbook(source, data_only=False)
    reference_workbook = load_workbook(source, data_only=False)
    fast_sheet = fast_workbook[_SHEET]
    reference_sheet = reference_workbook[_SHEET]
    # Deliberately materialize one far, unstyled, value-empty Cell.  Fast
    # rekeying must preserve that sparse key without rectangular tail growth.
    fast_sheet.cell(10, 10)
    reference_sheet.cell(10, 10)
    assert (10, 10) in fast_sheet._cells and (10, 10) in reference_sheet._cells
    _apply_test_sequence(fast_sheet, fast=True)
    _apply_test_sequence(reference_sheet, fast=False)
    return root, source, fast_workbook, reference_workbook


def test_fast_insert_delete_matches_openpyxl_supported_semantics():
    _root, _source, fast_workbook, reference_workbook = _prepare_fast_and_reference()
    try:
        fast_sheet = fast_workbook[_SHEET]
        reference_sheet = reference_workbook[_SHEET]
        assert _semantic_snapshot(fast_sheet) == _semantic_snapshot(reference_sheet)
        assert (10, 11) in fast_sheet._cells, sorted(fast_sheet._cells)[-8:]
        assert (10, 11) in reference_sheet._cells, sorted(reference_sheet._cells)[-8:]
        assert len(fast_sheet._cells) <= 20, len(fast_sheet._cells)
        assert fast_sheet["G3"].value == "=A3*2"
        assert fast_sheet["F2"].comment.text == "fast-column-comment"
        assert fast_sheet["F2"].hyperlink.target.endswith("/fast-column")
        assert fast_sheet.column_dimensions["F"].width == 23.0
        assert fast_sheet.column_dimensions["F"].hidden is True
        assert [str(item.sqref) for item in fast_sheet.data_validations.dataValidation] == ["F2:G4"]
        assert [str(key.sqref) for key in fast_sheet.conditional_formatting._cf_rules] == ["F2:G4"]
        assert {str(value) for value in fast_sheet.merged_cells.ranges} == {"F5:G5"}
    finally:
        _close_workbook(fast_workbook)
        _close_workbook(reference_workbook)


def test_delete_range_starting_inside_block_shifts_surviving_tail_to_anchor():
    workbook = Workbook()
    reference_workbook = Workbook()
    try:
        for candidate in (workbook.active, reference_workbook.active):
            candidate.title = _SHEET
            candidate["C5"] = "merged-tail"
            candidate.merge_cells("C5:E5")
            validation = DataValidation(type="list", formula1='"x,y"')
            validation.add("C2:E4")
            candidate.add_data_validation(validation)
            candidate.conditional_formatting.add(
                "C2:E4",
                CellIsRule(operator="equal", formula=["1"]),
            )
        _apply_metadata_aware_shift(
            workbook[_SHEET], fast=True, anchor=2, count=2, insert=False
        )
        _apply_metadata_aware_shift(
            reference_workbook[_SHEET], fast=False, anchor=2, count=2, insert=False
        )
        for candidate in (workbook[_SHEET], reference_workbook[_SHEET]):
            assert [
                str(item.sqref) for item in candidate.data_validations.dataValidation
            ] == ["B2:C4"]
            assert [
                str(key.sqref) for key in candidate.conditional_formatting._cf_rules
            ] == ["B2:C4"]
            assert {str(value) for value in candidate.merged_cells.ranges} == {"B5:C5"}
    finally:
        _close_workbook(workbook)
        _close_workbook(reference_workbook)


def test_fast_final_state_matches_excel_native_log_replay():
    _root, source, fast_workbook, reference_workbook = _prepare_fast_and_reference()
    output = os.path.join(os.path.dirname(source), "native-output.xlsx")
    try:
        fast_sheet = fast_workbook[_SHEET]
        operations = _column_ops()
        assert [
            operation["kind"]
            for operation in smt._validated_structural_replay_operations([], operations)
        ] == ["insert_cols", "delete_cols"]
        ok = smt._build_manual_merge_output_with_excel(
            source,
            output,
            {},
            row_ops=[],
            column_ops=operations,
        )
        if not ok:
            return "skipped"
        assert smt._excel_reopen_validate(output), "Excel could not reopen native parity output"
        output_workbook = load_workbook(output, data_only=False)
        try:
            output_sheet = output_workbook[_SHEET]
            _assert_native_semantics_match(output_sheet, fast_sheet)
            assert output_sheet["G3"].value == fast_sheet["G3"].value == "=A3*2"
            assert output_sheet["F2"].value == fast_sheet["F2"].value == "decorated"
            assert output_sheet["F2"].comment.text == fast_sheet["F2"].comment.text
            assert output_sheet["F2"].hyperlink.target == fast_sheet["F2"].hyperlink.target
            assert {str(value) for value in output_sheet.merged_cells.ranges} == {"F5:G5"}
        finally:
            _close_workbook(output_workbook)
    finally:
        _close_workbook(fast_workbook)
        _close_workbook(reference_workbook)


def main():
    tests = (
        test_fast_insert_delete_matches_openpyxl_supported_semantics,
        test_delete_range_starting_inside_block_shifts_surviving_tail_to_anchor,
        test_fast_final_state_matches_excel_native_log_replay,
    )
    passed = 0
    skipped = 0
    for test in tests:
        result = test()
        if result == "skipped":
            skipped += 1
            print(f"SKIP: {test.__name__} (Excel COM unavailable)")
        else:
            passed += 1
            print(f"PASS: {test.__name__}")
    print(f"PASS: fast column mutation parity ({passed} passed, {skipped} skipped)")


if __name__ == "__main__":
    main()

"""OpenSpec 4.3 regressions for convergence and semantic pre-merge."""

from __future__ import annotations

import inspect
import io
import os
import shutil
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

import sow_merge_tool as smt
from _test_temp_utils import make_temp_dir


def _make_base(path: str) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    worksheet.append(["id", "left", "right", "overlap"])
    worksheet.append(["row-2", "base-left", "base-right", "base-overlap"])
    workbook.save(path)
    workbook.close()


def _copy_with_changes(source: str, destination: str, changes: dict[str, object]) -> None:
    shutil.copy2(source, destination)
    if not changes:
        return
    # SVN sidecars end in ``.rN``/``.merge-left.rN``.  openpyxl rejects those
    # names by extension even though their bytes are a valid OOXML package.
    with open(destination, "rb") as handle:
        workbook = load_workbook(io.BytesIO(handle.read()))
    worksheet = workbook["Data"]
    for coordinate, value in changes.items():
        worksheet[coordinate] = value
    workbook.save(destination)
    workbook.close()


def _field(value, *names, default=None):
    if isinstance(value, dict):
        for name in names:
            if name in value:
                return value[name]
        return default
    for name in names:
        if hasattr(value, name):
            return getattr(value, name)
    return default


def _build_context(base, mine, theirs, merged, pristine=None):
    builder = getattr(smt, "build_merge_launch_context", None)
    assert callable(builder), "missing build_merge_launch_context API"
    signature = inspect.signature(builder)
    candidates = {
        "base_path": base,
        "source_base_path": base,
        "mine_path": mine,
        "theirs_path": theirs,
        "merged_path": merged,
        "target_pristine_path": pristine,
    }
    kwargs = {
        name: value
        for name, value in candidates.items()
        if name in signature.parameters
    }
    try:
        return builder(**kwargs)
    except TypeError:
        try:
            return builder(base, mine, theirs, merged, target_pristine_path=pristine)
        except TypeError:
            if pristine is None:
                return builder(base, mine, theirs, merged)
            return builder(base, mine, theirs, merged, pristine)


def _run_analysis(context):
    runner = getattr(smt, "run_startup_merge_analysis", None)
    assert callable(runner), "missing run_startup_merge_analysis API"
    result = runner(context)
    outcome = _field(result, "outcome", "startup_outcome")
    if outcome is None and isinstance(result, tuple):
        for item in result:
            if type(item).__name__ == "StartupMergeOutcome":
                outcome = item
                break
    if outcome is None and type(result).__name__ == "StartupMergeOutcome":
        outcome = result
    assert outcome is not None, f"analysis did not expose StartupMergeOutcome: {result!r}"
    return result, outcome


def _decision_text(outcome) -> str:
    values = (
        _field(
            outcome,
            "automatic_action",
            "action",
            "convergence_action",
            "initialized_from",
            "result_source",
        ),
        _field(outcome, "reason", "summary", "message"),
    )
    return " ".join(str(value) for value in values if value is not None).lower()


def _count(outcome, *names) -> int:
    value = _field(outcome, *names, default=0)
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return len(value or ())


def _candidate_path(analysis, outcome, expected_merged: str) -> str | None:
    for owner in (outcome, analysis):
        value = _field(
            owner,
            "merged_candidate_path",
            "candidate_path",
            "result_path",
            "initialized_path",
        )
        if value:
            return os.fspath(value)
    if os.path.exists(expected_merged):
        return expected_merged
    return None


def _candidate_workbook(analysis, outcome, expected_merged: str):
    for owner in (outcome, analysis):
        workbook = _field(
            owner,
            "merged_workbook",
            "candidate_workbook",
            "result_workbook",
            "workbook",
        )
        if workbook is not None and hasattr(workbook, "worksheets"):
            return workbook, False
    path = _candidate_path(analysis, outcome, expected_merged)
    assert path and os.path.exists(path), (
        "analysis must expose or persist the initialized merged candidate",
        analysis,
        outcome,
    )
    return load_workbook(path, data_only=False, keep_vba=path.lower().endswith(".xlsm")), True


def _assert_candidate_marker(analysis, outcome, merged, expected: str) -> None:
    workbook, owned = _candidate_workbook(analysis, outcome, merged)
    try:
        assert workbook["Data"]["B2"].value == expected
    finally:
        if owned:
            workbook.close()


def _convergence_case(
    *,
    prefix: str,
    base_changes: dict[str, object],
    mine_changes: dict[str, object],
    theirs_changes: dict[str, object],
    expected_role: str,
    expected_value: str,
    branch_names: bool = False,
) -> None:
    root = make_temp_dir(f"sow_convergence_{prefix}_")
    seed = os.path.join(root, "seed.xlsx")
    _make_base(seed)
    if branch_names:
        base = os.path.join(root, "Design.xlsx.merge-left.r20")
        theirs = os.path.join(root, "Design.xlsx.merge-right.r30")
    else:
        base = os.path.join(root, "Design.xlsx.r10")
        theirs = os.path.join(root, "Design.xlsx.r12")
    mine = os.path.join(root, "Design.xlsx")
    merged = os.path.join(root, "merged.xlsx")
    _copy_with_changes(seed, base, base_changes)
    _copy_with_changes(base, mine, mine_changes)
    _copy_with_changes(base, theirs, theirs_changes)

    context = _build_context(base, mine, theirs, merged)
    analysis, outcome = _run_analysis(context)
    decision = _decision_text(outcome)
    assert expected_role in decision, (expected_role, decision, outcome)
    assert _count(outcome, "unresolved_count", "unresolved_conflicts") == 0, outcome
    _assert_candidate_marker(analysis, outcome, merged, expected_value)


def test_whole_workbook_convergence_rules() -> None:
    # Equal inputs are generated with the same edit sequence.  The comparator
    # still proves equality from complete package bytes, never this test setup.
    _convergence_case(
        prefix="mine_base",
        base_changes={},
        mine_changes={},
        theirs_changes={"B2": "theirs-result"},
        expected_role="theirs",
        expected_value="theirs-result",
    )
    _convergence_case(
        prefix="theirs_base",
        base_changes={},
        mine_changes={"B2": "mine-result"},
        theirs_changes={},
        expected_role="mine",
        expected_value="mine-result",
    )
    _convergence_case(
        prefix="mine_theirs",
        base_changes={},
        mine_changes={"B2": "common-result"},
        theirs_changes={"B2": "common-result"},
        expected_role="mine",
        expected_value="common-result",
    )
    _convergence_case(
        prefix="empty_branch_delta",
        base_changes={"B2": "branch-left"},
        mine_changes={"B2": "target-mine"},
        theirs_changes={},
        expected_role="mine",
        expected_value="target-mine",
        branch_names=True,
    )


def test_non_overlapping_changes_are_premerged() -> None:
    root = make_temp_dir("sow_semantic_non_overlap_")
    base = os.path.join(root, "Design.xlsx.r10")
    mine = os.path.join(root, "Design.xlsx")
    theirs = os.path.join(root, "Design.xlsx.r12")
    merged = os.path.join(root, "merged.xlsx")
    _make_base(base)
    _copy_with_changes(base, mine, {"B2": "mine-left"})
    _copy_with_changes(base, theirs, {"C2": "theirs-right"})

    analysis, outcome = _run_analysis(_build_context(base, mine, theirs, merged))
    # ``merged_count`` is an operation/block count, not necessarily one count
    # per changed cell.  The candidate-value assertions below prove that both
    # independent logical changes were actually applied.
    assert _count(outcome, "auto_merged_count", "merged_count", "automatic_count") >= 1, outcome
    assert _count(outcome, "unresolved_count", "unresolved_conflicts") == 0, outcome
    workbook, owned = _candidate_workbook(analysis, outcome, merged)
    try:
        worksheet = workbook["Data"]
        assert worksheet["B2"].value == "mine-left"
        assert worksheet["C2"].value == "theirs-right"
    finally:
        if owned:
            workbook.close()


def _locations(analysis, outcome) -> str:
    values = (
        _field(outcome, "unresolved_cells", "conflicts", "unresolved_locations"),
        _field(analysis, "conflicts"),
        _field(analysis, "conflict_cells_by_sheet", "conflict_map"),
        _field(outcome, "summary", "message", "reason"),
    )
    return " ".join(str(value) for value in values if value is not None).lower()


def test_overlapping_change_remains_navigable_conflict() -> None:
    root = make_temp_dir("sow_semantic_overlap_")
    base = os.path.join(root, "Design.xlsx.r10")
    mine = os.path.join(root, "Design.xlsx")
    theirs = os.path.join(root, "Design.xlsx.r12")
    merged = os.path.join(root, "merged.xlsx")
    _make_base(base)
    _copy_with_changes(base, mine, {"D2": "mine-overlap"})
    _copy_with_changes(base, theirs, {"D2": "theirs-overlap"})

    analysis, outcome = _run_analysis(_build_context(base, mine, theirs, merged))
    assert _count(outcome, "unresolved_count", "unresolved_conflicts") >= 1, outcome
    locations = _locations(analysis, outcome)
    assert "d2" in locations or ("data" in locations and "2" in locations), (
        "overlap must remain available to conflict navigation",
        outcome,
    )


def test_effective_width_mismatch_defers_legacy_premerge_without_scanning() -> None:
    root = make_temp_dir("sow_semantic_width_preflight_")
    # This regression exercises the legacy update-conflict writer.  Branch
    # merges now use source-delta projection and may safely map a target-only
    # logical column instead of taking this physical-width fallback.
    base = os.path.join(root, "Design.xlsx.r10")
    mine = os.path.join(root, "Design.xlsx")
    theirs = os.path.join(root, "Design.xlsx.r12")
    merged = os.path.join(root, "merged.xlsx")
    _make_base(base)
    _copy_with_changes(base, mine, {"B2": "mine-result"})
    _copy_with_changes(base, theirs, {"C2": "theirs-result"})

    with open(mine, "rb") as handle:
        workbook = load_workbook(io.BytesIO(handle.read()))
    try:
        worksheet = workbook["Data"]
        worksheet["E1"] = "mine-only-column"
        worksheet["E2"] = "mine-only-value"
        workbook.save(mine)
    finally:
        workbook.close()

    blocker = smt._legacy_premerge_effective_width_mismatch(
        smt._ensure_xlsx_copy(base),
        smt._ensure_xlsx_copy(mine),
        smt._ensure_xlsx_copy(theirs),
    )
    assert blocker == ("Data", 4, 5, 4), blocker
    context = _build_context(base, mine, theirs, merged)
    with patch.object(
        smt,
        "_scan_three_way_conflicts",
        side_effect=AssertionError("width preflight must defer the exact scan to manual UI"),
    ):
        analysis, outcome = _run_analysis(context)

    assert _field(outcome, "automatic_action") == "manual-review", outcome
    assert not _field(analysis, "conflicts", default=[]), analysis
    candidate = _candidate_path(analysis, outcome, merged)
    assert candidate and os.path.isfile(candidate), outcome
    candidate_book = load_workbook(candidate, data_only=False, read_only=True)
    try:
        assert candidate_book["Data"]["E2"].value == "mine-only-value"
    finally:
        candidate_book.close()
    fallback = " ".join(_field(outcome, "fallback_reasons", default=[])).lower()
    assert "effective populated widths" in fallback and "data" in fallback, fallback


def main() -> None:
    tests = (
        test_whole_workbook_convergence_rules,
        test_non_overlapping_changes_are_premerged,
        test_overlapping_change_remains_navigable_conflict,
        test_effective_width_mismatch_defers_legacy_premerge_without_scanning,
    )
    for test in tests:
        test()
        print(f"PASS: {test.__name__}")
    print(f"PASS: automatic merge semantics ({len(tests)} tests)")


if __name__ == "__main__":
    main()

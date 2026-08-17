import os
import time

from openpyxl import Workbook, load_workbook

import sow_merge_tool as mod
from _test_temp_utils import make_temp_dir


def _make_formula_book(path: str, cached_value: int):
    raw = path + ".raw.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "S1"
    ws["A1"] = "=1"
    wb.save(raw)
    wb.close()
    mod._build_manual_merge_xlsx_via_zip(
        raw,
        path,
        {("S1", 1, 1): "=1"},
        cached_values={("S1", 1, 1): cached_value},
    )


def _pump(app, loops=80):
    for _ in range(loops):
        app.root.update_idletasks()
        app.root.update()
        time.sleep(0.02)


def main():
    root = make_temp_dir("sow_formula_cache_undo_")
    base = os.path.join(root, "base.xlsx")
    mine = os.path.join(root, "mine.xlsx")
    theirs = os.path.join(root, "theirs.xlsx")
    merged = os.path.join(root, "merged.xlsx")
    _make_formula_book(base, 3)
    _make_formula_book(mine, 1)
    _make_formula_book(theirs, 2)

    app = mod.SowMergeApp(mine, theirs, merge_mode=True, merged_path=merged, base_path=base)
    try:
        app.nb.select(app._sheet_containers["S1"])
        _pump(app)
        view = app.sheet_views["S1"]
        if not view._data_ready:
            view.refresh(row_only=None, rescan=True)
        view._show_formula_copy_skip_notice = lambda _count: None

        assert view._copy_selected_row("B2A", override_pair_idx=0, override_cols={1})
        assert app.ws_a_val("S1")["A1"].value == 2
        assert view._copy_selected_row("BASE2A", override_pair_idx=0, override_cols={1})
        assert app.ws_a_val("S1")["A1"].value == 3

        view._undo_last_action()
        assert app.ws_a_val("S1")["A1"].value == 2
        assert app.manual_a_formula_cache_ops[("S1", 1, 1)] == 2

        out = app.build_manual_merge_output_file()
        wb_formula = load_workbook(out, data_only=False)
        wb_value = load_workbook(out, data_only=True)
        try:
            assert wb_formula["S1"]["A1"].value == "=1"
            assert wb_value["S1"]["A1"].value == 2
        finally:
            wb_formula.close()
            wb_value.close()
    finally:
        try:
            app._shutdown_root()
        except Exception:
            pass

    print("SMOKE_FORMULA_CACHE_UNDO_OK")


if __name__ == "__main__":
    main()

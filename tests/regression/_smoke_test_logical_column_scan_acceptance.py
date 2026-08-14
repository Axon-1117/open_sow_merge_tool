"""File-level 2.1/2.2 regression gate for logical column conflict scanning."""

from __future__ import annotations

import os
import tempfile

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

import sow_merge_tool as mod


SHEET = "Data"
BASE_HEADERS = ("A", "B", "C", "D", "E")


def _row_values(headers, row_number, *, edit_col=None, edit_value=None):
    values = {
        name: f"{name.lower()}-{row_number}"
        for name in ("A", "B", "C", "D", "E", "M", "T", "X", "Y")
    }
    if edit_col is not None:
        values[edit_col] = edit_value
    return tuple(values[name] for name in headers)


def _rows(headers, *, edit_col=None, edit_value=None):
    return [
        _row_values(
            headers,
            row_number,
            edit_col=edit_col if row_number == 6 else None,
            edit_value=edit_value,
        )
        for row_number in range(1, 13)
    ]


def _save_book(path, headers, rows):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = SHEET
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    workbook.save(path)
    workbook.close()


def _scan_case(
    base_headers,
    mine_headers,
    theirs_headers,
    *,
    mine_edit_col=None,
    mine_edit_value=None,
    theirs_edit_col=None,
    theirs_edit_value=None,
    insert_mine_row=False,
):
    with tempfile.TemporaryDirectory(prefix="sow-logical-scan-") as temp_dir:
        base_path = os.path.join(temp_dir, "base.xlsx")
        mine_path = os.path.join(temp_dir, "mine.xlsx")
        theirs_path = os.path.join(temp_dir, "theirs.xlsx")
        base_rows = _rows(base_headers)
        mine_rows = _rows(
            mine_headers,
            edit_col=mine_edit_col,
            edit_value=mine_edit_value,
        )
        theirs_rows = _rows(
            theirs_headers,
            edit_col=theirs_edit_col,
            edit_value=theirs_edit_value,
        )
        if insert_mine_row:
            mine_rows.insert(
                4,
                tuple(f"{name.lower()}-new" for name in mine_headers),
            )
        _save_book(base_path, base_headers, base_rows)
        _save_book(mine_path, mine_headers, mine_rows)
        _save_book(theirs_path, theirs_headers, theirs_rows)

        first_conflicts, first_map = mod._scan_three_way_conflicts(
            base_path,
            mine_path,
            theirs_path,
        )
        first_analysis = mod._LAST_THREE_WAY_COLUMN_ANALYSIS[SHEET]
        second_conflicts, second_map = mod._scan_three_way_conflicts(
            base_path,
            mine_path,
            theirs_path,
        )
        second_analysis = mod._LAST_THREE_WAY_COLUMN_ANALYSIS[SHEET]
        assert first_conflicts == second_conflicts
        assert first_map == second_map
        assert first_analysis == second_analysis
        return first_conflicts, first_map, first_analysis


def _conflict_summary(analysis):
    return [
        (conflict.logical_col, conflict.kind, conflict.state)
        for conflict in analysis.structural_conflicts
    ]


def _test_two_way_foreground_worker_and_formula_cache():
    reads = {"cell": 0, "iter_rows": 0}
    original_cell = Worksheet.cell
    original_iter_rows = Worksheet.iter_rows

    def _forbid(kind):
        def _read(*_args, **_kwargs):
            reads[kind] += 1
            raise AssertionError(f"logical comparison called Worksheet.{kind}")

        return _read

    Worksheet.cell = _forbid("cell")
    Worksheet.iter_rows = _forbid("iter_rows")
    try:
        key = mod.ColumnModelCacheKey(SHEET, 3, 7, 11, 13, 17)
        mine = [("A", "B", "C", "D", "E", "F")]
        theirs = [("A", "X", "Y", "B", "D", "E", "F")]
        for row_number in range(1, 13):
            mine.append(tuple(f"{name.lower()}-{row_number}" for name in "ABCDEF"))
            theirs.append((
                f"a-{row_number}",
                f"x-{row_number}",
                f"y-{row_number}",
                f"b-{row_number}",
                f"d-{row_number}",
                "THEIRS-E" if row_number == 6 else f"e-{row_number}",
                f"f-{row_number}",
            ))
        cache = mod.build_logical_column_comparison_cache_2way(
            key,
            mine,
            theirs,
            mine,
            theirs,
            mine_max_col=6,
            theirs_max_col=7,
        )
        assert cache.model.cache_key == key
        assert cache.structural_diff_cols == frozenset((2, 3, 5))
        assert cache == mod.build_logical_column_comparison_cache_2way(
            key,
            mine,
            theirs,
            mine,
            theirs,
            mine_max_col=6,
            theirs_max_col=7,
        )

        worker_result = mod.compare_logical_row_2way(
            cache,
            mine[6],
            theirs[6],
            mine[6],
            theirs[6],
            mine_row=7,
            theirs_row=7,
        )
        assert worker_result.diff_cols == frozenset((2, 3, 5, 7))

        view = object.__new__(mod.SheetView)
        view.column_comparison_cache = cache
        view.max_col = 7
        _parts_a, _parts_b, foreground_cols = (
            view._build_row_parts_and_diff_pair_from_values(
                mine[6],
                theirs[6],
                ra=7,
                rb=7,
                row_a_edit_vals=mine[6],
                row_b_edit_vals=theirs[6],
            )
        )
        assert foreground_cols == set(worker_result.diff_cols)

        mine_values = (("id", "calc", "tail"), (1, None, 10))
        mine_edits = (("id", "calc", "tail"), (1, "=A2+1", 10))
        theirs_values = (("id", "calc", "new", "tail"), (1, None, "x", 10))
        theirs_edits = (("id", "calc", "new", "tail"), (1, "=A3+1", "x", 10))
        formula_cache = mod.build_logical_column_comparison_cache_2way(
            key,
            mine_values,
            theirs_values,
            mine_edits,
            theirs_edits,
            mine_max_col=3,
            theirs_max_col=4,
        )
        equivalent = mod.compare_logical_row_2way(
            formula_cache,
            mine_values[1],
            theirs_values[1],
            mine_edits[1],
            theirs_edits[1],
            mine_row=2,
            theirs_row=3,
        )
        assert equivalent.diff_cols == formula_cache.structural_diff_cols
        actual_formula_edit = list(theirs_edits[1])
        actual_formula_edit[1] = "=A3+2"
        edited = mod.compare_logical_row_2way(
            formula_cache,
            mine_values[1],
            theirs_values[1],
            mine_edits[1],
            tuple(actual_formula_edit),
            mine_row=2,
            theirs_row=3,
        )
        formula_slot = next(
            slot.logical_idx + 1
            for slot in formula_cache.model.slots
            if slot.mine_col == 2 and slot.theirs_col == 2
        )
        assert edited.diff_cols == frozenset(
            set(formula_cache.structural_diff_cols) | {formula_slot}
        )

        newer_key = mod.ColumnModelCacheKey(SHEET, 4, 8, 12, 14, 18)
        newer_cache = mod.build_logical_column_comparison_cache_2way(
            newer_key,
            mine,
            theirs,
            mine,
            theirs,
            mine_max_col=6,
            theirs_max_col=7,
        )
        assert newer_cache.model.cache_key == newer_key
        assert newer_cache.model.cache_key != cache.model.cache_key
        assert newer_cache.model.slots == cache.model.slots
    finally:
        Worksheet.cell = original_cell
        Worksheet.iter_rows = original_iter_rows

    assert reads == {"cell": 0, "iter_rows": 0}


def _test_typed_row_identity_control_characters_do_not_alias():
    """Separator-like text payloads must retain distinct row identities."""
    left_rows = [
        ("id:primary", "payload"),
        ("left\x1eright:typed", "tail\x1fvalue"),
    ]
    right_rows = [
        ("id:primary", "payload"),
        ("left\x1fright:typed", "tail\x1fvalue"),
    ]
    left_sigs, right_sigs = mod._row_signatures_from_unique_column_anchors(
        left_rows,
        right_rows,
        left_width=2,
        right_width=2,
        left_edit_rows=left_rows,
        right_edit_rows=right_rows,
    )
    assert left_sigs[0] == right_sigs[0]
    assert left_sigs[1] != right_sigs[1], (
        "typed row identity collision for text containing \\x1e/\\x1f/colon",
        left_sigs[1],
        right_sigs[1],
    )


def _test_ambiguous_mapping_is_a_structural_conflict():
    key = mod.ColumnModelCacheKey(SHEET, 1, 1)
    base = (("A", None, "Z"), ("a", None, "z"))
    mine = (("A", None, None, "Z"), ("a", None, None, "z"))
    cache = mod.build_logical_column_comparison_cache_3way(
        key,
        mine,
        base,
        base,
        mine,
        base,
        base,
        mine_max_col=4,
        base_max_col=3,
        theirs_max_col=3,
    )
    analysis = mod.classify_logical_columns_3way(cache)
    assert analysis.structural_conflicts
    assert all(
        conflict.kind == "unresolved-mapping"
        and mod.COLUMN_MAPPING_CAUSE_BLANK_COLUMN in conflict.cause_codes
        for conflict in analysis.structural_conflicts
    )
    assert analysis == mod.classify_logical_columns_3way(cache)


def _test_no_structure_preserves_cell_conflict():
    conflicts, conflict_map, analysis = _scan_case(
        BASE_HEADERS,
        BASE_HEADERS,
        BASE_HEADERS,
        mine_edit_col="B",
        mine_edit_value="MINE-B",
        theirs_edit_col="B",
        theirs_edit_value="THEIRS-B",
    )
    assert conflicts == [(SHEET, 7, 2, "MINE-B", "THEIRS-B")]
    assert conflict_map == {SHEET: {7: {2}}}
    assert not analysis.structural_conflicts


def _test_delete_modify_directions_and_both_delete():
    conflicts, conflict_map, analysis = _scan_case(
        ("A", "B", "C"),
        ("A", "C"),
        ("A", "B", "C"),
        theirs_edit_col="B",
        theirs_edit_value="THEIRS-B",
    )
    assert conflicts == [(SHEET, 7, 2, None, "THEIRS-B")]
    assert conflict_map == {SHEET: {7: {2}}}
    assert _conflict_summary(analysis) == [
        (2, "delete-versus-modify", "mine-deleted")
    ]

    conflicts, conflict_map, analysis = _scan_case(
        ("A", "B", "C"),
        ("A", "B", "C"),
        ("A", "C"),
        mine_edit_col="B",
        mine_edit_value="MINE-B",
    )
    assert conflicts == [(SHEET, 7, 2, "MINE-B", None)]
    assert conflict_map == {SHEET: {7: {2}}}
    assert _conflict_summary(analysis) == [
        (2, "delete-versus-modify", "theirs-deleted")
    ]

    conflicts, conflict_map, analysis = _scan_case(
        ("A", "B", "C"),
        ("A", "C"),
        ("A", "C"),
    )
    assert not conflicts and not conflict_map
    assert not analysis.structural_conflicts
    assert any(state.logical_col == 2 and state.state == "both-deleted" for state in analysis.states)


def _test_independent_and_competing_insertions():
    conflicts, conflict_map, analysis = _scan_case(
        ("A", "B", "C", "D"),
        ("A", "M", "B", "C", "D"),
        ("A", "B", "C", "T", "D"),
        theirs_edit_col="C",
        theirs_edit_value="THEIRS-C",
    )
    assert not conflicts and not conflict_map
    assert not analysis.structural_conflicts
    modified = [state for state in analysis.states if state.state == "modified"]
    assert len(modified) == 1
    assert modified[0].theirs_changed and not modified[0].mine_changed

    conflicts, conflict_map, analysis = _scan_case(
        ("A", "C"),
        ("A", "M", "C"),
        ("A", "T", "C"),
    )
    assert len(conflicts) == 2
    assert conflict_map == {SHEET: {1: {2, 3}}}
    assert _conflict_summary(analysis) == [
        (2, "competing-insertion", "inserted"),
        (3, "competing-insertion", "inserted"),
    ]


def _test_row_and_column_structure_compose_before_value_conflicts():
    mixed_headers = ("A", "X", "Y", "B", "D", "E")

    # Each isolated structural dimension is safe.
    conflicts, _conflict_map, analysis = _scan_case(
        BASE_HEADERS,
        mixed_headers,
        BASE_HEADERS,
        theirs_edit_col="E",
        theirs_edit_value="THEIRS-E",
    )
    assert not conflicts and not analysis.structural_conflicts

    conflicts, _conflict_map, analysis = _scan_case(
        BASE_HEADERS,
        BASE_HEADERS,
        BASE_HEADERS,
        theirs_edit_col="E",
        theirs_edit_value="THEIRS-E",
        insert_mine_row=True,
    )
    assert not conflicts and not analysis.structural_conflicts

    # Combining the same row insertion with insert-two/delete-one columns must
    # not regress to physical row/column comparison or cascade header conflicts.
    conflicts, conflict_map, analysis = _scan_case(
        BASE_HEADERS,
        mixed_headers,
        BASE_HEADERS,
        theirs_edit_col="E",
        theirs_edit_value="THEIRS-E",
        insert_mine_row=True,
    )
    assert not conflicts, (conflicts, conflict_map, analysis)
    assert not conflict_map
    assert not analysis.structural_conflicts
    modified = [state for state in analysis.states if state.state == "modified"]
    assert len(modified) == 1
    assert modified[0].theirs_changed and not modified[0].mine_changed


def main():
    _test_two_way_foreground_worker_and_formula_cache()
    _test_typed_row_identity_control_characters_do_not_alias()
    _test_ambiguous_mapping_is_a_structural_conflict()
    _test_no_structure_preserves_cell_conflict()
    _test_delete_modify_directions_and_both_delete()
    _test_independent_and_competing_insertions()
    _test_row_and_column_structure_compose_before_value_conflicts()
    print("SMOKE_TEST_LOGICAL_COLUMN_SCAN_ACCEPTANCE_OK")


if __name__ == "__main__":
    main()

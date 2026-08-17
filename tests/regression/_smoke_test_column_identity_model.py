"""Focused smoke tests for column identity model tasks 1.1 through 1.3."""

from dataclasses import FrozenInstanceError
import time

from openpyxl.worksheet.worksheet import Worksheet

import sow_merge_tool as mod


def _assert_frozen(instance, field_name, value):
    try:
        setattr(instance, field_name, value)
    except FrozenInstanceError:
        return
    raise AssertionError(f"{type(instance).__name__} is mutable")


def _test_immutable_model_and_lookups():
    key = mod.ColumnModelCacheKey("Data", 7, 3, 11, 13, 17)
    exact = mod.ColumnMappingConfidence(1.0, False, "exact-anchor", ("header", "content"))
    unresolved = mod.ColumnMappingConfidence(0.25, True, "duplicate-columns", ("blank",))
    slots = (
        mod.ColumnSlot(0, mine_col=1, base_col=1, theirs_col=1, confidence=exact),
        mod.ColumnSlot(1, mine_col=2, base_col=2, theirs_col=None, state="deleted", confidence=unresolved),
        mod.ColumnSlot(2, mine_col=None, base_col=None, theirs_col=2, state="inserted", confidence=unresolved),
        mod.ColumnSlot(3, mine_col=3, base_col=3, theirs_col=3, confidence=exact),
    )
    block = mod.ColumnBlock(0, (1, 2), "unresolved", unresolved)
    model = mod.ColumnModel.from_slots(
        key,
        slots,
        blocks=(block,),
        confidence=mod.ColumnMappingConfidence(0.75, True, "bounded-ambiguity"),
    )

    assert model.logical_for_physical("mine", 3) == 3
    assert model.logical_for_physical("base", 2) == 1
    assert model.logical_for_physical("theirs", 2) == 2
    assert model.physical_for_logical("A", 2) is None
    assert model.physical_for_logical("B", 2) == 2
    assert model.blocks[0].start_slot_idx == 1
    assert model.blocks[0].end_slot_idx == 2
    assert model.confidence.ambiguous

    _assert_frozen(key, "row_model_version", 8)
    _assert_frozen(slots[0], "mine_col", 9)
    _assert_frozen(block, "state", "retained")
    _assert_frozen(model, "slots", ())
    _assert_frozen(model.mine_physical_to_logical, "entries", ())


def _test_cache_key_versions_change_identity():
    original = mod.ColumnModelCacheKey("Data", 1, 1, 2, 3, 4)
    assert original != mod.ColumnModelCacheKey("Data", 2, 1, 2, 3, 4)
    assert original != mod.ColumnModelCacheKey("Data", 1, 2, 2, 3, 4)
    assert original != mod.ColumnModelCacheKey("Data", 1, 1, 5, 3, 4)

    snapshot = mod.build_column_signature_snapshot(
        original,
        (("header",), (1,)),
        (("header",), (1,)),
    )
    assert snapshot.cache_key == original
    _assert_frozen(snapshot, "cache_key", mod.ColumnModelCacheKey("Data", 9, 9))
    assert isinstance(hash(original), int)

    for bad_version in (True, -0.5, 1.9, "1"):
        try:
            mod.ColumnModelCacheKey("Data", bad_version, 1)
        except TypeError:
            pass
        else:
            raise AssertionError(f"accepted non-integer cache version: {bad_version!r}")
    try:
        mod.ColumnModelCacheKey(["Data"], 1, 1)
    except TypeError:
        pass
    else:
        raise AssertionError("accepted mutable sheet_name")


def _test_model_rejects_mutable_duck_records():
    class MutableSlot:
        def __init__(self):
            self.logical_idx = 0
            self.mine_col = 1
            self.base_col = None
            self.theirs_col = 1

    key = mod.ColumnModelCacheKey("Data", 1, 1)
    mutable = MutableSlot()
    try:
        mod.ColumnModel.from_slots(key, (mutable,))
    except TypeError:
        pass
    else:
        raise AssertionError("ColumnModel retained an externally mutable duck slot")

    slot = mod.ColumnSlot(0, mine_col=1, theirs_col=1)
    try:
        mod.ColumnModel.from_slots(key, (slot,), blocks=(mutable,))
    except TypeError:
        pass
    else:
        raise AssertionError("ColumnModel retained an externally mutable duck block")


def _test_inserted_column_signatures_and_context():
    base_rows = (
        ("id", "name"),
        (1, "alpha"),
        (2, "beta"),
        (3, "gamma"),
    )
    inserted_rows = (
        ("id", "new-field", "name"),
        (1, "x", "alpha"),
        (2, "y", "beta"),
        (3, "z", "gamma"),
    )
    base = mod.build_column_signatures_from_row_cache(base_rows, base_rows)
    inserted = mod.build_column_signatures_from_row_cache(inserted_rows, inserted_rows)

    assert len(base) == 2 and len(inserted) == 3
    assert base[0].intrinsic_key == inserted[0].intrinsic_key
    assert base[1].intrinsic_key == inserted[2].intrinsic_key
    assert inserted[1].intrinsic_key not in {base[0].intrinsic_key, base[1].intrinsic_key}
    assert inserted[0].right_context_key == inserted[1].intrinsic_key
    assert inserted[2].left_context_key == inserted[1].intrinsic_key


def _test_duplicate_and_blank_ambiguity_signals():
    rows = (
        ("left", None, None, "right"),
        (1, None, None, 9),
        (2, "", "", 10),
    )
    signatures = mod.build_column_signatures_from_row_cache(rows, rows)
    left_blank, right_blank = signatures[1], signatures[2]
    assert left_blank.is_blank and right_blank.is_blank
    assert left_blank.intrinsic_key == right_blank.intrinsic_key
    assert left_blank.ambiguous and right_blank.ambiguous
    assert left_blank.ambiguity_reason == "blank-column"
    # Neighbor context remains available to a later conservative aligner even
    # though the intrinsic blank identities are indistinguishable.
    assert left_blank.context_key != right_blank.context_key

    duplicate_rows = (
        ("same", "same"),
        (7, 7),
        (8, 8),
    )
    duplicates = mod.build_column_signatures_from_row_cache(duplicate_rows, duplicate_rows)
    assert duplicates[0].intrinsic_key == duplicates[1].intrinsic_key
    assert all(item.ambiguous for item in duplicates)
    assert all(item.ambiguity_reason == "duplicate-intrinsic-signature" for item in duplicates)


def _test_uncached_formula_and_no_worksheet_reads():
    value_rows = (
        ("formula", "blank"),
        (None, None),
        (None, None),
    )
    edit_rows = (
        ("formula", "blank"),
        ('=""', None),
        ("=A2+1", None),
    )

    original_cell = Worksheet.cell
    original_iter_rows = Worksheet.iter_rows

    def _unexpected_read(*_args, **_kwargs):
        raise AssertionError("column signatures must not read a worksheet")

    Worksheet.cell = _unexpected_read
    Worksheet.iter_rows = _unexpected_read
    try:
        signatures = mod.build_column_signatures_from_row_cache(
            value_rows,
            edit_rows,
            header_row_limit=1,
            representative_row_limit=2,
        )
    finally:
        Worksheet.cell = original_cell
        Worksheet.iter_rows = original_iter_rows

    formula_col, blank_col = signatures
    assert formula_col.non_empty_count == 3
    assert len(formula_col.formula_signals) == 2
    assert not formula_col.is_blank
    assert blank_col.non_empty_count == 1  # its header is a literal signal
    assert formula_col.intrinsic_key != blank_col.intrinsic_key
    assert len(formula_col.header_signals) == 1
    assert len(formula_col.representative_signals) <= 2
    assert len(formula_col.non_empty_pattern) <= 2


def _snapshot(rows, *, row_version=1, column_version=1):
    key = mod.ColumnModelCacheKey(
        "Data",
        row_version,
        column_version,
        mine_edit_version=column_version,
        theirs_edit_version=column_version,
    )
    max_col = max((len(row) for row in rows), default=0)
    return mod.build_column_signature_snapshot(key, rows, rows, max_col=max_col)


def _block_states(result):
    return [(block.state, block.slot_indices) for block in result.ranges]


def _test_two_way_middle_insert_and_delete_ranges():
    left_rows = (
        ("A", "B", "C", "D"),
        (1, 10, 100, 1000),
        (2, 20, 200, 2000),
        (3, 30, 300, 3000),
    )
    right_rows = (
        ("A", "X", "Y", "B", "C", "D"),
        (1, "x1", "y1", 10, 100, 1000),
        (2, "x2", "y2", 20, 200, 2000),
        (3, "x3", "y3", 30, 300, 3000),
    )
    inserted = mod.align_column_signatures_2way(
        _snapshot(left_rows),
        _snapshot(right_rows, column_version=2),
    )
    assert not inserted.has_unresolved
    assert _block_states(inserted) == [
        ("retained", (0,)),
        ("inserted", (1, 2)),
        ("retained", (3, 4, 5)),
    ], _block_states(inserted)
    assert inserted.model.logical_for_physical("mine", 2) == 3
    assert inserted.model.logical_for_physical("theirs", 4) == 3
    assert inserted.model.physical_for_logical("mine", 1) is None
    assert inserted.model.physical_for_logical("theirs", 1) == 2

    deleted = mod.align_column_signatures_2way(
        _snapshot(right_rows, column_version=2),
        _snapshot(left_rows),
    )
    assert not deleted.has_unresolved
    assert _block_states(deleted) == [
        ("retained", (0,)),
        ("deleted", (1, 2)),
        ("retained", (3, 4, 5)),
    ], _block_states(deleted)


def _test_high_confidence_edit_stays_with_logical_column():
    left_rows = tuple(
        [("A", "B", "C")]
        + [(idx, idx * 10, idx * 100) for idx in range(1, 13)]
    )
    right_mutable = [
        ["A", "X", "B", "C"],
        *[[idx, f"x-{idx}", idx * 10, idx * 100] for idx in range(1, 13)],
    ]
    right_mutable[6][2] = 99999  # one real value edit in retained logical B
    right_rows = tuple(tuple(row) for row in right_mutable)
    result = mod.align_column_signatures_2way(
        _snapshot(left_rows),
        _snapshot(right_rows, column_version=2),
    )
    assert not result.has_unresolved, _block_states(result)
    assert [slot.state for slot in result.model.slots] == [
        "retained", "inserted", "retained", "retained"
    ]
    assert result.model.logical_for_physical("mine", 2) == 2
    assert result.model.logical_for_physical("theirs", 3) == 2
    assert result.model.slots[2].confidence.reason == "high-confidence-anchor"


def _test_duplicate_blank_and_low_confidence_are_unresolved():
    duplicate_left = (
        ("A", None, None, "Z"),
        (1, None, None, 9),
        (2, None, None, 10),
    )
    duplicate_right = (
        ("A", None, "Z"),
        (1, None, 9),
        (2, None, 10),
    )
    duplicate = mod.align_column_signatures_2way(
        _snapshot(duplicate_left),
        _snapshot(duplicate_right, column_version=2),
    )
    assert duplicate.has_unresolved and duplicate.used_physical_fallback
    assert _block_states(duplicate) == [
        ("retained", (0,)),
        ("unresolved", (1,)),
        ("unresolved", (2,)),
        ("retained", (3,)),
    ], _block_states(duplicate)
    assert duplicate.model.slots[1].mine_col == 2
    assert duplicate.model.slots[1].theirs_col == 2
    assert duplicate.model.slots[2].mine_col == 3
    assert duplicate.model.slots[2].theirs_col is None

    changed_left = (("A", "Business-B", "Z"), (1, "old-1", 9), (2, "old-2", 10))
    changed_right = (("A", "Different-Q", "Z"), (1, "new-x", 9), (2, "new-y", 10))
    low = mod.align_column_signatures_2way(
        _snapshot(changed_left),
        _snapshot(changed_right, column_version=2),
    )
    repeated = mod.align_column_signatures_2way(
        _snapshot(changed_left),
        _snapshot(changed_right, column_version=2),
    )
    assert low == repeated  # deterministic fallback and block ordering
    assert low.has_unresolved
    assert [slot.state for slot in low.model.slots] == [
        "retained", "unresolved", "retained"
    ]
    unresolved_slot = low.model.slots[1]
    assert unresolved_slot.mine_col == 2 and unresolved_slot.theirs_col == 2
    assert low.model.logical_for_physical("mine", 2) == 1
    assert low.model.logical_for_physical("theirs", 2) == 1


def _test_incompatible_snapshot_uses_whole_physical_fallback():
    rows = (("A", "B"), (1, 2), (3, 4))
    result = mod.align_column_signatures_2way(
        _snapshot(rows, row_version=1),
        _snapshot(rows, row_version=2),
    )
    assert result.has_unresolved
    assert result.fallback_reason == "incompatible-signature-cache"
    assert [slot.state for slot in result.model.slots] == ["unresolved", "unresolved"]
    assert result.model.logical_for_physical("mine", 1) == 0
    assert result.model.logical_for_physical("theirs", 1) == 0


def _test_implicit_ragged_width_forces_safe_fallback():
    left = mod.build_column_signature_snapshot(
        mod.ColumnModelCacheKey("Data", 1, 1),
        (("A",), (1,)),
        (("A",), (1,)),
    )
    right = mod.build_column_signature_snapshot(
        mod.ColumnModelCacheKey("Data", 1, 2),
        (("A", None), (1, None)),
        (("A", None), (1, None)),
    )
    assert not left.width_is_explicit and not right.width_is_explicit
    assert left.max_col == 1 and right.max_col == 2
    result = mod.align_column_signatures_2way(left, right)
    assert result.has_unresolved
    assert result.fallback_reason == "implicit-column-boundary"
    assert all(slot.state == "unresolved" for slot in result.model.slots)

    explicit = mod.build_column_signature_snapshot(
        mod.ColumnModelCacheKey("Data", 1, 3),
        (("A",), (1,)),
        (("A",), (1,)),
        max_col=2,
    )
    assert explicit.width_is_explicit and explicit.max_col == 2


def _test_shifted_formula_identity_stays_unresolved():
    left_values = (
        ("id", "calc", "tail"),
        (1, None, 10),
        (2, None, 20),
        (3, None, 30),
    )
    left_edits = (
        ("id", "calc", "tail"),
        (1, "=A2+1", 10),
        (2, "=A3+1", 20),
        (3, "=A4+1", 30),
    )
    right_values = (
        ("new", "id", "calc", "tail"),
        ("x1", 1, None, 10),
        ("x2", 2, None, 20),
        ("x3", 3, None, 30),
    )
    right_edits = (
        ("new", "id", "calc", "tail"),
        ("x1", 1, "=B2+1", 10),
        ("x2", 2, "=B3+1", 20),
        ("x3", 3, "=B4+1", 30),
    )
    left = mod.build_column_signature_snapshot(
        mod.ColumnModelCacheKey("Data", 1, 1, mine_edit_version=1),
        left_values,
        left_edits,
        max_col=3,
    )
    right = mod.build_column_signature_snapshot(
        mod.ColumnModelCacheKey("Data", 1, 2, theirs_edit_version=2),
        right_values,
        right_edits,
        max_col=4,
    )
    result = mod.align_column_signatures_2way(left, right)
    states = [slot.state for slot in result.model.slots]
    assert states == ["inserted", "retained", "unresolved", "retained"], states
    formula_slot = result.model.slots[2]
    assert formula_slot.mine_col == 2 and formula_slot.theirs_col == 3
    assert formula_slot.confidence.ambiguous
    assert formula_slot.confidence.reason == "low-confidence-physical-fallback"


def _synthetic_signature(index, identity_prefix):
    signal = f"signal-{index}"
    return mod.ColumnSignature(
        physical_col=index + 1,
        row_count=4,
        non_empty_count=4,
        first_non_empty_row=1,
        last_non_empty_row=4,
        header_signals=(signal,),
        representative_signals=((1, signal),),
        non_empty_pattern=(1, 2, 3, 4),
        formula_signals=(),
        intrinsic_key=f"{identity_prefix}{index}",
    )


def _named_signatures(names):
    return tuple(
        mod.ColumnSignature(
            physical_col=index + 1,
            row_count=3,
            non_empty_count=3,
            first_non_empty_row=1,
            last_non_empty_row=3,
            header_signals=(name,),
            representative_signals=((1, name),),
            non_empty_pattern=(1, 2, 3),
            formula_signals=(),
            intrinsic_key=name,
            exact_content_key=name,
        )
        for index, name in enumerate(names)
    )


def _test_three_way_independent_insertions_use_base_boundaries():
    base = _named_signatures(("A", "B", "C", "D"))
    mine = _named_signatures(("A", "M", "B", "C", "D"))
    theirs = _named_signatures(("A", "B", "C", "T", "D"))

    result = mod.align_column_signatures_3way(mine, base, theirs)
    layout = tuple(
        (slot.mine_col, slot.base_col, slot.theirs_col, slot.state)
        for slot in result.model.slots
    )
    assert layout == (
        (1, 1, 1, "retained"),
        (2, None, None, "inserted"),
        (3, 2, 2, "retained"),
        (4, 3, 3, "retained"),
        (None, None, 4, "inserted"),
        (5, 4, 5, "retained"),
    ), layout
    assert not result.has_unresolved
    assert result.model.logical_for_physical("mine", 2) == 1
    assert result.model.logical_for_physical("base", 2) == 2
    assert result.model.logical_for_physical("theirs", 4) == 4
    assert result == mod.align_column_signatures_3way(mine, base, theirs)


def _test_three_way_same_boundary_exact_insertions_share_slots():
    base = _named_signatures(("A", "C"))
    mine = _named_signatures(("A", "X", "Y", "C"))
    theirs = _named_signatures(("A", "X", "Y", "C"))

    result = mod.align_column_signatures_3way(mine, base, theirs)
    layout = tuple(
        (slot.mine_col, slot.base_col, slot.theirs_col, slot.state)
        for slot in result.model.slots
    )
    assert layout == (
        (1, 1, 1, "retained"),
        (2, None, 2, "inserted"),
        (3, None, 3, "inserted"),
        (4, 2, 4, "retained"),
    ), layout
    assert result.fallback_slot_indices == ()
    assert not result.used_physical_fallback
    assert all(
        slot.base_boundary == 1
        and slot.origin_side == "both"
        and slot.confidence.reason == "common-side-insertion"
        for slot in result.model.slots[1:3]
    )


def _test_three_way_same_boundary_partial_match_stays_competing():
    base = _named_signatures(("A", "C"))
    mine = _named_signatures(("A", "X", "M", "C"))
    theirs = _named_signatures(("A", "X", "T", "C"))

    result = mod.align_column_signatures_3way(mine, base, theirs)
    layout = tuple(
        (slot.mine_col, slot.base_col, slot.theirs_col, slot.state)
        for slot in result.model.slots
    )
    assert layout == (
        (1, 1, 1, "retained"),
        (2, None, 2, "inserted"),
        (3, None, None, "unresolved"),
        (None, None, 3, "unresolved"),
        (4, 2, 4, "retained"),
    ), layout
    assert result.fallback_slot_indices == (2, 3)
    assert result.model.slots[1].origin_side == "both"
    assert all(
        result.model.slots[index].confidence.reason
        == "ambiguous-competing-insertions"
        for index in result.fallback_slot_indices
    )


def _test_three_way_base_slots_preserve_deletion_side():
    base = _named_signatures(("A", "B", "C", "D", "E", "F"))
    mine = _named_signatures(("A", "C", "E"))
    theirs = _named_signatures(("A", "B", "D", "E"))

    result = mod.align_column_signatures_3way(mine, base, theirs)
    base_states = tuple(
        (slot.base_col, slot.state)
        for slot in result.model.slots
        if slot.base_col is not None
    )
    assert base_states == (
        (1, "retained"),
        (2, "mine-deleted"),
        (3, "theirs-deleted"),
        (4, "mine-deleted"),
        (5, "retained"),
        (6, "both-deleted"),
    ), base_states
    assert tuple(block.state for block in result.model.blocks) == (
        "retained",
        "mine-deleted",
        "theirs-deleted",
        "mine-deleted",
        "retained",
        "both-deleted",
    )
    assert len(result.model.base_physical_to_logical) == len(base)


def _test_three_way_snapshot_versions_and_compatibility():
    signatures = _named_signatures(("A", "B"))

    def _snapshot(key):
        return mod.ColumnSignatureSnapshot(key, signatures, 8, 24, 2, True)

    mine = _snapshot(mod.ColumnModelCacheKey(
        "Data", 7, 10, mine_edit_version=11
    ))
    base = _snapshot(mod.ColumnModelCacheKey(
        "Data", 7, 20, base_edit_version=13
    ))
    theirs = _snapshot(mod.ColumnModelCacheKey(
        "Data", 7, 30, theirs_edit_version=17
    ))
    result = mod.align_column_signatures_3way(mine, base, theirs)
    assert result.model.cache_key == mod.ColumnModelCacheKey(
        "Data", 7, 30, 11, 13, 17
    )
    _assert_frozen(result, "fallback_reason", "changed")

    incompatible_theirs = _snapshot(mod.ColumnModelCacheKey(
        "Data", 8, 31, theirs_edit_version=18
    ))
    incompatible = mod.align_column_signatures_3way(
        mine, base, incompatible_theirs
    )
    assert incompatible.used_physical_fallback
    assert incompatible.theirs_to_base.fallback_reason == (
        "incompatible-signature-cache"
    )
    assert incompatible.has_unresolved


def _test_alignment_scale_guards():
    exact = tuple(_synthetic_signature(index, "exact-") for index in range(257))
    original_similarity = mod._column_signature_similarity_prepared

    def _unexpected_similarity(*_args):
        raise AssertionError("unique exact anchors entered approximate scoring")

    mod._column_signature_similarity_prepared = _unexpected_similarity
    try:
        started = time.perf_counter()
        exact_result = mod.align_column_signatures_2way(exact[:256], exact[:256])
        exact_elapsed = time.perf_counter() - started
    finally:
        mod._column_signature_similarity_prepared = original_similarity
    assert len(exact_result.model.slots) == 256
    assert all(slot.state == "retained" for slot in exact_result.model.slots)
    assert exact_elapsed < 0.25, f"256 exact columns took {exact_elapsed:.3f}s"

    started = time.perf_counter()
    over_limit_result = mod.align_column_signatures_2way(exact, exact)
    over_limit_elapsed = time.perf_counter() - started
    assert over_limit_result.used_physical_fallback
    assert over_limit_result.fallback_reason == "column-limit-exceeded"
    assert not over_limit_result.anchor_pairs
    assert len(over_limit_result.fallback_slot_indices) == 257
    assert over_limit_elapsed < 0.10, (
        f"257-column fallback took {over_limit_elapsed:.3f}s"
    )
    three_way_over_limit = mod.align_column_signatures_3way(exact, exact, exact)
    assert three_way_over_limit.used_physical_fallback
    assert three_way_over_limit.mine_to_base.fallback_reason == "column-limit-exceeded"
    assert three_way_over_limit.theirs_to_base.fallback_reason == "column-limit-exceeded"
    assert len(three_way_over_limit.model.base_physical_to_logical) == 257
    assert len(three_way_over_limit.fallback_slot_indices) == 257

    # A no-exact-anchor matrix is the conservative O(n^2) path.  This relaxed
    # ceiling catches a return to the former repeated full-score scans while
    # leaving headroom for slower CI hosts.
    left = tuple(_synthetic_signature(index, "left-") for index in range(256))
    right = tuple(_synthetic_signature(index, "right-") for index in range(256))
    started = time.perf_counter()
    approximate_result = mod.align_column_signatures_2way(left, right)
    approximate_elapsed = time.perf_counter() - started
    assert len(approximate_result.model.slots) == 256
    assert all(slot.state == "retained" for slot in approximate_result.model.slots)
    assert approximate_elapsed < 0.50, (
        f"256 approximate columns took {approximate_elapsed:.3f}s"
    )

    seventy = exact[:70]
    started = time.perf_counter()
    seventy_result = mod.align_column_signatures_2way(seventy, seventy)
    seventy_elapsed = time.perf_counter() - started
    assert len(seventy_result.model.slots) == 70
    assert seventy_elapsed < 0.10, f"70 exact columns took {seventy_elapsed:.3f}s"


def main():
    _test_immutable_model_and_lookups()
    _test_cache_key_versions_change_identity()
    _test_model_rejects_mutable_duck_records()
    _test_inserted_column_signatures_and_context()
    _test_duplicate_and_blank_ambiguity_signals()
    _test_uncached_formula_and_no_worksheet_reads()
    _test_two_way_middle_insert_and_delete_ranges()
    _test_high_confidence_edit_stays_with_logical_column()
    _test_duplicate_blank_and_low_confidence_are_unresolved()
    _test_incompatible_snapshot_uses_whole_physical_fallback()
    _test_implicit_ragged_width_forces_safe_fallback()
    _test_shifted_formula_identity_stays_unresolved()
    _test_three_way_independent_insertions_use_base_boundaries()
    _test_three_way_same_boundary_exact_insertions_share_slots()
    _test_three_way_same_boundary_partial_match_stays_competing()
    _test_three_way_base_slots_preserve_deletion_side()
    _test_three_way_snapshot_versions_and_compatibility()
    _test_alignment_scale_guards()
    print("SMOKE_TEST_COLUMN_IDENTITY_MODEL_OK")


if __name__ == "__main__":
    main()

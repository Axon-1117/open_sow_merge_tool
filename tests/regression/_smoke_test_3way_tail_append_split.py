import os
import time

from openpyxl import Workbook, load_workbook

import sow_merge_tool as mod
from _test_temp_utils import make_temp_dir


def _make_book(path: str, rows: list[list[object]]):
    wb = Workbook()
    ws = wb.active
    ws.title = "S1"
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx).value = value
    wb.save(path)
    wb.close()


def _pump(root, loops: int = 20, delay: float = 0.02):
    for _ in range(loops):
        root.update_idletasks()
        root.update()
        time.sleep(delay)


def main():
    root_dir = make_temp_dir("sow_tail_append_split_")
    base = os.path.join(root_dir, "base.xlsx")
    mine = os.path.join(root_dir, "mine.xlsx")
    theirs = os.path.join(root_dir, "theirs.xlsx")
    merged = os.path.join(root_dir, "merged.xlsx")

    base_rows = [["id", "name", "formula"], [1, "x", "=1"], [2, "y", "=1"]]
    mine_rows = [["id", "name", "formula"], [1, "x", "=1"], [2, "y", "=1"], [3, "a", "=1"]]
    theirs_rows = [["id", "name", "formula"], [1, "x", "=1"], [2, "y", "=1"], [4, "b", "=1"]]

    _make_book(base, base_rows)
    _make_book(mine, mine_rows)
    _make_book(theirs, theirs_rows)

    conflicts, conflict_map = mod._scan_three_way_conflicts(base, mine, theirs)
    assert conflicts == [], conflicts
    assert conflict_map == {}, conflict_map

    app = mod.SowMergeApp(mine, theirs, merge_mode=True, merged_path=merged, base_path=base)
    try:
        app.nb.select(app._sheet_containers["S1"])
        _pump(app.root, 60)
        view = app.sheet_views.get("S1")
        assert view is not None, "Expected sheet view to be created"
        if not getattr(view, "_data_ready", False):
            view.refresh(row_only=None, rescan=True)
            _pump(app.root, 10)

        expected_pairs = [(1, 1), (2, 2), (3, 3), (None, 4), (4, None)]
        assert view.row_pairs == expected_pairs, view.row_pairs

        ok = view._copy_selected_row("B2A", override_pair_idx=3)
        assert ok, "Expected B2A on theirs-only tail block to insert into mine"
        _pump(app.root, 10)

        assert view.app.ws_a_val("S1").cell(row=4, column=1).value == 4
        assert view.app.ws_a_val("S1").cell(row=4, column=2).value == "b"
        assert view.app.ws_a_val("S1").cell(row=5, column=1).value == 3
        assert view.app.ws_a_val("S1").cell(row=5, column=2).value == "a"

        out = app.build_manual_merge_output_file()
        wb_out = load_workbook(out, data_only=False)
        try:
            ws_out = wb_out["S1"]
            assert ws_out.cell(row=4, column=1).value == 4
            assert ws_out.cell(row=4, column=2).value == "b"
            assert ws_out.cell(row=5, column=1).value == 3
            assert ws_out.cell(row=5, column=2).value == "a"
            assert ws_out.cell(row=4, column=3).value == "=1"
            assert ws_out.cell(row=5, column=3).value == "=1"
        finally:
            wb_out.close()
    finally:
        try:
            app._shutdown_root()
        except Exception:
            pass

    print("SMOKE_3WAY_TAIL_APPEND_SPLIT_OK")


if __name__ == "__main__":
    main()

"""Headless integration checks against a disposable real Subversion repository.

No TortoiseSVN window is opened.  A small adapter translates the exact
TortoiseProc command contract emitted by BranchSubmitEngine into svn.exe
commands, so repository state and post-commit reconciliation are real while UI
selection remains deterministic.
"""

from __future__ import annotations

import importlib
import os
import shutil
import stat
import subprocess
import sys
import tempfile
import time
from contextlib import contextmanager
from pathlib import Path
from types import SimpleNamespace

from openpyxl import Workbook, load_workbook


def _find_svn_bin(repo_root: Path) -> Path:
    configured = os.environ.get("SOW_SVN_BIN", "").strip()
    candidates = [Path(configured)] if configured else []
    candidates.extend((repo_root / ".local" / "tools").glob("SlikSVN-*/portable/PFiles/bin"))
    for candidate in candidates:
        if (candidate / "svn.exe").is_file() and (candidate / "svnadmin.exe").is_file():
            return candidate.resolve()
    svn = shutil.which("svn")
    svnadmin = shutil.which("svnadmin")
    if svn and svnadmin:
        return Path(svn).resolve().parent
    raise RuntimeError("缺少 svn.exe/svnadmin.exe；先运行 tools/setup_svn_test_runtime.ps1")


REPO_ROOT = Path(__file__).resolve().parents[2]
SVN_BIN = _find_svn_bin(REPO_ROOT)
os.environ["PATH"] = os.fspath(SVN_BIN) + os.pathsep + os.environ.get("PATH", "")

bs = importlib.import_module("sow_merge_tool.branch_submit")
sp = importlib.import_module("sow_merge_tool.svn_status_provider")


SVN = SVN_BIN / "svn.exe"
SVNADMIN = SVN_BIN / "svnadmin.exe"
SVNLOOK = SVN_BIN / "svnlook.exe"


@contextmanager
def _temporary_test_dir(parent: Path):
    path = Path(tempfile.mkdtemp(prefix="sow-svn-e2e-", dir=parent))
    try:
        yield path
    finally:
        last_error: OSError | None = None

        def clear_readonly(function, target, _excinfo):
            os.chmod(target, stat.S_IWRITE)
            function(target)

        for _attempt in range(20):
            try:
                shutil.rmtree(path, onexc=clear_readonly)
                last_error = None
                break
            except FileNotFoundError:
                last_error = None
                break
            except OSError as exc:
                last_error = exc
                time.sleep(0.15)
        if last_error is not None:
            print(f"WARN deferred test cleanup: {path}: {last_error}", file=sys.stderr)


def _run(*args: os.PathLike | str, cwd: Path | None = None, check: bool = True):
    command = [os.fspath(arg) for arg in args]
    result = subprocess.run(
        command,
        cwd=os.fspath(cwd) if cwd else None,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        timeout=60,
        check=False,
    )
    if check and result.returncode != 0:
        raise AssertionError(
            f"command failed ({result.returncode}): {command}\n"
            f"stdout={result.stdout}\nstderr={result.stderr}"
        )
    return result


def _book(path: Path, value: int, *, extra: str = "") -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["id", "value", "extra"])
    ws.append(["row", value, extra])
    wb.save(path)
    wb.close()


def _value(path: Path) -> int:
    wb = load_workbook(path, read_only=True, data_only=False)
    try:
        return int(wb["Data"]["B2"].value)
    finally:
        wb.close()


def _repo_url(path: Path) -> str:
    return path.resolve().as_uri()


def _create_repository(root: Path, *, status_lab: bool = False) -> tuple[Path, str]:
    root.mkdir(parents=True, exist_ok=True)
    repository = root / "repository"
    seed = root / "seed"
    _run(SVNADMIN, "create", repository)
    if status_lab:
        lab = seed / "sheets" / "status_lab"
        for name in (
            "Normal.xlsx",
            "Missing.xlsx",
            "Deleted.xlsx",
            "Replaced.xlsx",
            "Modified.xlsx",
            "Conflict.xlsx",
            "Obstructed.xlsx",
            "Lock.xlsx",
            "Move.xlsx",
        ):
            _book(lab / name, 1)
        _book(lab / "switchdir" / "Switch.xlsx", 1)
        _book(seed / "sheets" / "switch_source" / "switchdir" / "Switch.xlsx", 2)
        _book(seed / "external_source" / "External.xlsx", 1)
    else:
        for branch in ("develop", "release", "sandbox", "master"):
            config = seed / "sheets" / branch / "config"
            modify_extra = "source-base" if branch == "develop" else "target-only" if branch == "release" else "common"
            _book(config / "Modify.xlsx", 1, extra=modify_extra)
            _book(config / "Delete.xlsx", 2, extra="common")
            _book(config / "Missing.xlsx", 3, extra="common")
            _book(config / "PartialA.xlsx", 4, extra="common")
            _book(config / "PartialB.xlsx", 5, extra="common")
            _book(config / "Dirty.xlsx", 6, extra="common")
    url = _repo_url(repository)
    for child in seed.iterdir():
        _run(
            SVN,
            "import",
            child,
            f"{url}/{child.name}",
            "-m",
            "初始化测试仓库",
            "--username",
            "tester",
            "--non-interactive",
        )
    return repository, url


def _checkout(url: str, destination: Path) -> None:
    _run(SVN, "checkout", url, destination, "--ignore-externals", "--non-interactive")


def _decode_tortoise_paths(args: list[str]) -> list[str]:
    for value in args:
        if value.startswith("/pathfile:"):
            return Path(value.split(":", 1)[1]).read_bytes().decode("utf-16-le").splitlines()
        if value.startswith("/path:"):
            return [value.split(":", 1)[1]]
    return []


def _read_log_message(args: list[str]) -> str:
    for value in args:
        if value.startswith("/logmsgfile:"):
            return Path(value.split(":", 1)[1]).read_text(encoding="utf-8-sig")
    return "无界面集成测试"


def _text_status(path: str) -> str:
    result = _run(SVN, "status", path, check=False)
    if result.returncode != 0:
        return "error"
    line = next((line for line in result.stdout.splitlines() if line.strip()), "")
    return line[:1] if line else " "


class HeadlessTortoiseRunner:
    def __init__(self, *, reported_commit_code: int = 0, first_commit_limit: int | None = None):
        self.reported_commit_code = reported_commit_code
        self.first_commit_limit = first_commit_limit
        self.commit_calls = 0
        self.commands: list[tuple[str, tuple[str, ...]]] = []

    def __call__(self, args: list[str], **_kwargs):
        command = next(value.split(":", 1)[1] for value in args if value.startswith("/command:"))
        paths = _decode_tortoise_paths(args)
        self.commands.append((command, tuple(paths)))
        if command == "update":
            result = _run(
                SVN,
                "update",
                *paths,
                "--ignore-externals",
                "--non-interactive",
                check=False,
            )
            return SimpleNamespace(returncode=result.returncode)
        if command == "revert":
            result = _run(SVN, "revert", *paths, "--non-interactive", check=False)
            return SimpleNamespace(returncode=result.returncode)
        if command != "commit":
            return SimpleNamespace(returncode=0)

        self.commit_calls += 1
        selected = list(paths)
        if self.commit_calls == 1 and self.first_commit_limit is not None:
            selected = selected[: self.first_commit_limit]
        for path in selected:
            status = _text_status(path)
            if status == "?":
                _run(SVN, "add", path, "--parents", "--non-interactive")
            elif status == "!":
                _run(SVN, "delete", path, "--force", "--non-interactive")
        result = _run(
            SVN,
            "commit",
            *selected,
            "-m",
            _read_log_message(args),
            "--username",
            "tester",
            "--non-interactive",
            check=False,
        )
        if result.returncode != 0:
            return SimpleNamespace(returncode=result.returncode)
        return SimpleNamespace(returncode=self.reported_commit_code)


def run_all_naturally_reproducible_svn_states(root: Path) -> None:
    _repository, url = _create_repository(root, status_lab=True)
    wc = root / "status-wc"
    peer = root / "status-peer"
    _checkout(f"{url}/sheets", wc)
    _checkout(f"{url}/sheets/status_lab", peer)
    lab = wc / "status_lab"

    _book(lab / "Unversioned.xlsx", 10)
    _book(lab / "Added.xlsx", 11)
    _run(SVN, "add", lab / "Added.xlsx")
    (lab / "Missing.xlsx").unlink()
    _run(SVN, "delete", lab / "Deleted.xlsx")
    _run(SVN, "delete", lab / "Replaced.xlsx")
    _book(lab / "Replaced.xlsx", 12)
    _run(SVN, "add", lab / "Replaced.xlsx")
    _book(lab / "Modified.xlsx", 13)
    _run(SVN, "changelist", "ignore-on-commit", lab / "Modified.xlsx")
    _run(SVN, "propset", "test:property", "changed", lab / "Normal.xlsx")
    _run(SVN, "lock", lab / "Lock.xlsx", "-m", "状态锁测试", "--username", "tester")
    _run(SVN, "move", lab / "Move.xlsx", lab / "Moved.xlsx")

    _book(peer / "Conflict.xlsx", 20)
    _run(SVN, "commit", peer / "Conflict.xlsx", "-m", "制造远端冲突", "--username", "peer")
    _book(lab / "Conflict.xlsx", 21)
    conflict_update = _run(
        SVN,
        "update",
        lab / "Conflict.xlsx",
        "--accept",
        "postpone",
        "--non-interactive",
        check=False,
    )
    assert conflict_update.returncode != 0 or "conflict" in (
        conflict_update.stdout + conflict_update.stderr
    ).lower()

    obstructed = lab / "Obstructed.xlsx"
    obstructed.unlink()
    obstructed.mkdir()
    _run(SVN, "switch", f"{url}/sheets/switch_source/switchdir", lab / "switchdir", "--ignore-ancestry")

    ignored = lab / "Ignored.xlsx"
    _run(SVN, "propset", "svn:ignore", ignored.name, lab)
    _book(ignored, 30)
    external_parent = lab / "external_parent"
    external_parent.mkdir()
    _run(SVN, "add", external_parent)
    _run(SVN, "propset", "svn:externals", f"{url}/external_source external_wc", external_parent)
    _run(SVN, "update", external_parent)
    assert (external_parent / "external_wc" / "External.xlsx").is_file()

    records = sp.scan_status(os.fspath(lab))
    by_name = {Path(record.path).name: record for record in records}
    expected = {
        "Normal.xlsx": "normal",  # property-only change is carried separately
        "Unversioned.xlsx": "unversioned",
        "Added.xlsx": "added",
        "Missing.xlsx": "missing",
        "Deleted.xlsx": "deleted",
        "Replaced.xlsx": "replaced",
        "Modified.xlsx": "modified",
        "Conflict.xlsx": "conflicted",
        "Obstructed.xlsx": "obstructed",
    }
    for name, status in expected.items():
        assert name in by_name, (name, sorted(by_name))
        assert by_name[name].node_status == status, (name, by_name[name])
    assert "Ignored.xlsx" not in by_name
    assert by_name["Modified.xlsx"].changelist == "ignore-on-commit"
    assert by_name["Conflict.xlsx"].conflicted
    assert any(record.switched for record in records)
    assert by_name["Lock.xlsx"].lock_owner == "tester"
    assert any(record.moved_from or record.moved_to for record in records)
    assert any(record.prop_status == "modified" for record in records)

    changes = bs.scan_changes(os.fspath(wc), "status_lab", os.fspath(lab))
    change_by_name = {Path(item.path).name: item for item in changes}
    assert all("external_wc" not in Path(item.path).parts for item in changes)
    assert not change_by_name["Conflict.xlsx"].selectable
    assert not change_by_name["Obstructed.xlsx"].selectable
    assert not change_by_name["Replaced.xlsx"].selectable
    assert not change_by_name["Modified.xlsx"].checked
    assert change_by_name["Missing.xlsx"].checked


def run_real_repository_multi_branch_flow(root: Path) -> None:
    repository, url = _create_repository(root)
    wc = root / "flow-wc"
    verify_wc = root / "verify-wc"
    _checkout(f"{url}/sheets", wc)
    develop = wc / "develop" / "config"
    release = wc / "release" / "config"

    _book(develop / "Modify.xlsx", 101, extra="source-base")
    _book(develop / "Added.xlsx", 102)
    _run(SVN, "add", develop / "Added.xlsx")
    _book(develop / "Unversioned.xlsx", 103)
    _run(SVN, "delete", develop / "Delete.xlsx")
    (develop / "Missing.xlsx").unlink()
    missing_target_hash = bs._sha256(os.fspath(release / "Missing.xlsx"))

    selected_names = {
        "Modify.xlsx",
        "Added.xlsx",
        "Unversioned.xlsx",
        "Delete.xlsx",
        "Missing.xlsx",
    }
    items = [
        item
        for item in bs.scan_changes(os.fspath(wc), "develop", os.fspath(develop))
        if Path(item.path).name in selected_names
    ]
    assert {Path(item.path).name for item in items} == selected_names
    for item in items:
        item.checked = True

    runner = HeadlessTortoiseRunner(reported_commit_code=2)
    engine = bs.BranchSubmitEngine(os.fspath(wc), runner=runner)
    batch = engine.preflight(
        "develop",
        ["release"],
        items,
        "真实仓库无界面多分支提交",
        scope_path=os.fspath(develop),
    )
    plans = {Path(plan.relative_path).name: plan for plan in batch.files}
    assert plans["Missing.xlsx"].operation == bs.SOURCE_ONLY_MISSING
    assert plans["Missing.xlsx"].actions["release"].state == "excluded"
    assert bs._sha256(os.fspath(release / "Missing.xlsx")) == missing_target_hash

    result = engine.commit(batch)
    assert result.source_status == "committed", result.error
    assert result.target_status["release"] == "committed", result.error
    assert runner.commit_calls == 2
    assert all(command != "TortoiseProc" for command, _paths in runner.commands)

    _checkout(f"{url}/sheets", verify_wc)
    verify_develop = verify_wc / "develop" / "config"
    verify_release = verify_wc / "release" / "config"
    assert _value(verify_release / "Modify.xlsx") == 101
    verify_book = load_workbook(verify_release / "Modify.xlsx", read_only=True, data_only=False)
    try:
        assert verify_book["Data"]["C2"].value == "target-only"
    finally:
        verify_book.close()
    assert _value(verify_release / "Added.xlsx") == 102
    assert _value(verify_release / "Unversioned.xlsx") == 103
    assert not (verify_release / "Delete.xlsx").exists()
    assert (verify_release / "Missing.xlsx").is_file()
    assert _value(verify_release / "Missing.xlsx") == 3
    assert not (verify_develop / "Missing.xlsx").exists()

    youngest = int(_run(SVNLOOK, "youngest", repository).stdout.strip())
    changed = _run(SVNLOOK, "changed", "-r", str(youngest), repository).stdout
    assert "sheets/release/config/Missing.xlsx" not in changed
    assert "sheets/release/config/Modify.xlsx" in changed


def run_real_repository_partial_selection_and_dirty_target(root: Path) -> None:
    _repository, url = _create_repository(root)
    wc = root / "partial-wc"
    _checkout(f"{url}/sheets", wc)
    develop = wc / "develop" / "config"
    release = wc / "release" / "config"

    _book(develop / "Dirty.xlsx", 61)
    _book(release / "Dirty.xlsx", 62)
    dirty_item = next(
        item
        for item in bs.scan_changes(os.fspath(wc), "develop", os.fspath(develop))
        if Path(item.path).name == "Dirty.xlsx"
    )
    engine = bs.BranchSubmitEngine(os.fspath(wc), runner=HeadlessTortoiseRunner())
    dirty_hash = bs._sha256(os.fspath(release / "Dirty.xlsx"))
    try:
        engine.preflight(
            "develop", ["release"], [dirty_item], "目标脏文件",
            scope_path=os.fspath(develop),
        )
    except RuntimeError as exc:
        assert "更新前工作副本检查失败" in str(exc)
    else:
        raise AssertionError("真实目标脏文件未阻止 Update")
    assert bs._sha256(os.fspath(release / "Dirty.xlsx")) == dirty_hash
    _run(SVN, "revert", develop / "Dirty.xlsx", release / "Dirty.xlsx")

    _book(develop / "PartialA.xlsx", 71)
    _book(develop / "PartialB.xlsx", 72)
    items = [
        item
        for item in bs.scan_changes(os.fspath(wc), "develop", os.fspath(develop))
        if Path(item.path).name in {"PartialA.xlsx", "PartialB.xlsx"}
    ]
    runner = HeadlessTortoiseRunner(first_commit_limit=1)
    engine = bs.BranchSubmitEngine(os.fspath(wc), runner=runner)
    batch = engine.preflight(
        "develop", ["release"], items, "部分勾选",
        scope_path=os.fspath(develop),
    )
    result = engine.commit(batch)
    assert result.source_status == "partial", result.error
    assert result.superseded_by.endswith("-committed")
    assert result.target_status["release"] == "ready"
    assert _value(release / "PartialA.xlsx") == 4
    assert _value(release / "PartialB.xlsx") == 5


def main() -> None:
    test_root = Path(os.environ.get("SOW_TEST_TMPDIR") or tempfile.gettempdir())
    test_root.mkdir(parents=True, exist_ok=True)
    with _temporary_test_dir(test_root) as root:
        run_all_naturally_reproducible_svn_states(root / "states")
        print("PASS real SVN status inventory")
    with _temporary_test_dir(test_root) as root:
        run_real_repository_multi_branch_flow(root / "flow")
        print("PASS real SVN multi-branch flow")
    with _temporary_test_dir(test_root) as root:
        run_real_repository_partial_selection_and_dirty_target(root / "partial")
        print("PASS real SVN partial selection and dirty-target guard")
    print(f"headless SVN end-to-end tests passed with {SVN.name} { _run(SVN, '--version', '--quiet').stdout.strip() }")


if __name__ == "__main__":
    main()

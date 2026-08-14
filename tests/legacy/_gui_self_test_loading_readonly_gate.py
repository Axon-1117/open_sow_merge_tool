"""Focused GUI regression for loading/read-only mutation gates.

The test intentionally calls command/event handlers directly.  A disabled Tk
button is not enough protection because row headers, the comparison panel,
keyboard commands, and app-level save commands can invoke the same mutations.
"""

from __future__ import annotations

import copy
import hashlib
import os
import threading
import time
from contextlib import contextmanager
from types import SimpleNamespace

from openpyxl import Workbook

import sow_merge_tool as smt
from _gui_self_test_logical_column_actions import _worksheet_snapshot
from _test_temp_utils import make_temp_dir


def _make_book(path: str, rows) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    for row in rows:
        worksheet.append(list(row))
    workbook.save(path)
    workbook.close()


def _pump(root, seconds: float = 0.05) -> None:
    deadline = time.time() + seconds
    while time.time() < deadline:
        root.update_idletasks()
        root.update()
        time.sleep(0.005)


def _wait_until(root, predicate, message: str, timeout: float = 15.0) -> None:
    deadline = time.time() + timeout
    while time.time() < deadline:
        _pump(root, 0.025)
        if predicate():
            return
    raise AssertionError(message)


def _open_ready_view():
    root_dir = make_temp_dir("sow_loading_readonly_gate_")
    mine = os.path.join(root_dir, "mine.xlsx")
    theirs = os.path.join(root_dir, "theirs.xlsx")
    _make_book(
        mine,
        (
            ("id", "value", "note"),
            ("row-2", "mine-2", "same"),
            ("row-3", "mine-3", "same"),
            ("row-4", "same-4", "same"),
        ),
    )
    _make_book(
        theirs,
        (
            ("id", "value", "note"),
            ("row-2", "theirs-2", "same"),
            ("row-3", "theirs-3", "same"),
            ("row-4", "same-4", "same"),
        ),
    )
    app = smt.SowMergeApp(mine, theirs)
    app.root.deiconify()
    app.root.geometry("940x760")
    app.nb.select(app._sheet_containers["Data"])

    _wait_until(
        app.root,
        lambda: (
            app.sheet_views.get("Data") is not None
            and bool(getattr(app.sheet_views["Data"], "_data_ready", False))
            and app._edit_workbooks_ready()
        ),
        "small test workbook did not finish initial loading",
    )
    view = app.sheet_views["Data"]

    # Establish one exact, formula-aware full-view baseline.  The files are
    # tiny, so this explicit foreground setup is bounded and deterministic.
    view.only_diff_var.set(0)
    view._last_only_diff_value = 0
    view._suppress_bg_apply = True
    try:
        view.refresh(row_only=None, rescan=True)
    finally:
        view._suppress_bg_apply = False
    view._row_model_exact = True
    view._cache_formula_aware = True
    view._lifecycle_error = None
    view._refresh_interaction_gate()
    _pump(app.root)
    assert view._lifecycle_state == "READY", view._lifecycle_state
    return app, view, (mine, theirs)


def _file_digest(path: str) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as handle:
        for chunk in iter(lambda: handle.read(64 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _mutation_snapshot(app, view, source_paths):
    """Capture every persistent/in-memory write surface named by the contract."""
    operation_attrs = (
        "manual_a_cell_ops",
        "manual_b_cell_ops",
        "manual_a_formula_cache_ops",
        "manual_b_formula_cache_ops",
        "manual_a_row_ops",
        "manual_b_row_ops",
        "manual_a_column_ops",
        "manual_b_column_ops",
        "manual_sheet_ops",
        "undo_stack",
    )
    return {
        # Read the owned workbooks directly: EDIT_LOADING is simulated by
        # overriding the readiness predicate, and calling ws_*_edit() there
        # would intentionally attempt the loading fallback.
        "mine_edit": _worksheet_snapshot(app._wb_a_edit["Data"]),
        "theirs_edit": _worksheet_snapshot(app._wb_b_edit["Data"]),
        "mine_value": _worksheet_snapshot(app._wb_a_val["Data"]),
        "theirs_value": _worksheet_snapshot(app._wb_b_val["Data"]),
        "operations": {
            name: copy.deepcopy(getattr(app, name))
            for name in operation_attrs
        },
        "modified": (
            bool(app.modified_a),
            bool(app.modified_b),
            frozenset(app.modified_sheets_a),
            frozenset(app.modified_sheets_b),
        ),
        "touched_rows": frozenset(view.touched_rows),
        "row_pairs": tuple(view.row_pairs),
        "pair_diff_cols": tuple(
            sorted(
                (int(pair_idx), tuple(sorted(cols)))
                for pair_idx, cols in view.pair_diff_cols.items()
            )
        ),
        "pair_base_diff_cols": tuple(
            sorted(
                (int(pair_idx), tuple(sorted(cols)))
                for pair_idx, cols in view.pair_base_diff_cols.items()
            )
        ),
        "source_hashes": tuple(_file_digest(path) for path in source_paths),
    }


def _assert_mutation_unchanged(before, app, view, source_paths, action: str) -> None:
    after = _mutation_snapshot(app, view, source_paths)
    assert after == before, f"{action} mutated state while Sheet was read-only"


def _event_for_text_index(widget, index: str):
    widget.see(index)
    widget.update_idletasks()
    box = widget.bbox(index)
    assert box is not None, f"no visible bbox for {widget!r} at {index}"
    x, y, width, height = box
    return SimpleNamespace(
        x=int(x + max(1, width // 2)),
        y=int(y + max(1, height // 2)),
    )


def _prepare_direct_handler_events(view):
    pair_idx = 1
    main_line = int(view.row_to_line[pair_idx])
    view._select_line(main_line)
    view._set_main_selected_cell(main_line, 2)
    view._cursor_cmp_sel_col = 2
    view._cursor_cmp_sel_line = 2
    view.hover_pair_idx = pair_idx
    view.hover_col_idx = 2
    view.hover_side = "B"
    view._last_cursor_cmp_pair_idx = pair_idx
    view._update_cursor_lines()
    _pump(view.root)

    row_header_event = _event_for_text_index(view.right_ln, f"{main_line}.0")
    line_text = view.cursor_cmp.get("2.0", "2.end")
    start, end = view._spans_for_line(line_text)[2]
    char_pos = start + 1 if end - start > 1 else start
    comparison_event = _event_for_text_index(view.cursor_cmp, f"2.{char_pos}")
    return pair_idx, row_header_event, comparison_event


@contextmanager
def _forced_lifecycle(view, state: str):
    app = view.app
    original = {
        "_data_ready": view._data_ready,
        "_row_model_exact": view._row_model_exact,
        "_cache_formula_aware": view._cache_formula_aware,
        "_lifecycle_error": view._lifecycle_error,
        "_lifecycle_canceled": view._lifecycle_canceled,
        "_suppress_bg_apply": view._suppress_bg_apply,
        "_edit_workbooks_ready": app._edit_workbooks_ready,
        "_interactive_event_set": app._interactive_action_event.is_set(),
    }
    view.only_diff_var.set(0)
    view._last_only_diff_value = 0
    view._data_ready = True
    view._row_model_exact = True
    view._cache_formula_aware = True
    view._lifecycle_error = None
    view._lifecycle_canceled = False
    view._suppress_bg_apply = True
    app._edit_workbooks_ready = original["_edit_workbooks_ready"]

    if state == "LOADING":
        view._data_ready = False
    elif state == "DIFFING":
        view._row_model_exact = False
    elif state == "EDIT_LOADING":
        app._edit_workbooks_ready = lambda: False
    elif state == "FAILED":
        view._lifecycle_error = "forced failure"
    elif state == "CANCELED":
        view._lifecycle_canceled = True
    elif state == "BUSY":
        app._interactive_action_event.set()
    else:
        raise AssertionError(f"unsupported forced lifecycle: {state}")

    view._refresh_interaction_gate()
    assert view._lifecycle_state == state, (
        state,
        view._lifecycle_state,
        view._mutation_block_message(),
    )
    try:
        yield
    finally:
        view._data_ready = original["_data_ready"]
        view._row_model_exact = original["_row_model_exact"]
        view._cache_formula_aware = original["_cache_formula_aware"]
        view._lifecycle_error = original["_lifecycle_error"]
        view._lifecycle_canceled = original["_lifecycle_canceled"]
        view._suppress_bg_apply = original["_suppress_bg_apply"]
        app._edit_workbooks_ready = original["_edit_workbooks_ready"]
        if original["_interactive_event_set"]:
            app._interactive_action_event.set()
        else:
            app._interactive_action_event.clear()
        view._refresh_interaction_gate()


def _assert_mutation_widgets_locked(view) -> None:
    for name in (
        "only_diff_cb",
        "force_align_cb",
        "use_left_btn",
        "use_right_btn",
        "undo_btn",
        "save_a_btn",
        "save_b_btn",
    ):
        widget = getattr(view, name, None)
        if widget is not None:
            assert str(widget.cget("state")) == "disabled", (
                name,
                widget.cget("state"),
                view._lifecycle_state,
            )


def test_non_ready_direct_handlers_have_no_write_side_effects():
    app, view, source_paths = _open_ready_view()
    original_confirm = app._confirm_overwrite
    try:
        pair_idx, row_header_event, comparison_event = _prepare_direct_handler_events(view)
        # A sentinel proves that a blocked undo handler does not merely no-op
        # because the stack happens to be empty.
        app.undo_stack.append({"kind": "gate-test-sentinel", "sheet": "Data"})
        app._confirm_overwrite = lambda *_args, **_kwargs: (
            (_ for _ in ()).throw(AssertionError("blocked save reached confirmation"))
        )

        actions = (
            ("toolbar row action", lambda: view._run_copy_action_by_mode("B2A")),
            (
                "direct row primitive",
                lambda: view._copy_selected_row(
                    "B2A",
                    override_pair_idx=pair_idx,
                    override_cols={2},
                ),
            ),
            ("direct region handler", lambda: view._copy_selected_region("B2A")),
            (
                "direct cell handler",
                lambda: view._copy_single_cell_by_pair(pair_idx, "B2A", 2),
            ),
            (
                "forged guarded cell primitive",
                lambda: view._copy_single_cell_by_pair(
                    pair_idx,
                    "B2A",
                    2,
                    _guarded=True,
                ),
            ),
            (
                "forged batch row primitive",
                lambda: view._copy_selected_row(
                    "B2A",
                    override_pair_idx=pair_idx,
                    override_cols={2},
                    suppress_refresh=True,
                ),
            ),
            (
                "row-header event handler",
                lambda: view._on_row_header_click(
                    view.right_ln,
                    row_header_event,
                    "B2A",
                ),
            ),
            (
                "C-area double-click handler",
                lambda: view._on_cursor_cmp_double_click(comparison_event),
            ),
            ("column action handler", lambda: view._on_column_action_button("B")),
            ("undo handler", view._undo_last_action),
            ("save mine handler", app.save_a_inplace),
            ("save theirs handler", app.save_b_inplace),
        )

        for state in (
            "LOADING",
            "DIFFING",
            "EDIT_LOADING",
            "FAILED",
            "CANCELED",
            "BUSY",
        ):
            with _forced_lifecycle(view, state):
                _assert_mutation_widgets_locked(view)
                before = _mutation_snapshot(app, view, source_paths)
                for action_name, action in actions:
                    assert view._derive_lifecycle_state() == state, (
                        state,
                        action_name,
                        view._derive_lifecycle_state(),
                    )
                    action()
                    _pump(app.root, 0.01)
                    _assert_mutation_unchanged(
                        before,
                        app,
                        view,
                        source_paths,
                        f"{state}: {action_name}",
                    )

        assert view._lifecycle_state == "READY", view._lifecycle_state
        assert view._guard_mutation_ready("READY probe", notify=False)
        for name in ("use_left_btn", "use_right_btn", "undo_btn"):
            assert str(getattr(view, name).cget("state")) == "normal", name

        # Remove the artificial undo entry, then prove READY genuinely unlocks
        # the same write path that was rejected in all three prior states.
        app.undo_stack.clear()
        assert app.ws_a_edit("Data").cell(2, 2).value == "mine-2"
        view._copy_single_cell_by_pair(pair_idx, "B2A", 2)
        assert app.ws_a_edit("Data").cell(2, 2).value == "theirs-2"
        assert app.modified_a is True
        assert "Data" in app.modified_sheets_a
        assert app.manual_a_cell_ops
        assert app.undo_stack
    finally:
        app._confirm_overwrite = original_confirm
        app._shutdown_root()


def test_only_diff_pending_stays_checked_locked_and_keeps_stable_view():
    app, view, source_paths = _open_ready_view()
    original_start = view._start_async_large_only_diff_build
    original_cache_from_maps = view._cache_only_diff_rows_from_exact_pair_maps
    try:
        view._invalidate_only_diff_snapshot_cache()
        # Full-view READY requires an exact pair map; only the dedicated
        # only-diff row snapshot is absent in this transition scenario.
        view._pair_diff_full_exact = True
        view._only_diff_rows_exact = False
        view._refresh_interaction_gate()
        assert view._lifecycle_state == "READY"

        stable = (
            tuple(view.display_rows),
            view.left.get("1.0", "end-1c"),
            view.right.get("1.0", "end-1c"),
        )
        starts = []

        def _start_pending_once(*, user_initiated=False):
            starts.append(view._current_only_diff_cache_key())
            assert user_initiated is True
            view._only_diff_async_building = True
            view._only_diff_async_requested_value = 1
            view._only_diff_async_build_key = view._current_only_diff_cache_key()
            view._only_diff_preview_full = True
            view._set_only_diff_pending_info()
            view._refresh_interaction_gate()
            return True

        view._start_async_large_only_diff_build = _start_pending_once
        view._cache_only_diff_rows_from_exact_pair_maps = lambda: False
        view.only_diff_var.set(1)
        view._toggle_only_diff()
        _pump(app.root)

        assert int(view.only_diff_var.get()) == 1
        # The stable preference is committed only after exact publication;
        # cancellation must still know it should return to full view.
        assert view._last_only_diff_value == 0
        assert view._lifecycle_state == "DIFFING", view._lifecycle_state
        assert str(view.only_diff_cb.cget("state")) == "disabled"
        assert str(view.only_diff_cb.cget("text")) == "只看差异内容"
        assert "正在后台生成精确差异行" in str(view.info.cget("text"))
        assert starts and len(starts) == 1
        assert (
            tuple(view.display_rows),
            view.left.get("1.0", "end-1c"),
            view.right.get("1.0", "end-1c"),
        ) == stable

        before = _mutation_snapshot(app, view, source_paths)
        view._copy_single_cell_by_pair(1, "B2A", 2)
        _assert_mutation_unchanged(
            before,
            app,
            view,
            source_paths,
            "only-diff pending cell handler",
        )

        # A repeated direct callback must preserve the checked request and may
        # neither start nor cancel a second transition.
        view._toggle_only_diff()
        _pump(app.root)
        assert int(view.only_diff_var.get()) == 1
        assert view._lifecycle_state == "DIFFING"
        assert len(starts) == 1
        assert view._only_diff_async_building is True

        exact_rows = sorted(
            pair_idx
            for pair_idx in range(len(view.row_pairs))
            if view._pair_has_visual_diff(pair_idx)
        )
        view._only_diff_async_building = False
        view._only_diff_async_build_key = None
        view._only_diff_rows_exact = True
        view._pair_diff_full_exact = True
        view._cache_only_diff_rows_snapshot(exact_rows)
        view._only_diff_preview_full = False
        view._refresh_mode_switch_preserving_selection(rescan=False)
        view._refresh_interaction_gate()
        _pump(app.root)

        assert view._lifecycle_state == "READY", view._lifecycle_state
        assert int(view.only_diff_var.get()) == 1
        assert str(view.only_diff_cb.cget("state")) == "normal"
        assert "计算中" not in str(view.only_diff_cb.cget("text"))
        assert tuple(view.display_rows) == tuple(exact_rows)
        assert str(view.use_right_btn.cget("state")) == "normal"
    finally:
        view._start_async_large_only_diff_build = original_start
        view._cache_only_diff_rows_from_exact_pair_maps = original_cache_from_maps
        app._shutdown_root()


def test_edit_ready_callback_never_calls_refresh_rescan_true():
    app, view, _source_paths = _open_ready_view()
    original_refresh = view.refresh
    original_enqueue = app._enqueue_sheet
    original_kick = app._kick_worker
    original_missing = view._is_missing_sheet_view
    try:
        refresh_calls = []
        enqueue_calls = []

        def _record_refresh(*args, **kwargs):
            refresh_calls.append((args, dict(kwargs)))

        view.refresh = _record_refresh
        app._enqueue_sheet = lambda *args, **kwargs: enqueue_calls.append(
            (args, dict(kwargs))
        )
        app._kick_worker = lambda: None

        # Ordinary materialized views must be requeued for prepared background
        # data rather than synchronously rescanned by the edit-ready callback.
        view._data_ready = True
        view._cache_formula_aware = False
        app._refresh_loaded_views_after_edit_ready()
        assert refresh_calls == [], refresh_calls
        assert enqueue_calls, "non-formula-aware view was not requeued"

        # The no-rescan contract is universal, including missing-Sheet views.
        # This branch historically performed refresh(rescan=True) directly.
        refresh_calls.clear()
        enqueue_calls.clear()
        view._data_ready = True
        view._cache_formula_aware = False
        view._is_missing_sheet_view = lambda: True
        app._refresh_loaded_views_after_edit_ready()
        assert not any(
            bool(kwargs.get("rescan"))
            for _args, kwargs in refresh_calls
        ), refresh_calls
    finally:
        view.refresh = original_refresh
        app._enqueue_sheet = original_enqueue
        app._kick_worker = original_kick
        view._is_missing_sheet_view = original_missing
        app._shutdown_root()


def test_exact_worker_failure_is_retryable_and_broker_is_serial():
    app, view, _source_paths = _open_ready_view()
    original_submit = app._submit_priority_exact
    original_enqueue = app._enqueue_sheet
    original_kick = app._kick_worker
    try:
        view._invalidate_only_diff_snapshot_cache()
        app._submit_priority_exact = lambda *_args, **_kwargs: None
        assert view._start_async_large_only_diff_build() is False
        assert view._derive_lifecycle_state() == "FAILED"
        assert "后台任务" in str(view._lifecycle_error)

        enqueued = []
        kicked = []
        app._enqueue_sheet = lambda *args, **kwargs: enqueued.append(
            (args, dict(kwargs))
        )
        app._kick_worker = lambda: kicked.append(True)
        view._manual_rescan()
        assert view._lifecycle_error is None
        assert view._data_ready is False
        assert enqueued and kicked

        # The app-level exact broker owns one thread and serializes a
        # replaceable pending request.
        app._submit_priority_exact = original_submit
        started = threading.Event()
        release = threading.Event()
        finished = threading.Event()
        lock = threading.Lock()
        active = 0
        max_active = 0
        ran = []

        def first_worker():
            nonlocal active, max_active
            with lock:
                active += 1
                max_active = max(max_active, active)
            started.set()
            release.wait(5.0)
            with lock:
                active -= 1
            ran.append("first")

        def second_worker():
            nonlocal active, max_active
            with lock:
                active += 1
                max_active = max(max_active, active)
            ran.append("second")
            with lock:
                active -= 1
            finished.set()

        view._only_diff_async_build_seq = 100
        first_thread = app._submit_priority_exact(view, 100, first_worker)
        assert first_thread is not None
        assert started.wait(2.0)
        view._only_diff_async_build_seq = 101
        second_thread = app._submit_priority_exact(view, 101, second_worker)
        assert second_thread is first_thread
        release.set()
        assert finished.wait(5.0)
        assert ran == ["first", "second"], ran
        assert max_active == 1, max_active
    finally:
        release = locals().get("release")
        if release is not None:
            release.set()
        app._submit_priority_exact = original_submit
        app._enqueue_sheet = original_enqueue
        app._kick_worker = original_kick
        app._shutdown_root()


def test_saved_baseline_and_missing_sheet_nested_guard():
    app, view, _source_paths = _open_ready_view()
    try:
        view._copy_single_cell_by_pair(1, "B2A", 2)
        assert view._has_user_edits_for_current_sheet()
        app._commit_saved_side_baseline("A")
        assert app.modified_a is False
        assert not app.modified_sheets_a
        assert not app.manual_a_cell_ops
        assert not app.manual_a_row_ops
        assert not app.manual_a_column_ops
        assert not app.manual_a_formula_cache_ops
        assert not view.touched_rows
        assert not view._has_user_edits_for_current_sheet()
    finally:
        app._shutdown_root()

    root_dir = make_temp_dir("sow_missing_sheet_guard_")
    mine = os.path.join(root_dir, "mine.xlsx")
    theirs = os.path.join(root_dir, "theirs.xlsx")
    _make_book(mine, (("id", "value"), ("row-1", "mine")))
    theirs_book = Workbook()
    data = theirs_book.active
    data.title = "Data"
    data.append(("id", "value"))
    data.append(("row-1", "theirs"))
    extra = theirs_book.create_sheet("Extra")
    extra.append(("id", "value"))
    extra.append(("extra-1", "theirs-extra"))
    theirs_book.save(theirs)
    theirs_book.close()

    app = smt.SowMergeApp(mine, theirs)
    try:
        app.nb.select(app._sheet_containers["Extra"])
        _wait_until(
            app.root,
            lambda: (
                app.sheet_views.get("Extra") is not None
                and bool(getattr(app.sheet_views["Extra"], "_data_ready", False))
                and app._edit_workbooks_ready()
            ),
            "missing-Sheet view did not finish loading",
        )
        missing_view = app.sheet_views["Extra"]
        missing_view.only_diff_var.set(0)
        missing_view._last_only_diff_value = 0
        missing_view._row_model_exact = True
        missing_view._cache_formula_aware = True
        missing_view._pair_diff_full_exact = True
        missing_view._lifecycle_error = None
        missing_view._lifecycle_canceled = False
        missing_view._refresh_interaction_gate()
        assert missing_view._derive_lifecycle_state() == "READY", (
            missing_view._derive_lifecycle_state(),
            missing_view._data_ready,
            missing_view._row_model_exact,
            missing_view._cache_formula_aware,
            missing_view._pair_diff_full_exact,
            missing_view._column_mapping_is_current(),
            app._edit_workbooks_ready(),
        )
        missing_view._copy_selected_row("B2A")
        assert "Extra" in app._wb_a_edit.sheetnames
        assert app._wb_a_edit["Extra"]["B2"].value == "theirs-extra"
        assert app.manual_sheet_ops
    finally:
        app._shutdown_root()


def test_stale_exact_generation_cannot_publish():
    app, view, _source_paths = _open_ready_view()
    original_queue_ui = app._queue_ui_task
    try:
        queued_ui = []
        before_diff = copy.deepcopy(view.pair_diff_cols)
        before_base_diff = copy.deepcopy(view.pair_base_diff_cols)
        view._invalidate_only_diff_snapshot_cache()
        view.only_diff_var.set(1)
        app._queue_ui_task = lambda fn: queued_ui.append(fn) or True
        assert view._start_async_large_only_diff_build() is True
        deadline = time.time() + 10.0
        while time.time() < deadline:
            with app._exact_broker_lock:
                running = bool(app._exact_broker_running)
            if not running and queued_ui:
                break
            time.sleep(0.01)
        assert queued_ui, "exact worker did not queue a result"

        # Supersede the worker after it computed but before Tk publication.
        view._only_diff_async_build_seq += 1
        for callback in queued_ui:
            callback()
        assert view.pair_diff_cols == before_diff
        assert view.pair_base_diff_cols == before_base_diff
        assert not view._has_valid_only_diff_snapshot_cache()
        assert view._derive_lifecycle_state() != "READY"
    finally:
        app._queue_ui_task = original_queue_ui
        app._shutdown_root()


def test_hidden_cache_completeness_is_retained_without_view_render():
    root_dir = make_temp_dir("sow_hidden_cache_completeness_")
    mine = os.path.join(root_dir, "mine.xlsx")
    theirs = os.path.join(root_dir, "theirs.xlsx")
    for path, suffix in ((mine, "mine"), (theirs, "theirs")):
        workbook = Workbook()
        data = workbook.active
        data.title = "Data"
        data.append(("id", "value"))
        data.append(("row-1", suffix))
        hidden = workbook.create_sheet("Hidden")
        hidden.append(("id", "value"))
        hidden.append(("hidden-1", suffix))
        workbook.save(path)
        workbook.close()

    app = smt.SowMergeApp(mine, theirs)
    try:
        app.nb.select(app._sheet_containers["Data"])
        _wait_until(
            app.root,
            lambda: "Hidden" in app._sheet_cache_store,
            "hidden Sheet cache was not retained",
        )
        assert app.sheet_views["Hidden"] is None
        cache = app._sheet_cache_store["Hidden"]
        completeness = cache.get("completeness") or {}
        assert set(completeness) >= {
            "formula_aware",
            "row_model_exact",
            "column_projection_exact",
            "ab_diff_exact",
            "base_diff_exact",
            "only_diff_rows_exact",
            "mode",
        }, completeness
        assert completeness["formula_aware"] is True
        assert completeness["row_model_exact"] is True
        assert completeness["column_projection_exact"] is True
        with app._compute_lock:
            app._compute_queue[:] = ["Hidden", "Data"]
        app._enqueue_sheet("Data", front=True)
        with app._compute_lock:
            assert app._compute_queue[0] == "Data", app._compute_queue
    finally:
        app._shutdown_root()


def main():
    original_settings_path = smt._SETTINGS_PATH
    settings_dir = make_temp_dir("sow_loading_gate_settings_")
    smt._SETTINGS_PATH = os.path.join(settings_dir, "settings.json")
    tests = (
        test_non_ready_direct_handlers_have_no_write_side_effects,
        test_only_diff_pending_stays_checked_locked_and_keeps_stable_view,
        test_edit_ready_callback_never_calls_refresh_rescan_true,
        test_exact_worker_failure_is_retryable_and_broker_is_serial,
        test_saved_baseline_and_missing_sheet_nested_guard,
        test_stale_exact_generation_cannot_publish,
        test_hidden_cache_completeness_is_retained_without_view_render,
    )
    try:
        for test in tests:
            test()
            print(f"PASS: {test.__name__}")
        print(f"PASS: loading/read-only gate regression ({len(tests)} tests)")
    finally:
        smt._SETTINGS_PATH = original_settings_path


if __name__ == "__main__":
    main()

"""Isolated Link.xlsx acceptance for adopting a theirs-side row deletion."""

from __future__ import annotations

import hashlib
import glob
import json
import os
import shutil
import tempfile

from openpyxl import load_workbook

import sow_merge_tool as smt
from _ux_5_3_final_acceptance import (
    _pump,
    _wait_for_stable_projection,
    _wait_for_view,
    wait_edit_ready,
    wait_view_ready,
)


SOURCE_ROOT = r"C:\GM15\design\sheets\release"
SHEET = "Link@design"


def _sha256(path: str) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest().upper()


def _column_values(worksheet, column=1):
    return [
        worksheet.cell(row=row, column=column).value
        for row in range(1, worksheet.max_row + 1)
    ]


def _mine_only_pairs(view):
    return [
        pair_idx
        for pair_idx, (row_a, row_b) in enumerate(view.row_pairs)
        if row_a is not None and row_b is None
    ]


def main():
    temp_root = tempfile.gettempdir()
    raw_base = os.path.join(SOURCE_ROOT, "Link.xlsx.merge-left.r36473")
    raw_theirs = os.path.join(SOURCE_ROOT, "Link.xlsx.merge-right.r36474")
    base_candidates = [raw_base] + glob.glob(
        os.path.join(temp_root, "sow_merge_tool_svncat_BASE_*_Link.xlsx")
    )
    theirs_candidates = [raw_theirs] + glob.glob(
        os.path.join(
            temp_root,
            "sow_merge_tool_svn_Link.xlsx.merge-right.r36474_*.xlsx",
        )
    )
    base_path = max(
        (path for path in base_candidates if os.path.isfile(path)),
        key=os.path.getmtime,
    )
    theirs_path = max(
        (path for path in theirs_candidates if os.path.isfile(path)),
        key=os.path.getmtime,
    )
    sources = {
        "mine": os.path.join(SOURCE_ROOT, "Link.xlsx"),
        "base": base_path,
        "theirs": theirs_path,
    }
    for path in sources.values():
        if not os.path.isfile(path):
            raise FileNotFoundError(path)
    source_hashes = {role: _sha256(path) for role, path in sources.items()}

    root_dir = tempfile.mkdtemp(prefix="sow-link-row-delete-")
    paths = {}
    for role, source in sources.items():
        destination = os.path.join(root_dir, role, "Link.xlsx")
        os.makedirs(os.path.dirname(destination), exist_ok=True)
        shutil.copy2(source, destination)
        paths[role] = destination
    settings_path = os.path.join(root_dir, "settings.json")
    with open(settings_path, "w", encoding="utf-8") as stream:
        json.dump({"only_diff": 0}, stream)

    original_settings_path = smt._SETTINGS_PATH
    smt._SETTINGS_PATH = settings_path
    app = None
    output = None
    try:
        app = smt.SowMergeApp(
            paths["mine"],
            paths["theirs"],
            merge_mode=True,
            merged_path=os.path.join(root_dir, "merged", "Link.xlsx"),
            base_path=paths["base"],
        )
        app._intended_window_state = "normal"
        app.root.state("normal")
        app.root.geometry("1500x900")
        app.nb.select(app._sheet_containers[SHEET])
        view = _wait_for_view(app, SHEET, timeout=120.0)
        wait_edit_ready(app, timeout=120.0)
        wait_view_ready(view, timeout=120.0)
        _wait_for_stable_projection(view, timeout=120.0, stable_for=0.3)

        view.only_diff_var.set(0)
        view._last_only_diff_value = 0
        view.refresh(row_only=None, rescan=True)
        _pump(app.root, 0.2)
        view._set_copy_scope_mode("region")

        mine_only = _mine_only_pairs(view)
        if not mine_only:
            raise AssertionError("Link fixture no longer contains a theirs-deleted row")
        anchor_pair = mine_only[0]
        target_block = view._logical_diff_pair_block_for_pair(anchor_pair)
        deleted_pairs = [
            pair_idx
            for pair_idx in target_block
            if view.row_pairs[pair_idx][0] is not None
            and view.row_pairs[pair_idx][1] is None
        ]
        if not deleted_pairs:
            raise AssertionError((anchor_pair, target_block, view.row_pairs[anchor_pair]))

        mine_ws = app.ws_a_edit(SHEET)
        mine_rows_before = int(mine_ws.max_row)
        deleted_row_numbers = [int(view.row_pairs[pair_idx][0]) for pair_idx in deleted_pairs]
        deleted_markers = [
            tuple(
                mine_ws.cell(row=row_idx, column=col_idx).value
                for col_idx in range(1, min(4, int(mine_ws.max_column or 1)) + 1)
            )
            for row_idx in deleted_row_numbers
        ]
        mine_col1_before = _column_values(mine_ws)
        undo_before = len(app.undo_stack)

        view._select_line(view.row_to_line[anchor_pair])
        view._run_copy_action_by_mode("B2A")
        _pump(app.root, 0.5)

        assert int(app.ws_a_edit(SHEET).max_row) == mine_rows_before - len(deleted_pairs)
        assert len(app.undo_stack) == undo_before + 1
        delete_ops = [
            op for op in app.manual_a_row_ops
            if op.get("sheet") == SHEET and op.get("kind") == "delete_rows"
        ]
        assert sum(int(op.get("count", 0)) for op in delete_ops) == len(deleted_pairs)
        remaining_mine_only = _mine_only_pairs(view)
        assert len(remaining_mine_only) == len(mine_only) - len(deleted_pairs)

        view._undo_last_action()
        _pump(app.root, 0.5)
        assert _column_values(app.ws_a_edit(SHEET)) == mine_col1_before
        assert len(app.undo_stack) == undo_before
        assert not [
            op for op in app.manual_a_row_ops
            if op.get("sheet") == SHEET and op.get("kind") == "delete_rows"
        ]

        mine_only = _mine_only_pairs(view)
        view._select_line(view.row_to_line[mine_only[0]])
        view._run_copy_action_by_mode("B2A")
        _pump(app.root, 0.5)
        assert len(_mine_only_pairs(view)) == len(mine_only) - len(deleted_pairs)

        output = app.build_manual_merge_output_file()
        saved = load_workbook(output, data_only=False, read_only=True)
        try:
            saved_ws = saved[SHEET]
            assert int(saved_ws.max_row) == mine_rows_before - len(deleted_pairs)
            saved_markers = {
                tuple(
                    saved_ws.cell(row=row_idx, column=col_idx).value
                    for col_idx in range(1, min(4, int(saved_ws.max_column or 1)) + 1)
                )
                for row_idx in range(1, int(saved_ws.max_row) + 1)
            }
            assert all(marker not in saved_markers for marker in deleted_markers)
        finally:
            saved.close()

        assert {
            role: _sha256(path) for role, path in sources.items()
        } == source_hashes
        print(
            "PASS: real Link theirs-row-delete "
            f"pairs={deleted_pairs} rows={deleted_row_numbers} "
            f"ops={[(op['row'], op['count']) for op in delete_ops]}"
        )
    finally:
        smt._SETTINGS_PATH = original_settings_path
        if app is not None:
            app._shutdown_root()
        if output and os.path.exists(output):
            os.remove(output)
        shutil.rmtree(root_dir, ignore_errors=True)


if __name__ == "__main__":
    main()

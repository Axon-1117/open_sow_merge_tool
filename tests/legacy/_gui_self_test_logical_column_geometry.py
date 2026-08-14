"""OpenSpec 2.6: logical-column GUI/projection geometry regression.

Run:
  python _gui_self_test_logical_column_geometry.py

The pure checks intentionally avoid creating Tk widgets.  The final check opens
one small real view to prove that background-cache replay and Tk geometry use
the same immutable logical-column projection.
"""

from __future__ import annotations

import os
import ast
import inspect
import textwrap
import time
from types import SimpleNamespace

from openpyxl import Workbook

import sow_merge_tool as smt
from _test_temp_utils import make_temp_dir


class _Var:
    def __init__(self, value=0):
        self.value = value

    def get(self):
        return self.value

    def set(self, value):
        self.value = value


def _rows(columns, count: int = 8, *, edit=None):
    columns = tuple(columns)
    result = [columns]
    for row_no in range(1, count + 1):
        row = [f"{name.lower()}-{row_no}-{'v' * 18}" for name in columns]
        if edit is not None and row_no == edit[0]:
            row[columns.index(edit[1])] = edit[2]
        result.append(tuple(row))
    return result


def _key(name: str, row_version: int = 1, column_version: int = 1):
    return smt.ColumnModelCacheKey(name, row_version, column_version)


def _cache_2way(name, mine, theirs):
    return smt.build_logical_column_comparison_cache_2way(
        _key(name),
        mine,
        theirs,
        mine,
        theirs,
        mine_max_col=len(mine[0]),
        theirs_max_col=len(theirs[0]),
    )


def _cache_3way(name, mine, base, theirs):
    return smt.build_logical_column_comparison_cache_3way(
        _key(name),
        mine,
        base,
        theirs,
        mine,
        base,
        theirs,
        mine_max_col=len(mine[0]),
        base_max_col=len(base[0]),
        theirs_max_col=len(theirs[0]),
    )


def _fake_view(cache, *, three_way=False, max_col=None):
    view = object.__new__(smt.SheetView)
    view.sheet = cache.model.cache_key.sheet_name
    view.app = SimpleNamespace(
        merge_mode=bool(three_way),
        has_base=bool(three_way),
        get_sheet_meta=lambda _sheet: {"view_mode": "normal"},
    )
    view.three_way_var = _Var(1 if three_way else 0)
    view.grid_overlay_var = _Var(0)
    view.max_col = int(max_col or max(
        (slot.mine_col or 0) for slot in cache.model.slots
    ) or 1)
    view.col_max_a = max((slot.mine_col or 0) for slot in cache.model.slots) or 1
    view.col_max_b = max((slot.theirs_col or 0) for slot in cache.model.slots) or 1
    view.col_max_base = max((slot.base_col or 0) for slot in cache.model.slots) or 1
    view.column_comparison_cache = cache
    view.column_projection = None
    view._column_projection_generation = 0
    view._row_model_version = cache.model.cache_key.row_model_version
    view._column_model_version = cache.model.cache_key.column_model_version
    view._mine_edit_version = cache.model.cache_key.mine_edit_version
    view._base_edit_version = cache.model.cache_key.base_edit_version
    view._theirs_edit_version = cache.model.cache_key.theirs_edit_version
    view._column_mapping_stale_reason = ""
    view._data_version = cache.model.cache_key.row_model_version
    view._col_widths_version = 0
    view._base_spans_cache = None
    view._base_spans_cache_key = None
    view.col_char_widths = {
        logical_col: 8 for logical_col in range(1, len(cache.model.slots) + 1)
    }
    view.row_pairs = []
    view.pair_diff_cols = {}
    view.pair_base_diff_cols = {}
    view._diff_map_cache = None
    view._diff_map_cache_version = None
    view._align_rows_enabled = True
    view._force_sequence_align = False
    view.touched_rows = set()
    view.row_a_to_pair_idx = {}
    view._only_diff_source_version = 0
    view._only_diff_rows_cache = None
    view._only_diff_rows_cache_key = None
    view._only_diff_async_build_key = None
    view._only_diff_async_building = False
    view._only_diff_async_build_seq = 0
    return view


def _fragments(view, line: str):
    return [
        line[start:end].rstrip()
        for _logical_col, (start, end) in view._spans_for_line(line).items()
    ]


def _assert_monotonic_spans(spans, expected_count: int):
    assert list(spans) == list(range(1, expected_count + 1)), spans
    previous_end = -1
    for start, end in spans.values():
        assert 0 <= start < end, (start, end)
        assert start > previous_end, (previous_end, start, end)
        previous_end = end


def _two_way_fixture():
    mine = _rows(("A", "B", "C", "D", "E", "F"))
    theirs = _rows(
        ("A", "X", "Y", "B", "D", "E", "F"),
        edit=(5, "E", "e-5-independent-edit"),
    )
    cache = _cache_2way("two-way-geometry", mine, theirs)
    return mine, theirs, cache


def test_two_way_projection_headers_spans_and_placeholders():
    mine, theirs, cache = _two_way_fixture()
    projection = smt.LogicalColumnProjection.from_model(cache.model)
    assert projection.slot_count == 8, cache.model.slots
    assert [projection.physical_col("A", col) for col in range(1, 9)] == [
        1, None, None, 2, 3, 4, 5, 6
    ]
    assert [projection.physical_col("B", col) for col in range(1, 9)] == [
        1, 2, 3, 4, None, 5, 6, 7
    ]
    assert [projection.logical_col("A", col) for col in range(1, 7)] == [1, 4, 5, 6, 7, 8]
    assert [projection.logical_col("B", col) for col in range(1, 8)] == [1, 2, 3, 4, 6, 7, 8]

    view = _fake_view(cache, max_col=7)
    view._projected_widths_from_cached_parts({0: mine[0]}, {0: theirs[0]})
    expected_a = ["A", smt._LOGICAL_COLUMN_PLACEHOLDER, smt._LOGICAL_COLUMN_PLACEHOLDER,
                  "B", "C", "D", "E", "F"]
    expected_b = ["A", "X", "Y", "B", smt._LOGICAL_COLUMN_PLACEHOLDER,
                  "D", "E", "F"]
    assert view._project_raw_parts(mine[0], "A") == expected_a
    assert view._project_raw_parts(theirs[0], "B") == expected_b

    line_a = view._render_line_from_raw_parts(list(mine[0]), "A")
    line_b = view._render_line_from_raw_parts(list(theirs[0]), "B")
    assert _fragments(view, line_a) == expected_a
    assert _fragments(view, line_b) == expected_b
    spans_a = view._spans_for_line(line_a)
    spans_b = view._spans_for_line(line_b)
    _assert_monotonic_spans(spans_a, 8)
    assert spans_a == spans_b == view._base_spans()

    for side in ("A", "B", "LOGICAL"):
        header = view._build_col_header_line(side)
        assert _fragments(view, header) == [
            projection.header_label(side, col) for col in range(1, 9)
        ]
        assert view._spans_for_line(header) == spans_a

    # Two missing Mine slots and one missing Theirs slot, on two rendered rows.
    assert len(view._padding_column_tag_args("A", 2)) == 2 * 2 * 2
    assert len(view._padding_column_tag_args("B", 2)) == 1 * 2 * 2


def test_three_way_independent_and_same_anchor_geometry():
    base = _rows(("A", "B", "C", "D"))
    mine = _rows(("A", "M", "B", "C", "D"))
    theirs = _rows(("A", "B", "C", "T", "D"))
    cache = _cache_3way("three-independent", mine, base, theirs)
    projection = smt.LogicalColumnProjection.from_model(cache.model)
    assert projection.slot_count == 6
    assert [(slot.base_boundary, slot.origin_side) for slot in projection.slots if slot.base_col is None] == [
        (1, "mine"), (3, "theirs")
    ]
    assert [projection.physical_col("A", col) for col in range(1, 7)] == [1, 2, 3, 4, None, 5]
    assert [projection.physical_col("BASE", col) for col in range(1, 7)] == [1, None, 2, 3, None, 4]
    assert [projection.physical_col("B", col) for col in range(1, 7)] == [1, None, 2, 3, 4, 5]

    view = _fake_view(cache, three_way=True, max_col=5)
    view._projected_widths_from_cached_parts({0: mine[0]}, {0: theirs[0]}, {0: base[0]})
    spans = view._base_spans()
    _assert_monotonic_spans(spans, 6)
    for side, row in (("A", mine[0]), ("BASE", base[0]), ("B", theirs[0])):
        line = view._render_line_from_raw_parts(list(row), side)
        assert len(_fragments(view, line)) == 6
        assert view._spans_for_line(line) == spans

    same_base = _rows(("A", "B", "C"))
    same_mine = _rows(("A", "M", "B", "C"))
    same_theirs = _rows(("A", "T", "B", "C"))
    competing = _cache_3way("three-same-anchor", same_mine, same_base, same_theirs)
    competing_projection = smt.LogicalColumnProjection.from_model(competing.model)
    assert competing_projection.slot_count == 5
    assert [
        (slot.mine_col, slot.base_col, slot.theirs_col, slot.base_boundary)
        for slot in competing_projection.slots[1:3]
    ] == [(2, None, None, 1), (None, None, 2, 1)]
    unresolved_blocks = [block for block in competing.model.blocks if block.state == "unresolved"]
    assert [block.slot_indices for block in unresolved_blocks] == [(1,), (2,)], (
        competing.model.blocks
    )
    analysis = smt.classify_logical_columns_3way(competing)
    assert {item.kind for item in analysis.structural_conflicts} == {"competing-insertion"}


def test_hit_selection_hover_and_mixed_row_column_mapping():
    mine, theirs, cache = _two_way_fixture()
    view = _fake_view(cache, max_col=7)
    view.col_char_widths = {col: 7 + col for col in range(1, 9)}
    spans = view._base_spans()
    for logical_col, (start, end) in spans.items():
        midpoint = start + ((end - start) // 2)
        assert view._col_from_char(midpoint) == logical_col
        assert view._hit_col_from_char(midpoint)[0] == logical_col
        if logical_col < 8:
            # Tooltip hit-testing owns separators on the left; click selection
            # accepts only actual cell spans.  Both rules are deterministic.
            assert view._hit_col_from_char(end)[0] == logical_col
            assert view._col_from_char(end) is None

    assert view._physical_col_for_logical("A", 2) is None
    assert view._physical_col_for_logical("B", 2) == 2
    assert view._physical_col_for_logical("A", 5) == 3
    assert view._physical_col_for_logical("B", 5) is None
    assert view._physical_col_for_logical("A", 7) == 5
    assert view._physical_col_for_logical("B", 7) == 6

    wb_a = Workbook()
    wb_b = Workbook()
    ws_a = wb_a.active
    ws_b = wb_b.active
    ws_a.title = ws_b.title = "S"
    for row in mine:
        ws_a.append(list(row))
    # Insert one physical row only on the right; the logical-column projection
    # must compose with row_pairs instead of drifting to a neighbour cell.
    ws_b.append(list(theirs[0]))
    ws_b.append(["a-new", "x-new", "y-new", "b-new", "d-new", "e-new", "f-new"])
    for row in theirs[1:]:
        ws_b.append(list(row))
    view.sheet = "S"
    view.app = SimpleNamespace(
        merge_mode=False,
        has_base=False,
        get_sheet_meta=lambda _sheet: {"view_mode": "normal"},
        ws_a_val=lambda _sheet: ws_a,
        ws_b_val=lambda _sheet: ws_b,
    )
    view.row_pairs = [(1, 1), (None, 2)] + [(row, row + 1) for row in range(2, len(mine) + 1)]

    inserted_col_payload = view._cmp_tooltip_payload_by_pair_col(2, 2, force_panel=True)
    assert inserted_col_payload is not None
    assert smt._LOGICAL_COLUMN_PLACEHOLDER in inserted_col_payload[0]
    assert "x-1-" in inserted_col_payload[0], inserted_col_payload[0]

    retained_payload = view._cmp_tooltip_payload_by_pair_col(2, 4, force_panel=True)
    assert retained_payload is not None
    assert retained_payload[0].count("b-1-") == 2, retained_payload[0]

    inserted_row_payload = view._cmp_tooltip_payload_by_pair_col(1, 4, force_panel=True)
    assert inserted_row_payload is not None
    assert "<missing>" in inserted_row_payload[0]
    assert "b-new" in inserted_row_payload[0]
    wb_a.close()
    wb_b.close()


def test_only_diff_minimap_uses_logical_slots_without_row_flooding():
    _mine, _theirs, cache = _two_way_fixture()
    view = _fake_view(cache, max_col=7)
    view.row_pairs = [(row, row) for row in range(1, 10)]
    structural = set(cache.structural_diff_cols)
    assert structural == {2, 3, 5}
    view.pair_diff_cols = {pair_idx: set(structural) for pair_idx in range(9)}
    # One real value edit in retained E: logical 7, physical A=5/B=6.
    view.pair_diff_cols[5].add(7)
    assert view._visual_diff_cols_for_pair(0) == set()
    assert view._visual_diff_cols_for_pair(5) == {7}
    total_pairs, diff_rows, diff_cols = view._compute_diff_map_data()
    assert total_pairs == 9
    assert diff_rows == [5], diff_rows
    assert diff_cols == [2, 3, 5, 7], diff_cols
    assert view._all_logical_diff_cols_for_pair(0) == {2, 3, 5}
    summary = view._column_structure_summary()
    assert "B:C" in summary and "新增列" in summary, summary
    assert "E" in summary and "删除列" in summary, summary


def test_three_way_common_base_missing_column_does_not_flood_visual_rows():
    slots = (
        smt.ColumnSlot(0, mine_col=1, base_col=1, theirs_col=1),
        smt.ColumnSlot(
            1,
            mine_col=2,
            base_col=None,
            theirs_col=2,
            state="inserted",
            base_boundary=1,
            origin_side="both",
        ),
        smt.ColumnSlot(2, mine_col=3, base_col=2, theirs_col=3),
    )
    model = smt.ColumnModel.from_slots(
        _key("three-way-common-base-missing"),
        slots,
        blocks=smt._build_column_blocks(slots),
    )
    cache = smt.LogicalColumnComparisonCache(model=model)
    view = _fake_view(cache, three_way=True, max_col=3)
    view.row_pairs = [(1, 1), (2, 2), (3, 3)]
    view.pair_diff_cols = {0: set(), 1: set(), 2: set()}
    view.pair_base_diff_cols = {0: {2}, 1: {2}, 2: {2}}
    assert any(2 in cols for cols in view.pair_base_diff_cols.values())
    assert all(
        view._visual_diff_cols_for_pair(pair_idx) == set()
        for pair_idx in range(3)
    )

    # A real current-side disagreement in the same logical slot remains a row
    # diff even though Base also lacks that common inserted column.
    view.pair_diff_cols[1] = {2}
    assert view._visual_diff_cols_for_pair(1) == {2}

    # An explicitly accepted whole-column copy may still look different when
    # row alignment pairs different physical rows.  Suppress that artifact
    # only while the copied physical cells remain equal; a later real edit
    # must immediately make the row visible again.
    wb_a_val = Workbook()
    wb_b_val = Workbook()
    wb_a_edit = Workbook()
    wb_b_edit = Workbook()
    for workbook in (wb_a_val, wb_b_val, wb_a_edit, wb_b_edit):
        workbook.active.title = view.sheet
        workbook[view.sheet].cell(2, 2).value = "accepted-copy"
    view.app = SimpleNamespace(
        merge_mode=True,
        has_base=True,
        get_sheet_meta=lambda _sheet: {"view_mode": "normal"},
        ws_a_val=lambda _sheet: wb_a_val[view.sheet],
        ws_b_val=lambda _sheet: wb_b_val[view.sheet],
        ws_a_edit=lambda _sheet: wb_a_edit[view.sheet],
        ws_b_edit=lambda _sheet: wb_b_edit[view.sheet],
    )
    view._accepted_common_insert_sources = {2: "B"}
    assert view._visual_diff_cols_for_pair(1) == set()
    wb_a_val[view.sheet].cell(2, 2).value = "later-real-edit"
    wb_a_edit[view.sheet].cell(2, 2).value = "later-real-edit"
    assert view._visual_diff_cols_for_pair(1) == {2}
    for workbook in (wb_a_val, wb_b_val, wb_a_edit, wb_b_edit):
        workbook.close()

    unresolved = smt.ColumnMappingConfidence(
        0.0,
        True,
        "low-confidence-physical-fallback",
        ("physical-order",),
        (smt.COLUMN_MAPPING_CAUSE_LOW_CONFIDENCE,),
    )
    unresolved_slots = (
        slots[0],
        smt.ColumnSlot(
            1,
            mine_col=2,
            base_col=None,
            theirs_col=2,
            state="unresolved",
            confidence=unresolved,
            base_boundary=1,
            origin_side="both",
        ),
        slots[2],
    )
    unresolved_model = smt.ColumnModel.from_slots(
        _key("three-way-common-base-missing-unresolved"),
        unresolved_slots,
        blocks=smt._build_column_blocks(unresolved_slots),
        confidence=unresolved,
    )
    unresolved_view = _fake_view(
        smt.LogicalColumnComparisonCache(
            model=unresolved_model,
            unresolved_cols=frozenset({2}),
        ),
        three_way=True,
        max_col=3,
    )
    unresolved_view.row_pairs = [(1, 1)]
    unresolved_view.pair_diff_cols = {0: set()}
    unresolved_view.pair_base_diff_cols = {0: {2}}
    assert unresolved_view._visual_diff_cols_for_pair(0) == {2}


def test_only_diff_toggle_after_column_edit_caches_current_visual_rows_synchronously():
    rows = _rows(("A", "B", "C"), count=39)
    cache = _cache_2way("only-diff-after-column-edit", rows, rows)
    view = _fake_view(cache, max_col=3)
    view.row_pairs = [(row, row) for row in range(1, 41)]
    view.row_a_to_pair_idx = {row: row - 1 for row in range(1, 41)}
    view.row_b_to_pair_idx = dict(view.row_a_to_pair_idx)
    expected = list(range(34))
    view.pair_diff_cols = {
        pair_idx: ({2} if pair_idx in expected else set())
        for pair_idx in range(40)
    }
    view.pair_base_diff_cols = {}
    view.only_diff_var = _Var(1)
    view._data_ready = True
    view._row_model_exact = True
    view._cache_formula_aware = True
    view._pair_diff_full_exact = True
    view._is_large_sheet = True
    view._full_render = True
    view._last_only_diff_value = 0
    view._only_diff_rows_cache = None
    view._only_diff_rows_cache_key = None
    view._active_column_projection()
    view.app.modified_sheets_a = {view.sheet}
    view.app.modified_sheets_b = set()
    view.app._start_background_thread = lambda *_args, **_kwargs: (_ for _ in ()).throw(
        AssertionError("disk async must not start after a column/user edit")
    )
    refresh_calls = []
    view._refresh_mode_switch_preserving_selection = (
        lambda *, rescan: refresh_calls.append(bool(rescan))
    )
    view._schedule_cached_only_diff_mode_switch = lambda _value: False
    view._refresh_interaction_gate = lambda: None
    view._persist_only_diff_setting_debounced = lambda: None

    smt.SheetView._toggle_only_diff(view)

    assert view._has_valid_only_diff_snapshot_cache()
    assert view._only_diff_rows_cache == expected, view._only_diff_rows_cache
    assert refresh_calls == [False], refresh_calls


def test_35_logical_slots_do_not_clip_to_34_physical_columns():
    slots = []
    for logical_col in range(1, 36):
        if logical_col == 18:
            slots.append(smt.ColumnSlot(logical_col - 1, mine_col=None, theirs_col=18, state="inserted"))
        elif logical_col < 18:
            slots.append(smt.ColumnSlot(logical_col - 1, mine_col=logical_col, theirs_col=logical_col))
        else:
            slots.append(smt.ColumnSlot(logical_col - 1, mine_col=logical_col - 1, theirs_col=logical_col))
    slots = tuple(slots)
    model = smt.ColumnModel.from_slots(
        _key("wide-35"), slots, blocks=smt._build_column_blocks(slots)
    )
    cache = smt.LogicalColumnComparisonCache(model, structural_diff_cols=frozenset((18,)))
    view = _fake_view(cache, max_col=35)
    view.col_char_widths = {col: 4 for col in range(1, 36)}
    projection = view._active_column_projection()
    assert projection.slot_count == 35
    assert projection.physical_col("A", 35) == 34
    assert projection.physical_col("B", 35) == 35
    assert len(view._project_raw_parts([f"A{col}" for col in range(1, 35)], "A")) == 35
    assert len(view._project_raw_parts([f"B{col}" for col in range(1, 36)], "B")) == 35
    _assert_monotonic_spans(view._base_spans(), 35)
    assert _fragments(view, view._build_col_header_line("A"))[-1] == "AH"
    assert _fragments(view, view._build_col_header_line("B"))[-1] == "AI"


def test_projection_cache_replay_and_invalidation_is_geometry_only():
    mine, theirs, first_cache = _two_way_fixture()
    view = _fake_view(first_cache, max_col=7)
    first_projection = view._active_column_projection()
    first_generation = view._column_projection_generation
    assert view._active_column_projection() is first_projection
    assert view._column_projection_generation == first_generation
    view._base_spans()
    view.row_pairs = [(1, 1)]
    view._cache_only_diff_rows_snapshot([0])
    assert view._has_valid_only_diff_snapshot_cache()

    identity_rows = _rows(("A", "B", "C", "D", "E", "F"))
    newer_cache = smt.build_logical_column_comparison_cache_2way(
        _key("two-way-geometry", 2, 2),
        identity_rows,
        identity_rows,
        identity_rows,
        identity_rows,
        mine_max_col=6,
        theirs_max_col=6,
    )
    # Simulate a background/new-version cache becoming authoritative.  The old
    # projection must not survive merely because a spans cache already exists.
    view.column_comparison_cache = newer_cache
    newer_projection = view._active_column_projection()
    assert newer_projection.model is view.column_comparison_cache.model
    assert newer_projection.model.slots == newer_cache.model.slots
    assert newer_projection.model.cache_key == view._expected_column_model_cache_key()
    assert newer_projection is not first_projection
    assert view._column_projection_generation == first_generation + 1
    assert view._base_spans_cache is None
    assert not view._has_valid_only_diff_snapshot_cache()
    assert view._logical_slot_count() == 6
    _assert_monotonic_spans(view._base_spans(), 6)

    # No-structure compatibility: identity mappings preserve prior behaviour.
    assert not newer_cache.structural_diff_cols
    assert view._project_raw_parts(identity_rows[2], "A") == list(identity_rows[2])
    assert view._project_raw_parts(identity_rows[2], "B") == list(identity_rows[2])
    view.pair_diff_cols = {0: set()}
    view.pair_base_diff_cols = {}
    view._diff_map_cache_version = None
    assert view._compute_diff_map_data() == (1, [], [])


def test_action_mapping_and_missing_or_unresolved_slots_are_safe():
    _mine, _theirs, cache = _two_way_fixture()
    view = _fake_view(cache, max_col=7)
    view._install_column_projection(cache)
    assert view._column_mapping_is_current()
    assert view._action_physical_columns("A2B", 4) == (2, 4)
    assert view._action_physical_columns("B2A", 4) == (4, 2)
    assert view._action_physical_columns("A2B", 7) == (5, 6)
    assert view._action_physical_columns("B2A", 7) == (6, 5)

    for direction, logical_col in (("A2B", 2), ("B2A", 5)):
        try:
            view._action_physical_columns(direction, logical_col)
        except RuntimeError as exc:
            assert "列结构操作" in str(exc), exc
        else:
            raise AssertionError((direction, logical_col, "missing-side action was not blocked"))

    base = _rows(("A", "B", "C"))
    mine = _rows(("A", "M", "B", "C"))
    theirs = _rows(("A", "T", "B", "C"))
    unresolved_cache = _cache_3way("unresolved-action", mine, base, theirs)
    unresolved_view = _fake_view(unresolved_cache, three_way=True, max_col=4)
    unresolved_view._install_column_projection(unresolved_cache)
    try:
        unresolved_view._action_physical_columns("A2B", 2)
    except RuntimeError as exc:
        assert "映射待确认" in str(exc), exc
    else:
        raise AssertionError("unresolved logical column action was not blocked")


def test_row_insertion_without_column_structure_stays_retained():
    # The Base row gap is a row-model concern.  It must not make the sole,
    # physically identical column ambiguous and block otherwise-safe BASE2A.
    mine = [("id",), ("A",), ("B",), ("C",)]
    base = [("id",), ("A",), (None,), ("C",)]
    theirs = list(mine)
    cache = smt.build_logical_column_comparison_cache_3way(
        _key("row-only-three-way"),
        mine,
        base,
        theirs,
        mine,
        base,
        theirs,
        mine_max_col=1,
        base_max_col=1,
        theirs_max_col=1,
    )
    assert len(cache.model.slots) == 1
    slot = cache.model.slots[0]
    assert (slot.mine_col, slot.base_col, slot.theirs_col) == (1, 1, 1)
    assert slot.state == "retained" and not slot.confidence.ambiguous, slot
    assert not cache.structural_diff_cols and not cache.unresolved_cols
    view = _fake_view(cache, three_way=True, max_col=1)
    view._install_column_projection(cache)
    assert view._action_physical_columns("BASE2A", 1) == (1, 1)


def test_value_edits_do_not_turn_a_same_order_column_into_structure():
    mine = [
        tuple([row_no, f"same-{row_no}"] + [f"payload-{col}-{row_no}" for col in range(3, 13)])
        for row_no in range(1, 11)
    ]
    theirs = [tuple(row) for row in mine]
    theirs = [list(row) for row in theirs]
    for excel_row in (2, 3, 7, 8):
        theirs[excel_row - 1][1] = f"changed-{excel_row}"
    theirs = [tuple(row) for row in theirs]
    cache = _cache_2way("value-edits-not-structure", mine, theirs)
    assert not cache.structural_diff_cols, cache.model.slots
    assert not cache.unresolved_cols, cache.model.slots
    assert all(slot.state == "retained" for slot in cache.model.slots), cache.model.slots

    view = _fake_view(cache, max_col=12)
    view.row_pairs = [(row, row) for row in range(1, 11)]
    view.pair_diff_cols = {}
    for pair_idx in range(10):
        result = smt.compare_logical_row_2way(
            cache,
            mine[pair_idx],
            theirs[pair_idx],
            mine[pair_idx],
            theirs[pair_idx],
            mine_row=pair_idx + 1,
            theirs_row=pair_idx + 1,
        )
        view.pair_diff_cols[pair_idx] = set(result.diff_cols)
    assert [idx for idx in range(10) if view._pair_has_visual_diff(idx)] == [1, 2, 6, 7]
    assert all(view._visual_diff_cols_for_pair(idx) == {2} for idx in (1, 2, 6, 7))


def test_unresolved_but_both_present_slot_keeps_raw_cell_differences_visible():
    confidence = smt.ColumnMappingConfidence(
        0.0,
        True,
        "low-confidence-physical-fallback",
        ("physical-order",),
        (smt.COLUMN_MAPPING_CAUSE_LOW_CONFIDENCE,),
    )
    slots = (
        smt.ColumnSlot(0, mine_col=1, theirs_col=1),
        smt.ColumnSlot(1, mine_col=2, theirs_col=2, state="unresolved", confidence=confidence),
        smt.ColumnSlot(2, mine_col=3, theirs_col=3),
    )
    model = smt.ColumnModel.from_slots(
        _key("unresolved-value-channel"),
        slots,
        blocks=smt._build_column_blocks(slots),
    )
    cache = smt.LogicalColumnComparisonCache(
        model=model,
        unresolved_cols=frozenset((2,)),
    )
    view = _fake_view(cache, max_col=3)
    view.row_pairs = [(1, 1), (2, 2)]
    view.pair_diff_cols = {0: {2}, 1: set()}
    assert view._all_logical_diff_cols_for_pair(0) == {2}
    assert view._visual_diff_cols_for_pair(0) == {2}
    assert view._pair_has_visual_diff(0)
    assert not view._pair_has_visual_diff(1)


def test_stale_lifecycle_rebuilds_before_geometry_and_append_has_no_free_rescan():
    mine, theirs, cache = _two_way_fixture()
    view = _fake_view(cache, max_col=7)
    view.row_pairs = [(row, row) for row in range(1, len(mine) + 1)]
    view.row_a_to_pair_idx = {row: row - 1 for row in range(1, len(mine) + 1)}
    view._install_column_projection(cache)
    view._base_spans()
    view._cache_only_diff_rows_snapshot([0])
    old_key = view._expected_column_model_cache_key()
    old_generation = view._column_projection_generation

    view._mark_column_mapping_stale(
        "synthetic-row-column-edit",
        row_structure=True,
        column_structure=True,
        edited_sides=("A", "B"),
    )
    expected_key = view._expected_column_model_cache_key()
    assert expected_key.row_model_version == old_key.row_model_version + 1
    assert expected_key.column_model_version == old_key.column_model_version + 1
    assert expected_key.mine_edit_version == old_key.mine_edit_version + 1
    assert expected_key.theirs_edit_version == old_key.theirs_edit_version + 1
    assert not view._column_mapping_is_current()
    assert not view._has_valid_only_diff_snapshot_cache()
    assert view._diff_map_cache is None and view._diff_map_cache_version is None
    try:
        view._ensure_column_projection_current("禁用重建测试", allow_rebuild=False)
    except RuntimeError as exc:
        assert "已过期" in str(exc), exc
    else:
        raise AssertionError("stale projection was accepted with rebuild disabled")

    wb_a = Workbook()
    wb_b = Workbook()
    ws_a = wb_a.active
    ws_b = wb_b.active
    ws_a.title = ws_b.title = "two-way-geometry"
    for row in mine:
        ws_a.append(list(row))
    for row in theirs:
        ws_b.append(list(row))
    view.app = SimpleNamespace(
        merge_mode=False,
        has_base=False,
        get_sheet_meta=lambda _sheet: {"view_mode": "normal"},
        ws_a_val=lambda _sheet: ws_a,
        ws_b_val=lambda _sheet: ws_b,
        ws_a_edit=lambda _sheet: ws_a,
        ws_b_edit=lambda _sheet: ws_b,
    )
    rebuilt = view._ensure_column_projection_current("合成重建测试")
    assert view._column_mapping_is_current()
    assert rebuilt.model.cache_key == expected_key
    assert view._column_projection_generation == old_generation + 1
    assert view._base_spans_cache is None

    # Once current, geometry/navigation must not touch worksheets again.
    def _poison(_sheet):
        raise AssertionError("current logical geometry unexpectedly reopened a worksheet")

    view.app.ws_a_val = _poison
    view.app.ws_b_val = _poison
    view.app.ws_a_edit = _poison
    view.app.ws_b_edit = _poison
    assert view._ensure_column_projection_current("缓存复用测试") is rebuilt
    view._base_spans()
    view._build_col_header_line("A")
    view._hit_col_from_char(0)
    wb_a.close()
    wb_b.close()

    # Guard the incremental-loading path against a previously regressed free
    # variable (``rescan`` existed in refresh but not in _append_rows).
    append_source = textwrap.dedent(inspect.getsource(smt.SheetView._append_rows))
    append_tree = ast.parse(append_source)
    loaded_rescan = any(
        isinstance(node, ast.Name)
        and node.id == "rescan"
        and isinstance(node.ctx, ast.Load)
        for node in ast.walk(append_tree)
    )
    assert not loaded_rescan or "rescan" in inspect.signature(smt.SheetView._append_rows).parameters, (
        "_append_rows reads undefined name 'rescan'"
    )


def test_undo_and_save_guard_mapping_before_consuming_operations():
    undo_events = []
    undo_view = object.__new__(smt.SheetView)
    undo_view._guard_mutation_ready = lambda _operation, **_kwargs: True
    undo_view._ensure_column_projection_current = lambda operation: undo_events.append(
        ("ensure", operation)
    )

    def _pop_undo():
        undo_events.append(("pop", None))
        return None

    undo_view.sheet = "S"
    undo_view.app = SimpleNamespace(
        pop_undo=_pop_undo,
        undo_stack=[{"sheet": "S", "kind": "fixture"}],
    )
    undo_view._undo_last_action()
    assert undo_events == [("ensure", "撤销操作"), ("pop", None)], undo_events

    class _SaveGuardReached(RuntimeError):
        pass

    save_events = []

    def _stop_at_save_guard(operation):
        save_events.append(operation)
        raise _SaveGuardReached(operation)

    guarded_view = SimpleNamespace(
        _data_ready=True,
        _ensure_column_projection_current=_stop_at_save_guard,
    )
    fake_app = object.__new__(smt.SowMergeApp)
    fake_app.sheet_views = {"S": guarded_view}
    try:
        fake_app.build_manual_merge_output_file()
    except _SaveGuardReached as exc:
        assert str(exc) == "构建合并输出", exc
    else:
        raise AssertionError("save continued without validating the stale mapping")
    assert save_events == ["构建合并输出"], save_events


def _save_book(path: str, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    for row in rows:
        ws.append(list(row))
    wb.save(path)
    wb.close()


def _pump(root, seconds: float = 0.05):
    deadline = time.time() + seconds
    while time.time() < deadline:
        root.update_idletasks()
        root.update()
        time.sleep(0.01)


def _wait_for_view(app, sheet="S", timeout=12.0):
    app.nb.select(app._sheet_containers[sheet])
    deadline = time.time() + timeout
    view = None
    while time.time() < deadline:
        _pump(app.root)
        view = app.sheet_views.get(sheet)
        if view is not None and getattr(view, "_data_ready", False):
            return view
    raise AssertionError(f"logical-column GUI view did not become ready: {view!r}")


def test_real_gui_background_replay_only_diff_sync_and_click_x():
    mine, theirs, expected_cache = _two_way_fixture()
    dir_a = make_temp_dir("sow_logical_geometry_a_")
    dir_b = make_temp_dir("sow_logical_geometry_b_")
    path_a = os.path.join(dir_a, "same.xlsx")
    path_b = os.path.join(dir_b, "same.xlsx")
    _save_book(path_a, mine)
    _save_book(path_b, theirs)

    app = smt.SowMergeApp(path_a, path_b)
    try:
        app.root.geometry("1000x760")
        view = _wait_for_view(app)
        _pump(app.root, 0.15)

        projection = view._active_column_projection()
        assert projection.model is view.column_comparison_cache.model
        assert projection.slot_count == 8, projection.slots
        expected_pairs = [
            (slot.mine_col, slot.theirs_col) for slot in expected_cache.model.slots
        ]
        assert [(slot.mine_col, slot.theirs_col) for slot in projection.slots] == expected_pairs

        # Never inherit the user's persisted only-diff preference in this test.
        if bool(view.only_diff_var.get()):
            view.only_diff_var.set(0)
            view._toggle_only_diff()
            view.refresh(row_only=None, rescan=False)
            _pump(app.root, 0.1)

        first_a = view.left.get("1.0", "1.end")
        first_b = view.right.get("1.0", "1.end")
        assert _fragments(view, first_a) == view._project_raw_parts(mine[0], "A"), (
            _fragments(view, first_a), view._project_raw_parts(mine[0], "A")
        )
        assert _fragments(view, first_b) == view._project_raw_parts(theirs[0], "B"), (
            _fragments(view, first_b), view._project_raw_parts(theirs[0], "B")
        )
        assert len(_fragments(view, first_a)) == len(_fragments(view, first_b)) == 8

        header_a = view.left_colhdr.get("1.0", "1.end")
        header_b = view.right_colhdr.get("1.0", "1.end")
        assert _fragments(view, header_a) == [projection.header_label("A", col) for col in range(1, 9)]
        assert _fragments(view, header_b) == [projection.header_label("B", col) for col in range(1, 9)]
        assert view._spans_for_line(first_a) == view._spans_for_line(first_b)
        assert view._spans_for_line(header_a) == view._spans_for_line(first_a)

        spans = view._base_spans()
        for logical_col in (2, 3):
            start, _end = spans[logical_col]
            assert "paddingcol" in view.left.tag_names(f"1.{start}"), (logical_col, view.left.tag_names(f"1.{start}"))
        start, _end = spans[5]
        assert "paddingcol" in view.right.tag_names(f"1.{start}"), view.right.tag_names(f"1.{start}")

        # Only the retained E edit is a row-level difference; structural slots
        # remain visible in headers/horizontal minimap without flooding all rows.
        edit_pair = next(
            idx for idx, pair in enumerate(view.row_pairs) if pair == (6, 6)
        )
        view._diff_map_cache_version = None
        _total, diff_rows, diff_cols = view._compute_diff_map_data()
        assert diff_rows == [edit_pair], (diff_rows, view.pair_diff_cols)
        assert set(diff_cols) == {2, 3, 5, 7}, diff_cols
        view.only_diff_var.set(1)
        view._toggle_only_diff()
        view.refresh(row_only=None, rescan=False)
        _pump(app.root, 0.15)
        assert view.display_rows == [edit_pair], view.display_rows

        # Shared logical width keeps all panes and their headers horizontally aligned.
        only_diff_line = view.row_to_line[edit_pair]
        view.left.mark_set("insert", f"{only_diff_line}.0")
        view.right.mark_set("insert", f"{only_diff_line}.0")
        view._highlight_selected_line(only_diff_line)
        view.selected_pair_idx = edit_pair
        view._update_cursor_lines()
        _pump(app.root, 0.05)
        view._sync_main_x_to_frac(0.42)
        view._sync_c_x_to_frac(0.42)
        _pump(app.root, 0.05)
        main_first = float(view.left.xview()[0])
        for widget in (view.right, view.left_colhdr, view.right_colhdr):
            assert abs(float(widget.xview()[0]) - main_first) < 0.025, (
                widget, widget.xview(), view.left.xview()
            )
        expected_c = view._map_xfirst_between_widgets(
            view.left,
            view.cursor_cmp,
            main_first,
        )
        for widget in (view.cursor_cmp, view.cursor_cmp_colhdr):
            assert abs(float(widget.xview()[0]) - expected_c) < 0.025, (
                widget, widget.xview(), expected_c
            )

        # Click a visible character after horizontal scrolling.  Tk converts
        # viewport x to content x; selection must remain logical column 7.
        view.only_diff_var.set(0)
        view._toggle_only_diff()
        view.refresh(row_only=None, rescan=False)
        _pump(app.root, 0.05)
        target_line = view.row_to_line[edit_pair]
        start, end = view._base_spans()[7]
        char_pos = start + max(0, (end - start) // 2)
        view.left.see(f"{target_line}.{char_pos}")
        _pump(app.root, 0.05)
        box = view.left.bbox(f"{target_line}.{char_pos}")
        assert box is not None, (target_line, char_pos, view.left.xview())
        x, y, width, height = box
        original_widget_line = view._widget_line
        view._widget_line = lambda _widget: target_line
        try:
            view._select_from_widget(
                view.left,
                SimpleNamespace(x=x + max(1, width // 2), y=y + max(1, height // 2)),
            )
        finally:
            view._widget_line = original_widget_line
        assert view._main_sel_col == 7
        assert view._physical_col_for_logical("A", view._main_sel_col) == 5
        assert view._physical_col_for_logical("B", view._main_sel_col) == 6

        payload = view._cmp_tooltip_payload_by_pair_col(edit_pair, 7, force_panel=True)
        assert payload is not None
        assert "e-5-independent-edit" in payload[0]
        assert "e-5-" in payload[0]

        # A stale edit version must force a rebuild before refresh renders or
        # navigates with the previous projection/spans generation.
        stale_generation = view._column_projection_generation
        view._mark_column_mapping_stale("gui-stale-lifecycle", edited_sides=("A",))
        assert not view._column_mapping_is_current()
        view.refresh(row_only=None, rescan=False)
        _pump(app.root, 0.1)
        assert view._column_mapping_is_current()
        assert view._column_projection_generation > stale_generation
    finally:
        try:
            app._shutdown_root()
        except Exception:
            pass


def main():
    tests = (
        test_two_way_projection_headers_spans_and_placeholders,
        test_three_way_independent_and_same_anchor_geometry,
        test_hit_selection_hover_and_mixed_row_column_mapping,
        test_only_diff_minimap_uses_logical_slots_without_row_flooding,
        test_three_way_common_base_missing_column_does_not_flood_visual_rows,
        test_only_diff_toggle_after_column_edit_caches_current_visual_rows_synchronously,
        test_35_logical_slots_do_not_clip_to_34_physical_columns,
        test_projection_cache_replay_and_invalidation_is_geometry_only,
        test_action_mapping_and_missing_or_unresolved_slots_are_safe,
        test_row_insertion_without_column_structure_stays_retained,
        test_value_edits_do_not_turn_a_same_order_column_into_structure,
        test_unresolved_but_both_present_slot_keeps_raw_cell_differences_visible,
        test_stale_lifecycle_rebuilds_before_geometry_and_append_has_no_free_rescan,
        test_undo_and_save_guard_mapping_before_consuming_operations,
        test_real_gui_background_replay_only_diff_sync_and_click_x,
    )
    for test in tests:
        test()
        print(f"PASS {test.__name__}")
    print("GUI_SELF_TEST_LOGICAL_COLUMN_GEOMETRY_OK")


if __name__ == "__main__":
    main()

"""Fresh-process, non-Excel UX hotspot profiler for real workbook fixtures."""

from __future__ import annotations

import argparse
import contextlib
import cProfile
import io
import json
import os
import pstats
import time
from collections import defaultdict

import sow_merge_tool as smt
import _ux_5_3_final_acceptance as ux


CASES = {
    "WorldMonster": "WorldMonster@design",
    "Dungeon": "Dungeon@design",
}
DEFAULT_ROOT = r"D:\Tools\sow_merge_tool_proj\.perf_ux_hotspot_isolate_20260723"


class Probe:
    def __init__(self, enabled: bool):
        self.enabled = bool(enabled)
        self.phase = "startup"
        self.counts = defaultdict(lambda: defaultdict(int))
        self.timings = defaultdict(lambda: defaultdict(float))
        self.profiles: dict[str, str] = {}

    @contextlib.contextmanager
    def section(self, name: str, *, profile: bool = False):
        previous = self.phase
        self.phase = name
        profiler = cProfile.Profile() if self.enabled and profile else None
        if profiler is not None:
            profiler.enable()
        try:
            yield
        finally:
            if profiler is not None:
                profiler.disable()
                stream = io.StringIO()
                pstats.Stats(profiler, stream=stream).strip_dirs().sort_stats(
                    "cumulative"
                ).print_stats(30)
                self.profiles[name] = stream.getvalue()
            self.phase = previous

    def add_call(self, name: str, elapsed: float = 0.0, count: int = 1):
        phase = self.phase
        self.counts[phase][name] += int(count)
        self.timings[phase][name] += float(elapsed)

    def install(self):
        if not self.enabled:
            return
        from openpyxl.worksheet._read_only import ReadOnlyWorksheet
        from openpyxl.worksheet.worksheet import Worksheet

        original_cell = Worksheet.cell

        def cell(worksheet, *args, **kwargs):
            self.add_call("Worksheet.cell")
            return original_cell(worksheet, *args, **kwargs)

        Worksheet.cell = cell

        def wrap_iter_rows(worksheet_type, label):
            original_iter_rows = worksheet_type.iter_rows

            def iter_rows(worksheet, *args, **kwargs):
                call_phase = self.phase
                self.counts[call_phase][f"{label}.iter_rows"] += 1
                source = original_iter_rows(worksheet, *args, **kwargs)

                def counted():
                    for row in source:
                        self.counts[call_phase][f"{label}.iter_rows.rows"] += 1
                        self.counts[call_phase][f"{label}.iter_rows.cells"] += len(row)
                        yield row

                return counted()

            worksheet_type.iter_rows = iter_rows

        wrap_iter_rows(Worksheet, "Worksheet")
        wrap_iter_rows(ReadOnlyWorksheet, "ReadOnlyWorksheet")

        self._wrap_module("load_workbook")
        self._wrap_module("_read_rows_into_cache")
        self._wrap_module("_read_rows_into_shared_cache")
        for name in (
            "_ensure_column_projection_current",
            "_rebuild_column_comparison_cache_from_worksheets",
            "_precompute_large_diff_by_blocks",
            "_quick_diff_cols_from_value_rows",
            "_build_row_parts_and_diff_pair_from_values",
            "_compute_base_diff_cols_from_values",
            "_recalc_row_diff_and_update",
            "_column_action_workbook_snapshot",
            "_prepare_column_copy_payload",
            "_apply_column_copy_payload",
        ):
            if hasattr(smt.SheetView, name):
                self._wrap_method(name)
        self._wrap_static_method("_capture_column_formula_transformations")

        original_refresh = smt.SheetView.refresh

        def refresh(view, row_only, rescan, *args, **kwargs):
            phase = self.phase
            self.counts[phase]["refresh.calls"] += 1
            self.counts[phase][f"refresh.rescan_{bool(rescan)}"] += 1
            self.counts[phase][
                "refresh.row_full" if row_only is None else "refresh.row_partial"
            ] += 1
            started = time.perf_counter()
            try:
                return original_refresh(view, row_only, rescan, *args, **kwargs)
            finally:
                self.timings[phase]["refresh"] += time.perf_counter() - started

        smt.SheetView.refresh = refresh

    def _wrap_module(self, name: str):
        original = getattr(smt, name)

        def wrapped(*args, **kwargs):
            phase = self.phase
            self.counts[phase][name] += 1
            started = time.perf_counter()
            try:
                return original(*args, **kwargs)
            finally:
                self.timings[phase][name] += time.perf_counter() - started

        setattr(smt, name, wrapped)

    def _wrap_method(self, name: str):
        original = getattr(smt.SheetView, name)

        def wrapped(view, *args, **kwargs):
            phase = self.phase
            self.counts[phase][name] += 1
            started = time.perf_counter()
            try:
                return original(view, *args, **kwargs)
            finally:
                self.timings[phase][name] += time.perf_counter() - started

        setattr(smt.SheetView, name, wrapped)

    def _wrap_static_method(self, name: str):
        original = getattr(smt.SheetView, name)

        def wrapped(*args, **kwargs):
            phase = self.phase
            self.counts[phase][name] += 1
            started = time.perf_counter()
            try:
                return original(*args, **kwargs)
            finally:
                self.timings[phase][name] += time.perf_counter() - started

        setattr(smt.SheetView, name, staticmethod(wrapped))

    def result(self):
        return {
            "counts": {phase: dict(values) for phase, values in self.counts.items()},
            "timings_ms": {
                phase: {name: round(value * 1000, 3) for name, value in values.items()}
                for phase, values in self.timings.items()
            },
            "cprofile": self.profiles,
        }


def _wait_initial_ready_event(app, timeout: float = 240.0) -> bool:
    event = getattr(app, "_initial_sheet_ready_event", None)
    if event is None:
        return False
    deadline = time.perf_counter() + timeout
    while time.perf_counter() < deadline:
        if event.is_set():
            return True
        ux._pump(app.root, 0.03)
    return bool(event.is_set())


def _run_worldmonster(view, app, probe: Probe):
    result = {}
    with probe.section("only_diff", profile=True):
        result["only_diff"] = ux.only_diff_metrics(view)
    pair_idx = next(
        (idx for idx, pair in enumerate(view.row_pairs) if pair == (3000, 3000)),
        None,
    )
    if pair_idx is None:
        raise AssertionError("WorldMonster row 3000 pair not found")
    mine_local_before = app.ws_a_edit(view.sheet).cell(3000, 2).value

    started = time.perf_counter()
    with probe.section("cell_apply", profile=True):
        view._copy_single_cell_by_pair(pair_idx, "B2A", 3)
        ux._pump(app.root, 0.2)
    result["cell_apply_ms"] = round((time.perf_counter() - started) * 1000, 3)

    started = time.perf_counter()
    with probe.section("undo", profile=True):
        view._undo_last_action()
        ux._pump(app.root, 0.2)
    result["undo_ms"] = round((time.perf_counter() - started) * 1000, 3)

    started = time.perf_counter()
    with probe.section("redo", profile=True):
        view._copy_single_cell_by_pair(pair_idx, "B2A", 3)
        ux._pump(app.root, 0.2)
    result["redo_ms"] = round((time.perf_counter() - started) * 1000, 3)
    result["adopted"] = app.ws_a_edit(view.sheet).cell(3000, 3).value
    result["mine_local_preserved"] = (
        app.ws_a_edit(view.sheet).cell(3000, 2).value == mine_local_before
    )
    return result


def _run_dungeon(view, app, probe: Probe):
    with probe.section("only_diff", profile=True):
        only_diff = ux.only_diff_metrics(view)
    started = time.perf_counter()
    with probe.section("column_actions", profile=True):
        actions = ux.apply_all_structural(
            view, app, source_side="B", validate_undo=True
        )
    return {
        "only_diff": only_diff,
        "actions": actions,
        "column_actions_total_ms": round((time.perf_counter() - started) * 1000, 3),
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--case", choices=tuple(CASES), required=True)
    parser.add_argument("--root", default=DEFAULT_ROOT)
    parser.add_argument("--deep-profile", action="store_true")
    args = parser.parse_args()

    probe = Probe(args.deep_profile)
    probe.install()
    case_root = os.path.join(args.root, args.case)
    mine = os.path.join(case_root, "mine", args.case + ".xlsx")
    theirs = os.path.join(case_root, "theirs", args.case + ".xlsx")
    if not os.path.isfile(mine) or not os.path.isfile(theirs):
        raise FileNotFoundError(case_root)

    app = None
    result = {"case": args.case, "pid": os.getpid(), "deep_profile": args.deep_profile}
    total_started = time.perf_counter()
    try:
        started = time.perf_counter()
        with probe.section("constructor"):
            app = smt.SowMergeApp(mine, theirs)
        result["constructor_ms"] = round((time.perf_counter() - started) * 1000, 3)
        app.root.withdraw()

        with probe.section("initial_ready_event"):
            event_ready = _wait_initial_ready_event(app)
        result["initial_ready_event_set"] = event_ready
        result["initial_ready_event_ms"] = round(
            (time.perf_counter() - total_started) * 1000, 3
        )

        stable_started = time.perf_counter()
        with probe.section("full_projection_stable"):
            view = ux._force_full_view(
                ux._wait_for_view(app, CASES[args.case], timeout=240.0)
            )
            ux.wait_edit_ready(app, timeout=240.0)
            ux.wait_view_ready(view, timeout=240.0)
            ux._wait_for_stable_projection(view, timeout=240.0, stable_for=0.5)
        result["full_projection_phase_ms"] = round(
            (time.perf_counter() - stable_started) * 1000, 3
        )
        result["full_projection_stable_ms"] = round(
            (time.perf_counter() - total_started) * 1000, 3
        )
        result["view_before"] = ux.view_metrics(view)

        if args.case == "WorldMonster":
            result.update(_run_worldmonster(view, app, probe))
        else:
            result.update(_run_dungeon(view, app, probe))
        result["probe"] = probe.result()
        print(json.dumps(result, ensure_ascii=False, sort_keys=True), flush=True)
    finally:
        ux.close_app(app)


if __name__ == "__main__":
    main()

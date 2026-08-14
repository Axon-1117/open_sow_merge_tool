"""Performance and cache-boundary guards for logical column alignment.

This test intentionally uses only in-memory row caches and immutable signature
snapshots.  It records timing and operation counts so a fast result cannot hide
worksheet rescans or repeated alignment work.
"""

from __future__ import annotations

import gc
import json
import math
import statistics
import threading
import time

from openpyxl.worksheet.worksheet import Worksheet

import sow_merge_tool as mod


_REPEATS = 3
_TWO_WAY_LIMIT_MS = {35: 50.0, 68: 50.0, 256: 250.0, 257: 5.0}
_THREE_WAY_LIMIT_MS = {35: 50.0, 68: 50.0, 256: 500.0, 257: 20.0}


class _SequentialRows:
    """One-pass row cache that rejects random/indexed access."""

    def __init__(self, rows, name: str):
        self._rows = tuple(tuple(row) for row in rows)
        self.name = name
        self.iterations = 0
        self.rows_yielded = 0

    def __iter__(self):
        self.iterations += 1
        if self.iterations > 1:
            raise AssertionError(f"{self.name} row cache was rescanned")
        for row in self._rows:
            self.rows_yielded += 1
            yield row

    def __getitem__(self, _index):
        raise AssertionError(f"{self.name} row cache was accessed by index")


class _WorksheetReadGuard:
    def __init__(self):
        self.cell_calls = 0
        self.iter_rows_calls = 0

    def __enter__(self):
        self._original_cell = Worksheet.cell
        self._original_iter_rows = Worksheet.iter_rows

        def _unexpected_cell(*_args, **_kwargs):
            self.cell_calls += 1
            raise AssertionError("column cache path called Worksheet.cell")

        def _unexpected_iter_rows(*_args, **_kwargs):
            self.iter_rows_calls += 1
            raise AssertionError("column cache path called Worksheet.iter_rows")

        Worksheet.cell = _unexpected_cell
        Worksheet.iter_rows = _unexpected_iter_rows
        return self

    def __exit__(self, exc_type, exc, traceback):
        Worksheet.cell = self._original_cell
        Worksheet.iter_rows = self._original_iter_rows

    @property
    def total_calls(self) -> int:
        return self.cell_calls + self.iter_rows_calls


class _RebuildGuard:
    """Fail if a populated comparison cache re-enters any build/alignment API."""

    _NAMES = (
        "build_column_signatures_from_row_cache",
        "build_column_signature_snapshot",
        "align_column_signatures_2way",
        "align_column_signatures_3way",
        "build_logical_column_comparison_cache_2way",
        "build_logical_column_comparison_cache_3way",
    )

    def __init__(self):
        self.calls = {name: 0 for name in self._NAMES}

    def __enter__(self):
        self._originals = {name: getattr(mod, name) for name in self._NAMES}
        for name in self._NAMES:
            def _unexpected(*_args, _name=name, **_kwargs):
                self.calls[_name] += 1
                raise AssertionError(f"cached replay re-entered {_name}")

            setattr(mod, name, _unexpected)
        return self

    def __exit__(self, exc_type, exc, traceback):
        for name, original in self._originals.items():
            setattr(mod, name, original)


def _high_similarity_signatures(width: int, identity_prefix: str):
    """No exact keys, but one unique high-similarity peer per column."""
    return tuple(
        mod.ColumnSignature(
            physical_col=index + 1,
            row_count=12,
            non_empty_count=12,
            first_non_empty_row=1,
            last_non_empty_row=12,
            header_signals=(f"header-{index:04d}",),
            representative_signals=(
                (1, f"value-{index:04d}"),
                (12, f"tail-{index:04d}"),
            ),
            non_empty_pattern=(1, 12),
            formula_signals=(),
            intrinsic_key=f"{identity_prefix}-{index:04d}",
            left_context_key=f"context-{index - 1:04d}" if index else None,
            right_context_key=f"context-{index + 1:04d}",
        )
        for index in range(width)
    )


def _measure_ms(callback, repeats: int = _REPEATS):
    samples = []
    last_result = None
    for _repeat in range(repeats):
        gc.collect()
        started = time.perf_counter()
        last_result = callback()
        samples.append((time.perf_counter() - started) * 1000)
    ordered = sorted(samples)
    return last_result, {
        "raw": samples,
        "median": statistics.median(samples),
        "p95_nearest_rank": ordered[math.ceil(0.95 * len(ordered)) - 1],
    }


def _count_similarity_calls(callback) -> tuple[object, int]:
    original = mod._column_signature_similarity_prepared
    calls = 0

    def _counted(*args, **kwargs):
        nonlocal calls
        calls += 1
        return original(*args, **kwargs)

    mod._column_signature_similarity_prepared = _counted
    try:
        return callback(), calls
    finally:
        mod._column_signature_similarity_prepared = original


def _rows_for_labels(labels, row_count: int = 32):
    labels = tuple(labels)
    return tuple(
        labels if row == 0 else tuple(f"{label}:{row}" for label in labels)
        for row in range(row_count)
    )


def _insert_labels(base_labels, boundary: int, inserted_labels):
    base_labels = tuple(base_labels)
    return (
        base_labels[:boundary]
        + tuple(inserted_labels)
        + base_labels[boundary:]
    )


def _sequential_cache(rows, name: str):
    return _SequentialRows(rows, name), _SequentialRows(rows, name + "_edit")


def _test_signature_uses_one_pass_sequential_cache(metrics: dict):
    rows = tuple(
        tuple(
            f"H{column}" if row == 0 else row * 1000 + column
            for column in range(35)
        )
        for row in range(64)
    )
    edit_rows = tuple(
        tuple(
            f"=R{row + 1}C{column + 1}" if row and column % 7 == 0 else value
            for column, value in enumerate(row_values)
        )
        for row, row_values in enumerate(rows)
    )
    values = _SequentialRows(rows, "value")
    edits = _SequentialRows(edit_rows, "edit")
    key = mod.ColumnModelCacheKey("PerfGuard", 1, 1)

    with _WorksheetReadGuard() as worksheet_guard:
        started = time.perf_counter()
        snapshot = mod.build_column_signature_snapshot(
            key,
            values,
            edits,
            max_col=35,
        )
        elapsed_ms = (time.perf_counter() - started) * 1000

    assert len(snapshot.signatures) == 35
    assert values.iterations == edits.iterations == 1
    assert values.rows_yielded == edits.rows_yielded == 64
    assert worksheet_guard.total_calls == 0
    metrics["sequential_signature"] = {
        "elapsed_ms": elapsed_ms,
        "value_cache_iterations": values.iterations,
        "edit_cache_iterations": edits.iterations,
        "value_rows_yielded": values.rows_yielded,
        "edit_rows_yielded": edits.rows_yielded,
        "worksheet_cell_calls": worksheet_guard.cell_calls,
        "worksheet_iter_rows_calls": worksheet_guard.iter_rows_calls,
    }


def _test_mapping_width_guards(metrics: dict):
    width_metrics = {}
    with _WorksheetReadGuard() as worksheet_guard:
        for width in (35, 68, 256, 257):
            left = _high_similarity_signatures(width, "left")
            right = _high_similarity_signatures(width, "right")
            base = _high_similarity_signatures(width, "base")
            mine = _high_similarity_signatures(width, "mine")
            theirs = _high_similarity_signatures(width, "theirs")

            two_way_result, two_way_calls = _count_similarity_calls(
                lambda: mod.align_column_signatures_2way(left, right)
            )
            three_way_result, three_way_calls = _count_similarity_calls(
                lambda: mod.align_column_signatures_3way(mine, base, theirs)
            )
            _last_two_way, two_way_timing = _measure_ms(
                lambda: mod.align_column_signatures_2way(left, right)
            )
            _last_three_way, three_way_timing = _measure_ms(
                lambda: mod.align_column_signatures_3way(mine, base, theirs)
            )

            if width <= 256:
                assert two_way_calls == width * width, (width, two_way_calls)
                assert three_way_calls == 2 * width * width, (
                    width,
                    three_way_calls,
                )
                assert not two_way_result.used_physical_fallback
                assert not three_way_result.used_physical_fallback
                assert len(two_way_result.anchor_pairs) == width
                assert len(three_way_result.model.slots) == width
            else:
                assert two_way_calls == three_way_calls == 0
                assert two_way_result.used_physical_fallback
                assert two_way_result.fallback_reason == "column-limit-exceeded"
                assert three_way_result.used_physical_fallback
                assert "mine-to-base:column-limit-exceeded" in (
                    three_way_result.fallback_reason
                )
                assert "theirs-to-base:column-limit-exceeded" in (
                    three_way_result.fallback_reason
                )

            assert (
                two_way_timing["p95_nearest_rank"] <= _TWO_WAY_LIMIT_MS[width]
            ), (width, two_way_timing)
            assert (
                three_way_timing["p95_nearest_rank"] <= _THREE_WAY_LIMIT_MS[width]
            ), (width, three_way_timing)
            width_metrics[str(width)] = {
                "similarity_calls_2way": two_way_calls,
                "similarity_calls_3way": three_way_calls,
                "two_way_timing_ms": two_way_timing,
                "three_way_timing_ms": three_way_timing,
                "two_way_limit_ms": _TWO_WAY_LIMIT_MS[width],
                "three_way_limit_ms": _THREE_WAY_LIMIT_MS[width],
                "fallback": width > 256,
            }

    assert worksheet_guard.total_calls == 0
    metrics["mapping_widths"] = width_metrics
    metrics["mapping_worksheet_reads"] = {
        "cell": worksheet_guard.cell_calls,
        "iter_rows": worksheet_guard.iter_rows_calls,
    }


def _build_replay_fixtures(metrics: dict):
    base_labels = tuple(f"C{index:02d}" for index in range(35))
    mine_labels_2way = base_labels
    theirs_labels_2way = _insert_labels(base_labels, 9, ("T_INS_A", "T_INS_B"))
    mine_rows_2way = _rows_for_labels(mine_labels_2way)
    theirs_rows_2way = _rows_for_labels(theirs_labels_2way)
    mine_values_2way, mine_edits_2way = _sequential_cache(
        mine_rows_2way, "2way_mine"
    )
    theirs_values_2way, theirs_edits_2way = _sequential_cache(
        theirs_rows_2way, "2way_theirs"
    )

    mine_labels_3way = _insert_labels(base_labels, 5, ("M_INS",))
    theirs_labels_3way = _insert_labels(base_labels, 24, ("T_INS",))
    mine_rows_3way = _rows_for_labels(mine_labels_3way)
    base_rows_3way = _rows_for_labels(base_labels)
    theirs_rows_3way = _rows_for_labels(theirs_labels_3way)
    mine_values_3way, mine_edits_3way = _sequential_cache(
        mine_rows_3way, "3way_mine"
    )
    base_values_3way, base_edits_3way = _sequential_cache(
        base_rows_3way, "3way_base"
    )
    theirs_values_3way, theirs_edits_3way = _sequential_cache(
        theirs_rows_3way, "3way_theirs"
    )

    with _WorksheetReadGuard() as worksheet_guard:
        started = time.perf_counter()
        cache_2way = mod.build_logical_column_comparison_cache_2way(
            mod.ColumnModelCacheKey("PerfReplay2", 1, 1),
            mine_values_2way,
            theirs_values_2way,
            mine_edits_2way,
            theirs_edits_2way,
            mine_max_col=len(mine_labels_2way),
            theirs_max_col=len(theirs_labels_2way),
        )
        build_2way_ms = (time.perf_counter() - started) * 1000
        started = time.perf_counter()
        cache_3way = mod.build_logical_column_comparison_cache_3way(
            mod.ColumnModelCacheKey("PerfReplay3", 1, 1),
            mine_values_3way,
            base_values_3way,
            theirs_values_3way,
            mine_edits_3way,
            base_edits_3way,
            theirs_edits_3way,
            mine_max_col=len(mine_labels_3way),
            base_max_col=len(base_labels),
            theirs_max_col=len(theirs_labels_3way),
        )
        build_3way_ms = (time.perf_counter() - started) * 1000

    sequential_inputs = (
        mine_values_2way,
        mine_edits_2way,
        theirs_values_2way,
        theirs_edits_2way,
        mine_values_3way,
        mine_edits_3way,
        base_values_3way,
        base_edits_3way,
        theirs_values_3way,
        theirs_edits_3way,
    )
    assert all(cache.iterations == 1 for cache in sequential_inputs)
    assert all(cache.rows_yielded == 32 for cache in sequential_inputs)
    assert worksheet_guard.total_calls == 0
    assert len(cache_2way.structural_diff_cols) == 2
    assert len(cache_3way.structural_diff_cols) == 2
    metrics["comparison_cache_build"] = {
        "two_way_ms": build_2way_ms,
        "three_way_ms": build_3way_ms,
        "sequential_inputs": len(sequential_inputs),
        "iterations_per_input": 1,
        "rows_per_input": 32,
        "worksheet_cell_calls": worksheet_guard.cell_calls,
        "worksheet_iter_rows_calls": worksheet_guard.iter_rows_calls,
    }
    return {
        "cache_2way": cache_2way,
        "cache_3way": cache_3way,
        "mine_row_2way": mine_rows_2way[7],
        "theirs_row_2way": theirs_rows_2way[7],
        "mine_row_3way": mine_rows_3way[7],
        "base_row_3way": base_rows_3way[7],
        "theirs_row_3way": theirs_rows_3way[7],
    }


def _batch(callback, operations: int):
    result = None
    for index in range(operations):
        result = callback(index)
    return result


def _measure_replay_path(callback, operations: int, limit_per_operation_ms: float):
    last, timing = _measure_ms(lambda: _batch(callback, operations))
    timing["operations_per_repeat"] = operations
    timing["p95_per_operation_ms"] = (
        timing["p95_nearest_rank"] / operations
    )
    timing["limit_per_operation_ms"] = limit_per_operation_ms
    assert timing["p95_per_operation_ms"] <= limit_per_operation_ms, timing
    return last, timing


def _test_cached_replay_and_interaction_paths(metrics: dict):
    fixture = _build_replay_fixtures(metrics)
    cache_2way = fixture["cache_2way"]
    cache_3way = fixture["cache_3way"]
    model = cache_3way.model
    structural = tuple(sorted(cache_3way.structural_diff_cols))
    assert structural

    replay_metrics = {}
    with _WorksheetReadGuard() as worksheet_guard, _RebuildGuard() as rebuild_guard:
        comparison_2way, replay_metrics["comparison_replay_2way"] = (
            _measure_replay_path(
                lambda _index: mod.compare_logical_row_2way(
                    cache_2way,
                    fixture["mine_row_2way"],
                    fixture["theirs_row_2way"],
                    fixture["mine_row_2way"],
                    fixture["theirs_row_2way"],
                    mine_row=8,
                    theirs_row=8,
                ),
                500,
                50.0,
            )
        )
        assert comparison_2way.diff_cols == cache_2way.structural_diff_cols

        comparison_3way, replay_metrics["comparison_replay_3way"] = (
            _measure_replay_path(
                lambda _index: mod.compare_logical_row_3way(
                    cache_3way,
                    fixture["mine_row_3way"],
                    fixture["base_row_3way"],
                    fixture["theirs_row_3way"],
                    fixture["mine_row_3way"],
                    fixture["base_row_3way"],
                    fixture["theirs_row_3way"],
                    mine_row=8,
                    base_row=8,
                    theirs_row=8,
                ),
                500,
                50.0,
            )
        )
        assert not comparison_3way.mine_changed_cols
        assert not comparison_3way.theirs_changed_cols
        assert not comparison_3way.conflict_cols

        analysis, replay_metrics["block_conflict_presentation"] = (
            _measure_replay_path(
                lambda _index: mod.classify_logical_columns_3way(cache_3way),
                500,
                50.0,
            )
        )
        assert sum(state.state == "inserted" for state in analysis.states) == 2
        assert not analysis.structural_conflicts

        selection, replay_metrics["selection_lookup"] = _measure_replay_path(
            lambda index: (
                model.physical_for_logical("mine", index % len(model.slots)),
                model.physical_for_logical("base", index % len(model.slots)),
                model.physical_for_logical("theirs", index % len(model.slots)),
            ),
            2000,
            50.0,
        )
        assert len(selection) == 3

        _scroll, replay_metrics["scroll_window"] = _measure_replay_path(
            lambda index: tuple(
                (
                    slot.logical_idx,
                    slot.mine_col,
                    slot.base_col,
                    slot.theirs_col,
                )
                for slot in model.slots[
                    index % max(1, len(model.slots) - 8) :
                    index % max(1, len(model.slots) - 8) + 8
                ]
            ),
            2000,
            250.0,
        )

        next_target, replay_metrics["navigation_target"] = _measure_replay_path(
            lambda index: next(
                (logical for logical in structural if logical > index % len(model.slots)),
                structural[0],
            ),
            2000,
            250.0,
        )
        assert next_target in structural

        block_rows, replay_metrics["block_presentation"] = _measure_replay_path(
            lambda _index: tuple(
                (
                    block.ordinal,
                    block.start_slot_idx,
                    block.end_slot_idx,
                    block.state,
                    block.confidence.reason,
                )
                for block in model.blocks
            ),
            2000,
            50.0,
        )
        assert block_rows

    assert worksheet_guard.total_calls == 0
    assert all(count == 0 for count in rebuild_guard.calls.values())
    metrics["cached_replay"] = {
        "paths": replay_metrics,
        "worksheet_reads": {
            "cell": worksheet_guard.cell_calls,
            "iter_rows": worksheet_guard.iter_rows_calls,
        },
        "rebuild_or_realign_calls": rebuild_guard.calls,
    }


def _test_single_owner_edit_preload(metrics: dict):
    """Concurrent callers share one parse and preserve an edited workbook."""
    calls = []
    call_lock = threading.Lock()

    class _Workbook:
        def __init__(self, path):
            self.path = path
            self.marker = "pristine"
            self.closed = False

        def close(self):
            self.closed = True

    original_load = mod.load_workbook

    def _fake_load(path, *, data_only=False):
        assert data_only is False
        with call_lock:
            calls.append(path)
        time.sleep(0.03)
        return _Workbook(path)

    app = object.__new__(mod.SowMergeApp)
    app.file_a = "mine.xlsx"
    app.file_b = "theirs.xlsx"
    app.base_path = None
    app.has_base = False
    app._is_closing = False
    app._wb_a_edit = None
    app._wb_b_edit = None
    app._wb_base_edit = None
    app._edit_fallback_lock = threading.Lock()
    app._edit_preload_active_event = threading.Event()
    results = []
    errors = []
    try:
        mod.load_workbook = _fake_load

        def _load():
            try:
                results.append(app._load_edit_workbooks_owned())
            except Exception as exc:
                errors.append(exc)

        workers = [threading.Thread(target=_load) for _index in range(2)]
        for worker in workers:
            worker.start()
        for worker in workers:
            worker.join(timeout=5)
        assert not any(worker.is_alive() for worker in workers)
        assert not errors, errors
        assert results == [True, True], results
        assert calls == ["mine.xlsx", "theirs.xlsx"], calls

        mine = app._wb_a_edit
        theirs = app._wb_b_edit
        mine.marker = "user-edited"
        assert app._load_edit_workbooks_owned() is True
        assert app._wb_a_edit is mine and app._wb_a_edit.marker == "user-edited"
        assert app._wb_b_edit is theirs
        assert calls == ["mine.xlsx", "theirs.xlsx"], calls
        assert not app._edit_preload_active_event.is_set()
    finally:
        mod.load_workbook = original_load
    metrics["single_owner_edit_preload"] = {
        "parse_calls": len(calls),
        "call_order": calls,
        "late_overwrite_blocked": True,
    }


def _test_compact_detached_column_diff_seed(metrics: dict):
    """Dense undo layers stay compact and cannot observe later dict mutation."""
    pair_count = 5000
    slot_count = 257
    dense = {
        pair_idx: ({-1} if pair_idx >= 4500 else {2, 64, 129, 257})
        for pair_idx in range(pair_count)
    }
    expected = {pair_idx: set(cols) for pair_idx, cols in dense.items()}
    packed = mod.SheetView._pack_exact_pair_diff_map(
        dense, pair_count, slot_count
    )
    # Simulate a late background apply / only-diff presentation rebuild.
    dense.clear()
    dense[0] = {1}
    restored = mod.SheetView._unpack_exact_pair_diff_map(packed)
    assert restored == expected
    stride = int(packed["stride"])
    assert len(packed["bits"]) == pair_count * stride
    five_layer_bytes = len(packed["bits"]) * 5
    # Even a 257-slot, 5k-row dense map remains below 1 MiB for five undo layers.
    assert five_layer_bytes < 1024 * 1024, five_layer_bytes
    metrics["compact_detached_diff_seed"] = {
        "pair_count": pair_count,
        "slot_count": slot_count,
        "bytes_per_layer": len(packed["bits"]),
        "five_layer_bytes": five_layer_bytes,
        "late_source_mutation_isolated": True,
    }


def main():
    metrics = {}
    _test_signature_uses_one_pass_sequential_cache(metrics)
    _test_mapping_width_guards(metrics)
    _test_cached_replay_and_interaction_paths(metrics)
    _test_single_owner_edit_preload(metrics)
    _test_compact_detached_column_diff_seed(metrics)
    print(json.dumps(metrics, ensure_ascii=False, indent=2, sort_keys=True))
    print("PERFORMANCE_TEST_COLUMN_STRUCTURE_GUARDS_OK")


if __name__ == "__main__":
    main()

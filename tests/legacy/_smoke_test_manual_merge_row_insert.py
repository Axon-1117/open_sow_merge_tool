import os
import time
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

import sow_merge_tool as mod
from _test_temp_utils import make_temp_dir


def _make_book(path: str, rows: list[list[object]]):
    wb = Workbook()
    ws = wb.active
    ws.title = "S1"
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx).value = value
    wb.save(path)
    wb.close()


def main():
    root = make_temp_dir("sow_manual_row_insert_")
    base = os.path.join(root, "base.xlsx")
    mine = os.path.join(root, "mine.xlsx")
    theirs = os.path.join(root, "theirs.xlsx")
    merged = os.path.join(root, "merged.xlsx")

    _make_book(base, [["id"], ["A"], ["C"]])
    _make_book(mine, [["id"], ["A"], ["C"]])
    _make_book(theirs, [["id"], ["A"], ["B"], ["C"]])

    for path in (base, mine, theirs):
        wb = load_workbook(path, data_only=False)
        ws = wb["S1"]
        for row in range(1, ws.max_row + 1):
            ws.cell(row=row, column=2).value = "=1"
        if path == theirs:
            ws.row_dimensions[3].height = 27
            ws.cell(row=3, column=1).number_format = "0000"
            ws.cell(row=3, column=1).fill = PatternFill(fill_type="solid", fgColor="00FF00")
            ws.cell(row=3, column=3).value = "=literal-text"
            ws.cell(row=3, column=3).data_type = "s"
            ws.cell(row=3, column=4).value = datetime(2026, 7, 21, 8, 30)
            ws.cell(row=3, column=4).number_format = "yyyy-mm-dd hh:mm"
        wb.save(path)
        wb.close()

    app = mod.SowMergeApp(mine, theirs, merge_mode=True, merged_path=merged, base_path=base)
    try:
        for _ in range(200):
            if app._edit_loaded_event.is_set():
                break
            app.root.update_idletasks()
            app.root.update()
            time.sleep(0.02)

        view = app.sheet_views.get("S1")
        if view is None:
            app.nb.select(app._sheet_containers["S1"])
            for _ in range(50):
                app.root.update_idletasks()
                app.root.update()
                time.sleep(0.02)
            view = app.sheet_views["S1"]

        view.force_align_var.set(1)
        view._toggle_force_align()
        view.refresh(row_only=None, rescan=True)

        insert_pair = None
        for pair_idx, (ra, rb) in enumerate(view.row_pairs):
            if ra is None and rb is not None and app.ws_b_val("S1").cell(rb, 1).value == "B":
                insert_pair = pair_idx
                break

        assert insert_pair is not None, f"did not find inserted row: {view.row_pairs}"

        # A Base-side failure must roll back both in-memory insertion and save ops.
        base_edit_ws = app.ws_base_edit("S1")
        original_base_insert = base_edit_ws.insert_rows
        original_showerror = mod.messagebox.showerror

        def _fail_base_insert(*_args, **_kwargs):
            raise RuntimeError("injected base insert failure")

        base_edit_ws.insert_rows = _fail_base_insert
        mod.messagebox.showerror = lambda *_args, **_kwargs: None
        try:
            assert not view._copy_selected_row("B2A", override_pair_idx=insert_pair)
        finally:
            base_edit_ws.insert_rows = original_base_insert
            mod.messagebox.showerror = original_showerror
        assert [app.ws_a_val("S1").cell(row=r, column=1).value for r in range(1, 4)] == ["id", "A", "C"]
        assert not app.manual_a_row_ops
        assert not app.manual_a_cell_ops
        assert not app.undo_stack

        assert view._copy_selected_row("B2A", override_pair_idx=insert_pair), "row insert copy failed"

        out = app.build_manual_merge_output_file()
        wb = load_workbook(out, data_only=False)
        try:
            ws = wb["S1"]
            values = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
            assert values == ["id", "A", "B", "C"], f"unexpected merged rows: {values}"
            assert ws.cell(row=3, column=2).value == "=1", "inserted identical formula was dropped"
            assert ws.row_dimensions[3].height == 27
            assert ws.cell(row=3, column=1).number_format == "0000"
            assert ws.cell(row=3, column=1).fill.fgColor.rgb in ("FF00FF00", "0000FF00", "00FF00")
            assert ws.cell(row=3, column=3).value == "=literal-text"
            assert ws.cell(row=3, column=3).data_type == "s"
            assert ws.cell(row=3, column=4).value == datetime(2026, 7, 21, 8, 30)
        finally:
            wb.close()
    finally:
        try:
            app._shutdown_root()
        except Exception:
            pass

    print("SMOKE_MANUAL_ROW_INSERT_OK")


if __name__ == "__main__":
    main()

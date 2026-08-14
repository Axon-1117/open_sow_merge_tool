"""GUI self-test: C-area hover tooltip shows full cell content.

Validates:
- 2-way mode: tooltip contains A/B full text lines.
- 3-way mode: tooltip contains BASE/A/B full text lines.
- real hover event path updates fixed hover-compare panel content.
- hover panel persists after leave and supports pin/clear + shift-wheel horizontal scroll.
"""

import os
from types import SimpleNamespace

from openpyxl import Workbook

import sow_merge_tool as mod
from _test_temp_utils import make_temp_dir


def _make_xlsx(path: str, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, v in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx).value = v
    wb.save(path)
    wb.close()


def _ensure_view(app: mod.SowMergeApp):
    sheet = app.common_sheets[0]
    view = app.sheet_views.get(sheet)
    if view is None:
        app.nb.select(app._sheet_containers[sheet])
        app.root.update_idletasks()
        app.root.update()
        view = app.sheet_views[sheet]
    return view


def _c_tooltip_text_by_col(view, c_line: int, col: int) -> str:
    spans = view._spans_for_line()
    assert col in spans, f"Column {col} not found in spans: {spans}"
    s, e = spans[col]
    char_pos = s + 1 if (e - s) > 1 else s
    payload = view._cursor_cmp_tooltip_payload(char_pos)
    assert payload is not None, f"Expected tooltip payload for c_line={c_line}, col={col}, span=({s},{e})"
    txt, _key = payload
    return str(txt)


def _main_tooltip_text_by_col(view, line_no: int, col: int) -> str:
    pair_idx = view._pair_idx_for_line(line_no)
    assert pair_idx is not None, f"Expected pair for line {line_no}"
    payload = view._cmp_tooltip_payload_by_pair_col(pair_idx, col)
    assert payload is not None, f"Expected main tooltip payload for line={line_no}, col={col}"
    txt, _key = payload
    return str(txt)


def _panel_text(view) -> str:
    try:
        return str(view.hover_cmp_text.get("1.0", "end-1c"))
    except Exception:
        return ""


def _panel_title(view) -> str:
    try:
        return str(view.hover_cmp_title_var.get())
    except Exception:
        return ""


def _has_tag(view, tag_name: str) -> bool:
    try:
        return len(view.hover_cmp_text.tag_ranges(tag_name)) >= 2
    except Exception:
        return False


def _motion_event_for_cell(text_widget, line_no: int, char_pos: int):
    index = f"{line_no}.{max(0, int(char_pos))}"
    # The Sheet-wide fixed-width model can place later logical columns outside
    # the initial viewport. Drive the same visibility step a user would get
    # from horizontal navigation before asking Tk for viewport coordinates.
    text_widget.see(index)
    text_widget.update_idletasks()
    text_widget.update()
    box = text_widget.bbox(index)
    assert box is not None, f"bbox is None for line={line_no}, char={char_pos}"
    x, y, w, h = box
    px = int(x + max(1, w // 2))
    py = int(y + max(1, h // 2))
    return SimpleNamespace(
        x=px,
        y=py,
        x_root=int(text_widget.winfo_rootx() + px),
        y_root=int(text_widget.winfo_rooty() + py),
    )


def _drive_main_hover(view, line_no: int, col: int, side: str = "A"):
    spans = view._spans_for_line()
    assert col in spans, f"Column {col} not found in spans: {spans}"
    s, e = spans[col]
    char_pos = s + 1 if (e - s) > 1 else s
    widget = view.left if side == "A" else (view.base if side == "BASE" else view.right)
    ev = _motion_event_for_cell(widget, line_no=line_no, char_pos=char_pos)
    view._on_cell_hover_tooltip(widget, ev, side)
    view.app.root.update_idletasks()
    view.app.root.update()


def _drive_c_hover(view, line_no: int, col: int):
    spans = view._spans_for_line()
    assert col in spans, f"Column {col} not found in spans: {spans}"
    s, e = spans[col]
    char_pos = s + 1 if (e - s) > 1 else s
    ev = _motion_event_for_cell(view.cursor_cmp, line_no=line_no, char_pos=char_pos)
    view._on_cursor_cmp_hover_tooltip(ev)
    view.app.root.update_idletasks()
    view.app.root.update()


def _run_2way():
    td_a = make_temp_dir(prefix="sow_c_tip_2a_")
    td_b = make_temp_dir(prefix="sow_c_tip_2b_")
    fa = os.path.join(td_a, "same.xlsx")
    fb = os.path.join(td_b, "same.xlsx")
    long_a = "A_FULL_" + ("x" * 80)
    long_b = "B_FULL_" + ("y" * 90)
    rows_a = [["h1", "h2", "h3", "h4", "h5"], [1, 2, 3, 4, long_a]]
    rows_b = [["h1", "h2", "h3", "h4", "h5"], [1, 2, 3, 4, long_b]]
    _make_xlsx(fa, rows_a)
    _make_xlsx(fb, rows_b)

    app = mod.SowMergeApp(fa, fb)
    view = _ensure_view(app)
    view.only_diff_var.set(0)
    view.refresh(row_only=None, rescan=True)
    view.left.mark_set("insert", "2.0")
    view.right.mark_set("insert", "2.0")
    view._update_cursor_lines()
    app.root.update_idletasks()
    app.root.update()

    txt = _c_tooltip_text_by_col(view, c_line=1, col=5)
    assert "base[" in txt and "mine[" in txt, txt
    assert "theirs[" not in txt, txt
    assert long_a in txt and long_b in txt, txt
    txt_main = _main_tooltip_text_by_col(view, line_no=2, col=5)
    assert "base[" in txt_main and "mine[" in txt_main, txt_main
    assert long_a in txt_main and long_b in txt_main, txt_main

    # Short, non-truncated cells should still update the fixed hover panel.
    _drive_main_hover(view, line_no=2, col=2, side="A")
    panel_short = _panel_text(view)
    assert "base[2]: 2" in panel_short and "mine[2]: 2" in panel_short, panel_short

    _drive_main_hover(view, line_no=2, col=5, side="A")
    panel = _panel_text(view)
    assert "base[" in panel and "mine[" in panel, panel
    assert long_a in panel and long_b in panel, panel
    assert (
        "Col: E(5)" in _panel_title(view)
        or "列 E(5)" in _panel_title(view)
    ), _panel_title(view)
    assert _has_tag(view, "hover_side_base"), "expected BASE-like row background in 2-way hover panel"
    assert _has_tag(view, "hover_side_mine"), "expected MINE-like row background in 2-way hover panel"
    assert _has_tag(view, "hover_diffchar"), "expected diff-char highlight in hover panel"

    _drive_c_hover(view, line_no=1, col=5)
    panel = _panel_text(view)
    assert "base[" in panel and "mine[" in panel, panel
    assert long_a in panel and long_b in panel, panel
    keep_panel = panel

    # Leave should not clear panel anymore (allow manual horizontal review).
    view._on_hover_compare_leave()
    view.app.root.update_idletasks()
    view.app.root.update()
    assert _panel_text(view) == keep_panel, _panel_text(view)

    # F4 hotkey toggles pin on/off (both direct handler and root key event).
    assert int(view.hover_cmp_pin_var.get()) == 0
    view._on_hover_compare_f4_toggle()
    assert int(view.hover_cmp_pin_var.get()) == 1
    view._on_hover_compare_f4_toggle()
    assert int(view.hover_cmp_pin_var.get()) == 0
    view.app.root.event_generate("<F4>")
    view.app.root.update_idletasks()
    view.app.root.update()
    assert int(view.hover_cmp_pin_var.get()) == 1
    view.app.root.event_generate("<F4>")
    view.app.root.update_idletasks()
    view.app.root.update()
    assert int(view.hover_cmp_pin_var.get()) == 0

    # Pin mode freezes auto updates.
    view.hover_cmp_pin_var.set(1)
    view._on_hover_compare_pin_toggle()
    view._set_hover_compare_panel(
        "base[9]: SHOULD_NOT_APPLY\nmine[9]: SHOULD_NOT_APPLY",
        ("S", "CMP", 999, 9, ("SHOULD_NOT_APPLY", "SHOULD_NOT_APPLY")),
    )
    assert _panel_text(view) == keep_panel, _panel_text(view)

    # Shift+wheel should move horizontal viewport.
    x0 = float((view.hover_cmp_text.xview() or (0.0, 1.0))[0])
    view._on_hover_cmp_shift_wheel(SimpleNamespace(delta=-120, state=0x1))
    view.app.root.update_idletasks()
    view.app.root.update()
    x1 = float((view.hover_cmp_text.xview() or (0.0, 1.0))[0])
    assert x1 >= x0, (x0, x1)

    # Clear button should clear even when pinned.
    view._on_hover_compare_clear_click()
    view.app.root.update_idletasks()
    view.app.root.update()
    assert _panel_text(view) == "", _panel_text(view)

    # Unpin then next update should apply.
    view.hover_cmp_pin_var.set(0)
    view._on_hover_compare_pin_toggle()
    _drive_main_hover(view, line_no=2, col=5, side="A")
    panel = _panel_text(view)
    assert "base[" in panel and "mine[" in panel, panel

    try:
        view._cancel_hover_compare_clear()
        app._shutdown_root()
    except Exception:
        pass


def _run_3way():
    td_base = make_temp_dir(prefix="sow_c_tip_3base_")
    td_mine = make_temp_dir(prefix="sow_c_tip_3mine_")
    td_theirs = make_temp_dir(prefix="sow_c_tip_3theirs_")
    fbase = os.path.join(td_base, "same.xlsx")
    fmine = os.path.join(td_mine, "same.xlsx")
    ftheirs = os.path.join(td_theirs, "same.xlsx")
    long_base = "BASE_FULL_" + ("b" * 70)
    long_mine = "MINE_FULL_" + ("m" * 75)
    long_theirs = "THEIRS_FULL_" + ("t" * 85)
    rows_base = [["h1", "h2", "h3", "h4", "h5"], [1, 2, 3, 4, long_base]]
    rows_mine = [["h1", "h2", "h3", "h4", "h5"], [1, 2, 3, 4, long_mine]]
    rows_theirs = [["h1", "h2", "h3", "h4", "h5"], [1, 2, 3, 4, long_theirs]]
    _make_xlsx(fbase, rows_base)
    _make_xlsx(fmine, rows_mine)
    _make_xlsx(ftheirs, rows_theirs)

    app = mod.SowMergeApp(fmine, ftheirs, merge_mode=True, base_path=fbase)
    view = _ensure_view(app)
    view.only_diff_var.set(0)
    view.refresh(row_only=None, rescan=True)
    view.left.mark_set("insert", "2.0")
    view.base.mark_set("insert", "2.0")
    view.right.mark_set("insert", "2.0")
    view._update_cursor_lines()
    app.root.update_idletasks()
    app.root.update()

    txt = _c_tooltip_text_by_col(view, c_line=2, col=5)
    assert "base[" in txt and "mine[" in txt and "theirs[" in txt, txt
    assert long_base in txt and long_mine in txt and long_theirs in txt, txt
    assert int(view.hover_cmp_text.cget("height")) >= 4, view.hover_cmp_text.cget("height")
    actual_h = int(view.hover_cmp_text.winfo_height())
    req_h = int(view.hover_cmp_text.winfo_reqheight())
    assert actual_h >= max(40, req_h - 20), (actual_h, req_h)
    txt_main = _main_tooltip_text_by_col(view, line_no=2, col=5)
    assert "base[" in txt_main and "mine[" in txt_main and "theirs[" in txt_main, txt_main
    assert long_base in txt_main and long_mine in txt_main and long_theirs in txt_main, txt_main

    _drive_main_hover(view, line_no=2, col=2, side="BASE")
    panel_short = _panel_text(view)
    assert "base[2]: 2" in panel_short and "mine[2]: 2" in panel_short and "theirs[2]: 2" in panel_short, panel_short

    _drive_main_hover(view, line_no=2, col=5, side="BASE")
    panel = _panel_text(view)
    assert "base[" in panel and "mine[" in panel and "theirs[" in panel, panel
    assert long_base in panel and long_mine in panel and long_theirs in panel, panel
    assert (
        "Col: E(5)" in _panel_title(view)
        or "列 E(5)" in _panel_title(view)
    ), _panel_title(view)
    assert _has_tag(view, "hover_side_base"), "expected BASE row background in 3-way hover panel"
    assert _has_tag(view, "hover_side_mine"), "expected MINE row background in 3-way hover panel"
    assert _has_tag(view, "hover_side_theirs"), "expected THEIRS row background in 3-way hover panel"
    assert _has_tag(view, "hover_diffchar"), "expected diff-char highlight in hover panel"

    _drive_c_hover(view, line_no=2, col=5)
    panel = _panel_text(view)
    assert "base[" in panel and "mine[" in panel and "theirs[" in panel, panel
    assert long_base in panel and long_mine in panel and long_theirs in panel, panel

    try:
        view._cancel_hover_compare_clear()
        app._shutdown_root()
    except Exception:
        pass


def _run_f4_routes_to_active_sheet():
    td_a = make_temp_dir(prefix="sow_c_tip_f4_a_")
    td_b = make_temp_dir(prefix="sow_c_tip_f4_b_")
    fa = os.path.join(td_a, "same.xlsx")
    fb = os.path.join(td_b, "same.xlsx")

    wb_a = Workbook()
    ws = wb_a.active
    ws.title = "S1"
    ws["A1"] = "A1"
    ws2 = wb_a.create_sheet("S2")
    ws2["A1"] = "A2"
    wb_a.save(fa)
    wb_a.close()

    wb_b = Workbook()
    ws = wb_b.active
    ws.title = "S1"
    ws["A1"] = "B1"
    ws2 = wb_b.create_sheet("S2")
    ws2["A1"] = "B2"
    wb_b.save(fb)
    wb_b.close()

    app = mod.SowMergeApp(fa, fb)
    assert len(app.common_sheets) >= 2, app.common_sheets
    s1, s2 = app.common_sheets[0], app.common_sheets[1]

    app.nb.select(app._sheet_containers[s1])
    app.root.update_idletasks()
    app.root.update()
    v1 = app.sheet_views[s1]

    app.nb.select(app._sheet_containers[s2])
    app.root.update_idletasks()
    app.root.update()
    v2 = app.sheet_views[s2]

    assert int(v1.hover_cmp_pin_var.get()) == 0
    assert int(v2.hover_cmp_pin_var.get()) == 0

    # Deterministic route check via app-level F4 handler.
    app._on_global_f4(None)
    app.root.update_idletasks()
    app.root.update()
    assert int(v1.hover_cmp_pin_var.get()) == 0, "inactive sheet should not be toggled by F4"
    assert int(v2.hover_cmp_pin_var.get()) == 1, "active sheet should be toggled by F4"

    # Also verify real key-event path.
    try:
        app.root.focus_force()
    except Exception:
        pass
    app.root.focus_set()
    app.root.update_idletasks()
    app.root.update()
    app.root.event_generate("<F4>")
    app.root.update_idletasks()
    app.root.update()
    assert int(v1.hover_cmp_pin_var.get()) == 0, "inactive sheet should remain unchanged"
    assert int(v2.hover_cmp_pin_var.get()) == 0, "active sheet should toggle back on second F4"

    try:
        v1._cancel_hover_compare_clear()
        v2._cancel_hover_compare_clear()
        app._shutdown_root()
    except Exception:
        pass


def main():
    _run_2way()
    _run_3way()
    _run_f4_routes_to_active_sheet()
    print("GUI_SELF_TEST_C_HOVER_TOOLTIP_OK")


if __name__ == "__main__":
    main()

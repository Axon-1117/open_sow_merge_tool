"""Regression coverage for safe global ambiguity handling and Sheet overwrite."""

from __future__ import annotations

import os

from openpyxl import load_workbook

import sow_merge_tool as smt
from _gui_self_test_openspec_section10 import _synthetic_view, _wait_until


def _equal_blank_rows():
    mine = [
        ("ID", "Value", None, None, "Tail", "Tail2"),
        (1, "mine", None, None, "same", "same"),
    ]
    theirs = [list(row) for row in mine]
    theirs[1][1] = "theirs"
    return mine, theirs


def _duplicate_column_rows():
    mine = [
        ("ID", "Value", "Spacer", "Value", "Tail"),
        (1, "same", None, "same", "t"),
        (2, "same", None, "same", "t"),
    ]
    theirs = [list(row) for row in mine]
    theirs[2][3] = "右侧中文"
    return mine, theirs


def _sheet_values(workbook, sheet="Data"):
    worksheet = workbook[sheet]
    return tuple(
        tuple(cell.value for cell in row)
        for row in worksheet.iter_rows(
            min_row=1,
            max_row=worksheet.max_row,
            min_col=1,
            max_col=worksheet.max_column,
        )
    )


def test_equal_unresolved_columns_do_not_block_global_apply():
    mine, theirs = _equal_blank_rows()
    with _synthetic_view(mine, theirs) as (app, view, _dialogs):
        cache = view._active_column_comparison_cache()
        assert cache.unresolved_cols, "fixture must contain unresolved blank columns"
        before = _sheet_values(app._wb_a_edit, "Data")
        view._set_copy_scope_mode("global")
        assert view._copy_all_safe_sheet_differences("B2A") is True
        after = _sheet_values(app._wb_a_edit, "Data")
        assert after != before
        assert app.ws_a_edit("Data")["B2"].value == "theirs"
        assert app.ws_a_edit("Data")["C2"].value is None
        assert app.ws_a_edit("Data")["D2"].value is None


def test_real_ambiguous_diff_blocks_and_reports_details():
    mine, theirs = _duplicate_column_rows()
    with _synthetic_view(mine, theirs) as (app, view, _dialogs):
        cache = view._active_column_comparison_cache()
        assert 4 in cache.unresolved_cols
        captured = {}

        def _cancel(direction, details):
            captured["direction"] = direction
            captured["details"] = details
            return "cancel"

        view._ask_global_ambiguous_action = _cancel
        before = _sheet_values(app._wb_a_edit, "Data")
        view._set_copy_scope_mode("global")
        assert view._copy_all_safe_sheet_differences("B2A") is False
        assert _sheet_values(app._wb_a_edit, "Data") == before
        assert captured["direction"] == "B2A"
        assert [item["logical_col"] for item in captured["details"]] == [4]
        assert "重复列特征" in str(captured["details"][0]["cause"])
        assert captured["details"][0]["count"] == 1
        assert captured["details"][0]["samples"]


def test_confirmed_whole_sheet_copy_is_undoable_and_preserves_unicode():
    mine, theirs = _duplicate_column_rows()
    with _synthetic_view(mine, theirs) as (app, view, _dialogs):
        before = _sheet_values(app._wb_a_edit, "Data")
        view._ask_global_ambiguous_action = lambda _direction, _details: "B2A_SHEET"
        view._set_copy_scope_mode("global")
        assert view._copy_all_safe_sheet_differences("B2A") is True
        assert app.ws_a_edit("Data")["D3"].value == "右侧中文"
        assert app.manual_sheet_ops[-1]["source_side"] == "B"
        assert len(app.undo_stack) == 1

        view._undo_last_action()
        assert _sheet_values(app._wb_a_edit, "Data") == before
        assert not app.manual_sheet_ops
        assert not app.undo_stack

        view._ask_global_ambiguous_action = lambda _direction, _details: "B2A_SHEET"
        view._set_copy_scope_mode("global")
        assert view._copy_all_safe_sheet_differences("B2A") is True
        output = app.build_manual_merge_output_file()
        try:
            valid, reason = smt._validate_xlsx_package(output)
            assert valid, reason
            workbook = load_workbook(output, data_only=False, read_only=True)
            try:
                assert workbook["Data"]["D3"].value == "右侧中文"
            finally:
                workbook.close()
        finally:
            if os.path.exists(output):
                os.remove(output)

    # Exercise the opposite explicit overwrite direction as well.
    mine, theirs = _duplicate_column_rows()
    with _synthetic_view(mine, theirs) as (app, view, _dialogs):
        before = _sheet_values(app._wb_b_edit, "Data")
        view._ask_global_ambiguous_action = lambda _direction, _details: "A2B_SHEET"
        view._set_copy_scope_mode("global")
        assert view._copy_all_safe_sheet_differences("A2B") is True
        assert app.ws_b_edit("Data")["D3"].value == "same"
        assert app.manual_sheet_ops[-1]["source_side"] == "A"
        assert len(app.undo_stack) == 1
        view._undo_last_action()
        assert _sheet_values(app._wb_b_edit, "Data") == before
        assert not app.manual_sheet_ops
        assert not app.undo_stack


def main():
    tests = (
        test_equal_unresolved_columns_do_not_block_global_apply,
        test_real_ambiguous_diff_blocks_and_reports_details,
        test_confirmed_whole_sheet_copy_is_undoable_and_preserves_unicode,
    )
    for test in tests:
        test()
        print(f"PASS {test.__name__}")


if __name__ == "__main__":
    main()

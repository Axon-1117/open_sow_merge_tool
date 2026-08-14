"""Real Item.xlsx acceptance for an explicitly rendered blank-row deletion."""

from __future__ import annotations

import glob
import hashlib
import json
import os
import shutil
import tempfile

from openpyxl import load_workbook

import sow_merge_tool as smt
from _test_temp_utils import visible_render_text
from _ux_5_3_final_acceptance import (
    _pump,
    _wait_for_stable_projection,
    _wait_for_view,
    wait_edit_ready,
    wait_view_ready,
)


SHEET = "Item@design"
DELETED_MINE_ROW = 812


def _sha256(path: str) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest().upper()


def _latest(pattern: str) -> str:
    candidates = [path for path in glob.glob(pattern) if os.path.isfile(path)]
    if not candidates:
        raise FileNotFoundError(pattern)
    return max(candidates, key=os.path.getmtime)


def _row_values(ws, row_idx: int, max_col: int):
    return tuple(ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, max_col + 1))


def _find_deleted_pair(view) -> int:
    candidates = [
        (
            pair_idx,
            row_a,
            row_b,
            view._base_row_for_pair(pair_idx),
            sorted(view.pair_diff_cols.get(pair_idx, set())),
            sorted(view.pair_base_diff_cols.get(pair_idx, set())),
        )
        for pair_idx, (row_a, row_b) in enumerate(view.row_pairs)
        if (
            row_a is not None and 805 <= int(row_a) <= 818
        ) or (
            row_b is not None and 805 <= int(row_b) <= 818
        ) or view._base_row_for_pair(pair_idx) == DELETED_MINE_ROW
    ]
    for pair_idx, row_a, row_b, _base_row, _lr, _base in candidates:
        if row_a == DELETED_MINE_ROW and row_b is None:
            return pair_idx
    raise AssertionError(f"deleted pair not found; candidates={candidates}")


def main():
    temp_root = tempfile.gettempdir()
    source_base = _latest(os.path.join(temp_root, "sow_merge_tool_svncat_BASE_*_Item.xlsx"))
    source_theirs = _latest(
        os.path.join(
            temp_root,
            "sow_merge_tool_svn_Item.xlsx.merge-right.r36474_*.xlsx",
        )
    )
    source_hashes = {
        source_base: _sha256(source_base),
        source_theirs: _sha256(source_theirs),
    }

    root_dir = tempfile.mkdtemp(prefix="sow-item-blank-row-delete-")
    paths = {
        "mine": os.path.join(root_dir, "mine", "Item.xlsx"),
        "base": os.path.join(root_dir, "base", "Item.xlsx"),
        "theirs": os.path.join(root_dir, "theirs", "Item.xlsx"),
    }
    for role, destination in paths.items():
        os.makedirs(os.path.dirname(destination), exist_ok=True)
        source = source_theirs if role == "theirs" else source_base
        shutil.copy2(source, destination)
    settings_path = os.path.join(root_dir, "settings.json")
    with open(settings_path, "w", encoding="utf-8") as stream:
        json.dump({"only_diff": 0}, stream)

    original_settings_path = smt._SETTINGS_PATH
    smt._SETTINGS_PATH = settings_path
    app = None
    output = None
    try:
        app = smt.SowMergeApp(
            paths["mine"],
            paths["theirs"],
            merge_mode=True,
            merged_path=os.path.join(root_dir, "merged", "Item.xlsx"),
            base_path=paths["base"],
        )
        app._intended_window_state = "normal"
        app.root.state("normal")
        app.root.geometry("1600x900")
        app.nb.select(app._sheet_containers[SHEET])
        view = _wait_for_view(app, SHEET, timeout=180.0)
        wait_edit_ready(app, timeout=180.0)
        wait_view_ready(view, timeout=180.0)
        _wait_for_stable_projection(view, timeout=180.0, stable_for=0.3)

        view.only_diff_var.set(0)
        view._last_only_diff_value = 0
        view.refresh(row_only=None, rescan=True)
        _pump(app.root, 0.3)
        view._set_copy_scope_mode("region")

        pair_idx = _find_deleted_pair(view)
        assert view._base_row_for_pair(pair_idx) == DELETED_MINE_ROW
        assert view._row_label_for_pair_idx(pair_idx, "B") == "缺行"
        assert "此侧缺行" in visible_render_text(
            view.pair_text_b[pair_idx],
            placeholder=smt._TK_INDEX_PLACEHOLDER,
        )
        assert view.pair_diff_cols[pair_idx] == {-1}

        mine_ws = app.ws_a_edit(SHEET)
        max_col = int(mine_ws.max_column or 1)
        blank_before = _row_values(mine_ws, DELETED_MINE_ROW, max_col)
        next_before = _row_values(mine_ws, DELETED_MINE_ROW + 1, max_col)
        assert not any(value is not None for value in blank_before)
        assert next_before[0] == "skin_avatar_privilege"
        undo_before = len(app.undo_stack)

        assert view._materialize_pair_for_navigation(pair_idx)
        view._select_line(view.row_to_line[pair_idx])
        view._run_copy_action_by_mode("B2A")
        _pump(app.root, 0.6)
        assert _row_values(app.ws_a_edit(SHEET), DELETED_MINE_ROW, max_col)[0] == "skin_avatar_privilege"
        assert len(app.undo_stack) == undo_before + 1
        delete_op = app.manual_a_row_ops[-1]
        assert (
            delete_op.get("kind"),
            int(delete_op.get("row", 0)),
            int(delete_op.get("count", 0)),
        ) == ("delete_rows", DELETED_MINE_ROW, 1)

        view._undo_last_action()
        _pump(app.root, 0.6)
        assert _row_values(app.ws_a_edit(SHEET), DELETED_MINE_ROW, max_col) == blank_before
        assert _row_values(app.ws_a_edit(SHEET), DELETED_MINE_ROW + 1, max_col) == next_before
        assert len(app.undo_stack) == undo_before

        pair_idx = _find_deleted_pair(view)
        assert view._materialize_pair_for_navigation(pair_idx)
        view._select_line(view.row_to_line[pair_idx])
        view._run_copy_action_by_mode("B2A")
        _pump(app.root, 0.6)
        output = app.build_manual_merge_output_file()

        saved = load_workbook(output, data_only=False, read_only=True, keep_links=False)
        theirs = load_workbook(paths["theirs"], data_only=False, read_only=True, keep_links=False)
        try:
            saved_ws = saved[SHEET]
            theirs_ws = theirs[SHEET]
            for row_idx in range(807, 816):
                assert _row_values(saved_ws, row_idx, 3) == _row_values(theirs_ws, row_idx, 3), (
                    row_idx,
                    _row_values(saved_ws, row_idx, 3),
                    _row_values(theirs_ws, row_idx, 3),
                )
        finally:
            saved.close()
            theirs.close()

        assert {
            path: _sha256(path) for path in source_hashes
        } == source_hashes
        print(
            "PASS: real Item blank-row deletion "
            f"pair={pair_idx} row={DELETED_MINE_ROW} "
            "render=explicit undo=ok native-save=ok"
        )
    finally:
        smt._SETTINGS_PATH = original_settings_path
        if app is not None:
            app._shutdown_root()
        if output and os.path.exists(output):
            os.remove(output)
        shutil.rmtree(root_dir, ignore_errors=True)


if __name__ == "__main__":
    main()

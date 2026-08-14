"""Real Guide 3-way regression for common side-insertion pairing.

Run from a desktop logon session because the save/reopen phase requires Excel
native replay::

    python _gui_self_test_real_guide_common_insert.py
"""

from __future__ import annotations

import hashlib
import json
import os
import shutil
import time

from openpyxl import load_workbook

import sow_merge_tool as smt
from _gui_self_test_logical_column_actions import (
    _force_full_view,
    _model_snapshot,
    _pump,
    _selection_snapshot,
    _wait_for_stable_projection,
    _wait_for_view,
    _worksheet_snapshot,
)
from _test_temp_utils import make_temp_dir


_FIXTURE_ROOT = r"C:\Users\dd\AppData\Local\Temp\sow_ux_5_3_20260723_001\cases\Guide"
_REPORT = r"C:\Users\dd\AppData\Local\Temp\sow_ux_5_3_20260723_001\acceptance_report.json"
_SHEET = "TGuideStep@design"
_MARKERS = ("__UX_INS1_R1", "__UX_INS2_R1")
def _sha256(path: str) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest().upper()


def _copy_sources():
    with open(_REPORT, "r", encoding="utf-8") as stream:
        report_hashes = json.load(stream)["prepare"]["cases"]["Guide"]["hashes"]
    expected_hashes = {
        side: str(report_hashes[side]).upper() for side in ("base", "mine", "theirs")
    }
    root = make_temp_dir("sow_real_guide_common_insert_")
    paths = {}
    originals = {}
    for side in ("base", "mine", "theirs"):
        source = os.path.join(_FIXTURE_ROOT, side, "Guide.xlsx")
        assert os.path.isfile(source), source
        assert _sha256(source) == expected_hashes[side], (side, _sha256(source))
        side_dir = os.path.join(root, side)
        os.makedirs(side_dir, exist_ok=True)
        target = os.path.join(side_dir, "Guide.xlsx")
        shutil.copy2(source, target)
        assert _sha256(target) == expected_hashes[side]
        paths[side] = target
        originals[side] = source
    paths["merged"] = os.path.join(root, "output", "Guide.xlsx")
    os.makedirs(os.path.dirname(paths["merged"]), exist_ok=True)
    return paths, originals, expected_hashes


def _wait_for_edit_books(app, timeout=180.0):
    deadline = time.time() + timeout
    while time.time() < deadline:
        _pump(app.root, 0.05)
        if app._edit_loaded_event.is_set() and app._edit_workbooks_ready():
            return
    raise AssertionError("Guide editable workbooks did not become ready")


def _force_mapping_rebuild(view, reason: str):
    view._mark_column_mapping_stale(reason, edited_sides=("A",))
    view.refresh(row_only=None, rescan=True)
    _wait_for_stable_projection(view, timeout=180.0, stable_for=0.5)


def _marker_slot_evidence(view):
    projection = view._ensure_column_projection_current("Guide共同新增列验收")
    ws_a = view.app.ws_a_edit(_SHEET)
    ws_b = view.app.ws_b_edit(_SHEET)
    evidence = []
    for slot in projection.model.slots:
        mine_header = (
            ws_a.cell(1, slot.mine_col).value if slot.mine_col is not None else None
        )
        theirs_header = (
            ws_b.cell(1, slot.theirs_col).value if slot.theirs_col is not None else None
        )
        if mine_header in _MARKERS or theirs_header in _MARKERS:
            evidence.append(
                {
                    "logical": slot.logical_idx + 1,
                    "mine_col": slot.mine_col,
                    "base_col": slot.base_col,
                    "theirs_col": slot.theirs_col,
                    "state": slot.state,
                    "ambiguous": slot.confidence.ambiguous,
                    "reason": slot.confidence.reason,
                    "score": slot.confidence.score,
                    "base_boundary": slot.base_boundary,
                    "origin_side": slot.origin_side,
                    "mine_header": mine_header,
                    "theirs_header": theirs_header,
                }
            )
    return evidence


def _assert_two_paired_common_insertions(view):
    evidence = _marker_slot_evidence(view)
    assert len(evidence) == 2, evidence
    assert [item["mine_header"] for item in evidence] == list(_MARKERS), evidence
    assert [item["theirs_header"] for item in evidence] == list(_MARKERS), evidence
    assert all(item["mine_col"] is not None for item in evidence), evidence
    assert all(item["theirs_col"] is not None for item in evidence), evidence
    assert all(item["base_col"] is None for item in evidence), evidence
    assert all(item["state"] == "inserted" for item in evidence), evidence
    assert all(item["origin_side"] == "both" for item in evidence), evidence
    assert all(item["reason"] == "common-side-insertion" for item in evidence), evidence
    assert all(item["score"] == 1.0 and not item["ambiguous"] for item in evidence), evidence
    assert len({item["base_boundary"] for item in evidence}) == 1, evidence
    assert not view.column_comparison_cache.structural_diff_cols, (
        view.column_comparison_cache.structural_diff_cols,
        evidence,
    )
    return evidence


def _worksheet_snapshot_diff(before, after):
    before_cells = {(row, col): payload for row, col, payload in before[2]}
    after_cells = {(row, col): payload for row, col, payload in after[2]}
    changed = []
    for coordinate in sorted(set(before_cells) | set(after_cells)):
        if before_cells.get(coordinate) != after_cells.get(coordinate):
            changed.append(
                (coordinate, before_cells.get(coordinate), after_cells.get(coordinate))
            )
    return {
        "before_shape": before[:2],
        "after_shape": after[:2],
        "before_cell_count": len(before_cells),
        "after_cell_count": len(after_cells),
        "changed_cell_count": len(changed),
        "changed_cell_samples": changed[:8],
        "column_dimensions_equal": before[3] == after[3],
        "validations_equal": before[4] == after[4],
        "conditional_formatting_equal": before[5] == after[5],
        "merged_cells_equal": before[6] == after[6],
    }


def main():
    paths, originals, expected_hashes = _copy_sources()
    identical_conflicts, identical_conflict_map = smt._scan_three_way_conflicts(
        paths["base"], paths["base"], paths["base"]
    )
    assert identical_conflicts == [] and identical_conflict_map == {}, (
        identical_conflicts[:12], identical_conflict_map
    )
    conflicts, conflict_map = smt._scan_three_way_conflicts(
        paths["base"], paths["mine"], paths["theirs"]
    )
    assert conflicts and conflict_map.get(_SHEET), "expected controlled Guide conflicts"
    app = smt.SowMergeApp(
        paths["mine"],
        paths["theirs"],
        merge_mode=True,
        merged_path=paths["merged"],
        base_path=paths["base"],
        raw_mine=paths["mine"],
        raw_base=paths["base"],
        raw_theirs=paths["theirs"],
        merge_conflict_cells_by_sheet=conflict_map,
        merge_conflict_mode=True,
    )
    output = None
    try:
        app.root.withdraw()
        view = _force_full_view(_wait_for_view(app, _SHEET, timeout=180.0))
        _wait_for_edit_books(app)
        _wait_for_stable_projection(view, timeout=180.0, stable_for=0.5)

        before_model = _model_snapshot(view)
        before_row_pairs = tuple(view.row_pairs)
        before_selection = _selection_snapshot(view)
        before_sheet = _worksheet_snapshot(app.ws_a_edit(_SHEET))
        before_headers = [app.ws_a_edit(_SHEET).cell(1, col).value for col in range(1, 20)]
        initial = _marker_slot_evidence(view)
        assert len(initial) == 2, initial
        assert all(item["mine_col"] is None for item in initial), initial
        assert [item["theirs_col"] for item in initial] == [12, 13], initial

        block = view._select_column_block_by_logical_col(12, "B")
        assert block is not None and tuple(block.slot_indices) == (11, 12), block
        selected_before_apply = _selection_snapshot(view)
        plan = view._apply_selected_column_block("B", "A", confirm_unresolved=True)
        assert plan.action_kind == "insert_copy" and plan.count == 2, plan
        assert tuple(plan.source_physical_cols) == (12, 13), plan
        cells_after_action = len(app.ws_a_edit(_SHEET)._cells)
        _force_mapping_rebuild(view, "real-guide-common-insert-after-action")
        assert tuple(view.row_pairs) == before_row_pairs, (
            len(before_row_pairs), len(view.row_pairs), tuple(view.row_pairs[:8])
        )
        cells_after_action_rebuild = len(app.ws_a_edit(_SHEET)._cells)
        after_apply = _assert_two_paired_common_insertions(view)
        marker_logical_cols = {item["logical"] for item in after_apply}
        base_marker_evidence = {
            pair_idx: sorted(marker_logical_cols & set(cols or ()))
            for pair_idx, cols in view.pair_base_diff_cols.items()
            if marker_logical_cols & set(cols or ())
        }
        assert base_marker_evidence, view.pair_base_diff_cols
        marker_visual_flood = {
            pair_idx: sorted(
                marker_logical_cols & view._visual_diff_cols_for_pair(pair_idx)
            )
            for pair_idx in range(len(view.row_pairs))
            if marker_logical_cols & view._visual_diff_cols_for_pair(pair_idx)
        }
        assert not marker_visual_flood, marker_visual_flood

        view._undo_last_action()
        cells_after_undo = len(app.ws_a_edit(_SHEET)._cells)
        _force_mapping_rebuild(view, "real-guide-common-insert-after-undo")
        assert tuple(view.row_pairs) == before_row_pairs, (
            len(before_row_pairs), len(view.row_pairs), tuple(view.row_pairs[:8])
        )
        cells_after_undo_rebuild = len(app.ws_a_edit(_SHEET)._cells)
        after_undo_sheet = _worksheet_snapshot(app.ws_a_edit(_SHEET))
        undo_diff = _worksheet_snapshot_diff(before_sheet, after_undo_sheet)
        undo_diff["cell_counts_by_stage"] = {
            "before": len(before_sheet[2]),
            "after_action": cells_after_action,
            "after_action_rebuild": cells_after_action_rebuild,
            "after_undo": cells_after_undo,
            "after_undo_rebuild": cells_after_undo_rebuild,
        }
        assert after_undo_sheet == before_sheet, undo_diff
        assert [app.ws_a_edit(_SHEET).cell(1, col).value for col in range(1, 20)] == before_headers
        after_undo_model = _model_snapshot(view)
        assert after_undo_model == before_model, [
            (index, before_model[index], after_undo_model[index])
            for index in range(len(before_model))
            if before_model[index] != after_undo_model[index]
        ]
        assert _selection_snapshot(view) in (before_selection, selected_before_apply)

        view._select_column_block_by_logical_col(12, "B")
        reapplied = view._apply_selected_column_block("B", "A", confirm_unresolved=True)
        assert reapplied.action_kind == "insert_copy" and reapplied.count == 2
        _force_mapping_rebuild(view, "real-guide-common-insert-before-save")
        assert tuple(view.row_pairs) == before_row_pairs, (
            len(before_row_pairs), len(view.row_pairs), tuple(view.row_pairs[:8])
        )
        before_save = _assert_two_paired_common_insertions(view)

        output = app.build_manual_merge_output_file()
        package_ok, package_error = smt._validate_xlsx_package(output)
        assert package_ok, package_error
    finally:
        app._shutdown_root()

    wb = load_workbook(output, data_only=False, read_only=True)
    try:
        ws = wb[_SHEET]
        assert [ws.cell(1, 12).value, ws.cell(1, 13).value] == list(_MARKERS)
    finally:
        wb.close()
    assert all(_sha256(path) == expected_hashes[side] for side, path in originals.items())
    print("GUI_SELF_TEST_REAL_GUIDE_COMMON_INSERT_OK")
    print(
        {
            "initial": initial,
            "after_apply": after_apply,
            "before_save": before_save,
            "output": output,
        }
    )


if __name__ == "__main__":
    main()

"""Adversarial smoke tests for the recoverable multi-branch submit workflow."""

from __future__ import annotations

import os
import shutil
import sqlite3
import tempfile
from types import SimpleNamespace
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

import sow_merge_tool as smt
from sow_merge_tool import branch_submit as bs
from sow_merge_tool import svn_status_provider as sp


def _book(path: str, value: int) -> None:
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["id", "value"])
    ws.append(["a", value])
    wb.save(path)
    wb.close()


def _create_wc(root: str, branches=("develop", "release", "sandbox", "master")) -> None:
    svn_dir = os.path.join(root, ".svn")
    os.makedirs(svn_dir, exist_ok=True)
    with sqlite3.connect(os.path.join(svn_dir, "wc.db")) as conn:
        conn.executescript(
            """
            create table REPOSITORY (id integer primary key, root text, uuid text);
            create table NODES (
              wc_id integer, local_relpath text, op_depth integer, parent_relpath text,
              repos_id integer, repos_path text, revision integer, presence text,
              moved_here integer, moved_to text, kind text, properties blob, depth text,
              checksum text, symlink_target text, changed_revision integer,
              changed_date integer, changed_author text, translated_size integer,
              last_mod_time integer, dav_cache blob, file_external integer,
              inherited_props blob
            );
            create table ACTUAL_NODE (
              wc_id integer, local_relpath text, parent_relpath text, properties blob,
              conflict_old text, conflict_new text, conflict_working text,
              prop_reject text, changelist text, text_mod text,
              tree_conflict_data text, conflict_data blob, older_checksum text,
              left_checksum text, right_checksum text
            );
            insert into REPOSITORY values (1, 'file:///repo', 'fixture-uuid');
            """
        )
        for branch in branches:
            os.makedirs(os.path.join(root, branch, "config"), exist_ok=True)
            _insert_node(conn, branch, "dir")
            _insert_node(conn, f"{branch}/config", "dir")


def _insert_node(conn, relpath: str, kind: str = "file", revision: int = 100) -> None:
    parent = relpath.rsplit("/", 1)[0] if "/" in relpath else ""
    conn.execute(
        """insert into NODES
        (wc_id,local_relpath,op_depth,parent_relpath,repos_id,repos_path,revision,presence,
         moved_here,moved_to,kind,changed_revision,changed_author,file_external)
        values (1,?,?,?,1,?,?, 'normal',0,'',?,?, 'tester',0)""",
        (relpath, 0, parent, "sheets/" + relpath, revision, kind, revision),
    )


def _set_branch_changed(root: str, branch: str, unix_seconds: int) -> None:
    with sqlite3.connect(os.path.join(root, ".svn", "wc.db")) as conn:
        conn.execute(
            "update NODES set changed_date=? where local_relpath=? or local_relpath like ?",
            (unix_seconds * 1_000_000, branch, branch + "/%"),
        )


def _delete_node(root: str, path: str) -> None:
    rel = os.path.relpath(path, root).replace("\\", "/")
    with sqlite3.connect(os.path.join(root, ".svn", "wc.db")) as conn:
        conn.execute("delete from NODES where local_relpath=?", (rel,))


def _ensure_node(root: str, path: str, revision: int = 200) -> None:
    rel = os.path.relpath(path, root).replace("\\", "/")
    with sqlite3.connect(os.path.join(root, ".svn", "wc.db")) as conn:
        row = conn.execute("select 1 from NODES where local_relpath=?", (rel,)).fetchone()
        if row:
            conn.execute("update NODES set revision=?,changed_revision=? where local_relpath=?", (revision, revision, rel))
        else:
            _insert_node(conn, rel, "file", revision)


class FakeCore:
    APP_NAME = "sow_merge_tool"

    @staticmethod
    def _find_svn_wc_root_for_path(path):
        current = os.path.abspath(path if os.path.isdir(path) else os.path.dirname(path))
        while current != os.path.dirname(current):
            if os.path.isfile(os.path.join(current, ".svn", "wc.db")):
                return current
            current = os.path.dirname(current)
        return None

    @staticmethod
    def _try_export_svn_base_from_working_copy(path):
        candidate = path + ".pristine.xlsx"
        return candidate if os.path.isfile(candidate) else None

    @staticmethod
    def _wc_node_metadata(root, relative):
        node = bs._node_for_path(root, os.path.join(root, *relative.split("/")))
        return (node.changed_revision if node else None), "tester", "fixture"

    @staticmethod
    def _detect_svn_conflict_files(_path):
        return None

    @staticmethod
    def _has_svn_conflict_artifacts(_path):
        return False

    @staticmethod
    def _cross_branch_source_delta_premerge(before, target, after):
        candidate = target + ".candidate.xlsx"
        shutil.copy2(target, candidate)
        before_wb = load_workbook(before, data_only=False)
        after_wb = load_workbook(after, data_only=False)
        target_wb = load_workbook(candidate, data_only=False)
        applied = already = 0
        conflicts = []
        for sheet in before_wb.sheetnames:
            b, a, t = before_wb[sheet], after_wb[sheet], target_wb[sheet]
            for row in range(1, max(b.max_row, a.max_row) + 1):
                for col in range(1, max(b.max_column, a.max_column) + 1):
                    old, new, current = b.cell(row, col).value, a.cell(row, col).value, t.cell(row, col).value
                    if old == new:
                        continue
                    if current == new:
                        already += 1
                    elif current == old:
                        t.cell(row, col).value = new
                        applied += 1
                    else:
                        conflicts.append((sheet, row, col, current, new))
        target_wb.save(candidate)
        before_wb.close(); after_wb.close(); target_wb.close()
        summary = {
            "incoming_count": applied + already + len(conflicts), "applied_count": applied,
            "already_present_count": already, "target_retained_count": 0,
            "unresolved_count": len(conflicts), "merged_count": applied,
        }
        return conflicts, candidate, {}, summary, None

    @staticmethod
    def _find_tortoise_proc_exe():
        return "TortoiseProc.exe"


class FixtureScanner:
    def __init__(self, root: str):
        self.root = root
        self.overrides: dict[str, sp.SvnStatusRecord] = {}

    def __call__(self, scope: str):
        records = []
        for path, record in self.overrides.items():
            if bs._is_within(path, scope):
                records.append(record)
        with sqlite3.connect(f"file:{os.path.join(self.root, '.svn', 'wc.db')}?mode=ro", uri=True) as conn:
            rows = conn.execute("select local_relpath,kind,revision from NODES where kind='file' and presence='normal'").fetchall()
        versioned = set()
        for rel, kind, revision in rows:
            path = os.path.join(self.root, *str(rel).split("/")); versioned.add(os.path.normcase(path))
            if not bs._is_within(path, scope) or os.path.normcase(path) in {os.path.normcase(x) for x in self.overrides}:
                continue
            pristine = path + ".pristine.xlsx"
            if not os.path.exists(path):
                records.append(sp.SvnStatusRecord(path=path, node_kind=kind, node_status="missing", versioned=True, revision=revision))
            elif os.path.isfile(pristine) and bs._sha256(path) != bs._sha256(pristine):
                records.append(sp.SvnStatusRecord(path=path, node_kind=kind, node_status="modified", text_status="modified", prop_status="normal", versioned=True, revision=revision))
        for folder, _dirs, files in os.walk(scope):
            if ".svn" in folder.split(os.sep):
                continue
            for name in files:
                path = os.path.join(folder, name)
                if name.endswith(".pristine.xlsx") or ".candidate.xlsx" in name:
                    continue
                if os.path.normcase(path) not in versioned and os.path.normcase(path) not in {os.path.normcase(x) for x in self.overrides}:
                    records.append(sp.SvnStatusRecord(path=path, node_kind="file", node_status="unversioned", versioned=False))
        return records


def _commit_paths(args):
    for value in args:
        if value.startswith("/pathfile:"):
            return open(value.split(":", 1)[1], "rb").read().decode("utf-16-le").splitlines()
        if value.startswith("/path:"):
            return [value.split(":", 1)[1]]
    return []


def _commit_fixture_paths(root: str, paths: list[str]) -> None:
    for path in paths:
        if os.path.isfile(path):
            shutil.copy2(path, path + ".pristine.xlsx")
            _ensure_node(root, path)
        else:
            _delete_node(root, path)
            try: os.remove(path + ".pristine.xlsx")
            except OSError: pass


def _item(path: str, root: str, status: str, versioned=True) -> bs.SvnChangeItem:
    rel = os.path.relpath(path, os.path.join(root, "develop")).replace("\\", "/")
    return bs.SvnChangeItem(
        path=path, relative_path=rel, extension=".xlsx", node_kind="file",
        node_status=status, text_status=status, prop_status="normal",
        versioned=versioned, revision=100, checked=True, selectable=True,
    )


def _fixture():
    root = tempfile.mkdtemp(prefix="branch-submit-v2-")
    _create_wc(root)
    old_appdata = os.environ.get("LOCALAPPDATA")
    os.environ["LOCALAPPDATA"] = os.path.join(root, "appdata")
    return root, old_appdata


def _cleanup(root, old_appdata):
    if old_appdata is None: os.environ.pop("LOCALAPPDATA", None)
    else: os.environ["LOCALAPPDATA"] = old_appdata
    shutil.rmtree(root, ignore_errors=True)


def test_dynamic_branches_context_and_defaults() -> None:
    root, old = _fixture()
    try:
        for branch in ("develop", "release", "sandbox", "master"):
            _book(os.path.join(root, branch, "config", "A.xlsx"), 1)
        for branch, changed in (("develop", 100), ("release", 200), ("sandbox", 300), ("master", 400)):
            _set_branch_changed(root, branch, changed)
        candidates = bs.discover_branch_candidates(root, favorites=("develop",))
        assert [item.name for item in candidates] == ["master", "sandbox", "release", "develop"]
        assert all(item.enabled for item in candidates)
        assert bs._validate_branch_name("master", [item.name for item in candidates]) == "master"
        context = bs.infer_context([os.path.join(root, "develop", "config", "A.xlsx")])
        assert context.source_branch == "develop"
        assert context.scope_path == os.path.join(root, "develop", "config")
    finally: _cleanup(root, old)


def test_status_xml_and_windows_abi() -> None:
    xml = """<?xml version='1.0'?><status><target path='.'><changelist name='ignore-on-commit'><entry path='A.xlsx'><wc-status item='modified' props='normal' revision='7' switched='true'><lock><owner>me</owner></lock></wc-status></entry></changelist></target></status>"""
    row = sp._parse_cli_status(xml, r"C:\wc")[0]
    assert row.node_status == "modified" and row.switched and row.lock_owner == "me" and row.changelist == "ignore-on-commit"
    if os.name == "nt":
        assert sp._SvnClientStatus.revision.offset == 80
        assert sp._SvnClientStatus.changelist.offset == 120


def test_scan_defaults_and_blockers() -> None:
    root, old = _fixture()
    try:
        scope = os.path.join(root, "develop", "config")
        paths = {
            "modified": os.path.join(scope, "中文修改.xlsx"),
            "unversioned": os.path.join(scope, "New.xlsx"),
            "property": os.path.join(scope, "Prop.xlsx"),
            "other": os.path.join(scope, "readme.txt"),
        }
        for path in paths.values():
            os.makedirs(os.path.dirname(path), exist_ok=True)
            open(path, "wb").close()
        records = [
            sp.SvnStatusRecord(path=paths["modified"], node_kind="file", node_status="modified", text_status="modified", prop_status="normal", versioned=True),
            sp.SvnStatusRecord(path=paths["unversioned"], node_kind="file", node_status="unversioned", versioned=False),
            sp.SvnStatusRecord(path=paths["property"], node_kind="file", node_status="modified", text_status="modified", prop_status="modified", versioned=True),
            sp.SvnStatusRecord(path=paths["other"], node_kind="file", node_status="modified", text_status="modified", prop_status="normal", versioned=True),
        ]
        with patch.object(bs, "scan_status", lambda *_args, **_kwargs: records):
            items = bs.scan_changes(root, "develop", scope)
        by_name = {os.path.basename(item.path): item for item in items}
        assert by_name["中文修改.xlsx"].checked and by_name["中文修改.xlsx"].selectable
        assert not by_name["New.xlsx"].checked and by_name["New.xlsx"].selectable
        assert not by_name["Prop.xlsx"].selectable and "属性修改" in by_name["Prop.xlsx"].reason
        assert not by_name["readme.txt"].selectable
    finally: _cleanup(root, old)


def test_high_confidence_rename_is_blocked() -> None:
    root, old = _fixture()
    try:
        scanner = FixtureScanner(root)
        old_path = os.path.join(root, "develop", "config", "Old.xlsx")
        new_path = os.path.join(root, "develop", "config", "New.xlsx")
        _book(old_path + ".pristine.xlsx", 4); _book(new_path, 4)
        with sqlite3.connect(os.path.join(root, ".svn", "wc.db")) as conn:
            _insert_node(conn, "develop/config/Old.xlsx")
        for branch in ("release", "sandbox", "master"):
            _book(os.path.join(root, branch, "config", "seed.xlsx"), 1)
        engine = bs.BranchSubmitEngine(root, status_scanner=scanner); engine.core = FakeCore
        try:
            engine.preflight(
                "develop", ["release"],
                [_item(old_path, root, "missing"), _item(new_path, root, "unversioned", False)],
                "rename", scope_path=os.path.dirname(new_path),
            )
        except RuntimeError as exc:
            assert "Repair Move" in str(exc)
        else:
            raise AssertionError("high-confidence rename was not blocked")
    finally: _cleanup(root, old)


def test_preflight_modify_add_delete_and_dirty_block() -> None:
    root, old = _fixture()
    try:
        scanner = FixtureScanner(root)
        items = []
        # modify
        src_m = os.path.join(root, "develop", "config", "M.xlsx")
        _book(src_m, 9); _book(src_m + ".pristine.xlsx", 1)
        items.append(_item(src_m, root, "modified"))
        # add
        src_a = os.path.join(root, "develop", "config", "A.xlsx")
        _book(src_a, 5); items.append(_item(src_a, root, "unversioned", False))
        # delete
        src_d = os.path.join(root, "develop", "config", "D.xlsx")
        _book(src_d + ".pristine.xlsx", 3); items.append(_item(src_d, root, "missing"))
        with sqlite3.connect(os.path.join(root, ".svn", "wc.db")) as conn:
            for rel in ("develop/config/M.xlsx", "develop/config/D.xlsx", "release/config/M.xlsx", "release/config/D.xlsx"):
                _insert_node(conn, rel)
        _book(os.path.join(root, "release", "config", "M.xlsx"), 1)
        _book(os.path.join(root, "release", "config", "M.xlsx.pristine.xlsx"), 1)
        _book(os.path.join(root, "release", "config", "D.xlsx"), 3)
        _book(os.path.join(root, "release", "config", "D.xlsx.pristine.xlsx"), 3)
        # Make all candidate branch roots discoverable.
        _book(os.path.join(root, "sandbox", "config", "seed.xlsx"), 1)
        _book(os.path.join(root, "master", "config", "seed.xlsx"), 1)
        engine = bs.BranchSubmitEngine(root, status_scanner=scanner)
        engine.core = FakeCore
        batch = engine.preflight("develop", ["release"], items, "配置调整", scope_path=os.path.join(root, "develop", "config"))
        assert batch.source_status == "ready", batch.error
        assert [plan.operation for plan in batch.files] == ["modify", "add", "delete"]
        assert all(plan.actions["release"].state == "planned" for plan in batch.files)
        dirty = os.path.join(root, "release", "config", "M.xlsx")
        scanner.overrides[dirty] = sp.SvnStatusRecord(path=dirty, node_kind="file", node_status="modified", text_status="modified", prop_status="normal", versioned=True)
        blocked = engine.preflight("develop", ["release"], [items[0]], "配置调整", scope_path=os.path.join(root, "develop", "config"))
        assert blocked.source_status == "pending"
        assert blocked.files[0].actions["release"].state == "blocked"
    finally: _cleanup(root, old)


def test_source_partial_selection_stops_propagation() -> None:
    root, old = _fixture()
    try:
        scanner = FixtureScanner(root); items=[]
        for name, value in (("A.xlsx", 7), ("B.xlsx", 8)):
            source=os.path.join(root,"develop","config",name);target=os.path.join(root,"release","config",name)
            _book(source,value);_book(source+".pristine.xlsx",1);_book(target,1);_book(target+".pristine.xlsx",1)
            with sqlite3.connect(os.path.join(root,".svn","wc.db")) as conn:
                _insert_node(conn,f"develop/config/{name}");_insert_node(conn,f"release/config/{name}")
            items.append(_item(source,root,"modified"))
        _book(os.path.join(root,"sandbox","config","seed.xlsx"),1);_book(os.path.join(root,"master","config","seed.xlsx"),1)
        calls=[]
        def runner(args, **_kwargs):
            calls.append(args)
            if "/command:commit" in args and len(calls)==1:
                _commit_fixture_paths(root,_commit_paths(args)[:1])
            return SimpleNamespace(returncode=0)
        engine=bs.BranchSubmitEngine(root,runner=runner,status_scanner=scanner);engine.core=FakeCore
        batch=engine.preflight("develop",["release"],items,"partial",scope_path=os.path.join(root,"develop","config"))
        engine.commit(batch)
        assert batch.source_status=="partial",batch.error
        assert [p.source_state for p in batch.files].count("committed")==1
        assert batch.target_status["release"]=="ready"
        assert batch.superseded_by.endswith("-committed")
        child=bs.BranchSubmitBatch.load(os.path.join(bs.settings_dir(),"batches",batch.superseded_by,"batch.json"))
        assert child.source_status=="committed" and len(child.files)==1
        assert len(calls)==1
    finally:_cleanup(root,old)


def test_server_success_beats_error_exit_code() -> None:
    root, old = _fixture()
    try:
        scanner=FixtureScanner(root)
        source=os.path.join(root,"develop","config","A.xlsx");target=os.path.join(root,"release","config","A.xlsx")
        _book(source,9);_book(source+".pristine.xlsx",1);_book(target,1);_book(target+".pristine.xlsx",1)
        with sqlite3.connect(os.path.join(root,".svn","wc.db")) as conn:
            _insert_node(conn,"develop/config/A.xlsx");_insert_node(conn,"release/config/A.xlsx")
        _book(os.path.join(root,"sandbox","config","seed.xlsx"),1);_book(os.path.join(root,"master","config","seed.xlsx"),1)
        commit_count=0
        def runner(args,**_kwargs):
            nonlocal commit_count
            if "/command:commit" in args:
                commit_count+=1;_commit_fixture_paths(root,_commit_paths(args));return SimpleNamespace(returncode=2)
            return SimpleNamespace(returncode=0)
        engine=bs.BranchSubmitEngine(root,runner=runner,status_scanner=scanner);engine.core=FakeCore
        batch=engine.preflight("develop",["release"],[_item(source,root,"modified")],"timeout",scope_path=os.path.dirname(source))
        engine.commit(batch)
        assert batch.source_status=="committed"
        assert batch.target_status["release"]=="committed",batch.error
        assert commit_count==2
    finally:_cleanup(root,old)


def test_target_partial_and_restore_guard() -> None:
    root, old = _fixture()
    try:
        scanner=FixtureScanner(root);items=[]
        for name,value in (("A.xlsx",7),("B.xlsx",8)):
            source=os.path.join(root,"develop","config",name);target=os.path.join(root,"release","config",name)
            _book(source,value);_book(source+".pristine.xlsx",1);_book(target,1);_book(target+".pristine.xlsx",1)
            with sqlite3.connect(os.path.join(root,".svn","wc.db")) as conn:
                _insert_node(conn,f"develop/config/{name}");_insert_node(conn,f"release/config/{name}")
            items.append(_item(source,root,"modified"))
        _book(os.path.join(root,"sandbox","config","seed.xlsx"),1);_book(os.path.join(root,"master","config","seed.xlsx"),1)
        commit_count=0
        def runner(args,**_kwargs):
            nonlocal commit_count
            if "/command:commit" in args:
                commit_count+=1
                paths=_commit_paths(args)
                _commit_fixture_paths(root,paths if commit_count==1 else paths[:1])
            return SimpleNamespace(returncode=0)
        engine=bs.BranchSubmitEngine(root,runner=runner,status_scanner=scanner);engine.core=FakeCore
        batch=engine.preflight("develop",["release"],items,"target partial",scope_path=os.path.join(root,"develop","config"))
        engine.commit(batch)
        assert batch.target_status["release"]=="partial",batch.error
        states=[p.actions["release"].state for p in batch.files]
        assert sorted(states)==["committed","prepared"]
        pending=next(p for p in batch.files if p.actions["release"].state=="prepared")
        target=os.path.join(root,"release",*pending.relative_path.split("/"))
        original_hash=pending.actions["release"].target_before_hash
        engine.restore_uncommitted(batch)
        assert pending.actions["release"].state=="restored"
        assert bs._sha256(target)==original_hash
    finally:_cleanup(root,old)


def test_resume_reconciles_commits_before_reopening_dialog() -> None:
    root, old = _fixture()
    try:
        scanner=FixtureScanner(root)
        source=os.path.join(root,"develop","config","A.xlsx");target=os.path.join(root,"release","config","A.xlsx")
        _book(source,9);_book(source+".pristine.xlsx",1);_book(target,1);_book(target+".pristine.xlsx",1)
        with sqlite3.connect(os.path.join(root,".svn","wc.db")) as conn:
            _insert_node(conn,"develop/config/A.xlsx");_insert_node(conn,"release/config/A.xlsx")
        for branch in ("sandbox","master"):_book(os.path.join(root,branch,"config","seed.xlsx"),1)
        calls=[]
        def runner(args,**_kwargs):
            calls.append(args)
            if "/command:commit" in args:_commit_fixture_paths(root,_commit_paths(args))
            return SimpleNamespace(returncode=0)
        engine=bs.BranchSubmitEngine(root,runner=runner,status_scanner=scanner);engine.core=FakeCore
        batch=engine.preflight("develop",["release"],[_item(source,root,"modified")],"resume",scope_path=os.path.dirname(source))
        # Simulate a process crash after the server accepted the source commit.
        _commit_fixture_paths(root,[source])
        engine.commit(batch)
        assert batch.source_status=="committed" and batch.target_status["release"]=="committed"
        assert sum("/command:commit" in args for args in calls)==1  # target only

        # Simulate a second batch whose target commit succeeded before state save.
        _book(source,10)
        source_item=_item(source,root,"modified")
        batch2=engine.preflight("develop",["release"],[source_item],"target crash",scope_path=os.path.dirname(source))
        _commit_fixture_paths(root,[source]);batch2.source_status="committed";batch2.source_revision_after=200
        plan=batch2.files[0]
        status_map=sp.records_by_path(scanner(os.path.join(root,"release")))
        engine._fresh_target_action(batch2,plan,"release",status_map)
        engine._prepare_target_action(batch2,plan,"release")
        _commit_fixture_paths(root,[target])
        calls.clear();engine.commit(batch2)
        assert batch2.target_status["release"]=="committed"
        assert not any("/command:commit" in args for args in calls)
    finally:_cleanup(root,old)


def test_write_intent_crash_restore_and_corrupt_state_detection() -> None:
    root, old = _fixture()
    try:
        scanner=FixtureScanner(root)
        source=os.path.join(root,"develop","config","A.xlsx");target=os.path.join(root,"release","config","A.xlsx")
        _book(source,9);_book(source+".pristine.xlsx",1);_book(target,1);_book(target+".pristine.xlsx",1)
        with sqlite3.connect(os.path.join(root,".svn","wc.db")) as conn:
            _insert_node(conn,"develop/config/A.xlsx");_insert_node(conn,"release/config/A.xlsx")
        for branch in ("sandbox","master"):_book(os.path.join(root,branch,"config","seed.xlsx"),1)
        engine=bs.BranchSubmitEngine(root,status_scanner=scanner);engine.core=FakeCore
        batch=engine.preflight("develop",["release"],[_item(source,root,"modified")],"intent",scope_path=os.path.dirname(source))
        action=batch.files[0].actions["release"]
        backup=bs._artifact_path(batch.folder,os.path.join("backups","release"),batch.files[0].relative_path)
        bs._safe_copy(target,backup);action.backup_path=backup;action.target_before_hash=bs._sha256(backup)
        batch.event("prepare-intent",target="release",path=batch.files[0].relative_path,operation="modify")
        bs._safe_copy(action.candidate_path,target)  # crash here: action remains planned
        engine.restore_uncommitted(batch)
        assert action.state=="restored" and bs._sha256(target)==action.target_before_hash
        broken=os.path.join(bs.settings_dir(),"batches","broken","batch.json")
        os.makedirs(os.path.dirname(broken),exist_ok=True)
        with open(broken,"w",encoding="utf-8") as stream:stream.write("{broken")
        assert broken in bs.list_corrupt_batch_files()
    finally:_cleanup(root,old)


def test_entrypoint_registry_scope_and_real_status_child() -> None:
    repo=os.path.dirname(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    )
    install=open(os.path.join(repo,"install_context_menu.bat"),encoding="utf-8").read()
    uninstall=open(os.path.join(repo,"uninstall_context_menu.bat"),encoding="utf-8").read()
    for key in (
        r"SystemFileAssociations\.xlsx\shell\SowMultiBranchSVNSubmit",
        r"Directory\shell\SowMultiBranchSVNSubmit",
        r"Directory\Background\shell\SowMultiBranchSVNSubmit",
    ):
        assert key in install and key in uninstall
    assert "--branch-submit" in install
    assert "$percent=[char]37" in install
    assert "Token='1'" in install and "Token='V'" in install
    assert "ASCII-only" in install
    assert "Remove-ItemProperty -LiteralPath $item.Path -Name 'Position'" in install
    assert "New-ItemProperty -LiteralPath $item.Path -Name 'Position'" not in install
    assert "TortoiseSVN" not in uninstall
    real_wc=r"C:\sow_main\excel"
    if os.environ.get("SOW_SKIP_REAL_WC_TESTS") == "1":
        print("SKIP real working-copy status scan (set by synthetic test profile)")
        return
    if os.path.isfile(os.path.join(real_wc,".svn","wc.db")):
        rows=sp.scan_status(real_wc)
        assert all(os.path.isabs(row.path) for row in rows)


def test_cli_handoff_for_folder() -> None:
    root, old = _fixture()
    try:
        _book(os.path.join(root,"develop","config","A.xlsx"),1)
        for branch in ("release","sandbox","master"):_book(os.path.join(root,branch,"config","seed.xlsx"),1)
        folder=os.path.join(root,"develop","config");captured=[]
        with patch.object(bs,"launch_ui",lambda paths=None:captured.append(list(paths or []))):
            with patch.object(__import__("sys"),"argv",["sow_merge_tool.py","--branch-submit",folder]):smt.main()
        assert captured==[[folder]]
    finally:_cleanup(root,old)


if __name__ == "__main__":
    tests=[
        test_dynamic_branches_context_and_defaults,
        test_status_xml_and_windows_abi,
        test_scan_defaults_and_blockers,
        test_high_confidence_rename_is_blocked,
        test_preflight_modify_add_delete_and_dirty_block,
        test_source_partial_selection_stops_propagation,
        test_server_success_beats_error_exit_code,
        test_target_partial_and_restore_guard,
        test_resume_reconciles_commits_before_reopening_dialog,
        test_write_intent_crash_restore_and_corrupt_state_detection,
        test_entrypoint_registry_scope_and_real_status_child,
        test_cli_handoff_for_folder,
    ]
    for test in tests:
        test();print("PASS",test.__name__)
    print("branch submit adversarial smoke tests passed")

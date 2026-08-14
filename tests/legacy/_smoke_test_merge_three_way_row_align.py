"""Validate that _merge_three_way uses row alignment (not physical row numbers).

Scenario with an inserted row in mine:
  base   : [id, A, B, C]
  mine   : [id, X, A, B, C]   (mine inserted row "X" after the header)
  theirs : [id, A, B, C2]     (theirs changed the last cell C -> C2)

With the old physical-row-number merge, mine[r] and theirs[r] would be compared
at the same row index, so the inserted row shifts everything and produces spurious
conflicts/diffs. With row alignment, A/B/C line up correctly: theirs' C->C2 change
applies cleanly onto mine and the inserted X is preserved, with no conflicts.
"""
import os

from openpyxl import Workbook, load_workbook

import sow_merge_tool as mod
from _test_temp_utils import make_temp_dir


def _make_book(path: str, rows: list[object]):
    wb = Workbook()
    ws = wb.active
    ws.title = "S1"
    for idx, value in enumerate(rows, start=1):
        ws.cell(row=idx, column=1).value = value
    wb.save(path)
    wb.close()


def main():
    root = make_temp_dir("sow_merge3w_rowalign_")
    base = os.path.join(root, "base.xlsx")
    mine = os.path.join(root, "mine.xlsx")
    theirs = os.path.join(root, "theirs.xlsx")
    merged = os.path.join(root, "merged.xlsx")

    _make_book(base, ["id", "A", "B", "C"])
    _make_book(mine, ["id", "X", "A", "B", "C"])
    _make_book(theirs, ["id", "A", "B", "C2"])

    # Scan (active UI path) must see no conflict: theirs changed C, mine didn't.
    scan_conflicts, _ = mod._scan_three_way_conflicts(base, mine, theirs)
    assert not scan_conflicts, scan_conflicts

    # Auto-merge must agree with the aligned scan: no conflicts.
    conflicts, preview, cmap = mod._merge_three_way(base, mine, theirs, merged, save_merged=True)
    assert not conflicts, ("expected no conflict from row-aligned merge", conflicts, cmap)

    # No conflicts -> merged_path written directly.
    assert os.path.exists(merged), merged
    wb = load_workbook(merged, data_only=False)
    try:
        ws = wb["S1"]
        values = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    finally:
        wb.close()

    # Mine's inserted row "X" preserved; theirs' C->C2 applied onto the aligned row.
    assert values == ["id", "X", "A", "B", "C2"], values

    print("SMOKE_MERGE_THREE_WAY_ROW_ALIGN_OK")


if __name__ == "__main__":
    main()

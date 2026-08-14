"""OpenSpec 4.4 GUI regressions for adaptive three-way presentation."""

from __future__ import annotations

import inspect
import os
import shutil
import time
from contextlib import ExitStack
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

import sow_merge_tool as smt
from _test_temp_utils import make_temp_dir


def _make_book(path: str, marker: str, *, rows: int = 60, columns: int = 18) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    for column in range(1, columns + 1):
        worksheet.cell(1, column).value = f"column-{column:02d}"
    for row in range(2, rows + 1):
        worksheet.cell(row, 1).value = f"row-{row:03d}"
        for column in range(2, columns + 1):
            worksheet.cell(row, column).value = f"{marker}-{row}-{column}"
    workbook.save(path)
    workbook.close()


def _field(value, *names, default=None):
    if isinstance(value, dict):
        for name in names:
            if name in value:
                return value[name]
        return default
    for name in names:
        if hasattr(value, name):
            return getattr(value, name)
    return default


def _scenario_name(value) -> str:
    if hasattr(value, "name"):
        value = value.name
    elif hasattr(value, "value"):
        value = value.value
    return str(value).strip().replace("-", "_").replace(" ", "_").upper()


def _build_context(base, mine, theirs, merged, pristine=None):
    builder = getattr(smt, "build_merge_launch_context", None)
    assert callable(builder), "missing build_merge_launch_context API"
    signature = inspect.signature(builder)
    candidates = {
        "base_path": base,
        "source_base_path": base,
        "mine_path": mine,
        "theirs_path": theirs,
        "merged_path": merged,
        "target_pristine_path": pristine,
    }
    kwargs = {
        name: value
        for name, value in candidates.items()
        if name in signature.parameters
    }
    try:
        return builder(**kwargs)
    except TypeError:
        try:
            return builder(base, mine, theirs, merged, target_pristine_path=pristine)
        except TypeError:
            if pristine is None:
                return builder(base, mine, theirs, merged)
            return builder(base, mine, theirs, merged, pristine)


def _analysis_outcome(context):
    runner = getattr(smt, "run_startup_merge_analysis", None)
    assert callable(runner), "missing run_startup_merge_analysis API"
    result = runner(context)
    outcome = _field(result, "outcome", "startup_outcome")
    if outcome is None and isinstance(result, tuple):
        for item in result:
            if type(item).__name__ == "StartupMergeOutcome":
                outcome = item
                break
    if outcome is None and type(result).__name__ == "StartupMergeOutcome":
        outcome = result
    assert outcome is not None, result
    return outcome


def _pump(root, seconds: float = 0.05) -> None:
    deadline = time.time() + seconds
    while time.time() < deadline:
        root.update_idletasks()
        root.update()
        time.sleep(0.005)


def _wait_until(root, predicate, message: str, timeout: float = 20.0) -> None:
    deadline = time.time() + timeout
    while time.time() < deadline:
        _pump(root, 0.025)
        if predicate():
            return
    raise AssertionError(message)


def _construct_app(
    mine: str,
    theirs: str,
    *,
    base: str | None,
    merged: str,
    context,
    outcome=None,
):
    raw_mine = mine
    raw_theirs = theirs
    raw_base = base
    ui_mine = mine
    ui_theirs = theirs
    ui_base = base
    if base and theirs and context is not None:
        # Production ``main()`` preserves raw sidecar identities in the
        # context, but passes stable ``.xlsx/.xlsm`` copies to openpyxl/UI.
        # Mirror that boundary so this test validates GUI behavior rather than
        # openpyxl's intentional rejection of a raw ``.rN`` filename.
        for role in ("base", "mine", "theirs"):
            identity = context.identity_for(role)
            if identity is not None and identity.path and not identity.stable_path:
                identity.stable_path = smt._ensure_xlsx_copy(identity.path)
        base_identity = context.identity_for("base")
        mine_identity = context.identity_for("mine")
        theirs_identity = context.identity_for("theirs")
        ui_base = base_identity.effective_path if base_identity else base
        ui_mine = (
            outcome.candidate_path
            if outcome is not None and outcome.candidate_path
            else (mine_identity.effective_path if mine_identity else mine)
        )
        ui_theirs = theirs_identity.effective_path if theirs_identity else theirs

    signature = inspect.signature(smt.SowMergeApp)
    optional = {
        "merge_mode": bool(base and theirs),
        "merged_path": merged,
        "base_path": ui_base,
        "raw_base": raw_base,
        "raw_mine": raw_mine,
        "raw_theirs": raw_theirs,
        "launch_context": context,
        "startup_outcome": outcome,
    }
    kwargs = {
        name: value
        for name, value in optional.items()
        if name in signature.parameters
    }

    with ExitStack() as stack:
        # Startup summaries are tested separately.  A GUI self-test must never
        # block waiting for a modal acknowledgement.
        for name in (
            "_show_startup_merge_outcome_dialog",
            "show_startup_merge_outcome_dialog",
        ):
            if hasattr(smt.SowMergeApp, name):
                stack.enter_context(
                    patch.object(smt.SowMergeApp, name, lambda *_args, **_kwargs: None)
                )
        app = smt.SowMergeApp(ui_mine, ui_theirs, **kwargs)

    assert _field(app, "launch_context", "merge_launch_context") is context, (
        "SowMergeApp must retain the complete launch context",
        kwargs,
    )
    try:
        app.root.deiconify()
        app.root.geometry("1180x820")
    except Exception:
        pass
    app.nb.select(app._sheet_containers["Data"])
    _wait_until(
        app.root,
        lambda: (
            app.sheet_views.get("Data") is not None
            and bool(getattr(app.sheet_views["Data"], "_data_ready", False))
            and bool(getattr(app.sheet_views["Data"], "row_pairs", None))
        ),
        "Data SheetView did not finish initial loading",
    )
    view = app.sheet_views["Data"]
    return app, view


def _rgb(widget, color) -> tuple[int, int, int]:
    return tuple(int(value) for value in widget.winfo_rgb(str(color)))


def _is_white(widget, color) -> bool:
    return _rgb(widget, color) == _rgb(widget, "#ffffff")


def _workspace_color(app):
    for name in (
        "workspace_color",
        "workspace_chrome_color",
        "workspace_bg",
        "chrome_color",
        "_workspace_color",
        "_workspace_bg",
        "_chrome_color",
    ):
        value = getattr(app, name, None)
        if value:
            return str(value)
    context = _field(app, "launch_context", "merge_launch_context")
    scenario = _field(context, "scenario", "merge_scenario")
    for mapping_name in (
        "WORKSPACE_COLORS",
        "WORKSPACE_CHROME_COLORS",
        "_WORKSPACE_COLORS",
    ):
        mapping = getattr(smt, mapping_name, None)
        if not isinstance(mapping, dict):
            continue
        for key in (scenario, _scenario_name(scenario), _scenario_name(scenario).lower()):
            if key in mapping:
                return str(mapping[key])
    raise AssertionError("GUI does not expose centralized workspace chrome color")


def _widget_background(widget):
    for option in ("background", "bg"):
        try:
            value = widget.cget(option)
        except Exception:
            continue
        if value:
            return str(value)
    return None


def _chrome_uses_color(app, expected: str, spreadsheet_widgets: set[object]) -> bool:
    expected_rgb = _rgb(app.root, expected)
    queue = [app.root]
    while queue:
        widget = queue.pop()
        try:
            queue.extend(widget.winfo_children())
        except Exception:
            pass
        if widget in spreadsheet_widgets:
            continue
        background = _widget_background(widget)
        if not background:
            continue
        try:
            if _rgb(app.root, background) == expected_rgb:
                return True
        except Exception:
            continue
    return False


def _mode_fixture(root: str, mode: str):
    mine = os.path.join(root, f"{mode}-Design.xlsx")
    merged = os.path.join(root, f"{mode}-merged.xlsx")
    if mode == "two_way":
        left = os.path.join(root, f"{mode}-left.xlsx")
        _make_book(left, "left", rows=6, columns=5)
        _make_book(mine, "right", rows=6, columns=5)
        return left, mine, None, merged
    if mode == "update":
        base = os.path.join(root, "Design.xlsx.r100")
        theirs = os.path.join(root, "Design.xlsx.r110")
    elif mode == "branch":
        base = os.path.join(root, "Design.xlsx.merge-left.r200")
        theirs = os.path.join(root, "Design.xlsx.merge-right.r220")
    else:
        base = os.path.join(root, "Design-old-copy.xlsx")
        theirs = os.path.join(root, "Design-incoming-copy.xlsx")
    _make_book(base, f"{mode}-base", rows=6, columns=5)
    _make_book(mine, f"{mode}-mine", rows=6, columns=5)
    _make_book(theirs, f"{mode}-theirs", rows=6, columns=5)
    return base, mine, theirs, merged


def test_mode_colors_are_chrome_only_and_spreadsheets_stay_white() -> None:
    root = make_temp_dir("sow_adaptive_workspace_colors_")
    colors = {}
    for mode in ("two_way", "update", "branch", "unknown"):
        base, mine, theirs, merged = _mode_fixture(root, mode)
        if mode == "two_way":
            context = _build_context(base, mine, None, merged)
            app, view = _construct_app(
                base,
                mine,
                base=None,
                merged=merged,
                context=context,
            )
        else:
            context = _build_context(base, mine, theirs, merged)
            app, view = _construct_app(
                mine,
                theirs,
                base=base,
                merged=merged,
                context=context,
            )
        try:
            _pump(app.root, 0.1)
            color = _workspace_color(app)
            colors[mode] = _rgb(app.root, color)
            spreadsheet_widgets = {
                view.left,
                view.right,
                view.base,
                view.cursor_cmp,
                view.cell_cmp_text,
            }
            assert _chrome_uses_color(app, color, spreadsheet_widgets), (
                mode,
                color,
                "no surrounding chrome widget used the scenario color",
            )
            for name in ("left", "right", "base"):
                widget = getattr(view, name)
                assert _is_white(widget, widget.cget("background")), (
                    mode,
                    name,
                    widget.cget("background"),
                )
            for widget in (view.left, view.right, view.base):
                for tag in ("diffcell", "selcell", "paddingrow", "paddingcol"):
                    tag_bg = widget.tag_cget(tag, "background")
                    if tag_bg:
                        assert _rgb(widget, tag_bg) != colors[mode], (
                            "scenario chrome color leaked into spreadsheet state rendering",
                            mode,
                            tag,
                            tag_bg,
                        )
        finally:
            app._shutdown_root()

    assert colors["two_way"] == colors["unknown"], colors
    assert colors["update"] != colors["two_way"], colors
    assert colors["branch"] != colors["two_way"], colors
    assert colors["branch"] != colors["update"], colors
    assert colors["update"] != _rgb_from_tuple_white(), colors
    assert colors["branch"] != _rgb_from_tuple_white(), colors


def _rgb_from_tuple_white() -> tuple[int, int, int]:
    # Tk uses 16-bit RGB values on Windows.
    return (65535, 65535, 65535)


def _pane_count(view) -> int:
    return len(tuple(view._main_paned.panes()))


def _toggle_text(view) -> str:
    widget = getattr(view, "three_way_cb", None)
    return str(widget.cget("text") if widget is not None else "")


def _view_state(app, view):
    return {
        "sheet": getattr(app, "selected_sheet", None),
        "selected_pair": getattr(view, "selected_pair_idx", None),
        "selected_line": getattr(view, "_main_sel_line", None),
        "selected_col": getattr(view, "_main_sel_col", None),
        "only_diff": int(view.only_diff_var.get()),
        "left_x": float((view.left.xview() or (0.0, 1.0))[0]),
        "left_y": float((view.left.yview() or (0.0, 1.0))[0]),
        "left_insert": str(view.left.index("insert")),
        "base_insert": str(view.base.index("insert")),
        "right_insert": str(view.right.index("insert")),
        "row_pairs": tuple(view.row_pairs),
        "pending_a": dict(app.manual_a_cell_ops),
        "pending_b": dict(app.manual_b_cell_ops),
        "undo": list(app.undo_stack),
        "candidate_b2": app._wb_a_edit["Data"]["B2"].value,
    }


def _assert_state_preserved(before, after) -> None:
    for key in (
        "sheet",
        "selected_pair",
        "selected_line",
        "selected_col",
        "only_diff",
        "left_insert",
        "base_insert",
        "right_insert",
        "row_pairs",
        "pending_a",
        "pending_b",
        "undo",
        "candidate_b2",
    ):
        assert after[key] == before[key], (key, before[key], after[key])
    assert abs(after["left_x"] - before["left_x"]) <= 0.03, (before, after)
    assert abs(after["left_y"] - before["left_y"]) <= 0.03, (before, after)


def test_proven_redundant_pane_folds_and_toggle_preserves_state() -> None:
    root = make_temp_dir("sow_adaptive_workspace_fold_")
    base = os.path.join(root, "Design.xlsx.r100")
    mine = os.path.join(root, "Design.xlsx")
    theirs = os.path.join(root, "Design.xlsx.r110")
    merged = os.path.join(root, "merged.xlsx")
    _make_book(base, "base")
    shutil.copy2(base, mine)
    _make_book(theirs, "theirs")

    context = _build_context(base, mine, theirs, merged)
    outcome = _analysis_outcome(context)
    folded_identity = _field(
        outcome,
        "folded_identity",
        "folded_role",
        "redundant_identity",
        "redundant_role",
    )
    assert folded_identity is not None, outcome

    app, view = _construct_app(
        mine,
        theirs,
        base=base,
        merged=merged,
        context=context,
        outcome=outcome,
    )
    try:
        _wait_until(
            app.root,
            lambda: (
                app._edit_workbooks_ready()
                and view._derive_lifecycle_state() == "READY"
            ),
            "fold fixture did not reach editable READY state",
        )
        _pump(app.root, 0.15)
        assert _pane_count(view) == 2, (
            "proven redundant pane must be folded by default",
            _pane_count(view),
            outcome,
        )
        assert "展开三方" in _toggle_text(view), _toggle_text(view)

        pair_idx = min(5, len(view.row_pairs) - 1)
        line = int(view.row_to_line.get(pair_idx, pair_idx + 1))
        view._highlight_selected_line(line)
        view._set_main_selected_cell(line, 3)
        view.selected_pair_idx = pair_idx
        if not view.only_diff_var.get():
            view.only_diff_var.set(1)
            view._toggle_only_diff()
            _wait_until(
                app.root,
                lambda: (
                    view.only_diff_var.get() == 1
                    and view._derive_lifecycle_state() == "READY"
                ),
                "only-difference state did not become exact before fold",
            )
        view.left.xview_moveto(0.25)
        view._yview_both("moveto", 0.35)
        view.left.mark_set("insert", f"{line}.3")
        view.base.mark_set("insert", f"{line}.4")
        view.right.mark_set("insert", f"{line}.5")
        app.manual_a_cell_ops[("Data", 2, 2)] = "pending-choice"
        app.manual_b_cell_ops[("Data", 3, 3)] = "pending-other-choice"
        app.undo_stack.append({"kind": "fold-state-sentinel", "sheet": "Data"})
        _pump(app.root, 0.1)
        folded_state = _view_state(app, view)

        refresh_calls = []
        original_refresh = view.refresh

        def _forbid_refresh(*args, **kwargs):
            refresh_calls.append((args, kwargs))
            raise AssertionError("fold/expand must not refresh or rescan the workbook model")

        view.refresh = _forbid_refresh
        try:
            view.three_way_var.set(1)
            view._toggle_three_way_view()
            _pump(app.root, 0.15)
            assert _pane_count(view) == 3
            expanded_state = _view_state(app, view)
            _assert_state_preserved(folded_state, expanded_state)

            view.three_way_var.set(0)
            view._toggle_three_way_view()
            _pump(app.root, 0.15)
            assert _pane_count(view) == 2
            assert "展开三方" in _toggle_text(view)
            refolded_state = _view_state(app, view)
            _assert_state_preserved(expanded_state, refolded_state)
            assert not refresh_calls, refresh_calls
        finally:
            view.refresh = original_refresh
    finally:
        app._shutdown_root()


def test_no_proven_equality_keeps_all_three_panes() -> None:
    root = make_temp_dir("sow_adaptive_workspace_no_fold_")
    base = os.path.join(root, "Design.xlsx.r100")
    mine = os.path.join(root, "Design.xlsx")
    theirs = os.path.join(root, "Design.xlsx.r110")
    merged = os.path.join(root, "merged.xlsx")
    _make_book(base, "base", rows=8, columns=6)
    _make_book(mine, "mine", rows=8, columns=6)
    _make_book(theirs, "theirs", rows=8, columns=6)
    context = _build_context(base, mine, theirs, merged)
    outcome = _analysis_outcome(context)
    app, view = _construct_app(
        mine,
        theirs,
        base=base,
        merged=merged,
        context=context,
        outcome=outcome,
    )
    try:
        _pump(app.root, 0.1)
        assert _pane_count(view) == 3, (
            "full three-pane layout is mandatory without proven equality",
            _pane_count(view),
            outcome,
        )
    finally:
        app._shutdown_root()


def test_mine_theirs_equivalence_folds_theirs_not_base() -> None:
    root = make_temp_dir("sow_adaptive_workspace_fold_theirs_")
    base = os.path.join(root, "Design.xlsx.r100")
    mine = os.path.join(root, "Design.xlsx")
    theirs = os.path.join(root, "Design.xlsx.r110")
    merged = os.path.join(root, "merged.xlsx")
    _make_book(base, "base", rows=8, columns=6)
    _make_book(mine, "common-result", rows=8, columns=6)
    shutil.copy2(mine, theirs)

    context = _build_context(base, mine, theirs, merged)
    outcome = _analysis_outcome(context)
    assert str(_field(outcome, "folded_identity", default="")).lower() == "theirs", outcome
    app, view = _construct_app(
        mine,
        theirs,
        base=base,
        merged=merged,
        context=context,
        outcome=outcome,
    )
    try:
        _pump(app.root, 0.1)
        panes = {str(pane) for pane in view._main_paned.panes()}
        assert str(view._mid_wrap) in panes, "Base must remain visible when Mine equals Theirs"
        assert str(view._left_wrap) in panes, "Mine/candidate must remain visible"
        assert str(view._right_wrap) not in panes, "only redundant Theirs should be folded"
        assert "展开三方" in _toggle_text(view), _toggle_text(view)
    finally:
        app._shutdown_root()


def test_cross_branch_actions_use_source_target_role_labels() -> None:
    root = make_temp_dir("sow_adaptive_workspace_branch_roles_")
    base = os.path.join(root, "Building.xlsx.merge-left.r100")
    mine = os.path.join(root, "Building.xlsx")
    theirs = os.path.join(root, "Building.xlsx.merge-right.r101")
    merged = os.path.join(root, "merged.xlsx")
    _make_book(base, "before", rows=8, columns=6)
    _make_book(mine, "target", rows=8, columns=6)
    _make_book(theirs, "after", rows=8, columns=6)
    context = _build_context(base, mine, theirs, merged)
    assert _scenario_name(_field(context, "scenario", "merge_scenario")) == "CROSS_BRANCH_MERGE"
    outcome = _analysis_outcome(context)
    app, view = _construct_app(
        mine,
        theirs,
        base=base,
        merged=merged,
        context=context,
        outcome=outcome,
    )
    try:
        _pump(app.root, 0.1)
        view._refresh_column_action_buttons()
        assert "Target Working" in view.left_title.cget("text")
        assert view.mid_title.cget("text") == "Source Before"
        assert view.right_title.cget("text") == "Source After"
        assert view.use_base_btn is not None
        assert view.use_base_btn.cget("text") == "保留Target Working"
        assert view.use_mine_col_btn.cget("text") == "保留Target Working列"
        assert view.use_base_col_btn.cget("text") == "采用Source Before列"
        assert view.use_theirs_col_btn.cget("text") == "采用Source After列"
        assert smt.merge_side_label(context, "A") == "Target Working"
        assert smt.merge_side_label(context, "BASE") == "Source Before"
        assert smt.merge_side_label(context, "B") == "Source After"
    finally:
        app._shutdown_root()


def main() -> None:
    tests = (
        test_mode_colors_are_chrome_only_and_spreadsheets_stay_white,
        test_proven_redundant_pane_folds_and_toggle_preserves_state,
        test_no_proven_equality_keeps_all_three_panes,
        test_mine_theirs_equivalence_folds_theirs_not_base,
        test_cross_branch_actions_use_source_target_role_labels,
    )
    for test in tests:
        test()
        print(f"PASS: {test.__name__}")
    print(f"PASS: adaptive three-way workspace GUI ({len(tests)} tests)")


if __name__ == "__main__":
    main()

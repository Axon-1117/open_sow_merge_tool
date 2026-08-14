"""GUI self-test: bottom bar remains aligned and visible.

Run:
  .venv\\Scripts\\python.exe _gui_self_test_bottom_bar_aligned.py
"""

import os
import time
from openpyxl import Workbook
from _test_temp_utils import make_temp_dir


def _make_xlsx(path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws.cell(row=1, column=1).value = "h1"
    ws.cell(row=2, column=1).value = "v1"
    wb.save(path)


def _pump(root, seconds: float = 0.8):
    end = time.time() + seconds
    while time.time() < end:
        root.update_idletasks()
        root.update()
        time.sleep(0.02)


def _run_2way():
    td = make_temp_dir(prefix="sow_merge_gui_test_bottom_align_")
    fb_dir = os.path.join(td, "normal")
    os.makedirs(fb_dir, exist_ok=True)

    # file_a under temp dir: SaveA button should be hidden by design.
    fa = os.path.join(td, "a.xlsx")
    fb = os.path.join(fb_dir, "b.xlsx")
    _make_xlsx(fa)
    _make_xlsx(fb)

    import sow_merge_tool as mod

    app = mod.SowMergeApp(fa, fb)
    sheet = app.common_sheets[0]
    view = app.sheet_views.get(sheet)
    if view is None:
        app.nb.select(app._sheet_containers[sheet])
        _pump(app.root, 0.2)
        view = app.sheet_views[sheet]

    _pump(app.root, 0.2)

    y_left = view.hsb_left.winfo_rooty()
    y_right = view.hsb_right.winfo_rooty()
    assert y_left == y_right, f"Expected aligned horizontal bars; left_y={y_left} right_y={y_right}"

    try:
        app.root.destroy()
    except Exception:
        pass

def _run_3way_bottom_nav_visible():
    td = make_temp_dir(prefix="sow_merge_gui_test_bottom_nav_3way_")
    base = os.path.join(td, "base.xlsx")
    mine = os.path.join(td, "mine.xlsx")
    theirs = os.path.join(td, "theirs.xlsx")
    for p in (base, mine, theirs):
        _make_xlsx(p)

    import sow_merge_tool as mod

    app = mod.SowMergeApp(mine, theirs, merge_mode=True, base_path=base)
    try:
        _pump(app.root, 1.2)
        assert app.bottom.winfo_ismapped() == 1, "Expected bottom sheet nav frame to be mapped in 3-way mode"
        assert app.nav_canvas.winfo_ismapped() == 1, "Expected bottom sheet nav canvas to be mapped in 3-way mode"
        root_top = app.root.winfo_rooty()
        root_h = app.root.winfo_height()
        bottom_top = app.bottom.winfo_rooty()
        bottom_h = app.bottom.winfo_height()
        bottom_end = bottom_top - root_top + bottom_h
        assert bottom_end <= root_h, (
            f"Expected bottom sheet nav within root bounds; root_h={root_h} bottom_end={bottom_end}"
        )
    finally:
        try:
            app.root.destroy()
        except Exception:
            pass


def main():
    _run_2way()
    _run_3way_bottom_nav_visible()
    print("GUI_SELF_TEST_BOTTOM_BAR_ALIGNED_OK")


if __name__ == "__main__":
    main()

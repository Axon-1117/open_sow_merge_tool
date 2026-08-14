"""Independent acceptance tests for OpenSpec section 10.

This module deliberately drives public GUI entry points and never saves an
input workbook.  It covers the centered difference navigator, the compact rich
structural status, Global Mode atomicity, the Sheet-wide equal-width model, and
Excel-style C-area headers.

Run:
  python _gui_self_test_openspec_section10.py
"""

from __future__ import annotations

import copy
import hashlib
import os
import time
import traceback
from contextlib import contextmanager
from dataclasses import replace

import tkinter as tk
import tkinter.font as tkfont
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

import sow_merge_tool as smt
from _gui_self_test_latest_gunships_feedback import _real_gunships_app
from _test_temp_utils import make_temp_dir


DEVELOP_ROOT = r"C:\GM15\design\sheets\develop"
RELEASE_ROOT = r"C:\GM15\design\sheets\release"
REAL_REPLAY_CANDIDATES = (
    "ActivityCWPuzzle纵横拼图.xlsx",
    "ActivityInfinite幸运无限礼包.xlsx",
)


def _pump(root, seconds: float = 0.08) -> None:
    deadline = time.time() + seconds
    while time.time() < deadline:
        root.update_idletasks()
        root.update()
        time.sleep(0.005)


def _wait_until(root, predicate, message: str, timeout: float = 20.0) -> None:
    deadline = time.time() + timeout
    last_error = None
    while time.time() < deadline:
        _pump(root, 0.03)
        try:
            if predicate():
                return
        except Exception as exc:  # pragma: no cover - reported below
            last_error = exc
    suffix = f"; last error={last_error!r}" if last_error else ""
    raise AssertionError(f"{message}{suffix}")


def _write_book(path: str, rows, *, sheet: str = "Data") -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet
    for row in rows:
        worksheet.append(list(row))
    workbook.save(path)
    workbook.close()


@contextmanager
def _quiet_dialogs():
    names = ("showerror", "showwarning", "showinfo", "askyesno", "askokcancel")
    originals = {
        name: getattr(smt.messagebox, name)
        for name in names
        if hasattr(smt.messagebox, name)
    }
    events = []

    def _record(name, result):
        def _dialog(*args, **kwargs):
            events.append((name, args, kwargs))
            return result

        return _dialog

    try:
        for name in originals:
            setattr(
                smt.messagebox,
                name,
                _record(name, name in ("askyesno", "askokcancel")),
            )
        yield events
    finally:
        for name, original in originals.items():
            setattr(smt.messagebox, name, original)


@contextmanager
def _synthetic_view(
    mine_rows,
    theirs_rows,
    *,
    base_rows=None,
    conflict_map=None,
    conflict_mode: bool = False,
    geometry: str = "1100x780",
):
    root_dir = make_temp_dir("sow_openspec_section10_")
    mine_path = os.path.join(root_dir, "mine.xlsx")
    theirs_path = os.path.join(root_dir, "theirs.xlsx")
    base_path = os.path.join(root_dir, "base.xlsx")
    merged_path = os.path.join(root_dir, "merged.xlsx")
    _write_book(mine_path, mine_rows)
    _write_book(theirs_path, theirs_rows)
    if base_rows is not None:
        _write_book(base_path, base_rows)

    app = None
    with _quiet_dialogs() as dialogs:
        try:
            app = smt.SowMergeApp(
                mine_path,
                theirs_path,
                merge_mode=base_rows is not None,
                merged_path=merged_path if base_rows is not None else None,
                base_path=base_path if base_rows is not None else None,
                merge_conflict_cells_by_sheet=copy.deepcopy(conflict_map or {}),
                merge_conflict_mode=bool(conflict_mode),
                raw_mine=mine_path if base_rows is not None else None,
                raw_base=base_path if base_rows is not None else None,
                raw_theirs=theirs_path if base_rows is not None else None,
            )
            app.root.deiconify()
            app.root.state("normal")
            app.root.geometry(geometry)
            app.nb.select(app._sheet_containers["Data"])
            _wait_until(
                app.root,
                lambda: (
                    app.sheet_views.get("Data") is not None
                    and bool(getattr(app.sheet_views["Data"], "_data_ready", False))
                ),
                "Data view did not load",
            )
            view = app.sheet_views["Data"]
            view._suppress_bg_apply = True
            view.only_diff_var.set(0)
            view._last_only_diff_value = 0
            view.refresh(row_only=None, rescan=True)
            _wait_until(
                app.root,
                lambda: view._derive_lifecycle_state() == "READY",
                "Data view did not reach READY",
            )
            yield app, view, dialogs
        finally:
            if app is not None:
                app._shutdown_root()


@contextmanager
def _existing_view(path_a: str, path_b: str, sheet: str):
    app = None
    with _quiet_dialogs() as dialogs:
        try:
            app = smt.SowMergeApp(path_a, path_b)
            app.root.deiconify()
            app.root.state("normal")
            app.root.geometry("1100x780")
            app.nb.select(app._sheet_containers[sheet])
            _wait_until(
                app.root,
                lambda: (
                    app.sheet_views.get(sheet) is not None
                    and bool(getattr(app.sheet_views[sheet], "_data_ready", False))
                ),
                f"real replay Sheet did not load: {sheet}",
                timeout=45.0,
            )
            view = app.sheet_views[sheet]
            view._suppress_bg_apply = True
            view.only_diff_var.set(0)
            view._last_only_diff_value = 0
            view.refresh(row_only=None, rescan=True)
            _wait_until(
                app.root,
                lambda: view._derive_lifecycle_state() == "READY",
                f"real replay Sheet did not reach READY: {sheet}",
                timeout=45.0,
            )
            yield app, view, dialogs
        finally:
            if app is not None:
                app._shutdown_root()


def _worksheet_state(worksheet):
    return {
        "max_row": int(worksheet.max_row),
        "max_column": int(worksheet.max_column),
        "merged": tuple(sorted(str(item) for item in worksheet.merged_cells.ranges)),
        "cells": tuple(
            tuple(
                (
                    worksheet.cell(row=row, column=column).value,
                    worksheet.cell(row=row, column=column).data_type,
                    worksheet.cell(row=row, column=column).style_id,
                    worksheet.cell(row=row, column=column).number_format,
                )
                for column in range(1, worksheet.max_column + 1)
            )
            for row in range(1, worksheet.max_row + 1)
        ),
    }


def _app_mutation_state(app, sheet: str = "Data"):
    workbooks = {
        "a_edit": _worksheet_state(app.ws_a_edit(sheet)),
        "a_val": _worksheet_state(app.ws_a_val(sheet)),
        "b_edit": _worksheet_state(app.ws_b_edit(sheet)),
        "b_val": _worksheet_state(app.ws_b_val(sheet)),
    }
    if getattr(app, "has_base", False):
        workbooks["base_edit"] = _worksheet_state(app.ws_base_edit(sheet))
        workbooks["base_val"] = _worksheet_state(app.ws_base_val(sheet))
    names = (
        "manual_a_cell_ops",
        "manual_b_cell_ops",
        "manual_a_formula_cache_ops",
        "manual_b_formula_cache_ops",
        "manual_a_row_ops",
        "manual_b_row_ops",
        "manual_a_column_ops",
        "manual_b_column_ops",
        "manual_sheet_ops",
        "undo_stack",
        "modified_a",
        "modified_b",
        "modified_sheets_a",
        "modified_sheets_b",
        "merge_conflict_cells_by_sheet",
        "user_touched_conflicts",
    )
    return {
        "workbooks": workbooks,
        "state": {
            name: copy.deepcopy(getattr(app, name))
            for name in names
        },
    }


def _sheet_values(worksheet):
    return tuple(
        tuple(
            worksheet.cell(row=row, column=column).value
            for column in range(1, worksheet.max_column + 1)
        )
        for row in range(1, worksheet.max_row + 1)
    )


def _menu_labels(menu) -> tuple[str, ...]:
    end = menu.index("end")
    if end is None:
        return ()
    return tuple(str(menu.entrycget(index, "label")) for index in range(int(end) + 1))


def _assert_global_menu(view) -> None:
    for menu in (view._use_left_menu, view._use_right_menu):
        labels = _menu_labels(menu)
        assert any("全局" in label or "Global" in label for label in labels), labels
    view._set_copy_scope_mode("global")
    assert view._copy_scope_mode == "global"
    assert view._copy_scope_var.get() == "global"
    assert "全局" in str(view.use_left_btn.cget("text"))
    assert "全局" in str(view.use_right_btn.cget("text"))


def _set_only_diff(view, enabled: bool) -> None:
    target = int(bool(enabled))
    if int(view.only_diff_var.get()) == target:
        return
    view.only_diff_var.set(target)
    view._toggle_only_diff()
    _wait_until(
        view.root,
        lambda: (
            not bool(getattr(view, "_mode_switch_pending", False))
            and view._derive_lifecycle_state() == "READY"
            and int(view.only_diff_var.get()) == target
        ),
        f"only-diff did not settle at {target}",
    )


def _widget_rect(widget) -> tuple[int, int, int, int]:
    left = int(widget.winfo_rootx())
    top = int(widget.winfo_rooty())
    return (
        left,
        top,
        left + int(widget.winfo_width()),
        top + int(widget.winfo_height()),
    )


def _nearest_common_ancestor(left_widget, right_widget):
    left_ancestors = []
    current = left_widget
    while current is not None:
        left_ancestors.append(current)
        current = getattr(current, "master", None)
    right_ancestors = set()
    current = right_widget
    while current is not None:
        right_ancestors.add(current)
        current = getattr(current, "master", None)
    return next(
        ancestor
        for ancestor in left_ancestors
        if ancestor in right_ancestors
    )


def _descendants(widget):
    for child in widget.winfo_children():
        yield child
        yield from _descendants(child)


def _widget_text(widget) -> str:
    if isinstance(widget, tk.Text):
        return str(widget.get("1.0", "end-1c"))
    try:
        return str(widget.cget("text"))
    except Exception:
        pass
    try:
        variable_name = str(widget.cget("textvariable"))
        return str(widget.getvar(variable_name)) if variable_name else ""
    except Exception:
        return ""


def _is_red(widget, color: str) -> bool:
    if not color:
        return False
    red, green, blue = widget.winfo_rgb(color)
    return red > 36000 and red > green * 1.35 and red > blue * 1.25


def _font_is_bold(widget, font_spec) -> bool:
    if not font_spec:
        return False
    try:
        return str(tkfont.Font(root=widget, font=font_spec).actual("weight")) == "bold"
    except Exception:
        return "bold" in str(font_spec).lower()


def _rich_status_text(container) -> str:
    if isinstance(container, tk.Text):
        return container.get("1.0", "end-1c")
    leaves = [
        child
        for child in _descendants(container)
        if bool(child.winfo_ismapped()) and not child.winfo_children()
    ]
    if not leaves:
        leaves = [container]
    leaves.sort(key=lambda item: _widget_rect(item)[0])
    return "".join(_widget_text(item) for item in leaves)


def _assert_rich_t_status_immediately_before_buttons(view) -> None:
    expected = "待处理 T 已自动选｜可执行"
    bar = view.column_action_bar
    candidate_names = (
        "column_action_rich_status",
        "column_action_selection_group",
        "column_action_status_group",
        "column_action_status_label",
    )
    candidates = [
        getattr(view, name)
        for name in candidate_names
        if getattr(view, name, None) is not None
    ]
    candidates.extend(
        child
        for child in bar.winfo_children()
        if child not in candidates
    )
    candidates = [
        item
        for item in candidates
        if bool(item.winfo_ismapped()) and _rich_status_text(item).strip() == expected
    ]
    assert candidates, (
        "the complete actionable status must be one compact rich group; "
        f"status={view.column_action_status_var.get()!r}, "
        f"selection={getattr(view, 'column_action_selection_var', None) and view.column_action_selection_var.get()!r}"
    )
    status = candidates[0]
    status_rect = _widget_rect(status)
    buttons_rect = _widget_rect(view.column_action_button_group)
    assert status_rect[2] <= buttons_rect[0], (status_rect, buttons_rect)
    assert 0 <= buttons_rect[0] - status_rect[2] <= 12, (status_rect, buttons_rect)
    assert status_rect[2] - status_rect[0] <= int(status.winfo_reqwidth()) + 8, (
        status_rect,
        status.winfo_reqwidth(),
    )

    token = "T"
    if isinstance(status, tk.Text):
        text = status.get("1.0", "end-1c")
        start = text.index(token)
        token_tags = None
        for offset in range(start, start + len(token)):
            tags = set(status.tag_names(f"1.{offset}"))
            token_tags = tags if token_tags is None else token_tags & tags
        rich_tags = [
            tag
            for tag in (token_tags or ())
            if _is_red(status, str(status.tag_cget(tag, "foreground")))
            and _font_is_bold(status, status.tag_cget(tag, "font"))
        ]
        assert rich_tags, (text, token_tags)
        for offset in tuple(range(0, start)) + tuple(range(start + len(token), len(text))):
            assert not set(rich_tags) & set(status.tag_names(f"1.{offset}")), (
                text,
                offset,
                rich_tags,
            )
        return

    leaves = [
        child
        for child in _descendants(status)
        if bool(child.winfo_ismapped()) and not child.winfo_children()
    ]
    if not leaves:
        leaves = [status]
    token_widgets = [item for item in leaves if _widget_text(item).strip() == token]
    assert len(token_widgets) == 1, [
        (_widget_text(item), item.winfo_class()) for item in leaves
    ]
    token_widget = token_widgets[0]
    assert _is_red(token_widget, str(token_widget.cget("foreground"))), token_widget.cget(
        "foreground"
    )
    assert _font_is_bold(token_widget, token_widget.cget("font")), token_widget.cget(
        "font"
    )
    for item in leaves:
        if item is token_widget:
            continue
        try:
            assert not _is_red(item, str(item.cget("foreground"))), (
                _widget_text(item),
                item.cget("foreground"),
            )
        except tk.TclError:
            pass


def test_centered_navigation_and_compact_rich_t_status_at_1450_and_1024():
    with _real_gunships_app() as (app, view, _analysis):
        assert view.selected_column_logical_range == (14, 14)
        result = view._apply_selected_column_block("BASE", "A")
        assert result.logical_start == result.logical_end == 14
        _wait_until(
            app.root,
            lambda: (
                view._derive_lifecycle_state() == "READY"
                and view.selected_column_logical_range == (20, 20)
            ),
            "real Gunships did not advance from L14 to L20",
            timeout=35.0,
        )
        assert view.column_action_status_var.get() == "待处理 T 已自动选｜可执行"

        for geometry in ("1450x860", "1024x760"):
            app.root.geometry(geometry)
            _pump(app.root, 0.2)
            nav_group = _widget_rect(view.diff_nav_group)
            column_group = _widget_rect(view.column_action_button_group)
            status_group = _widget_rect(view.column_action_status_group)
            structural_group = (
                min(status_group[0], column_group[0]),
                min(status_group[1], column_group[1]),
                max(status_group[2], column_group[2]),
                max(status_group[3], column_group[3]),
            )
            action_row_widget = _nearest_common_ancestor(
                view.diff_nav_group,
                view.column_action_button_group,
            )
            assert action_row_widget not in (view.frame, app.root)
            action_row = _widget_rect(action_row_widget)
            assert (
                action_row[0] <= nav_group[0] < nav_group[2] <= action_row[2]
            ), (
                geometry,
                action_row,
                nav_group,
            )
            assert (
                nav_group[3] >= structural_group[1]
                and structural_group[3] >= nav_group[1]
            ), (
                geometry,
                nav_group,
                structural_group,
            )
            assert nav_group[2] + 6 <= structural_group[0], (
                geometry,
                nav_group,
                structural_group,
            )
            centered_left = (
                action_row[0]
                + (action_row[2] - action_row[0] - (nav_group[2] - nav_group[0]))
                // 2
            )
            centered_right = centered_left + (nav_group[2] - nav_group[0])
            if centered_right + 6 <= structural_group[0]:
                assert abs(
                    (nav_group[0] + nav_group[2])
                    - (action_row[0] + action_row[2])
                ) <= 4, (geometry, action_row, nav_group)
            _assert_rich_t_status_immediately_before_buttons(view)


def _ordinary_global_rows():
    mine = [
        ("ID", "Name", "Count", "Note"),
        (1, "alpha", 10, "same"),
        (2, "mine-two", 20, "same"),
        (3, "same", 30, "mine-three"),
        (4, "same", 40, "same"),
        (5, "mine-five", 50, "mine-tail"),
    ]
    theirs = [list(row) for row in mine]
    theirs[2][1] = "theirs-two"
    theirs[3][3] = "theirs-three"
    theirs[5][1] = "theirs-five"
    theirs[5][3] = "theirs-tail"
    return mine, theirs


def test_global_mode_both_menus_full_sheet_only_diff_independent_and_one_undo():
    mine, theirs = _ordinary_global_rows()
    applied_states = []
    for only_diff in (False, True):
        with _synthetic_view(mine, theirs) as (app, view, _dialogs):
            _assert_global_menu(view)
            assert view._left_copy_direction == "A2B"
            assert view._right_copy_direction == "B2A"
            _set_only_diff(view, only_diff)
            if only_diff:
                assert len(view.display_rows) < len(view.row_pairs), (
                    view.display_rows,
                    view.row_pairs,
                )
            before = _app_mutation_state(app)
            undo_before = len(app.undo_stack)
            expected = _sheet_values(app.ws_b_edit("Data"))
            view._run_copy_action_by_mode(view._right_copy_direction)
            _wait_until(
                app.root,
                lambda: _sheet_values(app.ws_a_edit("Data")) == expected,
                "Global B2A did not apply all safe Sheet differences",
            )
            applied_states.append(_sheet_values(app.ws_a_edit("Data")))
            assert len(app.undo_stack) == undo_before + 1, app.undo_stack
            view._undo_last_action()
            _wait_until(
                app.root,
                lambda: _app_mutation_state(app) == before,
                "one undo did not restore the exact pre-Global state",
            )
    assert applied_states[0] == applied_states[1]


def _three_way_rows():
    base = [
        ("ID", "Value", "Note"),
        (1, "base-one", "base-note-one"),
        (2, "base-two", "base-note-two"),
        (3, "base-three", "base-note-three"),
    ]
    mine = [list(row) for row in base]
    theirs = [list(row) for row in base]
    mine[1][1] = "mine-one"
    mine[2][2] = "mine-note-two"
    mine[3][1] = "mine-three"
    theirs[1][1] = "theirs-one"
    theirs[2][2] = "theirs-note-two"
    theirs[3][1] = "theirs-three"
    return mine, base, theirs


def test_three_way_global_left_is_base_and_right_is_theirs_not_mine():
    mine, base, theirs = _three_way_rows()
    with _synthetic_view(mine, theirs, base_rows=base) as (app, view, _dialogs):
        _assert_global_menu(view)
        assert view._left_copy_direction == "BASE2A"
        assert view._right_copy_direction == "B2A"
        original = _app_mutation_state(app)
        base_values = _sheet_values(app.ws_base_edit("Data"))
        theirs_values = _sheet_values(app.ws_b_edit("Data"))
        assert base_values != _sheet_values(app.ws_a_edit("Data"))
        assert theirs_values != _sheet_values(app.ws_a_edit("Data"))

        undo_before = len(app.undo_stack)
        view._run_copy_action_by_mode(view._left_copy_direction)
        _wait_until(
            app.root,
            lambda: _sheet_values(app.ws_a_edit("Data")) == base_values,
            "3-way left Global used Mine instead of Base",
        )
        assert len(app.undo_stack) == undo_before + 1
        view._undo_last_action()
        _wait_until(app.root, lambda: _app_mutation_state(app) == original, "BASE2A undo failed")

        view._set_copy_scope_mode("global")
        view._run_copy_action_by_mode(view._right_copy_direction)
        _wait_until(
            app.root,
            lambda: _sheet_values(app.ws_a_edit("Data")) == theirs_values,
            "3-way right Global did not use Theirs",
        )
        assert len(app.undo_stack) == undo_before + 1
        view._undo_last_action()
        _wait_until(app.root, lambda: _app_mutation_state(app) == original, "B2A undo failed")


def test_global_merge_conflict_mode_apply_undo_and_failure_restore_metadata():
    mine, base, theirs = _three_way_rows()
    conflicts = {"Data": {2: {2}, 3: {3}, 4: {2}}}
    with _synthetic_view(
        mine,
        theirs,
        base_rows=base,
        conflict_map=conflicts,
        conflict_mode=True,
    ) as (app, view, _dialogs):
        _assert_global_menu(view)
        assert app.merge_conflict_mode is True
        original = _app_mutation_state(app)
        assert app.user_touched_conflicts is False

        for direction, expected_ws in (
            ("BASE2A", app.ws_base_edit("Data")),
            ("B2A", app.ws_b_edit("Data")),
        ):
            expected = _sheet_values(expected_ws)
            view._set_copy_scope_mode("global")
            view._run_copy_action_by_mode(direction)
            _wait_until(
                app.root,
                lambda: _sheet_values(app.ws_a_edit("Data")) == expected,
                f"conflict-mode Global {direction} was rejected or incomplete",
            )
            assert app.merge_conflict_cells_by_sheet == {}, (
                direction,
                app.merge_conflict_cells_by_sheet,
            )
            assert app.user_touched_conflicts is True
            assert len(app.undo_stack) == len(original["state"]["undo_stack"]) + 1
            view._undo_last_action()
            _wait_until(
                app.root,
                lambda: _app_mutation_state(app) == original,
                f"conflict-mode {direction} undo did not restore values/map/touched flag",
            )

        original_copy_row = view._copy_selected_row
        calls = []

        def _fail_after_first_committed_row(*args, **kwargs):
            result = original_copy_row(*args, **kwargs)
            if result and kwargs.get("suppress_refresh"):
                calls.append(kwargs.get("override_pair_idx"))
                raise RuntimeError("section10 injected conflict-mode failure")
            return result

        view._copy_selected_row = _fail_after_first_committed_row
        try:
            view._set_copy_scope_mode("global")
            view._run_copy_action_by_mode("B2A")
            _pump(app.root, 0.15)
        finally:
            view._copy_selected_row = original_copy_row
        assert calls, "failure injection did not reach the Global commit loop"
        assert _app_mutation_state(app) == original


def _assert_zero_write(view, app, direction: str = "B2A"):
    before = _app_mutation_state(app)
    view._set_copy_scope_mode("global")
    original_chooser = getattr(view, "_ask_global_ambiguous_action", None)
    view._ask_global_ambiguous_action = lambda _direction, _details: "cancel"
    try:
        view._run_copy_action_by_mode(direction)
        _pump(app.root, 0.12)
    finally:
        if original_chooser is not None:
            view._ask_global_ambiguous_action = original_chooser
    assert _app_mutation_state(app) == before


def test_global_structural_ambiguity_stale_and_mid_commit_failure_are_zero_write():
    mine, theirs = _ordinary_global_rows()

    structural_theirs = [
        ("ID", "Inserted", "Name", "Count", "Note"),
        (1, "x", "alpha", 10, "same"),
        (2, "x", "theirs-two", 20, "same"),
        (3, "x", "same", 30, "theirs-three"),
        (4, "x", "same", 40, "same"),
        (5, "x", "theirs-five", 50, "theirs-tail"),
    ]
    with _synthetic_view(mine, structural_theirs) as (app, view, _dialogs):
        assert view._active_column_comparison_cache().structural_diff_cols
        _assert_zero_write(view, app)

    with _synthetic_view(mine, theirs) as (app, view, _dialogs):
        cache = view._active_column_comparison_cache()
        view.column_comparison_cache = replace(cache, unresolved_cols=frozenset({2}))
        assert view._active_column_comparison_cache().unresolved_cols == frozenset({2})
        _assert_zero_write(view, app)

    with _synthetic_view(mine, theirs) as (app, view, _dialogs):
        view._mark_column_mapping_stale(
            "section10-explicit-stale",
            edited_sides=("A",),
        )
        assert not view._column_mapping_is_current()
        _assert_zero_write(view, app)

    with _synthetic_view(mine, theirs) as (app, view, _dialogs):
        before = _app_mutation_state(app)
        original_copy_row = view._copy_selected_row
        calls = []

        def _fail_after_first_committed_row(*args, **kwargs):
            result = original_copy_row(*args, **kwargs)
            if result and kwargs.get("suppress_refresh"):
                calls.append(kwargs.get("override_pair_idx"))
                raise RuntimeError("section10 injected mid-commit failure")
            return result

        view._copy_selected_row = _fail_after_first_committed_row
        try:
            view._set_copy_scope_mode("global")
            view._run_copy_action_by_mode("B2A")
            _pump(app.root, 0.15)
        finally:
            view._copy_selected_row = original_copy_row
        assert calls, "failure injection did not reach the Global commit loop"
        assert _app_mutation_state(app) == before


def test_global_many_small_conflict_blocks_refreshes_once_and_stays_atomic():
    row_count = 700
    base = [("ID", "Value")]
    mine = [("ID", "Value")]
    theirs = [("ID", "Value")]
    conflict_rows = {}
    for index in range(1, row_count + 1):
        base_value = f"base-{index}"
        if index % 2:
            mine_value = f"mine-{index}"
            theirs_value = f"theirs-{index}"
            conflict_rows[index + 1] = {2}
        else:
            mine_value = theirs_value = base_value
        base.append((index, base_value))
        mine.append((index, mine_value))
        theirs.append((index, theirs_value))
    expected_blocks = (row_count + 1) // 2
    assert len(conflict_rows) == expected_blocks

    with _synthetic_view(
        mine,
        theirs,
        base_rows=base,
        conflict_map={"Data": conflict_rows},
        conflict_mode=True,
        geometry="1100x780",
    ) as (app, view, _dialogs):
        _assert_global_menu(view)
        before = _app_mutation_state(app)
        expected = _sheet_values(app.ws_b_edit("Data"))
        undo_before = len(app.undo_stack)
        refresh_calls = []
        original_refresh = view.refresh

        def _counted_refresh(*args, **kwargs):
            refresh_calls.append((args, dict(kwargs)))
            return original_refresh(*args, **kwargs)

        view.refresh = _counted_refresh
        started = time.perf_counter()
        try:
            view._run_copy_action_by_mode("B2A")
            elapsed = time.perf_counter() - started
        finally:
            view.refresh = original_refresh

        assert _sheet_values(app.ws_a_edit("Data")) == expected
        assert app.merge_conflict_cells_by_sheet == {}
        assert app.user_touched_conflicts is True
        assert len(app.undo_stack) == undo_before + 1, app.undo_stack
        assert len(refresh_calls) <= 3, (
            "Global must publish in one bounded batch instead of refreshing "
            f"per small block: calls={len(refresh_calls)}"
        )
        assert elapsed < 20.0, (
            f"700-row/{expected_blocks}-block Global apply took {elapsed:.3f}s"
        )
        print(
            "GLOBAL_MANY_BLOCKS "
            f"rows={row_count} blocks={expected_blocks} "
            f"refresh_calls={len(refresh_calls)} elapsed_sec={elapsed:.3f}"
        )

        view._undo_last_action()
        _wait_until(
            app.root,
            lambda: _app_mutation_state(app) == before,
            "many-block Global one-step undo did not restore the exact state",
            timeout=30.0,
        )


def _wide_rows(prefix: str, *, changed=None):
    headers = tuple(get_column_letter(index) for index in range(1, 28))
    rows = [headers]
    for row in range(1, 6):
        values = [f"{prefix}-{row}-{header}" for header in headers]
        if row == 1:
            values[26] = f"{prefix}-AA-a-content-value-longer-than-neighbors"
        rows.append(values)
    if changed is not None:
        row, column, value = changed
        rows[row][column - 1] = value
    return rows


def _line_span_widths(view, line: str):
    spans = view._spans_for_line(line)
    return tuple(end - start for start, end in spans.values()), spans


def _assert_all_rendered_widths_equal(view) -> int:
    slot_count = view._active_column_projection().slot_count
    widths = dict(view.col_char_widths)
    assert tuple(widths) == tuple(range(1, slot_count + 1)), widths
    assert len(set(widths.values())) == 1, widths
    uniform = next(iter(widths.values()))

    main_widgets = [view.left, view.right]
    header_widgets = [view.left_colhdr, view.right_colhdr, view.cursor_cmp_colhdr]
    if view._is_three_way_enabled():
        main_widgets.insert(1, view.base)
        header_widgets.insert(1, view.base_colhdr)
    for widget in main_widgets:
        for line_no in range(1, len(view.display_rows) + 1):
            line = widget.get(f"{line_no}.0", f"{line_no}.end")
            span_widths, spans = _line_span_widths(view, line)
            assert len(span_widths) == slot_count, (widget, line_no, spans)
            assert set(span_widths) == {uniform}, (widget, line_no, span_widths)
            assert max(end for _start, end in spans.values()) <= len(line)
    for widget in header_widgets:
        line = widget.get("1.0", "1.end")
        span_widths, spans = _line_span_widths(view, line)
        assert len(span_widths) == slot_count
        assert set(span_widths) == {uniform}, (widget, span_widths)
        assert max(end for _start, end in spans.values()) <= len(line)
    for line_no in range(1, 4):
        line = view.cursor_cmp.get(f"{line_no}.0", f"{line_no}.end")
        if not line:
            continue
        span_widths, spans = _line_span_widths(view, line)
        assert len(span_widths) == slot_count
        assert set(span_widths) == {uniform}, (line_no, span_widths)
        assert max(end for _start, end in spans.values()) <= len(line)
    return uniform


def _header_fragments(view, widget):
    line = widget.get("1.0", "1.end")
    return [
        line[start:end].strip()
        for start, end in view._spans_for_line(line).values()
    ]


def test_sheet_wide_uniform_width_is_stable_and_c_headers_are_a_through_aa():
    base = _wide_rows("base")
    mine = _wide_rows("base", changed=(3, 4, "mine-D3"))
    theirs = _wide_rows("base", changed=(5, 26, "theirs-Z5"))
    with _synthetic_view(mine, theirs, base_rows=base) as (app, view, _dialogs):
        assert view._active_column_projection().slot_count == 27
        first_diff_pair = next(
            pair_idx
            for pair_idx in range(len(view.row_pairs))
            if view._all_logical_diff_cols_for_pair(pair_idx)
        )
        view._select_line(view.row_to_line[first_diff_pair])
        view._update_cursor_lines()
        _pump(app.root, 0.08)

        expected = [get_column_letter(index) for index in range(1, 28)]
        assert _header_fragments(view, view.left_colhdr) == expected
        assert _header_fragments(view, view.base_colhdr) == expected
        assert _header_fragments(view, view.right_colhdr) == expected
        assert _header_fragments(view, view.cursor_cmp_colhdr) == expected
        assert expected[-2:] == ["Z", "AA"]
        assert "L27" not in view.cursor_cmp_colhdr.get("1.0", "1.end")

        _assert_all_rendered_widths_equal(view)
        width_state = (
            dict(view.col_char_widths),
            tuple(view._logical_column_widths),
        )

        view._sync_main_x_to_frac(0.73)
        view._sync_c_x_to_frac(0.73)
        _pump(app.root, 0.05)
        assert (dict(view.col_char_widths), tuple(view._logical_column_widths)) == width_state

        last_pair = view.display_rows[-1]
        view._select_line(view.row_to_line[last_pair])
        view._update_cursor_lines()
        _pump(app.root, 0.05)
        assert (dict(view.col_char_widths), tuple(view._logical_column_widths)) == width_state

        _set_only_diff(view, True)
        _assert_all_rendered_widths_equal(view)
        assert (dict(view.col_char_widths), tuple(view._logical_column_widths)) == width_state
        _set_only_diff(view, False)
        _assert_all_rendered_widths_equal(view)
        assert (dict(view.col_char_widths), tuple(view._logical_column_widths)) == width_state

        generation = view._column_projection_generation
        view._mark_column_mapping_stale(
            "section10-width-projection-rebuild",
            edited_sides=("A",),
        )
        view.refresh(row_only=None, rescan=False)
        _wait_until(
            app.root,
            lambda: (
                view._column_mapping_is_current()
                and view._column_projection_generation > generation
            ),
            "projection change did not rebuild the uniform width model",
        )
        _assert_all_rendered_widths_equal(view)
        assert dict(view.col_char_widths) == width_state[0]


def _sha256(path: str) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _choose_real_cell_only_replay():
    for name in REAL_REPLAY_CANDIDATES:
        develop_path = os.path.join(DEVELOP_ROOT, name)
        release_path = os.path.join(RELEASE_ROOT, name)
        if not os.path.isfile(develop_path) or not os.path.isfile(release_path):
            continue
        develop = load_workbook(develop_path, read_only=True, data_only=False)
        release = load_workbook(release_path, read_only=True, data_only=False)
        try:
            for sheet in develop.sheetnames:
                if sheet not in release.sheetnames:
                    continue
                ws_develop = develop[sheet]
                ws_release = release[sheet]
                if (
                    ws_develop.max_row != ws_release.max_row
                    or ws_develop.max_column != ws_release.max_column
                ):
                    continue
                diffs = []
                for row in range(1, ws_develop.max_row + 1):
                    for column in range(1, ws_develop.max_column + 1):
                        left = ws_release.cell(row=row, column=column).value
                        right = ws_develop.cell(row=row, column=column).value
                        if left != right:
                            diffs.append((row, column, left, right))
                if diffs:
                    return develop_path, release_path, sheet, tuple(diffs)
        finally:
            develop.close()
            release.close()
    raise AssertionError(
        "no cell-only develop/release replay candidate is available: "
        f"{REAL_REPLAY_CANDIDATES!r}"
    )


def test_real_develop_release_global_replay_is_read_only_and_hash_guarded():
    develop_path, release_path, sheet, diffs = _choose_real_cell_only_replay()
    hashes_before = {
        develop_path: _sha256(develop_path),
        release_path: _sha256(release_path),
    }
    try:
        with _existing_view(release_path, develop_path, sheet) as (app, view, _dialogs):
            _assert_global_menu(view)
            before = _app_mutation_state(app, sheet)
            expected = _sheet_values(app.ws_b_edit(sheet))
            assert any(
                app.ws_a_edit(sheet).cell(row=row, column=column).value != right
                for row, column, _left, right in diffs
            )
            view._run_copy_action_by_mode("B2A")
            _wait_until(
                app.root,
                lambda: _sheet_values(app.ws_a_edit(sheet)) == expected,
                f"real develop/release Global replay was incomplete: {sheet}",
                timeout=30.0,
            )
            assert len(app.undo_stack) == len(before["state"]["undo_stack"]) + 1
            view._undo_last_action()
            _wait_until(
                app.root,
                lambda: _app_mutation_state(app, sheet) == before,
                "real replay one-step undo did not restore the in-memory candidate",
            )
    finally:
        hashes_after = {
            develop_path: _sha256(develop_path),
            release_path: _sha256(release_path),
        }
        assert hashes_after == hashes_before, (hashes_before, hashes_after)


def main() -> None:
    tests = (
        test_centered_navigation_and_compact_rich_t_status_at_1450_and_1024,
        test_global_mode_both_menus_full_sheet_only_diff_independent_and_one_undo,
        test_three_way_global_left_is_base_and_right_is_theirs_not_mine,
        test_global_merge_conflict_mode_apply_undo_and_failure_restore_metadata,
        test_global_structural_ambiguity_stale_and_mid_commit_failure_are_zero_write,
        test_global_many_small_conflict_blocks_refreshes_once_and_stays_atomic,
        test_sheet_wide_uniform_width_is_stable_and_c_headers_are_a_through_aa,
        test_real_develop_release_global_replay_is_read_only_and_hash_guarded,
    )
    failures = []
    for test in tests:
        try:
            test()
            print(f"PASS {test.__name__}")
        except Exception:
            failures.append(test.__name__)
            traceback.print_exc()
    if failures:
        raise SystemExit(f"OPENSPEC_SECTION10_FAILED: {failures}")
    print(f"PASS: OpenSpec section 10 acceptance ({len(tests)} tests)")


if __name__ == "__main__":
    main()

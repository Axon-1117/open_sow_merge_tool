"""Demo test: export a human-readable report for missing-sheet blank views.

Run:
  .venv\\Scripts\\python.exe _demo_test_missing_sheet_view.py

Outputs:
  A temp folder containing markdown reports and, when available, screenshots.
"""

import os
import subprocess
import time

from openpyxl import Workbook

import sow_merge_tool as mod
from _test_temp_utils import make_temp_dir


def _make_book(path: str, sheets: list[tuple[str, list[list[object]]]]):
    wb = Workbook()
    ws = wb.active
    ws.title = sheets[0][0]
    for r_idx, row in enumerate(sheets[0][1], start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx).value = value
    for title, rows in sheets[1:]:
        ws2 = wb.create_sheet(title)
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, value in enumerate(row, start=1):
                ws2.cell(row=r_idx, column=c_idx).value = value
    wb.save(path)
    wb.close()


def _open_sheet(app, sheet: str):
    app.nb.select(app._sheet_containers[sheet])
    for _ in range(60):
        app.root.update_idletasks()
        app.root.update()
        time.sleep(0.02)
    view = app.sheet_views.get(sheet)
    if view is None:
        raise RuntimeError(f"view missing for {sheet}")
    view.refresh(row_only=None, rescan=True)
    app.root.update_idletasks()
    app.root.update()
    return view


def _grab_text(widget, max_lines: int = 8) -> str:
    try:
        text = widget.get("1.0", "end-1c")
    except Exception:
        return "<unavailable>"
    lines = text.splitlines()
    if len(lines) > max_lines:
        lines = lines[:max_lines] + ["..."]
    return "\n".join(lines)


def _try_screenshot(app, out_path: str) -> str | None:
    try:
        from PIL import ImageGrab  # type: ignore
    except Exception:
        ImageGrab = None
    try:
        app.root.update_idletasks()
        app.root.update()
        try:
            app.root.lift()
            app.root.attributes("-topmost", True)
            app.root.update()
        except Exception:
            pass
        time.sleep(0.15)
        x = app.root.winfo_rootx()
        y = app.root.winfo_rooty()
        w = app.root.winfo_width()
        h = app.root.winfo_height()
        if w <= 0 or h <= 0:
            return None
        if ImageGrab is not None:
            try:
                img = ImageGrab.grab(bbox=(x, y, x + w, y + h))
                img.save(out_path)
                return out_path
            except Exception:
                pass
        save_path_ps = out_path.replace("'", "''")
        ps = (
            "Add-Type -AssemblyName System.Drawing;"
            f"$x={int(x)};$y={int(y)};$w={int(w)};$h={int(h)};"
            "$bmp=New-Object System.Drawing.Bitmap $w,$h;"
            "$g=[System.Drawing.Graphics]::FromImage($bmp);"
            "$g.CopyFromScreen($x,$y,0,0,$bmp.Size);"
            f"$bmp.Save('{save_path_ps}', [System.Drawing.Imaging.ImageFormat]::Png);"
            "$g.Dispose();"
            "$bmp.Dispose();"
        )
        r = subprocess.run(
            ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", ps],
            capture_output=True,
            text=True,
            timeout=30,
        )
        if r.returncode == 0 and os.path.exists(out_path):
            return out_path
        return None
    except Exception:
        return None
    finally:
        try:
            app.root.attributes("-topmost", False)
        except Exception:
            pass


def _write_case_report(app, view, out_md: str, title: str):
    tabs = [app.nb.tab(tab_id, "text") for tab_id in app.nb.tabs()]
    meta = app.get_sheet_meta(view.sheet)
    screenshot_path = _try_screenshot(app, os.path.splitext(out_md)[0] + ".png")

    lines = [
        f"# {title}",
        "",
        "## Tabs",
        "",
        f"`{tabs}`",
        "",
        "## Sheet Meta",
        "",
        f"`{meta}`",
        "",
        "## Pane Titles",
        "",
        f"- left: `{view.left_title.cget('text')}`",
        f"- base: `{view.mid_title.cget('text')}`",
        f"- right: `{view.right_title.cget('text')}`",
        "",
        "## Path Labels",
        "",
        f"- left/path A: `{view.path_label_a.cget('text')}`",
        f"- base/path: `{view.path_label_base.cget('text')}`",
        f"- right/path B: `{view.path_label_b.cget('text')}`",
        "",
        "## Left Pane Snapshot",
        "",
        "```text",
        _grab_text(view.left),
        "```",
        "",
        "## Base Pane Snapshot",
        "",
        "```text",
        _grab_text(view.base),
        "```",
        "",
        "## Right Pane Snapshot",
        "",
        "```text",
        _grab_text(view.right),
        "```",
        "",
        "## Row Headers",
        "",
        "```text",
        "\n".join([
            "left:",
            _grab_text(view.left_ln),
            "",
            "base:",
            _grab_text(view.base_ln),
            "",
            "right:",
            _grab_text(view.right_ln),
        ]),
        "```",
        "",
        "## Screenshot",
        "",
        f"- available: `{bool(screenshot_path)}`",
        f"- path: `{screenshot_path or 'not captured'}`",
        "",
    ]
    with open(out_md, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def _run_2way_case(root: str) -> str:
    a = os.path.join(root, "2way_a.xlsx")
    b = os.path.join(root, "2way_b.xlsx")
    _make_book(a, [
        ("Common", [["id", "value"], ["A", "same"]]),
        ("OnlyA", [["id", "value"], ["A1", "left only"], ["A2", "still left only"]]),
    ])
    _make_book(b, [
        ("Common", [["id", "value"], ["A", "same"]]),
    ])
    app = mod.SowMergeApp(a, b)
    try:
        view = _open_sheet(app, "OnlyA")
        assert view._is_missing_sheet_view()
        out_md = os.path.join(root, "demo_missing_sheet_2way.md")
        _write_case_report(app, view, out_md, "2-way Missing Sheet Blank View")
        return out_md
    finally:
        try:
            app._shutdown_root()
        except Exception:
            pass


def _run_3way_case(root: str) -> str:
    base = os.path.join(root, "3way_base.xlsx")
    mine = os.path.join(root, "3way_mine.xlsx")
    theirs = os.path.join(root, "3way_theirs.xlsx")
    merged = os.path.join(root, "3way_merged.xlsx")
    _make_book(base, [
        ("Common", [["id", "value"], ["A", "base"]]),
        ("BaseOnly", [["id", "value"], ["B1", "from base"], ["B2", "base row 2"]]),
    ])
    _make_book(mine, [
        ("Common", [["id", "value"], ["A", "mine"]]),
    ])
    _make_book(theirs, [
        ("Common", [["id", "value"], ["A", "theirs"]]),
        ("Added", [["id", "value"], ["T1", "from theirs"], ["T2", "theirs row 2"]]),
    ])
    app = mod.SowMergeApp(mine, theirs, merge_mode=True, merged_path=merged, base_path=base)
    try:
        view = _open_sheet(app, "Added")
        assert view._is_missing_sheet_view()
        out_md = os.path.join(root, "demo_missing_sheet_3way_theirs_only.md")
        _write_case_report(app, view, out_md, "3-way Missing Sheet Blank View (Theirs Only)")

        view2 = _open_sheet(app, "BaseOnly")
        assert view2._is_missing_sheet_view()
        out_md2 = os.path.join(root, "demo_missing_sheet_3way_base_only.md")
        _write_case_report(app, view2, out_md2, "3-way Missing Sheet Blank View (Base Only)")
        return out_md + "\n" + out_md2
    finally:
        try:
            app._shutdown_root()
        except Exception:
            pass


def main():
    root = make_temp_dir("sow_demo_missing_sheet_")
    out_2way = _run_2way_case(root)
    out_3way = _run_3way_case(root)
    print("DEMO_MISSING_SHEET_VIEW_OK")
    print(out_2way)
    print(out_3way)


if __name__ == "__main__":
    main()

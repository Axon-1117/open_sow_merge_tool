#!/usr/bin/env python3
"""Reproducible, non-destructive column-structure performance baseline.

The three project workbooks are opened read-only for cache/signature/mapping
measurements.  Structural variants used by the mapping benchmark exist only as
Python row-cache tuples, so retained formula text is never rewritten by
openpyxl.  The action/save phase works exclusively on disposable copies under
``C:\\tmp`` and is a throughput fixture, not a fidelity/correctness oracle.

Run from the repository root, for example::

    python benchmarks/column_structure_baseline.py --repeats 3

Each measured case/phase/repeat is launched in a fresh Python process.  The
driver writes a self-contained JSON report under ``C:\\tmp``.
"""

from __future__ import annotations

import argparse
from collections import Counter
from datetime import datetime
import gc
import hashlib
import json
import math
import os
from pathlib import Path
import platform
import shutil
import statistics
import subprocess
import sys
import threading
import time

import openpyxl
import psutil
from openpyxl import Workbook, load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))
DEFAULT_REAL_ROOT = Path(r"C:\GM15\design\sheets\develop")
DEFAULT_OUTPUT_DIR = Path(r"C:\tmp\column_structure_baseline_v2")
REAL_CASES = {
    "Guide": ("Guide.xlsx", "TGuideStep@design"),
    "Skill": ("Skill.xlsx", "SkillTimeline@design"),
    "Dungeon": ("Dungeon.xlsx", "Dungeon@design"),
}
SYNTHETIC_ROWS = 4000
SYNTHETIC_COLS = 96
DEFAULT_SCALE_WIDTHS = (68, 128, 255, 256, 257, 512, 513)


def _require_tmp_child(path: Path) -> Path:
    candidate = path.resolve()
    temp_root = Path(r"C:\tmp").resolve()
    if candidate == temp_root:
        raise ValueError("output directory must be a child of C:\\tmp, not C:\\tmp itself")
    try:
        candidate.relative_to(temp_root)
    except ValueError as exc:
        raise ValueError(f"all copies and outputs must stay under C:\\tmp: {candidate}") from exc
    return candidate


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


class _PeakMemory:
    """Poll fresh-process resident memory without distorting timed Python work."""

    def __init__(self, interval_seconds: float = 0.005):
        self._process = psutil.Process()
        self._interval = interval_seconds
        self._stop = threading.Event()
        self._thread = threading.Thread(target=self._poll, daemon=True)
        self.rss_start = 0
        self.rss_peak = 0

    def _poll(self):
        while not self._stop.is_set():
            try:
                self.rss_peak = max(self.rss_peak, self._process.memory_info().rss)
            except psutil.Error:
                pass
            self._stop.wait(self._interval)

    def __enter__(self):
        gc.collect()
        self.rss_start = self._process.memory_info().rss
        self.rss_peak = self.rss_start
        self._thread.start()
        return self

    def __exit__(self, exc_type, exc, traceback):
        self._stop.set()
        self._thread.join(timeout=1.0)
        try:
            self.rss_peak = max(self.rss_peak, self._process.memory_info().rss)
        except psutil.Error:
            pass

    def as_dict(self) -> dict:
        mib = 1024 * 1024
        return {
            "rss_start_mb": self.rss_start / mib,
            "rss_peak_mb": self.rss_peak / mib,
            "rss_delta_mb": max(0, self.rss_peak - self.rss_start) / mib,
            "rss_poll_interval_ms": self._interval * 1000,
        }


def _make_synthetic_cache(rows: int = SYNTHETIC_ROWS, cols: int = SYNTHETIC_COLS):
    value_rows = []
    edit_rows = []
    for row_idx in range(1, rows + 1):
        value_row = []
        edit_row = []
        for col_idx in range(1, cols + 1):
            if row_idx == 1:
                value = f"COL_{col_idx:03d}"
                edit = value
            elif col_idx % 7 == 0:
                value = row_idx * 1000 + col_idx
                edit = f"=R{row_idx}C{col_idx}+{col_idx}"
            else:
                value = (row_idx * 131 + col_idx * 17) % 100003
                edit = value
            value_row.append(value)
            edit_row.append(edit)
        value_rows.append(tuple(value_row))
        edit_rows.append(tuple(edit_row))
    return tuple(value_rows), tuple(edit_rows), rows, cols


def _read_real_cache(path: Path, sheet_name: str):
    started = time.perf_counter()
    wb_value = load_workbook(path, data_only=True, read_only=True, keep_links=True)
    wb_edit = load_workbook(path, data_only=False, read_only=True, keep_links=True)
    open_ms = (time.perf_counter() - started) * 1000
    try:
        if sheet_name not in wb_value.sheetnames or sheet_name not in wb_edit.sheetnames:
            raise KeyError(f"worksheet {sheet_name!r} not present in {path}")
        ws_value = wb_value[sheet_name]
        ws_edit = wb_edit[sheet_name]
        max_row = max(ws_value.max_row, ws_edit.max_row)
        max_col = max(ws_value.max_column, ws_edit.max_column)
        started = time.perf_counter()
        value_rows = tuple(
            tuple(row)
            for row in ws_value.iter_rows(
                min_row=1,
                max_row=max_row,
                min_col=1,
                max_col=max_col,
                values_only=True,
            )
        )
        edit_rows = tuple(
            tuple(row)
            for row in ws_edit.iter_rows(
                min_row=1,
                max_row=max_row,
                min_col=1,
                max_col=max_col,
                values_only=True,
            )
        )
        cache_ms = (time.perf_counter() - started) * 1000
    finally:
        wb_value.close()
        wb_edit.close()
    return value_rows, edit_rows, max_row, max_col, open_ms, cache_ms


def _variant_positions(width: int) -> tuple[int, int]:
    insert_at = max(2, width // 3)
    delete_original = min(width, max(insert_at + 3, (width * 2) // 3))
    return insert_at, delete_original


def _make_in_memory_variant(value_rows, edit_rows, width: int):
    """Insert two and delete one column without parsing or rewriting formulas."""
    insert_at, delete_original = _variant_positions(width)
    prefix_end = insert_at - 1
    deleted_offset = delete_original - 1
    right_values = []
    right_edits = []
    formula_payload_preserved = True
    for row_idx, (value_row, edit_row) in enumerate(zip(value_rows, edit_rows), start=1):
        value_row = tuple(value_row)
        edit_row = tuple(edit_row)
        new_values = (f"BENCH_INSERT_A_{row_idx}", f"BENCH_INSERT_B_{row_idx}")
        right_value = (
            value_row[:prefix_end]
            + new_values
            + value_row[prefix_end:deleted_offset]
            + value_row[deleted_offset + 1 :]
        )
        right_edit = (
            edit_row[:prefix_end]
            + new_values
            + edit_row[prefix_end:deleted_offset]
            + edit_row[deleted_offset + 1 :]
        )
        expected_retained = edit_row[:deleted_offset] + edit_row[deleted_offset + 1 :]
        observed_retained = right_edit[:prefix_end] + right_edit[prefix_end + 2 :]
        formula_payload_preserved &= expected_retained == observed_retained
        right_values.append(right_value)
        right_edits.append(right_edit)
    return (
        tuple(right_values),
        tuple(right_edits),
        width + 1,
        {
            "insert_at_1based": insert_at,
            "inserted_columns": 2,
            "deleted_original_1based": delete_original,
            "deleted_columns": 1,
            "retained_edit_payload_exact": formula_payload_preserved,
            "construction": "in-memory tuple splice; no workbook write or formula rewrite",
        },
    )


def _load_case_cache(case: str, real_root: Path):
    if case == "synthetic":
        started = time.perf_counter()
        values, edits, rows, cols = _make_synthetic_cache()
        cache_ms = (time.perf_counter() - started) * 1000
        return values, edits, rows, cols, 0.0, cache_ms, "synthetic_formula_cache"
    filename, sheet_name = REAL_CASES[case]
    values, edits, rows, cols, open_ms, cache_ms = _read_real_cache(
        real_root / filename, sheet_name
    )
    return values, edits, rows, cols, open_ms, cache_ms, sheet_name


def _build_and_align(case: str, values, edits, rows: int, cols: int):
    from sow_merge_tool import (
        ColumnModelCacheKey,
        align_column_signatures_2way,
        build_column_signature_snapshot,
    )

    variant_started = time.perf_counter()
    right_values, right_edits, right_cols, variant_info = _make_in_memory_variant(
        values, edits, cols
    )
    variant_ms = (time.perf_counter() - variant_started) * 1000
    left_key = ColumnModelCacheKey(case, 1, 1, mine_edit_version=0)
    right_key = ColumnModelCacheKey(case, 1, 2, theirs_edit_version=1)
    started = time.perf_counter()
    left_snapshot = build_column_signature_snapshot(
        left_key, values, edits, max_col=cols
    )
    left_signature_ms = (time.perf_counter() - started) * 1000
    started = time.perf_counter()
    right_snapshot = build_column_signature_snapshot(
        right_key, right_values, right_edits, max_col=right_cols
    )
    right_signature_ms = (time.perf_counter() - started) * 1000
    started = time.perf_counter()
    result = align_column_signatures_2way(left_snapshot, right_snapshot)
    mapping_ms = (time.perf_counter() - started) * 1000
    state_counts = Counter(slot.state for slot in result.model.slots)
    block_counts = Counter(block.state for block in result.model.blocks)
    return {
        "timing_ms": {
            "variant_cache_construction": variant_ms,
            "left_signature_construction": left_signature_ms,
            "right_signature_construction": right_signature_ms,
            "signature_construction_total": left_signature_ms + right_signature_ms,
            "mapping_2way_current_1_3_api": mapping_ms,
        },
        "shape": {
            "rows": rows,
            "left_columns": cols,
            "right_columns": right_cols,
        },
        "variant": variant_info,
        "mapping": {
            "slot_states": dict(sorted(state_counts.items())),
            "block_states": dict(sorted(block_counts.items())),
            "slots": len(result.model.slots),
            "blocks": len(result.model.blocks),
            "anchors": len(result.anchor_pairs),
            "unresolved_slots": len(result.fallback_slot_indices),
            "used_physical_fallback": result.used_physical_fallback,
            "fallback_reason": result.fallback_reason,
        },
    }


def _measure_cold(case: str, real_root: Path) -> dict:
    total_started = time.perf_counter()
    with _PeakMemory() as memory:
        values, edits, rows, cols, open_ms, cache_ms, sheet = _load_case_cache(
            case, real_root
        )
        measured = _build_and_align(case, values, edits, rows, cols)
    measured["timing_ms"]["workbook_read_only_open"] = open_ms
    measured["timing_ms"]["sequential_row_cache_construction"] = cache_ms
    measured["timing_ms"]["cold_first_ready_proxy_total"] = (
        time.perf_counter() - total_started
    ) * 1000
    measured["memory"] = memory.as_dict()
    measured["sheet"] = sheet
    measured["scope"] = (
        "fresh-worker proxy: elapsed timing starts after Python, benchmark-script, "
        "openpyxl, and psutil startup; it includes the first sow_merge_tool import, "
        "read-only workbook open, sequential row-cache construction, in-memory "
        "structural variant, two signatures, and current 1.3 2-way mapping"
    )
    return measured


def _measure_signature_mapping(case: str, real_root: Path) -> dict:
    values, edits, rows, cols, _open_ms, _cache_ms, sheet = _load_case_cache(
        case, real_root
    )
    gc.collect()
    with _PeakMemory() as memory:
        measured = _build_and_align(case, values, edits, rows, cols)
    measured["memory"] = memory.as_dict()
    measured["sheet"] = sheet
    measured["scope"] = (
        "fresh process with populated sequential row caches; measures in-memory "
        "variant construction, signature construction, and current 1.3 mapping"
    )
    return measured


def _create_synthetic_workbook(path: Path):
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Synthetic")
    for row_idx in range(1, SYNTHETIC_ROWS + 1):
        row = []
        for col_idx in range(1, SYNTHETIC_COLS + 1):
            if row_idx == 1:
                row.append(f"COL_{col_idx:03d}")
            elif col_idx % 7 == 0:
                row.append(f"=ROW()*1000+COLUMN()+{col_idx}")
            else:
                row.append((row_idx * 131 + col_idx * 17) % 100003)
        ws.append(row)
    wb.save(path)
    wb.close()


def _measure_action_save(
    case: str,
    repeat: int,
    real_root: Path,
    output_dir: Path,
) -> dict:
    work_dir = output_dir / "work" / f"{case}_{repeat}_{os.getpid()}"
    work_dir.mkdir(parents=True, exist_ok=False)
    source_copy = work_dir / "source.xlsx"
    saved_copy = work_dir / "saved.xlsx"
    if case == "synthetic":
        _create_synthetic_workbook(source_copy)
        sheet_name = "Synthetic"
    else:
        filename, sheet_name = REAL_CASES[case]
        shutil.copy2(real_root / filename, source_copy)
    try:
        with _PeakMemory() as memory:
            started = time.perf_counter()
            wb = load_workbook(source_copy, data_only=False, read_only=False, keep_links=True)
            load_ms = (time.perf_counter() - started) * 1000
            try:
                ws = wb[sheet_name]
                rows = ws.max_row
                cols = ws.max_column
                insert_at, delete_original = _variant_positions(cols)
                started = time.perf_counter()
                ws.insert_cols(insert_at, amount=2)
                for row_idx in range(1, rows + 1):
                    ws.cell(row_idx, insert_at, f"BENCH_INSERT_A_{row_idx}")
                    ws.cell(row_idx, insert_at + 1, f"BENCH_INSERT_B_{row_idx}")
                ws.delete_cols(delete_original + 2, amount=1)
                action_ms = (time.perf_counter() - started) * 1000
                started = time.perf_counter()
                wb.save(saved_copy)
                save_ms = (time.perf_counter() - started) * 1000
            finally:
                wb.close()
        return {
            "sheet": sheet_name,
            "shape": {
                "rows": rows,
                "left_columns": cols,
                "right_columns": cols + 1,
            },
            "timing_ms": {
                "writable_copy_open": load_ms,
                "column_action_insert_2_delete_1": action_ms,
                "openpyxl_save": save_ms,
            },
            "memory": memory.as_dict(),
            "scope": (
                "disposable C:\\tmp copy only; openpyxl insert/delete/save throughput "
                "fixture, not a formula-cache, formula-reference, OOXML fidelity, or "
                "mapping-correctness golden"
            ),
            "saved_size_bytes": saved_copy.stat().st_size,
        }
    finally:
        shutil.rmtree(work_dir, ignore_errors=True)


def _measure_mapping_scale(width: int) -> dict:
    from sow_merge_tool import (
        ColumnModelCacheKey,
        align_column_signatures_2way,
        build_column_signature_snapshot,
    )

    left_rows = tuple(
        tuple(f"W{col_idx:04d}_R{row_idx:02d}" for col_idx in range(1, width + 1))
        for row_idx in range(1, 13)
    )
    right_rows = left_rows[:-1] + (
        tuple(f"W{col_idx:04d}_RIGHT" for col_idx in range(1, width + 1)),
    )
    key_left = ColumnModelCacheKey(f"scale_{width}", 1, 1)
    key_right = ColumnModelCacheKey(f"scale_{width}", 1, 2, theirs_edit_version=1)
    left = build_column_signature_snapshot(
        key_left, left_rows, left_rows, max_col=width
    )
    right = build_column_signature_snapshot(
        key_right, right_rows, right_rows, max_col=width
    )
    configured_limit = int(align_column_signatures_2way.__kwdefaults__["max_columns"])
    gc.collect()
    with _PeakMemory(interval_seconds=0.001) as memory:
        started = time.perf_counter()
        result = align_column_signatures_2way(left, right)
        mapping_ms = (time.perf_counter() - started) * 1000
    return {
        "shape": {
            "rows": len(left_rows),
            "left_columns": width,
            "right_columns": width,
        },
        "timing_ms": {"mapping_2way_current_1_3_api": mapping_ms},
        "memory": memory.as_dict(),
        "mapping": {
            "configured_max_columns": configured_limit,
            "anchors": len(result.anchor_pairs),
            "unresolved_slots": len(result.fallback_slot_indices),
            "used_physical_fallback": result.used_physical_fallback,
            "fallback_reason": result.fallback_reason,
        },
        "scope": (
            "prebuilt actual ColumnSignatureSnapshot inputs with no exact intrinsic-key "
            "pairs but a unique high-similarity same-column candidate; fresh-process "
            f"mapping-only width curve around max_columns={configured_limit}"
        ),
    }


def _worker(args) -> int:
    output_dir = _require_tmp_child(Path(args.output_dir))
    output_dir.mkdir(parents=True, exist_ok=True)
    real_root = Path(args.real_root).resolve()
    if args.phase == "cold_first_ready":
        payload = _measure_cold(args.case, real_root)
    elif args.phase == "signature_mapping":
        payload = _measure_signature_mapping(args.case, real_root)
    elif args.phase == "action_save":
        payload = _measure_action_save(
            args.case, args.repeat, real_root, output_dir
        )
    elif args.phase == "mapping_scale":
        payload = _measure_mapping_scale(args.width)
    else:
        raise ValueError(args.phase)
    payload.update({
        "case": args.case,
        "phase": args.phase,
        "repeat": args.repeat,
        "implementation_sha256": _sha256(REPO_ROOT / "sow_merge_tool.py"),
        "script_sha256": _sha256(Path(__file__).resolve()),
    })
    print(json.dumps(payload, ensure_ascii=False, sort_keys=True))
    return 0


def _cpu_model() -> str:
    if os.name == "nt":
        try:
            import winreg

            with winreg.OpenKey(
                winreg.HKEY_LOCAL_MACHINE,
                r"HARDWARE\DESCRIPTION\System\CentralProcessor\0",
            ) as key:
                return str(winreg.QueryValueEx(key, "ProcessorNameString")[0]).strip()
        except OSError:
            pass
    return platform.processor().strip() or "unknown"


def _git_metadata() -> dict:
    def run(*arguments: str) -> str:
        completed = subprocess.run(
            ["git", *arguments],
            cwd=REPO_ROOT,
            check=True,
            capture_output=True,
            text=True,
        )
        return completed.stdout.strip()

    try:
        status = run("status", "--porcelain")
        return {
            "head": run("rev-parse", "HEAD"),
            "branch": run("branch", "--show-current"),
            "dirty": bool(status),
            "porcelain_sha256": hashlib.sha256(status.encode("utf-8")).hexdigest(),
        }
    except (OSError, subprocess.CalledProcessError) as exc:
        return {"error": str(exc)}


def _environment() -> dict:
    implementation_path = REPO_ROOT / "sow_merge_tool.py"
    return {
        "captured_at": datetime.now().astimezone().isoformat(),
        "timezone": str(datetime.now().astimezone().tzinfo),
        "python_version": sys.version,
        "python_version_info": list(sys.version_info),
        "python_executable": sys.executable,
        "python_implementation": platform.python_implementation(),
        "architecture": platform.architecture(),
        "platform": platform.platform(),
        "system": platform.system(),
        "release": platform.release(),
        "os_version": platform.version(),
        "machine": platform.machine(),
        "cpu_model": _cpu_model(),
        "cpu_logical_count": psutil.cpu_count(logical=True),
        "cpu_physical_count": psutil.cpu_count(logical=False),
        "total_memory_bytes": psutil.virtual_memory().total,
        "total_memory_gib": psutil.virtual_memory().total / (1024**3),
        "openpyxl_version": openpyxl.__version__,
        "psutil_version": psutil.__version__,
        "working_directory": str(Path.cwd().resolve()),
        "script": str(Path(__file__).resolve()),
        "script_sha256": _sha256(Path(__file__).resolve()),
        "implementation_file": str(implementation_path),
        "implementation_sha256": _sha256(implementation_path),
        "git": _git_metadata(),
    }


def _source_metadata(real_root: Path) -> dict:
    result = {}
    for case, (filename, sheet) in REAL_CASES.items():
        path = real_root / filename
        stat = path.stat()
        result[case] = {
            "path": str(path),
            "sheet": sheet,
            "size_bytes": stat.st_size,
            "mtime": datetime.fromtimestamp(stat.st_mtime).astimezone().isoformat(),
            "sha256": _sha256(path),
            "access": "read-only; never copied back or saved in place",
        }
    return result


def _nearest_rank(values: list[float], percentile: float) -> float:
    ordered = sorted(values)
    return ordered[max(0, math.ceil(percentile * len(ordered)) - 1)]


def _flatten_metrics(payload: dict, prefix: str = ""):
    for key, value in payload.items():
        name = f"{prefix}.{key}" if prefix else key
        if isinstance(value, dict):
            yield from _flatten_metrics(value, name)
        elif isinstance(value, (int, float)) and not isinstance(value, bool):
            yield name, float(value)


def _aggregate(raw: list[dict]) -> dict:
    grouped: dict[tuple[str, str], list[dict]] = {}
    for record in raw:
        grouped.setdefault((record["case"], record["phase"]), []).append(record)
    aggregate = {}
    allowed_prefixes = ("timing_ms.", "memory.")
    for (case, phase), records in sorted(grouped.items()):
        metrics: dict[str, list[float]] = {}
        for record in records:
            for name, value in _flatten_metrics(record):
                if name.startswith(allowed_prefixes):
                    metrics.setdefault(name, []).append(value)
        aggregate.setdefault(case, {})[phase] = {
            name: {
                "n": len(values),
                "min": min(values),
                "median": statistics.median(values),
                "p95_nearest_rank": _nearest_rank(values, 0.95),
                "max": max(values),
            }
            for name, values in sorted(metrics.items())
        }
    return aggregate


def _run_worker(
    case: str,
    phase: str,
    repeat: int,
    real_root: Path,
    output_dir: Path,
    *,
    width: int = 0,
) -> dict:
    command = [
        sys.executable,
        str(Path(__file__).resolve()),
        "_worker",
        "--case",
        case,
        "--phase",
        phase,
        "--repeat",
        str(repeat),
        "--real-root",
        str(real_root),
        "--output-dir",
        str(output_dir),
    ]
    if width:
        command.extend(("--width", str(width)))
    environment = os.environ.copy()
    environment["PYTHONHASHSEED"] = "0"
    completed = subprocess.run(
        command,
        cwd=REPO_ROOT,
        env=environment,
        check=False,
        capture_output=True,
        text=True,
    )
    if completed.returncode:
        raise RuntimeError(
            "worker failed with exit code "
            f"{completed.returncode}: {' '.join(command)}\n"
            f"stdout:\n{completed.stdout}\nstderr:\n{completed.stderr}"
        )
    lines = [line for line in completed.stdout.splitlines() if line.strip()]
    if not lines:
        raise RuntimeError(f"worker produced no JSON: {' '.join(command)}")
    return json.loads(lines[-1])


def _driver(args) -> int:
    if args.repeats < 1:
        raise ValueError("--repeats must be positive")
    output_dir = _require_tmp_child(Path(args.output_dir))
    output_dir.mkdir(parents=True, exist_ok=True)
    real_root = Path(args.real_root).resolve()
    missing = [
        str(real_root / filename)
        for filename, _sheet in REAL_CASES.values()
        if not (real_root / filename).is_file()
    ]
    if missing:
        raise FileNotFoundError(f"missing real workbook(s): {missing}")
    scale_widths = tuple(
        int(item.strip()) for item in args.scale_widths.split(",") if item.strip()
    )
    from sow_merge_tool import align_column_signatures_2way

    mapping_default_max_columns = int(
        align_column_signatures_2way.__kwdefaults__["max_columns"]
    )
    raw = []
    cases = ("synthetic", *REAL_CASES.keys())
    phases = ("cold_first_ready", "signature_mapping", "action_save")
    for case in cases:
        for phase in phases:
            for repeat in range(1, args.repeats + 1):
                print(f"[{case}] {phase} repeat {repeat}/{args.repeats}", flush=True)
                raw.append(
                    _run_worker(
                        case, phase, repeat, real_root, output_dir
                    )
                )
    for width in scale_widths:
        case = f"scale_{width}"
        for repeat in range(1, args.repeats + 1):
            print(f"[{case}] mapping_scale repeat {repeat}/{args.repeats}", flush=True)
            raw.append(
                _run_worker(
                    case,
                    "mapping_scale",
                    repeat,
                    real_root,
                    output_dir,
                    width=width,
                )
            )
    command = (
        f'"{sys.executable}" "{Path(__file__).resolve()}" '
        f'--repeats {args.repeats} --real-root "{real_root}" '
        f'--output-dir "{output_dir}" --scale-widths "{args.scale_widths}"'
    )
    environment_metadata = _environment()
    worker_implementation_hashes = {
        record["implementation_sha256"] for record in raw
    }
    worker_script_hashes = {record["script_sha256"] for record in raw}
    if worker_implementation_hashes != {environment_metadata["implementation_sha256"]}:
        raise RuntimeError(
            "sow_merge_tool.py changed during the run; discard this mixed-code report"
        )
    if worker_script_hashes != {environment_metadata["script_sha256"]}:
        raise RuntimeError("benchmark script changed during the run")
    report = {
        "schema_version": 2,
        "environment": environment_metadata,
        "configuration": {
            "repeats": args.repeats,
            "fresh_process_per_case_phase_repeat": True,
            "python_hash_seed_for_workers": 0,
            "real_root": str(real_root),
            "output_dir": str(output_dir),
            "synthetic_rows": SYNTHETIC_ROWS,
            "synthetic_columns": SYNTHETIC_COLS,
            "scale_widths": list(scale_widths),
            "mapping_default_max_columns": mapping_default_max_columns,
            "percentile_method": "p95 nearest-rank; with n=3 this equals max",
            "reproduction_command": command,
        },
        "source_workbooks": _source_metadata(real_root),
        "methodology": {
            "mapping_fixture": (
                "source caches are transformed in memory by exact tuple splicing; "
                "retained formulas are not parsed, translated, recalculated, or saved"
            ),
            "cold_first_ready": (
                "proxy because GUI integration is not implemented in task 1.3; includes "
                "first sow_merge_tool import, read-only open/cache construction, "
                "signatures, and current 1.3 mapping. It excludes Python/benchmark "
                "script/openpyxl/psutil startup and GUI rendering"
            ),
            "action_save_limitation": (
                "openpyxl variants exist only under C:\\tmp and measure throughput. "
                "They are not correctness/fidelity goldens because formula cached values, "
                "references, and OOXML details may be rewritten or lost"
            ),
            "memory": (
                "fresh-process RSS sampled every 5 ms (1 ms for scale mapping); "
                "RSS delta is peak minus phase start. tracemalloc is intentionally not "
                "enabled because its instrumentation materially distorts these timings"
            ),
        },
        "raw": raw,
        "aggregates": _aggregate(raw),
    }
    report_path = output_dir / "baseline_results.json"
    report_path.write_text(
        json.dumps(report, ensure_ascii=False, indent=2, sort_keys=True),
        encoding="utf-8",
    )
    print(f"report: {report_path}")
    return 0


def _parse_args(argv=None):
    parser = argparse.ArgumentParser(description=__doc__)
    subparsers = parser.add_subparsers(dest="command")
    worker = subparsers.add_parser("_worker", help=argparse.SUPPRESS)
    worker.add_argument("--case", required=True)
    worker.add_argument("--phase", required=True)
    worker.add_argument("--repeat", required=True, type=int)
    worker.add_argument("--real-root", required=True)
    worker.add_argument("--output-dir", required=True)
    worker.add_argument("--width", type=int, default=0)
    parser.add_argument("--repeats", type=int, default=3)
    parser.add_argument("--real-root", default=str(DEFAULT_REAL_ROOT))
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR))
    parser.add_argument(
        "--scale-widths",
        default=",".join(str(width) for width in DEFAULT_SCALE_WIDTHS),
    )
    return parser.parse_args(argv)


def main(argv=None) -> int:
    args = _parse_args(argv)
    if args.command == "_worker":
        return _worker(args)
    return _driver(args)


if __name__ == "__main__":
    raise SystemExit(main())

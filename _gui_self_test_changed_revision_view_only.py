"""Deterministic view-only, quiet-window and edit-preload regression coverage.

This intentionally drives both a small only-diff result (the real Dungeon
case can be a few dozen rows) and a virtual result.  Once exact caches have
published, every observation path is guarded against value/edit worksheet
access while an explicit mutation request holds an injected eight-second
editable-loader delay in the background.
"""

import argparse
import os
import sys
import tempfile
import threading
import time
import hashlib
import json
from contextlib import contextmanager
from types import SimpleNamespace

from openpyxl import Workbook
from openpyxl.worksheet._read_only import ReadOnlyWorksheet
from openpyxl.worksheet.worksheet import Worksheet

import sow_merge_tool as sm


_MUTATION_RETRY_2WAY_CASE = "mutation-retry-2way"
_MUTATION_RETRY_3WAY_CASE = "mutation-retry-3way"
_SAVE_RETRY_2WAY_CASE = "save-retry-2way"
_SAVE_RETRY_3WAY_CASE = "save-retry-3way"
_QUIET_WINDOW_PREEMPT_2WAY_CASE = "quiet-window-preempt-2way"
_QUIET_WINDOW_PREEMPT_3WAY_CASE = "quiet-window-preempt-3way"
_HIDDEN_SNAPSHOT_TECHNICAL_FAILURE_CASE = "hidden-snapshot-technical-failure"

# Each selector is one fresh, bounded natural case.  Keep the default to one
# lightweight demand/retry case; a matrix runner must launch selectors in
# separate processes rather than reviving the old environment-driven aggregate.
_CASES = (
    _MUTATION_RETRY_2WAY_CASE,
    _MUTATION_RETRY_3WAY_CASE,
    _SAVE_RETRY_2WAY_CASE,
    _SAVE_RETRY_3WAY_CASE,
    _QUIET_WINDOW_PREEMPT_2WAY_CASE,
    _QUIET_WINDOW_PREEMPT_3WAY_CASE,
    _HIDDEN_SNAPSHOT_TECHNICAL_FAILURE_CASE,
)


def _make_book(path: str, side: str) -> None:
    wb = Workbook(write_only=True)
    for sheet, rows in (
        ("Dungeon", 1_650),
        ("MonsterGroup", 3_000),
        ("Archive", 360),
        # Large enough for the snapshot route but with only 18 changed rows:
        # this isolates the short-and-wide result surface from vertical
        # virtualization without turning the narrow MonsterGroup case wide.
        ("ShortWide", 250),
    ):
        ws = wb.create_sheet(sheet)
        is_wide = sheet in {"Dungeon", "ShortWide"}
        wide_headers = [f"wide_{col}" for col in range(4, 70)] if is_wide else []
        ws.append(["id@id", "value", "note", *wide_headers])
        for row in range(1, rows + 1):
            value = f"{sheet}-{row}"
            note = f"note-{row % 7}"
            wide_values = (
                [f"{sheet}-{row}-{col}" for col in range(4, 70)]
                if is_wide
                else []
            )
            if sheet == "Dungeon":
                # Force a virtual only-diff result while keeping row identity
                # exact and stable in every mode.
                if side == "theirs":
                    value = f"{value}-theirs"
                elif side == "base":
                    value = f"{value}-base"
                # A deterministic off-screen logical field proves that a
                # bounded column surface keeps physical/Base targets exact.
                if wide_values and side == "theirs":
                    wide_values[-1] = f"{wide_values[-1]}-theirs"
                elif wide_values and side == "base":
                    wide_values[-1] = f"{wide_values[-1]}-base"
            elif sheet == "ShortWide" and row <= 18 and side in ("theirs", "base"):
                value = f"{value}-{side}"
                if wide_values:
                    wide_values[-1] = f"{wide_values[-1]}-{side}"
            elif (
                sheet in {"MonsterGroup", "Archive"}
                and side in ("theirs", "base")
                and row <= 340
                and row % 17 == 0
            ):
                # The normal result path must remain non-virtual even when
                # the sheet itself is substantial.
                value = f"{value}-{side}"
            ws.append([row, value, note, *wide_values])
    wb.save(path)
    wb.close()


def _sha256(path: str) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _pump(root, seconds: float) -> None:
    deadline = time.monotonic() + seconds
    while time.monotonic() < deadline:
        root.update_idletasks()
        root.update()
        time.sleep(0.01)


def _wait(root, predicate, timeout: float, label: str) -> None:
    deadline = time.monotonic() + timeout
    while time.monotonic() < deadline:
        _pump(root, 0.04)
        if predicate():
            return
    raise AssertionError(f"timeout: {label}")


def _show_only_diff_from_exact_cache(view) -> None:
    assert view._cache_only_diff_rows_from_exact_pair_maps(), "exact only-diff cache unavailable"
    view.only_diff_var.set(1)
    view._last_only_diff_value = 1
    view.refresh(row_only=None, rescan=False)
    assert view._only_diff_rows_exact and view._has_valid_only_diff_snapshot_cache()


def _select_tab(app, name: str, *, timeout: float = 15.0):
    app.nb.select(app._sheet_containers[name])
    # The production handler is intentionally a build-ui closure bound to the
    # notebook event; pump Tk rather than reaching into an implementation-only
    # attribute.
    _pump(app.root, 0.05)
    _wait(
        app.root,
        lambda: (
            app.sheet_views.get(name) is not None
            and app.sheet_views[name]._data_ready
            and app._is_sheet_exact_current(name)
        ),
        float(timeout),
        f"{name} exact current",
    )
    return app.sheet_views[name]


def _assert_exact_operation_targets(view, *, three_way: bool) -> None:
    """Prove exact pair/physical mappings used by a later cell operation."""
    changed_pair = next(
        pair_idx for pair_idx, cols in view.pair_diff_cols.items() if 2 in cols
    )
    pair = view.row_pairs[changed_pair]
    assert view._row_for_side(pair, "A") is not None
    assert view._row_for_side(pair, "B") is not None
    assert view._physical_col_for_logical("A", 2) == 2
    assert view._physical_col_for_logical("B", 2) == 2
    assert view._action_physical_columns("B2A", 2) == (2, 2)
    assert view._prepared_value_for_logical_cell(changed_pair, "B", 2).endswith("-theirs")
    if three_way:
        base_pair = next(
            pair_idx
            for pair_idx, cols in view.pair_base_diff_cols.items()
            if 2 in cols and view._base_row_for_pair(pair_idx, view.row_pairs[pair_idx]) is not None
        )
        base_pair_rows = view.row_pairs[base_pair]
        assert view._base_row_for_pair(base_pair, base_pair_rows) is not None
        assert view._physical_col_for_logical("BASE", 2) == 2
        assert view._action_physical_columns("BASE2A", 2) == (2, 2)
        assert view._prepared_value_for_logical_cell(base_pair, "BASE", 2).endswith("-base")


def _assert_wide_column_window(app, view, *, three_way: bool) -> None:
    """Check first/middle/last logical columns without Text-local remapping."""
    assert view._wide_column_virtual_active(), "69-column Dungeon must use wide viewport"
    total = view._logical_slot_count()
    assert total == 69, total
    pair_idx = next(index for index, cols in view.pair_diff_cols.items() if 69 in cols)
    pair = view.row_pairs[pair_idx]
    excel_row = view._row_for_side(pair, "A") or view._row_for_side(pair, "B")
    assert excel_row is not None

    # Horizontal scrollbar/main controller and minimap both select windows in
    # complete logical coordinates.  Header, main, and C spans must carry the
    # same global keys even though their rendered strings are short.
    def _main_hscroll(*args):
        # Exercise the actual ttk.Scrollbar callback registered by production
        # UI construction, rather than calling the logical controller directly.
        command = str(view.hsb_left.cget("command"))
        assert command, "main horizontal scrollbar command missing"
        view.hsb_left.tk.call(command, *args)

    for label, fraction, expected_last in (
        # Start at a changing target; an initial first->first no-op cannot
        # dilute end-to-end latency evidence.
        ("last", 1.0, total),
        ("first", 0.0, None),
        ("middle", 0.50, None),
    ):
        completed_before = len(view._viewport_request_completed)
        _main_hscroll("moveto", str(fraction))
        _pump(app.root, 0.12)
        columns = view._rendered_logical_columns()
        assert 1 <= len(columns) <= sm._VIRTUAL_VIEWPORT_MAX_COLUMNS
        if label == "first":
            assert columns[0] == 1, columns
        if expected_last is not None:
            assert columns[-1] == expected_last, columns
        main_spans = view._spans_for_line(view.left.get("1.0", "1.end"))
        header_spans = view._spans_for_line(view.left_colhdr.get("1.0", "1.end"))
        c_spans = view._spans_for_line(view.cursor_cmp.get("1.0", "1.end"))
        assert tuple(main_spans) == columns, (label, main_spans, columns)
        assert tuple(header_spans) == columns, (label, header_spans, columns)
        assert tuple(c_spans) == columns, (label, c_spans, columns)
        completed_now = list(view._viewport_request_completed)[completed_before:]
        hthumb = [
            item for item in completed_now
            if item.get("reason") == "hthumb" and item.get("counted")
        ]
        assert hthumb, (label, completed_now)
        assert hthumb[-1].get("actual_rendered_columns") == columns, hthumb[-1]
        assert float(hthumb[-1].get("elapsed_ms", 9999.0)) <= 33.0, hthumb[-1]
        view._redraw_diff_viewport()
        coords = list(view.hdiff_left.coords("vpbox"))
        assert len(coords) == 4, (label, coords)
        expected_x = int(view._wide_column_scroll_fractions()[0] * view.hdiff_left.winfo_width())
        assert abs(int(coords[0]) - expected_x) <= 2, (label, coords, expected_x)

    # A bounded preflight mode isolates the actual Scrollbar command and the
    # single after(0) coalescer before the complete A/B GUI matrix runs. Each
    # sample is a changing logical column window; no initial edge no-op can
    # dilute request→complete evidence.
    stress_count = max(
        0,
        int(os.environ.get("SOW_CHANGED_REVISION_SHORTWIDE_HSCROLL_STRESS", "0") or 0),
    )
    if stress_count and str(getattr(view, "sheet", "")) == "ShortWide":
        completed_before = len(view._viewport_request_completed)
        fractions = ("1.0", "0.0", "0.50")
        for index in range(stress_count):
            _main_hscroll("moveto", fractions[index % len(fractions)])
            deadline = time.monotonic() + 1.0
            while time.monotonic() < deadline:
                app.root.update_idletasks()
                app.root.update()
                fresh = list(view._viewport_request_completed)[completed_before:]
                if len([item for item in fresh if item.get("reason") == "hthumb" and item.get("counted")]) >= index + 1:
                    break
            else:
                raise AssertionError(f"ShortWide hthumb stress request {index} did not complete")
        stress_records = [
            item
            for item in list(view._viewport_request_completed)[completed_before:]
            if item.get("reason") == "hthumb" and item.get("counted")
        ]
        assert len(stress_records) == stress_count, stress_records
        for item in stress_records:
            assert item.get("status") == "complete", item
            assert int(item["row_start"]) == int(item["actual_row_start"]), item
            assert int(item["column_start"]) == int(item["actual_column_start"]), item
            actual_columns = tuple(item.get("actual_rendered_columns") or ())
            assert actual_columns, item
            assert actual_columns[0] == int(item["actual_column_start"]) + 1, item
            assert actual_columns[-1] == min(
                total,
                int(item["actual_column_start"]) + sm._VIRTUAL_VIEWPORT_MAX_COLUMNS,
            ), item
        elapsed = [float(item["elapsed_ms"]) for item in stress_records]
        queue_wait = [float(item.get("queue_wait_ms", 0.0)) for item in stress_records]
        publish = [float(item.get("publish_ms", 0.0)) for item in stress_records]
        assert app._p95(elapsed) <= 33.0, stress_records
        assert max(elapsed) <= 66.0, stress_records
        assert {str(item.get("drain_trigger") or "") for item in stress_records} == {"after0"}
        print(
            "SHORTWIDE_HSCROLL_STRESS "
            f"n={len(stress_records)} p95={app._p95(elapsed):.3f} max={max(elapsed):.3f} "
            f"queue_p95={app._p95(queue_wait):.3f} queue_max={max(queue_wait):.3f} "
            f"publish_p95={app._p95(publish):.3f} publish_max={max(publish):.3f} "
            "trigger=after0"
        )

    view._on_hdiff_map_click(
        SimpleNamespace(x=max(1, view.hdiff_left.winfo_width()), y=1), "left"
    )
    _pump(app.root, 0.12)
    assert view._rendered_logical_columns()[-1] == total

    # Off-screen focus is a view-only operation: it must bring the complete
    # physical target into the window without a Worksheet read or a recycled
    # visible column index.  Exercise first/middle/last explicitly.
    for logical_col in (1, (total + 1) // 2, total):
        assert view.focus_logical_cell(excel_row, logical_col)
        columns = view._rendered_logical_columns()
        assert logical_col in columns, (logical_col, columns)
        assert logical_col in view._spans_for_line(view.left.get("1.0", "1.end"))
        assert logical_col in view._spans_for_line(view.left_colhdr.get("1.0", "1.end"))
        assert logical_col in view._spans_for_line(view.cursor_cmp.get("1.0", "1.end"))

    assert view._prepared_value_for_logical_cell(pair_idx, "B", 69).endswith("-theirs")
    assert view._action_physical_columns("B2A", 69) == (69, 69)
    if three_way:
        base_pair = next(index for index, cols in view.pair_base_diff_cols.items() if 69 in cols)
        assert view._prepared_value_for_logical_cell(base_pair, "BASE", 69).endswith("-base")
        assert view._action_physical_columns("BASE2A", 69) == (69, 69)


def _move_prepared_raw_to_hidden_pending(view):
    """Put a real exact raw cache back into the hidden staged-cache shape."""
    raw_a = dict(view.pair_raw_parts_a)
    raw_b = dict(view.pair_raw_parts_b)
    raw_base = dict(view.pair_raw_parts_base)
    widths = dict(view.col_char_widths)
    assert raw_a or raw_b, (raw_a, raw_b)
    view.pair_raw_parts_a = {}
    view.pair_raw_parts_b = {}
    view.pair_raw_parts_base = {}
    view.pair_text_a = {}
    view.pair_text_b = {}
    view.pair_text_base = {}
    view._prepared_text_lru.clear()
    view._stage_cached_pair_parts(raw_a, raw_b, raw_base, widths, replace=True)
    assert view._pending_pair_parts_cache is not None
    assert not view.pair_raw_parts_a and not view.pair_raw_parts_b
    return (len(raw_a), len(raw_b), len(raw_base))


def _assert_pending_tab_promotion_is_view_only(app, view, other) -> None:
    """Promote a genuine hidden-pending raw cache without worksheet access."""
    gateway_names = (
        "ws_a_val", "ws_b_val", "ws_base_val",
        "ws_a_edit", "ws_b_edit", "ws_base_edit", "_request_edit_preload",
    )
    originals = {}
    hits = []

    def _forbidden(name):
        def _raise(*_args, **_kwargs):
            hits.append(name)
            raise AssertionError(f"pending tab promotion accessed {name}")

        return _raise

    raw_before = _move_prepared_raw_to_hidden_pending(view)
    install_samples_before = len(view._snapshot_install_only_samples_ms)
    for name in gateway_names:
        if hasattr(app, name):
            originals[name] = getattr(app, name)
            setattr(app, name, _forbidden(name))
    try:
        # Model a complete cache that finished while hidden: raw fragments are
        # pending, no raw/text cache is installed, and the notebook callback
        # must transfer only raw evidence before it starts the bounded surface.
        view._pending_exact_render = True
        app.nb.select(app._sheet_containers[other.sheet])
        _pump(app.root, 0.06)
        app.nb.select(app._sheet_containers[view.sheet])
        _wait(
            app.root,
            lambda: (
                not bool(view._pending_exact_render)
                and app._is_sheet_exact_current(view.sheet)
                and bool(view.display_rows)
            ),
            8,
            "pending exact tab promotion",
        )
    finally:
        for name, original in originals.items():
            setattr(app, name, original)
    assert not hits, hits
    assert raw_before == (
        len(view.pair_raw_parts_a),
        len(view.pair_raw_parts_b),
        len(view.pair_raw_parts_base),
    )
    assert view._pending_pair_parts_cache is None
    assert len(view._snapshot_install_only_samples_ms) == install_samples_before + 1
    assert view._snapshot_install_only_samples_ms[-1] >= 0.0
    # A short wide result is vertically complete, while the text cache remains
    # bounded to its actual viewport rows rather than eagerly formatting raw.
    assert len(view.pair_text_a) <= sm._VIRTUAL_VIEWPORT_MAX_ROWS
    assert len(view.pair_text_b) <= sm._VIRTUAL_VIEWPORT_MAX_ROWS


def _guard_view_only(app, primary, other, *, expect_virtual: bool) -> None:
    """Exercise interaction callbacks with every workbook gateway forbidden."""
    old_methods = {}
    hits = []
    metrics = {
        "sheet": str(getattr(primary, "sheet", "")),
        "virtual": bool(expect_virtual),
    }

    def _forbidden(name):
        def _raise(*_args, **_kwargs):
            hits.append(name)
            raise AssertionError(f"view-only callback accessed {name}")

        return _raise

    for name in (
        "ws_a_val", "ws_b_val", "ws_base_val",
        "ws_a_edit", "ws_b_edit", "ws_base_edit", "_request_edit_preload",
    ):
        if hasattr(app, name):
            old_methods[name] = getattr(app, name)
            setattr(app, name, _forbidden(name))
    old_cell, old_iter = Worksheet.cell, Worksheet.iter_rows
    old_ro_iter = ReadOnlyWorksheet.iter_rows
    Worksheet.cell = _forbidden("Worksheet.cell")
    Worksheet.iter_rows = _forbidden("Worksheet.iter_rows")
    ReadOnlyWorksheet.iter_rows = _forbidden("ReadOnlyWorksheet.iter_rows")
    try:
        assert primary.display_rows, "no prepared display row"
        # Every callback below must be attributed to the visible primary
        # Sheet. The preceding guard may have ended on its "other" tab.
        app.nb.select(app._sheet_containers[primary.sheet])
        _wait(
            app.root,
            lambda: str(getattr(app, "selected_sheet", "")) == str(primary.sheet),
            2,
            "primary tab activation before view-only callbacks",
        )
        has_vertical_viewport = bool(
            len(primary._full_display_rows) > sm._VIRTUAL_VIEWPORT_MAX_ROWS
        )
        has_wide_viewport = bool(primary._wide_column_virtual_active())
        # P95 is measured from a real input receipt to the latest fully
        # published logical surface.  The older internal render samples remain
        # diagnostic only and deliberately do not gate this test.
        primary._viewport_request_samples_ms.clear()
        primary._interaction_request_samples_ms.clear()
        primary._viewport_request_completed.clear()
        primary._viewport_request_terminal.clear()
        primary._interaction_request_completed.clear()
        primary._viewport_request_active = None
        primary._interaction_request_active = None
        primary.selected_pair_idx = int(primary.display_rows[0])
        primary.hover_pair_idx = int(primary.display_rows[0])
        primary.hover_col_idx = 2
        primary._last_hover_payload_request_key = None
        primary._last_hover_payload_request_value = None
        calls = 0
        original_payload = primary._cmp_tooltip_payload_by_pair_col

        def _count_payload(*args, **kwargs):
            nonlocal calls
            calls += 1
            return original_payload(*args, **kwargs)

        primary._cmp_tooltip_payload_by_pair_col = _count_payload
        try:
            primary.update_hover_driven_panels(primary.hover_pair_idx, 2, "A")
            primary.update_hover_driven_panels(primary.hover_pair_idx, 2, "A")
        finally:
            primary._cmp_tooltip_payload_by_pair_col = original_payload
        assert calls == 1, f"same hover target rebuilt payload {calls} times"

        skipped_before = int(getattr(primary, "_c_area_same_row_skips", 0))
        primary._update_cursor_lines()
        primary._update_cursor_lines()
        assert int(getattr(primary, "_c_area_same_row_skips", 0)) > skipped_before, (
            "same logical C-area row rebuilt instead of short-circuiting"
        )

        # Small and virtual routes share these input handlers; all must only
        # publish prepared data.  Keep their terminal evidence separately
        # from repaint P95: an edge/no-op must still finish promptly, but it
        # must not be counted as a real surface change merely to pad P95.
        route_terminal_records = {}

        def _exercise_view_route(reason, invoke, *, expect_terminal):
            terminal_before = len(primary._viewport_request_terminal)
            invoke()
            _pump(app.root, 0.10)
            if not expect_terminal:
                return None
            records = [
                item
                for item in list(primary._viewport_request_terminal)[terminal_before:]
                if str(item.get("reason") or "") == reason
            ]
            assert records, ("missing viewport route terminal", reason)
            record = records[-1]
            assert record.get("status") == "complete", record
            route_terminal_records.setdefault(reason, []).append(record)
            if not record.get("surface_changed"):
                assert float(record.get("elapsed_ms", 9999.0)) <= 33.0, record
                return record
            assert int(record["row_start"]) == int(record["actual_row_start"]), record
            if primary._wide_column_virtual_active():
                assert int(record["column_start"]) == int(
                    record["actual_column_start"]
                ), record
            return record

        wheel_record = _exercise_view_route(
            "wheel",
            lambda: primary._on_mousewheel(SimpleNamespace(delta=-120, num=None)),
            expect_terminal=has_vertical_viewport,
        )
        # Preserve the boundary no-op as a latency sample, then exercise the
        # same physical wheel route in the reverse direction.  Dungeon starts
        # at the last logical row in this case, so a negative wheel alone is
        # correctly a no-op and cannot prove its changed-surface path.
        if has_vertical_viewport and not wheel_record.get("surface_changed"):
            _exercise_view_route(
                "wheel",
                lambda: primary._on_mousewheel(SimpleNamespace(delta=120, num=None)),
                expect_terminal=True,
            )
        vpage_record = _exercise_view_route(
            "vpage",
            lambda: primary._yview_both("scroll", "1", "pages"),
            expect_terminal=has_vertical_viewport,
        )
        if has_vertical_viewport and not vpage_record.get("surface_changed"):
            _exercise_view_route(
                "vpage",
                lambda: primary._yview_both("scroll", "-1", "pages"),
                expect_terminal=True,
            )
        _exercise_view_route(
            "vthumb",
            lambda: primary._yview_both("moveto", "0.35"),
            expect_terminal=has_vertical_viewport,
        )
        _exercise_view_route(
            "vminimap",
            lambda: primary._on_vdiff_map_click(
                SimpleNamespace(y=max(1, primary.vdiff_map.winfo_height()))
            ),
            expect_terminal=has_vertical_viewport,
        )
        _exercise_view_route(
            "hminimap",
            lambda: primary._on_hdiff_map_click(
                SimpleNamespace(x=max(1, primary.hdiff_left.winfo_width()), y=1), "left"
            ),
            expect_terminal=has_wide_viewport,
        )
        _exercise_view_route(
            "hthumb",
            lambda: primary._xview_cell_cmp("moveto", "0.20"),
            expect_terminal=has_wide_viewport,
        )
        # Deterministic same-turn h->v/v->h inputs prove latest-wins.  Unlike
        # an edge wheel/unit event, both inputs below are forced to target a
        # different logical coordinate, so the first pending request must have
        # an auditable newer-input terminal record rather than an artificial
        # completion.  Both directions still publish only the final 2D state.
        if has_vertical_viewport and has_wide_viewport:
            row_total = max(0, len(primary._full_display_rows) - sm._VIRTUAL_VIEWPORT_MAX_ROWS)
            col_total = max(0, primary._logical_slot_count() - sm._VIRTUAL_VIEWPORT_MAX_COLUMNS)

            def _different_fraction(current, total, choices):
                for candidate in choices:
                    if int(float(candidate) * total) != int(current):
                        return candidate
                raise AssertionError(("no changed viewport target", current, total, choices))

            publications_before = int(primary._virtual_scroll_publications)
            terminal_before = len(primary._viewport_request_terminal)
            before_row = int(primary._virtual_window_start)
            before_col = int(primary._virtual_column_window_start)
            h_fraction = _different_fraction(before_col, col_total, ("0.91", "0.07", "0.53"))
            primary._xview_cell_cmp("moveto", h_fraction)
            first_h_request = dict(primary._viewport_request_active or {})
            assert first_h_request.get("status") == "pending", first_h_request
            assert int(first_h_request["column_start"]) != before_col, first_h_request
            v_fraction = _different_fraction(before_row, row_total, ("0.73", "0.19", "0.51"))
            primary._yview_both("moveto", v_fraction)
            final_hv_request = dict(primary._viewport_request_active or {})
            assert final_hv_request.get("status") == "pending", final_hv_request
            assert int(final_hv_request["id"]) != int(first_h_request["id"]), (
                first_h_request, final_hv_request,
            )
            assert int(final_hv_request["row_start"]) != before_row, final_hv_request
            assert int(final_hv_request["column_start"]) == int(first_h_request["column_start"]), (
                first_h_request, final_hv_request,
            )
            first_terminal = [
                item for item in list(primary._viewport_request_terminal)[terminal_before:]
                if int(item.get("id", -1)) == int(first_h_request["id"])
            ]
            assert first_terminal and first_terminal[-1].get("superseded_reason") == "newer-input", first_terminal
            _pump(app.root, 0.12)
            assert int(primary._virtual_scroll_publications) == publications_before + 1, (
                "h->v must publish one newest 2D surface",
                publications_before,
                primary._virtual_scroll_publications,
            )
            completed_hv = [
                item for item in primary._viewport_request_completed
                if int(item.get("id", -1)) == int(final_hv_request["id"])
            ]
            assert completed_hv and completed_hv[-1].get("status") == "complete", completed_hv
            assert int(completed_hv[-1]["row_start"]) == int(completed_hv[-1]["actual_row_start"]), completed_hv[-1]
            assert int(completed_hv[-1]["column_start"]) == int(completed_hv[-1]["actual_column_start"]), completed_hv[-1]

            publications_before = int(primary._virtual_scroll_publications)
            terminal_before = len(primary._viewport_request_terminal)
            before_row = int(primary._virtual_window_start)
            before_col = int(primary._virtual_column_window_start)
            v_fraction = _different_fraction(before_row, row_total, ("0.25", "0.83", "0.11"))
            primary._yview_both("moveto", v_fraction)
            first_v_request = dict(primary._viewport_request_active or {})
            assert first_v_request.get("status") == "pending", first_v_request
            assert int(first_v_request["row_start"]) != before_row, first_v_request
            h_fraction = _different_fraction(before_col, col_total, ("0.15", "0.87", "0.39"))
            primary._xview_cell_cmp("moveto", h_fraction)
            final_vh_request = dict(primary._viewport_request_active or {})
            assert final_vh_request.get("status") == "pending", final_vh_request
            assert int(final_vh_request["id"]) != int(first_v_request["id"]), (
                first_v_request, final_vh_request,
            )
            assert int(final_vh_request["column_start"]) != before_col, final_vh_request
            assert int(final_vh_request["row_start"]) == int(first_v_request["row_start"]), (
                first_v_request, final_vh_request,
            )
            first_terminal = [
                item for item in list(primary._viewport_request_terminal)[terminal_before:]
                if int(item.get("id", -1)) == int(first_v_request["id"])
            ]
            assert first_terminal and first_terminal[-1].get("superseded_reason") == "newer-input", first_terminal
            _pump(app.root, 0.12)
            assert int(primary._virtual_scroll_publications) == publications_before + 1, (
                "v->h must publish one newest 2D surface",
                publications_before,
                primary._virtual_scroll_publications,
            )
            completed_vh = [
                item for item in primary._viewport_request_completed
                if int(item.get("id", -1)) == int(final_vh_request["id"])
            ]
            assert completed_vh and completed_vh[-1].get("status") == "complete", completed_vh
            assert int(completed_vh[-1]["row_start"]) == int(completed_vh[-1]["actual_row_start"]), completed_vh[-1]
            assert int(completed_vh[-1]["column_start"]) == int(completed_vh[-1]["actual_column_start"]), completed_vh[-1]
            # Sustain genuine vertical/horizontal scrollbar commands for more
            # than a second. Tk timers drive injection, rather than sleeping
            # in the UI thread, so records measure product input→complete
            # latency instead of the platform's sleep quantisation.
            completed_before = len(primary._viewport_request_completed)
            app._ui_heartbeat_gaps_ms.clear()
            stream_started = time.perf_counter()
            stream = {"running": True, "tick": 0}
            fractions = ("0.08", "0.29", "0.51", "0.73", "0.91")
            vcommand = str(primary.vsb_left.cget("command"))
            hcommand = str(primary.hsb_left.cget("command"))
            assert vcommand and hcommand

            def _inject_stream_input():
                if not stream["running"]:
                    return
                fraction = fractions[int(stream["tick"]) % len(fractions)]
                primary.vsb_left.tk.call(vcommand, "moveto", fraction)
                primary.hsb_left.tk.call(hcommand, "moveto", fraction)
                stream["tick"] = int(stream["tick"]) + 1
                app.root.after(1, _inject_stream_input)

            app.root.after(0, _inject_stream_input)
            app.root.after(1050, lambda: stream.__setitem__("running", False))
            stream_deadline = time.monotonic() + 3.0
            while (
                stream["running"]
                or bool(
                    primary._viewport_request_active
                    and primary._viewport_request_active.get("status") == "pending"
                )
            ):
                if time.monotonic() >= stream_deadline:
                    raise AssertionError(
                        "continuous viewport stream did not drain: "
                        f"active={primary._viewport_request_active!r} "
                        f"completed={list(primary._viewport_request_completed)[-4:]!r}"
                    )
                app.root.update_idletasks()
                app.root.update()
            stream_elapsed_ms = (time.perf_counter() - stream_started) * 1000.0
            assert stream_elapsed_ms >= 1000.0, stream_elapsed_ms
            stream_records = [
                item
                for item in list(primary._viewport_request_completed)[completed_before:]
                if (
                    item.get("status") == "complete"
                    and item.get("counted")
                    and item.get("surface_changed")
                    and int(item.get("generation", -1))
                    == int(app._sheet_compute_generation[primary.sheet])
                    and str(item.get("selected_sheet", "")) == str(primary.sheet)
                )
            ]
            assert len(stream_records) >= 30, stream_records
            for previous, current in zip(stream_records, stream_records[1:]):
                assert (
                    int(previous["actual_row_start"]), int(previous["actual_column_start"])
                ) != (
                    int(current["actual_row_start"]), int(current["actual_column_start"])
                ), (previous, current)
            for item in stream_records:
                assert int(item["row_start"]) == int(item["actual_row_start"]), item
                assert int(item["column_start"]) == int(item["actual_column_start"]), item
            assert len({int(item["actual_row_start"]) for item in stream_records}) > 1
            assert len({int(item["actual_column_start"]) for item in stream_records}) > 1
            # The after(0) coalescer must make its single latest-wins wakeup
            # auditable. It runs on the next event-loop turn without waiting
            # behind unrelated Tcl idle callbacks.
            triggers = {str(item.get("drain_trigger") or "") for item in stream_records}
            assert triggers == {"after0"}, stream_records
            stream_elapsed = [float(item["elapsed_ms"]) for item in stream_records]
            assert app._p95(stream_elapsed) <= 33.0, stream_records
            assert max(stream_elapsed) <= 66.0, stream_records
            assert stream_elapsed[0] <= 33.0, stream_records[0]

            # Stop with a guaranteed different 2D target. This final record
            # must be current, exact, and within the strict per-request gate.
            final_fraction = "0.13" if fractions[-1] != "0.13" else "0.87"
            primary.vsb_left.tk.call(vcommand, "moveto", final_fraction)
            primary.hsb_left.tk.call(hcommand, "moveto", final_fraction)
            active = primary._viewport_request_active
            assert active is not None
            final_request_id = int(active["id"])
            row_total = max(0, len(primary._full_display_rows) - sm._VIRTUAL_VIEWPORT_MAX_ROWS)
            col_total = max(0, primary._logical_slot_count() - sm._VIRTUAL_VIEWPORT_MAX_COLUMNS)
            expected_row = int(float(final_fraction) * row_total)
            expected_col = int(float(final_fraction) * col_total)
            final_deadline = time.monotonic() + 2.0
            final_records = []
            while time.monotonic() < final_deadline:
                app.root.update_idletasks()
                app.root.update()
                final_records = [
                    item for item in primary._viewport_request_completed
                    if int(item.get("id", -1)) == final_request_id
                ]
                if final_records:
                    break
            assert final_records and final_records[-1].get("counted"), final_records
            final_record = final_records[-1]
            assert float(final_record.get("elapsed_ms", 9999.0)) <= 33.0, final_record
            assert int(final_record["row_start"]) == expected_row, final_record
            assert int(final_record["column_start"]) == expected_col, final_record
            assert int(final_record["actual_row_start"]) == expected_row, final_record
            assert int(final_record["actual_column_start"]) == expected_col, final_record
            stream_heartbeat = list(app._ui_heartbeat_gaps_ms)
            assert len(stream_heartbeat) >= 10, stream_heartbeat
            assert app._p95(stream_heartbeat) <= 200.0, stream_heartbeat
            assert max(stream_heartbeat or [0.0]) <= 200.0, stream_heartbeat
            metrics["continuous"] = {
                "n": len(stream_records),
                "p95_ms": round(app._p95(stream_elapsed), 3),
                "max_ms": round(max(stream_elapsed), 3),
                "first_ms": round(stream_elapsed[0], 3),
                "final_ms": round(float(final_record["elapsed_ms"]), 3),
                "heartbeat_n": len(stream_heartbeat),
                "heartbeat_p95_ms": round(app._p95(stream_heartbeat), 3),
                "heartbeat_max_ms": round(max(stream_heartbeat), 3),
                "trigger": "after0",
            }
        # A selection callback must not suppress an already queued viewport
        # publication.  Queue vertical and horizontal changes, immediately
        # click main/C, then drain the original latest viewport intent.
        if has_vertical_viewport or has_wide_viewport:
            if has_vertical_viewport:
                primary._on_mousewheel(SimpleNamespace(delta=-120, num=None))
            else:
                primary._xview_cell_cmp("moveto", "0.35")
            if has_vertical_viewport and has_wide_viewport:
                primary._xview_cell_cmp("moveto", "0.35")
            pending_id = int(primary._viewport_request_active["id"])
            primary._select_from_widget(primary.left, SimpleNamespace(x=1, y=1))
            primary._on_cursor_cmp_click(SimpleNamespace(x=1, y=1))
            _pump(app.root, 0.12)
            terminal = [
                item for item in primary._viewport_request_terminal
                if int(item.get("id", -1)) == pending_id
            ]
            assert terminal and terminal[-1].get("status") == "complete", terminal
        primary._select_from_widget(primary.left, SimpleNamespace(x=1, y=1))
        primary._on_cursor_cmp_click(SimpleNamespace(x=1, y=1))

        # A pending request invalidated by a new comparison generation must
        # leave an explicit terminal reason rather than silently disappearing
        # from latency evidence.
        if has_vertical_viewport or has_wide_viewport:
            if has_vertical_viewport:
                primary._on_mousewheel(SimpleNamespace(delta=-120, num=None))
            else:
                primary._xview_cell_cmp("moveto", "0.65")
            stale_id = int(primary._viewport_request_active["id"])
            entry = app._sheet_exact_entry(primary.sheet)
            old_entry_generation = int(entry.get("generation", 0))
            old_compute_generation = int(app._sheet_compute_generation[primary.sheet])
            entry["generation"] = old_entry_generation + 1
            app._sheet_compute_generation[primary.sheet] = old_compute_generation + 1
            _pump(app.root, 0.08)
            terminal = [
                item for item in primary._viewport_request_terminal
                if int(item.get("id", -1)) == stale_id
            ]
            assert terminal and terminal[-1].get("superseded_reason") == "generation", terminal
            entry["generation"] = old_entry_generation
            app._sheet_compute_generation[primary.sheet] = old_compute_generation

        # Tab activation must use retained prepared cache, not `refresh(rescan=True)`.
        pending_tab_request = None
        if has_vertical_viewport or has_wide_viewport:
            if has_vertical_viewport:
                primary._on_mousewheel(SimpleNamespace(delta=-120, num=None))
            else:
                primary._xview_cell_cmp("moveto", "0.25")
            pending_tab_request = int(primary._viewport_request_active["id"])
        app.nb.select(app._sheet_containers[other.sheet])
        _pump(app.root, 0.12)
        if pending_tab_request is not None:
            terminal = [
                item for item in primary._viewport_request_terminal
                if int(item.get("id", -1)) == pending_tab_request
            ]
            assert terminal, "tab switch silently dropped queued viewport request"
            assert terminal[-1].get("status") == "complete" or terminal[-1].get(
                "superseded_reason"
            ) in {"selected-sheet", "generation"}, terminal
        app.nb.select(app._sheet_containers[primary.sheet])
        _pump(app.root, 0.05)
        _pump(app.root, 0.12)
        _wait(
            app.root,
            lambda: str(getattr(app, "selected_sheet", "")) == str(primary.sheet),
            2,
            "primary tab activation before viewport telemetry",
        )

        if expect_virtual:
            assert primary._virtual_mode_active(), "expected virtual result path"
            if primary._wide_column_virtual_active():
                _assert_wide_column_window(
                    app, primary, three_way=primary._is_three_way_enabled()
                )
            primary._queue_virtual_window(
                max(0, len(primary._full_display_rows) - sm._VIRTUAL_VIEWPORT_MAX_ROWS)
            )
            _pump(app.root, 0.12)
            completed = list(primary._viewport_request_completed)
            samples = list(primary._viewport_request_samples_ms)
            interaction_samples = list(primary._interaction_request_samples_ms)
            repaint_completed = [item for item in completed if item.get("counted")]
            interaction_completed = list(primary._interaction_request_completed)
            assert repaint_completed and samples, "no completed end-to-end viewport requests"
            assert all(item.get("status") == "complete" for item in completed), completed
            expected_reasons = set()
            if has_vertical_viewport:
                expected_reasons.update({"wheel", "vpage", "vthumb", "vminimap"})
            if has_wide_viewport:
                expected_reasons.update({"hminimap", "hthumb"})
            assert expected_reasons.issubset(route_terminal_records), (
                expected_reasons,
                route_terminal_records,
            )
            # The complete Dungeon gate is deliberately stronger than generic
            # route wiring: both axes are movable, so every listed route must
            # have produced at least one actual prepared surface change.
            if has_vertical_viewport and has_wide_viewport:
                changed_routes = {
                    reason
                    for reason, records in route_terminal_records.items()
                    if any(record.get("surface_changed") for record in records)
                }
                assert expected_reasons.issubset(changed_routes), (
                    expected_reasons,
                    route_terminal_records,
                )
            assert {"main-click", "c-area-click"}.issubset(
                {item.get("reason") for item in interaction_completed}
            ), interaction_completed
            assert interaction_samples and app._p95(interaction_samples) <= 33.0, (
                interaction_samples, interaction_completed,
            )
            for item in repaint_completed:
                expected_rows = tuple(
                    primary._full_display_rows[
                        int(item["row_start"]): int(item["row_start"])
                        + sm._VIRTUAL_VIEWPORT_MAX_ROWS
                    ]
                )
                assert item.get("surface_changed"), item
                assert item.get("actual_row_start") == item.get("row_start"), item
                assert item.get("actual_display_rows") == expected_rows, item
                if primary._wide_column_virtual_active():
                    expected_cols = tuple(
                        range(
                            int(item["column_start"]) + 1,
                            min(
                                primary._logical_slot_count(),
                                int(item["column_start"])
                                + sm._VIRTUAL_VIEWPORT_MAX_COLUMNS,
                            ) + 1,
                        )
                    )
                    assert item.get("actual_column_start") == item.get("column_start"), item
                    assert item.get("actual_rendered_columns") == expected_cols, item
            if has_vertical_viewport and has_wide_viewport:
                assert any(
                    item.get("previous_row_start") != item.get("actual_row_start")
                    and item.get("previous_column_start") != item.get("actual_column_start")
                    for item in repaint_completed
                ), repaint_completed
            assert app._p95(samples) <= 33.0, (samples, repaint_completed)
            metrics["viewport"] = {
                "n": len(samples),
                "p95_ms": round(app._p95(samples), 3),
                "max_ms": round(max(samples), 3),
                "interaction_p95_ms": round(app._p95(interaction_samples), 3),
            }
        else:
            assert not primary._virtual_mode_active(), "small result unexpectedly used virtual path"
    finally:
        Worksheet.cell, Worksheet.iter_rows = old_cell, old_iter
        ReadOnlyWorksheet.iter_rows = old_ro_iter
        for name, original in old_methods.items():
            setattr(app, name, original)
    assert not hits, hits
    return metrics


def _assert_base_override_mapping_rules() -> None:
    """A None override must not hide a proven physical Base mapping."""

    class _Probe:
        _base_row_for_pair = sm.SheetView._base_row_for_pair

        def _is_three_way_enabled(self):
            return True

        def _is_missing_sheet_view(self):
            return False

    probe = _Probe()
    probe.app = SimpleNamespace(has_base=True)
    probe.row_pairs = [(11, 21), (12, 22), (13, 23)]
    probe.mine_to_base_row = {11: 101, 12: 102}
    probe.theirs_to_base_row = {21: 201, 22: 202}
    # This mirrors an old/stale cache. It must behave like an absent key.
    probe.pair_base_row_override = {0: None, 1: 901}
    assert probe._base_row_for_pair(0, probe.row_pairs[0]) == 101
    assert probe._base_row_for_pair(1, probe.row_pairs[1]) == 901
    assert probe._base_row_for_pair(2, probe.row_pairs[2]) is None


def _assert_accepted_common_insertion_equality_cache() -> None:
    """Mutation-built proofs preserve formula/shift semantics without view I/O."""

    class _Projection:
        _columns = {
            ("A", 2): 2,
            ("B", 2): 4,
            ("A", 3): 3,
            ("B", 3): 5,
        }

        def physical_col(self, side, logical_col):
            return self._columns.get((str(side).upper(), int(logical_col)))

    class _Probe:
        _accepted_common_insertion_equality_scope = (
            sm.SheetView._accepted_common_insertion_equality_scope
        )
        _clear_accepted_common_insertion_equalities = (
            sm.SheetView._clear_accepted_common_insertion_equalities
        )
        _compute_accepted_common_insertion_equalities_from_edit_workbooks = (
            sm.SheetView._compute_accepted_common_insertion_equalities_from_edit_workbooks
        )
        _finalize_accepted_common_equality_cache_after_mutation = (
            sm.SheetView._finalize_accepted_common_equality_cache_after_mutation
        )
        _accepted_common_insertion_row_is_unchanged = (
            sm.SheetView._accepted_common_insertion_row_is_unchanged
        )
        _all_logical_diff_cols_for_pair = sm.SheetView._all_logical_diff_cols_for_pair
        _build_visual_diff_surface_context = (
            sm.SheetView._build_visual_diff_surface_context
        )
        _visual_diff_cols_for_pair = sm.SheetView._visual_diff_cols_for_pair

        def _is_three_way_enabled(self):
            return True

        def _active_column_projection(self):
            return _Projection()

        def _active_column_comparison_cache(self):
            return self._comparison_cache

    def _book():
        return Workbook().active

    probe = _Probe()
    probe.sheet = "AcceptedCommon"
    probe.app = SimpleNamespace(
        has_base=True,
        _sheet_compute_generation={probe.sheet: 7},
    )
    # Pair 0 deliberately has shifted physical rows.  Source A must compare
    # row B=4 on both sides, while source B must compare row A=2 on both.
    probe.row_pairs = [(2, 4), (5, 7)]
    probe._accepted_common_insert_sources = {2: "A", 3: "B"}
    probe._accepted_common_insertion_equalities = {}
    probe._accepted_common_insertion_mutation_generation = 0
    probe._data_version = 11
    probe._mine_edit_version = 2
    probe._base_edit_version = 3
    probe._theirs_edit_version = 5
    probe._column_projection_generation = 13
    probe._visual_diff_surface_context_build_count = 0
    common_slot = sm.ColumnSlot(
        logical_idx=1,
        mine_col=2,
        base_col=None,
        theirs_col=4,
        state="inserted",
        base_boundary=1,
        origin_side="both",
    )
    probe._comparison_cache = sm.LogicalColumnComparisonCache(
        model=sm.ColumnModel.from_slots(
            sm.ColumnModelCacheKey(probe.sheet, 1, 1), (common_slot,)
        )
    )
    # A real offset pair has a retained row-content diff at the accepted
    # insertion.  The renderer may suppress it only with the immutable proof.
    probe.pair_diff_cols = {0: {2}}
    probe.pair_base_diff_cols = {}

    a_val, b_val, a_edit, b_edit = (_book() for _ in range(4))
    # Source A, logical C2: physical A:B and B:D.  The equivalent formulas
    # become the same logical formula only after their physical column shift.
    for row in (4, 7):
        a_val.cell(row=row, column=2).value = row * 10
        b_val.cell(row=row, column=4).value = row * 10
        a_edit.cell(row=row, column=2).value = f"=A{row}"
        b_edit.cell(row=row, column=4).value = f"=C{row}"
    # Source B, logical C3: physical A:C and B:E.  These use the Mine row
    # (2/5), proving the source-side physical-row rule independently.
    for row in (2, 5):
        a_val.cell(row=row, column=3).value = row * 100
        b_val.cell(row=row, column=5).value = row * 100
        a_edit.cell(row=row, column=3).value = f"=A{row}"
        b_edit.cell(row=row, column=5).value = f"=C{row}"
    # A normal alignment would inspect these values for source A pair 0; they
    # intentionally differ so the assertion below depends on physical row 4.
    a_val.cell(row=2, column=2).value = "wrong-normal-row"
    b_val.cell(row=2, column=4).value = "different-normal-row"

    assert probe._compute_accepted_common_insertion_equalities_from_edit_workbooks(
        a_val, b_val, a_edit, b_edit,
    ) == 4
    for pair_idx in range(2):
        assert probe._accepted_common_insertion_row_is_unchanged(pair_idx, 2, "A")
        assert probe._accepted_common_insertion_row_is_unchanged(pair_idx, 3, "B")
    assert probe._visual_diff_cols_for_pair(0) == set()

    # The surface-context builder is a pure view operation.  Keep hard
    # worksheet sentinels installed for every context parity assertion below.
    ws_hits = []

    def _forbid_view_ws(name):
        def _raise(_sheet):
            ws_hits.append(name)
            raise AssertionError(f"view read {name}")
        return _raise

    probe.app.ws_a_val = _forbid_view_ws("A")
    probe.app.ws_b_val = _forbid_view_ws("B")
    probe.app.ws_a_edit = _forbid_view_ws("A-edit")
    probe.app.ws_b_edit = _forbid_view_ws("B-edit")

    def _assert_surface_context_parity(expected: set[int]):
        """One surface build is pure and exactly matches the legacy helper."""
        before = probe._visual_diff_surface_context_build_count
        context = probe._build_visual_diff_surface_context((0,))
        assert context is not None
        assert probe._visual_diff_surface_context_build_count == before + 1
        assert probe._visual_diff_cols_for_pair(0) == expected
        assert probe._visual_diff_cols_for_pair(
            0, surface_context=context
        ) == expected
        return context

    initial_context = _assert_surface_context_parity(set())
    # The equality snapshot is deliberately bounded to this one viewport pair;
    # pair 1 also has a true proof but may not leak into the surface context.
    assert all(pair_idx == 0 for pair_idx, _col, _source in (
        initial_context.accepted_common_equality_true_keys
    ))
    assert not ws_hits

    # Version scope is part of every proof key.  A fresh source comparison may
    # not reuse a formerly true result merely because its pair index matches.
    probe._data_version += 1
    assert not probe._accepted_common_insertion_row_is_unchanged(0, 2, "A")
    probe._data_version -= 1
    assert probe._accepted_common_insertion_row_is_unchanged(0, 2, "A")

    # The render lookup must not recover missing data from a workbook.  A
    # cache clear is conservative even if app workbook access is forbidden.
    probe._clear_accepted_common_insertion_equalities("test-view-only")
    assert not probe._accepted_common_insertion_row_is_unchanged(0, 2, "A")
    assert probe._visual_diff_cols_for_pair(0) == {2}
    _assert_surface_context_parity({2})
    assert not ws_hits

    # A failed explicit finalizer records bounded diagnostic state and leaves
    # no proof behind; raw values, physical projection, and ordinary workbook
    # save remain intact for the completed mutation/save pipeline.
    probe.app._edit_workbooks_ready = lambda: True
    assert probe._finalize_accepted_common_equality_cache_after_mutation() == 0
    assert probe._accepted_common_insertion_equalities == {}
    assert "AssertionError" in str(probe._accepted_common_insertion_equality_error)
    assert not probe._accepted_common_insertion_row_is_unchanged(0, 2, "A")
    # The explicit mutation finalizer is allowed to try the ready backend; it
    # is intentionally separate from the zero-I/O surface-context assertions.
    assert ws_hits == ["A"]
    ws_hits.clear()
    assert _Projection().physical_col("A", 2) == 2
    assert a_val.cell(row=4, column=2).value == 40
    with tempfile.TemporaryDirectory(prefix="sow_accepted_common_save_") as save_dir:
        a_val.parent.save(os.path.join(save_dir, "mutation-survives.xlsx"))

    # Explicit mutation rebuild sees an edited target immediately; it must not
    # accidentally retain the former equality proof.
    b_edit.cell(row=4, column=4).value = "=D4"
    assert probe._compute_accepted_common_insertion_equalities_from_edit_workbooks(
        a_val, b_val, a_edit, b_edit,
    ) == 4
    assert not probe._accepted_common_insertion_row_is_unchanged(0, 2, "A")
    assert probe._accepted_common_insertion_row_is_unchanged(0, 3, "B")

    # Rollback drops the accepted source proof; undo restores it only through
    # another explicit build after the workbook snapshot is back in place.
    probe._accepted_common_insert_sources = {}
    probe._clear_accepted_common_insertion_equalities("test-rollback")
    assert not probe._accepted_common_insertion_row_is_unchanged(0, 2, "A")
    b_edit.cell(row=4, column=4).value = "=C4"
    probe._accepted_common_insert_sources = {2: "A", 3: "B"}
    assert probe._compute_accepted_common_insertion_equalities_from_edit_workbooks(
        a_val, b_val, a_edit, b_edit,
    ) == 4
    assert probe._accepted_common_insertion_row_is_unchanged(0, 2, "A")
    assert probe._visual_diff_cols_for_pair(0) == set()

    # Context parity must retain all special channels.  Common Base-only
    # evidence is suppressed as structural content, while -1 is preserved;
    # unresolved/ambiguous slots stay row diffs and a structural positive slot
    # is removed.  The source spelling also verifies the context's A/B
    # normalization matches the legacy cache-only proof lookup.
    unresolved_slot = sm.ColumnSlot(
        logical_idx=2,
        mine_col=3,
        base_col=None,
        theirs_col=5,
        state="unresolved",
    )
    ambiguous_slot = sm.ColumnSlot(
        logical_idx=3,
        mine_col=4,
        base_col=None,
        theirs_col=6,
        state="inserted",
        confidence=sm.ColumnMappingConfidence(ambiguous=True),
        base_boundary=3,
        origin_side="both",
    )
    probe._comparison_cache = sm.LogicalColumnComparisonCache(
        model=sm.ColumnModel.from_slots(
            sm.ColumnModelCacheKey(probe.sheet, 2, 2),
            (common_slot, unresolved_slot, ambiguous_slot),
        ),
        structural_diff_cols=frozenset({5}),
    )
    probe._accepted_common_insert_sources = {2: "a", 3: "A", 4: "B"}
    probe.pair_diff_cols = {0: {2, 3, 4, 5}}
    probe.pair_base_diff_cols = {0: {-1, 2}}
    scope = probe._accepted_common_insertion_equality_scope()
    probe._accepted_common_insertion_equalities = {
        (scope, 0, 2, "A"): True,
        # These intentionally cannot suppress their unresolved/ambiguous slots.
        (scope, 0, 3, "A"): True,
        (scope, 0, 4, "B"): True,
    }
    _assert_surface_context_parity({-1, 3, 4})
    probe.pair_diff_cols = {0: set()}
    probe.pair_base_diff_cols = {0: {-1, 2}}
    _assert_surface_context_parity({-1})

    # A defensive context-build failure returns None.  Publishing with that
    # result must take the unchanged legacy helper route, not a partial
    # context-filtered result.
    probe.pair_diff_cols = {0: {2, 3, 4, 5}}
    probe.pair_base_diff_cols = {0: {-1, 2}}
    legacy_after_failed_context = probe._visual_diff_cols_for_pair(0)
    original_scope = probe._accepted_common_insertion_equality_scope
    probe._accepted_common_insertion_equality_scope = (
        lambda: (_ for _ in ()).throw(RuntimeError("surface-context-scope"))
    )
    assert probe._build_visual_diff_surface_context((0,)) is None
    probe._accepted_common_insertion_equality_scope = original_scope
    assert probe._visual_diff_cols_for_pair(
        0, surface_context=None
    ) == legacy_after_failed_context
    assert not ws_hits


def _assert_diffcell_surface_context_parity() -> None:
    """Visible tag projection keeps legacy ranges while avoiding per-row maps."""

    class _Projection:
        slot_count = 3
        _physical = {
            ("A", 1): 1,
            ("A", 3): None,
            ("BASE", 1): 2,
            ("BASE", 3): None,
            ("B", 1): 1,
            ("B", 3): 4,
        }

        def physical_col(self, side, logical_col):
            return self._physical.get((str(side).upper(), int(logical_col)))

    class _Probe:
        _build_diffcell_surface_context = sm.SheetView._build_diffcell_surface_context
        _diffcell_tag_args_for_line = sm.SheetView._diffcell_tag_args_for_line
        _slot_exists_on_side = sm.SheetView._slot_exists_on_side
        _physical_col_for_logical = sm.SheetView._physical_col_for_logical

        def __init__(self):
            self.row_pairs = [(7, 8)]
            self._projection = _Projection()
            self._active_calls = 0
            self._base_span_calls = 0
            self._diffcell_surface_context_build_count = 0

        def _is_three_way_enabled(self):
            return True

        def _base_row_for_pair(self, _pair_idx, _pair):
            return 9

        def _active_column_projection(self):
            self._active_calls += 1
            return self._projection

        def _base_spans(self):
            self._base_span_calls += 1
            return {1: (0, 2), 3: (5, 7)}

        def _spans_for_line(self, line=""):
            base = self._base_spans()
            text = str(line or "")
            if not text:
                return dict(base)
            return {
                col: (min(start, len(text)), min(end, len(text)))
                for col, (start, end) in base.items()
            }

    probe = _Probe()
    rendered = frozenset({1, 3})
    # Column 2 and -1 are full logical differences but are intentionally not
    # materialized in this surface.  Column 3 is missing on A/Base, proving
    # side presence remains exact without a per-cell projection lookup.
    full_visual = (1, 2, 3, -1)
    visible_visual = (1, 3)
    context = probe._build_diffcell_surface_context(rendered)
    assert context is not None
    assert probe._diffcell_surface_context_build_count == 1
    assert context.mine_present_columns == frozenset({1})
    assert context.base_present_columns == frozenset({1})
    assert context.theirs_present_columns == frozenset({1, 3})

    def _assert_same(line_a, line_base, line_b):
        legacy = probe._diffcell_tag_args_for_line(
            4, 0, line_a, line_base, line_b,
            visual_diff_cols=full_visual,
            rendered_logical_columns=rendered,
        )
        active_before_context = probe._active_calls
        spans_before_context = probe._base_span_calls
        bounded = probe._diffcell_tag_args_for_line(
            4, 0, line_a, line_base, line_b,
            visual_diff_cols=full_visual,
            rendered_logical_columns=rendered,
            visible_diff_cols=visible_visual,
            surface_context=context,
        )
        assert bounded == legacy
        # The context path contains no active projection/base-span access for
        # this row; line clamping derives only from its frozen visible spans.
        assert probe._active_calls == active_before_context
        assert probe._base_span_calls == spans_before_context

    _assert_same("AA   CC", "BB   DD", "EE   FF")
    # Concrete short Text lines clamp the off-end C3 range to an empty span;
    # this must match the legacy `_spans_for_line` behavior exactly.
    _assert_same("A", "B", "C")


def _assert_hidden_preemption_state_rules() -> None:
    """A CALCULATING UI-cache wait is never runnable worker work.

    This is deliberately pure state coverage: it catches the cache-await
    spin before any GUI/child-process timing can hide it.  Archive is selected
    and still CALCULATING, while MonsterGroup is the active hidden item.
    """
    selected = "Archive"
    active = "MonsterGroup"
    calculating = sm._SHEET_EXACT_CALCULATING
    for queue in ((), ("ShortWide", "MonsterGroup")):
        assert not sm._selected_sheet_is_runnable_queue_front(selected, queue)
        assert not sm._selected_sheet_should_preempt_background(
            active,
            selected,
            calculating,
            queue,
        )
        assert sm._hidden_interrupt_must_defer(calculating, False)

    # A genuinely front-queued selected request is the only permitted
    # immediate-continue case.
    queue = (selected, "MonsterGroup")
    assert sm._selected_sheet_is_runnable_queue_front(selected, queue)
    assert sm._selected_sheet_should_preempt_background(
        active,
        selected,
        calculating,
        queue,
    )
    assert not sm._hidden_interrupt_must_defer(calculating, True)


def _make_ambiguous_book(path: str, salt: str) -> None:
    """Create duplicate declared records without a whole-file identity shortcut."""
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Ambiguous")
    ws.append(["id@id", "value"])
    # Stay above the accelerated alignment threshold: duplicate declared keys
    # must then take the conservative unresolved path rather than a small-sheet
    # physical-coordinate shortcut.
    for _ in range(260):
        ws.append([7, "duplicate"])
    salt_ws = wb.create_sheet("Salt")
    salt_ws.append(["salt"])
    salt_ws.append([salt])
    wb.save(path)
    wb.close()


def _assert_ambiguous_snapshot_is_blocked(tmp: str) -> None:
    mine = os.path.join(tmp, "ambiguous-mine.xlsx")
    theirs = os.path.join(tmp, "ambiguous-theirs.xlsx")
    base = os.path.join(tmp, "ambiguous-base.xlsx")
    _make_ambiguous_book(mine, "mine")
    _make_ambiguous_book(theirs, "theirs")
    _make_ambiguous_book(base, "base")
    app = sm.SowMergeApp(
        mine,
        theirs,
        merge_mode=True,
        base_path=base,
        initial_sheet="Ambiguous",
    )
    try:
        _wait(
            app.root,
            lambda: (
                app._sheet_exact_entry("Ambiguous").get("state")
                == sm._SHEET_EXACT_UNRESOLVED
            ),
            15,
            "ambiguous snapshot explicit UNRESOLVED",
        )
        view = app.sheet_views["Ambiguous"]
        assert view._derive_lifecycle_state() == "UNRESOLVED"
        assert not app._is_sheet_exact_current("Ambiguous")
        assert not view._guard_mutation_ready("ambiguous-target", notify=False)
    finally:
        app._shutdown_root()


def _run_hidden_snapshot_technical_failure_case(tmp: str) -> None:
    """One bounded hidden-failure case with real, finite view-only routes."""
    started = time.monotonic()
    global_deadline = started + 90.0
    missing = object()

    def _remaining_timeout(stage: str, cap: float = 15.0) -> float:
        remaining = global_deadline - time.monotonic()
        if remaining <= 0.0:
            raise AssertionError(
                f"hidden technical case exceeded 90s before {stage}: "
                f"elapsed={time.monotonic() - started:.3f}s"
            )
        return min(float(cap), remaining)

    def _wait_stage(predicate, stage: str, *, cap: float = 15.0) -> None:
        _wait(app.root, predicate, _remaining_timeout(stage, cap), stage)

    def _file_state(path: str) -> tuple[bool, bytes | None]:
        exists = os.path.exists(path)
        if exists:
            assert os.path.isfile(path), f"settings path is not a file: {path}"
            with open(path, "rb") as source:
                return True, source.read()
        return False, None

    def _canonical(value):
        """Pure container snapshot; it never indexes a Workbook or Tk widget."""
        if value is missing:
            return ("missing",)
        if value is None or isinstance(value, (bool, int, float, str, bytes)):
            return value
        if isinstance(value, dict):
            return tuple(sorted(((repr(key), _canonical(item)) for key, item in value.items()), key=lambda item: item[0]))
        if isinstance(value, (list, tuple)):
            return tuple(_canonical(item) for item in value)
        if isinstance(value, (set, frozenset)):
            return tuple(sorted((_canonical(item) for item in value), key=repr))
        fields = getattr(type(value), "__dataclass_fields__", None)
        if fields:
            return (type(value).__qualname__, tuple((name, _canonical(getattr(value, name))) for name in fields))
        return ("opaque", type(value).__qualname__, id(value))

    def _owner_fields(owner, names):
        return tuple((name, _canonical(getattr(owner, name, missing))) for name in names)

    def _handle_state(value):
        if value is None:
            return None
        return (type(value).__qualname__, id(value), bool(getattr(value, "read_only", False)))

    def _hard_snapshot(view):
        """Operations/model/raw maps are hard; viewport presentation is separately diagnostic."""
        app_fields = (
            "manual_a_cell_ops", "manual_b_cell_ops", "manual_a_formula_cache_ops", "manual_b_formula_cache_ops",
            "manual_a_row_ops", "manual_b_row_ops", "manual_a_column_ops", "manual_b_column_ops",
            "manual_sheet_ops", "auto_sheet_ops", "sheet_level_conflicts", "undo_stack", "redo_stack",
            "modified_a", "modified_b", "modified_sheets_a", "modified_sheets_b", "sheet_operation_overlays",
            "user_touched_conflicts", "_manual_column_action_seq", "_manual_column_op_seq",
            "_manual_structural_op_seq", "_sheet_compute_generation",
        )
        view_fields = (
            "row_pairs", "base_rows", "row_a_to_pair_idx", "row_b_to_pair_idx", "mine_to_base_row",
            "theirs_to_base_row", "base_to_pair_idx", "pair_base_row_overrides", "_missing_base_row_map",
            "pair_raw_parts_a", "pair_raw_parts_b", "pair_raw_parts_base", "pair_diff_cols", "pair_base_diff_cols",
            "_pair_diff_full_exact", "_base_diff_full_exact",
            "_pending_pair_parts_cache", "_comparison_cache", "_column_comparison_cache", "_column_model_cache",
            "col_char_widths", "max_row", "max_col", "col_max_a", "col_max_b", "col_max_base",
            "_sheet_structural_diff", "_align_rows_enabled", "_data_version", "_topology_generation",
            "_mutation_generation", "_only_diff_source_version", "_prepared_complete", "_prepared_rows_complete",
            "_data_ready", "touched_rows",
        )
        return {
            "input_sha": {"mine": _sha256(mine), "theirs": _sha256(theirs)},
            "handles": tuple((name, _handle_state(getattr(app, name, None))) for name in (
                "_wb_a_val", "_wb_b_val", "_wb_base_val", "_wb_a_edit", "_wb_b_edit", "_wb_base_edit",
            )),
            "app": _owner_fields(app, app_fields),
            "view": _owner_fields(view, view_fields),
        }

    def _presentation_snapshot(view):
        return {
            "sheet": str(getattr(view, "sheet", "")), "selected_sheet": str(getattr(app, "selected_sheet", "")),
            "display_rows": _canonical(getattr(view, "display_rows", missing)),
            "full_display_rows": _canonical(getattr(view, "_full_display_rows", missing)),
            "row_to_line": _canonical(getattr(view, "row_to_line", missing)),
            "virtual_window_start": _canonical(getattr(view, "_virtual_window_start", missing)),
            "selected_pair_idx": _canonical(getattr(view, "selected_pair_idx", missing)),
            "hover_pair_idx": _canonical(getattr(view, "hover_pair_idx", missing)),
            "hover_col_idx": _canonical(getattr(view, "hover_col_idx", missing)),
            "viewport_active": _canonical(getattr(view, "_viewport_request_active", missing)),
        }

    def _field_diffs(before, after, path=""):
        if type(before) is not type(after):
            return [(path or "<root>", repr(before)[:240], repr(after)[:240])]
        if isinstance(before, dict):
            result = []
            for key in sorted(set(before) | set(after), key=repr):
                child = f"{path}.{key}" if path else str(key)
                if key not in before or key not in after:
                    result.append((child, repr(before.get(key, missing))[:240], repr(after.get(key, missing))[:240]))
                else:
                    result.extend(_field_diffs(before[key], after[key], child))
            return result
        if isinstance(before, tuple):
            result = []
            if len(before) != len(after):
                result.append((path or "<root>", f"len={len(before)}", f"len={len(after)}"))
            for index, (left, right) in enumerate(zip(before, after)):
                result.extend(_field_diffs(left, right, f"{path}[{index}]"))
            return result
        return [] if before == after else [(path or "<root>", repr(before)[:240], repr(after)[:240])]

    def _assert_no_legacy_fallback(stage: str) -> None:
        assert not legacy_load_calls, (f"forbidden legacy fallback during {stage}", legacy_load_calls)

    def _install_forbidden(owner, names, label, restore, hits):
        for name in names:
            if not hasattr(owner, name):
                continue
            original = getattr(owner, name)

            def _forbidden(*_args, _name=name, **_kwargs):
                hits.append(f"{label}.{_name}")
                raise AssertionError(f"view-only route accessed forbidden {label}.{_name}")

            setattr(owner, name, _forbidden)
            restore.append((owner, name, original))

    @contextmanager
    def _view_only_action_scope(view, stage: str):
        """Public routes plus one Tk turn must not reach I/O, mutable, or compare boundaries."""
        hard_before = _hard_snapshot(view)
        presentation_before = _presentation_snapshot(view)
        restore, hits = [], []
        try:
            _install_forbidden(app, (
                "ws_a_val", "ws_b_val", "ws_base_val", "ws_a_edit", "ws_b_edit", "ws_base_edit",
                "_request_edit_preload", "_load_edit_workbooks_owned", "_ensure_edit_loaded", "_start_background_thread",
                "_atomic_save", "_atomic_save_with_retry", "_atomic_replace_file", "_atomic_replace_file_with_retry",
                "_try_alt_save", "build_manual_merge_output_file", "build_manual_b_output_file", "save_a_inplace",
                "save_b_inplace", "save_merged_and_exit", "apply_sheet_operation_overlay",
                "revert_sheet_operation_overlay", "mark_sheet_operation_overlay_structural", "_request_sheet_compute",
            ), "app", restore, hits)
            _install_forbidden(view, (
                "refresh", "_manual_rescan", "_refresh_mode_switch_preserving_selection", "_start_async_large_only_diff_build",
                "_toggle_only_diff", "_run_copy_action_by_mode", "_refresh_pair_indices_exact",
            ), f"view:{view.sheet}", restore, hits)
            for name in ("_align_selected_sheet_snapshots", "_compare_selected_sheet_snapshots"):
                if not hasattr(sm, name):
                    continue
                original = getattr(sm, name)

                def _forbidden_module(*_args, _name=name, **_kwargs):
                    hits.append(f"module.{_name}")
                    raise AssertionError(f"view-only route invoked forbidden module.{_name}")

                setattr(sm, name, _forbidden_module)
                restore.append((sm, name, original))
            yield
            app.root.update_idletasks()
            app.root.update()
            _assert_no_legacy_fallback(f"{stage} Tk turn")
            assert not hits, (stage, hits)
            hard_after = _hard_snapshot(view)
            assert hard_after == hard_before, {
                "stage": stage,
                "hard_diff": _field_diffs(hard_before, hard_after),
                "presentation_before": presentation_before,
                "presentation_after": _presentation_snapshot(view),
            }
        finally:
            for owner, name, original in reversed(restore):
                setattr(owner, name, original)

    def _cancel_settings_debounces():
        assert app is not None
        for owner in [app, *[view for view in app.sheet_views.values() if view is not None]]:
            after_id = getattr(owner, "_settings_save_id", None)
            if not after_id:
                continue
            scheduler = getattr(owner, "frame", None) or app.root
            try:
                scheduler.after_cancel(after_id)
            finally:
                setattr(owner, "_settings_save_id", None)

    mine = os.path.join(tmp, "hidden-failure-mine.xlsx")
    theirs = os.path.join(tmp, "hidden-failure-theirs.xlsx")
    settings_path_before = str(sm._SETTINGS_PATH)
    settings_before = _file_state(settings_path_before)
    temp_settings_path = os.path.join(tmp, "hidden-technical-settings.json")
    assert os.path.commonpath((os.path.abspath(tmp), os.path.abspath(temp_settings_path))) == os.path.abspath(tmp)
    print("CHANGED_REVISION_HIDDEN_STAGE fixture", flush=True)
    _make_book(mine, "mine")
    _make_book(theirs, "theirs")
    with open(temp_settings_path, "w", encoding="utf-8") as settings_file:
        json.dump({"only_diff": 0}, settings_file, ensure_ascii=False)
    input_before = {"mine": _sha256(mine), "theirs": _sha256(theirs)}
    app = None
    original_load_workbook = sm.load_workbook
    previous_failure_sheet = os.environ.get("SOW_TEST_SNAPSHOT_TECHNICAL_FAILURE_SHEET")
    # The case itself starts with no child hook, irrespective of the caller env.
    os.environ.pop("SOW_TEST_SNAPSHOT_TECHNICAL_FAILURE_SHEET", None)
    legacy_load_calls = []
    primary_error, cleanup_errors = None, []

    def _cleanup_check(label, callback):
        try:
            callback()
        except BaseException as exc:
            cleanup_errors.append((label, exc))

    try:
        sm._SETTINGS_PATH = temp_settings_path
        app = sm.SowMergeApp(mine, theirs, initial_sheet="Dungeon")
        assert int(app.only_diff_default) == 0
        print("CHANGED_REVISION_HIDDEN_STAGE dungeon-exact", flush=True)
        _wait_stage(lambda: (
            app.sheet_views.get("Dungeon") is not None and app.sheet_views["Dungeon"]._data_ready
            and app._is_sheet_exact_current("Dungeon")
        ), "hidden-failure Dungeon exact current")
        dungeon = app.sheet_views["Dungeon"]
        # This deliberately begins with no child-failure injection.  The
        # foreground sequence must observe a genuine MonsterGroup child, tab
        # confirmation, retained request clock, cancellation, and requeue
        # before the deterministic technical failure is armed for its later
        # foreground-resume attempt.
        monster_sheet = "MonsterGroup"
        archive_sheet = "Archive"
        monster_entry_before = dict(app._sheet_exact_entry(monster_sheet))
        monster_generation = int(app._sheet_compute_generation[monster_sheet])
        assert monster_entry_before.get("state") == sm._SHEET_EXACT_PENDING, monster_entry_before
        assert os.environ.get("SOW_TEST_SNAPSHOT_TECHNICAL_FAILURE_SHEET") is None
        foreground_before = len(app._foreground_resume_events)
        child_before = len(app._snapshot_child_events)
        hidden_worker_before = len(app._hidden_worker_events)
        yield_before = len(app._background_yield_events)

        print("CHANGED_REVISION_HIDDEN_STAGE monster-foreground-request", flush=True)
        app.nb.select(app._sheet_containers[monster_sheet])

        def _monster_tab_confirmed() -> bool:
            events = list(app._foreground_resume_events)[foreground_before:]
            return any(
                item.get("event") == "tab-confirmed"
                and isinstance(item.get("visit"), dict)
                and item["visit"].get("sheet") == monster_sheet
                and int(item["visit"].get("generation", -1)) == monster_generation
                and app.selected_sheet == monster_sheet
                and app._sheet_exact_entry(monster_sheet).get("state") in (
                    sm._SHEET_EXACT_PENDING,
                    sm._SHEET_EXACT_CALCULATING,
                )
                for item in events
            )

        _wait_stage(_monster_tab_confirmed, "MonsterGroup tab-confirmed pending/calculating", cap=5.0)
        monster_tab_event = next(
            item for item in list(app._foreground_resume_events)[foreground_before:]
            if item.get("event") == "tab-confirmed"
            and isinstance(item.get("visit"), dict)
            and item["visit"].get("sheet") == monster_sheet
            and int(item["visit"].get("generation", -1)) == monster_generation
        )
        monster_visit = dict(monster_tab_event["visit"])

        def _monster_request_and_child_ready() -> bool:
            events = list(app._foreground_resume_events)[foreground_before:]
            request_events = [
                item for item in events
                if item.get("event") == "request-ledger"
                and isinstance(item.get("request"), dict)
                and item["request"].get("sheet") == monster_sheet
                and int(item["request"].get("generation", -1)) == monster_generation
            ]
            if not request_events:
                return False
            with app._snapshot_child_lock:
                owner = app._snapshot_child_owner
                process = owner.get("process") if isinstance(owner, dict) else None
                return bool(
                    isinstance(owner, dict)
                    and owner.get("sheet") == monster_sheet
                    and int(owner.get("generation", -1)) == monster_generation
                    and process is not None
                    and process.is_alive()
                )

        _wait_stage(_monster_request_and_child_ready, "MonsterGroup ledger + live child", cap=12.0)
        monster_request_event = next(
            item for item in list(app._foreground_resume_events)[foreground_before:]
            if item.get("event") == "request-ledger"
            and isinstance(item.get("request"), dict)
            and item["request"].get("sheet") == monster_sheet
            and int(item["request"].get("generation", -1)) == monster_generation
        )
        monster_request = dict(monster_request_event["request"])
        assert int(monster_request["tab_seq"]) == int(monster_visit["tab_seq"]), (monster_request, monster_visit)
        monster_live_entry = dict(app._sheet_exact_entry(monster_sheet))
        assert float(monster_request["request_started_at"]) == float(monster_live_entry["request_started_at"]), (
            monster_request,
            monster_live_entry,
        )
        with app._snapshot_child_lock:
            monster_owner = app._snapshot_child_owner
            assert isinstance(monster_owner, dict), monster_owner
            monster_process = monster_owner.get("process")
            assert monster_process is not None and monster_process.is_alive(), monster_owner
            monster_pid = int(monster_process.pid)
            monster_child_token = str(monster_owner.get("token") or "")
        assert monster_pid > 0 and monster_child_token, (monster_pid, monster_child_token)

        print("CHANGED_REVISION_HIDDEN_STAGE archive-preempt", flush=True)
        app.nb.select(app._sheet_containers[archive_sheet])

        def _archive_selected_pending() -> bool:
            entry = app._sheet_exact_entry(archive_sheet)
            return bool(
                app.selected_sheet == archive_sheet
                and entry.get("state") in (sm._SHEET_EXACT_PENDING, sm._SHEET_EXACT_CALCULATING)
            )

        _wait_stage(_archive_selected_pending, "Archive selected pending/calculating", cap=5.0)

        def _preempt_evidence():
            terminals = list(app._snapshot_child_events)[child_before:]
            requeues = list(app._hidden_worker_events)[hidden_worker_before:]
            yields = list(app._background_yield_events)[yield_before:]
            terminal = next((
                item for item in terminals
                if item.get("event") == "terminated"
                and item.get("reason") == "cancel-or-preempt"
                and item.get("sheet") == monster_sheet
                and int(item.get("generation", -1)) == monster_generation
                and int(item.get("pid", 0) or 0) == monster_pid
            ), None)
            requeue = next((
                item for item in requeues
                if item.get("event") == "requeue"
                and item.get("sheet") == monster_sheet
                and item.get("selected_sheet") == archive_sheet
            ), None)
            checkpoint_yield = next((
                item for item in yields
                if len(item) >= 4
                and item[1] in {"selected-preempt", "ui-activity"}
                and item[2] == monster_sheet
                and item[3] == archive_sheet
            ), None)
            with app._snapshot_child_lock:
                owner = app._snapshot_child_owner
                process = owner.get("process") if isinstance(owner, dict) else None
                try:
                    owner_alive = bool(process is not None and process.is_alive())
                except BaseException:
                    owner_alive = False
                owner_summary = {
                    "sheet": owner.get("sheet") if isinstance(owner, dict) else None,
                    "generation": owner.get("generation") if isinstance(owner, dict) else None,
                    "token": owner.get("token") if isinstance(owner, dict) else None,
                    "pid": int(getattr(process, "pid", 0) or 0),
                    "alive": owner_alive,
                }
            with app._compute_lock:
                queue_snapshot = tuple(app._compute_queue)
            return {
                "terminals": terminals,
                "requeues": requeues,
                "yields": yields,
                "terminal": terminal,
                "requeue": requeue,
                "checkpoint_yield": checkpoint_yield,
                "checkpoint_reason": (str(checkpoint_yield[1]) if checkpoint_yield is not None else None),
                "entry": dict(app._sheet_exact_entry(monster_sheet)),
                "owner": owner_summary,
                "queue_snapshot": queue_snapshot,
            }

        def _monster_preempted_and_requeued() -> bool:
            evidence = _preempt_evidence()
            terminal = evidence["terminal"]
            entry = evidence["entry"]
            return bool(
                terminal is not None
                and terminal.get("request_token") == monster_child_token
                and terminal.get("result_exists_after_cleanup") is False
                and terminal.get("partial_exists_after_cleanup") is False
                and evidence["requeue"] is not None
                and evidence["checkpoint_yield"] is not None
                and int(entry.get("generation", -1)) == monster_generation
                and entry.get("state") in (sm._SHEET_EXACT_PENDING, sm._SHEET_EXACT_CALCULATING)
            )

        try:
            _wait_stage(_monster_preempted_and_requeued, "MonsterGroup child preempt + requeue", cap=10.0)
        except BaseException:
            preempt_diagnostic = _preempt_evidence()
            print(
                "CHANGED_REVISION_HIDDEN_PREEMPT_DIAGNOSTIC "
                + json.dumps(_canonical(preempt_diagnostic), ensure_ascii=False, sort_keys=True),
                flush=True,
            )
            raise
        preempt_evidence = _preempt_evidence()
        monster_terminal = preempt_evidence["terminal"]
        monster_requeue = preempt_evidence["requeue"]
        checkpoint_yield = preempt_evidence["checkpoint_yield"]
        assert monster_terminal is not None and monster_requeue is not None and checkpoint_yield is not None, preempt_evidence
        assert preempt_evidence["checkpoint_reason"] in {"selected-preempt", "ui-activity"}, preempt_evidence
        assert float(checkpoint_yield[0]) <= float(monster_terminal["at"]) <= float(monster_requeue["at"]), (
            checkpoint_yield,
            monster_terminal,
            monster_requeue,
        )
        assert monster_terminal.get("request_token") == monster_child_token, monster_terminal
        assert monster_terminal.get("result_exists_after_cleanup") is False, monster_terminal
        assert monster_terminal.get("partial_exists_after_cleanup") is False, monster_terminal
        owner_after_preempt = preempt_evidence["owner"]
        assert not (
            owner_after_preempt["sheet"] == monster_sheet
            and int(owner_after_preempt["generation"] if owner_after_preempt["generation"] is not None else -1) == monster_generation
            and owner_after_preempt["alive"]
        ), ("Monster child was already restarted before the controlled hook", preempt_evidence)

        def _slow_legacy_sentinel(*args, **kwargs):
            legacy_load_calls.append(args[0] if args else "unknown")
            raise AssertionError("hidden technical path attempted legacy workbook load")

        # Archive is now the foreground owner; arm only the resumed hidden
        # MonsterGroup attempt.  The parent-side sentinel proves no legacy
        # compatibility fallback was taken by the later failure path.
        os.environ["SOW_TEST_SNAPSHOT_TECHNICAL_FAILURE_SHEET"] = monster_sheet
        sm.load_workbook = _slow_legacy_sentinel
        print("CHANGED_REVISION_HIDDEN_STAGE archive-full-terminal", flush=True)
        _wait_stage(lambda: (
            app.selected_sheet == archive_sheet
            and app.sheet_views.get(archive_sheet) is not None
            and app.sheet_views[archive_sheet]._data_ready
            and app._is_sheet_exact_current(archive_sheet)
            and bool(app._sheet_exact_entry(archive_sheet).get("full_detail_terminal", False))
        ), "Archive selected full terminal", cap=15.0)
        archive = app.sheet_views[archive_sheet]

        def _monster_priority_fired() -> bool:
            events = list(app._foreground_resume_events)[foreground_before:]
            queued = [
                item for item in events
                if item.get("event") == "priority-queued"
                and isinstance(item.get("request"), dict)
                and item["request"].get("sheet") == monster_sheet
                and int(item["request"].get("generation", -1)) == monster_generation
            ]
            fired = [
                item for item in events
                if item.get("event") == "priority-fire"
                and isinstance(item.get("request"), dict)
                and item["request"].get("sheet") == monster_sheet
                and int(item["request"].get("generation", -1)) == monster_generation
            ]
            return len(queued) == 1 and len(fired) == 1

        _wait_stage(_monster_priority_fired, "MonsterGroup priority queued/fire", cap=5.0)
        foreground_events = list(app._foreground_resume_events)[foreground_before:]
        priority_queued = next(
            item for item in foreground_events
            if item.get("event") == "priority-queued"
            and isinstance(item.get("request"), dict)
            and item["request"].get("sheet") == monster_sheet
            and int(item["request"].get("generation", -1)) == monster_generation
        )
        priority_fired = next(
            item for item in foreground_events
            if item.get("event") == "priority-fire"
            and isinstance(item.get("request"), dict)
            and item["request"].get("sheet") == monster_sheet
            and int(item["request"].get("generation", -1)) == monster_generation
        )
        for priority_event in (priority_queued, priority_fired):
            priority_request = priority_event["request"]
            assert priority_request == monster_request, (priority_event, monster_request)
            assert int(priority_event.get("ticket_epoch", -1)) >= 0, priority_event
            assert int(priority_event.get("ticket_seq", -1)) >= 1, priority_event
            assert priority_event.get("after_delay_ms") == 0, priority_event
        assert (priority_queued["ticket_epoch"], priority_queued["ticket_seq"]) == (
            priority_fired["ticket_epoch"], priority_fired["ticket_seq"]
        ), (priority_queued, priority_fired)
        assert priority_fired.get("queue_front") is True and priority_fired.get("worker_kicked") is True, priority_fired
        archive_terminal = next(
            item for item in foreground_events
            if item.get("event") == "terminal"
            and item.get("sheet") == archive_sheet
            and int(item.get("generation", -1)) == int(app._sheet_compute_generation[archive_sheet])
        )
        assert float(monster_terminal["at"]) <= float(archive_terminal["at"]) <= float(priority_queued["at"]) <= float(priority_fired["at"]), (
            monster_terminal,
            archive_terminal,
            priority_queued,
            priority_fired,
        )

        print("CHANGED_REVISION_HIDDEN_STAGE resumed-hidden-failure", flush=True)
        _wait_stage(lambda: (
            int(app._sheet_exact_entry(monster_sheet).get("generation", -1)) == monster_generation
            and app._sheet_exact_entry(monster_sheet).get("state") == sm._SHEET_EXACT_FAILED
        ), "resumed MonsterGroup technical failure", cap=15.0)
        monster_final = dict(app._sheet_exact_entry(monster_sheet))
        _assert_no_legacy_fallback("resumed hidden-failure terminal")
        assert not app._is_sheet_exact_current(monster_sheet)
        assert not bool(monster_final.get("full_detail_terminal", False)), monster_final
        assert float(monster_final.get("request_started_at")) == float(monster_request["request_started_at"]), (
            monster_final,
            monster_request,
        )
        # Selecting MonsterGroup above necessarily constructs an empty view.
        # It must remain non-actionable; a technical failure cannot apply stale
        # rows or invoke a legacy rendering path.
        monster_view = app.sheet_views.get(monster_sheet)
        assert monster_view is not None and not monster_view._data_ready and not monster_view.row_pairs, monster_view
        assert monster_sheet not in app._sheet_cache_store, app._sheet_cache_store
        ledger = app._foreground_resume_ledger
        assert (monster_sheet, monster_generation) not in ledger.entries
        assert (monster_sheet, monster_generation) not in ledger.visits

        # The preemption/failure sequence is now terminal.  Prime two ordinary
        # heartbeat callbacks before beginning the route measurement, then
        # discard only those settled samples.  Do not touch the app heartbeat
        # clock: every later sample is produced by its real scheduled callback.
        heartbeat_settle_counter0 = int(app._ui_heartbeat_samples)
        _wait_stage(
            lambda: int(app._ui_heartbeat_samples) >= heartbeat_settle_counter0 + 2,
            "hidden technical heartbeat settle ticks",
            cap=3.0,
        )
        heartbeat_settle_counter1 = int(app._ui_heartbeat_samples)
        app._ui_heartbeat_gaps_ms.clear()
        heartbeat_measurement_counter0 = int(app._ui_heartbeat_samples)
        assert heartbeat_measurement_counter0 >= heartbeat_settle_counter0 + 2
        heartbeat_measurement_started_at = time.monotonic()
        heartbeat_routes = []
        print(
            "CHANGED_REVISION_HIDDEN_HEARTBEAT_SETTLE "
            + json.dumps(
                {
                    "counter0": heartbeat_settle_counter0,
                    "counter_after_two_ticks": heartbeat_settle_counter1,
                    "measurement_counter0": heartbeat_measurement_counter0,
                    "measurement_wall_started_at": heartbeat_measurement_started_at,
                },
                ensure_ascii=False,
                sort_keys=True,
            ),
            flush=True,
        )

        def _post_route_heartbeat(stage: str, prior_counter: int) -> int:
            _wait_stage(
                lambda: int(app._ui_heartbeat_samples) >= prior_counter + 1,
                f"{stage} post-route heartbeat tick",
                cap=3.0,
            )
            current_counter = int(app._ui_heartbeat_samples)
            heartbeat_routes.append(
                {
                    "stage": stage,
                    "counter_before": prior_counter,
                    "counter_after": current_counter,
                    "post_tick_at": time.monotonic(),
                }
            )
            return current_counter

        def _exercise_route(view, stage: str, invoke) -> None:
            with _view_only_action_scope(view, f"{stage} tab-select"):
                app.nb.select(app._sheet_containers[view.sheet])
                _wait_stage(lambda: (
                    app.selected_sheet == view.sheet and view._data_ready and app._is_sheet_exact_current(view.sheet)
                ), f"{stage} selected current", cap=5.0)
            terminal_before = len(view._viewport_request_terminal)
            active_before = dict(getattr(view, "_viewport_request_active", {}) or {})
            with _view_only_action_scope(view, stage):
                callback_started = time.perf_counter()
                invoke()
                callback_elapsed_ms = (time.perf_counter() - callback_started) * 1000.0
                assert callback_elapsed_ms <= 33.0, (stage, callback_elapsed_ms)
                _assert_no_legacy_fallback(f"{stage} callback")
                active = dict(getattr(view, "_viewport_request_active", {}) or {})
                request_id = active.get("id") if active.get("status") == "pending" else None
                if request_id is None:
                    records = list(view._viewport_request_terminal)[terminal_before:]
                    if records:
                        terminal = records[-1]
                        assert terminal.get("status") == "complete", (stage, terminal)
                        assert int(terminal.get("generation", -1)) == int(app._sheet_compute_generation[view.sheet]), (stage, terminal)
                    else:
                        assert active == active_before or active.get("status") != "pending", (stage, active_before, active)
                    return
                expected_generation = int(app._sheet_compute_generation[view.sheet])

                def _terminal_for_current_request():
                    _assert_no_legacy_fallback(f"{stage} pending")
                    records = [item for item in list(view._viewport_request_terminal)[terminal_before:]
                               if int(item.get("id", -1)) == int(request_id)]
                    if not records:
                        return False
                    terminal = records[-1]
                    assert terminal.get("status") == "complete", (stage, terminal)
                    assert int(terminal.get("generation", -1)) == expected_generation, (stage, terminal)
                    assert str(terminal.get("selected_sheet", "")) == str(view.sheet), (stage, terminal)
                    return True

                _wait_stage(_terminal_for_current_request, f"{stage} request terminal", cap=5.0)
                _assert_no_legacy_fallback(f"{stage} terminal")

        heartbeat_counter = heartbeat_measurement_counter0
        print("CHANGED_REVISION_HIDDEN_STAGE archive-wheel", flush=True)
        _exercise_route(archive, "archive-wheel", lambda: archive._on_mousewheel(SimpleNamespace(delta=-120, num=None)))
        heartbeat_counter = _post_route_heartbeat("archive-wheel", heartbeat_counter)
        print("CHANGED_REVISION_HIDDEN_STAGE dungeon-wheel", flush=True)
        _exercise_route(dungeon, "dungeon-wheel", lambda: dungeon._on_mousewheel(SimpleNamespace(delta=120, num=None)))
        heartbeat_counter = _post_route_heartbeat("dungeon-wheel", heartbeat_counter)
        print("CHANGED_REVISION_HIDDEN_STAGE dungeon-viewport", flush=True)
        _exercise_route(dungeon, "dungeon-viewport", lambda: dungeon._yview_both("moveto", "0.50"))
        heartbeat_counter = _post_route_heartbeat("dungeon-viewport", heartbeat_counter)
        heartbeat_measurement = list(app._ui_heartbeat_gaps_ms)
        heartbeat_counter_after = int(app._ui_heartbeat_samples)
        heartbeat_delta = heartbeat_counter_after - heartbeat_measurement_counter0
        assert heartbeat_delta >= 3 and len(heartbeat_measurement) >= 3, {
            "measurement_counter0": heartbeat_measurement_counter0,
            "measurement_counter_after": heartbeat_counter_after,
            "delta": heartbeat_delta,
            "samples": heartbeat_measurement,
            "routes": heartbeat_routes,
        }
        heartbeat_diagnostic = {
            "measurement_counter0": heartbeat_measurement_counter0,
            "measurement_counter_after": heartbeat_counter_after,
            "measurement_delta": heartbeat_delta,
            "measurement_wall_elapsed_ms": (time.monotonic() - heartbeat_measurement_started_at) * 1000.0,
            "samples_ms": heartbeat_measurement,
            "p95_ms": app._p95(heartbeat_measurement),
            "max_ms": max(heartbeat_measurement or [0.0]),
            "routes": heartbeat_routes,
        }
        print("CHANGED_REVISION_HIDDEN_HEARTBEAT_DIAGNOSTIC " + json.dumps(heartbeat_diagnostic, ensure_ascii=False, sort_keys=True), flush=True)
        assert app._p95(heartbeat_measurement) <= 200.0
        assert max(heartbeat_measurement or [0.0]) <= 200.0
        assert app._is_sheet_exact_current("Dungeon") and app._is_sheet_exact_current("Archive")
    except BaseException as exc:
        primary_error = exc
        raise
    finally:
        _cleanup_check("restore load_workbook", lambda: setattr(sm, "load_workbook", original_load_workbook))
        if previous_failure_sheet is None:
            _cleanup_check("restore failure-sheet env", lambda: os.environ.pop("SOW_TEST_SNAPSHOT_TECHNICAL_FAILURE_SHEET", None))
        else:
            _cleanup_check("restore failure-sheet env", lambda: os.environ.__setitem__("SOW_TEST_SNAPSHOT_TECHNICAL_FAILURE_SHEET", previous_failure_sheet))
        if app is not None:
            _cleanup_check("cancel temp settings debounce", _cancel_settings_debounces)
            _cleanup_check("shutdown app", app._shutdown_root)

        input_after = {"mine": None, "theirs": None}

        def _verify_input_sha() -> None:
            input_after["mine"] = _sha256(mine)
            input_after["theirs"] = _sha256(theirs)
            assert input_after == input_before, (input_before, input_after)

        _cleanup_check("verify hidden technical input SHA", _verify_input_sha)
        print(
            "CHANGED_REVISION_HIDDEN_INPUT_SHA "
            + json.dumps(
                {
                    "before": input_before,
                    "after": input_after,
                    "settings_temp_exists": os.path.exists(temp_settings_path),
                    "elapsed_sec": round(time.monotonic() - started, 3),
                },
                ensure_ascii=False,
                sort_keys=True,
            ),
            flush=True,
        )

        def _verify_settings():
            # Keep the temporary path live through debounce cancellation and
            # shutdown, but restore the user path even when any inspection or
            # JSON validation below fails.  A secondary verification failure
            # must not conceal a primary test exception.
            verification_error = None
            try:
                temp_exists, temp_bytes = _file_state(temp_settings_path)
                assert temp_exists and temp_bytes is not None, "temporary settings vanished before context cleanup"
                assert json.loads(temp_bytes.decode("utf-8")).get("only_diff") == 0
                assert _file_state(settings_path_before) == settings_before, "user settings changed while temp path active"
            except BaseException as exc:
                verification_error = exc
            finally:
                try:
                    sm._SETTINGS_PATH = settings_path_before
                    assert _file_state(settings_path_before) == settings_before, "user settings changed after path restore"
                except BaseException as restore_exc:
                    if verification_error is None:
                        verification_error = restore_exc
                    else:
                        verification_error.add_note(
                            f"settings restore secondary failure: {type(restore_exc).__name__}: {restore_exc}"
                        )
            if verification_error is not None:
                raise verification_error

        _cleanup_check("verify and restore settings path", _verify_settings)
        if primary_error is not None:
            for label, exc in cleanup_errors:
                primary_error.add_note(f"secondary cleanup failure [{label}]: {type(exc).__name__}: {exc}")
        elif cleanup_errors:
            label, exc = cleanup_errors[0]
            raise AssertionError(f"hidden technical cleanup failure [{label}]: {exc}") from exc


def _stable_canonical(value):
    """Canonicalize pure test state without indexing workbook or Tk objects."""
    if value is None or isinstance(value, (bool, int, float, str, bytes)):
        return value
    if isinstance(value, dict):
        return tuple(sorted(
            ((repr(key), _stable_canonical(item)) for key, item in value.items()),
            key=lambda item: item[0],
        ))
    if isinstance(value, (list, tuple)):
        return tuple(_stable_canonical(item) for item in value)
    if isinstance(value, (set, frozenset)):
        return tuple(sorted((_stable_canonical(item) for item in value), key=repr))
    fields = getattr(type(value), "__dataclass_fields__", None)
    if fields:
        return (
            type(value).__qualname__,
            tuple((name, _stable_canonical(getattr(value, name))) for name in fields),
        )
    return ("opaque", type(value).__qualname__, id(value))


def _stable_file_state(path: str) -> tuple[bool, bytes | None]:
    exists = os.path.exists(path)
    if exists:
        assert os.path.isfile(path), f"settings path is not a file: {path}"
        with open(path, "rb") as source:
            return True, source.read()
    return False, None


def _stable_input_sha(paths: dict[str, str]) -> dict[str, str]:
    return {name: _sha256(path) for name, path in sorted(paths.items())}


def _stable_operation_snapshot(app) -> dict[str, object]:
    """Hard no-write state used before a guarded retry becomes accepted."""
    app_fields = (
        "manual_a_cell_ops", "manual_b_cell_ops",
        "manual_a_formula_cache_ops", "manual_b_formula_cache_ops",
        "manual_a_row_ops", "manual_b_row_ops",
        "manual_a_column_ops", "manual_b_column_ops",
        "manual_sheet_ops", "auto_sheet_ops", "sheet_level_conflicts",
        "undo_stack", "redo_stack", "modified_a", "modified_b",
        "modified_sheets_a", "modified_sheets_b", "sheet_operation_overlays",
        "user_touched_conflicts", "_manual_column_action_seq",
        "_manual_column_op_seq", "_manual_structural_op_seq",
    )
    return {
        name: _stable_canonical(getattr(app, name, None))
        for name in app_fields
    }


def _stable_assert_no_write(before, app, paths, input_before, label: str) -> None:
    after = _stable_operation_snapshot(app)
    assert after == before, (label, before, after)
    assert _stable_input_sha(paths) == input_before, label


def _stable_cancel_settings_debounces(app) -> None:
    for owner in [app, *[item for item in app.sheet_views.values() if item is not None]]:
        after_id = getattr(owner, "_settings_save_id", None)
        if not after_id:
            continue
        try:
            owner.frame.after_cancel(after_id)
        finally:
            setattr(owner, "_settings_save_id", None)


def _stable_public_select_right_cell(app, view):
    """Select an actionable row only through the production Text binding."""
    pair_idx = next(
        int(pair)
        for pair in view.display_rows
        if 2 in set(view.pair_diff_cols.get(int(pair), set()))
    )
    line = int(view.row_to_line[pair_idx])
    logical_col = 2
    spans = view._spans_for_line()
    start, end = (int(part) for part in spans[logical_col])
    index = f"{line}.{start}"
    view.right.see(index)
    _pump(app.root, 0.05)
    _wait(
        app.root,
        lambda: view.right.bbox(index) is not None,
        2.0,
        "public right-cell bbox",
    )
    box = view.right.bbox(index)
    assert box is not None, (index, view.right.dlineinfo(index))
    x, y, width, height = (int(part) for part in box)
    event_x = max(1, x + max(1, width // 2))
    event_y = max(1, y + max(1, height // 2))
    event_line, event_col = (
        int(part) for part in str(view.right.index(f"@{event_x},{event_y}")).split(".")
    )
    assert event_line == line and start <= event_col < end, (
        index,
        event_line,
        event_col,
        start,
        end,
    )
    assert int(view._col_from_char(event_col)) == logical_col
    assert str(view.right.bind("<Button-1>") or "").strip()
    original_widget_line = view._widget_line
    right_calls, other_calls = [], []

    def _right_event_line(widget):
        # event_generate supplies event coordinates but cannot move the physical
        # pointer.  Production's legacy helper reads that pointer for the row;
        # bridge only this one verified public right-pane event to its line.
        if widget is view.right:
            right_calls.append({"widget": "right", "line": line})
            return line
        other_calls.append(type(widget).__name__)
        return original_widget_line(widget)

    try:
        view._widget_line = _right_event_line
        view.right.event_generate("<Button-1>", x=event_x, y=event_y)
    finally:
        view._widget_line = original_widget_line
    assert right_calls == [{"widget": "right", "line": line}], right_calls
    assert not other_calls, other_calls
    _pump(app.root, 0.08)
    assert int(view.selected_pair_idx) == pair_idx, (
        pair_idx,
        view.selected_pair_idx,
        view.right.index(f"@{event_x},{event_y}"),
    )
    assert int(view._main_sel_line) == line and int(view._main_sel_col) == logical_col
    expected_cursor_line = 3 if view._is_three_way_enabled() else 2
    assert int(view._cursor_cmp_sel_line) == expected_cursor_line
    assert int(view._cursor_cmp_sel_col) == logical_col
    assert view.has_explicit_cell_selection(), "public Text click did not establish a cell selection"
    return pair_idx


def _stable_selector_kind(case: str) -> tuple[str, bool]:
    case = str(case)
    if case in (_MUTATION_RETRY_2WAY_CASE, _MUTATION_RETRY_3WAY_CASE):
        return "mutation", case.endswith("3way")
    if case in (_SAVE_RETRY_2WAY_CASE, _SAVE_RETRY_3WAY_CASE):
        return "save", case.endswith("3way")
    if case in (_QUIET_WINDOW_PREEMPT_2WAY_CASE, _QUIET_WINDOW_PREEMPT_3WAY_CASE):
        return "quiet", case.endswith("3way")
    raise ValueError(f"unsupported stable selector: {case}")


def _assert_readiness_warning(calls, *, action: str, sheet: str, first_deferred: bool) -> None:
    """Assert the real app modal, while keeping its Tk call non-blocking."""
    assert calls, (action, sheet, calls)
    title, message, kwargs = calls[-1]
    assert title == "正在加载可编辑工作簿", (title, message)
    assert action in message and f"Sheet：{sheet}" in message, message
    assert "不会执行或自动重试" in message, message
    assert "加载完成后请手动重试" in message, message
    assert kwargs.get("parent") is not None, kwargs
    if first_deferred:
        assert "本次首次操作已启动后台加载" in message, message
    else:
        assert "本次首次操作已启动后台加载" not in message, message


def _run_public_mutation_retry(app, view, paths, wait_stage, replace):
    assert str(view._derive_lifecycle_state()) == "EDIT_DEFERRED"
    assert str(view.use_right_btn.cget("state")) == "normal"
    pair_idx = _stable_public_select_right_cell(app, view)
    before = _stable_operation_snapshot(app)
    input_before = _stable_input_sha(paths)
    request_before = len(app._edit_load_requests)
    warning_calls = []
    loader_entered, loader_release = threading.Event(), threading.Event()
    original_loader = app._load_edit_workbooks_owned

    def _gated_real_loader():
        loader_entered.set()
        assert loader_release.wait(30.0), "test gate never released the real edit loader"
        return original_loader()

    def _record_warning(title, message, **kwargs):
        warning_calls.append((str(title), str(message), dict(kwargs)))
        return "ok"

    replace(app, "_load_edit_workbooks_owned", _gated_real_loader)
    replace(sm.messagebox, "showwarning", _record_warning)
    try:
        # The first and second public invokes both reach the handler guard. The
        # Event gate proves one genuine background owner, not a fake READY state.
        view.use_right_btn.invoke()
        wait_stage(loader_entered.is_set, "mutation first real loader entered", cap=10.0)
        owner = app._edit_preload_thread
        assert owner is not None, "first public mutation did not create an edit owner"
        assert len(app._edit_load_requests) == request_before + 1, app._edit_load_requests
        first_request = app._edit_load_requests[-1]
        assert first_request.get("reason") == "mutation:采用所选内容", first_request
        assert first_request.get("caller") == "SheetView._guard_mutation_ready", first_request
        _assert_readiness_warning(
            warning_calls,
            action="采用所选内容",
            sheet=view.sheet,
            first_deferred=True,
        )
        # _show_exact_readiness_modal consumes the one-shot deferred marker.
        assert not bool(getattr(view, "_last_mutation_started_edit_preload", False))
        _stable_assert_no_write(before, app, paths, input_before, "first public mutation rejection")

        view.use_right_btn.invoke()
        assert app._edit_preload_thread is owner, "repeat created a second edit owner"
        assert len(app._edit_load_requests) == request_before + 2, app._edit_load_requests
        assert app._edit_load_requests[-1].get("reason") == "mutation:采用所选内容"
        assert len(warning_calls) == 2, warning_calls
        _assert_readiness_warning(
            warning_calls,
            action="采用所选内容",
            sheet=view.sheet,
            first_deferred=False,
        )
        _stable_assert_no_write(before, app, paths, input_before, "loading mutation retry")

        loader_release.set()
        wait_stage(app._edit_workbooks_ready, "mutation real loader ready", cap=30.0)
        wait_stage(
            lambda: str(view._derive_lifecycle_state()) == "READY",
            "mutation control projects READY",
            cap=15.0,
        )
        assert app._edit_preload_thread is owner
        warning_count_before_manual_retry = len(warning_calls)
        _pump(app.root, 0.12)
        assert len(warning_calls) == warning_count_before_manual_retry, warning_calls
        assert len(app._edit_load_requests) == request_before + 2, app._edit_load_requests
        _stable_assert_no_write(before, app, paths, input_before, "mutation backend completion")

        pair = view.row_pairs[pair_idx]
        direction = str(getattr(view, "_right_copy_direction", "B2A"))
        assert direction == "B2A", direction
        assert str(getattr(view, "_copy_scope_mode", "")) == "row"
        # This exact visual map is evidence for the selected public row, not
        # the row-action write domain.  Row mode deliberately adopts every
        # retained logical slot; the immutable projection below is its oracle.
        assert set(view.pair_diff_cols.get(pair_idx, set())) == {2, 69}
        target_row = view._row_for_side(pair, "A")
        assert target_row is not None
        projection = view._active_column_projection()
        comparison_cache = view._active_column_comparison_cache()
        assert projection.model is comparison_cache.model
        assert tuple(projection.slots) == tuple(comparison_cache.model.slots)
        assert not comparison_cache.unresolved_cols, comparison_cache
        logical_domain = tuple(range(1, int(projection.slot_count) + 1))
        assert logical_domain == tuple(range(1, view._logical_slot_count() + 1))
        assert logical_domain == tuple(range(1, 70)), logical_domain
        raw_a_before = tuple(view.pair_raw_parts_a[pair_idx])
        raw_b_source = tuple(view.pair_raw_parts_b[pair_idx])
        three_way = bool(view._is_three_way_enabled())
        raw_base_before = tuple(view.pair_raw_parts_base.get(pair_idx, ()))
        assert len(raw_a_before) >= len(logical_domain)
        assert len(raw_b_source) >= len(logical_domain)
        if three_way:
            assert len(raw_base_before) >= len(logical_domain)
        assert all(
            not str(value).startswith("=")
            for value in raw_a_before + raw_b_source + raw_base_before
        )

        def _typed_fixture_value(logical_col: int, raw_value):
            """Recover this fixture's documented id/value scalar types from raw cache."""
            raw_value = str(raw_value)
            if logical_col == 1:
                assert raw_value.isdecimal(), raw_value
                return int(raw_value)
            return raw_value

        expected_manual, expected_before = {}, {}
        for logical_col in logical_domain:
            slot = projection.slot(logical_col)
            assert slot is not None and int(slot.logical_idx) == logical_col - 1, slot
            assert str(slot.state) == "retained" and not bool(slot.confidence.ambiguous), slot
            source_col = projection.physical_col("theirs", logical_col)
            destination_col = projection.physical_col("mine", logical_col)
            assert source_col is not None and destination_col is not None, (logical_col, slot)
            source_col, destination_col = int(source_col), int(destination_col)
            assert 1 <= source_col <= len(raw_b_source)
            assert 1 <= destination_col <= len(raw_a_before)
            if three_way:
                base_col = projection.physical_col("base", logical_col)
                assert base_col is not None and 1 <= int(base_col) <= len(raw_base_before), (
                    logical_col,
                    slot,
                )
            key = (view.sheet, int(target_row), destination_col)
            assert key not in expected_manual, (logical_col, key)
            expected_manual[key] = _typed_fixture_value(
                logical_col, raw_b_source[source_col - 1]
            )
            expected_before[key] = _typed_fixture_value(
                logical_col, raw_a_before[destination_col - 1]
            )
        source_value = expected_manual[(view.sheet, int(target_row), 2)]
        manual_before = dict(app.manual_a_cell_ops)
        manual_formula_before = dict(app.manual_a_formula_cache_ops)
        manual_b_before = dict(app.manual_b_cell_ops)
        manual_b_formula_before = dict(app.manual_b_formula_cache_ops)
        assert not manual_before and not manual_formula_before, (manual_before, manual_formula_before)
        overlay_before = dict(
            getattr(app.sheet_operation_overlays.get(view.sheet), "cells", {})
        )
        assert not overlay_before, overlay_before
        # Base identity is immutable for a content-only B->A row adoption.
        # Keep it separate from the derived Base-vs-Mine diff map: the
        # row-only renderer deliberately removes a no-Base empty entry in
        # 2-way mode, while it recomputes the selected Base diff in 3-way.
        base_identity_before = _stable_canonical({
            "raw": getattr(view, "pair_raw_parts_base", {}),
            "mine_to_base": getattr(view, "mine_to_base_row", {}),
            "theirs_to_base": getattr(view, "theirs_to_base_row", {}),
            "override": getattr(view, "pair_base_row_override", {}),
            "missing": getattr(view, "_missing_base_row_map", {}),
            "manual_base_cells": getattr(app, "manual_base_cell_ops", {}),
            "manual_base_formulas": getattr(app, "manual_base_formula_cache_ops", {}),
            "manual_base_rows": getattr(app, "manual_base_row_ops", ()),
            "manual_base_columns": getattr(app, "manual_base_column_ops", ()),
        })
        base_diff_before = {
            int(index): frozenset(int(col) for col in cols)
            for index, cols in getattr(view, "pair_base_diff_cols", {}).items()
        }
        pair_domain = set(range(len(view.row_pairs)))
        assert set(base_diff_before) == pair_domain, (
            "initial exact Base diff domain",
            len(view.row_pairs),
            tuple(sorted(base_diff_before)),
        )
        assert all(not cols for cols in base_diff_before.values()), base_diff_before
        base_diff_exact_before = bool(getattr(view, "_base_diff_full_exact", False))
        expected_target_base_diff = frozenset()
        if three_way:
            # Base=Mine before the action.  Derive the post-action Base diff
            # independently from immutable Base raw data plus the exact values
            # the public B->A row action must install in Mine.
            expected_target_base_diff = frozenset(
                logical_col
                for logical_col in logical_domain
                if _typed_fixture_value(
                    logical_col,
                    raw_base_before[
                        int(projection.physical_col("base", logical_col)) - 1
                    ],
                )
                != expected_manual[
                    (
                        view.sheet,
                        int(target_row),
                        int(projection.physical_col("mine", logical_col)),
                    )
                ]
            )
            assert expected_target_base_diff == frozenset({2, 69}), (
                expected_target_base_diff,
                pair_idx,
            )
        undo_before = len(app.undo_stack)
        view.use_right_btn.invoke()
        _pump(app.root, 0.12)
        key = (view.sheet, int(target_row), 2)
        assert app.manual_a_cell_ops.get(key) == source_value, (
            key,
            source_value,
            app.manual_a_cell_ops.get(key),
        )
        added_keys = set(app.manual_a_cell_ops).difference(manual_before)
        assert added_keys == set(expected_manual), (added_keys, expected_manual)
        assert dict(app.manual_a_cell_ops) == expected_manual
        assert dict(app.manual_a_formula_cache_ops) == manual_formula_before
        assert dict(app.manual_b_cell_ops) == manual_b_before
        assert dict(app.manual_b_formula_cache_ops) == manual_b_formula_before
        assert len(app.undo_stack) == undo_before + 1, app.undo_stack[-2:]
        undo_action = app.undo_stack[-1]
        assert undo_action.get("sheet") == view.sheet and undo_action.get("target") == "A", undo_action
        undo_cells = {
            (view.sheet, int(row), int(physical_col)): (old_edit, old_value)
            for row, physical_col, old_edit, old_value in undo_action.get("cells", ())
        }
        assert set(undo_cells) == set(expected_manual), undo_cells
        assert undo_cells == {
            key: (expected_before[key], expected_before[key])
            for key in expected_manual
        }, undo_cells
        overlay = app.sheet_operation_overlays.get(view.sheet)
        assert overlay is not None
        overlay_cells = tuple(overlay.cells.values())
        assert len(overlay_cells) == len(expected_manual), overlay_cells
        overlay_values = {
            (view.sheet, int(delta.physical_row), int(delta.physical_col)): (
                str(delta.side).upper(), delta.before, delta.after,
            )
            for delta in overlay_cells
        }
        assert set(overlay_values) == set(expected_manual), overlay_values
        assert overlay_values == {
            key: ("A", expected_before[key], expected_manual[key])
            for key in expected_manual
        }, overlay_values
        base_identity_after = _stable_canonical({
            "raw": getattr(view, "pair_raw_parts_base", {}),
            "mine_to_base": getattr(view, "mine_to_base_row", {}),
            "theirs_to_base": getattr(view, "theirs_to_base_row", {}),
            "override": getattr(view, "pair_base_row_override", {}),
            "missing": getattr(view, "_missing_base_row_map", {}),
            "manual_base_cells": getattr(app, "manual_base_cell_ops", {}),
            "manual_base_formulas": getattr(app, "manual_base_formula_cache_ops", {}),
            "manual_base_rows": getattr(app, "manual_base_row_ops", ()),
            "manual_base_columns": getattr(app, "manual_base_column_ops", ()),
        })
        assert base_identity_after == base_identity_before, (
            "Base identity changed during B2A row adoption",
            base_identity_before,
            base_identity_after,
        )
        base_diff_after = {
            int(index): frozenset(int(col) for col in cols)
            for index, cols in getattr(view, "pair_base_diff_cols", {}).items()
        }
        assert bool(getattr(view, "_base_diff_full_exact", False)) == base_diff_exact_before
        if three_way:
            expected_base_diff_after = dict(base_diff_before)
            expected_base_diff_after[pair_idx] = expected_target_base_diff
        else:
            # A two-way row-only refresh has no Base channel.  It removes only
            # the selected pair's synthetic empty Base entry; every other
            # empty entry is retained from the exact snapshot adapter.
            expected_base_diff_after = {
                index: cols
                for index, cols in base_diff_before.items()
                if index != pair_idx
            }
        assert base_diff_after == expected_base_diff_after, (
            "derived Base diff refresh",
            {"before": base_diff_before, "after": base_diff_after},
            expected_base_diff_after,
        )
        assert _stable_input_sha(paths) == input_before
    finally:
        loader_release.set()


def _run_public_save_retry(app, view, paths, wait_stage, replace, *, three_way: bool):
    assert str(view._derive_lifecycle_state()) == "EDIT_DEFERRED"
    button = view.save_a_btn if three_way else view.save_b_btn
    action = "保存 Merged" if three_way else "保存 B"
    assert button is not None and str(button.cget("state")) == "normal"
    before = _stable_operation_snapshot(app)
    input_before = _stable_input_sha(paths)
    request_before = len(app._edit_load_requests)
    original_merged_path = app.merged_path
    merged_output_path = None
    if three_way:
        # `save_merged_and_exit` intentionally rejects an unspecified target
        # before it reaches the shared readiness guard.  Supply only a
        # disposable, initially absent case-local target so this test drives
        # that public guard; the first actual output boundary remains trapped.
        merged_output_path = os.path.abspath(os.path.join(
            os.path.dirname(os.path.abspath(paths["mine"])),
            "save-retry-3way-merged-output.xlsx",
        ))
        assert original_merged_path is None, original_merged_path
        assert os.path.isdir(os.path.dirname(merged_output_path)), merged_output_path
        assert not os.path.exists(merged_output_path), merged_output_path
        app.merged_path = merged_output_path
        assert os.path.abspath(str(app.merged_path)) == merged_output_path
    loader_entered, loader_release = threading.Event(), threading.Event()
    original_loader = app._load_edit_workbooks_owned
    warning_calls, boundary_hits, forbidden_hits = [], [], []
    boundary_reached = threading.Event()
    primary_error = None

    def _gated_real_loader():
        loader_entered.set()
        assert loader_release.wait(30.0), "test gate never released the real save loader"
        return original_loader()

    def _record_warning(title, message, **kwargs):
        warning_calls.append((str(title), str(message), dict(kwargs)))
        return "ok"

    class _FirstIrreversibleSave(RuntimeError):
        pass

    def _save_boundary(*args, **kwargs):
        boundary_hits.append((args, kwargs))
        boundary_reached.set()
        raise _FirstIrreversibleSave("test sentinel at first irreversible save boundary")

    def _forbidden_write(name):
        def _raise(*_args, **_kwargs):
            forbidden_hits.append(name)
            raise AssertionError(f"unexpected write path after sentinel selection: {name}")

        return _raise

    replace(app, "_load_edit_workbooks_owned", _gated_real_loader)
    replace(sm.messagebox, "showwarning", _record_warning)
    replace(sm.messagebox, "askyesno", lambda *_args, **_kwargs: True)
    replace(sm.messagebox, "showerror", lambda *_args, **_kwargs: None)
    replace(sm.messagebox, "showinfo", lambda *_args, **_kwargs: None)
    try:
        if three_way:
            assert app.merged_path == merged_output_path
            assert not os.path.exists(merged_output_path), merged_output_path
        button.invoke()
        wait_stage(loader_entered.is_set, "save first real loader entered", cap=10.0)
        owner = app._edit_preload_thread
        assert owner is not None, "first public Save did not create an edit owner"
        assert len(app._edit_load_requests) == request_before + 1, app._edit_load_requests
        first_request = app._edit_load_requests[-1]
        assert first_request.get("reason") == f"save:{action}", first_request
        assert first_request.get("caller") == "_guard_save_readiness", first_request
        _assert_readiness_warning(
            warning_calls,
            action=action,
            sheet=view.sheet,
            first_deferred=False,
        )
        _stable_assert_no_write(before, app, paths, input_before, "first public save demand")

        button.invoke()
        assert app._edit_preload_thread is owner, "save retry created a second edit owner"
        assert len(app._edit_load_requests) == request_before + 2, app._edit_load_requests
        assert app._edit_load_requests[-1].get("reason") == f"save:{action}"
        assert len(warning_calls) == 2, warning_calls
        _assert_readiness_warning(
            warning_calls,
            action=action,
            sheet=view.sheet,
            first_deferred=False,
        )
        _stable_assert_no_write(before, app, paths, input_before, "loading save retry")

        loader_release.set()
        wait_stage(app._edit_workbooks_ready, "save real loader ready", cap=30.0)
        wait_stage(
            lambda: str(view._derive_lifecycle_state()) == "READY",
            "save control projects READY",
            cap=15.0,
        )
        assert app._edit_preload_thread is owner
        wait_stage(lambda: not owner.is_alive(), "save preload owner released", cap=10.0)
        warning_count_before_manual_retry = len(warning_calls)
        _pump(app.root, 0.12)
        assert len(warning_calls) == warning_count_before_manual_retry, warning_calls
        assert len(app._edit_load_requests) == request_before + 2, app._edit_load_requests
        _stable_assert_no_write(before, app, paths, input_before, "save backend completion")

        # Save is intentionally stopped at its first possible file-producing
        # boundary.  The public retry is real, but no disposable input is ever
        # written, replaced, or redirected to an alternate path by this gate.
        if three_way:
            replace(app, "build_manual_merge_output_file", _save_boundary)
        else:
            replace(app, "_atomic_save", _save_boundary)
        for name in (
            "_atomic_save", "_atomic_save_with_retry", "_atomic_replace_file",
            "_atomic_replace_file_with_retry", "_try_alt_save",
            "build_manual_merge_output_file", "build_manual_b_output_file",
        ):
            if (three_way and name == "build_manual_merge_output_file") or (not three_way and name == "_atomic_save"):
                continue
            if hasattr(app, name):
                replace(app, name, _forbidden_write(name))
        button.invoke()
        wait_stage(boundary_reached.is_set, "public save reached first irreversible sentinel", cap=20.0)
        assert boundary_reached.is_set() and len(boundary_hits) == 1, boundary_hits
        boundary_args, boundary_kwargs = boundary_hits[0]
        assert not boundary_kwargs, boundary_hits
        if three_way:
            assert boundary_args == (), boundary_hits
        else:
            assert len(boundary_args) == 2, boundary_hits
            assert boundary_args[0] is app._wb_b_edit, boundary_hits
            assert os.path.abspath(str(boundary_args[1])) == os.path.abspath(str(app.file_b)), boundary_hits
        assert not forbidden_hits, forbidden_hits
        assert len(app._edit_load_requests) == request_before + 2, app._edit_load_requests
        assert not owner.is_alive(), "save sentinel recreated or retained edit owner"
        _stable_assert_no_write(before, app, paths, input_before, "save sentinel no-write boundary")
        if three_way:
            assert not os.path.exists(merged_output_path), merged_output_path
    except BaseException as exc:
        primary_error = exc
        raise
    finally:
        loader_release.set()
        if three_way:
            output_exists = os.path.exists(merged_output_path)
            app.merged_path = original_merged_path
            if output_exists:
                cleanup_error = AssertionError(
                    f"three-way save test unexpectedly created output: {merged_output_path}"
                )
                if primary_error is None:
                    raise cleanup_error
                primary_error.add_note(str(cleanup_error))


def _run_quiet_window_preempt(app, view, paths, wait_stage, replace, *, three_way: bool):
    assert str(view._derive_lifecycle_state()) == "EDIT_DEFERRED"
    input_before = _stable_input_sha(paths)
    operations_before = _stable_operation_snapshot(app)
    requests_before = len(app._edit_load_requests)
    monster_sheet, archive_sheet = "MonsterGroup", "Archive"
    monster_generation = int(app._sheet_compute_generation[monster_sheet])
    quiet_window_sec = float(sm._BACKGROUND_SHEET_QUIET_WINDOW_SEC)
    assert 1.0 <= quiet_window_sec <= 2.0, quiet_window_sec
    forbidden_edit_hits = []

    def _forbid_edit_demand(name):
        def _raise(*_args, **_kwargs):
            forbidden_edit_hits.append(name)
            raise AssertionError(f"quiet/preempt unexpectedly demanded editable workbooks: {name}")

        return _raise

    replace(app, "_request_edit_preload", _forbid_edit_demand("_request_edit_preload"))
    replace(app, "_load_edit_workbooks_owned", _forbid_edit_demand("_load_edit_workbooks_owned"))

    # Real bound wheel input continually refreshes the quiet window.  No
    # controller, cache refresh, or direct handler is used here.
    for _ in range(4):
        activity_before = int(getattr(app, "_ui_activity_seq", 0))
        view_activity_calls, app_activity_calls = [], []
        original_note_view_activity = view._note_view_activity
        original_note_ui_activity = app._note_ui_activity

        def _spy_note_view_activity(reason):
            view_activity_calls.append(str(reason))
            return original_note_view_activity(reason)

        def _spy_note_ui_activity(reason):
            app_activity_calls.append(str(reason))
            return original_note_ui_activity(reason)

        view._note_view_activity = _spy_note_view_activity
        app._note_ui_activity = _spy_note_ui_activity
        try:
            view.left.event_generate("<MouseWheel>", delta=-120)
            assert view_activity_calls == ["wheel", "vscroll"], view_activity_calls
            assert app_activity_calls == ["wheel", "vscroll", "viewport"], app_activity_calls
        finally:
            view._note_view_activity = original_note_view_activity
            app._note_ui_activity = original_note_ui_activity
        assert int(getattr(app, "_ui_activity_seq", 0)) == activity_before + 3
        assert str(getattr(app, "_ui_activity_reason", "")) == "viewport"
        _pump(app.root, 0.27)
        assert not app._is_sheet_exact_current(monster_sheet), "hidden work ignored the quiet window"
        _stable_assert_no_write(operations_before, app, paths, input_before, "quiet-window wheel input")
    with app._ui_activity_lock:
        final_activity_at = float(app._ui_last_activity_at)
        final_activity_seq = int(app._ui_activity_seq)
    assert final_activity_at > 0.0 and final_activity_seq > 0, (
        final_activity_at,
        final_activity_seq,
    )

    def _live_monster_owner():
        if getattr(app, "_active_compute_sheet", None) != monster_sheet:
            return None
        if monster_sheet not in getattr(app, "_compute_inflight", set()):
            return None
        with app._snapshot_child_lock:
            owner = app._snapshot_child_owner
            if not isinstance(owner, dict):
                return None
            process = owner.get("process")
            if (
                owner.get("sheet") != monster_sheet
                or int(owner.get("generation", -1)) != monster_generation
                or process is None
                or not process.is_alive()
            ):
                return None
            return {
                "pid": int(process.pid),
                "token": str(owner.get("token") or ""),
                "generation": int(owner.get("generation", -1)),
                "started": float(owner.get("started", -1.0)),
            }

    wait_stage(lambda: _live_monster_owner() is not None, "quiet-window live hidden owner", cap=15.0)
    monster_owner = _live_monster_owner()
    assert monster_owner is not None and monster_owner["token"], monster_owner
    with app._ui_activity_lock:
        activity_seq_before_select = int(app._ui_activity_seq)
    assert activity_seq_before_select == final_activity_seq, (
        final_activity_seq,
        activity_seq_before_select,
    )
    clock_epsilon_sec = 0.005
    assert monster_owner["started"] >= (
        final_activity_at + quiet_window_sec - clock_epsilon_sec
    ), (monster_owner, final_activity_at, quiet_window_sec, clock_epsilon_sec)
    assert not forbidden_edit_hits, forbidden_edit_hits

    # Selection itself is the public foreground-preemption input.
    child_before = len(app._snapshot_child_events)
    yield_before = len(app._background_yield_events)
    hidden_before = len(app._hidden_worker_events)
    select_started_at = time.monotonic()
    app.nb.select(app._sheet_containers[archive_sheet])
    wait_stage(
        lambda: (
            str(getattr(app, "selected_sheet", "")) == archive_sheet
            and app._sheet_exact_entry(archive_sheet).get("state")
            in (sm._SHEET_EXACT_PENDING, sm._SHEET_EXACT_CALCULATING)
        ),
        "Archive selected pending",
        cap=8.0,
    )

    def _preempt_records():
        terminals = list(app._snapshot_child_events)[child_before:]
        yields = list(app._background_yield_events)[yield_before:]
        requeues = list(app._hidden_worker_events)[hidden_before:]
        selected_preempt = next((
            item for item in yields
            if len(item) >= 4
            and float(item[0]) >= select_started_at
            and item[1] == "selected-preempt"
            and item[2] == monster_sheet
            and item[3] == archive_sheet
        ), None)
        terminal = next((
            item for item in terminals
            if item.get("event") == "terminated"
            and float(item.get("at", -1.0)) >= select_started_at
            and item.get("reason") == "cancel-or-preempt"
            and item.get("sheet") == monster_sheet
            and int(item.get("generation", -1)) == monster_generation
            and int(item.get("pid", 0) or 0) == monster_owner["pid"]
            and str(item.get("request_token") or "") == monster_owner["token"]
            and item.get("result_exists_after_cleanup") is False
            and item.get("partial_exists_after_cleanup") is False
        ), None)
        requeue = next((
            item for item in requeues
            if item.get("event") == "requeue"
            and float(item.get("at", -1.0)) >= select_started_at
            and item.get("sheet") == monster_sheet
            and item.get("selected_sheet") == archive_sheet
        ), None)
        return selected_preempt, terminal, requeue

    wait_stage(
        lambda: all(item is not None for item in _preempt_records()),
        "selected-preempt terminal plus requeue",
        cap=15.0,
    )
    selected_preempt, terminal, requeue = _preempt_records()
    assert selected_preempt is not None and terminal is not None and requeue is not None
    assert float(selected_preempt[0]) <= float(terminal["at"]) <= float(requeue["at"])
    entry = app._sheet_exact_entry(monster_sheet)
    assert int(entry.get("generation", -1)) == monster_generation, entry
    assert entry.get("state") in (sm._SHEET_EXACT_PENDING, sm._SHEET_EXACT_CALCULATING), entry

    wait_stage(
        lambda: (
            app.sheet_views.get(archive_sheet) is not None
            and app.sheet_views[archive_sheet]._data_ready
            and app._is_sheet_exact_current(archive_sheet)
        ),
        "Archive full exact after selected preempt",
        cap=15.0,
    )
    wait_stage(
        lambda: app._is_sheet_exact_current(monster_sheet),
        "same-generation MonsterGroup requeue terminal",
        cap=20.0,
    )
    assert int(app._sheet_exact_entry(monster_sheet).get("generation", -1)) == monster_generation
    assert len(app._edit_load_requests) == requests_before, app._edit_load_requests
    assert not forbidden_edit_hits, forbidden_edit_hits
    assert not app._edit_workbooks_ready()
    assert getattr(app, "_wb_a_edit", None) is None and getattr(app, "_wb_b_edit", None) is None
    if three_way:
        assert getattr(app, "_wb_base_edit", None) is None
    _stable_assert_no_write(operations_before, app, paths, input_before, "quiet/preempt view-only route")
    assert _stable_input_sha(paths) == input_before


def _run_stable_view_only_selector(tmp: str, case: str) -> None:
    """Run one selector with its own app, temporary settings, and 90s budget."""
    started = time.monotonic()
    deadline = started + 90.0
    kind, three_way = _stable_selector_kind(case)
    mode = "3way" if three_way else "2way"
    mine = os.path.join(tmp, f"{case}-mine.xlsx")
    theirs = os.path.join(tmp, f"{case}-theirs.xlsx")
    base = os.path.join(tmp, f"{case}-base.xlsx")
    settings_path_before = str(sm._SETTINGS_PATH)
    settings_before = _stable_file_state(settings_path_before)
    temp_settings_path = os.path.join(tmp, f"{case}-settings.json")
    app = None
    input_paths = {"mine": mine, "theirs": theirs}
    patches, cleanup_errors = [], []
    primary_error = None
    input_before = None

    def _remaining(stage: str, cap: float = 15.0) -> float:
        remaining = deadline - time.monotonic()
        if remaining <= 0.0:
            raise AssertionError(
                f"{case} exceeded 90s before {stage}: elapsed={time.monotonic() - started:.3f}s"
            )
        return min(float(cap), remaining)

    def _wait_stage(predicate, stage: str, *, cap: float = 15.0) -> None:
        assert app is not None
        _wait(app.root, predicate, _remaining(stage, cap), stage)

    def _replace(owner, name: str, value) -> None:
        original = getattr(owner, name)
        patches.append((owner, name, original))
        setattr(owner, name, value)

    def _cleanup(label: str, callback) -> None:
        try:
            callback()
        except BaseException as exc:
            cleanup_errors.append((label, exc))

    try:
        _make_book(mine, "mine")
        _make_book(theirs, "theirs")
        if three_way:
            # Base=Mine makes the public B->A row adoption and public merged
            # Save path conflict-free while still retaining genuine 3-way data.
            _make_book(base, "mine")
            input_paths["base"] = base
        with open(temp_settings_path, "w", encoding="utf-8") as settings_file:
            json.dump({"only_diff": 0}, settings_file, ensure_ascii=False)
        input_before = _stable_input_sha(input_paths)
        sm._SETTINGS_PATH = temp_settings_path
        app = sm.SowMergeApp(
            mine,
            theirs,
            merge_mode=three_way,
            base_path=(base if three_way else None),
            initial_sheet="Dungeon",
        )
        _wait_stage(
            lambda: (
                app.sheet_views.get("Dungeon") is not None
                and app.sheet_views["Dungeon"]._data_ready
                and bool(getattr(app.sheet_views["Dungeon"], "_prepared_complete", False))
                and app._is_sheet_exact_current("Dungeon")
            ),
            f"{mode} Dungeon exact immutable detail",
            cap=15.0,
        )
        view = app.sheet_views["Dungeon"]
        assert str(view._derive_lifecycle_state()) == "EDIT_DEFERRED"
        assert not app._edit_workbooks_ready()
        if kind == "mutation":
            _run_public_mutation_retry(app, view, input_paths, _wait_stage, _replace)
        elif kind == "save":
            _run_public_save_retry(app, view, input_paths, _wait_stage, _replace, three_way=three_way)
        else:
            _run_quiet_window_preempt(app, view, input_paths, _wait_stage, _replace, three_way=three_way)
        assert _stable_input_sha(input_paths) == input_before, (case, input_before, _stable_input_sha(input_paths))
        print(
            "CHANGED_REVISION_STABLE_SELECTOR "
            + json.dumps(
                {
                    "case": case,
                    "mode": mode,
                    "elapsed_sec": round(time.monotonic() - started, 3),
                    "input_sha": input_before,
                    "edit_requests": len(app._edit_load_requests),
                },
                ensure_ascii=False,
                sort_keys=True,
            ),
            flush=True,
        )
    except BaseException as exc:
        primary_error = exc
        raise
    finally:
        for owner, name, original in reversed(patches):
            _cleanup(f"restore {type(owner).__name__}.{name}", lambda owner=owner, name=name, original=original: setattr(owner, name, original))
        if app is not None:
            _cleanup("cancel stable selector settings debounce", lambda: _stable_cancel_settings_debounces(app))
            _cleanup("shutdown stable selector app", app._shutdown_root)
        if input_before is not None:
            _cleanup(
                "verify stable selector input SHA",
                lambda: assert_input_equal(_stable_input_sha(input_paths), input_before),
            )

        def _restore_settings_and_verify():
            verification_error = None
            try:
                temp_exists, temp_bytes = _stable_file_state(temp_settings_path)
                assert temp_exists and temp_bytes is not None
                assert json.loads(temp_bytes.decode("utf-8")).get("only_diff") == 0
                assert _stable_file_state(settings_path_before) == settings_before
            except BaseException as exc:
                verification_error = exc
            finally:
                try:
                    sm._SETTINGS_PATH = settings_path_before
                    assert _stable_file_state(settings_path_before) == settings_before
                except BaseException as restore_exc:
                    if verification_error is None:
                        verification_error = restore_exc
                    else:
                        verification_error.add_note(
                            f"settings restore secondary failure: {type(restore_exc).__name__}: {restore_exc}"
                        )
            if verification_error is not None:
                raise verification_error

        _cleanup("verify and restore stable selector settings", _restore_settings_and_verify)
        if primary_error is not None:
            for label, exc in cleanup_errors:
                primary_error.add_note(f"secondary cleanup failure [{label}]: {type(exc).__name__}: {exc}")
        elif cleanup_errors:
            label, exc = cleanup_errors[0]
            raise AssertionError(f"stable selector cleanup failure [{label}]: {exc}") from exc


def assert_input_equal(actual, expected) -> None:
    assert actual == expected, (expected, actual)


def _run_mode(tmp: str, *, three_way: bool) -> None:
    mode = "3way" if three_way else "2way"
    mine = os.path.join(tmp, f"{mode}-mine.xlsx")
    theirs = os.path.join(tmp, f"{mode}-theirs.xlsx")
    base = os.path.join(tmp, f"{mode}-base.xlsx")
    _make_book(mine, "mine")
    _make_book(theirs, "theirs")
    if three_way:
        _make_book(base, "base")
    app = sm.SowMergeApp(
        mine,
        theirs,
        merge_mode=three_way,
        base_path=(base if three_way else None),
        initial_sheet="Dungeon",
    )
    try:
        _wait(
            app.root,
            lambda: (
                app.sheet_views.get("Dungeon") is not None
                and app.sheet_views["Dungeon"]._data_ready
                and app._is_sheet_exact_current("Dungeon")
            ),
            15,
            f"{mode} Dungeon exact current",
        )
        dungeon = app.sheet_views["Dungeon"]
        _show_only_diff_from_exact_cache(dungeon)
        assert dungeon._full_display_rows == dungeon._only_diff_rows_cache
        assert dungeon._virtual_mode_active(), (
            "Dungeon must cover virtual viewport path",
            len(dungeon._full_display_rows),
            len(dungeon.display_rows),
            len(dungeon.row_pairs),
            sum(1 for cols in dungeon.pair_diff_cols.values() if cols),
        )

        # Activity continually extends the quiet window.  It must keep hidden
        # MonsterGroup unopened before the 1.2s threshold.
        for _ in range(3):
            dungeon._on_mousewheel(SimpleNamespace(delta=-120, num=None))
            _pump(app.root, 0.35)
            assert not app._is_sheet_exact_current("MonsterGroup"), (
                "hidden exact work ignored the quiet window/UI activity"
            )
        # The loop emits real wheel input. A queued main-pane motion event may
        # legitimately be the later activity marker; it is also a view-only
        # interaction that must extend the quiet window.
        assert app._ui_activity_reason in ("wheel", "vscroll", "viewport", "main-hover"), app._ui_activity_reason

        # Then let the hidden worker really start. A foreground wheel must
        # interrupt/requeue it at a bounded checkpoint. Selecting Archive
        # immediately after proves a new selected sheet is front-prioritized.
        _wait(
            app.root,
            lambda: (
                getattr(app, "_active_compute_sheet", None) == "MonsterGroup"
                and "MonsterGroup" in app._compute_inflight
            ),
            12,
            "hidden MonsterGroup worker start after quiet window",
        )
        yields_before = len(app._background_yield_events)
        resume_events_before = len(app._hidden_resume_events)
        worker_events_before = len(app._hidden_worker_events)
        dungeon._on_mousewheel(SimpleNamespace(delta=-120, num=None))
        _wait(
            app.root,
            lambda: len(app._background_yield_events) > yields_before,
            8,
            "hidden worker yield after visible activity",
        )
        assert app._background_yield_events[-1][1] == "ui-activity", app._background_yield_events
        _wait(
            app.root,
            lambda: (
                getattr(app, "_active_compute_sheet", None) != "MonsterGroup"
                and "MonsterGroup" in app._compute_queue
            ),
            8,
            "hidden worker completed cooperative interruption/requeue",
        )

        archive = _select_tab(app, "Archive")
        _show_only_diff_from_exact_cache(archive)
        assert archive._logical_slot_count() == 3
        assert app._is_sheet_exact_current("Archive"), "selected tab did not preempt hidden work"
        _wait(
            app.root,
            lambda: app._is_sheet_exact_current("MonsterGroup"),
            30,
            "requeued MonsterGroup terminal exact map",
        )
        resume_events = list(app._hidden_resume_events)[resume_events_before:]
        tk_thread_id = threading.get_ident()
        scheduled_resume_events = [
            event for event in resume_events if event.get("event") == "schedule"
        ]
        fired_resume_events = [
            event for event in resume_events if event.get("event") == "fire"
        ]
        assert scheduled_resume_events, resume_events
        assert fired_resume_events, resume_events
        assert len(scheduled_resume_events) <= 2, resume_events
        assert len(fired_resume_events) <= 2, resume_events
        assert all(
            int(event.get("thread_id", -1)) == tk_thread_id
            for event in [*scheduled_resume_events, *fired_resume_events]
        ), resume_events
        assert app._sheet_exact_entry("MonsterGroup").get("state") in sm._SHEET_EXACT_TERMINAL
        worker_events = [
            event
            for event in list(app._hidden_worker_events)[worker_events_before:]
            if event.get("sheet") == "MonsterGroup"
        ]
        assert any(event.get("event") == "requeue" for event in worker_events), worker_events
        defer_events = [
            event for event in worker_events if event.get("event") == "defer"
        ]
        assert defer_events, worker_events
        assert all(
            event.get("current_sheet") == "MonsterGroup"
            and not bool(event.get("selected_is_queue_front"))
            for event in defer_events
        ), defer_events
        assert sum(event.get("event") == "start" for event in worker_events) <= 2, worker_events
        assert sum(event.get("event") == "requeue" for event in worker_events) <= 2, worker_events
        assert len(defer_events) <= 3, worker_events
        hidden_resume_metrics = {
            "schedule_count": len(scheduled_resume_events),
            "fire_count": len(fired_resume_events),
            "resume_thread_ids": sorted(
                {
                    int(event["thread_id"])
                    for event in [*scheduled_resume_events, *fired_resume_events]
                }
            ),
            "tk_thread_id": tk_thread_id,
            "worker_start_count": sum(event.get("event") == "start" for event in worker_events),
            "worker_requeue_count": sum(event.get("event") == "requeue" for event in worker_events),
            "worker_defer_count": sum(event.get("event") == "defer" for event in worker_events),
        }
        monster = _select_tab(app, "MonsterGroup")
        _show_only_diff_from_exact_cache(monster)
        assert monster._logical_slot_count() == 3
        assert not monster._virtual_mode_active(), "MonsterGroup must cover non-virtual path"
        _assert_exact_operation_targets(monster, three_way=three_way)
        short_wide = _select_tab(app, "ShortWide")
        _show_only_diff_from_exact_cache(short_wide)
        assert short_wide._only_diff_rows_exact
        assert len(short_wide._only_diff_rows_cache) == 18
        assert short_wide._full_display_rows == short_wide._only_diff_rows_cache
        assert short_wide.display_rows == short_wide._full_display_rows
        assert short_wide._wide_column_virtual_active(), (
            "short but 69-column result must retain the bounded column surface"
        )
        _assert_exact_operation_targets(short_wide, three_way=three_way)
        _assert_wide_column_window(app, short_wide, three_way=three_way)
        if str(os.environ.get("SOW_CHANGED_REVISION_SHORTWIDE_ONLY", "") or "").strip() == "1":
            return
        _assert_pending_tab_promotion_is_view_only(app, short_wide, archive)
        assert app._is_sheet_exact_current("Dungeon")
        assert app._is_sheet_exact_current("MonsterGroup")
        assert app._is_sheet_exact_current("Archive")
        assert app._is_sheet_exact_current("ShortWide")
        for sheet in ("Dungeon", "MonsterGroup", "Archive", "ShortWide"):
            entry = app._sheet_exact_entry(sheet)
            assert entry.get("state") in sm._SHEET_EXACT_TERMINAL, entry
            assert int(entry.get("generation", -1)) == int(app._sheet_compute_generation[sheet])

        # Arm, but do not start, an eight-second editable-loader sentinel.
        # Phase A below must remain entirely view-only: every input route has
        # to stay independent of this worker and never request it.
        delayed_started = threading.Event()
        original_loader = app._load_edit_workbooks_owned
        loader_calls = []

        def _delayed_loader():
            loader_calls.append(time.perf_counter())
            delayed_started.set()
            time.sleep(8.0)
            return False

        app._load_edit_workbooks_owned = _delayed_loader
        requests_before = len(app._edit_load_requests)
        # Phase A: 8s of continuous read-only input while the sentinel remains
        # uncalled.  Its heartbeat evidence is the hard view-only gate and is
        # intentionally not mixed with the explicit-mutation phase below.
        phase_a_started = time.perf_counter()
        app._ui_heartbeat_gaps_ms.clear()
        phase_a_guards = [
            _guard_view_only(app, monster, dungeon, expect_virtual=False),
            _guard_view_only(app, dungeon, monster, expect_virtual=True),
            _guard_view_only(app, short_wide, archive, expect_virtual=True),
        ]
        view_deadline = time.monotonic() + 8.10
        while time.monotonic() < view_deadline:
            monster._on_mousewheel(SimpleNamespace(delta=-120, num=None))
            dungeon._on_mousewheel(SimpleNamespace(delta=-120, num=None))
            dungeon._xview_cell_cmp("moveto", "0.55")
            _pump(app.root, 0.05)
        assert not loader_calls and not delayed_started.is_set(), loader_calls
        assert len(app._edit_load_requests) == requests_before, app._edit_load_requests
        assert app._p95(app._ui_heartbeat_gaps_ms) <= 200.0, list(app._ui_heartbeat_gaps_ms)
        assert max(app._ui_heartbeat_gaps_ms or [0.0]) <= 200.0, list(app._ui_heartbeat_gaps_ms)
        phase_a_heartbeat = list(app._ui_heartbeat_gaps_ms)
        phase_a_metrics = {
            "duration_ms": round((time.perf_counter() - phase_a_started) * 1000.0, 3),
            "loader_calls": len(loader_calls),
            "edit_request_delta": len(app._edit_load_requests) - requests_before,
            "heartbeat_n": len(phase_a_heartbeat),
            "heartbeat_p95_ms": round(app._p95(phase_a_heartbeat), 3),
            "heartbeat_max_ms": round(max(phase_a_heartbeat or [0.0]), 3),
            "guards": phase_a_guards,
        }

        # Phase B: a real cell-copy entry is explicitly rejected, creates one
        # async owner, and shows the existing readiness modal.  Substitute the
        # native blocking message box only at its boundary so the production
        # modal construction/path is still exercised while this GUI test can
        # prove a Tk callback paints within 100ms.
        modal_calls = []
        original_showwarning = sm.messagebox.showwarning

        def _record_warning(title, message, **kwargs):
            modal_calls.append((str(title), str(message), time.perf_counter()))
            app.show_nonblocking_notice(str(title), warning=True, duration_ms=1000)

        sm.messagebox.showwarning = _record_warning
        mutation_before = (
            dict(app.manual_a_cell_ops),
            dict(app.manual_b_cell_ops),
            list(app.manual_a_row_ops),
            list(app.manual_b_row_ops),
            list(app.undo_stack),
            set(app.modified_sheets_a),
            set(app.modified_sheets_b),
        )
        repaint_at = []
        started = time.perf_counter()
        try:
            app.root.after(50, lambda: repaint_at.append(time.perf_counter()))
            pair_idx = int(monster.display_rows[0])
            assert not monster._copy_single_cell_by_pair(pair_idx, "A2B", 2), (
                "first explicit mutation was not rejected while edit is deferred"
            )
            returned_ms = (time.perf_counter() - started) * 1000.0
            assert returned_ms <= 100.0, returned_ms
            assert delayed_started.wait(1.0), "async edit preload did not start"
            _wait(app.root, lambda: bool(repaint_at), 0.30, "guarded mutation modal repaint")
            assert (repaint_at[0] - started) * 1000.0 <= 150.0, repaint_at
            assert modal_calls and modal_calls[-1][0] == "正在加载可编辑工作簿", modal_calls
            assert bool(app.notice_frame.winfo_ismapped()), "guard notice did not paint"
            assert len(app._edit_load_requests) == requests_before + 1
            assert mutation_before == (
                dict(app.manual_a_cell_ops),
                dict(app.manual_b_cell_ops),
                list(app.manual_a_row_ops),
                list(app.manual_b_row_ops),
                list(app.undo_stack),
                set(app.modified_sheets_a),
                set(app.modified_sheets_b),
            )
        finally:
            sm.messagebox.showwarning = original_showwarning

        # The delayed loader models a failed backend attempt. It must not
        # auto-apply the rejected copy, but the next explicit user retry must
        # create a new single-owner request rather than being permanently
        # blocked by the old `_edit_loading_started` bit.
        _pump(app.root, 8.15)
        assert app._edit_loaded_event.is_set(), "injected eight-second loader did not finish"
        assert not app._edit_loading_started, "failed loader left retry guard stuck"
        assert len(loader_calls) == 1, loader_calls
        retry_modal_calls = []
        original_showwarning = sm.messagebox.showwarning
        sm.messagebox.showwarning = lambda title, message, **kwargs: retry_modal_calls.append(
            (str(title), str(message))
        )
        try:
            assert not monster._copy_single_cell_by_pair(int(monster.display_rows[0]), "A2B", 2)
            _wait(app.root, lambda: len(loader_calls) == 2, 1.0, "second edit loader owner start")
            assert not app._edit_loaded_event.is_set(), "new owner did not reset edit-ready event"
        finally:
            sm.messagebox.showwarning = original_showwarning
        assert retry_modal_calls and retry_modal_calls[-1][0] == "正在加载可编辑工作簿", retry_modal_calls
        assert len(app._edit_load_requests) == requests_before + 2, app._edit_load_requests
        assert mutation_before == (
            dict(app.manual_a_cell_ops),
            dict(app.manual_b_cell_ops),
            list(app.manual_a_row_ops),
            list(app.manual_b_row_ops),
            list(app.undo_stack),
            set(app.modified_sheets_a),
            set(app.modified_sheets_b),
        )

        _pump(app.root, 8.15)
        assert app._edit_loaded_event.is_set(), "second injected loader did not finish"
        assert not app._edit_loading_started, "second failed loader left retry guard stuck"
        assert len(loader_calls) == 2, loader_calls
        failed_owner_count = len(loader_calls)
        assert mutation_before == (
            dict(app.manual_a_cell_ops),
            dict(app.manual_b_cell_ops),
            list(app.manual_a_row_ops),
            list(app.manual_b_row_ops),
            list(app.undo_stack),
            set(app.modified_sheets_a),
            set(app.modified_sheets_b),
        )

        # Restore the genuine background loader. A third explicit click still
        # rejects immediately while it loads; after READY a fourth explicit
        # click alone performs the operation, proving no failed attempt was
        # silently replayed and the physical target is preserved.
        app._load_edit_workbooks_owned = original_loader
        ready_modal_calls = []
        original_showwarning = sm.messagebox.showwarning
        sm.messagebox.showwarning = lambda title, message, **kwargs: ready_modal_calls.append(
            (str(title), str(message))
        )
        real_load_started = time.perf_counter()
        try:
            assert not monster._copy_single_cell_by_pair(int(monster.display_rows[0]), "A2B", 2)
        finally:
            sm.messagebox.showwarning = original_showwarning
        assert ready_modal_calls and ready_modal_calls[-1][0] == "正在加载可编辑工作簿", ready_modal_calls
        assert len(app._edit_load_requests) == requests_before + 3, app._edit_load_requests
        _wait(app.root, app._edit_workbooks_ready, 30, "real edit loader ready after retry")
        assert mutation_before == (
            dict(app.manual_a_cell_ops),
            dict(app.manual_b_cell_ops),
            list(app.manual_a_row_ops),
            list(app.manual_b_row_ops),
            list(app.undo_stack),
            set(app.modified_sheets_a),
            set(app.modified_sheets_b),
        )
        pair_idx = int(monster.display_rows[0])
        pair = monster.row_pairs[pair_idx]
        source_row = monster._row_for_side(pair, "A")
        target_row = monster._row_for_side(pair, "B")
        assert source_row is not None and target_row is not None
        expected_value = app.ws_a_val(monster.sheet).cell(row=source_row, column=2).value
        monster._copy_single_cell_by_pair(pair_idx, "A2B", 2)
        assert app.manual_b_cell_ops[(monster.sheet, target_row, 2)] == expected_value
        assert (
            app._is_sheet_exact_current("Dungeon")
            and app._is_sheet_exact_current("MonsterGroup")
            and app._is_sheet_exact_current("Archive")
        )
        phase_b_metrics = {
            "first_return_ms": round(returned_ms, 3),
            "first_modal_repaint_ms": round((repaint_at[0] - started) * 1000.0, 3),
            "failed_owner_count": failed_owner_count,
            "edit_request_delta": len(app._edit_load_requests) - requests_before,
            "real_ready_ms": round((time.perf_counter() - real_load_started) * 1000.0, 3),
            "manual_retry_applied": bool(
                app.manual_b_cell_ops.get((monster.sheet, target_row, 2)) == expected_value
            ),
            "target": [monster.sheet, int(target_row), 2],
        }
        print(
            "SOW_CHANGED_REVISION_METRICS "
            + json.dumps(
                {
                    "mode": "3way" if three_way else "2way",
                    "hidden_resume": hidden_resume_metrics,
                    "continuous": next(
                        (
                            item.get("continuous")
                            for item in phase_a_guards
                            if item.get("continuous")
                        ),
                        None,
                    ),
                    "phase_a": phase_a_metrics,
                    "phase_b": phase_b_metrics,
                },
                ensure_ascii=False,
                sort_keys=True,
            )
        )
        app._load_edit_workbooks_owned = original_loader
    finally:
        try:
            app._load_edit_workbooks_owned = original_loader
        except Exception:
            pass
        app._shutdown_root()


def main(argv=None) -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--list-cases", action="store_true")
    parser.add_argument("--case", choices=_CASES)
    args = parser.parse_args(argv)
    if args.list_cases:
        if args.case:
            parser.error("--list-cases cannot be combined with --case")
        print("\n".join(_CASES), flush=True)
        return

    selected_cases = (args.case,) if args.case else (_MUTATION_RETRY_2WAY_CASE,)
    source_path = os.path.abspath(sm.__file__)
    with open(source_path, "rb") as source_file:
        source_sha256 = hashlib.sha256(source_file.read()).hexdigest()
    test_path = os.path.abspath(__file__)
    with open(test_path, "rb") as test_file:
        test_sha256 = hashlib.sha256(test_file.read()).hexdigest()
    print(
        "SOW_CHANGED_REVISION_CONFIG "
        + json.dumps(
            {
                "cases": selected_cases,
                "source": source_path,
                "source_sha256": source_sha256,
                "test_sha256": test_sha256,
            },
            ensure_ascii=False,
            sort_keys=True,
        )
    )
    old_threshold = sm._LARGE_SHEET_ROW_THRESHOLD
    old_snapshot_enabled = sm._LARGE_SHEET_SNAPSHOT_ENGINE_ENABLED
    sm._LARGE_SHEET_ROW_THRESHOLD = 200
    sm._LARGE_SHEET_SNAPSHOT_ENGINE_ENABLED = True
    try:
        for selected_case in selected_cases:
            temp_root = None
            try:
                with tempfile.TemporaryDirectory(
                    prefix=f"sow_changed_revision_{selected_case}_"
                ) as tmp:
                    temp_root = str(tmp)
                    if selected_case == _HIDDEN_SNAPSHOT_TECHNICAL_FAILURE_CASE:
                        _run_hidden_snapshot_technical_failure_case(tmp)
                    else:
                        _run_stable_view_only_selector(tmp, selected_case)
            finally:
                if temp_root is not None and os.path.lexists(temp_root):
                    cleanup_error = AssertionError(
                        f"changed-revision TemporaryDirectory not removed: {temp_root}"
                    )
                    active_error = sys.exc_info()[1]
                    if active_error is not None:
                        active_error.add_note(str(cleanup_error))
                    else:
                        raise cleanup_error
    finally:
        sm._LARGE_SHEET_ROW_THRESHOLD = old_threshold
        sm._LARGE_SHEET_SNAPSHOT_ENGINE_ENABLED = old_snapshot_enabled
    print("GUI_SELF_TEST_CHANGED_REVISION_VIEW_ONLY_OK")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"GUI_SELF_TEST_CHANGED_REVISION_VIEW_ONLY_FAIL: {exc}", file=sys.stderr)
        raise

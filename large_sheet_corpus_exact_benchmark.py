"""Corpus benchmark whose final state is guarded by a direct legacy fallback.

Snapshots are measured candidates only.  A same-copy candidate that is
unresolved or reports a difference/conflict is diagnostic evidence, never a
published exact result: the worker immediately runs a direct read-only legacy
cell/formula oracle against the already-disposable sides.  Its time remains in
the 15-second request-to-final-exact budget.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import subprocess
import sys
import tempfile
import time
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path

import psutil
from openpyxl import load_workbook

from large_sheet_corpus_benchmark import ROOT, SCHEMA, append_jsonl, copy_sides, inventory, rankings, signature, write_json, write_report
from sow_merge_tool import _formula_text, _special_formula_signature

EXACT_SCHEMA = "large-sheet-corpus-exact-v4"


def _sha256(path):
    digest = hashlib.sha256()
    with Path(path).open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _run_provenance():
    repo = Path(__file__).resolve().parent
    try:
        diff = subprocess.run(
            ["git", "diff", "--binary", "--no-ext-diff"],
            cwd=repo,
            text=False,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            check=False,
        )
        diff_hash = hashlib.sha256(diff.stdout).hexdigest() if diff.returncode == 0 else None
    except OSError:
        diff_hash = None
    return {
        "sow_merge_tool_sha256": _sha256(repo / "sow_merge_tool.py"),
        "harness_sha256": _sha256(Path(__file__).resolve()),
        "git_diff_sha256": diff_hash,
    }


def _stable_formula_token(value):
    """Serialize formula values without process-specific object addresses.

    openpyxl represents array and data-table formulas as objects whose repr()
    contains a memory address. The direct fallback is an exact Oracle, so a
    same-copy comparison must use the production formula identity instead.
    """
    special = _special_formula_signature(value)
    if special is not None:
        return ("special", special)
    text = _formula_text(value)
    if text is not None:
        return ("formula-text", text)
    if value is None or isinstance(value, (bool, int, float, str, bytes)):
        return ("scalar", type(value).__module__, type(value).__qualname__, value)
    attributes = getattr(value, "__dict__", None)
    if isinstance(attributes, dict):
        return (
            "attributes",
            type(value).__module__,
            type(value).__qualname__,
            tuple(sorted((str(key), _stable_formula_token(item)) for key, item in attributes.items())),
        )
    return ("opaque-formula-class", type(value).__module__, type(value).__qualname__)


def token(value_cell, formula_cell):
    return (
        str(value_cell.data_type or ""),
        repr(value_cell.value),
        str(formula_cell.data_type or ""),
        _stable_formula_token(formula_cell.value),
    )


def direct_legacy_self_oracle(mine_path, theirs_path, base_path, sheet):
    """Streaming normal legacy value/formula equality for identical copies.

    This is purposely independent of immutable snapshot alignment.  It is a
    conservative final oracle for self comparison, not a replacement runtime
    comparison engine or a mutation path.
    """
    paths = [Path(mine_path), Path(theirs_path)] + ([Path(base_path)] if base_path else [])
    value_books = [load_workbook(path, read_only=True, data_only=True, keep_vba=path.suffix.casefold() == ".xlsm") for path in paths]
    formula_books = [load_workbook(path, read_only=True, data_only=False, keep_vba=path.suffix.casefold() == ".xlsm") for path in paths]
    try:
        value_sheets, formula_sheets = [book[sheet] for book in value_books], [book[sheet] for book in formula_books]
        max_row = max(int(ws.max_row or 1) for ws in value_sheets + formula_sheets)
        max_col = max(int(ws.max_column or 1) for ws in value_sheets + formula_sheets)
        row_iters = [ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col) for ws in value_sheets]
        formula_iters = [ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col) for ws in formula_sheets]
        differences = conflicts = formulas = 0
        for value_rows, formula_rows in zip(zip(*row_iters), zip(*formula_iters)):
            for cells, formula_cells in zip(zip(*value_rows), zip(*formula_rows)):
                row_tokens = [token(cells[index], formula_cells[index]) for index in range(len(paths))]
                formulas += str(formula_cells[0].data_type or "") == "f"
                if row_tokens[0] != row_tokens[1]: differences += 1
                if len(row_tokens) == 3:
                    if row_tokens[0] != row_tokens[2] or row_tokens[1] != row_tokens[2]: differences += 1
                    if row_tokens[0] != row_tokens[2] and row_tokens[1] != row_tokens[2] and row_tokens[0] != row_tokens[1]: conflicts += 1
        return {"dimensions": {"rows": max_row, "columns": max_col}, "formula_cells": formulas,
                "diff_cell_count": differences, "conflict_cell_count": conflicts,
                "zero_difference": differences == 0, "zero_conflict": conflicts == 0}
    finally:
        for book in value_books + formula_books: book.close()


def sheet_worker(args):
    import sow_merge_tool as sm
    started = time.perf_counter(); process = psutil.Process(); before = process.memory_info().rss
    paths = [Path(args.mine), Path(args.theirs)] + ([Path(args.base)] if args.base else [])
    output = {"schema": EXACT_SCHEMA, "mode": args.mode, "source_path": args.source, "sheet": args.sheet,
              "gate_seconds": args.gate, "snapshot_candidate_state": "FAILED", "snapshot_candidate_diff_pairs": None,
              "snapshot_candidate_conflict_pairs": None, "snapshot_candidate_unresolved": None, "candidate_mismatch": True,
              "engine": "snapshot-paired-read-only", "fallback_reason": None, "fallback_engine": None, "final_state": "FAILED",
              "status": "FAILED", "oracle_zero_difference": False, "oracle_zero_conflict": False, "error": None}
    try:
        if any(path.resolve().is_relative_to(ROOT.resolve()) for path in paths): raise AssertionError("source path supplied as a test side")
        request = time.perf_counter(); candidate_error = None
        try:
            with ThreadPoolExecutor(max_workers=len(paths)) as pool:
                future = [pool.submit(sm._stream_selected_sheet_snapshot, str(path), str(path), args.sheet, side) for path, side in zip(paths, ("mine", "theirs", "base"))]
                mine, theirs = future[0].result(), future[1].result(); base = future[2].result() if len(future) == 3 else None
            candidate = sm._compare_selected_sheet_snapshots(mine, theirs, base)
            diffs = sum(bool(value) for value in candidate.pair_diff_cols) + sum(bool(value) for value in candidate.pair_base_diff_cols)
            conflicts = sum(bool(value) for value in candidate.conflict_cols)
            candidate_same = diffs == conflicts == 0 and not candidate.unresolved
            output.update({"snapshot_candidate_state": "EXACT_SAME" if candidate_same else "UNRESOLVED" if candidate.unresolved else "EXACT_CHANGED",
                           "snapshot_candidate_diff_pairs": diffs, "snapshot_candidate_conflict_pairs": conflicts,
                           "snapshot_candidate_unresolved": bool(candidate.unresolved), "candidate_mismatch": not candidate_same,
                           "dimensions": {"rows": mine.max_row, "columns": mine.max_col},
                           "formula_cells": sum(cell.formula_kind in {"formula", "special"} for row in mine.rows for cell in row.cells)})
        except Exception as exc:
            candidate_error = f"{type(exc).__name__}: {exc}"; output["snapshot_candidate_error"] = candidate_error
        if output["candidate_mismatch"]:
            reasons = []
            if candidate_error: reasons.append("snapshot-error")
            if output["snapshot_candidate_unresolved"]: reasons.append("snapshot-unresolved")
            if output["snapshot_candidate_diff_pairs"]: reasons.append("snapshot-nonzero-diff")
            if output["snapshot_candidate_conflict_pairs"]: reasons.append("snapshot-nonzero-conflict")
            output.update(engine="snapshot-candidate+legacy-fallback", fallback_engine="legacy-direct-read-only-cell-formula", fallback_reason=",".join(reasons) or "candidate-not-exact-same")
            fallback = direct_legacy_self_oracle(args.mine, args.theirs, args.base, args.sheet)
            output.update({"legacy_oracle": fallback, "dimensions": fallback["dimensions"], "formula_cells": fallback["formula_cells"],
                           "oracle_zero_difference": fallback["zero_difference"], "oracle_zero_conflict": fallback["zero_conflict"]})
        else:
            output.update(oracle_zero_difference=True, oracle_zero_conflict=True)
        elapsed = (time.perf_counter()-request)*1000
        exact_same = output["oracle_zero_difference"] and output["oracle_zero_conflict"]
        output.update(request_to_final_exact_ms=round(elapsed, 3), final_state="EXACT_SAME" if exact_same else "EXACT_CHANGED",
                      status="PASS" if exact_same and elapsed <= args.gate*1000 else "TIMEOUT" if elapsed > args.gate*1000 else "ORACLE_FAILURE")
        revisit = time.perf_counter(); _ = output["final_state"]; output["revisit_ms"] = round((time.perf_counter()-revisit)*1000, 3)
    except Exception as exc:
        output["error"] = f"{type(exc).__name__}: {exc}"
    finally:
        after = process.memory_info().rss
        output.update(rss_before_bytes=before, rss_after_bytes=after, rss_delta_bytes=after-before, rss_peak_bytes=after,
                      startup_and_request_ms=round((time.perf_counter()-started)*1000, 3))
    return output


def _kill_process_tree(child):
    """Kill a timed-out worker and all descendants, then reap every process."""
    processes = []
    try:
        root = psutil.Process(child.pid)
        processes = root.children(recursive=True)
    except (psutil.Error, OSError):
        root = None
    for process in reversed(processes):
        try:
            process.kill()
        except (psutil.Error, OSError):
            pass
    try:
        child.kill()
    except OSError:
        pass
    targets = processes + ([root] if root is not None else [])
    _gone, alive = psutil.wait_procs(targets, timeout=5.0)
    return "killed-tree-and-waited" if not alive else "killed-tree-pending"


def sheet_child(args, mine, theirs, base, sheet):
    command = [sys.executable, str(Path(__file__).resolve()), "--sheet-worker", "--source", args.source, "--mode", args.mode, "--sheet", sheet,
               "--mine", str(mine), "--theirs", str(theirs), "--gate", str(args.gate)] + (["--base", str(base)] if base else [])
    child = subprocess.Popen(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
    try: stdout, stderr = child.communicate(timeout=args.gate + args.process_grace)
    except subprocess.TimeoutExpired as exc:
        cleanup = _kill_process_tree(child); stdout, stderr = child.communicate()
        return {"schema": EXACT_SCHEMA, "mode": args.mode, "source_path": args.source, "sheet": sheet, "status": "PROCESS_TIMEOUT", "final_state": "FAILED", "engine": "snapshot-candidate+legacy-fallback", "fallback_reason": "process-watchdog", "request_to_final_exact_ms": None, "oracle_zero_difference": False, "oracle_zero_conflict": False, "child_pid": child.pid, "child_cleanup": cleanup, "error": str(exc)}
    lines = [line for line in stdout.splitlines() if line.strip()]
    if child.returncode or not lines:
        return {"schema": EXACT_SCHEMA, "mode": args.mode, "source_path": args.source, "sheet": sheet, "status": "WORKER_FAILURE", "final_state": "FAILED", "engine": "snapshot-candidate+legacy-fallback", "request_to_final_exact_ms": None, "oracle_zero_difference": False, "oracle_zero_conflict": False, "child_pid": child.pid, "child_cleanup": "waited", "error": (stderr or stdout or f"exit {child.returncode}")[-4000:]}
    data = json.loads(lines[-1]); data.update(child_pid=child.pid, child_cleanup="waited"); return data


def file_worker(args):
    started = time.perf_counter(); proc = psutil.Process(); before = proc.memory_info().rss
    result = {"schema": EXACT_SCHEMA, "source_path": args.source, "relative_path": args.relative_path, "mode": args.mode, "status": "FAILED", "sheets": [], "error": None, "rss_before_bytes": before}
    try:
        Path(args.jsonl).parent.mkdir(parents=True, exist_ok=True)
        import sow_merge_tool
        imported = time.perf_counter(); catalog = time.perf_counter()
        book = load_workbook(args.source, read_only=True, data_only=False, keep_vba=Path(args.source).suffix.casefold()==".xlsm")
        try: sheets = list(book.sheetnames)
        finally: book.close()
        result.update(startup_ms=round((imported-started)*1000,3), catalog_ms=round((time.perf_counter()-catalog)*1000,3))
        with tempfile.TemporaryDirectory(prefix="sow_corpus_exact_file_") as temporary:
            mine, theirs, base = copy_sides(args.source, temporary, args.mode); result["disposable_sides"] = True
            for sheet in sheets:
                row = sheet_child(args, mine, theirs, base, sheet); result["sheets"].append(row)
                append_jsonl(args.jsonl, {"event":"sheet-complete", "relative_path":args.relative_path, "mode":args.mode, "sheet_result":row})
        result["status"] = "PASS" if all(row["status"]=="PASS" for row in result["sheets"]) else "FAIL"
    except Exception as exc: result["error"] = f"{type(exc).__name__}: {exc}"
    finally:
        after=proc.memory_info().rss; result.update(whole_summary_ms=round((time.perf_counter()-started)*1000,3), rss_after_bytes=after, rss_delta_bytes=after-before, rss_peak_bytes=after)
    return result


def file_child(source, relative, mode, args):
    command=[sys.executable,str(Path(__file__).resolve()),"--file-worker","--source",str(source),"--relative-path",relative,"--mode",mode,"--gate",str(args.gate),"--process-grace",str(args.process_grace),"--jsonl",str(args.jsonl)]
    child = subprocess.Popen(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
    try:
        stdout, stderr = child.communicate(timeout=args.file_process_timeout)
    except subprocess.TimeoutExpired as exc:
        cleanup = _kill_process_tree(child); stdout, stderr = child.communicate()
        return {"schema":EXACT_SCHEMA,"source_path":str(source),"relative_path":relative,"mode":mode,"status":"FILE_PROCESS_TIMEOUT","sheets":[],"child_pid":child.pid,"child_cleanup":cleanup,"error":str(exc)}
    lines=[line for line in stdout.splitlines() if line.strip()]
    return json.loads(lines[-1]) if not child.returncode and lines else {"schema":EXACT_SCHEMA,"source_path":str(source),"relative_path":relative,"mode":mode,"status":"FILE_WORKER_FAILURE","sheets":[],"child_pid":child.pid,"child_cleanup":"waited","error":(stderr or stdout or f"exit {child.returncode}")[-4000:]}


def controller(args):
    root,out=Path(args.source_root),Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    Path(args.jsonl).parent.mkdir(parents=True, exist_ok=True)
    if args.resume and out.exists(): evidence=json.loads(out.read_text(encoding="utf-8")); evidence.setdefault("runs",[])
    else: evidence={"schema":EXACT_SCHEMA,"created_at_epoch":time.time(),"gate_seconds":args.gate,"fresh_process":True,"provenance":_run_provenance(),"inventory_before":inventory(root),"runs":[],"source_signature_before":{},"source_signature_after":{},"slowest_sheets":{"2way":[],"3way":[]}}
    done={(item.get("relative_path"),item.get("mode")) for item in evidence["runs"]} if args.resume else set(); supported=[item for item in evidence["inventory_before"]["files"] if item["classification"]=="supported"]
    if args.max_files is not None: supported=supported[:args.max_files]
    for item in supported:
        source,relative=Path(item["source_path"]),item["relative_path"]
        if relative not in evidence["source_signature_before"]: evidence["source_signature_before"][relative]=signature(source)
        for mode in ("2way","3way"):
            if (relative,mode) in done: continue
            run=file_child(source,relative,mode,args); evidence["runs"].append(run); evidence["slowest_sheets"]={name:rankings(evidence["runs"],name) for name in ("2way","3way")}
            append_jsonl(args.jsonl,{"event":"file-mode-checkpoint","relative_path":relative,"mode":mode,"status":run.get("status")}); write_json(out,evidence)
    for item in supported: evidence["source_signature_after"][item["relative_path"]]=signature(Path(item["source_path"]))
    evidence.update(inventory_after=inventory(root),source_unchanged=evidence["source_signature_before"]==evidence["source_signature_after"],slowest_sheets={name:rankings(evidence["runs"],name) for name in ("2way","3way")})
    write_json(out,evidence); write_report(out.with_name(out.stem+"_slowest.md"),evidence); failed=[row for run in evidence["runs"] for row in run.get("sheets",[]) if row.get("status")!="PASS"]
    print(json.dumps({"out":str(out),"runs":len(evidence["runs"]),"sheet_failures":len(failed),"source_unchanged":evidence["source_unchanged"]},ensure_ascii=False)); return 0 if not failed and evidence["source_unchanged"] else 1


def main():
    parser=argparse.ArgumentParser(); parser.add_argument("--source-root",default=str(ROOT)); parser.add_argument("--out",default="large_sheet_corpus_exact.json"); parser.add_argument("--jsonl",default="large_sheet_corpus_exact.jsonl"); parser.add_argument("--gate",type=float,default=15.0); parser.add_argument("--process-grace",type=float,default=30.0); parser.add_argument("--file-process-timeout",type=float,default=900.0); parser.add_argument("--resume",action="store_true"); parser.add_argument("--max-files",type=int); parser.add_argument("--file-worker",action="store_true"); parser.add_argument("--sheet-worker",action="store_true"); parser.add_argument("--source"); parser.add_argument("--relative-path",default=""); parser.add_argument("--mode",choices=("2way","3way")); parser.add_argument("--sheet"); parser.add_argument("--mine"); parser.add_argument("--theirs"); parser.add_argument("--base")
    args=parser.parse_args()
    if args.sheet_worker: print(json.dumps(sheet_worker(args),ensure_ascii=False,sort_keys=True)); return
    if args.file_worker: print(json.dumps(file_worker(args),ensure_ascii=False,sort_keys=True)); return
    raise SystemExit(controller(args))


if __name__=="__main__": main()

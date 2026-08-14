"""OpenSpec 5.1/5.3/5.4/6.1 source-delta merge regressions."""

from __future__ import annotations

import os
import shutil
import sqlite3
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

import sow_merge_tool as smt
from _test_temp_utils import make_temp_dir


def _book(path: str, rows: list[list[object]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    workbook.close()


def _copy(source: str, destination: str) -> None:
    shutil.copy2(source, destination)


def _replace_zip_member(path: str, member: str, transform) -> None:
    replacement = path + ".repacked"
    with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(
        replacement, "w", compression=zipfile.ZIP_DEFLATED
    ) as destination:
        for info in source.infolist():
            payload = source.read(info.filename)
            if info.filename == member:
                payload = transform(payload)
            destination.writestr(info, payload)
    os.replace(replacement, path)


def _materialize_shared_strings(path: str, *, reverse_table: bool = False) -> None:
    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    package_ns = "http://schemas.openxmlformats.org/package/2006/relationships"
    office_rel_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    content_ns = "http://schemas.openxmlformats.org/package/2006/content-types"
    with zipfile.ZipFile(path, "r") as source:
        payloads = {info.filename: source.read(info.filename) for info in source.infolist()}
    worksheet = ET.fromstring(payloads["xl/worksheets/sheet1.xml"])
    strings = []
    for cell in worksheet.findall(f".//{{{main_ns}}}c"):
        if cell.attrib.get("t") != "inlineStr":
            continue
        text = "".join(node.text or "" for node in cell.findall(f".//{{{main_ns}}}t"))
        strings.append(text)
    unique = list(dict.fromkeys(strings))
    if reverse_table:
        unique.reverse()
    indexes = {value: index for index, value in enumerate(unique)}
    for cell in worksheet.findall(f".//{{{main_ns}}}c"):
        if cell.attrib.get("t") != "inlineStr":
            continue
        text = "".join(node.text or "" for node in cell.findall(f".//{{{main_ns}}}t"))
        for child in list(cell):
            cell.remove(child)
        cell.attrib["t"] = "s"
        value = ET.SubElement(cell, f"{{{main_ns}}}v")
        value.text = str(indexes[text])
    payloads["xl/worksheets/sheet1.xml"] = ET.tostring(worksheet, encoding="utf-8", xml_declaration=True)
    shared = ET.Element(f"{{{main_ns}}}sst", {"count": str(len(strings)), "uniqueCount": str(len(unique))})
    for value in unique:
        item = ET.SubElement(shared, f"{{{main_ns}}}si")
        text = ET.SubElement(item, f"{{{main_ns}}}t")
        text.text = value
    payloads["xl/sharedStrings.xml"] = ET.tostring(shared, encoding="utf-8", xml_declaration=True)
    rels = ET.fromstring(payloads["xl/_rels/workbook.xml.rels"])
    ET.SubElement(rels, f"{{{package_ns}}}Relationship", {
        "Id": "rIdSharedStrings",
        "Type": f"{office_rel_ns}/sharedStrings",
        "Target": "sharedStrings.xml",
    })
    payloads["xl/_rels/workbook.xml.rels"] = ET.tostring(rels, encoding="utf-8", xml_declaration=True)
    content_types = ET.fromstring(payloads["[Content_Types].xml"])
    ET.SubElement(content_types, f"{{{content_ns}}}Override", {
        "PartName": "/xl/sharedStrings.xml",
        "ContentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml",
    })
    payloads["[Content_Types].xml"] = ET.tostring(content_types, encoding="utf-8", xml_declaration=True)
    replacement = path + ".shared-strings"
    with zipfile.ZipFile(replacement, "w", compression=zipfile.ZIP_DEFLATED) as destination:
        for name, payload in payloads.items():
            destination.writestr(name, payload)
    os.replace(replacement, path)


def _book_sheets(path: str, sheets: dict[str, list[list[object]]]) -> None:
    workbook = Workbook()
    # These fixtures represent revisions of one logical workbook. Keep the
    # generated package timestamps stable across wall-clock boundaries.
    fixed_time = datetime(2020, 1, 1, 0, 0, 0)
    workbook.properties.created = fixed_time
    workbook.properties.modified = fixed_time
    first = True
    for name, rows in sheets.items():
        sheet = workbook.active if first else workbook.create_sheet()
        first = False
        sheet.title = name
        for row in rows:
            sheet.append(row)
    workbook.save(path)
    workbook.close()


def _cell(path: str, coordinate: str):
    workbook = load_workbook(path, data_only=False)
    try:
        return workbook["Data"][coordinate].value
    finally:
        workbook.close()


def _delta(before: str, target: str, after: str):
    conflicts, candidate, _conflict_map, summary, reason = smt._cross_branch_source_delta_premerge(
        before, target, after
    )
    assert candidate and os.path.isfile(candidate), (candidate, summary, reason)
    assert summary["incoming_count"] == (
        summary["applied_count"]
        + summary["already_present_count"]
        + summary["unresolved_count"]
    ), (summary, conflicts, reason)
    return conflicts, candidate, summary, reason


def test_wc_conflict_fixture_uses_exact_source_nodes() -> None:
    root = make_temp_dir("sow_wc_conflict_fixture_")
    os.makedirs(os.path.join(root, ".svn"), exist_ok=True)
    db_path = os.path.join(root, ".svn", "wc.db")
    chinese_path = "sheets/开发/楼宇.xlsx"
    path_bytes = chinese_path.encode("utf-8")
    uuid = b"509b88cb-e3bb-49fc-85e7-49e888d66b00"
    fixture = (
        b"((merge ((subversion http://svn.example/SOW "
        + str(len(uuid)).encode()
        + b" "
        + uuid
        + b" "
        + str(len(path_bytes)).encode()
        + b" "
        + path_bytes
        + b" 5 37073 file) (subversion http://svn.example/SOW "
        + str(len(uuid)).encode()
        + b" "
        + uuid
        + b" "
        + str(len(path_bytes)).encode()
        + b" "
        + path_bytes
        + b" 5 37074 file))) ((text (ignored))))"
    )
    with sqlite3.connect(db_path) as connection:
        connection.execute(
            "create table ACTUAL_NODE (wc_id integer, local_relpath text, conflict_data blob)"
        )
        connection.execute(
            "insert into ACTUAL_NODE values (?, ?, ?)",
            (1, "sheets/release/楼宇.xlsx", fixture),
        )
    sources, reason = smt._read_cross_branch_sources_from_wc(
        root, "sheets/release/楼宇.xlsx"
    )
    assert reason == "wc-conflict-data"
    assert sources == [(chinese_path, 37073), (chinese_path, 37074)]

    malformed = fixture.replace(b"37074", b"37073")
    assert smt._extract_svn_conflict_source_nodes(malformed) == [
        (chinese_path, 37073),
        (chinese_path, 37073),
    ]


def test_source_delta_cell_states_and_target_retention() -> None:
    root = make_temp_dir("sow_source_delta_cells_")
    before = os.path.join(root, "before.xlsx")
    after = os.path.join(root, "after.xlsx")
    _book(before, [["id", "v"], ["a", 1], ["b", 2]])
    _book(after, [["id", "v"], ["a", 9], ["b", 2]])

    target_apply = os.path.join(root, "target-apply.xlsx")
    _copy(before, target_apply)
    _conflicts, candidate, summary, reason = _delta(before, target_apply, after)
    assert reason is None and summary == {
        "incoming_count": 1, "applied_count": 1, "already_present_count": 0,
        "target_retained_count": 0, "unresolved_count": 0, "merged_count": 1,
        "unresolved_structural_count": 0,
    }
    assert _cell(candidate, "B2") == 9

    target_already = os.path.join(root, "target-already.xlsx")
    _copy(after, target_already)
    _conflicts, candidate, summary, reason = _delta(before, target_already, after)
    assert reason is None
    assert summary["incoming_count"] == 1 and summary["already_present_count"] == 1
    assert summary["applied_count"] == summary["unresolved_count"] == 0
    assert smt._sha256_file(candidate) == smt._sha256_file(target_already)

    target_third = os.path.join(root, "target-third.xlsx")
    _book(target_third, [["id", "v"], ["a", 7], ["b", 2]])
    _conflicts, candidate, summary, reason = _delta(before, target_third, after)
    assert reason is None and summary["unresolved_count"] == 1
    assert summary["applied_count"] == summary["already_present_count"] == 0
    assert smt._sha256_file(candidate) == smt._sha256_file(target_third)

    target_retained = os.path.join(root, "target-retained.xlsx")
    _book(target_retained, [["id", "v"], ["a", 1], ["b-target-only", 2]])
    _conflicts, candidate, summary, reason = _delta(before, target_retained, after)
    assert reason is None and summary["applied_count"] == 1
    assert summary["target_retained_count"] >= 1
    assert _cell(candidate, "A3") == "b-target-only"
    assert _cell(candidate, "B2") == 9


def test_target_only_structure_is_retained_and_mapped_cell_delta_applies() -> None:
    root = make_temp_dir("sow_source_delta_target_structure_")
    before = os.path.join(root, "before.xlsx")
    after = os.path.join(root, "after.xlsx")
    _book(before, [["id", "v"], ["a", 1], ["b", 2]])
    _book(after, [["id", "v"], ["a", 9], ["b", 2]])

    target_column = os.path.join(root, "target-column.xlsx")
    _book(target_column, [["id", "target-only", "v"], ["a", "x", 1], ["b", "y", 2]])
    _conflicts, candidate, summary, reason = _delta(before, target_column, after)
    assert reason is None and summary["incoming_count"] == summary["applied_count"] == 1
    assert summary["target_retained_count"] == 1
    workbook = load_workbook(candidate, data_only=False)
    try:
        rows = list(workbook["Data"].iter_rows(values_only=True))
        assert rows == [("id", "target-only", "v"), ("a", "x", 9), ("b", "y", 2)]
    finally:
        workbook.close()

    unchanged_after = os.path.join(root, "after-unchanged.xlsx")
    _copy(before, unchanged_after)
    target_row = os.path.join(root, "target-row.xlsx")
    _book(target_row, [["id", "v"], ["a", 1], ["target-only", 8], ["b", 2]])
    _conflicts, candidate, summary, reason = _delta(before, target_row, unchanged_after)
    assert reason is None
    assert summary["incoming_count"] == summary["applied_count"] == 0
    assert summary["target_retained_count"] == 1
    assert smt._sha256_file(candidate) == smt._sha256_file(target_row)


def test_source_row_and_column_structure_stay_manual() -> None:
    root = make_temp_dir("sow_source_delta_manual_structure_")
    before = os.path.join(root, "before.xlsx")
    target = os.path.join(root, "target.xlsx")
    _book(before, [["id", "v"], ["a", 1], ["b", 2]])
    _copy(before, target)

    row_after = os.path.join(root, "row-after.xlsx")
    _book(row_after, [["id", "v"], ["a", 1], ["new", 8], ["b", 2]])
    _conflicts, candidate, summary, reason = _delta(before, target, row_after)
    assert reason is None and summary["incoming_count"] == summary["unresolved_count"] == 1
    assert summary["applied_count"] == summary["already_present_count"] == 0
    assert smt._sha256_file(candidate) == smt._sha256_file(target)

    column_after = os.path.join(root, "column-after.xlsx")
    _book(column_after, [["id", "new", "v"], ["a", "x", 1], ["b", "y", 2]])
    _conflicts, candidate, summary, reason = _delta(before, target, column_after)
    assert reason is None and summary["incoming_count"] == summary["unresolved_count"] == 1
    assert summary["applied_count"] == summary["already_present_count"] == 0
    assert smt._sha256_file(candidate) == smt._sha256_file(target)


def test_same_shape_target_row_replacement_is_not_projected_by_position() -> None:
    root = make_temp_dir("sow_source_delta_same_shape_row_")
    before = os.path.join(root, "before.xlsx")
    after = os.path.join(root, "after.xlsx")
    target = os.path.join(root, "target.xlsx")
    _book(before, [["id", "v"], ["a", 1], ["b", 2]])
    _book(after, [["id", "v"], ["a", 9], ["b", 2]])
    # The first target record was replaced in place.  Equal dimensions must
    # not make it look like source record "a".
    _book(target, [["id", "v"], ["x", 1], ["b", 2]])
    _conflicts, candidate, summary, reason = _delta(before, target, after)
    assert reason is None
    assert summary["incoming_count"] == summary["unresolved_count"] == 1
    assert summary["applied_count"] == summary["already_present_count"] == 0
    assert _cell(candidate, "A2") == "x" and _cell(candidate, "B2") == 1
    assert smt._sha256_file(candidate) == smt._sha256_file(target)


def test_same_width_source_column_replacement_stays_manual() -> None:
    root = make_temp_dir("sow_source_delta_same_width_column_")
    before = os.path.join(root, "before.xlsx")
    after = os.path.join(root, "after.xlsx")
    target = os.path.join(root, "target.xlsx")
    _book(before, [["id", "A", "B"], ["a", 1, 2], ["b", 3, 4]])
    # Source column A was replaced by X without changing physical width.
    _book(after, [["id", "X", "B"], ["a", 9, 2], ["b", 3, 4]])
    _copy(before, target)
    _conflicts, candidate, summary, reason = _delta(before, target, after)
    assert reason is None
    assert summary["incoming_count"] == summary["unresolved_count"] == 1
    assert summary["applied_count"] == summary["already_present_count"] == 0
    assert _cell(candidate, "B1") == "A" and _cell(candidate, "B2") == 1
    assert smt._sha256_file(candidate) == smt._sha256_file(target)


def test_unrepresentable_source_changes_are_counted_as_incoming_unresolved() -> None:
    root = make_temp_dir("sow_source_delta_unrepresentable_")
    before = os.path.join(root, "before.xlsx")
    target = os.path.join(root, "target.xlsx")
    _book(before, [["id", "v"], ["a", 1]])
    _copy(before, target)

    style_after = os.path.join(root, "style-after.xlsx")
    _copy(before, style_after)
    workbook = load_workbook(style_after)
    workbook["Data"]["B2"].fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    workbook.save(style_after)
    workbook.close()
    _conflicts, candidate, summary, reason = _delta(before, target, style_after)
    assert reason is not None
    assert summary["incoming_count"] == summary["unresolved_count"] == 1
    assert smt._sha256_file(candidate) == smt._sha256_file(target)

    sheet_added = os.path.join(root, "sheet-added.xlsx")
    _book_sheets(sheet_added, {
        "Data": [["id", "v"], ["a", 1]],
        "Added": [["id"], ["new"]],
    })
    _conflicts, candidate, summary, reason = _delta(before, target, sheet_added)
    assert reason is not None
    assert summary["incoming_count"] == summary["unresolved_count"] == 1
    assert smt._sha256_file(candidate) == smt._sha256_file(target)

    before_with_deleted_sheet = os.path.join(root, "before-with-deleted-sheet.xlsx")
    target_with_deleted_sheet = os.path.join(root, "target-with-deleted-sheet.xlsx")
    _book_sheets(before_with_deleted_sheet, {
        "Data": [["id", "v"], ["a", 1]],
        "Deleted": [["id"], ["old"]],
    })
    _copy(before_with_deleted_sheet, target_with_deleted_sheet)
    _conflicts, candidate, summary, reason = _delta(
        before_with_deleted_sheet, target_with_deleted_sheet, before
    )
    assert reason is not None
    assert summary["incoming_count"] == summary["unresolved_count"] == 1
    assert smt._sha256_file(candidate) == smt._sha256_file(target_with_deleted_sheet)


def test_cell_delta_does_not_hide_sheet_document_or_vba_package_changes() -> None:
    root = make_temp_dir("sow_source_delta_package_audit_")
    before = os.path.join(root, "before.xlsx")
    after = os.path.join(root, "after.xlsx")
    target = os.path.join(root, "target.xlsx")
    _book(before, [["id", "v"], ["a", 1]])
    workbook = load_workbook(before)
    workbook.properties.title = "before-title"
    workbook["Data"].sheet_properties.tabColor = "FF0000"
    workbook.save(before)
    workbook.close()
    _copy(before, after)
    _copy(before, target)
    workbook = load_workbook(after)
    workbook["Data"]["B2"] = 9
    workbook["Data"].sheet_properties.tabColor = "00FF00"
    workbook.properties.title = "after-title"
    workbook.save(after)
    workbook.close()

    _conflicts, candidate, summary, reason = _delta(before, target, after)
    assert reason is not None and "OOXML" in reason
    # One value delta plus the unrepresentable package change must be manual;
    # no candidate may pretend the tab color/title merged successfully.
    assert summary["incoming_count"] == summary["unresolved_count"] >= 2
    assert summary["applied_count"] == summary["already_present_count"] == 0
    assert _cell(candidate, "B2") == 1
    candidate_book = load_workbook(candidate)
    try:
        assert candidate_book.properties.title == "before-title"
        assert candidate_book["Data"].sheet_properties.tabColor.rgb == "00FF0000"
    finally:
        candidate_book.close()

    vba_after = os.path.join(root, "vba-after.xlsx")
    _copy(before, vba_after)
    workbook = load_workbook(vba_after)
    workbook["Data"]["B2"] = 9
    workbook.save(vba_after)
    workbook.close()
    with zipfile.ZipFile(vba_after, "a", compression=zipfile.ZIP_DEFLATED) as package:
        package.writestr("xl/vbaProject.bin", b"source-after-vba-change")
    _conflicts, candidate, summary, reason = _delta(before, target, vba_after)
    assert reason is not None and "vbaProject.bin" in reason
    assert summary["incoming_count"] == summary["unresolved_count"] >= 2
    assert summary["applied_count"] == summary["already_present_count"] == 0
    assert _cell(candidate, "B2") == 1


def test_representable_blank_string_and_formula_cache_cell_deltas_pass_audit() -> None:
    root = make_temp_dir("sow_source_delta_payload_audit_")

    blank_before = os.path.join(root, "blank-before.xlsx")
    blank_after = os.path.join(root, "blank-after.xlsx")
    blank_target = os.path.join(root, "blank-target.xlsx")
    _book(blank_before, [["id", "v"], ["a", None]])
    _book(blank_after, [["id", "v"], ["a", 5]])
    _copy(blank_before, blank_target)
    _conflicts, candidate, summary, reason = _delta(blank_before, blank_target, blank_after)
    assert reason is None and summary["incoming_count"] == summary["applied_count"] == 1
    assert _cell(candidate, "B2") == 5

    cleared_after = os.path.join(root, "blank-cleared-after.xlsx")
    _copy(blank_after, cleared_after)
    workbook = load_workbook(cleared_after)
    workbook["Data"]["B2"] = None
    workbook.save(cleared_after)
    workbook.close()
    cleared_target = os.path.join(root, "blank-cleared-target.xlsx")
    _copy(blank_after, cleared_target)
    _conflicts, candidate, summary, reason = _delta(blank_after, cleared_target, cleared_after)
    assert reason is None and summary["incoming_count"] == summary["applied_count"] == 1
    assert _cell(candidate, "B2") is None

    string_before = os.path.join(root, "string-before.xlsx")
    string_after = os.path.join(root, "string-after.xlsx")
    string_target = os.path.join(root, "string-target.xlsx")
    _book(string_before, [["id", "v"], ["a", "old"], ["b", "old"]])
    _book(string_after, [["id", "v"], ["a", "new"], ["b", "old"]])
    _copy(string_before, string_target)
    # Use true shared-string tables and intentionally reverse Source After's
    # table order: unchanged "old" must compare by resolved text, not by its
    # package-local index.
    _materialize_shared_strings(string_before)
    _materialize_shared_strings(string_after, reverse_table=True)
    _materialize_shared_strings(string_target)
    _conflicts, candidate, summary, reason = _delta(string_before, string_target, string_after)
    assert reason is None and summary["incoming_count"] == summary["applied_count"] == 1
    assert _cell(candidate, "B2") == "new" and _cell(candidate, "B3") == "old"

    formula_before = os.path.join(root, "formula-before.xlsx")
    formula_after = os.path.join(root, "formula-after.xlsx")
    formula_target = os.path.join(root, "formula-target.xlsx")
    _book(formula_before, [["id", "v", "calc"], ["a", 1, "=B2*2"]])
    _copy(formula_before, formula_after)
    _copy(formula_before, formula_target)
    workbook = load_workbook(formula_after)
    workbook["Data"]["B2"] = 2
    workbook.save(formula_after)
    workbook.close()
    # Simulate Excel updating only an unchanged formula's cached <v> value.
    _replace_zip_member(
        formula_after,
        "xl/worksheets/sheet1.xml",
        lambda payload: payload.replace(b"<v></v>", b"<v>4</v>", 1),
    )
    _conflicts, candidate, summary, reason = _delta(formula_before, formula_target, formula_after)
    assert reason is None and summary["incoming_count"] == summary["applied_count"] == 1
    assert _cell(candidate, "B2") == 2


def test_target_only_and_deleted_sheets_are_retained_without_incoming() -> None:
    root = make_temp_dir("sow_source_delta_target_sheet_structure_")
    before = os.path.join(root, "before.xlsx")
    after = os.path.join(root, "after.xlsx")
    _book_sheets(before, {
        "Data": [["id", "v"], ["a", 1]],
        "Stable": [["id", "v"], ["s", 2]],
    })
    _copy(before, after)

    target_extra = os.path.join(root, "target-extra.xlsx")
    _book_sheets(target_extra, {
        "Data": [["id", "v"], ["a", 1]],
        "Stable": [["id", "v"], ["s", 2]],
        "TargetOnly": [["id"], ["keep"]],
    })
    _conflicts, candidate, summary, reason = _delta(before, target_extra, after)
    assert reason is None and summary["incoming_count"] == 0
    assert summary["target_retained_count"] >= 1
    assert smt._sha256_file(candidate) == smt._sha256_file(target_extra)

    target_missing = os.path.join(root, "target-missing.xlsx")
    _book_sheets(target_missing, {"Data": [["id", "v"], ["a", 1]]})
    _conflicts, candidate, summary, reason = _delta(before, target_missing, after)
    assert reason is None and summary["incoming_count"] == 0
    assert summary["target_retained_count"] >= 1
    assert smt._sha256_file(candidate) == smt._sha256_file(target_missing)

    changed_after = os.path.join(root, "after-missing-target-sheet.xlsx")
    _book_sheets(changed_after, {
        "Data": [["id", "v"], ["a", 1]],
        "Stable": [["id", "v"], ["s", 9]],
    })
    _conflicts, candidate, summary, reason = _delta(before, target_missing, changed_after)
    assert reason is None
    assert summary["incoming_count"] == summary["unresolved_count"] == 1
    assert summary["target_retained_count"] >= 1
    assert smt._sha256_file(candidate) == smt._sha256_file(target_missing)


def test_cross_branch_exact_source_before_and_after_still_count_delta() -> None:
    root = make_temp_dir("sow_source_delta_exact_startup_")
    seed = os.path.join(root, "seed.xlsx")
    after_seed = os.path.join(root, "after-seed.xlsx")
    _book(seed, [["id", "v"], ["a", 1], ["b", 2]])
    _book(after_seed, [["id", "v"], ["a", 9], ["b", 2]])
    before = os.path.join(root, "Building.xlsx.merge-left.r37073")
    after = os.path.join(root, "Building.xlsx.merge-right.r37074")
    _copy(seed, before)
    _copy(after_seed, after)

    target_before = os.path.join(root, "TargetBefore.xlsx")
    _copy(seed, target_before)
    applied = smt.run_startup_merge_analysis(
        smt.build_merge_launch_context(before, target_before, after, target_before)
    ).outcome
    assert applied.automatic_action == "cross-branch-source-delta"
    assert (
        applied.incoming_count,
        applied.applied_count,
        applied.already_present_count,
        applied.unresolved_count,
        applied.merged_count,
    ) == (1, 1, 0, 0, 1)
    assert _cell(applied.candidate_path, "B2") == 9

    target_after = os.path.join(root, "TargetAfter.xlsx")
    _copy(after_seed, target_after)
    already = smt.run_startup_merge_analysis(
        smt.build_merge_launch_context(before, target_after, after, target_after)
    ).outcome
    assert already.automatic_action == "incoming-already-present"
    assert (
        already.incoming_count,
        already.applied_count,
        already.already_present_count,
        already.unresolved_count,
        already.merged_count,
    ) == (1, 0, 1, 0, 0)
    assert smt._sha256_file(already.candidate_path) == smt._sha256_file(target_after)


def test_cross_branch_startup_already_present_and_update_regression() -> None:
    root = make_temp_dir("sow_source_delta_startup_")
    seed = os.path.join(root, "seed.xlsx")
    _book(seed, [["id", "v"], ["a", 1], ["b", 2]])
    before = os.path.join(root, "Building.xlsx.merge-left.r37073")
    after = os.path.join(root, "Building.xlsx.merge-right.r37074")
    target = os.path.join(root, "Building.xlsx")
    _copy(seed, before)
    _book(after + ".xlsx", [["id", "v"], ["a", 9], ["b", 2]])
    _copy(after + ".xlsx", after)
    _book(target, [["id", "v"], ["a", 9], ["b-target-only", 2]])
    context = smt.build_merge_launch_context(before, target, after, target)
    analysis = smt.run_startup_merge_analysis(context)
    outcome = analysis.outcome
    assert outcome.automatic_action == "incoming-already-present"
    assert (outcome.incoming_count, outcome.applied_count, outcome.already_present_count, outcome.unresolved_count) == (1, 0, 1, 0)
    assert outcome.merged_count == 0 and outcome.target_retained_count >= 1
    assert smt._sha256_file(outcome.candidate_path) == smt._sha256_file(target)

    update_before = os.path.join(root, "Update.xlsx.r1")
    update_after = os.path.join(root, "Update.xlsx.r2")
    update_target = os.path.join(root, "Update.xlsx")
    _copy(seed, update_before)
    _copy(after + ".xlsx", update_after)
    _copy(seed, update_target)
    update = smt.run_startup_merge_analysis(
        smt.build_merge_launch_context(update_before, update_target, update_after, update_target)
    ).outcome
    assert update.automatic_action == "initialize-from-theirs"
    assert update.incoming_count == update.applied_count == update.already_present_count == 0
    assert _cell(update.candidate_path, "B2") == 9


def main() -> None:
    tests = [
        test_wc_conflict_fixture_uses_exact_source_nodes,
        test_source_delta_cell_states_and_target_retention,
        test_target_only_structure_is_retained_and_mapped_cell_delta_applies,
        test_source_row_and_column_structure_stay_manual,
        test_same_shape_target_row_replacement_is_not_projected_by_position,
        test_same_width_source_column_replacement_stays_manual,
        test_unrepresentable_source_changes_are_counted_as_incoming_unresolved,
        test_cell_delta_does_not_hide_sheet_document_or_vba_package_changes,
        test_representable_blank_string_and_formula_cache_cell_deltas_pass_audit,
        test_target_only_and_deleted_sheets_are_retained_without_incoming,
        test_cross_branch_exact_source_before_and_after_still_count_delta,
        test_cross_branch_startup_already_present_and_update_regression,
    ]
    for test in tests:
        test()
        print(f"PASS {test.__name__}")


if __name__ == "__main__":
    main()

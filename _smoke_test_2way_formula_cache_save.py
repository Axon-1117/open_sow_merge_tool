"""Two-way formula-cache copy, atomic save, and reopen regression."""

from __future__ import annotations

import hashlib
import json
import os
import sys
import tempfile
import time
from contextlib import contextmanager
from types import SimpleNamespace

from openpyxl import Workbook, load_workbook

import sow_merge_tool as mod


def _sha256(path: str) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _make_formula_book(path: str, cached_value: int):
    raw = path + ".raw.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "S1"
    worksheet.append(["id@id", "formula"])
    worksheet.append(["string", "formula"])
    worksheet.append(["row-1", "=1"])
    workbook.save(raw)
    workbook.close()
    mod._build_manual_merge_xlsx_via_zip(
        raw,
        path,
        {("S1", 3, 2): "=1"},
        cached_values={("S1", 3, 2): cached_value},
    )


@contextmanager
def _owned_case():
    original_settings_path = mod._SETTINGS_PATH
    original_settings_exists = os.path.lexists(original_settings_path)
    if original_settings_exists:
        with open(original_settings_path, "rb") as stream:
            original_settings_bytes = stream.read()
    else:
        original_settings_bytes = None
    temporary = tempfile.TemporaryDirectory(prefix="sow_2way_formula_cache_")
    root_dir = temporary.name
    settings_path = os.path.join(root_dir, "settings.json")
    with open(settings_path, "w", encoding="utf-8") as stream:
        json.dump({"only_diff": 0}, stream)
    original_prompt_scheduler = mod.SowMergeApp._schedule_formula_cache_prompt
    state = SimpleNamespace(app=None, immutable_hashes={}, root_dir=root_dir)
    mod._SETTINGS_PATH = settings_path
    mod.SowMergeApp._schedule_formula_cache_prompt = lambda _self: None
    primary = None
    try:
        yield state
    except BaseException as exc:
        primary = exc
        raise
    finally:
        cleanup_errors = []
        if state.app is not None:
            try:
                for candidate in (state.app, *getattr(state.app, "sheet_views", {}).values()):
                    if candidate is None:
                        continue
                    after_id = getattr(candidate, "_settings_save_id", None)
                    if after_id is not None:
                        state.app.root.after_cancel(after_id)
                        candidate._settings_save_id = None
                state.app._shutdown_root()
            except BaseException as exc:
                cleanup_errors.append(f"shutdown: {exc!r}")
        try:
            mod.SowMergeApp._schedule_formula_cache_prompt = original_prompt_scheduler
            mod._SETTINGS_PATH = original_settings_path
            if original_settings_exists:
                with open(original_settings_path, "rb") as stream:
                    assert stream.read() == original_settings_bytes
            else:
                assert not os.path.lexists(original_settings_path)
        except BaseException as exc:
            cleanup_errors.append(f"settings/prompt restore: {exc!r}")
        for path, before_hash in state.immutable_hashes.items():
            try:
                assert _sha256(path) == before_hash, path
            except BaseException as exc:
                cleanup_errors.append(f"immutable input SHA {path!r}: {exc!r}")
        try:
            temporary.cleanup()
            assert not os.path.lexists(root_dir), root_dir
        except BaseException as exc:
            cleanup_errors.append(f"owned temporary root: {exc!r}")
        if cleanup_errors:
            message = "two-way formula-cache cleanup failed: " + "; ".join(cleanup_errors)
            if primary is not None:
                primary.add_note(message)
            else:
                raise AssertionError(message)


def _wait_operation_ready(app, sheet: str, deadline: float):
    view = None
    while time.monotonic() < deadline:
        app.root.update_idletasks()
        app.root.update()
        view = app.sheet_views.get(sheet)
        if view is not None and view._data_ready and app._is_sheet_exact_current(sheet) and app._edit_workbooks_ready():
            return view
        time.sleep(0.01)
    raise AssertionError(("view not operation-ready", app._sheet_exact_entry(sheet), view))


def main():
    deadline = time.monotonic() + 90.0
    with _owned_case() as owned:
        mine = os.path.join(owned.root_dir, "mine.xlsx")
        theirs = os.path.join(owned.root_dir, "theirs.xlsx")
        _make_formula_book(mine, 1)
        _make_formula_book(theirs, 2)
        mine_before = _sha256(mine)
        owned.immutable_hashes = {theirs: _sha256(theirs)}

        app = mod.SowMergeApp(mine, theirs)
        owned.app = app
        app.nb.select(app._sheet_containers["S1"])
        app._request_edit_preload()
        view = _wait_operation_ready(app, "S1", deadline)
        row_a = next(
            row for row in range(1, app.ws_a_val("S1").max_row + 1)
            if app.ws_a_val("S1").cell(row, 1).value == "row-1"
        )
        pair_idx = view.row_a_to_pair_idx[row_a]
        logical_col = view._logical_col_for_physical("A", 2)
        assert logical_col == 2 and view.row_pairs[pair_idx][0] == row_a
        view._show_formula_copy_skip_notice = lambda _count: None

        assert view._copy_selected_row("B2A", override_pair_idx=pair_idx, override_cols={logical_col})
        assert app.manual_a_formula_cache_ops[("S1", row_a, 2)] == 2

        app._atomic_save(app._wb_a_edit, mine)
        warning = app._post_save_refresh("A", mine)
        if warning is not None:
            assert "公式缓存刷新失败" in warning, warning
        assert not app.manual_a_formula_cache_ops
        assert _sha256(mine) != mine_before
        assert _sha256(theirs) == owned.immutable_hashes[theirs]

        formula_book = load_workbook(mine, data_only=False)
        value_book = load_workbook(mine, data_only=True)
        try:
            assert formula_book["S1"].cell(row_a, 2).value == "=1"
            assert value_book["S1"].cell(row_a, 2).value == 2
        finally:
            formula_book.close()
            value_book.close()

    print("SMOKE_2WAY_FORMULA_CACHE_SAVE_OK", flush=True)


if __name__ == "__main__":
    main()

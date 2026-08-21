"""Focused GUI regression for region-mode target resolution and interaction."""

from __future__ import annotations

import os
import argparse
import json
import sys
import tempfile
import time
from contextlib import contextmanager
from types import SimpleNamespace

from openpyxl import Workbook, load_workbook

import sow_merge_tool as smt
from _test_temp_utils import make_temp_dir, visible_render_text


def _sha256(path: str) -> str:
    import hashlib

    digest = hashlib.sha256()
    with open(path, "rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


@contextmanager
def _typed_region_case(prefix: str):
    """Own the fidelity-only typed region fixture and all of its side state."""
    original_settings_path = smt._SETTINGS_PATH
    original_settings_exists = os.path.lexists(original_settings_path)
    if original_settings_exists:
        with open(original_settings_path, "rb") as stream:
            original_settings_bytes = stream.read()
    else:
        original_settings_bytes = None
    temporary = tempfile.TemporaryDirectory(prefix=prefix)
    root_dir = temporary.name
    isolated_settings_path = os.path.join(root_dir, "settings.json")
    with open(isolated_settings_path, "w", encoding="utf-8") as stream:
        json.dump({"only_diff": 0}, stream)
    original_prompt_scheduler = smt.SowMergeApp._schedule_formula_cache_prompt
    state = SimpleNamespace(app=None, input_hashes={}, root_dir=root_dir)
    smt._SETTINGS_PATH = isolated_settings_path
    smt.SowMergeApp._schedule_formula_cache_prompt = lambda _self: None
    primary = None
    try:
        yield state
    except BaseException as exc:
        primary = exc
        raise
    finally:
        cleanup_errors = []
        if state.app is not None:
            try:
                for candidate in (state.app, *getattr(state.app, "sheet_views", {}).values()):
                    if candidate is None:
                        continue
                    after_id = getattr(candidate, "_settings_save_id", None)
                    if after_id is not None:
                        state.app.root.after_cancel(after_id)
                        candidate._settings_save_id = None
                state.app._shutdown_root()
            except BaseException as exc:
                cleanup_errors.append(f"shutdown: {exc!r}")
        try:
            smt.SowMergeApp._schedule_formula_cache_prompt = original_prompt_scheduler
            smt._SETTINGS_PATH = original_settings_path
            if original_settings_exists:
                with open(original_settings_path, "rb") as stream:
                    assert stream.read() == original_settings_bytes
            else:
                assert not os.path.lexists(original_settings_path)
        except BaseException as exc:
            cleanup_errors.append(f"settings/prompt restore: {exc!r}")
        for path, before_hash in state.input_hashes.items():
            try:
                assert _sha256(path) == before_hash, path
            except BaseException as exc:
                cleanup_errors.append(f"input SHA {path!r}: {exc!r}")
        try:
            temporary.cleanup()
            assert not os.path.lexists(root_dir), root_dir
        except BaseException as exc:
            cleanup_errors.append(f"owned temporary root: {exc!r}")
        if cleanup_errors:
            message = "typed region fixture cleanup failed: " + "; ".join(cleanup_errors)
            if primary is not None:
                primary.add_note(message)
            else:
                raise AssertionError(message)


def _make_book(path: str, rows) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    for row in rows:
        worksheet.append(list(row))
    workbook.save(path)
    workbook.close()


def _pump(root, seconds=0.08) -> None:
    deadline = time.time() + seconds
    while time.time() < deadline:
        root.update_idletasks()
        root.update()
        time.sleep(0.005)


def _request_edit_backend_for_completed_operation(app, sheet: str = "Data") -> None:
    """Test-only setup for cases that assert a completed mutation/undo path."""
    deadline = time.time() + 30.0
    while time.time() < deadline:
        # The initial catalog callback may publish the deferred-preload target
        # after the exact snapshot reaches terminal state. Retry the idempotent
        # demand signal until that callback is installed.
        app._request_edit_preload()
        _pump(app.root, 0.03)
        view = app.sheet_views.get(sheet)
        if (
            app._edit_workbooks_ready()
            and view is not None
            and view._derive_lifecycle_state() == "READY"
        ):
            return
    view = app.sheet_views.get(sheet)
    assert app._edit_workbooks_ready() and view is not None and view._derive_lifecycle_state() == "READY", (
        "editable backend did not become ready",
        app._sheet_exact_entry(sheet),
        getattr(view, "_lifecycle_state", None),
    )


def _open_view(mine_rows, theirs_rows, *, rescan=True, owned_case=None):
    root_dir = owned_case.root_dir if owned_case is not None else make_temp_dir("sow_region_mode_interaction_")
    mine = os.path.join(root_dir, "mine.xlsx")
    theirs = os.path.join(root_dir, "theirs.xlsx")
    _make_book(mine, mine_rows)
    _make_book(theirs, theirs_rows)
    if owned_case is not None:
        owned_case.input_hashes = {mine: _sha256(mine), theirs: _sha256(theirs)}
    app = smt.SowMergeApp(mine, theirs)
    if owned_case is not None:
        owned_case.app = app
    app.root.deiconify()
    app.root.geometry("900x760")
    app.nb.select(app._sheet_containers["Data"])
    deadline = time.time() + 15.0
    view = None
    while time.time() < deadline:
        _pump(app.root, 0.03)
        view = app.sheet_views.get("Data")
        if (
            view is not None
            and getattr(view, "_data_ready", False)
            and app._is_sheet_exact_current("Data")
        ):
            break
    assert view is not None and view._data_ready and app._is_sheet_exact_current("Data"), (
        "Data did not reach exact readiness", app._sheet_exact_entry("Data"),
    )
    # Do not leave the legacy stale-result test suppression enabled across a
    # fresh rescan: the current worker deliberately skips applying that cache
    # while suppression is set, which would keep this fixture CALCULATING.
    view._suppress_bg_apply = False
    view.only_diff_var.set(0)
    if rescan:
        view.refresh(row_only=None, rescan=True)
    deadline = time.time() + 15.0
    while time.time() < deadline and not app._is_sheet_exact_current("Data"):
        _pump(app.root, 0.03)
    assert app._is_sheet_exact_current("Data"), (
        "Data did not recover exact readiness after test rescan",
        app._sheet_exact_entry("Data"),
    )
    _pump(app.root)
    return app, view


def _open_three_way_view(mine_rows, base_rows, theirs_rows, *, rescan=True):
    root_dir = make_temp_dir("sow_region_mode_three_way_")
    mine = os.path.join(root_dir, "mine.xlsx")
    base = os.path.join(root_dir, "base.xlsx")
    theirs = os.path.join(root_dir, "theirs.xlsx")
    merged = os.path.join(root_dir, "merged.xlsx")
    _make_book(mine, mine_rows)
    _make_book(base, base_rows)
    _make_book(theirs, theirs_rows)
    app = smt.SowMergeApp(
        mine,
        theirs,
        merge_mode=True,
        merged_path=merged,
        base_path=base,
    )
    app.root.deiconify()
    app.root.geometry("1000x760")
    app.nb.select(app._sheet_containers["Data"])
    deadline = time.time() + 15.0
    view = None
    while time.time() < deadline:
        _pump(app.root, 0.03)
        view = app.sheet_views.get("Data")
        if (
            view is not None
            and getattr(view, "_data_ready", False)
            and app._is_sheet_exact_current("Data")
        ):
            break
    assert view is not None and view._data_ready and app._is_sheet_exact_current("Data"), (
        "Data did not reach exact readiness", app._sheet_exact_entry("Data"),
    )
    view._suppress_bg_apply = False
    view.only_diff_var.set(0)
    if rescan:
        view.refresh(row_only=None, rescan=True)
    deadline = time.time() + 15.0
    while time.time() < deadline and not app._is_sheet_exact_current("Data"):
        _pump(app.root, 0.03)
    assert app._is_sheet_exact_current("Data"), (
        "Data did not recover exact readiness after test rescan",
        app._sheet_exact_entry("Data"),
    )
    _pump(app.root)
    return app, view


def _ordinary_diff_rows(*, typed_schema=False):
    if typed_schema:
        # Fidelity-only declared records: stable identifiers and the matching
        # type row let the exact gate prove every row before a region action.
        mine = [
            ("id@id", "value"),
            ("string", "string"),
            *((f"id-{idx}", f"same-{idx}") for idx in range(1, 11)),
        ]
        theirs = [list(row) for row in mine]
        for identifier in ("id-2", "id-3", "id-7", "id-8"):
            logical_id = int(identifier.rsplit("-", 1)[1])
            row_index = next(
                index for index, row in enumerate(theirs) if row[0] == identifier
            )
            theirs[row_index][1] = f"changed-{logical_id}"
        return mine, theirs
    mine = [(idx, f"same-{idx}") for idx in range(1, 11)]
    theirs = [list(row) for row in mine]
    for excel_row in (2, 3, 7, 8):
        theirs[excel_row - 1][1] = f"changed-{excel_row}"
    return mine, theirs


def _block_pairs(block):
    return tuple(int(value) for value in block.pair_indices)


def _clear_region_anchor(view, *, insertion_pair_idx=None) -> None:
    view.clear_explicit_cell_selection()
    view.selected_pair_idx = None
    view.hover_pair_idx = None
    view.hover_col_idx = None
    view.hover_side = None
    view._last_cursor_cmp_pair_idx = None
    view._last_selected_line = None
    if insertion_pair_idx is not None:
        line = int(view.row_to_line[insertion_pair_idx])
        for widget in (view.left, view.base, view.right):
            widget.mark_set("insert", f"{line}.0")


def _assert_region_buttons_enabled(view) -> None:
    view._set_copy_scope_mode("region")
    assert view._copy_scope_mode == "region"
    assert view._copy_scope_var.get() == "region"
    assert str(view.use_left_btn.cget("state")) == "normal"
    assert str(view.use_right_btn.cget("state")) == "normal"
    assert "区" in str(view.use_left_btn.cget("text"))
    assert "区" in str(view.use_right_btn.cget("text"))
    for menu in (view._use_left_menu, view._use_right_menu):
        end = menu.index("end")
        labels = [
            str(menu.entrycget(index, "label"))
            for index in range(int(end) + 1)
        ]
        assert any("区域" in label for label in labels), labels
        assert any("全局" in label for label in labels), labels


def _column_values(worksheet, column=1):
    return [
        worksheet.cell(row=row, column=column).value
        for row in range(1, worksheet.max_row + 1)
    ]


def test_region_target_resolver_uses_nearest_block_and_earlier_tie_break():
    mine, theirs = _ordinary_diff_rows()
    app, view = _open_view(mine, theirs)
    try:
        assert view._logical_diff_pair_block_for_pair(1) == [1, 2]
        assert view._logical_diff_pair_block_for_pair(6) == [6, 7]

        block, relocated = view._resolve_region_action_target("B2A", 4)
        assert _block_pairs(block) == (1, 2), _block_pairs(block)
        assert relocated is True

        block, relocated = view._resolve_region_action_target("B2A", 6)
        assert _block_pairs(block) == (6, 7), _block_pairs(block)
        assert relocated is False

        block, relocated = view._resolve_region_action_target("B2A", None)
        assert _block_pairs(block) == (1, 2), _block_pairs(block)
        assert relocated is True
    finally:
        app._shutdown_root()


def test_region_target_resolver_filters_blocks_by_copy_direction():
    mine = [("id-1",), ("id-2",), ("mine-only",), ("id-3",), ("id-4",), ("id-5",), ("id-6",)]
    theirs = [("id-1",), ("id-2",), ("id-3",), ("id-4",), ("id-5",), ("theirs-only",), ("id-6",)]
    app, view = _open_view(mine, theirs)
    try:
        mine_only_pair = next(
            idx for idx, (row_a, row_b) in enumerate(view.row_pairs)
            if row_a is not None and row_b is None
        )
        theirs_only_pair = next(
            idx for idx, (row_a, row_b) in enumerate(view.row_pairs)
            if row_a is None and row_b is not None
        )
        block, relocated = view._resolve_region_action_target("B2A", mine_only_pair)
        assert _block_pairs(block) == (mine_only_pair,)
        assert relocated is False
        block, relocated = view._resolve_region_action_target("A2B", theirs_only_pair)
        assert _block_pairs(block) == (theirs_only_pair,)
        assert relocated is False
        _assert_region_buttons_enabled(view)
    finally:
        app._shutdown_root()


def test_explicit_applicable_region_writes_immediately():
    mine, theirs = _ordinary_diff_rows()
    app, view = _open_view(mine, theirs)
    try:
        _assert_region_buttons_enabled(view)
        target_pair = 1
        view._select_line(view.row_to_line[target_pair])
        undo_before = len(app.undo_stack)

        view._run_copy_action_by_mode("B2A")
        _pump(app.root)

        assert view.selected_pair_idx == target_pair
        assert app.ws_a_edit("Data").cell(2, 2).value == "changed-2"
        assert app.ws_a_edit("Data").cell(3, 2).value == "changed-3"
        assert len(app.undo_stack) == undo_before + 1
        action = app.undo_stack[-1]
        children = action.get("actions", ()) if action.get("kind") == "compound" else (action,)
        overlay_actions = [
            child for child in children
            if child.get("target") == "A" and child.get("overlay_transaction") is not None
        ]
        assert len(overlay_actions) == 1
        assert app.sheet_operation_overlay("Data").cells
        view._undo_last_action()
        _pump(app.root)
        assert app.ws_a_val("Data").cell(2, 2).value == "same-2"
        assert app.redo_stack
        view._redo_last_action()
        _pump(app.root)
        assert app.ws_a_val("Data").cell(2, 2).value == "changed-2"
        assert app.sheet_operation_overlay("Data").cells
        assert not view.pair_diff_cols.get(1) and not view.pair_diff_cols.get(2)
        assert view.pair_diff_cols.get(6) and view.pair_diff_cols.get(7)
    finally:
        app._shutdown_root()


def test_explicit_theirs_deleted_region_deletes_mine_rows_and_undo_reapplies():
    mine = [
        ("id-1",),
        ("id-2",),
        ("theirs-deleted-1",),
        ("theirs-deleted-2",),
        ("id-3",),
        ("id-4",),
    ]
    theirs = [("id-1",), ("id-2",), ("id-3",), ("id-4",)]
    app, view = _open_view(mine, theirs)
    output = None
    try:
        _assert_region_buttons_enabled(view)
        mine_only_pairs = [
            idx for idx, (row_a, row_b) in enumerate(view.row_pairs)
            if row_a is not None and row_b is None
        ]
        assert len(mine_only_pairs) == 2, (view.row_pairs, mine_only_pairs)
        assert view._logical_diff_pair_block_for_pair(mine_only_pairs[0]) == mine_only_pairs
        assert view._row_label_for_pair_idx(mine_only_pairs[0], "B") == "缺行"
        assert "此侧缺行" in visible_render_text(
            view.pair_text_b[mine_only_pairs[0]],
            placeholder=smt._TK_INDEX_PLACEHOLDER,
        )
        undo_before = len(app.undo_stack)

        view._select_line(view.row_to_line[mine_only_pairs[0]])
        view._run_copy_action_by_mode("B2A")
        _pump(app.root)
        assert _column_values(app.ws_a_edit("Data")) == [row[0] for row in theirs]
        assert len(app.undo_stack) == undo_before + 1
        assert app.manual_a_row_ops[-1]["kind"] == "delete_rows"
        assert app.manual_a_row_ops[-1]["row"] == 3
        assert app.manual_a_row_ops[-1]["count"] == 2

        view._undo_last_action()
        _pump(app.root)
        assert _column_values(app.ws_a_edit("Data")) == [row[0] for row in mine]
        assert len(app.undo_stack) == undo_before
        assert app.manual_a_row_ops == []

        mine_only_pairs = [
            idx for idx, (row_a, row_b) in enumerate(view.row_pairs)
            if row_a is not None and row_b is None
        ]
        view._select_line(view.row_to_line[mine_only_pairs[0]])
        view._run_copy_action_by_mode("B2A")
        _pump(app.root)
        assert _column_values(app.ws_a_edit("Data")) == [row[0] for row in theirs]

        output = app.build_manual_merge_output_file()
        saved = load_workbook(output, data_only=False, read_only=True)
        try:
            assert _column_values(saved["Data"]) == [row[0] for row in theirs]
        finally:
            saved.close()
    finally:
        app._shutdown_root()
        if output and os.path.exists(output):
            os.remove(output)


def test_explicit_mine_deleted_region_symmetrically_deletes_theirs_rows():
    mine = [("id-1",), ("id-2",), ("id-3",), ("id-4",)]
    theirs = [
        ("id-1",),
        ("id-2",),
        ("mine-deleted-1",),
        ("mine-deleted-2",),
        ("id-3",),
        ("id-4",),
    ]
    app, view = _open_view(mine, theirs)
    try:
        _request_edit_backend_for_completed_operation(app)
        _assert_region_buttons_enabled(view)
        theirs_only_pairs = [
            idx for idx, (row_a, row_b) in enumerate(view.row_pairs)
            if row_a is None and row_b is not None
        ]
        assert len(theirs_only_pairs) == 2
        assert view._row_label_for_pair_idx(theirs_only_pairs[0], "A") == "缺行"
        assert "此侧缺行" in visible_render_text(
            view.pair_text_a[theirs_only_pairs[0]],
            placeholder=smt._TK_INDEX_PLACEHOLDER,
        )
        view._select_line(view.row_to_line[theirs_only_pairs[0]])
        undo_before = len(app.undo_stack)

        view._run_copy_action_by_mode("A2B")
        _pump(app.root)
        assert _column_values(app.ws_b_edit("Data")) == [row[0] for row in mine]
        assert len(app.undo_stack) == undo_before + 1
        assert app.manual_b_row_ops[-1]["kind"] == "delete_rows"
        assert app.manual_b_row_ops[-1]["count"] == 2

        view._undo_last_action()
        _pump(app.root)
        assert _column_values(app.ws_b_edit("Data")) == [row[0] for row in theirs]
        assert app.manual_b_row_ops == []
    finally:
        app._shutdown_root()


def test_region_fallback_first_click_only_locates_second_click_applies_and_undo_reapplies(*, typed_schema=False):
    typed_case = _typed_region_case("sow_region_fidelity_") if typed_schema else None
    owned = typed_case.__enter__() if typed_case is not None else None
    app = None
    try:
        mine, theirs = _ordinary_diff_rows(typed_schema=typed_schema)
        app, view = _open_view(mine, theirs, owned_case=owned)
        _request_edit_backend_for_completed_operation(app)
        _assert_region_buttons_enabled(view)

        early_ids = ("id-2", "id-3") if typed_schema else (2, 3)
        anchor_id = "id-6" if typed_schema else 6
        target_ids = ("id-7", "id-8") if typed_schema else (7, 8)

        def position_for(identifier):
            worksheet = app.ws_a_val("Data")
            physical_row = next(
                row
                for row in range(1, worksheet.max_row + 1)
                if worksheet.cell(row, 1).value == identifier
            )
            pair_idx = view.row_a_to_pair_idx[physical_row]
            return physical_row, pair_idx

        early_positions = tuple(position_for(identifier) for identifier in early_ids)
        anchor_row, anchor_pair = position_for(anchor_id)
        target_positions = tuple(position_for(identifier) for identifier in target_ids)
        target_rows = tuple(position[0] for position in target_positions)
        target_pairs = tuple(position[1] for position in target_positions)
        target_pair = target_pairs[0]
        assert tuple(view._logical_diff_pair_block_for_pair(target_pair)) == target_pairs
        assert anchor_pair not in target_pairs and anchor_row < target_rows[0]

        # The ID-derived equal row is nearest to the later changed block.
        _clear_region_anchor(view, insertion_pair_idx=anchor_pair)
        undo_before = len(app.undo_stack)
        view.info.configure(text="")
        view._run_copy_action_by_mode("B2A")
        _pump(app.root)

        target_line = int(view.row_to_line[target_pair])
        assert view.selected_pair_idx == target_pair
        assert view.hover_pair_idx == target_pair
        assert view._last_cursor_cmp_pair_idx == target_pair
        for widget in (view.left, view.right):
            assert int(str(widget.index("insert")).split(".")[0]) == target_line
            assert widget.bbox(f"{target_line}.0") is not None
        assert "changed-7" in view.cursor_cmp.get("1.0", "end")
        assert tuple(
            app.ws_a_edit("Data").cell(row, 2).value for row in target_rows
        ) == ("same-7", "same-8")
        assert all(view.pair_diff_cols.get(pair_idx) for _row, pair_idx in early_positions)
        assert all(view.pair_diff_cols.get(pair_idx) for pair_idx in target_pairs)
        assert len(app.undo_stack) == undo_before
        message = str(view.info.cget("text"))
        assert "区域" in message and "再次" in message, message

        # The located block is now an explicit selection, so the second click
        # applies it without another relocation or intermediate undo entry.
        view._run_copy_action_by_mode("B2A")
        _pump(app.root)
        assert view.selected_pair_idx == target_pair
        assert tuple(
            app.ws_a_edit("Data").cell(row, 2).value for row in target_rows
        ) == ("changed-7", "changed-8")
        assert all(not view.pair_diff_cols.get(pair_idx) for pair_idx in target_pairs)
        assert len(app.undo_stack) == undo_before + 1

        view._undo_last_action()
        _pump(app.root)
        assert tuple(
            app.ws_a_edit("Data").cell(row, 2).value for row in target_rows
        ) == ("same-7", "same-8")
        assert all(view.pair_diff_cols.get(pair_idx) for pair_idx in target_pairs)

        # Reapply is the existing redo-equivalent interaction; it must keep the
        # selected target and must not jump to the earlier still-different block.
        view._run_copy_action_by_mode("B2A")
        _pump(app.root)
        assert view.selected_pair_idx == target_pair
        assert tuple(
            app.ws_a_edit("Data").cell(row, 2).value for row in target_rows
        ) == ("changed-7", "changed-8")
        assert len(app.undo_stack) == undo_before + 1
        _assert_region_buttons_enabled(view)
    finally:
        if typed_case is not None:
            typed_case.__exit__(*sys.exc_info())
        elif app is not None:
            app._shutdown_root()


def test_region_action_with_no_applicable_diff_is_nonmodal_silent_noop():
    rows = [(idx, f"same-{idx}") for idx in range(1, 6)]
    app, view = _open_view(rows, rows)
    original_showerror = smt.messagebox.showerror
    bells = []
    errors = []
    try:
        _assert_region_buttons_enabled(view)
        view.root.bell = lambda: bells.append(True)
        smt.messagebox.showerror = lambda *args, **kwargs: errors.append((args, kwargs))
        _clear_region_anchor(view, insertion_pair_idx=0)
        undo_before = list(app.undo_stack)
        mine_before = [
            app.ws_a_edit("Data").cell(row=row, column=2).value
            for row in range(1, 6)
        ]

        assert view._resolve_region_action_target("B2A", None) is None
        view.info.configure(text="")
        view._run_copy_action_by_mode("B2A")
        _pump(app.root)
        message = str(view.info.cget("text"))
        assert "区域" in message and ("无" in message or "没有" in message), message
        assert bells == [], "no-applicable-region feedback must not ring the bell"
        assert errors == [], errors
        assert app.undo_stack == undo_before
        assert [
            app.ws_a_edit("Data").cell(row=row, column=2).value
            for row in range(1, 6)
        ] == mine_before
        _assert_region_buttons_enabled(view)
    finally:
        smt.messagebox.showerror = original_showerror
        app._shutdown_root()


def test_mixed_structural_and_cell_region_is_one_atomic_undo():
    mine = [
        ("ID", "值"),
        ("id-1", "same"),
        ("deleted-by-theirs", "remove"),
        ("id-2", "mine-value"),
        ("id-3", "same"),
    ]
    theirs = [
        ("ID", "值"),
        ("id-1", "same"),
        ("id-2", "theirs-value"),
        ("id-3", "same"),
    ]
    app, view = _open_view(mine, theirs)
    try:
        structural_pair = next(
            idx
            for idx, (row_a, row_b) in enumerate(view.row_pairs)
            if row_a is not None and row_b is None
        )
        block = view._logical_diff_pair_block_for_pair(structural_pair)
        assert len(block) == 2, (view.row_pairs, block)
        undo_before = len(app.undo_stack)
        view._set_copy_scope_mode("region")
        view._select_line(view.row_to_line[structural_pair])
        view._run_copy_action_by_mode("B2A")
        _pump(app.root)

        assert _column_values(app.ws_a_edit("Data"), 1) == [
            row[0] for row in theirs
        ]
        assert app.ws_a_edit("Data").cell(3, 2).value == "theirs-value"
        assert len(app.undo_stack) == undo_before + 1
        grouped = app.undo_stack[-1]
        assert grouped.get("kind") == "compound", grouped
        assert len(grouped.get("actions") or ()) == 2, grouped

        view._undo_last_action()
        _pump(app.root)
        assert _column_values(app.ws_a_edit("Data"), 1) == [
            row[0] for row in mine
        ]
        assert app.ws_a_edit("Data").cell(4, 2).value == "mine-value"
        assert len(app.undo_stack) == undo_before
    finally:
        app._shutdown_root()


def test_region_column_delete_is_guided_before_any_cell_write():
    base = [
        ("id@id", "part", "part_move", "model", "quality"),
        ("uint32", "map<int,string>", "int32", "string", "int32"),
        ("ID", "部件", "镜头移动", "旧模型描述", "品质"),
        (1, "head", 4, "old-model-a", 1),
        (2, "tail", 2, "old-model-b", 2),
        (3, "wing", 1, "old-model-c", 3),
    ]
    mine = list(base)
    theirs = [
        ("id@id", "part", "model", "quality"),
        ("uint32", "map<int,string>", "string", "int32"),
        ("ID", "车辆部件", "新模型描述", "品质"),
        (1, "weapon", "new-model-x", 1),
        (2, "wheel", "new-model-y", 2),
        (3, "trunk", "new-model-z", 3),
    ]
    app, view = _open_three_way_view(mine, base, theirs)
    original_showwarning = smt.messagebox.showwarning
    original_showerror = smt.messagebox.showerror
    warnings = []
    errors = []
    try:
        projection = view._active_column_projection()
        assert (
            projection.slot(3).mine_col,
            projection.slot(3).base_col,
            projection.slot(3).theirs_col,
            projection.slot(3).state,
        ) == (3, 3, None, "theirs-deleted")
        assert (
            projection.slot(4).mine_col,
            projection.slot(4).base_col,
            projection.slot(4).theirs_col,
            projection.slot(4).state,
        ) == (4, 4, 3, "retained")

        structural_pair = next(
            pair_idx
            for pair_idx, cols in view.pair_diff_cols.items()
            if 3 in cols
        )
        before_rows = [
            tuple(
                app.ws_a_edit("Data").cell(row=row, column=col).value
                for col in range(1, 6)
            )
            for row in range(1, 7)
        ]
        undo_before = list(app.undo_stack)
        manual_cells_before = dict(app.manual_a_cell_ops)
        manual_rows_before = list(app.manual_a_row_ops)
        smt.messagebox.showwarning = lambda *args, **kwargs: warnings.append(
            (args, kwargs)
        )
        smt.messagebox.showerror = lambda *args, **kwargs: errors.append(
            (args, kwargs)
        )

        view._set_copy_scope_mode("region")
        view._select_line(view.row_to_line[structural_pair])
        view._run_copy_action_by_mode("B2A")
        _pump(app.root)
        # A toolbar command without a retained explicit cell selection first
        # locates the applicable visual block by design.  Its second click is
        # the mutating attempt and must be intercepted by the structure gate.
        if not warnings:
            view._run_copy_action_by_mode("B2A")
            _pump(app.root)

        after_rows = [
            tuple(
                app.ws_a_edit("Data").cell(row=row, column=col).value
                for col in range(1, 6)
            )
            for row in range(1, 7)
        ]
        assert after_rows == before_rows
        assert app.undo_stack == undo_before
        assert app.manual_a_cell_ops == manual_cells_before
        assert app.manual_a_row_ops == manual_rows_before
        assert errors == [], errors
        assert warnings, "structure guidance must be visible"
        warning_text = str(warnings[-1][0])
        assert "C" in warning_text and "part_move" in warning_text, warning_text
        assert "采用Theirs列" in warning_text, warning_text
        assert "本次未写入任何内容" in warning_text, warning_text
        assert view.selected_column_logical_range is not None
        start_col, end_col = view.selected_column_logical_range
        assert int(start_col) <= 3 <= int(end_col)
        assert "本次未写入" in str(view.info.cget("text"))
    finally:
        smt.messagebox.showwarning = original_showwarning
        smt.messagebox.showerror = original_showerror
        app._shutdown_root()


def test_base_row_actions_physically_restore_and_delete_with_undo():
    base = [
        ("ID", "值"),
        ("id-1", "same"),
        ("base-only", "restore"),
        ("id-2", "same"),
    ]
    mine = [
        ("ID", "值"),
        ("id-1", "same"),
        ("mine-only", "delete"),
        ("id-2", "same"),
    ]
    app, view = _open_three_way_view(mine, base, base)
    try:
        restore_pair = next(
            idx
            for idx, pair in enumerate(view.row_pairs)
            if (
                view._row_for_side(pair, "A") is None
                and view._base_row_for_pair(idx, pair) is not None
            )
        )
        assert view._copy_selected_row(
            "BASE2A",
            override_pair_idx=restore_pair,
        )
        _pump(app.root)
        assert "base-only" in _column_values(app.ws_a_edit("Data"), 1)
        assert app.undo_stack[-1]["target"] == "A_INSERT_ROW"
        view._undo_last_action()
        _pump(app.root)
        assert _column_values(app.ws_a_edit("Data"), 1) == [
            row[0] for row in mine
        ]

        delete_pair = next(
            idx
            for idx, pair in enumerate(view.row_pairs)
            if (
                view._row_for_side(pair, "A") is not None
                and view._base_row_for_pair(idx, pair) is None
            )
        )
        assert view._copy_selected_row(
            "BASE2A",
            override_pair_idx=delete_pair,
        )
        _pump(app.root)
        assert "mine-only" not in _column_values(app.ws_a_edit("Data"), 1)
        assert app.undo_stack[-1]["target"] == "A_DELETE_ROW"
        view._undo_last_action()
        _pump(app.root)
        assert _column_values(app.ws_a_edit("Data"), 1) == [
            row[0] for row in mine
        ]
    finally:
        app._shutdown_root()


def test_initial_missing_marker_and_row_header_clicks_are_safe():
    mine = [
        ("ID", "值"),
        ("id-1", "same"),
        ("mine-only", "keep"),
        ("id-2", "same"),
    ]
    theirs = [
        ("ID", "值"),
        ("id-1", "same"),
        ("id-2", "same"),
    ]
    app, view = _open_view(mine, theirs, rescan=False)
    try:
        structural_pair = next(
            idx
            for idx, (row_a, row_b) in enumerate(view.row_pairs)
            if row_a is not None and row_b is None
        )
        assert "此侧" in visible_render_text(
            view.pair_text_b[structural_pair],
            placeholder=smt._TK_INDEX_PLACEHOLDER,
        )
        line = int(view.row_to_line[structural_pair])
        before = _column_values(app.ws_a_edit("Data"), 1)
        undo_before = list(app.undo_stack)

        class Event:
            x = 2
            y = 2

        bbox = view.left_ln.bbox(f"{line}.0")
        assert bbox is not None
        Event.y = bbox[1] + 1
        assert view._on_row_header_click(
            view.left_ln, Event(), "A2B"
        ) == "break"
        assert view._on_row_header_click(
            view.right_ln, Event(), "B2A"
        ) == "break"
        _pump(app.root)
        assert _column_values(app.ws_a_edit("Data"), 1) == before
        assert app.undo_stack == undo_before
        assert view.selected_pair_idx == structural_pair
        assert "已选择" in str(view.info.cget("text"))
    finally:
        app._shutdown_root()


def test_visible_row_header_arrow_applies_blank_source_row():
    base_side = [
        ("ID", "空值", "其他"),
        ("id-1", None, "base"),
        ("id-2", "same", "same"),
    ]
    mine_side = [
        ("ID", "空值", "其他"),
        ("id-1", "=A2", "mine"),
        ("id-2", "same", "same"),
    ]
    app, view = _open_view(base_side, mine_side)
    try:
        pair_idx = view.row_a_to_pair_idx[2]
        assert view.pair_diff_cols[pair_idx] == {2, 3}

        # Reproduce the reported sequence: first adopt one blank Base cell,
        # then use the visible row-number arrow for the remaining row diff.
        view._copy_single_cell_by_pair(pair_idx, "A2B", 2)
        _pump(app.root)
        assert app.ws_b_edit("Data").cell(2, 2).value is None
        assert view.pair_diff_cols[pair_idx] == {3}

        line = int(view.row_to_line[pair_idx])
        bbox = view.left_ln.bbox(f"{line}.0")
        assert bbox is not None

        class Event:
            x = bbox[0] + 1
            y = bbox[1] + 1

        view._on_row_header_hover(view.left_ln, Event(), "A2B")
        assert view._hover_ln_line_left == line
        assert view._on_row_header_click(
            view.left_ln,
            Event(),
            "A2B",
        ) == "break"
        _pump(app.root)

        assert app.ws_b_edit("Data").cell(2, 2).value is None
        assert app.ws_b_val("Data").cell(2, 2).value is None
        assert app.ws_b_edit("Data").cell(2, 3).value == "base"
        assert ("Data", 2, 2) in app.manual_b_cell_ops
        assert app.manual_b_cell_ops[("Data", 2, 2)] is None
        assert app.undo_stack[-1]["target"] == "B"

        view._undo_last_action()
        _pump(app.root)
        assert app.ws_b_edit("Data").cell(2, 2).value is None
        assert app.ws_b_edit("Data").cell(2, 3).value == "mine"
        view._undo_last_action()
        _pump(app.root)
        assert app.ws_b_edit("Data").cell(2, 2).value == "=A2"
    finally:
        app._shutdown_root()


def test_row_delete_transforms_recorded_formula_and_native_output():
    mine = [
        ("ID", "公式"),
        ("id-1", None),
        ("deleted-by-theirs", None),
        ("id-2", None),
        ("id-3", "=A4"),
    ]
    theirs = [
        ("ID", "公式"),
        ("id-1", None),
        ("id-2", None),
        ("id-3", "=A3"),
    ]
    app, view = _open_view(mine, theirs)
    output = None
    try:
        app.record_manual_a_cell("Data", 5, 2, "=A4")
        app.ws_a_edit("Data").cell(5, 2).value = "=A4"
        structural_pair = next(
            idx
            for idx, (row_a, row_b) in enumerate(view.row_pairs)
            if row_a is not None and row_b is None
        )
        assert view._copy_selected_row(
            "B2A",
            override_pair_idx=structural_pair,
        )
        _pump(app.root)
        assert ("Data", 5, 2) not in app.manual_a_cell_ops
        assert app.manual_a_cell_ops[("Data", 4, 2)] == "=A3"

        output = app.build_manual_merge_output_file()
        saved = load_workbook(output, data_only=False, read_only=True)
        try:
            assert saved["Data"].cell(4, 2).value == "=A3"
        finally:
            saved.close()

        view._undo_last_action()
        _pump(app.root)
        assert app.manual_a_cell_ops[("Data", 5, 2)] == "=A4"
        assert app.ws_a_edit("Data").cell(5, 2).value == "=A4"
    finally:
        app._shutdown_root()
        if output and os.path.exists(output):
            os.remove(output)


def test_region_failure_after_structural_commit_rolls_back_atomically():
    mine = [
        ("ID", "值"),
        ("id-1", "same"),
        ("deleted-by-theirs", "remove"),
        ("id-2", "mine-value"),
        ("id-3", "same"),
    ]
    theirs = [
        ("ID", "值"),
        ("id-1", "same"),
        ("id-2", "theirs-value"),
        ("id-3", "same"),
    ]
    app, view = _open_view(mine, theirs)
    original_delete = view._batch_delete_rows_for_missing_source
    original_showerror = smt.messagebox.showerror
    errors = []
    try:
        structural_pair = next(
            idx
            for idx, (row_a, row_b) in enumerate(view.row_pairs)
            if row_a is not None and row_b is None
        )
        view._set_copy_scope_mode("region")
        view._select_line(view.row_to_line[structural_pair])
        before_rows = [
            tuple(
                app.ws_a_edit("Data").cell(row=row, column=col).value
                for col in range(1, 3)
            )
            for row in range(1, app.ws_a_edit("Data").max_row + 1)
        ]
        undo_before = list(app.undo_stack)

        def fail_after_commit(*args, **kwargs):
            assert original_delete(*args, **kwargs)
            raise RuntimeError("injected post-commit failure")

        view._batch_delete_rows_for_missing_source = fail_after_commit
        smt.messagebox.showerror = lambda *args, **kwargs: errors.append(
            (args, kwargs)
        )
        view._run_copy_action_by_mode("B2A")
        _pump(app.root)

        after_rows = [
            tuple(
                app.ws_a_edit("Data").cell(row=row, column=col).value
                for col in range(1, 3)
            )
            for row in range(1, app.ws_a_edit("Data").max_row + 1)
        ]
        assert after_rows == before_rows
        assert app.undo_stack == undo_before
        assert app.manual_a_row_ops == []
        assert errors and "injected post-commit failure" in str(errors[-1][0])
    finally:
        view._batch_delete_rows_for_missing_source = original_delete
        smt.messagebox.showerror = original_showerror
        app._shutdown_root()


def test_conflict_dialog_has_location_and_goto_button():
    base = [
        ("ID", "值"),
        ("id-1", "base"),
    ]
    mine = [
        ("ID", "值"),
        ("id-1", "mine"),
    ]
    theirs = [
        ("ID", "值"),
        ("id-1", "theirs"),
    ]
    app, _view = _open_three_way_view(mine, base, theirs)
    captured_text = []
    try:
        app.initial_conflict_cell_count = 3

        def _descendants(widget):
            result = []
            for child in widget.winfo_children():
                result.append(child)
                result.extend(_descendants(child))
            return result

        def _invoke_goto():
            for widget in _descendants(app.root):
                try:
                    text = str(widget.cget("text"))
                except Exception:
                    continue
                if text:
                    captured_text.append(text)
                if text == "前往首个冲突":
                    widget.invoke()
                    return
            app.root.after(25, _invoke_goto)

        app.root.after(50, _invoke_goto)
        action = app._show_unresolved_conflict_save_dialog(
            3,
            ("Data", 2, 2),
        )
        assert action == "goto"
        all_text = "\n".join(captured_text)
        assert "前往首个冲突" in all_text
        assert "Sheet：Data" in all_text
        assert "位置：B2" in all_text
        assert "行号：2" in all_text
        assert "列号：B" in all_text
    finally:
        app._shutdown_root()


def test_focus_logical_conflict_cell_selects_row_column_and_c_area():
    base = [
        ("ID", "值", "尾列"),
        ("id-1", "base", "same"),
        ("id-2", "same", "same"),
    ]
    mine = [
        ("ID", "值", "尾列"),
        ("id-1", "mine", "same"),
        ("id-2", "same", "same"),
    ]
    theirs = [
        ("ID", "值", "尾列"),
        ("id-1", "theirs", "same"),
        ("id-2", "same", "same"),
    ]
    app, view = _open_three_way_view(mine, base, theirs)
    try:
        assert view.focus_logical_cell(2, 2)
        _pump(app.root)
        pair_idx = view.row_a_to_pair_idx[2]
        line = int(view.row_to_line[pair_idx])
        assert view.selected_pair_idx == pair_idx
        assert view._main_sel_line == line
        assert view._main_sel_col == 2
        assert view._cursor_cmp_sel_col == 2
        assert view._cursor_cmp_sel_line == 2
        assert view.left.tag_ranges("selcell")
        assert view.base.tag_ranges("selcell")
        assert view.right.tag_ranges("selcell")
        assert view.cursor_cmp.tag_ranges("cselcell")
        assert str(view.left.index("insert")).startswith(f"{line}.")
        info = str(view.info.cget("text"))
        assert "B2" in info
        assert "第 2 行" in info
        assert "Excel 列 B" in info
    finally:
        app._shutdown_root()


def main(argv=None):
    tests = (
        test_region_target_resolver_uses_nearest_block_and_earlier_tie_break,
        test_region_target_resolver_filters_blocks_by_copy_direction,
        test_explicit_applicable_region_writes_immediately,
        test_explicit_theirs_deleted_region_deletes_mine_rows_and_undo_reapplies,
        test_explicit_mine_deleted_region_symmetrically_deletes_theirs_rows,
        test_region_fallback_first_click_only_locates_second_click_applies_and_undo_reapplies,
        test_region_action_with_no_applicable_diff_is_nonmodal_silent_noop,
        test_mixed_structural_and_cell_region_is_one_atomic_undo,
        test_region_column_delete_is_guided_before_any_cell_write,
        test_base_row_actions_physically_restore_and_delete_with_undo,
        test_initial_missing_marker_and_row_header_clicks_are_safe,
        test_visible_row_header_arrow_applies_blank_source_row,
        test_row_delete_transforms_recorded_formula_and_native_output,
        test_region_failure_after_structural_commit_rolls_back_atomically,
        test_conflict_dialog_has_location_and_goto_button,
        test_focus_logical_conflict_cell_selects_row_column_and_c_area,
    )
    parser = argparse.ArgumentParser()
    parser.add_argument("--case", choices=tuple(test.__name__ for test in tests))
    args = parser.parse_args(argv)
    selected = tuple(test for test in tests if args.case in (None, test.__name__))
    original_prompt_scheduler = smt.SowMergeApp._schedule_formula_cache_prompt
    smt.SowMergeApp._schedule_formula_cache_prompt = lambda _self: None
    try:
        for test in selected:
            print(f"START: {test.__name__}", flush=True)
            test()
            print(f"PASS: {test.__name__}", flush=True)
    finally:
        smt.SowMergeApp._schedule_formula_cache_prompt = original_prompt_scheduler
    print(f"PASS: region mode interaction regression ({len(selected)} tests)", flush=True)


if __name__ == "__main__":
    main()

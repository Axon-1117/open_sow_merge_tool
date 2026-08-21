"""Read-only real-workbook parity plus disposable selected-sheet mutations."""

from __future__ import annotations

import shutil
import tempfile
from pathlib import Path
import argparse
import zipfile
import time
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import coordinate_from_string

import sow_merge_tool as sm
from _large_sheet_direct_oracle import capture
from _large_sheet_oracle_fixtures import REAL_FIXTURES, copy_real_fixture
from _large_sheet_snapshot_oracle import compare_manifests


def _assert_parity(mine_path: Path, theirs_path: Path, sheet: str, label: str) -> None:
    started = time.perf_counter()
    print(f"REAL_ORACLE_START {label} direct", flush=True)
    legacy = capture(str(mine_path), str(theirs_path), sheet)
    print(f"REAL_ORACLE_DONE {label} direct_ms={(time.perf_counter() - started) * 1000:.1f}", flush=True)
    snapshot_started = time.perf_counter()
    print(f"REAL_ORACLE_START {label} snapshot", flush=True)
    mine = sm._stream_selected_sheet_snapshot(str(mine_path), str(mine_path), sheet, "A")
    theirs = sm._stream_selected_sheet_snapshot(str(theirs_path), str(theirs_path), sheet, "B")
    result = sm._compare_selected_sheet_snapshots(mine, theirs)
    print(f"REAL_ORACLE_DONE {label} snapshot_ms={(time.perf_counter() - snapshot_started) * 1000:.1f}", flush=True)
    assert not result.unresolved, f"{label}: snapshot gate unexpectedly unresolved"
    parity = compare_manifests(
        legacy, sm.snapshot_comparison_oracle_manifest(mine, theirs, result)
    )
    assert parity["exact"], f"{label}: {parity['mismatches'][:1]}"


def _save_variant(path: Path, sheet: str, mutate) -> None:
    wb = load_workbook(path, data_only=False)
    try:
        mutate(wb[sheet])
        wb.save(path)
    finally:
        wb.close()


def _worksheet_member(members: dict[str, bytes], sheet: str) -> str:
    main = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    rel = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
    package_rel = "{http://schemas.openxmlformats.org/package/2006/relationships}"
    workbook = ET.fromstring(members["xl/workbook.xml"])
    relationship_id = next(node.attrib[rel + "id"] for node in workbook.iter(main + "sheet") if node.attrib.get("name") == sheet)
    relationships = ET.fromstring(members["xl/_rels/workbook.xml.rels"])
    target = next(node.attrib["Target"] for node in relationships.iter(package_rel + "Relationship") if node.attrib.get("Id") == relationship_id)
    return "xl/" + target.lstrip("/")


def _zip_sheet_mutate(path: Path, sheet: str, mutate) -> None:
    """Change only target worksheet XML; avoids slow whole-workbook re-save."""
    with zipfile.ZipFile(path, "r") as source:
        members = {name: source.read(name) for name in source.namelist()}
    member = _worksheet_member(members, sheet)
    root = ET.fromstring(members[member])
    mutate(root)
    members[member] = ET.tostring(root, encoding="utf-8", xml_declaration=True)
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as target:
        for name, payload in members.items():
            target.writestr(name, payload)


def _sheet_xml_names():
    return "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}", "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"


def _append_row_xml(row_number: int, key: int, text: str):
    main, _unused = _sheet_xml_names()
    def _mutate(root):
        data = root.find(main + "sheetData")
        row = ET.SubElement(data, main + "row", {"r": str(row_number)})
        left = ET.SubElement(row, main + "c", {"r": f"A{row_number}", "t": "n"})
        ET.SubElement(left, main + "v").text = str(key)
        right = ET.SubElement(row, main + "c", {"r": f"B{row_number}", "t": "inlineStr"})
        inline = ET.SubElement(right, main + "is")
        ET.SubElement(inline, main + "t").text = text
    return _mutate


def _append_column_xml():
    main, _unused = _sheet_xml_names()
    def _mutate(root):
        existing_columns = []
        for cell in root.iter(main + "c"):
            try:
                letters, _row = coordinate_from_string(cell.attrib.get("r", ""))
                existing_columns.append(column_index_from_string(letters))
            except ValueError:
                pass
        at_column = max(existing_columns, default=1) + 1
        data = root.find(main + "sheetData")
        by_row = {int(row.attrib.get("r", "0")): row for row in data.findall(main + "row")}
        for row_number, value in ((1, "oracle_added@pm"), (2, "string"), (5, "oracle-column-insert")):
            row = by_row.get(row_number)
            if row is None:
                continue
            cell = ET.SubElement(row, main + "c", {"r": f"{get_column_letter(at_column)}{row_number}", "t": "inlineStr"})
            inline = ET.SubElement(cell, main + "is")
            ET.SubElement(inline, main + "t").text = value
    return _mutate


def main(argv=None) -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--case", choices=("all", "self", "cell", "row", "column"), default="all")
    args = parser.parse_args(argv)
    root = Path(tempfile.mkdtemp(prefix="sow_real_snapshot_oracle_"))
    skill = next(item for item in REAL_FIXTURES if item.name == "Skill")
    world = next(item for item in REAL_FIXTURES if item.name == "WorldMonster")
    skill_left, skill_sheet = copy_real_fixture(skill, root / "skill")
    world_left, world_sheet = copy_real_fixture(world, root / "world")
    _assert_parity(skill_left, skill_left, skill_sheet, "Skill-self")
    print("REAL_SNAPSHOT_DIRECT_ORACLE_SKILL_SELF_OK", flush=True)
    _assert_parity(world_left, world_left, world_sheet, "WorldMonster-self")
    print("REAL_SNAPSHOT_DIRECT_ORACLE_WORLDMONSTER_SELF_OK", flush=True)
    if args.case == "self":
        return

    if args.case in ("all", "cell"):
        cell_right = root / "skill-cell.xlsx"
        shutil.copy2(skill_left, cell_right)
        _save_variant(cell_right, skill_sheet, lambda ws: setattr(ws.cell(5, 2), "value", "oracle-cell-change"))
        _assert_parity(skill_left, cell_right, skill_sheet, "Skill-cell")
        print("REAL_SNAPSHOT_DIRECT_ORACLE_CELL_OK", flush=True)

    if args.case in ("all", "row"):
        row_right = root / "skill-row.xlsx"
        shutil.copy2(skill_left, row_right)
        _zip_sheet_mutate(row_right, skill_sheet, _append_row_xml(3462, 987654321, "oracle-row-insert"))
        _assert_parity(skill_left, row_right, skill_sheet, "Skill-row")
        print("REAL_SNAPSHOT_DIRECT_ORACLE_ROW_OK", flush=True)

    if args.case in ("all", "column"):
        column_right = root / "skill-column.xlsx"
        shutil.copy2(skill_left, column_right)
        _zip_sheet_mutate(column_right, skill_sheet, _append_column_xml())
        _assert_parity(skill_left, column_right, skill_sheet, "Skill-column")
        print("REAL_SNAPSHOT_DIRECT_ORACLE_COLUMN_OK", flush=True)
    print("REAL_SNAPSHOT_DIRECT_ORACLE_OK")


if __name__ == "__main__":
    main()

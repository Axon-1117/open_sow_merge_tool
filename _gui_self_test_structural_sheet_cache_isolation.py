"""Natural public structural-column cache-isolation gate.

The one stable selector proves that a real S1 column adoption rebuilds only
S1.  S2 remains an immutable exact sibling throughout; no private mutation,
stale-map injection, save, or rescan route is used.
"""

from __future__ import annotations

import argparse
import json
import os
import shutil
import tempfile
import time

from openpyxl import Workbook

import sow_merge_tool as sm


_CASE = "structural-sheet-cache-isolation"
_DEADLINE_SECONDS = 90.0
_ROWS = 40


def _abs(path: str) -> str:
    return os.path.normcase(os.path.abspath(os.fspath(path)))


def _sha(path: str) -> str:
    return sm._sha256_file(path)


def _note_or_raise(primary: BaseException | None, errors: list[str]) -> None:
    if not errors:
        return
    detail = " | ".join(errors)
    if primary is not None:
        try:
            primary.add_note(f"structural cache cleanup: {detail}")
        except Exception:
            pass
        return
    raise AssertionError(detail)


def _pump(app, deadline: float, duration: float = 0.025) -> None:
    until = min(deadline, time.monotonic() + duration)
    while time.monotonic() < until:
        app.root.update_idletasks()
        app.root.update()
        time.sleep(0.003)


def _wait(app, predicate, label: str, deadline: float) -> None:
    while time.monotonic() < deadline:
        _pump(app, deadline)
        if predicate():
            return
    selected = str(getattr(app, "selected_sheet", "") or "")
    view = getattr(app, "sheet_views", {}).get(selected)
    diagnostic = {
        "selected": selected,
        "entry": app._sheet_exact_entry(selected) if selected else None,
        "queue": tuple(getattr(app, "_compute_queue", ()) or ()),
        "inflight": tuple(getattr(app, "_compute_inflight", ()) or ()),
        "edit_requests": tuple(getattr(app, "_edit_load_requests", ()) or ()),
        "edit_ready": bool(getattr(app, "_edit_workbooks_ready", lambda: False)()),
        "view": {
            "prepared": getattr(view, "_prepared_complete", None),
            "data": getattr(view, "_data_ready", None),
            "pending": getattr(view, "_pending_exact_render", None),
            "lifecycle": getattr(view, "_lifecycle_state", None),
        },
    }
    raise AssertionError(f"{label}: {json.dumps(diagnostic, ensure_ascii=False, default=str)}")


def _full_current(app, sheet: str, expected: str | None = None) -> bool:
    view = getattr(app, "sheet_views", {}).get(sheet)
    entry = app._sheet_exact_entry(sheet)
    state = str(entry.get("state") or "")
    return bool(
        view
        and app.selected_sheet == sheet
        and state in sm._SHEET_EXACT_TERMINAL
        and (expected is None or state == expected)
        and bool(entry.get("full_detail_terminal"))
        and bool(view._prepared_complete)
        and bool(view._data_ready)
        and not bool(view._pending_exact_render)
        and bool(view._row_model_exact)
        and bool(view._pair_diff_full_exact)
    )


def _select_sheet(app, sheet: str, expected: str, deadline: float) -> None:
    app.nb.select(app._sheet_containers[sheet])
    _wait(
        app,
        lambda: app.nb.tab(app.nb.select(), "text") == sheet
        and _full_current(app, sheet, expected),
        f"{sheet} public Notebook full terminal",
        deadline,
    )


def _pair_parts(mapping) -> tuple:
    return tuple(sorted((int(index), tuple(parts)) for index, parts in dict(mapping or {}).items()))


def _pair_maps(mapping) -> tuple:
    return tuple(sorted((int(index), tuple(sorted(int(col) for col in cols))) for index, cols in dict(mapping or {}).items()))


def _sibling_semantics(view) -> dict:
    cache = view.column_comparison_cache
    projection = view.column_projection
    assert isinstance(cache, sm.LogicalColumnComparisonCache)
    assert isinstance(projection, sm.LogicalColumnProjection)
    assert projection.model is cache.model
    return {
        "row_pairs": tuple(view.row_pairs),
        "full_rows": tuple(view._full_display_rows),
        "raw_a": _pair_parts(view.pair_raw_parts_a),
        "raw_b": _pair_parts(view.pair_raw_parts_b),
        "raw_base": _pair_parts(view.pair_raw_parts_base),
        "row_a_to_pair": tuple(sorted((int(row), int(pair)) for row, pair in view.row_a_to_pair_idx.items())),
        "row_b_to_pair": tuple(sorted((int(row), int(pair)) for row, pair in view.row_b_to_pair_idx.items())),
        "mine_to_base": tuple(sorted((int(row), int(base)) for row, base in view.mine_to_base_row.items())),
        "theirs_to_base": tuple(sorted((int(row), int(base)) for row, base in view.theirs_to_base_row.items())),
        "base_override": tuple(sorted((int(pair), int(base)) for pair, base in view.pair_base_row_override.items())),
        "pair_diff": _pair_maps(view.pair_diff_cols),
        "base_diff": _pair_maps(view.pair_base_diff_cols),
        "slots": tuple(
            (int(slot.logical_idx), slot.mine_col, slot.base_col, slot.theirs_col, str(slot.state))
            for slot in cache.model.slots
        ),
        "blocks": tuple(
            (int(block.ordinal), tuple(int(index) for index in block.slot_indices))
            for block in cache.model.blocks
        ),
        "unresolved": tuple(sorted(int(value) for value in cache.unresolved_cols)),
        "structural": tuple(sorted(int(value) for value in cache.structural_diff_cols)),
        "versions": (
            int(view._row_model_version),
            int(view._column_model_version),
            int(view._mine_edit_version),
            int(view._theirs_edit_version),
        ),
    }


def _queued_for(records, sheet: str) -> bool:
    for record in tuple(records or ()):
        if isinstance(record, dict) and str(record.get("sheet") or "") == sheet:
            return True
        if isinstance(record, (tuple, list)) and record and str(record[0]) == sheet:
            return True
        if str(record) == sheet:
            return True
    return False


def _assert_sibling_unchanged(app, sibling_view, sibling_snapshot, sibling_entry, sibling_semantics, sibling_generation: int, event_count: int) -> None:
    view = app.sheet_views["S2"]
    assert view is sibling_view
    assert app.selected_sheet_snapshot("A", "S2") is sibling_snapshot
    assert dict(app._sheet_exact_entry("S2")) == sibling_entry
    assert int(app._sheet_compute_generation["S2"]) == sibling_generation
    assert _sibling_semantics(view) == sibling_semantics
    assert not _queued_for(getattr(app, "_compute_queue", ()), "S2")
    assert not _queued_for(getattr(app, "_compute_inflight", ()), "S2")
    new_events = tuple(getattr(app, "_snapshot_child_events", ()) or ())[event_count:]
    assert not any(str(event.get("sheet") or "") == "S2" for event in new_events if isinstance(event, dict))


def _header_select_inserted_column(app, view, deadline: float) -> None:
    widget = view.right_colhdr
    assert widget.bind("<Button-1>")
    spans = view._spans_for_line(widget.get("1.0", "1.end"))
    assert 2 in spans, spans
    start, end = spans[2]
    char = int(start + max(0, (end - start - 1) // 2))
    index = f"1.{char}"
    widget.see(index)
    _pump(app, deadline)
    box = widget.bbox(index)
    assert box is not None, (index, spans, widget.get("1.0", "1.end"))
    x, y, width, height = box
    x, y = int(x + max(1, width // 2)), int(y + max(1, height // 2))
    resolved = widget.index(f"@{x},{y}")
    resolved_char = int(str(resolved).split(".", 1)[1])
    assert int(start) <= resolved_char < int(end), (index, resolved, spans)
    widget.event_generate("<Button-1>", x=x, y=y)
    _pump(app, deadline)
    assert view.selected_column_logical_range == (2, 2)
    assert view.selected_column_source_side == "B"


def _cancel_debounces(app) -> None:
    owners = (app, *tuple(getattr(app, "sheet_views", {}).values()))
    for owner in owners:
        if owner is None:
            continue
        for attr in ("_settings_save_id", "_hover_debounce_id", "_diff_map_debounce_id"):
            after_id = getattr(owner, attr, None)
            if after_id is None:
                continue
            try:
                app.root.after_cancel(after_id)
            except Exception:
                pass
            finally:
                setattr(owner, attr, None)


class _Fixture:
    def __init__(self):
        self.previous_settings_path = sm._SETTINGS_PATH
        self.temporary = None
        self.root = None
        self.user_settings_exists = False
        self.user_settings_bytes = None
        self.settings_path = None
        self.input_hashes: dict[str, str] = {}
        try:
            self.temporary = tempfile.TemporaryDirectory(prefix="sow_structural_cache_")
            self.root = self.temporary.name
            self.user_settings_exists = os.path.lexists(self.previous_settings_path)
            if self.user_settings_exists:
                with open(self.previous_settings_path, "rb") as stream:
                    self.user_settings_bytes = stream.read()
            self.settings_path = os.path.join(self.root, "settings.json")
            with open(self.settings_path, "w", encoding="utf-8") as stream:
                json.dump({"only_diff": 0}, stream)
            sm._SETTINGS_PATH = self.settings_path
        except BaseException as exc:
            cleanup_errors: list[str] = []
            try:
                sm._SETTINGS_PATH = self.previous_settings_path
            except BaseException as restore_exc:
                cleanup_errors.append(f"settings restore: {type(restore_exc).__name__}: {restore_exc}")
            try:
                if self.temporary is not None:
                    self.temporary.cleanup()
                if self.root is not None:
                    assert not os.path.lexists(self.root)
            except BaseException as cleanup_exc:
                cleanup_errors.append(f"temporary root: {type(cleanup_exc).__name__}: {cleanup_exc}")
            if cleanup_errors:
                try:
                    exc.add_note("structural fixture constructor cleanup: " + " | ".join(cleanup_errors))
                except Exception:
                    pass
            raise

    def track(self, path: str) -> str:
        normalized = _abs(path)
        assert os.path.isfile(normalized), normalized
        self.input_hashes[normalized] = _sha(normalized)
        return normalized

    def assert_inputs(self) -> None:
        for path, digest in self.input_hashes.items():
            assert os.path.isfile(path) and _sha(path) == digest, path

    def restore_user_settings(self) -> None:
        sm._SETTINGS_PATH = self.previous_settings_path
        if self.user_settings_exists:
            with open(self.previous_settings_path, "rb") as stream:
                assert stream.read() == self.user_settings_bytes
        else:
            assert not os.path.lexists(self.previous_settings_path)


def _make_workbook(path: str, *, theirs: bool) -> None:
    wb = Workbook()
    for index, sheet in enumerate(("S1", "S2")):
        ws = wb.active if index == 0 else wb.create_sheet(sheet)
        ws.title = sheet
        if sheet == "S1" and theirs:
            ws.append(("id@id", "inserted", "value"))
            ws.append(("int32", "string", "string"))
            for number in range(1, _ROWS + 1):
                ws.append((number, f"new-{number}", f"value-{number}"))
        else:
            ws.append(("id@id", "value"))
            ws.append(("int32", "string"))
            for number in range(1, _ROWS + 1):
                ws.append((number, f"{sheet.lower()}-value-{number}" if sheet == "S2" else f"value-{number}"))
    wb.save(path)
    wb.close()


def _make_inputs(fixture: _Fixture) -> tuple[str, str]:
    mine_book = os.path.join(fixture.root, "mine.xlsx")
    theirs_book = os.path.join(fixture.root, "theirs.xlsx")
    _make_workbook(mine_book, theirs=False)
    _make_workbook(theirs_book, theirs=True)
    mine = os.path.join(fixture.root, "mine.xlsx.r39265")
    theirs = os.path.join(fixture.root, "theirs.xlsx.r39264")
    shutil.copy2(mine_book, mine)
    shutil.copy2(theirs_book, theirs)
    return fixture.track(mine), fixture.track(theirs)


def _assert_effective_inputs(app, fixture: _Fixture, startup_ledger: set[str], mine: str, theirs: str) -> set[str]:
    effective = {_abs(app.file_a), _abs(app.file_b)}
    raw = {_abs(mine), _abs(theirs)}
    assert effective.isdisjoint(raw) and len(effective) == 2
    assert startup_ledger is app._owned_startup_temp_paths
    assert {_abs(path) for path in startup_ledger} == effective
    assert _sha(app.file_a) == fixture.input_hashes[_abs(mine)]
    assert _sha(app.file_b) == fixture.input_hashes[_abs(theirs)]
    return effective


def _expected_s1_rows(*, adopted: bool) -> tuple[tuple[object, ...], ...]:
    header = ("id@id", "inserted", "value") if adopted else ("id@id", "value")
    type_row = ("int32", "string", "string") if adopted else ("int32", "string")
    data = tuple(
        (number, f"new-{number}", f"value-{number}")
        if adopted else (number, f"value-{number}")
        for number in range(1, _ROWS + 1)
    )
    return (header, type_row, *data)


def _assert_s1_typed_values(ws, *, adopted: bool) -> None:
    expected_rows = _expected_s1_rows(adopted=adopted)
    assert int(ws.max_row) == len(expected_rows)
    assert int(ws.max_column) == len(expected_rows[0])
    for row_number, expected in enumerate(expected_rows, start=1):
        actual = tuple(ws.cell(row=row_number, column=column).value for column in range(1, len(expected) + 1))
        assert actual == expected, (row_number, actual, expected)


def _assert_no_manual_cell_or_row_ops(app) -> None:
    assert not app.manual_a_cell_ops
    assert not app.manual_b_cell_ops
    assert not app.manual_a_formula_cache_ops
    assert not app.manual_b_formula_cache_ops
    assert not app.manual_a_row_ops
    assert not app.manual_b_row_ops


def _assert_column_applied(app, view, overlay_before: tuple[int, int], deadline: float) -> tuple[dict, ...]:
    _wait(app, lambda: _full_current(app, "S1", sm._SHEET_EXACT_CHANGED), "S1 inserted column exact CHANGED", deadline)
    ws = app.ws_a_val("S1")
    _assert_s1_typed_values(ws, adopted=True)
    _assert_no_manual_cell_or_row_ops(app)
    assert all(not cols for cols in view.pair_diff_cols.values())
    operations = tuple(dict(op) for op in app.manual_a_column_ops)
    assert [(op["kind"], op["sheet"], op["target_side"], op["source_side"], int(op["target_logical_slot"]), int(op["target_physical_anchor"]), int(op["count"]), tuple(op["source_physical_cols"])) for op in operations] == [
        ("insert_cols", "S1", "A", "B", 2, 2, 1, (2,)),
        ("copy_cols", "S1", "A", "B", 2, 2, 1, (2,)),
    ], operations
    assert len({str(op["action_id"]) for op in operations}) == 1
    assert all(str(op["batch_id"]) == str(operations[0]["action_id"]) for op in operations)
    assert app.manual_b_column_ops == []
    overlay = app.sheet_operation_overlay("S1")
    assert overlay.cells == {}
    assert (int(overlay.topology_generation), int(overlay.mutation_generation)) == (overlay_before[0] + 1, overlay_before[1] + 1)
    assert len(app.undo_stack) == 1 and not app.redo_stack
    return operations


def _run_case() -> dict:
    assert 0 < _DEADLINE_SECONDS <= 90.0
    deadline = time.monotonic() + _DEADLINE_SECONDS
    fixture = None
    app = None
    startup_ledger: set[str] = set()
    expected_owned: set[str] = set()
    warnings: list[tuple[str, str]] = []
    errors: list[tuple[str, str]] = []
    confirms: list[tuple[str, str]] = []
    original_warning = sm.messagebox.showwarning
    original_error = sm.messagebox.showerror
    original_confirm = sm.messagebox.askyesno
    primary = None
    try:
        fixture = _Fixture()
        mine, theirs = _make_inputs(fixture)
        sm.messagebox.showwarning = lambda title, message, **_kwargs: warnings.append((str(title), str(message)))
        sm.messagebox.showerror = lambda title, message, **_kwargs: errors.append((str(title), str(message)))
        sm.messagebox.askyesno = lambda title, message, **_kwargs: confirms.append((str(title), str(message))) or False
        app = sm.SowMergeApp(
            mine,
            theirs,
            initial_sheet="S1",
            startup_owned_paths=startup_ledger,
        )
        expected_owned = _assert_effective_inputs(app, fixture, startup_ledger, mine, theirs)
        assert app.only_diff_default == 0 and int(app.settings.get("only_diff", -1)) == 0
        with open(fixture.settings_path, "r", encoding="utf-8") as stream:
            assert json.load(stream) == {"only_diff": 0}

        _wait(app, lambda: _full_current(app, "S1", sm._SHEET_EXACT_CHANGED), "S1 initial typed exact CHANGED", deadline)
        _select_sheet(app, "S2", sm._SHEET_EXACT_SAME, deadline)
        s2_view = app.sheet_views["S2"]
        s2_snapshot = app.selected_sheet_snapshot("A", "S2")
        s2_entry = dict(app._sheet_exact_entry("S2"))
        s2_semantics = _sibling_semantics(s2_view)
        s2_generation = int(app._sheet_compute_generation["S2"])
        s2_event_count = len(tuple(app._snapshot_child_events or ()))
        assert s2_snapshot is not None

        _select_sheet(app, "S1", sm._SHEET_EXACT_CHANGED, deadline)
        view = app.sheet_views["S1"]
        assert int(view.only_diff_var.get()) == 0 and int(s2_view.only_diff_var.get()) == 0
        assert view._derive_lifecycle_state() == "EDIT_DEFERRED"
        assert not app._edit_workbooks_ready()
        assert tuple(app._edit_load_requests) == () and app._edit_preload_thread is None
        _header_select_inserted_column(app, view, deadline)
        overlay_before = app.sheet_operation_overlay("S1")
        overlay_before_state = (int(overlay_before.topology_generation), int(overlay_before.mutation_generation))
        before_manual = tuple(app.manual_a_column_ops)
        before_undo = tuple(app.undo_stack)
        modified_a_before = bool(app.modified_a)
        modified_sheets_a_before = set(app.modified_sheets_a)

        view.use_theirs_col_btn.invoke()
        _pump(app, deadline)
        assert len(warnings) == 1 and not errors and not confirms, (warnings, errors, confirms)
        title, message = warnings[0]
        assert title == "正在加载可编辑工作簿"
        assert "列结构操作未执行" in message and "首次操作已启动后台加载" in message
        assert "不会执行或自动重试" in message and "请手动重试" in message
        requests = tuple(app._edit_load_requests)
        assert len(requests) == 1
        assert requests[0]["reason"] == "mutation:列结构操作"
        assert requests[0]["caller"] == "SheetView._guard_mutation_ready" and requests[0]["ready"] is False
        owner = app._edit_preload_thread
        assert owner is not None and app.manual_a_column_ops == list(before_manual) and app.undo_stack == list(before_undo)
        assert app.sheet_operation_overlay("S1").cells == {}
        _wait(app, lambda: app._edit_workbooks_ready(), "single edit preload ready", deadline)
        assert app._edit_preload_thread is owner and tuple(app._edit_load_requests) == requests
        _wait(app, lambda: view._derive_lifecycle_state() == "READY", "S1 READY after edit preload", deadline)

        view.use_theirs_col_btn.invoke()
        operations = _assert_column_applied(app, view, overlay_before_state, deadline)
        assert app.modified_a is True and "S1" in app.modified_sheets_a
        assert not errors and not confirms and tuple(app._edit_load_requests) == requests
        _assert_sibling_unchanged(app, s2_view, s2_snapshot, s2_entry, s2_semantics, s2_generation, s2_event_count)

        view.undo_btn.invoke()
        _wait(app, lambda: _full_current(app, "S1", sm._SHEET_EXACT_CHANGED) and not app.manual_a_column_ops and not app.undo_stack and len(app.redo_stack) == 1, "public undo structural column", deadline)
        undo_overlay = app.sheet_operation_overlay("S1")
        assert undo_overlay.cells == {}
        assert (int(undo_overlay.topology_generation), int(undo_overlay.mutation_generation)) == (overlay_before_state[0] + 2, overlay_before_state[1] + 2)
        ws_undo = app.ws_a_val("S1")
        _assert_s1_typed_values(ws_undo, adopted=False)
        _assert_no_manual_cell_or_row_ops(app)
        assert app.manual_b_column_ops == []
        assert bool(app.modified_a) == modified_a_before
        assert set(app.modified_sheets_a) == modified_sheets_a_before
        assert "S1" not in app.modified_sheets_a
        _assert_sibling_unchanged(app, s2_view, s2_snapshot, s2_entry, s2_semantics, s2_generation, s2_event_count)

        view.redo_btn.invoke()
        _wait(app, lambda: _full_current(app, "S1", sm._SHEET_EXACT_CHANGED) and tuple(app.manual_a_column_ops) == operations and len(app.undo_stack) == 1 and not app.redo_stack, "public redo structural column", deadline)
        redo_overlay = app.sheet_operation_overlay("S1")
        assert redo_overlay.cells == {}
        assert (int(redo_overlay.topology_generation), int(redo_overlay.mutation_generation)) == (overlay_before_state[0] + 3, overlay_before_state[1] + 3)
        _assert_s1_typed_values(app.ws_a_val("S1"), adopted=True)
        _assert_no_manual_cell_or_row_ops(app)
        assert app.manual_b_column_ops == []
        assert app.modified_a is True and "S1" in app.modified_sheets_a
        _assert_sibling_unchanged(app, s2_view, s2_snapshot, s2_entry, s2_semantics, s2_generation, s2_event_count)
        detail = {
            "sheet": "S1",
            "action": "public-right-column-header/use-theirs-column",
            "manual_operations": [op["kind"] for op in operations],
            "s2_generation": s2_generation,
            "input_hashes": dict(fixture.input_hashes),
        }
        assert time.monotonic() <= deadline, (
            "structural case exceeded original deadline",
            time.monotonic(),
            deadline,
        )
        return detail
    except BaseException as exc:
        primary = exc
        raise
    finally:
        cleanup_errors: list[str] = []
        try:
            if app is not None:
                _cancel_debounces(app)
                app._shutdown_root()
                evidence = tuple(app._owned_startup_temp_cleanup_evidence)
                assert not startup_ledger and not app._owned_startup_temp_paths
                assert {_abs(item["path"]) for item in evidence} == expected_owned
                assert all(item["removed"] and not item["exists_after"] and not item["error"] for item in evidence)
            elif startup_ledger:
                expected = {_abs(path) for path in startup_ledger}
                evidence: list[dict] = []
                sm._consume_owned_startup_temp_paths(startup_ledger, evidence)
                assert not startup_ledger
                assert {_abs(item["path"]) for item in evidence} == expected
                assert all(item["removed"] and not item["exists_after"] and not item["error"] for item in evidence)
        except BaseException as exc:
            cleanup_errors.append(f"shutdown/ledger: {type(exc).__name__}: {exc}")
        finally:
            sm.messagebox.showwarning = original_warning
            sm.messagebox.showerror = original_error
            sm.messagebox.askyesno = original_confirm
        if fixture is not None:
            try:
                with open(fixture.settings_path, "r", encoding="utf-8") as stream:
                    assert json.load(stream) == {"only_diff": 0}
            except BaseException as exc:
                cleanup_errors.append(f"temporary settings: {type(exc).__name__}: {exc}")
            try:
                fixture.assert_inputs()
            except BaseException as exc:
                cleanup_errors.append(f"input SHA: {type(exc).__name__}: {exc}")
            try:
                fixture.restore_user_settings()
            except BaseException as exc:
                cleanup_errors.append(f"settings restore: {type(exc).__name__}: {exc}")
            root = fixture.root
            try:
                fixture.temporary.cleanup()
                assert not os.path.lexists(root)
            except BaseException as exc:
                cleanup_errors.append(f"temporary root: {type(exc).__name__}: {exc}")
        _note_or_raise(primary, cleanup_errors)


def main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--case", choices=(_CASE,))
    parser.add_argument("--list-cases", action="store_true")
    args = parser.parse_args(argv)
    if args.list_cases:
        print(_CASE)
        return
    selected = (args.case,) if args.case else (_CASE,)
    for case in selected:
        assert case == _CASE
        detail = _run_case()
        print(f"GUI_SELF_TEST_STRUCTURAL_SHEET_CACHE_ISOLATION_OK {json.dumps(detail, ensure_ascii=False, sort_keys=True)}")


if __name__ == "__main__":
    main()

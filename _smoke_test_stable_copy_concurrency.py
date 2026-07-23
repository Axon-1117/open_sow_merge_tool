"""Concurrent stable-copy names must never collide across worker processes."""

from __future__ import annotations

import concurrent.futures
import os

from openpyxl import Workbook, load_workbook

from _test_temp_utils import make_temp_dir


def _copy_in_worker(source: str, temp_root: str) -> str:
    import sow_merge_tool as smt

    # Keep the test sandbox-local while making the source qualify as a temp
    # artifact for the production stable-copy path.
    smt.tempfile.tempdir = temp_root
    return smt._ensure_stable_copy(source)


def main():
    temp_root = make_temp_dir("sow_stable_copy_concurrency_")
    source = os.path.join(temp_root, "same-name.xlsx")
    workbook = Workbook()
    workbook.active["A1"] = "stable-copy-sentinel"
    workbook.save(source)
    workbook.close()

    worker_count = 8
    with concurrent.futures.ProcessPoolExecutor(max_workers=worker_count) as executor:
        paths = list(executor.map(
            _copy_in_worker,
            [source] * worker_count,
            [temp_root] * worker_count,
        ))

    assert len(paths) == worker_count
    assert len(set(paths)) == worker_count, paths
    assert all(os.path.basename(path).startswith("sow_merge_tool_stable_") for path in paths)
    assert all(path.lower().endswith(".xlsx") for path in paths)
    for path in paths:
        assert path != source and os.path.exists(path), path
        copied = load_workbook(path, read_only=True, data_only=True)
        try:
            assert copied.active["A1"].value == "stable-copy-sentinel"
        finally:
            copied.close()
    print("SMOKE_TEST_STABLE_COPY_CONCURRENCY_OK")


if __name__ == "__main__":
    main()

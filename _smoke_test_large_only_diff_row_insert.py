"""Large immutable-only-diff regression with a declared row-insert fixture."""

from __future__ import annotations

import copy
import hashlib
import json
import os
import sys
import tempfile
import time
from contextlib import contextmanager
from types import SimpleNamespace

from openpyxl import Workbook

import sow_merge_tool as mod


def _sha256(path: str) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _make_book(path: str, rows: list[tuple[str, str]]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "S1"
    for row in rows:
        worksheet.append(list(row))
    workbook.save(path)
    workbook.close()


def _pump(root, seconds: float) -> None:
    deadline = time.monotonic() + seconds
    while time.monotonic() < deadline:
        root.update_idletasks()
        root.update()
        time.sleep(0.01)


@contextmanager
def _owned_case():
    original_settings_path = mod._SETTINGS_PATH
    original_settings_exists = os.path.lexists(original_settings_path)
    if original_settings_exists:
        with open(original_settings_path, "rb") as stream:
            original_settings_bytes = stream.read()
    else:
        original_settings_bytes = None
    temporary = tempfile.TemporaryDirectory(prefix="sow_large_onlydiff_insert_")
    root_dir = temporary.name
    settings_path = os.path.join(root_dir, "settings.json")
    with open(settings_path, "w", encoding="utf-8") as stream:
        json.dump({"only_diff": 0}, stream)
    original_prompt_scheduler = mod.SowMergeApp._schedule_formula_cache_prompt
    state = SimpleNamespace(app=None, input_hashes={}, root_dir=root_dir)
    mod._SETTINGS_PATH = settings_path
    mod.SowMergeApp._schedule_formula_cache_prompt = lambda _self: None
    primary = None
    try:
        yield state
    except BaseException as exc:
        primary = exc
        raise
    finally:
        cleanup_errors = []
        if state.app is not None:
            try:
                for candidate in (state.app, *getattr(state.app, "sheet_views", {}).values()):
                    if candidate is None:
                        continue
                    after_id = getattr(candidate, "_settings_save_id", None)
                    if after_id is not None:
                        state.app.root.after_cancel(after_id)
                        candidate._settings_save_id = None
                state.app._shutdown_root()
            except BaseException as exc:
                cleanup_errors.append(f"shutdown: {exc!r}")
        try:
            mod.SowMergeApp._schedule_formula_cache_prompt = original_prompt_scheduler
            mod._SETTINGS_PATH = original_settings_path
            if original_settings_exists:
                with open(original_settings_path, "rb") as stream:
                    assert stream.read() == original_settings_bytes
            else:
                assert not os.path.lexists(original_settings_path)
        except BaseException as exc:
            cleanup_errors.append(f"settings/prompt restore: {exc!r}")
        for path, before_hash in state.input_hashes.items():
            try:
                assert _sha256(path) == before_hash, path
            except BaseException as exc:
                cleanup_errors.append(f"input SHA {path!r}: {exc!r}")
        try:
            temporary.cleanup()
            assert not os.path.lexists(root_dir), root_dir
        except BaseException as exc:
            cleanup_errors.append(f"owned temporary root: {exc!r}")
        if cleanup_errors:
            message = "large only-diff cleanup failed: " + "; ".join(cleanup_errors)
            if primary is not None:
                primary.add_note(message)
            else:
                raise AssertionError(message)


def _rows() -> tuple[list[tuple[str, str]], list[tuple[str, str]]]:
    mine = [("id@id", "value"), ("string", "string")]
    mine.extend((f"row-{index:04d}", f"value-{index:04d}") for index in range(1, 2401))
    theirs = list(mine)
    mod_index = next(index for index, row in enumerate(theirs) if row[0] == "row-0752")
    theirs[mod_index] = ("row-0752", "value-0752-mod")
    insert_after = next(index for index, row in enumerate(theirs) if row[0] == "row-0752")
    theirs.insert(insert_after + 1, ("row-0753-new", "value-0753-new"))
    return mine, theirs


def _zero_write_snapshot(app, view):
    def frozen_mapping(mapping):
        return tuple(sorted((repr(key), repr(value)) for key, value in mapping.items()))

    overlay = app.sheet_operation_overlay("S1")
    return {
        "manual_a_cell": frozen_mapping(app.manual_a_cell_ops),
        "manual_b_cell": frozen_mapping(app.manual_b_cell_ops),
        "manual_a_row": tuple(copy.deepcopy(app.manual_a_row_ops)),
        "manual_b_row": tuple(copy.deepcopy(app.manual_b_row_ops)),
        "undo": tuple(copy.deepcopy(app.undo_stack)),
        "redo": tuple(copy.deepcopy(app.redo_stack)),
        "modified": (
            app.modified_a,
            app.modified_b,
            tuple(sorted(app.modified_sheets_a)),
            tuple(sorted(app.modified_sheets_b)),
        ),
        "overlay": frozen_mapping(overlay.cells),
        "row_pairs": tuple(view.row_pairs),
        "pair_diff": tuple(sorted((key, tuple(sorted(value))) for key, value in view.pair_diff_cols.items())),
        "pair_base_diff": tuple(sorted((key, tuple(sorted(value))) for key, value in view.pair_base_diff_cols.items())),
        "raw_a": frozen_mapping(view.pair_raw_parts_a),
        "raw_b": frozen_mapping(view.pair_raw_parts_b),
        "raw_base": frozen_mapping(view.pair_raw_parts_base),
        "edit_ready": app._edit_workbooks_ready(),
    }


@contextmanager
def _forbid_edit_preload(app):
    calls = []
    originals = {}

    def forbidden(name):
        def _raise(*args, **kwargs):
            calls.append((name, args, kwargs))
            raise AssertionError(f"immutable only-diff route called {name}")
        return _raise

    for name in ("_request_edit_preload", "_load_edit_workbooks_owned"):
        if hasattr(app, name):
            originals[name] = getattr(app, name)
            setattr(app, name, forbidden(name))
    try:
        yield calls
    finally:
        for name, original in originals.items():
            setattr(app, name, original)


def _switch_diagnostic(app, view, sheet: str):
    return {
        "exact": app._sheet_exact_entry(sheet),
        "build": {
            "active": getattr(view, "_only_diff_async_building", None),
            "seq": getattr(view, "_only_diff_async_build_seq", None),
            "key": getattr(view, "_only_diff_async_build_key", None),
        },
        "pending": {
            "mode": getattr(view, "_mode_switch_pending", None),
            "requested": getattr(view, "_mode_switch_requested_value", None),
            "exact_render": getattr(view, "_pending_exact_render", None),
        },
        "display": tuple(getattr(view, "display_rows", ()) or ()),
        "full": tuple(getattr(view, "_full_display_rows", ()) or ()),
        "cache": tuple(getattr(view, "_only_diff_rows_cache", ()) or ()),
    }


def _wait_immutable_ready(app, sheet: str, deadline: float):
    view = None
    while time.monotonic() < deadline:
        _pump(app.root, 0.02)
        view = app.sheet_views.get(sheet)
        entry = app._sheet_exact_entry(sheet)
        if (
            view is not None
            and app._is_sheet_exact_current(sheet)
            and bool(entry.get("full_detail_terminal"))
            and bool(getattr(view, "_prepared_complete", False))
            and bool(getattr(view, "_data_ready", False))
            and view._is_exact_immutable_view_ready()
            and view._derive_lifecycle_state() == "EDIT_DEFERRED"
            and not app._edit_workbooks_ready()
        ):
            return view
    raise AssertionError(("immutable S1 did not become ready", _switch_diagnostic(app, view, sheet)))


def _wait_mode_switch(app, view, value: int, expected_rows: tuple[int, ...], deadline: float):
    while time.monotonic() < deadline:
        _pump(app.root, 0.02)
        if (
            int(view.only_diff_var.get()) == value
            and int(getattr(view, "_last_only_diff_value", -1)) == value
            and not bool(getattr(view, "_only_diff_async_building", False))
            and not bool(getattr(view, "_mode_switch_pending", False))
        ):
            if value:
                if (
                    view._has_valid_only_diff_snapshot_cache()
                    and bool(getattr(view, "_only_diff_rows_exact", False))
                    and tuple(view._only_diff_rows_cache or ()) == expected_rows
                    and tuple(view.display_rows) == expected_rows
                ):
                    return
            elif tuple(view._full_display_rows) == expected_rows:
                return
    raise AssertionError(
        (
            "only-diff switch did not settle",
            value,
            getattr(view, "_only_diff_async_building", None),
            getattr(view, "_mode_switch_pending", None),
            _switch_diagnostic(app, view, "S1"),
        )
    )


def main():
    case_deadline = time.monotonic() + 90.0
    with _owned_case() as owned:
        file_a = os.path.join(owned.root_dir, "a.xlsx")
        file_b = os.path.join(owned.root_dir, "b.xlsx")
        rows_a, rows_b = _rows()
        _make_book(file_a, rows_a)
        _make_book(file_b, rows_b)
        owned.input_hashes = {file_a: _sha256(file_a), file_b: _sha256(file_b)}

        app = mod.SowMergeApp(file_a, file_b)
        owned.app = app
        app.nb.select(app._sheet_containers["S1"])
        view = _wait_immutable_ready(app, "S1", case_deadline)
        assert str(view.only_diff_cb.cget("state")) == "normal"
        assert view._virtual_mode_active()
        assert view._full_display_rows == list(range(len(view.row_pairs)))
        full_rows = tuple(view._full_display_rows)
        assert len(view.row_pairs) == 2403, len(view.row_pairs)
        assert len(full_rows) == 2403, len(full_rows)

        def position_for(identifier: str, side: str):
            worksheet = app.ws_a_val("S1") if side == "A" else app.ws_b_val("S1")
            physical_row = next(
                row
                for row in range(1, worksheet.max_row + 1)
                if worksheet.cell(row, 1).value == identifier
            )
            pair_idx = (
                view.row_a_to_pair_idx[physical_row]
                if side == "A"
                else view.row_b_to_pair_idx[physical_row]
            )
            return physical_row, pair_idx

        mod_a_row, mod_pair = position_for("row-0752", "A")
        mod_b_row, mod_pair_b = position_for("row-0752", "B")
        insert_b_row, insert_pair = position_for("row-0753-new", "B")
        assert mod_pair == mod_pair_b
        assert view.row_pairs[mod_pair] == (mod_a_row, mod_b_row)
        assert view.row_pairs[insert_pair] == (None, insert_b_row)
        assert view.pair_diff_cols.get(mod_pair) == {2}
        assert view.pair_diff_cols.get(insert_pair) == {-1}
        expected_only_diff = tuple(sorted((mod_pair, insert_pair)))
        zero_write_before = _zero_write_snapshot(app, view)

        with _forbid_edit_preload(app) as preload_calls:
            t0 = time.perf_counter()
            view.only_diff_cb.invoke()
            first_toggle_seconds = time.perf_counter() - t0
            assert first_toggle_seconds < 1.5, first_toggle_seconds
            _wait_mode_switch(app, view, 1, expected_only_diff, case_deadline)
            assert _zero_write_snapshot(app, view) == zero_write_before
            assert _sha256(file_a) == owned.input_hashes[file_a]
            assert _sha256(file_b) == owned.input_hashes[file_b]

            view.only_diff_cb.invoke()
            _wait_mode_switch(app, view, 0, full_rows, case_deadline)
            assert tuple(view._full_display_rows) == full_rows
            assert _zero_write_snapshot(app, view) == zero_write_before

            t1 = time.perf_counter()
            view.only_diff_cb.invoke()
            cached_toggle_seconds = time.perf_counter() - t1
            assert cached_toggle_seconds < 0.5, cached_toggle_seconds
            _wait_mode_switch(app, view, 1, expected_only_diff, case_deadline)
            assert _zero_write_snapshot(app, view) == zero_write_before
            assert _sha256(file_a) == owned.input_hashes[file_a]
            assert _sha256(file_b) == owned.input_hashes[file_b]
            assert preload_calls == [], preload_calls

    print("SMOKE_LARGE_ONLY_DIFF_ROW_INSERT_OK", flush=True)


if __name__ == "__main__":
    main()

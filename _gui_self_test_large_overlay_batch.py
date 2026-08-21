"""Public 1,000-row immutable-to-overlay batch regression."""

from __future__ import annotations

import argparse
from contextlib import contextmanager
import hashlib
import json
import math
import os
import sys
import tempfile
import time
from pathlib import Path

from openpyxl import Workbook

import sow_merge_tool as sm


_CASE = "large-overlay-batch"
_CASES = (_CASE,)
_ROWS = 1_000
_SCHEMA_ROWS = 2
_CASE_TIMEOUT = 90.0
_SHEET = "Data"


def _sha(path: str | os.PathLike) -> str:
    digest = hashlib.sha256()
    with Path(path).open("rb") as stream:
        for block in iter(lambda: stream.read(1 << 20), b""):
            digest.update(block)
    return digest.hexdigest()


def _setting(path: str | os.PathLike) -> tuple[bool, bytes]:
    candidate = Path(path)
    return (True, candidate.read_bytes()) if candidate.exists() else (False, b"")


def _norm(path: str | os.PathLike) -> str:
    return os.path.normcase(os.path.abspath(os.fspath(path)))


def _canon(value):
    if value is None or isinstance(value, (str, int, float, bool, bytes)):
        return value
    if isinstance(value, dict):
        return tuple(sorted(((_canon(key), _canon(item)) for key, item in value.items()), key=repr))
    if isinstance(value, (tuple, list, set, frozenset)):
        items = tuple(_canon(item) for item in value)
        return tuple(sorted(items, key=repr)) if isinstance(value, (set, frozenset)) else items
    return (type(value).__name__, repr(value))


def _same(before, after, label: str) -> None:
    assert before == after, f"{label}: immutable identity changed\nbefore={before!r}\nafter={after!r}"


def _pump(root) -> None:
    root.update()
    root.update_idletasks()


def _wait(root, predicate, deadline: float, label: str) -> None:
    while time.monotonic() < deadline:
        _pump(root)
        if predicate():
            return
        time.sleep(0.005)
    raise AssertionError(f"timeout {label}")


def _book(path: Path, side: str) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _SHEET
    sheet.append(["id@id", "value"])
    sheet.append(["int32", "string"])
    for ident in range(1, _ROWS + 1):
        sheet.append([ident, f"{side}-{ident}"])
    workbook.save(path)
    workbook.close()


def _exact_deferred(app, view) -> bool:
    entry = dict(app._sheet_exact_entry(_SHEET) or {})
    return bool(
        app.selected_sheet == _SHEET
        and app._is_sheet_exact_current(_SHEET)
        and entry.get("state") == "EXACT_CHANGED"
        and entry.get("full_detail_terminal")
        and view._prepared_complete
        and view._data_ready
        and view._row_model_exact
        and view._derive_lifecycle_state() == "EDIT_DEFERRED"
        and not app._edit_workbooks_ready()
    )


def _immutable_identity(app, view, inputs, *, include_ready_handles: bool) -> tuple:
    cache = view.column_comparison_cache
    model = cache.model
    projection = view.column_projection
    return _canon({
        "input_sha": tuple(sorted((name, _sha(path)) for name, path in inputs.items())),
        "rows": (view.row_pairs, view.row_a_to_pair_idx, view.row_b_to_pair_idx),
        "base_maps": (view.mine_to_base_row, view.theirs_to_base_row, view.pair_base_row_override),
        "raw": (view.pair_raw_parts_a, view.pair_raw_parts_b, view.pair_raw_parts_base),
        "projection": (
            projection.model is model,
            (model.cache_key.sheet_name, model.cache_key.row_model_version, model.cache_key.column_model_version),
            tuple((slot.logical_idx, slot.mine_col, slot.base_col, slot.theirs_col, str(slot.state), _canon(slot.confidence), slot.base_boundary, slot.origin_side) for slot in model.slots),
            tuple((block.ordinal, tuple(block.slot_indices), str(block.state), _canon(block.confidence)) for block in model.blocks),
            tuple(projection.block_ordinal_by_slot),
            tuple(model.mine_physical_to_logical.entries), tuple(model.base_physical_to_logical.entries), tuple(model.theirs_physical_to_logical.entries),
            tuple(model.mine_logical_to_physical.entries), tuple(model.base_logical_to_physical.entries), tuple(model.theirs_logical_to_physical.entries),
            _canon(model.confidence), tuple(cache.structural_diff_cols), tuple(cache.unresolved_cols),
        ),
        "semantic_model_versions": (view._row_model_version, view._column_model_version),
        "base": (app.has_base, app.base_path),
        "value_handles": (
            tuple((name, id(getattr(app, name, None)), getattr(getattr(app, name, None), "read_only", None)) for name in ("_wb_a_val", "_wb_b_val", "_wb_base_val"))
            if include_ready_handles else ()
        ),
        "edit_handles": (
            tuple((name, id(getattr(app, name, None)), getattr(getattr(app, name, None), "read_only", None)) for name in ("_wb_a_edit", "_wb_b_edit", "_wb_base_edit"))
            if include_ready_handles else ()
        ),
    })


def _assert_input_sha(inputs, before) -> None:
    assert {name: _sha(path) for name, path in inputs.items()} == before


def _expected_domain(view) -> tuple[dict[tuple[str, int, int], str], dict[tuple[str, int], str]]:
    expected = {}
    original = {}
    for ident in range(1, _ROWS + 1):
        physical_row = _SCHEMA_ROWS + ident
        pair_idx = view.row_b_to_pair_idx.get(physical_row)
        assert pair_idx is not None, (ident, physical_row, view.row_b_to_pair_idx)
        row_a, row_b = view.row_pairs[int(pair_idx)]
        assert row_a == row_b == physical_row, (ident, pair_idx, row_a, row_b)
        raw_a = tuple(view.pair_raw_parts_a[int(pair_idx)])
        raw_b = tuple(view.pair_raw_parts_b[int(pair_idx)])
        assert raw_a == (str(ident), f"mine-{ident}"), raw_a
        assert raw_b == (str(ident), f"theirs-{ident}"), raw_b
        assert set(view.pair_diff_cols.get(int(pair_idx), set())) == {2}, (pair_idx, view.pair_diff_cols.get(int(pair_idx)))
        expected[(_SHEET, physical_row, 2)] = raw_b[1]
        original[(_SHEET, physical_row)] = raw_a[1]
    assert len(expected) == _ROWS and len(original) == _ROWS
    return expected, original


def _assert_manual_shape(app, expected: dict[tuple[str, int, int], str], *, label: str) -> None:
    assert dict(app.manual_a_cell_ops) == expected, (label, len(app.manual_a_cell_ops), len(expected))
    assert app.manual_a_formula_cache_ops == {}, (label, app.manual_a_formula_cache_ops)
    assert app.manual_b_cell_ops == {} and app.manual_b_formula_cache_ops == {}, label
    assert app.manual_a_row_ops == [] and app.manual_b_row_ops == [], label
    assert app.manual_a_column_ops == [] and app.manual_b_column_ops == [], label
    assert app.manual_sheet_ops == [] and app.auto_sheet_ops == [], label


def _assert_sheet_values(app, expected: dict[tuple[str, int, int], str], *, label: str) -> None:
    edit = app.ws_a_edit(_SHEET)
    value = app.ws_a_val(_SHEET)
    actual_edit = {(_SHEET, row, 2): edit.cell(row=row, column=2).value for row in range(_SCHEMA_ROWS + 1, _SCHEMA_ROWS + _ROWS + 1)}
    actual_value = {(_SHEET, row, 2): value.cell(row=row, column=2).value for row in range(_SCHEMA_ROWS + 1, _SCHEMA_ROWS + _ROWS + 1)}
    assert actual_edit == expected and actual_value == expected, (label, actual_edit, actual_value)


def _assert_overlay(overlay, expected: dict[tuple[str, int, int], str], *, label: str) -> None:
    assert len(overlay.cells) == _ROWS, (label, len(overlay.cells))
    actual = {(_SHEET, int(delta.physical_row), int(delta.physical_col)): delta.after for delta in overlay.cells.values()}
    assert actual == expected, (label, actual, expected)
    assert {str(delta.side) for delta in overlay.cells.values()} == {"A"}, label


def _cancel_debounces(app) -> None:
    if app is None:
        return
    for owner in (app, *getattr(app, "sheet_views", {}).values()):
        if owner is None:
            continue
        for attr in ("_settings_save_id", "_hover_debounce_id", "_diff_map_debounce_id"):
            after_id = getattr(owner, attr, None)
            if after_id is None:
                continue
            try:
                app.root.after_cancel(after_id)
            except Exception:
                pass
            finally:
                setattr(owner, attr, None)


def _assert_cleanup_evidence(evidence, expected_paths) -> None:
    expected = {_norm(path) for path in expected_paths}
    actual = {_norm(item.get("path", "")) for item in evidence}
    assert actual == expected, (actual, expected, evidence)
    assert all(item.get("removed") is True and item.get("exists_after") is False and not item.get("error") for item in evidence), evidence


def _assert_app_startup_cleanup(app, expected_paths, ledger) -> None:
    _assert_cleanup_evidence(app._owned_startup_temp_cleanup_evidence, expected_paths)
    assert not ledger, ledger


def _consume_preapp_startup_cleanup(ledger) -> None:
    expected_paths = set(ledger)
    evidence = []
    sm._consume_owned_startup_temp_paths(ledger, evidence)
    _assert_cleanup_evidence(evidence, expected_paths)
    assert not ledger, ledger


def _projection_triplet(view):
    cache = view.column_comparison_cache
    model = cache.model
    projection = view.column_projection
    assert cache.model is model and projection.model is model
    return cache, model, projection


def _assert_rebuilt_projection(current, *prior_triples) -> None:
    _cache, _model, _projection = current
    assert _cache.model is _model and _projection.model is _model
    for prior in prior_triples:
        assert all(now is not old for now, old in zip(current, prior)), (tuple(id(item) for item in current), tuple(id(item) for item in prior))


def _assert_cache_key_edit_versions(cache_key, baseline, mine_offset: int) -> None:
    assert (
        cache_key.sheet_name,
        int(cache_key.row_model_version),
        int(cache_key.column_model_version),
    ) == (
        baseline.sheet_name,
        int(baseline.row_model_version),
        int(baseline.column_model_version),
    )
    assert (
        int(cache_key.mine_edit_version),
        int(cache_key.base_edit_version),
        int(cache_key.theirs_edit_version),
    ) == (
        int(baseline.mine_edit_version) + int(mine_offset),
        int(baseline.base_edit_version),
        int(baseline.theirs_edit_version),
    )


@contextmanager
def _post_ready_traps(app, view):
    hits = []
    originals = []

    def forbidden(name):
        def fail(*_args, **_kwargs):
            hits.append(name)
            raise AssertionError(f"forbidden after ready: {name}")
        return fail

    targets = (
        (app, "_request_edit_preload"), (app, "_load_edit_workbooks_owned"),
        (app, "_start_background_thread"), (app, "_atomic_save"),
        (app, "_atomic_save_with_retry"), (app, "_atomic_replace_file_with_retry"),
        (app, "_try_alt_save"), (app, "save_a_inplace"), (app, "save_b_inplace"),
        (app, "save_merged_and_exit"), (view, "_manual_rescan"),
        (view, "_start_async_large_only_diff_build"),
        (sm, "_align_selected_sheet_snapshots"),
        (sm, "_compare_selected_sheet_snapshots"), (sm, "_atomic_save_wb"),
    )
    try:
        for owner, name in targets:
            if hasattr(owner, name):
                originals.append((owner, name, getattr(owner, name)))
                setattr(owner, name, forbidden(f"{type(owner).__name__}.{name}"))
        yield hits
    finally:
        for owner, name, original in reversed(originals):
            setattr(owner, name, original)


def _select_region_menu(view) -> None:
    menu = view._use_right_menu
    end = menu.index("end")
    assert end is not None
    region_index = next(
        index for index in range(int(end) + 1)
        if str(menu.entrycget(index, "value")) == "region"
    )
    menu.invoke(region_index)
    assert view._copy_scope_var.get() == "region" and view._copy_scope_mode == "region"
    assert "区" in str(view.use_right_btn.cget("text"))


def _select_offscreen_right_cell(app, view, pair_idx: int, deadline: float) -> dict:
    full_rows = tuple(view._full_display_rows)
    target_pos = full_rows.index(pair_idx)
    cap = min(sm._VIRTUAL_VIEWPORT_MAX_ROWS, len(full_rows))
    wanted = max(0, min(target_pos, max(0, len(full_rows) - cap)))
    height = max(1, int(view.vdiff_map.winfo_height()))
    fraction = float(wanted) / float(max(1, len(full_rows) - cap))
    before_seq = int(view._viewport_request_seq)
    assert str(view.vdiff_map.bind("<Button-1>") or "").strip()
    view.vdiff_map.event_generate("<Button-1>", x=1, y=max(1, min(height, int(round(fraction * height)))))
    _wait(
        app.root,
        lambda: pair_idx in tuple(view.display_rows) and int(view._viewport_request_seq) > before_seq,
        deadline,
        "public vminimap target",
    )
    request_id = int(view._viewport_request_seq)
    _wait(
        app.root,
        lambda: any(int(item.get("id", -1)) == request_id and item.get("status") == "complete" for item in tuple(view._viewport_request_terminal)),
        deadline,
        "public vminimap terminal",
    )
    line = int(view.row_to_line[pair_idx])
    spans = view._spans_for_line()
    start, end = spans[2]
    index = f"{line}.{start}"
    view.right.see(index)
    _pump(app.root)
    box = view.right.bbox(index)
    lineinfo = view.right.dlineinfo(index)
    assert box is not None and lineinfo is not None, (index, box, lineinfo)
    x, y, width, height = map(int, box)
    event_x = max(x, x + max(0, width // 2))
    event_y = max(y, y + max(0, height // 2))
    actual_line, actual_col = map(int, str(view.right.index(f"@{event_x},{event_y}")).split("."))
    assert actual_line == line and start <= actual_col < end, (actual_line, actual_col, line, start, end)
    original_widget_line = view._widget_line
    calls = []

    def bridge(widget):
        if widget is view.right:
            calls.append({"line": line, "index": index, "event_line": actual_line, "event_col": actual_col})
            return line
        return original_widget_line(widget)

    try:
        view._widget_line = bridge
        view.right.event_generate("<Button-1>", x=event_x, y=event_y)
    finally:
        view._widget_line = original_widget_line
    _pump(app.root)
    assert calls == [{"line": line, "index": index, "event_line": actual_line, "event_col": actual_col}], calls
    assert int(view.selected_pair_idx) == pair_idx and int(view._main_sel_col) == 2, (view.selected_pair_idx, view._main_sel_col)
    return {"pair": pair_idx, "line": line, "request_id": request_id, "row_start": int(view._virtual_window_start)}


def _run_case() -> None:
    deadline = time.monotonic() + _CASE_TIMEOUT
    original_settings = os.fspath(sm._SETTINGS_PATH)
    original_setting = _setting(original_settings)
    original_prompt = sm.SowMergeApp._schedule_formula_cache_prompt
    temporary = tempfile.TemporaryDirectory(prefix="sow_large_overlay_batch_")
    root = Path(temporary.name)
    settings = root / "settings.json"
    mine, theirs = root / "mine.xlsx", root / "theirs.xlsx"
    app = None
    view = None
    primary = None
    startup_ledger: set[str] = set()
    expected_stable: set[str] = set()
    inputs = {}
    input_before = {}
    refresh_original = None
    try:
        settings.write_text(json.dumps({"only_diff": 0}) + "\n", encoding="utf-8")
        sm._SETTINGS_PATH = os.fspath(settings)
        sm.SowMergeApp._schedule_formula_cache_prompt = lambda _self: None
        _book(mine, "mine")
        _book(theirs, "theirs")
        inputs = {"mine": mine, "theirs": theirs}
        input_before = {name: _sha(path) for name, path in inputs.items()}
        print(f"START {_CASE}", flush=True)
        app = sm.SowMergeApp(os.fspath(mine), os.fspath(theirs), initial_sheet=_SHEET, startup_owned_paths=startup_ledger)
        _wait(app.root, lambda: (app.sheet_views.get(_SHEET) is not None and _exact_deferred(app, app.sheet_views[_SHEET])), deadline, "immutable exact deferred")
        view = app.sheet_views[_SHEET]
        assert len(view.row_pairs) == _ROWS + _SCHEMA_ROWS and len(view._full_display_rows) == _ROWS + _SCHEMA_ROWS
        assert view._virtual_mode_active() and len(view.display_rows) <= sm._VIRTUAL_VIEWPORT_MAX_ROWS
        assert app._owned_startup_temp_paths is startup_ledger
        expected_stable = {_norm(app.file_a), _norm(app.file_b)}
        assert expected_stable == {_norm(path) for path in startup_ledger}
        assert expected_stable.isdisjoint({_norm(mine), _norm(theirs)})
        assert _sha(app.file_a) == input_before["mine"] and _sha(app.file_b) == input_before["theirs"]
        deferred_hard = _immutable_identity(app, view, inputs, include_ready_handles=False)
        expected_after, expected_before = _expected_domain(view)
        assert app.manual_a_cell_ops == {} and app.manual_b_cell_ops == {}
        assert app.manual_a_formula_cache_ops == {} and app.manual_b_formula_cache_ops == {}
        assert app.sheet_operation_overlay(_SHEET).cells == {}
        target_pair = view.row_b_to_pair_idx[_SCHEMA_ROWS + 900]
        selection = _select_offscreen_right_cell(app, view, int(target_pair), deadline)
        _select_region_menu(view)
        assert str(view.use_right_btn.cget("state")) == "normal"
        request_before = len(app._edit_load_requests)
        mutation_before = _canon((app.manual_a_cell_ops, app.manual_a_formula_cache_ops, app.manual_b_cell_ops, app.manual_b_formula_cache_ops, app.undo_stack, app.redo_stack, app.sheet_operation_overlay(_SHEET).cells))
        modal_calls = []
        original_warning = sm.messagebox.showwarning

        def record_warning(*args, **kwargs):
            modal_calls.append((args, kwargs))

        try:
            sm.messagebox.showwarning = record_warning
            view.use_right_btn.invoke()
        finally:
            sm.messagebox.showwarning = original_warning
        _pump(app.root)
        assert len(app._edit_load_requests) == request_before + 1
        request = app._edit_load_requests[-1]
        assert request.get("reason") == "mutation:采用所选内容" and request.get("caller") == "SheetView._guard_mutation_ready", request
        owner = app._edit_preload_thread
        assert owner is not None and app._edit_loading_started and not app._edit_workbooks_ready()
        assert len(modal_calls) == 1 and modal_calls[0][0][0] == "正在加载可编辑工作簿", modal_calls
        modal_text = str(modal_calls[0][0][1])
        assert modal_text.startswith("采用所选内容未执行。") and "不会执行或自动重试。" in modal_text and "加载完成后请手动重试。" in modal_text, modal_text
        _same(mutation_before, _canon((app.manual_a_cell_ops, app.manual_a_formula_cache_ops, app.manual_b_cell_ops, app.manual_b_formula_cache_ops, app.undo_stack, app.redo_stack, app.sheet_operation_overlay(_SHEET).cells)), "first public guard")
        _same(deferred_hard, _immutable_identity(app, view, inputs, include_ready_handles=False), "first public guard deferred")
        _wait(app.root, lambda: app._edit_workbooks_ready() and view._derive_lifecycle_state() == "READY", deadline, "single edit owner ready")
        assert app._edit_preload_thread is owner and len(app._edit_load_requests) == request_before + 1
        _same(deferred_hard, _immutable_identity(app, view, inputs, include_ready_handles=False), "post-edit-ready deferred")
        mutation_hard = _immutable_identity(app, view, inputs, include_ready_handles=True)
        ready_projection = _projection_triplet(view)
        ready_cache_key = ready_projection[1].cache_key
        projection_generation0 = int(view._column_projection_generation)
        refresh_calls = []
        refresh_original = view.refresh

        def tracked_refresh(*args, **kwargs):
            refresh_calls.append((tuple(args), dict(kwargs)))
            return refresh_original(*args, **kwargs)

        view.refresh = tracked_refresh
        try:
            with _post_ready_traps(app, view) as hits:
                print("OVERLAY_STAGE public-region-retry", flush=True)
                view.use_right_btn.invoke()
                _wait(app.root, lambda: len(app.undo_stack) == 1 and dict(app.manual_a_cell_ops) == expected_after, deadline, "public region applied")
                assert refresh_calls == [((), {"row_only": None, "rescan": False})], refresh_calls
                child_group = app.undo_stack[-1]
                children = tuple(child_group.get("actions") or ())
                assert len(children) == 1, child_group
                child = children[0]
                assert child.get("target") == "A" and child.get("sheet") == _SHEET
                cells = tuple(child.get("cells") or ())
                assert len(cells) == _ROWS and {(int(row), int(col)) for row, col, _old_edit, _old_value in cells} == {(row, 2) for _sheet, row, _col in expected_after}
                transaction = child.get("overlay_transaction")
                assert isinstance(transaction, sm.OverlayTransaction) and len(transaction.deltas) == _ROWS
                assert {(int(delta.physical_row), int(delta.physical_col)): delta.after for delta in transaction.deltas} == {(int(row), 2): value for (_sheet, row, _col), value in expected_after.items()}
                _assert_manual_shape(app, expected_after, label="apply")
                _assert_sheet_values(app, expected_after, label="apply")
                _assert_overlay(app.sheet_operation_overlay(_SHEET), expected_after, label="apply")
                _same(mutation_hard, _immutable_identity(app, view, inputs, include_ready_handles=True), "apply mutation hard")
                assert int(view._column_projection_generation) == projection_generation0 + 1
                apply_projection = _projection_triplet(view)
                _assert_rebuilt_projection(apply_projection, ready_projection)
                _assert_cache_key_edit_versions(apply_projection[1].cache_key, ready_cache_key, 1)
                _assert_input_sha(inputs, input_before)
                assert not hits, hits
                refresh_calls.clear()
                assert str(view.undo_btn.cget("state")) == "normal"
                view.undo_btn.invoke()
                _wait(app.root, lambda: len(app.undo_stack) == 0 and len(app.redo_stack) == 1, deadline, "public undo")
                assert refresh_calls == [((), {"row_only": None, "rescan": False})], refresh_calls
                expected_undo = {(_SHEET, row, 2): value for (_sheet, row), value in expected_before.items()}
                _assert_manual_shape(app, expected_undo, label="undo")
                _assert_sheet_values(app, expected_undo, label="undo")
                assert app.sheet_operation_overlay(_SHEET).cells == {}
                _same(mutation_hard, _immutable_identity(app, view, inputs, include_ready_handles=True), "undo mutation hard")
                assert int(view._column_projection_generation) == projection_generation0 + 2
                undo_projection = _projection_triplet(view)
                _assert_rebuilt_projection(undo_projection, ready_projection, apply_projection)
                _assert_cache_key_edit_versions(undo_projection[1].cache_key, ready_cache_key, 2)
                _assert_input_sha(inputs, input_before)
                assert not hits, hits
                refresh_calls.clear()
                assert str(view.redo_btn.cget("state")) == "normal"
                view.redo_btn.invoke()
                _wait(app.root, lambda: len(app.undo_stack) == 1 and len(app.redo_stack) == 0 and dict(app.manual_a_cell_ops) == expected_after, deadline, "public redo")
                assert refresh_calls == [((), {"row_only": None, "rescan": False})], refresh_calls
                _assert_manual_shape(app, expected_after, label="redo")
                _assert_sheet_values(app, expected_after, label="redo")
                _assert_overlay(app.sheet_operation_overlay(_SHEET), expected_after, label="redo")
                _same(mutation_hard, _immutable_identity(app, view, inputs, include_ready_handles=True), "redo mutation hard")
                assert int(view._column_projection_generation) == projection_generation0 + 3
                redo_projection = _projection_triplet(view)
                _assert_rebuilt_projection(redo_projection, ready_projection, apply_projection, undo_projection)
                _assert_cache_key_edit_versions(redo_projection[1].cache_key, ready_cache_key, 3)
                _assert_input_sha(inputs, input_before)
                assert not hits, hits
        finally:
            if refresh_original is not None:
                view.refresh = refresh_original
                refresh_original = None
        print("LARGE_OVERLAY_BATCH_OK " + json.dumps({"selection": selection, "rows": _ROWS, "undo_groups": len(app.undo_stack), "manual_cells": len(app.manual_a_cell_ops)}, sort_keys=True), flush=True)
    except BaseException as exc:
        primary = exc
        raise
    finally:
        errors = []

        def check(label, callback):
            try:
                callback()
            except BaseException as exc:
                errors.append((label, exc))

        if view is not None and refresh_original is not None:
            check("refresh restore", lambda: setattr(view, "refresh", refresh_original))
        check("settings debounce", lambda: _cancel_debounces(app))
        if app is not None:
            check("shutdown", app._shutdown_root)
            check("startup ledger clear", lambda: _assert_app_startup_cleanup(app, expected_stable, startup_ledger))
        elif startup_ledger:
            check("pre-app startup cleanup", lambda: _consume_preapp_startup_cleanup(startup_ledger))
        if inputs:
            check("inputs", lambda: _assert_input_sha(inputs, input_before))
        check("prompt restore", lambda: setattr(sm.SowMergeApp, "_schedule_formula_cache_prompt", original_prompt))
        check("settings path restore", lambda: setattr(sm, "_SETTINGS_PATH", original_settings))
        check("user settings", lambda: _same(original_setting, _setting(original_settings), "user settings"))
        try:
            temporary.cleanup()
            if os.path.lexists(root):
                raise AssertionError(root)
        except BaseException as exc:
            errors.append(("owned temp", exc))
        if errors:
            message = "; ".join(f"{label}: {type(exc).__name__}: {exc}" for label, exc in errors)
            if primary is not None:
                primary.add_note("cleanup secondary: " + message)
            else:
                raise AssertionError(message)


def main(argv=None) -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--case", choices=_CASES)
    parser.add_argument("--list-cases", action="store_true")
    args = parser.parse_args(argv)
    if args.list_cases:
        for case in _CASES:
            print(case)
        return
    selected = (args.case,) if args.case else (_CASE,)
    for case in selected:
        assert case == _CASE
        _run_case()
    print(f"SUITE_OK ({len(selected)} cases)", flush=True)


if __name__ == "__main__":
    main()

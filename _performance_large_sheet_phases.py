"""Fresh-process performance capture for large-sheet comparison phases.

Default measurements are deliberately read-only: startup, selected-Sheet
readiness, cached revisit and virtual scrolling.  Operation/save phases are
opt-in because they mutate only a disposable copy and may require Excel.
Every sample runs in a new Python process so RSS is comparable.
"""

from __future__ import annotations

import argparse
import json
import os
import shutil
import statistics
import subprocess
import sys
import tempfile
import time
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

import psutil
from openpyxl import Workbook, load_workbook

from _large_sheet_oracle_fixtures import REAL_FIXTURES, copy_real_fixture


DEFAULT_PHASES = ("startup", "selected_ready", "cached_revisit", "scroll")
OPTIONAL_PHASES = ("action_1000", "undo", "redo", "save")


def _move_sheet_to_catalog_front(path: str, sheet: str) -> None:
    """Make a disposable fixture's requested Sheet the workbook's first tab.

    This changes only ``xl/workbook.xml`` ordering.  It is used solely to
    exercise the real production entry point which has no initial-sheet hint;
    source workbooks remain read-only and unmodified.
    """
    main = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    with zipfile.ZipFile(path, "r") as source:
        members = {name: source.read(name) for name in source.namelist()}
    workbook = ET.fromstring(members["xl/workbook.xml"])
    sheets = workbook.find(main + "sheets")
    if sheets is None:
        raise AssertionError("fixture workbook has no sheet catalog")
    nodes = list(sheets)
    target = next((node for node in nodes if node.attrib.get("name") == sheet), None)
    if target is None:
        raise AssertionError(f"fixture Sheet missing from catalog: {sheet}")
    sheets.remove(target)
    sheets.insert(0, target)
    members["xl/workbook.xml"] = ET.tostring(
        workbook, encoding="utf-8", xml_declaration=True
    )
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as target_zip:
        for name, payload in members.items():
            target_zip.writestr(name, payload)


def _make_synthetic_pair(root: Path, rows: int = 20_000) -> tuple[str, str, str]:
    paths = []
    for side, value in (("mine", "mine"), ("theirs", "theirs")):
        path = root / f"{side}.xlsx"
        wb = Workbook(write_only=True)
        ws = wb.create_sheet("S1")
        ws.append(("id@id", "value"))
        ws.append(("int", "string"))
        for row in range(1, rows + 1):
            ws.append((row, f"{value}-{row}"))
        wb.save(path)
        wb.close()
        paths.append(str(path))
    return paths[0], paths[1], "S1"


def _pump(root, until: float) -> None:
    while time.monotonic() < until:
        root.update_idletasks()
        root.update()
        time.sleep(0.005)


def _wait_ready(app, sheet: str, timeout: float) -> object:
    deadline = time.monotonic() + timeout
    while time.monotonic() < deadline:
        _pump(app.root, min(deadline, time.monotonic() + 0.05))
        view = app.sheet_views.get(sheet)
        if view is not None and bool(getattr(view, "_data_ready", False)) and bool(getattr(view, "_pair_diff_full_exact", False)):
            return view
    raise TimeoutError(f"Sheet did not become exact READY: {sheet}")


def _phase_worker(args: argparse.Namespace) -> dict[str, object]:
    import sow_merge_tool as sm

    temp = Path(tempfile.mkdtemp(prefix="sow_large_sheet_phase_"))
    if args.real:
        item = next((fixture for fixture in REAL_FIXTURES if fixture.name == args.real), None)
        if item is None:
            raise ValueError(f"unknown real fixture: {args.real}")
        mine, sheet = copy_real_fixture(item, temp / "mine")
        theirs, resolved_theirs = copy_real_fixture(item, temp / "theirs")
        assert sheet == resolved_theirs
        base = None
        if args.three_way:
            base, resolved_base = copy_real_fixture(item, temp / "base")
            assert sheet == resolved_base
        if args.default_to_fixture_sheet:
            _move_sheet_to_catalog_front(str(mine), sheet)
            _move_sheet_to_catalog_front(str(theirs), sheet)
            if args.three_way:
                _move_sheet_to_catalog_front(str(base), sheet)
        mine, theirs = str(mine), str(theirs)
    else:
        mine, theirs, sheet = _make_synthetic_pair(temp, args.rows)
    process = psutil.Process()
    rss_before = process.memory_info().rss
    timings: dict[str, float] = {}
    app = None
    try:
        started = time.perf_counter()
        app_kwargs = {
            "merge_mode": bool(args.three_way),
            "base_path": (str(base) if args.three_way else None),
        }
        # Production callers normally do not know a review Sheet beforehand.
        # Keep that path measurable instead of treating an explicit hint as the
        # only supported lightweight-startup scenario.
        if not args.no_initial_sheet:
            app_kwargs["initial_sheet"] = sheet
        app = sm.SowMergeApp(mine, theirs, **app_kwargs)
        app.root.withdraw()
        timings["startup"] = (time.perf_counter() - started) * 1000
        ready_started = time.perf_counter()
        # The application deliberately starts only its active tab.  Selecting
        # this fixture Sheet is part of the selected-Sheet readiness phase.
        app._select_tab(sheet)
        view = _wait_ready(app, sheet, args.timeout)
        timings["selected_ready"] = (time.perf_counter() - ready_started) * 1000
        # Startup parsing is intentionally excluded from interaction heartbeat
        # evidence; establish a fresh idle baseline before revisit/scroll.
        _pump(app.root, time.monotonic() + 0.11)
        app._ui_heartbeat_max_gap = 0.0
        app._ui_heartbeat_samples = 0
        if args.phase == "cached_revisit":
            started = time.perf_counter()
            if hasattr(view, "_publish_virtual_window"):
                view._publish_virtual_window(int(getattr(view, "_virtual_window_start", 0)))
            else:
                view.refresh(rescan=False)
            timings["cached_revisit"] = (time.perf_counter() - started) * 1000
            _pump(app.root, time.monotonic() + 0.05)
        elif args.phase == "scroll":
            originals = (app.ws_a_val, app.ws_b_val)
            app.ws_a_val = lambda *_: (_ for _ in ()).throw(AssertionError("scroll read Mine worksheet"))
            app.ws_b_val = lambda *_: (_ for _ in ()).throw(AssertionError("scroll read Theirs worksheet"))
            started = time.perf_counter()
            view._yview_both("moveto", "0.75")
            _pump(app.root, time.monotonic() + 0.12)
            # The queued request itself is intentionally cheap; the acceptance
            # quantity is the recorded bounded publication after it settles.
            timings["scroll"] = float(getattr(view, "_last_virtual_render_ms", 0.0) or 0.0)
            timings["scroll_request_settle"] = (time.perf_counter() - started) * 1000
            app.ws_a_val, app.ws_b_val = originals
        elif args.phase in OPTIONAL_PHASES:
            if args.phase == "save":
                # This is an actual accepted cell operation followed by the
                # production safe-save entry point.  Both inputs and its
                # output remain beneath this worker's disposable root.
                app._request_edit_preload()
                edit_deadline = time.monotonic() + args.timeout
                while not app._edit_workbooks_ready() and time.monotonic() < edit_deadline:
                    _pump(app.root, min(edit_deadline, time.monotonic() + 0.05))
                if not app._edit_workbooks_ready():
                    raise TimeoutError("editable backend did not become ready for save phase")
                pair_idx = next(
                    (
                        index
                        for index, pair in enumerate(view.row_pairs)
                        if pair[0] is not None and pair[1] is not None and index >= 2
                    ),
                    None,
                )
                if pair_idx is None:
                    raise RuntimeError("save fixture has no common data pair")
                action_started = time.perf_counter()
                view._copy_single_cell_by_pair(pair_idx, "B2A", 2)
                timings["save_action"] = (time.perf_counter() - action_started) * 1000
                save_started = time.perf_counter()
                output = app.build_manual_merge_output_file()
                timings["save"] = (time.perf_counter() - save_started) * 1000
                output_path = Path(output).resolve()
                disposable_roots = (temp.resolve(), Path(tempfile.gettempdir()).resolve())
                if not any(output_path.is_relative_to(root) for root in disposable_roots):
                    raise AssertionError(f"save output escaped disposable roots: {output_path}")
                package_ok, package_error = sm._validate_xlsx_package(str(output_path))
                if not package_ok:
                    raise AssertionError(f"saved package invalid: {package_error}")
                reopened = load_workbook(output_path, read_only=True, data_only=False)
                reopened.close()
                # The production save helper uses the OS temp root for a
                # staged output.  It has already passed its reopen check, so
                # remove this benchmark-only disposable artifact promptly.
                output_path.unlink()
            if args.phase != "save":
                overlay = app.sheet_operation_overlay(sheet)
                deltas = [
                    sm.OverlayCellDelta(
                        record_key=("benchmark", str(index)),
                        field_key=("logical", "2"), side="A",
                        physical_row=index + 3, physical_col=2,
                        before=f"mine-{index + 1}", after=f"theirs-{index + 1}",
                    )
                    for index in range(1000)
                ]
                if args.phase == "action_1000":
                    started = time.perf_counter()
                    overlay.apply_batch(deltas)
                    timings["action_1000"] = (time.perf_counter() - started) * 1000
                elif args.phase == "undo":
                    transaction = overlay.apply_batch(deltas)
                    started = time.perf_counter()
                    overlay.revert_transaction(transaction)
                    timings["undo"] = (time.perf_counter() - started) * 1000
                else:  # redo
                    transaction = overlay.apply_batch(deltas)
                    overlay.revert_transaction(transaction)
                    started = time.perf_counter()
                    overlay.apply_batch(deltas)
                    timings["redo"] = (time.perf_counter() - started) * 1000
        else:
            timings[args.phase] = timings[args.phase]
        return {
            "status": "ok",
            "phase": args.phase,
            "timings_ms": {name: round(value, 3) for name, value in timings.items()},
            "rss_delta_bytes": process.memory_info().rss - rss_before,
            "rss_after_bytes": process.memory_info().rss,
            "rss_peak_bytes": process.memory_info().rss,
            "fixture": args.real or f"synthetic-{args.rows}",
            "sheet": sheet,
            "snapshot_metrics_ms": dict(getattr(view, "_snapshot_metrics_ms", {}) or {}),
            "virtual_publish_ms": round(float(getattr(view, "_last_virtual_render_ms", 0.0) or 0.0), 3),
            "virtual_publish_phases_ms": dict(getattr(view, "_last_virtual_render_phases_ms", {}) or {}),
            "ui_heartbeat_max_gap_ms": round(float(getattr(app, "_ui_heartbeat_max_gap", 0.0) or 0.0) * 1000.0, 3),
            "pid": os.getpid(),
        }
    finally:
        if app is not None:
            app._shutdown_root()
        shutil.rmtree(temp, ignore_errors=True)


def _sample(args: argparse.Namespace, phase: str) -> dict[str, object]:
    command = [sys.executable, os.path.abspath(__file__), "--worker", "--phase", phase, "--rows", str(args.rows), "--timeout", str(args.timeout)]
    if args.real:
        command.extend(("--real", args.real))
    if args.three_way:
        command.append("--three-way")
    if args.no_initial_sheet:
        command.append("--no-initial-sheet")
    if args.default_to_fixture_sheet:
        command.append("--default-to-fixture-sheet")
    completed = subprocess.run(command, check=True, capture_output=True, text=True)
    return json.loads(completed.stdout.strip().splitlines()[-1])


def _percentile(values: list[float], fraction: float) -> float:
    if not values:
        return 0.0
    ordered = sorted(values)
    return ordered[min(len(ordered) - 1, max(0, round((len(ordered) - 1) * fraction)))]


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--worker", action="store_true")
    parser.add_argument("--phase", choices=DEFAULT_PHASES + OPTIONAL_PHASES)
    parser.add_argument("--phases", default=",".join(DEFAULT_PHASES))
    parser.add_argument("--runs", type=int, default=3)
    parser.add_argument("--rows", type=int, default=20_000)
    parser.add_argument("--real", choices=tuple(item.name for item in REAL_FIXTURES))
    parser.add_argument("--three-way", action="store_true")
    parser.add_argument("--no-initial-sheet", action="store_true")
    parser.add_argument("--default-to-fixture-sheet", action="store_true")
    parser.add_argument("--timeout", type=float, default=120.0)
    parser.add_argument("--out")
    args = parser.parse_args()
    if args.worker:
        print(json.dumps(_phase_worker(args), ensure_ascii=False, sort_keys=True), flush=True)
        return
    phases = tuple(name.strip() for name in args.phases.split(",") if name.strip())
    unknown = set(phases) - set(DEFAULT_PHASES + OPTIONAL_PHASES)
    if unknown:
        parser.error("unknown phase(s): " + ", ".join(sorted(unknown)))
    evidence: dict[str, object] = {"schema": "large-sheet-phase-baseline-v1", "fresh_process": True, "runs": args.runs, "fixture": args.real or f"synthetic-{args.rows}", "phases": {}}
    for phase in phases:
        samples = []
        errors = []
        for _ in range(args.runs):
            try:
                samples.append(_sample(args, phase))
            except Exception as exc:
                errors.append(str(exc))
        values = [float(item["timings_ms"].get(phase, 0)) for item in samples]
        viewport_values = [
            float(item.get("virtual_publish_ms", 0.0) or 0.0)
            for item in samples
        ]
        heartbeat_values = [
            float(item.get("ui_heartbeat_max_gap_ms", 0.0) or 0.0)
            for item in samples
        ]
        rss_values = [
            int(item.get("rss_peak_bytes", 0) or 0)
            for item in samples
        ]
        evidence["phases"][phase] = {
            "samples": samples,
            "p50_ms": round(statistics.median(values), 3) if values else None,
            "p95_ms": round(_percentile(values, .95), 3) if values else None,
            "viewport_publish_p95_ms": round(_percentile(viewport_values, .95), 3) if viewport_values else None,
            "heartbeat_max_gap_p95_ms": round(_percentile(heartbeat_values, .95), 3) if heartbeat_values else None,
            "rss_peak_p95_bytes": _percentile(rss_values, .95) if rss_values else None,
            "errors": errors,
        }
    rendered = json.dumps(evidence, ensure_ascii=False, sort_keys=True, indent=2)
    if args.out:
        Path(args.out).write_text(rendered + "\n", encoding="utf-8")
    print(rendered)


if __name__ == "__main__":
    main()

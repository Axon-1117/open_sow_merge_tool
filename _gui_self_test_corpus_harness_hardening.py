import ctypes
import json
import os
import subprocess
import sys
import tempfile
import threading
import time
from pathlib import Path

from openpyxl import Workbook, load_workbook

import large_sheet_corpus_gui_benchmark as gui


def _assert_atomic_json_hardening(root):
    missing = root / "missing" / "nested" / "checkpoint.json"
    gui._atomic_json(missing, {"generation": 1})
    assert json.loads(missing.read_text(encoding="utf-8")) == {"generation": 1}

    if os.name != "nt":
        return
    kernel32 = ctypes.WinDLL("kernel32", use_last_error=True)
    kernel32.CreateFileW.restype = ctypes.c_void_p
    kernel32.CloseHandle.argtypes = [ctypes.c_void_p]
    handle = kernel32.CreateFileW(
        str(missing),
        0x80000000,
        0x00000001 | 0x00000002,
        None,
        3,
        0x00000080,
        None,
    )
    assert handle not in (0, ctypes.c_void_p(-1).value), ctypes.get_last_error()

    def release_reader():
        time.sleep(0.15)
        kernel32.CloseHandle(handle)

    release = threading.Thread(target=release_reader, daemon=True)
    release.start()
    started = time.perf_counter()
    gui._atomic_json(missing, {"generation": 2})
    elapsed = time.perf_counter() - started
    release.join(timeout=2.0)
    assert elapsed >= 0.10, elapsed
    assert json.loads(missing.read_text(encoding="utf-8")) == {"generation": 2}


def _make_workbook(path):
    workbook = Workbook()
    data = workbook.active
    data.title = "Data"
    data["A1"] = "id"
    data["A2"] = 1
    workbook.create_sheet("Empty")
    styled = workbook.create_sheet("StyledBlank")
    styled["A1"].number_format = "0.00"
    workbook.save(path)
    workbook.close()


def _run_sheet_worker(root, source, mode, sheet):
    scratch = root / f"scratch-{mode}-{sheet}"
    scratch.mkdir()
    mine, theirs, base = gui.copy_sides(source, scratch, mode)
    package_sha256 = {"A": gui._sha256(mine), "B": gui._sha256(theirs)}
    if base:
        package_sha256["BASE"] = gui._sha256(base)
    command = [
        sys.executable,
        str(Path(gui.__file__).resolve()),
        "--sheet-worker",
        "--source",
        str(source),
        "--mode",
        mode,
        "--sheet",
        sheet,
        "--mine",
        str(mine),
        "--theirs",
        str(theirs),
        "--mine-sha256",
        package_sha256["A"],
        "--theirs-sha256",
        package_sha256["B"],
        "--scratch",
        str(scratch),
        "--gate",
        "15",
    ]
    if base:
        command.extend(["--base", str(base), "--base-sha256", package_sha256["BASE"]])
    completed = subprocess.run(command, text=True, capture_output=True, timeout=45, check=True)
    lines = [line for line in completed.stdout.splitlines() if line.strip()]
    assert lines, completed.stderr
    return json.loads(lines[-1])


def _empty_profile(*, styled_blank_cells=0, images=0, drawings=0):
    return {
        "rows": 1,
        "columns": 1,
        "populated_cells": 0,
        "formula_cells": 0,
        "asset_profile": {
            "physical_cell_records": styled_blank_cells,
            "styled_cells": styled_blank_cells,
            "styled_blank_cells": styled_blank_cells,
            "row_dimension_records": 0,
            "column_dimension_records": 0,
            "merged_ranges": 0,
            "hyperlinks": 0,
            "tables": 0,
            "auto_filters": 0,
            "data_validations": 0,
            "conditional_formattings": 0,
            "comments": 0,
            "images": images,
            "drawings": drawings,
            "charts": 0,
            "relationship_type_counts": {"drawing": drawings} if drawings else {},
        },
    }


def _assert_cell_target_applicability_gate():
    missing = {"ok": False, "applicable": True, "reason": "missing physical cache"}
    for three_way in (False, True):
        required = ("A", "B", "BASE") if three_way else ("A", "B")
        hashes = {side: "same-package" for side in required}
        # Image/drawing-only Sheets have no meaningful cell target and remain
        # eligible for N/A; the asset profile itself is still retained.
        profiles = {side: _empty_profile(images=1, drawings=1) for side in required}
        allowed = gui._cell_target_applicability(
            missing,
            final_state=gui.sm._SHEET_EXACT_SAME,
            package_sha256=hashes,
            side_sheet_profiles=profiles,
            three_way=three_way,
        )
        assert allowed["ok"] and allowed["applicable"] is False, allowed
        assert allowed["reason"] == "not_applicable_no_cell_or_structural_assets_identical_inputs", allowed

        unequal = dict(hashes)
        unequal["B"] = "different-package"
        rejected = gui._cell_target_applicability(
            missing,
            final_state=gui.sm._SHEET_EXACT_SAME,
            package_sha256=unequal,
            side_sheet_profiles=profiles,
            three_way=three_way,
        )
        assert not rejected["ok"] and rejected["applicable"] is True, rejected

        with_value = {side: _empty_profile() for side in required}
        with_value["A"]["populated_cells"] = 1
        rejected = gui._cell_target_applicability(
            missing,
            final_state=gui.sm._SHEET_EXACT_SAME,
            package_sha256=hashes,
            side_sheet_profiles=with_value,
            three_way=three_way,
        )
        assert not rejected["ok"] and rejected["applicable"] is True, rejected

        with_formula = {side: _empty_profile() for side in required}
        with_formula["A"]["formula_cells"] = 1
        rejected = gui._cell_target_applicability(
            missing,
            final_state=gui.sm._SHEET_EXACT_SAME,
            package_sha256=hashes,
            side_sheet_profiles=with_formula,
            three_way=three_way,
        )
        assert not rejected["ok"] and rejected["applicable"] is True, rejected

        # Any stored/styled cell or non-image structural/semantic asset still
        # requires a real physical target, even when values/formulas are zero.
        blockers = (
            "physical_cell_records",
            "styled_cells",
            "styled_blank_cells",
            "row_dimension_records",
            "column_dimension_records",
            "merged_ranges",
            "hyperlinks",
            "tables",
            "auto_filters",
            "data_validations",
            "conditional_formattings",
            "comments",
            "charts",
        )
        for blocker in blockers:
            structured = {side: _empty_profile() for side in required}
            structured["A"]["asset_profile"][blocker] = 1
            rejected = gui._cell_target_applicability(
                missing,
                final_state=gui.sm._SHEET_EXACT_SAME,
                package_sha256=hashes,
                side_sheet_profiles=structured,
                three_way=three_way,
            )
            assert not rejected["ok"] and rejected["applicable"] is True, (blocker, rejected)

        other_relationship = {side: _empty_profile() for side in required}
        other_relationship["A"]["asset_profile"]["relationship_type_counts"] = {"oleObject": 1}
        rejected = gui._cell_target_applicability(
            missing,
            final_state=gui.sm._SHEET_EXACT_SAME,
            package_sha256=hashes,
            side_sheet_profiles=other_relationship,
            three_way=three_way,
        )
        assert not rejected["ok"] and rejected["applicable"] is True, rejected


def _assert_empty_sheet_gate(root):
    source = root / "empty-gate.xlsx"
    _make_workbook(source)
    empty_profile = gui._sheet_content_profile(source, "Empty")
    styled_profile = gui._sheet_content_profile(source, "StyledBlank")
    assert empty_profile["populated_cells"] == empty_profile["formula_cells"] == 0
    assert styled_profile["populated_cells"] == styled_profile["formula_cells"] == 0
    assert styled_profile["asset_profile"]["styled_blank_cells"] == 1, styled_profile

    for mode in ("2way", "3way"):
        for sheet in ("Empty", "StyledBlank"):
            row = _run_sheet_worker(root, source, mode, sheet)
            assert row["status"] == "PASS", row
            assert row["final_state"] == "EXACT_SAME", row
            assert row["comparison_detail_ready"], row
            assert row["cell_target_applicability"]["ok"], row
            assert row["cell_target_applicability"]["reason"] in {
                "physical_cell_targets_ready",
                "not_applicable_no_cell_or_structural_assets_identical_inputs",
            }, row
            assert len(set(row["input_package_sha256"].values())) == 1, row
            assert all(
                profile["populated_cells"] == 0 and profile["formula_cells"] == 0
                for profile in row["side_sheet_profiles"].values()
            ), row
            if sheet == "StyledBlank":
                assert all(
                    profile["asset_profile"]["styled_blank_cells"] == 1
                    for profile in row["side_sheet_profiles"].values()
                ), row
            assert row["mutation_backend_state"] == "DEFERRED", row
            assert row["no_calculation_surface"], row


def main():
    with tempfile.TemporaryDirectory(prefix="sow_gui_harness_hardening_") as temporary:
        root = Path(temporary)
        _assert_atomic_json_hardening(root)
        _assert_cell_target_applicability_gate()
        _assert_empty_sheet_gate(root)
    print("PASS: atomic retry/fsync and strict cell-target applicability GUI gates")


if __name__ == "__main__":
    main()

"""Focused no-GUI regression for startup-created workbook ownership."""
from __future__ import annotations

import argparse
import copy
import hashlib
import os
import sqlite3
import stat
import sys
import tempfile
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch
from xml.etree import ElementTree as ET

from openpyxl import Workbook

import sow_merge_tool as sm


_CASE = "startup-sidecar-ownership"
_FIXTURE_UTC = datetime(2024, 1, 2, 3, 4, 5, tzinfo=timezone.utc)
_FIXTURE_MODIFIED_TEXT = "2024-01-02T03:04:05Z"
_CORE_PROPERTIES_MEMBER = "docProps/core.xml"
_DCTERMS_NS = "http://purl.org/dc/terms/"
_CORE_METADATA_FIELDS = (
    "date_time",
    "compress_type",
    "comment",
    "extra",
    "internal_attr",
    "external_attr",
    "create_system",
    "create_version",
    "extract_version",
    "flag_bits",
    "volume",
)


def _sha256(path: Path | str) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _write_workbook(path: Path, marker: str) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet["A1"] = marker
    workbook.save(path)
    workbook.close()


def _write_typed_workbook(path: Path, rows: tuple[tuple[object, ...], ...]) -> None:
    workbook = Workbook()
    workbook.properties.creator = "sow-startup-sidecar-test"
    workbook.properties.lastModifiedBy = "sow-startup-sidecar-test"
    workbook.properties.created = _FIXTURE_UTC
    workbook.properties.modified = _FIXTURE_UTC
    workbook.properties.revision = "1"
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    for row_index, row in enumerate(rows, start=1):
        for column_index, value in enumerate(row, start=1):
            worksheet.cell(row=row_index, column=column_index).value = value
    workbook.save(path)
    workbook.close()
    _canonicalize_core_modified(path)


def _core_properties_bytes(path: Path) -> bytes:
    with zipfile.ZipFile(path) as package:
        return package.read(_CORE_PROPERTIES_MEMBER)


def _zipinfo_metadata(info: zipfile.ZipInfo) -> tuple[object, ...]:
    return tuple(getattr(info, field) for field in _CORE_METADATA_FIELDS)


def _canonicalize_core_modified(path: Path) -> None:
    """Rewrite only core modified time while preserving the workbook ZIP shape."""
    temporary_fd, temporary_name = tempfile.mkstemp(
        prefix=f".{path.name}.core-",
        suffix=".zip",
        dir=path.parent,
    )
    os.close(temporary_fd)
    temporary = Path(temporary_name)
    try:
        with zipfile.ZipFile(path, "r") as original:
            names = original.namelist()
            assert names.count(_CORE_PROPERTIES_MEMBER) == 1, names
            archive_comment = original.comment
            infos = tuple(original.infolist())
            contents = {info.filename: original.read(info.filename) for info in infos}
        core_root = ET.fromstring(contents[_CORE_PROPERTIES_MEMBER])
        modified = core_root.findall(f"{{{_DCTERMS_NS}}}modified")
        assert len(modified) == 1, modified
        modified[0].text = _FIXTURE_MODIFIED_TEXT
        contents[_CORE_PROPERTIES_MEMBER] = ET.tostring(
            core_root,
            encoding="utf-8",
            xml_declaration=False,
        )
        with zipfile.ZipFile(temporary, "w") as rebuilt:
            rebuilt.comment = archive_comment
            for info in infos:
                preserved = copy.copy(info)
                rebuilt.writestr(
                    preserved,
                    contents[info.filename],
                    compress_type=info.compress_type,
                )
        with zipfile.ZipFile(temporary, "r") as rebuilt:
            rebuilt_infos = tuple(rebuilt.infolist())
            assert rebuilt.namelist() == names, (rebuilt.namelist(), names)
            assert rebuilt.comment == archive_comment
            assert len(rebuilt_infos) == len(infos)
            for before, after in zip(infos, rebuilt_infos):
                assert after.filename == before.filename
                assert _zipinfo_metadata(after) == _zipinfo_metadata(before), (
                    before.filename,
                    _zipinfo_metadata(before),
                    _zipinfo_metadata(after),
                )
                if before.filename != _CORE_PROPERTIES_MEMBER:
                    assert rebuilt.read(after.filename) == contents[before.filename], before.filename
            rebuilt_root = ET.fromstring(rebuilt.read(_CORE_PROPERTIES_MEMBER))
            rebuilt_modified = rebuilt_root.findall(f"{{{_DCTERMS_NS}}}modified")
            assert len(rebuilt_modified) == 1
            assert rebuilt_modified[0].text == _FIXTURE_MODIFIED_TEXT
        os.replace(temporary, path)
    except BaseException:
        if temporary.exists():
            temporary.unlink()
        raise


def _write_sidecar(root: Path, name: str, marker: str) -> Path:
    direct = root / f"{name}.xlsx"
    sidecar = root / f"{name}.xlsx.r123"
    _write_workbook(direct, marker)
    os.replace(direct, sidecar)
    return sidecar


def _make_wc_pristine(root: Path, marker: str) -> tuple[Path, Path]:
    working = root / "working.xlsx"
    _write_workbook(working, marker)
    payload = working.read_bytes()
    digest = hashlib.sha1(payload).hexdigest()
    pristine = root / ".svn" / "pristine" / digest[:2] / f"{digest}.svn-base"
    pristine.parent.mkdir(parents=True)
    pristine.write_bytes(payload)
    db_path = root / ".svn" / "wc.db"
    with sqlite3.connect(db_path) as conn:
        conn.execute(
            "create table NODES (local_relpath text, op_depth integer, kind text, "
            "presence text, checksum text)"
        )
        conn.execute(
            "insert into NODES values (?, ?, ?, ?, ?)",
            ("working.xlsx", 0, "file", "normal", f"$sha1${digest}"),
        )
    return working, pristine


def _assert_success_paths_are_owned_once(root: Path) -> None:
    root.mkdir(parents=True, exist_ok=True)
    ledger: set[str] = set()
    sidecar = _write_sidecar(root, "source", "sidecar-source")
    sidecar_sha = _sha256(sidecar)
    os.chmod(sidecar, stat.S_IREAD)
    neighbor = root / "unowned-neighbor.xlsx"
    _write_workbook(neighbor, "neighbor")
    neighbor_sha = _sha256(neighbor)
    candidate_source = root / "candidate-source.xlsx"
    _write_workbook(candidate_source, "candidate-source")
    candidate_source_sha = _sha256(candidate_source)
    working, pristine_blob = _make_wc_pristine(root, "wc-pristine")
    working_sha = _sha256(working)
    pristine_sha = _sha256(pristine_blob)

    with patch.object(sm.tempfile, "gettempdir", lambda: str(root)):
        effective = sm._ensure_xlsx_copy(str(sidecar), owned_paths=ledger)
        candidate = sm._create_startup_candidate_copy(
            str(candidate_source),
            "mine",
            owned_paths=ledger,
        )
        pristine = sm._try_export_svn_base_from_working_copy(
            str(working),
            owned_paths=ledger,
        )

    assert pristine is not None
    expected = {
        os.path.normcase(os.path.abspath(effective)),
        os.path.normcase(os.path.abspath(candidate)),
        os.path.normcase(os.path.abspath(pristine)),
    }
    assert ledger == expected, (ledger, expected)
    assert len(expected) == 3
    assert _sha256(effective) == sidecar_sha
    assert _sha256(candidate) == candidate_source_sha
    assert _sha256(pristine) == pristine_sha
    assert os.path.normcase(os.path.abspath(neighbor)) not in ledger

    shutdown_evidence: list[dict] = []
    first = sm._consume_owned_startup_temp_paths(ledger, shutdown_evidence)
    assert len(first) == 3, first
    assert tuple(shutdown_evidence) == first, shutdown_evidence
    assert {
        fact["path"] for fact in first
    } == expected, first
    assert all(
        fact["removed"] and not fact["exists_after"] and not fact["error"]
        for fact in first
    ), first
    assert all(not os.path.exists(path) for path in expected)
    assert ledger == set(), ledger
    assert _sha256(sidecar) == sidecar_sha
    assert _sha256(neighbor) == neighbor_sha
    assert _sha256(candidate_source) == candidate_source_sha
    assert _sha256(working) == working_sha
    assert _sha256(pristine_blob) == pristine_sha

    second = sm._consume_owned_startup_temp_paths(ledger, shutdown_evidence)
    assert second == (), second
    assert tuple(shutdown_evidence) == first, shutdown_evidence
    os.chmod(sidecar, stat.S_IWRITE)


def _assert_copy_validation_failure_rolls_back(root: Path) -> None:
    root.mkdir(parents=True, exist_ok=True)
    sidecar = _write_sidecar(root, "bad-source", "bad-source")
    sidecar_sha = _sha256(sidecar)
    ledger: set[str] = set()
    with (
        patch.object(sm.tempfile, "gettempdir", lambda: str(root)),
        patch.object(sm, "_workbook_package_ready", lambda _path: False),
    ):
        try:
            sm._ensure_xlsx_copy(str(sidecar), owned_paths=ledger)
        except RuntimeError:
            pass
        else:
            raise AssertionError("invalid sidecar copy unexpectedly succeeded")
    assert ledger == set(), ledger
    assert _sha256(sidecar) == sidecar_sha
    assert tuple(root.glob(f"{sm.APP_NAME}_svn_*")) == ()


def _assert_main_preapp_failure_cleans_exact_ledger(root: Path) -> None:
    root.mkdir(parents=True, exist_ok=True)
    base = _write_sidecar(root, "merge-left", "base")
    mine = root / "mine.xlsx"
    theirs = root / "theirs.xlsx"
    merged = root / "merged.xlsx"
    _write_workbook(mine, "mine")
    _write_workbook(theirs, "theirs")
    _write_workbook(merged, "merged")
    before = {path: _sha256(path) for path in (base, mine, theirs, merged)}
    observed = SimpleNamespace(effective=None, evidence=(), app_ledger=None)

    def _analysis(context, *, owned_startup_paths=None):
        observed.effective = sm._ensure_xlsx_copy(
            context.identity_for("base").path,
            owned_paths=owned_startup_paths,
        )
        return sm.StartupMergeAnalysis(
            context,
            sm.EquivalenceMatrix(),
            sm.StartupMergeOutcome(),
            [],
            {},
        )

    class _FailApp:
        def __init__(self, *_args, **kwargs):
            observed.app_ledger = kwargs.get("startup_owned_paths")
            raise RuntimeError("forced-app-construction-failure")

    original_cleanup = sm._cleanup_unclaimed_startup_temp_paths

    def _observe_cleanup(paths, *, reason):
        observed.evidence = original_cleanup(paths, reason=reason)
        return observed.evidence

    def _startup(_title, _message, worker):
        return worker(lambda *_args, **_kwargs: None)

    argv = [
        "sow_merge_tool.py",
        "--base",
        str(base),
        "--mine",
        str(mine),
        "--theirs",
        str(theirs),
        "--merged",
        str(merged),
    ]
    with (
        patch.object(sys, "argv", argv),
        patch.object(sm.tempfile, "gettempdir", lambda: str(root)),
        patch.object(sm, "run_startup_merge_analysis", _analysis),
        patch.object(sm, "_run_startup_progress_task", _startup),
        patch.object(sm, "SowMergeApp", _FailApp),
        patch.object(sm, "_cleanup_unclaimed_startup_temp_paths", _observe_cleanup),
        patch.object(sm.messagebox, "showerror", lambda *_args, **_kwargs: None),
        patch.object(sm, "_SETTINGS_PATH", str(root / "settings.json")),
    ):
        try:
            sm.main()
        except SystemExit as exc:
            assert exc.code == 1, exc.code
        else:
            raise AssertionError("forced app construction failure did not stop main")

    effective = os.path.normcase(os.path.abspath(observed.effective))
    assert observed.app_ledger is not None
    assert observed.app_ledger == set(), observed.app_ledger
    assert len(observed.evidence) == 1, observed.evidence
    fact = observed.evidence[0]
    assert fact["path"] == effective and fact["removed"], fact
    assert not fact["exists_after"] and not fact["error"], fact
    assert not os.path.exists(effective)
    assert {path: _sha256(path) for path in before} == before


def _assert_textdiff_sidecars_are_main_owned(root: Path) -> None:
    root.mkdir(parents=True, exist_ok=True)
    left = _write_sidecar(root, "left", "left")
    right = _write_sidecar(root, "right", "right")
    neighbor = root / "unowned-neighbor.xlsx"
    _write_workbook(neighbor, "neighbor")
    before = {path: _sha256(path) for path in (left, right, neighbor)}
    observed = SimpleNamespace(sources=[], cleanup=(), tortoise=())
    original_cleanup = sm._cleanup_unclaimed_startup_temp_paths

    def _excel_to_text(source, target, **_kwargs):
        observed.sources.append(os.path.normcase(os.path.abspath(source)))
        assert source.lower().endswith(".xlsx"), source
        assert os.path.isfile(source), source
        assert target.lower().endswith(".txt"), target

    def _observe_cleanup(paths, *, reason):
        assert reason == "main-pre-app", reason
        observed.cleanup = original_cleanup(paths, reason=reason)
        return observed.cleanup

    def _open_tortoise(first, second, **_kwargs):
        observed.tortoise = (first, second)

    argv = ["sow_merge_tool.py", str(left), str(right), "--textdiff"]
    with (
        patch.object(sys, "argv", argv),
        patch.object(sm.tempfile, "gettempdir", lambda: str(root)),
        patch.dict(os.environ, {"LOCALAPPDATA": str(root)}, clear=False),
        patch.object(sm, "resolve_svn_author_metadata", lambda context: context.identities),
        patch.object(sm, "excel_to_text", _excel_to_text),
        patch.object(
            sm,
            "open_tortoise_merge",
            _open_tortoise,
        ),
        patch.object(sm, "_cleanup_unclaimed_startup_temp_paths", _observe_cleanup),
        patch.object(
            sm,
            "SowMergeApp",
            lambda *_args, **_kwargs: (_ for _ in ()).throw(
                AssertionError("textdiff must not construct SowMergeApp")
            ),
        ),
    ):
        sm.main()

    assert len(observed.sources) == 2, observed.sources
    assert len(set(observed.sources)) == 2, observed.sources
    assert len(observed.tortoise) == 2, observed.tortoise
    assert len(observed.cleanup) == 2, observed.cleanup
    assert {fact["path"] for fact in observed.cleanup} == set(observed.sources)
    assert all(
        fact["removed"] and not fact["exists_after"] and not fact["error"]
        for fact in observed.cleanup
    ), observed.cleanup
    assert all(not os.path.exists(path) for path in observed.sources)
    assert {path: _sha256(path) for path in before} == before


def _assert_scan_merge_and_cross_candidate_share_ledger(root: Path) -> None:
    root.mkdir(parents=True, exist_ok=True)
    base = root / "base.xlsx"
    mine = root / "mine.xlsx"
    theirs = root / "theirs.xlsx"
    merged = root / "merged.xlsx"
    stable_rows = (("id@id", "value"), ("string", "string"), ("row-1", "old"))
    changed_rows = (("id@id", "value"), ("string", "string"), ("row-1", "new"))
    for path in (base, mine, theirs):
        _write_typed_workbook(path, stable_rows)
    ledger: set[str] = set()
    ensure_calls: list[set[str] | None] = []
    original_ensure = sm._ensure_xlsx_copy

    def _record_ensure(path, *, owned_paths=None):
        ensure_calls.append(owned_paths)
        return original_ensure(path, owned_paths=owned_paths)

    with (
        patch.object(sm.tempfile, "gettempdir", lambda: str(root)),
        patch.object(sm, "_ensure_xlsx_copy", _record_ensure),
    ):
        _conflicts, preview, _map = sm._merge_three_way(
            str(base),
            str(mine),
            str(theirs),
            str(merged),
            save_merged=False,
            owned_startup_paths=ledger,
        )
    assert len(ensure_calls) >= 6 and all(call is ledger for call in ensure_calls), ensure_calls
    assert preview is not None
    preview_path = os.path.normcase(os.path.abspath(preview))
    assert preview_path in ledger and os.path.isfile(preview_path), (preview_path, ledger)
    assert all(
        fact["removed"] and not fact["exists_after"]
        for fact in sm._cleanup_owned_startup_temp_paths(ledger)
    )
    ledger.clear()

    _write_typed_workbook(theirs, changed_rows)
    assert _core_properties_bytes(base) == _core_properties_bytes(theirs), (
        "cross-premerge fixture must differ only in the intended worksheet cell",
        _core_properties_bytes(base),
        _core_properties_bytes(theirs),
    )
    ensure_calls.clear()
    with (
        patch.object(sm.tempfile, "gettempdir", lambda: str(root)),
        patch.object(sm, "_ensure_xlsx_copy", _record_ensure),
    ):
        _conflicts, candidate, _map, _summary, reason = sm._cross_branch_source_delta_premerge(
            str(base),
            str(mine),
            str(theirs),
            owned_startup_paths=ledger,
        )
    assert reason is None and candidate is not None, (reason, candidate)
    candidate_path = os.path.normcase(os.path.abspath(candidate))
    assert os.path.basename(candidate_path).startswith(
        f"{sm.APP_NAME}_cross_branch_candidate_"
    ), candidate_path
    assert candidate_path in ledger and os.path.isfile(candidate_path), ledger
    assert len(ensure_calls) == 3 and all(call is ledger for call in ensure_calls), ensure_calls
    assert all(
        fact["removed"] and not fact["exists_after"]
        for fact in sm._cleanup_owned_startup_temp_paths(ledger)
    )
    ledger.clear()

    with (
        patch.object(sm.tempfile, "gettempdir", lambda: str(root)),
        patch.object(sm, "_atomic_save_wb", side_effect=RuntimeError("forced-atomic-failure")),
    ):
        _conflicts, fallback, _map, _summary, reason = sm._cross_branch_source_delta_premerge(
            str(base),
            str(mine),
            str(theirs),
            owned_startup_paths=ledger,
        )
    assert reason is not None and fallback is not None, (reason, fallback)
    assert tuple(root.glob(f"{sm.APP_NAME}_cross_branch_candidate_*")) == ()
    assert all(
        fact["removed"] and not fact["exists_after"]
        for fact in sm._cleanup_owned_startup_temp_paths(ledger)
    )


def run_case() -> None:
    with tempfile.TemporaryDirectory(prefix="sow_startup_sidecar_ownership_") as raw_root:
        root = Path(raw_root)
        _assert_success_paths_are_owned_once(root / "success")
        _assert_copy_validation_failure_rolls_back(root / "rollback")
        _assert_main_preapp_failure_cleans_exact_ledger(root / "preapp")
        _assert_textdiff_sidecars_are_main_owned(root / "textdiff")
        _assert_scan_merge_and_cross_candidate_share_ledger(root / "premerge")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--case", choices=(_CASE,))
    parser.add_argument("--list-cases", action="store_true")
    args = parser.parse_args()
    if args.list_cases:
        print(_CASE)
        return
    run_case()
    print("PASS " + (args.case or _CASE))


if __name__ == "__main__":
    main()

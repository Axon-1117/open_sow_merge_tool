import os
import time

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

import sow_merge_tool as mod
from _test_temp_utils import make_temp_dir


def _make_book(path, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    for row in rows:
        ws.append(row)
    wb.save(path)
    wb.close()


def _pump(root, seconds=0.2):
    deadline = time.time() + seconds
    while time.time() < deadline:
        root.update_idletasks()
        root.update()
        time.sleep(0.01)


def _prepare_view(app, *, only_diff=True):
    app.root.deiconify()
    app.nb.select(app._sheet_containers["Data"])
    _pump(app.root)
    view = app.sheet_views["Data"]
    view._suppress_bg_apply = True
    view.only_diff_var.set(1 if only_diff else 0)
    view.refresh(row_only=None, rescan=True)
    _pump(app.root, 0.05)
    return view


def _send_shortcut(view, widget, sequence):
    widget.focus_force()
    view.root.update_idletasks()
    widget.event_generate(sequence, when="tail")
    _pump(view.root, 0.05)


def _text_line_count(widget):
    return int(str(widget.index("end-1c")).split(".")[0])


def _assert_materialized_panes(view, widgets, target_line, expected_rows):
    expected_text_lines = int(expected_rows) + 1
    view.root.update_idletasks()
    for widget in widgets:
        assert _text_line_count(widget) == expected_text_lines, (
            widget,
            _text_line_count(widget),
            expected_text_lines,
        )
        bbox = widget.bbox(f"{target_line}.0")
        first, last = (float(value) for value in widget.yview())
        target_fraction = (int(target_line) - 1) / max(1, int(expected_rows))
        assert bbox is not None or first <= target_fraction <= last, (
            widget,
            target_line,
            widget.yview(),
        )
        assert "blockstart" in widget.tag_names(f"{target_line}.0"), widget
    yviews = [float(widget.yview()[0]) for widget in widgets]
    assert max(yviews) - min(yviews) < 0.002, yviews


def _diff_channel_summary(view):
    raw = {idx: sorted(cols) for idx, cols in view.pair_diff_cols.items() if cols}
    base = {idx: sorted(cols) for idx, cols in view.pair_base_diff_cols.items() if cols}
    visual = [
        idx for idx in range(len(view.row_pairs)) if view._pair_has_visual_diff(idx)
    ]
    projection = view._active_column_projection()
    return {
        "row_pairs": len(view.row_pairs),
        "full_display_rows": len(view._full_display_rows),
        "raw_nonempty": len(raw),
        "raw_samples": list(raw.items())[:4] + list(raw.items())[-4:],
        "base_nonempty": len(base),
        "base_samples": list(base.items())[:4] + list(base.items())[-4:],
        "visual_count": len(visual),
        "visual_samples": visual[:4] + visual[-4:],
        "structural": sorted(view.column_comparison_cache.structural_diff_cols),
        "unresolved": sorted(view.column_comparison_cache.unresolved_cols),
        "slots": [
            (slot.logical_idx + 1, slot.mine_col, slot.base_col, slot.theirs_col, slot.state)
            for slot in projection.model.slots
        ],
        "pair_diff_full_exact": bool(view._pair_diff_full_exact),
        "snapshot_only_diff": bool(view.snapshot_only_diff),
        "seed_last": getattr(view, "_column_diff_seed_last", None),
    }


def _test_two_way_block_presentation(root_dir):
    mine = os.path.join(root_dir, "two-mine.xlsx")
    theirs = os.path.join(root_dir, "two-theirs.xlsx")
    rows_a = [
        [idx, f"same-{idx}"]
        + [f"payload-{col}-{idx}-" + ("x" * 80) for col in range(3, 13)]
        for idx in range(1, 11)
    ]
    rows_b = [list(row) for row in rows_a]
    for excel_row in (2, 3, 7, 8):
        rows_b[excel_row - 1][1] = f"changed-{excel_row}"
    _make_book(mine, rows_a)
    _make_book(theirs, rows_b)

    app = mod.SowMergeApp(mine, theirs)
    try:
        app.root.state("normal")
        app.root.geometry("900x1000")
        view = _prepare_view(app)
        blocks = view._ensure_full_diff_blocks()
        assert [(b.start_pair_idx, b.end_pair_idx) for b in blocks] == [(1, 2), (6, 7)]
        assert "1/2" in view.diff_block_status_var.get(), view.diff_block_status_var.get()
        assert "待处理 2" in view.diff_block_status_var.get(), view.diff_block_status_var.get()
        assert view.diff_block_status.winfo_manager() == "pack"
        assert view.left_ln.get("1.0", "1.end").lstrip().startswith("[1]")
        assert view.left_ln.get("3.0", "3.end").lstrip().startswith("[2]")

        # Calculating and empty states must not publish a provisional count.
        view._only_diff_async_building = True
        view._update_diff_nav_state()
        assert view.diff_block_status_var.get() == "差异块 计算中..."
        assert str(view.prev_diff_btn.cget("state")) == "disabled"
        assert str(view.next_diff_btn.cget("state")) == "disabled"
        view._only_diff_async_building = False
        saved_full_rows = view._full_display_rows
        view._full_display_rows = []
        view._invalidate_diff_block_model()
        view._update_diff_nav_state()
        assert view.diff_block_status_var.get() == "差异块 -/0 · 待处理 0"
        view._full_display_rows = saved_full_rows
        view._invalidate_diff_block_model()
        view._update_diff_nav_state()

        # Hover is a preview only and must not switch the active block counter.
        initial_status = view.diff_block_status_var.get()
        view.hover_pair_idx = 6
        view._update_diff_nav_state()
        assert view.diff_block_status_var.get() == initial_status

        # Row-header hover text and row-only redraw both preserve the marker.
        rn_w = view._rownum_render_width()
        view._set_row_header_text(
            view.left_ln,
            1,
            view._format_main_row_header(1, "A", rn_w, arrow="->"),
        )
        assert view.left_ln.get("1.0", "1.end").lstrip().startswith("[1]")
        assert "blockmarker" in view.left_ln.tag_names("1.0")
        view.refresh(row_only=2, rescan=False)
        assert view.left_ln.get("1.0", "1.end").lstrip().startswith("[1]")

        for widget in (view.left, view.right, view.left_ln, view.right_ln):
            assert int(widget.tag_cget("blockstart", "spacing1")) == 8
            starts = [str(index) for index in widget.tag_ranges("blockstart")]
            assert starts[:2] == ["3.0", "4.0"], (widget, starts)

        # Resolve one member, then adopt the stable block. The resolved member
        # is skipped and the next non-contiguous block must stay untouched.
        assert view._copy_selected_row("B2A", override_pair_idx=1, override_cols={2})
        assert app.ws_a_edit("Data").cell(row=2, column=2).value == "changed-2"
        view._select_line(view.row_to_line[1])
        view.hover_pair_idx = 6
        view._copy_selected_region("B2A")
        assert app.ws_a_edit("Data").cell(row=3, column=2).value == "changed-3"
        assert app.ws_a_edit("Data").cell(row=7, column=2).value == "same-7"
        assert "1/2" in view.diff_block_status_var.get(), view.diff_block_status_var.get()
        assert "待处理 1" in view.diff_block_status_var.get(), view.diff_block_status_var.get()
        assert "已处理" in view.diff_block_status_var.get(), view.diff_block_status_var.get()
        assert view.selected_pair_idx == 1

        view._undo_last_action()
        assert app.ws_a_edit("Data").cell(row=3, column=2).value == "same-3"
        assert [(b.start_pair_idx, b.end_pair_idx) for b in view._ensure_full_diff_blocks()] == [(1, 2), (6, 7)]

        view._sync_main_x_to_frac(0.35)
        _pump(app.root, 0.05)
        saved_x = float(view.left.xview()[0])
        assert saved_x > 0.01, view.left.xview()
        # Exercise the real widget bindings. Ctrl+N moves forward, remains at
        # the last endpoint, and does not disturb synchronized horizontal view.
        _send_shortcut(view, view.left, "<Control-n>")
        assert view.selected_pair_idx == 6, view.selected_pair_idx
        assert "2/2" in view.diff_block_status_var.get(), view.diff_block_status_var.get()
        assert str(view.next_diff_btn.cget("state")) == "disabled"
        assert str(view.prev_diff_btn.cget("state")) == "normal"
        assert abs(float(view.left.xview()[0]) - saved_x) < 0.02, (saved_x, view.left.xview())
        _send_shortcut(view, view.left, "<Control-n>")
        assert view.selected_pair_idx == 6
        assert abs(float(view.left.xview()[0]) - saved_x) < 0.02
        assert view._on_next_diff_block_shortcut() == "break"
        assert view.selected_pair_idx == 6

        _send_shortcut(view, view.left, "<Control-p>")
        assert view.selected_pair_idx == 1
        assert "1/2" in view.diff_block_status_var.get()
        assert str(view.prev_diff_btn.cget("state")) == "disabled"
        _send_shortcut(view, view.left, "<Shift-F4>")
        assert view.selected_pair_idx == 1
        assert abs(float(view.left.xview()[0]) - saved_x) < 0.02
        assert view._on_prev_diff_block_shortcut() == "break"
        assert view.selected_pair_idx == 1

        # Plain F4 keeps its existing hover-panel pin/unpin behavior.
        pin_before = int(view.hover_cmp_pin_var.get())
        _send_shortcut(view, view.left, "<F4>")
        assert int(view.hover_cmp_pin_var.get()) == (0 if pin_before else 1)

        view.only_diff_var.set(0)
        view._refresh_mode_switch_preserving_selection(rescan=False)
        _pump(app.root, 0.05)
        assert not view.diff_block_status.winfo_manager()
        assert not view.left.tag_ranges("blockstart")
        assert not view.left_ln.tag_ranges("blockmarker")
    finally:
        app._shutdown_root()


def _test_three_way_base_only_and_processed_status(root_dir):
    base = os.path.join(root_dir, "three-base.xlsx")
    mine = os.path.join(root_dir, "three-mine.xlsx")
    theirs = os.path.join(root_dir, "three-theirs.xlsx")
    rows_base = [
        [idx, f"same-{idx}"]
        + [f"payload-{col}-{idx}-" + ("x" * 80) for col in range(3, 13)]
        for idx in range(1, 9)
    ]
    rows_mine = [list(row) for row in rows_base]
    rows_theirs = [list(row) for row in rows_base]
    rows_base[1][1] = "base-old"
    rows_mine[1][1] = "shared-new"
    rows_theirs[1][1] = "shared-new"
    rows_theirs[6][1] = "theirs-change"
    rows_theirs.append([9, "theirs-only"])
    _make_book(base, rows_base)
    _make_book(mine, rows_mine)
    _make_book(theirs, rows_theirs)

    app = mod.SowMergeApp(mine, theirs, merge_mode=True, base_path=base)
    try:
        app.root.state("normal")
        app.root.geometry("900x1000")
        view = _prepare_view(app)
        blocks = view._ensure_full_diff_blocks()
        assert [(b.start_pair_idx, b.end_pair_idx) for b in blocks] == [(1, 1), (6, 6), (8, 8)]
        assert view.pair_base_diff_cols.get(1), view.pair_base_diff_cols
        assert view.pair_diff_cols.get(8) == {-1}, view.pair_diff_cols.get(8)
        for widget in (view.left, view.base, view.right, view.left_ln, view.base_ln, view.right_ln):
            assert "blockstart" in widget.tag_names("2.0"), widget
            assert "blockstart" in widget.tag_names("3.0"), widget

        view.selected_pair_idx = 1
        view._last_selected_line = 1
        view.pair_base_diff_cols[1] = set()
        view._invalidate_render_cache()
        view._update_diff_nav_state()
        assert "1/3" in view.diff_block_status_var.get(), view.diff_block_status_var.get()
        assert "待处理 2" in view.diff_block_status_var.get(), view.diff_block_status_var.get()
        assert "已处理" in view.diff_block_status_var.get(), view.diff_block_status_var.get()

        # Ordinary 3-way mode receives the same bindings on its Base pane.
        view.selected_pair_idx = None
        view._last_selected_line = None
        view._sync_main_x_to_frac(0.35)
        _pump(app.root, 0.05)
        saved_x = float(view.left.xview()[0])
        assert saved_x > 0.01, view.left.xview()
        _send_shortcut(view, view.base, "<Control-n>")
        assert view.selected_pair_idx == 6
        _send_shortcut(view, view.base, "<Control-n>")
        assert view.selected_pair_idx == 8
        assert str(view.next_diff_btn.cget("state")) == "disabled"
        _send_shortcut(view, view.base, "<Control-n>")
        assert view.selected_pair_idx == 8
        assert abs(float(view.left.xview()[0]) - saved_x) < 0.02
        _send_shortcut(view, view.base, "<Control-p>")
        assert view.selected_pair_idx == 6
        _send_shortcut(view, view.base, "<Shift-F4>")
        assert view.selected_pair_idx == 1
        assert str(view.prev_diff_btn.cget("state")) == "disabled"
        _send_shortcut(view, view.base, "<Shift-F4>")
        assert view.selected_pair_idx == 1
        assert abs(float(view.base.xview()[0]) - saved_x) < 0.02

        # The structural third block adopts only its one-sided row; the
        # noncontiguous theirs-change block immediately before it stays intact.
        view._select_line(view.row_to_line[8])
        view._copy_selected_region("B2A")
        assert app.ws_a_edit("Data").cell(row=9, column=2).value == "theirs-only"
        assert app.ws_a_edit("Data").cell(row=7, column=2).value == "same-7"
    finally:
        app._shutdown_root()


def _test_large_cross_render_limit_navigation(root_dir):
    mine = os.path.join(root_dir, "large-mine.xlsx")
    theirs = os.path.join(root_dir, "large-theirs.xlsx")
    rows_a = []
    rows_b = []
    for idx in range(1, 2101):
        tail = [f"payload-{col}-{idx}-" + ("x" * 24) for col in range(3, 11)]
        rows_a.append([idx, f"same-{idx}"] + tail)
        changed = idx <= 450 or 700 <= idx <= 1100 or 1800 <= idx <= 1801
        rows_b.append([idx, f"changed-{idx}" if changed else f"same-{idx}"] + tail)
    _make_book(mine, rows_a)
    _make_book(theirs, rows_b)

    app = mod.SowMergeApp(mine, theirs)
    try:
        app.root.state("normal")
        app.root.geometry("800x600")
        view = _prepare_view(app)
        blocks = view._ensure_full_diff_blocks()
        assert len(blocks) == 3, [(b.start_pair_idx, b.end_pair_idx) for b in blocks]
        assert blocks[2].start_pair_idx == 1799
        assert len(view._full_display_rows) == 853, len(view._full_display_rows)
        assert len(view.display_rows) < 800, len(view.display_rows)

        rescan_calls = []
        original_refresh = view.refresh

        def _tracked_refresh(row_only, rescan):
            rescan_calls.append(bool(rescan))
            return original_refresh(row_only, rescan)

        view.refresh = _tracked_refresh
        view._sync_main_x_to_frac(0.35)
        _pump(app.root, 0.05)
        saved_x = float(view.left.xview()[0])
        assert saved_x > 0.01, view.left.xview()
        started = time.perf_counter()
        view._goto_full_diff_block(2)
        elapsed = time.perf_counter() - started
        assert view.selected_pair_idx == 1799, view.selected_pair_idx
        assert 1799 in view.row_to_line
        assert view.display_rows == [1799, 1800], view.display_rows
        assert len(view.display_rows) <= mod._LARGE_DIFF_NAV_PREVIEW_ROWS
        assert True not in rescan_calls, rescan_calls
        assert abs(float(view.left.xview()[0]) - saved_x) < 0.02, (saved_x, view.left.xview())
        assert "3/3" in view.diff_block_status_var.get(), view.diff_block_status_var.get()
        assert elapsed < 0.25, elapsed
        target_line = view.row_to_line[1799]
        panes = (view.left, view.right, view.left_ln, view.right_ln)
        _assert_materialized_panes(view, panes, target_line, len(view.display_rows))
        assert view.left_ln.get(f"{target_line}.0", f"{target_line}.end").lstrip().startswith("[3]")
        assert view.left_ln.get(f"{target_line}.0", f"{target_line}.end").rstrip().endswith("1800")
        assert "blockmarker" in view.left_ln.tag_names(f"{target_line}.0")
    finally:
        app._shutdown_root()


def _test_three_way_cached_base_navigation_no_workbook_reads(root_dir):
    base = os.path.join(root_dir, "large-three-base.xlsx")
    mine = os.path.join(root_dir, "large-three-mine.xlsx")
    theirs = os.path.join(root_dir, "large-three-theirs.xlsx")
    rows_base = []
    rows_mine = []
    rows_theirs = []
    for idx in range(1, 2101):
        tail = [f"payload-{col}-{idx}-" + ("x" * 24) for col in range(3, 11)]
        base_row = [idx, f"base-{idx}"] + tail
        rows_base.append(base_row)
        rows_mine.append(list(base_row))
        changed = idx <= 450 or 700 <= idx <= 1100 or 1800 <= idx <= 1801
        rows_theirs.append([idx, f"theirs-{idx}" if changed else f"base-{idx}"] + tail)
    _make_book(base, rows_base)
    _make_book(mine, rows_mine)
    _make_book(theirs, rows_theirs)

    app = mod.SowMergeApp(mine, theirs, merge_mode=True, base_path=base)
    try:
        app.root.state("normal")
        app.root.geometry("800x600")
        view = _prepare_view(app)
        blocks = view._ensure_full_diff_blocks()
        assert len(blocks) == 3, (
            [(b.start_pair_idx, b.end_pair_idx) for b in blocks],
            _diff_channel_summary(view),
        )
        assert blocks[2].start_pair_idx == 1799
        assert len(view._full_display_rows) == 853, len(view._full_display_rows)
        assert len(view.display_rows) < 800, len(view.display_rows)
        assert set(view._full_display_rows).issubset(view.pair_text_base), (
            len(view._full_display_rows),
            len(view.pair_text_base),
        )
        assert set(view._full_display_rows).issubset(view.pair_base_diff_cols)
        view._sync_main_x_to_frac(0.35)
        _pump(app.root, 0.05)
        saved_x = float(view.left.xview()[0])
        assert saved_x > 0.01, view.left.xview()

        counters = {"cell": 0, "iter_rows": 0, "rescan": 0}
        original_cell = Worksheet.cell
        original_iter_rows = Worksheet.iter_rows
        original_refresh = view.refresh

        def _tracked_cell(self, *args, **kwargs):
            counters["cell"] += 1
            return original_cell(self, *args, **kwargs)

        def _tracked_iter_rows(self, *args, **kwargs):
            counters["iter_rows"] += 1
            return original_iter_rows(self, *args, **kwargs)

        def _tracked_refresh(row_only, rescan):
            counters["rescan"] += int(bool(rescan))
            return original_refresh(row_only, rescan)

        Worksheet.cell = _tracked_cell
        Worksheet.iter_rows = _tracked_iter_rows
        view.refresh = _tracked_refresh
        try:
            # Ordinary selection must render Base and the C-area entirely from cache.
            view._select_line(view.row_to_line[0])
            assert view.cursor_cmp.get("1.0", "1.end") == view.pair_text_base[0]
            assert counters == {"cell": 0, "iter_rows": 0, "rescan": 0}, counters

            counters.update(cell=0, iter_rows=0, rescan=0)
            view._goto_full_diff_block(2)
            assert view.selected_pair_idx == 1799, view.selected_pair_idx
            assert view.display_rows == [1799, 1800], view.display_rows
            assert len(view.display_rows) <= mod._LARGE_DIFF_NAV_PREVIEW_ROWS
            assert view._base_row_for_pair(1799) == 1800
            target_line = view.row_to_line[1799]
            assert view.base.get(f"{target_line}.0", f"{target_line}.end") == view.pair_text_base[1799]
            assert "base-1800" in view.pair_text_base[1799]
            assert view.cursor_cmp.get("1.0", "1.end") == view.pair_text_base[1799], (
                repr(view.cursor_cmp.get("1.0", "1.end")),
                repr(view.pair_text_base[1799]),
                view._last_cursor_cmp_pair_idx,
            )
            panes = (
                view.left,
                view.base,
                view.right,
                view.left_ln,
                view.base_ln,
                view.right_ln,
            )
            _assert_materialized_panes(view, panes, target_line, len(view.display_rows))
            assert view.left_ln.get(f"{target_line}.0", f"{target_line}.end").lstrip().startswith("[3]")
            assert view.left_ln.get(f"{target_line}.0", f"{target_line}.end").rstrip().endswith("1800")
            assert view.base_ln.get(f"{target_line}.0", f"{target_line}.end").rstrip().endswith("1800")
            assert view.right_ln.get(f"{target_line}.0", f"{target_line}.end").rstrip().endswith("1800")
            assert "blockmarker" in view.left_ln.tag_names(f"{target_line}.0")
            rn_w = view._sync_row_header_width_widgets()
            expected_right_header = view._format_main_row_header(1799, "B", rn_w)
            hover_right_header = view._format_main_row_header(
                1799,
                "B",
                rn_w,
                arrow=mod._ROW_ARROW_LEFT,
            )
            view._set_row_header_text(view.right_ln, target_line, hover_right_header)
            view._hover_ln_line_right = target_line
            view._clear_row_header_hover(view.right_ln)
            assert view.right_ln.get(f"{target_line}.0", f"{target_line}.end") == expected_right_header
            assert "blockstart" in view.right_ln.tag_names(f"{target_line}.0")
            for widget in (view.left, view.base, view.right):
                assert abs(float(widget.xview()[0]) - saved_x) < 0.02, (widget, saved_x, widget.xview())
            assert counters == {"cell": 0, "iter_rows": 0, "rescan": 0}, counters
        finally:
            Worksheet.cell = original_cell
            Worksheet.iter_rows = original_iter_rows
            view.refresh = original_refresh
    finally:
        app._shutdown_root()


def main():
    root_dir = make_temp_dir("sow_diff_blocks_")
    _test_two_way_block_presentation(root_dir)
    _test_three_way_base_only_and_processed_status(root_dir)
    _test_large_cross_render_limit_navigation(root_dir)
    _test_three_way_cached_base_navigation_no_workbook_reads(root_dir)
    print("GUI_SELF_TEST_DIFF_BLOCKS_OK")


if __name__ == "__main__":
    main()

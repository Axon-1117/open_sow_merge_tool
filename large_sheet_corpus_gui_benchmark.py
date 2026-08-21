"""Fresh-process GUI corpus timing for current-generation exact Sheet results.

This intentionally exercises ``SowMergeApp`` rather than the direct Oracle
harness.  It measures constructor-start through the selected Sheet's final
generation-matched state, then immediately tears the GUI down so background
work for unrelated Sheets cannot affect the next measurement.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import os
import posixpath
import re
import subprocess
import sys
import tempfile
import threading
import time
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

import psutil
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

import sow_merge_tool as sm
from large_sheet_corpus_benchmark import ROOT, copy_sides, inventory, signature


SCHEMA = "large-sheet-corpus-gui-v2"
EXACT_FINAL = {sm._SHEET_EXACT_SAME, sm._SHEET_EXACT_CHANGED}
NONEXACT_FINAL = {sm._SHEET_EXACT_UNRESOLVED, sm._SHEET_EXACT_FAILED}

_DOCUMENT_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_CELL_REF_RE = re.compile(r"^\$?([A-Z]+)\$?(\d+)$")


def _sha256(path):
    digest = hashlib.sha256()
    with Path(path).open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _atomic_json(path, value):
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    staged = path.with_name(
        f"{path.name}.{os.getpid()}.{threading.get_ident()}.next"
    )
    payload = json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True) + "\n"
    with staged.open("w", encoding="utf-8", newline="\n") as stream:
        stream.write(payload)
        stream.flush()
        os.fsync(stream.fileno())
    delay = 0.01
    try:
        for attempt in range(10):
            try:
                os.replace(staged, path)
                return
            except OSError as exc:
                retryable = isinstance(exc, PermissionError) or getattr(exc, "winerror", None) in {5, 32}
                if not retryable or attempt == 9:
                    raise
                time.sleep(delay)
                delay = min(delay * 2.0, 0.25)
    finally:
        try:
            staged.unlink(missing_ok=True)
        except OSError:
            pass


def _append_jsonl(path, value):
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as stream:
        stream.write(json.dumps(value, ensure_ascii=False, sort_keys=True) + "\n")
        stream.flush()
        os.fsync(stream.fileno())
    _atomic_json(path.with_suffix(path.suffix + ".status.json"), {
        "schema": "large-sheet-corpus-gui-heartbeat-v1",
        "updated_at_epoch": time.time(),
        "last_event": value,
    })


def _kill_process_tree(child):
    processes = []
    try:
        root = psutil.Process(child.pid)
        processes = root.children(recursive=True)
    except (psutil.Error, OSError):
        root = None
    for process in reversed(processes):
        try:
            process.kill()
        except (psutil.Error, OSError):
            pass
    try:
        child.kill()
    except OSError:
        pass
    _gone, alive = psutil.wait_procs(processes + ([root] if root is not None else []), timeout=5.0)
    return "killed-tree-and-waited" if not alive else "killed-tree-pending"


def _xml_local_name(tag):
    return str(tag).rsplit("}", 1)[-1]


def _package_target(source_part, target):
    target = str(target or "").replace("\\", "/")
    if target.startswith("/"):
        return target.lstrip("/")
    return posixpath.normpath(posixpath.join(posixpath.dirname(source_part), target))


def _relationship_part(source_part):
    directory, name = posixpath.split(source_part)
    return posixpath.join(directory, "_rels", f"{name}.rels")


def _sheet_package_part(package, sheet):
    workbook_part = "xl/workbook.xml"
    workbook_root = ET.fromstring(package.read(workbook_part))
    relationship_id = None
    for element in workbook_root.iter():
        if _xml_local_name(element.tag) == "sheet" and element.get("name") == sheet:
            relationship_id = element.get(f"{{{_DOCUMENT_REL_NS}}}id")
            break
    if not relationship_id:
        raise KeyError(f"Sheet not found in workbook package: {sheet}")
    rels_root = ET.fromstring(package.read(_relationship_part(workbook_part)))
    for element in rels_root.iter():
        if _xml_local_name(element.tag) == "Relationship" and element.get("Id") == relationship_id:
            if str(element.get("TargetMode") or "").casefold() == "external":
                break
            part = _package_target(workbook_part, element.get("Target"))
            if part in package.namelist():
                return part
            break
    raise KeyError(f"Sheet package part not found: {sheet}")


def _column_index(letters):
    result = 0
    for char in str(letters or "").upper():
        result = result * 26 + ord(char) - ord("A") + 1
    return result


def _drawing_asset_counts(package, drawing_part):
    counts = {"images": 0, "charts": 0}
    if drawing_part not in package.namelist():
        return counts
    with package.open(drawing_part) as stream:
        for _event, element in ET.iterparse(stream, events=("end",)):
            local = _xml_local_name(element.tag)
            if local == "pic":
                counts["images"] += 1
            elif local == "chart":
                counts["charts"] += 1
            element.clear()
    return counts


def _sheet_content_profile(path, sheet):
    """Stream one OOXML Sheet and retain cell semantics plus asset evidence."""
    with zipfile.ZipFile(path, "r") as package:
        sheet_part = _sheet_package_part(package, sheet)
        asset = {
            "physical_cell_records": 0,
            "styled_cells": 0,
            "styled_blank_cells": 0,
            "comments": 0,
            "hyperlinks": 0,
            "merged_ranges": 0,
            "row_dimension_records": 0,
            "column_dimension_records": 0,
            "tables": 0,
            "auto_filters": 0,
            "data_validations": 0,
            "conditional_formattings": 0,
            "drawings": 0,
            "images": 0,
            "charts": 0,
        }
        populated_cells = 0
        formula_cells = 0
        max_row = 1
        max_column = 1
        dimension_ref = None
        with package.open(sheet_part) as stream:
            for _event, element in ET.iterparse(stream, events=("end",)):
                local = _xml_local_name(element.tag)
                if local == "dimension":
                    dimension_ref = element.get("ref")
                elif local == "c":
                    asset["physical_cell_records"] += 1
                    styled = element.get("s") is not None
                    formula = any(_xml_local_name(child.tag) == "f" for child in element)
                    value = any(_xml_local_name(child.tag) in {"v", "is"} for child in element)
                    populated = formula or value
                    if formula:
                        formula_cells += 1
                    if populated:
                        populated_cells += 1
                    if styled:
                        asset["styled_cells"] += 1
                        if not populated:
                            asset["styled_blank_cells"] += 1
                    match = _CELL_REF_RE.match(str(element.get("r") or ""))
                    if match:
                        max_column = max(max_column, _column_index(match.group(1)))
                        max_row = max(max_row, int(match.group(2)))
                elif local == "row":
                    try:
                        max_row = max(max_row, int(element.get("r") or 1))
                    except ValueError:
                        pass
                    if set(element.attrib) - {"r", "spans"}:
                        asset["row_dimension_records"] += 1
                elif local == "col":
                    asset["column_dimension_records"] += 1
                elif local == "mergeCell":
                    asset["merged_ranges"] += 1
                elif local == "hyperlink":
                    asset["hyperlinks"] += 1
                elif local == "tablePart":
                    asset["tables"] += 1
                elif local == "autoFilter":
                    asset["auto_filters"] += 1
                elif local == "dataValidation":
                    asset["data_validations"] += 1
                elif local == "conditionalFormatting":
                    asset["conditional_formattings"] += 1
                element.clear()

        if dimension_ref:
            try:
                _min_col, _min_row, dim_max_col, dim_max_row = range_boundaries(dimension_ref)
                max_row = max(max_row, int(dim_max_row or 1))
                max_column = max(max_column, int(dim_max_col or 1))
            except (TypeError, ValueError):
                pass

        relationship_counts = {}
        rels_part = _relationship_part(sheet_part)
        if rels_part in package.namelist():
            rels_root = ET.fromstring(package.read(rels_part))
            for element in rels_root.iter():
                if _xml_local_name(element.tag) != "Relationship":
                    continue
                relation_type = str(element.get("Type") or "").rsplit("/", 1)[-1]
                relationship_counts[relation_type] = relationship_counts.get(relation_type, 0) + 1
                if relation_type == "comments":
                    asset["comments"] += 1
                elif relation_type == "drawing" and str(element.get("TargetMode") or "").casefold() != "external":
                    asset["drawings"] += 1
                    drawing = _drawing_asset_counts(
                        package,
                        _package_target(sheet_part, element.get("Target")),
                    )
                    asset["images"] += drawing["images"]
                    asset["charts"] += drawing["charts"]
        asset["relationship_type_counts"] = relationship_counts
        return {
            "rows": int(max_row),
            "columns": int(max_column),
            "populated_cells": int(populated_cells),
            "formula_cells": int(formula_cells),
            "asset_profile": asset,
        }


def _side_sheet_profiles(args, package_sha256):
    paths = {"A": args.mine, "B": args.theirs}
    if args.mode == "3way":
        paths["BASE"] = args.base
    required = tuple(paths)
    identical = bool(required) and all(package_sha256.get(side) for side in required) and len(
        {package_sha256[side] for side in required}
    ) == 1
    if identical:
        mine = _sheet_content_profile(paths["A"], args.sheet)
        return {
            side: dict(mine, asset_profile=dict(mine["asset_profile"]))
            for side in required
        }
    return {side: _sheet_content_profile(path, args.sheet) for side, path in paths.items()}


def _pump_until_terminal(app, sheet, generation, timeout):
    deadline = time.monotonic() + timeout
    last = app._sheet_exact_entry(sheet)
    while time.monotonic() < deadline:
        app.root.update_idletasks()
        app.root.update()
        last = app._sheet_exact_entry(sheet)
        state = str(last.get("state") or "")
        if int(last.get("generation", -1)) == generation and state in EXACT_FINAL | NONEXACT_FINAL:
            return last
        time.sleep(0.005)
    raise TimeoutError(f"GUI exact state did not finish: {sheet} {last}")


def _physical_target_validation(view, *, three_way: bool):
    """Prove the published cache can address real physical operation targets."""
    cache = getattr(view, "column_comparison_cache", None)
    model = getattr(cache, "model", None)
    row_pairs = tuple(getattr(view, "row_pairs", ()) or ())
    pair_index = next(
        (index for index, pair in enumerate(row_pairs) if pair[0] is not None and pair[1] is not None),
        None,
    )
    if model is None or pair_index is None:
        return {"ok": False, "applicable": True, "reason": "missing physical cache or paired row"}
    pair = row_pairs[pair_index]
    logical = next(
        (slot.logical_idx for slot in model.slots if slot.mine_col and slot.theirs_col),
        None,
    )
    if logical is None:
        return {"ok": False, "applicable": True, "reason": "missing common physical column"}
    result = {
        "ok": bool(
            model.physical_for_logical("A", logical)
            and model.physical_for_logical("B", logical)
            and pair[0] > 0 and pair[1] > 0
        ),
        "pair_index": int(pair_index),
        "mine_row": int(pair[0]),
        "theirs_row": int(pair[1]),
        "logical_column": int(logical),
        "mine_column": model.physical_for_logical("A", logical),
        "theirs_column": model.physical_for_logical("B", logical),
        "applicable": True,
        "reason": "physical_cell_targets_ready",
    }
    if three_way:
        base_row = getattr(view, "_base_row_for_pair", lambda *_args: None)(pair_index, pair)
        base_column = model.physical_for_logical("BASE", logical)
        result.update(base_row=base_row, base_column=base_column)
        result["ok"] = bool(result["ok"] and base_row and base_column)
    return result


def _cell_target_applicability(
    operation_targets,
    *,
    final_state,
    package_sha256,
    side_sheet_profiles,
    three_way,
):
    """Allow target N/A only for proven identical inputs with no cell content."""
    required = ("A", "B", "BASE") if three_way else ("A", "B")
    hashes = [str(package_sha256.get(side) or "") for side in required]
    identical_inputs = bool(hashes) and all(hashes) and len(set(hashes)) == 1
    profiles_complete = all(side in side_sheet_profiles for side in required)
    no_populated_cells = bool(profiles_complete) and all(
        int(side_sheet_profiles[side].get("populated_cells", -1)) == 0
        and int(side_sheet_profiles[side].get("formula_cells", -1)) == 0
        for side in required
    )
    blocking_asset_fields = (
        "physical_cell_records",
        "styled_cells",
        "styled_blank_cells",
        "row_dimension_records",
        "column_dimension_records",
        "merged_ranges",
        "hyperlinks",
        "tables",
        "auto_filters",
        "data_validations",
        "conditional_formattings",
        "comments",
        "charts",
    )
    side_asset_safety = {}
    for side in required:
        asset = dict(side_sheet_profiles.get(side, {}).get("asset_profile") or {})
        blocking_counts = {
            name: int(asset.get(name, 0) or 0)
            for name in blocking_asset_fields
            if int(asset.get(name, 0) or 0) != 0
        }
        disallowed_relationships = sorted(
            str(name)
            for name, count in dict(asset.get("relationship_type_counts") or {}).items()
            if int(count or 0) != 0 and str(name) not in {"drawing", "image"}
        )
        side_asset_safety[side] = {
            "ok": not blocking_counts and not disallowed_relationships,
            "blocking_asset_counts": blocking_counts,
            "disallowed_relationship_types": disallowed_relationships,
        }
    no_cell_or_structural_assets = bool(profiles_complete) and all(
        side_asset_safety[side]["ok"] for side in required
    )
    if operation_targets.get("ok"):
        return {
            "ok": True,
            "applicable": True,
            "reason": "physical_cell_targets_ready",
            "identical_inputs": identical_inputs,
            "side_populated_cells": {
                side: side_sheet_profiles.get(side, {}).get("populated_cells")
                for side in required
            },
            "side_formula_cells": {
                side: side_sheet_profiles.get(side, {}).get("formula_cells")
                for side in required
            },
            "side_asset_safety": side_asset_safety,
        }
    if (
        final_state == sm._SHEET_EXACT_SAME
        and identical_inputs
        and no_populated_cells
        and no_cell_or_structural_assets
    ):
        return {
            "ok": True,
            "applicable": False,
            "reason": "not_applicable_no_cell_or_structural_assets_identical_inputs",
            "identical_inputs": True,
            "side_populated_cells": {side: 0 for side in required},
            "side_formula_cells": {side: 0 for side in required},
            "side_asset_safety": side_asset_safety,
        }
    return {
        "ok": False,
        "applicable": True,
        "reason": str(operation_targets.get("reason") or "physical cell targets required"),
        "identical_inputs": identical_inputs,
        "profiles_complete": profiles_complete,
        "no_populated_cells": no_populated_cells,
        "no_cell_or_structural_assets": no_cell_or_structural_assets,
        "side_populated_cells": {
            side: side_sheet_profiles.get(side, {}).get("populated_cells")
            for side in required
        },
        "side_formula_cells": {
            side: side_sheet_profiles.get(side, {}).get("formula_cells")
            for side in required
        },
        "side_asset_safety": side_asset_safety,
    }


def gui_sheet_worker(args):
    started = time.perf_counter()
    process = psutil.Process()
    before = process.memory_info().rss
    sample_stop = threading.Event()
    rss_samples = [before]

    def _sample_comparison_rss():
        while not sample_stop.wait(0.02):
            try:
                rss_samples.append(process.memory_info().rss)
            except (psutil.Error, OSError):
                return

    sampler = threading.Thread(target=_sample_comparison_rss, name="sow-gui-rss", daemon=True)
    sampler.start()
    result = {
        "schema": SCHEMA, "source_path": args.source, "mode": args.mode, "sheet": args.sheet,
        "gate_seconds": args.gate, "status": "FAILED", "final_state": "FAILED", "error": None,
    }
    app = None
    original_askyesno = sm.messagebox.askyesno
    try:
        # Test state only: decline the optional formula-cache recalculation dialog.
        sm.messagebox.askyesno = lambda *unused_args, **unused_kwargs: False
        constructor_started = time.perf_counter()
        app = sm.SowMergeApp(
            args.mine,
            args.theirs,
            merge_mode=args.mode == "3way",
            merged_path=str(Path(args.scratch) / "merged.xlsx") if args.mode == "3way" else None,
            base_path=args.base,
            initial_sheet=args.sheet,
        )
        constructor_ms = (time.perf_counter() - constructor_started) * 1000.0
        try:
            app.root.withdraw()
        except Exception:
            pass
        generation = int(app._sheet_compute_generation.get(args.sheet, 0))
        selected_initial = app._sheet_exact_entry(args.sheet)
        state_counts = {}
        for name in app.compare_sheets:
            state = str(app._sheet_exact_entry(name).get("state") or "")
            state_counts[state] = state_counts.get(state, 0) + 1
        terminal = _pump_until_terminal(app, args.sheet, generation, args.gate)
        total_ms = (time.perf_counter() - constructor_started) * 1000.0
        view = app.sheet_views.get(args.sheet)
        metrics = dict(getattr(view, "_snapshot_metrics_ms", {}) or {}) if view is not None else {}
        final_state = str(terminal.get("state") or "FAILED")
        package_sha256 = {
            "A": str(args.mine_sha256 or ""),
            "B": str(args.theirs_sha256 or ""),
        }
        if args.mode == "3way":
            package_sha256["BASE"] = str(args.base_sha256 or "")
        side_sheet_profiles = _side_sheet_profiles(args, package_sha256)
        dimensions = {
            key: side_sheet_profiles["A"][key]
            for key in ("rows", "columns", "populated_cells", "formula_cells")
        }
        operation_targets = _physical_target_validation(view, three_way=args.mode == "3way")
        target_applicability = _cell_target_applicability(
            operation_targets,
            final_state=final_state,
            package_sha256=package_sha256,
            side_sheet_profiles=side_sheet_profiles,
            three_way=args.mode == "3way",
        )
        no_calculation_surface = bool(
            view is not None
            and not bool(getattr(view, "_pending_exact_render", False))
            and not bool(getattr(view, "_only_diff_async_building", False))
        )
        comparison_detail_ready = bool(
            final_state in EXACT_FINAL
            and view is not None
            and bool(getattr(view, "_prepared_complete", False))
            and bool(getattr(view, "_data_ready", False))
            and bool(target_applicability.get("ok"))
            and no_calculation_surface
        )
        result.update({
            "constructor_file_open_ms": round(constructor_ms, 3),
            "constructor_to_final_exact_ms": round(total_ms, 3),
            "post_constructor_ready_ms": round(max(0.0, total_ms - constructor_ms), 3),
            "current_generation": generation,
            "initial_selected_state": selected_initial,
            "initial_state_counts": state_counts,
            "final_exact_entry": terminal,
            "cache_source": getattr(view, "_cache_source", None) if view is not None else None,
            "snapshot_metrics_ms": metrics,
            "prepared_complete": bool(getattr(view, "_prepared_complete", False)),
            "lifecycle_state": getattr(view, "_derive_lifecycle_state", lambda: "")() if view is not None else "",
            "physical_operation_targets": operation_targets,
            "cell_target_applicability": target_applicability,
            "comparison_detail_ready": comparison_detail_ready,
            "no_calculation_surface": no_calculation_surface,
            # Intentionally separate from comparison timing/RSS: this is
            # loaded only by a real copy/overwrite/column/save request, whose
            # first rejected click shows the readiness modal and starts one
            # owner loader.
            "mutation_backend_state": (
                "READY" if app._edit_workbooks_ready() else "DEFERRED"
            ),
            "input_package_sha256": package_sha256,
            "side_sheet_profiles": side_sheet_profiles,
            "dimensions": dimensions,
            "final_state": final_state,
            "status": "PASS" if (
                final_state == sm._SHEET_EXACT_SAME
                and comparison_detail_ready
                and total_ms <= args.gate * 1000
            ) else "TIMEOUT" if total_ms > args.gate * 1000 else "ORACLE_FAILURE",
        })
    except Exception as exc:
        result["error"] = f"{type(exc).__name__}: {exc}"
    finally:
        try:
            result["view_only_rss_pre_shutdown_bytes"] = process.memory_info().rss
            rss_samples.append(result["view_only_rss_pre_shutdown_bytes"])
        except (psutil.Error, OSError):
            result["view_only_rss_pre_shutdown_bytes"] = None
        sample_stop.set()
        sampler.join(timeout=1.0)
        sm.messagebox.askyesno = original_askyesno
        if app is not None:
            try:
                app._shutdown_root()
            except Exception as exc:
                result["shutdown_error"] = f"{type(exc).__name__}: {exc}"
        after = process.memory_info().rss
        result.update({
            "rss_before_bytes": before, "rss_after_bytes": after, "rss_delta_bytes": after - before,
            "rss_peak_bytes": max(rss_samples) if rss_samples else after,
            "view_only_rss_peak_bytes": max(rss_samples) if rss_samples else after,
            "rss_sample_count": len(rss_samples),
            "worker_elapsed_ms": round((time.perf_counter() - started) * 1000.0, 3),
        })
    return result


def operation_sheet_worker(args):
    """Separate first-operation phase; never part of comparison SLA timing."""
    started = time.perf_counter()
    result = {"schema": SCHEMA, "phase": "first-operation", "mode": args.mode, "sheet": args.sheet,
              "status": "FAILED", "error": None}
    app = None
    original_warning = sm.messagebox.showwarning
    original_yesno = sm.messagebox.askyesno
    original_info = sm.messagebox.showinfo
    original_error = sm.messagebox.showerror
    try:
        modals = []
        sm.messagebox.showwarning = lambda *a, **k: modals.append((a, k))
        sm.messagebox.askyesno = lambda *a, **k: False
        sm.messagebox.showinfo = lambda *a, **k: None
        sm.messagebox.showerror = lambda *a, **k: None
        app = sm.SowMergeApp(args.mine, args.theirs, merge_mode=args.mode == "3way",
                             merged_path=str(Path(args.scratch) / "merged.xlsx") if args.mode == "3way" else None,
                             base_path=args.base, initial_sheet=args.sheet)
        generation = int(app._sheet_compute_generation.get(args.sheet, 0))
        terminal = _pump_until_terminal(app, args.sheet, generation, args.gate)
        view = app.sheet_views[args.sheet]
        target = _physical_target_validation(view, three_way=args.mode == "3way")
        logical = int(target["logical_column"]) + 1
        pair_index = int(target["pair_index"])
        before = (dict(app.manual_a_cell_ops), list(app.undo_stack))
        if view._derive_lifecycle_state() != "EDIT_DEFERRED":
            raise AssertionError(f"expected EDIT_DEFERRED, got {view._derive_lifecycle_state()}")
        load_started = time.perf_counter()
        first = view._copy_single_cell_by_pair(pair_index, "B2A", logical)
        assert first is False and modals and app._edit_loading_started
        assert (dict(app.manual_a_cell_ops), list(app.undo_stack)) == before
        modal_title = str(modals[0][0][0]) if modals[0][0] else ""
        modal_text = "\n".join(str(part) for part in modals[0][0][1:])
        if (
            modal_title != "正在加载可编辑工作簿"
            or "此前尚未加载" not in modal_text
            or "不会执行或自动重试" not in modal_text
            or "精确比较已完成" not in modal_text
        ):
            raise AssertionError(f"inconsistent deferred modal: {modal_title} {modal_text}")
        deadline = time.monotonic() + max(args.gate, 30.0)
        while time.monotonic() < deadline and not app._edit_workbooks_ready():
            app.root.update_idletasks(); app.root.update(); time.sleep(0.005)
        if not app._edit_workbooks_ready():
            raise TimeoutError("first-operation editable backend did not load")
        load_ms = (time.perf_counter() - load_started) * 1000.0
        action_started = time.perf_counter()
        retry = view._copy_single_cell_by_pair(pair_index, "B2A", logical)
        if retry is not None or not app.undo_stack:
            raise AssertionError("retry operation was not accepted/undoable")
        action_ms = (time.perf_counter() - action_started) * 1000.0
        view._undo_last_action()
        app._confirm_overwrite = lambda *_a, **_k: True
        app.save_a_inplace()
        reopened = load_workbook(args.mine, data_only=False)
        reopened.close()
        result.update(status="PASS", final_exact_entry=terminal, physical_operation_targets=target,
                      first_click_modal_count=len(modals), first_click_mutated=False,
                      first_click_modal_title=modal_title, first_click_modal_text=modal_text,
                      first_operation_load_ms=round(load_ms, 3), retry_action_ms=round(action_ms, 3),
                      undo_ok=not app.undo_stack, save_reopen_ok=True)
    except Exception as exc:
        result["error"] = f"{type(exc).__name__}: {exc}"
    finally:
        sm.messagebox.showwarning = original_warning
        sm.messagebox.askyesno = original_yesno
        sm.messagebox.showinfo = original_info
        sm.messagebox.showerror = original_error
        if app is not None:
            app._shutdown_root()
        result["worker_elapsed_ms"] = round((time.perf_counter() - started) * 1000.0, 3)
    return result


def _sheet_child(args, mine, theirs, base, sheet, scratch, package_sha256):
    command = [sys.executable, str(Path(__file__).resolve()), "--sheet-worker", "--source", args.source,
               "--mode", args.mode, "--sheet", sheet, "--mine", str(mine), "--theirs", str(theirs),
               "--mine-sha256", package_sha256["A"], "--theirs-sha256", package_sha256["B"],
               "--scratch", str(scratch), "--gate", str(args.gate)]
    if base:
        command.extend(["--base", str(base), "--base-sha256", package_sha256["BASE"]])
    child = subprocess.Popen(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
    try:
        stdout, stderr = child.communicate(timeout=args.gate + args.process_grace)
    except subprocess.TimeoutExpired as exc:
        cleanup = _kill_process_tree(child)
        stdout, stderr = child.communicate()
        return {"schema": SCHEMA, "source_path": args.source, "mode": args.mode, "sheet": sheet,
                "status": "PROCESS_TIMEOUT", "final_state": "FAILED", "child_pid": child.pid,
                "child_cleanup": cleanup, "error": str(exc)}
    lines = [line for line in stdout.splitlines() if line.strip()]
    if child.returncode or not lines:
        return {"schema": SCHEMA, "source_path": args.source, "mode": args.mode, "sheet": sheet,
                "status": "WORKER_FAILURE", "final_state": "FAILED", "child_pid": child.pid,
                "child_cleanup": "waited", "error": (stderr or stdout or f"exit {child.returncode}")[-4000:]}
    row = json.loads(lines[-1]); row.update(child_pid=child.pid, child_cleanup="waited")
    return row


def file_worker(args):
    started = time.perf_counter()
    result = {"schema": SCHEMA, "source_path": args.source, "relative_path": args.relative_path,
              "mode": args.mode, "status": "FAILED", "sheets": [], "error": None}
    try:
        Path(args.jsonl).parent.mkdir(parents=True, exist_ok=True)
        workbook = load_workbook(args.source, read_only=True, data_only=False, keep_vba=Path(args.source).suffix.casefold() == ".xlsm")
        try:
            sheets = list(workbook.sheetnames)
        finally:
            workbook.close()
        if args.only_sheet:
            sheets = [sheet for sheet in sheets if sheet == args.only_sheet]
            if not sheets:
                raise KeyError(f"requested sheet not found: {args.only_sheet}")
        with tempfile.TemporaryDirectory(prefix="sow_gui_corpus_file_") as temporary:
            mine, theirs, base = copy_sides(args.source, temporary, args.mode)
            result["disposable_sides"] = True
            package_sha256 = {"A": _sha256(mine), "B": _sha256(theirs)}
            if base:
                package_sha256["BASE"] = _sha256(base)
            result["input_package_sha256"] = package_sha256
            for sheet in sheets:
                row = _sheet_child(
                    args,
                    mine,
                    theirs,
                    base,
                    sheet,
                    temporary,
                    package_sha256,
                )
                result["sheets"].append(row)
                _append_jsonl(args.jsonl, {"event": "gui-sheet-complete", "relative_path": args.relative_path, "mode": args.mode, "sheet_result": row})
        result["status"] = "PASS" if all(row.get("status") == "PASS" for row in result["sheets"]) else "FAIL"
    except Exception as exc:
        result["error"] = f"{type(exc).__name__}: {exc}"
    result["whole_file_elapsed_ms"] = round((time.perf_counter() - started) * 1000.0, 3)
    return result


def _file_child(source, relative, mode, args):
    command = [sys.executable, str(Path(__file__).resolve()), "--file-worker", "--source", str(source),
               "--relative-path", relative, "--mode", mode, "--gate", str(args.gate),
               "--process-grace", str(args.process_grace), "--file-process-timeout", str(args.file_process_timeout),
               "--jsonl", str(args.jsonl)]
    if args.only_sheet:
        command.extend(["--only-sheet", args.only_sheet])
    child = subprocess.Popen(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
    try:
        stdout, stderr = child.communicate(timeout=args.file_process_timeout)
    except subprocess.TimeoutExpired as exc:
        cleanup = _kill_process_tree(child)
        stdout, stderr = child.communicate()
        return {"schema": SCHEMA, "source_path": str(source), "relative_path": relative, "mode": mode,
                "status": "FILE_PROCESS_TIMEOUT", "sheets": [], "child_pid": child.pid,
                "child_cleanup": cleanup, "error": str(exc)}
    lines = [line for line in stdout.splitlines() if line.strip()]
    if child.returncode or not lines:
        return {"schema": SCHEMA, "source_path": str(source), "relative_path": relative, "mode": mode,
                "status": "FILE_WORKER_FAILURE", "sheets": [], "child_pid": child.pid,
                "child_cleanup": "waited", "error": (stderr or stdout or f"exit {child.returncode}")[-4000:]}
    return json.loads(lines[-1])


def _rankings(runs, mode):
    rows = [dict(row, relative_path=run["relative_path"]) for run in runs if run.get("mode") == mode for row in run.get("sheets", [])]
    return sorted(rows, key=lambda row: float(row.get("constructor_to_final_exact_ms") or float("inf")), reverse=True)


def controller(args):
    root, out = Path(args.source_root), Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    Path(args.jsonl).parent.mkdir(parents=True, exist_ok=True)
    if args.resume and out.exists():
        evidence = json.loads(out.read_text(encoding="utf-8"))
    else:
        evidence = {"schema": SCHEMA, "created_at_epoch": time.time(), "gate_seconds": args.gate,
                    "inventory_before": inventory(root), "runs": [], "source_signature_before": {},
                    "source_signature_after": {}, "provenance": {"sow_merge_tool_sha256": _sha256(Path(sm.__file__)),
                    "gui_harness_sha256": _sha256(Path(__file__).resolve())}}
    done = {(row.get("relative_path"), row.get("mode")) for row in evidence["runs"]}
    supported = [row for row in evidence["inventory_before"]["files"] if row["classification"] == "supported"]
    if args.only_relative_path:
        supported = [row for row in supported if row["relative_path"] == args.only_relative_path]
    if args.max_files is not None:
        supported = supported[:args.max_files]
    for item in supported:
        source, relative = Path(item["source_path"]), item["relative_path"]
        evidence["source_signature_before"].setdefault(relative, signature(source))
        for mode in ("2way", "3way"):
            if (relative, mode) in done:
                continue
            run = _file_child(source, relative, mode, args)
            evidence["runs"].append(run)
            _append_jsonl(args.jsonl, {"event": "gui-file-mode-checkpoint", "relative_path": relative, "mode": mode, "status": run.get("status")})
            _atomic_json(out, evidence)
    for item in supported:
        evidence["source_signature_after"][item["relative_path"]] = signature(Path(item["source_path"]))
    evidence["source_unchanged"] = evidence["source_signature_before"] == evidence["source_signature_after"]
    evidence["slowest_sheets"] = {mode: _rankings(evidence["runs"], mode) for mode in ("2way", "3way")}
    _atomic_json(out, evidence)
    failures = [row for run in evidence["runs"] for row in run.get("sheets", []) if row.get("status") != "PASS"]
    print(json.dumps({"out": str(out), "runs": len(evidence["runs"]), "sheet_failures": len(failures), "source_unchanged": evidence["source_unchanged"]}, ensure_ascii=False))
    return 0 if not failures and evidence["source_unchanged"] else 1


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-root", default=str(ROOT)); parser.add_argument("--out", default="large_sheet_corpus_gui.json")
    parser.add_argument("--jsonl", default="large_sheet_corpus_gui.jsonl"); parser.add_argument("--gate", type=float, default=15.0)
    parser.add_argument("--process-grace", type=float, default=30.0); parser.add_argument("--file-process-timeout", type=float, default=900.0)
    parser.add_argument("--max-files", type=int); parser.add_argument("--only-relative-path"); parser.add_argument("--only-sheet"); parser.add_argument("--resume", action="store_true")
    parser.add_argument("--file-worker", action="store_true"); parser.add_argument("--sheet-worker", action="store_true"); parser.add_argument("--operation-worker", action="store_true")
    parser.add_argument("--source"); parser.add_argument("--relative-path", default=""); parser.add_argument("--mode", choices=("2way", "3way"))
    parser.add_argument("--sheet"); parser.add_argument("--mine"); parser.add_argument("--theirs"); parser.add_argument("--base"); parser.add_argument("--scratch")
    parser.add_argument("--mine-sha256"); parser.add_argument("--theirs-sha256"); parser.add_argument("--base-sha256")
    args = parser.parse_args()
    if args.sheet_worker:
        print(json.dumps(gui_sheet_worker(args), ensure_ascii=False, sort_keys=True)); return
    if args.operation_worker:
        print(json.dumps(operation_sheet_worker(args), ensure_ascii=False, sort_keys=True)); return
    if args.file_worker:
        print(json.dumps(file_worker(args), ensure_ascii=False, sort_keys=True)); return
    raise SystemExit(controller(args))


if __name__ == "__main__":
    main()

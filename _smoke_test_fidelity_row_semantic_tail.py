"""Focused gate regression for row mutations after style-only worksheet tails."""

from __future__ import annotations

import argparse
import hashlib
import os
import shutil
import tempfile
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

import _large_sheet_excel_fidelity_gate as gate


_CASE = "fidelity-row-semantic-tail"
_SHEET = "S1"
_STYLE_TAIL_ROWS = tuple(range(8, 13))


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for block in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _write_typed_book(path: Path, *, style_tail: bool) -> None:
    workbook = Workbook()
    try:
        sheet = workbook.active
        sheet.title = _SHEET
        sheet.cell(1, 1).value = "id@id"
        sheet.cell(1, 2).value = "value@pm"
        sheet.cell(2, 1).value = "string"
        sheet.cell(2, 2).value = "string"
        sheet.cell(3, 1).value = "record-1"
        sheet.cell(3, 2).value = "seed"
        if style_tail:
            for row in _STYLE_TAIL_ROWS:
                sheet.cell(row, 1).fill = PatternFill(
                    fill_type="solid",
                    fgColor="00D9EAD3",
                )
        workbook.save(path)
    finally:
        workbook.close()


def _assert_target_and_style_tail(
    path: Path,
    target: gate.MutationTarget,
    *,
    marker: str,
    style_tail: bool,
) -> None:
    workbook = load_workbook(path, data_only=False)
    try:
        sheet = workbook[_SHEET]
        expected_row = 4
        assert (target.kind, target.row, target.col, target.marker) == (
            "row", expected_row, 2, marker,
        )
        assert gate._worksheet_semantic_bounds(sheet) == (expected_row, 2)
        assert sheet.cell(expected_row, 1).value == "__SOW_FIDELITY_ROW_KEY__"
        assert sheet.cell(expected_row, 2).value == marker
        if style_tail:
            assert int(sheet.max_row or 0) >= max(_STYLE_TAIL_ROWS)
            for row in _STYLE_TAIL_ROWS:
                assert sheet.cell(row, 1).value is None
                assert sheet.cell(row, 2).value is None
        else:
            assert int(sheet.max_row or 0) == expected_row
    finally:
        workbook.close()


def _assert_direct_snapshot_exact(
    mine: Path,
    theirs: Path,
    target_mine: gate.MutationTarget,
    target_theirs: gate.MutationTarget,
    *,
    label: str,
) -> None:
    legacy = gate._assert_direct_pair_parity(mine, theirs, _SHEET, label)
    snapshot, unresolved = gate._snapshot_manifest(mine, theirs, None, _SHEET)
    assert unresolved is False
    parity = gate._compare_manifests_by_stable_identity(legacy, snapshot)
    assert parity["exact"], parity
    for manifest in (legacy, snapshot):
        gate._assert_target(manifest, target_mine, side="mine")
        gate._assert_target(manifest, target_theirs, side="theirs")
        records = tuple(manifest["records"])
        assert len(records) == 1
        record = records[0]
        assert (record["mine_row"], record["theirs_row"]) == (4, 4)
        assert record["row_structure"] is False
        assert record["diff_cols"] == [2]
        assert manifest["only_diff_rows"] == [record["pair"]]


def _run_variant(root: Path, *, style_tail: bool) -> None:
    label = "style-tail" if style_tail else "no-style-tail"
    source = root / f"{label}-source.xlsx"
    mine = root / f"{label}-mine.xlsx"
    theirs = root / f"{label}-theirs.xlsx"
    _write_typed_book(source, style_tail=style_tail)
    source_sha = _sha256(source)
    if style_tail:
        workbook = load_workbook(source, data_only=False)
        try:
            sheet = workbook[_SHEET]
            assert int(sheet.max_row or 0) >= max(_STYLE_TAIL_ROWS)
            assert gate._worksheet_semantic_bounds(sheet) == (3, 2)
        finally:
            workbook.close()
    shutil.copy2(source, mine)
    shutil.copy2(source, theirs)
    target_mine = gate._mutate_row(mine, _SHEET, f"{label}-mine")
    target_theirs = gate._mutate_row(theirs, _SHEET, f"{label}-theirs")
    _assert_target_and_style_tail(
        mine,
        target_mine,
        marker=f"{label}-mine",
        style_tail=style_tail,
    )
    _assert_target_and_style_tail(
        theirs,
        target_theirs,
        marker=f"{label}-theirs",
        style_tail=style_tail,
    )
    _assert_direct_snapshot_exact(
        mine,
        theirs,
        target_mine,
        target_theirs,
        label=label,
    )
    assert _sha256(source) == source_sha


def _run_case() -> None:
    temporary = tempfile.TemporaryDirectory(prefix="sow_fidelity_row_semantic_tail_")
    root = Path(temporary.name)
    primary: BaseException | None = None
    cleanup_errors: list[str] = []
    try:
        _run_variant(root, style_tail=True)
        _run_variant(root, style_tail=False)
    except BaseException as exc:
        primary = exc
        raise
    finally:
        try:
            temporary.cleanup()
        except BaseException as exc:
            cleanup_errors.append(f"temporary cleanup: {type(exc).__name__}: {exc}")
        if os.path.lexists(root):
            cleanup_errors.append(f"owned root remains: {root}")
        if cleanup_errors:
            detail = "; ".join(cleanup_errors)
            if primary is not None:
                primary.add_note(detail)
            else:
                raise AssertionError(detail)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--case")
    parser.add_argument("--list-cases", action="store_true")
    args = parser.parse_args()
    if args.list_cases:
        print(_CASE)
        return
    selected = args.case or _CASE
    if selected != _CASE:
        raise SystemExit(f"unknown case: {selected}")
    _run_case()
    print(f"SMOKE_FIDELITY_ROW_SEMANTIC_TAIL_OK {selected}")


if __name__ == "__main__":
    main()

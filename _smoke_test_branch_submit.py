"""Smoke tests for the multi-branch SVN submit workflow."""

from __future__ import annotations

import json
import os
import shutil
import tempfile
from types import SimpleNamespace

from openpyxl import Workbook, load_workbook

import branch_submit as bs


def _book(path: str, value: int) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["id", "value"])
    ws.append(["a", value])
    wb.save(path)
    wb.close()


class FakeCore:
    APP_NAME = "sow_merge_tool"

    @staticmethod
    def _find_svn_wc_root_for_path(path):
        current = os.path.abspath(path)
        while current != os.path.dirname(current):
            if os.path.isfile(os.path.join(current, ".svn", "wc.db")):
                return current
            current = os.path.dirname(current)
        return None

    @staticmethod
    def _try_export_svn_base_from_working_copy(path):
        return path + ".pristine.xlsx"

    @staticmethod
    def _wc_node_metadata(root, relative):
        return 100, "tester", "fake"

    @staticmethod
    def _detect_svn_conflict_files(path):
        return None

    @staticmethod
    def _has_svn_conflict_artifacts(path):
        return False

    @staticmethod
    def _cross_branch_source_delta_premerge(before, target, after):
        candidate = target + ".candidate.xlsx"
        shutil.copy2(target, candidate)
        before_wb = load_workbook(before, data_only=False)
        after_wb = load_workbook(after, data_only=False)
        target_wb = load_workbook(candidate, data_only=False)
        changed = 0
        for sheet in before_wb.sheetnames:
            b = before_wb[sheet]
            a = after_wb[sheet]
            t = target_wb[sheet]
            for row in range(1, max(b.max_row, a.max_row) + 1):
                for col in range(1, max(b.max_column, a.max_column) + 1):
                    if b.cell(row, col).value != a.cell(row, col).value:
                        if t.cell(row, col).value not in (b.cell(row, col).value, a.cell(row, col).value):
                            before_wb.close(); after_wb.close(); target_wb.close()
                            return [(sheet, row, col, t.cell(row, col).value, a.cell(row, col).value)], candidate, {}, {"incoming_count": 1, "applied_count": 0, "already_present_count": 0, "target_retained_count": 0, "unresolved_count": 1}, None
                        t.cell(row, col).value = a.cell(row, col).value
                        changed += 1
        target_wb.save(candidate)
        before_wb.close(); after_wb.close(); target_wb.close()
        summary = {"incoming_count": changed, "applied_count": changed, "already_present_count": 0, "target_retained_count": 0, "unresolved_count": 0}
        return [], candidate, {}, summary, None

    @staticmethod
    def _find_tortoise_proc_exe():
        return "TortoiseProc.exe"


def test_discover_and_validate() -> None:
    root = tempfile.mkdtemp(prefix="branch-submit-discover-")
    try:
        os.makedirs(os.path.join(root, ".svn"))
        for branch in ("develop", "release", "sandbox", "master", "notes"):
            os.makedirs(os.path.join(root, branch))
        assert bs.discover_branches(root) == ["develop", "release", "sandbox"]
        assert bs._validate_relative_file("configs/A.xlsx") == "configs/A.xlsx"
        try:
            bs._validate_relative_file("../escape.xlsx")
        except ValueError:
            pass
        else:
            raise AssertionError("path traversal accepted")
    finally:
        shutil.rmtree(root, ignore_errors=True)


def test_preflight_persists_delta_and_blocks_conflict() -> None:
    root = tempfile.mkdtemp(prefix="branch-submit-preflight-")
    old_appdata = os.environ.get("LOCALAPPDATA")
    os.environ["LOCALAPPDATA"] = os.path.join(root, "appdata")
    try:
        os.makedirs(os.path.join(root, ".svn"))
        for branch in ("develop", "release", "sandbox"):
            os.makedirs(os.path.join(root, branch, "config"))
        source = os.path.join(root, "develop", "config", "A.xlsx")
        release = os.path.join(root, "release", "config", "A.xlsx")
        sandbox = os.path.join(root, "sandbox", "config", "A.xlsx")
        _book(source, 9)
        _book(release, 1)
        _book(sandbox, 7)
        shutil.copy2(source, source + ".pristine.xlsx")
        # The fake pristine must represent source-before.
        _book(source + ".pristine.xlsx", 1)
        engine = bs.BranchSubmitEngine(root)
        engine.core = FakeCore
        batch = engine.preflight("develop", ["release", "sandbox"], [source], "调整配置")
        assert batch.source_status == "pending", batch.error
        assert batch.target_status == {"release": "ready", "sandbox": "blocked"}
        assert batch.files[0].target_summaries["release"]["applied_count"] == 1
        assert os.path.isfile(batch.state_path)
        with open(batch.state_path, "r", encoding="utf-8") as stream:
            saved = json.load(stream)
        assert saved["source_branch"] == "develop"
    finally:
        if old_appdata is None:
            os.environ.pop("LOCALAPPDATA", None)
        else:
            os.environ["LOCALAPPDATA"] = old_appdata
        shutil.rmtree(root, ignore_errors=True)


def test_pathfile_is_utf16_without_bom_and_footer_is_generated() -> None:
    captured = []
    observed = {}
    root = tempfile.mkdtemp(prefix="branch-submit-runner-")
    try:
        def runner(args, **kwargs):
            captured.append(args)
            pathfile = next(item.split(":", 1)[1] for item in args if item.startswith("/pathfile:"))
            logmsg = next(item.split(":", 1)[1] for item in args if item.startswith("/logmsgfile:"))
            observed["pathfile"] = open(pathfile, "rb").read()
            observed["logmsg"] = open(logmsg, "rb").read()
            return SimpleNamespace(returncode=0)
        engine = bs.BranchSubmitEngine(root, runner=runner)
        engine.core = FakeCore
        engine._tortoise("commit", [r"C:\测试\A.xlsx", r"C:\测试\B.xlsx"], message="说明\n\n[MultiBranchSync] batch=x source=develop@r1")
        assert observed["pathfile"] == "C:\\测试\\A.xlsx\nC:\\测试\\B.xlsx".encode("utf-16-le")
        assert observed["logmsg"].decode("utf-8-sig").startswith("说明")
        assert "/deletepathfile" in captured[0]
        assert any(item.startswith("/logmsgfile:") for item in captured[0])
    finally:
        shutil.rmtree(root, ignore_errors=True)


def test_source_changes_after_preflight_are_rejected() -> None:
    root = tempfile.mkdtemp(prefix="branch-submit-source-drift-")
    old_appdata = os.environ.get("LOCALAPPDATA")
    os.environ["LOCALAPPDATA"] = os.path.join(root, "appdata")
    try:
        os.makedirs(os.path.join(root, ".svn"))
        for branch in ("develop", "release"):
            os.makedirs(os.path.join(root, branch, "config"))
        source = os.path.join(root, "develop", "config", "A.xlsx")
        target = os.path.join(root, "release", "config", "A.xlsx")
        _book(source, 9); _book(target, 1); _book(source + ".pristine.xlsx", 1)
        engine = bs.BranchSubmitEngine(root, runner=lambda args, **kwargs: SimpleNamespace(returncode=0))
        engine.core = FakeCore
        batch = engine.preflight("develop", ["release"], [source], "说明")
        assert batch.source_status == "ready"
        _book(source, 10)
        engine.commit(batch)
        assert batch.source_status == "failed"
        assert "偏离预览内容" in batch.error
    finally:
        if old_appdata is None:
            os.environ.pop("LOCALAPPDATA", None)
        else:
            os.environ["LOCALAPPDATA"] = old_appdata
        shutil.rmtree(root, ignore_errors=True)


if __name__ == "__main__":
    test_discover_and_validate()
    test_preflight_persists_delta_and_blocks_conflict()
    test_pathfile_is_utf16_without_bom_and_footer_is_generated()
    test_source_changes_after_preflight_are_rejected()
    print("branch submit smoke tests passed")

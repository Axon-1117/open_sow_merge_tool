"""Focused regression for fidelity column mutations next to semantic content."""

from __future__ import annotations

import argparse
import shutil
import tempfile
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

import _large_sheet_excel_fidelity_gate as gate


_CASE = "fidelity-column-semantic-tail"
_SHEET = "Data"
_MARKER = "__SOW_FIDELITY_SEMANTIC_TAIL__"


def _close(workbook) -> None:
    if workbook is not None:
        workbook.close()


def _write_styled_tail(path: Path, *, semantic: bool) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _SHEET
    if semantic:
        sheet["A1"] = "id@id"
        sheet["B1"] = "value@pm"
        sheet["C1"] = "formula@pm"
        sheet["A2"] = "string"
        sheet["B2"] = "string"
        sheet["C2"] = "formula"
        sheet["A3"] = "record-1"
        sheet["B3"] = "payload"
        # A formula must be part of the semantic horizon even if it has no
        # data-only cache value in a paired reader.
        sheet["C3"] = "=40+2"
    # This allocates a physical blank tail at F while leaving D/E/F without a
    # semantic value.  The mutation must use D, not G.
    sheet["F1"].fill = PatternFill(fill_type="solid", fgColor="00FF00")
    workbook.save(path)
    _close(workbook)


def _test_styled_tail_uses_semantic_adjacent_column(root: Path) -> None:
    base = root / "base.xlsx"
    left = root / "left.xlsx"
    right = root / "right.xlsx"
    _write_styled_tail(base, semantic=True)
    shutil.copy2(base, left)
    shutil.copy2(base, right)
    assert gate._sha256(base) == gate._sha256(left) == gate._sha256(right)

    workbook = load_workbook(left, data_only=False)
    try:
        sheet = workbook[_SHEET]
        assert sheet.max_column == 6
        assert gate._worksheet_semantic_bounds(sheet) == (3, 3)
    finally:
        _close(workbook)

    target = gate._mutate_column(right, _SHEET, _MARKER)
    assert target.kind == "column"
    assert (target.row, target.col, target.marker) == (3, 4, _MARKER)
    direct = gate._assert_direct_pair_parity(left, right, _SHEET, "semantic-tail")
    gate._assert_target(direct, target, side="theirs")

    workbook = load_workbook(right, data_only=False)
    try:
        sheet = workbook[_SHEET]
        assert sheet.max_column == 6
        assert sheet.cell(1, 4).value == f"fidelity_{_MARKER}@pm"
        assert sheet.cell(2, 4).value == "string"
        assert sheet.cell(3, 4).value == _MARKER
        assert sheet["F1"].value is None and sheet["F1"].has_style
        assert gate._worksheet_semantic_bounds(sheet) == (3, 4)
    finally:
        _close(workbook)


def _test_semantic_empty_sheet_fails_closed(root: Path) -> None:
    path = root / "styled-only.xlsx"
    _write_styled_tail(path, semantic=False)
    before = gate._sha256(path)
    try:
        gate._mutate_column(path, _SHEET, _MARKER)
    except AssertionError as error:
        assert "semantic worksheet content" in str(error)
    else:
        raise AssertionError("style-only worksheet accepted a column mutation")
    assert gate._sha256(path) == before


def _test_scalar_semantic_bounds() -> None:
    workbook = Workbook()
    try:
        sheet = workbook.active
        # Neither falsey scalar is padding; only an empty string is ignored.
        sheet.cell(4, 3).value = 0
        sheet.cell(5, 4).value = False
        sheet.cell(9, 8).value = ""
        assert gate._worksheet_semantic_bounds(sheet) == (5, 4)
    finally:
        _close(workbook)


def run_case() -> None:
    with tempfile.TemporaryDirectory(prefix="sow_fidelity_column_semantic_tail_") as raw_root:
        root = Path(raw_root)
        _test_styled_tail_uses_semantic_adjacent_column(root)
        _test_semantic_empty_sheet_fails_closed(root)
        _test_scalar_semantic_bounds()


def main(argv=None) -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--case", choices=(_CASE,))
    parser.add_argument("--list-cases", action="store_true")
    args = parser.parse_args(argv)
    if args.list_cases:
        print(_CASE)
        return
    run_case()
    print(f"PASS {_CASE}")


if __name__ == "__main__":
    main()

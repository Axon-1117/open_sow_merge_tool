"""Focused logic matrix for OpenSpec column-alignment task 1.5."""

from openpyxl.worksheet.worksheet import Worksheet

import sow_merge_tool as mod


_worksheet_reads = {"cell": 0, "iter_rows": 0}


def _forbid_worksheet_read(kind):
    def _forbidden(*_args, **_kwargs):
        _worksheet_reads[kind] += 1
        raise AssertionError(f"column alignment called Worksheet.{kind}")

    return _forbidden


def _named_signatures(names):
    return tuple(
        mod.ColumnSignature(
            physical_col=index + 1,
            row_count=3,
            non_empty_count=3,
            first_non_empty_row=1,
            last_non_empty_row=3,
            header_signals=(name,),
            representative_signals=((1, name),),
            non_empty_pattern=(1, 2, 3),
            formula_signals=(),
            intrinsic_key=name,
            exact_content_key=name,
        )
        for index, name in enumerate(names)
    )


def _snapshot(
    value_rows,
    edit_rows=None,
    *,
    sheet_name="Data",
    row_version=1,
    column_version=1,
    max_col=None,
):
    if max_col is None:
        max_col = max((len(row) for row in value_rows), default=0)
    key = mod.ColumnModelCacheKey(
        sheet_name,
        row_version,
        column_version,
        mine_edit_version=column_version,
        base_edit_version=column_version,
        theirs_edit_version=column_version,
    )
    return mod.build_column_signature_snapshot(
        key,
        value_rows,
        edit_rows if edit_rows is not None else value_rows,
        max_col=max_col,
    )


def _implicit_snapshot(
    value_rows,
    *,
    row_version=1,
    column_version=1,
):
    key = mod.ColumnModelCacheKey(
        "Data",
        row_version,
        column_version,
        mine_edit_version=column_version,
        base_edit_version=column_version,
        theirs_edit_version=column_version,
    )
    return mod.build_column_signature_snapshot(key, value_rows, value_rows)


def _expected_blocks(states):
    blocks = []
    start = 0
    for index in range(1, len(states) + 1):
        if index < len(states) and states[index] == states[start]:
            continue
        blocks.append((states[start], tuple(range(start, index))))
        start = index
    return blocks


def _block_layout(result):
    return [(block.state, block.slot_indices) for block in result.model.blocks]


def _assert_integrity(result, mine_count, theirs_count, *, base_count=None):
    slots = result.model.slots
    assert [slot.logical_idx for slot in slots] == list(range(len(slots)))

    expected_counts = {
        "mine_col": mine_count,
        "theirs_col": theirs_count,
    }
    if base_count is None:
        assert all(slot.base_col is None for slot in slots)
        assert all(slot.base_boundary is None for slot in slots)
        assert all(slot.origin_side is None for slot in slots)
    else:
        expected_counts["base_col"] = base_count
        for slot in slots:
            if slot.base_col is not None:
                assert slot.base_boundary is None
                assert slot.origin_side is None
                continue
            assert 0 <= slot.base_boundary <= base_count
            assert slot.origin_side in ("mine", "theirs")
            if slot.origin_side == "mine":
                assert slot.mine_col is not None and slot.theirs_col is None
            else:
                assert slot.mine_col is None and slot.theirs_col is not None

    side_names = {
        "mine_col": "mine",
        "base_col": "base",
        "theirs_col": "theirs",
    }
    for field_name, count in expected_counts.items():
        physical_columns = [
            getattr(slot, field_name)
            for slot in slots
            if getattr(slot, field_name) is not None
        ]
        assert physical_columns == list(range(1, count + 1)), (
            field_name,
            physical_columns,
        )
        assert len(physical_columns) == len(set(physical_columns))
        side = side_names[field_name]
        for slot in slots:
            physical_col = getattr(slot, field_name)
            if physical_col is None:
                continue
            assert result.model.logical_for_physical(side, physical_col) == slot.logical_idx
            assert result.model.physical_for_logical(side, slot.logical_idx) == physical_col

    assert [block.ordinal for block in result.model.blocks] == list(
        range(len(result.model.blocks))
    )
    assert [
        slot_index
        for block in result.model.blocks
        for slot_index in block.slot_indices
    ] == list(range(len(slots)))
    for block in result.model.blocks:
        assert block.state == slots[block.start_slot_idx].state

    expected_fallback = tuple(
        slot.logical_idx
        for slot in slots
        if slot.state == "unresolved" or slot.confidence.ambiguous
    )
    assert result.fallback_slot_indices == expected_fallback


def _assert_cause_on_slots(result, cause_code, slot_indices):
    slot_indices = tuple(slot_indices)
    assert slot_indices
    for slot_index in slot_indices:
        assert cause_code in result.model.slots[slot_index].confidence.cause_codes
    assert cause_code in result.model.confidence.cause_codes
    covering_blocks = [
        block
        for block in result.model.blocks
        if any(slot_index in block.slot_indices for slot_index in slot_indices)
    ]
    assert covering_blocks
    assert all(
        cause_code in block.confidence.cause_codes
        for block in covering_blocks
    )


def _align_2way(left, right):
    result = mod.align_column_signatures_2way(left, right)
    assert result == mod.align_column_signatures_2way(left, right)
    _assert_integrity(result, len(left.signatures if hasattr(left, "signatures") else left),
                      len(right.signatures if hasattr(right, "signatures") else right))
    return result


def _align_3way(mine, base, theirs):
    result = mod.align_column_signatures_3way(mine, base, theirs)
    assert result == mod.align_column_signatures_3way(mine, base, theirs)
    _assert_integrity(
        result,
        len(mine.signatures if hasattr(mine, "signatures") else mine),
        len(theirs.signatures if hasattr(theirs, "signatures") else theirs),
        base_count=len(base.signatures if hasattr(base, "signatures") else base),
    )
    return result


def _assert_retained_name_pairs(result, left_names, right_names):
    for name in set(left_names).intersection(right_names):
        left_col = left_names.index(name) + 1
        right_col = right_names.index(name) + 1
        matching = [slot for slot in result.model.slots if slot.mine_col == left_col]
        assert len(matching) == 1
        assert matching[0].theirs_col == right_col, (name, matching[0])
        assert matching[0].state == "retained", (name, matching[0])


def _test_two_way_edge_insertions_and_deletions():
    base = tuple("ABCDEF")
    insertion_cases = (
        ("leading-single", ("X",) + base, ["inserted"] + ["retained"] * 6),
        ("middle-single", base[:2] + ("X",) + base[2:], ["retained"] * 2 + ["inserted"] + ["retained"] * 4),
        ("tail-single", base + ("X",), ["retained"] * 6 + ["inserted"]),
        ("leading-multiple", ("X", "Y") + base, ["inserted"] * 2 + ["retained"] * 6),
        ("middle-multiple", base[:2] + ("X", "Y") + base[2:], ["retained"] * 2 + ["inserted"] * 2 + ["retained"] * 4),
        ("tail-multiple", base + ("X", "Y"), ["retained"] * 6 + ["inserted"] * 2),
    )
    deletion_cases = (
        ("leading-single", base[1:], ["deleted"] + ["retained"] * 5),
        ("middle-single", base[:2] + base[3:], ["retained"] * 2 + ["deleted"] + ["retained"] * 3),
        ("tail-single", base[:-1], ["retained"] * 5 + ["deleted"]),
        ("leading-multiple", base[2:], ["deleted"] * 2 + ["retained"] * 4),
        ("middle-multiple", base[:2] + base[4:], ["retained"] * 2 + ["deleted"] * 2 + ["retained"] * 2),
        ("tail-multiple", base[:-2], ["retained"] * 4 + ["deleted"] * 2),
    )

    left = _named_signatures(base)
    for label, variant, expected_states in insertion_cases + deletion_cases:
        result = _align_2way(left, _named_signatures(variant))
        actual_states = [slot.state for slot in result.model.slots]
        assert actual_states == expected_states, (label, actual_states)
        assert _block_layout(result) == _expected_blocks(expected_states), label
        assert not result.has_unresolved, label
        _assert_retained_name_pairs(result, base, variant)


def _test_two_way_mixed_insert_delete_keeps_later_anchors():
    base = tuple("ABCDEF")
    variant = ("A", "X", "Y", "B", "D", "E", "F")
    result = _align_2way(_named_signatures(base), _named_signatures(variant))
    expected_states = [
        "retained",
        "inserted",
        "inserted",
        "retained",
        "deleted",
        "retained",
        "retained",
        "retained",
    ]
    assert [slot.state for slot in result.model.slots] == expected_states
    assert _block_layout(result) == _expected_blocks(expected_states)
    _assert_retained_name_pairs(result, base, variant)
    tail = result.model.slots[-1]
    assert (tail.mine_col, tail.theirs_col, tail.state) == (6, 7, "retained")


def _test_unique_header_prefix_anchors_delete_before_replaced_payload():
    base_rows = (
        ("id@id", "part", "part_move", "model", "quality"),
        ("uint32", "map<int,string>", "int32", "string", "int32"),
        ("ID", "部件", "镜头移动", "旧模型描述", "品质"),
        (1, "head", 4, "old-model-a", 1),
        (2, "tail", 2, "old-model-b", 2),
        (3, "wing", 1, "old-model-c", 3),
    )
    theirs_rows = (
        ("id@id", "part", "model", "quality"),
        ("uint32", "map<int,string>", "string", "int32"),
        ("ID", "车辆部件", "新模型描述", "品质"),
        (1, "weapon", "new-model-x", 1),
        (2, "wheel", "new-model-y", 2),
        (3, "trunk", "new-model-z", 3),
    )
    base_formula_rows = tuple(
        row
        if row_idx < 3
        else row[:3] + (f'=A{row_idx + 1}&"-old"',) + row[4:]
        for row_idx, row in enumerate(base_rows)
    )
    theirs_formula_rows = tuple(
        row
        if row_idx < 3
        else row[:2] + (f'=A{row_idx + 1}&"-new"',) + row[3:]
        for row_idx, row in enumerate(theirs_rows)
    )

    for label, base_edits, theirs_edits in (
        ("value-payload", None, None),
        ("uncached-formula-payload", base_formula_rows, theirs_formula_rows),
    ):
        base = _snapshot(
            base_rows,
            base_edits,
            column_version=4,
        )
        theirs = _snapshot(
            theirs_rows,
            theirs_edits,
            column_version=4,
        )

        two_way = _align_2way(base, theirs)
        assert [
            (slot.mine_col, slot.theirs_col, slot.state)
            for slot in two_way.model.slots
        ] == [
            (1, 1, "retained"),
            (2, 2, "retained"),
            (3, None, "deleted"),
            (4, 3, "retained"),
            (5, 4, "retained"),
        ], label
        assert not two_way.has_unresolved, label
        model_slot = two_way.model.slots[3]
        assert model_slot.confidence.reason == "unique-header-prefix-anchor", (
            label,
            model_slot.confidence,
        )
        assert "schema-header" in model_slot.confidence.evidence, label

        three_way = _align_3way(base, base, theirs)
        assert [
            (slot.mine_col, slot.base_col, slot.theirs_col, slot.state)
            for slot in three_way.model.slots
        ] == [
            (1, 1, 1, "retained"),
            (2, 2, 2, "retained"),
            (3, 3, None, "theirs-deleted"),
            (4, 4, 3, "retained"),
            (5, 5, 4, "retained"),
        ], label
        assert not three_way.has_unresolved, label


def _test_duplicate_and_blank_ranges_are_bounded():
    cases = (
        (
            "blank",
            mod.COLUMN_MAPPING_CAUSE_BLANK_COLUMN,
            (("A", None, None, "Z"), ("a", None, None, "z")),
            (("A", None, "Z"), ("a", None, "z")),
        ),
        (
            "duplicate",
            mod.COLUMN_MAPPING_CAUSE_DUPLICATE_SIGNATURE,
            (("A", "X", "X", "Z"), ("a", "x", "x", "z")),
            (("A", "X", "Z"), ("a", "x", "z")),
        ),
    )
    for label, cause_code, left_rows, right_rows in cases:
        left = _snapshot(left_rows, column_version=2)
        base = _snapshot(right_rows)
        result = _align_2way(
            left,
            base,
        )
        assert [slot.state for slot in result.model.slots] == [
            "retained",
            "unresolved",
            "unresolved",
            "retained",
        ], label
        assert _block_layout(result) == [
            ("retained", (0,)),
            ("unresolved", (1,)),
            ("unresolved", (2,)),
            ("retained", (3,)),
        ], label
        later_anchor = result.model.slots[3]
        assert (later_anchor.mine_col, later_anchor.theirs_col) == (4, 3)
        _assert_cause_on_slots(result, cause_code, (1, 2))

        three_way = _align_3way(
            left,
            base,
            _snapshot(right_rows, column_version=3),
        )
        caused_slots = tuple(
            slot.logical_idx
            for slot in three_way.model.slots
            if slot.confidence.ambiguous
        )
        _assert_cause_on_slots(three_way, cause_code, caused_slots)

    left_tail = (("A", None, "X", "X"), ("a", None, "x", "x"))
    right_tail = (("A", None, "X"), ("a", None, "x"))
    tail = _align_2way(
        _snapshot(left_tail),
        _snapshot(right_tail, column_version=2),
    )
    assert [slot.state for slot in tail.model.slots] == [
        "retained",
        "unresolved",
        "unresolved",
        "unresolved",
    ]
    assert _block_layout(tail) == [
        ("retained", (0,)),
        ("unresolved", (1, 2)),
        ("unresolved", (3,)),
    ]
    _assert_cause_on_slots(
        tail,
        mod.COLUMN_MAPPING_CAUSE_BLANK_COLUMN,
        (1, 2, 3),
    )
    _assert_cause_on_slots(
        tail,
        mod.COLUMN_MAPPING_CAUSE_DUPLICATE_SIGNATURE,
        (1, 2, 3),
    )


def _test_formula_shift_and_literal_edit_are_conservative():
    base_values = [("id", "calc", "note", "tail")]
    base_edits = [("id", "calc", "note", "tail")]
    variant_values = [("new", "id", "calc", "note", "tail")]
    variant_edits = [("new", "id", "calc", "note", "tail")]
    for value in range(1, 13):
        row = value + 1
        note = f"note-{value}"
        variant_note = "EDITED" if value == 6 else note
        base_values.append((value, value + 1, note, value * 100))
        base_edits.append((value, f"=A{row}+1", note, value * 100))
        variant_values.append((f"new-{value}", value, value + 1, variant_note, value * 100))
        variant_edits.append((f"new-{value}", value, f"=B{row}+1", variant_note, value * 100))

    base = _snapshot(tuple(base_values), tuple(base_edits))
    variant = _snapshot(
        tuple(variant_values),
        tuple(variant_edits),
        column_version=2,
    )
    result = _align_2way(base, variant)
    assert [slot.state for slot in result.model.slots] == [
        "inserted",
        "retained",
        "unresolved",
        "retained",
        "retained",
    ]
    formula_slot = result.model.slots[2]
    assert (formula_slot.mine_col, formula_slot.theirs_col) == (2, 3)
    assert formula_slot.confidence.ambiguous
    _assert_cause_on_slots(
        result,
        mod.COLUMN_MAPPING_CAUSE_FORMULA_MISMATCH,
        (formula_slot.logical_idx,),
    )
    edited_literal_slot = result.model.slots[3]
    assert (edited_literal_slot.mine_col, edited_literal_slot.theirs_col) == (3, 4)
    assert edited_literal_slot.state == "retained"
    assert edited_literal_slot.confidence.reason == "high-confidence-anchor"

    three_way = _align_3way(
        variant,
        base,
        _snapshot(tuple(base_values), tuple(base_edits), column_version=3),
    )
    three_way_formula = next(
        slot for slot in three_way.model.slots if slot.base_col == 2
    )
    assert three_way_formula.state == "unresolved"
    _assert_cause_on_slots(
        three_way,
        mod.COLUMN_MAPPING_CAUSE_FORMULA_MISMATCH,
        (three_way_formula.logical_idx,),
    )


def _test_three_way_independent_and_edge_insertions():
    base = _named_signatures(("A", "B", "C", "D"))
    mine = _named_signatures(("A", "M", "B", "C", "D"))
    theirs = _named_signatures(("A", "B", "C", "T", "D"))
    middle = _align_3way(mine, base, theirs)
    assert [
        (slot.mine_col, slot.base_col, slot.theirs_col, slot.state)
        for slot in middle.model.slots
    ] == [
        (1, 1, 1, "retained"),
        (2, None, None, "inserted"),
        (3, 2, 2, "retained"),
        (4, 3, 3, "retained"),
        (None, None, 4, "inserted"),
        (5, 4, 5, "retained"),
    ]
    assert not middle.has_unresolved
    mine_insertion = middle.model.slots[1]
    theirs_insertion = middle.model.slots[4]
    assert (mine_insertion.base_boundary, mine_insertion.origin_side) == (1, "mine")
    assert (theirs_insertion.base_boundary, theirs_insertion.origin_side) == (
        3,
        "theirs",
    )

    leading_mine = _named_signatures(("M", "A", "B", "C", "D"))
    trailing_theirs = _named_signatures(("A", "B", "C", "D", "T"))
    edges = _align_3way(leading_mine, base, trailing_theirs)
    assert [
        (slot.mine_col, slot.base_col, slot.theirs_col, slot.state)
        for slot in edges.model.slots
    ] == [
        (1, None, None, "inserted"),
        (2, 1, 1, "retained"),
        (3, 2, 2, "retained"),
        (4, 3, 3, "retained"),
        (5, 4, 4, "retained"),
        (None, None, 5, "inserted"),
    ]
    assert not edges.has_unresolved
    assert (edges.model.slots[0].base_boundary, edges.model.slots[0].origin_side) == (
        0,
        "mine",
    )
    assert (edges.model.slots[-1].base_boundary, edges.model.slots[-1].origin_side) == (
        4,
        "theirs",
    )


def _test_three_way_same_anchor_collision_splits_by_side_presence():
    base_rows = tuple((f"A{row}", f"C{row}") for row in range(1, 101))
    mine_rows = tuple(
        (f"A{row}", "mine-only" if row == 50 else "bounded-same", f"C{row}")
        for row in range(1, 101)
    )
    theirs_rows = tuple(
        (f"A{row}", "theirs-only" if row == 50 else "bounded-same", f"C{row}")
        for row in range(1, 101)
    )
    base = _snapshot(base_rows)
    mine = _snapshot(mine_rows, column_version=2)
    theirs = _snapshot(theirs_rows, column_version=3)
    assert mine_rows[49][1] != theirs_rows[49][1]
    assert mine.signatures[1].intrinsic_key == theirs.signatures[1].intrinsic_key

    result = _align_3way(mine, base, theirs)
    side_only = [slot for slot in result.model.slots if slot.base_col is None]
    assert [(slot.mine_col, slot.theirs_col) for slot in side_only] == [
        (2, None),
        (None, 2),
    ]
    assert all(
        slot.state == "unresolved" and slot.confidence.ambiguous
        for slot in side_only
    )
    assert [
        (slot.base_boundary, slot.origin_side)
        for slot in side_only
    ] == [(1, "mine"), (1, "theirs")]
    assert not any(
        slot.mine_col == 2 and slot.theirs_col == 2
        for slot in result.model.slots
    )
    competing_blocks = [
        block
        for block in result.model.blocks
        if block.state == "unresolved"
        and block.confidence.reason == "ambiguous-competing-insertions"
    ]
    assert [block.slot_indices for block in competing_blocks] == [(1,), (2,)]


def _test_three_way_deletion_directions():
    base = _named_signatures(("A", "B", "C", "D", "E", "F"))
    mine = _named_signatures(("A", "C", "E"))
    theirs = _named_signatures(("A", "B", "D", "E"))
    result = _align_3way(mine, base, theirs)
    assert [
        (slot.base_col, slot.state)
        for slot in result.model.slots
        if slot.base_col is not None
    ] == [
        (1, "retained"),
        (2, "mine-deleted"),
        (3, "theirs-deleted"),
        (4, "mine-deleted"),
        (5, "retained"),
        (6, "both-deleted"),
    ]
    assert [block.state for block in result.model.blocks] == [
        "retained",
        "mine-deleted",
        "theirs-deleted",
        "mine-deleted",
        "retained",
        "both-deleted",
    ]


def _test_three_way_one_side_cache_incompatibility():
    rows = (("A", "B", "C"), ("a", "b", "c"))
    base = _snapshot(rows, row_version=10, column_version=20)
    mine = _snapshot(rows, row_version=11, column_version=21)
    theirs = _snapshot(rows, row_version=10, column_version=22)
    two_way = _align_2way(mine, base)
    assert two_way.fallback_reason == "incompatible-signature-cache"
    assert all(slot.state == "unresolved" for slot in two_way.model.slots)
    _assert_cause_on_slots(
        two_way,
        mod.COLUMN_MAPPING_CAUSE_INCOMPATIBLE_CACHE,
        range(len(two_way.model.slots)),
    )

    result = _align_3way(mine, base, theirs)
    assert result.mine_to_base.used_physical_fallback
    assert result.mine_to_base.fallback_reason == "incompatible-signature-cache"
    assert not result.theirs_to_base.used_physical_fallback
    assert not result.theirs_to_base.has_unresolved
    assert result.theirs_to_base.anchor_pairs == ((1, 1), (2, 2), (3, 3))
    assert all(slot.theirs_col == slot.base_col for slot in result.model.slots)
    assert result.has_unresolved and result.used_physical_fallback
    assert "mine-to-base:incompatible-signature-cache" in result.fallback_reason
    assert "theirs-to-base" not in result.fallback_reason
    _assert_cause_on_slots(
        result,
        mod.COLUMN_MAPPING_CAUSE_INCOMPATIBLE_CACHE,
        range(len(result.model.slots)),
    )


def _test_implicit_boundary_and_low_confidence_causes():
    base_rows = (("A", "B"), (1, 2))
    wider_rows = (("A", "B", "C"), (1, 2, 3))
    base = _implicit_snapshot(base_rows)
    wider = _implicit_snapshot(wider_rows, column_version=2)
    theirs = _implicit_snapshot(base_rows, column_version=3)

    implicit = _align_2way(wider, base)
    assert implicit.fallback_reason == "implicit-column-boundary"
    assert all(slot.state == "unresolved" for slot in implicit.model.slots)
    _assert_cause_on_slots(
        implicit,
        mod.COLUMN_MAPPING_CAUSE_IMPLICIT_BOUNDARY,
        range(len(implicit.model.slots)),
    )

    implicit_three_way = _align_3way(wider, base, theirs)
    assert implicit_three_way.mine_to_base.fallback_reason == (
        "implicit-column-boundary"
    )
    assert not implicit_three_way.theirs_to_base.has_unresolved
    _assert_cause_on_slots(
        implicit_three_way,
        mod.COLUMN_MAPPING_CAUSE_IMPLICIT_BOUNDARY,
        range(len(implicit_three_way.model.slots)),
    )

    mine_rows = (
        ("A", "Business-B", "Z"),
        (1, "old-1", 9),
        (2, "old-2", 10),
    )
    base_rows = (
        ("A", "Different-Q", "Z"),
        (1, "new-x", 9),
        (2, "new-y", 10),
    )
    mine = _snapshot(mine_rows, column_version=2)
    base = _snapshot(base_rows)
    theirs = _snapshot(base_rows, column_version=3)

    low = _align_2way(mine, base)
    assert [slot.state for slot in low.model.slots] == [
        "retained",
        "unresolved",
        "retained",
    ]
    _assert_cause_on_slots(
        low,
        mod.COLUMN_MAPPING_CAUSE_LOW_CONFIDENCE,
        (1,),
    )

    low_three_way = _align_3way(mine, base, theirs)
    low_slot = next(slot for slot in low_three_way.model.slots if slot.base_col == 2)
    assert low_slot.state == "unresolved"
    _assert_cause_on_slots(
        low_three_way,
        mod.COLUMN_MAPPING_CAUSE_LOW_CONFIDENCE,
        (low_slot.logical_idx,),
    )


def _test_empty_side_force_fallbacks_remain_unresolved():
    def _assert_whole_fallback(result, fallback_reason, cause_code, count):
        assert result.used_physical_fallback
        assert result.fallback_reason == fallback_reason
        assert len(result.model.slots) == count
        assert all(slot.state == "unresolved" for slot in result.model.slots)
        assert not any(
            slot.state in ("inserted", "deleted")
            for slot in result.model.slots
        )
        _assert_cause_on_slots(result, cause_code, range(count))

    wide = _named_signatures(tuple(f"C{index}" for index in range(1, 258)))
    _assert_whole_fallback(
        _align_2way((), wide),
        "column-limit-exceeded",
        mod.COLUMN_MAPPING_CAUSE_COLUMN_LIMIT,
        257,
    )
    _assert_whole_fallback(
        _align_2way(wide, ()),
        "column-limit-exceeded",
        mod.COLUMN_MAPPING_CAUSE_COLUMN_LIMIT,
        257,
    )

    rows = (("A", "B"), (1, 2))
    populated = _snapshot(rows, row_version=1)
    incompatible_empty = _snapshot((), row_version=2, column_version=2, max_col=0)
    for left, right in (
        (incompatible_empty, populated),
        (populated, incompatible_empty),
    ):
        _assert_whole_fallback(
            _align_2way(left, right),
            "incompatible-signature-cache",
            mod.COLUMN_MAPPING_CAUSE_INCOMPATIBLE_CACHE,
            2,
        )

    implicit_populated = _implicit_snapshot(rows)
    implicit_empty = _implicit_snapshot((), column_version=2)
    for left, right in (
        (implicit_empty, implicit_populated),
        (implicit_populated, implicit_empty),
    ):
        _assert_whole_fallback(
            _align_2way(left, right),
            "implicit-column-boundary",
            mod.COLUMN_MAPPING_CAUSE_IMPLICIT_BOUNDARY,
            2,
        )


def _test_three_way_256_257_boundary():
    signatures = _named_signatures(tuple(f"C{index}" for index in range(1, 258)))
    two_way_at_limit = _align_2way(signatures[:256], signatures[:256])
    assert not two_way_at_limit.has_unresolved
    assert not two_way_at_limit.used_physical_fallback

    at_limit = _align_3way(signatures[:256], signatures[:256], signatures[:256])
    assert len(at_limit.model.slots) == 256
    assert not at_limit.has_unresolved
    assert not at_limit.used_physical_fallback
    assert all(slot.state == "retained" for slot in at_limit.model.slots)

    two_way_over_limit = _align_2way(signatures, signatures)
    assert two_way_over_limit.fallback_reason == "column-limit-exceeded"
    assert all(slot.state == "unresolved" for slot in two_way_over_limit.model.slots)
    _assert_cause_on_slots(
        two_way_over_limit,
        mod.COLUMN_MAPPING_CAUSE_COLUMN_LIMIT,
        range(257),
    )

    over_limit = _align_3way(signatures, signatures, signatures)
    assert len(over_limit.model.slots) == 257
    assert over_limit.has_unresolved and over_limit.used_physical_fallback
    assert over_limit.mine_to_base.fallback_reason == "column-limit-exceeded"
    assert over_limit.theirs_to_base.fallback_reason == "column-limit-exceeded"
    assert all(
        slot.state == "unresolved" and slot.confidence.ambiguous
        for slot in over_limit.model.slots
    )
    _assert_cause_on_slots(
        over_limit,
        mod.COLUMN_MAPPING_CAUSE_COLUMN_LIMIT,
        range(257),
    )


def main():
    original_cell = Worksheet.cell
    original_iter_rows = Worksheet.iter_rows
    Worksheet.cell = _forbid_worksheet_read("cell")
    Worksheet.iter_rows = _forbid_worksheet_read("iter_rows")
    try:
        _test_two_way_edge_insertions_and_deletions()
        _test_two_way_mixed_insert_delete_keeps_later_anchors()
        _test_unique_header_prefix_anchors_delete_before_replaced_payload()
        _test_duplicate_and_blank_ranges_are_bounded()
        _test_formula_shift_and_literal_edit_are_conservative()
        _test_three_way_independent_and_edge_insertions()
        _test_three_way_same_anchor_collision_splits_by_side_presence()
        _test_three_way_deletion_directions()
        _test_three_way_one_side_cache_incompatibility()
        _test_implicit_boundary_and_low_confidence_causes()
        _test_empty_side_force_fallbacks_remain_unresolved()
        _test_three_way_256_257_boundary()
    finally:
        Worksheet.cell = original_cell
        Worksheet.iter_rows = original_iter_rows

    assert _worksheet_reads == {"cell": 0, "iter_rows": 0}
    print("SMOKE_TEST_COLUMN_ALIGNMENT_MATRIX_OK")


if __name__ == "__main__":
    main()

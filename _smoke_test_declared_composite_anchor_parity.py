"""Focused parity proof for explicit composite row identities."""

from __future__ import annotations

import argparse
import hashlib
import os
import tempfile
from pathlib import Path

from openpyxl import Workbook, load_workbook

import sow_merge_tool as sm
from _large_sheet_direct_oracle import capture as capture_direct
from _large_sheet_snapshot_oracle import compare_manifests


_CASE = "declared-composite-anchor-parity"
_SHEET = "Composite"


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _write_book(
    path: Path,
    *,
    presentation: str,
    changed_identity: bool = False,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _SHEET
    sheet.append(("building@id", "name@pm", "level@id"))
    sheet.append(("string", "string", "int"))
    # A nonblank descriptor with a blank continuation is the real composite
    # shape that must be aligned identically by direct and immutable paths.
    sheet.append(("descriptor-building", "description", "descriptor-level"))
    sheet.append((None, None, None))
    sheet.append((
        "building-a" if not changed_identity else "building-a-renamed",
        presentation,
        1,
    ))
    sheet.append(("building-a", "name-two", 2))
    workbook.save(path)
    workbook.close()


def _write_no_raw_anchor_book(path: Path, *, right_side: bool) -> None:
    """Write matching declared keys whose raw header rows cannot anchor."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _SHEET
    if right_side:
        sheet.append(("building@const", "name@right", "level@const"))
    else:
        sheet.append(("building@id", "name@left", "level@id"))
    # All type cells are intentionally duplicated, so row 2 cannot make a
    # raw unique anchor.  The explicit declared identities remain clear.
    sheet.append(("string", "string", "string"))
    if right_side:
        sheet.append(("right-descriptor", "right-description", "right-level"))
        sheet.append((None, None, None))
        sheet.append(("right-lead", "leading", "0"))
        sheet.append(("building-a", "name-one", "1"))
        sheet.append(("building-b", "name-two", "2"))
    else:
        sheet.append(("left-descriptor", "left-description", "left-level"))
        sheet.append((None, None, None))
        sheet.append(("building-a", "name-one", "1"))
        sheet.append(("building-b", "name-two", "2"))
        # Retain a seventh physical row without introducing a raw shared
        # token: the right side's b/2 record is deliberately one row later.
        sheet.append((None, None, None))
    workbook.save(path)
    workbook.close()


def _snapshot_manifest(mine: Path, theirs: Path) -> dict:
    left = sm._stream_selected_sheet_snapshot(str(mine), str(mine), _SHEET, "A")
    right = sm._stream_selected_sheet_snapshot(str(theirs), str(theirs), _SHEET, "B")
    result = sm._compare_selected_sheet_snapshots(left, right)
    if result.unresolved:
        raise AssertionError(("unexpected unresolved", result.unresolved_reason))
    return sm.snapshot_comparison_oracle_manifest(left, right, result)


def _assert_exact_parity(mine: Path, theirs: Path) -> tuple[dict, dict]:
    direct = capture_direct(str(mine), str(theirs), _SHEET)
    snapshot = _snapshot_manifest(mine, theirs)
    parity = compare_manifests(direct, snapshot)
    if not parity["exact"]:
        raise AssertionError(parity["mismatches"])
    if direct["columns"] != snapshot["columns"]:
        raise AssertionError(("columns", direct["columns"], snapshot["columns"]))
    if direct["records"] != snapshot["records"]:
        raise AssertionError(("records", direct["records"], snapshot["records"]))
    if direct["only_diff_rows"] != snapshot["only_diff_rows"]:
        raise AssertionError(("only-diff", direct["only_diff_rows"], snapshot["only_diff_rows"]))
    return direct, snapshot


def _raw_header_anchors(left_rows, right_rows, width: int) -> tuple[tuple[int, int], ...]:
    """Mirror only the legacy raw-header anchor observation for this gate."""
    anchors = []
    best_score = (-1, -1)
    for header_idx in range(min(8, len(left_rows), len(right_rows))):
        def _unique_positions(row):
            positions = {}
            duplicates = set()
            for offset, value in enumerate(row[:width]):
                token = sm._row_alignment_cell_token(value)
                if token == "BLANK:":
                    continue
                if token in positions:
                    duplicates.add(token)
                else:
                    positions[token] = offset
            for token in duplicates:
                positions.pop(token, None)
            return positions

        left_positions = _unique_positions(left_rows[header_idx])
        right_positions = _unique_positions(right_rows[header_idx])
        candidate = tuple(sorted(
            (
                (left_positions[token], right_offset)
                for token, right_offset in right_positions.items()
                if token in left_positions
            ),
            key=lambda item: item[1],
        ))
        score = (len(candidate), len(left_positions) + len(right_positions))
        if score > best_score:
            anchors = list(candidate)
            best_score = score
    return tuple(anchors)


def _assert_no_anchor_declared_signature_gate(mine: Path, theirs: Path) -> None:
    def _values(path: Path):
        workbook = load_workbook(path, data_only=True, read_only=True)
        try:
            return [tuple(row) for row in workbook[_SHEET].iter_rows(values_only=True)]
        finally:
            workbook.close()

    left_rows, right_rows = _values(mine), _values(theirs)
    assert _raw_header_anchors(left_rows, right_rows, 3) == ()
    original = sm._compute_row_pairs_from_signatures
    captured = []

    def _recording_pairs(left_signatures, right_signatures):
        captured.append((tuple(left_signatures), tuple(right_signatures)))
        return original(left_signatures, right_signatures)

    sm._compute_row_pairs_from_signatures = _recording_pairs
    try:
        direct = capture_direct(str(mine), str(theirs), _SHEET)
    finally:
        sm._compute_row_pairs_from_signatures = original
    assert len(captured) == 1, captured
    left_signatures, _right_signatures = captured[0]
    assert all(
        signature.startswith(sm._UNIQUE_ROW_KEY_SIGNATURE_PREFIX)
        for signature in left_signatures[4:6]
    ), left_signatures
    snapshot = _snapshot_manifest(mine, theirs)
    parity = compare_manifests(direct, snapshot)
    assert parity["exact"], parity["mismatches"]


def _assert_owned_cleanup(
    temporary: tempfile.TemporaryDirectory[str],
    root: Path,
    expected_hashes: dict[Path, str],
    primary: BaseException | None,
) -> None:
    failures = []
    for path, expected in expected_hashes.items():
        if path.exists() and _sha256(path) != expected:
            failures.append(f"input SHA changed: {path}")
    try:
        temporary.cleanup()
    except BaseException as exc:  # preserve a preceding assertion as primary
        failures.append(f"owned TemporaryDirectory cleanup: {exc!r}")
    if os.path.lexists(root):
        failures.append(f"owned root remains: {root}")
    if failures:
        secondary = AssertionError("; ".join(failures))
        if primary is not None:
            primary.add_note(str(secondary))
        else:
            raise secondary


def run_case() -> None:
    temporary = tempfile.TemporaryDirectory(prefix="sow_declared_composite_anchor_")
    root = Path(temporary.name)
    mine = root / "mine.xlsx"
    theirs = root / "theirs.xlsx"
    primary: BaseException | None = None
    expected_hashes: dict[Path, str] = {}
    try:
        _write_book(mine, presentation="name-one")
        _write_book(theirs, presentation="changed-name")
        expected_hashes = {mine: _sha256(mine), theirs: _sha256(theirs)}
        direct, _snapshot = _assert_exact_parity(mine, theirs)
        assert len(direct["records"]) == 1, direct["records"]
        record = direct["records"][0]
        assert (record["mine_row"], record["theirs_row"]) == (5, 5), record
        assert record["row_structure"] is False, record
        assert record["diff_cols"] == [2], record
        assert record["base_diff_cols"] == [], record
        assert record["cells"] == {
            "2": {
                "mine": {"present": True, "cached": {"type": "str", "value": "name-one"}, "cached_data_type": "s", "formula": None, "formula_data_type": "s"},
                "theirs": {"present": True, "cached": {"type": "str", "value": "changed-name"}, "cached_data_type": "s", "formula": None, "formula_data_type": "s"},
                "base": {"present": False},
            },
        }, record
        assert direct["only_diff_rows"] == [record["pair"]], direct

        # Changing an explicit identity must not make the test assume a
        # physical pairing.  High-confidence distinct keys must surface as
        # separate structural records, while direct/snapshot remain exact.
        _write_book(theirs, presentation="name-one", changed_identity=True)
        expected_hashes[theirs] = _sha256(theirs)
        direct, _snapshot = _assert_exact_parity(mine, theirs)
        one_sided = [
            record for record in direct["records"]
            if int(record.get("mine_row") or 0) == 5
            or int(record.get("theirs_row") or 0) == 5
        ]
        assert {(record["mine_row"], record["theirs_row"]) for record in one_sided} == {
            (5, None),
            (None, 5),
        }, one_sided
        assert all(record["row_structure"] is True for record in one_sided), one_sided

        no_anchor_mine = root / "no-anchor-mine.xlsx"
        no_anchor_theirs = root / "no-anchor-theirs.xlsx"
        _write_no_raw_anchor_book(no_anchor_mine, right_side=False)
        _write_no_raw_anchor_book(no_anchor_theirs, right_side=True)
        expected_hashes[no_anchor_mine] = _sha256(no_anchor_mine)
        expected_hashes[no_anchor_theirs] = _sha256(no_anchor_theirs)
        _assert_no_anchor_declared_signature_gate(no_anchor_mine, no_anchor_theirs)
    except BaseException as exc:
        primary = exc
        raise
    finally:
        _assert_owned_cleanup(temporary, root, expected_hashes, primary)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--case", choices=(_CASE,))
    parser.add_argument("--list-cases", action="store_true")
    args = parser.parse_args()
    if args.list_cases:
        print(_CASE)
        return
    run_case()
    print(f"PASS {_CASE}")


if __name__ == "__main__":
    main()

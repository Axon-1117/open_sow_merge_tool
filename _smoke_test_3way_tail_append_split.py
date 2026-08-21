"""Bounded 3-way tail insertion/save regression with an exact READY fixture."""

from __future__ import annotations

import argparse
import hashlib
import shutil
import tempfile
import time
from pathlib import Path

from openpyxl import Workbook, load_workbook

import sow_merge_tool as mod
from _test_temp_utils import make_temp_dir


_CASE = "theirs-tail-b2a-save"
_SHEET = "S1"


def _make_book(path: Path, rows: list[list[object]]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = _SHEET
    for row in rows:
        worksheet.append(row)
    workbook.save(path)
    workbook.close()


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _pump(app) -> None:
    app.root.update_idletasks()
    app.root.update()


def _new_app(mine: Path, theirs: Path, base: Path, merged: Path):
    original_scheduler = mod.SowMergeApp._schedule_formula_cache_prompt
    mod.SowMergeApp._schedule_formula_cache_prompt = lambda _self: None
    try:
        return mod.SowMergeApp(
            str(mine),
            str(theirs),
            merge_mode=True,
            merged_path=str(merged),
            base_path=str(base),
        ), original_scheduler
    except Exception:
        mod.SowMergeApp._schedule_formula_cache_prompt = original_scheduler
        raise


def _close_app(app, original_scheduler) -> None:
    try:
        if app is not None:
            app._shutdown_root()
    finally:
        if original_scheduler is not None:
            mod.SowMergeApp._schedule_formula_cache_prompt = original_scheduler


def _wait_for_operation_ready(app, *, timeout: float = 15.0):
    app.nb.select(app._sheet_containers[_SHEET])
    deadline = time.monotonic() + timeout
    view = None
    while time.monotonic() < deadline:
        app._request_edit_preload()
        _pump(app)
        view = app.sheet_views.get(_SHEET)
        if (
            view is not None
            and app.selected_sheet == _SHEET
            and view._data_ready
            and app._is_sheet_exact_current(_SHEET)
            and app._edit_workbooks_ready()
            and view._derive_lifecycle_state() == "READY"
        ):
            return view
        time.sleep(0.01)
    raise AssertionError(
        "tail insertion fixture did not reach selected current READY within 15s: "
        f"view={view!r} exact={app._sheet_exact_entry(_SHEET)!r} "
        f"edit_ready={app._edit_workbooks_ready()!r}"
    )


def _theirs_tail_b2a_save(root: Path) -> None:
    base = root / "base.xlsx"
    mine = root / "mine.xlsx"
    theirs = root / "theirs.xlsx"
    merged = root / "merged.xlsx"
    schema = [["id@id", "name@pm", "formula@pm"], ["int", "string", "int"]]
    _make_book(base, schema + [[1, "x", "=1"], [2, "y", "=1"]])
    _make_book(mine, schema + [[1, "x", "=1"], [2, "y", "=1"], [3, "a", "=1"]])
    _make_book(theirs, schema + [[1, "x", "=1"], [2, "y", "=1"], [4, "b", "=1"]])
    before_hashes = tuple(_sha256(path) for path in (base, mine, theirs))

    app = None
    original_scheduler = None
    output = None
    try:
        app, original_scheduler = _new_app(mine, theirs, base, merged)
        print("SMOKE_3WAY_TAIL_APPEND_SPLIT_STAGE ready", flush=True)
        view = _wait_for_operation_ready(app)
        assert app._sheet_exact_entry(_SHEET).get("state") == mod._SHEET_EXACT_CHANGED
        theirs_pair = next(
            pair_index
            for pair_index, (row_a, row_b) in enumerate(view.row_pairs)
            if row_a is None and row_b is not None and app.ws_b_val(_SHEET).cell(row=row_b, column=1).value == 4
        )

        print("SMOKE_3WAY_TAIL_APPEND_SPLIT_STAGE b2a", flush=True)
        assert view._copy_selected_row("B2A", override_pair_idx=theirs_pair), "B2A on theirs-only tail failed"
        assert [app.ws_a_val(_SHEET).cell(row=row, column=1).value for row in range(1, 7)] == [
            "id@id", "int", 1, 2, 4, 3,
        ]
        assert app.ws_a_val(_SHEET).cell(row=5, column=2).value == "b"
        assert app.ws_a_val(_SHEET).cell(row=6, column=2).value == "a"

        print("SMOKE_3WAY_TAIL_APPEND_SPLIT_STAGE save-reopen", flush=True)
        output = Path(app.build_manual_merge_output_file())
        temp_root = Path(tempfile.gettempdir()).resolve()
        assert output.is_file() and output.resolve().parent == temp_root, output
        workbook = load_workbook(output, data_only=False)
        try:
            worksheet = workbook[_SHEET]
            assert [worksheet.cell(row=row, column=1).value for row in range(1, 7)] == [
                "id@id", "int", 1, 2, 4, 3,
            ]
            assert worksheet.cell(row=5, column=2).value == "b"
            assert worksheet.cell(row=6, column=2).value == "a"
            assert worksheet.cell(row=5, column=3).value == "=1"
            assert worksheet.cell(row=6, column=3).value == "=1"
        finally:
            workbook.close()
        assert tuple(_sha256(path) for path in (base, mine, theirs)) == before_hashes
    finally:
        _close_app(app, original_scheduler)
        if output is not None:
            output.unlink(missing_ok=True)


def main(argv=None) -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--list-cases", action="store_true")
    parser.add_argument("--case", choices=(_CASE,))
    args = parser.parse_args(argv)
    if args.list_cases:
        if args.case:
            parser.error("--list-cases cannot be combined with --case")
        print(_CASE, flush=True)
        return

    selected = args.case or _CASE
    root = Path(make_temp_dir("sow_tail_append_split_"))
    try:
        print(f"SMOKE_3WAY_TAIL_APPEND_SPLIT_CASE_START {selected}", flush=True)
        _theirs_tail_b2a_save(root)
        print(f"SMOKE_3WAY_TAIL_APPEND_SPLIT_CASE_OK {selected}", flush=True)
    finally:
        shutil.rmtree(root, ignore_errors=True)
    print("SMOKE_3WAY_TAIL_APPEND_SPLIT_OK", flush=True)


if __name__ == "__main__":
    main()

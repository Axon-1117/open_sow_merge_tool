"""Bounded three-way alignment regressions with independently selectable cases."""

from __future__ import annotations

import argparse
import hashlib
import os
import shutil
import time
from pathlib import Path

from openpyxl import Workbook, load_workbook

import sow_merge_tool as mod
from _test_temp_utils import make_temp_dir


_CASES = (
    "preflight-no-conflict",
    "unresolved-guard",
    "base-row-mapping-action-save",
)


def _make_book(path: Path, rows: list[object]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "S1"
    for row_index, row in enumerate(rows, start=1):
        if isinstance(row, (list, tuple)):
            ws.append(list(row))
        else:
            ws.cell(row=row_index, column=1).value = row
    wb.save(path)
    wb.close()


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _pump(app) -> None:
    app.root.update_idletasks()
    app.root.update()


def _new_app(mine: Path, theirs: Path, base: Path, merged: Path):
    """Create a noninteractive app while retaining production modal behavior."""
    original_scheduler = mod.SowMergeApp._schedule_formula_cache_prompt
    mod.SowMergeApp._schedule_formula_cache_prompt = lambda _self: None
    try:
        return mod.SowMergeApp(
            str(mine),
            str(theirs),
            merge_mode=True,
            merged_path=str(merged),
            base_path=str(base),
        ), original_scheduler
    except Exception:
        mod.SowMergeApp._schedule_formula_cache_prompt = original_scheduler
        raise


def _close_app(app, original_scheduler) -> None:
    try:
        if app is not None:
            app._shutdown_root()
    finally:
        if original_scheduler is not None:
            mod.SowMergeApp._schedule_formula_cache_prompt = original_scheduler


def _wait_for_unresolved(app, *, timeout: float = 15.0):
    app.nb.select(app._sheet_containers["S1"])
    deadline = time.monotonic() + timeout
    view = None
    while time.monotonic() < deadline:
        _pump(app)
        view = app.sheet_views.get("S1")
        entry = app._sheet_exact_entry("S1")
        if view is not None and entry.get("state") == mod._SHEET_EXACT_UNRESOLVED:
            return view
        time.sleep(0.01)
    raise AssertionError(
        "ambiguous fixture did not reach current UNRESOLVED state within 15s: "
        f"view={view!r} exact={app._sheet_exact_entry('S1')!r}"
    )


def _wait_for_operation_ready(app, *, timeout: float = 15.0):
    """Explicit test demand before asserting a completed mutation."""
    app.nb.select(app._sheet_containers["S1"])
    deadline = time.monotonic() + timeout
    view = None
    while time.monotonic() < deadline:
        app._request_edit_preload()
        _pump(app)
        view = app.sheet_views.get("S1")
        if (
            view is not None
            and app.selected_sheet == "S1"
            and view._data_ready
            and app._is_sheet_exact_current("S1")
            and app._edit_workbooks_ready()
            and view._derive_lifecycle_state() == "READY"
        ):
            return view
        time.sleep(0.01)
    raise AssertionError(
        "exact three-way fixture did not reach operation READY within 15s: "
        f"view={view!r} exact={app._sheet_exact_entry('S1')!r} "
        f"edit={app._edit_workbooks_ready()!r}"
    )


def _preflight_no_conflict(root: Path) -> None:
    base, mine, theirs = (root / "preflight-base.xlsx", root / "preflight-mine.xlsx", root / "preflight-theirs.xlsx")
    _make_book(base, ["id", "A", "C"])
    _make_book(mine, ["id", "A", "B", "C"])
    _make_book(theirs, ["id", "A", "X"])
    conflicts, cmap = mod._scan_three_way_conflicts(str(base), str(mine), str(theirs))
    assert not conflicts, (conflicts, cmap)


def _unresolved_guard(root: Path) -> None:
    """The old ambiguous fixture must remain terminal and non-actionable."""
    base, mine, theirs = (root / "ambiguous-base.xlsx", root / "ambiguous-mine.xlsx", root / "ambiguous-theirs.xlsx")
    _make_book(base, ["id", "A", "C"])
    _make_book(mine, ["id", "A", "B", "C"])
    _make_book(theirs, ["id", "A", "B", "C"])
    before_hashes = tuple(_sha256(path) for path in (base, mine, theirs))
    app = None
    original_scheduler = None
    try:
        app, original_scheduler = _new_app(mine, theirs, base, root / "ambiguous-merged.xlsx")
        view = _wait_for_unresolved(app)
        before_values = tuple(app.ws_a_val("S1").cell(row=row, column=1).value for row in range(1, 5))
        before_ops = (
            list(app.manual_a_cell_ops.items()),
            list(app.manual_a_row_ops),
            list(app.undo_stack),
        )
        # `notify=False` exercises the production readiness predicate without
        # presenting the interactive modal that a smoke runner cannot dismiss.
        assert not view._guard_mutation_ready("three-way ambiguity regression", notify=False)
        assert view._derive_lifecycle_state() == "UNRESOLVED"
        assert tuple(app.ws_a_val("S1").cell(row=row, column=1).value for row in range(1, 5)) == before_values
        assert (
            list(app.manual_a_cell_ops.items()),
            list(app.manual_a_row_ops),
            list(app.undo_stack),
        ) == before_ops
        assert tuple(_sha256(path) for path in (base, mine, theirs)) == before_hashes
    finally:
        _close_app(app, original_scheduler)


def _base_row_mapping_action_save(root: Path) -> None:
    """Exercise a current READY mapping using valid declarations and unique keys."""
    base, mine, theirs = (root / "exact-base.xlsx", root / "exact-mine.xlsx", root / "exact-theirs.xlsx")
    headers = [["id@id", "value@pm"], ["int", "string"]]
    _make_book(base, headers + [[1, "A"], [3, "C"]])
    _make_book(mine, headers + [[1, "A"], [2, "B"], [3, "C"]])
    _make_book(theirs, headers + [[1, "A"], [2, "B"], [3, "C"]])
    app = None
    original_scheduler = None
    output = None
    try:
        app, original_scheduler = _new_app(mine, theirs, base, root / "exact-merged.xlsx")
        view = _wait_for_operation_ready(app)
        assert app._sheet_exact_entry("S1").get("state") == mod._SHEET_EXACT_CHANGED

        insert_pair = next(
            pair_index
            for pair_index, (row_a, _row_b) in enumerate(view.row_pairs)
            if row_a is not None and app.ws_a_val("S1").cell(row=row_a, column=1).value == 2
        )
        mapped_pair = view.row_a_to_pair_idx[5]
        assert view.row_pairs[mapped_pair][0] == 5
        assert view.mine_to_base_row.get(5) == 4
        assert view._base_row_for_pair(mapped_pair, view.row_pairs[mapped_pair]) == 4
        assert view._action_physical_columns("BASE2A", 1) == (1, 1)
        assert "此侧" in view._build_base_line(insert_pair).replace(mod._TK_INDEX_PLACEHOLDER, "")

        # Prove the mapped Base physical row restores Mine row 5 before the
        # separate BASE2A structural deletion of the common inserted row.
        app.ws_a_val("S1").cell(row=5, column=1).value = "wrong"
        app.ws_a_edit("S1").cell(row=5, column=1).value = "wrong"
        view._copy_single_cell_by_pair(mapped_pair, "BASE2A", 1)
        assert app.ws_a_val("S1").cell(row=5, column=1).value == 3
        assert app.ws_a_edit("S1").cell(row=5, column=1).value == 3

        assert view._copy_selected_row("BASE2A", override_pair_idx=insert_pair), "BASE2A delete failed"
        output = Path(app.build_manual_merge_output_file())
        wb = load_workbook(output, data_only=False)
        try:
            ws = wb["S1"]
            assert [ws.cell(row=row, column=1).value for row in range(1, ws.max_row + 1)] == ["id@id", "int", 1, 3]
        finally:
            wb.close()
    finally:
        _close_app(app, original_scheduler)
        if output is not None:
            output.unlink(missing_ok=True)


_CASE_FUNCS = {
    "preflight-no-conflict": _preflight_no_conflict,
    "unresolved-guard": _unresolved_guard,
    "base-row-mapping-action-save": _base_row_mapping_action_save,
}


def main(argv=None) -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--list-cases", action="store_true")
    parser.add_argument("--case", choices=_CASES)
    args = parser.parse_args(argv)
    if args.list_cases:
        if args.case:
            parser.error("--list-cases cannot be combined with --case")
        for case in _CASES:
            print(case, flush=True)
        return

    selected = (args.case,) if args.case else _CASES
    root = Path(make_temp_dir("sow_3way_align_"))
    try:
        for case in selected:
            print(f"SMOKE_3WAY_ALIGNMENT_CASE_START {case}", flush=True)
            _CASE_FUNCS[case](root)
            print(f"SMOKE_3WAY_ALIGNMENT_CASE_OK {case}", flush=True)
    finally:
        shutil.rmtree(root, ignore_errors=True)
    print("SMOKE_3WAY_ALIGNMENT_OK", flush=True)


if __name__ == "__main__":
    main()

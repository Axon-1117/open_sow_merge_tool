import os
import sys
import tempfile
import time

from openpyxl import Workbook
from openpyxl.worksheet._read_only import ReadOnlyWorksheet
from openpyxl.worksheet.worksheet import Worksheet

import sow_merge_tool as sm


def _make_book(path: str, suffix: str):
    wb = Workbook()
    first = wb.active
    first.title = "S1"
    second = wb.create_sheet("S2")
    for ws in (first, second):
        ws.append(["id @id", "value"])
        for row in range(1, 401):
            ws.append([row, f"{ws.title}-{row}-{suffix}"])
    wb.save(path)
    wb.close()


def main():
    with tempfile.TemporaryDirectory(prefix="sow_sheet_cache_") as tmp:
        previous_snapshot_flag = sm._LARGE_SHEET_SNAPSHOT_ENGINE_ENABLED
        previous_large_sheet_threshold = sm._LARGE_SHEET_ROW_THRESHOLD
        sm._LARGE_SHEET_SNAPSHOT_ENGINE_ENABLED = True
        sm._LARGE_SHEET_ROW_THRESHOLD = 200
        left = os.path.join(tmp, "Book.xlsx")
        right = os.path.join(tmp, "Book-right.xlsx")
        _make_book(left, "A")
        _make_book(right, "B")

        app = sm.SowMergeApp(left, right, initial_sheet="S2")
        # The explicit initial Sheet is the only one eligible for snapshot
        # parsing; sibling worksheet XML must stay unopened.
        deadline = time.monotonic() + 0.5
        while time.monotonic() < deadline:
            app.root.update()
            time.sleep(0.02)
        assert "S1" not in app._sheet_cache_store, "unopened Sheet was eagerly scanned"
        assert app.sheet_views.get("S1") is None

        deadline = time.monotonic() + 20
        view = None
        while time.monotonic() < deadline:
            app.root.update()
            view = app.sheet_views.get("S2")
            if view is not None and view._data_ready:
                break
            time.sleep(0.02)
        assert view is not None and view._data_ready, "selected Sheet was not prepared"
        assert getattr(view, "_cache_source", "") == "snapshot", "selected Sheet did not use snapshot cache"
        assert "S1" not in app._sheet_cache_store
        snapshot = app.selected_sheet_snapshot("A", "S2")
        assert snapshot is not None and snapshot.sheet == "S2"
        old_cell, old_iter_rows = Worksheet.cell, Worksheet.iter_rows
        old_ro_iter_rows = ReadOnlyWorksheet.iter_rows
        Worksheet.cell = lambda *_args, **_kwargs: (_ for _ in ()).throw(
            AssertionError("cached revisit called Worksheet.cell")
        )
        Worksheet.iter_rows = lambda *_args, **_kwargs: (_ for _ in ()).throw(
            AssertionError("cached revisit called Worksheet.iter_rows")
        )
        ReadOnlyWorksheet.iter_rows = lambda *_args, **_kwargs: (_ for _ in ()).throw(
            AssertionError("cached revisit iterated a read-only worksheet")
        )
        try:
            assert app.selected_sheet_snapshot("A", "S2") is snapshot, "unchanged snapshot was not reused"
        finally:
            Worksheet.cell, Worksheet.iter_rows = old_cell, old_iter_rows
            ReadOnlyWorksheet.iter_rows = old_ro_iter_rows
        # Snapshot readiness must not require editable workbooks.  The first
        # explicit mutation/access request starts one background owner and is
        # rejected for retry; it must never parse normal workbooks on Tk.
        assert not app._edit_workbooks_ready(), "editable books loaded before snapshot demand"
        try:
            app.ws_a_edit("S2")
            raise AssertionError("first editable access synchronously loaded workbooks")
        except RuntimeError as exc:
            assert "后台准备" in str(exc)
        assert app._edit_loading_started, "first editable request did not start async owner"
        deadline = time.monotonic() + 20
        while time.monotonic() < deadline and not app._edit_workbooks_ready():
            app.root.update()
            time.sleep(0.02)
        assert app._edit_workbooks_ready(), "async editable loader did not become ready"
        assert app.ws_a_edit("S2")["B2"].value == "S2-1-A"
        with app._compute_lock:
            assert "S2" not in app._compute_queue
            assert "S2" not in app._compute_inflight

        app._shutdown_root()
        sm._LARGE_SHEET_SNAPSHOT_ENGINE_ENABLED = previous_snapshot_flag
        sm._LARGE_SHEET_ROW_THRESHOLD = previous_large_sheet_threshold
    print("GUI_SELF_TEST_SHEET_CACHE_REUSE_OK")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"GUI_SELF_TEST_SHEET_CACHE_REUSE_FAIL: {exc}", file=sys.stderr)
        raise

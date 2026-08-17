"""Recoverable multi-branch SVN submission workbench for Excel configs.

The workflow deliberately keeps every repository-changing action behind a
visible TortoiseSVN dialog.  This module owns discovery, semantic projection,
write-ahead recovery, and post-dialog reconciliation; it never treats a
TortoiseProc exit code as proof that a commit happened.
"""

from __future__ import annotations

import copy
import hashlib
import json
import os
import queue
import re
import shutil
import sqlite3
import subprocess
import sys
import tempfile
import threading
import time
import uuid
from collections.abc import Callable, Iterable
from dataclasses import asdict, dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import ClassVar

from .fast_branch_merge import analyze_source as fast_analyze_source
from .fast_branch_merge import analyze_target as fast_analyze_target
from .fast_branch_merge import apply_source_change_plan
from .svn_status_provider import (
    SvnStatusRecord,
    records_by_path,
    scan_status,
)
from .ui_foundation import THEME, UiTaskRunner, configure_ttk_style

SUPPORTED_EXTENSIONS = (".xlsx",)
DEFAULT_BRANCHES = ("develop", "release", "sandbox")
HIDDEN_BRANCH_NAMES = {".svn", "tool", "tools", "sow_merge_tool", "excel_merge_tool"}
TERMINAL_ACTION_STATES = {"committed", "already_applied", "excluded", "restored"}
BLOCKING_NODE_STATES = {"conflicted", "obstructed", "replaced", "incomplete", "status-callback-failed"}
SOURCE_CHANGE_STATES = {"modified", "added", "deleted", "missing", "unversioned"}
SOURCE_ONLY_MISSING = "source_only_missing"
STATE_VERSION = 6

SVN_OPERATION_POLICIES = (
    ("modified", "默认勾选", "同步源分支的单元格修改", "保留目标分支其他内容"),
    ("added", "默认勾选", "目标路径不存在时新增", "目标已有不同文件则阻断"),
    ("unversioned", "默认不勾选", "用户勾选后按新增处理", "目标已有文件则阻断"),
    ("deleted", "默认勾选", "同步明确的 SVN 删除", "目标有独立内容时需人工确认"),
    ("missing", "默认勾选", "仅交给源分支 TortoiseSVN", "不更新、不写入任何目标分支"),
    ("conflicted / obstructed / replaced", "不可勾选", "安全阻断", "先在 TortoiseSVN 中修复状态"),
    ("switched / external / 属性修改", "不可勾选", "安全阻断", "避免跨工作副本或属性误提交"),
)


def _multi_branch_policy_text(node_status: str, reason: str = "") -> str:
    if reason:
        return "不进入批次"
    return {
        "modified": "同步修改到目标",
        "added": "在目标新增文件",
        "unversioned": "勾选后在目标新增",
        "deleted": "同步删除到目标",
        "missing": "仅源分支；目标不变",
    }.get(node_status, "不进入批次")


def _sha256(path: str) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _safe_copy(src: str, dst: str) -> None:
    os.makedirs(os.path.dirname(os.path.abspath(dst)), exist_ok=True)
    tmp = dst + f".tmp-{os.getpid()}-{uuid.uuid4().hex}"
    shutil.copy2(src, tmp)
    os.replace(tmp, dst)


def _safe_json_write(path: str, payload: dict) -> None:
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    tmp = path + f".tmp-{os.getpid()}-{uuid.uuid4().hex}"
    with open(tmp, "w", encoding="utf-8") as stream:
        json.dump(payload, stream, ensure_ascii=False, indent=2)
    os.replace(tmp, path)


def _artifact_path(root: str, category: str, relative_path: str) -> str:
    return os.path.join(root, category, *relative_path.replace("\\", "/").split("/"))


def _is_within(path: str, parent: str) -> bool:
    try:
        return os.path.normcase(os.path.commonpath((os.path.abspath(path), os.path.abspath(parent)))) == os.path.normcase(os.path.abspath(parent))
    except ValueError:
        return False


def settings_dir() -> str:
    root = os.environ.get("LOCALAPPDATA") or tempfile.gettempdir()
    return os.path.join(root, "SowMergeTool", "branch_submit")


def load_settings() -> dict:
    try:
        with open(os.path.join(settings_dir(), "settings.json"), "r", encoding="utf-8") as stream:
            value = json.load(stream)
        return value if isinstance(value, dict) else {}
    except (OSError, ValueError):
        return {}


def save_settings(data: dict) -> None:
    _safe_json_write(os.path.join(settings_dir(), "settings.json"), data)


@dataclass
class _WcNode:
    local_relpath: str
    repos_id: int | None
    repos_path: str
    revision: int | None
    presence: str
    kind: str
    moved_here: bool = False
    moved_to: str = ""
    changed_revision: int | None = None
    repo_root: str = ""
    repo_uuid: str = ""


def _wc_db_path(wc_root: str) -> str:
    return os.path.join(os.path.abspath(wc_root), ".svn", "wc.db")


def _node_for_path(wc_root: str, path: str) -> _WcNode | None:
    rel = os.path.relpath(os.path.abspath(path), os.path.abspath(wc_root)).replace("\\", "/")
    if rel == ".":
        rel = ""
    try:
        with sqlite3.connect(f"file:{_wc_db_path(wc_root)}?mode=ro", uri=True) as conn:
            row = conn.execute(
                """
                select n.local_relpath, n.repos_id, coalesce(n.repos_path,''), n.revision,
                       n.presence, n.kind, coalesce(n.moved_here,0), coalesce(n.moved_to,''),
                       n.changed_revision, coalesce(r.root,''), coalesce(r.uuid,'')
                from NODES n left join REPOSITORY r on r.id=n.repos_id
                where n.local_relpath=?
                order by n.op_depth desc limit 1
                """,
                (rel,),
            ).fetchone()
    except sqlite3.Error as exc:
        raise RuntimeError(f"无法读取 SVN 工作副本数据库：{exc}") from exc
    if not row:
        return None
    return _WcNode(
        local_relpath=str(row[0] or ""), repos_id=row[1], repos_path=str(row[2] or ""),
        revision=int(row[3]) if row[3] is not None else None,
        presence=str(row[4] or ""), kind=str(row[5] or ""), moved_here=bool(row[6]),
        moved_to=str(row[7] or ""), changed_revision=int(row[8]) if row[8] is not None else None,
        repo_root=str(row[9] or "").rstrip("/"), repo_uuid=str(row[10] or ""),
    )


def repository_metadata(wc_root: str) -> tuple[str, str]:
    try:
        with sqlite3.connect(f"file:{_wc_db_path(wc_root)}?mode=ro", uri=True) as conn:
            row = conn.execute("select root, uuid from REPOSITORY order by id limit 1").fetchone()
    except sqlite3.Error as exc:
        raise RuntimeError(f"无法读取 SVN 仓库信息：{exc}") from exc
    if not row or not row[0] or not row[1]:
        raise RuntimeError("SVN 工作副本缺少仓库 URL 或 UUID")
    return str(row[0]).rstrip("/"), str(row[1])


@dataclass
class BranchCandidate:
    name: str
    path: str
    url: str
    repo_root: str
    repo_uuid: str
    enabled: bool = True
    reason: str = ""
    favorite: bool = False
    last_changed_at: float = 0.0


def _load_workbook(*args, **kwargs):
    """Load openpyxl only when semantic preflight actually needs it."""

    from openpyxl import load_workbook

    return load_workbook(*args, **kwargs)


def _branch_last_changed_times(wc_root: str, repo_root: str, repo_uuid: str) -> dict[str, float]:
    """Return each branch's newest SVN changed_date as a Unix timestamp."""
    try:
        with sqlite3.connect(f"file:{_wc_db_path(wc_root)}?mode=ro", uri=True) as conn:
            rows = conn.execute(
                """
                select case
                         when instr(n.local_relpath, '/')=0 then n.local_relpath
                         else substr(n.local_relpath, 1, instr(n.local_relpath, '/')-1)
                       end as branch_name,
                       max(coalesce(n.changed_date, 0))
                from NODES n join REPOSITORY r on r.id=n.repos_id
                where n.local_relpath<>'' and n.op_depth=0 and n.presence='normal'
                  and coalesce(n.file_external,0)=0 and r.root=? and r.uuid=?
                group by branch_name
                """,
                (repo_root, repo_uuid),
            ).fetchall()
    except sqlite3.Error:
        return {}
    return {
        str(name): (float(changed_date) / 1_000_000.0)
        for name, changed_date in rows if name and changed_date
    }


def discover_branch_candidates(wc_root: str, *, favorites: Iterable[str] = ()) -> list[BranchCandidate]:
    root = os.path.abspath(wc_root)
    if not os.path.isfile(_wc_db_path(root)):
        raise ValueError(f"不是 SVN 工作副本根目录：{root}")
    favorite_set = {str(item) for item in favorites}
    expected_root, expected_uuid = repository_metadata(root)
    changed_times = _branch_last_changed_times(root, expected_root, expected_uuid)
    candidates: list[BranchCandidate] = []
    for entry in os.scandir(root):
        if not entry.is_dir(follow_symlinks=False) or entry.name.lower() in HIDDEN_BRANCH_NAMES or entry.is_symlink():
            continue
        node = _node_for_path(root, entry.path)
        if (
            not node or node.kind != "dir" or node.presence != "normal"
            or node.repo_root != expected_root or node.repo_uuid != expected_uuid
        ):
            continue
        url = node.repo_root + ("/" + node.repos_path.strip("/") if node.repos_path else "")
        candidates.append(BranchCandidate(
            name=entry.name, path=os.path.abspath(entry.path), url=url,
            repo_root=node.repo_root, repo_uuid=node.repo_uuid,
            favorite=entry.name in favorite_set,
            last_changed_at=changed_times.get(entry.name, entry.stat(follow_symlinks=False).st_mtime),
        ))
    candidates.sort(key=lambda item: (-item.last_changed_at, item.name.lower()))
    return candidates


def discover_branches(wc_root: str, allowed: Iterable[str] | None = None) -> list[str]:
    """Compatibility API returning dynamic, versioned branch directory names."""
    try:
        names = [item.name for item in discover_branch_candidates(wc_root)]
    except (RuntimeError, ValueError, sqlite3.Error):
        # Tiny legacy unit fixtures contain only ``.svn`` without a wc.db.
        names = sorted(
            entry.name for entry in os.scandir(wc_root)
            if entry.is_dir() and entry.name in DEFAULT_BRANCHES
        )
    if allowed is not None:
        allowed_set = {str(item).strip() for item in allowed if str(item).strip()}
        names = [name for name in names if name in allowed_set]
    return sorted(dict.fromkeys(names), key=lambda item: (item != "develop", item.lower()))


def _validate_branch_name(branch: str, branches: Iterable[str]) -> str:
    value = str(branch or "").strip()
    if value not in set(branches):
        raise ValueError(f"分支不在已验证候选中：{value or '<empty>'}")
    return value


def _validate_relative_file(relative_path: str) -> str:
    value = str(relative_path or "").replace("\\", "/").strip("/")
    if not value or value.startswith("../") or "/../" in f"/{value}/":
        raise ValueError(f"非法配置相对路径：{relative_path!r}")
    if Path(value).suffix.lower() not in SUPPORTED_EXTENSIONS:
        raise ValueError(f"只支持 .xlsx：{relative_path}")
    return value


@dataclass
class BranchContext:
    wc_root: str
    source_branch: str
    scope_path: str
    initial_paths: list[str] = field(default_factory=list)
    candidates: list[BranchCandidate] | None = None


def _find_wc_root(path: str) -> str:
    probe = os.path.abspath(path if os.path.isdir(path) else os.path.dirname(path))
    while probe:
        if os.path.isfile(os.path.join(probe, ".svn", "wc.db")):
            return probe
        parent = os.path.dirname(probe)
        if parent == probe:
            break
        probe = parent
    raise ValueError(f"路径不在 SVN 工作副本中：{path}")


def infer_context(initial_paths: Iterable[str]) -> BranchContext:
    paths = [os.path.abspath(str(path)) for path in initial_paths if str(path).strip()]
    if not paths:
        raise ValueError("没有可用的右键路径")
    first = paths[0]
    if not os.path.exists(first):
        raise ValueError(f"右键路径不存在：{first}")
    wc_root = _find_wc_root(first)
    source = ""
    for path in paths:
        if not os.path.exists(path):
            raise ValueError(f"右键路径不存在：{path}")
        current_root = _find_wc_root(path)
        rel = os.path.relpath(path, current_root).replace("\\", "/")
        parts = rel.split("/")
        if not parts or parts[0] in ("", "."):
            raise ValueError("请在具体分支目录内启动多分支提交")
        current_source = parts[0]
        if not source:
            source = current_source
        if os.path.normcase(current_root) != os.path.normcase(wc_root) or current_source != source:
            raise ValueError("右键路径必须位于同一 SVN 工作副本和同一源分支")
    candidates = discover_branch_candidates(wc_root)
    _validate_branch_name(source, [item.name for item in candidates])
    scope = first if os.path.isdir(first) else os.path.dirname(first)
    if any(not _is_within(path, scope) for path in paths):
        scope = os.path.commonpath(paths)
        if os.path.isfile(scope):
            scope = os.path.dirname(scope)
    return BranchContext(
        wc_root=wc_root,
        source_branch=source,
        scope_path=os.path.abspath(scope),
        initial_paths=paths,
        candidates=candidates,
    )


def infer_context_from_files(initial_paths: Iterable[str]) -> tuple[str, str, list[str]]:
    context = infer_context(initial_paths)
    files = [path for path in context.initial_paths if os.path.isfile(path)]
    return context.wc_root, context.source_branch, files


@dataclass
class SvnChangeItem:
    path: str
    relative_path: str
    extension: str
    node_kind: str
    node_status: str
    text_status: str
    prop_status: str
    versioned: bool
    conflicted: bool = False
    switched: bool = False
    file_external: bool = False
    wc_locked: bool = False
    lock_owner: str = ""
    changelist: str = ""
    moved_from: str = ""
    moved_to: str = ""
    revision: int | None = None
    checked: bool = False
    selectable: bool = False
    reason: str = ""


def _record_reason(record: SvnStatusRecord, extension: str) -> str:
    if record.conflicted or record.node_status == "conflicted" or record.prop_status == "conflicted":
        return "存在 SVN 冲突"
    if record.node_status in BLOCKING_NODE_STATES:
        return f"不支持的 SVN 状态：{record.node_status}"
    if record.wc_locked:
        return "SVN 工作副本被锁定，请先执行 TortoiseSVN Cleanup"
    if record.switched:
        return "路径已 switched，禁止自动跨分支提交"
    if record.file_external or record.node_status == "external":
        return "svn:externals 不进入批次"
    if record.prop_status not in {"none", "normal"} and record.node_status != "added":
        return f"存在属性修改：{record.prop_status}"
    if extension not in SUPPORTED_EXTENSIONS:
        return "仅显示，不支持提交此文件类型"
    if os.path.basename(record.path).startswith("~$"):
        return "Excel 临时锁文件不可提交"
    if record.node_kind != "file":
        return "仅普通文件可进入批次；目录、符号链接和未知节点只显示"
    if record.node_status not in SOURCE_CHANGE_STATES:
        return f"不是可提交的 Excel 变更：{record.node_status}"
    return ""


def scan_changes(wc_root: str, source_branch: str, scope_path: str, *, cancel_event: threading.Event | None = None) -> list[SvnChangeItem]:
    branch_root = os.path.abspath(os.path.join(wc_root, source_branch))
    scope = os.path.abspath(scope_path)
    if not _is_within(scope, branch_root):
        raise ValueError("扫描范围必须位于源分支内")
    records = scan_status(scope, cancel_event=cancel_event)
    items: list[SvnChangeItem] = []
    for record in records:
        if not _is_within(record.path, scope) or not _is_within(record.path, branch_root):
            continue
        if record.node_status in {"ignored", "external"} or record.file_external:
            continue
        relative = os.path.relpath(record.path, branch_root).replace("\\", "/")
        extension = Path(record.path).suffix.lower()
        reason = _record_reason(record, extension)
        selectable = not reason and extension in SUPPORTED_EXTENSIONS and record.node_kind != "dir"
        checked = selectable and record.versioned and record.node_status in {"modified", "added", "deleted", "missing"}
        if record.changelist == "ignore-on-commit":
            checked = False
        items.append(SvnChangeItem(
            path=os.path.abspath(record.path), relative_path=relative, extension=extension,
            node_kind=record.node_kind, node_status=record.node_status,
            text_status=record.text_status, prop_status=record.prop_status,
            versioned=record.versioned, conflicted=record.conflicted,
            switched=record.switched, file_external=record.file_external,
            wc_locked=record.wc_locked, lock_owner=record.lock_owner,
            changelist=record.changelist, moved_from=record.moved_from,
            moved_to=record.moved_to, revision=record.revision,
            checked=checked, selectable=selectable, reason=reason,
        ))
    items.sort(key=lambda item: (item.node_kind == "dir", item.relative_path.lower()))
    return items


def read_recent_messages(repo_uuid: str, limit: int = 25) -> list[str]:
    if os.name != "nt" or not repo_uuid:
        return []
    import winreg
    keys = [
        rf"Software\TortoiseSVN\History\commit{repo_uuid}",
        r"Software\TortoiseSVN\History\commit",
    ]
    messages: list[tuple[int, str]] = []
    for key_name in keys:
        try:
            with winreg.OpenKey(winreg.HKEY_CURRENT_USER, key_name) as key:
                index = 0
                while True:
                    try:
                        name, value, _kind = winreg.EnumValue(key, index)
                    except OSError:
                        break
                    index += 1
                    match = re.fullmatch(r"logmsgs(\d+)", str(name), re.IGNORECASE)
                    text = str(value or "").strip()
                    if match and text:
                        messages.append((int(match.group(1)), text))
        except OSError:
            continue
        if messages:
            break
    seen: set[str] = set()
    result: list[str] = []
    for _index, message in sorted(messages, key=lambda pair: pair[0]):
        if message not in seen:
            seen.add(message)
            result.append(message)
        if len(result) >= max(1, int(limit)):
            break
    return result


def _json_atom(value):
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return repr(value)


def _workbook_semantic_digest(path: str) -> str:
    """Hash configuration semantics while ignoring OOXML timestamps/caches."""
    wb = _load_workbook(path, read_only=False, data_only=False, keep_links=False)
    try:
        digest = hashlib.sha256()
        for ws in wb.worksheets:
            digest.update(json.dumps(["sheet", ws.title, ws.sheet_state], ensure_ascii=False).encode("utf-8"))
            digest.update(json.dumps(sorted(str(item) for item in ws.merged_cells.ranges), ensure_ascii=False).encode("utf-8"))
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    payload = [cell.coordinate, cell.data_type, _json_atom(cell.value), cell.number_format]
                    digest.update(json.dumps(payload, ensure_ascii=False, sort_keys=True).encode("utf-8"))
        return digest.hexdigest()
    finally:
        wb.close()


def _semantic_equal(left: str, right: str) -> bool:
    try:
        return _workbook_semantic_digest(left) == _workbook_semantic_digest(right)
    except Exception as exc:
        raise RuntimeError(f"Excel 语义比较失败：{exc}") from exc


def _wc_revision(core, path: str) -> int | None:
    root = core._find_svn_wc_root_for_path(path)
    if not root:
        return None
    relative = os.path.relpath(os.path.abspath(path), root).replace("\\", "/")
    revision, _author, _reason = core._wc_node_metadata(root, relative)
    return revision


def _has_conflict(core, path: str) -> bool:
    try:
        return bool(core._detect_svn_conflict_files(path)) or bool(core._has_svn_conflict_artifacts(path))
    except Exception as exc:
        raise RuntimeError(f"SVN 冲突探测失败：{path}：{exc}") from exc


@dataclass
class BatchFileAction:
    branch: str
    relative_path: str
    operation: str
    state: str = "planned"
    reason: str = ""
    target_before_hash: str = ""
    candidate_hash: str = ""
    backup_path: str = ""
    candidate_path: str = ""
    preview_path: str = ""
    preview_hash: str = ""
    preview_target_hash: str = ""
    preview_created_at: str = ""
    revision_before: int | None = None
    revision_after: int | None = None
    prepared_at: str = ""
    disposition: str = ""
    reason_code: str = ""
    manual_result: str = ""  # v3 state compatibility; no longer used by the workflow
    confirmed: bool = False
    confirmation_target_hash: str = ""


@dataclass
class FilePlan:
    relative_path: str
    operation: str = "modify"
    source_before: str = ""
    source_after: str = ""
    source_before_hash: str = ""
    source_after_hash: str = ""
    source_revision: int | None = None
    source_state: str = "planned"
    source_committed_revision: int | None = None
    actions: dict[str, BatchFileAction] = field(default_factory=dict)
    target_summaries: dict[str, dict] = field(default_factory=dict)
    target_details: dict[str, list[dict]] = field(default_factory=dict)


@dataclass
class BranchSubmitBatch:
    batch_id: str
    wc_root: str
    source_branch: str
    target_branches: list[str]
    files: list[FilePlan]
    message: str
    scope_path: str = ""
    source_status: str = "pending"
    target_status: dict[str, str] = field(default_factory=dict)
    created_at: str = field(default_factory=lambda: datetime.now().isoformat(timespec="seconds"))
    updated_at: str = ""
    source_revision_after: int | None = None
    error: str = ""
    abandoned: bool = False
    superseded_by: str = ""
    journal: list[dict] = field(default_factory=list)

    @property
    def folder(self) -> str:
        return os.path.join(settings_dir(), "batches", self.batch_id)

    @property
    def state_path(self) -> str:
        return os.path.join(self.folder, "batch.json")

    def event(self, kind: str, **details) -> None:
        self.journal.append({"time": datetime.now().isoformat(timespec="seconds"), "kind": kind, **details})
        self.save()

    def save(self) -> None:
        self.updated_at = datetime.now().isoformat(timespec="seconds")
        payload = asdict(self)
        payload["state_version"] = STATE_VERSION
        _safe_json_write(self.state_path, payload)

    @classmethod
    def load(cls, path: str) -> "BranchSubmitBatch":
        with open(path, "r", encoding="utf-8") as stream:
            payload = json.load(stream)
        version = int(payload.pop("state_version", 1))
        if version not in {2, 3, 4, 5, STATE_VERSION}:
            raise ValueError(f"批次状态版本不兼容：{version}")
        plans: list[FilePlan] = []
        for raw in payload.pop("files", []):
            actions = {name: BatchFileAction(**action) for name, action in raw.pop("actions", {}).items()}
            plans.append(FilePlan(actions=actions, **raw))
        batch = cls(files=plans, **payload)
        # Older batches did not classify source-change actions.  Keep them
        # readable for recovery; a fresh target validation fills new fields.
        if version in {2, 3}:
            for plan in batch.files:
                for action in plan.actions.values():
                    action.disposition = action.disposition or ("already_applied" if action.state == "already_applied" else "")
        return batch


def list_unfinished_batches() -> list[BranchSubmitBatch]:
    root = os.path.join(settings_dir(), "batches")
    if not os.path.isdir(root):
        return []
    result: list[BranchSubmitBatch] = []
    for entry in os.scandir(root):
        state_path = os.path.join(entry.path, "batch.json")
        if not entry.is_dir() or not os.path.isfile(state_path):
            continue
        try:
            batch = BranchSubmitBatch.load(state_path)
        except Exception:
            continue
        if batch_requires_recovery(batch):
            result.append(batch)
    return sorted(result, key=lambda item: item.updated_at, reverse=True)


_READ_ONLY_BATCH_EVENTS = {
    "preflight",
    "preflight-failed",
    "target-update-start",
    "target-update-complete",
    "target-update-skipped",
    "source-change-confirmed",
    "target-file-excluded",
    "target-preview-created",
    "abandoned",
}


def batch_requires_recovery(batch: BranchSubmitBatch) -> bool:
    """Return whether a batch may have repository or working-copy effects.

    A preflight may run SVN Update on the selected target paths, then creates
    candidates and previews only inside the batch folder.  An ordinary SVN
    synchronization has no tool-owned candidate to roll back, so those records
    must not interrupt the next launch.  Recovery is reserved for a batch that
    entered the commit flow, wrote a target working copy, reached a
    non-preflight source state, or contains an effect-bearing action left by an
    older state version.
    """
    if batch.abandoned or batch.superseded_by:
        return False
    complete = batch.source_status == "committed" and all(
        state in {"committed", "already_present", "skipped"}
        for state in batch.target_status.values()
    )
    if complete:
        return False
    if batch.source_status not in {"pending", "ready"}:
        return True
    event_kinds = {
        str(entry.get("kind", "") or "")
        for entry in batch.journal
        if isinstance(entry, dict)
    }
    if any(kind and kind not in _READ_ONLY_BATCH_EVENTS for kind in event_kinds):
        return True
    effect_states = {"prepared", "committed", "partial", "unknown", "restored"}
    return any(
        action.state in effect_states
        for plan in batch.files
        for action in plan.actions.values()
    )


def list_corrupt_batch_files() -> list[str]:
    root = os.path.join(settings_dir(), "batches")
    if not os.path.isdir(root):
        return []
    corrupt: list[str] = []
    for entry in os.scandir(root):
        state_path = os.path.join(entry.path, "batch.json")
        if not entry.is_dir() or not os.path.isfile(state_path):
            continue
        try:
            BranchSubmitBatch.load(state_path)
        except Exception:
            corrupt.append(state_path)
    return sorted(corrupt)


def _status_for_exact_path(wc_root: str, path: str, status_map: dict[str, SvnStatusRecord]) -> SvnStatusRecord:
    record = status_map.get(os.path.normcase(os.path.abspath(path)))
    if record:
        return record
    node = _node_for_path(wc_root, path)
    exists = os.path.exists(path)
    if node and node.presence == "normal":
        return SvnStatusRecord(
            path=os.path.abspath(path), node_kind=node.kind, node_status="normal" if exists else "missing",
            text_status="normal" if exists else "none", prop_status="normal",
            versioned=True, revision=node.revision, repos_root_url=node.repo_root,
            repos_uuid=node.repo_uuid, repos_relpath=node.repos_path,
            moved_to=node.moved_to,
        )
    return SvnStatusRecord(
        path=os.path.abspath(path), node_kind="file" if exists else "none",
        node_status="unversioned" if exists else "none", versioned=False,
    )


def _status_block_reason(record: SvnStatusRecord, *, require_clean: bool) -> str:
    if record.conflicted or record.node_status == "conflicted" or record.prop_status == "conflicted":
        return "存在 SVN 冲突"
    if record.node_status in BLOCKING_NODE_STATES:
        return f"不支持的 SVN 状态：{record.node_status}"
    if record.wc_locked:
        return "SVN 工作副本被锁定，请先执行 TortoiseSVN Cleanup"
    if record.switched:
        return "路径已 switched"
    if record.file_external or record.node_status == "external":
        return "路径属于 svn:externals"
    if record.prop_status not in {"none", "normal"} and record.node_status != "added":
        return f"存在属性修改：{record.prop_status}"
    if require_clean and record.node_status not in {"normal", "none"}:
        return f"目标工作副本不干净：{record.node_status}"
    return ""


class BranchSubmitEngine:
    def __init__(
        self,
        wc_root: str,
        *,
        allowed_branches: Iterable[str] | None = None,
        runner: Callable | None = None,
        status_scanner: Callable[[str], list[SvnStatusRecord]] | None = None,
        candidates: Iterable[BranchCandidate] | None = None,
    ):
        self.wc_root = os.path.abspath(wc_root)
        self.allowed_branches = tuple(allowed_branches) if allowed_branches is not None else None
        self.core = None
        self.runner = runner or self._default_runner
        self.status_scanner = status_scanner or scan_status
        self.candidates = list(candidates) if candidates is not None else None
        self._fast_delta_cache: dict[tuple[str, str], object] = {}

    def _load_core(self):
        if self.core is None:
            import sow_merge_tool as core
            self.core = core
        return self.core

    @staticmethod
    def _default_runner(args, *, timeout=300):
        return subprocess.run(args, capture_output=True, text=True, errors="replace", timeout=timeout, check=False)

    def _tortoise(self, command: str, paths: list[str], *, message: str | None = None) -> int:
        core = self._load_core()
        exe = core._find_tortoise_proc_exe()
        if not exe or (os.path.dirname(exe) and not os.path.isfile(exe)):
            raise RuntimeError("未找到 TortoiseProc.exe")
        temp_paths: list[str] = []
        args = [exe, f"/command:{command}"]
        if len(paths) == 1:
            args.append(f"/path:{paths[0]}")
        else:
            pathfile = os.path.join(tempfile.gettempdir(), f"sow_pathfile_{uuid.uuid4().hex}.txt")
            with open(pathfile, "w", encoding="utf-16-le", newline="") as stream:
                stream.write("\n".join(paths))
            temp_paths.append(pathfile)
            args.extend((f"/pathfile:{pathfile}", "/deletepathfile"))
        if message is not None:
            logmsg = os.path.join(tempfile.gettempdir(), f"sow_logmsg_{uuid.uuid4().hex}.txt")
            with open(logmsg, "w", encoding="utf-8-sig", newline="") as stream:
                stream.write(message)
            temp_paths.append(logmsg)
            args.append(f"/logmsgfile:{logmsg}")
        args.append("/closeonend:1")
        try:
            result = self.runner(args, timeout=3600 if command == "commit" else 900)
            return int(getattr(result, "returncode", 1))
        finally:
            for path in temp_paths:
                try:
                    os.remove(path)
                except OSError:
                    pass

    def show_log(self, path: str) -> int:
        return self._tortoise("log", [path])

    def show_diff(self, path: str) -> int:
        return self._tortoise("diff", [path])

    def open_excel_comparison(
        self,
        batch: BranchSubmitBatch,
        plan: FilePlan,
        target: str,
    ) -> None:
        """Compare target-before with a target-derived after preview.

        The source workbook is never used as one side of this comparison.  A
        preview answers the product question directly: which cells would be
        changed in this target branch if the source delta were accepted?
        """
        target_path = os.path.join(batch.wc_root, target, *plan.relative_path.split("/"))
        preview_path = self.ensure_target_preview(batch, plan, target)
        if not os.path.isfile(target_path) or not os.path.isfile(preview_path):
            raise RuntimeError(f"{target}/{plan.relative_path}：缺少目标修改前文件或修改后预览")
        command = [sys.executable]
        if not getattr(sys, "frozen", False):
            command.extend(("-m", "sow_merge_tool"))
        command.extend((target_path, preview_path))
        subprocess.Popen(command)

    def ensure_target_preview(
        self,
        batch: BranchSubmitBatch,
        plan: FilePlan,
        target: str,
    ) -> str:
        """Materialize an artifact-only target-after preview on demand."""
        if plan.operation != "modify":
            raise RuntimeError("只有修改文件支持查看目标修改点")
        action = plan.actions.get(target)
        if action is None:
            raise RuntimeError(f"预检查结果中不存在 {target}/{plan.relative_path}")
        if action.state in {"blocked", "excluded", "unknown", "failed"}:
            raise RuntimeError(action.reason or "该项未通过安全检查，不能生成修改预览")
        target_path = os.path.join(batch.wc_root, target, *plan.relative_path.split("/"))
        if not os.path.isfile(target_path):
            raise RuntimeError(f"{target}/{plan.relative_path}：目标文件不存在")
        status_map = records_by_path(self.status_scanner(os.path.join(batch.wc_root, target)))
        record = _status_for_exact_path(batch.wc_root, target_path, status_map)
        reason = _status_block_reason(record, require_clean=True)
        if reason or not record.versioned or record.node_status != "normal":
            raise RuntimeError(f"{target}/{plan.relative_path}：{reason or '目标文件不是干净的已版本化文件'}")
        if _has_conflict(self._load_core(), target_path):
            raise RuntimeError(f"{target}/{plan.relative_path}：目标文件存在 SVN 冲突或冲突残留")
        current_hash = _sha256(target_path)
        if action.target_before_hash and current_hash != action.target_before_hash:
            raise RuntimeError(f"{target}/{plan.relative_path}：预检查后目标内容已变化，请重新预检查")
        if (
            action.preview_path
            and action.preview_target_hash == current_hash
            and os.path.isfile(action.preview_path)
            and action.preview_hash
            and _sha256(action.preview_path) == action.preview_hash
        ):
            return action.preview_path
        cache_key = (plan.source_before, plan.source_after)
        delta = self._fast_delta_cache.get(cache_key)
        if delta is None:
            delta = fast_analyze_source(plan.source_before, plan.source_after)
            self._fast_delta_cache[cache_key] = delta
        decision = fast_analyze_target(delta, target_path)
        plan.target_summaries[target] = dict(decision.summary)
        plan.target_details[target] = list(decision.details)
        if decision.disposition == "already_applied":
            raise RuntimeError("目标已经包含全部源修改，没有新的目标修改点")
        if decision.disposition == "unsupported":
            raise RuntimeError(decision.reason or "无法安全生成目标修改预览")
        preview = _artifact_path(batch.folder, os.path.join("previews", target), plan.relative_path)
        apply_source_change_plan(
            plan.source_before,
            plan.source_after,
            target_path,
            preview,
            decision,
            # The preview is hypothetical and is specifically used to decide
            # whether an overlapping source change should be accepted.
            confirmed=decision.disposition == "confirmation_required",
        )
        action.preview_path = preview
        action.preview_hash = _sha256(preview)
        action.preview_target_hash = current_hash
        action.preview_created_at = datetime.now().isoformat(timespec="seconds")
        batch.event(
            "target-preview-created",
            target=target,
            path=plan.relative_path,
            target_hash=current_hash,
            preview_hash=action.preview_hash,
        )
        return preview

    def _update(self, paths: list[str]) -> None:
        if paths and self._tortoise("update", sorted(dict.fromkeys(paths))) != 0:
            raise RuntimeError("SVN update 未成功或被取消")

    @staticmethod
    def _target_update_paths(batch: BranchSubmitBatch, target: str) -> list[str]:
        """Return the smallest safe SVN Update scope for a target branch.

        Existing files can be updated directly.  A source-side add has no
        target path yet, so its versioned parent must be updated to learn
        whether the repository now contains a colliding file.
        """
        result: list[str] = []
        seen: set[str] = set()
        for plan in batch.files:
            if plan.operation == SOURCE_ONLY_MISSING:
                continue
            target_path = os.path.join(
                batch.wc_root, target, *plan.relative_path.split("/")
            )
            update_path = os.path.dirname(target_path) if plan.operation == "add" else target_path
            key = os.path.normcase(os.path.abspath(update_path))
            if key not in seen:
                seen.add(key)
                result.append(os.path.abspath(update_path))
        return sorted(result, key=os.path.normcase)

    @staticmethod
    def _verify_target_update_scope_clean(
        target: str,
        update_paths: list[str],
        status_map: dict[str, SvnStatusRecord],
    ) -> None:
        """Prevent SVN Update from merging into pre-existing local changes."""
        for record in status_map.values():
            record_path = os.path.abspath(record.path)
            affected = any(
                os.path.normcase(record_path) == os.path.normcase(path)
                or (os.path.isdir(path) and _is_within(record_path, path))
                or (
                    (record.node_kind == "dir" or os.path.isdir(record_path))
                    and _is_within(path, record_path)
                )
                for path in update_paths
            )
            if not affected:
                continue
            reason = _status_block_reason(record, require_clean=True)
            if reason:
                raise RuntimeError(
                    f"目标分支 {target} 更新前工作副本检查失败："
                    f"{record_path}（{reason}）"
                )

    @staticmethod
    def _refresh_target_status(batch: BranchSubmitBatch, target: str) -> str:
        actions = [plan.actions[target] for plan in batch.files if target in plan.actions]
        states = {action.state for action in actions}
        if states & {"blocked", "unknown", "failed"}:
            state = "blocked"
        elif "confirmation_required" in states:
            state = "confirmation_required"
        else:
            active = [
                action for action in actions
                if action.state not in {"already_applied", "excluded", "restored"}
            ]
            if active:
                state = "ready"
            elif any(action.state == "excluded" for action in actions):
                state = "skipped"
            else:
                state = "already_present"
        batch.target_status[target] = state
        return state

    def confirm_source_changes(
        self,
        batch: BranchSubmitBatch,
        target: str,
        relative_path: str,
    ) -> BatchFileAction:
        """Accept source results for one target/file conflict set.

        Confirmation is only a decision.  It never writes the target working
        copy and it never creates a whole-file merge result.
        """
        plan = next((item for item in batch.files if item.relative_path == relative_path), None)
        if plan is None or target not in plan.actions:
            raise RuntimeError(f"预检查结果中不存在 {target}/{relative_path}")
        action = plan.actions[target]
        if action.state != "confirmation_required":
            if action.confirmed and action.state == "ready":
                return action
            raise RuntimeError(f"{target}/{relative_path} 当前不需要人工确认")
        target_path = os.path.join(batch.wc_root, target, *relative_path.split("/"))
        status_map = records_by_path(self.status_scanner(os.path.join(batch.wc_root, target)))
        record = _status_for_exact_path(batch.wc_root, target_path, status_map)
        reason = _status_block_reason(record, require_clean=True)
        if reason:
            raise RuntimeError(f"{target}/{relative_path}：{reason}")
        if not os.path.isfile(target_path) or not record.versioned or record.node_status != "normal":
            raise RuntimeError(f"{target}/{relative_path}：目标文件不是干净的已版本化文件")
        if _has_conflict(self._load_core(), target_path):
            raise RuntimeError(f"{target}/{relative_path}：目标文件存在 SVN 冲突或冲突残留")
        current_hash = _sha256(target_path)
        if action.target_before_hash and current_hash != action.target_before_hash:
            raise RuntimeError(f"{target}/{relative_path}：预检查后目标内容已变化，请重新预检查")
        action.confirmed = True
        action.confirmation_target_hash = current_hash
        action.disposition = "confirmed"
        action.state = "ready"
        action.reason = "已确认采用源分支修改；目标分支其他内容保持不变"
        self._refresh_target_status(batch, target)
        batch.event("source-change-confirmed", target=target, path=relative_path)
        return action

    def exclude_target_file(
        self,
        batch: BranchSubmitBatch,
        target: str,
        relative_path: str,
    ) -> BatchFileAction:
        """Explicitly remove one conflicted target/file pair from the batch."""
        plan = next((item for item in batch.files if item.relative_path == relative_path), None)
        if plan is None or target not in plan.actions:
            raise RuntimeError(f"预检查结果中不存在 {target}/{relative_path}")
        action = plan.actions[target]
        if action.state != "confirmation_required":
            raise RuntimeError(f"{target}/{relative_path} 当前不能从确认列表移除")
        action.state = "excluded"
        action.disposition = "excluded"
        action.reason = "用户明确从该目标分支的同步批次中移除"
        self._refresh_target_status(batch, target)
        batch.event("target-file-excluded", target=target, path=relative_path)
        return action

    def _source_snapshot(self, batch: BranchSubmitBatch, item: SvnChangeItem) -> FilePlan:
        core = self._load_core()
        relative = _validate_relative_file(item.relative_path)
        source_path = os.path.join(batch.wc_root, batch.source_branch, *relative.split("/"))
        if item.reason or not item.selectable:
            raise RuntimeError(f"{relative}：{item.reason or '不可提交'}")
        if item.moved_from or item.moved_to:
            raise RuntimeError(f"{relative}：检测到移动/重命名，请先在 TortoiseSVN 中 Repair Move")
        if item.node_status == "modified":
            operation = "modify"
        elif item.node_status in {"added", "unversioned"}:
            operation = "add"
        elif item.node_status == "deleted":
            operation = "delete"
        elif item.node_status == "missing":
            operation = SOURCE_ONLY_MISSING
        else:
            raise RuntimeError(f"{relative}：不支持的源状态 {item.node_status}")
        plan = FilePlan(relative_path=relative, operation=operation, source_revision=item.revision)
        if operation in {"modify", "delete", SOURCE_ONLY_MISSING}:
            before = core._try_export_svn_base_from_working_copy(source_path)
            if not before:
                raise RuntimeError(
                    f"{relative}：无法读取源文件 SVN pristine；"
                    "请先执行 TortoiseSVN Cleanup，并检查工作副本基线是否完整"
                )
            plan.source_before = _artifact_path(batch.folder, "source-before", relative)
            _safe_copy(before, plan.source_before)
            plan.source_before_hash = _sha256(plan.source_before)
        if operation in {"modify", "add"}:
            if not os.path.isfile(source_path):
                raise RuntimeError(f"{relative}：源文件不存在")
            plan.source_after = _artifact_path(batch.folder, "source-after", relative)
            _safe_copy(source_path, plan.source_after)
            plan.source_after_hash = _sha256(plan.source_after)
        return plan

    @staticmethod
    def _detect_rename_pairs(plans: list[FilePlan]) -> None:
        added = [plan for plan in plans if plan.operation == "add"]
        deleted = [
            plan for plan in plans
            if plan.operation in {"delete", SOURCE_ONLY_MISSING}
        ]
        for old in deleted:
            for new in added:
                if _semantic_equal(old.source_before, new.source_after):
                    raise RuntimeError(
                        f"检测到高置信度重命名：{old.relative_path} → {new.relative_path}；"
                        "请使用 TortoiseSVN Repair Move 保留历史"
                    )

    def _preflight_target_action(
        self,
        batch: BranchSubmitBatch,
        plan: FilePlan,
        target: str,
        status_map: dict[str, SvnStatusRecord],
    ) -> BatchFileAction:
        core = self._load_core()
        target_path = os.path.join(batch.wc_root, target, *plan.relative_path.split("/"))
        record = _status_for_exact_path(batch.wc_root, target_path, status_map)
        action = BatchFileAction(branch=target, relative_path=plan.relative_path, operation=plan.operation)
        if plan.operation == SOURCE_ONLY_MISSING:
            action.state = "excluded"
            action.disposition = "source_only"
            action.reason = "源文件为 missing：仅由源分支 TortoiseSVN 处理，目标分支保持不变"
            return action
        reason = _status_block_reason(record, require_clean=plan.operation != "add" or record.versioned)
        if reason:
            action.state, action.reason = "blocked", reason
            return action
        if _has_conflict(core, target_path):
            action.state, action.reason = "blocked", "目标文件存在 SVN 冲突或冲突残留"
            return action
        action.revision_before = record.revision
        if plan.operation == "add":
            if os.path.isfile(target_path):
                if record.versioned and record.node_status == "normal" and _semantic_equal(plan.source_after, target_path):
                    action.state = "already_applied"
                    action.disposition = "already_applied"
                    action.reason = "目标已经存在相同的新文件"
                    return action
                action.disposition = "unsupported"
                action.state, action.reason = (
                    "blocked",
                    "源分支新增了整个文件，但目标路径已有不同文件；没有共同基线，禁止整文件覆盖",
                )
                return action
            parent = os.path.dirname(target_path)
            parent_node = _node_for_path(batch.wc_root, parent)
            parent_record = _status_for_exact_path(batch.wc_root, parent, status_map)
            parent_reason = _status_block_reason(parent_record, require_clean=True)
            if not parent_node or parent_node.kind != "dir" or parent_node.presence != "normal" or parent_reason:
                action.state, action.reason = "blocked", parent_reason or "目标父目录未版本化"
                return action
            action.state = "ready"
            action.disposition = "direct"
            action.reason = "目标路径不存在，可直接新增文件"
            return action
        if not os.path.isfile(target_path):
            if plan.operation == "delete" and not record.versioned:
                action.state = "already_applied"
            else:
                action.state, action.reason = "blocked", "目标文件不存在"
            return action
        if not record.versioned or record.node_status != "normal":
            action.state, action.reason = "blocked", f"目标文件不干净：{record.node_status}"
            return action
        action.target_before_hash = _sha256(target_path)
        if plan.operation == "delete":
            if not _semantic_equal(plan.source_before, target_path):
                action.state = "confirmation_required"
                action.disposition = "confirmation_required"
                action.reason = "源分支删除了文件，但目标文件已有独立内容；需确认是否删除目标文件"
                plan.target_summaries[target] = {"confirmation": 1}
                plan.target_details[target] = [{
                    "kind": "confirmation",
                    "apply_kind": "delete_file",
                    "sheet": "（整个文件）",
                    "key": plan.relative_path,
                    "field": "文件",
                    "before": plan.source_before_hash[:12],
                    "source": "删除",
                    "target": action.target_before_hash[:12],
                    "reason": action.reason,
                }]
            else:
                action.state = "ready"
                action.disposition = "direct"
                action.reason = "目标文件未独立变化，可直接同步删除"
            return action
        # Semantic analysis is deliberately a fast, read-only classification.
        # It must not turn an Excel structure that TortoiseSVN can commit into
        # a source-branch submission blocker.  Candidate materialization is
        # deferred until the target is actually processed.
        try:
            cache_key = (plan.source_before, plan.source_after)
            delta = getattr(self, "_fast_delta_cache", {}).get(cache_key)
            if delta is None:
                delta = fast_analyze_source(plan.source_before, plan.source_after)
                self._fast_delta_cache[cache_key] = delta
            decision = fast_analyze_target(delta, target_path)
        except Exception as exc:
            action.disposition = "unsupported"
            action.state, action.reason = "blocked", f"源修改分析失败，已安全阻断：{exc}"
            action.reason_code = "fast-analysis-error"
            plan.target_summaries[target] = {"blocked": 1, "reason": str(exc)}
            return action
        action.disposition = decision.disposition
        action.reason_code = decision.disposition
        plan.target_summaries[target] = dict(decision.summary)
        plan.target_details[target] = list(decision.details)
        if decision.disposition == "already_applied":
            action.state = "already_applied"
            action.reason = decision.reason
        elif decision.disposition == "confirmation_required":
            action.state, action.reason = "confirmation_required", decision.reason
        elif decision.disposition == "unsupported":
            action.state, action.reason = "blocked", decision.reason
        else:
            action.state = "ready"
            action.reason = decision.reason
        return action

    def preflight(
        self,
        source_branch: str,
        target_branches: Iterable[str],
        selected_files: Iterable[SvnChangeItem | str],
        message: str,
        *,
        scope_path: str | None = None,
    ) -> BranchSubmitBatch:
        candidates = list(self.candidates or discover_branch_candidates(self.wc_root))
        branch_names = [item.name for item in candidates]
        if self.allowed_branches is not None:
            branch_names = [name for name in branch_names if name in self.allowed_branches]
        source = _validate_branch_name(source_branch, branch_names)
        targets: list[str] = []
        for raw in target_branches:
            target = _validate_branch_name(raw, branch_names)
            if target == source:
                raise ValueError("目标分支不能与源分支相同")
            if target not in targets:
                targets.append(target)
        if not targets:
            raise ValueError("至少选择一个目标分支")
        frozen_message = str(message or "").strip()
        raw_selected = list(selected_files)
        if not raw_selected:
            raise ValueError("至少选择一个 Excel 变更")
        if all(isinstance(item, SvnChangeItem) for item in raw_selected):
            items = [item for item in raw_selected if isinstance(item, SvnChangeItem)]
        else:
            paths = [os.path.abspath(str(item)) for item in raw_selected]
            scan_scope = scope_path or os.path.commonpath([path if os.path.isdir(path) else os.path.dirname(path) for path in paths])
            all_items = scan_changes(self.wc_root, source, scan_scope)
            selected_norm = {os.path.normcase(path) for path in paths}
            items = [item for item in all_items if os.path.normcase(item.path) in selected_norm]
        batch = BranchSubmitBatch(
            batch_id=datetime.now().strftime("%Y%m%d-%H%M%S") + "-" + uuid.uuid4().hex[:8],
            wc_root=self.wc_root, source_branch=source, target_branches=targets,
            files=[], message=frozen_message,
            scope_path=os.path.abspath(scope_path or os.path.join(self.wc_root, source)),
            target_status={target: "pending" for target in targets},
        )
        try:
            for item in items:
                batch.files.append(self._source_snapshot(batch, item))
            self._detect_rename_pairs(batch.files)
            target_maps: dict[str, dict[str, SvnStatusRecord]] = {}
            for target in targets:
                target_scope = os.path.join(self.wc_root, target)
                update_paths = self._target_update_paths(batch, target)
                if not update_paths:
                    target_maps[target] = {}
                    batch.event(
                        "target-update-skipped",
                        target=target,
                        reason="所选文件仅需源分支处理",
                    )
                    continue
                before_update = records_by_path(self.status_scanner(target_scope))
                self._verify_target_update_scope_clean(target, update_paths, before_update)
                batch.event("target-update-start", target=target, paths=len(update_paths))
                try:
                    self._update(update_paths)
                except Exception as exc:
                    raise RuntimeError(f"目标分支 {target} 更新失败或被取消：{exc}") from exc
                batch.event("target-update-complete", target=target, paths=len(update_paths))
                target_maps[target] = records_by_path(self.status_scanner(target_scope))
            blocked = False
            for plan in batch.files:
                for target in targets:
                    action = self._preflight_target_action(batch, plan, target, target_maps[target])
                    plan.actions[target] = action
                    if action.state == "blocked":
                        blocked = True
                        batch.target_status[target] = "blocked"
            for target in targets:
                self._refresh_target_status(batch, target)
            if blocked:
                batch.error = "预检查未通过：任一不兼容项都会阻止整批提交"
            else:
                batch.source_status = "ready"
            batch.event("preflight", source_status=batch.source_status, target_status=dict(batch.target_status))
            return batch
        except Exception as exc:
            batch.error = str(exc)
            batch.event("preflight-failed", error=batch.error)
            raise

    def _source_status_map(self, batch: BranchSubmitBatch) -> dict[str, SvnStatusRecord]:
        scope = batch.scope_path if _is_within(batch.scope_path, os.path.join(batch.wc_root, batch.source_branch)) else os.path.join(batch.wc_root, batch.source_branch)
        return records_by_path(self.status_scanner(scope))

    def _verify_source_before_commit(self, batch: BranchSubmitBatch) -> None:
        status_map = self._source_status_map(batch)
        for plan in batch.files:
            path = os.path.join(batch.wc_root, batch.source_branch, *plan.relative_path.split("/"))
            record = _status_for_exact_path(batch.wc_root, path, status_map)
            reason = _status_block_reason(record, require_clean=False)
            if reason:
                raise RuntimeError(f"{plan.relative_path}：{reason}")
            if plan.operation == "modify":
                valid = os.path.isfile(path) and _sha256(path) == plan.source_after_hash and record.node_status == "modified"
            elif plan.operation == "add":
                valid = os.path.isfile(path) and _sha256(path) == plan.source_after_hash and record.node_status in {"added", "unversioned"}
            elif plan.operation == SOURCE_ONLY_MISSING:
                valid = not os.path.exists(path) and record.node_status == "missing"
            else:
                valid = not os.path.exists(path) and record.node_status == "deleted"
            if not valid:
                raise RuntimeError(f"源文件已偏离预检查结果：{plan.relative_path}")

    def _reconcile_source(self, batch: BranchSubmitBatch) -> tuple[int, int, int]:
        status_map = self._source_status_map(batch)
        committed = pending = unknown = 0
        revisions: list[int] = []
        core = self._load_core()
        for plan in batch.files:
            path = os.path.join(batch.wc_root, batch.source_branch, *plan.relative_path.split("/"))
            record = _status_for_exact_path(batch.wc_root, path, status_map)
            node = _node_for_path(batch.wc_root, path)
            if plan.operation in {"modify", "add"}:
                pristine = core._try_export_svn_base_from_working_copy(path) if os.path.isfile(path) else None
                if (
                    os.path.isfile(path) and record.node_status == "normal" and record.versioned
                    and pristine and _sha256(path) == plan.source_after_hash and _sha256(pristine) == plan.source_after_hash
                ):
                    plan.source_state = "committed"
                elif os.path.isfile(path) and _sha256(path) == plan.source_after_hash and record.node_status in {"modified", "added", "unversioned"}:
                    plan.source_state = "prepared"
                else:
                    plan.source_state = "unknown"
            else:
                if not os.path.exists(path) and (not node or node.presence != "normal") and record.node_status not in {"deleted", "missing"}:
                    plan.source_state = "committed"
                elif not os.path.exists(path) and record.node_status in {"deleted", "missing"}:
                    plan.source_state = "prepared"
                else:
                    plan.source_state = "unknown"
            if plan.source_state == "committed":
                committed += 1
                revision = node.changed_revision if node else None
                plan.source_committed_revision = revision
                if revision:
                    revisions.append(revision)
            elif plan.source_state == "prepared":
                pending += 1
            else:
                unknown += 1
        batch.source_revision_after = max(revisions, default=None)
        return committed, pending, unknown

    def _create_committed_sub_batch(self, batch: BranchSubmitBatch) -> BranchSubmitBatch:
        """Split a source partial commit into one explicitly resumable child."""
        if batch.superseded_by:
            existing = os.path.join(settings_dir(), "batches", batch.superseded_by, "batch.json")
            return BranchSubmitBatch.load(existing)
        committed_plans = [copy.deepcopy(plan) for plan in batch.files if plan.source_state == "committed"]
        if not committed_plans:
            raise RuntimeError("源分支没有可拆分的已提交文件")
        child_id = batch.batch_id + "-committed"
        child_status = {}
        for target in batch.target_branches:
            states = {plan.actions[target].state for plan in committed_plans}
            if states <= {"excluded"}:
                child_status[target] = "skipped"
            elif states <= {"already_applied", "excluded"}:
                child_status[target] = "already_present"
            else:
                child_status[target] = "ready"
        child = BranchSubmitBatch(
            batch_id=child_id, wc_root=batch.wc_root, source_branch=batch.source_branch,
            target_branches=list(batch.target_branches), files=committed_plans,
            message=batch.message, scope_path=batch.scope_path, source_status="committed",
            target_status=child_status, source_revision_after=batch.source_revision_after,
        )
        child.event("created-from-partial-source", parent_batch=batch.batch_id)
        batch.superseded_by = child.batch_id
        batch.event("partial-source-split", child_batch=child.batch_id, committed_files=len(committed_plans))
        return child

    @staticmethod
    def _has_prepare_intent(batch: BranchSubmitBatch, target: str, relative_path: str) -> bool:
        return any(
            event.get("kind") == "prepare-intent"
            and event.get("target") == target
            and event.get("path") == relative_path
            for event in batch.journal
        )

    def _audit_incomplete_intents(self, batch: BranchSubmitBatch) -> None:
        """Recover the crash window between a filesystem write and state save."""
        changed = False
        for plan in batch.files:
            for target, action in plan.actions.items():
                if action.state != "planned" or not self._has_prepare_intent(batch, target, plan.relative_path):
                    continue
                path = os.path.join(batch.wc_root, target, *plan.relative_path.split("/"))
                if plan.operation in {"modify", "add"} and os.path.isfile(path) and action.candidate_hash and _sha256(path) == action.candidate_hash:
                    action.state = "prepared"; changed = True
                elif plan.operation == "delete" and not os.path.exists(path) and action.backup_path:
                    action.state = "prepared"; changed = True
                elif action.backup_path and os.path.isfile(path) and _sha256(path) == action.target_before_hash:
                    # The intent was durable but the working-copy write never happened.
                    continue
                elif plan.operation == "add" and not os.path.exists(path):
                    continue
                else:
                    action.state, action.reason = "unknown", "写前日志存在，但工作副本无法匹配候选或备份"
                    changed = True
        if changed:
            batch.event("prepare-intents-audited")

    def _fresh_target_action(self, batch: BranchSubmitBatch, plan: FilePlan, target: str, status_map: dict[str, SvnStatusRecord]) -> BatchFileAction:
        action = plan.actions[target]
        target_path = os.path.join(batch.wc_root, target, *plan.relative_path.split("/"))
        record = _status_for_exact_path(batch.wc_root, target_path, status_map)
        if action.state in TERMINAL_ACTION_STATES:
            return action
        # A candidate left by this same batch is the sole non-clean state that
        # can be resumed without an update/reprojection.
        if action.state == "prepared":
            if plan.operation == "delete" and not os.path.exists(target_path):
                return action
            if os.path.isfile(target_path) and action.candidate_hash and _sha256(target_path) == action.candidate_hash:
                return action
            action.state, action.reason = "unknown", "工作副本不再等于本批次候选"
            return action
        was_confirmed = bool(action.confirmed)
        confirmation_hash = action.confirmation_target_hash
        reason = _status_block_reason(record, require_clean=plan.operation != "add" or record.versioned)
        if reason:
            action.state, action.reason = "blocked", reason
            return action
        action.revision_before = record.revision
        if plan.operation == "add":
            if os.path.isfile(target_path):
                if record.versioned and record.node_status == "normal" and _semantic_equal(plan.source_after, target_path):
                    action.state, action.disposition = "already_applied", "already_applied"
                    action.reason = "目标已经存在相同的新文件"
                else:
                    action.state, action.disposition = "blocked", "unsupported"
                    action.reason = "目标路径已有不同文件；没有共同基线，禁止整文件覆盖"
                return action
            candidate = _artifact_path(batch.folder, os.path.join("candidates", target), plan.relative_path)
            _safe_copy(plan.source_after, candidate)
            action.candidate_path, action.candidate_hash = candidate, _sha256(candidate)
            action.state, action.disposition = "ready", "direct"
            action.reason = "目标路径不存在，可直接新增文件"
            return action
        if plan.operation == "delete":
            if not os.path.exists(target_path) and not record.versioned:
                action.state, action.disposition = "already_applied", "already_applied"
                action.reason = "目标文件已经不存在"
            elif not os.path.isfile(target_path) or record.node_status != "normal":
                action.state, action.reason = "blocked", "更新后目标文件状态不满足删除条件"
            else:
                current_hash = _sha256(target_path)
                action.target_before_hash = current_hash
                if _semantic_equal(plan.source_before, target_path):
                    action.state, action.disposition = "ready", "direct"
                    action.reason = "目标文件未独立变化，可直接同步删除"
                elif was_confirmed and confirmation_hash == current_hash:
                    action.state, action.disposition = "ready", "confirmed"
                    action.reason = "已确认删除目标分支的独立版本"
                else:
                    action.confirmed = False
                    action.confirmation_target_hash = ""
                    action.state, action.disposition = "confirmation_required", "confirmation_required"
                    action.reason = "目标文件已有独立内容；需确认是否同步源分支的删除"
            return action
        if not os.path.isfile(target_path) or record.node_status != "normal":
            action.state, action.reason = "blocked", f"更新后目标状态不是 normal：{record.node_status}"
            return action
        current_hash = _sha256(target_path)
        action.target_before_hash = current_hash
        try:
            cache_key = (plan.source_before, plan.source_after)
            delta = self._fast_delta_cache.get(cache_key)
            if delta is None:
                delta = fast_analyze_source(plan.source_before, plan.source_after)
                self._fast_delta_cache[cache_key] = delta
            decision = fast_analyze_target(delta, target_path)
            plan.target_summaries[target] = dict(decision.summary)
            plan.target_details[target] = list(decision.details)
            action.disposition = decision.disposition
            action.reason_code = decision.disposition
            if decision.disposition == "already_applied":
                action.state, action.reason = "already_applied", decision.reason
            elif decision.disposition == "unsupported":
                action.state, action.reason = "blocked", decision.reason
                action.confirmed = False
                action.confirmation_target_hash = ""
            elif decision.disposition == "confirmation_required" and not (
                was_confirmed and confirmation_hash == current_hash
            ):
                action.state, action.reason = "confirmation_required", decision.reason
                action.confirmed = False
                action.confirmation_target_hash = ""
            else:
                candidate_copy = _artifact_path(batch.folder, os.path.join("candidates", target), plan.relative_path)
                if (
                    action.preview_path
                    and action.preview_target_hash == current_hash
                    and action.preview_hash
                    and os.path.isfile(action.preview_path)
                    and _sha256(action.preview_path) == action.preview_hash
                ):
                    _safe_copy(action.preview_path, candidate_copy)
                else:
                    apply_source_change_plan(
                        plan.source_before,
                        plan.source_after,
                        target_path,
                        candidate_copy,
                        decision,
                        confirmed=decision.disposition == "confirmation_required",
                    )
                action.candidate_path, action.candidate_hash = candidate_copy, _sha256(candidate_copy)
                action.state = "ready"
                if decision.disposition == "confirmation_required":
                    action.disposition = "confirmed"
                    action.reason = "已确认采用源分支修改；候选仅修改冲突位置"
                else:
                    action.reason = decision.reason
        except Exception as exc:
            action.state, action.reason = "blocked", f"无法安全生成目标补丁候选：{exc}"
            action.disposition = "unsupported"
            action.reason_code = "materialize-error"
        return action

    def _prepare_target_action(self, batch: BranchSubmitBatch, plan: FilePlan, target: str) -> None:
        action = plan.actions[target]
        if action.state in TERMINAL_ACTION_STATES or action.state == "prepared":
            return
        path = os.path.join(batch.wc_root, target, *plan.relative_path.split("/"))
        backup = _artifact_path(batch.folder, os.path.join("backups", target), plan.relative_path)
        if os.path.isfile(path):
            _safe_copy(path, backup)
            action.backup_path = backup
            action.target_before_hash = _sha256(backup)
        action.state = "planned"
        batch.event("prepare-intent", target=target, path=plan.relative_path, operation=plan.operation)
        if plan.operation == "delete":
            os.remove(path)
        else:
            if (
                not action.candidate_path
                or not os.path.isfile(action.candidate_path)
                or _sha256(action.candidate_path) != action.candidate_hash
            ):
                candidate_kind = "从目标文件生成的补丁" if plan.operation == "modify" else "新增文件"
                raise RuntimeError(
                    f"{target}/{plan.relative_path}：缺少有效的{candidate_kind}候选；禁止回退为整文件覆盖"
                )
            source = action.candidate_path
            _safe_copy(source, path)
            action.candidate_hash = _sha256(path)
        action.prepared_at = datetime.now().isoformat(timespec="seconds")
        action.state = "prepared"
        batch.event("prepared", target=target, path=plan.relative_path, operation=plan.operation)

    def _reconcile_target(self, batch: BranchSubmitBatch, target: str) -> tuple[int, int, int]:
        status_map = records_by_path(self.status_scanner(os.path.join(batch.wc_root, target)))
        committed = pending = unknown = 0
        core = self._load_core()
        for plan in batch.files:
            action = plan.actions[target]
            if action.state != "prepared":
                continue
            path = os.path.join(batch.wc_root, target, *plan.relative_path.split("/"))
            record = _status_for_exact_path(batch.wc_root, path, status_map)
            node = _node_for_path(batch.wc_root, path)
            if plan.operation in {"modify", "add"}:
                pristine = core._try_export_svn_base_from_working_copy(path) if os.path.isfile(path) and record.versioned else None
                if (
                    os.path.isfile(path) and record.node_status == "normal" and record.versioned
                    and action.candidate_hash and _sha256(path) == action.candidate_hash
                    and pristine and _sha256(pristine) == action.candidate_hash
                ):
                    action.state = "committed"
                elif os.path.isfile(path) and action.candidate_hash and _sha256(path) == action.candidate_hash and record.node_status in {"modified", "added", "unversioned"}:
                    action.state = "prepared"
                else:
                    action.state, action.reason = "unknown", "提交后内容或 SVN 状态无法与候选对账"
            else:
                if not os.path.exists(path) and (not node or node.presence != "normal") and record.node_status not in {"deleted", "missing"}:
                    action.state = "committed"
                elif not os.path.exists(path) and record.node_status in {"deleted", "missing"}:
                    action.state = "prepared"
                else:
                    action.state, action.reason = "unknown", "删除提交后状态无法确认"
            if action.state == "committed":
                committed += 1
                action.revision_after = node.changed_revision if node else batch.source_revision_after
            elif action.state == "prepared":
                pending += 1
            else:
                unknown += 1
        batch.event("target-reconciled", target=target, committed=committed, pending=pending, unknown=unknown)
        return committed, pending, unknown

    def commit(self, batch: BranchSubmitBatch, *, stop_on_failure: bool = True) -> BranchSubmitBatch:
        if batch.source_status not in {"ready", "committed"}:
            raise RuntimeError(f"批次不可提交：source_status={batch.source_status}")
        if not str(batch.message or "").strip():
            raise RuntimeError("开始提交前必须填写 SVN 提交说明")
        pending_confirmations = [
            f"{target}/{plan.relative_path}"
            for plan in batch.files
            for target, action in plan.actions.items()
            if action.state == "confirmation_required"
        ]
        if pending_confirmations:
            raise RuntimeError(
                f"仍有 {len(pending_confirmations)} 个目标文件需要人工确认，不能开始提交"
            )
        source_paths = [os.path.join(batch.wc_root, batch.source_branch, *plan.relative_path.split("/")) for plan in batch.files]
        if batch.source_status == "ready":
            try:
                exit_code = None
                # Reconcile first: the previous process may have died after
                # the server committed but before batch.json was updated.
                committed, pending, unknown = self._reconcile_source(batch)
                if committed == len(batch.files):
                    batch.source_status = "committed"
                    batch.event("source-recovered-after-crash", committed=committed)
                elif committed:
                    batch.source_status = "partial"
                    child = self._create_committed_sub_batch(batch)
                    batch.error = f"源分支只提交了部分文件；已生成可继续的子批次 {child.batch_id}"
                    batch.save()
                    return batch
                elif unknown:
                    batch.source_status = "unknown"
                    batch.error = "源分支提交结果无法确认；已停止传播"
                    batch.save()
                    return batch
                else:
                    self._verify_source_before_commit(batch)
                    batch.event("source-commit-open", paths=len(source_paths))
                    exit_code = self._tortoise("commit", source_paths, message=batch.message)
                    committed, pending, unknown = self._reconcile_source(batch)
                if committed == len(batch.files):
                    batch.source_status = "committed"
                elif committed:
                    batch.source_status = "partial"
                    child = self._create_committed_sub_batch(batch)
                    batch.error = f"源分支只提交了部分文件；已生成可继续的子批次 {child.batch_id}"
                elif unknown:
                    batch.source_status = "unknown"
                    batch.error = "源分支提交结果无法确认；已停止传播"
                else:
                    batch.source_status = "cancelled" if exit_code == 1 else "failed"
                    batch.error = f"源分支没有文件完成提交（TortoiseSVN 退出码 {exit_code}）"
                batch.event("source-reconciled", committed=committed, pending=pending, unknown=unknown, exit_code=exit_code)
                if batch.source_status != "committed":
                    return batch
            except Exception as exc:
                batch.source_status, batch.error = "unknown", str(exc)
                batch.event("source-error", error=batch.error)
                return batch
        footer = f"[MultiBranchSync] batch={batch.batch_id} source={batch.source_branch}@r{batch.source_revision_after or 'unknown'}"
        target_message = batch.message.rstrip() + "\n\n" + footer
        self._audit_incomplete_intents(batch)
        for target in batch.target_branches:
            if batch.target_status.get(target) in {"committed", "already_present", "skipped"}:
                continue
            try:
                actions = [plan.actions[target] for plan in batch.files]
                nonterminal = [action for action in actions if action.state not in TERMINAL_ACTION_STATES]
                if nonterminal and all(action.state == "prepared" for action in nonterminal):
                    committed, pending, unknown = self._reconcile_target(batch, target)
                    if unknown:
                        batch.target_status[target] = "unknown"
                        batch.error = f"目标分支 {target} 的中断现场无法确认"
                        break
                    if pending == 0:
                        batch.target_status[target] = "committed"
                        batch.event("target-recovered-after-crash", target=target, committed=committed)
                        continue
                resume_ready = any(action.state == "prepared" for action in actions)
                if not resume_ready:
                    self._update(self._target_update_paths(batch, target))
                status_map = records_by_path(self.status_scanner(os.path.join(batch.wc_root, target)))
                for plan in batch.files:
                    action = self._fresh_target_action(batch, plan, target, status_map)
                    if action.state == "confirmation_required":
                        batch.target_status[target] = "confirmation_required"
                        batch.error = (
                            f"目标分支 {target} 更新后出现新的内容重叠；"
                            "请确认采用源修改后继续当前批次"
                        )
                        batch.event(
                            "target-confirmation-required",
                            target=target,
                            path=plan.relative_path,
                            reason=action.reason,
                        )
                        return batch
                    if action.state in {"blocked", "unknown"}:
                        batch.target_status[target] = action.state
                        batch.error = f"{target}/{plan.relative_path}：{action.reason}"
                        batch.event("target-validation-stopped", target=target, error=batch.error)
                        return batch
                for plan in batch.files:
                    self._prepare_target_action(batch, plan, target)
                commit_paths = [
                    os.path.join(batch.wc_root, target, *plan.relative_path.split("/"))
                    for plan in batch.files if plan.actions[target].state == "prepared"
                ]
                if not commit_paths:
                    batch.target_status[target] = "already_present"
                    batch.event("target-already-present", target=target)
                    continue
                batch.event("target-commit-open", target=target, paths=len(commit_paths))
                exit_code = self._tortoise("commit", commit_paths, message=target_message)
                committed, pending, unknown = self._reconcile_target(batch, target)
                active = len([
                    action for action in actions
                    if action.state not in {"already_applied", "excluded", "restored"}
                ])
                if committed == active and unknown == 0:
                    batch.target_status[target] = "committed"
                elif committed:
                    batch.target_status[target] = "partial"
                    batch.error = f"目标分支 {target} 只提交了部分文件，已停止后续分支"
                elif unknown:
                    batch.target_status[target] = "unknown"
                    batch.error = f"目标分支 {target} 提交结果无法确认"
                else:
                    batch.target_status[target] = "cancelled" if exit_code == 1 else "failed"
                    batch.error = f"目标分支 {target} 没有文件完成提交（退出码 {exit_code}）"
                batch.save()
                if batch.target_status[target] != "committed" and stop_on_failure:
                    break
            except Exception as exc:
                batch.target_status[target] = "unknown"
                batch.error = str(exc)
                batch.event("target-error", target=target, error=batch.error)
                if stop_on_failure:
                    break
        return batch

    def restore_uncommitted(self, batch: BranchSubmitBatch) -> BranchSubmitBatch:
        """Restore only target files still proven to be this batch's candidate."""
        self._audit_incomplete_intents(batch)
        # A client may have stopped after the server accepted the commit but
        # before batch.json was updated.  Reconcile first so recovery never
        # creates a reverse local change for content already in the repository.
        for target in batch.target_branches:
            if any(
                plan.actions.get(target) is not None
                and plan.actions[target].state == "prepared"
                for plan in batch.files
            ):
                self._reconcile_target(batch, target)
        for plan in batch.files:
            for target, action in plan.actions.items():
                if action.state != "prepared":
                    continue
                path = os.path.join(batch.wc_root, target, *plan.relative_path.split("/"))
                try:
                    if plan.operation == "modify":
                        if not os.path.isfile(path) or _sha256(path) != action.candidate_hash or not os.path.isfile(action.backup_path):
                            raise RuntimeError("当前文件不再等于候选或备份缺失")
                        _safe_copy(action.backup_path, path)
                    elif plan.operation == "delete":
                        if os.path.exists(path) or not os.path.isfile(action.backup_path):
                            raise RuntimeError("删除现场已变化或备份缺失")
                        status_map = records_by_path(
                            self.status_scanner(os.path.dirname(path))
                        )
                        status = _status_for_exact_path(batch.wc_root, path, status_map)
                        if status.node_status == "deleted":
                            if self._tortoise("revert", [path]) != 0:
                                raise RuntimeError("SVN 删除计划未撤销；请在 TortoiseSVN 中 Revert")
                        elif status.node_status != "missing":
                            raise RuntimeError(
                                f"删除现场 SVN 状态为 {status.node_status}，不能自动恢复"
                            )
                        if not os.path.isfile(path):
                            _safe_copy(action.backup_path, path)
                        restored_status = _status_for_exact_path(
                            batch.wc_root,
                            path,
                            records_by_path(self.status_scanner(os.path.dirname(path))),
                        )
                        if (
                            restored_status.node_status != "normal"
                            or _sha256(path) != action.target_before_hash
                        ):
                            raise RuntimeError("撤销删除后未恢复为原始干净文件")
                    else:
                        if not os.path.isfile(path) or _sha256(path) != action.candidate_hash:
                            raise RuntimeError("新增文件不再等于候选")
                        status = _status_for_exact_path(batch.wc_root, path, records_by_path(self.status_scanner(os.path.dirname(path))))
                        if status.node_status == "added":
                            if self._tortoise("revert", [path]) != 0:
                                raise RuntimeError("SVN Add 未撤销；请在 TortoiseSVN 中 Undo Add")
                            status = _status_for_exact_path(
                                batch.wc_root,
                                path,
                                records_by_path(self.status_scanner(os.path.dirname(path))),
                            )
                        if status.node_status != "unversioned":
                            raise RuntimeError(f"新增文件状态为 {status.node_status}，不能自动移除")
                        os.remove(path)
                    action.state, action.reason = "restored", ""
                    batch.event("restored", target=target, path=plan.relative_path)
                except Exception as exc:
                    action.state, action.reason = "unknown", str(exc)
                    batch.event("restore-failed", target=target, path=plan.relative_path, error=str(exc))
        if all(action.state in TERMINAL_ACTION_STATES for plan in batch.files for action in plan.actions.values()):
            batch.abandoned = True
            batch.event("restore-complete")
        return batch

    @staticmethod
    def abandon(batch: BranchSubmitBatch) -> None:
        batch.abandoned = True
        batch.event("abandoned")


def _format_batch_result(batch: BranchSubmitBatch) -> str:
    targets = ", ".join(f"{name}={state}" for name, state in batch.target_status.items())
    return f"批次 {batch.batch_id}\n源分支：{batch.source_status}\n目标分支：{targets}\n{batch.error}".strip()


def _choose_recovery_action(root, batches: list[BranchSubmitBatch]):
    import tkinter as tk
    from tkinter import ttk
    result = {"value": None}
    win = tk.Toplevel(root)
    win.title("检测到未完成的多分支提交")
    win.geometry("850x330")
    # ``root`` is intentionally withdrawn during startup.  Making a dialog
    # transient to a withdrawn owner can leave the dialog with no visible
    # window on Windows Explorer launches, even though the process is alive.
    # Keep it as an independent, explicitly raised recovery window instead.
    win.resizable(True, True)
    frame = ttk.Frame(win, padding=12)
    frame.pack(fill="both", expand=True)
    ttk.Label(frame, text="以下批次没有完整结束。继续或恢复前会再次核对 SVN 状态。", font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(0, 8))
    tree = ttk.Treeview(frame, columns=("time", "source", "targets", "state"), show="headings", height=8)
    for key, title, width in (("time", "更新时间", 150), ("source", "源分支", 100), ("targets", "目标分支", 280), ("state", "状态", 220)):
        tree.heading(key, text=title); tree.column(key, width=width, anchor="w")
    for batch in batches:
        tree.insert("", "end", iid=batch.batch_id, values=(batch.updated_at, batch.source_branch, ", ".join(batch.target_branches), batch.source_status))
    tree.pack(fill="both", expand=True)
    if batches:
        tree.selection_set(batches[0].batch_id)
    buttons = ttk.Frame(frame); buttons.pack(fill="x", pady=(10, 0))
    def selected():
        values = tree.selection()
        return next((batch for batch in batches if values and batch.batch_id == values[0]), None)
    def finish(action):
        batch = selected()
        if batch:
            result["value"] = (action, batch)
            win.destroy()
    ttk.Button(buttons, text="继续批次", command=lambda: finish("continue")).pack(side="left")
    ttk.Button(buttons, text="恢复未提交工作副本", command=lambda: finish("restore")).pack(side="left", padx=6)
    ttk.Button(buttons, text="放弃批次", command=lambda: finish("abandon")).pack(side="left")
    ttk.Button(buttons, text="查看批次目录", command=lambda: os.startfile(selected().folder) if selected() else None).pack(side="left", padx=6)
    ttk.Button(buttons, text="稍后处理", command=win.destroy).pack(side="right")
    win.protocol("WM_DELETE_WINDOW", win.destroy)
    win.update_idletasks()
    win.deiconify()
    win.lift()
    try:
        win.attributes("-topmost", True)
        win.after(350, lambda: win.attributes("-topmost", False) if win.winfo_exists() else None)
    except tk.TclError:
        pass
    win.focus_force()
    win.grab_set(); root.wait_window(win)
    return result["value"]


class BranchSubmitWorkbench:
    """TortoiseSVN-inspired Tk workbench; no repository writes happen here."""

    STATUS_TEXT: ClassVar[dict[str, str]] = {
        "modified": "已修改（modified）", "added": "新增（added）", "deleted": "已删除（deleted）", "missing": "缺失（missing）",
        "unversioned": "未版本化（unversioned）", "conflicted": "冲突（conflicted）", "normal": "正常（normal）",
    }

    def __init__(self, root, context: BranchContext, *, resume_batch: BranchSubmitBatch | None = None):
        import tkinter as tk
        from tkinter import ttk
        self.tk, self.ttk, self.root = tk, ttk, root
        self.context = context
        self.settings = load_settings()
        favorites = self.settings.get("favorite_branches", list(DEFAULT_BRANCHES))
        favorite_set = set(favorites)
        self.candidates = list(context.candidates or discover_branch_candidates(context.wc_root, favorites=favorites))
        for candidate in self.candidates:
            candidate.favorite = candidate.name in favorite_set
        self.items: list[SvnChangeItem] = []
        self.scan_generation = 0
        self.scan_cancel = threading.Event()
        self.scan_results: queue.Queue = queue.Queue()
        self.scan_polling = False
        self.closing = False
        self.current_batch = resume_batch
        self._loaded_resume_batch_id = resume_batch.batch_id if resume_batch else ""
        self.engine = BranchSubmitEngine(context.wc_root, candidates=self.candidates)
        self.target_vars: dict[str, object] = {}
        self._target_selection: dict[str, bool] = {}
        self._target_rows: dict[str, str] = {}
        self._target_status_map: dict[str, str] = {}
        self._target_rebuild_after = None
        self._item_rows: dict[str, SvnChangeItem] = {}
        self._preflight_generation = 0
        self._selection_generation = 0
        self._approved_preflight_signature: tuple | None = None
        self._scan_active = False
        self._preflight_active = False
        self._confirmation_active = False
        self._confirmation_dialog = None
        self._confirmation_dialog_tree = None
        self._confirmation_dialog_button = None
        self._confirmation_exclude_button = None
        self._confirmation_detail = None
        self._confirmation_dialog_summary_var = None
        self._confirmation_dialog_rows: dict[tuple[str, str], str] = {}
        self._commit_active = False
        self.ui_tasks = UiTaskRunner(root)
        self.source_var = tk.StringVar(value=context.source_branch)
        self.scope_var = tk.StringVar(value=context.scope_path)
        self.repo_url_var = tk.StringVar()
        self.target_search_var = tk.StringVar()
        self.show_unversioned_var = tk.BooleanVar(value=True)
        self.status_var = tk.StringVar(value="正在读取 SVN 状态…")
        self.count_var = tk.StringVar(value="0 个文件")
        self._build_style()
        self._build_ui()
        self._refresh_source_metadata()
        self._rebuild_targets()
        if resume_batch:
            self.message.delete("1.0", tk.END); self.message.insert("1.0", resume_batch.message)
            self._apply_resume_targets(resume_batch)
            self.submit_button.state(["disabled"])
            self.status_var.set(f"已载入未完成批次 {resume_batch.batch_id}；请先重新预检查")
        self._start_scan(preserve_batch=bool(resume_batch))

    def _build_style(self):
        style = configure_ttk_style(self.root)
        style.configure("Title.TLabel", background=THEME.window_bg, foreground=THEME.text, font=(THEME.font_family, 10, "bold"))
        # Keep the three work areas visibly separated on both light and dark
        # Windows themes.  A subtle solid border reads closer to TortoiseSVN
        # than a collection of borderless frames.
        style.configure("Panel.TLabelframe", background=THEME.panel_bg, relief="solid", borderwidth=1)
        style.configure("Panel.TLabelframe.Label", background=THEME.panel_bg, foreground=THEME.text, font=(THEME.font_family, 9, "bold"))
        style.configure("Status.App.TLabel", background=THEME.window_bg, foreground=THEME.secondary_text, padding=(4, 3))
        style.configure("Status.Error.TLabel", background=THEME.window_bg, foreground=THEME.error, padding=(4, 3), font=(THEME.font_family, 9, "bold"))
        style.configure("Status.Footer.TLabel", background=THEME.panel_bg, foreground=THEME.secondary_text, padding=(4, 3))
        style.configure("Status.FooterError.TLabel", background=THEME.panel_bg, foreground=THEME.error, padding=(4, 3), font=(THEME.font_family, 9, "bold"))
        style.configure("Danger.TButton", foreground=THEME.error, padding=(10, 5), font=(THEME.font_family, 9, "bold"))
        style.configure("Workbench.Header.TFrame", background="#EAF3FB", relief="solid", borderwidth=1)
        style.configure("Workbench.HeaderTitle.TLabel", background="#EAF3FB", foreground="#173A5E", font=(THEME.font_family, 14, "bold"))
        style.configure("Workbench.HeaderSub.TLabel", background="#EAF3FB", foreground="#4E6578", font=(THEME.font_family, 9))
        style.configure("Workbench.Step.TLabel", background="#D8EAF8", foreground="#174A7E", padding=(9, 4), font=(THEME.font_family, 9, "bold"))
        style.configure("Workbench.Summary.TLabel", background="#EEF6FC", foreground="#174A7E", padding=(8, 5), font=(THEME.font_family, 9, "bold"))
        style.configure("Workbench.Hint.TLabel", background=THEME.panel_bg, foreground=THEME.secondary_text, font=(THEME.font_family, 9))
        style.configure("Workbench.Bottom.TFrame", background=THEME.panel_bg, relief="solid", borderwidth=1)
        style.configure("Workbench.Treeview", rowheight=27, background=THEME.panel_bg, fieldbackground=THEME.panel_bg)
        style.configure("Workbench.Treeview.Heading", font=(THEME.font_family, 9, "bold"))

    def _build_ui(self):
        tk, ttk = self.tk, self.ttk
        self.root.title("Excel 合并器 · 多分支 SVN 提交")
        self.root.geometry(self.settings.get("window_geometry", "1120x760"))
        self.root.minsize(900, 620)
        outer = ttk.Frame(self.root, padding=12, style="App.TFrame"); outer.pack(fill="both", expand=True)
        # The root owns the window geometry.  Do not let the combined natural
        # height of header, trees and footer enlarge this frame beyond the
        # visible client area and clip bottom actions.
        outer.pack_propagate(False)
        header = ttk.Frame(outer, padding=(14, 10), style="Workbench.Header.TFrame")
        header.pack(fill="x", pady=(0, 10))
        header_copy = ttk.Frame(header, style="Workbench.Header.TFrame")
        header_copy.pack(side="left", fill="x", expand=True)
        ttk.Label(header_copy, text="多分支 SVN 提交", style="Workbench.HeaderTitle.TLabel").pack(anchor="w")
        ttk.Label(
            header_copy,
            text="将源分支已修改内容安全同步到所选目标分支，再逐分支打开 TortoiseSVN 提交",
            style="Workbench.HeaderSub.TLabel",
        ).pack(anchor="w", pady=(2, 0))
        steps = ttk.Frame(header, style="Workbench.Header.TFrame")
        steps.pack(side="right", padx=(16, 0))
        for text in ("1 选择内容", "2 预检查", "3 开始提交"):
            ttk.Label(steps, text=text, style="Workbench.Step.TLabel").pack(side="left", padx=(5, 0))

        info = ttk.LabelFrame(outer, text="提交范围", padding=10, style="Panel.TLabelframe"); info.pack(fill="x")
        ttk.Label(info, text="源分支").grid(row=0, column=0, sticky="w")
        sources = [item.name for item in self.candidates if item.enabled]
        self.source_box = ttk.Combobox(info, textvariable=self.source_var, values=sources, state="readonly", width=22)
        self.source_box.grid(row=0, column=1, sticky="w", padx=(8, 24)); self.source_box.bind("<<ComboboxSelected>>", self._source_changed)
        ttk.Label(info, text="提交到").grid(row=0, column=2, sticky="w")
        ttk.Label(info, textvariable=self.repo_url_var, foreground=THEME.accent).grid(row=0, column=3, sticky="w", padx=(8, 0))
        ttk.Label(info, text="扫描范围").grid(row=1, column=0, sticky="w", pady=(8, 0))
        ttk.Entry(info, textvariable=self.scope_var, state="readonly").grid(row=1, column=1, columnspan=3, sticky="ew", padx=(8, 0), pady=(8, 0))
        info.columnconfigure(3, weight=1)

        paned = ttk.Panedwindow(outer, orient="horizontal"); paned.pack(fill="both", expand=True, pady=(10, 0))
        target_box = ttk.LabelFrame(paned, text="目标分支", padding=10, style="Panel.TLabelframe"); paned.add(target_box, weight=2)
        self.target_summary_var = tk.StringVar(value="尚未预检查 · 完成预检查后才能开始多分支提交")
        ttk.Label(target_box, textvariable=self.target_summary_var, style="Workbench.Summary.TLabel").pack(fill="x", pady=(0, 8))
        search_row = ttk.Frame(target_box); search_row.pack(fill="x")
        ttk.Label(search_row, text="筛选").pack(side="left", padx=(0, 6))
        ttk.Entry(search_row, textvariable=self.target_search_var).pack(side="left", fill="x", expand=True)
        self.target_search_var.trace_add("write", lambda *_: self._schedule_rebuild_targets())
        quick = ttk.Frame(target_box); quick.pack(fill="x", pady=6)
        ttk.Button(quick, text="全选", command=lambda: self._set_targets("all")).pack(side="left")
        ttk.Button(quick, text="全不选", command=lambda: self._set_targets("none")).pack(side="left", padx=4)
        ttk.Button(quick, text="常用", command=lambda: self._set_targets("favorite")).pack(side="left")
        canvas_holder = ttk.Frame(target_box, style="Panel.TFrame"); canvas_holder.pack(fill="both", expand=True, pady=(6, 0))
        target_scroll = ttk.Scrollbar(canvas_holder, orient="vertical")
        self.target_tree = ttk.Treeview(
            canvas_holder,
            columns=("check", "favorite", "branch", "state", "changed"),
            show="headings",
            selectmode="browse",
            height=4,
            yscrollcommand=target_scroll.set,
            style="Workbench.Treeview",
        )
        target_scroll.configure(command=self.target_tree.yview)
        for key, title, width in (("check", "✓", 34), ("favorite", "", 28), ("branch", "分支", 125), ("state", "处理状态", 90), ("changed", "最近修改", 112)):
            self.target_tree.heading(key, text=title)
            self.target_tree.column(key, width=width, anchor="center" if key in {"check", "favorite"} else "w", stretch=key == "branch")
        self.target_tree.tag_configure("ready", foreground=THEME.success)
        self.target_tree.tag_configure("confirmation", foreground=THEME.error, background="#FDE7E9", font=(THEME.font_family, 9, "bold"))
        self.target_tree.tag_configure("blocked", foreground=THEME.error)
        self.target_tree.tag_configure("committed", foreground=THEME.success)
        self.target_tree.pack(side="left", fill="both", expand=True); target_scroll.pack(side="right", fill="y")
        self.target_tree.bind("<Button-1>", self._target_tree_click)
        self.target_tree.bind("<space>", self._target_space)
        self.target_tree.bind("<Double-1>", self._target_tree_double_click)

        main = ttk.Frame(paned, style="App.TFrame"); paned.add(main, weight=5)
        message_box = ttk.LabelFrame(main, text="提交说明", padding=10, style="Panel.TLabelframe"); message_box.pack(fill="x")
        message_tools = ttk.Frame(message_box); message_tools.pack(fill="x", pady=(0, 6))
        ttk.Button(message_tools, text="最近提交消息", command=self._show_recent_messages).pack(side="left")
        ttk.Button(message_tools, text="粘贴文件名", command=self._paste_filenames).pack(side="left", padx=5)
        ttk.Button(message_tools, text="显示日志", command=self._show_log).pack(side="left")
        self.message = tk.Text(message_box, height=4, wrap="word", font=("Segoe UI", 9), undo=True)
        self.message.pack(fill="x")
        self.message.bind("<KeyRelease>", self._message_changed)
        ttk.Label(
            message_box,
            text="可先留空进行预检查；真正开始提交前必须填写。修改说明不会让预检查结果失效。",
            style="Workbench.Hint.TLabel",
        ).pack(anchor="w", pady=(6, 0))

        changes = ttk.LabelFrame(main, text="文件变更（双击查看差异）", padding=10, style="Panel.TLabelframe"); changes.pack(fill="both", expand=True, pady=(10, 0))
        filters = ttk.Frame(changes); filters.pack(fill="x", pady=(0, 6))
        ttk.Label(filters, text="选择：", style="Title.TLabel").pack(side="left")
        for text, mode in (("全部", "all"), ("无", "none"), ("已版本化", "versioned"), ("新增", "added"), ("删除", "deleted"), ("修改", "modified"), ("文件", "files")):
            label = tk.Label(filters, text=text, fg="#0645ad", cursor="hand2", font=("Segoe UI", 9, "underline"))
            label.pack(side="left", padx=(8, 0)); label.bind("<Button-1>", lambda _e, value=mode: self._quick_check(value))
        ttk.Button(filters, text="处理规则…", command=self._show_operation_policies).pack(side="right", padx=(6, 0))
        ttk.Checkbutton(filters, text="显示未版本化文件", variable=self.show_unversioned_var, command=self._render_items).pack(side="right")
        self.scan_stop_button = ttk.Button(filters, text="停止扫描", command=self._stop_scan)
        self.scan_stop_button.pack(side="right", padx=6)
        ttk.Button(filters, text="刷新", command=self._start_scan).pack(side="right")
        columns = ("check", "path", "handling", "extension", "status", "property", "lock", "switched", "changelist")
        self.tree = ttk.Treeview(changes, columns=columns, show="headings", selectmode="extended", height=4, style="Workbench.Treeview")
        self.tree.tag_configure("alternate", background=THEME.row_alt)
        self.tree.tag_configure("not_selectable", foreground=THEME.disabled)
        headings = {"check":"✓", "path":"路径", "handling":"多分支处理", "extension":"扩展名", "status":"状态", "property":"属性状态", "lock":"锁定", "switched":"已切换", "changelist":"变更列表"}
        widths = {"check":38, "path":320, "handling":170, "extension":72, "status":150, "property":105, "lock":80, "switched":70, "changelist":120}
        widths.update({key: int(value) for key, value in self.settings.get("column_widths", {}).items() if key in widths})
        for key in columns:
            self.tree.heading(key, text=headings[key]); self.tree.column(key, width=widths[key], stretch=key == "path", anchor="w" if key != "check" else "center")
        yscroll = ttk.Scrollbar(changes, orient="vertical", command=self.tree.yview)
        xscroll = ttk.Scrollbar(changes, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        self.tree.pack(side="left", fill="both", expand=True); yscroll.pack(side="right", fill="y"); xscroll.pack(side="bottom", fill="x")
        self.tree.bind("<Button-1>", self._tree_click); self.tree.bind("<Double-1>", self._tree_double_click); self.tree.bind("<Button-3>", self._tree_menu)

        # Reserve a real footer before the expandable panes.  The previous
        # packing order allowed the new header and tall trees to push the
        # preflight/submit buttons below the client area on 768px displays.
        self.footer_host = ttk.Frame(outer, style="App.TFrame")
        self.footer_host.pack(side="bottom", fill="x", before=paned)
        self.confirmation_alert = tk.Frame(
            self.footer_host,
            background="#FDE7E9",
            highlightbackground=THEME.error,
            highlightthickness=1,
            padx=10,
            pady=8,
        )
        self.confirmation_alert_var = tk.StringVar(value="")
        tk.Label(
            self.confirmation_alert,
            textvariable=self.confirmation_alert_var,
            background="#FDE7E9",
            foreground=THEME.error,
            font=(THEME.font_family, 9, "bold"),
            anchor="w",
        ).pack(side="left", fill="x", expand=True)
        self.confirmation_button = ttk.Button(
            self.confirmation_alert,
            text="查看并确认…",
            style="Danger.TButton",
            command=self._open_confirmation_dialog,
        )
        self.confirmation_button.pack(side="right", padx=(12, 0))

        bottom = ttk.Frame(self.footer_host, padding=(10, 8), style="Workbench.Bottom.TFrame"); bottom.pack(side="bottom", fill="x", pady=(10, 0))
        self.bottom_bar = bottom
        actions = ttk.Frame(bottom, style="Panel.TFrame")
        actions.pack(side="right", padx=(12, 0))
        ttk.Button(actions, text="取消", command=self._close).pack(side="right")
        self.submit_button = ttk.Button(actions, text="② 开始提交", style="Primary.TButton", command=self._submit); self.submit_button.pack(side="right", padx=6); self.submit_button.state(["disabled"])
        self.preflight_button = ttk.Button(actions, text="① 预检查（必需）", style="App.TButton", command=self._preflight); self.preflight_button.pack(side="right")
        status_area = ttk.Frame(bottom, style="Panel.TFrame")
        status_area.pack(side="left", fill="x", expand=True)
        self.scan_progress = ttk.Progressbar(status_area, mode="indeterminate", length=110)
        self.scan_progress.pack(side="left", padx=(0, 8))
        ttk.Label(status_area, textvariable=self.count_var, background=THEME.panel_bg).pack(side="left")
        self.status_label = ttk.Label(status_area, textvariable=self.status_var, style="Status.Footer.TLabel")
        self.status_label.pack(side="left", fill="x", expand=True, padx=(18, 0))
        self.root.protocol("WM_DELETE_WINDOW", self._close)

    def _candidate_for_source(self):
        return next((item for item in self.candidates if item.name == self.source_var.get()), None)

    def _refresh_source_metadata(self):
        candidate = self._candidate_for_source()
        self.repo_url_var.set(candidate.url if candidate else "")

    def _source_changed(self, _event=None):
        if self._confirmation_active or self._commit_active:
            return
        source = self.source_var.get()
        self.context.source_branch = source
        self.context.scope_path = os.path.join(self.context.wc_root, source)
        self.scope_var.set(self.context.scope_path)
        self._invalidate_batch()
        self._target_selection.clear()
        self._refresh_source_metadata(); self._rebuild_targets(); self._start_scan()

    def _schedule_rebuild_targets(self):
        if self._target_rebuild_after is not None:
            try:
                self.root.after_cancel(self._target_rebuild_after)
            except Exception:
                pass
        self._target_rebuild_after = self.root.after(120, self._rebuild_targets)

    def _rebuild_targets(self):
        tk = self.tk
        self._target_rebuild_after = None
        self._target_selection.update({name: bool(var.get()) for name, var in self.target_vars.items()})
        remembered = set(self.settings.get("last_targets", {}).get(self.source_var.get(), []))
        for iid in self.target_tree.get_children():
            self.target_tree.delete(iid)
        self.target_vars = {}
        self._target_rows = {}
        query = self.target_search_var.get().strip().lower()
        for index, candidate in enumerate(self.candidates):
            if candidate.name == self.source_var.get() or (query and query not in candidate.name.lower()):
                continue
            selected = self._target_selection.get(candidate.name, candidate.name in remembered)
            var = tk.BooleanVar(value=selected and candidate.enabled)
            self.target_vars[candidate.name] = var
            iid = f"branch-{index}"
            self._target_rows[iid] = candidate.name
            changed = time.strftime("%Y-%m-%d %H:%M", time.localtime(candidate.last_changed_at)) if candidate.last_changed_at else "—"
            state = self._target_status_map.get(candidate.name, "待检查")
            tag = {"可直接同步": "ready", "需人工确认": "confirmation", "安全阻断": "blocked", "已提交": "committed"}.get(state, "")
            self.target_tree.insert("", "end", iid=iid, values=("☑" if var.get() else "☐", "★" if candidate.favorite else "☆", candidate.name, state, changed), tags=(tag,))

    def _target_tree_click(self, event):
        if self._confirmation_active or self._commit_active:
            return "break"
        iid = self.target_tree.identify_row(event.y)
        if not iid:
            return
        self.target_tree.selection_set(iid)
        column = self.target_tree.identify_column(event.x)
        name = self._target_rows.get(iid)
        if not name:
            return
        if column == "#2":
            self._toggle_favorite(name)
        elif column in {"#1", "#3"}:
            var = self.target_vars[name]
            var.set(not bool(var.get()))
            self._target_selection[name] = bool(var.get())
            self.target_tree.set(iid, "check", "☑" if var.get() else "☐")
            self._invalidate_batch()
        return "break"

    def _target_space(self, _event=None):
        if self._confirmation_active or self._commit_active:
            return "break"
        selection = self.target_tree.selection()
        if selection:
            name = self._target_rows.get(selection[0])
            if name:
                var = self.target_vars[name]
                var.set(not bool(var.get()))
                self._target_selection[name] = bool(var.get())
                self.target_tree.set(selection[0], "check", "☑" if var.get() else "☐")
                self._invalidate_batch()
        return "break"

    def _target_tree_double_click(self, event):
        if self._confirmation_active or self._commit_active:
            return "break"
        iid = self.target_tree.identify_row(event.y)
        name = self._target_rows.get(iid)
        if name and self._target_status_map.get(name) == "需人工确认":
            self._open_confirmation_dialog(preferred_target=name)
            return "break"

    def _toggle_favorite(self, name: str):
        favorites = set(self.settings.get("favorite_branches", list(DEFAULT_BRANCHES)))
        if name in favorites: favorites.remove(name)
        else: favorites.add(name)
        self.settings["favorite_branches"] = sorted(favorites)
        for candidate in self.candidates:
            candidate.favorite = candidate.name in favorites
        self._rebuild_targets()

    def _set_targets(self, mode: str):
        if self._confirmation_active or self._commit_active:
            return
        for candidate in self.candidates:
            if candidate.name == self.source_var.get():
                continue
            selected = bool(candidate.enabled and (mode == "all" or (mode == "favorite" and candidate.favorite)))
            self._target_selection[candidate.name] = selected
            var = self.target_vars.get(candidate.name)
            if var is not None:
                var.set(selected)
            iid = next((row for row, value in self._target_rows.items() if value == candidate.name), None)
            if iid:
                self.target_tree.set(iid, "check", "☑" if selected else "☐")
        self._invalidate_batch()

    def _apply_resume_targets(self, batch: BranchSubmitBatch):
        for name, var in self.target_vars.items():
            var.set(name in batch.target_branches)
            self._target_selection[name] = bool(var.get())
            iid = next((row for row, value in self._target_rows.items() if value == name), None)
            if iid:
                self.target_tree.set(iid, "check", "☑" if var.get() else "☐")

    def _start_scan(self, *, preserve_batch: bool = False):
        if self._confirmation_active or self._commit_active or self._preflight_active:
            self.status_var.set("当前操作完成前不能刷新 SVN 状态")
            return
        if not preserve_batch:
            self._invalidate_batch("SVN 状态已刷新，请重新预检查", refresh_targets=False)
        self.scan_generation += 1; generation = self.scan_generation
        self.scan_cancel.set(); self.scan_cancel = threading.Event()
        source_branch = self.source_var.get()
        scope_path = self.scope_var.get()
        cancel_event = self.scan_cancel
        self._scan_active = True
        self.status_var.set("正在递归读取 SVN 状态…"); self.preflight_button.state(["disabled"])
        self.submit_button.state(["disabled"])
        self.scan_stop_button.state(["!disabled"]); self.scan_progress.start(12)
        def worker():
            try:
                result = scan_changes(self.context.wc_root, source_branch, scope_path, cancel_event=cancel_event)
                error = None
            except Exception as exc:
                result, error = [], str(exc)
            self.scan_results.put((generation, result, error))
        threading.Thread(target=worker, daemon=True, name="branch-submit-status").start()
        if not self.scan_polling:
            self.scan_polling = True
            self.root.after(30, self._poll_scan_results)

    def _poll_scan_results(self):
        if self.closing:
            self.scan_polling = False
            return
        try:
            while True:
                generation, items, error = self.scan_results.get_nowait()
                self._finish_scan(generation, items, error)
        except queue.Empty:
            pass
        self.root.after(30, self._poll_scan_results)

    def _stop_scan(self):
        self.scan_generation += 1
        self.scan_cancel.set()
        self._scan_active = False
        self.scan_progress.stop(); self.scan_progress.configure(value=0); self.scan_stop_button.state(["disabled"])
        self.status_var.set("已停止等待本次扫描结果；可点击刷新重新扫描")
        self._refresh_primary_button()

    def _finish_scan(self, generation: int, items: list[SvnChangeItem], error: str | None):
        if generation != self.scan_generation:
            return
        self._scan_active = False
        self.scan_progress.stop(); self.scan_progress.configure(value=0); self.scan_stop_button.state(["disabled"])
        if error:
            self.items = []; self._render_items(); self.status_var.set(error); self._refresh_primary_button(); return
        self.items = items; self._render_items()
        self.status_var.set("状态扫描完成；灰色或带原因的项目不会进入批次")
        self._refresh_primary_button()

    def _render_items(self):
        for iid in self.tree.get_children(): self.tree.delete(iid)
        visible = [item for item in self.items if self.show_unversioned_var.get() or item.node_status != "unversioned"]
        self._item_rows = {}
        for index, item in enumerate(visible):
            mark = "☑" if item.checked else "☐" if item.selectable else "—"
            status = self.STATUS_TEXT.get(item.node_status, item.node_status)
            if item.reason: status += f" · {item.reason}"
            iid = f"item-{index}"
            self._item_rows[iid] = item
            tags = []
            if index % 2:
                tags.append("alternate")
            if not item.selectable:
                tags.append("not_selectable")
            self.tree.insert(
                "", "end", iid=iid,
                values=(
                    mark,
                    item.relative_path,
                    _multi_branch_policy_text(item.node_status, item.reason),
                    item.extension,
                    status,
                    item.prop_status,
                    item.lock_owner or ("已锁定" if item.wc_locked else ""),
                    "是" if item.switched else "",
                    item.changelist,
                ),
                tags=tuple(tags),
            )
        selected = sum(item.checked for item in self.items)
        self.count_var.set(f"已选 {selected} 个，显示 {len(visible)} 个，共 {len(self.items)} 个")
        self._refresh_primary_button()

    def _tree_click(self, event):
        if self.tree.identify_region(event.x, event.y) != "cell" or self.tree.identify_column(event.x) != "#1": return
        iid = self.tree.identify_row(event.y)
        if not iid: return
        item = self._item_rows.get(iid)
        if item is None:
            return
        if item.selectable:
            item.checked = not item.checked
            self.tree.set(iid, "check", "☑" if item.checked else "☐")
            self._invalidate_batch()
        return "break"

    def _tree_double_click(self, event):
        iid = self.tree.identify_row(event.y)
        if iid:
            item = self._item_rows.get(iid)
            if item is None:
                return
            if item.node_status == "modified" and os.path.isfile(item.path):
                threading.Thread(target=lambda: self.engine.show_diff(item.path), daemon=True).start()

    def _tree_menu(self, event):
        tk = self.tk; iid = self.tree.identify_row(event.y)
        if not iid: return
        self.tree.selection_set(iid); item = self._item_rows.get(iid)
        if item is None:
            return
        menu = tk.Menu(self.root, tearoff=False)
        menu.add_command(label="查看差异", command=lambda: threading.Thread(target=lambda: self.engine.show_diff(item.path), daemon=True).start(), state="normal" if item.node_status == "modified" else "disabled")
        menu.add_command(label="打开文件", command=lambda: os.startfile(item.path), state="normal" if os.path.isfile(item.path) else "disabled")
        menu.add_command(label="打开所在目录", command=lambda: subprocess.Popen(["explorer.exe", "/select,", item.path] if os.path.exists(item.path) else ["explorer.exe", os.path.dirname(item.path)]))
        menu.add_command(label="复制路径", command=lambda: (self.root.clipboard_clear(), self.root.clipboard_append(item.path)))
        menu.tk_popup(event.x_root, event.y_root)

    def _quick_check(self, mode: str):
        if self._confirmation_active or self._commit_active:
            return
        for item in self.items:
            if not item.selectable: continue
            item.checked = mode == "all" or (mode == "versioned" and item.versioned) or (mode == "files" and item.node_kind != "dir") or item.node_status == mode
            if mode == "none": item.checked = False
        self._invalidate_batch(); self._render_items()

    def _message_changed(self, _event=None):
        # The commit message does not affect branch/file safety analysis.
        # Keep a valid preflight while the user writes or revises the message;
        # the final text is frozen into the batch immediately before commit.
        self._refresh_primary_button()

    def _invalidate_batch(
        self,
        reason: str = "选择已变化，请重新预检查",
        *,
        refresh_targets: bool = True,
    ):
        stale_batch = self.current_batch
        self._selection_generation += 1
        self.current_batch = None
        self._approved_preflight_signature = None
        if self._preflight_generation:
            self.ui_tasks.cancel(self._preflight_generation)
        self._preflight_active = False
        if (
            stale_batch is not None
            and stale_batch.batch_id != self._loaded_resume_batch_id
            and stale_batch.source_status == "ready"
            and not self._commit_active
            and not any(
                action.state in {"prepared", "committed"}
                for plan in stale_batch.files
                for action in plan.actions.values()
            )
        ):
            try:
                self.engine.abandon(stale_batch)
            except Exception:
                pass
        self._target_status_map.clear()
        self._update_confirmation_alert()
        if refresh_targets and hasattr(self, "target_tree"):
            self._render_target_statuses()
        if reason and not self._scan_active:
            self.status_var.set(reason)
        self._refresh_primary_button()

    def _can_preflight(self) -> bool:
        return bool(self.items and self._selected_items() and self._selected_targets())

    def _request_signature(self) -> tuple:
        return (
            self.source_var.get(),
            os.path.normcase(os.path.abspath(self.scope_var.get())),
            tuple(self._selected_targets()),
            tuple(
                sorted(
                    (item.relative_path, item.node_status, item.revision)
                    for item in self._selected_items()
                )
            ),
        )

    def _has_valid_preflight(self) -> bool:
        batch = self.current_batch
        if batch is None:
            return False
        committed_resume = batch.source_status == "committed"
        if not committed_resume:
            if self._approved_preflight_signature is None:
                return False
            if self._approved_preflight_signature != self._request_signature():
                return False
        if not committed_resume and (
            batch.source_branch != self.source_var.get()
            or os.path.normcase(os.path.abspath(batch.scope_path))
            != os.path.normcase(os.path.abspath(self.scope_var.get()))
            or tuple(batch.target_branches) != tuple(self._selected_targets())
            or {plan.relative_path for plan in batch.files}
            != {item.relative_path for item in self._selected_items()}
        ):
            return False
        if batch.source_status not in {"ready", "committed"}:
            return False
        return all(
            action.state not in {"confirmation_required", "blocked", "unknown", "failed", "cancelled"}
            for plan in batch.files
            for action in plan.actions.values()
        )

    def _can_start(self) -> bool:
        """Start requires a valid safety result plus a non-empty commit message."""
        try:
            message_ready = bool(self.message.get("1.0", self.tk.END).strip())
        except Exception:
            message_ready = False
        return self._has_valid_preflight() and message_ready

    def _refresh_primary_button(self):
        if not hasattr(self, "submit_button"):
            return
        busy = (
            self._commit_active
            or self._preflight_active
            or self._confirmation_active
            or self._scan_active
            or self.closing
        )
        selected_targets = self._selected_targets()
        if not selected_targets:
            self.preflight_button.configure(text="① 预检查（多分支）")
            self.submit_button.configure(text="SVN 单分支提交")
            self.preflight_button.state(["disabled"])
            if busy or not self._selected_items():
                self.submit_button.state(["disabled"])
            else:
                self.submit_button.state(["!disabled"])
            return
        self.preflight_button.configure(text="① 预检查（必需）")
        self.submit_button.configure(text="② 开始提交")
        if busy or not self._can_preflight():
            self.preflight_button.state(["disabled"])
        else:
            self.preflight_button.state(["!disabled"])
        if busy or not self._can_start():
            self.submit_button.state(["disabled"])
        else:
            self.submit_button.state(["!disabled"])

    def _selected_targets(self):
        self._target_selection.update({name: bool(var.get()) for name, var in self.target_vars.items()})
        return [candidate.name for candidate in self.candidates if self._target_selection.get(candidate.name, False) and candidate.name != self.source_var.get()]

    def _render_target_statuses(self):
        labels = {
            "ready": "可直接同步", "confirmation_required": "需人工确认", "blocked": "安全阻断",
            "already_present": "已同步", "committed": "已提交", "partial": "部分成功",
            "skipped": "目标不处理", "unknown": "未知", "cancelled": "已取消",
            "failed": "失败", "pending": "待处理",
        }
        self._target_status_map = {}
        if self.current_batch:
            for target, state in self.current_batch.target_status.items():
                actions = [plan.actions.get(target) for plan in self.current_batch.files]
                actions = [action for action in actions if action is not None]
                if any(action.state == "confirmation_required" for action in actions):
                    label = "需人工确认"
                elif state == "ready" and any(action.state in {"ready", "prepared"} for action in actions):
                    label = "可直接同步"
                else:
                    label = labels.get(state, state)
                self._target_status_map[target] = label
        selected = self._selected_targets()
        counts = {"可直接同步": 0, "需人工确认": 0, "安全阻断": 0, "已提交": 0}
        for name in selected:
            counts[self._target_status_map.get(name, "待检查")] = counts.get(self._target_status_map.get(name, "待检查"), 0) + 1
        self.target_summary_var.set(
            f"已选 {len(selected)} · 可直接 {counts['可直接同步']} · 待确认 {counts['需人工确认']} · 阻断 {counts['安全阻断']}"
        )
        if not selected:
            self.target_summary_var.set(
                "未选目标分支 · 将直接打开源分支 TortoiseSVN 提交，不执行多分支预检查"
            )
        elif self.current_batch is None:
            self.target_summary_var.set(
                f"已选 {len(selected)} · 尚未预检查 · 完成预检查后才能开始多分支提交"
            )
        self._update_confirmation_alert()
        self._rebuild_targets()

    def _confirmation_entries(self) -> list[tuple[str, FilePlan, BatchFileAction]]:
        if self.current_batch is None:
            return []
        entries: list[tuple[str, FilePlan, BatchFileAction]] = []
        for plan in self.current_batch.files:
            for target in self.current_batch.target_branches:
                action = plan.actions.get(target)
                if action is not None and action.state == "confirmation_required":
                    entries.append((target, plan, action))
        return entries

    def _update_confirmation_alert(self) -> None:
        if not hasattr(self, "confirmation_alert"):
            return
        entries = self._confirmation_entries()
        if entries:
            conflict_count = sum(
                int(plan.target_summaries.get(target, {}).get("confirmation", 1))
                for target, plan, _action in entries
            )
            self.confirmation_alert_var.set(
                f"⚠ {len(entries)} 个文件存在 {conflict_count} 项内容重叠。确认采用源修改或移出批次后才能提交。"
            )
            if not self.confirmation_alert.winfo_manager():
                self.confirmation_alert.pack(
                    fill="x",
                    pady=(10, 0),
                    before=self.bottom_bar,
                )
            self.status_label.configure(style="Status.FooterError.TLabel")
            self.confirmation_button.state(
                ["disabled"] if self._confirmation_active else ["!disabled"]
            )
        else:
            if self.confirmation_alert.winfo_manager():
                self.confirmation_alert.pack_forget()
            self.status_label.configure(style="Status.Footer.TLabel")

    def _open_confirmation_dialog(self, *, preferred_target: str | None = None):
        from tkinter import messagebox

        if self._confirmation_active or self._commit_active:
            return
        existing = self._confirmation_dialog
        if existing is not None:
            try:
                if existing.winfo_exists():
                    existing.deiconify(); existing.lift(); existing.focus_force()
                    return
            except self.tk.TclError:
                pass
        batch = self.current_batch
        if not batch or not (
            batch.source_status == "committed"
            or self._approved_preflight_signature == self._request_signature()
        ):
            messagebox.showwarning(
                "请先预检查",
                "当前选择尚未完成有效预检查，请先点击“① 预检查（必需）”。",
                parent=self.root,
            )
            return
        entries = self._confirmation_entries()
        if preferred_target:
            entries.sort(key=lambda item: item[0] != preferred_target)
        if not entries:
            messagebox.showinfo("人工确认", "当前没有待确认的内容重叠项。", parent=self.root)
            return
        tk, ttk = self.tk, self.ttk
        win = tk.Toplevel(self.root)
        win.title("确认源分支修改")
        win.geometry("980x620")
        win.transient(self.root)
        self._confirmation_dialog = win
        self._confirmation_dialog_rows = {}
        frame = ttk.Frame(win, padding=10); frame.pack(fill="both", expand=True)
        self._confirmation_dialog_summary_var = tk.StringVar(
            value=f"待确认 {len(entries)} 个文件 · 只应用源修改位置，不替换整个目标文件"
        )
        ttk.Label(
            frame,
            textvariable=self._confirmation_dialog_summary_var,
            style="Title.TLabel",
        ).pack(anchor="w", pady=(0, 8))
        ttk.Label(
            frame,
            text="选择文件可查看修改前、源修改和目标值。确认结果保留到窗口关闭；此时目标工作副本仍不会被修改。",
            style="Muted.App.TLabel",
        ).pack(anchor="w", pady=(0, 8))
        tree = ttk.Treeview(
            frame,
            columns=("branch", "file", "state", "reason"),
            show="headings",
            selectmode="browse",
            height=9,
        )
        for key, title, width in (
            ("branch", "目标分支", 130),
            ("file", "文件", 250),
            ("state", "确认状态", 100),
            ("reason", "内容重叠原因", 430),
        ):
            tree.heading(key, text=title); tree.column(key, width=width, anchor="w")
        self._confirmation_dialog_tree = tree
        tree.tag_configure("pending", foreground=THEME.error, background="#FDE7E9")
        tree.tag_configure("processing", foreground=THEME.warning, background="#FFF4CE")
        tree.tag_configure("completed", foreground=THEME.success, background="#DFF6DD")
        tree.tag_configure("excluded", foreground=THEME.secondary_text, background="#F2F2F2")
        tree.tag_configure("failed", foreground=THEME.error, background="#FDE7E9")
        row_entries: dict[str, tuple[str, FilePlan]] = {}
        for index, (target, plan, action) in enumerate(entries):
            iid = f"confirmation-{index}"
            row_entries[iid] = (target, plan)
            self._confirmation_dialog_rows[(target, plan.relative_path)] = iid
            tree.insert(
                "",
                "end",
                iid=iid,
                values=(target, plan.relative_path, "待确认", action.reason),
                tags=("pending",),
            )
        tree.pack(fill="x")
        detail_box = ttk.LabelFrame(
            frame,
            text="修改明细（修改前 → 源分支新值；目标当前值）",
            padding=6,
            style="Panel.TLabelframe",
        )
        detail_box.pack(fill="both", expand=True, pady=(8, 0))
        detail = tk.Text(detail_box, height=10, wrap="none", state="disabled", font=("Consolas", 9))
        detail.pack(fill="both", expand=True)
        self._confirmation_detail = detail

        def render_details(_event=None):
            selection = tree.selection()
            lines: list[str] = []
            if selection:
                target, plan = row_entries[selection[0]]
                conflicts = [
                    item for item in plan.target_details.get(target, [])
                    if item.get("kind") == "confirmation"
                ]
                for index, item in enumerate(conflicts, 1):
                    location = "/".join(filter(None, (
                        str(item.get("sheet", "")),
                        str(item.get("key", "")),
                        str(item.get("field", "")),
                    )))
                    lines.extend((
                        f"{index}. {location}",
                        f"   修改前：{item.get('before', '—')}",
                        f"   源修改：{item.get('source', '—')}",
                        f"   目标值：{item.get('target', '—')}",
                        f"   原因：{item.get('reason', '')}",
                    ))
            detail.configure(state="normal")
            detail.delete("1.0", tk.END)
            detail.insert("1.0", "\n".join(lines))
            detail.configure(state="disabled")
            self._refresh_confirmation_dialog_controls()
        first = next(iter(row_entries), None)
        if first:
            tree.selection_set(first); tree.focus(first)
        buttons = ttk.Frame(frame); buttons.pack(fill="x", pady=(8, 0))

        def open_selected(_event=None):
            selection = tree.selection()
            if not selection:
                return
            state = tree.set(selection[0], "state")
            if state not in {"待确认", "未完成"} or self._confirmation_active:
                return
            target, plan = row_entries[selection[0]]
            if not messagebox.askyesno(
                "确认采用源修改",
                f"将对 {target}/{plan.relative_path} 的全部重叠项采用源分支修改。\n"
                "目标文件其他记录和字段保持不变。是否确认？",
                parent=win,
            ):
                return
            self._set_confirmation_dialog_row(
                target,
                plan.relative_path,
                "processing",
                "正在核对目标文件状态…",
            )
            self._start_confirmation(target, plan.relative_path)

        def exclude_selected():
            selection = tree.selection()
            if not selection:
                return
            state = tree.set(selection[0], "state")
            if state not in {"待确认", "未完成"} or self._confirmation_active:
                return
            target, plan = row_entries[selection[0]]
            if not messagebox.askyesno(
                "移出本批次",
                f"{target}/{plan.relative_path} 将不会同步源修改，也不会进入该目标分支提交窗口。\n是否移出？",
                parent=win,
            ):
                return
            try:
                self.engine.exclude_target_file(batch, target, plan.relative_path)
            except Exception as exc:
                messagebox.showwarning("无法移出", str(exc), parent=win)
                return
            self._set_confirmation_dialog_row(
                target, plan.relative_path, "excluded", "已从该目标分支批次移除"
            )
            self._render_target_statuses()
            self._refresh_primary_button()

        def close_dialog():
            if self._confirmation_active:
                self.status_var.set("正在核对确认项，请稍候")
                return
            try:
                win.grab_release()
            except tk.TclError:
                pass
            self._confirmation_dialog = None
            self._confirmation_dialog_tree = None
            self._confirmation_dialog_button = None
            self._confirmation_exclude_button = None
            self._confirmation_detail = None
            self._confirmation_dialog_summary_var = None
            self._confirmation_dialog_rows = {}
            win.destroy()

        ttk.Button(buttons, text="关闭", command=close_dialog).pack(side="right")
        self._confirmation_dialog_button = ttk.Button(
            buttons,
            text="确认采用源修改",
            style="Danger.TButton",
            command=open_selected,
        )
        self._confirmation_dialog_button.pack(side="right", padx=6)
        self._confirmation_exclude_button = ttk.Button(
            buttons, text="移出本批次", command=exclude_selected
        )
        self._confirmation_exclude_button.pack(side="right")
        tree.bind("<<TreeviewSelect>>", render_details)
        tree.bind("<Double-1>", open_selected)
        win.protocol("WM_DELETE_WINDOW", close_dialog)
        render_details()
        win.grab_set()

    def _set_confirmation_dialog_row(
        self,
        target: str,
        relative_path: str,
        state: str,
        reason: str = "",
    ) -> None:
        tree = self._confirmation_dialog_tree
        iid = self._confirmation_dialog_rows.get((target, relative_path))
        if tree is None or iid is None:
            return
        try:
            if not tree.winfo_exists() or not tree.exists(iid):
                return
            labels = {
                "pending": "待确认",
                "processing": "核对中…",
                "completed": "已确认",
                "excluded": "已移除",
                "failed": "未完成",
            }
            tree.set(iid, "state", labels.get(state, state))
            if reason:
                tree.set(iid, "reason", reason)
            tree.item(iid, tags=(state,))
            self._refresh_confirmation_dialog_controls()
        except self.tk.TclError:
            pass

    def _refresh_confirmation_dialog_controls(self) -> None:
        tree = self._confirmation_dialog_tree
        button = self._confirmation_dialog_button
        exclude = self._confirmation_exclude_button
        if tree is None or button is None or exclude is None:
            return
        try:
            rows = list(tree.get_children())
            completed = sum(tree.set(iid, "state") in {"已确认", "已移除"} for iid in rows)
            remaining = len(rows) - completed
            if self._confirmation_dialog_summary_var is not None:
                if remaining:
                    self._confirmation_dialog_summary_var.set(
                        f"待确认 {remaining} 个文件 · 已处理 {completed} 个 · 完成项将在关闭窗口前保留"
                    )
                else:
                    self._confirmation_dialog_summary_var.set(
                        f"全部 {completed} 个文件均已处理，可以关闭窗口并开始提交"
                    )
            selection = tree.selection()
            selected_state = tree.set(selection[0], "state") if selection else ""
            if self._confirmation_active or selected_state not in {"待确认", "未完成"}:
                button.state(["disabled"])
                exclude.state(["disabled"])
            else:
                button.state(["!disabled"])
                exclude.state(["!disabled"])
        except self.tk.TclError:
            pass

    def _start_confirmation(self, target: str, relative_path: str) -> None:
        from tkinter import messagebox

        batch = self.current_batch
        if batch is None or not (
            batch.source_status == "committed"
            or self._approved_preflight_signature == self._request_signature()
        ):
            self.status_var.set("当前预检查结果已失效，请重新预检查")
            self._refresh_primary_button()
            return
        if self._confirmation_active or self._commit_active:
            return
        self._confirmation_active = True
        self.source_box.configure(state="disabled")
        self.message.configure(state="disabled")
        self.status_var.set(f"正在核对确认项：{target}/{relative_path}")
        self._update_confirmation_alert()
        self._refresh_confirmation_dialog_controls()
        self._refresh_primary_button()

        def worker(_cancel_event):
            return self.engine.confirm_source_changes(batch, target, relative_path)

        def done(_action, error, _generation):
            self._confirmation_active = False
            self.source_box.configure(state="readonly")
            self.message.configure(state="normal")
            if error:
                self.status_var.set(str(error))
                self._set_confirmation_dialog_row(target, relative_path, "failed", str(error))
                self._update_confirmation_alert()
                self._refresh_primary_button()
                parent = self._confirmation_dialog or self.root
                messagebox.showwarning("人工确认未完成", str(error), parent=parent)
                return
            self._set_confirmation_dialog_row(
                target,
                relative_path,
                "completed",
                "已确认采用源分支修改",
            )
            self._render_target_statuses()
            remaining = len(self._confirmation_entries())
            if remaining:
                self.status_var.set(f"已确认；仍有 {remaining} 个文件需要处理")
            else:
                self.status_var.set("所有内容重叠项均已处理；目标工作副本仍未修改，可以开始提交")
            self._refresh_primary_button()

        self.ui_tasks.submit(worker, done)

    def _selected_items(self):
        return [item for item in self.items if item.checked and item.selectable]

    def _show_operation_policies(self):
        """Show the exact source-status to target-operation contract."""
        tk, ttk = self.tk, self.ttk
        win = tk.Toplevel(self.root)
        win.title("多分支提交 · SVN 状态处理规则")
        win.geometry("900x430")
        win.minsize(760, 360)
        win.transient(self.root)
        frame = ttk.Frame(win, padding=14, style="App.TFrame")
        frame.pack(fill="both", expand=True)
        ttk.Label(frame, text="SVN 状态处理规则", style="Title.App.TLabel").pack(anchor="w")
        ttk.Label(
            frame,
            text="源分支仍由原生 TortoiseSVN 提交；下表只决定是否以及如何传播到目标分支。",
            style="Muted.App.TLabel",
        ).pack(anchor="w", pady=(3, 10))
        tree = ttk.Treeview(
            frame,
            columns=("status", "selection", "target", "safety"),
            show="headings",
            selectmode="browse",
            style="Workbench.Treeview",
        )
        for key, title, width in (
            ("status", "源 SVN 状态", 200),
            ("selection", "默认选择", 105),
            ("target", "目标分支动作", 265),
            ("safety", "安全规则", 300),
        ):
            tree.heading(key, text=title)
            tree.column(key, width=width, anchor="w", stretch=key in {"target", "safety"})
        for index, values in enumerate(SVN_OPERATION_POLICIES):
            tree.insert("", "end", values=values, tags=("alternate",) if index % 2 else ())
        tree.tag_configure("alternate", background=THEME.row_alt)
        tree.pack(fill="both", expand=True)
        note = (
            "重点：missing 只表示工作文件在磁盘上缺失，并不等于已经执行 SVN Delete。"
            "因此工具不会更新或修改任何目标分支；如需跨分支删除，请先在源分支明确执行 TortoiseSVN → Delete。"
        )
        ttk.Label(frame, text=note, style="Workbench.Hint.TLabel", wraplength=850).pack(
            anchor="w", pady=(10, 0)
        )
        ttk.Button(frame, text="关闭", command=win.destroy).pack(anchor="e", pady=(10, 0))

    def _show_recent_messages(self):
        tk, ttk = self.tk, self.ttk
        try: _root_url, repo_uuid = repository_metadata(self.context.wc_root); messages = read_recent_messages(repo_uuid)
        except Exception as exc: messages = []; self.status_var.set(str(exc))
        win = tk.Toplevel(self.root); win.title("最近提交消息"); win.geometry("720x420"); win.transient(self.root)
        frame = ttk.Frame(win, padding=10); frame.pack(fill="both", expand=True)
        query = tk.StringVar(); ttk.Entry(frame, textvariable=query).pack(fill="x", pady=(0, 6))
        listing = tk.Listbox(frame, exportselection=False); listing.pack(fill="both", expand=True)
        def render(*_):
            listing.delete(0, tk.END); needle=query.get().lower()
            for value in messages:
                if not needle or needle in value.lower(): listing.insert(tk.END, value)
        def use():
            if listing.curselection():
                value=listing.get(listing.curselection()[0]); self.message.delete("1.0", tk.END); self.message.insert("1.0", value); self._refresh_primary_button(); win.destroy()
        query.trace_add("write", render); render(); listing.bind("<Double-1>", lambda _e: use())
        ttk.Button(frame, text="使用所选消息", command=use).pack(anchor="e", pady=(6,0))

    def _paste_filenames(self):
        names = "\n".join(item.relative_path for item in self._selected_items())
        if names:
            self.message.insert(self.tk.END, ("\n" if self.message.get("1.0", self.tk.END).strip() else "") + names)
            self._refresh_primary_button()

    def _show_log(self):
        threading.Thread(target=lambda: self.engine.show_log(self.scope_var.get()), daemon=True).start()

    def _matrix_dialog(self, batch: BranchSubmitBatch) -> bool:
        tk, ttk = self.tk, self.ttk
        result = {"ok": False}
        win = tk.Toplevel(self.root)
        win.title("多分支提交 · 预检查结果")
        win.geometry("1080x650")
        win.minsize(860, 560)
        win.configure(background=THEME.window_bg)
        win.transient(self.root)
        frame = ttk.Frame(win, padding=14, style="App.TFrame")
        frame.pack(fill="both", expand=True)

        header = ttk.Frame(frame, style="App.TFrame")
        header.pack(fill="x")
        ttk.Label(header, text="预检查结果", style="Title.App.TLabel").pack(anchor="w")
        ttk.Label(
            header,
            text=f"批次 {batch.batch_id}  ·  提交按分支依次执行，可中断、可恢复",
            style="Muted.App.TLabel",
        ).pack(anchor="w", pady=(3, 10))

        states = [action.state for plan in batch.files for action in plan.actions.values()]
        summary_bar = tk.Frame(frame, background=THEME.window_bg)
        summary_bar.pack(fill="x", pady=(0, 10))
        summary_items = (
            ("目标分支", len(batch.target_branches), "#E8F1FB", THEME.accent),
            ("文件", len(batch.files), "#EEF0F2", THEME.text),
            ("可直接同步", states.count("ready"), "#E7F4E4", THEME.success),
            ("需人工确认", states.count("confirmation_required"), "#FFF4CE", THEME.warning),
            ("安全阻断", states.count("blocked"), "#FDE7E9", THEME.error),
        )
        for label, value, background, foreground in summary_items:
            chip = tk.Label(
                summary_bar,
                text=f"  {label}  {value}  ",
                background=background,
                foreground=foreground,
                font=(THEME.font_family, 9, "bold"),
                padx=4,
                pady=5,
            )
            chip.pack(side="left", padx=(0, 8))

        result_box = ttk.LabelFrame(frame, text="目标分支 × 文件", padding=8, style="Panel.TLabelframe")
        result_box.pack(fill="both", expand=True)
        tree_holder = ttk.Frame(result_box, style="Panel.TFrame")
        tree_holder.pack(fill="both", expand=True)
        tree = ttk.Treeview(
            tree_holder,
            columns=("branch", "file", "operation", "state", "reason"),
            show="headings",
            selectmode="browse",
        )
        columns = (
            ("branch", "目标分支", 135, False),
            ("file", "文件", 270, True),
            ("operation", "动作", 72, False),
            ("state", "处理状态", 115, False),
            ("reason", "说明", 390, True),
        )
        for key, title, width, stretch in columns:
            tree.heading(key, text=title)
            tree.column(key, width=width, minwidth=60, stretch=stretch, anchor="w")
        yscroll = ttk.Scrollbar(tree_holder, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=yscroll.set)
        tree.pack(side="left", fill="both", expand=True)
        yscroll.pack(side="right", fill="y")
        tree.tag_configure("ready", foreground=THEME.success, background="#F3FAF1")
        tree.tag_configure("confirmation", foreground=THEME.error, background="#FDE7E9", font=(THEME.font_family, 9, "bold"))
        tree.tag_configure("blocked", foreground=THEME.error, background="#FFF3F3")
        tree.tag_configure("already", foreground=THEME.secondary_text, background=THEME.row_alt)
        tree.tag_configure("excluded", foreground=THEME.disabled, background=THEME.row_alt)

        state_labels = {
            "ready": "可直接同步", "confirmation_required": "需人工确认",
            "excluded": "目标不处理", "already_applied": "已包含修改",
            "blocked": "安全阻断", "prepared": "已准备", "committed": "已提交",
        }
        operation_labels = {
            "modify": "修改",
            "add": "新增",
            "delete": "删除",
            SOURCE_ONLY_MISSING: "仅源分支",
        }
        row_map: dict[str, tuple[str, FilePlan]] = {}
        for row_index, (plan, target) in enumerate(
            pair for plan in batch.files for pair in ((plan, target) for target in batch.target_branches)
        ):
            action = plan.actions[target]
            summary = plan.target_summaries.get(target, {})
            reason = action.reason or (
                f"将修改 {summary.get('direct', 0)} 项，保留目标分支其他内容"
                if plan.operation == "modify" else ""
            )
            tag = {
                "ready": "ready", "confirmation_required": "confirmation",
                "blocked": "blocked", "already_applied": "already", "excluded": "excluded",
            }.get(action.state, "")
            iid = f"matrix-{row_index}"
            row_map[iid] = (target, plan)
            tree.insert(
                "", "end", iid=iid,
                values=(
                    target, plan.relative_path, operation_labels.get(plan.operation, plan.operation),
                    state_labels.get(action.state, action.state), reason,
                ),
                tags=(tag,),
            )

        detail_box = ttk.LabelFrame(frame, text="所选项", padding=9, style="Panel.TLabelframe")
        detail_box.pack(fill="x", pady=(10, 0))
        detail_var = tk.StringVar(value="选择一项查看判断依据。")
        ttk.Label(detail_box, textvariable=detail_var, style="App.TLabel", wraplength=1000).pack(anchor="w")
        compare_hint = ttk.Label(
            detail_box,
            text="查看内容：目标修改前  ↔  应用源分支修改后的目标预览（不会修改工作副本）",
            foreground=THEME.accent,
        )
        compare_hint.pack(anchor="w", pady=(5, 0))

        footer = ttk.Frame(frame, style="App.TFrame")
        footer.pack(fill="x", pady=(10, 0))
        preview_status = tk.StringVar(value="")
        ttk.Label(footer, textvariable=preview_status, style="Muted.App.TLabel").pack(side="left", padx=(0, 10))
        preview_button = ttk.Button(footer, text="查看目标修改点", style="App.TButton")
        preview_button.pack(side="left")
        preview_busy = {"value": False}
        preview_results: queue.Queue = queue.Queue()

        def selected_entry() -> tuple[str, FilePlan] | None:
            selection = tree.selection()
            return row_map.get(selection[0]) if selection else None

        def update_selection(_event=None) -> None:
            entry = selected_entry()
            if not entry:
                detail_var.set("选择一项查看判断依据。")
                preview_button.state(["disabled"])
                return
            target, plan = entry
            action = plan.actions[target]
            item_summary = plan.target_summaries.get(target, {})
            counts = " · ".join(
                text for text in (
                    f"直接 {item_summary.get('direct', 0)}" if item_summary.get("direct") else "",
                    f"已包含 {item_summary.get('already', 0)}" if item_summary.get("already") else "",
                    f"需确认 {item_summary.get('confirmation', 0)}" if item_summary.get("confirmation") else "",
                ) if text
            )
            detail_var.set(
                f"{target} / {plan.relative_path}  ·  {state_labels.get(action.state, action.state)}"
                f"{('  ·  ' + counts) if counts else ''}\n{action.reason or '未提供补充说明'}"
            )
            can_preview = plan.operation == "modify" and action.state in {"ready", "confirmation_required"}
            preview_button.state(["!disabled"] if can_preview and not preview_busy["value"] else ["disabled"])

        def compare_selected(_event=None) -> None:
            from tkinter import messagebox
            entry = selected_entry()
            if not entry or preview_busy["value"]:
                return
            target, plan = entry
            action = plan.actions[target]
            if plan.operation != "modify" or action.state not in {"ready", "confirmation_required"}:
                return
            preview_busy["value"] = True
            preview_button.state(["disabled"])
            preview_status.set("正在生成目标修改后预览…")

            def worker() -> None:
                error = None
                try:
                    self.engine.open_excel_comparison(batch, plan, target)
                except Exception as exc:  # noqa: BLE001 - surfaced on the Tk thread
                    error = exc
                preview_results.put(error)

            def poll_result() -> None:
                try:
                    error = preview_results.get_nowait()
                except queue.Empty:
                    if preview_busy["value"] and win.winfo_exists():
                        win.after(40, poll_result)
                    return
                preview_busy["value"] = False
                if error:
                    preview_status.set("预览生成失败")
                    messagebox.showwarning("无法查看目标修改点", str(error), parent=win)
                else:
                    preview_status.set("已打开：目标修改前 ↔ 目标修改后预览")
                update_selection()

            threading.Thread(target=worker, name="sow-target-preview", daemon=True).start()
            win.after(40, poll_result)

        def accept() -> None:
            result["ok"] = True
            win.destroy()

        def open_confirmations() -> None:
            win.destroy()
            self.root.after_idle(self._open_confirmation_dialog)

        preview_button.configure(command=compare_selected)
        ttk.Button(footer, text="关闭", command=win.destroy).pack(side="right")
        if batch.source_status == "ready":
            ttk.Button(footer, text="保留预检查结果", style="Primary.TButton", command=accept).pack(side="right", padx=6)
        if self._confirmation_entries():
            ttk.Button(footer, text="处理人工确认", style="Danger.TButton", command=open_confirmations).pack(side="right", padx=6)
        tree.bind("<<TreeviewSelect>>", update_selection)
        tree.bind("<Double-1>", compare_selected)
        first = tree.get_children()
        if first:
            tree.selection_set(first[0])
            tree.focus(first[0])
            update_selection()
        win.grab_set()
        self.root.wait_window(win)
        return result["ok"]

    def _preflight(self, *, auto_start: bool = False):
        from tkinter import messagebox
        if self._commit_active or self._preflight_active or self._confirmation_active:
            return
        if not self._can_preflight():
            messagebox.showwarning(
                "无法预检查",
                "请先选择目标分支和需要同步的 Excel 变更。\n提交说明可以在预检查完成后再填写。",
                parent=self.root,
            )
            return
        selected = list(self._selected_items())
        targets = list(self._selected_targets())
        message = self.message.get("1.0", self.tk.END)
        source = self.source_var.get()
        scope = self.scope_var.get()
        selection_generation = self._selection_generation
        request_signature = self._request_signature()
        self.current_batch = None
        self._approved_preflight_signature = None
        self._preflight_active = True
        self._update_confirmation_alert()
        self.submit_button.state(["disabled"])
        self.preflight_button.state(["disabled"])
        self.status_var.set(
            f"正在更新 {len(targets)} 个目标分支的相关路径，随后生成预检查矩阵…"
        )
        self.root.update_idletasks()

        def worker(cancel_event):
            if cancel_event.is_set():
                return None
            result = self.engine.preflight(source, targets, selected, message, scope_path=scope)
            if cancel_event.is_set():
                self.engine.abandon(result)
                return None
            return result

        def done(batch, error, _generation):
            self._preflight_active = False
            if selection_generation != self._selection_generation:
                self.status_var.set("预检查期间选择已变化，请重新执行预检查")
                self._refresh_primary_button()
                return
            if error:
                self.status_var.set(str(error))
                self._refresh_primary_button()
                messagebox.showerror("预检查失败", str(error), parent=self.root)
                return
            if batch is None:
                self.status_var.set("预检查已取消")
                self._refresh_primary_button()
                return
            self.current_batch = batch
            self._approved_preflight_signature = request_signature
            self._render_target_statuses()
            if not auto_start:
                self._matrix_dialog(batch)
            if batch.source_status == "ready":
                ready_count = sum(a.state == "ready" for p in batch.files for a in p.actions.values())
                confirmation_count = sum(a.state == "confirmation_required" for p in batch.files for a in p.actions.values())
                if confirmation_count:
                    self.status_var.set(
                        f"预检查完成：可直接同步 {ready_count} 个文件，需人工确认 {confirmation_count} 个文件"
                    )
                else:
                    if self.message.get("1.0", self.tk.END).strip():
                        self.status_var.set(f"预检查通过：可直接同步 {ready_count} 个文件；现在可以开始提交")
                    else:
                        self.status_var.set(f"预检查通过：可直接同步 {ready_count} 个文件；填写提交说明后可开始提交")
            else:
                self.status_var.set(batch.error or "预检查结果未确认")
            self._refresh_primary_button()

        self._preflight_generation = self.ui_tasks.submit(worker, done)

    def _submit(self):
        from tkinter import messagebox
        if not self._selected_targets():
            self._submit_single_branch()
            return
        batch=self.current_batch
        if self._commit_active:return
        if not batch or not self._has_valid_preflight():
            self.status_var.set("多分支提交必须先完成预检查，并处理所有人工确认或阻断项")
            messagebox.showwarning(
                "请先预检查",
                "多分支提交必须先点击“① 预检查（必需）”。\n"
                "如出现红色“需人工确认”，还需确认采用源修改或将该项移出批次。\n\n"
                "原生 TortoiseSVN 单分支提交不受此门禁影响。",
                parent=self.root,
            )
            return
        message = self.message.get("1.0", self.tk.END).strip()
        if not message:
            self.status_var.set("预检查已通过；填写提交说明后才能开始提交")
            messagebox.showwarning(
                "请填写提交说明",
                "预检查已经完成，不需要重新检查。\n请填写 SVN 提交说明后再开始提交。",
                parent=self.root,
            )
            try:
                self.message.focus_set()
            except Exception:
                pass
            return
        if not messagebox.askyesno("开始分步提交","将依次打开源分支和目标分支的 TortoiseSVN 提交窗口。\n任何取消、部分勾选或未知结果都会停止后续分支。\n\n继续吗？",parent=self.root):return
        batch.message = message
        batch.event("commit-message-frozen", length=len(message))
        batch.save()
        self.submit_button.state(["disabled"])
        self.preflight_button.state(["disabled"])
        self._commit_active = True
        self.status_var.set("等待 TortoiseSVN 提交与逐文件对账…")
        self.root.update_idletasks()

        def done(result, error, _generation):
            self._commit_active = False
            self._approved_preflight_signature = None
            if error:
                self.status_var.set(str(error))
                self._refresh_primary_button()
                messagebox.showerror("提交失败", str(error), parent=self.root)
                return
            self.current_batch = result
            self._render_target_statuses()
            self.status_var.set(_format_batch_result(result).replace("\n", " · "))
            messagebox.showinfo("批次结果", _format_batch_result(result), parent=self.root)
            if result.superseded_by:
                child_path = os.path.join(settings_dir(), "batches", result.superseded_by, "batch.json")
                if os.path.isfile(child_path) and messagebox.askyesno("载入已提交子批次", f"源分支部分文件已成功提交。\n是否载入子批次 {result.superseded_by}，稍后再次点击“开始提交”传播这些已提交文件？", parent=self.root):
                    self.current_batch = BranchSubmitBatch.load(child_path)
                    self.status_var.set(f"已载入子批次 {result.superseded_by}；继续前请重新预检查")
            self._refresh_primary_button()

        self.ui_tasks.submit(lambda _cancel_event: self.engine.commit(batch), done)

    def _submit_single_branch(self):
        """Use the native TortoiseSVN commit path when no target is selected."""
        from tkinter import messagebox
        if self._commit_active or self._preflight_active or self._confirmation_active:
            return
        selected = list(self._selected_items())
        if not selected:
            messagebox.showwarning("请选择文件", "请先选择要提交的文件。", parent=self.root)
            return
        source_paths = [item.path for item in selected]
        message = self.message.get("1.0", self.tk.END).strip()
        self.current_batch = None
        self._approved_preflight_signature = None
        self._commit_active = True
        self.submit_button.state(["disabled"])
        self.preflight_button.state(["disabled"])
        self.status_var.set("正在打开源分支原生 TortoiseSVN 提交窗口…")
        self.root.update_idletasks()

        def worker(cancel_event):
            if cancel_event.is_set():
                return None
            return self.engine._tortoise(
                "commit",
                source_paths,
                message=message or None,
            )

        def done(exit_code, error, _generation):
            self._commit_active = False
            if error:
                self.status_var.set(str(error))
                messagebox.showerror("单分支提交失败", str(error), parent=self.root)
                self._refresh_primary_button()
                return
            if exit_code is None:
                self.status_var.set("单分支提交已取消")
            elif exit_code == 0:
                self.status_var.set("原生 SVN 提交窗口已关闭；正在刷新源分支状态…")
            else:
                self.status_var.set(
                    f"原生 SVN 提交窗口已关闭（退出码 {exit_code}）；正在刷新状态…"
                )
            self._start_scan(preserve_batch=False)

        self.ui_tasks.submit(worker, done)

    def _close(self):
        if self._commit_active:
            self.status_var.set("提交进行中，当前窗口不能关闭；请等待 TortoiseSVN 对账完成")
            return
        if self._confirmation_active:
            self.status_var.set("人工确认正在核对目标状态，请稍候")
            return
        self.closing = True
        self.scan_cancel.set()
        self.ui_tasks.close()
        if self.current_batch is not None and not batch_requires_recovery(self.current_batch):
            try:
                self.engine.abandon(self.current_batch)
            except Exception:
                pass
        data=dict(self.settings);data["wc_root"]=self.context.wc_root;data["window_geometry"]=self.root.geometry()
        data["column_widths"]={key:self.tree.column(key,"width") for key in self.tree["columns"]}
        data.setdefault("last_targets",{})[self.source_var.get()]=self._selected_targets()
        data["favorite_branches"]=list(dict.fromkeys(data.get("favorite_branches",[])))
        save_settings(data);self.root.destroy()


def launch_ui(initial_paths: Iterable[str] | None = None) -> None:
    if os.name == "nt":
        try:
            import ctypes
            ctypes.windll.shcore.SetProcessDpiAwareness(1)
        except Exception:
            try:
                ctypes.windll.user32.SetProcessDPIAware()
            except Exception:
                pass
    import tkinter as tk
    from tkinter import filedialog, messagebox
    root = tk.Tk(); root.withdraw()
    resume_batch = None
    corrupt = list_corrupt_batch_files()
    if corrupt:
        messagebox.showwarning(
            "发现损坏的批次状态",
            "以下批次状态无法解析，工具不会猜测或自动处理：\n\n" + "\n".join(corrupt[:8]),
            parent=root,
        )
    unfinished = list_unfinished_batches()
    if unfinished:
        choice = _choose_recovery_action(root, unfinished)
        if choice:
            action, batch = choice
            engine = BranchSubmitEngine(batch.wc_root)
            if action == "restore":
                engine.restore_uncommitted(batch); messagebox.showinfo("恢复结果", _format_batch_result(batch), parent=root)
            elif action == "abandon":
                engine.abandon(batch)
            elif action == "continue":
                resume_batch = batch
    paths = [os.path.abspath(str(path)) for path in (initial_paths or []) if str(path).strip()]
    try:
        if resume_batch:
            context = BranchContext(resume_batch.wc_root, resume_batch.source_branch, resume_batch.scope_path or os.path.join(resume_batch.wc_root, resume_batch.source_branch))
        elif paths:
            context = infer_context(paths)
        else:
            settings = load_settings(); initial = settings.get("wc_root") or os.getcwd()
            folder = filedialog.askdirectory(parent=root, title="选择某个 SVN 分支内的文件夹", initialdir=initial)
            if not folder:
                root.destroy(); return
            context = infer_context([folder])
    except Exception as exc:
        messagebox.showerror("无法打开多分支提交", str(exc), parent=root); root.destroy(); return
    root.deiconify()
    try:
        BranchSubmitWorkbench(root, context, resume_batch=resume_batch)
        root.mainloop()
    except Exception as exc:
        messagebox.showerror("多分支提交启动失败", str(exc), parent=root)
        root.destroy()


def prompt_mode() -> str:
    import tkinter as tk
    from tkinter import ttk
    root=tk.Tk();root.title("Excel 合并器");result={"mode":"legacy"}
    frame=ttk.Frame(root,padding=20);frame.pack(fill="both",expand=True)
    ttk.Label(frame,text="请选择工作模式",font=("Microsoft YaHei",12,"bold")).pack(pady=(0,16))
    def choose(mode):result["mode"]=mode;root.destroy()
    ttk.Button(frame,text="Excel 差异 / 冲突合并",command=lambda:choose("legacy"),width=28).pack(pady=5)
    ttk.Button(frame,text="多分支 SVN 提交",command=lambda:choose("branch"),width=28).pack(pady=5)
    root.protocol("WM_DELETE_WINDOW",root.destroy);root.mainloop();return result["mode"]

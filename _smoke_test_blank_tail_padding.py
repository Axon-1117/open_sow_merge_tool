"""Focused regression for value-empty tail coordinate padding alignment."""

import os

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

import sow_merge_tool as mod
from _test_temp_utils import make_temp_dir


SHEET = "Data"


def _save_tail_book(path: str, *, append_rows=(), formula_row=None):
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    for row_idx in range(1, 141):
        ws.cell(row_idx, 1).value = f"same-{row_idx}"

    # Match the Guide shape: style/row-dimension records extend beyond the
    # value boundary.  Style-only structure intentionally remains outside the
    # tool's existing value/formula diff semantics.
    fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    for row_idx in range(141, 152):
        ws.row_dimensions[row_idx].height = 14.25
    for row_idx in range(141, 144):
        ws.cell(row_idx, 8).fill = fill

    if formula_row is not None:
        ws.cell(formula_row, 2).value = '=""'
    for row_idx, value in append_rows:
        ws.cell(row_idx, 1).value = value
    wb.save(path)
    wb.close()


def _open_pair(path_a: str, path_b: str):
    val_a = load_workbook(path_a, read_only=True, data_only=True)
    val_b = load_workbook(path_b, read_only=True, data_only=True)
    edit_a = load_workbook(path_a, read_only=True, data_only=False)
    edit_b = load_workbook(path_b, read_only=True, data_only=False)
    return val_a, val_b, edit_a, edit_b


def _normalized_pairs(val_a, val_b, edit_a, edit_b):
    ws_a_val = val_a[SHEET]
    ws_b_val = val_b[SHEET]
    ws_a_edit = edit_a[SHEET]
    ws_b_edit = edit_b[SHEET]
    max_a, col_a = mod._effective_bounds_with_edit(ws_a_val, ws_a_edit)
    max_b, col_b = mod._effective_bounds_with_edit(ws_b_val, ws_b_edit)
    max_col = max(col_a, col_b)

    foreground = mod._compute_row_pairs_generic(
        ws_a_val,
        ws_b_val,
        max_col,
        max_row_a=max_a,
        max_row_b=max_b,
    )
    foreground = mod._collapse_one_sided_blank_tail_padding(
        foreground,
        ws_a_val,
        ws_b_val,
        ws_a_edit,
        ws_b_edit,
        max_a,
        max_b,
        max_col,
    )

    # Background alignment uses precomputed signatures but must produce the
    # same normalized pair model.
    background = mod._compute_row_pairs_from_signatures(
        mod._row_sig_list_for_ws(ws_a_val, max_a, max_col),
        mod._row_sig_list_for_ws(ws_b_val, max_b, max_col),
    )
    background = mod._collapse_one_sided_blank_tail_padding(
        background,
        ws_a_val,
        ws_b_val,
        ws_a_edit,
        ws_b_edit,
        max_a,
        max_b,
        max_col,
    )
    assert foreground == background
    return foreground, (max_a, max_b, max_col)


def _row_has_value_or_formula(ws_val, ws_edit, row_idx: int, max_col: int) -> bool:
    values = next(
        ws_val.iter_rows(
            min_row=row_idx,
            max_row=row_idx,
            min_col=1,
            max_col=max_col,
            values_only=True,
        ),
        (),
    )
    edits = next(
        ws_edit.iter_rows(
            min_row=row_idx,
            max_row=row_idx,
            min_col=1,
            max_col=max_col,
            values_only=True,
        ),
        (),
    )
    return any(
        value not in (None, "") or edit_value not in (None, "")
        for value, edit_value in zip(values, edits)
    )


def _test_guide_shape_two_way_and_three_way(root_dir: str):
    base_path = os.path.join(root_dir, "base.xlsx")
    mine_path = os.path.join(root_dir, "mine.xlsx")
    theirs_path = os.path.join(root_dir, "theirs.xlsx")
    _save_tail_book(base_path)
    _save_tail_book(mine_path)
    _save_tail_book(
        theirs_path,
        append_rows=((150, "theirs-tail-1"), (151, "theirs-tail-2")),
    )

    val_m, val_t, edit_m, edit_t = _open_pair(mine_path, theirs_path)
    val_b = edit_b = None
    try:
        pairs, (_max_m, _max_t, max_col) = _normalized_pairs(
            val_m, val_t, edit_m, edit_t
        )
        structural = [pair for pair in pairs if (pair[0] is None) != (pair[1] is None)]
        assert structural[-2:] == [(None, 150), (None, 151)], structural[-12:]
        assert not any(pair == (None, row) for row in range(141, 150) for pair in pairs)

        blank_diff_count = sum(
            1
            for left_row, right_row in structural
            if right_row is not None
            and not _row_has_value_or_formula(
                val_t[SHEET], edit_t[SHEET], right_row, max_col
            )
        )
        assert blank_diff_count == 0

        # Three-way Base maps use the same normalizer.  The real appended rows
        # remain unmapped/structural while blank coordinate padding disappears.
        val_b = load_workbook(base_path, read_only=True, data_only=True)
        edit_b = load_workbook(base_path, read_only=True, data_only=False)
        theirs_base_pairs, _bounds = _normalized_pairs(val_t, val_b, edit_t, edit_b)
        theirs_to_base = mod._row_map_from_pairs(theirs_base_pairs)
        assert 150 not in theirs_to_base and 151 not in theirs_to_base
        assert all(row not in theirs_to_base for row in range(141, 150))

        diff_pair_indices = [
            idx
            for idx, pair in enumerate(pairs)
            if (pair[0] is None) != (pair[1] is None)
        ]
        assert diff_pair_indices[-1] == diff_pair_indices[-2] + 1
        assert [pairs[idx] for idx in diff_pair_indices[-2:]] == [
            (None, 150),
            (None, 151),
        ]
    finally:
        for wb in (val_m, val_t, edit_m, edit_t, val_b, edit_b):
            if wb is not None:
                wb.close()


def _test_internal_blank_insert_is_preserved():
    wb_a = Workbook()
    ws_a = wb_a.active
    ws_a.title = SHEET
    wb_b = Workbook()
    ws_b = wb_b.active
    ws_b.title = SHEET
    for row_idx in range(1, 8):
        ws_a.cell(row_idx, 1).value = f"row-{row_idx}"
    ws_b.cell(1, 1).value = "row-1"
    ws_b.cell(2, 1).value = "row-2"
    for source_row in range(3, 8):
        ws_b.cell(source_row + 1, 1).value = f"row-{source_row}"

    max_a, col_a = mod._effective_bounds_with_edit(ws_a, ws_a)
    max_b, col_b = mod._effective_bounds_with_edit(ws_b, ws_b)
    pairs = mod._compute_row_pairs_generic(ws_a, ws_b, max(col_a, col_b))
    normalized = mod._collapse_one_sided_blank_tail_padding(
        pairs, ws_a, ws_b, ws_a, ws_b, max_a, max_b, max(col_a, col_b)
    )
    assert normalized == pairs
    assert (None, 3) in normalized, normalized
    wb_a.close()
    wb_b.close()


def _test_uncached_empty_formula_is_not_padding(root_dir: str):
    mine_path = os.path.join(root_dir, "formula-mine.xlsx")
    theirs_path = os.path.join(root_dir, "formula-theirs.xlsx")
    _save_tail_book(mine_path)
    _save_tail_book(
        theirs_path,
        formula_row=145,
        append_rows=((150, "real-tail"),),
    )
    val_m, val_t, edit_m, edit_t = _open_pair(mine_path, theirs_path)
    try:
        assert val_t[SHEET].cell(145, 2).value is None
        assert edit_t[SHEET].cell(145, 2).value == '=""'
        pairs, _bounds = _normalized_pairs(val_m, val_t, edit_m, edit_t)
        assert (None, 145) in pairs, pairs[-12:]
        assert (None, 150) in pairs, pairs[-12:]
        assert all((None, row) not in pairs for row in range(141, 145))
    finally:
        for wb in (val_m, val_t, edit_m, edit_t):
            wb.close()


def main():
    root_dir = make_temp_dir("sow_blank_tail_padding_")
    _test_guide_shape_two_way_and_three_way(root_dir)
    _test_internal_blank_insert_is_preserved()
    _test_uncached_empty_formula_is_not_padding(root_dir)
    print("SMOKE_TEST_BLANK_TAIL_PADDING_OK")


if __name__ == "__main__":
    main()

"""Release gate for OpenSpec 10.1--10.3 large-sheet Excel fidelity.

This is deliberately a *disposable-copy* harness.  ``C:\\GM15`` is input
only: every copy, mutation, save candidate, manifest and report is created
beneath a directory returned by :func:`_test_temp_utils.make_temp_dir`.

It complements the focussed smoke tests instead of replacing them.  The gate
keeps the evidence needed to reject a performance optimisation which happens
to be fast but changes an exact comparison target or damages an OOXML package.

Examples::

    python _large_sheet_excel_fidelity_gate.py --list-cases
    python _large_sheet_excel_fidelity_gate.py --case fixture --variant Skill:value --out gate.json
    python _large_sheet_excel_fidelity_gate.py --case corpus-shard --variant 1/5 --out corpus-reopen.json

``--real-excel`` is opt-in because it starts Excel through the application's
existing COM validation hook.  A missing COM installation is a non-PASS case
result and can never be silently counted as a successful Excel reopen.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import shutil
import stat
import sys
import tempfile
import time
import traceback
import zipfile
from dataclasses import dataclass
from pathlib import Path
from types import SimpleNamespace
from typing import Callable
from xml.etree import ElementTree as ET

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula

import sow_merge_tool as sm
from _large_sheet_direct_oracle import capture as capture_direct_legacy
from _large_sheet_oracle_fixtures import REAL_FIXTURES, REAL_SOURCE_ROOT, copy_real_fixture
from _large_sheet_snapshot_oracle import capture_legacy, compare_manifests
from _test_temp_utils import make_temp_dir


_MAIN_NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
_DOC_REL_NS = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
_PACKAGE_REL_NS = "{http://schemas.openxmlformats.org/package/2006/relationships}"
_SUPPORTED_SUFFIXES = {".xlsx", ".xlsm"}
_CASE_BUDGET_SECONDS = 90.0
_OUTER_CASE_TIMEOUT_SECONDS = 115.0
_CASE_FINALIZE_RESERVE_SECONDS = 3.0
_WORKSPACE_ROOT = Path(os.path.abspath(__file__)).parent


@dataclass(frozen=True)
class MutationTarget:
    kind: str
    row: int
    col: int
    marker: str


@dataclass(frozen=True)
class FidelityCasePlan:
    """One natural fidelity case selected before any copy or workbook work."""

    case: str
    variant: str | None
    input_paths: tuple[Path, ...]
    work_items: tuple[tuple[str, str | None], ...]


@dataclass(frozen=True)
class _ApprovedReportDestination:
    """An unused JSON destination proven safe before a case starts."""

    raw_path: Path
    path: Path
    partial_path: Path
    approved_parent: Path


class _CaseDeadlineExceeded(TimeoutError):
    """One natural gate case consumed its single 90-second budget."""


def _remaining_case_seconds(
    absolute_deadline: float | None,
    label: str,
    *,
    reserve_seconds: float = 0.0,
) -> float | None:
    """Return the remaining shared case budget or fail before a new phase."""
    if absolute_deadline is None:
        return None
    remaining = float(absolute_deadline) - time.monotonic() - float(reserve_seconds)
    if remaining <= 0:
        raise _CaseDeadlineExceeded(f"case deadline expired before {label}")
    return remaining


def _case_temp_parent() -> Path:
    configured = (os.environ.get("SOW_TEST_TMPDIR") or "").strip()
    base = Path(configured) if configured else _WORKSPACE_ROOT / "tmp" / "test_tmp"
    return base.resolve(strict=False)


def _is_safe_owned_case_root(root: Path) -> bool:
    """Accept only the direct root allocated for this gate's disposable case."""
    root = Path(root)
    try:
        configured_parent = _case_temp_parent()
        resolved_root = root.resolve(strict=False)
        resolved_real_source = REAL_SOURCE_ROOT.resolve()
        return (
            root.is_absolute()
            and root.name.startswith("sow_large_sheet_excel_fidelity_")
            and root.parent.resolve(strict=False) == configured_parent
            and configured_parent.is_dir()
            and not _has_link_or_reparse_component(configured_parent)
            and root.is_dir()
            and not _has_link_or_reparse_component(root)
            and not resolved_root.is_relative_to(resolved_real_source)
        )
    except (OSError, ValueError):
        return False


def _validate_owned_case_root(root: Path) -> None:
    root = Path(root)
    if not _is_safe_owned_case_root(root):
        raise AssertionError(f"unsafe disposable case root: {root}")


def _cleanup_owned_case_root(root: Path) -> dict:
    """Remove exactly the validated direct case root; never scan by prefix."""
    root = Path(root)
    if not _is_safe_owned_case_root(root):
        raise AssertionError(f"refusing cleanup for unsafe disposable case root: {root}")
    shutil.rmtree(root)
    absent = not os.path.lexists(str(root))
    if not absent:
        raise AssertionError(f"disposable case root retained: {root}")
    return {"disposable_root_deleted": True, "disposable_root_absent": True}


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _is_within(path: Path, parent: Path) -> bool:
    try:
        path.relative_to(parent)
        return True
    except ValueError:
        return False


def _component_is_link_or_reparse(path: Path) -> bool:
    """Treat every Windows link/junction/reparse component as unsafe."""
    try:
        if not path.exists() or path.is_symlink():
            return True
        is_junction = getattr(path, "is_junction", None)
        if callable(is_junction) and is_junction():
            return True
        attributes = getattr(path.lstat(), "st_file_attributes", None)
        reparse_flag = getattr(stat, "FILE_ATTRIBUTE_REPARSE_POINT", 0x0400)
        return isinstance(attributes, int) and bool(attributes & reparse_flag)
    except (OSError, ValueError):
        return True


def _has_link_or_reparse_component(path: Path) -> bool:
    """Reject aliases rather than resolving an output through them."""
    path = Path(path)
    if not path.is_absolute():
        return True
    current = Path(path.anchor)
    if _component_is_link_or_reparse(current):
        return True
    for part in path.parts[1:]:
        current /= part
        if _component_is_link_or_reparse(current):
            return True
    return False


def _approved_report_parents() -> tuple[Path, ...]:
    """Resolve only direct, non-reparse report parents on both path chains."""
    raw_parents = (
        _WORKSPACE_ROOT / "benchmark_results",
        Path(tempfile.gettempdir()),
    )
    resolved_parents = []
    for raw_parent in raw_parents:
        if _has_link_or_reparse_component(raw_parent):
            raise ValueError("approved report parent contains a symlink, junction, or reparse point")
        try:
            resolved_parent = raw_parent.resolve(strict=True)
        except (OSError, RuntimeError) as exc:
            raise ValueError(f"approved report parent cannot be resolved: {raw_parent}") from exc
        if _has_link_or_reparse_component(resolved_parent):
            raise ValueError("resolved report parent contains a symlink, junction, or reparse point")
        resolved_parents.append(resolved_parent)
    return tuple(resolved_parents)


def _approve_report_destination(
    raw_value: str | os.PathLike[str], *, forbidden_paths: tuple[Path, ...]
) -> _ApprovedReportDestination:
    """Fail closed before a report can create or replace any file.

    Reports are deliberately limited to a direct workspace ``benchmark_results``
    child or the direct system Temp directory. Neither a symlink/alias nor an
    existing destination/partial is accepted, which keeps success and error
    reporting away from source, export, and revision inputs.
    """
    raw_path = Path(raw_value)
    if not raw_path.is_absolute():
        raise ValueError("--out must be an absolute path")
    if raw_path.suffix.lower() != ".json":
        raise ValueError("--out must use the .json suffix")
    if raw_path.exists():
        raise FileExistsError(f"--out must name a new report: {raw_path}")
    if _has_link_or_reparse_component(raw_path.parent):
        raise ValueError("--out parent cannot contain a symlink, junction, or alias")
    if not raw_path.parent.is_dir():
        raise ValueError("--out parent must already exist")

    approved_parents = _approved_report_parents()
    try:
        resolved_parent = raw_path.parent.resolve(strict=True)
    except (OSError, RuntimeError) as exc:
        raise ValueError("--out parent cannot be resolved") from exc
    if _has_link_or_reparse_component(resolved_parent):
        raise ValueError("resolved --out parent cannot contain a symlink, junction, or alias")
    if resolved_parent not in approved_parents:
        allowed = ", ".join(str(path) for path in approved_parents)
        raise ValueError(f"--out parent must be exactly one of: {allowed}")
    resolved_path = raw_path.resolve(strict=False)
    if resolved_path.parent != resolved_parent:
        raise ValueError("--out resolve changed its approved parent")

    forbidden = tuple(path.resolve(strict=False) for path in forbidden_paths)
    forbidden += (REAL_SOURCE_ROOT.resolve(strict=False),)
    if any(resolved_path == path for path in forbidden):
        raise ValueError("--out may not replace a source, input, or runtime path")
    if _is_within(resolved_path, REAL_SOURCE_ROOT.resolve(strict=False)):
        raise ValueError("--out may not be within a revision/export source path")

    partial_path = resolved_path.with_name(f"{resolved_path.name}.{os.getpid()}.partial")
    if partial_path.exists():
        raise FileExistsError(f"report partial path already exists: {partial_path}")
    if _has_link_or_reparse_component(partial_path.parent):
        raise ValueError("report partial parent cannot contain a symlink, junction, or alias")
    return _ApprovedReportDestination(raw_path, resolved_path, partial_path, resolved_parent)


def _write_json_atomic(destination: _ApprovedReportDestination, payload: dict) -> None:
    """Write one complete case report without exposing a partial JSON file."""
    if not isinstance(destination, _ApprovedReportDestination):
        raise TypeError("report destination must be approved before writing")
    report_path = destination.path
    partial = destination.partial_path
    if _has_link_or_reparse_component(destination.raw_path.parent):
        raise ValueError("approved report raw parent changed to a symlink, junction, or alias")
    if _has_link_or_reparse_component(report_path.parent):
        raise ValueError("approved report resolved parent changed to a symlink, junction, or alias")
    if report_path.parent.resolve(strict=True) != destination.approved_parent:
        raise ValueError("approved report parent changed after validation")
    if report_path.exists() or partial.exists():
        raise FileExistsError("approved report destination was occupied before write")
    try:
        with partial.open("x", encoding="utf-8", newline="\n") as stream:
            json.dump(payload, stream, ensure_ascii=False, sort_keys=True, indent=2)
        # Hard-link publication fails if another writer created the final name.
        # It is therefore atomic without os.replace's overwrite behavior.
        os.link(partial, report_path)
        partial.unlink()
    finally:
        if partial.exists():
            partial.unlink()


def _path_fact(path: Path) -> dict:
    path = Path(path)
    return {
        "path": str(path.resolve()),
        "size_bytes": int(path.stat().st_size),
        "sha256": _sha256(path),
    }


def _input_facts(paths: tuple[Path, ...]) -> list[dict]:
    """Record only immutable source identities for a single natural case."""
    return [
        _path_fact(path)
        for path in paths
    ]


def _safe_name(value: str) -> str:
    return "".join(char if char.isalnum() else "_" for char in value)[:80]


def _close_workbook(workbook) -> None:
    if workbook is None:
        return
    try:
        workbook.close()
    finally:
        archive = getattr(workbook, "vba_archive", None)
        if archive is not None:
            archive.close()


def _worksheet_member(members: dict[str, bytes], sheet: str) -> str:
    workbook = ET.fromstring(members["xl/workbook.xml"])
    relation_id = next(
        node.attrib[_DOC_REL_NS + "id"]
        for node in workbook.iter(_MAIN_NS + "sheet")
        if node.attrib.get("name") == sheet
    )
    relationships = ET.fromstring(members["xl/_rels/workbook.xml.rels"])
    target = next(
        node.attrib["Target"]
        for node in relationships.iter(_PACKAGE_REL_NS + "Relationship")
        if node.attrib.get("Id") == relation_id
    )
    return "xl/" + target.lstrip("/")


def _rewrite_member(path: Path, member: str, payload: bytes) -> None:
    temporary = path.with_suffix(path.suffix + ".fidelity-rewrite")
    with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(temporary, "w", zipfile.ZIP_DEFLATED) as target:
        for info in source.infolist():
            target.writestr(info, payload if info.filename == member else source.read(info.filename))
    os.replace(temporary, path)


def _set_formula_cache(path: Path, sheet: str, row: int, col: int, formula: str, cached: object) -> None:
    """Set a formula and its cached ``<v>`` without saving the source workbook."""
    with zipfile.ZipFile(path, "r") as source:
        members = {name: source.read(name) for name in source.namelist()}
    member = _worksheet_member(members, sheet)
    root = ET.fromstring(members[member])
    ref = f"{get_column_letter(col)}{row}"
    cell = next((node for node in root.iter(_MAIN_NS + "c") if node.attrib.get("r") == ref), None)
    if cell is None:
        data = root.find(_MAIN_NS + "sheetData")
        if data is None:
            data = ET.SubElement(root, _MAIN_NS + "sheetData")
        row_node = next((node for node in data.findall(_MAIN_NS + "row") if int(node.attrib.get("r", "0")) == row), None)
        if row_node is None:
            row_node = ET.SubElement(data, _MAIN_NS + "row", {"r": str(row)})
        cell = ET.SubElement(row_node, _MAIN_NS + "c", {"r": ref})
    cell.attrib.pop("t", None)
    formula_node = cell.find(_MAIN_NS + "f")
    if formula_node is None:
        formula_node = ET.SubElement(cell, _MAIN_NS + "f")
    formula_node.attrib.clear()
    formula_node.text = formula.lstrip("=")
    value_node = cell.find(_MAIN_NS + "v")
    if value_node is None:
        value_node = ET.SubElement(cell, _MAIN_NS + "v")
    value_node.text = str(cached)
    _rewrite_member(path, member, ET.tostring(root, encoding="utf-8", xml_declaration=True))


def _sheet_dimensions(path: Path, sheet: str) -> tuple[int, int]:
    workbook = load_workbook(path, read_only=True, data_only=False, keep_vba=path.suffix.lower() == ".xlsm")
    try:
        ws = workbook[sheet]
        return max(1, int(ws.max_row or 1)), max(1, int(ws.max_column or 1))
    finally:
        _close_workbook(workbook)


def _value_mutation_target(rows: int, cols: int, marker: str) -> MutationTarget:
    """Return the established value-mutation coordinate without a workbook write."""
    return MutationTarget(
        "value",
        min(max(3, int(rows or 3)), 5),
        min(max(1, int(cols or 1)), 2),
        marker,
    )


def _select_overlay_target(path: Path, sheet: str, marker: str) -> MutationTarget:
    """Select the operations target from an immutable fixture copy."""
    rows, cols = _sheet_dimensions(path, sheet)
    return _value_mutation_target(rows, cols, marker)


def _mutate_value(path: Path, sheet: str, marker: str) -> MutationTarget:
    workbook = load_workbook(path, data_only=False, keep_vba=path.suffix.lower() == ".xlsm")
    try:
        ws = workbook[sheet]
        target = _value_mutation_target(ws.max_row, ws.max_column, marker)
        ws.cell(target.row, target.col).value = marker
        workbook.save(path)
    finally:
        _close_workbook(workbook)
    return target


def _mutate_formula_cache(path: Path, sheet: str, marker: str) -> MutationTarget:
    rows, cols = _sheet_dimensions(path, sheet)
    row, col = min(max(3, rows), 5), min(max(1, cols), 2)
    # Formula text and cached value must both change.  A numeric cache is
    # intentional: it verifies the paired value/formula readers separately.
    _set_formula_cache(path, sheet, row, col, "=40+2", 42)
    return MutationTarget("formula_cache", row, col, marker)


def _mutate_row(path: Path, sheet: str, marker: str) -> MutationTarget:
    workbook = load_workbook(path, data_only=False, keep_vba=path.suffix.lower() == ".xlsm")
    try:
        ws = workbook[sheet]
        semantic_row, _semantic_col = _worksheet_semantic_bounds(ws)
        row = int(semantic_row) + 1
        if any(
            int(existing_row) == row
            and getattr(cell, "value", None) is not None
            and not (
                isinstance(getattr(cell, "value", None), str)
                and getattr(cell, "value", None) == ""
            )
            for (existing_row, _existing_col), cell in getattr(ws, "_cells", {}).items()
        ):
            raise AssertionError("row mutation target already has semantic content")
        # Do not duplicate declared @id values.  A textual key remains a
        # deterministic one-sided structural record even when the source key
        # column is numeric.
        ws.cell(row, 1).value = "__SOW_FIDELITY_ROW_KEY__"
        ws.cell(row, min(max(1, int(ws.max_column or 1)), 2)).value = marker
        workbook.save(path)
        col = min(max(1, int(ws.max_column or 1)), 2)
    finally:
        _close_workbook(workbook)
    return MutationTarget("row", row, col, marker)


def _worksheet_semantic_bounds(ws) -> tuple[int, int]:
    """Return the value/formula horizon without style-only worksheet padding.

    ``Worksheet.max_row``/``max_column`` include allocated cells that contain
    formatting only.  A structural fidelity mutation must be adjacent to real
    worksheet content, not beyond a styled empty tail.  Formula text is stored
    in ``Cell.value`` by the editable reader and is therefore semantic content.
    """
    semantic_cells = [
        (int(row), int(col))
        for (row, col), cell in getattr(ws, "_cells", {}).items()
        if (
            getattr(cell, "value", None) is not None
            and not (
                isinstance(getattr(cell, "value", None), str)
                and getattr(cell, "value", None) == ""
            )
        )
    ]
    if not semantic_cells:
        raise AssertionError("column mutation requires semantic worksheet content")
    return (
        max(row for row, _col in semantic_cells),
        max(col for _row, col in semantic_cells),
    )


def _mutate_column(path: Path, sheet: str, marker: str) -> MutationTarget:
    workbook = load_workbook(path, data_only=False, keep_vba=path.suffix.lower() == ".xlsm")
    try:
        ws = workbook[sheet]
        semantic_row, semantic_col = _worksheet_semantic_bounds(ws)
        col = int(semantic_col) + 1
        if any(
            int(existing_col) == col
            and getattr(cell, "value", None) is not None
            and not (
                isinstance(getattr(cell, "value", None), str)
                and getattr(cell, "value", None) == ""
            )
            for (_existing_row, existing_col), cell in getattr(ws, "_cells", {}).items()
        ):
            raise AssertionError("column mutation target already has semantic content")
        ws.cell(1, col).value = f"fidelity_{marker}@pm"
        ws.cell(2, col).value = "string"
        row = min(max(3, int(semantic_row)), 5)
        ws.cell(row, col).value = marker
        workbook.save(path)
    finally:
        _close_workbook(workbook)
    return MutationTarget("column", row, col, marker)


_MUTATORS: tuple[tuple[str, Callable[[Path, str, str], MutationTarget]], ...] = (
    ("value", _mutate_value),
    ("formula", _mutate_formula_cache),
    ("row", _mutate_row),
    ("column", _mutate_column),
)
_MUTATORS_BY_LABEL = dict(_MUTATORS)
_FIXTURE_VARIANTS = tuple((*_MUTATORS_BY_LABEL, "operations"))


def _snapshot_cache_key(path: Path, sheet: str, side: str) -> tuple[str, str, str]:
    """Return one case-local immutable snapshot key without content aliasing."""
    return (
        os.path.normcase(os.path.realpath(os.path.abspath(os.fspath(path)))),
        str(sheet),
        str(side),
    )


def _manifest_snapshot(
    path: Path,
    sheet: str,
    side: str,
    *,
    absolute_deadline: float | None,
    snapshot_cache: dict[tuple[str, str, str], sm.SheetSnapshot] | None,
) -> sm.SheetSnapshot:
    """Stream, or fail-closed reuse, one immutable case-local snapshot."""
    _remaining_case_seconds(absolute_deadline, f"snapshot {side} ingress for {sheet}")
    if snapshot_cache is None:
        return sm._stream_selected_sheet_snapshot(str(path), str(path), sheet, side)
    if not isinstance(snapshot_cache, dict):
        raise TypeError("snapshot_cache must be a dictionary when supplied")
    key = _snapshot_cache_key(path, sheet, side)
    entry = snapshot_cache.get(key)
    if entry is not None:
        current_version = sm._selected_sheet_snapshot_version(str(path))
        if (
            not isinstance(entry, sm.SheetSnapshot)
            or entry.side != str(side)
            or entry.sheet != str(sheet)
            or entry.version != current_version
        ):
            raise AssertionError(
                f"case-local snapshot cache identity/version mismatch: {key}"
            )
        return entry
    snapshot = sm._stream_selected_sheet_snapshot(str(path), str(path), sheet, side)
    current_version = sm._selected_sheet_snapshot_version(str(path))
    if (
        not isinstance(snapshot, sm.SheetSnapshot)
        or snapshot.side != str(side)
        or snapshot.sheet != str(sheet)
        or snapshot.version != current_version
    ):
        raise AssertionError(f"completed snapshot identity/version mismatch: {key}")
    snapshot_cache[key] = snapshot
    return snapshot


def _snapshot_manifest(
    mine: Path,
    theirs: Path,
    base: Path | None,
    sheet: str,
    *,
    absolute_deadline: float | None = None,
    snapshot_cache: dict[tuple[str, str, str], sm.SheetSnapshot] | None = None,
) -> tuple[dict, bool]:
    _remaining_case_seconds(absolute_deadline, f"snapshot ingestion for {sheet}")
    mine_snapshot = _manifest_snapshot(
        mine,
        sheet,
        "A",
        absolute_deadline=absolute_deadline,
        snapshot_cache=snapshot_cache,
    )
    theirs_snapshot = _manifest_snapshot(
        theirs,
        sheet,
        "B",
        absolute_deadline=absolute_deadline,
        snapshot_cache=snapshot_cache,
    )
    base_snapshot = (
        _manifest_snapshot(
            base,
            sheet,
            "BASE",
            absolute_deadline=absolute_deadline,
            snapshot_cache=snapshot_cache,
        )
        if base is not None else None
    )
    result = sm._compare_selected_sheet_snapshots(mine_snapshot, theirs_snapshot, base_snapshot)
    _remaining_case_seconds(absolute_deadline, f"snapshot comparison for {sheet}")
    return sm.snapshot_comparison_oracle_manifest(mine_snapshot, theirs_snapshot, result, base_snapshot), bool(result.unresolved)


def _direct_legacy_manifest(
    mine: Path,
    theirs: Path,
    base: Path | None,
    sheet: str,
    *,
    absolute_deadline: float | None = None,
) -> dict:
    """Fast immutable-input legacy reference; suitable for focused local runs."""
    if base is not None:
        # The direct helper intentionally supports two way only.  Three-way
        # needs the frozen UI worker because it is the authoritative base map.
        raise ValueError("three-way fidelity checks require the frozen legacy Oracle")
    _remaining_case_seconds(absolute_deadline, f"direct legacy capture for {sheet}")
    result = capture_direct_legacy(str(mine), str(theirs), sheet)
    _remaining_case_seconds(absolute_deadline, f"direct legacy capture for {sheet}")
    return result


def _stable_record_identity(record: dict) -> str:
    """A manifest record identity that cannot depend on renderer pair ordinal."""
    item = {key: value for key, value in record.items() if key != "pair"}
    return json.dumps(item, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def _stable_only_diff_identities(raw: dict) -> list[str]:
    """Resolve only-diff membership through physical/logical record contents.

    A pair ordinal is display state and may legitimately differ between bounded
    aligners.  Requiring every only-diff ordinal to resolve to a record makes
    an incomplete manifest fail closed instead of accidentally preserving a
    hidden pair dependency.
    """
    by_pair = {int(record["pair"]): record for record in raw.get("records", ())}
    result = []
    for pair in raw.get("only_diff_rows", ()):
        record = by_pair.get(int(pair))
        if record is None:
            raise AssertionError(f"only-diff pair has no complete record: {pair}")
        result.append(_stable_record_identity(record))
    return sorted(result)


def _assert_structural_column_target(manifest: dict, target: MutationTarget, *, side: str) -> None:
    """Prove a column target through its structural logical slot, not a row badge."""
    slots = [
        item for item in manifest.get("columns", ())
        if int(item.get(side) or 0) == int(target.col)
    ]
    if not slots:
        raise AssertionError(f"column target has no {side} physical slot: {target}")
    structural = [str(item.get("state") or "").strip().casefold() for item in slots]
    if not any(state.startswith(("insert", "delete", "reorder")) for state in structural):
        raise AssertionError(f"column target lacks structural state: {target}; {slots!r}")


def _assert_nonactionable_ambiguous_column_marker(
    manifest: dict,
    target: MutationTarget,
    *,
    side: str,
) -> dict:
    """Prove a physical Language column write without claiming an action target."""
    if target.kind != "column" or side != "theirs":
        raise AssertionError(f"expected a Theirs column marker: {target}; {side}")
    slots = [
        item for item in manifest.get("columns", ())
        if int(item.get(side) or 0) == int(target.col)
    ]
    if len(slots) != 1:
        raise AssertionError(f"column marker has non-unique {side} slot: {target}; {slots!r}")
    slot = slots[0]
    logical = int(slot.get("logical") or 0)
    if logical <= 0 or slot.get("mine") is not None or slot.get("base") is not None:
        raise AssertionError(f"column marker has an actionable cross-side slot: {target}; {slot!r}")
    if str(slot.get("state") or "") != "unresolved" or slot.get("ambiguous") is not True:
        raise AssertionError(f"column marker must remain unresolved and ambiguous: {target}; {slot!r}")
    records = [
        record for record in manifest.get("records", ())
        if int(record.get(f"{side}_row") or 0) == int(target.row)
    ]
    if len(records) != 1:
        raise AssertionError(f"column marker has non-unique {side} row: {target}; {records!r}")
    record = records[0]
    if logical not in {int(value) for value in record.get("diff_cols", ())}:
        raise AssertionError(f"column marker logical slot is not a direct diff: {target}; {record!r}")
    token = (record.get("cells") or {}).get(str(logical), {}).get(side)
    expected_token = {
        "present": True,
        "cached": {"type": "str", "value": target.marker},
        "cached_data_type": "s",
        "formula_data_type": "s",
        "formula": None,
    }
    if token != expected_token:
        raise AssertionError(f"column marker token changed: {target}; {token!r}")
    return {
        "target": dict(target.__dict__),
        "side": side,
        "logical": logical,
        "state": str(slot["state"]),
        "ambiguous": bool(slot["ambiguous"]),
        "token": token,
    }


def _assert_target(manifest: dict, target: MutationTarget, *, side: str) -> None:
    """Require the marker to remain attached to the intended physical target."""
    for record in manifest.get("records", ()):
        if record.get(f"{side}_row") != target.row:
            continue
        if target.kind == "row" and record.get("row_structure"):
            if _stable_record_identity(record) not in set(_stable_only_diff_identities(manifest)):
                raise AssertionError("physical row target missing from only-diff membership")
            return
        for logical, cells in (record.get("cells") or {}).items():
            token = cells.get(side) or {}
            text = json.dumps(token, ensure_ascii=False, sort_keys=True)
            if target.marker in text or (target.kind == "formula_cache" and "40+2" in text):
                slot = next((item for item in manifest.get("columns", ()) if int(item["logical"]) == int(logical)), None)
                if slot is None or int(slot.get(side) or 0) != target.col:
                    raise AssertionError((target, side, logical, slot))
                if target.kind == "column":
                    _assert_structural_column_target(manifest, target, side=side)
                    return
                if _stable_record_identity(record) not in set(_stable_only_diff_identities(manifest)):
                    raise AssertionError("physical target missing from only-diff membership")
                return
    raise AssertionError(f"mutation target not found in {side} legacy Oracle: {target}")


def _compare_manifests_by_stable_identity(legacy: dict, candidate: dict) -> dict:
    """Compare Oracle results by durable physical/logical identities.

    ``pair`` is a renderer/display ordinal.  It is not a stable operation
    target: different bounded aligners can place the same one-sided tail row
    at a different ordinal while retaining identical Mine/Base/Theirs physical
    coordinates, cell tokens and only-difference membership.  The release
    gate deliberately compares those durable identities instead.
    """
    def canonical(raw: dict) -> dict:
        records = [_stable_record_identity(record) for record in raw.get("records", ())]
        return {
            "sheet": raw.get("sheet"), "three_way": bool(raw.get("three_way")),
            "columns": raw.get("columns"),
            "records": sorted(records),
            "only_diff_identities": _stable_only_diff_identities(raw),
        }
    expected, actual = canonical(legacy), canonical(candidate)
    mismatches = [key for key in expected if expected[key] != actual[key]]
    return {"exact": not mismatches, "mismatches": mismatches, "legacy": expected, "candidate": actual}

def _assert_direct_pair_parity(
    mine: Path,
    theirs: Path,
    sheet: str,
    label: str,
    *,
    absolute_deadline: float | None = None,
    snapshot_cache: dict[tuple[str, str, str], sm.SheetSnapshot] | None = None,
) -> dict:
    """Compare the paired read-only legacy oracle with the new snapshot path."""
    legacy = _direct_legacy_manifest(
        mine, theirs, None, sheet, absolute_deadline=absolute_deadline
    )
    candidate, unresolved = _snapshot_manifest(
        mine,
        theirs,
        None,
        sheet,
        absolute_deadline=absolute_deadline,
        snapshot_cache=snapshot_cache,
    )
    if unresolved:
        raise AssertionError(f"{label}: snapshot comparison is unresolved; cannot claim parity")
    parity = _compare_manifests_by_stable_identity(legacy, candidate)
    if not parity["exact"]:
        raise AssertionError(f"{label}: direct Oracle mismatch: {parity['mismatches'][:1]}")
    return legacy


def _assert_frozen_three_way_parity(
    mine: Path,
    theirs: Path,
    base: Path,
    sheet: str,
    label: str,
    *,
    timeout: float,
    absolute_deadline: float | None = None,
    snapshot_cache: dict[tuple[str, str, str], sm.SheetSnapshot] | None = None,
) -> tuple[dict, dict]:
    """Compare the candidate against the real frozen Base-anchored legacy view."""
    worker_timeout = _remaining_case_seconds(
        absolute_deadline, f"frozen legacy capture for {sheet}"
    )
    legacy = capture_legacy(SimpleNamespace(
        mine=str(mine), theirs=str(theirs), base=str(base), sheet=sheet,
        timeout=float(timeout if worker_timeout is None else min(float(timeout), worker_timeout)),
        absolute_deadline=absolute_deadline,
    ))
    candidate, unresolved = _snapshot_manifest(
        mine,
        theirs,
        base,
        sheet,
        absolute_deadline=absolute_deadline,
        snapshot_cache=snapshot_cache,
    )
    if unresolved:
        raise AssertionError(f"{label}: three-way snapshot is unresolved")
    parity = _compare_manifests_by_stable_identity(legacy, candidate)
    if not parity["exact"]:
        raise AssertionError(f"{label}: frozen three-way Oracle mismatch: {parity['mismatches'][:1]}")
    return legacy, candidate


def _expected_safe_unresolved_snapshot(
    mine: Path,
    theirs: Path,
    base: Path | None,
    sheet: str,
    *,
    absolute_deadline: float | None = None,
) -> dict:
    """Prove a deliberately unpublishable snapshot without forming a manifest.

    A fixture whose repeated blank schema fields contain data cannot be mapped
    exactly without a positional guess.  Keep the conservative snapshot result
    unpublishable while retaining direct/frozen legacy target evidence.
    """
    _remaining_case_seconds(absolute_deadline, f"safe-unresolved snapshot ingestion for {sheet}")
    mine_snapshot = sm._stream_selected_sheet_snapshot(str(mine), str(mine), sheet, "A")
    theirs_snapshot = sm._stream_selected_sheet_snapshot(str(theirs), str(theirs), sheet, "B")
    base_snapshot = (
        sm._stream_selected_sheet_snapshot(str(base), str(base), sheet, "BASE")
        if base is not None else None
    )
    alignment = sm._align_selected_sheet_snapshots(mine_snapshot, theirs_snapshot)
    if base_snapshot is None:
        duplicate_proof = sm._try_snapshot_duplicate_field_identity_proof(
            mine_snapshot, theirs_snapshot, alignment
        )
    else:
        mine_base_alignment = sm._align_selected_sheet_snapshots(mine_snapshot, base_snapshot)
        theirs_base_alignment = sm._align_selected_sheet_snapshots(theirs_snapshot, base_snapshot)
        duplicate_proof = sm._try_snapshot_duplicate_field_identity_proof(
            mine_snapshot,
            theirs_snapshot,
            alignment,
            base_snapshot,
            mine_base_alignment,
            theirs_base_alignment,
        )
    result = sm._compare_selected_sheet_snapshots(
        mine_snapshot, theirs_snapshot, base_snapshot
    )
    _remaining_case_seconds(absolute_deadline, f"safe-unresolved snapshot comparison for {sheet}")
    if duplicate_proof is not None or not result.unresolved:
        raise AssertionError(
            f"{sheet}: expected non-actionable duplicate-field snapshot result"
        )

    nonblank_groups = []
    for snapshot in (mine_snapshot, theirs_snapshot, base_snapshot):
        if snapshot is None:
            continue
        fields = sm._snapshot_field_groups(snapshot).get(("", ""), ())
        if len(fields) < 2:
            raise AssertionError(f"{sheet}: expected repeated blank schema fields")
        evidence = []
        for field in fields:
            for row in snapshot.rows:
                cell = row.cells[int(field.physical_col) - 1]
                if cell.cached_value not in (None, "") or cell.formula_value not in (None, ""):
                    evidence.append({
                        "row": int(row.physical_row),
                        "col": int(field.physical_col),
                        "cached": cell.cached_value not in (None, ""),
                        "formula": cell.formula_value not in (None, ""),
                    })
                    break
        if not evidence:
            raise AssertionError(f"{sheet}: repeated blank schema group has no content")
        nonblank_groups.append(tuple(evidence))
    return {
        "state": "expected-safe-unresolved",
        "blank_duplicate_proof": "rejected-nonblank-content",
        "nonblank_duplicate_groups": tuple(nonblank_groups),
    }


def _assert_language_expected_safe_unresolved(
    label: str,
    mine: Path,
    theirs: Path,
    base: Path,
    sheet: str,
    target_mine: MutationTarget | None,
    target_theirs: MutationTarget,
    *,
    absolute_deadline: float | None = None,
) -> dict:
    """Keep Language's known ambiguous schema safely non-actionable."""
    expected_kind = {
        "value": "value",
        "formula": "formula_cache",
        "row": "row",
        "column": "column",
    }.get(str(label))
    if expected_kind is None or target_theirs.kind != expected_kind:
        raise AssertionError(f"Language:{label} target kind changed: {target_theirs}")
    if label == "column":
        if target_mine is not None:
            raise AssertionError("Language:column must not create a Mine structural target")
    elif target_mine is None or target_mine.kind != expected_kind:
        raise AssertionError(f"Language:{label} Mine target changed: {target_mine}")
    two_way_legacy = _direct_legacy_manifest(
        base, theirs, None, sheet, absolute_deadline=absolute_deadline
    )
    column_marker = None
    if label == "column":
        column_marker = _assert_nonactionable_ambiguous_column_marker(
            two_way_legacy, target_theirs, side="theirs"
        )
    else:
        _assert_target(two_way_legacy, target_theirs, side="theirs")
    two_way = _expected_safe_unresolved_snapshot(
        base, theirs, None, sheet, absolute_deadline=absolute_deadline
    )
    direct_marker_pairs = 1
    if target_mine is not None:
        mine_target_legacy = _direct_legacy_manifest(
            base, mine, None, sheet, absolute_deadline=absolute_deadline
        )
        # The second direct manifest names its right input "theirs" by ABI.
        # It is marker-only evidence, not a synthetic three-way action claim.
        _assert_target(mine_target_legacy, target_mine, side="theirs")
        direct_marker_pairs += 1
    three_way = _expected_safe_unresolved_snapshot(
        mine, theirs, base, sheet, absolute_deadline=absolute_deadline
    )
    return {
        "two_way": two_way,
        "three_way": three_way,
        "legacy": "two-way-targets-verified-only",
        "direct_marker_pairs": direct_marker_pairs,
        "column_marker": column_marker,
    }


def _assert_dual_column_conflict_blocked(
    mine: Path,
    theirs: Path,
    base: Path,
    sheet: str,
    *,
    timeout: float,
    absolute_deadline: float | None = None,
    snapshot_cache: dict[tuple[str, str, str], sm.SheetSnapshot] | None = None,
) -> None:
    """A dual independent column insertion must stay unresolved and un-actionable."""
    worker_timeout = _remaining_case_seconds(
        absolute_deadline, f"frozen legacy conflict capture for {sheet}"
    )
    legacy = capture_legacy(SimpleNamespace(
        mine=str(mine), theirs=str(theirs), base=str(base), sheet=sheet,
        timeout=float(timeout if worker_timeout is None else min(float(timeout), worker_timeout)),
        absolute_deadline=absolute_deadline,
    ))
    candidate, _result_unresolved = _snapshot_manifest(
        mine,
        theirs,
        base,
        sheet,
        absolute_deadline=absolute_deadline,
        snapshot_cache=snapshot_cache,
    )
    parity = _compare_manifests_by_stable_identity(legacy, candidate)
    if not parity["exact"]:
        raise AssertionError(f"dual column unresolved Oracle mismatch: {parity['mismatches'][:1]}")
    for manifest, label in ((legacy, "legacy"), (candidate, "candidate")):
        blocked_slots = [
            slot for slot in manifest.get("columns", ())
            if str(slot.get("state") or "").casefold() == "unresolved"
            or bool(slot.get("ambiguous"))
        ]
        if not blocked_slots:
            raise AssertionError(f"dual column conflict lacks {label} blocked structural slot")
    # No target assertion is allowed here: unresolved slots must remain blocked.


def _assert_stable_identity_contract() -> None:
    """Positive/negative proof that pair renumbering cannot mask bad targets."""
    record = {
        "pair": 7, "mine_row": 5, "base_row": 5, "theirs_row": 5,
        "row_structure": False, "diff_cols": [2], "base_diff_cols": [2], "conflicts": [2],
        "cells": {"2": {"mine": {"present": True, "cached": {"value": "mine"}},
                           "base": {"present": True, "cached": {"value": "base"}},
                           "theirs": {"present": True, "cached": {"value": "theirs"}}}},
    }
    legacy = {"sheet": "S", "three_way": True, "columns": [{"logical": 2, "mine": 2, "base": 2, "theirs": 2}], "records": [record], "only_diff_rows": [7]}
    renumbered = {**legacy, "records": [{**record, "pair": 91}], "only_diff_rows": [91]}
    assert _compare_manifests_by_stable_identity(legacy, renumbered)["exact"]
    wrong_row = {**renumbered, "records": [{**renumbered["records"][0], "mine_row": 6}]}
    assert not _compare_manifests_by_stable_identity(legacy, wrong_row)["exact"]
    wrong_token = {**renumbered, "records": [{**renumbered["records"][0], "cells": {"2": {"mine": {"present": True, "cached": {"value": "WRONG"}}, "base": {"present": True, "cached": {"value": "base"}}, "theirs": {"present": True, "cached": {"value": "theirs"}}}}}]}
    assert not _compare_manifests_by_stable_identity(legacy, wrong_token)["exact"]
    wrong_col = {**renumbered, "columns": [{"logical": 2, "mine": 2, "base": 2, "theirs": 3, "state": "insert"}]}
    assert not _compare_manifests_by_stable_identity(legacy, wrong_col)["exact"]
    wrong_kind = {**renumbered, "columns": [{"logical": 2, "mine": 2, "base": 2, "theirs": 2, "state": "reorder"}]}
    assert not _compare_manifests_by_stable_identity(legacy, wrong_kind)["exact"]


def _package_probe(
    path: Path, *, real_excel: bool, absolute_deadline: float | None = None
) -> dict:
    """Read-only package/openpyxl/optional Excel reopening inspection."""
    _remaining_case_seconds(absolute_deadline, f"package probe for {path.name}")
    with zipfile.ZipFile(path, "r") as archive:
        bad = archive.testzip()
        names = sorted(archive.namelist())
        relationship_parts = [name for name in names if name.endswith(".rels")]
        has_vba = "xl/vbaProject.bin" in names
        comments = [name for name in names if name.startswith("xl/comments") or name.startswith("xl/threadedComments/")]
        links = [name for name in names if name.startswith("xl/externalLinks/")]
    workbook = values = None
    try:
        # Normal-mode metadata is required here only for a disposable no-op copy;
        # comparison paths remain read-only and are tested separately.
        kwargs = {"read_only": False, "data_only": False, "keep_vba": path.suffix.lower() == ".xlsm"}
        workbook = load_workbook(path, **kwargs)
        values = load_workbook(path, read_only=True, data_only=True, keep_vba=path.suffix.lower() == ".xlsm")
        sheets = []
        for sheet in workbook.sheetnames:
            formula_count = 0
            comment_count = hyperlink_count = 0
            ws = workbook[sheet]
            value_ws = values[sheet]
            for formula_row, value_row in zip(ws.iter_rows(), value_ws.iter_rows()):
                for formula_cell, value_cell in zip(formula_row, value_row):
                    if isinstance(formula_cell.value, str) and formula_cell.value.startswith("="):
                        formula_count += 1
                    comment_count += int(getattr(formula_cell, "comment", None) is not None)
                    hyperlink_count += int(getattr(formula_cell, "hyperlink", None) is not None)
            sheets.append({
                "sheet": sheet, "rows": int(ws.max_row or 1), "cols": int(ws.max_column or 1),
                "formulas": formula_count, "comments": comment_count, "hyperlinks": hyperlink_count,
                "row_dimensions": len(ws.row_dimensions), "column_dimensions": len(ws.column_dimensions),
            })
    finally:
        _close_workbook(workbook)
        _close_workbook(values)
    excel = "not_requested"
    if real_excel:
        reopen_kwargs = {}
        if absolute_deadline is not None:
            reopen_kwargs["absolute_deadline"] = float(absolute_deadline)
        excel = "ok" if sm._excel_reopen_validate(str(path), **reopen_kwargs) else "failed"
    _remaining_case_seconds(absolute_deadline, f"package probe for {path.name}")
    if bad is not None:
        raise AssertionError(f"invalid ZIP member {bad}: {path}")
    if excel == "failed":
        raise AssertionError(f"Excel cannot reopen disposable package: {path}")
    return {
        "zip_valid": True, "parts": len(names), "relationship_parts": relationship_parts,
        "vba": has_vba, "comment_parts": comments, "external_link_parts": links,
        "excel_reopen": excel, "sheets": sheets,
    }


def _corpus_sources() -> tuple[Path, ...]:
    """Return the deterministic supported-input order for one corpus shard."""
    sources = tuple(sorted(
        path for path in REAL_SOURCE_ROOT.rglob("*")
        if path.is_file() and not path.name.startswith("~$") and path.suffix.lower() in _SUPPORTED_SUFFIXES
    ))
    if not sources:
        raise FileNotFoundError(f"no supported workbooks under {REAL_SOURCE_ROOT}")
    return sources


def _no_op_corpus_shard(
    root: Path,
    *,
    ordinal: int,
    real_excel: bool,
    absolute_deadline: float | None = None,
) -> dict:
    """Validate exactly one disposable no-op corpus member."""
    sources = _corpus_sources()
    assert 1 <= ordinal <= len(sources), (ordinal, len(sources))
    source = sources[ordinal - 1]
    source_hash = _sha256(source)
    target = root / "no-op" / f"{ordinal:04d}_{source.name}"
    target.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, target)
    probe = _package_probe(
        target, real_excel=real_excel, absolute_deadline=absolute_deadline
    )
    if _sha256(source) != source_hash:
        raise AssertionError(f"source changed during no-op validation: {source}")
    print(f"NO_OP_OK {ordinal}/{len(sources)} {source.name}", flush=True)
    return {
        "ordinal": ordinal,
        "total": len(sources),
        "source": str(source),
        "copy": str(target),
        "source_sha256": source_hash,
        **probe,
    }


def _special_formula_fidelity(
    root: Path,
    *,
    timeout: float,
    real_excel: bool,
    absolute_deadline: float | None = None,
) -> dict:
    """Exercise array, data-table and external formulas through both Oracles.

    This is deliberately synthetic because the real corpus has no stable
    guarantee that every special formula flavor is present.  It remains a
    disposable workbook and uses the exact same snapshot/legacy/save paths as
    the real-fixture checks.
    """
    root.mkdir(parents=True, exist_ok=True)

    def make_book(path: Path, *, array_text: str, table_r1: str, external: str) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Special"
        worksheet.append(["id@id", "array", "table", "external"])
        worksheet.append(["int", "int", "int", "string"])
        worksheet.append([1, ArrayFormula("B3:B4", array_text), None, external])
        worksheet.append([2, None, None, None])
        worksheet.append([3, None, DataTableFormula("C5:C6", ca=True, dt2D=True, r1=table_r1, r2="A4"), None])
        worksheet.append([4, None, None, None])
        workbook.save(path)
        workbook.close()

    base, mine, theirs = root / "base.xlsx", root / "mine.xlsx", root / "theirs.xlsx"
    make_book(base, array_text="=A3:A4*2", table_r1="A3", external="=[Other.xlsx]S1!A1")
    make_book(mine, array_text="=A3:A4*2", table_r1="A3", external="=[Other.xlsx]S1!A1")
    make_book(theirs, array_text="=A3:A4*3", table_r1="A4", external="=[Other.xlsx]S1!A2")
    base_hash = _sha256(base)
    two_way = _assert_direct_pair_parity(
        base, theirs, "Special", "special-2way", absolute_deadline=absolute_deadline
    )
    legacy_three, candidate_three = _assert_frozen_three_way_parity(
        mine, theirs, base, "Special", "special-3way", timeout=timeout,
        absolute_deadline=absolute_deadline,
    )
    snapshot = sm._stream_selected_sheet_snapshot(str(theirs), str(theirs), "Special", "theirs")
    assert snapshot.rows[2].cells[1].formula_kind == "special"
    assert snapshot.rows[4].cells[2].formula_kind == "special"
    assert snapshot.rows[2].cells[3].external_link
    # Cell/row adoption of a multi-cell formula is deliberately refused.  A
    # cache-only ZIP update is allowed and must preserve its special formula.
    try:
        sm._ensure_formula_copy_supported(ArrayFormula("B3:B4", "=A3:A4*2"), None)
    except RuntimeError:
        pass
    else:
        raise AssertionError("array formula adoption was not rejected")
    saved = root / "special-cache-only.xlsx"
    sm._build_manual_merge_xlsx_via_zip(
        str(base), str(saved), {("Special", 3, 2): "=A3:A4*2"},
        cached_values={("Special", 3, 2): 42}, cache_only_keys={("Special", 3, 2)},
    )
    probe = _package_probe(
        saved, real_excel=real_excel, absolute_deadline=absolute_deadline
    )
    saved_snapshot = sm._stream_selected_sheet_snapshot(str(saved), str(saved), "Special", "saved")
    assert sm._special_formula_signature(saved_snapshot.rows[2].cells[1].formula_value) == sm._special_formula_signature(
        sm._stream_selected_sheet_snapshot(str(base), str(base), "Special", "base").rows[2].cells[1].formula_value
    )
    assert _sha256(base) == base_hash
    return {
        "two_way": "exact-direct", "three_way": "exact-frozen-legacy",
        "legacy_three_way_conflicts": sum(len(row.get("conflicts", ())) for row in legacy_three.get("records", ())),
        "candidate_three_way_conflicts": sum(len(row.get("conflicts", ())) for row in candidate_three.get("records", ())),
        "array_copy": "rejected", "cache_only_save": "ok", "package": probe,
    }


def _exercise_overlay_save(
    root: Path,
    source: Path,
    sheet: str,
    target: MutationTarget,
    *,
    real_excel: bool,
    absolute_deadline: float | None = None,
) -> dict:
    """Exercise accepted cell/region/undo/redo/save plus safe failure recovery."""
    root.mkdir(parents=True, exist_ok=True)
    baseline_hash = _sha256(source)
    working = root / "operations-source.xlsx"
    output = root / "operations-output.xlsx"
    atomic_target = root / "atomic-target.xlsx"
    shutil.copy2(source, working)
    shutil.copy2(source, atomic_target)
    if _sha256(working) != baseline_hash:
        raise AssertionError("operations working copy diverged before ZIP build")
    before_atomic = _sha256(atomic_target)
    prebuild_excel_reopen = "not_requested"
    if real_excel:
        reopen_kwargs = {}
        if absolute_deadline is not None:
            reopen_kwargs["absolute_deadline"] = float(absolute_deadline)
        if not sm._excel_reopen_validate(str(working), **reopen_kwargs):
            raise AssertionError(f"Excel cannot reopen pre-build fixture: {working}")
        prebuild_excel_reopen = "ok"
    overlay = sm.SheetOperationOverlay()
    deltas = tuple(
        sm.OverlayCellDelta(
            record_key=("fidelity", str(target.row + offset)), field_key=("physical", str(target.col)), side="A",
            physical_row=target.row + offset, physical_col=target.col, before=None,
            after=f"__SOW_FIDELITY_REGION_{offset}__",
        )
        for offset in range(3)
    )
    first = overlay.apply_batch(deltas[:1])
    region = overlay.apply_batch(deltas[1:])
    assert len(overlay.cells) == 3
    overlay.revert_transaction(region)  # undo region
    assert len(overlay.cells) == 1
    overlay.apply_batch(deltas[1:])  # redo region
    assert len(overlay.cells) == 3
    assert overlay.cells[(deltas[0].record_key, deltas[0].field_key, "A")].physical_row == target.row

    manual = {(sheet, delta.physical_row, delta.physical_col): delta.after for delta in deltas}
    sm._build_manual_merge_xlsx_via_zip(str(working), str(output), manual)
    probe = _package_probe(
        output, real_excel=real_excel, absolute_deadline=absolute_deadline
    )
    reopened = load_workbook(output, data_only=False, read_only=True)
    try:
        assert all(reopened[sheet].cell(delta.physical_row, delta.physical_col).value == delta.after for delta in deltas)
    finally:
        _close_workbook(reopened)

    # This is only a ZIP-helper no-op probe.  It is deliberately retained as a
    # narrow safety check, but must never be reported as public atomic recovery
    # or real application-operation coverage.
    failed = sm._build_manual_merge_output_with_openpyxl(
        str(working), str(root / "should-not-exist.xlsx"), {},
        row_ops=[{"kind": "insert_rows", "sheet": "__missing__", "row": 1, "count": 1}],
    )
    assert failed is True  # ignored invalid Sheet is a safe no-op staging run
    assert _sha256(atomic_target) == before_atomic
    assert _sha256(source) == baseline_hash
    assert _sha256(working) == baseline_hash
    return {
        "overlay_unit": "ok", "zip_cell_patch": "ok",
        "actual_app_cell_row_region_column_undo_redo": "not_covered",
        "overlay_generation": overlay.mutation_generation, "manual_targets": sorted(map(str, manual)),
        "atomic_failure_retry": "not_covered", "output": str(output), "package": probe,
        "prebuild_excel_reopen": prebuild_excel_reopen,
        "structural_record": "not_executed",
        "native": "requires actual app-operation/native replay gate",
    }


def _run_fixture(
    root: Path,
    fixture,
    *,
    timeout: float,
    real_excel: bool,
    absolute_deadline: float | None = None,
    variant_labels: tuple[str, ...] | None = None,
    include_operations: bool = True,
) -> dict:
    source_hash = _sha256(fixture.path)
    source, sheet = copy_real_fixture(fixture, root / _safe_name(fixture.name) / "source")
    item = {"fixture": fixture.name, "source": str(fixture.path), "sheet": sheet, "variants": []}
    for label, mutator in _MUTATORS:
        if variant_labels is not None and label not in variant_labels:
            continue
        base = root / _safe_name(fixture.name) / label / "base.xlsx"
        mine = root / _safe_name(fixture.name) / label / "mine.xlsx"
        theirs = root / _safe_name(fixture.name) / label / "theirs.xlsx"
        base.parent.mkdir(parents=True, exist_ok=True)
        for copy in (base, mine, theirs):
            shutil.copy2(source, copy)
        source_copy_hash = _sha256(source)
        target_theirs = mutator(theirs, sheet, f"__SOW_FIDELITY_{fixture.name}_{label}_THEIRS__")
        # Independent column insertions carry incompatible declaration
        # identities and are correctly unresolved in a 3-way merge.  This
        # variant instead needs a deterministic, actionable structural target:
        # Mine remains Base while Theirs owns the inserted logical slot.
        target_mine = None if label == "column" else mutator(
            mine, sheet, f"__SOW_FIDELITY_{fixture.name}_{label}_MINE__"
        )

        # The Language structural fixture deliberately ends in its expected
        # safe-unresolved route and has no dual-conflict Oracle.  Every other
        # column route constructs all four disposable inputs before the first
        # immutable snapshot so its case-local cache can never observe a write.
        mine_conflict = None
        snapshot_cache = None
        if label == "column" and fixture.name != "Language":
            mine_conflict = root / _safe_name(fixture.name) / label / "mine-dual-conflict.xlsx"
            shutil.copy2(source, mine_conflict)
            mutator(
                mine_conflict,
                sheet,
                f"__SOW_FIDELITY_{fixture.name}_{label}_MINE_CONFLICT__",
            )
            snapshot_cache = {}

        if fixture.name == "Language" and label in _MUTATORS_BY_LABEL:
            if _sha256(source) != source_copy_hash:
                raise AssertionError(f"Language:{label} changed the immutable fixture copy")
            expected_unresolved = _assert_language_expected_safe_unresolved(
                label,
                mine,
                theirs,
                base,
                sheet,
                target_mine,
                target_theirs,
                absolute_deadline=absolute_deadline,
            )
            item["variants"].append({
                "kind": label,
                "comparison_mode": "expected-safe-unresolved",
                "two_way": "expected-safe-unresolved",
                "three_way": "expected-safe-unresolved",
                "snapshot": expected_unresolved,
                "theirs_target": target_theirs.__dict__,
                "mine_target": target_mine.__dict__ if target_mine is not None else None,
            })
            print(f"VARIANT_EXPECTED_SAFE_UNRESOLVED {fixture.name} {label}", flush=True)
            continue

        two_way = _assert_direct_pair_parity(
            base,
            theirs,
            sheet,
            f"{fixture.name}-{label}-2way",
            absolute_deadline=absolute_deadline,
            snapshot_cache=snapshot_cache,
        )
        _assert_target(two_way, target_theirs, side="theirs")
        three_way_legacy, three_way = _assert_frozen_three_way_parity(
            mine, theirs, base, sheet, f"{fixture.name}-{label}-3way", timeout=timeout,
            absolute_deadline=absolute_deadline,
            snapshot_cache=snapshot_cache,
        )
        _assert_target(three_way_legacy, target_theirs, side="theirs")
        if target_mine is not None:
            _assert_target(three_way_legacy, target_mine, side="mine")
        _assert_target(three_way, target_theirs, side="theirs")
        if target_mine is not None:
            _assert_target(three_way, target_mine, side="mine")
        if mine_conflict is not None:
            _assert_dual_column_conflict_blocked(
                mine_conflict, theirs, base, sheet, timeout=timeout,
                absolute_deadline=absolute_deadline,
                snapshot_cache=snapshot_cache,
            )
        item["variants"].append({
            "kind": label, "two_way": "exact-direct", "three_way": "exact-frozen-legacy",
            "two_way_conflicts": sum(len(record.get("conflicts", ())) for record in two_way.get("records", ())),
            "three_way_conflicts": sum(len(record.get("conflicts", ())) for record in three_way.get("records", ())),
            "theirs_target": target_theirs.__dict__,
            "mine_target": target_mine.__dict__ if target_mine is not None else None,
        })
        print(f"VARIANT_ORACLE_OK {fixture.name} {label}", flush=True)
    if include_operations:
        source_copy_hash = _sha256(source)
        value_target = _select_overlay_target(
            source, sheet, "__SOW_FIDELITY_OPERATION_TARGET__"
        )
        item["operations"] = _exercise_overlay_save(
            root / _safe_name(fixture.name) / "operations",
            source,
            sheet,
            value_target,
            real_excel=real_excel,
            absolute_deadline=absolute_deadline,
        )
        if _sha256(source) != source_copy_hash:
            raise AssertionError("operations selector changed the fixture copy")
    if _sha256(fixture.path) != source_hash:
        raise AssertionError(f"source changed: {fixture.path}")
    return item


def _run_fixture_variant(
    root: Path,
    fixture,
    variant: str,
    *,
    timeout: float,
    real_excel: bool,
    absolute_deadline: float | None = None,
) -> dict:
    """Run one fixture mutation or the independent overlay/save natural case."""
    assert variant in _FIXTURE_VARIANTS, variant
    if variant == "operations":
        return _run_fixture(
            root,
            fixture,
            timeout=timeout,
            real_excel=real_excel,
            absolute_deadline=absolute_deadline,
            variant_labels=(),
            include_operations=True,
        )
    return _run_fixture(
        root,
        fixture,
        timeout=timeout,
        real_excel=real_excel,
        absolute_deadline=absolute_deadline,
        variant_labels=(variant,),
        include_operations=False,
    )


def _fixture_by_name(name: str):
    return next((item for item in REAL_FIXTURES if item.name == name), None)


def _parse_fixture_variant(value: str):
    fixture_name, separator, variant = str(value or "").partition(":")
    fixture = _fixture_by_name(fixture_name)
    if not separator or fixture is None or variant not in _FIXTURE_VARIANTS:
        choices = ", ".join(
            f"{item.name}:{variant_name}"
            for item in REAL_FIXTURES
            for variant_name in _FIXTURE_VARIANTS
        )
        raise ValueError(f"fixture --variant must be one of: {choices}")
    return fixture, variant


def _parse_corpus_shard(value: str) -> tuple[int, Path]:
    ordinal_text, separator, total_text = str(value or "").partition("/")
    sources = _corpus_sources()
    if not separator:
        raise ValueError("corpus-shard --variant must use <ordinal>/<total>")
    ordinal, total = int(ordinal_text), int(total_text)
    if total != len(sources) or not 1 <= ordinal <= total:
        raise ValueError(
            f"corpus-shard --variant must be within 1/{len(sources)} through "
            f"{len(sources)}/{len(sources)}"
        )
    return ordinal, sources[ordinal - 1]


def _case_catalog() -> dict:
    sources = _corpus_sources()
    return {
        "schema": "large-sheet-excel-fidelity-gate-cases-v1",
        "case_budget_seconds": _CASE_BUDGET_SECONDS,
        "outer_timeout_recommendation_seconds": _OUTER_CASE_TIMEOUT_SECONDS,
        "cases": {
            "special-formulas": [None],
            "fixture": [
                f"{item.name}:{variant}"
                for item in REAL_FIXTURES
                for variant in _FIXTURE_VARIANTS
            ],
            "corpus-shard": [
                f"{ordinal}/{len(sources)}"
                for ordinal in range(1, len(sources) + 1)
            ],
        },
    }


def _plan_case(case: str, variant: str | None) -> FidelityCasePlan:
    """Resolve exactly one natural case without opening or copying a workbook."""
    if case == "special-formulas":
        if variant not in (None, ""):
            raise ValueError("special-formulas does not accept --variant")
        return FidelityCasePlan(
            case=case,
            variant=None,
            input_paths=(),
            work_items=((case, None),),
        )
    if case == "fixture":
        fixture, fixture_variant = _parse_fixture_variant(str(variant or ""))
        return FidelityCasePlan(
            case=case,
            variant=f"{fixture.name}:{fixture_variant}",
            input_paths=(fixture.path,),
            work_items=((case, f"{fixture.name}:{fixture_variant}"),),
        )
    if case == "corpus-shard":
        ordinal, source = _parse_corpus_shard(str(variant or ""))
        normalized = f"{ordinal}/{len(_corpus_sources())}"
        return FidelityCasePlan(
            case=case,
            variant=normalized,
            input_paths=(source,),
            work_items=((case, normalized),),
        )
    raise AssertionError(case)


def _validate_case_timeout(value: float) -> float:
    timeout = float(value)
    if not 0.0 < timeout <= _CASE_BUDGET_SECONDS:
        raise ValueError(f"--timeout must be > 0 and <= {_CASE_BUDGET_SECONDS}")
    return timeout


def _run_single_case(
    root: Path,
    *,
    case: str,
    variant: str | None,
    timeout: float,
    real_excel: bool,
    absolute_deadline: float | None = None,
) -> dict:
    _remaining_case_seconds(absolute_deadline, "case dispatch")
    _assert_stable_identity_contract()
    if case == "special-formulas":
        if variant not in (None, ""):
            raise ValueError("special-formulas does not accept --variant")
        return {
            "stable_identity_contract": "ok",
            "special_formulas": _special_formula_fidelity(
                root / "special-formulas",
                timeout=timeout,
                real_excel=real_excel,
                absolute_deadline=absolute_deadline,
            ),
        }
    if case == "fixture":
        fixture, fixture_variant = _parse_fixture_variant(str(variant or ""))
        return {
            "stable_identity_contract": "ok",
            "fixture": _run_fixture_variant(
                root,
                fixture,
                fixture_variant,
                timeout=timeout,
                real_excel=real_excel,
                absolute_deadline=absolute_deadline,
            ),
        }
    if case == "corpus-shard":
        ordinal, _source = _parse_corpus_shard(str(variant or ""))
        return {
            "stable_identity_contract": "ok",
            "corpus_shard": _no_op_corpus_shard(
                root,
                ordinal=ordinal,
                real_excel=real_excel,
                absolute_deadline=absolute_deadline,
            ),
        }
    raise AssertionError(case)


def main(argv=None) -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--list-cases", action="store_true")
    parser.add_argument(
        "--case",
        choices=("special-formulas", "fixture", "corpus-shard"),
        help="one natural fidelity case; aggregate release runs are intentionally disabled",
    )
    parser.add_argument("--variant", help="case-specific selector from --list-cases")
    parser.add_argument("--real-excel", action="store_true", help="require the application Excel COM reopen hook")
    parser.add_argument("--timeout", type=float)
    parser.add_argument("--out")
    # Keep legacy aggregate flags parseable only to fail closed with an
    # actionable explanation instead of silently creating an over-120s run.
    parser.add_argument("--real", nargs="*", metavar="FIXTURE", help=argparse.SUPPRESS)
    parser.add_argument("--no-op-corpus", action="store_true", help=argparse.SUPPRESS)
    parser.add_argument("--no-fixtures", action="store_true", help=argparse.SUPPRESS)
    args = parser.parse_args(argv)
    if args.list_cases:
        execution_arguments = (
            args.case is not None,
            args.variant is not None,
            args.real_excel,
            args.timeout is not None,
            args.out is not None,
            args.real is not None,
            args.no_op_corpus,
            args.no_fixtures,
        )
        if any(execution_arguments):
            parser.error("--list-cases cannot be combined with any execution selector or option")
        print(json.dumps(_case_catalog(), ensure_ascii=False, sort_keys=True, indent=2))
        return
    if args.real is not None or args.no_op_corpus or args.no_fixtures:
        parser.error("aggregate selectors are disabled; use exactly one --case/--variant")
    if not args.case or not args.out:
        parser.error("one --case and --out are required; use --list-cases first")
    try:
        case_timeout = _validate_case_timeout(
            _CASE_BUDGET_SECONDS if args.timeout is None else args.timeout
        )
    except (TypeError, ValueError) as exc:
        parser.error(str(exc))

    started = time.monotonic()
    case_deadline = started + case_timeout
    work_deadline = case_deadline - _CASE_FINALIZE_RESERVE_SECONDS
    try:
        _remaining_case_seconds(work_deadline, "case planning")
        plan = _plan_case(args.case, args.variant)
        _remaining_case_seconds(work_deadline, "case planning")
    except (TypeError, ValueError) as exc:
        parser.error(str(exc))
    inputs = plan.input_paths
    runtime_sources = (
        Path(sm.__file__).resolve(),
        Path(__file__).resolve(),
    )
    try:
        destination = _approve_report_destination(
            args.out,
            forbidden_paths=(*inputs, *runtime_sources),
        )
    except (OSError, ValueError) as exc:
        parser.error(str(exc))
    before = _input_facts(inputs)
    runtime_before = _input_facts(runtime_sources)
    _remaining_case_seconds(work_deadline, "input identity capture")
    root: Path | None = None
    report = {
        "schema": "large-sheet-excel-fidelity-gate-case-v4",
        "case": plan.case,
        "variant": plan.variant,
        "planned_work_items": [list(item) for item in plan.work_items],
        "case_budget_seconds": case_timeout,
        "case_deadline_seconds": round(case_deadline - started, 3),
        "finalize_reserve_seconds": _CASE_FINALIZE_RESERVE_SECONDS,
        "outer_timeout_recommendation_seconds": _OUTER_CASE_TIMEOUT_SECONDS,
        "real_excel_requested": bool(args.real_excel),
        "inputs_before": before,
        "runtime_sources_before": runtime_before,
    }
    failure = None
    failure_traceback = None
    try:
        _remaining_case_seconds(work_deadline, "disposable root allocation")
        root = Path(make_temp_dir("sow_large_sheet_excel_fidelity_"))
        report["disposable_root"] = str(root)
        _validate_owned_case_root(root)
        _remaining_case_seconds(work_deadline, "disposable root validation")
        report["result"] = _run_single_case(
            root,
            case=args.case,
            variant=args.variant,
            timeout=case_timeout,
            real_excel=bool(args.real_excel),
            absolute_deadline=work_deadline,
        )
        _remaining_case_seconds(work_deadline, "case result finalization")
        report["status"] = "PASS"
    except BaseException as exc:
        failure = exc
        failure_traceback = exc.__traceback__
        report["status"] = "FAILED"
        report["error"] = f"{type(exc).__name__}: {exc}"
        report["traceback"] = traceback.format_exc(limit=30)
    finally:
        try:
            after = _input_facts(inputs)
            report["inputs_after"] = after
            if after != before:
                raise AssertionError(("source input changed", before, after))
            runtime_after = _input_facts(runtime_sources)
            report["runtime_sources_after"] = runtime_after
            if runtime_after != runtime_before:
                raise AssertionError(("runtime source changed", runtime_before, runtime_after))
        except BaseException as cleanup_exc:
            report.setdefault("secondary_failures", []).append(
                f"input-integrity: {type(cleanup_exc).__name__}: {cleanup_exc}"
            )
            if failure is None:
                failure = cleanup_exc
                failure_traceback = cleanup_exc.__traceback__
                report["status"] = "FAILED"
                report["error"] = f"{type(cleanup_exc).__name__}: {cleanup_exc}"
                report["traceback"] = traceback.format_exc(limit=30)
        if root is not None:
            try:
                report["cleanup"] = _cleanup_owned_case_root(root)
            except BaseException as cleanup_exc:
                report.setdefault("secondary_failures", []).append(
                    f"disposable-cleanup: {type(cleanup_exc).__name__}: {cleanup_exc}"
                )
                if failure is None:
                    failure = cleanup_exc
                    failure_traceback = cleanup_exc.__traceback__
                    report["status"] = "FAILED"
                    report["error"] = f"{type(cleanup_exc).__name__}: {cleanup_exc}"
                    report["traceback"] = traceback.format_exc(limit=30)
        else:
            report["cleanup"] = {"disposable_root_deleted": True, "disposable_root_absent": True}
        elapsed = time.monotonic() - started
        try:
            if elapsed > case_timeout:
                raise _CaseDeadlineExceeded(
                    f"case budget exceeded after cleanup: {elapsed:.3f}s > {case_timeout:.3f}s"
                )
        except BaseException as cleanup_exc:
            report.setdefault("secondary_failures", []).append(
                f"elapsed-budget: {type(cleanup_exc).__name__}: {cleanup_exc}"
            )
            if failure is None:
                failure = cleanup_exc
                failure_traceback = cleanup_exc.__traceback__
                report["status"] = "FAILED"
                report["error"] = f"{type(cleanup_exc).__name__}: {cleanup_exc}"
                report["traceback"] = traceback.format_exc(limit=30)
        report["elapsed_seconds"] = round(elapsed, 3)
        _write_json_atomic(destination, report)
    print(
        f"LARGE_SHEET_EXCEL_FIDELITY_CASE_{report['status']} "
        f"case={plan.case} variant={plan.variant or '-'} out={destination.path}"
    )
    if failure is not None:
        raise failure.with_traceback(failure_traceback)


if __name__ == "__main__":
    main()

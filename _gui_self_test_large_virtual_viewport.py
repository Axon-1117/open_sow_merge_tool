"""B3 immutable public 2D viewport regressions."""
from __future__ import annotations

import argparse
from contextlib import contextmanager
import hashlib
import json
import math
import os
import sys
import tempfile
import time
from pathlib import Path
from types import SimpleNamespace

from openpyxl import Workbook
from openpyxl.worksheet._read_only import ReadOnlyWorksheet
from openpyxl.worksheet.worksheet import Worksheet

import sow_merge_tool as sm

_CASE_20K = "20k-viewport-actions"
_CASE_WIDE = "wide-3way-2d-public-viewport"
_CASES = (_CASE_20K, _CASE_WIDE)
_ROWS = 20_000
_WIDE_ROWS = 2200
_SCHEMA = 2
_WIDE_DATA_COLUMNS = 56
_CASE_TIMEOUT = 90.0
_REQUEST_TIMEOUT = 2.0


def _sha(path):
    digest = hashlib.sha256()
    with Path(path).open("rb") as stream:
        for block in iter(lambda: stream.read(1 << 20), b""):
            digest.update(block)
    return digest.hexdigest()


def _setting(path):
    file_path = Path(path)
    return (True, file_path.read_bytes()) if file_path.exists() else (False, b"")


def _settings_snapshot(path):
    """Return serializable, read-only evidence for the case-local setting."""
    file_path = Path(path)
    assert file_path.is_file(), file_path
    contents = file_path.read_bytes()
    content = json.loads(contents.decode("utf-8"))
    assert isinstance(content, dict), content
    return {
        "mtime_ns": int(file_path.stat().st_mtime_ns),
        "content_sha256": hashlib.sha256(contents).hexdigest(),
        "content": content,
    }


def _wait_for_only_diff_settings_flush(app, path, before, deadline, label):
    """Keep the one-second settings debounce outside the scored route window."""
    after = None

    def flushed():
        nonlocal after
        candidate = _settings_snapshot(path)
        if (
            int(candidate["mtime_ns"]) > int(before["mtime_ns"])
            and candidate["content"].get("only_diff") == 0
        ):
            after = candidate
            return True
        return False

    _wait(app.root, flushed, deadline, label)
    assert after is not None
    return {"before": before, "after": after}


def _book(path, side):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "S1"
    sheet.append(["id@id", "value"])
    sheet.append(["int32", "string"])
    middle = _ROWS // 2
    for row_id in range(1, _ROWS + 1):
        value = f"value-{row_id}" if side == "mine" or row_id == middle else f"value-{row_id}-changed"
        sheet.append([row_id, value])
    workbook.save(path)
    workbook.close()


def _wide_book(path, side):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "S1"
    headers = ["id@id"] + [f"c{column:02d}" for column in range(1, _WIDE_DATA_COLUMNS + 1)]
    types = ["int32"] + ["string"] * _WIDE_DATA_COLUMNS
    if side != "base":
        # An explicit Mine/Base/Theirs structural slot: Mine and Theirs own it,
        # while Base intentionally has no physical column for its logical slot.
        headers.append("base-struct-slot")
        types.append("string")
    sheet.append(headers)
    sheet.append(types)
    for row_id in range(1, _WIDE_ROWS + 1):
        row = [row_id] + [f"v{column:02d}-{row_id}" for column in range(1, _WIDE_DATA_COLUMNS + 1)]
        if side != "base":
            row.append(f"struct-{row_id}")
        sheet.append(row)
    workbook.save(path)
    workbook.close()


def _pump(root):
    root.update()
    root.update_idletasks()


def _wait(root, predicate, deadline, label, fail=None):
    while time.monotonic() < deadline:
        root.update()
        if callable(fail) and (failure := fail()):
            raise AssertionError(f"{label}: {failure}")
        if predicate():
            return
        root.update_idletasks()
        if callable(fail) and (failure := fail()):
            raise AssertionError(f"{label}: {failure}")
        if predicate():
            return
        time.sleep(0.005)
    raise AssertionError(f"timeout {label}")


def _canon(value):
    if value is None or isinstance(value, (str, int, float, bool, bytes)):
        return value
    if isinstance(value, dict):
        return tuple(sorted(((_canon(key), _canon(item)) for key, item in value.items()), key=repr))
    if isinstance(value, (tuple, list, set, frozenset)):
        items = tuple(_canon(item) for item in value)
        return tuple(sorted(items, key=repr)) if isinstance(value, (set, frozenset)) else items
    return ("opaque", type(value).__name__, id(value))


def _digest(value):
    return hashlib.sha256(repr(_canon(value)).encode()).hexdigest()


def _fact(value):
    try:
        size = len(value)
    except Exception:
        size = None
    return {"type": type(value).__name__, "len": size, "digest": _digest(value), "preview": repr(_canon(value))[:128]}


def _diff(before, after, path="hard"):
    if before == after:
        return []
    if isinstance(before, dict) and isinstance(after, dict):
        changes = []
        for key in sorted(set(before) | set(after), key=repr):
            child = f"{path}[{key!r}]"
            if key not in before or key not in after:
                changes.append({"path": child, "before": _fact(before.get(key)), "after": _fact(after.get(key))})
            else:
                changes.extend(_diff(before[key], after[key], child))
        return changes
    return [{"path": path, "before": _fact(before), "after": _fact(after)}]


def _same(before, after, label):
    changes = _diff(before, after)
    assert not changes, f"{label}: immutable mutation {json.dumps(changes, ensure_ascii=False, sort_keys=True)}"


def _hard(app, view, inputs):
    overlay = (app.sheet_operation_overlays or {}).get(view.sheet)
    cache = view.column_comparison_cache
    model = cache.model
    projection = view.column_projection
    operations = tuple(getattr(app, name) for name in (
        "manual_a_cell_ops", "manual_b_cell_ops", "manual_a_formula_cache_ops", "manual_b_formula_cache_ops",
        "manual_a_row_ops", "manual_b_row_ops", "manual_a_column_ops", "manual_b_column_ops",
        "manual_sheet_ops", "auto_sheet_ops",
    ))
    return {
        "input": tuple(sorted((name, _sha(path)) for name, path in inputs.items())),
        "ops": _canon(operations),
        "undo": _canon((app.undo_stack, app.redo_stack)),
        "modified": _canon((app.modified_a, app.modified_b, app.modified_sheets_a, app.modified_sheets_b, app.user_touched_conflicts, view.touched_rows)),
        "overlay": _canon((getattr(overlay, "cells", None), getattr(overlay, "topology_generation", None), getattr(overlay, "mutation_generation", None))),
        "prepared": _canon((view.row_pairs, view.row_a_to_pair_idx, view.row_b_to_pair_idx, view.mine_to_base_row, view.theirs_to_base_row, view.pair_base_row_override, view.pair_raw_parts_a, view.pair_raw_parts_b, view.pair_raw_parts_base, view.pair_diff_cols, view.pair_base_diff_cols, view._prepared_complete, view._data_ready, view._row_model_exact, view._pair_diff_full_exact, view._base_diff_full_exact)),
        "model": _canon((id(cache), id(model), id(projection), projection.model is model, tuple((slot.logical_idx, slot.mine_col, slot.base_col, slot.theirs_col, str(slot.state)) for slot in model.slots), tuple(projection.block_ordinal_by_slot), tuple(cache.structural_diff_cols), tuple(cache.unresolved_cols))),
        "gens": _canon((app._sheet_compute_generation.get(view.sheet), view._row_model_version, view._column_model_version, view._column_projection_generation, view._virtual_column_window_generation, view.max_row, view.max_col, view.col_max_a, view.col_max_b, view.col_max_base)),
        "handles": tuple((name, id(getattr(app, name, None)), type(getattr(app, name, None)).__name__, getattr(getattr(app, name, None), "read_only", None)) for name in ("_wb_a_edit", "_wb_b_edit", "_wb_base_edit")),
    }


def _wide_hard(app, view, inputs):
    # Window start/generation, display strings and selection are presentation;
    # every immutable model/projection/operation field remains locked.
    fact = _hard(app, view, inputs)
    generations = tuple(fact["gens"])
    fact["gens"] = generations[:4] + generations[5:]
    return fact


def _cancel_settings(app):
    if app is None:
        return
    for view in app.sheet_views.values():
        if view is None:
            continue
        for attr in ("_settings_save_id", "_hover_debounce_id", "_diff_map_debounce_id"):
            after_id = getattr(view, attr, None)
            if after_id is not None:
                view.frame.after_cancel(after_id)
                setattr(view, attr, None)


@contextmanager
def _traps(app, view):
    hits = []
    originals = []

    def bad(name):
        def fail(*_args, **_kwargs):
            hits.append(name)
            raise AssertionError(f"immutable viewport accessed {name}")
        return fail

    targets = (
        (app, "ws_a_val"), (app, "ws_b_val"), (app, "ws_base_val"),
        (app, "ws_a_edit"), (app, "ws_b_edit"), (app, "ws_base_edit"),
        (app, "_request_edit_preload"), (app, "_ensure_edit_loaded"),
        (app, "_load_edit_workbooks_owned"), (app, "_start_background_thread"),
        (app, "_atomic_save"), (app, "_atomic_save_with_retry"),
        (app, "_atomic_replace_file_with_retry"), (app, "_try_alt_save"),
        (app, "build_manual_b_output_file"), (app, "build_manual_merge_output_file"),
        (app, "save_a_inplace"), (app, "save_b_inplace"), (app, "save_merged_and_exit"),
        (view, "refresh"), (view, "_refresh_mode_switch_preserving_selection"),
        (view, "_start_async_large_only_diff_build"), (view, "_run_copy_action_by_mode"),
        (view, "_apply_global_sheet_overwrite"), (view, "_apply_selected_column_block"),
        (sm, "_align_selected_sheet_snapshots"), (sm, "_compare_selected_sheet_snapshots"), (sm, "_atomic_save_wb"),
    )
    old_cell, old_iter, old_ro_iter = Worksheet.cell, Worksheet.iter_rows, ReadOnlyWorksheet.iter_rows
    try:
        for owner, name in targets:
            if hasattr(owner, name):
                originals.append((owner, name, getattr(owner, name)))
                setattr(owner, name, bad(f"{type(owner).__name__}.{name}"))
        Worksheet.cell = bad("Worksheet.cell")
        Worksheet.iter_rows = bad("Worksheet.iter_rows")
        ReadOnlyWorksheet.iter_rows = bad("ReadOnlyWorksheet.iter_rows")
        yield hits
    finally:
        Worksheet.cell, Worksheet.iter_rows, ReadOnlyWorksheet.iter_rows = old_cell, old_iter, old_ro_iter
        for owner, name, original in reversed(originals):
            setattr(owner, name, original)


@contextmanager
def _publisher(app, view):
    calls = []
    original = view._publish_prepared_cache_surface

    def wrap(*args, **kwargs):
        entry = dict(app._sheet_exact_entry(view.sheet) or {})
        record = {
            "before": {
                "sheet": view.sheet, "selected": app.selected_sheet, "gen": app._sheet_compute_generation[view.sheet],
                "exact": entry.get("generation"), "ticket": view._mode_switch_seq,
                "requested": view._mode_switch_requested_value, "pending": view._mode_switch_pending,
            },
            "rows": tuple(kwargs.get("prepared_rows") or ()),
        }
        calls.append(record)
        record["result"] = original(*args, **kwargs)
        return record["result"]

    view._publish_prepared_cache_surface = wrap
    try:
        yield calls
    finally:
        view._publish_prepared_cache_surface = original


def _terminal(view, request_id):
    terminals = [dict(record) for record in tuple(getattr(view, "_viewport_request_terminal", ()) or ()) if int(record.get("id", -1)) == request_id]
    return terminals[-1] if terminals else None


def _route(app, view, name, callback, generation, deadline):
    sequence = int(getattr(view, "_viewport_request_seq", 0))
    started = time.perf_counter()
    callback()
    callback_ms = (time.perf_counter() - started) * 1000.0
    assert callback_ms <= 33.0, (name, callback_ms)
    active = dict(getattr(view, "_viewport_request_active", {}) or {})
    request_id = active.get("id")
    if request_id is None or int(request_id) <= sequence:
        return {"route": name, "kind": "no-request", "callback_ms": callback_ms}
    request_id = int(request_id)
    _wait(app.root, lambda: _terminal(view, request_id) is not None, min(deadline, time.monotonic() + _REQUEST_TIMEOUT), f"{name}/{request_id}")
    record = _terminal(view, request_id)
    assert record and record.get("status") == "complete" and record.get("reason") == name, record
    assert int(record.get("generation", -1)) == generation and record.get("selected_sheet") == view.sheet, record
    assert int(record.get("row_start", -1)) == int(record.get("actual_row_start", -2)), record
    assert bool(record.get("counted")) == bool(record.get("surface_changed")), record
    if record.get("counted"):
        publication = dict(record.get("publication") or {})
        context = dict(publication.get("render_context") or {})
        assert int(publication.get("request_id", -1)) == request_id, record
        assert context.get("enabled") is True and context.get("fallback") is False and int(context.get("readiness_scan_count", -1)) == 0, record
    else:
        assert float(record.get("elapsed_ms", 9999)) <= 33.0, record
    record.update(route=name, kind="complete", callback_ms=callback_ms)
    return record


def _route_diagnostic(record):
    """Bounded terminal evidence without serializing full viewport rows."""
    publication = dict(record.get("publication") or {})
    context = dict(publication.get("render_context") or {})
    return {
        "route": record.get("route"),
        "position": record.get("position"),
        "id": record.get("id"),
        "reason": record.get("reason"),
        "generation": record.get("generation"),
        "target": {
            "row_start": record.get("row_start"),
            "column_start": record.get("column_start"),
        },
        "actual": {
            "row_start": record.get("actual_row_start"),
            "column_start": record.get("actual_column_start"),
        },
        "callback_ms": round(float(record.get("callback_ms", 0.0)), 3),
        "e2e_ms": round(float(record.get("elapsed_ms", 0.0)), 3),
        "queue_wait_ms": round(float(record.get("queue_wait_ms", 0.0)), 3),
        "publish_ms": round(float(record.get("publish_ms", 0.0)), 3),
        "phases": {
            "publish": dict(record.get("publish_phases_ms") or {}),
            "publication": dict(publication.get("phase_ms") or {}),
        },
        "publication": {
            "request_id": publication.get("request_id"),
            "window": dict(publication.get("window") or {}),
            "render_context": {
                "enabled": context.get("enabled"),
                "fallback": context.get("fallback"),
                "readiness_scan_count": context.get("readiness_scan_count"),
            },
        },
        "counted": bool(record.get("counted")),
        "surface_changed": bool(record.get("surface_changed")),
    }


def _mode(app, view, value):
    return bool(
        int(view.only_diff_var.get()) == value and not view._mode_switch_pending and not view._only_diff_async_building
        and app._is_sheet_exact_current(view.sheet) and app._sheet_exact_entry(view.sheet).get("full_detail_terminal")
        and view._prepared_complete and view._data_ready and view._row_model_exact
        and view._derive_lifecycle_state() == "EDIT_DEFERRED" and not app._edit_workbooks_ready()
        and view.only_diff_cb.cget("state") == "normal"
    )


def _minimap(view, fraction):
    height = max(1, int(view.vdiff_map.winfo_height()))
    view._on_vdiff_map_click(SimpleNamespace(y=max(1, min(height, int(round(fraction * height))))))


def _assert_settings_path(actual, expected):
    assert os.fspath(actual) == os.fspath(expected), (actual, expected)


def _assert_inputs(inputs, before):
    assert {name: _sha(path) for name, path in inputs.items()} == before


def _assert_setting(path, state):
    assert _setting(path) == state


def _run_20k():
    start = time.monotonic()
    deadline = start + _CASE_TIMEOUT
    original_settings = os.fspath(sm._SETTINGS_PATH)
    original_state = _setting(original_settings)
    app = None
    primary = None
    root = None
    try:
        with tempfile.TemporaryDirectory(prefix="sow_virtual_viewport_") as raw:
            root = Path(raw)
            mine, theirs, settings = root / "mine.xlsx", root / "theirs.xlsx", root / "settings.json"
            _book(mine, "mine")
            _book(theirs, "theirs")
            inputs = {"mine": mine, "theirs": theirs}
            before_inputs = {name: _sha(path) for name, path in inputs.items()}
            settings.write_text(json.dumps({"only_diff": 0}) + "\n", encoding="utf-8")
            sm._SETTINGS_PATH = os.fspath(settings)
            print(f"START {_CASE_20K}", flush=True)
            try:
                app = sm.SowMergeApp(os.fspath(mine), os.fspath(theirs))
                _wait(app.root, lambda: (view := app.sheet_views.get("S1")) is not None and app.selected_sheet == "S1" and app._is_sheet_exact_current("S1") and app._sheet_exact_entry("S1").get("full_detail_terminal") and view._prepared_complete and view._data_ready and view._row_model_exact and view._derive_lifecycle_state() == "EDIT_DEFERRED" and not app._edit_workbooks_ready() and not view._pending_exact_render and not view._virtual_publishing and view.only_diff_cb.cget("state") == "normal" and not view.only_diff_var.get(), deadline, "immutable exact/full")
                view = app.sheet_views["S1"]
                assert len(view.row_pairs) == _ROWS + _SCHEMA and len(view._full_display_rows) == _ROWS + _SCHEMA and view._virtual_mode_active() and len(view.display_rows) <= view._virtual_viewport_row_capacity()
                hard = _hard(app, view, inputs)
                requests = tuple(app._edit_load_requests)
                generation = int(app._sheet_compute_generation[view.sheet])
                print("VIEWPORT_STAGE immutable-public-cache-only", flush=True)
                with _traps(app, view) as hits, _publisher(app, view) as publications:
                    enable_ticket = view._mode_switch_seq + 1
                    only_rows = tuple(view._only_diff_rows_with_touched(view._only_diff_rows_cache))
                    assert only_rows and view._has_valid_only_diff_snapshot_cache()
                    started_callback = time.perf_counter()
                    view.only_diff_cb.invoke()
                    assert (time.perf_counter() - started_callback) * 1000 <= 33
                    _wait(app.root, lambda: _mode(app, view, 1), deadline, "public full-to-onlydiff")
                    assert len(publications) == 1 and publications[0]["result"] is True and publications[0]["rows"] == only_rows and publications[0]["before"] == {"sheet": view.sheet, "selected": view.sheet, "gen": generation, "exact": generation, "ticket": enable_ticket, "requested": 1, "pending": True}, publications
                    assert tuple(view._full_display_rows) == only_rows
                    _same(hard, _hard(app, view, inputs), "0-to-1")
                    assert tuple(app._edit_load_requests) == requests and not app._edit_workbooks_ready()
                    disable_ticket = view._mode_switch_seq + 1
                    settings_before_disable = _settings_snapshot(settings)
                    started_callback = time.perf_counter()
                    view.only_diff_cb.invoke()
                    assert (time.perf_counter() - started_callback) * 1000 <= 33
                    _wait(app.root, lambda: _mode(app, view, 0), deadline, "public onlydiff-to-full")
                    full_rows = tuple(range(len(view.row_pairs)))
                    assert len(publications) == 2 and publications[1]["result"] is True and publications[1]["rows"] == full_rows and publications[1]["before"] == {"sheet": view.sheet, "selected": view.sheet, "gen": generation, "exact": generation, "ticket": disable_ticket, "requested": 0, "pending": True}, publications
                    assert tuple(view._full_display_rows) == full_rows and len(view.display_rows) <= view._virtual_viewport_row_capacity()
                    _same(hard, _hard(app, view, inputs), "1-to-0")
                    assert tuple(app._edit_load_requests) == requests and not app._edit_workbooks_ready()
                    settings_flush = _wait_for_only_diff_settings_flush(
                        app, settings, settings_before_disable, deadline,
                        "20k only-diff settings flush",
                    )
                    heartbeat_before = int(getattr(app, "_ui_heartbeat_samples", 0) or 0)
                    _wait(app.root, lambda: int(getattr(app, "_ui_heartbeat_samples", 0) or 0) >= heartbeat_before + 2, deadline, "heartbeat prime")
                    render_cursor = len(view._viewport_render_samples_ms)
                    heartbeat_cursor = len(app._ui_heartbeat_gaps_ms)
                    heartbeat_counter = int(getattr(app, "_ui_heartbeat_samples", 0) or 0)
                    records = {name: [] for name in ("vthumb", "vpage", "wheel", "vminimap", "diff-block")}
                    route_log = []
                    changed = []
                    primary_groups = (("0.08", .60), ("0.28", .18), ("0.48", .76), ("0.68", .34), ("0.86", .94))
                    extra_groups = (("0.12", .52), ("0.32", .22), ("0.52", .82), ("0.72", .42), ("0.92", .98))
                    def group(thumb, mini):
                        routes = (
                            _route(app, view, "vthumb", lambda fraction=thumb: view._yview_both("moveto", fraction), generation, deadline),
                            _route(app, view, "vpage", lambda: view._yview_both("scroll", "1", "pages"), generation, deadline),
                            _route(app, view, "wheel", lambda: view._on_mousewheel(SimpleNamespace(delta=-120, num=None)), generation, deadline),
                            _route(app, view, "vminimap", lambda fraction=mini: _minimap(view, fraction), generation, deadline),
                        )
                        for item in routes:
                            records[item["route"]].append(item)
                            route_log.append(item)
                            if item.get("counted") and item.get("surface_changed"):
                                changed.append(item)
                        return routes[0]
                    cold_vthumb = group(*primary_groups[0])
                    for thumb, mini in primary_groups[1:]:
                        group(thumb, mini)
                    home = _route(app, view, "vthumb", lambda: view._yview_both("moveto", "0.0"), generation, deadline)
                    records["vthumb"].append(home)
                    route_log.append(home)
                    assert home.get("counted") and home.get("surface_changed")
                    changed.append(home)
                    blocks = view._ensure_full_diff_blocks()
                    assert len(blocks) == 2 and view.next_diff_btn.cget("state") == "normal"
                    active = view._active_full_diff_block_index()
                    assert active is not None and active + 1 < len(blocks)
                    expected = int(blocks[active + 1].start_pair_idx)
                    next_diff = _route(app, view, "diff-block", view.next_diff_btn.invoke, generation, deadline)
                    assert next_diff.get("counted") and next_diff.get("surface_changed") and int(view.selected_pair_idx) == expected, next_diff
                    records["diff-block"].append(next_diff)
                    route_log.append(next_diff)
                    changed.append(next_diff)
                    for thumb, mini in extra_groups:
                        group(thumb, mini)
                    _wait(app.root, lambda: int(getattr(app, "_ui_heartbeat_samples", 0) or 0) >= heartbeat_counter + 3, deadline, "heartbeat suffix")
                    assert not hits, hits
                    assert tuple(app._edit_load_requests) == requests and not app._edit_workbooks_ready()
                    _same(hard, _hard(app, view, inputs), "routes")
                    _assert_inputs(inputs, before_inputs)
                    for name in ("vthumb", "vpage", "wheel", "vminimap"):
                        assert any(item.get("counted") and item.get("surface_changed") for item in records[name]), (name, records[name])
                    actions = [float(item["elapsed_ms"]) for item in changed]
                    callbacks = [float(item.get("callback_ms", float("nan"))) for item in route_log]
                    renders = list(view._viewport_render_samples_ms)[render_cursor:]
                    beats = list(app._ui_heartbeat_gaps_ms)[heartbeat_cursor:]
                    heartbeat_end = int(getattr(app, "_ui_heartbeat_samples", 0) or 0)
                    diagnostic = {"changed": len(changed), "action": {"count": len(actions), "p95": round(app._p95(actions), 3), "max": round(max(actions), 3), "samples": [round(item, 3) for item in actions]}, "render": {"count": len(renders), "p95": round(app._p95(renders), 3), "max": round(max(renders), 3), "samples": [round(float(item), 3) for item in renders]}, "heartbeat": {"count": len(beats), "p95": round(app._p95(beats), 3), "max": round(max(beats), 3), "samples": [round(float(item), 3) for item in beats], "delta": heartbeat_end - heartbeat_counter}, "measurement": {"settings_flush": settings_flush, "heartbeat_baseline": {"before": heartbeat_before, "after_prime": heartbeat_counter, "required_delta": 2}, "cursors": {"render": render_cursor, "heartbeat": heartbeat_cursor}}, "tickets": [enable_ticket, disable_ticket], "generation": generation, "routes": [_route_diagnostic(item) for item in route_log], "route_counts": {name: len(items) for name, items in records.items()}}
                    print("VIEWPORT_ACTION_DIAGNOSTICS " + json.dumps(diagnostic, sort_keys=True), flush=True)
                    assert len(changed) >= 42 and len(route_log) == 42 and len(callbacks) == 42 and all(math.isfinite(item) and 0 <= item <= 33 for item in callbacks), diagnostic
                    assert cold_vthumb.get("counted") and cold_vthumb.get("surface_changed") and float(cold_vthumb["elapsed_ms"]) <= 66, diagnostic
                    assert next_diff.get("counted") and next_diff.get("surface_changed") and float(next_diff["elapsed_ms"]) <= 66, diagnostic
                    assert all(math.isfinite(item) and item > 0 for item in actions + renders), diagnostic
                    assert app._p95(actions) <= 33 and max(actions) <= 66 and len(renders) >= len(changed) and app._p95(renders) <= 33 and max(renders) <= 66, diagnostic
                    assert heartbeat_end - heartbeat_counter >= 3 and len(beats) >= 3 and app._p95(beats) <= 200 and max(beats) <= 200, diagnostic
            except BaseException as exc:
                primary = exc
                raise
            finally:
                errors = []
                def check(name, callback):
                    try:
                        callback()
                    except BaseException as exc:
                        errors.append((name, exc))
                check("input", lambda: _assert_inputs(inputs, before_inputs))
                check("settings path", lambda: _assert_settings_path(sm._SETTINGS_PATH, settings))
                check("settings debounce", lambda: _cancel_settings(app))
                check("shutdown", lambda: app._shutdown_root() if app else None)
                check("user settings", lambda: _assert_setting(original_settings, original_state))
                if errors:
                    text = "; ".join(f"{name}: {type(exc).__name__}: {exc}" for name, exc in errors)
                    if primary is not None:
                        primary.add_note("cleanup secondary: " + text)
                    else:
                        raise AssertionError(text)
    finally:
        sm._SETTINGS_PATH = original_settings
        try:
            _assert_setting(original_settings, original_state)
            if root is not None:
                assert not root.exists(), root
        except BaseException as exc:
            if primary is not None:
                primary.add_note(f"outer cleanup secondary: {type(exc).__name__}: {exc}")
            else:
                raise


def _wide_ready(app, view):
    return bool(
        app.selected_sheet == view.sheet and app._is_sheet_exact_current(view.sheet)
        and app._sheet_exact_entry(view.sheet).get("full_detail_terminal")
        and view._prepared_complete and view._data_ready and view._row_model_exact
        and view._is_exact_immutable_view_ready() and not view._pending_exact_render and not view._virtual_publishing
        and view._derive_lifecycle_state() == "EDIT_DEFERRED" and not app._edit_workbooks_ready()
        and not view.only_diff_var.get()
    )


def _wide_structural_slot(view):
    projection = view.column_projection
    slots = []
    for logical_col in range(1, projection.slot_count + 1):
        slot = projection.slot(logical_col)
        if slot is not None and slot.base_col is None and slot.mine_col is not None and slot.theirs_col is not None:
            slots.append((logical_col, slot))
    assert len(slots) == 1, [(logical, slot.state, slot.mine_col, slot.base_col, slot.theirs_col) for logical, slot in slots]
    logical, slot = slots[0]
    assert logical == projection.slot_count and int(slot.mine_col) == _WIDE_DATA_COLUMNS + 2 and int(slot.theirs_col) == _WIDE_DATA_COLUMNS + 2
    assert projection.is_missing("BASE", logical)
    return logical, slot


def _wide_tail_pair(view):
    target_row = _SCHEMA + _WIDE_ROWS
    candidates = [
        pair_idx for pair_idx, pair in enumerate(view.row_pairs)
        if tuple(pair) == (target_row, target_row)
        and tuple(view.pair_raw_parts_a.get(pair_idx, ()))[:2] == (str(_WIDE_ROWS), "v01-2200")
        and tuple(view.pair_raw_parts_b.get(pair_idx, ()))[:2] == (str(_WIDE_ROWS), "v01-2200")
        and tuple(view.pair_raw_parts_base.get(pair_idx, ()))[:2] == (str(_WIDE_ROWS), "v01-2200")
    ]
    assert len(candidates) == 1, candidates
    return int(candidates[0])


def _event_on_text(root, widget, line, start, end):
    index = f"{line}.{start + 1 if end - start > 1 else start}"
    widget.see(index)
    _pump(root)
    box = widget.bbox(index) or widget.dlineinfo(index)
    assert box is not None, (index, widget.index("@0,0"))
    x, y, width, height = box
    event = SimpleNamespace(x=int(x + max(1, width // 2)), y=int(y + max(1, height // 2)))
    found_line, found_col = map(int, str(widget.index(f"@{event.x},{event.y}")).split("."))
    assert found_line == line and start <= found_col < max(start + 1, end), (found_line, found_col, line, start, end)
    return event


def _wide_hdiff_event(view, fraction):
    canvas = view.hdiff_right
    assert str(canvas.bind("<Button-1>") or "").strip()
    width = max(2, int(canvas.winfo_width()))
    return canvas, max(1, min(width, int(round(float(fraction) * width))))


def _wide_vdiff_event(view):
    canvas = view.vdiff_map
    assert str(canvas.bind("<Button-1>") or "").strip()
    return canvas, max(1, int(canvas.winfo_height()))


def _route_hminimap(app, view, fraction, generation, deadline):
    canvas, x = _wide_hdiff_event(view, fraction)
    return _route(app, view, "hminimap", lambda: canvas.event_generate("<Button-1>", x=x, y=max(1, int(canvas.winfo_height()) // 2)), generation, deadline)


def _route_vminimap_tail(app, view, generation, deadline):
    canvas, y = _wide_vdiff_event(view)
    return _route(app, view, "vminimap", lambda: canvas.event_generate("<Button-1>", x=1, y=y), generation, deadline)


def _wide_headers_match(view, logical_col, slot):
    projection = view.column_projection
    rendered = tuple(view._rendered_logical_columns())
    assert logical_col in rendered and slot.base_col is None and projection.is_missing("BASE", logical_col), (rendered, logical_col, slot)
    labels = {
        "A": projection.header_label("A", logical_col),
        "BASE": projection.header_label("BASE", logical_col),
        "B": projection.header_label("B", logical_col),
        "LOGICAL": projection.header_label("LOGICAL", logical_col),
    }
    assert all(isinstance(label, str) and label for label in labels.values()), labels
    actual = {
        "A": view.left_colhdr.get("1.0", "end-1c"),
        "BASE": view.base_colhdr.get("1.0", "end-1c"),
        "B": view.right_colhdr.get("1.0", "end-1c"),
        "LOGICAL": view.cursor_cmp_colhdr.get("1.0", "end-1c"),
    }
    assert all(labels[side] in actual[side] for side in labels), (labels, actual)
    return {"logical": logical_col, "mine_physical": slot.mine_col, "base_physical": slot.base_col, "theirs_physical": slot.theirs_col, "labels": labels}


def _run_wide_3way():
    deadline = time.monotonic() + _CASE_TIMEOUT
    original_settings = os.fspath(sm._SETTINGS_PATH)
    original_state = _setting(original_settings)
    app = None
    root = None
    primary = None
    try:
        with tempfile.TemporaryDirectory(prefix="sow_wide_3way_viewport_") as raw:
            root = Path(raw)
            mine, theirs, base, settings = root / "mine.xlsx", root / "theirs.xlsx", root / "base.xlsx", root / "settings.json"
            _wide_book(mine, "mine")
            _wide_book(theirs, "theirs")
            _wide_book(base, "base")
            inputs = {"mine": mine, "theirs": theirs, "base": base}
            before_inputs = {name: _sha(path) for name, path in inputs.items()}
            settings.write_text(json.dumps({"only_diff": 0}) + "\n", encoding="utf-8")
            sm._SETTINGS_PATH = os.fspath(settings)
            try:
                print(f"WIDE_3WAY_STAGE open-current-exact case={_CASE_WIDE}", flush=True)
                app = sm.SowMergeApp(os.fspath(mine), os.fspath(theirs), merge_mode=True, base_path=os.fspath(base), initial_sheet="S1")
                _wait(app.root, lambda: (view := app.sheet_views.get("S1")) is not None and _wide_ready(app, view), deadline, "wide immutable exact/full")
                view = app.sheet_views["S1"]
                assert len(view.row_pairs) == _WIDE_ROWS + _SCHEMA and view._virtual_mode_active() and view._wide_column_virtual_active()
                assert int(view._logical_slot_count()) > sm._VIRTUAL_VIEWPORT_MAX_COLUMNS and int(view._logical_slot_count()) >= 48
                structural_logical, structural_slot = _wide_structural_slot(view)
                tail_pair = _wide_tail_pair(view)
                generation = int(app._sheet_compute_generation[view.sheet])
                hard = _wide_hard(app, view, inputs)
                request_baseline = tuple(app._edit_load_requests)
                print("WIDE_3WAY_STAGE public-horizontal-first-middle-last", flush=True)
                with _traps(app, view) as hits:
                    view.only_diff_cb.invoke()
                    _wait(app.root, lambda: _mode(app, view, 1), deadline, "wide public full-to-onlydiff")
                    _same(hard, _wide_hard(app, view, inputs), "wide 0-to-1")
                    assert tuple(app._edit_load_requests) == request_baseline and not app._edit_workbooks_ready()
                    settings_before_disable = _settings_snapshot(settings)
                    view.only_diff_cb.invoke()
                    _wait(app.root, lambda: _mode(app, view, 0), deadline, "wide public onlydiff-to-full")
                    _same(hard, _wide_hard(app, view, inputs), "wide 1-to-0")
                    assert tuple(app._edit_load_requests) == request_baseline and not app._edit_workbooks_ready()
                    settings_flush = _wait_for_only_diff_settings_flush(
                        app, settings, settings_before_disable, deadline,
                        "wide only-diff settings flush",
                    )
                    heartbeat_before = int(getattr(app, "_ui_heartbeat_samples", 0) or 0)
                    _wait(app.root, lambda: int(getattr(app, "_ui_heartbeat_samples", 0) or 0) >= heartbeat_before + 2, deadline, "wide heartbeat prime")
                    render_cursor = len(view._viewport_render_samples_ms)
                    heartbeat_cursor = len(app._ui_heartbeat_gaps_ms)
                    heartbeat_counter = int(getattr(app, "_ui_heartbeat_samples", 0) or 0)
                    routes = []
                    changed = []
                    # Real hdiff-map Button-1 bindings: mid -> last -> first;
                    # each next action changes the bounded logical window.
                    for cycle in range(14):
                        for label, fraction in (("middle", 0.50), ("last", 1.00), ("first", 0.00)):
                            record = _route_hminimap(app, view, fraction, generation, deadline)
                            record["position"] = label
                            routes.append(record)
                            assert record.get("kind") == "complete" and record.get("counted") and record.get("surface_changed"), record
                            changed.append(record)
                            _pump(app.root)
                            _same(hard, _wide_hard(app, view, inputs), f"hminimap-{cycle}-{label}")
                            _assert_inputs(inputs, before_inputs)
                            assert tuple(app._edit_load_requests) == request_baseline and not app._edit_workbooks_ready()
                    # A real tail-row vdiff-map action followed by a real last
                    # horizontal hdiff-map action creates the off-screen 2D window.
                    print("WIDE_3WAY_STAGE public-combined-row-column", flush=True)
                    vertical = _route_vminimap_tail(app, view, generation, deadline)
                    routes.append(vertical)
                    assert vertical.get("kind") == "complete" and vertical.get("counted") and vertical.get("surface_changed"), vertical
                    _pump(app.root)
                    _same(hard, _wide_hard(app, view, inputs), "vminimap-tail")
                    assert tail_pair in tuple(view.display_rows), (tail_pair, tuple(view.display_rows))
                    horizontal = _route_hminimap(app, view, 1.00, generation, deadline)
                    horizontal["position"] = "last-combined"
                    routes.append(horizontal)
                    assert horizontal.get("kind") == "complete" and horizontal.get("counted") and horizontal.get("surface_changed"), horizontal
                    changed.extend((vertical, horizontal))
                    _pump(app.root)
                    _same(hard, _wide_hard(app, view, inputs), "combined-last")
                    assert tail_pair in tuple(view.display_rows)
                    total = int(view._logical_slot_count())
                    column_cap = min(sm._VIRTUAL_VIEWPORT_MAX_COLUMNS, total)
                    assert int(view._virtual_column_window_start) == total - column_cap
                    assert tuple(view._rendered_logical_columns()) == tuple(range(total - column_cap + 1, total + 1))
                    header_identity = _wide_headers_match(view, structural_logical, structural_slot)
                    line = int(view.row_to_line[tail_pair])
                    spans = view._spans_for_line()
                    start, end = spans[structural_logical]
                    assert str(view.right.bind("<Button-1>") or "").strip()
                    main_event = _event_on_text(app.root, view.right, line, start, end)
                    event_line, event_col = map(int, str(view.right.index(f"@{main_event.x},{main_event.y}")).split("."))
                    assert event_line == line and start <= event_col < end, (event_line, event_col, line, start, end)
                    original_widget_line = view._widget_line
                    right_widget_line_calls = []
                    other_widget_line_calls = []

                    def _wide_right_event_line(widget):
                        if widget is view.right:
                            right_widget_line_calls.append(
                                {"event_line": event_line, "event_col": event_col, "returned_line": line}
                            )
                            return line
                        other_widget_line_calls.append(type(widget).__name__)
                        return original_widget_line(widget)

                    callback_started = time.perf_counter()
                    try:
                        # Tk event_generate does not move the physical pointer, while
                        # the legacy selection helper intentionally reads it.  Scope
                        # the deterministic bridge to this one verified right-pane
                        # public Button-1 event; production binding/handler stays live.
                        view._widget_line = _wide_right_event_line
                        view.right.event_generate("<Button-1>", x=main_event.x, y=main_event.y)
                    finally:
                        view._widget_line = original_widget_line
                    assert (time.perf_counter() - callback_started) * 1000.0 <= 33.0
                    _pump(app.root)
                    selection_actual = {
                        "expected": {"pair": tail_pair, "line": line, "logical_col": structural_logical},
                        "actual": {
                            "pair": view.selected_pair_idx,
                            "main_line": view._main_sel_line,
                            "main_col": view._main_sel_col,
                            "cursor_line": view._cursor_cmp_sel_line,
                            "cursor_col": view._cursor_cmp_sel_col,
                            "row_a": view.selected_excel_row_a,
                            "row_b": view.selected_excel_row_b,
                            "row_window": int(view._virtual_window_start),
                            "column_window": int(view._virtual_column_window_start),
                        },
                        "wrapper": {
                            "right_calls": tuple(right_widget_line_calls),
                            "other_calls": tuple(other_widget_line_calls),
                        },
                    }
                    print("WIDE_3WAY_SELECTION=" + json.dumps(selection_actual, ensure_ascii=False, sort_keys=True), flush=True)
                    assert right_widget_line_calls == [{"event_line": event_line, "event_col": event_col, "returned_line": line}]
                    assert not other_widget_line_calls, other_widget_line_calls
                    assert int(view.selected_pair_idx) == tail_pair and int(view._main_sel_line) == line and int(view._main_sel_col) == structural_logical
                    assert int(view._cursor_cmp_sel_line) == 3 and int(view._cursor_cmp_sel_col) == structural_logical
                    assert int(view.selected_excel_row_a) == _SCHEMA + _WIDE_ROWS and int(view.selected_excel_row_b) == _SCHEMA + _WIDE_ROWS
                    selection_window = (int(view._virtual_window_start), int(view._virtual_column_window_start), tuple(view._rendered_logical_columns()))
                    c_click_before = {
                        "window": selection_window,
                        "request_seq": int(view._viewport_request_seq),
                        "terminal": _canon(tuple(view._viewport_request_terminal)),
                        "publication": int(view._virtual_publish_generation),
                    }
                    cursor_spans = view._spans_for_line(view.cursor_cmp.get("1.0", "1.end"))
                    cursor_start, cursor_end = cursor_spans[structural_logical]
                    assert str(view.cursor_cmp.bind("<Button-1>") or "").strip()
                    cursor_event = _event_on_text(app.root, view.cursor_cmp, 1, cursor_start, cursor_end)
                    callback_started = time.perf_counter()
                    view.cursor_cmp.event_generate("<Button-1>", x=cursor_event.x, y=cursor_event.y)
                    assert (time.perf_counter() - callback_started) * 1000.0 <= 33.0
                    _pump(app.root)
                    assert int(view.selected_pair_idx) == tail_pair and int(view._main_sel_col) == structural_logical and int(view._cursor_cmp_sel_col) == structural_logical
                    c_click_after = {
                        "window": (int(view._virtual_window_start), int(view._virtual_column_window_start), tuple(view._rendered_logical_columns())),
                        "request_seq": int(view._viewport_request_seq),
                        "terminal": _canon(tuple(view._viewport_request_terminal)),
                        "publication": int(view._virtual_publish_generation),
                    }
                    print("WIDE_3WAY_C_CLICK_DELTA=" + json.dumps({"before": c_click_before, "after": c_click_after}, ensure_ascii=False, sort_keys=True), flush=True)
                    assert c_click_after == c_click_before, (c_click_before, c_click_after)
                    _same(hard, _wide_hard(app, view, inputs), "public main/C selection")
                    assert tuple(app._edit_load_requests) == request_baseline and not app._edit_workbooks_ready()
                    _assert_inputs(inputs, before_inputs)
                    _wait(app.root, lambda: int(getattr(app, "_ui_heartbeat_samples", 0) or 0) >= heartbeat_counter + 3, deadline, "wide heartbeat suffix")
                    assert not hits, hits
                    actions = [float(record["elapsed_ms"]) for record in changed]
                    callbacks = [float(record["callback_ms"]) for record in routes]
                    renders = list(view._viewport_render_samples_ms)[render_cursor:]
                    beats = list(app._ui_heartbeat_gaps_ms)[heartbeat_cursor:]
                    heartbeat_end = int(getattr(app, "_ui_heartbeat_samples", 0) or 0)
                    diagnostic = {
                        "case": _CASE_WIDE,
                        "structural_slot": header_identity,
                        "tail_pair": tail_pair,
                        "windows": {"row_start": int(view._virtual_window_start), "column_start": int(view._virtual_column_window_start), "columns": tuple(view._rendered_logical_columns())},
                        "measurement": {"settings_flush": settings_flush, "heartbeat_baseline": {"before": heartbeat_before, "after_prime": heartbeat_counter, "required_delta": 2}, "cursors": {"render": render_cursor, "heartbeat": heartbeat_cursor}},
                        "routes": [_route_diagnostic(record) for record in routes],
                        "action": {"count": len(actions), "p95": round(app._p95(actions), 3), "max": round(max(actions), 3), "samples": [round(value, 3) for value in actions]},
                        "render": {"count": len(renders), "p95": round(app._p95(renders), 3), "max": round(max(renders), 3), "samples": [round(float(value), 3) for value in renders]},
                        "heartbeat": {"count": len(beats), "p95": round(app._p95(beats), 3), "max": round(max(beats), 3), "counter_delta": heartbeat_end - heartbeat_counter, "samples": [round(float(value), 3) for value in beats]},
                    }
                    print("WIDE_3WAY_2D_DIAGNOSTICS " + json.dumps(diagnostic, sort_keys=True), flush=True)
                    assert len(changed) >= 40 and len(callbacks) >= 40 and all(math.isfinite(value) and 0.0 <= value <= 33.0 for value in callbacks), diagnostic
                    assert all(math.isfinite(value) and value > 0.0 for value in actions + renders), diagnostic
                    assert app._p95(actions) <= 33.0 and max(actions) <= 66.0, diagnostic
                    assert len(renders) >= len(changed) and app._p95(renders) <= 33.0 and max(renders) <= 66.0, diagnostic
                    assert heartbeat_end - heartbeat_counter >= 3 and len(beats) >= 3 and app._p95(beats) <= 200.0 and max(beats) <= 200.0, diagnostic
            except BaseException as exc:
                primary = exc
                raise
            finally:
                errors = []
                def check(name, callback):
                    try:
                        callback()
                    except BaseException as exc:
                        errors.append((name, exc))
                check("input", lambda: _assert_inputs(inputs, before_inputs))
                check("temporary settings path", lambda: _assert_settings_path(sm._SETTINGS_PATH, settings))
                check("settings debounce", lambda: _cancel_settings(app))
                check("shutdown", lambda: app._shutdown_root() if app else None)
                check("user settings", lambda: _assert_setting(original_settings, original_state))
                if errors:
                    text = "; ".join(f"{name}: {type(exc).__name__}: {exc}" for name, exc in errors)
                    if primary is not None:
                        primary.add_note("cleanup secondary: " + text)
                    else:
                        raise AssertionError(text)
    finally:
        sm._SETTINGS_PATH = original_settings
        active = primary if primary is not None else sys.exc_info()[1]
        def outer_check(label, callback):
            try:
                callback()
            except BaseException as exc:
                if active is not None:
                    active.add_note(f"outer cleanup secondary {label}: {type(exc).__name__}: {exc}")
                else:
                    raise
        outer_check("user settings", lambda: _assert_setting(original_settings, original_state))
        if root is not None:
            outer_check("temporary root removed", lambda: _assert_root_absent(root))


def _assert_root_absent(path):
    assert not path.exists(), path


def main(argv=None):
    parser = argparse.ArgumentParser()
    parser.add_argument("--list-cases", action="store_true")
    parser.add_argument("--case", choices=_CASES)
    args = parser.parse_args(argv)
    if args.list_cases:
        if args.case:
            parser.error("--list-cases cannot be combined with --case")
        for case in _CASES:
            print(case, flush=True)
        return
    selected = (args.case,) if args.case else (_CASE_20K,)
    for case in selected:
        print(f"START {case}", flush=True)
        if case == _CASE_20K:
            _run_20k()
        else:
            _run_wide_3way()
        print(f"CASE_OK {case}", flush=True)
    print(f"SUITE_OK ({len(selected)} cases)", flush=True)


if __name__ == "__main__":
    main()

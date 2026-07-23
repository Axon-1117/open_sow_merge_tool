import os
import sys
import tempfile
import time

from openpyxl import Workbook, load_workbook

import sow_merge_tool as sm


def _make_book(path: str, value: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "id"
    ws["B1"] = "value"
    ws["A2"] = 1
    ws["B2"] = value
    wb.save(path)
    wb.close()


def main():
    startup_updates = []

    def _startup_task(report):
        report("阶段一", "读取测试数据", 20)
        startup_updates.append("stage1")
        time.sleep(0.12)
        report("阶段二", "完成", 100)
        startup_updates.append("stage2")
        return "STARTUP_OK"

    result = sm._run_startup_progress_task("进度自测", "正在测试启动反馈...", _startup_task)
    assert result == "STARTUP_OK"
    assert startup_updates == ["stage1", "stage2"]

    with tempfile.TemporaryDirectory(prefix="sow_progress_test_") as tmp:
        original_settings_path = sm._SETTINGS_PATH
        sm._SETTINGS_PATH = os.path.join(tmp, "settings.json")
        left = os.path.join(tmp, "Book.xlsx")
        right = os.path.join(tmp, "Book-right.xlsx")
        merged = os.path.join(tmp, "Book-merged.xlsx")
        _make_book(left, "mine")
        _make_book(right, "theirs")

        app = sm.SowMergeApp(left, right, merge_mode=False, merged_path=merged)
        app.root.update()
        assert tuple(int(v) for v in app.root.resizable()) == (1, 1)
        if sys.platform.startswith("win"):
            assert app.root.state() == "zoomed", app.root.state()

        app.show_nonblocking_notice("未检测到直接冲突，继续进入手动处理。", duration_ms=0)
        app.root.update()
        assert app.notice_frame.winfo_manager() == "pack"
        assert "未检测到直接冲突" in app.notice_var.get()
        assert app.root.grab_current() is None
        app._hide_nonblocking_notice()
        app.root.update()
        assert not app.notice_frame.winfo_manager()

        app._set_task_status("正在加载 Sheet：Data（1/3）", active=True, current=0, total=3)
        app.root.update()
        assert "Data" in app.task_status_var.get()
        assert str(app.task_progress.cget("mode")) == "determinate"
        app._set_task_status("数据加载完成：已计算 3/3 个 Sheet", active=False, current=3, total=3)
        app.root.update()
        assert float(app.task_progress.cget("value")) == 3.0

        view = app.sheet_views.get("Data")
        assert view is not None
        view._suppress_bg_apply = True
        view._show_loading("正在后台精确计算 Data...")
        assert view.loading_progress.winfo_manager()
        app.root.update_idletasks()
        view._hide_loading()
        app.root.update()
        assert not view.loading_progress.winfo_manager()
        view._suppress_bg_apply = False

        # Editable preload completion may update readiness, but it must never
        # trigger a foreground rescan on the Tk thread.
        original_refresh = view.refresh
        edit_ready_refreshes = []
        view.refresh = lambda *args, **kwargs: edit_ready_refreshes.append(
            (args, dict(kwargs))
        )
        try:
            view._cache_formula_aware = True
            app._refresh_loaded_views_after_edit_ready()
        finally:
            view.refresh = original_refresh
        assert not any(
            bool(kwargs.get("rescan"))
            for _args, kwargs in edit_ready_refreshes
        ), edit_ready_refreshes

        # A checkbox click while the Sheet cache is still pending must never
        # run refresh(rescan=True) on the Tk callback.
        old_ready = view._data_ready
        old_enqueue = app._enqueue_sheet
        old_kick = app._kick_worker
        old_refresh = view.refresh
        refresh_calls = []
        app._enqueue_sheet = lambda *args, **kwargs: None
        app._kick_worker = lambda: None
        view.refresh = lambda row_only, rescan: refresh_calls.append((row_only, rescan))
        view._data_ready = False
        view.only_diff_var.set(1)
        started = time.monotonic()
        view._toggle_only_diff()
        elapsed = time.monotonic() - started
        assert elapsed < 0.2, elapsed
        assert refresh_calls == [], refresh_calls
        view.refresh = old_refresh
        view._data_ready = old_ready
        app._enqueue_sheet = old_enqueue
        app._kick_worker = old_kick

        ticks = []

        def _tick():
            ticks.append(time.monotonic())
            if len(ticks) < 100:
                app.root.after(20, _tick)

        app.root.after(0, _tick)
        progress_window_state = str(app.root.state())
        progress_window_geometry = str(app.root.geometry())

        def _background_task(report):
            report("后台任务", "执行中", 35)
            time.sleep(0.35)
            report("后台任务", "完成", 100)
            return 42

        value = app._with_progress(
            "后台进度自测",
            "验证 UI 事件循环保持响应",
            _background_task,
            run_in_background=True,
            pass_reporter=True,
        )
        assert value == 42
        assert len(ticks) >= 5, f"progress dialog blocked Tk event loop: ticks={len(ticks)}"
        tick_gaps = [
            later - earlier
            for earlier, later in zip(ticks, ticks[1:])
        ]
        assert not tick_gaps or max(tick_gaps) < 0.2, max(tick_gaps)
        assert str(app.root.state()) == progress_window_state
        assert str(app.root.geometry()) == progress_window_geometry

        # A user-selected normal window must not be maximized, normalized, or
        # repositioned by a cached only-diff transition.
        if sys.platform.startswith("win"):
            app.root.state("normal")
            app.root.geometry("940x760+120+90")
            app.root.update()
            normal_state = str(app.root.state())
            normal_geometry = str(app.root.geometry())
            view._row_model_exact = True
            view._cache_formula_aware = True
            view._pair_diff_full_exact = True
            view._only_diff_rows_exact = True
            view._cache_only_diff_rows_snapshot(
                pair_idx
                for pair_idx in range(len(view.row_pairs))
                if view._pair_has_visual_diff(pair_idx)
            )
            view._refresh_interaction_gate()
            view.only_diff_var.set(1)
            view._toggle_only_diff()
            app.root.update()
            assert str(app.root.state()) == normal_state
            assert str(app.root.geometry()) == normal_geometry

        old_showinfo = sm.messagebox.showinfo
        old_showwarning = sm.messagebox.showwarning
        old_showerror = sm.messagebox.showerror
        sm.messagebox.showinfo = lambda *args, **kwargs: None
        sm.messagebox.showwarning = lambda *args, **kwargs: None
        sm.messagebox.showerror = lambda *args, **kwargs: None
        try:
            try:
                app.save_merged_and_exit(auto=True)
            except SystemExit:
                pass
        finally:
            sm.messagebox.showinfo = old_showinfo
            sm.messagebox.showwarning = old_showwarning
            sm.messagebox.showerror = old_showerror

        assert sm._workbook_package_ready(merged)
        wb = load_workbook(merged, data_only=False)
        try:
            assert wb["Data"]["B2"].value == "mine"
        finally:
            wb.close()
        sm._SETTINGS_PATH = original_settings_path

    print("GUI_SELF_TEST_PROGRESS_FEEDBACK_OK")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"GUI_SELF_TEST_PROGRESS_FEEDBACK_FAIL: {exc}", file=sys.stderr)
        raise

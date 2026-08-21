"""Focused regression coverage for Unicode saves and blank-value diffs."""

from __future__ import annotations

import hashlib
import json
import os
import re
import subprocess
import tempfile
import zipfile
from types import SimpleNamespace

from openpyxl import Workbook, load_workbook

import sow_merge_tool as mod


SHEET = "GunshipsMaster@design"
REPORTED_COORDINATES = ("F5", "H5", "H6", "H7", "H9", "H10", "H11")


def _sha256(path: str) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _record_input_hash(input_hashes: dict[str, str], path: str):
    path = os.path.abspath(path)
    assert os.path.isfile(path), path
    input_hashes[path] = _sha256(path)


def _make_book(path: str, *, blank_kind: str):
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws["A1"] = "标题"
    ws["B1"] = "中文内容"
    ws["F5"] = None if blank_kind == "none" else ""
    for address in ("H5", "H6", "H7", "H9", "H10", "H11"):
        ws[address] = None if blank_kind == "none" else ""
    ws["A3"] = "结构行"
    wb.save(path)
    wb.close()


def _assert_valid_xlsx(path: str):
    valid, reason = mod._validate_xlsx_package(path)
    assert valid, reason
    with zipfile.ZipFile(path, "r") as zf:
        assert zf.testzip() is None


def _assert_unicode_round_trip(path: str, expected: str):
    wb = load_workbook(path, data_only=False, read_only=True)
    try:
        ws = wb[SHEET]
        assert ws["A1"].value == "标题"
        assert ws["B1"].value == expected
    finally:
        wb.close()


def _write_without_worksheet_dimension(source: str, output: str):
    with zipfile.ZipFile(source, "r") as src, zipfile.ZipFile(output, "w") as dst:
        for info in src.infolist():
            payload = src.read(info.filename)
            if info.filename.startswith("xl/worksheets/") and info.filename.endswith(".xml"):
                payload = re.sub(rb"<dimension\b[^>]*/>", b"", payload)
            dst.writestr(info, payload)


def _test_missing_dimension_recovers_actual_scan_bounds(
    root: str, input_hashes: dict[str, str]
):
    source = os.path.join(root, "normal-dimension.xlsx")
    dimensionless = os.path.join(root, "missing-dimension.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "FaceUi@design"
    ws["A1"] = "id"
    ws["E1"] = "last-column"
    ws["E53"] = "last-row"
    wb.save(source)
    wb.close()
    _write_without_worksheet_dimension(source, dimensionless)
    _record_input_hash(input_hashes, source)
    _record_input_hash(input_hashes, dimensionless)

    loaded = mod.load_workbook(dimensionless, read_only=True, data_only=False)
    try:
        sheet = loaded["FaceUi@design"]
        assert sheet.max_row is None and sheet.max_column is None
        assert mod._worksheet_scan_bounds(sheet) == (53, 5)
        assert mod._effective_bounds(sheet) == (53, 5)
    finally:
        loaded.close()


def _test_utf8_native_payload_reader(root: str, input_hashes: dict[str, str]):
    payload_path = os.path.join(root, "ops.json")
    with open(payload_path, "w", encoding="utf-8", newline="") as stream:
        json.dump({"value": "中文保存 / 护山神兽"}, stream, ensure_ascii=False)
    _record_input_hash(input_hashes, payload_path)

    script = (
        "$ErrorActionPreference='Stop';"
        "[Console]::OutputEncoding=[System.Text.Encoding]::UTF8;"
        "$p='" + payload_path.replace("'", "''") + "';"
        "$x=ConvertFrom-Json -InputObject ([System.IO.File]::ReadAllText($p,[System.Text.Encoding]::UTF8));"
        "[Console]::Write($x.value);"
    )
    result = subprocess.run(
        ["powershell", "-NoProfile", "-STA", "-ExecutionPolicy", "Bypass", "-Command", script],
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        check=False,
        timeout=30,
    )
    assert result.returncode == 0, result.stderr
    assert result.stdout == "中文保存 / 护山神兽", repr(result.stdout)


def _test_native_script_uses_explicit_utf8(root: str, input_hashes: dict[str, str]):
    source = os.path.join(root, "native-script-source.xlsx")
    output = os.path.join(root, "native-output.xlsx")
    _make_book(source, blank_kind="none")
    _record_input_hash(input_hashes, source)
    captured = {}
    original_runner = mod._run_excel_powershell_with_transient_retry

    def _capture(script: str, *, timeout: int):
        captured["script"] = script
        return SimpleNamespace(returncode=1, stderr="captured test runner")

    mod._run_excel_powershell_with_transient_retry = _capture
    try:
        assert not mod._build_manual_merge_output_with_excel(
            source,
            output,
            {(SHEET, 1, 2): "中文保存"},
        )
    finally:
        mod._run_excel_powershell_with_transient_retry = original_runner
    assert "ReadAllText($opsPath,[System.Text.Encoding]::UTF8)" in captured["script"]


def _test_native_sheet_copy_payload_is_utf8_and_positioned(
    root: str, input_hashes: dict[str, str]
):
    source = os.path.join(root, "sheet-copy-source.xlsx")
    source_b = os.path.join(root, "sheet-copy-source-b.xlsx")
    output = os.path.join(root, "sheet-copy-native.xlsx")
    _make_book(source, blank_kind="none")
    _make_book(source_b, blank_kind="none")
    wb = load_workbook(source_b)
    try:
        wb[SHEET]["B1"] = "整 Sheet 中文"
        wb.save(source_b)
    finally:
        wb.close()
    _record_input_hash(input_hashes, source)
    _record_input_hash(input_hashes, source_b)

    captured = {}
    original_runner = mod._run_excel_powershell_with_transient_retry

    def _capture(script: str, *, timeout: int):
        captured["script"] = script
        match = re.search(r"\$opsPath='([^']*)';", script)
        assert match, script
        with open(match.group(1), "r", encoding="utf-8") as stream:
            captured["payload"] = json.load(stream)
        return SimpleNamespace(returncode=1, stderr="captured native sheet test")

    mod._run_excel_powershell_with_transient_retry = _capture
    try:
        assert not mod._build_manual_merge_output_with_excel(
            source,
            output,
            {},
            sheet_ops=[
                {
                    "kind": "copy_sheet",
                    "sheet": SHEET,
                    "source_side": "B",
                    "target_side": "A",
                    "target_index": 0,
                }
            ],
            source_paths={"B": source_b},
        )
    finally:
        mod._run_excel_powershell_with_transient_retry = original_runner
    assert "ReadAllText($opsPath,[System.Text.Encoding]::UTF8)" in captured["script"]
    assert captured["payload"]["sheet_ops"][0]["target_index"] == 0
    assert "targetIndex" in captured["script"]


def _test_cell_only_and_fallback_unicode_saves(
    root: str, input_hashes: dict[str, str]
):
    source = os.path.join(root, "unicode-save-source.xlsx")
    cell_only_2way = os.path.join(root, "cell-only-2way.xlsx")
    cell_only_3way = os.path.join(root, "cell-only-3way.xlsx")
    fallback = os.path.join(root, "fallback.xlsx")
    _make_book(source, blank_kind="none")
    _record_input_hash(input_hashes, source)

    for output in (cell_only_2way, cell_only_3way):
        mod._build_manual_merge_xlsx_via_zip(
            source,
            output,
            {(SHEET, 1, 2): "中文保存"},
        )
        _assert_valid_xlsx(output)
        _assert_unicode_round_trip(output, "中文保存")

    assert mod._build_manual_merge_output_with_openpyxl(
        source,
        fallback,
        {(SHEET, 2, 2): "三路中文"},
        row_ops=[{"kind": "insert_rows", "sheet": SHEET, "row": 2, "count": 1}],
    )
    _assert_valid_xlsx(fallback)
    wb = load_workbook(fallback, data_only=False, read_only=True)
    try:
        assert wb[SHEET]["B2"].value == "三路中文"
    finally:
        wb.close()


def _test_blank_equivalence_and_conservative_diffs():
    assert mod._merge_cmp_value(None) == mod._merge_cmp_value("")
    assert mod._cell_display_and_equal_from_values(None, "", None, "")[2]
    assert not mod._cell_display_and_equal_from_values(None, " ", None, " ")[2]
    assert not mod._cell_display_and_equal_from_values(None, "x", None, "x")[2]
    assert not mod._cell_display_and_equal_from_values(None, "=A1", None, "=A1")[2]

    key = mod.ColumnModelCacheKey(SHEET, 1, 1)
    cache_2way = mod.build_logical_column_comparison_cache_2way(
        key,
        [(None,)] + [(None,)] * 10,
        [("x",)] + [("x",)] * 10,
        [(None,)] + [(None,)] * 10,
        [("x",)] + [("x",)] * 10,
        mine_max_col=1,
        theirs_max_col=1,
    )
    result_2way = mod.compare_logical_row_2way(
        cache_2way,
        (None,),
        ("",),
        (None,),
        ("",),
    )
    assert not result_2way.has_diff, result_2way

    cache_3way = mod.build_logical_column_comparison_cache_3way(
        key,
        [(None,)] * 11,
        [("",)] * 11,
        [(None,)] * 11,
        [(None,)] * 11,
        [("",)] * 11,
        [(None,)] * 11,
        mine_max_col=1,
        base_max_col=1,
        theirs_max_col=1,
    )
    result_3way = mod.compare_logical_row_3way(
        cache_3way,
        (None,),
        ("",),
        (None,),
        (None,),
        ("",),
        (None,),
    )
    assert not result_3way.mine_changed_cols
    assert not result_3way.theirs_changed_cols
    assert not result_3way.conflict_cols

    for address in REPORTED_COORDINATES:
        assert mod._cell_display_and_equal_from_values(None, "", None, "")[2], address


def main():
    temporary = tempfile.TemporaryDirectory(prefix="sow_save_diff_fidelity_")
    root = temporary.name
    input_hashes: dict[str, str] = {}
    primary = None
    try:
        _test_missing_dimension_recovers_actual_scan_bounds(root, input_hashes)
        _test_utf8_native_payload_reader(root, input_hashes)
        _test_native_script_uses_explicit_utf8(root, input_hashes)
        _test_native_sheet_copy_payload_is_utf8_and_positioned(root, input_hashes)
        _test_cell_only_and_fallback_unicode_saves(root, input_hashes)
        _test_blank_equivalence_and_conservative_diffs()
        print("SMOKE_SAVE_AND_DIFF_FIDELITY_OK")
    except BaseException as exc:
        primary = exc
        raise
    finally:
        cleanup_errors = []
        for path, before_hash in input_hashes.items():
            try:
                assert _sha256(path) == before_hash, path
            except BaseException as exc:
                cleanup_errors.append(f"input SHA {path!r}: {exc!r}")
        try:
            temporary.cleanup()
            assert not os.path.lexists(root), root
        except BaseException as exc:
            cleanup_errors.append(f"owned temporary root: {exc!r}")
        if cleanup_errors:
            message = "save-and-diff fidelity cleanup failed: " + "; ".join(cleanup_errors)
            if primary is not None:
                primary.add_note(message)
            else:
                raise AssertionError(message)


if __name__ == "__main__":
    main()

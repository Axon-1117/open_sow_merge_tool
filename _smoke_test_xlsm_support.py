import os

from openpyxl import Workbook

import sow_merge_tool as mod
from _test_temp_utils import make_temp_dir


def _make_book(path: str, value: object):
    wb = Workbook()
    ws = wb.active
    ws.title = "S1"
    ws["A1"] = "id"
    ws["A2"] = value
    wb.save(path)
    wb.close()


def main():
    root = make_temp_dir("sow_xlsm_support_")
    a = os.path.join(root, "a.xlsm")
    b = os.path.join(root, "b.xlsm")
    out = os.path.join(root, "out.xlsm")
    _make_book(a, "left")
    _make_book(b, "right")

    real_loader = mod._openpyxl_load_workbook
    calls = []

    def _spy_loader(filename, *args, **kwargs):
        calls.append((str(filename), dict(kwargs)))
        return real_loader(filename, *args, **kwargs)

    mod._openpyxl_load_workbook = _spy_loader
    app = None
    try:
        app = mod.SowMergeApp(a, b)
        assert any(path.lower().endswith(".xlsm") and kwargs.get("keep_vba") is True for path, kwargs in calls), calls
        app._ensure_edit_loaded()
        app._atomic_save(app._wb_a_edit, out)
        assert os.path.exists(out), f"save output missing: {out}"
    finally:
        mod._openpyxl_load_workbook = real_loader
        if app is not None:
            try:
                app._shutdown_root()
            except Exception:
                pass

    print("SMOKE_XLSM_SUPPORT_OK")


if __name__ == "__main__":
    main()

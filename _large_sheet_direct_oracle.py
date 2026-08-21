"""Direct read-only two-way Oracle for one selected Sheet.

Unlike the GUI oracle this opens only the requested Sheet through paired
value/formula readers.  It deliberately reuses the production legacy row and
logical-column functions, then emits the common normalized manifest schema.
"""

from __future__ import annotations

import argparse
import json

from openpyxl import load_workbook

import sow_merge_tool as sm
from _large_sheet_legacy_oracle import _cell_token, _three_way_conflicts


def _trimmed_rows(ws):
    """Read a ReadOnlyWorksheet once, matching the worker's effective horizon."""
    max_row, max_col = sm._worksheet_scan_bounds(ws)
    rows, last_row, last_col = [], 1, 1
    found = False
    for row_index, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True),
        start=1,
    ):
        frozen = tuple(row or ())
        rows.append(frozen)
        used = [col for col, value in enumerate(frozen, start=1) if value not in (None, "")]
        if used:
            found = True
            last_row = row_index
            last_col = max(last_col, max(used))
    if not found:
        return 1, max(1, max_col), rows
    return max(1, last_row), max(1, last_col), rows


def _pairs_from_rows(values_a, values_b, formulas_a, formulas_b, max_ra, max_rb, max_col):
    if not sm._should_auto_row_align(max_ra, max_rb, force=False):
        return [(row if row <= max_ra else None, row if row <= max_rb else None)
                for row in range(1, max(max_ra, max_rb) + 1)]
    signatures_a, signatures_b = sm._row_signatures_from_unique_column_anchors(
        values_a[:max_ra], values_b[:max_rb], max_col, max_col,
        formulas_a[:max_ra], formulas_b[:max_rb],
    )
    return sm._compute_row_pairs_from_signatures(signatures_a, signatures_b)


def capture(mine_path: str, theirs_path: str, sheet: str) -> dict:
    books = []
    try:
        mine_val = load_workbook(mine_path, data_only=True, read_only=True)
        theirs_val = load_workbook(theirs_path, data_only=True, read_only=True)
        mine_edit = load_workbook(mine_path, data_only=False, read_only=True)
        theirs_edit = load_workbook(theirs_path, data_only=False, read_only=True)
        books.extend((mine_val, theirs_val, mine_edit, theirs_edit))
        ws_a, ws_b = mine_val[sheet], theirs_val[sheet]
        ws_a_edit, ws_b_edit = mine_edit[sheet], theirs_edit[sheet]
        max_ra, max_ca, value_rows_a = _trimmed_rows(ws_a)
        max_rb, max_cb, value_rows_b = _trimmed_rows(ws_b)
        edit_ra, edit_ca, formula_rows_a = _trimmed_rows(ws_a_edit)
        edit_rb, edit_cb, formula_rows_b = _trimmed_rows(ws_b_edit)
        max_ra, max_ca = max(max_ra, edit_ra), max(max_ca, edit_ca)
        max_rb, max_cb = max(max_rb, edit_rb), max(max_cb, edit_cb)
        max_col = max(max_ca, max_cb)
        pairs = _pairs_from_rows(
            value_rows_a, value_rows_b, formula_rows_a, formula_rows_b,
            max_ra, max_rb, max_col,
        )
        pairs = sm._collapse_one_sided_blank_tail_padding(
            pairs, ws_a, ws_b, ws_a_edit, ws_b_edit, max_ra, max_rb, max_col
        )
        rows_a = sm._read_rows_into_cache(ws_a, [a for a, _ in pairs], max_col)
        rows_b = sm._read_rows_into_cache(ws_b, [b for _, b in pairs], max_col)
        rows_a_edit = sm._read_rows_into_cache(ws_a_edit, [a for a, _ in pairs], max_col)
        rows_b_edit = sm._read_rows_into_cache(ws_b_edit, [b for _, b in pairs], max_col)
        aligned_a = [sm._row_from_cache(rows_a, a, max_col) for a, _ in pairs]
        aligned_b = [sm._row_from_cache(rows_b, b, max_col) for _, b in pairs]
        aligned_af = [sm._row_from_cache(rows_a_edit, a, max_col) for a, _ in pairs]
        aligned_bf = [sm._row_from_cache(rows_b_edit, b, max_col) for _, b in pairs]
        cache = sm.build_logical_column_comparison_cache_2way(
            sm.ColumnModelCacheKey(sheet, 1, 1), aligned_a, aligned_b, aligned_af, aligned_bf,
            mine_max_col=max_ca, theirs_max_col=max_cb,
        )
        projection = sm.LogicalColumnProjection.from_model(cache.model)
        columns = [
            {"logical": slot.logical_idx + 1, "mine": slot.mine_col, "base": slot.base_col,
             "theirs": slot.theirs_col, "state": slot.state, "ambiguous": bool(slot.confidence.ambiguous)}
            for slot in projection.model.slots
        ]
        records, only_diff_rows = [], []
        structural = {int(col) for col in cache.structural_diff_cols if int(col) > 0}
        common_insertions = {
            slot.logical_idx + 1 for slot in cache.model.slots
            if slot.base_col is None and slot.mine_col is not None and slot.theirs_col is not None
            and slot.state != "unresolved" and not slot.confidence.ambiguous
        }
        for index, (mine_row, theirs_row) in enumerate(pairs):
            comparison = sm.compare_logical_row_2way(
                cache, aligned_a[index], aligned_b[index], aligned_af[index], aligned_bf[index],
                mine_row=mine_row, theirs_row=theirs_row,
                mine_present=mine_row is not None, theirs_present=theirs_row is not None,
            )
            changed = sorted(int(col) for col in comparison.diff_cols if int(col) != -1)
            visual = set(comparison.diff_cols)
            visual.difference_update(structural)
            visual.difference_update(col for col in common_insertions if col not in comparison.diff_cols)
            if visual:
                only_diff_rows.append(index)
            if not changed and mine_row is not None and theirs_row is not None and -1 not in comparison.diff_cols:
                continue
            cells = {}
            for logical in changed:
                slot = projection.slot(logical)
                if slot is not None:
                    cells[str(logical)] = {
                        "mine": _cell_token(ws_a, ws_a_edit, mine_row, slot.mine_col),
                        "theirs": _cell_token(ws_b, ws_b_edit, theirs_row, slot.theirs_col),
                        "base": {"present": False},
                    }
            records.append({
                "pair": index, "mine_row": mine_row, "theirs_row": theirs_row, "base_row": None,
                "row_structure": -1 in comparison.diff_cols, "diff_cols": changed,
                "base_diff_cols": [], "conflicts": _three_way_conflicts(cells), "cells": cells,
            })
        return {
            "schema": "legacy-large-sheet-oracle-v1", "sheet": sheet, "three_way": False,
            "sides": ["mine", "theirs"], "columns": columns,
            "only_diff_rows": only_diff_rows, "records": records,
        }
    finally:
        for book in books:
            book.close()


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--mine", required=True)
    parser.add_argument("--theirs", required=True)
    parser.add_argument("--sheet", required=True)
    parser.add_argument("--out", required=True)
    args = parser.parse_args()
    with open(args.out, "w", encoding="utf-8", newline="\n") as output:
        json.dump(capture(args.mine, args.theirs, args.sheet), output, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


if __name__ == "__main__":
    main()

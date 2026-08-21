"""Read-only full-corpus exact-readiness benchmark.

The source tree is deliberately never used as a mutation root.  Each Sheet
worker copies its source file to Mine/Theirs (and Base for merge) below a new
``TemporaryDirectory`` before opening it.  The controller treats every source
file as an input record, including lock files and unsupported formats, and
persists incomplete/timeout results rather than silently skipping them.

Run the complete corpus (the default has no fixture filter)::

    python _large_sheet_corpus_benchmark.py --out benchmark_results/corpus.json

Each measured Sheet runs in its own fresh Python process.  The 15-second gate
is measured from Sheet request to immutable exact result; an enclosing process
timeout preserves a concrete error if a reader fails to return at all.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import shutil
import subprocess
import sys
import tempfile
import time
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from typing import Any

import psutil
from openpyxl import load_workbook


SOURCE_ROOT = Path(r"C:\GM15\design\sheets\develop")
SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm"}
SCHEMA = "large-sheet-corpus-v1"
GATE_SECONDS = 15.0


def _json_dump(path: Path, payload: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    os.replace(temporary, path)


def _source_signature(path: Path) -> dict[str, Any]:
    stat = path.stat()
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for block in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(block)
    return {"size_bytes": stat.st_size, "mtime_ns": stat.st_mtime_ns, "sha256": digest.hexdigest()}


def inventory(root: Path) -> dict[str, Any]:
    """Recursively classify every regular input without opening it for write."""
    files: list[dict[str, Any]] = []
    for path in sorted(root.rglob("*"), key=lambda item: str(item).casefold()):
        if not path.is_file():
            continue
        relative = str(path.relative_to(root))
        suffix = path.suffix.casefold()
        temporary = path.name.startswith("~$")
        classification = "temporary" if temporary else "supported" if suffix in SUPPORTED_EXTENSIONS else "unsupported"
        item: dict[str, Any] = {"relative_path": relative, "source_path": str(path), "extension": suffix, "classification": classification}
        try:
            stat = path.stat()
            item.update({"size_bytes": stat.st_size, "mtime_ns": stat.st_mtime_ns})
        except OSError as exc:
            item["classification"] = "inventory_error"
            item["error"] = f"{type(exc).__name__}: {exc}"
        files.append(item)
    return {
        "source_root": str(root),
        "read_only_contract": "source files are inventory/read inputs only; every side is copied below a disposable temporary root",
        "files": files,
        "counts": {kind: sum(item["classification"] == kind for item in files) for kind in ("supported", "temporary", "unsupported", "inventory_error")},
    }


def _catalog(path: Path) -> list[str]:
    book = load_workbook(path, read_only=True, data_only=False, keep_vba=path.suffix.casefold() == ".xlsm")
    try:
        return list(book.sheetnames)
    finally:
        book.close()


def _copy_sides(source: Path, root: Path, three_way: bool) -> tuple[Path, Path, Path | None]:
    source_root = SOURCE_ROOT.resolve()
    if root.resolve().is_relative_to(source_root):
        raise AssertionError("disposable root must not be inside source corpus")
    mine, theirs = root / "Mine" / source.name, root / "Theirs" / source.name
    base = root / "Base" / source.name if three_way else None
    for target in (mine, theirs, base):
        if target is not None:
            target.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(source, target)
    return mine, theirs, base


def _formula_count(snapshot: Any) -> int:
    return sum(cell.formula_kind != "literal" for row in snapshot.rows for cell in row.cells)


def _sheet_worker(args: argparse.Namespace) -> dict[str, Any]:
    """One selected-Sheet fresh process; all comparison inputs are copies."""
    import sow_merge_tool as sm

    source = Path(args.source)
    started = time.perf_counter()
    process = psutil.Process()
    rss_before = process.memory_info().rss
    payload: dict[str, Any] = {
        "schema": SCHEMA, "mode": args.mode, "source_path": str(source), "sheet": args.sheet,
        "engine": "snapshot-paired-read-only", "fallback": None, "gate_seconds": args.gate,
        "status": "FAILED", "final_state": "FAILED", "oracle_zero_difference": False,
        "oracle_zero_conflict": False, "request_to_final_exact_ms": None, "revisit_ms": None,
        "error": None,
    }
    try:
        with tempfile.TemporaryDirectory(prefix="sow_corpus_sheet_") as temporary:
            disposable = Path(temporary)
            mine, theirs, base = _copy_sides(source, disposable, args.mode == "3way")
            request_started = time.perf_counter()
            # This intentionally parallels only independent ZIP readers.  No
            # workbook/worksheet object crosses threads after snapshot creation.
            with ThreadPoolExecutor(max_workers=3 if base else 2) as executor:
                future_mine = executor.submit(sm._stream_selected_sheet_snapshot, str(mine), str(mine), args.sheet, "mine")
                future_theirs = executor.submit(sm._stream_selected_sheet_snapshot, str(theirs), str(theirs), args.sheet, "theirs")
                future_base = executor.submit(sm._stream_selected_sheet_snapshot, str(base), str(base), args.sheet, "base") if base else None
                mine_snapshot, theirs_snapshot = future_mine.result(), future_theirs.result()
                base_snapshot = future_base.result() if future_base else None
            result = sm._compare_selected_sheet_snapshots(mine_snapshot, theirs_snapshot, base_snapshot)
            elapsed_ms = (time.perf_counter() - request_started) * 1000.0
            diff_count = sum(bool(item) for item in result.pair_diff_cols) + sum(bool(item) for item in result.pair_base_diff_cols)
            conflict_count = sum(bool(item) for item in result.conflict_cols)
            exact_same = diff_count == 0 and conflict_count == 0 and not result.unresolved
            payload.update({
                "dimensions": {"rows": mine_snapshot.max_row, "columns": mine_snapshot.max_col},
                "formula_cells": _formula_count(mine_snapshot),
                "request_to_final_exact_ms": round(elapsed_ms, 3),
                "diff_pair_count": diff_count,
                "conflict_pair_count": conflict_count,
                "unresolved": bool(result.unresolved),
                "oracle_zero_difference": diff_count == 0,
                "oracle_zero_conflict": conflict_count == 0,
                "final_state": "EXACT_SAME" if exact_same else "UNRESOLVED" if result.unresolved else "EXACT_CHANGED",
                "status": "PASS" if exact_same and elapsed_ms <= args.gate * 1000.0 else "TIMEOUT" if elapsed_ms > args.gate * 1000.0 else "ORACLE_FAILURE",
            })
            revisit_started = time.perf_counter()
            # Existing immutable snapshots are the cached revisit payload.
            _ = result.row_pairs, mine_snapshot.rows, theirs_snapshot.rows
            payload["revisit_ms"] = round((time.perf_counter() - revisit_started) * 1000.0, 3)
    except Exception as exc:
        payload["error"] = f"{type(exc).__name__}: {exc}"
        payload["fallback"] = "none: snapshot engine did not produce an exact result"
    finally:
        payload["startup_and_request_ms"] = round((time.perf_counter() - started) * 1000.0, 3)
        payload["rss_before_bytes"] = rss_before
        payload["rss_after_bytes"] = process.memory_info().rss
        payload["rss_delta_bytes"] = payload["rss_after_bytes"] - rss_before
        payload["rss_peak_bytes"] = payload["rss_after_bytes"]
    return payload


def _run_sheet_child(args: argparse.Namespace, sheet: str) -> dict[str, Any]:
    command = [sys.executable, os.path.abspath(__file__), "--sheet-worker", "--source", args.source, "--mode", args.mode, "--sheet", sheet, "--gate", str(args.gate)]
    try:
        completed = subprocess.run(command, capture_output=True, text=True, timeout=args.gate + args.process_grace, check=False)
    except subprocess.TimeoutExpired as exc:
        return {"schema": SCHEMA, "mode": args.mode, "source_path": args.source, "sheet": sheet, "status": "PROCESS_TIMEOUT", "final_state": "FAILED", "engine": "snapshot-paired-read-only", "fallback": "process watchdog", "error": f"process exceeded {args.gate + args.process_grace:.1f}s envelope: {exc}", "request_to_final_exact_ms": None, "revisit_ms": None, "oracle_zero_difference": False, "oracle_zero_conflict": False}
    lines = [line for line in completed.stdout.splitlines() if line.strip()]
    if completed.returncode or not lines:
        return {"schema": SCHEMA, "mode": args.mode, "source_path": args.source, "sheet": sheet, "status": "WORKER_FAILURE", "final_state": "FAILED", "engine": "snapshot-paired-read-only", "fallback": "worker-crash", "error": (completed.stderr or completed.stdout or f"exit code {completed.returncode}")[-4000:], "request_to_final_exact_ms": None, "revisit_ms": None, "oracle_zero_difference": False, "oracle_zero_conflict": False}
    try:
        return json.loads(lines[-1])
    except json.JSONDecodeError as exc:
        return {"schema": SCHEMA, "mode": args.mode, "source_path": args.source, "sheet": sheet, "status": "WORKER_PROTOCOL_FAILURE", "final_state": "FAILED", "engine": "snapshot-paired-read-only", "fallback": "worker-protocol", "error": f"{exc}: {lines[-1][-500:]}", "request_to_final_exact_ms": None, "revisit_ms": None, "oracle_zero_difference": False, "oracle_zero_conflict": False}


def _file_worker(args: argparse.Namespace) -> dict[str, Any]:
    started = time.perf_counter()
    process = psutil.Process()
    rss_before = process.memory_info().rss
    source = Path(args.source)
    result: dict[str, Any] = {"schema": SCHEMA, "source_path": str(source), "relative_path": args.relative_path, "mode": args.mode, "status": "FAILED", "startup_ms": None, "catalog_ms": None, "whole_summary_ms": None, "rss_before_bytes": rss_before, "sheets": [], "error": None}
    try:
        import sow_merge_tool  # noqa: F401 -- makes application import part of startup evidence
        import_done = time.perf_counter()
        catalog_started = time.perf_counter()
        sheets = _catalog(source)
        result["startup_ms"] = round((import_done - started) * 1000.0, 3)
        result["catalog_ms"] = round((time.perf_counter() - catalog_started) * 1000.0, 3)
        for sheet in sheets:
            result["sheets"].append(_run_sheet_child(args, sheet))
        failures = [sheet for sheet in result["sheets"] if sheet["status"] != "PASS"]
        result["status"] = "PASS" if not failures else "FAIL"
    except Exception as exc:
        result["error"] = f"{type(exc).__name__}: {exc}"
    finally:
        result["whole_summary_ms"] = round((time.perf_counter() - started) * 1000.0, 3)
        result["rss_after_bytes"] = process.memory_info().rss
        result["rss_delta_bytes"] = result["rss_after_bytes"] - rss_before
        result["rss_peak_bytes"] = result["rss_after_bytes"]
    return result


def _run_file_child(source: Path, relative_path: str, mode: str, gate: float, process_grace: float) -> dict[str, Any]:
    command = [sys.executable, os.path.abspath(__file__), "--file-worker", "--source", str(source), "--relative-path", relative_path, "--mode", mode, "--gate", str(gate), "--process-grace", str(process_grace)]
    try:
        completed = subprocess.run(command, capture_output=True, text=True, check=False)
    except Exception as exc:
        return {"source_path": str(source), "relative_path": relative_path, "mode": mode, "status": "PROCESS_FAILURE", "error": f"{type(exc).__name__}: {exc}", "sheets": []}
    lines = [line for line in completed.stdout.splitlines() if line.strip()]
    if completed.returncode or not lines:
        return {"source_path": str(source), "relative_path": relative_path, "mode": mode, "status": "PROCESS_FAILURE", "error": (completed.stderr or completed.stdout or f"exit code {completed.returncode}")[-4000:], "sheets": []}
    return json.loads(lines[-1])


def _rankings(results: list[dict[str, Any]], mode: str) -> list[dict[str, Any]]:
    flat = [
        {"relative_path": file_result["relative_path"], "sheet": sheet["sheet"], "status": sheet["status"], "request_to_final_exact_ms": sheet.get("request_to_final_exact_ms"), "final_state": sheet.get("final_state"), "error": sheet.get("error")}
        for file_result in results if file_result.get("mode") == mode for sheet in file_result.get("sheets", [])
    ]
    return sorted(flat, key=lambda item: (item["request_to_final_exact_ms"] is None, -(item["request_to_final_exact_ms"] or -1)))


def _write_rankings(path: Path, evidence: dict[str, Any]) -> None:
    lines = ["# Large-sheet corpus slowest-Sheet rankings", ""]
    for mode in ("2way", "3way"):
        lines.extend((f"## {mode}", "", "| Rank | Workbook | Sheet | Exact-ready ms | State | Status |", "| ---: | --- | --- | ---: | --- | --- |"))
        for index, item in enumerate(evidence["slowest_sheets"][mode], start=1):
            duration = "-" if item["request_to_final_exact_ms"] is None else f"{item['request_to_final_exact_ms']:.3f}"
            lines.append(f"| {index} | {item['relative_path']} | {item['sheet']} | {duration} | {item['final_state']} | {item['status']} |")
        lines.append("")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def run(args: argparse.Namespace) -> int:
    source_root = Path(args.source_root)
    if not source_root.is_dir():
        raise FileNotFoundError(f"source corpus missing: {source_root}")
    before = inventory(source_root)
    supported = [item for item in before["files"] if item["classification"] == "supported"]
    if args.max_files is not None:
        supported = supported[:args.max_files]
    evidence: dict[str, Any] = {"schema": SCHEMA, "created_at_epoch": time.time(), "gate_seconds": args.gate, "fresh_process": True, "inventory_before": before, "runs": [], "source_signature_before": {}, "source_signature_after": {}, "slowest_sheets": {"2way": [], "3way": []}}
    out = Path(args.out)
    for item in supported:
        source = Path(item["source_path"])
        try:
            evidence["source_signature_before"][item["relative_path"]] = _source_signature(source)
        except Exception as exc:
            evidence["runs"].append({"source_path": str(source), "relative_path": item["relative_path"], "status": "SOURCE_SIGNATURE_FAILURE", "error": f"{type(exc).__name__}: {exc}", "sheets": []})
            continue
        for mode in ("2way", "3way"):
            run_result = _run_file_child(source, item["relative_path"], mode, args.gate, args.process_grace)
            evidence["runs"].append(run_result)
            evidence["slowest_sheets"][mode] = _rankings(evidence["runs"], mode)
            _json_dump(out, evidence)  # checkpoint every file/mode, including failures
    for item in supported:
        source = Path(item["source_path"])
        try:
            evidence["source_signature_after"][item["relative_path"]] = _source_signature(source)
        except Exception as exc:
            evidence.setdefault("source_signature_errors", {})[item["relative_path"]] = f"{type(exc).__name__}: {exc}"
    evidence["source_unchanged"] = evidence["source_signature_before"] == evidence["source_signature_after"] and not evidence.get("source_signature_errors")
    evidence["inventory_after"] = inventory(source_root)
    evidence["slowest_sheets"] = {mode: _rankings(evidence["runs"], mode) for mode in ("2way", "3way")}
    _json_dump(out, evidence)
    _write_rankings(out.with_name(out.stem + "_slowest.md"), evidence)
    failures = [sheet for run_result in evidence["runs"] for sheet in run_result.get("sheets", []) if sheet.get("status") != "PASS"]
    print(json.dumps({"out": str(out), "runs": len(evidence["runs"]), "sheet_failures": len(failures), "source_unchanged": evidence["source_unchanged"]}, ensure_ascii=False), flush=True)
    return 0 if not failures and evidence["source_unchanged"] else 1


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-root", default=str(SOURCE_ROOT))
    parser.add_argument("--out", default="benchmark_results/large_sheet_corpus.json")
    parser.add_argument("--gate", type=float, default=GATE_SECONDS)
    parser.add_argument("--process-grace", type=float, default=30.0)
    parser.add_argument("--max-files", type=int)
    parser.add_argument("--file-worker", action="store_true")
    parser.add_argument("--sheet-worker", action="store_true")
    parser.add_argument("--source")
    parser.add_argument("--relative-path", default="")
    parser.add_argument("--mode", choices=("2way", "3way"))
    parser.add_argument("--sheet")
    args = parser.parse_args()
    if args.sheet_worker:
        print(json.dumps(_sheet_worker(args), ensure_ascii=False, sort_keys=True), flush=True)
        return
    if args.file_worker:
        print(json.dumps(_file_worker(args), ensure_ascii=False, sort_keys=True), flush=True)
        return
    raise SystemExit(run(args))


if __name__ == "__main__":
    main()

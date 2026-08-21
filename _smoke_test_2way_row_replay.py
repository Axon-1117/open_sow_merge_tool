import os
import time

from openpyxl import Workbook, load_workbook

import sow_merge_tool as mod
from _test_temp_utils import make_temp_dir


def _make_book(path, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "S1"
    for row in rows:
        ws.append(row)
    wb.save(path)
    wb.close()


def main():
    root = make_temp_dir("sow_2way_row_replay_")
    mine = os.path.join(root, "mine.xlsx")
    theirs = os.path.join(root, "theirs.xlsx")
    schema = [["id@id", "formula"], ["string", "formula"]]
    _make_book(mine, schema + [["A", None], ["NEW", "=A4"], ["C", None]])
    _make_book(theirs, schema + [["A", None], ["C", None]])

    # This is a non-interactive smoke test.  Keep the production confirmation
    # behavior unchanged while preventing a modal formula-cache prompt from
    # blocking the automated Tk pump.
    original_prompt_scheduler = mod.SowMergeApp._schedule_formula_cache_prompt
    mod.SowMergeApp._schedule_formula_cache_prompt = lambda _self: None
    try:
        app = mod.SowMergeApp(mine, theirs)
    except Exception:
        mod.SowMergeApp._schedule_formula_cache_prompt = original_prompt_scheduler
        raise
    original_excel_builder = mod._build_manual_merge_output_with_excel
    captured = {}
    try:
        # This exercise verifies the completed row operation.  Deferred edit
        # loading is covered separately; request it explicitly, then wait for
        # the exact snapshot and editable workbooks to be ready together.
        app._request_edit_preload()
        view = None
        exact_entry = None
        for _ in range(300):
            app.root.update_idletasks()
            app.root.update()
            view = app.sheet_views.get("S1")
            if (
                view is not None
                and view._data_ready
                and app._is_sheet_exact_current("S1")
                and app._edit_workbooks_ready()
            ):
                break
            time.sleep(0.02)
        exact_entry = app._sheet_exact_entry("S1")
        assert (
            view is not None
            and view._data_ready
            and app._is_sheet_exact_current("S1")
            and app._edit_workbooks_ready()
        ), (
            f"view did not become exact-ready: lifecycle={getattr(view, '_lifecycle_state', None)!r} "
            f"entry={exact_entry!r} edit_ready={app._edit_workbooks_ready()!r}"
        )
        view.force_align_var.set(1)
        view._toggle_force_align()
        insert_pair = next(
            pair_idx for pair_idx, (row_a, row_b) in enumerate(view.row_pairs)
            if row_a == 4 and row_b is None
        )
        assert view._copy_selected_row("A2B", override_pair_idx=insert_pair)
        assert app.manual_b_row_ops[0]["source_side"] == "A"
        assert app.manual_b_cell_ops[("S1", 4, 2)] == "=A4"

        # Exercise actual App row undo/redo before validating the saved replay.
        view._undo_last_action()
        assert [app.ws_b_edit("S1").cell(row, 1).value for row in range(1, 5)] == ["id@id", "string", "A", "C"]
        assert app.manual_b_row_ops == []
        view._redo_last_action()
        replayed_rows = [app.ws_b_edit("S1").cell(row, 1).value for row in range(1, 6)]
        assert replayed_rows == ["id@id", "string", "A", "NEW", "C"], replayed_rows
        assert app.manual_b_row_ops[0]["source_side"] == "A"
        assert app.manual_b_cell_ops[("S1", 4, 2)] == "=A4"

        def _fake_excel(
            src,
            out,
            manual_ops,
            row_ops,
            sheet_ops=None,
            source_paths=None,
            column_ops=None,
        ):
            captured["src"] = src
            captured["row_ops"] = list(row_ops or [])
            captured["column_ops"] = list(column_ops or [])
            captured["source_paths"] = dict(source_paths or {})
            return mod._build_manual_merge_output_with_openpyxl(
                src,
                out,
                manual_ops,
                row_ops,
                sheet_ops=sheet_ops,
                source_paths=source_paths,
            )

        mod._build_manual_merge_output_with_excel = _fake_excel
        out = app.build_manual_b_output_file()
        assert captured["src"] == theirs
        assert captured["column_ops"] == []
        assert captured["source_paths"]["A"] == mine
        assert captured["row_ops"][0]["source_side"] == "A"

        wb = load_workbook(out, data_only=False)
        try:
            ws = wb["S1"]
            assert [ws.cell(row=row, column=1).value for row in range(1, 6)] == ["id@id", "string", "A", "NEW", "C"]
            assert ws["B4"].value == "=A4"
        finally:
            wb.close()
    finally:
        mod.SowMergeApp._schedule_formula_cache_prompt = original_prompt_scheduler
        mod._build_manual_merge_output_with_excel = original_excel_builder
        app._shutdown_root()

    print("SMOKE_2WAY_ROW_REPLAY_OK")


if __name__ == "__main__":
    main()

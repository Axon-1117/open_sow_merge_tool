"""Focused proof for every Language expected-safe-unresolved mutator path."""

from __future__ import annotations

import argparse
import hashlib
import shutil
import tempfile
import time
import zipfile
from xml.etree import ElementTree as ET
from pathlib import Path
from types import SimpleNamespace

from openpyxl import Workbook, load_workbook

import _large_sheet_excel_fidelity_gate as gate
import sow_merge_tool as sm


_CASE = "fidelity-language-expected-safe-unresolved"
_SHEET = "Language"
_MARKER_MINE = "__SOW_FIDELITY_Language_value_MINE__"
_MARKER_THEIRS = "__SOW_FIDELITY_Language_value_THEIRS__"
_WORKBOOK_MEMBER = "xl/workbook.xml"
_WORKBOOK_RELS_MEMBER = "xl/_rels/workbook.xml.rels"


def _member_hashes(members: dict[str, bytes]) -> dict[str, str]:
    return {name: hashlib.sha256(payload).hexdigest() for name, payload in members.items()}


def _canonicalize_owned_fixture_worksheet_relation(path: Path) -> None:
    """Make the synthetic workbook use the relative sheet target real fixtures use."""
    with zipfile.ZipFile(path, "r") as archive:
        names_before = tuple(archive.namelist())
        assert names_before.count(_WORKBOOK_MEMBER) == 1
        assert names_before.count(_WORKBOOK_RELS_MEMBER) == 1
        assert archive.testzip() is None
        members_before = {name: archive.read(name) for name in names_before}

    workbook = ET.fromstring(members_before[_WORKBOOK_MEMBER])
    relation_id = next(
        node.attrib[gate._DOC_REL_NS + "id"]
        for node in workbook.iter(gate._MAIN_NS + "sheet")
        if node.attrib.get("name") == _SHEET
    )
    relationships = ET.fromstring(members_before[_WORKBOOK_RELS_MEMBER])
    worksheet_relationships = [
        node
        for node in relationships.iter(gate._PACKAGE_REL_NS + "Relationship")
        if node.attrib.get("Id") == relation_id
    ]
    assert len(worksheet_relationships) == 1
    worksheet_relationship = worksheet_relationships[0]
    original_target = worksheet_relationship.get("Target")
    assert original_target == "/xl/worksheets/sheet1.xml"
    worksheet_member = original_target.lstrip("/")
    assert worksheet_member in members_before
    canonical_target = worksheet_member.removeprefix("xl/")
    assert canonical_target == "worksheets/sheet1.xml"
    before_targets = tuple(
        (node.attrib.get("Id"), node.attrib.get("Target"))
        for node in relationships.iter(gate._PACKAGE_REL_NS + "Relationship")
    )
    worksheet_relationship.set("Target", canonical_target)
    gate._rewrite_member(
        path,
        _WORKBOOK_RELS_MEMBER,
        ET.tostring(relationships, encoding="utf-8", xml_declaration=True),
    )

    with zipfile.ZipFile(path, "r") as archive:
        names_after = tuple(archive.namelist())
        assert names_after == names_before
        assert archive.testzip() is None
        members_after = {name: archive.read(name) for name in names_after}
    hashes_before = _member_hashes(members_before)
    hashes_after = _member_hashes(members_after)
    assert all(
        hashes_after[name] == hashes_before[name]
        for name in names_before
        if name != _WORKBOOK_RELS_MEMBER
    )
    relationships_after = ET.fromstring(members_after[_WORKBOOK_RELS_MEMBER])
    after_targets = tuple(
        (node.attrib.get("Id"), node.attrib.get("Target"))
        for node in relationships_after.iter(gate._PACKAGE_REL_NS + "Relationship")
    )
    assert sum(before != after for before, after in zip(before_targets, after_targets)) == 1
    assert after_targets == tuple(
        (identifier, canonical_target if identifier == relation_id else target)
        for identifier, target in before_targets
    )
    assert gate._worksheet_member(members_after, _SHEET) == worksheet_member
    reopened = load_workbook(path, read_only=True, data_only=False)
    try:
        assert _SHEET in reopened.sheetnames
        assert reopened[_SHEET].title == _SHEET
    finally:
        reopened.close()


def _write_fixture(
    path: Path, *, duplicate_payload: bool, retain_blank_tail: bool = False
) -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.title = _SHEET
    for col, name in enumerate(("id@id", "value", "version", "text", "note", "flags"), 1):
        ws.cell(1, col).value = name
        ws.cell(2, col).value = "string"
    for row in range(3, 8):
        ws.cell(row, 1).value = f"row-{row}"
        ws.cell(row, 2).value = f"value-{row}"
        ws.cell(row, 3).value = "1.0"
    if duplicate_payload:
        ws.cell(3, 13).value = "=A3&\"-formula\""  # M, blank schema identity.
        for col in range(20, 25):  # T:X, same blank schema identity.
            ws.cell(5, col).value = f"payload-{col}"
        # Real Language has a physical Z payload without a declared schema
        # anchor.  Preserve that ambiguous shape while forcing AA append.
        ws.cell(5, 26).value = "payload-26"
    if retain_blank_tail:
        # Keep a G:Z physical envelope without adding content to G:Y's blank
        # duplicate schema group.  Z is the unique right-hand anchor.
        ws.cell(1, 26).value = "tail"
        ws.cell(2, 26).value = "string"
        ws.cell(3, 26).value = "tail-sentinel"
    workbook.save(path)
    workbook.close()
    _canonicalize_owned_fixture_worksheet_relation(path)


def _copy(source: Path, destination: Path) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, destination)


def _test_ambiguous_language_shape_is_unpublishable(root: Path) -> None:
    source = root / "source.xlsx"
    _write_fixture(source, duplicate_payload=True)
    before = gate._sha256(source)

    original_manifest = sm.snapshot_comparison_oracle_manifest
    original_direct = gate._assert_direct_pair_parity
    original_frozen = gate._assert_frozen_three_way_parity
    original_capture_legacy = gate.capture_legacy
    original_direct_manifest = gate._direct_legacy_manifest
    original_assert_target = gate._assert_target
    original_dual_column = gate._assert_dual_column_conflict_blocked
    direct_calls = []
    target_calls = []
    sm.snapshot_comparison_oracle_manifest = lambda *_args, **_kwargs: (_ for _ in ()).throw(
        AssertionError("expected-safe-unresolved must not publish a candidate manifest")
    )
    gate._assert_direct_pair_parity = lambda *_args, **_kwargs: (_ for _ in ()).throw(
        AssertionError("expected-safe-unresolved must not call exact direct parity")
    )
    gate._assert_frozen_three_way_parity = lambda *_args, **_kwargs: (_ for _ in ()).throw(
        AssertionError("expected-safe-unresolved must not call exact frozen parity")
    )
    gate.capture_legacy = lambda *_args, **_kwargs: (_ for _ in ()).throw(
        AssertionError("expected-safe-unresolved must not wait for frozen legacy READY")
    )
    gate._assert_dual_column_conflict_blocked = lambda *_args, **_kwargs: (_ for _ in ()).throw(
        AssertionError("Language:column expected-safe-unresolved must not enter dual conflict")
    )

    def traced_direct(mine_path, theirs_path, base_path, sheet, *, absolute_deadline=None):
        direct_calls.append((Path(mine_path), Path(theirs_path), base_path, sheet, absolute_deadline))
        return original_direct_manifest(
            mine_path, theirs_path, base_path, sheet, absolute_deadline=absolute_deadline
        )

    def traced_target(manifest, target, *, side):
        target_calls.append((target, side))
        return original_assert_target(manifest, target, side=side)

    gate._direct_legacy_manifest = traced_direct
    gate._assert_target = traced_target
    expected_direct = []
    expected_targets = []
    deadline = time.monotonic() + 45.0
    try:
        for label, mutator in gate._MUTATORS:
            if label == "column":
                continue
            case_root = root / label
            base, mine, theirs = (
                case_root / "base.xlsx",
                case_root / "mine.xlsx",
                case_root / "theirs.xlsx",
            )
            for destination in (base, mine, theirs):
                _copy(source, destination)
            target_theirs = mutator(
                theirs, _SHEET, f"__SOW_FIDELITY_Language_{label}_THEIRS__"
            )
            target_mine = mutator(
                mine, _SHEET, f"__SOW_FIDELITY_Language_{label}_MINE__"
            )
            evidence = gate._assert_language_expected_safe_unresolved(
                label,
                mine,
                theirs,
                base,
                _SHEET,
                target_mine,
                target_theirs,
                absolute_deadline=deadline,
            )
            assert evidence["legacy"] == "two-way-targets-verified-only"
            assert evidence["direct_marker_pairs"] == 2
            assert evidence["column_marker"] is None
            for route in ("two_way", "three_way"):
                assert evidence[route]["state"] == "expected-safe-unresolved"
                assert evidence[route]["blank_duplicate_proof"] == "rejected-nonblank-content"
                assert evidence[route]["nonblank_duplicate_groups"]
            expected_direct.append((base, theirs, None, _SHEET))
            expected_targets.append((target_theirs, "theirs"))
            if target_mine is not None:
                expected_direct.append((base, mine, None, _SHEET))
                expected_targets.append((target_mine, "theirs"))
            assert gate._sha256(source) == before
    finally:
        sm.snapshot_comparison_oracle_manifest = original_manifest
        gate._assert_direct_pair_parity = original_direct
        gate._assert_frozen_three_way_parity = original_frozen
        gate.capture_legacy = original_capture_legacy
        gate._direct_legacy_manifest = original_direct_manifest
        gate._assert_target = original_assert_target
        gate._assert_dual_column_conflict_blocked = original_dual_column
    assert [(left, right, base_path, sheet) for left, right, base_path, sheet, _deadline in direct_calls] == expected_direct
    assert target_calls == expected_targets
    assert len(direct_calls) == len(target_calls) == 6
    assert gate._sha256(source) == before


def _test_column_mutation_stays_raw_and_nonactionable(root: Path) -> None:
    source = root / "column-raw-source.xlsx"
    marker = "__SOW_FIDELITY_Language_column_THEIRS__"
    _write_fixture(source, duplicate_payload=True)
    target = gate._mutate_column(source, _SHEET, marker)
    assert target == gate.MutationTarget("column", 5, 27, marker)
    reopened = load_workbook(source, read_only=True, data_only=False)
    try:
        ws = reopened[_SHEET]
        assert ws.cell(1, 27).value == f"fidelity_{marker}@pm"
        assert ws.cell(2, 27).value == "string"
        assert ws.cell(5, 27).value == marker
    finally:
        reopened.close()


def _test_nonactionable_column_marker_rejects_invalid_manifest() -> None:
    target = gate.MutationTarget("column", 5, 27, "__SOW_FIDELITY_Language_column_THEIRS__")
    token = {
        "present": True,
        "cached": {"type": "str", "value": target.marker},
        "cached_data_type": "s",
        "formula_data_type": "s",
        "formula": None,
    }

    def manifest(*, state="unresolved", ambiguous=True, row=5, logical=27, cell_logical=None):
        cell_logical = logical if cell_logical is None else cell_logical
        return {
            "columns": [{
                "logical": logical,
                "mine": None,
                "base": None,
                "theirs": 27,
                "state": state,
                "ambiguous": ambiguous,
            }],
            "records": [{
                "theirs_row": row,
                "diff_cols": [logical],
                "cells": {str(cell_logical): {"theirs": token}},
            }],
        }

    accepted = gate._assert_nonactionable_ambiguous_column_marker(
        manifest(), target, side="theirs"
    )
    assert accepted["logical"] == 27 and accepted["token"] == token
    for invalid in (
        manifest(state="inserted"),
        manifest(ambiguous=False),
        manifest(row=6),
        manifest(logical=28, cell_logical=27),
    ):
        try:
            gate._assert_nonactionable_ambiguous_column_marker(invalid, target, side="theirs")
        except AssertionError:
            pass
        else:
            raise AssertionError(f"invalid non-actionable column evidence passed: {invalid!r}")


def _test_all_empty_duplicate_tail_stays_resolved(root: Path) -> None:
    source = root / "empty-source.xlsx"
    _write_fixture(source, duplicate_payload=False, retain_blank_tail=True)
    base, mine, theirs = root / "empty-base.xlsx", root / "empty-mine.xlsx", root / "empty-theirs.xlsx"
    for destination in (base, mine, theirs):
        _copy(source, destination)
    gate._mutate_value(mine, _SHEET, _MARKER_MINE)
    gate._mutate_value(theirs, _SHEET, _MARKER_THEIRS)
    snapshots = tuple(
        sm._stream_selected_sheet_snapshot(str(path), str(path), _SHEET, side)
        for path, side in ((mine, "A"), (theirs, "B"), (base, "BASE"))
    )
    assert tuple(snapshot.max_col for snapshot in snapshots) == (26, 26, 26)
    assert all(len(sm._snapshot_field_groups(snapshot).get(("", ""), ())) == 19 for snapshot in snapshots)
    two_alignment = sm._align_selected_sheet_snapshots(snapshots[0], snapshots[1])
    two_proof = sm._try_snapshot_duplicate_field_identity_proof(
        snapshots[0], snapshots[1], two_alignment
    )
    assert two_proof is not None
    mine_base_alignment = sm._align_selected_sheet_snapshots(snapshots[0], snapshots[2])
    theirs_base_alignment = sm._align_selected_sheet_snapshots(snapshots[1], snapshots[2])
    three_proof = sm._try_snapshot_duplicate_field_identity_proof(
        snapshots[0],
        snapshots[1],
        two_alignment,
        snapshots[2],
        mine_base_alignment,
        theirs_base_alignment,
    )
    assert three_proof is not None
    two_way = sm._compare_selected_sheet_snapshots(snapshots[0], snapshots[1])
    three_way = sm._compare_selected_sheet_snapshots(*snapshots)
    assert not two_way.unresolved
    assert not three_way.unresolved


def _test_run_fixture_column_append_is_non_actionable(root: Path) -> None:
    source = root / "column-source.xlsx"
    _write_fixture(source, duplicate_payload=True)
    fixture = SimpleNamespace(name="Language", path=source)
    original_copy = gate.copy_real_fixture
    original_mutators = gate._MUTATORS
    original_mutators_by_label = gate._MUTATORS_BY_LABEL
    original_safe = gate._assert_language_expected_safe_unresolved
    original_dual = gate._assert_dual_column_conflict_blocked
    original_exact = gate._assert_direct_pair_parity
    original_operations = gate._exercise_overlay_save
    helper_calls = []

    def copy_fixture(item, target_root):
        destination = Path(target_root) / "source.xlsx"
        _copy(item.path, destination)
        return destination, _SHEET

    def safe(label, mine, theirs, base, sheet, target_mine, target_theirs, **_kwargs):
        helper_calls.append((label, Path(mine), Path(theirs), Path(base), sheet, target_mine, target_theirs))
        assert label == "column" and target_mine is None and target_theirs.kind == "column"
        assert target_theirs.col == 27
        return {
            "two_way": {"state": "expected-safe-unresolved"},
            "three_way": {"state": "expected-safe-unresolved"},
            "legacy": "two-way-targets-verified-only",
            "direct_marker_pairs": 1,
            "column_marker": {
                "target": target_theirs.__dict__,
                "side": "theirs",
                "logical": 27,
                "state": "unresolved",
                "ambiguous": True,
                "token": {"synthetic": "report-evidence"},
            },
        }

    def forbidden(name):
        return lambda *_args, **_kwargs: (_ for _ in ()).throw(AssertionError(name))

    gate.copy_real_fixture = copy_fixture
    gate._MUTATORS = (("column", gate._mutate_column),)
    gate._MUTATORS_BY_LABEL = {"column": gate._mutate_column}
    gate._assert_language_expected_safe_unresolved = safe
    gate._assert_dual_column_conflict_blocked = forbidden("column safe path entered dual conflict")
    gate._assert_direct_pair_parity = forbidden("column safe path entered exact direct parity")
    gate._exercise_overlay_save = forbidden("column selector invoked operations")
    try:
        result = gate._run_fixture(
            root / "column-run",
            fixture,
            timeout=30.0,
            real_excel=False,
            variant_labels=("column",),
            include_operations=False,
        )
    finally:
        gate.copy_real_fixture = original_copy
        gate._MUTATORS = original_mutators
        gate._MUTATORS_BY_LABEL = original_mutators_by_label
        gate._assert_language_expected_safe_unresolved = original_safe
        gate._assert_dual_column_conflict_blocked = original_dual
        gate._assert_direct_pair_parity = original_exact
        gate._exercise_overlay_save = original_operations

    assert len(helper_calls) == 1
    variant = result["variants"]
    assert len(variant) == 1
    assert variant[0]["kind"] == "column"
    assert variant[0]["mine_target"] is None
    assert variant[0]["theirs_target"]["kind"] == "column"
    assert variant[0]["comparison_mode"] == "expected-safe-unresolved"
    assert variant[0]["snapshot"]["column_marker"]["logical"] == 27


def run_case() -> None:
    with tempfile.TemporaryDirectory(prefix="sow_fidelity_language_unresolved_") as raw_root:
        root = Path(raw_root)
        _test_ambiguous_language_shape_is_unpublishable(root)
        _test_column_mutation_stays_raw_and_nonactionable(root)
        _test_nonactionable_column_marker_rejects_invalid_manifest()
        _test_all_empty_duplicate_tail_stays_resolved(root)
        _test_run_fixture_column_append_is_non_actionable(root)


def main(argv=None) -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--case", choices=(_CASE,))
    parser.add_argument("--list-cases", action="store_true")
    args = parser.parse_args(argv)
    if args.list_cases:
        print(_CASE)
        return
    run_case()
    print(f"PASS {_CASE}")


if __name__ == "__main__":
    main()

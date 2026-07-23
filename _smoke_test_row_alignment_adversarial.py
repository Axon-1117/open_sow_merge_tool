"""Adversarial regressions for shared row identity and structural alignment."""

from __future__ import annotations

import time

from openpyxl import Workbook

import sow_merge_tool as mod


def _sheet(rows):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    for row in rows:
        worksheet.append(list(row))
    return workbook, worksheet


def _pairs(left_rows, right_rows):
    left_book, left = _sheet(left_rows)
    right_book, right = _sheet(right_rows)
    try:
        return mod._compute_row_pairs_generic(
            left,
            right,
            max(left.max_column, right.max_column),
            force=True,
            max_row_a=left.max_row,
            max_row_b=right.max_row,
            ws_a_edit=left,
            ws_b_edit=right,
        )
    finally:
        left_book.close()
        right_book.close()


def _paired_values(pairs, left_rows, right_rows, column=0):
    return [
        (
            left_rows[left - 1][column] if left is not None else None,
            right_rows[right - 1][column] if right is not None else None,
        )
        for left, right in pairs
    ]


def _test_numeric_ordinal_cannot_displace_id():
    left = [("序号", "ID", "值")] + [
        (idx, f"id-{idx}", f"value-{idx}") for idx in range(1, 9)
    ]
    right = [("序号", "ID", "值")] + [
        (1, "id-1", "value-1"),
        (2, "new-id", "new-value"),
        *[
            (idx + 1, f"id-{idx}", f"value-{idx}")
            for idx in range(2, 9)
        ],
    ]
    values = _paired_values(_pairs(left, right), left, right, column=1)
    assert (None, "new-id") in values, values
    assert all(
        left_id == right_id
        for left_id, right_id in values
        if left_id is not None and right_id is not None
    ), values


def _test_text_slot_cannot_displace_composite_key():
    left = [("槽位", "类型", "名称", "值")] + [
        (f"S{idx:02d}", "monster", f"name-{idx}", f"value-{idx}")
        for idx in range(1, 13)
    ]
    right = [("槽位", "类型", "名称", "值")] + [
        (f"S{idx:02d}", "monster", f"name-{source}", f"value-{source}")
        for idx, source in enumerate(
            [1, 2, 3, 5, 6, 7, 8, 9, 10, 11, 12],
            start=1,
        )
    ]
    values = _paired_values(_pairs(left, right), left, right, column=2)
    assert ("name-4", None) in values, values
    assert all(
        left_name == right_name
        for left_name, right_name in values
        if left_name is not None and right_name is not None
    ), values


def _test_three_column_composite_key_middle_delete():
    left = [("组", "阶段", "名称", "值")] + [
        ("world", idx // 4, f"monster-{idx % 4}", f"value-{idx}")
        for idx in range(1, 25)
    ]
    right = [row for row in left if row[:3] != ("world", 3, "monster-1")]
    values = _paired_values(_pairs(left, right), left, right, column=3)
    assert ("value-13", None) in values, values
    assert sum(1 for a, b in values if (a is None) != (b is None)) == 1, values


def _test_duplicate_id_payload_edit_is_not_structural():
    left = [
        ("ID", "名称", "值"),
        ("same", "alpha", 1),
        ("dup", "first", 10),
        ("dup", "second", 20),
        ("tail", "omega", 99),
    ]
    right = [list(row) for row in left]
    right[2][2] = 11
    pairs = _pairs(left, right)
    assert all(a is not None and b is not None for a, b in pairs), pairs
    assert len(pairs) == len(left), pairs


def _test_unique_replacement_and_independent_additions_split():
    left = [
        ("ID", "名称"),
        ("id-1", "one"),
        ("id-2", "two"),
        ("id-3", "three"),
        ("id-4", "four"),
    ]
    right = [
        ("ID", "名称"),
        ("new-id", "new"),
        ("id-2", "two"),
        ("id-3", "three"),
        ("id-4", "four"),
    ]
    values = _paired_values(_pairs(left, right), left, right)
    assert ("id-1", None) in values and (None, "new-id") in values, values
    assert ("id-1", "new-id") not in values, values

    base = [
        ("ID", "名称"),
        ("id-a", "A"),
        ("id-d", "D"),
        ("id-e", "E"),
    ]
    mine = [base[0], base[1], ("id-b", "B"), base[2], base[3]]
    theirs = [base[0], base[1], ("id-c", "C"), base[2], base[3]]
    mine_base = _pairs(mine, base)
    theirs_base = _pairs(theirs, base)
    display = _pairs(mine, theirs)
    reconciled = mod._reconcile_three_way_row_pairs_by_base(
        display,
        mod._row_map_from_pairs(mine_base),
        mod._row_map_from_pairs(theirs_base),
    )
    values = _paired_values(reconciled, mine, theirs)
    assert ("id-b", "id-c") not in values, values
    assert ("id-b", None) in values and (None, "id-c") in values, values


def _test_different_base_identities_never_pair():
    base = [("ID", "名称"), ("id-b", "B"), ("id-c", "C"), ("id-z", "Z")]
    mine = [base[0], base[2], base[3]]
    theirs = [base[0], base[1], base[3]]
    display = _pairs(mine, theirs)
    reconciled = mod._reconcile_three_way_row_pairs_by_base(
        display,
        mod._row_map_from_pairs(_pairs(mine, base)),
        mod._row_map_from_pairs(_pairs(theirs, base)),
    )
    values = _paired_values(reconciled, mine, theirs)
    assert ("id-c", "id-b") not in values, values
    assert ("id-c", None) in values and (None, "id-b") in values, values


def _test_row_formula_transform_and_replay_map():
    assert mod._transform_formula_for_row_structure(
        "=A9+A10+$B$11+SUM(C8:C12)+SUM(10:12)",
        formula_sheet="Data",
        target_sheet="Data",
        anchor=10,
        count=1,
        insert=True,
    ) == "=A9+A11+$B$12+SUM(C8:C13)+SUM(11:13)"
    assert mod._transform_formula_for_row_structure(
        "=Data!A9+Data!A10+Data!A12+SUM(Data!C8:C12)",
        formula_sheet="Other",
        target_sheet="Data",
        anchor=10,
        count=2,
        insert=False,
    ) == "=Data!A9+Data!#REF!+Data!A10+SUM(Data!C8:C10)"

    app = mod.SowMergeApp.__new__(mod.SowMergeApp)
    app.manual_a_cell_ops = {
        ("Data", 20, 1): "=A10",
        ("Other", 1, 1): "=Data!B12",
        ("Other", 2, 1): "=Unrelated!B11",
    }
    app.manual_b_cell_ops = {}
    app.transform_manual_formulas_for_row_action(
        "A", "Data", 10, 2, insert=False
    )
    assert app.manual_a_cell_ops[("Data", 20, 1)] == "=#REF!"
    assert app.manual_a_cell_ops[("Other", 1, 1)] == "=Data!B10"
    assert app.manual_a_cell_ops[("Other", 2, 1)] == "=Unrelated!B11"


def _test_large_sheet_boundary_and_common_performance():
    assert mod._should_auto_row_align(50001, 50000)
    left = [("ID", "名称")] + [
        (f"id-{idx}", f"name-{idx}") for idx in range(1, 10001)
    ]
    right = left[:5001] + [("new-id", "new-name")] + left[5001:]
    started = time.perf_counter()
    pairs = _pairs(left, right)
    elapsed = time.perf_counter() - started
    assert (None, 5002) in pairs, pairs[4995:5008]
    assert elapsed < 8.0, elapsed


def main():
    _test_numeric_ordinal_cannot_displace_id()
    _test_text_slot_cannot_displace_composite_key()
    _test_three_column_composite_key_middle_delete()
    _test_duplicate_id_payload_edit_is_not_structural()
    _test_unique_replacement_and_independent_additions_split()
    _test_different_base_identities_never_pair()
    _test_row_formula_transform_and_replay_map()
    _test_large_sheet_boundary_and_common_performance()
    print("SMOKE_ROW_ALIGNMENT_ADVERSARIAL_OK")


if __name__ == "__main__":
    main()

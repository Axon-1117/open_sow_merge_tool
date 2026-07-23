"""Replay only-diff block navigation on real WorldMonster workbook copies.

Example:
  python _gui_replay_diff_blocks_real.py --base <WorldMonster-r36162.xlsx>
"""

import argparse
import json
import os
import shutil
import time

from openpyxl import load_workbook

import sow_merge_tool as mod
from _test_temp_utils import make_temp_dir


DEFAULT_WORKING = r"C:\GM15\design\sheets\develop\WorldMonster.xlsx"
DEFAULT_SHEET = "WorldMonster@design"
LOCAL_EDIT_ROW = 3000
LOCAL_EDIT_COL = 2


def _pump(root, seconds=0.15):
    deadline = time.time() + seconds
    while time.time() < deadline:
        root.update_idletasks()
        root.update()
        time.sleep(0.01)


def _cell_value(path, sheet, row, col):
    wb = load_workbook(path, read_only=True, data_only=False)
    try:
        return wb[sheet].cell(row=row, column=col).value
    finally:
        wb.close()


def _build_replay_files(base_source, working_source, sheet):
    root_dir = make_temp_dir("sow_real_diff_blocks_")
    base = os.path.join(root_dir, "base.xlsx")
    mine = os.path.join(root_dir, "mine.xlsx")
    theirs = os.path.join(root_dir, "theirs.xlsx")
    merged = os.path.join(root_dir, "merged.xlsx")
    shutil.copy2(base_source, base)
    shutil.copy2(working_source, theirs)

    original = _cell_value(base, sheet, LOCAL_EDIT_ROW, LOCAL_EDIT_COL)
    local_value = f"{original} [local-mine-replay]"
    mod._build_manual_merge_xlsx_via_zip(
        base,
        mine,
        {(sheet, LOCAL_EDIT_ROW, LOCAL_EDIT_COL): local_value},
    )
    return root_dir, base, mine, theirs, merged, local_value


def _pair_rows(view, block):
    rows = set()
    for pair_idx in block.pair_indices:
        pair = view.row_pairs[pair_idx]
        row_a = view._row_for_side(pair, "A")
        if row_a is not None:
            rows.add(int(row_a))
    return rows


def _text_line_count(widget):
    return int(str(widget.index("end-1c")).split(".")[0])


def _tag_covers_line(widget, tag, line):
    ranges = widget.tag_ranges(tag)
    for start, end in zip(ranges[0::2], ranges[1::2]):
        if widget.compare(start, "<=", f"{line}.0") and widget.compare(f"{line}.0", "<", end):
            return True
    return False


def _json_cell_value(value):
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return repr(value)


def _diff_model_snapshot(view, pair_indices):
    """Capture compact raw, visual, Base, and structural replay evidence."""
    cache = view._active_column_comparison_cache()
    pairs = [int(pair_idx) for pair_idx in pair_indices]
    grouped = {}
    for pair_idx in pairs:
        raw = tuple(sorted(view.pair_diff_cols.get(pair_idx, set())))
        base = tuple(sorted(view.pair_base_diff_cols.get(pair_idx, set())))
        visual = tuple(sorted(view._visual_diff_cols_for_pair(pair_idx)))
        key = (raw, base, visual, bool(view._pair_has_visual_diff(pair_idx)))
        grouped.setdefault(key, []).append(pair_idx)
    samples = sorted(set(pairs[:2] + pairs[-2:] + ([pairs[len(pairs) // 2]] if pairs else [])))

    def _sample(pair_idx):
        row_a, row_b = view.row_pairs[pair_idx]
        logical_cols = sorted(
            set(view.pair_diff_cols.get(pair_idx, set()))
            | set(view.pair_base_diff_cols.get(pair_idx, set()))
        )
        cells = {}
        for logical_col in logical_cols:
            if logical_col <= 0 or logical_col > len(cache.model.slots):
                continue
            slot = cache.model.slots[logical_col - 1]
            base_row = view._base_row_for_pair(pair_idx, (row_a, row_b))
            cells[str(logical_col)] = {
                "mine_val": _json_cell_value(
                    view.app.ws_a_val(view.sheet).cell(row_a, slot.mine_col).value
                    if row_a is not None and slot.mine_col is not None else None
                ),
                "mine_edit": _json_cell_value(
                    view.app.ws_a_edit(view.sheet).cell(row_a, slot.mine_col).value
                    if row_a is not None and slot.mine_col is not None else None
                ),
                "base_val": _json_cell_value(
                    view.app.ws_base_val(view.sheet).cell(base_row, slot.base_col).value
                    if base_row is not None and slot.base_col is not None else None
                ),
                "theirs_val": _json_cell_value(
                    view.app.ws_b_val(view.sheet).cell(row_b, slot.theirs_col).value
                    if row_b is not None and slot.theirs_col is not None else None
                ),
                "theirs_edit": _json_cell_value(
                    view.app.ws_b_edit(view.sheet).cell(row_b, slot.theirs_col).value
                    if row_b is not None and slot.theirs_col is not None else None
                ),
            }
        return {
            "rows": [row_a, row_b],
            "raw": sorted(view.pair_diff_cols.get(pair_idx, set())),
            "base": sorted(view.pair_base_diff_cols.get(pair_idx, set())),
            "visual": sorted(view._visual_diff_cols_for_pair(pair_idx)),
            "has_visual": bool(view._pair_has_visual_diff(pair_idx)),
            "cells": cells,
        }

    return {
        "pair_count": len(pairs),
        "groups": [
            {
                "raw": list(key[0]),
                "base": list(key[1]),
                "visual": list(key[2]),
                "has_visual": key[3],
                "count": len(group_pairs),
                "first_pair": group_pairs[0],
                "last_pair": group_pairs[-1],
            }
            for key, group_pairs in grouped.items()
        ],
        "samples": {str(pair_idx): _sample(pair_idx) for pair_idx in samples},
        "structural_diff_cols": sorted(cache.structural_diff_cols),
        "unresolved_cols": sorted(cache.unresolved_cols),
        "slots": [
            {
                "logical": int(slot.logical_idx) + 1,
                "mine": slot.mine_col,
                "base": slot.base_col,
                "theirs": slot.theirs_col,
                "state": slot.state,
                "origin": slot.origin_side,
                "base_boundary": slot.base_boundary,
            }
            for slot in cache.model.slots
        ],
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--base", required=True, help="Historical WorldMonster Base workbook")
    parser.add_argument("--working", default=DEFAULT_WORKING)
    parser.add_argument("--sheet", default=DEFAULT_SHEET)
    args = parser.parse_args()

    root_dir, base, mine, theirs, merged, local_value = _build_replay_files(
        os.path.abspath(args.base),
        os.path.abspath(args.working),
        args.sheet,
    )

    started = time.perf_counter()
    app = mod.SowMergeApp(
        mine,
        theirs,
        merge_mode=True,
        merged_path=merged,
        base_path=base,
    )
    open_seconds = time.perf_counter() - started
    try:
        app.root.state("normal")
        app.root.geometry("1600x900")
        app.nb.select(app._sheet_containers[args.sheet])
        _pump(app.root)
        view = app.sheet_views[args.sheet]
        view._suppress_bg_apply = True
        view.only_diff_var.set(1)

        started = time.perf_counter()
        view.refresh(row_only=None, rescan=True)
        precise_seconds = time.perf_counter() - started
        blocks = view._ensure_full_diff_blocks()
        assert len(blocks) >= 2, [(b.start_pair_idx, b.end_pair_idx) for b in blocks]
        assert len(view._pair_to_full_diff_block) == len(view._full_display_rows)

        target_idx = next(
            idx
            for idx, block in enumerate(blocks[:-1])
            if len(block.pair_indices) > 1
            and any(view.pair_diff_cols.get(pair_idx) for pair_idx in block.pair_indices)
        )
        target_block = blocks[target_idx]
        next_block = blocks[target_idx + 1]
        target_rows = _pair_rows(view, target_block)
        next_pair = next_block.start_pair_idx

        view._goto_full_diff_block(target_idx)
        initial_x = float(view.left.xview()[0])
        rescan_calls = []
        original_refresh = view.refresh

        def _tracked_refresh(row_only, rescan):
            rescan_calls.append(bool(rescan))
            return original_refresh(row_only, rescan)

        view.refresh = _tracked_refresh
        started = time.perf_counter()
        view._goto_full_diff_block(len(blocks) - 1)
        _pump(app.root, 0.1)
        navigate_ms = (time.perf_counter() - started) * 1000.0
        assert view.selected_pair_idx == blocks[-1].start_pair_idx
        assert True not in rescan_calls, rescan_calls
        assert abs(float(view.left.xview()[0]) - initial_x) < 0.02

        # A model-only jump is not sufficient: the target must exist in every
        # active Text widget, carry its block marker/tag, and be in the shared
        # visible vertical range. This guards the historical >800-row bug where
        # data panes expanded but row-number panes remained at the first page.
        target_pair = blocks[-1].start_pair_idx
        target_line = view.row_to_line[target_pair]
        panes = [view.left, view.right, view.left_ln, view.right_ln]
        if view._is_three_way_enabled():
            panes.extend([view.base, view.base_ln])
        assert all(_text_line_count(widget) >= target_line for widget in panes), (
            target_line,
            [_text_line_count(widget) for widget in panes],
        )
        assert view.left.bbox(f"{target_line}.0") is not None
        assert view.left_ln.bbox(f"{target_line}.0") is not None
        assert view.left_ln.get(f"{target_line}.0", f"{target_line}.end").lstrip().startswith(
            f"[{blocks[-1].ordinal}]"
        )
        assert all(_tag_covers_line(widget, "blockstart", target_line) for widget in panes)
        yviews = [widget.yview() for widget in panes]
        assert max(float(v[0]) for v in yviews) - min(float(v[0]) for v in yviews) < 0.002, yviews

        view._goto_full_diff_block(target_idx)
        view.hover_pair_idx = next_pair
        direct_before_next = set(view.pair_diff_cols.get(next_pair, set()))
        assert direct_before_next
        evidence_pairs = list(target_block.pair_indices) + [next_pair]
        model_before_adoption = _diff_model_snapshot(view, evidence_pairs)
        formula_notices = []
        original_showinfo = mod.messagebox.showinfo
        mod.messagebox.showinfo = lambda title, message: formula_notices.append((title, message))
        started = time.perf_counter()
        try:
            view._copy_selected_region("B2A")
            region_seconds = time.perf_counter() - started
        finally:
            mod.messagebox.showinfo = original_showinfo

        written_keys = {
            key
            for key in (
                set(app.manual_a_cell_ops)
                | set(app.manual_a_formula_cache_ops)
            )
            if key[0] == args.sheet
        }
        written_rows = {int(row) for _sheet, row, _col in written_keys}
        model_after_adoption = _diff_model_snapshot(view, evidence_pairs)
        assert written_rows, "Expected real region adoption to record mine writes"
        assert written_rows <= target_rows, (min(written_rows), max(written_rows), target_rows)
        assert not any(view.pair_diff_cols.get(pair_idx) for pair_idx in target_block.pair_indices), (
            "Target block retained raw diffs after adoption",
            json.dumps(
                {
                    "before": model_before_adoption,
                    "after": model_after_adoption,
                },
                ensure_ascii=False,
            ),
        )
        assert view.pair_diff_cols.get(next_pair), "Adjacent block was unexpectedly adopted"
        assert app.ws_a_edit(args.sheet).cell(LOCAL_EDIT_ROW, LOCAL_EDIT_COL).value == local_value

        result = {
            "workspace": root_dir,
            "sheet": args.sheet,
            "open_seconds": round(open_seconds, 3),
            "precise_only_diff_seconds": round(precise_seconds, 3),
            "diff_rows": len(view._full_display_rows),
            "block_count": len(blocks),
            "block_pair_ranges": [
                [block.start_pair_idx, block.end_pair_idx] for block in blocks
            ],
            "cross_block_navigation_ms": round(navigate_ms, 2),
            "navigation_rescan_calls": rescan_calls,
            "adopted_block": target_block.ordinal,
            "adopted_rows": [min(written_rows), max(written_rows)],
            "adopted_cell_count": len(written_keys),
            "region_adoption_seconds": round(region_seconds, 3),
            "formula_dependency_notice_count": len(formula_notices),
            "adjacent_block_unchanged": True,
            "local_mine_edit_preserved": True,
            "model_before_adoption": model_before_adoption,
            "model_after_adoption": model_after_adoption,
        }
        print("REAL_DIFF_BLOCK_REPLAY_OK")
        print(json.dumps(result, ensure_ascii=False, indent=2))
    finally:
        app._shutdown_root()


if __name__ == "__main__":
    main()

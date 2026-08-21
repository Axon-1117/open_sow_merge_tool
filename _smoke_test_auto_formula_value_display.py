"""Regression: formula cells without a saved cache request an explicit refresh.

This test intentionally replaces Excel COM with a local stand-in.  It verifies
that an uncached formula is detected before a refresh is offered, the refreshed
temporary copy is selected only on request, and a refresh failure remains safe.
"""

import os
import tempfile

from openpyxl import Workbook

import sow_merge_tool as smt


def _make_uncached_formula(path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 2
    ws["B1"] = "=A1+3"
    wb.save(path)
    wb.close()


def main() -> None:
    with tempfile.TemporaryDirectory() as temp_dir:
        source = os.path.join(temp_dir, "source.xlsx")
        refreshed = os.path.join(temp_dir, "refreshed.xlsx")
        _make_uncached_formula(source)

        # A freshly saved openpyxl formula has no Excel result cache, so the
        # grid's safe fallback is the formula text before refresh.
        wb_values = smt.load_workbook(source, data_only=True)
        wb_edit = smt.load_workbook(source, data_only=False)
        try:
            assert wb_values["Sheet1"]["B1"].value is None
            assert smt._cell_display_from_values(
                wb_values["Sheet1"]["B1"].value,
                wb_edit["Sheet1"]["B1"].value,
            ) == "=A1+3"
        finally:
            wb_values.close()
            wb_edit.close()

        # Simulate Excel's calculated copy.  Detection must not recalculate on
        # its own; the later, user-confirmed path selects this temporary copy.
        result = Workbook()
        sheet = result.active
        sheet.title = "Sheet1"
        sheet["A1"] = 2
        sheet["B1"] = 5
        result.save(refreshed)
        result.close()

        original_recalc = smt._recalc_with_excel
        original_scan = smt._scan_formula_cache
        try:
            smt._recalc_with_excel = lambda path: refreshed
            smt._scan_formula_cache = lambda path: (True, path == source)
            assert smt._find_missing_formula_cache_paths([
                ("source", source),
                ("cached", refreshed),
            ]) == [("source", source)]
            assert smt._prepare_val_path(source) == source
            assert smt._recalc_and_prepare_val_path(source) == refreshed
        finally:
            smt._recalc_with_excel = original_recalc
            smt._scan_formula_cache = original_scan

        # Failure to launch Excel must not prevent a user from opening the
        # comparison; it simply preserves the formula-text fallback.
        try:
            smt._recalc_with_excel = lambda path: None
            assert smt._recalc_and_prepare_val_path(source) is None
        finally:
            smt._recalc_with_excel = original_recalc

    print("PASS: uncached formulas require confirmation and refresh safely")


if __name__ == "__main__":
    main()

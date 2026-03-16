"""Smoke test for read-only bounds trimming regression.

Validates:
- _effective_bounds keeps full max_col on fallback path.
- _save_values_only_from_wb does not truncate wide columns in read-only mode.
"""

import os

from openpyxl import Workbook, load_workbook

import sow_merge_tool as mod
from _test_temp_utils import make_temp_dir


def _build_sample(path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "S1"
    ws.cell(row=1, column=5).value = "WIDE_COL_VALUE"
    ws.cell(row=10, column=2).value = "TAIL_ROW_VALUE"
    wb.save(path)
    wb.close()


def main():
    root = make_temp_dir(prefix="sow_bounds_trim_")
    src = os.path.join(root, "src.xlsx")
    dst = os.path.join(root, "dst.xlsx")
    _build_sample(src)

    wb_ro = load_workbook(src, data_only=True, read_only=True)
    ws_ro = wb_ro["S1"]

    max_r, max_c = mod._effective_bounds(ws_ro)
    assert (max_r, max_c) == (10, 5), f"Unexpected bounds: {(max_r, max_c)}"

    mod._save_values_only_from_wb(wb_ro, dst)
    wb_ro.close()

    wb_out = load_workbook(dst, data_only=True)
    ws_out = wb_out["S1"]
    assert ws_out.cell(row=1, column=5).value == "WIDE_COL_VALUE", "Column 5 value was truncated"
    assert ws_out.cell(row=10, column=2).value == "TAIL_ROW_VALUE", "Tail row value was truncated"
    wb_out.close()

    print("SMOKE_BOUNDS_TRIM_OK")


if __name__ == "__main__":
    main()

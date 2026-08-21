"""Fresh-process normalized oracle for the current large-sheet engine.

The tool is deliberately read-only: it accepts Mine/Base/Theirs paths and
writes only the requested JSON result.  Fixture mutations belong in a caller's
temporary directory (see the OpenSpec baseline evidence).
"""

from __future__ import annotations

import argparse
import json
import os
import subprocess
import sys
import time
from datetime import date, datetime, time as datetime_time

from openpyxl import load_workbook

import sow_merge_tool as sm


def _typed(value):
    if value is None:
        return {"type": "blank", "value": None}
    if isinstance(value, (datetime, date, datetime_time)):
        return {"type": type(value).__name__, "value": value.isoformat()}
    if isinstance(value, bytes):
        return {"type": "bytes", "value": value.hex()}
    return {"type": type(value).__name__, "value": str(value)}


def _cell_token(value_ws, formula_ws, row, col):
    if row is None or col is None:
        return {"present": False}
    cached = value_ws.cell(int(row), int(col))
    formula = formula_ws.cell(int(row), int(col))
    special = sm._special_formula_signature(formula.value)
    formula_text = sm._formula_text(formula.value)
    return {
        "present": True,
        "cached": _typed(cached.value),
        "cached_data_type": str(cached.data_type or ""),
        "formula_data_type": str(formula.data_type or ""),
        "formula": repr(special) if special is not None else formula_text,
    }


def _token_key(token):
    """Compare the existing engine's typed value/formula view deterministically."""
    return json.dumps(token, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def _three_way_conflicts(cells):
    """Classify only direct, same-cell three-way conflicts from Oracle tokens."""
    conflicts = []
    for logical, values in (cells or {}).items():
        mine, base, theirs = values["mine"], values["base"], values["theirs"]
        if not (mine.get("present") and base.get("present") and theirs.get("present")):
            continue
        if _token_key(mine) != _token_key(base) and _token_key(theirs) != _token_key(base) and _token_key(mine) != _token_key(theirs):
            conflicts.append(int(logical))
    return conflicts


def _pump(app, seconds=0.05):
    end = time.monotonic() + seconds
    while time.monotonic() < end:
        app.root.update_idletasks()
        app.root.update()
        time.sleep(0.01)


def _worker(args):
    # The frozen Oracle is a non-interactive subprocess.  Keep the production
    # confirmation unchanged, but prevent a missing formula-cache prompt from
    # blocking its Tk event pump forever.  This affects this test worker only.
    original_prompt_scheduler = sm.SowMergeApp._schedule_formula_cache_prompt
    sm.SowMergeApp._schedule_formula_cache_prompt = lambda _self: None
    try:
        app = sm.SowMergeApp(args.mine, args.theirs, merge_mode=bool(args.base), base_path=args.base)
    except Exception:
        sm.SowMergeApp._schedule_formula_cache_prompt = original_prompt_scheduler
        raise
    value_books = []
    formula_books = []
    try:
        # Oracle callers can request any tab, not just workbook.sheetnames[0].
        # Dispatch it through the same lazy selected-Sheet lifecycle that a
        # real user uses before waiting for its exact cache.
        app._select_tab(args.sheet)
        deadline = time.monotonic() + args.timeout
        view = None
        while time.monotonic() < deadline:
            _pump(app)
            view = app.sheet_views.get(args.sheet)
            if view is not None and view._data_ready and view._pair_diff_full_exact:
                break
        if view is None or not view._data_ready or not view._pair_diff_full_exact:
            raise RuntimeError("legacy exact view did not become READY")
        paths = [("mine", args.mine), ("theirs", args.theirs)]
        if args.base:
            paths.insert(1, ("base", args.base))
        values = {}
        formulas = {}
        for side, path in paths:
            values[side] = load_workbook(path, data_only=True, read_only=False)
            formulas[side] = load_workbook(path, data_only=False, read_only=False)
            value_books.append(values[side])
            formula_books.append(formulas[side])
        projection = view._active_column_projection()
        slots = [
            {
                "logical": slot.logical_idx + 1,
                "mine": slot.mine_col,
                "base": slot.base_col,
                "theirs": slot.theirs_col,
                "state": slot.state,
                "ambiguous": bool(slot.confidence.ambiguous),
            }
            for slot in projection.model.slots
        ]
        records = []
        for index, (mine_row, theirs_row) in enumerate(view.row_pairs):
            changed = sorted(int(c) for c in view.pair_diff_cols.get(index, ()) if int(c) != -1)
            base_changed = sorted(int(c) for c in view.pair_base_diff_cols.get(index, ()) if int(c) != -1)
            if not changed and not base_changed and mine_row is not None and theirs_row is not None:
                continue
            cells = {}
            for logical in sorted(set(changed) | set(base_changed)):
                slot = projection.slot(logical)
                if slot is None:
                    continue
                cells[str(logical)] = {
                    "mine": _cell_token(values["mine"][args.sheet], formulas["mine"][args.sheet], mine_row, slot.mine_col),
                    "theirs": _cell_token(values["theirs"][args.sheet], formulas["theirs"][args.sheet], theirs_row, slot.theirs_col),
                    "base": (
                        _cell_token(
                            values["base"][args.sheet], formulas["base"][args.sheet],
                            view._base_row_for_pair(index, (mine_row, theirs_row)),
                            slot.base_col,
                        )
                        if args.base else {"present": False}
                    ),
                }
            records.append({
                "pair": index,
                "mine_row": mine_row,
                "theirs_row": theirs_row,
                "base_row": view._base_row_for_pair(index, (mine_row, theirs_row)) if args.base else None,
                "row_structure": -1 in view.pair_diff_cols.get(index, set()),
                "diff_cols": changed,
                "base_diff_cols": base_changed,
                "conflicts": _three_way_conflicts(cells) if args.base else [],
                "cells": cells,
            })
        manifest = {
            "schema": "legacy-large-sheet-oracle-v1",
            "sheet": args.sheet,
            "three_way": bool(args.base),
            "sides": [side for side, _path in paths],
            "columns": slots,
            "only_diff_rows": list(view._only_diff_rows_cache or ()),
            "records": records,
        }
        with open(args.out, "w", encoding="utf-8", newline="\n") as f:
            json.dump(manifest, f, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    finally:
        for wb in value_books + formula_books:
            wb.close()
        try:
            app._shutdown_root()
        finally:
            sm.SowMergeApp._schedule_formula_cache_prompt = original_prompt_scheduler


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--mine", required=True)
    parser.add_argument("--theirs", required=True)
    parser.add_argument("--base")
    parser.add_argument("--sheet", required=True)
    parser.add_argument("--out", required=True)
    parser.add_argument("--timeout", type=float, default=90.0)
    parser.add_argument("--worker", action="store_true")
    args = parser.parse_args()
    if args.worker:
        _worker(args)
        return
    command = [sys.executable, os.path.abspath(__file__), "--worker"]
    for name in ("mine", "theirs", "sheet", "out", "base", "timeout"):
        value = getattr(args, name)
        if value is not None:
            command.extend((f"--{name}", str(value)))
    subprocess.run(command, check=True)


if __name__ == "__main__":
    main()

"""OpenSpec 3.5: logical-column action, rollback, and undo regression.

Run:
  python _gui_self_test_logical_column_actions.py

The contract checks are intentionally widget-free.  The GUI checks use small
real workbooks so insert/delete/copy actions and one-step undo are verified
against openpyxl state rather than mocks.
"""

from __future__ import annotations

import copy
import os
import time
import hashlib
import gc
import weakref
from dataclasses import fields
from types import SimpleNamespace

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.xml.functions import tostring

import sow_merge_tool as smt
from _test_temp_utils import make_temp_dir


_GUIDE_ORIGINAL = r"C:\tmp\column_alignment_baseline\Guide\original.xlsx"
_GUIDE_INSERT2_DELETE1 = r"C:\tmp\column_alignment_baseline\Guide\insert2_delete1.xlsx"
_GUIDE_SHEET = "TGuideStep@design"


_OP_FIELDS = {
    "kind",
    "sheet",
    "target_side",
    "target_logical_slot",
    "target_physical_anchor",
    "count",
    "source_side",
    "source_physical_cols",
    "metadata_scope",
    "batch_id",
    "action_id",
    "order",
}


class _Var:
    def __init__(self, value=0):
        self.value = value

    def get(self):
        return self.value


class _ButtonState:
    """Tiny Tk-button substitute for readiness/selection contract checks."""

    def __init__(self):
        self.options = {}

    def configure(self, **kwargs):
        self.options.update(kwargs)


class _CountingCells(dict):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.scan_count = 0

    def __iter__(self):
        self.scan_count += 1
        return super().__iter__()


class _CountingWorksheet:
    def __init__(self, rows):
        self.rows = {int(row): tuple(values) for row, values in rows.items()}
        self._cells = _CountingCells({
            (row, col): value
            for row, values in self.rows.items()
            for col, value in enumerate(values, start=1)
            if value is not None
        })
        self.max_column_reads = 0
        self.iter_rows_calls = 0

    @property
    def max_column(self):
        self.max_column_reads += 1
        return max((col for _row, col in self._cells), default=1)

    def iter_rows(self, *, min_row, max_row, min_col, max_col, values_only):
        assert values_only is True
        self.iter_rows_calls += 1
        for row in range(int(min_row), int(max_row) + 1):
            values = self.rows.get(row, ())
            yield tuple(
                values[col - 1] if col - 1 < len(values) else None
                for col in range(int(min_col), int(max_col) + 1)
            )


def _rows(columns, count: int = 8):
    columns = tuple(columns)
    result = [columns]
    for row_no in range(1, count + 1):
        result.append(tuple(f"{name.lower()}-{row_no}-{'v' * 18}" for name in columns))
    return result


def _key(name: str):
    return smt.ColumnModelCacheKey(name, 1, 1)


def _cache_2way(name, mine, theirs):
    return smt.build_logical_column_comparison_cache_2way(
        _key(name),
        mine,
        theirs,
        mine,
        theirs,
        mine_max_col=len(mine[0]),
        theirs_max_col=len(theirs[0]),
    )


def _cache_3way(name, mine, base, theirs):
    return smt.build_logical_column_comparison_cache_3way(
        _key(name),
        mine,
        base,
        theirs,
        mine,
        base,
        theirs,
        mine_max_col=len(mine[0]),
        base_max_col=len(base[0]),
        theirs_max_col=len(theirs[0]),
    )


def _fake_action_view(cache, *, three_way=False):
    view = object.__new__(smt.SheetView)
    view.sheet = cache.model.cache_key.sheet_name
    view.app = SimpleNamespace(
        merge_mode=bool(three_way),
        has_base=bool(three_way),
        _manual_column_action_seq=0,
        next_manual_column_action_id=lambda: "column-action-fake",
    )
    view.three_way_var = _Var(1 if three_way else 0)
    view.column_comparison_cache = cache
    view.column_projection = smt.LogicalColumnProjection.from_model(cache.model)
    view._row_model_version = cache.model.cache_key.row_model_version
    view._column_model_version = cache.model.cache_key.column_model_version
    view._mine_edit_version = cache.model.cache_key.mine_edit_version
    view._base_edit_version = cache.model.cache_key.base_edit_version
    view._theirs_edit_version = cache.model.cache_key.theirs_edit_version
    view._column_mapping_stale_reason = ""
    view._column_projection_generation = 1
    view.selected_column_block_ordinal = None
    view.selected_column_logical_range = None
    view.selected_column_source_side = None
    view._selected_column_projection_generation = None
    return view


def _raises_runtime(call, contains: str):
    try:
        call()
    except RuntimeError as exc:
        assert contains in str(exc), str(exc)
        return str(exc)
    raise AssertionError(f"expected RuntimeError containing {contains!r}")


def _cell_snapshot(cell):
    hyperlink = cell.hyperlink
    return (
        cell.value,
        cell.data_type,
        tuple(cell._style) if cell.has_style and cell._style is not None else None,
        cell.number_format,
        (cell.comment.text, cell.comment.author) if cell.comment is not None else None,
        (
            hyperlink.target,
            hyperlink.location,
            hyperlink.tooltip,
            hyperlink.display,
        ) if hyperlink is not None else None,
    )


def _worksheet_snapshot(ws):
    max_row = max(1, int(ws.max_row or 1))
    max_col = max(1, int(ws.max_column or 1))
    # Iterate only already-existing cells.  Calling ws.cell() across the
    # rectangular dimension would itself materialize blank tail cells and
    # make the regression hide exactly the corruption it is meant to catch.
    cell_entries = tuple(
        (int(row), int(col), _cell_snapshot(cell))
        for (row, col), cell in sorted(ws._cells.items())
    )
    dimensions = tuple(
        (
            key,
            dim.width,
            bool(dim.hidden),
            dim.style_id,
            dim.outlineLevel,
            bool(dim.collapsed),
            dim.bestFit,
            dim.min,
            dim.max,
        )
        for key, dim in sorted(ws.column_dimensions.items())
    )
    validations = tuple(sorted(
        (
            str(dv.sqref),
            dv.type,
            dv.operator,
            dv.formula1,
            dv.formula2,
            dv.allow_blank,
            dv.showErrorMessage,
            dv.error,
            dv.errorTitle,
            dv.prompt,
            dv.promptTitle,
        )
        for dv in ws.data_validations.dataValidation
    ))
    # Preserve conditional-formatting container grouping and rule order.
    # A single ConditionalFormatting object may intentionally own many sqref
    # ranges; splitting it into one object per range is not an exact undo.
    conditional = tuple(
        (
            str(getattr(container, "sqref", "")),
            tuple(tostring(rule.to_tree()) for rule in rules),
        )
        for container, rules in ws.conditional_formatting._cf_rules.items()
    )
    return (
        max_row,
        max_col,
        cell_entries,
        dimensions,
        validations,
        conditional,
        tuple(sorted(str(value) for value in ws.merged_cells.ranges)),
    )


def _column_snapshot(ws, col: int):
    col = int(col)
    letter = get_column_letter(col)
    dimension = ws.column_dimensions.get(letter)
    validation_ranges = []
    for dv in ws.data_validations.dataValidation:
        for rng in dv.ranges.ranges:
            if rng.min_col <= col <= rng.max_col:
                validation_ranges.append((
                    rng.min_row,
                    rng.max_row,
                    dv.type,
                    dv.operator,
                    dv.formula1,
                    dv.formula2,
                    dv.allow_blank,
                    dv.error,
                    dv.errorTitle,
                ))
    return (
        tuple(_cell_snapshot(ws.cell(row=row, column=col)) for row in range(1, ws.max_row + 1)),
        (
            dimension.width,
            bool(dimension.hidden),
            dimension.style_id,
            dimension.outlineLevel,
            bool(dimension.collapsed),
            dimension.bestFit,
        ) if dimension is not None else None,
        tuple(sorted(validation_ranges)),
    )


def _selection_snapshot(view):
    return (
        getattr(view, "selected_column_block_ordinal", None),
        getattr(view, "selected_column_logical_range", None),
        getattr(view, "selected_column_source_side", None),
        getattr(view, "_main_sel_col", None),
        getattr(view, "_main_sel_line", None),
        getattr(view, "selected_pair_idx", None),
        getattr(view, "selected_excel_row", None),
        getattr(view, "selected_excel_row_a", None),
        getattr(view, "selected_excel_row_b", None),
    )


def _model_snapshot(view):
    projection = view._ensure_column_projection_current("测试列模型快照")
    return (
        tuple(projection.model.slots),
        tuple(projection.model.blocks),
        frozenset(view.column_comparison_cache.structural_diff_cols),
        frozenset(view.column_comparison_cache.unresolved_cols),
        tuple(view.row_pairs),
        tuple(sorted((idx, tuple(sorted(cols))) for idx, cols in view.pair_diff_cols.items())),
        tuple(sorted((idx, tuple(sorted(cols))) for idx, cols in view.pair_base_diff_cols.items())),
        tuple(getattr(view, "logical_column_states", ()) or ()),
        tuple(getattr(view, "logical_column_structural_conflicts", ()) or ()),
    )


def _pump(root, seconds=0.03):
    deadline = time.time() + seconds
    while time.time() < deadline:
        root.update()
        time.sleep(0.005)


def _wait_for_view(app, sheet="Sheet1", timeout=12.0):
    app.nb.select(app._sheet_containers[sheet])
    deadline = time.time() + timeout
    view = None
    while time.time() < deadline:
        _pump(app.root)
        view = app.sheet_views.get(sheet)
        if view is not None and getattr(view, "_data_ready", False):
            return view
    raise AssertionError(f"column action GUI view did not become ready: {view!r}")


def _force_full_view(view):
    if bool(view.only_diff_var.get()):
        view.only_diff_var.set(0)
        view._toggle_only_diff()
    view.refresh(row_only=None, rescan=False)
    _pump(view.root, 0.05)
    return view


def _stable_projection_fingerprint(view):
    projection = view._active_column_projection()
    return (
        tuple(projection.model.slots),
        tuple(projection.model.blocks),
        frozenset(view.column_comparison_cache.structural_diff_cols),
        frozenset(view.column_comparison_cache.unresolved_cols),
    )


def _wait_for_stable_projection(view, timeout=12.0, stable_for=0.75):
    deadline = time.time() + timeout
    previous = None
    stable_since = None
    while time.time() < deadline:
        _pump(view.root, 0.05)
        if not view._column_mapping_is_current():
            previous = None
            stable_since = None
            continue
        current = _stable_projection_fingerprint(view)
        if current == previous:
            if stable_since is not None and time.time() - stable_since >= stable_for:
                return current
        else:
            previous = current
            stable_since = time.time()
    raise AssertionError("logical column projection did not become stable")


def _sha256(path: str):
    digest = hashlib.sha256()
    with open(path, "rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _populate_action_sheet(ws, headers, *, decorated=()):
    decorated = set(int(value) for value in decorated)
    for col, header in enumerate(headers, start=1):
        ws.cell(1, col).value = header
        for row in range(2, 10):
            cell = ws.cell(row, col)
            # Keep the formula structurally identical across retained columns;
            # physical-reference shifts would intentionally trigger the
            # conservative alignment fallback before this action test begins.
            cell.value = "=1+ROW()" if row == 3 else f"{header}-{row}"
        if col in decorated:
            letter = get_column_letter(col)
            ws.column_dimensions[letter].width = 17.0 + col
            ws.column_dimensions[letter].hidden = (col % 2 == 0)
            target = ws.cell(2, col)
            target.fill = PatternFill("solid", fgColor="33AA77")
            target.font = Font(name="Calibri", bold=True, color="FFFFFF")
            target.alignment = Alignment(horizontal="center")
            target.border = Border(left=Side(style="thin", color="FF0000"))
            target.number_format = "0.00"
            target.comment = Comment(f"comment-{header}", "Codex")
            target.hyperlink = f"https://example.test/{header}"
            validation = DataValidation(type="list", formula1='"one,two,three"', allow_blank=True)
            validation.error = f"error-{header}"
            validation.errorTitle = f"title-{header}"
            validation.add(f"{letter}2:{letter}9")
            ws.add_data_validation(validation)
            ws.conditional_formatting.add(
                f"{letter}2:{letter}9",
                CellIsRule(operator="equal", formula=["1"], fill=PatternFill("solid", fgColor="FFFF00")),
            )


def _write_action_book(path: str, headers, *, decorated=()):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    _populate_action_sheet(ws, headers, decorated=decorated)
    wb.save(path)


def _write_formula_cache_asymmetry_book(path: str, headers, *, cached_formula=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for col, header in enumerate(headers, start=1):
        ws.cell(1, col).value = header
        if header == "id":
            ws.cell(2, col).value = 1
        elif header == "calc":
            ws.cell(2, col).value = "=A2+1"
        elif header == "tail":
            ws.cell(2, col).value = "tail-2"
        else:
            ws.cell(2, col).value = "new-2"
    if cached_formula is None:
        wb.save(path)
        wb.close()
        return
    raw = path + ".raw.xlsx"
    wb.save(raw)
    wb.close()
    calc_col = tuple(headers).index("calc") + 1
    smt._build_manual_merge_xlsx_via_zip(
        raw,
        path,
        {("Sheet1", 2, calc_col): "=A2+1"},
        cached_values={("Sheet1", 2, calc_col): cached_formula},
    )


def _write_excel_semantic_formula_structure_book(path: str, mode: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws2 = wb.create_sheet("Sheet2")
    for row in range(1, 8):
        ws2.cell(row, 2).value = row * 10

    if mode == "original":
        headers = ("id", "source", "moved", "fixed", "conservative")
    elif mode == "inserted":
        headers = ("id", "new", "source", "moved", "fixed", "conservative")
    elif mode == "deleted":
        headers = ("id", "moved", "fixed", "conservative")
    else:
        raise ValueError(mode)
    ws.append(headers)

    for row in range(2, 8):
        if mode == "original":
            values = (
                row,
                f"=A{row}+1",
                f"=B{row}*2",
                f"=A{row}*3",
                f"=Sheet2!B{row}+NamedRange+SUM(Table1[Amount])",
            )
        elif mode == "inserted":
            values = (
                row,
                f"new-{row}",
                f"=A{row}+1",
                f"=C{row}*2",
                f"=A{row}*3",
                f"=Sheet2!B{row}+NamedRange+SUM(Table1[Amount])",
            )
        else:
            values = (
                row,
                "=#REF!*2",
                f"=A{row}*3",
                f"=Sheet2!B{row}+NamedRange+SUM(Table1[Amount])",
            )
        ws.append(values)
    wb.save(path)
    wb.close()


def _append_action_sheet(path: str, sheet: str, headers, *, decorated=()):
    wb = load_workbook(path, data_only=False)
    ws = wb.create_sheet(sheet)
    _populate_action_sheet(ws, headers, decorated=decorated)
    wb.save(path)


def _real_two_way_app(*, decorated_a=(), decorated_b=(2, 3)):
    dir_a = make_temp_dir("sow_column_action_a_")
    dir_b = make_temp_dir("sow_column_action_b_")
    path_a = os.path.join(dir_a, "same.xlsx")
    path_b = os.path.join(dir_b, "same.xlsx")
    _write_action_book(
        path_a,
        ("A", "B", "C", "D", "E", "F"),
        decorated=decorated_a,
    )
    _write_action_book(
        path_b,
        ("A", "X", "Y", "B", "D", "E", "F"),
        decorated=decorated_b,
    )
    app = smt.SowMergeApp(path_a, path_b)
    return app, _force_full_view(_wait_for_view(app))


def _real_three_way_app(
    mine_headers,
    base_headers,
    theirs_headers,
    *,
    decorated_mine=(),
    decorated_base=(),
    decorated_theirs=(),
    conflict_map=None,
):
    dir_mine = make_temp_dir("sow_column_action_mine_")
    dir_base = make_temp_dir("sow_column_action_base_")
    dir_theirs = make_temp_dir("sow_column_action_theirs_")
    mine_path = os.path.join(dir_mine, "same.xlsx")
    base_path = os.path.join(dir_base, "same.xlsx")
    theirs_path = os.path.join(dir_theirs, "same.xlsx")
    _write_action_book(mine_path, mine_headers, decorated=decorated_mine)
    _write_action_book(base_path, base_headers, decorated=decorated_base)
    _write_action_book(theirs_path, theirs_headers, decorated=decorated_theirs)
    app = smt.SowMergeApp(
        mine_path,
        theirs_path,
        merge_mode=True,
        base_path=base_path,
        raw_mine=mine_path,
        raw_base=base_path,
        raw_theirs=theirs_path,
        merge_conflict_cells_by_sheet=conflict_map,
        merge_conflict_mode=bool(conflict_map),
    )
    return app, _force_full_view(_wait_for_view(app))


def _real_two_sheet_app():
    dir_a = make_temp_dir("sow_column_action_two_sheet_a_")
    dir_b = make_temp_dir("sow_column_action_two_sheet_b_")
    path_a = os.path.join(dir_a, "same.xlsx")
    path_b = os.path.join(dir_b, "same.xlsx")
    mine = ("A", "B", "C", "D", "E", "F")
    theirs = ("A", "X", "Y", "B", "D", "E", "F")
    _write_action_book(path_a, mine)
    _write_action_book(path_b, theirs, decorated=(2, 3))
    _append_action_sheet(path_a, "Sheet2", mine)
    _append_action_sheet(path_b, "Sheet2", theirs, decorated=(2, 3))
    app = smt.SowMergeApp(path_a, path_b)
    view1 = _force_full_view(_wait_for_view(app, "Sheet1"))
    view2 = _force_full_view(_wait_for_view(app, "Sheet2"))
    _wait_for_stable_projection(view1)
    _wait_for_stable_projection(view2)
    return app, view1, view2


def _app_for_manual_column_records():
    return SimpleNamespace(
        manual_a_column_ops=[],
        manual_b_column_ops=[],
        _manual_column_action_seq=0,
        _manual_column_op_seq=0,
    )


def _record(app, side: str, operations):
    return smt.SowMergeApp.record_manual_column_operations(app, side, operations)


def _action_id(app):
    return smt.SowMergeApp.next_manual_column_action_id(app)


def _plan(
    *,
    action_id: str,
    kind: str,
    source: str,
    target: str,
    logical_start: int,
    logical_end: int,
    source_cols=(),
    target_cols=(),
    anchor: int,
    unresolved: bool = False,
):
    return smt.ColumnBlockActionPlan(
        action_id=action_id,
        sheet="Sheet1",
        block_ordinal=1,
        logical_start=logical_start,
        logical_end=logical_end,
        source_side=source,
        target_side=target,
        source_physical_cols=tuple(source_cols),
        target_physical_cols=tuple(target_cols),
        target_physical_anchor=anchor,
        count=logical_end - logical_start + 1,
        action_kind=kind,
        unresolved=unresolved,
    )


def test_column_action_plan_and_operation_shape():
    expected_plan_fields = {
        "action_id",
        "sheet",
        "block_ordinal",
        "logical_start",
        "logical_end",
        "source_side",
        "target_side",
        "source_physical_cols",
        "target_physical_cols",
        "target_physical_anchor",
        "count",
        "action_kind",
        "unresolved",
        "metadata_scope",
    }
    assert {field.name for field in fields(smt.ColumnBlockActionPlan)} == expected_plan_fields

    app = _app_for_manual_column_records()
    action_id = _action_id(app)
    plan = _plan(
        action_id=action_id,
        kind="insert_copy",
        source="B",
        target="A",
        logical_start=2,
        logical_end=3,
        source_cols=(2, 3),
        anchor=2,
    )
    raw = plan.operation_records()
    assert [op["kind"] for op in raw] == ["insert_cols", "copy_cols"]
    assert all("order" not in op for op in raw)
    recorded = _record(app, "A", raw)
    assert recorded == app.manual_a_column_ops
    assert app.manual_b_column_ops == []
    assert [op["order"] for op in recorded] == [1, 2]
    for op in recorded:
        assert set(op) == _OP_FIELDS, op
        assert op["sheet"] == "Sheet1"
        assert op["target_side"] == "A"
        assert op["target_logical_slot"] == 2
        assert op["target_physical_anchor"] == 2
        assert op["count"] == 2
        assert op["source_side"] == "B"
        assert op["source_physical_cols"] == [2, 3]
        assert op["metadata_scope"] == list(smt._COLUMN_ACTION_METADATA_SCOPE)
        assert op["batch_id"] == op["action_id"] == action_id


def test_delete_copy_retain_and_global_operation_order():
    app = _app_for_manual_column_records()

    delete_id = _action_id(app)
    deleted = _record(
        app,
        "B",
        _plan(
            action_id=delete_id,
            kind="delete",
            source="A",
            target="B",
            logical_start=5,
            logical_end=5,
            target_cols=(5,),
            anchor=5,
        ).operation_records(),
    )
    copy_id = _action_id(app)
    copied = _record(
        app,
        "A",
        _plan(
            action_id=copy_id,
            kind="copy",
            source="BASE",
            target="A",
            logical_start=4,
            logical_end=4,
            source_cols=(3,),
            target_cols=(4,),
            anchor=4,
        ).operation_records(),
    )
    retain_id = _action_id(app)
    retained = _plan(
        action_id=retain_id,
        kind="retain",
        source="A",
        target="A",
        logical_start=2,
        logical_end=2,
        source_cols=(2,),
        target_cols=(2,),
        anchor=2,
    )

    assert [op["kind"] for op in deleted] == ["delete_cols"]
    assert [op["kind"] for op in copied] == ["copy_cols"]
    assert retained.operation_records() == ()
    assert app.manual_b_column_ops[0]["order"] == 1
    assert app.manual_a_column_ops[0]["order"] == 2
    assert delete_id != copy_id != retain_id


def test_consecutive_columns_stay_one_batch_and_do_not_expand_neighbors():
    app = _app_for_manual_column_records()
    action_id = _action_id(app)
    plan = _plan(
        action_id=action_id,
        kind="insert_copy",
        source="B",
        target="A",
        logical_start=7,
        logical_end=9,
        source_cols=(6, 7, 8),
        anchor=6,
    )
    recorded = _record(app, "A", plan.operation_records())

    assert len(recorded) == 2
    assert [op["count"] for op in recorded] == [3, 3]
    assert [op["target_logical_slot"] for op in recorded] == [7, 7]
    assert [op["target_physical_anchor"] for op in recorded] == [6, 6]
    assert [op["source_physical_cols"] for op in recorded] == [[6, 7, 8], [6, 7, 8]]
    assert {op["action_id"] for op in recorded} == {action_id}
    assert {op["batch_id"] for op in recorded} == {action_id}


def test_invalid_column_operation_batch_is_rejected_atomically():
    app = _app_for_manual_column_records()
    action_id = _action_id(app)
    good = dict(
        kind="insert_cols",
        sheet="Sheet1",
        target_side="A",
        target_logical_slot=2,
        target_physical_anchor=2,
        count=2,
        source_side="B",
        source_physical_cols=[2, 3],
        metadata_scope=list(smt._COLUMN_ACTION_METADATA_SCOPE),
        batch_id=action_id,
        action_id=action_id,
    )

    malformed = []
    malformed.append([{**good, "kind": "move_cols"}])
    malformed.append([{**good, "target_side": "B"}])
    malformed.append([{**good, "count": 0}])
    malformed.append([{**good, "target_physical_anchor": 0}])
    malformed.append([{**good, "target_logical_slot": 0}])
    malformed.append([{**good, "source_side": "UNKNOWN"}])
    malformed.append([{**good, "batch_id": "different"}])
    malformed.append([
        good,
        {**good, "kind": "copy_cols", "action_id": "different", "batch_id": "different"},
    ])
    malformed.append([
        {**good, "kind": "copy_cols", "source_physical_cols": [2]},
    ])

    for operations in malformed:
        before_a = list(app.manual_a_column_ops)
        before_b = list(app.manual_b_column_ops)
        try:
            _record(app, "A", operations)
        except ValueError:
            pass
        else:
            raise AssertionError(f"invalid column operation was accepted: {operations!r}")
        assert app.manual_a_column_ops == before_a
        assert app.manual_b_column_ops == before_b


def test_column_action_plan_validation():
    base = dict(
        action_id="column-action-000001",
        sheet="Sheet1",
        block_ordinal=0,
        logical_start=2,
        logical_end=3,
        source_side="B",
        target_side="A",
        source_physical_cols=(2, 3),
        target_physical_cols=(),
        target_physical_anchor=2,
        count=2,
        action_kind="insert_copy",
    )
    invalid = [
        {**base, "action_kind": "move"},
        {**base, "source_side": "THEIRS"},
        {**base, "target_side": "BASE"},
        {**base, "logical_start": 0},
        {**base, "logical_end": 1},
        {**base, "count": 1},
        {**base, "target_physical_anchor": 0},
    ]
    for kwargs in invalid:
        try:
            smt.ColumnBlockActionPlan(**kwargs)
        except ValueError:
            pass
        else:
            raise AssertionError(f"invalid action plan was accepted: {kwargs!r}")


def test_row_cache_reuses_physical_width_scan_for_thousands_of_calls():
    ws = _CountingWorksheet({
        row: (f"a-{row}", f"b-{row}", f"c-{row}", f"d-{row}")
        for row in range(1, 21)
    })
    for index in range(2000):
        row = index % 20 + 1
        cached = smt._read_rows_into_cache(ws, (row,), 12)
        assert len(cached[row]) == 12
        assert cached[row][:4] == (
            f"a-{row}", f"b-{row}", f"c-{row}", f"d-{row}",
        )
        assert cached[row][4:] == (None,) * 8
    assert ws._cells.scan_count == 1, ws._cells.scan_count
    assert ws.max_column_reads == 0, ws.max_column_reads
    assert ws.iter_rows_calls == 2000

    # Size change invalidates the cheap token automatically.
    ws._cells[(1, 8)] = "tail-8"
    assert smt._worksheet_cached_physical_max_column(ws) == 8
    assert ws._cells.scan_count == 2

    # A same-size structural move needs the explicit invalidation hook.
    del ws._cells[(1, 8)]
    ws._cells[(1, 10)] = "tail-10"
    smt._invalidate_worksheet_read_horizon(ws)
    assert smt._worksheet_cached_physical_max_column(ws) == 10
    assert ws._cells.scan_count == 3


def test_real_worksheet_wide_reads_do_not_materialize_tail_and_invalidation_updates_width():
    wb = Workbook()
    ws = wb.active
    for row in range(1, 6):
        for col in range(1, 5):
            ws.cell(row, col).value = f"r{row}c{col}"
    cells_before = set(ws._cells)
    assert smt._worksheet_cached_physical_max_column(ws) == 4
    for _ in range(50):
        rows = smt._read_rows_into_cache(ws, range(1, 6), 25)
        assert all(len(row) == 25 for row in rows.values())
        assert all(row[4:] == (None,) * 21 for row in rows.values())
    assert set(ws._cells) == cells_before
    assert ws.max_column == 4

    # openpyxl moves the same cell objects/keys for structural edits, so cache
    # invalidation is explicit even when the cell-map size does not change.
    ws.insert_cols(2, 3)
    smt._invalidate_worksheet_read_horizon(ws)
    assert smt._worksheet_cached_physical_max_column(ws) == 7
    inserted = smt._read_rows_into_cache(ws, (1,), 10)[1]
    assert inserted[:7] == ("r1c1", None, None, None, "r1c2", "r1c3", "r1c4")
    assert inserted[7:] == (None,) * 3

    ws.delete_cols(2, 2)
    smt._invalidate_worksheet_read_horizon(ws)
    assert smt._worksheet_cached_physical_max_column(ws) == 5
    deleted = smt._read_rows_into_cache(ws, (1,), 10)[1]
    assert deleted[:5] == ("r1c1", None, "r1c2", "r1c3", "r1c4")
    assert deleted[5:] == (None,) * 5


def test_sparse_regular_and_read_only_row_cache_are_equivalent_without_materialization():
    root_dir = make_temp_dir("sow_sparse_row_cache_")
    path = os.path.join(root_dir, "sparse.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Sparse"
    ws["A1"] = "head"
    ws["M50"] = "middle"
    ws["AG885"] = "tail"
    wb.save(path)

    regular = load_workbook(path, data_only=False)
    readonly = load_workbook(path, data_only=False, read_only=True)
    try:
        regular_ws = regular["Sparse"]
        readonly_ws = readonly["Sparse"]
        keys_before = set(regular_ws._cells)
        requested_rows = (1, 2, 49, 50, 51, 884, 885)
        regular_rows = smt._read_rows_into_cache(
            regular_ws, requested_rows, 33
        )
        readonly_rows = smt._read_rows_into_cache(
            readonly_ws, requested_rows, 33
        )
        assert regular_rows == readonly_rows
        assert regular_rows[1][0] == "head"
        assert regular_rows[50][12] == "middle"
        assert regular_rows[885][32] == "tail"
        assert all(
            value is None
            for row_idx, row in regular_rows.items()
            for col_idx, value in enumerate(row, start=1)
            if (row_idx, col_idx) not in {(1, 1), (50, 13), (885, 33)}
        )
        assert set(regular_ws._cells) == keys_before, (
            len(keys_before), len(regular_ws._cells)
        )
    finally:
        regular.close()
        readonly.close()
        wb.close()


def test_formula_cache_asymmetry_keeps_identity_rows_and_column_only_diff_actions():
    root_dir = make_temp_dir("sow_formula_cache_column_structure_")
    mine = os.path.join(root_dir, "mine.xlsx")
    theirs = os.path.join(root_dir, "theirs.xlsx")
    _write_formula_cache_asymmetry_book(
        mine, ("id", "calc", "tail"), cached_formula=2
    )
    _write_formula_cache_asymmetry_book(
        theirs, ("id", "new", "calc", "tail"), cached_formula=None
    )
    # Prove the intended asymmetric workbook state before exercising the GUI.
    mine_value = load_workbook(mine, data_only=True, read_only=True)
    theirs_value = load_workbook(theirs, data_only=True, read_only=True)
    try:
        assert mine_value["Sheet1"]["B2"].value == 2
        assert theirs_value["Sheet1"]["C2"].value is None
    finally:
        mine_value.close()
        theirs_value.close()

    app = smt.SowMergeApp(mine, theirs)
    try:
        view = _force_full_view(_wait_for_view(app))
        view._suppress_bg_apply = True
        view.only_diff_var.set(1)
        view.refresh(row_only=None, rescan=True)
        _pump(app.root, 0.05)
        identity_pairs = ((1, 1), (2, 2))
        assert tuple(view.row_pairs) == identity_pairs, view.row_pairs
        assert view.column_comparison_cache.structural_diff_cols == frozenset({2}), (
            view.column_comparison_cache.structural_diff_cols,
            _model_snapshot(view),
        )
        assert all(
            not view._visual_diff_cols_for_pair(pair_idx)
            for pair_idx in range(len(view.row_pairs))
        ), view.pair_diff_cols
        assert view._full_display_rows == [], view._full_display_rows

        view.only_diff_var.set(0)
        view.refresh(row_only=None, rescan=False)
        before_a = _worksheet_snapshot(app.ws_a_edit("Sheet1"))
        before_b = _worksheet_snapshot(app.ws_b_edit("Sheet1"))
        before_model = _model_snapshot(view)

        view._select_column_block_by_logical_col(2, "B")
        inserted = view._apply_selected_column_block("B", "A")
        assert inserted.action_kind == "insert_copy" and inserted.count == 1
        assert tuple(view.row_pairs) == identity_pairs
        view._undo_last_action()
        assert tuple(view.row_pairs) == identity_pairs
        assert _worksheet_snapshot(app.ws_a_edit("Sheet1")) == before_a
        assert _model_snapshot(view) == before_model

        view._select_column_block_by_logical_col(2, "A")
        deleted = view._apply_selected_column_block("A", "B")
        assert deleted.action_kind == "delete" and deleted.count == 1
        assert tuple(view.row_pairs) == identity_pairs
        view._undo_last_action()
        assert tuple(view.row_pairs) == identity_pairs
        assert _worksheet_snapshot(app.ws_b_edit("Sheet1")) == before_b
        assert _model_snapshot(view) == before_model
    finally:
        app._shutdown_root()


def test_real_gui_cell_apply_and_undo_reuse_mapping_without_full_scan():
    root_dir = make_temp_dir("sow_nonstructural_cell_fastpath_")
    mine = os.path.join(root_dir, "mine.xlsx")
    theirs = os.path.join(root_dir, "theirs.xlsx")
    headers = ("A", "B", "C", "D")
    _write_action_book(mine, headers)
    _write_action_book(theirs, headers)
    wb = load_workbook(theirs, data_only=False)
    wb["Sheet1"]["B3"] = "theirs-edited"
    wb.save(theirs)
    wb.close()

    app = smt.SowMergeApp(mine, theirs)
    try:
        view = _force_full_view(_wait_for_view(app))
        _wait_for_stable_projection(view)
        view._suppress_bg_apply = True
        view.only_diff_var.set(0)
        view.refresh(row_only=None, rescan=False)

        pair_idx = view.row_a_to_pair_idx[3]
        logical_col = view._logical_col_for_physical("B", 2)
        assert logical_col is not None
        before_cols = set(view.pair_diff_cols[pair_idx])
        assert logical_col in before_cols, (logical_col, before_cols)
        old_cache = view.column_comparison_cache
        old_model = old_cache.model
        old_versions = (
            view._row_model_version,
            view._column_model_version,
            view._mine_edit_version,
            view._theirs_edit_version,
        )
        old_lookups = (
            old_model.mine_physical_to_logical,
            old_model.base_physical_to_logical,
            old_model.theirs_physical_to_logical,
            old_model.mine_logical_to_physical,
            old_model.base_logical_to_physical,
            old_model.theirs_logical_to_physical,
        )

        rebuild_calls = []
        refresh_calls = []
        exact_calls = []
        original_refresh = view.refresh
        original_exact = view._refresh_pair_indices_exact

        def forbidden_rebuild(*args, **kwargs):
            rebuild_calls.append((args, kwargs))
            raise AssertionError("ordinary cell apply/undo rebuilt column mapping")

        def tracked_refresh(row_only, rescan, **kwargs):
            refresh_calls.append((row_only, bool(rescan)))
            return original_refresh(row_only=row_only, rescan=rescan, **kwargs)

        def tracked_exact(pair_indices):
            pair_indices = tuple(pair_indices)
            exact_calls.append(pair_indices)
            return original_exact(pair_indices)

        view._rebuild_column_comparison_cache_from_worksheets = forbidden_rebuild
        view.refresh = tracked_refresh
        view._refresh_pair_indices_exact = tracked_exact

        # Foreground apply/undo must stay on sparse point reads. Any iter_rows
        # call here would be a regression to a rectangular/full worksheet scan.
        for ws in (
            app.ws_a_val("Sheet1"),
            app.ws_b_val("Sheet1"),
            app.ws_a_edit("Sheet1"),
            app.ws_b_edit("Sheet1"),
        ):
            ws.iter_rows = lambda *args, **kwargs: (_ for _ in ()).throw(
                AssertionError("ordinary cell apply/undo called iter_rows")
            )

        view._copy_single_cell_by_pair(pair_idx, "B2A", logical_col)
        assert app.ws_a_edit("Sheet1").cell(3, 2).value == "theirs-edited"
        assert logical_col not in view.pair_diff_cols[pair_idx]
        applied_model = view.column_comparison_cache.model
        assert applied_model.slots is old_model.slots
        assert applied_model.blocks is old_model.blocks
        assert all(
            current is previous
            for current, previous in zip((
                applied_model.mine_physical_to_logical,
                applied_model.base_physical_to_logical,
                applied_model.theirs_physical_to_logical,
                applied_model.mine_logical_to_physical,
                applied_model.base_logical_to_physical,
                applied_model.theirs_logical_to_physical,
            ), old_lookups)
        )
        assert (view._row_model_version, view._column_model_version) == old_versions[:2]
        assert view._mine_edit_version == old_versions[2] + 1
        assert view._theirs_edit_version == old_versions[3]
        assert view._column_mapping_is_current()

        view._undo_last_action()
        assert app.ws_a_edit("Sheet1").cell(3, 2).value != "theirs-edited"
        assert set(view.pair_diff_cols[pair_idx]) == before_cols
        undone_model = view.column_comparison_cache.model
        assert undone_model.slots is old_model.slots
        assert all(
            current is previous
            for current, previous in zip((
                undone_model.mine_physical_to_logical,
                undone_model.base_physical_to_logical,
                undone_model.theirs_physical_to_logical,
                undone_model.mine_logical_to_physical,
                undone_model.base_logical_to_physical,
                undone_model.theirs_logical_to_physical,
            ), old_lookups)
        )
        assert (view._row_model_version, view._column_model_version) == old_versions[:2]
        assert view._mine_edit_version == old_versions[2] + 2
        assert exact_calls and exact_calls[-1] == (pair_idx,), exact_calls
        assert rebuild_calls == []
        assert refresh_calls and all(not rescan for _row, rescan in refresh_calls), refresh_calls
        assert view._column_mapping_is_current()
    finally:
        app._shutdown_root()


def test_real_gui_region_apply_and_undo_reuse_mapping_without_full_scan():
    root_dir = make_temp_dir("sow_nonstructural_region_fastpath_")
    mine = os.path.join(root_dir, "mine.xlsx")
    theirs = os.path.join(root_dir, "theirs.xlsx")
    headers = ("A", "B", "C", "D")
    _write_action_book(mine, headers)
    _write_action_book(theirs, headers)
    workbook = load_workbook(theirs, data_only=False)
    workbook["Sheet1"]["B3"] = "theirs-region-3"
    workbook["Sheet1"]["B4"] = "theirs-region-4"
    workbook.save(theirs)
    workbook.close()

    app = smt.SowMergeApp(mine, theirs)
    try:
        view = _force_full_view(_wait_for_view(app))
        _wait_for_stable_projection(view)
        view._suppress_bg_apply = True
        view.only_diff_var.set(0)
        view.refresh(row_only=None, rescan=False)

        pair_indices = tuple(view.row_a_to_pair_idx[row] for row in (3, 4))
        logical_col = view._logical_col_for_physical("B", 2)
        assert logical_col is not None
        assert all(
            logical_col in set(view.pair_diff_cols[pair_idx])
            for pair_idx in pair_indices
        )
        assert view._logical_diff_pair_block_for_pair(pair_indices[0]) == list(pair_indices)

        old_cache = view.column_comparison_cache
        old_model = old_cache.model
        old_versions = (
            view._row_model_version,
            view._column_model_version,
            view._mine_edit_version,
            view._theirs_edit_version,
        )
        old_geometry = (
            old_model.slots,
            old_model.blocks,
            old_model.mine_physical_to_logical,
            old_model.base_physical_to_logical,
            old_model.theirs_physical_to_logical,
            old_model.mine_logical_to_physical,
            old_model.base_logical_to_physical,
            old_model.theirs_logical_to_physical,
        )

        rebuild_calls = []
        refresh_calls = []
        exact_calls = []
        original_refresh = view.refresh
        original_exact = view._refresh_pair_indices_exact

        def forbidden_rebuild(*args, **kwargs):
            rebuild_calls.append((args, kwargs))
            raise AssertionError("ordinary region apply/undo rebuilt column mapping")

        def tracked_refresh(row_only, rescan, **kwargs):
            refresh_calls.append((row_only, bool(rescan)))
            return original_refresh(row_only=row_only, rescan=rescan, **kwargs)

        def tracked_exact(indices):
            indices = tuple(indices)
            exact_calls.append(indices)
            return original_exact(indices)

        view._rebuild_column_comparison_cache_from_worksheets = forbidden_rebuild
        view.refresh = tracked_refresh
        view._refresh_pair_indices_exact = tracked_exact
        for worksheet in (
            app.ws_a_val("Sheet1"),
            app.ws_b_val("Sheet1"),
            app.ws_a_edit("Sheet1"),
            app.ws_b_edit("Sheet1"),
        ):
            worksheet.iter_rows = lambda *args, **kwargs: (_ for _ in ()).throw(
                AssertionError("ordinary region apply/undo called iter_rows")
            )

        view.selected_pair_idx = pair_indices[0]
        view._last_selected_line = view.row_to_line[pair_indices[0]]
        view._copy_selected_region("B2A")
        assert app.ws_a_edit("Sheet1").cell(3, 2).value == "theirs-region-3"
        assert app.ws_a_edit("Sheet1").cell(4, 2).value == "theirs-region-4"
        assert all(not view.pair_diff_cols.get(pair_idx) for pair_idx in pair_indices)
        applied_model = view.column_comparison_cache.model
        assert view.column_projection.model is applied_model
        assert (
            applied_model.slots,
            applied_model.blocks,
            applied_model.mine_physical_to_logical,
            applied_model.base_physical_to_logical,
            applied_model.theirs_physical_to_logical,
            applied_model.mine_logical_to_physical,
            applied_model.base_logical_to_physical,
            applied_model.theirs_logical_to_physical,
        ) == old_geometry
        assert all(
            current is previous
            for current, previous in zip((
                applied_model.slots,
                applied_model.blocks,
                applied_model.mine_physical_to_logical,
                applied_model.base_physical_to_logical,
                applied_model.theirs_physical_to_logical,
                applied_model.mine_logical_to_physical,
                applied_model.base_logical_to_physical,
                applied_model.theirs_logical_to_physical,
            ), old_geometry)
        )
        assert (view._row_model_version, view._column_model_version) == old_versions[:2]
        assert view._mine_edit_version == old_versions[2] + 1
        assert view._theirs_edit_version == old_versions[3]
        assert view._column_mapping_is_current()

        view._undo_last_action()
        assert app.ws_a_edit("Sheet1").cell(3, 2).value != "theirs-region-3"
        assert app.ws_a_edit("Sheet1").cell(4, 2).value != "theirs-region-4"
        assert all(
            logical_col in set(view.pair_diff_cols[pair_idx])
            for pair_idx in pair_indices
        )
        undone_model = view.column_comparison_cache.model
        assert view.column_projection.model is undone_model
        assert all(
            current is previous
            for current, previous in zip((
                undone_model.slots,
                undone_model.blocks,
                undone_model.mine_physical_to_logical,
                undone_model.base_physical_to_logical,
                undone_model.theirs_physical_to_logical,
                undone_model.mine_logical_to_physical,
                undone_model.base_logical_to_physical,
                undone_model.theirs_logical_to_physical,
            ), old_geometry)
        )
        assert (view._row_model_version, view._column_model_version) == old_versions[:2]
        assert view._mine_edit_version == old_versions[2] + 2
        assert exact_calls and exact_calls[-1] == pair_indices, exact_calls
        assert rebuild_calls == []
        assert refresh_calls and all(not rescan for _row, rescan in refresh_calls)
        assert view._column_mapping_is_current()
    finally:
        app._shutdown_root()


def test_live_formula_references_follow_excel_insert_delete_and_undo_exactly():
    root_dir = make_temp_dir("sow_live_formula_column_structure_")
    original = os.path.join(root_dir, "original.xlsx")
    inserted = os.path.join(root_dir, "inserted.xlsx")
    deleted = os.path.join(root_dir, "deleted.xlsx")
    _write_excel_semantic_formula_structure_book(original, "original")
    _write_excel_semantic_formula_structure_book(inserted, "inserted")
    _write_excel_semantic_formula_structure_book(deleted, "deleted")

    app = smt.SowMergeApp(original, inserted)
    try:
        view = _force_full_view(_wait_for_view(app))
        before_edit = _worksheet_snapshot(app.ws_a_edit("Sheet1"))
        before_value = _worksheet_snapshot(app.ws_a_val("Sheet1"))
        before_model = _model_snapshot(view)
        selected = view._select_column_block_by_logical_col(2, "B")
        assert selected is not None and selected.state == "inserted", selected
        plan = view._apply_selected_column_block("B", "A")
        assert plan.action_kind == "insert_copy" and plan.target_physical_anchor == 2
        assert tuple(
            app.ws_a_edit("Sheet1").cell(1, col).value
            for col in range(1, 7)
        ) == ("id", "new", "source", "moved", "fixed", "conservative")
        for row in range(2, 8):
            assert app.ws_a_edit("Sheet1").cell(row, 3).value == f"=A{row}+1"
            assert app.ws_a_edit("Sheet1").cell(row, 4).value == f"=C{row}*2"
            assert app.ws_a_edit("Sheet1").cell(row, 5).value == f"=A{row}*3"
            assert app.ws_a_edit("Sheet1").cell(row, 6).value == (
                f"=Sheet2!B{row}+NamedRange+SUM(Table1[Amount])"
            )
        view.refresh(row_only=None, rescan=True)
        assert not view.column_comparison_cache.structural_diff_cols
        assert not view.column_comparison_cache.unresolved_cols
        assert not any(view.pair_diff_cols.values()), view.pair_diff_cols

        view._undo_last_action()
        assert _worksheet_snapshot(app.ws_a_edit("Sheet1")) == before_edit
        assert _worksheet_snapshot(app.ws_a_val("Sheet1")) == before_value
        assert _model_snapshot(view) == before_model
    finally:
        app._shutdown_root()

    app = smt.SowMergeApp(original, deleted)
    try:
        view = _force_full_view(_wait_for_view(app))
        natural_before_model = _model_snapshot(view)
        assert view.column_comparison_cache.unresolved_cols, natural_before_model
        # Raw formula identity intentionally keeps this deletion conservative:
        # Mine C ``=B*2`` versus Excel-semantic Theirs B ``=#REF!*2`` cannot be
        # auto-anchored. Install the user-confirmed physical intent explicitly
        # so this test exercises the real live mutation/undo path, not a
        # speculative scanner decision.
        unresolved = smt.ColumnMappingConfidence(
            0.0,
            True,
            "formula-identity-mismatch",
            ("explicit-test-intent",),
            (smt.COLUMN_MAPPING_CAUSE_FORMULA_MISMATCH,),
        )
        slots = (
            smt.ColumnSlot(0, mine_col=1, theirs_col=1),
            smt.ColumnSlot(
                1,
                mine_col=2,
                theirs_col=None,
                state="unresolved",
                confidence=unresolved,
            ),
            smt.ColumnSlot(2, mine_col=3, theirs_col=2),
            smt.ColumnSlot(3, mine_col=4, theirs_col=3),
            smt.ColumnSlot(4, mine_col=5, theirs_col=4),
        )
        model = smt.ColumnModel.from_slots(
            view._expected_column_model_cache_key(),
            slots,
            blocks=smt._build_column_blocks(slots),
            confidence=unresolved,
        )
        view._install_column_projection(smt.LogicalColumnComparisonCache(
            model=model,
            structural_diff_cols=frozenset({2}),
            unresolved_cols=frozenset({2}),
        ))
        before_edit = _worksheet_snapshot(app.ws_a_edit("Sheet1"))
        before_value = _worksheet_snapshot(app.ws_a_val("Sheet1"))
        selected = view._select_column_block_by_logical_col(2, "A")
        assert selected is not None and selected.state == "unresolved", selected
        plan = view._apply_selected_column_block(
            "B", "A", confirm_unresolved=True
        )
        assert plan.action_kind == "delete" and plan.target_physical_anchor == 2, plan
        assert tuple(
            app.ws_a_edit("Sheet1").cell(1, col).value
            for col in range(1, 5)
        ) == ("id", "moved", "fixed", "conservative")
        for row in range(2, 8):
            # The source-column formula was deleted; the following formula now
            # occupies its cell and must expose Excel's #REF! semantics.
            assert app.ws_a_edit("Sheet1").cell(row, 2).value == "=#REF!*2"
            assert app.ws_a_edit("Sheet1").cell(row, 3).value == f"=A{row}*3"
            assert app.ws_a_edit("Sheet1").cell(row, 4).value == (
                f"=Sheet2!B{row}+NamedRange+SUM(Table1[Amount])"
            )
        view.refresh(row_only=None, rescan=True)
        assert not view.column_comparison_cache.structural_diff_cols
        assert not view.column_comparison_cache.unresolved_cols
        assert not any(view.pair_diff_cols.values()), view.pair_diff_cols

        view._undo_last_action()
        assert _worksheet_snapshot(app.ws_a_edit("Sheet1")) == before_edit
        assert _worksheet_snapshot(app.ws_a_val("Sheet1")) == before_value
        assert _model_snapshot(view) == natural_before_model
    finally:
        app._shutdown_root()


def test_two_way_insert_delete_copy_and_retain_plans_use_logical_block():
    cache = _cache_2way(
        "Sheet1",
        _rows(("A", "B", "C", "D", "E", "F")),
        _rows(("A", "X", "Y", "B", "D", "E", "F")),
    )
    view = _fake_action_view(cache)

    inserted = view._select_column_block_by_logical_col(2, "B")
    assert inserted is not None and inserted.slot_indices == (1, 2)
    plan = view._plan_selected_column_block_action(
        "B", "A", action_id="column-action-insert"
    )
    assert plan.action_kind == "insert_copy"
    assert (plan.logical_start, plan.logical_end, plan.count) == (2, 3, 2)
    assert plan.source_physical_cols == (2, 3)
    assert plan.target_physical_cols == ()
    assert plan.target_physical_anchor == 2

    reverse = view._plan_selected_column_block_action(
        "A", "B", action_id="column-action-delete-inserted"
    )
    assert reverse.action_kind == "delete"
    assert reverse.target_physical_cols == (2, 3)
    assert reverse.target_physical_anchor == 2

    deleted = view._select_column_block_by_logical_col(5, "A")
    assert deleted is not None and deleted.slot_indices == (4,)
    restore = view._plan_selected_column_block_action(
        "A", "B", action_id="column-action-restore-deleted"
    )
    assert restore.action_kind == "insert_copy"
    assert restore.source_physical_cols == (3,)
    assert restore.target_physical_cols == ()
    assert restore.target_physical_anchor == 5

    adopt_absence = view._plan_selected_column_block_action(
        "B", "A", action_id="column-action-adopt-delete"
    )
    assert adopt_absence.action_kind == "delete"
    assert adopt_absence.target_physical_cols == (3,)
    assert adopt_absence.target_physical_anchor == 3

    retain = view._plan_selected_column_block_action(
        "A", "A", action_id="column-action-retain"
    )
    assert retain.action_kind == "retain"
    assert retain.operation_records() == ()


def test_manual_formula_ops_follow_excel_column_reference_semantics():
    def _app_with_ops(cell_ops, cache_ops=None):
        app = object.__new__(smt.SowMergeApp)
        app.manual_a_cell_ops = dict(cell_ops)
        app.manual_a_formula_cache_ops = dict(cache_ops or {})
        app.manual_b_cell_ops = {}
        app.manual_b_formula_cache_ops = {}
        return app

    insert_cache = _cache_2way(
        "Sheet1",
        _rows(("A", "B", "C")),
        _rows(("A", "X", "B", "C")),
    )
    insert_view = _fake_action_view(insert_cache)
    insert_view._select_column_block_by_logical_col(2, "B")
    insert_plan = insert_view._plan_selected_column_block_action(
        "B", "A", action_id="formula-insert"
    )
    insert_app = _app_with_ops(
        {
            ("Sheet1", 1, 3): "=B1*2",
            ("Sheet1", 2, 3): "=$B$1+SUM(B1:C2)+SUM(B:B)",
            ("Sheet1", 3, 3): "=A1*2",
            ("Sheet1", 4, 3): "=Sheet2!B1+SUM(Table1[Amount])+NamedRange",
            ("Other", 1, 3): "=B1*2",
        },
        {("Sheet1", 1, 3): 2},
    )
    smt.SowMergeApp.remap_manual_cell_operations_for_column_action(
        insert_app, "A", insert_plan
    )
    expected_insert_ops = {
        ("Sheet1", 1, 4): "=C1*2",
        ("Sheet1", 2, 4): "=$C$1+SUM(C1:D2)+SUM(C:C)",
        ("Sheet1", 3, 4): "=A1*2",
        ("Sheet1", 4, 4): "=Sheet2!B1+SUM(Table1[Amount])+NamedRange",
        ("Other", 1, 3): "=B1*2",
    }
    assert insert_app.manual_a_cell_ops == expected_insert_ops, (
        insert_app.manual_a_cell_ops,
        expected_insert_ops,
    )
    assert insert_app.manual_a_formula_cache_ops == {("Sheet1", 1, 4): 2}

    delete_cache = _cache_2way(
        "Sheet1",
        _rows(("A", "B", "C")),
        _rows(("A", "C")),
    )
    delete_view = _fake_action_view(delete_cache)
    delete_view._select_column_block_by_logical_col(2, "A")
    delete_plan = delete_view._plan_selected_column_block_action(
        "B", "A", action_id="formula-delete"
    )
    delete_app = _app_with_ops({
        ("Sheet1", 1, 3): "=B1*2",
        ("Sheet1", 2, 3): "=C1*2",
        ("Sheet1", 3, 3): "=A1*2",
        ("Sheet1", 4, 3): "=Sheet2!B1+SUM(Table1[Amount])+NamedRange",
        ("Sheet1", 5, 2): "=B1",
    })
    smt.SowMergeApp.remap_manual_cell_operations_for_column_action(
        delete_app, "A", delete_plan
    )
    expected_delete_ops = {
        ("Sheet1", 1, 2): "=#REF!*2",
        ("Sheet1", 2, 2): "=B1*2",
        ("Sheet1", 3, 2): "=A1*2",
        ("Sheet1", 4, 2): "=Sheet2!B1+SUM(Table1[Amount])+NamedRange",
    }
    assert delete_app.manual_a_cell_ops == expected_delete_ops, (
        delete_app.manual_a_cell_ops,
        expected_delete_ops,
    )


def test_formula_capture_prefilter_keeps_target_and_qualified_references_only():
    cache = _cache_2way(
        "Target",
        _rows(("A", "B", "C")),
        _rows(("A", "X", "B", "C")),
    )
    view = _fake_action_view(cache)
    view._select_column_block_by_logical_col(2, "B")
    plan = view._plan_selected_column_block_action(
        "B", "A", action_id="formula-prefilter-insert"
    )

    workbook = Workbook()
    target = workbook.active
    target.title = "Target"
    other = workbook.create_sheet("Other")
    target["F1"] = "=B1"
    target["F2"] = "=SUM(Table1[Amount])+NamedRange+[Book2.xlsx]Other!B1"
    other["A1"] = "=Target!B1"
    other["A2"] = "='Target'!B1"
    other["A3"] = "=B1"
    other["A4"] = "=SUM(Table1[Amount])"
    other["A5"] = "=NamedRange"
    other["A6"] = "=[Book2.xlsx]Target!B1"
    try:
        records = smt.SheetView._capture_column_formula_transformations(
            workbook, plan
        )
    finally:
        workbook.close()

    actual = {
        (sheet, row_idx, col_idx): (old_value, new_value)
        for sheet, row_idx, col_idx, old_value, _old_type, new_value in records
    }
    assert actual == {
        ("Target", 1, 6): ("=B1", "=C1"),
        ("Other", 1, 1): ("=Target!B1", "=Target!C1"),
        ("Other", 2, 1): ("='Target'!B1", "='Target'!C1"),
    }, actual


def test_three_way_mine_base_theirs_plans_and_same_anchor_confirmation():
    independent = _cache_3way(
        "Sheet1",
        _rows(("A", "M", "B", "C", "D")),
        _rows(("A", "B", "C", "D")),
        _rows(("A", "B", "C", "T", "D")),
    )
    view = _fake_action_view(independent, three_way=True)

    view._select_column_block_by_logical_col(2, "A")
    keep_mine = view._plan_selected_column_block_action(
        "A", "A", action_id="column-action-mine"
    )
    adopt_base_absence = view._plan_selected_column_block_action(
        "BASE", "A", action_id="column-action-base-absence"
    )
    adopt_theirs_absence = view._plan_selected_column_block_action(
        "B", "A", action_id="column-action-theirs-absence"
    )
    assert keep_mine.action_kind == "retain"
    assert adopt_base_absence.action_kind == "delete"
    assert adopt_theirs_absence.action_kind == "delete"

    view._select_column_block_by_logical_col(5, "B")
    adopt_theirs = view._plan_selected_column_block_action(
        "B", "A", action_id="column-action-theirs"
    )
    assert adopt_theirs.action_kind == "insert_copy"
    assert adopt_theirs.source_physical_cols == (4,)
    assert adopt_theirs.target_physical_anchor == 5

    mine_deleted = _cache_3way(
        "Sheet1",
        _rows(("A", "C", "D")),
        _rows(("A", "B", "C", "D")),
        _rows(("A", "B", "C", "D")),
    )
    view = _fake_action_view(mine_deleted, three_way=True)
    view._select_column_block_by_logical_col(2, "BASE")
    adopt_base = view._plan_selected_column_block_action(
        "BASE", "A", action_id="column-action-base"
    )
    assert adopt_base.action_kind == "insert_copy"
    assert adopt_base.source_physical_cols == (2,)
    assert adopt_base.target_physical_anchor == 2

    competing = _cache_3way(
        "Sheet1",
        _rows(("A", "M", "B", "C")),
        _rows(("A", "B", "C")),
        _rows(("A", "T", "B", "C")),
    )
    view = _fake_action_view(competing, three_way=True)
    unresolved = view._select_column_block_by_logical_col(2, "A")
    assert unresolved is not None and unresolved.state == "unresolved"
    before = (
        view.selected_column_block_ordinal,
        view.selected_column_logical_range,
        view.selected_column_source_side,
    )
    _raises_runtime(
        lambda: view._plan_selected_column_block_action(
            "A", "A", action_id="column-action-needs-confirmation"
        ),
        "必须显式确认",
    )
    assert (
        view.selected_column_block_ordinal,
        view.selected_column_logical_range,
        view.selected_column_source_side,
    ) == before
    # Explicit confirmation crosses the ambiguity gate. Competing same-anchor
    # insertions are separate one-sided blocks, so choosing Mine inserts only
    # M before the adjacent unresolved Theirs-only T block.
    confirmed = view._plan_selected_column_block_action(
        "A",
        "B",
        confirm_unresolved=True,
        action_id="column-action-confirmed-mine",
    )
    assert confirmed.action_kind == "insert_copy"
    assert confirmed.source_physical_cols == (2,)
    assert confirmed.target_physical_cols == ()
    assert confirmed.target_physical_anchor == 2
    assert confirmed.count == 1


def test_ready_auto_selects_first_real_structural_block_only():
    gunships_columns = tuple(f"column-{idx:02d}" for idx in range(1, 27))
    gunships_mine_columns = tuple(
        value for idx, value in enumerate(gunships_columns, start=1)
        if idx not in (14, 20)
    )
    mine_deleted = _cache_3way(
        "GunshipsModify@design",
        _rows(gunships_mine_columns),
        _rows(gunships_columns),
        _rows(gunships_columns),
    )
    view = _fake_action_view(mine_deleted, three_way=True)
    view._lifecycle_state = "READY"
    view.use_mine_col_btn = _ButtonState()
    view.use_base_col_btn = _ButtonState()
    view.use_theirs_col_btn = _ButtonState()

    view._refresh_column_action_buttons()

    assert view.selected_column_logical_range == (14, 14)
    assert view.selected_column_source_side == "LOGICAL"
    assert view._selected_column_block().state == "mine-deleted"
    assert all(
        button.options.get("state") == "normal"
        for button in (
            view.use_mine_col_btn,
            view.use_base_col_btn,
            view.use_theirs_col_btn,
        )
    )

    # Applying L14 rebuilds the projection. The next generation must select
    # L20 automatically, instead of leaving all three column-action buttons
    # disabled while a structural block remains.
    after_first_apply = _cache_3way(
        "GunshipsModify@design",
        _rows(
            tuple(
                value for idx, value in enumerate(gunships_columns, start=1)
                if idx != 20
            )
        ),
        _rows(gunships_columns),
        _rows(gunships_columns),
    )
    view._install_column_projection(after_first_apply)
    view.selected_column_block_ordinal = None
    view.selected_column_logical_range = None
    view.selected_column_source_side = None
    view._selected_column_projection_generation = None
    view._refresh_column_action_buttons()

    assert view.selected_column_logical_range == (20, 20)
    assert view._selected_column_block().state == "mine-deleted"
    assert all(
        button.options.get("state") == "normal"
        for button in (
            view.use_mine_col_btn,
            view.use_base_col_btn,
            view.use_theirs_col_btn,
        )
    )

    # A deliberate click on a regular column clears the current projection's
    # selection and must not immediately select L20 again.
    assert view._select_column_block_by_logical_col(1) is None
    assert view.selected_column_block_ordinal is None
    assert all(
        button.options.get("state") == "disabled"
        for button in (
            view.use_mine_col_btn,
            view.use_base_col_btn,
            view.use_theirs_col_btn,
        )
    )

    regular_base = _rows(("field", "type", "comment", "default", "data"))
    regular_mine = [list(row) for row in regular_base]
    regular_mine[1][4] = "target-only-cell-difference"
    regular = _cache_3way(
        "GunshipsConfig@column",
        regular_mine,
        regular_base,
        regular_base,
    )
    assert not regular.structural_diff_cols and not regular.unresolved_cols
    regular_view = _fake_action_view(regular, three_way=True)
    regular_view._lifecycle_state = "READY"
    regular_view.use_mine_col_btn = _ButtonState()
    regular_view.use_base_col_btn = _ButtonState()
    regular_view.use_theirs_col_btn = _ButtonState()

    regular_view._refresh_column_action_buttons()

    assert regular_view.selected_column_block_ordinal is None
    assert all(
        button.options.get("state") == "disabled"
        for button in (
            regular_view.use_mine_col_btn,
            regular_view.use_base_col_btn,
            regular_view.use_theirs_col_btn,
        )
    )


def test_first_direct_row_apply_revalidates_false_pending_mapping():
    root_dir = make_temp_dir("sow_first_row_mapping_revalidate_")
    mine = os.path.join(root_dir, "mine.xlsx")
    theirs = os.path.join(root_dir, "theirs.xlsx")
    _write_action_book(mine, ("A", "B", "C"))
    _write_action_book(theirs, ("A", "B", "C"))
    workbook = load_workbook(theirs)
    workbook["Sheet1"]["C2"] = "theirs-diff"
    workbook.save(theirs)
    workbook.close()

    app = smt.SowMergeApp(mine, theirs)
    try:
        view = _force_full_view(_wait_for_view(app))
        deadline = time.time() + 12.0
        while time.time() < deadline and view._derive_lifecycle_state() != "READY":
            _pump(app.root, 0.03)
        assert view._derive_lifecycle_state() == "READY"
        natural = view._active_column_comparison_cache()
        assert not natural.unresolved_cols

        pending = smt.ColumnMappingConfidence(
            0.2,
            True,
            "low-confidence-physical-fallback",
            ("provisional-background-cache",),
            (smt.COLUMN_MAPPING_CAUSE_LOW_CONFIDENCE,),
        )
        slots = list(natural.model.slots)
        first = slots[0]
        slots[0] = smt.ColumnSlot(
            logical_idx=first.logical_idx,
            mine_col=first.mine_col,
            base_col=first.base_col,
            theirs_col=first.theirs_col,
            state="unresolved",
            confidence=pending,
            base_boundary=first.base_boundary,
            origin_side=first.origin_side,
        )
        false_model = smt.ColumnModel.from_slots(
            view._expected_column_model_cache_key(),
            tuple(slots),
            blocks=smt._build_column_blocks(tuple(slots)),
            confidence=pending,
        )
        view._install_column_projection(smt.LogicalColumnComparisonCache(
            model=false_model,
            two_way_alignment=natural.two_way_alignment,
            structural_diff_cols=natural.structural_diff_cols,
            unresolved_cols=frozenset({1}),
        ))
        assert view._column_mapping_is_current()
        assert view._active_column_projection().slot(1).state == "unresolved"

        revalidation_calls = []
        original_revalidate = view._authoritative_revalidate_column_projection_for_action

        def counted_revalidate(*args, **kwargs):
            revalidation_calls.append(1)
            return original_revalidate(*args, **kwargs)

        view._authoritative_revalidate_column_projection_for_action = counted_revalidate
        errors = []
        original_showerror = smt.messagebox.showerror
        smt.messagebox.showerror = lambda *args, **kwargs: errors.append((args, kwargs))
        try:
            pair_idx = view.row_a_to_pair_idx[2]
            assert view._copy_selected_row("B2A", override_pair_idx=pair_idx)
        finally:
            smt.messagebox.showerror = original_showerror
            view._authoritative_revalidate_column_projection_for_action = original_revalidate

        assert errors == []
        assert revalidation_calls == [1]
        assert app.ws_a_edit("Sheet1")["C2"].value == "theirs-diff"
        assert view._active_column_projection().slot(1).state != "unresolved"
        assert 1 not in view._active_column_comparison_cache().unresolved_cols
    finally:
        app._shutdown_root()


def test_consecutive_row_apply_ignores_style_only_blank_tail_columns():
    root_dir = make_temp_dir("sow_row_style_tail_mapping_")
    mine = os.path.join(root_dir, "mine.xlsx")
    theirs = os.path.join(root_dir, "theirs.xlsx")

    for path, side in ((mine, "mine"), (theirs, "theirs")):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Sheet1"
        for col in range(1, 14):
            worksheet.cell(1, col).value = f"H{col}"
            worksheet.cell(2, col).value = (
                f"{side}-row-2-diff" if col == 5 else f"row-2-value-{col}"
            )
            worksheet.cell(3, col).value = (
                f"{side}-row-3-diff" if col == 5 else f"row-3-value-{col}"
            )
        # Match Link@design's misleading physical dimension: columns 14-18
        # contain styles only and are not part of the logical data structure.
        for col in range(14, 19):
            worksheet.cell(474, col).fill = PatternFill(
                "solid", fgColor="FF0000"
            )
        workbook.save(path)
        workbook.close()

    app = smt.SowMergeApp(mine, theirs)
    try:
        view = _wait_for_view(app)
        _wait_for_stable_projection(view)
        deadline = time.time() + 12.0
        while time.time() < deadline and view._derive_lifecycle_state() != "READY":
            _pump(app.root, 0.03)
        assert view._derive_lifecycle_state() == "READY"
        assert (view.col_max_a, view.col_max_b, view.max_col) == (13, 13, 13)
        assert view._logical_slot_count() == 13
        assert not view._active_column_comparison_cache().unresolved_cols

        errors = []
        original_showerror = smt.messagebox.showerror
        smt.messagebox.showerror = lambda *args, **kwargs: errors.append(
            (args, kwargs)
        )
        try:
            first_pair_idx = view.row_a_to_pair_idx[2]
            assert view._copy_selected_row(
                "B2A", override_pair_idx=first_pair_idx
            )
            assert (view.col_max_a, view.col_max_b, view.max_col) == (13, 13, 13)
            assert view._logical_slot_count() == 13
            assert not view._active_column_comparison_cache().unresolved_cols

            # Exercise the opposite direction on the next still-different row.
            second_pair_idx = view.row_a_to_pair_idx[3]
            assert view._copy_selected_row(
                "A2B", override_pair_idx=second_pair_idx
            )
        finally:
            smt.messagebox.showerror = original_showerror

        assert errors == []
        assert (
            app.ws_b_edit("Sheet1").cell(3, 5).value
            == "mine-row-3-diff"
        )
        assert (view.col_max_a, view.col_max_b, view.max_col) == (13, 13, 13)
        assert view._logical_slot_count() == 13
        assert not view._active_column_comparison_cache().unresolved_cols
        # The style-only cells remain physically intact without entering the
        # logical action range.
        assert app.ws_a_edit("Sheet1").max_column == 18
        assert app.ws_b_edit("Sheet1").max_column == 18
    finally:
        app._shutdown_root()


def test_full_row_apply_skips_unresolved_blank_gap_before_real_formula_column():
    root_dir = make_temp_dir("sow_row_blank_gap_formula_")
    mine = os.path.join(root_dir, "mine.xlsx")
    theirs = os.path.join(root_dir, "theirs.xlsx")

    for path, side in ((mine, "mine"), (theirs, "theirs")):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Sheet1"
        for col in range(1, 14):
            worksheet.cell(1, col).value = f"H{col}"
            worksheet.cell(2, col).value = (
                f"{side}-row-2-diff" if col == 5 else f"row-2-value-{col}"
            )
            worksheet.cell(3, col).value = (
                f"{side}-row-3-diff" if col == 5 else f"row-3-value-{col}"
            )
        # L14-L17 are intentionally empty on both sides. L18 is a real
        # formula-bearing column, so the semantic width must remain 18.
        worksheet.cell(1, 18).value = "formula-tail"
        if side == "mine":
            worksheet.cell(2, 18).value = "=A2"
            worksheet.cell(3, 18).value = "=C3"
        else:
            worksheet.cell(2, 18).value = "=B2"
            worksheet.cell(3, 18).value = "=D3"
        workbook.save(path)
        workbook.close()

    app = smt.SowMergeApp(mine, theirs)
    try:
        view = _wait_for_view(app)
        _wait_for_stable_projection(view)
        deadline = time.time() + 12.0
        while time.time() < deadline and view._derive_lifecycle_state() != "READY":
            _pump(app.root, 0.03)
        assert view._derive_lifecycle_state() == "READY"
        assert (view.col_max_a, view.col_max_b, view.max_col) == (18, 18, 18)
        assert view._logical_slot_count() == 18
        assert view._active_column_projection().slot(18).state == "retained"
        assert view._active_column_comparison_cache().unresolved_cols == frozenset(
            {14, 15, 16, 17}
        )

        errors = []
        original_showerror = smt.messagebox.showerror
        smt.messagebox.showerror = lambda *args, **kwargs: errors.append(
            (args, kwargs)
        )
        try:
            assert view._copy_selected_row(
                "B2A", override_pair_idx=view.row_a_to_pair_idx[2]
            )
            assert view._copy_selected_row(
                "A2B", override_pair_idx=view.row_a_to_pair_idx[3]
            )
        finally:
            smt.messagebox.showerror = original_showerror

        assert errors == []
        assert app.ws_a_edit("Sheet1").cell(2, 18).value == "=B2"
        assert app.ws_b_edit("Sheet1").cell(3, 18).value == "=C3"
        assert (view.col_max_a, view.col_max_b, view.max_col) == (18, 18, 18)
        # Blank-gap ambiguity remains visible to structural column operations;
        # only the row-specific no-op writes are skipped.
        assert view._active_column_comparison_cache().unresolved_cols == frozenset(
            {14, 15, 16, 17}
        )
    finally:
        app._shutdown_root()


def test_ordinary_cell_action_rejects_missing_or_unresolved_slot():
    cache = _cache_2way(
        "Sheet1",
        _rows(("A", "B", "C")),
        _rows(("A", "X", "B", "C")),
    )
    view = _fake_action_view(cache)
    _raises_runtime(lambda: view._action_physical_columns("B2A", 2), "不能按单元格覆盖")

    competing = _cache_3way(
        "Sheet1",
        _rows(("A", "M", "B", "C")),
        _rows(("A", "B", "C")),
        _rows(("A", "T", "B", "C")),
    )
    view = _fake_action_view(competing, three_way=True)
    _raises_runtime(lambda: view._action_physical_columns("B2A", 2), "映射待确认")


def test_real_gui_insert_block_and_one_step_undo_full_fidelity():
    app, view = _real_two_way_app()
    try:
        before_a_edit = _worksheet_snapshot(app.ws_a_edit("Sheet1"))
        before_a_value = _worksheet_snapshot(app.ws_a_val("Sheet1"))
        before_b_edit = _worksheet_snapshot(app.ws_b_edit("Sheet1"))
        before_b_value = _worksheet_snapshot(app.ws_b_val("Sheet1"))
        before_ops = list(app.manual_a_column_ops)
        before_undo_count = len(app.undo_stack)
        before_model = _model_snapshot(view)

        selected_pair = next(
            idx for idx, pair in enumerate(view.row_pairs) if pair == (3, 3)
        )
        view.selected_pair_idx = selected_pair
        view.selected_excel_row = 3
        view.selected_excel_row_a = 3
        view.selected_excel_row_b = 3
        view._main_sel_col = 6
        view._main_sel_line = view.row_to_line[selected_pair]

        block = view._select_column_block_by_logical_col(2, "B")
        assert block is not None and block.slot_indices == (1, 2)
        selected_before_action = _selection_snapshot(view)
        source_x = _column_snapshot(app.ws_b_edit("Sheet1"), 2)
        source_y = _column_snapshot(app.ws_b_edit("Sheet1"), 3)

        plan = view._apply_selected_column_block("B", "A")
        assert plan.action_kind == "insert_copy"
        assert [app.ws_a_edit("Sheet1").cell(1, col).value for col in range(1, 9)] == [
            "A", "X", "Y", "B", "C", "D", "E", "F",
        ]
        assert _column_snapshot(app.ws_a_edit("Sheet1"), 2) == source_x
        assert _column_snapshot(app.ws_a_edit("Sheet1"), 3) == source_y
        # The formulas, styles, comments, hyperlinks, validations, width and
        # hidden state are all included in the semantic column snapshots.
        assert app.ws_a_edit("Sheet1").cell(3, 2).value == "=1+ROW()"
        assert app.ws_a_edit("Sheet1").column_dimensions["B"].width == 19.0
        assert app.ws_a_edit("Sheet1").column_dimensions["B"].hidden is True
        assert app.ws_a_edit("Sheet1").cell(2, 2).comment.text == "comment-X"
        assert app.ws_a_edit("Sheet1").cell(2, 2).hyperlink.target.endswith("/X")
        assert len(app.ws_a_edit("Sheet1").data_validations.dataValidation) == 2

        recorded = app.manual_a_column_ops[len(before_ops):]
        assert [op["kind"] for op in recorded] == ["insert_cols", "copy_cols"]
        assert [op["count"] for op in recorded] == [2, 2]
        assert len({op["action_id"] for op in recorded}) == 1
        assert len(app.undo_stack) == before_undo_count + 1
        assert app.undo_stack[-1]["kind"] == "column_action"
        assert app.undo_stack[-1]["plan"] == plan
        remaining_special = (
            view.column_comparison_cache.structural_diff_cols
            or view.column_comparison_cache.unresolved_cols
        )
        selected_after_action = view._selected_column_block()
        if selected_after_action is not None:
            assert view._column_block_is_structural(selected_after_action)
        elif not remaining_special:
            assert _selection_snapshot(view)[:3] == (None, None, None)

        after_model = _model_snapshot(view)
        assert len(after_model[0]) == 8
        assert [(slot.mine_col, slot.theirs_col) for slot in after_model[0]][:4] == [
            (1, 1), (2, 2), (3, 3), (4, 4),
        ]

        view._undo_last_action()
        assert len(app.undo_stack) == before_undo_count
        assert app.manual_a_column_ops == before_ops
        assert _worksheet_snapshot(app.ws_a_edit("Sheet1")) == before_a_edit
        assert _worksheet_snapshot(app.ws_a_val("Sheet1")) == before_a_value
        # The non-target side must not acquire phantom empty columns during
        # refresh/rebuild; otherwise the restored model gains a tail slot.
        assert _worksheet_snapshot(app.ws_b_edit("Sheet1")) == before_b_edit
        assert _worksheet_snapshot(app.ws_b_val("Sheet1")) == before_b_value
        assert _selection_snapshot(view) == selected_before_action
        assert _model_snapshot(view) == before_model
    finally:
        app._shutdown_root()


def test_real_gui_delete_block_preserves_adjacent_columns_and_undo():
    app, view = _real_two_way_app(decorated_a=(2, 3, 4))
    try:
        before_edit = _worksheet_snapshot(app.ws_a_edit("Sheet1"))
        before_value = _worksheet_snapshot(app.ws_a_val("Sheet1"))
        before_model = _model_snapshot(view)
        selected = view._select_column_block_by_logical_col(5, "B")
        assert selected is not None and selected.slot_indices == (4,)
        selection_before = _selection_snapshot(view)
        left_neighbor = _column_snapshot(app.ws_a_edit("Sheet1"), 2)
        right_neighbor = _column_snapshot(app.ws_a_edit("Sheet1"), 4)

        plan = view._apply_selected_column_block("B", "A")
        assert plan.action_kind == "delete"
        assert plan.target_physical_anchor == 3 and plan.count == 1
        assert [app.ws_a_edit("Sheet1").cell(1, col).value for col in range(1, 6)] == [
            "A", "B", "D", "E", "F",
        ]
        assert _column_snapshot(app.ws_a_edit("Sheet1"), 2) == left_neighbor
        assert _column_snapshot(app.ws_a_edit("Sheet1"), 3) == right_neighbor
        assert [op["kind"] for op in app.manual_a_column_ops] == ["delete_cols"]
        assert app.manual_a_column_ops[0]["count"] == 1
        assert len(app.undo_stack) == 1

        view._undo_last_action()
        assert app.manual_a_column_ops == []
        assert app.undo_stack == []
        assert _worksheet_snapshot(app.ws_a_edit("Sheet1")) == before_edit
        assert _worksheet_snapshot(app.ws_a_val("Sheet1")) == before_value
        assert _selection_snapshot(view) == selection_before
        assert _model_snapshot(view) == before_model
    finally:
        app._shutdown_root()


def test_real_gui_failure_injection_is_atomic_at_every_mutating_stage():
    app, view = _real_two_way_app()
    try:
        before_a_edit = _worksheet_snapshot(app.ws_a_edit("Sheet1"))
        before_a_value = _worksheet_snapshot(app.ws_a_val("Sheet1"))
        before_b_edit = _worksheet_snapshot(app.ws_b_edit("Sheet1"))
        before_b_value = _worksheet_snapshot(app.ws_b_val("Sheet1"))
        before_model = _model_snapshot(view)
        view._select_column_block_by_logical_col(2, "B")
        selection_before = _selection_snapshot(view)
        before_flags = (
            app.modified_a,
            app.modified_b,
            set(app.modified_sheets_a),
            set(app.modified_sheets_b),
        )

        for stage in ("after_structure", "after_copy", "after_refresh"):
            def fail_at(current_stage, _plan, wanted=stage):
                if current_stage == wanted:
                    raise RuntimeError(f"injected-{wanted}")

            _raises_runtime(
                lambda: view._apply_selected_column_block(
                    "B", "A", _failure_injector=fail_at
                ),
                f"injected-{stage}",
            )
            assert _worksheet_snapshot(app.ws_a_edit("Sheet1")) == before_a_edit
            assert _worksheet_snapshot(app.ws_a_val("Sheet1")) == before_a_value
            assert _worksheet_snapshot(app.ws_b_edit("Sheet1")) == before_b_edit
            assert _worksheet_snapshot(app.ws_b_val("Sheet1")) == before_b_value
            assert app.manual_a_column_ops == []
            assert app.manual_b_column_ops == []
            assert app.undo_stack == []
            assert (
                app.modified_a,
                app.modified_b,
                set(app.modified_sheets_a),
                set(app.modified_sheets_b),
            ) == before_flags
            assert _selection_snapshot(view) == selection_before
            assert _model_snapshot(view) == before_model
    finally:
        app._shutdown_root()


def test_real_gui_repeated_adopt_undo_has_no_state_drift():
    app, view = _real_two_way_app()
    try:
        before_a_edit = _worksheet_snapshot(app.ws_a_edit("Sheet1"))
        before_a_value = _worksheet_snapshot(app.ws_a_val("Sheet1"))
        before_b_edit = _worksheet_snapshot(app.ws_b_edit("Sheet1"))
        before_b_value = _worksheet_snapshot(app.ws_b_val("Sheet1"))
        before_model = _model_snapshot(view)
        view._select_column_block_by_logical_col(2, "B")
        selection_before = _selection_snapshot(view)
        action_ids = []
        operation_orders = []

        for _cycle in range(3):
            plan = view._apply_selected_column_block("B", "A")
            action_ids.append(plan.action_id)
            operation_orders.extend(op["order"] for op in app.manual_a_column_ops)
            assert len(app.undo_stack) == 1
            view._undo_last_action()
            assert app.undo_stack == []
            assert app.manual_a_column_ops == []
            assert _worksheet_snapshot(app.ws_a_edit("Sheet1")) == before_a_edit
            assert _worksheet_snapshot(app.ws_a_val("Sheet1")) == before_a_value
            assert _worksheet_snapshot(app.ws_b_edit("Sheet1")) == before_b_edit
            assert _worksheet_snapshot(app.ws_b_val("Sheet1")) == before_b_value
            assert _selection_snapshot(view) == selection_before
            assert _model_snapshot(view) == before_model

        assert len(set(action_ids)) == 3, action_ids
        assert operation_orders == [1, 2, 3, 4, 5, 6], operation_orders
    finally:
        app._shutdown_root()


def test_column_snapshot_is_released_and_repeated_cycles_have_bounded_rss():
    import psutil

    app, view = _real_two_way_app(decorated_a=(2, 3, 4))
    try:
        process = psutil.Process(os.getpid())
        view._select_column_block_by_logical_col(5, "B")
        before = _worksheet_snapshot(app.ws_a_edit("Sheet1"))
        before_model = _model_snapshot(view)

        # Warm allocators/Tk rendering before the RSS baseline.
        view._apply_selected_column_block("B", "A")
        view._undo_last_action()
        gc.collect()
        baseline_rss = process.memory_info().rss
        rss_samples = []

        for _cycle in range(12):
            view._apply_selected_column_block("B", "A")
            snapshot = app.undo_stack[-1]["snapshot"]
            saved_dimension = snapshot["edit_sheet"]["column_dimensions"][2]
            saved_validation = snapshot["edit_sheet"]["data_validations"].dataValidation[0]
            dimension_ref = weakref.ref(saved_dimension)
            validation_ref = weakref.ref(saved_validation)
            del saved_dimension, saved_validation, snapshot

            view._undo_last_action()
            gc.collect()
            assert dimension_ref() is None, "undo stack retained a column-dimension snapshot"
            assert validation_ref() is None, "undo stack retained a data-validation snapshot"
            assert app.undo_stack == []
            assert app.manual_a_column_ops == []
            assert _worksheet_snapshot(app.ws_a_edit("Sheet1")) == before
            assert _model_snapshot(view) == before_model
            rss_samples.append(process.memory_info().rss)

        # Windows/Python allocators retain arenas, so use a generous bound
        # while still catching one full worksheet snapshot leaked per cycle.
        limit = 32 * 1024 * 1024
        assert rss_samples[-1] <= baseline_rss + limit, (baseline_rss, rss_samples)
        assert max(rss_samples[3:]) - min(rss_samples[3:]) <= limit, rss_samples
        print(
            "MEMORY column snapshot cycles "
            f"baseline={baseline_rss} final={rss_samples[-1]} "
            f"tail_span={max(rss_samples[3:]) - min(rss_samples[3:])}"
        )
    finally:
        app._shutdown_root()


def test_real_gui_three_way_adopt_mine_theirs_and_retain():
    app, view = _real_three_way_app(
        ("A", "M", "B", "C", "D"),
        ("A", "B", "C", "D"),
        ("A", "B", "C", "T", "D"),
        decorated_mine=(2,),
        decorated_theirs=(4,),
    )
    try:
        before_a = _worksheet_snapshot(app.ws_a_edit("Sheet1"))
        before_b = _worksheet_snapshot(app.ws_b_edit("Sheet1"))
        before_base = _worksheet_snapshot(app.ws_base_edit("Sheet1"))
        before_model = _model_snapshot(view)
        initial_widths = (
            app.ws_a_edit("Sheet1").max_column,
            app.ws_a_val("Sheet1").max_column,
            app.ws_base_edit("Sheet1").max_column,
            app.ws_base_val("Sheet1").max_column,
            app.ws_b_edit("Sheet1").max_column,
            app.ws_b_val("Sheet1").max_column,
        )

        view._select_column_block_by_logical_col(2, "A")
        selection_mine = _selection_snapshot(view)
        retained_before = set(view._retained_column_decisions)
        retain = view._apply_selected_column_block("A")
        assert retain.action_kind == "retain" and retain.target_side == "A"
        assert app.manual_a_column_ops == []
        assert _worksheet_snapshot(app.ws_a_edit("Sheet1")) == before_a
        assert len(view._retained_column_decisions) == len(retained_before) + 1
        assert len(app.undo_stack) == 1
        view._undo_last_action()
        assert view._retained_column_decisions == retained_before
        assert _selection_snapshot(view) == selection_mine
        assert _model_snapshot(view) == before_model

        # Mine can also be adopted to the other editable side through the
        # stable API; repeat the full-rescan undo path so Mine/Base/Theirs diff
        # channels and structural-conflict state cannot drift between cycles.
        mine_source = _column_snapshot(app.ws_a_edit("Sheet1"), 2)
        mine_to_theirs_ids = []
        for _cycle in range(3):
            mine_to_theirs = view._apply_selected_column_block("A", "B")
            mine_to_theirs_ids.append(mine_to_theirs.action_id)
            assert mine_to_theirs.action_kind == "insert_copy"
            assert [app.ws_b_edit("Sheet1").cell(1, col).value for col in range(1, 7)] == [
                "A", "M", "B", "C", "T", "D",
            ]
            assert _column_snapshot(app.ws_b_edit("Sheet1"), 2) == mine_source
            assert [op["kind"] for op in app.manual_b_column_ops] == ["insert_cols", "copy_cols"]
            assert (
                app.ws_a_edit("Sheet1").max_column,
                app.ws_a_val("Sheet1").max_column,
                app.ws_base_edit("Sheet1").max_column,
                app.ws_base_val("Sheet1").max_column,
            ) == initial_widths[:4]
            view._undo_last_action()
            assert app.manual_b_column_ops == []
            assert _worksheet_snapshot(app.ws_a_edit("Sheet1")) == before_a
            assert _worksheet_snapshot(app.ws_b_edit("Sheet1")) == before_b
            assert _worksheet_snapshot(app.ws_base_edit("Sheet1")) == before_base
            assert _model_snapshot(view) == before_model
        assert len(set(mine_to_theirs_ids)) == 3, mine_to_theirs_ids

        view._select_column_block_by_logical_col(5, "B")
        theirs_source = _column_snapshot(app.ws_b_edit("Sheet1"), 4)
        theirs_to_mine = view._apply_selected_column_block("B", "A")
        assert theirs_to_mine.action_kind == "insert_copy"
        assert [app.ws_a_edit("Sheet1").cell(1, col).value for col in range(1, 7)] == [
            "A", "M", "B", "C", "T", "D",
        ]
        assert _column_snapshot(app.ws_a_edit("Sheet1"), 5) == theirs_source
        assert [op["kind"] for op in app.manual_a_column_ops] == ["insert_cols", "copy_cols"]
        assert (
            app.ws_base_edit("Sheet1").max_column,
            app.ws_base_val("Sheet1").max_column,
            app.ws_b_edit("Sheet1").max_column,
            app.ws_b_val("Sheet1").max_column,
        ) == initial_widths[2:]
        view._undo_last_action()
        assert app.manual_a_column_ops == []
        assert _worksheet_snapshot(app.ws_a_edit("Sheet1")) == before_a
        assert _worksheet_snapshot(app.ws_b_edit("Sheet1")) == before_b
        assert _worksheet_snapshot(app.ws_base_edit("Sheet1")) == before_base
        assert _model_snapshot(view) == before_model
        assert (
            app.ws_a_edit("Sheet1").max_column,
            app.ws_a_val("Sheet1").max_column,
            app.ws_base_edit("Sheet1").max_column,
            app.ws_base_val("Sheet1").max_column,
            app.ws_b_edit("Sheet1").max_column,
            app.ws_b_val("Sheet1").max_column,
        ) == initial_widths
    finally:
        app._shutdown_root()


def test_real_gui_three_way_repeated_action_undo_restores_exact_diff_state():
    app, view = _real_three_way_app(
        ("A", "M", "B", "C", "D"),
        ("A", "B", "C", "D"),
        ("A", "B", "C", "T", "D"),
        decorated_mine=(2,),
        decorated_theirs=(4,),
    )
    try:
        before = _model_snapshot(view)
        before_projection = (
            tuple(view._active_column_projection().model.slots),
            tuple(view._active_column_projection().model.blocks),
        )
        before_pair_diff = copy.deepcopy(view.pair_diff_cols)
        before_pair_base_diff = copy.deepcopy(view.pair_base_diff_cols)
        before_conflicts = tuple(view.logical_column_structural_conflicts)
        assert view._is_three_way_enabled()

        for _cycle in range(3):
            view._select_column_block_by_logical_col(5, "B")
            plan = view._apply_selected_column_block("B", "A")
            assert plan.action_kind == "insert_copy"
            view._undo_last_action()
            assert _model_snapshot(view) == before
            assert (
                tuple(view._active_column_projection().model.slots),
                tuple(view._active_column_projection().model.blocks),
            ) == before_projection
            assert view.pair_diff_cols == before_pair_diff
            assert view.pair_base_diff_cols == before_pair_base_diff
            assert tuple(view.logical_column_structural_conflicts) == before_conflicts

            view._select_column_block_by_logical_col(2, "A")
            plan = view._apply_selected_column_block("A", "B")
            assert plan.action_kind == "insert_copy"
            view._undo_last_action()
            assert _model_snapshot(view) == before
            assert (
                tuple(view._active_column_projection().model.slots),
                tuple(view._active_column_projection().model.blocks),
            ) == before_projection
            assert view.pair_diff_cols == before_pair_diff
            assert view.pair_base_diff_cols == before_pair_base_diff
            assert tuple(view.logical_column_structural_conflicts) == before_conflicts
    finally:
        app._shutdown_root()


def test_real_gui_adjacent_unresolved_side_presence_splits_and_delete_undo_is_exact():
    app, view = _real_three_way_app(
        ("A", "M", "B", "C"),
        ("A", "B", "C"),
        ("A", "T", "B", "C"),
        decorated_mine=(2,),
        decorated_theirs=(2,),
    )
    try:
        projection = view._active_column_projection()
        unresolved_blocks = [
            block for block in projection.model.blocks if block.state == "unresolved"
        ]
        assert [tuple(block.slot_indices) for block in unresolved_blocks] == [
            (1,), (2,)
        ], unresolved_blocks
        assert (
            projection.model.slots[1].confidence.reason
            == projection.model.slots[2].confidence.reason
            == "ambiguous-competing-insertions"
        )
        assert (
            projection.model.slots[1].mine_col is not None
            and projection.model.slots[1].theirs_col is None
            and projection.model.slots[2].mine_col is None
            and projection.model.slots[2].theirs_col is not None
        )

        before_a = _worksheet_snapshot(app.ws_a_edit("Sheet1"))
        before_b = _worksheet_snapshot(app.ws_b_edit("Sheet1"))
        before_model = _model_snapshot(view)
        selected = view._select_column_block_by_logical_col(2, "B")
        assert selected is not None and tuple(selected.slot_indices) == (1,)
        selection_before_apply = _selection_snapshot(view)
        plan = view._apply_selected_column_block(
            "B", "A", confirm_unresolved=True
        )
        assert plan.action_kind == "delete" and plan.count == 1
        assert [app.ws_a_edit("Sheet1").cell(1, col).value for col in range(1, 4)] == [
            "A", "B", "C"
        ]
        assert [app.ws_b_edit("Sheet1").cell(1, col).value for col in range(1, 5)] == [
            "A", "T", "B", "C"
        ]

        view._undo_last_action()
        assert _worksheet_snapshot(app.ws_a_edit("Sheet1")) == before_a
        assert _worksheet_snapshot(app.ws_b_edit("Sheet1")) == before_b
        assert _model_snapshot(view) == before_model
        assert _selection_snapshot(view) == selection_before_apply
    finally:
        app._shutdown_root()


def test_real_gui_conflict_mode_column_apply_undo_preserves_full_row_model_and_map():
    controlled_conflicts = {"Sheet1": {5: {1}}}
    app, view = _real_three_way_app(
        ("A", "B", "C"),
        ("A", "B", "C"),
        ("A", "T", "B", "C"),
        decorated_theirs=(2,),
        conflict_map=controlled_conflicts,
    )
    try:
        before_model = _model_snapshot(view)
        before_pairs = tuple(view.row_pairs)
        before_conflicts = copy.deepcopy(app.merge_conflict_cells_by_sheet)
        assert len(before_pairs) == 9, before_pairs
        assert before_conflicts == controlled_conflicts

        view._select_column_block_by_logical_col(2, "B")
        plan = view._apply_selected_column_block("B", "A")
        assert plan.action_kind == "insert_copy" and plan.count == 1
        assert tuple(view.row_pairs) == before_pairs
        assert app.merge_conflict_cells_by_sheet == before_conflicts
        assert set(app.merge_conflict_cells_by_sheet["Sheet1"][5]) == {1}

        view._undo_last_action()
        assert tuple(view.row_pairs) == before_pairs
        assert app.merge_conflict_cells_by_sheet == before_conflicts
        assert _model_snapshot(view) == before_model
    finally:
        app._shutdown_root()


def test_real_gui_three_way_adopt_base_restores_deleted_mine_column():
    app, view = _real_three_way_app(
        ("A", "C", "D"),
        ("A", "B", "C", "D"),
        ("A", "B", "C", "D"),
        decorated_base=(2,),
    )
    try:
        before_a_edit = _worksheet_snapshot(app.ws_a_edit("Sheet1"))
        before_a_value = _worksheet_snapshot(app.ws_a_val("Sheet1"))
        before_base = _worksheet_snapshot(app.ws_base_edit("Sheet1"))
        before_b = _worksheet_snapshot(app.ws_b_edit("Sheet1"))
        before_model = _model_snapshot(view)
        block = view._select_column_block_by_logical_col(2, "BASE")
        assert block is not None
        selection_before = _selection_snapshot(view)
        base_source = _column_snapshot(app.ws_base_edit("Sheet1"), 2)

        plan = view._apply_selected_column_block("BASE", "A")
        assert plan.action_kind == "insert_copy"
        assert [app.ws_a_edit("Sheet1").cell(1, col).value for col in range(1, 5)] == [
            "A", "B", "C", "D",
        ]
        assert _column_snapshot(app.ws_a_edit("Sheet1"), 2) == base_source
        assert [op["source_side"] for op in app.manual_a_column_ops] == ["BASE", "BASE"]
        assert not view.column_comparison_cache.structural_diff_cols

        view._undo_last_action()
        assert app.manual_a_column_ops == []
        assert _worksheet_snapshot(app.ws_a_edit("Sheet1")) == before_a_edit
        assert _worksheet_snapshot(app.ws_a_val("Sheet1")) == before_a_value
        assert _worksheet_snapshot(app.ws_base_edit("Sheet1")) == before_base
        assert _worksheet_snapshot(app.ws_b_edit("Sheet1")) == before_b
        assert _selection_snapshot(view) == selection_before
        assert _model_snapshot(view) == before_model
    finally:
        app._shutdown_root()


def test_real_gui_column_undo_preserves_prior_cell_and_row_actions():
    app, view = _real_two_way_app()
    try:
        sheet = "Sheet1"
        ws_edit = app.ws_a_edit(sheet)
        ws_value = app.ws_a_val(sheet)

        # Seed a real explicit cell action before the structural actions.
        old_edit = ws_edit.cell(2, 6).value
        old_value = ws_value.cell(2, 6).value
        ws_edit.cell(2, 6).value = "manual-cell-before-column"
        ws_value.cell(2, 6).value = "manual-cell-before-column"
        app.record_manual_a_cell(sheet, 2, 6, "manual-cell-before-column")
        app.push_undo({"sheet": sheet, "target": "A", "cells": [(2, 6, old_edit, old_value)]})

        # Seed an actual appended/inserted row plus its normal row operation
        # and undo record.  The column action must remain exactly one later
        # stack item and its undo must not consume or rewrite these records.
        insert_row = ws_edit.max_row + 1
        ws_edit.insert_rows(insert_row, 1)
        ws_value.insert_rows(insert_row, 1)
        for col, header in enumerate(("A", "B", "C", "D", "E", "F"), start=1):
            value = f"mixed-row-{header}"
            ws_edit.cell(insert_row, col).value = value
            ws_value.cell(insert_row, col).value = value
        app.record_manual_a_row_insert(sheet, insert_row, 1)
        app.push_undo({
            "sheet": sheet,
            "target": "A_INSERT_ROW",
            "row": insert_row,
            "count": 1,
            "base_inserted": False,
        })
        app.modified_a = True
        app.modified_sheets_a.add(sheet)
        view._mark_column_mapping_stale(
            "mixed-cell-row-column-test",
            row_structure=True,
            edited_sides=("A",),
        )
        view.refresh(row_only=None, rescan=True)

        before_column_edit = _worksheet_snapshot(app.ws_a_edit(sheet))
        before_column_value = _worksheet_snapshot(app.ws_a_val(sheet))
        before_cell_ops = dict(app.manual_a_cell_ops)
        before_row_ops = list(app.manual_a_row_ops)
        before_undo = list(app.undo_stack)
        before_model = _model_snapshot(view)
        view._select_column_block_by_logical_col(2, "B")
        selection_before = _selection_snapshot(view)

        plan = view._apply_selected_column_block("B", "A")
        assert plan.action_kind == "insert_copy"
        assert len(app.undo_stack) == len(before_undo) + 1
        assert app.undo_stack[-1]["kind"] == "column_action"
        view._undo_last_action()

        assert app.undo_stack == before_undo
        assert app.manual_a_cell_ops == before_cell_ops
        assert app.manual_a_row_ops == before_row_ops
        assert app.manual_a_column_ops == []
        assert _worksheet_snapshot(app.ws_a_edit(sheet)) == before_column_edit
        assert _worksheet_snapshot(app.ws_a_val(sheet)) == before_column_value
        assert app.ws_a_edit(sheet).cell(2, 6).value == "manual-cell-before-column"
        assert [app.ws_a_edit(sheet).cell(insert_row, col).value for col in range(1, 7)] == [
            "mixed-row-A", "mixed-row-B", "mixed-row-C",
            "mixed-row-D", "mixed-row-E", "mixed-row-F",
        ]
        assert _selection_snapshot(view) == selection_before
        assert _model_snapshot(view) == before_model
    finally:
        app._shutdown_root()


def test_real_gui_same_anchor_partial_conflict_requires_confirmation_and_stays_safe():
    app, view = _real_three_way_app(
        ("A", "M", "B", "C"),
        ("A", "B", "C"),
        ("A", "T", "B", "C"),
        decorated_mine=(2,),
        decorated_theirs=(2,),
    )
    try:
        before_a = _worksheet_snapshot(app.ws_a_edit("Sheet1"))
        before_base = _worksheet_snapshot(app.ws_base_edit("Sheet1"))
        before_b = _worksheet_snapshot(app.ws_b_edit("Sheet1"))
        before_model = _model_snapshot(view)
        block = view._select_column_block_by_logical_col(2, "A")
        assert block is not None and block.state == "unresolved"
        assert block.slot_indices == (1,)
        selection_before = _selection_snapshot(view)

        _raises_runtime(
            lambda: view._apply_selected_column_block("A", "A"),
            "必须显式确认",
        )
        plan = view._apply_selected_column_block(
            "A", "B", confirm_unresolved=True
        )
        assert plan.action_kind == "insert_copy" and plan.count == 1
        assert plan.source_physical_cols == (2,)
        assert plan.target_physical_cols == ()
        assert plan.target_physical_anchor == 2
        assert tuple(
            app.ws_b_edit("Sheet1").cell(1, col).value
            for col in range(1, 6)
        ) == ("A", "M", "T", "B", "C")
        assert _worksheet_snapshot(app.ws_a_edit("Sheet1")) == before_a
        assert _worksheet_snapshot(app.ws_base_edit("Sheet1")) == before_base
        assert [op["kind"] for op in app.manual_b_column_ops] == [
            "insert_cols", "copy_cols"
        ], (
            app.manual_a_column_ops,
            app.manual_b_column_ops,
        )
        assert len(app.undo_stack) == 1

        view._undo_last_action()
        assert _worksheet_snapshot(app.ws_a_edit("Sheet1")) == before_a
        assert _worksheet_snapshot(app.ws_base_edit("Sheet1")) == before_base
        assert _worksheet_snapshot(app.ws_b_edit("Sheet1")) == before_b
        assert app.manual_a_column_ops == []
        assert app.manual_b_column_ops == []
        assert app.undo_stack == []
        assert _selection_snapshot(view) == selection_before
        assert _model_snapshot(view) == before_model
        _raises_runtime(lambda: view._action_physical_columns("B2A", 2), "映射待确认")
    finally:
        app._shutdown_root()


def test_real_guide_projection_stays_bounded_across_apply_and_undo():
    assert os.path.exists(_GUIDE_ORIGINAL), _GUIDE_ORIGINAL
    assert os.path.exists(_GUIDE_INSERT2_DELETE1), _GUIDE_INSERT2_DELETE1
    hashes_before = {
        _GUIDE_ORIGINAL: _sha256(_GUIDE_ORIGINAL),
        _GUIDE_INSERT2_DELETE1: _sha256(_GUIDE_INSERT2_DELETE1),
    }
    app = smt.SowMergeApp(_GUIDE_ORIGINAL, _GUIDE_INSERT2_DELETE1)
    try:
        view = _wait_for_view(app, _GUIDE_SHEET, timeout=30.0)
        initial_fingerprint = _wait_for_stable_projection(view)
        initial_model = _model_snapshot(view)
        initial_selection = None
        assert len(initial_fingerprint[0]) == 21, len(initial_fingerprint[0])
        assert initial_fingerprint[2] == frozenset({12, 13})
        assert initial_fingerprint[3] == frozenset()
        initial_widths = (
            app.ws_a_edit(_GUIDE_SHEET).max_column,
            app.ws_a_val(_GUIDE_SHEET).max_column,
            app.ws_b_edit(_GUIDE_SHEET).max_column,
            app.ws_b_val(_GUIDE_SHEET).max_column,
        )
        initial_a_edit = _worksheet_snapshot(app.ws_a_edit(_GUIDE_SHEET))
        initial_a_value = _worksheet_snapshot(app.ws_a_val(_GUIDE_SHEET))
        block = view._select_column_block_by_logical_col(12, "B")
        assert block is not None and block.slot_indices == (11, 12)
        initial_selection = _selection_snapshot(view)

        started = time.perf_counter()
        plan = view._apply_selected_column_block("B", "A")
        apply_seconds = time.perf_counter() - started
        after_apply = _wait_for_stable_projection(view)
        assert plan.action_kind == "insert_copy" and plan.count == 2
        assert plan.source_physical_cols == (12, 13)
        assert plan.target_physical_anchor == 12
        # The real sheet has a formatted/blank physical tail (33/34 columns)
        # but only 21 logical data columns.  Adopting L12:L13 must never turn
        # that physical tail into 35 logical slots or unresolved noise.
        assert len(after_apply[0]) == 21, len(after_apply[0])
        assert after_apply[2] == frozenset(), after_apply[2]
        assert after_apply[3] == frozenset(), after_apply[3]
        assert app.ws_a_edit(_GUIDE_SHEET).max_column == initial_widths[0] + 2
        assert app.ws_a_val(_GUIDE_SHEET).max_column == initial_widths[1] + 2
        assert app.ws_b_edit(_GUIDE_SHEET).max_column == initial_widths[2]
        assert app.ws_b_val(_GUIDE_SHEET).max_column == initial_widths[3]

        started = time.perf_counter()
        view._undo_last_action()
        undo_seconds = time.perf_counter() - started
        after_undo = _wait_for_stable_projection(view)
        assert after_undo == initial_fingerprint
        assert _model_snapshot(view) == initial_model
        assert _selection_snapshot(view) == initial_selection
        assert app.manual_a_column_ops == []
        assert _worksheet_snapshot(app.ws_a_edit(_GUIDE_SHEET)) == initial_a_edit
        assert _worksheet_snapshot(app.ws_a_val(_GUIDE_SHEET)) == initial_a_value
        assert (
            app.ws_a_edit(_GUIDE_SHEET).max_column,
            app.ws_a_val(_GUIDE_SHEET).max_column,
            app.ws_b_edit(_GUIDE_SHEET).max_column,
            app.ws_b_val(_GUIDE_SHEET).max_column,
        ) == initial_widths
        print(
            "TIMING real Guide L12:L13 "
            f"apply={apply_seconds:.3f}s undo={undo_seconds:.3f}s"
        )
    finally:
        app._shutdown_root()
    assert _sha256(_GUIDE_ORIGINAL) == hashes_before[_GUIDE_ORIGINAL]
    assert _sha256(_GUIDE_INSERT2_DELETE1) == hashes_before[_GUIDE_INSERT2_DELETE1]


def test_real_gui_cross_sheet_column_undo_routes_to_action_sheet():
    app, view1, view2 = _real_two_sheet_app()
    try:
        initial_sheet1 = _worksheet_snapshot(app.ws_a_edit("Sheet1"))
        initial_sheet2 = _worksheet_snapshot(app.ws_a_edit("Sheet2"))
        initial_model1 = _model_snapshot(view1)
        initial_model2 = _model_snapshot(view2)

        view1._select_column_block_by_logical_col(2, "B")
        selection1 = _selection_snapshot(view1)
        first = view1._apply_selected_column_block("B", "A")
        _wait_for_stable_projection(view1)
        adopted_sheet1 = _worksheet_snapshot(app.ws_a_edit("Sheet1"))
        adopted_model1 = _model_snapshot(view1)
        assert adopted_sheet1 != initial_sheet1

        view2._select_column_block_by_logical_col(2, "B")
        selection2 = _selection_snapshot(view2)
        second = view2._apply_selected_column_block("B", "A")
        _wait_for_stable_projection(view2)
        assert _worksheet_snapshot(app.ws_a_edit("Sheet2")) != initial_sheet2
        assert [action["plan"].sheet for action in app.undo_stack] == ["Sheet1", "Sheet2"]

        # Invoke undo from the other tab/view.  The global latest action is on
        # Sheet2, so workbook restore, cache rebuild and selection restoration
        # must all be routed to view2 rather than corrupting view1 state.
        view1._undo_last_action()
        _wait_for_stable_projection(view1)
        _wait_for_stable_projection(view2)
        assert len(app.undo_stack) == 1 and app.undo_stack[0]["plan"] == first
        assert _worksheet_snapshot(app.ws_a_edit("Sheet1")) == adopted_sheet1
        assert _worksheet_snapshot(app.ws_a_edit("Sheet2")) == initial_sheet2
        assert _model_snapshot(view1) == adopted_model1
        assert _model_snapshot(view2) == initial_model2
        assert _selection_snapshot(view2) == selection2
        assert [op["sheet"] for op in app.manual_a_column_ops] == ["Sheet1", "Sheet1"]

        view1._undo_last_action()
        _wait_for_stable_projection(view1)
        _wait_for_stable_projection(view2)
        assert app.undo_stack == []
        assert app.manual_a_column_ops == []
        assert _worksheet_snapshot(app.ws_a_edit("Sheet1")) == initial_sheet1
        assert _worksheet_snapshot(app.ws_a_edit("Sheet2")) == initial_sheet2
        assert _model_snapshot(view1) == initial_model1
        assert _model_snapshot(view2) == initial_model2
        assert _selection_snapshot(view1) == selection1
        assert first.action_id != second.action_id
    finally:
        app._shutdown_root()


def main():
    tests = [
        test_column_action_plan_and_operation_shape,
        test_delete_copy_retain_and_global_operation_order,
        test_consecutive_columns_stay_one_batch_and_do_not_expand_neighbors,
        test_invalid_column_operation_batch_is_rejected_atomically,
        test_column_action_plan_validation,
        test_row_cache_reuses_physical_width_scan_for_thousands_of_calls,
        test_real_worksheet_wide_reads_do_not_materialize_tail_and_invalidation_updates_width,
        test_sparse_regular_and_read_only_row_cache_are_equivalent_without_materialization,
        test_formula_cache_asymmetry_keeps_identity_rows_and_column_only_diff_actions,
        test_real_gui_cell_apply_and_undo_reuse_mapping_without_full_scan,
        test_real_gui_region_apply_and_undo_reuse_mapping_without_full_scan,
        test_live_formula_references_follow_excel_insert_delete_and_undo_exactly,
        test_two_way_insert_delete_copy_and_retain_plans_use_logical_block,
        test_manual_formula_ops_follow_excel_column_reference_semantics,
        test_formula_capture_prefilter_keeps_target_and_qualified_references_only,
        test_three_way_mine_base_theirs_plans_and_same_anchor_confirmation,
        test_ready_auto_selects_first_real_structural_block_only,
        test_first_direct_row_apply_revalidates_false_pending_mapping,
        test_consecutive_row_apply_ignores_style_only_blank_tail_columns,
        test_full_row_apply_skips_unresolved_blank_gap_before_real_formula_column,
        test_ordinary_cell_action_rejects_missing_or_unresolved_slot,
        test_real_gui_insert_block_and_one_step_undo_full_fidelity,
        test_real_gui_delete_block_preserves_adjacent_columns_and_undo,
        test_real_gui_failure_injection_is_atomic_at_every_mutating_stage,
        test_real_gui_repeated_adopt_undo_has_no_state_drift,
        test_column_snapshot_is_released_and_repeated_cycles_have_bounded_rss,
        test_real_gui_three_way_adopt_mine_theirs_and_retain,
        test_real_gui_three_way_repeated_action_undo_restores_exact_diff_state,
        test_real_gui_adjacent_unresolved_side_presence_splits_and_delete_undo_is_exact,
        test_real_gui_conflict_mode_column_apply_undo_preserves_full_row_model_and_map,
        test_real_gui_three_way_adopt_base_restores_deleted_mine_column,
        test_real_gui_column_undo_preserves_prior_cell_and_row_actions,
        test_real_gui_same_anchor_partial_conflict_requires_confirmation_and_stays_safe,
        test_real_guide_projection_stays_bounded_across_apply_and_undo,
        test_real_gui_cross_sheet_column_undo_routes_to_action_sheet,
    ]
    for test in tests:
        test()
        print(f"PASS: {test.__name__}")
    print(f"PASS: logical column action regression ({len(tests)} tests)")


if __name__ == "__main__":
    main()

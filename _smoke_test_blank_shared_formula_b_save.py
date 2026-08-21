"""Blank Base-side cells must safely clear shared formulas on B-side save."""

from __future__ import annotations

import hashlib
import json
import os
import sys
import tempfile
import time
import zipfile
import xml.etree.ElementTree as ET
from contextlib import contextmanager
from types import SimpleNamespace

from openpyxl import Workbook, load_workbook

import sow_merge_tool as mod


SHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


def _sha256(path: str) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _make_book(path: str, *, formulas: bool):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    worksheet.append(["id@id", "名称", "输入", "公式"])
    worksheet.append(["string", "string", "number", "formula"])
    worksheet.append(["row-1", "one", 10, "=C3" if formulas else None])
    worksheet.append(["row-2", "two", 20, "=C4" if formulas else None])
    workbook.save(path)
    workbook.close()


def _inject_shared_formula(path: str):
    part = "xl/worksheets/sheet1.xml"
    with zipfile.ZipFile(path, "r") as source:
        infos = source.infolist()
        payloads = {info.filename: source.read(info.filename) for info in infos}
    root = ET.fromstring(payloads[part])
    q = lambda name: f"{{{SHEET_NS}}}{name}"
    cells = {node.attrib.get("r"): node for node in root.iter(q("c")) if node.attrib.get("r")}
    master = cells["D3"].find(q("f"))
    follower = cells["D4"].find(q("f"))
    assert master is not None and follower is not None
    master.attrib.clear()
    master.attrib.update({"t": "shared", "ref": "D3:D4", "si": "0"})
    master.text = "C3"
    follower.attrib.clear()
    follower.attrib.update({"t": "shared", "si": "0"})
    follower.text = None
    payloads[part] = ET.tostring(root, encoding="utf-8", xml_declaration=True)
    rewritten = path + ".rewrite"
    with zipfile.ZipFile(rewritten, "w") as target:
        for info in infos:
            target.writestr(info, payloads[info.filename])
    os.replace(rewritten, path)


@contextmanager
def _owned_case():
    original_settings_path = mod._SETTINGS_PATH
    original_settings_exists = os.path.lexists(original_settings_path)
    if original_settings_exists:
        with open(original_settings_path, "rb") as stream:
            original_settings_bytes = stream.read()
    else:
        original_settings_bytes = None
    temporary = tempfile.TemporaryDirectory(prefix="sow_blank_shared_formula_")
    root_dir = temporary.name
    settings_path = os.path.join(root_dir, "settings.json")
    with open(settings_path, "w", encoding="utf-8") as stream:
        json.dump({"only_diff": 0}, stream)
    original_prompt_scheduler = mod.SowMergeApp._schedule_formula_cache_prompt
    state = SimpleNamespace(app=None, input_hashes={}, outputs=[], root_dir=root_dir)
    mod._SETTINGS_PATH = settings_path
    mod.SowMergeApp._schedule_formula_cache_prompt = lambda _self: None
    primary = None
    try:
        yield state
    except BaseException as exc:
        primary = exc
        raise
    finally:
        cleanup_errors = []
        if state.app is not None:
            try:
                for candidate in (state.app, *getattr(state.app, "sheet_views", {}).values()):
                    if candidate is None:
                        continue
                    after_id = getattr(candidate, "_settings_save_id", None)
                    if after_id is not None:
                        state.app.root.after_cancel(after_id)
                        candidate._settings_save_id = None
                state.app._shutdown_root()
            except BaseException as exc:
                cleanup_errors.append(f"shutdown: {exc!r}")
        try:
            mod.SowMergeApp._schedule_formula_cache_prompt = original_prompt_scheduler
            mod._SETTINGS_PATH = original_settings_path
            if original_settings_exists:
                with open(original_settings_path, "rb") as stream:
                    assert stream.read() == original_settings_bytes
            else:
                assert not os.path.lexists(original_settings_path)
        except BaseException as exc:
            cleanup_errors.append(f"settings/prompt restore: {exc!r}")
        for path, before_hash in state.input_hashes.items():
            try:
                assert _sha256(path) == before_hash, path
            except BaseException as exc:
                cleanup_errors.append(f"input SHA {path!r}: {exc!r}")
        for output in state.outputs:
            try:
                if os.path.lexists(output):
                    os.remove(output)
                assert not os.path.lexists(output), output
            except BaseException as exc:
                cleanup_errors.append(f"output cleanup {output!r}: {exc!r}")
        try:
            temporary.cleanup()
            assert not os.path.lexists(root_dir), root_dir
        except BaseException as exc:
            cleanup_errors.append(f"owned temporary root: {exc!r}")
        if cleanup_errors:
            message = "blank shared-formula cleanup failed: " + "; ".join(cleanup_errors)
            if primary is not None:
                primary.add_note(message)
            else:
                raise AssertionError(message)


def _wait_operation_ready(app, sheet: str, deadline: float):
    app.nb.select(app._sheet_containers[sheet])
    view = None
    while time.monotonic() < deadline:
        app.root.update_idletasks()
        app.root.update()
        view = app.sheet_views.get(sheet)
        if (
            view is not None
            and view._data_ready
            and app._is_sheet_exact_current(sheet)
            and app._edit_workbooks_ready()
            and view._derive_lifecycle_state() == "READY"
        ):
            return view
        time.sleep(0.01)
    raise AssertionError(("Data did not become operation-ready", app._sheet_exact_entry(sheet), view))


def main():
    deadline = time.monotonic() + 90.0
    with _owned_case() as owned:
        base = os.path.join(owned.root_dir, "base.xlsx")
        mine = os.path.join(owned.root_dir, "mine.xlsx")
        _make_book(base, formulas=False)
        _make_book(mine, formulas=True)
        _inject_shared_formula(mine)
        owned.input_hashes = {base: _sha256(base), mine: _sha256(mine)}

        app = mod.SowMergeApp(base, mine)
        owned.app = app
        app._request_edit_preload()
        view = _wait_operation_ready(app, "Data", deadline)
        base_row = next(
            row for row in range(1, app.ws_a_val("Data").max_row + 1)
            if app.ws_a_val("Data").cell(row, 1).value == "row-1"
        )
        pair_idx = view.row_a_to_pair_idx[base_row]
        mine_row = view.row_pairs[pair_idx][1]
        logical_col = view._logical_col_for_physical("A", 4)
        assert mine_row is not None and logical_col == 4
        assert logical_col in view.pair_diff_cols[pair_idx]
        assert view._copy_selected_row(
            "A2B",
            override_pair_idx=pair_idx,
            override_cols={logical_col},
        )
        assert app.ws_b_edit("Data").cell(mine_row, 4).value is None
        assert ("Data", mine_row, 4) in app.manual_b_cell_ops

        native_save_enabled = mod._EXCEL_NATIVE_SAVE_ON_MERGE
        mod._EXCEL_NATIVE_SAVE_ON_MERGE = False
        try:
            output = app.build_manual_b_output_file()
            owned.outputs.append(output)
        finally:
            mod._EXCEL_NATIVE_SAVE_ON_MERGE = native_save_enabled
        saved = load_workbook(output, read_only=True, data_only=False)
        try:
            assert saved["Data"].cell(mine_row, 4).value is None
            assert saved["Data"].cell(mine_row + 1, 4).value == "=C4"
        finally:
            saved.close()
        assert _sha256(base) == owned.input_hashes[base]
        assert _sha256(mine) == owned.input_hashes[mine]

    print("SMOKE_BLANK_SHARED_FORMULA_B_SAVE_OK", flush=True)


if __name__ == "__main__":
    main()
